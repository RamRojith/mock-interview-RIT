import re
from django.shortcuts import render,redirect
from course_management.models import CourseandexaminationFunction
from user_accounts.decorators import check_permission, faculty_login_required, no_cache, is_super_user
from user_accounts.models import Role
from library_management.models import Library_Permissions
from library_management.decorators import library_management
from django.contrib import messages
from faculty_management.models import general_information

@faculty_login_required
@no_cache
@is_super_user('library_management')
def library_assign_permission(request):
    if request.method == 'POST':  
        permissions = request.POST
        for role_name, role_permissions in permissions.items():
            if role_name.startswith('permissions'):
                try:

                    extract_data = list(re.findall(r'\[([^\]]+)\]', role_name))
                    if len(extract_data) < 2: 
                        messages.warning(request,f"Invalid format in role_name: {role_name}. Skipping.")
                        continue
                    
                    extract_data.append(role_permissions)


                    try:
                        role = Role.objects.using("rit_approval_system").get(role=extract_data[0])

                    except Role.DoesNotExist:
                        messages.error(request,f"Role {extract_data[0]} does not exist.")
                        messages.error(request, f"Role '{extract_data[0]}' does not exist. Skipping this entry.")
                        continue
                    

                    if isinstance(role_permissions, list):
                        role_permissions = role_permissions[0]
                    

                    permission = extract_data[2] == 'true'
       
                    

                    permission_obj = Library_Permissions.objects.filter(
                        role=role, function=extract_data[1]
                    ).first()
                    
                    if permission_obj:
                        permission_obj.permission = permission
                        permission_obj.save()
                    else:

                        Library_Permissions.objects.create(
                            role=role,
                            function=extract_data[1],
                            permission=permission
                        )
                except Exception as e:

                    messages.error(request,f"Error processing role '{role_name}': {str(e)}")
                    messages.error(request, f"An error occurred while processing '{role_name}': {str(e)}")


    messages.success(request,"The permission changes have been successfully applied.")
    return redirect('library_management')
 


@library_management
def library_home(request):

    request.session['current_page'] = 'library_home'

    return redirect('home')

@library_management
@check_permission("library_hello")
def library_hello(request):
    return render(request, "library_hello.html")





from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from datetime import date
from io import BytesIO
import re

from django.http import HttpResponse
from django.core.paginator import Paginator
from django.db.models import Count, Max, Q
from django.views.decorators.http import require_POST
from django.db import transaction
from django.utils.dateparse import parse_date

import openpyxl

from library_management.decorators import library_management as library_mgmt_required
from library_management.models import BookType, LibraryBook
from faculty_management.models import general_information, Add_Department, Degree



# =========================================================
# BOOK TYPE (MASTER) - ADD / EDIT / DELETE
# =========================================================


def book_type_page(request):
    if request.method == "POST":
        book_type = (request.POST.get("book_type") or "").strip()

        if not book_type:
            messages.error(request, "Book Type cannot be empty.")
            return redirect("book_type_page")

        normalized = book_type.upper()

        if BookType.objects.filter(book_type__iexact=normalized).exists():
            messages.warning(request, "This Book Type already exists.")
            return redirect("book_type_page")

        BookType.objects.create(book_type=normalized)
        messages.success(request, "Book Type saved successfully!")
        return redirect("book_type_page")

    book_types = BookType.objects.all().order_by("-id")
    return render(request, "library_management/admin/book_type.html", {
        "book_types": book_types,
        "edit_mode": False
    })



def edit_book_type(request, pk):
    obj = get_object_or_404(BookType, pk=pk)

    if request.method == "POST":
        book_type = (request.POST.get("book_type") or "").strip()

        if not book_type:
            messages.error(request, "Book Type cannot be empty.")
            return redirect("edit_book_type", pk=obj.id)

        normalized = book_type.upper()

        if BookType.objects.filter(book_type__iexact=normalized).exclude(id=obj.id).exists():
            messages.warning(request, "This Book Type already exists.")
            return redirect("edit_book_type", pk=obj.id)

        obj.book_type = normalized
        obj.save()
        messages.success(request, "Book Type updated successfully!")
        return redirect("book_type_page")

    book_types = BookType.objects.all().order_by("-id")
    return render(request, "library_management/admin/book_type.html", {
        "book_types": book_types,
        "edit_mode": True,
        "edit_obj": obj
    })



def delete_book_type(request, pk):
    BookType.objects.filter(pk=pk).delete()
    messages.success(request, "Book Type deleted successfully!")
    return redirect("book_type_page")


# =========================================================
# ACADEMIC YEAR HELPER
# =========================================================

def get_academic_year():
    today = date.today()
    y = today.year
    return f"{y}-{y+1}" if today.month >= 6 else f"{y-1}-{y}"


# =========================================================
# EXCEL UPLOAD CONFIG
# =========================================================

REQUIRED_COLUMNS = [
    "dept_code",
    "accession_book",
    "first_edition_year",
    "title",
    "title_no",
    "authors",
    "publisher",
    "book_type",
]

ALIASES = {
    "sl no": "sl_no", "slno": "sl_no", "sl.no": "sl_no", "slno.": "sl_no",
    "dept code": "dept_code", "deptcode": "dept_code",
    "accession book": "accession_book", "accessionbook": "accession_book",
    "first edition year": "first_edition_year", "firsteditionyear": "first_edition_year",

    "title": "title",

    "title number": "title_no", "titlenumber": "title_no",
    "title no": "title_no", "titleno": "title_no",
    "titile number": "title_no", "titilenumber": "title_no",
    "titile no": "title_no", "titileno": "title_no",

    "volume": "volume",
    "authors": "authors",
    "publisher": "publisher",
    "address": "address",

    "mobile no": "mobile_no", "mobileno": "mobile_no", "mobile number": "mobile_no",

    "type": "book_type", "book type": "book_type", "booktype": "book_type",
}


def normalize_header_cell(val):
    s = str(val or "").strip().lower()
    s = s.replace("\n", " ").replace("\r", " ").replace("\u00a0", " ")
    s = s.replace("_", " ").replace("-", " ").replace("/", " ")
    for ch in [".", ":", ";", ",", "(", ")", "[", "]", "{", "}", "'", '"']:
        s = s.replace(ch, "")
    return " ".join(s.split())


def _cell(v):
    return "" if v is None else str(v).strip()


def get_merged_value(ws, cell):
    if cell.value is not None:
        return cell.value
    for r in ws.merged_cells.ranges:
        if cell.coordinate in r:
            return ws.cell(r.min_row, r.min_col).value
    return None


def clean_mobile(v: str) -> str:
    s = (v or "").strip()
    digits = re.sub(r"\D", "", s)
    return digits[:11]


# =========================================================
# ✅ DELETE ALL BOOKS (POST)
# =========================================================


def delete_all_library_books(request):
    try:
        with transaction.atomic():
            deleted_count, _ = LibraryBook.objects.all().delete()
        messages.success(request, f"✅ All books deleted successfully! ({deleted_count} records removed)")
    except Exception as e:
        messages.error(request, f"❌ Error deleting records: {str(e)}")
    return redirect("library_book_entry")


# =========================================================
# ✅ DELETE SINGLE BOOK (POST)
# =========================================================


def delete_library_book(request, pk):
    deleted, _ = LibraryBook.objects.filter(pk=pk).delete()

    if deleted:
        messages.success(request, "✅ Book deleted successfully!")
    else:
        messages.error(request, "⚠️ Book not found (already deleted).")

    return redirect("library_book_entry")

# =========================================================
# LIBRARY BOOK ENTRY (UPLOAD + LIST + FILTER + 1..N ORDER)
# =========================================================


def library_book_entry(request):
    book_types = BookType.objects.all().order_by("book_type")

    # ✅ same faculty/department logic
    try:
        faculty = general_information.objects.get(faculty_id=request.user.Employee_id)
        department = faculty.department
    except general_information.DoesNotExist:
        messages.error(request, "General Information (general_information) not found for this login.")
        return render(request, "library_management/admin/library_book_form.html", {
            "book_types": book_types,
            "books": [],
            "title_choices": [],
            "selected_title": "",
        })

    # ===================== POST: UPLOAD =====================
    if request.method == "POST":
        excel = request.FILES.get("excel_file")

        if not excel:
            messages.error(request, "Please choose an Excel file (.xlsx).")
            return redirect("library_book_entry")

        if not excel.name.lower().endswith(".xlsx"):
            messages.error(request, "Only .xlsx files are supported.")
            return redirect("library_book_entry")

        try:
            wb = openpyxl.load_workbook(excel, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, "Invalid Excel file.")
            return redirect("library_book_entry")

        header_row_num = None
        header_tokens = None

        scan_upto = min(ws.max_row, 60)
        for r in range(1, scan_upto + 1):
            tokens = []
            for c in ws[r]:
                raw_val = get_merged_value(ws, c)
                raw = normalize_header_cell(raw_val)
                raw_compact = raw.replace(" ", "")

                mapped = ALIASES.get(raw) or ALIASES.get(raw_compact)

                if not mapped:
                    if (("title" in raw) or ("titile" in raw)) and (("number" in raw) or ("no" in raw)):
                        mapped = "title_no"
                    elif ("mobile" in raw) and (("no" in raw) or ("number" in raw)):
                        mapped = "mobile_no"
                    elif ("accession" in raw) and (("book" in raw) or ("no" in raw) or ("number" in raw)):
                        mapped = "accession_book"
                    elif ("dept" in raw) and ("code" in raw):
                        mapped = "dept_code"
                    elif ("first" in raw) and ("edition" in raw) and ("year" in raw):
                        mapped = "first_edition_year"
                    elif raw:
                        mapped = raw.replace(" ", "_")

                if mapped in ("title_number", "titile_number"):
                    mapped = "title_no"

                tokens.append(mapped or "")

            found = set(t for t in tokens if t)
            if len(found.intersection(set(REQUIRED_COLUMNS))) >= 5:
                header_row_num = r
                header_tokens = tokens
                break

        if not header_row_num:
            messages.error(request, "Header row not detected. Please use sample excel format.")
            return redirect("library_book_entry")

        idx = {}
        for i, key in enumerate(header_tokens):
            if key and key not in idx:
                idx[key] = i

        missing = [c for c in REQUIRED_COLUMNS if c not in idx]
        if missing:
            messages.error(request, f"Excel columns missing: {', '.join(missing)}")
            return redirect("library_book_entry")

        created = 0
        skipped = 0
        errors = 0

        for r in range(header_row_num + 1, ws.max_row + 1):
            row = ws[r]

            def g(colname):
                if colname not in idx:
                    return ""
                i = idx[colname]
                if i >= len(row):
                    return ""
                return _cell(row[i].value)

            dept_code = g("dept_code")
            accession_book = g("accession_book")
            year_raw = g("first_edition_year")
            title = g("title")
            title_no = g("title_no")
            volume = g("volume")
            authors = g("authors")
            publisher = g("publisher")
            address = g("address")
            mobile_no = clean_mobile(g("mobile_no"))
            book_type_raw = g("book_type")

            if not any([dept_code, accession_book, title, title_no, book_type_raw]):
                continue

            if not all([dept_code, accession_book, year_raw, title, title_no, authors, publisher, book_type_raw]):
                skipped += 1
                continue

            try:
                year_int = int(str(year_raw).strip()[:4])
            except Exception:
                skipped += 1
                continue

            bt = BookType.objects.filter(book_type__iexact=book_type_raw).first()
            if not bt:
                bt = BookType.objects.create(book_type=book_type_raw.strip().upper())

            if LibraryBook.objects.filter(accession_book__iexact=accession_book).exists():
                skipped += 1
                continue

            try:
                LibraryBook.objects.create(
                    dept_code=dept_code,
                    accession_book=accession_book,
                    first_edition_year=year_int,
                    title=title,
                    title_no=title_no,
                    volume=volume or None,
                    authors=authors,
                    publisher=publisher,
                    address=address or None,
                    mobile_no=mobile_no or None,
                    book_type=bt,
                    department=department,
                    faculty=faculty,
                    academic_year=get_academic_year(),
                )
                created += 1
            except Exception as e:
                errors += 1
                # print("❌ ERROR row", r, "->", str(e))

        messages.success(request, f"Upload done ✅ Created: {created}, Skipped: {skipped}, Errors: {errors}")
        return redirect("library_book_entry")

    # ===================== GET: LIST + FILTER + FIRST->LAST =====================
    selected_title = (request.GET.get("title") or "").strip()

    books_qs = LibraryBook.objects.select_related("book_type").order_by("id")  # ✅ ASC

    if selected_title:
        books_qs = books_qs.filter(title__icontains=selected_title)

    title_choices = list(
        LibraryBook.objects.exclude(title__isnull=True)
        .exclude(title__exact="")
        .values_list("title", flat=True)
        .distinct()
        .order_by("title")
    )

    paginator = Paginator(books_qs, 50)
    page_number = request.GET.get("page")
    books = paginator.get_page(page_number)

    return render(request, "library_management/admin/library_book_form.html", {
        "book_types": book_types,
        "books": books,
        "title_choices": title_choices,
        "selected_title": selected_title,
    })


# =========================================================
# SAMPLE EXCEL DOWNLOAD
# =========================================================


def library_books_sample_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append([
        "SL.NO", "DEPT CODE", "ACCESSION BOOK", "FIRST EDITION YEAR", "TITLE", "TITILE NUMBER",
        "VOLUME", "AUTHORS", "PUBLISHER", "ADDRESS", "MOBILE NO", "TYPE"
    ])

    ws.append([
        "1", "RIT/AI&DS/001", "1", "2022-DEC", "COMPUTER NETWORKS", "1",
        "1", "DR.J.SATHYA PRIYA", "MAGNUS PUBLICATIONS", "CHENNAI", "9840788450", "BOOK"
    ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="library_books_sample.xlsx"'
    return resp




from django.shortcuts import render
from django.db.models import Q, Count
from django.utils.dateparse import parse_date

from library_management.models import LibraryBook, LibraryBookRequest, BookType
from faculty_management.models import Degree, Add_Department


def library_analytics_dashboard(request):
    qs = LibraryBook.objects.select_related("book_type", "department", "department__degree").all()

    ay = (request.GET.get("ay") or "").strip()
    bt = (request.GET.get("type") or "").strip()
    q = (request.GET.get("q") or "").strip()
    book_name = (request.GET.get("book_name") or "").strip()
    publisher = (request.GET.get("publisher") or "").strip()

    degree_id = (request.GET.get("degree") or "").strip()
    department_id = (request.GET.get("department") or "").strip()

    issue_date = parse_date((request.GET.get("issue_date") or "").strip())
    return_date = parse_date((request.GET.get("return_date") or "").strip())

    if ay:
        qs = qs.filter(academic_year=ay)

    if bt:
        qs = qs.filter(book_type_id=bt)

    if q:
        qs = qs.filter(
            Q(title__icontains=q) |
            Q(authors__icontains=q) |
            Q(accession_book__icontains=q)
        )

    if book_name:
        qs = qs.filter(title__icontains=book_name)

    if publisher:
        qs = qs.filter(publisher__iexact=publisher)

    # ✅ degree comes through department
    if degree_id:
        qs = qs.filter(department__degree_id=degree_id)

    # ✅ direct department filter
    if department_id:
        qs = qs.filter(department_id=department_id)

    if issue_date:
        qs = qs.filter(issue_date=issue_date)

    if return_date:
        qs = qs.filter(return_date=return_date)

    total_books = qs.count()
    total_book_types = BookType.objects.count()

    total_publishers = (
        qs.exclude(publisher__isnull=True)
          .exclude(publisher__exact="")
          .values("publisher")
          .distinct()
          .count()
    )

    total_book_only = qs.filter(book_type__book_type__iexact="BOOK").count()
    total_xerox_only = qs.filter(book_type__book_type__iexact="XEROX").count()

    publisher_counts = (
        qs.exclude(publisher__isnull=True)
          .exclude(publisher__exact="")
          .values("publisher")
          .annotate(count=Count("id"))
          .order_by("-count", "publisher")
    )

    all_books = qs.order_by("id")

    academic_years = (
        LibraryBook.objects.exclude(academic_year__isnull=True)
        .exclude(academic_year__exact="")
        .values_list("academic_year", flat=True)
        .distinct()
        .order_by("academic_year")
    )

    book_types = BookType.objects.all().order_by("book_type")
    degrees = Degree.objects.all().order_by("degree")
    departments = Add_Department.objects.all().order_by("Department")

    if degree_id:
        departments = departments.filter(degree_id=degree_id)

    # --------------------------------------------------
    # Extra queryset for student issue / return details
    # --------------------------------------------------
    book_transactions = (
        LibraryBookRequest.objects
        .select_related(
            "book",
            "department",
            "book__book_type",
            "book__department",
            "book__department__degree",
        )
        .filter(book__isnull=False)
        .order_by("-id")
    )

    if ay:
        book_transactions = book_transactions.filter(book__academic_year=ay)

    if bt:
        book_transactions = book_transactions.filter(book__book_type_id=bt)

    if q:
        book_transactions = book_transactions.filter(
            Q(book__title__icontains=q) |
            Q(book__authors__icontains=q) |
            Q(book__accession_book__icontains=q) |
            Q(student_name__icontains=q) |
            Q(student_rollno__icontains=q)
        )

    if book_name:
        book_transactions = book_transactions.filter(book__title__icontains=book_name)

    if publisher:
        book_transactions = book_transactions.filter(book__publisher__iexact=publisher)

    if degree_id:
        book_transactions = book_transactions.filter(book__department__degree_id=degree_id)

    if department_id:
        book_transactions = book_transactions.filter(book__department_id=department_id)

    if issue_date:
        book_transactions = book_transactions.filter(issued_on__date=issue_date)

    if return_date:
        book_transactions = book_transactions.filter(returned_on__date=return_date)

    context = {
        "total_books": total_books,
        "total_publishers": total_publishers,
        "total_book_types": total_book_types,
        "total_book_only": total_book_only,
        "total_xerox_only": total_xerox_only,
        "publisher_counts": publisher_counts,
        "all_books": all_books,
        "book_transactions": book_transactions,
        "academic_years": list(academic_years),
        "book_types": book_types,
        "selected_publisher": publisher,
        "degrees": degrees,
        "departments": departments,
        "selected_degree": degree_id,
        "selected_department": department_id,
    }

    return render(request, "library_management/admin/library_analytics_dashboard.html", context)




import os
from django.http import HttpResponse
from django.conf import settings
from django.contrib.staticfiles import finders
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.models import Q

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from library_management.models import LibraryBook


def library_analytics_dashboard_pdf(request):
    # ✅ SAME LOGIC AS YOUR DASHBOARD (NO CHANGE)
    qs = LibraryBook.objects.select_related("book_type").all()

    ay = (request.GET.get("ay") or "").strip()
    bt = (request.GET.get("type") or "").strip()
    q = (request.GET.get("q") or "").strip()
    book_name = (request.GET.get("book_name") or "").strip()
    publisher = (request.GET.get("publisher") or "").strip()

    issue_date = parse_date((request.GET.get("issue_date") or "").strip())
    return_date = parse_date((request.GET.get("return_date") or "").strip())

    if ay:
        qs = qs.filter(academic_year=ay)

    if bt:
        qs = qs.filter(book_type_id=bt)

    if q:
        qs = qs.filter(
            Q(title__icontains=q) |
            Q(authors__icontains=q) |
            Q(accession_book__icontains=q)
        )

    if book_name:
        qs = qs.filter(title__icontains=book_name)

    if publisher:
        qs = qs.filter(publisher__iexact=publisher)

    if issue_date:
        qs = qs.filter(issue_date=issue_date)

    if return_date:
        qs = qs.filter(return_date=return_date)

    qs = qs.order_by("id")

    # ✅ Open in browser (new tab)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="library_analytics_report.pdf"'

    pagesize = A4  # ✅ smaller than landscape A4

    HEADER_HEIGHT = 44 * mm

    doc = SimpleDocTemplate(
        response,
        pagesize=pagesize,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=16 * mm + HEADER_HEIGHT,
        bottomMargin=16 * mm,
        title="Library Analytics Report",
    )

    styles = getSampleStyleSheet()

    # ✅ styles
    header_style = ParagraphStyle(
        "hdr", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=8.2,
        textColor=colors.white, alignment=TA_CENTER, leading=9.6,
        wordWrap="CJK"
    )
    cell_left = ParagraphStyle(
        "cell_left", parent=styles["Normal"],
        fontName="Helvetica", fontSize=7.8,
        textColor=colors.black, alignment=TA_LEFT, leading=9.5,
        wordWrap="CJK"
    )
    cell_center = ParagraphStyle(
        "cell_center", parent=cell_left, alignment=TA_CENTER
    )

    def s(v):
        return "" if v is None else str(v).strip()

    def wrap_slash(v):
        # ✅ makes Dept code visible in portrait
        return s(v).replace("/", "/<br/>")

    # -----------------------------
    # ✅ BODY CONTENT
    # -----------------------------
    story = []

    story.append(Paragraph(f"<b>Total Books:</b> {qs.count()}", styles["Normal"]))

    # ✅ ADD THIS: show selected publisher in PDF
    if publisher:
        story.append(Spacer(1, 2))
        story.append(Paragraph(f"<b>Publisher:</b> {publisher}", styles["Normal"]))

    story.append(Spacer(1, 6))

    # ✅ Portrait PDF: Keep important columns only (clear + readable)
    data = [[
        Paragraph("Dept", header_style),
        Paragraph("Accession", header_style),
        Paragraph("Title", header_style),
        Paragraph("Type", header_style),
        Paragraph("Academic Year", header_style),
        Paragraph("Issue Date", header_style),
        Paragraph("Return Date", header_style),
    ]]

    for b in qs:
        data.append([
            Paragraph(wrap_slash(getattr(b, "dept_code", "")), cell_left),
            Paragraph(s(getattr(b, "accession_book", "")), cell_center),
            Paragraph(s(getattr(b, "title", "")), cell_left),
            Paragraph(s(getattr(getattr(b, "book_type", None), "book_type", "")), cell_center),
            Paragraph(s(getattr(b, "academic_year", "")), cell_center),
            Paragraph(b.issue_date.strftime("%d-%m-%Y") if getattr(b, "issue_date", None) else "", cell_center),
            Paragraph(b.return_date.strftime("%d-%m-%Y") if getattr(b, "return_date", None) else "", cell_center),
        ])

    W = doc.width
    col_widths = [
        W * 0.18,  # Dept
        W * 0.10,  # Accession
        W * 0.32,  # Title
        W * 0.10,  # Type
        W * 0.14,  # Academic Year
        W * 0.08,  # Issue
        W * 0.08,  # Return
    ]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f2f57")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    story.append(table)

    # -----------------------------
    # Header/Footer (same style)
    # -----------------------------
    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    ACCENT_RED = colors.HexColor("#b91c1c")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    # ✅ keep publisher available inside header/footer
    selected_publisher = publisher

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = pagesize

        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 8 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            try:
                canvas.drawImage(
                    ImageReader(logo_path),
                    left, top_y - 20 * mm,
                    width=30 * mm, height=18 * mm,
                    preserveAspectRatio=True, mask="auto"
                )
            except Exception:
                pass

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawCentredString(center_x, top_y - 6 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "An Autonomous Institution")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 8.2)
        canvas.drawCentredString(center_x, top_y - 18.5 * mm, "Approved by AICTE, New Delhi")
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(center_x, top_y - 23 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
        canvas.drawCentredString(center_x, top_y - 27.5 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(center_x, top_y - 35 * mm, "LIBRARY ANALYTICS REPORT")

        # ✅ ADD THIS: show publisher in header when filtered
        if selected_publisher:
            canvas.setFillColor(MEDIUM_GRAY)
            canvas.setFont("Helvetica-Bold", 9)
            canvas.drawCentredString(center_x, top_y - 40 * mm, f"Publisher: {selected_publisher}")

        header_line_y = page_h - HEADER_HEIGHT + 6 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.8)
        canvas.line(left, header_line_y, right, header_line_y)

        footer_y = 16 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.8)
        canvas.line(left, footer_y + 7 * mm, right, footer_y + 7 * mm)

        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 8)
        gen_time = timezone.localtime(timezone.now()).strftime("%d %b %Y, %I:%M %p")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawRightString(right, footer_y, f"Page {canvas.getPageNumber()}")

        canvas.restoreState()

    doc.build(story, onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)
    return response















from datetime import date
import json
import logging

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.db.models.fields.related import ForeignObjectRel
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

# faculty_management in DEFAULT DB
from faculty_management.models import Add_Department, general_information

# library_management in DEFAULT DB
from library_management.models import (
    LibraryBook,
    LibraryBookRequest,
    LibraryNotification,
    LibraryRequestApprovers,
    LibraryRequestApproversData,
)

# SectionMaster (keep your path)
from course_management.models import SectionMaster

# Role in rit_approval_system DB
from user_accounts.models import Role

# decorators (keep your path)
from user_accounts.decorators import no_cache, is_super_user

logger = logging.getLogger(__name__)
APPROVAL_DB = "rit_approval_system"


# ============================================================
# HELPERS (NO WORKFLOW LOGIC CHANGE)
# ============================================================

def _safe_employee_id(request):
    emp = getattr(getattr(request, "user", None), "Employee_id", None)
    if emp not in (None, "", "null", "None", "undefined"):
        return emp

    for k in ("faculty_id", "emp_id", "employee_id", "staff_id", "user_id", "login_id"):
        v = request.session.get(k)
        if v not in (None, "", "null", "None", "undefined"):
            return v

    return None


def _book_is_currently_issued(book: LibraryBook) -> bool:
    return bool(getattr(book, "issue_date", None) and not getattr(book, "return_date", None))


def _notify(to_faculty_id: int, title: str, message: str):
    if not to_faculty_id:
        return
    try:
        LibraryNotification.objects.create(
            to_faculty_id=to_faculty_id,
            title=title,
            message=message
        )
    except Exception:
        logger.exception("Notification failed")


def _select_related_safe(qs, *fields):
    safe = []
    model = qs.model
    for f in fields:
        try:
            field_obj = model._meta.get_field(f)
            if getattr(field_obj, "is_relation", False) and not isinstance(field_obj, ForeignObjectRel):
                safe.append(f)
        except Exception:
            continue
    try:
        return qs.select_related(*safe)
    except Exception:
        return qs


def _is_int_string(s: str) -> bool:
    try:
        int(str(s).strip())
        return True
    except Exception:
        return False


def _safe_login_identity(request):
    for key in (
        "faculty_id", "emp_id", "employee_id", "staff_id",
        "user_id", "login_id",
        "user_name", "name",
        "college_email", "personal_email", "email",
        "phone", "mobile", "mobile_no",
    ):
        val = request.session.get(key)
        if val not in (None, "", "null", "None", "undefined"):
            return str(val).strip()

    u = getattr(request, "user", None)
    if u and getattr(u, "is_authenticated", False):
        uname = (getattr(u, "username", "") or "").strip()
        if uname:
            return uname

        emp = getattr(u, "Employee_id", None)
        if emp not in (None, "", "null", "None", "undefined"):
            return str(emp).strip()

        mail = (getattr(u, "email", "") or "").strip()
        if mail:
            return mail

    return None


def _resolve_logged_in_general_info(request):
    """
    ✅ Resolves logged-in user into faculty_management.general_information row safely.
    (faculty_id / name / email / phone)
    """
    ident = _safe_login_identity(request)
    if not ident:
        return None
    ident = str(ident).strip()

    # 1) Number -> faculty_id
    if _is_int_string(ident):
        fid = int(ident)
        gi = general_information.objects.filter(faculty_id=fid).first()
        if gi:
            return gi

    # 2) Email
    if "@" in ident:
        gi = general_information.objects.filter(college_email__iexact=ident).first()
        if gi:
            return gi
        gi = general_information.objects.filter(personal_email__iexact=ident).first()
        if gi:
            return gi

    # 3) Phone digits
    digits_only = "".join(ch for ch in ident if ch.isdigit())
    if digits_only and len(digits_only) >= 8:
        try:
            gi = general_information.objects.filter(phone=int(digits_only)).first()
            if gi:
                return gi
        except Exception:
            pass

    # 4) Name exact
    gi = general_information.objects.filter(name__iexact=ident).first()
    if gi:
        return gi

    # 5) Name contains
    gi = general_information.objects.filter(name__icontains=ident).order_by("id").first()
    if gi:
        return gi

    return None


# ============================================================
# APPROVER FINDING (KEEP YOUR LOGIC)
# ============================================================

def _find_first_user_for_role_and_department(approver_role: Role, department_id: int):
    role_name = (getattr(approver_role, "role", "") or "").strip().lower()

    if "hod" in role_name:
        keywords = ["hod"]
    elif "library" in role_name or "incharge" in role_name:
        keywords = [
            "department library incharge",
            "library incharge",
            "library",
            "incharge",
            "librarian",
        ]
    else:
        keywords = [role_name] if role_name else []

    qs = general_information.objects.filter(department_id=department_id)
    qs = _select_related_safe(qs, "designation", "department")

    for kw in keywords:
        user = qs.filter(designation__designation_name__icontains=kw).order_by("id").first()
        if user:
            return user

    return None


def _get_creator_role_and_chain_by_role_name(creator_role_name: str):
    creator_role = Role.objects.using(APPROVAL_DB).filter(role__iexact=creator_role_name).first()
    if not creator_role:
        return None, []

    chain = list(
        LibraryRequestApprovers.objects.filter(creator_role_id=creator_role.id).order_by("approver_level")
    )
    return creator_role, chain


# ============================================================
# ✅ LOGGED-IN STUDENT RESOLUTION
# ============================================================

def _get_logged_in_student(request):
    user = getattr(request, "user", None)
    if not user or not getattr(user, "is_authenticated", False):
        return None

    reg_no = getattr(user, "reg_no", None) or getattr(user, "RegNo", None)
    email = getattr(user, "email", None)

    if not reg_no:
        reg_no = getattr(user, "username", None)

    if not reg_no:
        reg_no = request.session.get("reg_no") or request.session.get("student_reg_no")

    if not email:
        email = request.session.get("email") or request.session.get("student_email")

    from student_management.models import StudentDetails

    qs = StudentDetails.objects.all().select_related("department")

    if reg_no:
        st = qs.filter(reg_no__iexact=str(reg_no).strip()).first()
        if st:
            return st

    if email:
        st = qs.filter(email__iexact=str(email).strip()).first()
        if st:
            return st

    return None


# ============================================================
# ✅ STUDENT REQUEST PAGE (KEEP YOUR LOGIC)
# ============================================================

def student_request_book(request):
    student = _get_logged_in_student(request)
    if not student:
        messages.error(request, "Student profile not found for your login. Please update your student details.")
        return redirect("student_request_book")

    if not student.department_id:
        messages.error(request, "Your department is not set in StudentDetails. Please update your profile.")
        return redirect("student_request_book")

    dept = get_object_or_404(Add_Department, id=student.department_id)

    books_qs = LibraryBook.objects.all()
    books = _select_related_safe(books_qs, "department").filter(department_id=dept.id).order_by("title")

    reqs_qs = LibraryBookRequest.objects.filter(student_rollno__iexact=(student.reg_no or ""))
    reqs_qs = _select_related_safe(reqs_qs, "book", "department").order_by("-id")

    paginator = Paginator(reqs_qs, 30)
    page = request.GET.get("page")
    my_requests_page = paginator.get_page(page)

    def _is_office_department(dept_id):
        if not dept_id:
            return False
        try:
            return Add_Department.objects.filter(id=dept_id, Department__iexact="OFFICE").exists()
        except Exception:
            return False

    def _find_first_user_for_role_any_department(approver_role: Role):
        role_name = (getattr(approver_role, "role", "") or "").strip().lower()

        if "hod" in role_name:
            keywords = ["hod"]
        elif "library" in role_name or "incharge" in role_name:
            keywords = [
                "department library incharge",
                "library incharge",
                "library",
                "incharge",
                "librarian",
            ]
        else:
            keywords = [role_name] if role_name else []

        qs = general_information.objects.all()
        qs = _select_related_safe(qs, "designation", "department")

        for kw in keywords:
            user = qs.filter(designation__designation_name__icontains=kw).order_by("id").first()
            if user:
                return user
        return None

    if request.method == "POST":
        student_name = (student.name or "").strip()
        student_rollno = (student.reg_no or "").strip()
        book_id = (request.POST.get("book_id") or "").strip()

        if not (student_name and student_rollno and dept.id and book_id):
            messages.error(request, "Fill all required fields.")
            return redirect("student_request_book")

        book = get_object_or_404(LibraryBook, id=book_id)

        if str(book.department_id) != str(dept.id):
            messages.error(request, "You can request only books from your department library.")
            return redirect("student_request_book")

        if _book_is_currently_issued(book):
            messages.error(request, "This book is already issued and not yet returned.")
            return redirect("student_request_book")

        if LibraryBookRequest.objects.filter(
            student_name__iexact=student_name,
            student_rollno=(student_rollno or None),
            department=dept,
            book=book,
            status="PENDING",
        ).exists():
            messages.warning(request, "This request is already pending.")
            return redirect("student_request_book")

        creator_role, chain = _get_creator_role_and_chain_by_role_name("student")
        if not creator_role:
            messages.error(request, "Student role not found in Role table (rit_approval_system).")
            return redirect("student_request_book")

        if not chain:
            messages.error(request, "No approval hierarchy configured for Student role.")
            return redirect("student_request_book")

        resolved_levels = []
        hod_user = None

        for row in chain:
            approver_role = Role.objects.using(APPROVAL_DB).filter(id=row.approver_role_id).first()
            if not approver_role:
                messages.error(request, f"Approver role not found (id={row.approver_role_id}).")
                return redirect("student_request_book")

            if (row.is_cross_department_approver == "YES") and row.approver_department_id and _is_office_department(row.approver_department_id):
                approver_user = _find_first_user_for_role_any_department(approver_role)
                if not approver_user:
                    role_name = getattr(approver_role, "role", None) or f"Role#{row.approver_role_id}"
                    messages.error(request, f"No approver user found for {role_name} (OFFICE cross-department).")
                    return redirect("student_request_book")
            else:
                target_dept_id = dept.id
                if (row.is_cross_department_approver == "YES") and row.approver_department_id:
                    target_dept_id = row.approver_department_id

                approver_user = _find_first_user_for_role_and_department(approver_role, target_dept_id)
                if not approver_user:
                    role_name = getattr(approver_role, "role", None) or f"Role#{row.approver_role_id}"
                    messages.error(request, f"No approver user found for {role_name} in department id {target_dept_id}.")
                    return redirect("student_request_book")

            resolved_levels.append((row.approver_level, approver_user))

            if "hod" in (getattr(approver_role, "role", "") or "").lower() and not hod_user:
                hod_user = approver_user

        resolved_levels.sort(key=lambda x: x[0])
        first_approver = resolved_levels[0][1]

        req = LibraryBookRequest.objects.create(
            student_name=student_name,
            student_rollno=student_rollno or None,
            department=dept,
            book=book,
            status="PENDING",
            requested_on=date.today(),
            incharge_faculty_id=getattr(first_approver, "faculty_id", None),
            hod_faculty_id=getattr(hod_user, "faculty_id", None) if hod_user else None,
        )

        for lvl, approver_user in resolved_levels:
            LibraryRequestApproversData.objects.create(
                request=req,
                approver_id=approver_user,
                creator_id=None,
                approver_level=lvl,
                status=LibraryRequestApproversData.Status.PENDING,
                reason=None,
                acted_on=timezone.now(),
            )

        _notify(
            to_faculty_id=getattr(first_approver, "faculty_id", None),
            title="Library Book Issue Request",
            message=f"{student_name} ({student_rollno or '-'}) requested '{book.title}'. Request ID: {req.id}"
        )

        messages.success(request, "✅ Request sent for approval (Level 1).")
        return redirect("student_request_book")

    return render(request, "library_management/admin/student_request_book.html", {
        "student": student,
        "student_dept": dept,
        "books": books,
        "requests": my_requests_page,
    })


# ============================================================
# ✅ STUDENT RETURN REQUEST
# ============================================================

def student_return_request(request, pk):
    req = get_object_or_404(LibraryBookRequest, pk=pk)

    student = _get_logged_in_student(request)
    if not student:
        messages.error(request, "Student profile not found for your login.")
        return redirect("student_request_book")

    if (req.student_rollno or "").strip().lower() != (student.reg_no or "").strip().lower():
        messages.error(request, "You are not allowed to return this request.")
        return redirect("student_request_book")

    if req.status != "APPROVED":
        messages.warning(request, "Return allowed only after approval (Issued).")
        return redirect("student_request_book")

    if req.status == "RETURN_REQUESTED":
        messages.warning(request, "Return request already sent.")
        return redirect("student_request_book")

    if req.status == "RETURNED":
        messages.warning(request, "This book is already returned.")
        return redirect("student_request_book")

    req.status = "RETURN_REQUESTED"
    req.return_requested_on = date.today()
    req.save(update_fields=["status", "return_requested_on"])

    _notify(
        to_faculty_id=req.incharge_faculty_id,
        title="Library Return Request",
        message=f"{req.student_name} ({req.student_rollno or '-'}) requested RETURN for '{req.book.title}' "
                f"(Acc: {getattr(req.book, 'accession_book', None) or '-'}). Request ID: {req.id}"
    )

    messages.success(request, "✅ Return request sent to Library Incharge.")
    return redirect("student_request_book")


# ============================================================
# ✅ INCHARGE: REQUESTS DASHBOARD
# (Keep your visibility logic, but UI will restrict buttons)
# Template: library_management/admin/library_incharge_requests.html
# ============================================================

@transaction.atomic
def library_incharge_requests(request):
    incharge = _resolve_logged_in_general_info(request)
# =========================================================
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse
from user_accounts.decorators import no_cache
from django.views.decorators.csrf import csrf_exempt
from datetime import date
import json
import logging

logger = logging.getLogger(__name__)


def library_incharge_requests(request):
    
    emp_id = _safe_employee_id(request)

    incharge = general_information.objects.select_related(
        "department", "designation"
    ).filter(
        faculty_id=emp_id
    ).first() if emp_id else None

    if not incharge:
        messages.error(
            request,
            "Library Incharge profile not found for this login. "
            "Your login must match a record in app_general_information (faculty_id/name/email/phone)."
        )
        return render(request, "library_management/admin/library_incharge_requests.html", {
            "incharge": None,
            "requests": [],
            "selected_status": "",
            "search_q": "",
            "notifications": [],
            "year_val": "",
            "section_val": "",
            "sections": [],
            "year_choices": [],
        })

    dept_id = getattr(incharge, "department_id", None)

    base_q = Q()

    # ✅ Keep your original logic:
    # show requests where this incharge is an approver OR dept requests OR assigned incharge_faculty_id
    try:
        base_q |= Q(approval_entries__approver_id=incharge)
    except Exception:
        pass

    if dept_id:
        base_q |= Q(department_id=dept_id)

    if getattr(incharge, "faculty_id", None):
        base_q |= Q(incharge_faculty_id=incharge.faculty_id)

    qs = LibraryBookRequest.objects.filter(base_q).distinct()
    qs = _select_related_safe(qs, "book", "department").order_by("-id")
    qs = LibraryBookRequest.objects.select_related(
        "book", "department"
    ).filter(
        incharge_faculty_id=incharge.faculty_id
    ).order_by("-id")

    status = (request.GET.get("status") or "").strip()
    q = (request.GET.get("q") or "").strip()
    year_val = (request.GET.get("year") or "").strip()
    section_val = (request.GET.get("section") or "").strip()

    if status:
        qs = qs.filter(status=status)

    if q:
        qs = qs.filter(
            Q(student_name__icontains=q) |
            Q(student_rollno__icontains=q) |
            Q(book__title__icontains=q) |
            Q(book__accession_book__icontains=q) |
            Q(book__publisher__icontains=q)
        )

    if year_val:
        try:
            qs = qs.filter(year=year_val)
        except Exception:
            try:
                qs = qs.filter(student_year=year_val)
            except Exception:
                pass

    if section_val:
        try:
            qs = qs.filter(section=section_val)
        except Exception:
            try:
                qs = qs.filter(student_section=section_val)
            except Exception:
                pass

    paginator = Paginator(qs, 30)
    page = request.GET.get("page")
    requests_page = paginator.get_page(page)

    notifications = LibraryNotification.objects.filter(
        to_faculty_id=getattr(incharge, "faculty_id", None)
    ).order_by("-id")[:10]

    sections = SectionMaster.objects.all().order_by("section")
    year_choices = ["1", "2", "3", "4"]

    return render(request, "library_management/admin/library_incharge_requests.html", {
        "incharge": incharge,
        "requests": requests_page,
        "selected_status": status,
        "search_q": q,
        "notifications": notifications,
        "year_val": year_val,
        "section_val": section_val,
        "sections": sections,
        "year_choices": year_choices,
    })


# ============================================================
# ✅ INCHARGE: APPROVE (POST ONLY) — STRICT ASSIGN CHECK
# ============================================================
@require_POST
# =========================================================
# INCHARGE APPROVE BOOK REQUEST
# =========================================================
@transaction.atomic
def incharge_approve_book_request(request, pk):

    incharge = _resolve_logged_in_general_info(request)
    emp_id = _safe_employee_id(request)

    incharge = general_information.objects.filter(
        faculty_id=emp_id
    ).first() if emp_id else None

    if not incharge:
        messages.error(request, "Library Incharge profile not found.")
        return redirect("library_incharge_requests")

    # ✅ allow only if assigned to this incharge
    if getattr(req, "incharge_faculty_id", None) != getattr(incharge, "faculty_id", None):
        messages.error(request, "Unauthorized action (no approval assigned to you).")
        return redirect("library_incharge_requests")
    req = get_object_or_404(
        LibraryBookRequest.objects.select_related("book"),
        pk=pk,
        incharge_faculty_id=incharge.faculty_id
    )

    # ✅ allow correction from REJECTED -> APPROVED also
    if req.status not in ["PENDING", "REJECTED"]:
        messages.warning(request, f"This request cannot be approved from '{req.status}' status.")
        return redirect("library_incharge_requests")

    if _book_is_currently_issued(req.book):
        messages.error(request, "This book is already issued and not yet returned.")
        return redirect("library_incharge_requests")

    today = date.today()
    old_status = req.status

    req.status = "APPROVED"
    req.approved_on = today
    req.issued_on = today
    req.remarks = None
    req.save(update_fields=["status", "approved_on", "issued_on", "remarks"])

    req.book.issue_date = today
    req.book.return_date = None
    req.book.save(update_fields=["issue_date", "return_date"])

    try:
        # first try pending entry
        entry = LibraryRequestApproversData.objects.filter(
            request=req,
            approver_id=incharge,
            status=LibraryRequestApproversData.Status.PENDING
        ).order_by("-id").first()

        # if request was rejected before, allow correction of that entry too
        if not entry:
            entry = LibraryRequestApproversData.objects.filter(
                request=req,
                approver_id=incharge,
                status=LibraryRequestApproversData.Status.REJECTED
            ).order_by("-id").first()

        if entry:
            entry.status = LibraryRequestApproversData.Status.APPROVED
            entry.reason = None
            entry.acted_on = timezone.now()
            entry.save(update_fields=["status", "reason", "acted_on"])
    except Exception:
        pass

    if getattr(req, "hod_faculty_id", None):
        _notify(
            to_faculty_id=req.hod_faculty_id,
            title="Library Book Approved (Incharge)",
            message=f"✅ Approved by Incharge: {req.student_name} ({req.student_rollno or '-'}) "
                    f"Book: '{req.book.title}' (Acc: {getattr(req.book, 'accession_book', None) or '-'}). "
                    f"Request ID: {req.id}. Issue Date: {today.strftime('%d-%m-%Y')}"
        )

    if old_status == "REJECTED":
        messages.success(request, "✅ Rejected request changed to Approved successfully.")
    else:
        messages.success(request, "✅ Approved. Issue date stored. Student can take the book.")

    # Notify HOD
    if req.hod_faculty_id:
        _notify(
            to_faculty_id=req.hod_faculty_id,
            title="Library Book Approved (Incharge)",
            message=f"Approved by Incharge: {req.student_name} ({req.student_rollno or '-'}) "
                    f"Book: '{req.book.title}' "
                    f"(Acc: {req.book.accession_book or '-'}) "
                    f"Request ID: {req.id} "
                    f"Issue Date: {today.strftime('%d-%m-%Y')}"
        )

    messages.success(request, "Approved successfully. Student can take the book.")
    return redirect("library_incharge_requests")
 

# ============================================================
# ✅ INCHARGE: REJECT (POST ONLY) — STRICT ASSIGN CHECK
# ============================================================

@require_POST
# =========================================================
# INCHARGE REJECT BOOK REQUEST
# =========================================================
@transaction.atomic
def incharge_reject_book_request(request, pk):

    incharge = _resolve_logged_in_general_info(request)
    emp_id = _safe_employee_id(request)

    incharge = general_information.objects.filter(
        faculty_id=emp_id
    ).first() if emp_id else None

    if not incharge:
        messages.error(request, "Library Incharge profile not found.")
        return redirect("library_incharge_requests")

    # ✅ allow only if assigned to this incharge
    if getattr(req, "incharge_faculty_id", None) != getattr(incharge, "faculty_id", None):
        messages.error(request, "Unauthorized action (no approval assigned to you).")
        return redirect("library_incharge_requests")
    req = get_object_or_404(
        LibraryBookRequest,
        pk=pk,
        incharge_faculty_id=incharge.faculty_id
    )

    # ✅ allow correction from APPROVED -> REJECTED also
    if req.status not in ["PENDING", "APPROVED"]:
        messages.warning(request, f"This request cannot be rejected from '{req.status}' status.")
        return redirect("library_incharge_requests")

    remarks = (request.POST.get("remarks") or "").strip()
    if not remarks:
        messages.error(request, "Remarks is required to reject.")
        return redirect("library_incharge_requests")

    old_status = req.status

    req.status = "REJECTED"
    req.remarks = remarks or None

    # optional cleanup if previously approved
    if old_status == "APPROVED":
        req.approved_on = None
        req.issued_on = None

    req.save(update_fields=["status", "remarks", "approved_on", "issued_on"])

    # if previously approved, clear book issue details
    if old_status == "APPROVED":
        req.book.issue_date = None
        req.book.return_date = None
        req.book.save(update_fields=["issue_date", "return_date"])

    try:
        # first try pending entry
        entry = LibraryRequestApproversData.objects.filter(
            request=req,
            approver_id=incharge,
            status=LibraryRequestApproversData.Status.PENDING
        ).order_by("-id").first()

        # if request was approved before, allow correction of that entry too
        if not entry:
            entry = LibraryRequestApproversData.objects.filter(
                request=req,
                approver_id=incharge,
                status=LibraryRequestApproversData.Status.APPROVED
            ).order_by("-id").first()

        if entry:
            entry.status = LibraryRequestApproversData.Status.REJECTED
            entry.reason = req.remarks
            entry.acted_on = timezone.now()
            entry.save(update_fields=["status", "reason", "acted_on"])
    except Exception:
        pass

    req.status = "REJECTED"
    req.remarks = remarks or None
    req.save(update_fields=["status", "remarks"])

    if req.hod_faculty_id:
        _notify(
            to_faculty_id=req.hod_faculty_id,
            title="Library Book Rejected (Incharge)",
            message=f"❌ Rejected by Incharge: {req.student_name} ({req.student_rollno or '-'}) "
                    f"Book: '{req.book.title}' (Acc: {getattr(req.book, 'accession_book', None) or '-'}). "
                    f"Request ID: {req.id}. Remarks: {req.remarks or '-'}"
        )

    if old_status == "APPROVED":
        messages.success(request, "✅ Approved request changed to Rejected successfully.")
    else:
        messages.success(request, "✅ Rejected.")

        _notify(
            to_faculty_id=req.hod_faculty_id,
            title="Library Book Rejected (Incharge)",
            message=f"❌ Rejected by Incharge: {req.student_name} ({req.student_rollno or '-'}) "
                    f"Book: '{req.book.title}' (Acc: {getattr(req.book, 'accession_book', None) or '-'}). "
                    f"Request ID: {req.id}. Remarks: {req.remarks or '-'}"
        )

    messages.success(request, "Request rejected.")
    return redirect("library_incharge_requests")

# ============================================================
# ✅ INCHARGE: CONFIRM RETURN (POST ONLY) — STRICT ASSIGN CHECK
# ============================================================

@require_POST
# =========================================================
# INCHARGE CONFIRM BOOK RETURN
# =========================================================
@transaction.atomic
def incharge_confirm_return(request, pk):
    incharge = _resolve_logged_in_general_info(request)
    emp_id = _safe_employee_id(request)

    incharge = general_information.objects.filter(
        faculty_id=emp_id
    ).first() if emp_id else None

    if not incharge:
        messages.error(request, "Library Incharge profile not found.")
        return redirect("library_incharge_requests")

    req = get_object_or_404(
        LibraryBookRequest.objects.select_related("book"),
        pk=pk,
        incharge_faculty_id=incharge.faculty_id
    )

    if getattr(req, "incharge_faculty_id", None) != getattr(incharge, "faculty_id", None):
        messages.error(request, "Unauthorized action.")
        return redirect("library_incharge_requests")

    if req.status != "RETURN_REQUESTED":
        messages.warning(request, "Return not requested yet.")
        return redirect("library_incharge_requests")

    today = date.today()

    req.status = "RETURNED"
    req.returned_on = today
    req.save(update_fields=["status", "returned_on"])

    req.book.return_date = today
    req.book.save(update_fields=["return_date"])

    if req.hod_faculty_id:
        _notify(
            to_faculty_id=req.hod_faculty_id,
            title="Library Book Returned (Incharge Confirmed)",
            message=(
                f"✅ Return confirmed: {req.student_name} ({req.student_rollno or '-'}) "
                f"Book: '{req.book.title}' "
                f"(Acc: {getattr(req.book, 'accession_book', None) or '-'}) "
                f"Request ID: {req.id}. "
                f"Return Date: {today.strftime('%d-%m-%Y')}"
            )
        )

    messages.success(request, "Return confirmed successfully.")
    return redirect("library_incharge_requests")
# ============================================================
# ✅ HOD: VIEW ONLY
# ============================================================
 
# =========================================================
# HOD: VIEW ONLY
# Template: library_management/admin/hod_library_requests.html
# =========================================================
def hod_library_requests(request):
    emp_id = _safe_employee_id(request)
    hod = general_information.objects.select_related("department", "designation").filter(
        faculty_id=emp_id
    ).first() if emp_id else None

@transaction.atomic
def hod_library_requests(request):
    hod = _resolve_logged_in_general_info(request)
    if not hod:
        messages.error(
            request,
            "HOD profile not found for this login. "
            "Your login must match a record in app_general_information (faculty_id/name/email/phone)."
        )
        return render(request, "library_management/admin/hod_library_requests.html", {
            "hod": None,
            "requests": [],
            "selected_status": "",
            "search_q": "",
            "notifications": [],
            "year_val": "",
            "section_val": "",
            "sections": [],
            "year_choices": [],
        })

    qs = LibraryBookRequest.objects.filter(hod_faculty_id=hod.faculty_id)
    qs = _select_related_safe(qs, "book", "department").order_by("-id")

    status = (request.GET.get("status") or "").strip()
    q = (request.GET.get("q") or "").strip()
    year_val = (request.GET.get("year") or "").strip()
    section_val = (request.GET.get("section") or "").strip()

    if status:
        qs = qs.filter(status=status)

    if q:
        qs = qs.filter(
            Q(student_name__icontains=q) |
            Q(student_rollno__icontains=q) |
            Q(book__title__icontains=q) |
            Q(book__accession_book__icontains=q) |
            Q(book__publisher__icontains=q)
        )

    if year_val:
        try:
            qs = qs.filter(year=year_val)
        except Exception:
            try:
                qs = qs.filter(student_year=year_val)
            except Exception:
                pass

    if section_val:
        try:
            qs = qs.filter(section=section_val)
        except Exception:
            try:
                qs = qs.filter(student_section=section_val)
            except Exception:
                pass

    paginator = Paginator(qs, 30)
    page = request.GET.get("page")
    requests_page = paginator.get_page(page)

    notifications = LibraryNotification.objects.filter(
        to_faculty_id=hod.faculty_id
    ).order_by("-id")[:10]

    sections = SectionMaster.objects.all().order_by("section")
    year_choices = ["1", "2", "3", "4"]

    return render(request, "library_management/admin/hod_library_requests.html", {
        "hod": hod,
        "requests": requests_page,
        "selected_status": status,
        "search_q": q,
        "notifications": notifications,
        "year_val": year_val,
        "section_val": section_val,
        "sections": sections,
        "year_choices": year_choices,
    })


# ============================================================
# HOD: BACKWARD COMPATIBILITY (kept same)
# ============================================================

def hod_approve_book_request(request, pk):
    messages.error(request, "Approval is handled by Department Library Incharge.")
    return redirect("hod_library_requests")


def hod_reject_book_request(request, pk):
    messages.error(request, "Rejection is handled by Department Library Incharge.")
    return redirect("hod_library_requests")


def hod_confirm_return(request, pk):
    messages.error(request, "Return confirmation is handled by Department Library Incharge.")
    return redirect("hod_library_requests")


# ============================================================
# APPROVAL MANAGEMENT PAGE (kept same)
# ============================================================

def _safe_role_id(request):
    for k in ("role_id", "user_role_id", "creator_role_id", "approver_role_id", "role"):
        v = request.session.get(k)
        try:
            if v is not None and str(v).strip() != "":
                return int(v)
        except Exception:
            pass
    return None


def _office_cross_department_enabled_for_current_user(request):
    role_id = _safe_role_id(request)

    if not role_id:
        role_name = (request.session.get("user_role") or request.session.get("role_name") or "").strip()
        if role_name:
            role_id = (
                Role.objects.using(APPROVAL_DB)
                .filter(role__iexact=role_name)
                .values_list("id", flat=True)
                .first()
            )

    if not role_id:
        return False

    return LibraryRequestApprovers.objects.filter(
        approver_role_id=role_id,
        is_cross_department_approver=LibraryRequestApprovers.DefaultApprover.YES,
        approver_department__Department__iexact="OFFICE",
    ).exists()


@is_super_user("library_management")
def library_request_approval_management(request):
    if request.method == "GET":
        roles = Role.objects.using(APPROVAL_DB).all()
        departments = Add_Department.objects.all()
        return render(request, "library_management/admin/library_request_management.html", {
            "roles": roles,
            "departments": departments,
        })

    if request.method == "POST":
        try:
            raw = (request.body or b"").decode("utf-8").strip()
            if not raw:
                return JsonResponse({"error": "Empty body"}, status=400)

            data = json.loads(raw)

            creator_role_id_raw = data.get("creatorRole")
            role_hierarchy = data.get("roleHierarchy", [])

            try:
                creator_role_id = int(creator_role_id_raw)
            except (TypeError, ValueError):
                return JsonResponse({"error": "Invalid creatorRole"}, status=400)

            if not isinstance(role_hierarchy, list):
                return JsonResponse({"error": "roleHierarchy must be a list"}, status=400)

            def _to_bool(v):
                if isinstance(v, bool):
                    return v
                if v is None:
                    return False
                if isinstance(v, (int, float)):
                    return bool(v)
                s = str(v).strip().lower()
                return s in ("1", "true", "yes", "y", "on")

            def _to_int_or_none(v):
                if v in (None, "", "null", "None", "undefined"):
                    return None
                try:
                    return int(v)
                except (TypeError, ValueError):
                    return None

            LibraryRequestApprovers.objects.filter(creator_role_id=creator_role_id).delete()

            for index, role_data in enumerate(role_hierarchy):
                role_id_raw = role_data.get("id")
                try:
                    approver_role_id = int(role_id_raw)
                except (TypeError, ValueError):
                    return JsonResponse({"error": f"Invalid role id: {role_id_raw}"}, status=400)

                is_cross_department = _to_bool(role_data.get("isCrossDepartment", False))
                dept_id = _to_int_or_none(role_data.get("departmentId"))

                if is_cross_department and not dept_id:
                    return JsonResponse({
                        "error": f"Department is required for cross-department approver (role_id={approver_role_id})."
                    }, status=400)

                department_obj = Add_Department.objects.filter(id=dept_id).first() if dept_id else None
                if is_cross_department and dept_id and not department_obj:
                    return JsonResponse({"error": f"Department not found: {dept_id}"}, status=404)

                LibraryRequestApprovers.objects.create(
                    creator_role_id=creator_role_id,
                    approver_role_id=approver_role_id,
                    approver_level=index + 1,
                    is_cross_department_approver="YES" if is_cross_department else "NO",
                    approver_department=department_obj
                )

            return JsonResponse({"message": "Library request roles submitted successfully"}, status=200)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)
        except Exception as e:
            logger.exception("Error in library request approval management POST")
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request method"}, status=405)
   



from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from library_management.models import LibraryBook, BookType

@csrf_exempt
def edit_library_book(request, pk):

    if request.method == "POST":

        try:
            book = LibraryBook.objects.get(pk=pk)

            # Read POST data instead of JSON
            dept_code = request.POST.get("dept_code")
            accession_book = request.POST.get("accession_book")
            title = request.POST.get("title")
            authors = request.POST.get("authors")
            book_type_name = request.POST.get("book_type")

            if dept_code is not None:
                book.dept_code = dept_code

            if accession_book is not None:
                book.accession_book = accession_book

            if title is not None:
                book.title = title

            if authors is not None:
                book.authors = authors

            if book_type_name:
                bt = BookType.objects.filter(book_type__iexact=book_type_name).first()
                if bt:
                    book.book_type = bt

            book.save()

            return JsonResponse({"success": True})

        except LibraryBook.DoesNotExist:
            return JsonResponse({"error": "Book not found"}, status=404)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Invalid request"}, status=405)







import io
import os
from datetime import datetime

from django.conf import settings
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.staticfiles import finders

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame,
    Table, TableStyle, Paragraph, Spacer
)

from library_management.models import LibraryBookRequest


def student_request_book_pdf(request, pk):
    req = get_object_or_404(
        LibraryBookRequest.objects.select_related("book", "department", "book__book_type"),
        pk=pk
    )

    buffer = io.BytesIO()
    styles = getSampleStyleSheet()

    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1d4f91")
    ACCENT_RED = colors.HexColor("#c62828")
    DARK = colors.HexColor("#1f2937")
    MID = colors.HexColor("#4b5563")
    LIGHT = colors.HexColor("#6b7280")
    BORDER = colors.HexColor("#cfd8e3")
    ROW_BG = colors.HexColor("#f8fbff")
    LABEL_BG = colors.HexColor("#eef4fb")

    slip_title_style = ParagraphStyle(
        "slip_title_style",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=17,
        leading=20,
        textColor=PRIMARY_BLUE,
        alignment=TA_CENTER,
        spaceAfter=4,
    )

    table_header_style = ParagraphStyle(
        "table_header_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9.5,
        leading=11,
        textColor=colors.white,
        alignment=TA_CENTER,
    )

    label_style = ParagraphStyle(
        "label_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9.5,
        leading=11.5,
        textColor=DARK,
        alignment=TA_LEFT,
    )

    value_style = ParagraphStyle(
        "value_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9.5,
        leading=11.5,
        textColor=DARK,
        alignment=TA_LEFT,
        wordWrap="CJK",
    )

    signature_style_left = ParagraphStyle(
        "signature_style_left",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=14,
        textColor=DARK,
        alignment=TA_LEFT,
    )

    signature_style_right = ParagraphStyle(
        "signature_style_right",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=14,
        textColor=DARK,
        alignment=TA_RIGHT,
    )

    doc = BaseDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"Library E-Slip {req.id}",
        showBoundary=0,
    )

    def info_table(rows, widths):
        data = [[Paragraph("Field", table_header_style), Paragraph("Value", table_header_style)]]
        for k, v in rows:
            data.append([
                Paragraph(str(k), label_style),
                Paragraph(str(v) if v not in (None, "") else "-", value_style)
            ])

        t = Table(data, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
            ("BACKGROUND", (0, 1), (0, -1), LABEL_BG),
            ("ROWBACKGROUNDS", (1, 1), (-1, -1), [colors.white, ROW_BG]),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return t

    HEADER_HEIGHT = 34 * mm

    def draw_header_footer(canvas, doc_):
        canvas.saveState()

        page_w, page_h = A4
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 6 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(
                ImageReader(logo_path),
                left,
                top_y - 18 * mm,
                width=24 * mm,
                height=15 * mm,
                preserveAspectRatio=True,
                mask="auto"
            )

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(center_x, top_y - 5 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 9)
        canvas.drawCentredString(center_x, top_y - 11 * mm, "An Autonomous Institution")

        canvas.setFillColor(MID)
        canvas.setFont("Helvetica", 7.5)
        canvas.drawCentredString(center_x, top_y - 16 * mm, "Approved by AICTE, New Delhi")
        canvas.drawCentredString(center_x, top_y - 20 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
        canvas.drawCentredString(center_x, top_y - 24 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

        canvas.setStrokeColor(BORDER)
        canvas.setLineWidth(0.7)
        canvas.line(left, top_y - 27 * mm, right, top_y - 27 * mm)

        footer_y = 10 * mm
        canvas.line(left, footer_y + 6 * mm, right, footer_y + 6 * mm)
        canvas.setFillColor(LIGHT)
        canvas.setFont("Helvetica", 7.5)
        canvas.drawString(left, footer_y, f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")

        canvas.restoreState()

    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin + 6 * mm,
        doc.width,
        doc.height - HEADER_HEIGHT - 4 * mm,
        leftPadding=0,
        rightPadding=0,
        topPadding=0,
        bottomPadding=0,
        id="normal"
    )

    doc.addPageTemplates([
        PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)
    ])

    book_type_val = "-"
    if req.book and getattr(req.book, "book_type", None):
        book_type_val = getattr(req.book.book_type, "book_type", "-") or "-"

    issued_on = str(req.issued_on) if getattr(req, "issued_on", None) else "-"
    returned_on = str(req.returned_on) if getattr(req, "returned_on", None) else "-"
    requested_on = str(req.requested_on) if getattr(req, "requested_on", None) else "-"

    elements = []

    elements.append(Spacer(1, 2 * mm))
    elements.append(Paragraph("LIBRARY E-SLIP", slip_title_style))
    elements.append(Spacer(1, 3 * mm))

    top_slip = Table([[
        Paragraph(f"<b>Request No:</b> {req.id}", value_style),
        Paragraph(f"<b>Requested Date:</b> {requested_on}", value_style),
        Paragraph(f"<b>Status:</b> {req.status or '-'}", value_style),
    ]], colWidths=[doc.width * 0.28, doc.width * 0.34, doc.width * 0.38])

    top_slip.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eef4fb")),
        ("BOX", (0, 0), (-1, -1), 0.7, BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(top_slip)
    elements.append(Spacer(1, 5 * mm))

    student_rows = [
        ("Student Name", req.student_name or "-"),
        ("Register No", req.student_rollno or "-"),
        ("Department", str(req.department) if req.department else "-"),
    ]

    book_rows = [
        ("Book Title", req.book.title if req.book else "-"),
        ("Accession No", req.book.accession_book if req.book and req.book.accession_book else "-"),
        ("Authors", req.book.authors if req.book and req.book.authors else "-"),
        ("Book Type", book_type_val),
    ]

    student_tbl = info_table(student_rows, [34 * mm, (doc.width / 2) - 34 * mm - 3 * mm])
    book_tbl = info_table(book_rows, [34 * mm, (doc.width / 2) - 34 * mm - 3 * mm])

    dual = Table([[student_tbl, book_tbl]], colWidths=[doc.width / 2, doc.width / 2])
    dual.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(dual)
    elements.append(Spacer(1, 5 * mm))

    issue_rows = [
        ("Issued On", issued_on),
        ("Returned On", returned_on),
        ("Remarks", getattr(req, "remarks", None) or "-"),
    ]
    elements.append(info_table(issue_rows, [36 * mm, doc.width - 36 * mm]))
    elements.append(Spacer(1, 18 * mm))

    sign_tbl = Table(
        [[
            Paragraph("Student Signature", signature_style_left),
            Paragraph("Department Library Incharge Signature", signature_style_right),
        ]],
        colWidths=[doc.width / 2, doc.width / 2]
    )

    sign_tbl.setStyle(TableStyle([
        ("TOPPADDING", (0, 0), (-1, -1), 42),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
    ]))
    elements.append(sign_tbl)

    try:
        doc.build(elements)
    except Exception as e:
        return HttpResponse(f"PDF generation failed: {e}", status=500)

    buffer.seek(0)
    return FileResponse(buffer, content_type="application/pdf")
   

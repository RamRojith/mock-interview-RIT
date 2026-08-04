from django.shortcuts import render,redirect,get_object_or_404
from user_accounts.decorators import no_cache,is_super_user
from django.contrib import messages
from course_management.models import PeriodAllocation, CourseHours, CourseEnrollment, AssignSubjectFaculty 
from examination_management.models import *
from django.shortcuts import render,redirect,get_object_or_404
from user_accounts.decorators import no_cache,is_super_user
from django.contrib import messages
from course_management.models import PeriodAllocation, CourseHours, CourseEnrollment, AssignSubjectFaculty 
from examination_management.models import *
from user_accounts.models import Add_Department, StudentDetails
from faculty_management.models import ProfessionalSociety, ProgramType
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
from django.core.paginator import Paginator
from django.shortcuts import render
from datetime import datetime, date, timedelta
from user_accounts.decorators import no_cache, check_permission
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.utils import timezone
from django.db.models import Q, Max
import os
import logging
import json

from django.shortcuts import render
from datetime import datetime
from user_accounts.decorators import no_cache, check_permission

from faculty_management.models import InventoryItemType, MaterialIssueProofApprovalData, general_information
from faculty_management.models import general_information, ProgramOrganizationRecord, ProgramOrganizationStudentMark, ProgramOutcomeMapping, program_org_data_Permission

def get_academic_year():
    """Get current academic year based on June boundary"""
    today = date.today()
    if today.month >= 6:
        return f"{today.year}-{today.year + 1}"
    return f"{today.year - 1}-{today.year}"



def get_program_org_permissions(user):
    """
    Get program organization permissions for a user
    Returns tuple: (can_view_all, can_view_dept, permission_obj)
    """
    user_role_id = getattr(user, "role_id", None)
    if not user_role_id:
        return False, False, None
    
    perm = program_org_data_Permission.objects.filter(role_id=user_role_id).first()
    if not perm:
        return False, False, None
    
    can_view_all = bool(getattr(perm, "can_view_all_program_org_data", False))
    can_view_dept = bool(getattr(perm, "can_view_department_program_org_data", False))
    
    return can_view_all, can_view_dept, perm


@no_cache
@check_permission("faculty_timetable")
def faculty_timetable(request):
    faculty_id = request.user.Employee_id  # assuming faculty login
    faculty = general_information.objects.get(faculty_id=faculty_id)
    # Get all assigned subjects for this faculty
    assigned_courses_all = AssignSubjectFaculty.objects.filter(
        Q(faculty=faculty) | Q(skilled_faculty=faculty)
    ).select_related("course")

    # Academic year wise filter
    academic_year_options = list(
        assigned_courses_all.exclude(academic_year__isnull=True)
        .exclude(academic_year="")
        .values_list("academic_year", flat=True)
        .distinct()
    )
    academic_year_options = sorted(academic_year_options, reverse=True)

    selected_academic_year = (request.GET.get("academic_year") or "").strip()
    if not selected_academic_year and academic_year_options:
        selected_academic_year = academic_year_options[0]

    assigned_courses = assigned_courses_all
    if selected_academic_year:
        assigned_courses = assigned_courses.filter(academic_year=selected_academic_year)

    # Build a mapping of course_code → course details
    course_map = {a.course.course_code: a.course for a in assigned_courses if a.course}

    # Get department, semester, section info from assignments
    timetable_allocations = PeriodAllocation.objects.filter(
        department__id__in=[a.department_id for a in assigned_courses if a.department_id],
        semester__in=[a.course.semester for a in assigned_courses if a.course],
        section__in=[a.section for a in assigned_courses if a.section],
    )

    # Define all periods
    all_periods = [
        "first_period", "second_period", "third_period", "fourth_period",
        "fifth_period", "sixth_period", "seventh_period", "eighth_period",
        "nineth_period", "tenth_period"
    ]

    # Group allocations by section, year, and semester
    grouped_allocations = {}
    subjects_set = set()
    subjects_list = []

    for alloc in timetable_allocations:
        # Create a unique key for grouping
        group_key = (alloc.section, alloc.year, alloc.semester)
        if group_key not in grouped_allocations:
            grouped_allocations[group_key] = []

        period_display = {}
        for p in all_periods:
            course_code = getattr(alloc, p)
            if course_code in course_map:
                course = course_map[course_code]
                period_display[p] = {
                    "course_code": course.course_code,
                    "title": course.title,
                    "batch": alloc.year,
                    "section": alloc.section,
                    "semester": alloc.semester,
                }
                # Add unique subject to list
                subject_key = (
                    course.course_code,
                    course.title,
                    alloc.year,
                    alloc.section,
                    alloc.semester,
                    selected_academic_year or "-"
                )
                if subject_key not in subjects_set:
                    subjects_set.add(subject_key)
                    subjects_list.append({
                        "course_code": course.course_code,
                        "title": course.title,
                        "batch": alloc.year,
                        "section": alloc.section,
                        "semester": alloc.semester,
                        "academic_year": selected_academic_year or "-",
                    })
            else:
                period_display[p] = None
        alloc.period_display = period_display
        grouped_allocations[group_key].append(alloc)

    # Sort allocations within each group by weekday order
    weekday_order = {
        "Monday": 1, "Tuesday": 2, "Wednesday": 3,
        "Thursday": 4, "Friday": 5, "Saturday": 6, "Sunday": 7,
    }
    for group_key in grouped_allocations:
        grouped_allocations[group_key] = sorted(
            grouped_allocations[group_key], key=lambda x: weekday_order.get(x.day, 99)
        )

    grouped_allocations_list = [
        {
            "section": key[0],
            "year": key[1],
            "semester": key[2],
            "allocations": allocations,
            "visible_periods": [
                p for p in all_periods if any(getattr(a, p) for a in allocations)
            ]
        }
        for key, allocations in grouped_allocations.items()
    ]

    total_classes = 0
    for group in grouped_allocations_list:
        for alloc in group["allocations"]:
            for p in group["visible_periods"]:
                if alloc.period_display.get(p):
                    total_classes += 1

    # Prepare context with grouped allocations
    context = {
        "grouped_allocations": grouped_allocations_list,
        "current_day": datetime.today().strftime("%A"),
        "today": datetime.today(),
        "subjects_list": sorted(subjects_list, key=lambda x: (x["section"], x["batch"], x["semester"])),
        "academic_year_options": academic_year_options,
        "selected_academic_year": selected_academic_year,
        "total_classes": total_classes,
    }

    return render(request, "faculty_management/faculty_timetable.html", context)



#from faculty_management.models import PeriodSubstitution
# @no_cache
# @check_permission("faculty_timetable")
# def faculty_timetable(request):
#     faculty_id = request.user.Employee_id

#     try:
#         faculty = general_information.objects.get(faculty_id=faculty_id)
#     except general_information.DoesNotExist:
#         messages.error(request, "Faculty profile not found.")
#         return render(request, "faculty_management/faculty_timetable.html", {})

#     current_academic_year = get_academic_year()

#     # Faculty's assigned subjects this year
#     assigned = AssignSubjectFaculty.objects.filter(
#         faculty=faculty,
#         academic_year=current_academic_year
#     ).select_related('course')

#     if not assigned.exists():
#         messages.warning(request, "No subjects assigned to you this academic year.")
#         # You can still continue or return early — your choice

#     # Departments this faculty is teaching in
#     faculty_depts = assigned.values_list('department_id', flat=True).distinct()

#     # ────────────────────────────────────────────────
#     # Substitute candidates: only active teachers in same dept(s)
#     # ────────────────────────────────────────────────
#     active_faculty_ids = AssignSubjectFaculty.objects.filter(
#         academic_year=current_academic_year,
#         department_id__in=faculty_depts,
#     ).values_list('faculty_id', flat=True).distinct()

#     possible_substitutes = general_information.objects.filter(
#         id__in=active_faculty_ids
#     ).exclude(id=faculty.id).select_related('department').order_by('name')

#     # ────────────────────────────────────────────────
#     # POST: Assign substitute for TODAY only
#     # ────────────────────────────────────────────────
#     if request.method == "POST" and request.POST.get("action") == "assign_substitute":
#         alloc_id = request.POST.get("allocation_id")
#         period_field = request.POST.get("period_field")
#         substitute_id = request.POST.get("substitute_faculty")
#         reason = request.POST.get("reason", "absent")
#         remarks = request.POST.get("remarks", "")

#         try:
#             allocation = PeriodAllocation.objects.get(id=alloc_id)
#             substitute = general_information.objects.get(id=substitute_id)

#             today_str = date.today().strftime("%A")

#             if allocation.day != today_str:
#                 messages.error(request, "Substitution can only be assigned for today.")
#             elif not getattr(allocation, period_field, None):
#                 messages.error(request, "No class scheduled in this period.")
#             else:
#                 already_exists = PeriodSubstitution.objects.filter(
#                     original_allocation=allocation,
#                     substitution_date=date.today(),
#                     period_field=period_field
#                 ).exists()

#                 if already_exists:
#                     messages.warning(request, "Substitution already assigned for this slot today.")
#                 else:
#                     PeriodSubstitution.objects.create(
#                         original_allocation=allocation,
#                         substitution_date=date.today(),
#                         period_field=period_field,
#                         original_faculty=faculty,
#                         substitute_faculty=substitute,
#                         reason=reason,
#                         remarks=remarks,
#                         # created_by=faculty,
#                     )
#                     messages.success(request, f"Substitution assigned to {substitute.name} for today.")

#         except PeriodAllocation.DoesNotExist:
#             messages.error(request, "Invalid timetable entry.")
#         except general_information.DoesNotExist:
#             messages.error(request, "Selected faculty not found.")
#         except Exception as e:
#             messages.error(request, f"Error: {str(e)}")

#         return redirect('faculty_timetable')

#     # ────────────────────────────────────────────────
#     # GET: Show only assigned years/sections
#     # ────────────────────────────────────────────────
#     assigned_combinations = assigned.values(
#         'course__year',
#         'course__semester',
#         'section'
#     ).distinct()

#     assigned_groups = [
#         (item['course__year'], item['course__semester'], item['section'])
#         for item in assigned_combinations
#         if item['course__year'] and item['course__semester'] and item['section']
#     ]

#     # Sort: even sem first, then higher year first
#     def sort_key(group):
#         year, sem, _ = group
#         year_num = int(year) if year.isdigit() else {"I":1, "II":2, "III":3, "IV":4}.get(year.upper(), 0)
#         sem_num = int(sem) if sem.isdigit() else 0
#         parity = sem_num % 2  # 0=even, 1=odd → even first
#         return (parity, -year_num, sem_num)

#     assigned_groups.sort(key=sort_key)

#     # Build Q filter for only assigned groups
#     q_filters = Q()
#     for yr, sem, sec in assigned_groups:
#         q_filters |= Q(year=yr, semester=sem, section=sec)

#     timetable_allocations = PeriodAllocation.objects.filter(q_filters).select_related('department')

#     course_map = {a.course.course_code: a.course for a in assigned if a.course}

#     all_periods = [
#         "first_period", "second_period", "third_period", "fourth_period",
#         "fifth_period", "sixth_period", "seventh_period", "eighth_period",
#         "nineth_period", "tenth_period"
#     ]

#     weekday_order = {
#         "Monday": 1, "Tuesday": 2, "Wednesday": 3,
#         "Thursday": 4, "Friday": 5, "Saturday": 6, "Sunday": 7,
#     }

#     grouped_allocations = {}
#     subjects_set = set()
#     subjects_list = []

#     today = date.today()
#     current_day_name = today.strftime("%A")

#     for alloc in timetable_allocations:
#         group_key = (alloc.section, alloc.year, alloc.semester)
#         if group_key not in grouped_allocations:
#             grouped_allocations[group_key] = []

#         period_display = {}
#         for p in all_periods:
#             course_code = getattr(alloc, p)
#             if course_code and course_code in course_map:
#                 course = course_map[course_code]
#                 period_display[p] = {
#                     "course_code": course.course_code,
#                     "title": course.title or course.course_code,
#                     "batch": alloc.year,
#                     "section": alloc.section,
#                     "semester": alloc.semester,
#                     "is_substitution": False,
#                 }

#                 subject_key = (course.course_code, alloc.year, alloc.section, alloc.semester)
#                 if subject_key not in subjects_set:
#                     subjects_set.add(subject_key)
#                     subjects_list.append({
#                         "course_code": course.course_code,
#                         "title": course.title or course.course_code,
#                         "batch": alloc.year,
#                         "section": alloc.section,
#                         "semester": alloc.semester,
#                     })
#             else:
#                 period_display[p] = None

#         alloc.period_display = period_display
#         grouped_allocations[group_key].append(alloc)

#     # Sort days
#     for key in grouped_allocations:
#         grouped_allocations[key] = sorted(
#             grouped_allocations[key],
#             key=lambda x: weekday_order.get(x.day, 99)
#         )

#     # Today's substitutions
#     today_substitutions = PeriodSubstitution.objects.filter(
#         substitution_date=today
#     ).select_related('original_allocation', 'substitute_faculty', 'original_faculty')

#     substitution_map = {}
#     for sub in today_substitutions:
#         alloc_id = sub.original_allocation_id
#         substitution_map.setdefault(alloc_id, {})[sub.period_field] = sub

#     for group in grouped_allocations.values():
#         for alloc in group:
#             if alloc.day == current_day_name:
#                 subs = substitution_map.get(alloc.id, {})
#                 for period_name, sub in subs.items():
#                     if period_name in alloc.period_display and alloc.period_display[period_name]:
#                         alloc.period_display[period_name].update({
#                             "is_substitution": True,
#                             "substitute_faculty": sub.substitute_faculty,
#                             "original_faculty": sub.original_faculty,
#                             "reason": sub.get_reason_display(),
#                             "remarks": sub.remarks,
#                         })

#     # Final grouped list
#     grouped_list = [
#         {
#             "section": key[0] or "—",
#             "year": key[1],
#             "semester": key[2],
#             "allocations": allocations,
#             "visible_periods": [
#                 p for p in all_periods
#                 if any(getattr(a, p) for a in allocations)
#             ]
#         }
#         for key, allocations in sorted(
#             grouped_allocations.items(),
#             key=lambda item: sort_key((item[0][1], item[0][2], item[0][0]))
#         )
#     ]

#     context = {
#         "grouped_allocations": grouped_list,
#         "current_day": current_day_name,
#         "today": today,
#         "subjects_list": sorted(subjects_list, key=lambda x: (x["section"], x["batch"], x["semester"])),
#         "today_substitutions": today_substitutions,
#         "all_faculties": possible_substitutes,
#         "faculty": faculty,
#         "academic_year": current_academic_year,
#     }

#     return render(request, "faculty_management/faculty_timetable.html", context)



from user_accounts.models import *
from course_management.models import SectionMaster
from django.utils.safestring import mark_safe

import pprint
import logging

logger = logging.getLogger(__name__)
pp = pprint.PrettyPrinter(indent=2, width=120, compact=False)

def _coerce_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None

@check_permission("assign_section")
def assign_section(request):
    # ---------- Dropdown Data ----------
    faculty = general_information.objects.get(faculty_id=request.user.Employee_id)
    department = faculty.department
    batches = StudentDetails.objects.values_list("batch", flat=True).distinct().order_by("batch")
    sections = SectionMaster.objects.all().order_by("section")
    selected_batch = request.GET.get("batch")

    students = []
    personal_details_map = {}   # Aadhaar_Number -> PersonalDetails.id
    aadhaar_to_mark = {}        # Aadhaar_Number -> VOC cutoff (float or None)
    pid_to_cutoff = {}          # PersonalDetails.id -> cutoff
    pid_to_quota = {}           # PersonalDetails.id -> Quota (string)

    # ---------- Handle Bulk Section Assignment (POST) ----------
    if request.method == "POST":
        action = request.POST.get("action")
        selected_batch_for_redirect = request.GET.get("batch") or request.POST.get("batch") or selected_batch

        if action == "bulk_assign":
            student_ids = request.POST.getlist("student_ids")
            section_id = request.POST.get("section_id")

            if not student_ids:
                messages.warning(request, "No students selected.")
                return redirect(f"{request.path}?batch={selected_batch_for_redirect or ''}")

            if not section_id:
                messages.error(request, "Select a section to assign.")
                return redirect(f"{request.path}?batch={selected_batch_for_redirect or ''}")

            # Validate Section
            section_obj = get_object_or_404(SectionMaster, id=section_id)

            # Scope update to the selected students (optional: also to the current batch)
            qs = StudentDetails.objects.filter(id__in=student_ids,
        is_active=True)
            # Optional safety: ensure they belong to the currently viewed batch
            if selected_batch_for_redirect:
                qs = qs.filter(batch=selected_batch_for_redirect)

            # Do a single bulk update
            updated_count = qs.update(section=section_obj.section)

            messages.success(
                request,
                mark_safe(f"✅ Assigned section <strong>{section_obj.section}</strong> to <strong>{updated_count}</strong> student(s).")
            )
            return redirect(f"{request.path}?batch={selected_batch_for_redirect or ''}")

        # Unknown POST action
        messages.error(request, "Invalid action.")
        return redirect(f"{request.path}?batch={selected_batch or ''}")

    # ---------- Build GET context (listing) ----------
    if selected_batch:
        students = list(
            StudentDetails.objects
            .filter(batch=selected_batch,department=department)
            .only("id", "name", "reg_no", "aadhar_number", "section", "gender")
            .order_by("reg_no")
        )

        aadhaars = [s.aadhar_number for s in students if s.aadhar_number]

        # # print("\n[DEBUG] Selected batch:", selected_batch)
        # # print("[DEBUG] Students count:", len(students))
        # # print("[DEBUG] Aadhaar list (non-empty):", aadhaars)

        # 1) Map Aadhaar -> PersonalDetails.id (admissionform1 DB)
        if aadhaars:
            p_qs = (
                PersonalDetails.objects.using("admissionform1")
                .filter(Aadhaar_Number__in=aadhaars)
                .values("Aadhaar_Number", "id")
            )
            for r in p_qs:
                personal_details_map[r["Aadhaar_Number"]] = r["id"]

        # # print("[DEBUG] Aadhaar -> PersonalDetails.id map:")
        # pp.pprint(personal_details_map)
        # logger.debug("Aadhaar->PID map: %s", personal_details_map)

        # 2) Fetch HSC rows by PersonalDetails ids; build pid -> cutoff
        pid_list = list(personal_details_map.values())
        if pid_list:
            hsc_rows = (
                HSCDetails.objects.using("admissionform1")
                .filter(school_details_id__in=pid_list)
                .values("school_details_id", "twelfth_std_aca_cut_off_mark")
            )
            for row in hsc_rows:
                pid = row["school_details_id"]
                pid_to_cutoff[pid] = _coerce_float(row["twelfth_std_aca_cut_off_mark"])

        # # print("[DEBUG] PersonalDetails.id -> VOC cutoff map:")
        # pp.pprint(pid_to_cutoff)
        # logger.debug("PID->cutoff map: %s", pid_to_cutoff)

        # 2b) Fetch AdmissionRecords quota by PersonalDetails ids; build pid -> quota
        if pid_list:
            ar_rows = (
                AdmissionRecords.objects.using("admissionform1")
                .filter(PersonalDetailsId_id__in=pid_list)
                .values("PersonalDetailsId_id", "Quota")
            )
            for r in ar_rows:
                pid = r["PersonalDetailsId_id"]
                quota_raw = r.get("Quota") or ""
                pid_to_quota[pid] = str(quota_raw).strip()

        # # print("[DEBUG] PersonalDetails.id -> Quota map:")
        # pp.pprint(pid_to_quota)
        # logger.debug("PID->quota map: %s", pid_to_quota)

        # 3) Attach cutoff and quota to student objects
        for s in students:
            pid = personal_details_map.get(s.aadhar_number)
            cutoff = pid_to_cutoff.get(pid) if pid else None
            quota = pid_to_quota.get(pid) if pid else ""
            s.voc_cutoff = cutoff
            s.quota = quota
            if s.aadhar_number:
                aadhaar_to_mark[s.aadhar_number] = cutoff

            # # print(f"[DEBUG] Student {s.id} | {s.name} | Aadhaar={s.aadhar_number} | PID={pid} | VOC cutoff={cutoff} | Quota={quota}")
            # logger.debug("Student %s (%s): Aadhaar=%s, PID=%s, cutoff=%s, quota=%s",
            #              s.id, s.name, s.aadhar_number, pid, cutoff, quota)

    # # print("[DEBUG] FINAL: aadhaar_to_mark (Aadhaar -> cutoff):")
    # pp.pprint(aadhaar_to_mark)

    # ---------- GENDER COUNTS (overall for the selected batch) ----------
    gender_counts = {"Male": 0, "Female": 0, "Other": 0}
    for s in students:
        g = (getattr(s, "gender", "") or "").strip()
        if not g:
            gender_counts["Other"] += 1
        else:
            gl = g.lower()
            if gl in ("m", "male"):
                gender_counts["Male"] += 1
            elif gl in ("f", "female"):
                gender_counts["Female"] += 1
            else:
                gender_counts["Other"] += 1

    # ---------- QUOTA COUNTS (overall for the selected batch) ----------
    quota_counts = {}
    for s in students:
        q = (getattr(s, "quota", "") or "").strip()
        if not q:
            q = "Unspecified"
        quota_counts[q] = quota_counts.get(q, 0) + 1

    # Build an ordered list of quotas for stable display (place Unspecified last)
    quotas_list = sorted([q for q in quota_counts.keys() if q != "Unspecified"])
    if "Unspecified" in quota_counts:
        quotas_list.append("Unspecified")

    # Build pairs for template-friendly iteration: [(quota, count), ...]
    quotas_pairs = [(q, quota_counts[q]) for q in quotas_list]

    quotas_list_json = json.dumps(quotas_list)

    # Sort: value first, then desc cutoff, then reg_no
    students.sort(
        key=lambda s: (
            s.voc_cutoff is None,
            -s.voc_cutoff if s.voc_cutoff is not None else 0.0,
            str(s.reg_no)
        )
    )

    context = {
        "batches": batches,
        "sections": sections,
        "selected_batch": selected_batch,
        "students": students,
        "aadhaar_to_mark": aadhaar_to_mark,
        "gender_counts": gender_counts,
        "quota_counts": quota_counts,
        "quotas_list": quotas_list,
        "quotas_list_json": quotas_list_json,
        "quotas_pairs": quotas_pairs,   # safe for template loops
    }
    return render(request, "course_management/faculty/assign_section.html", context)




from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

from examination_management.models import CourseOutcome, BloomsLevel, Assessments
from user_accounts.models import Add_Department, Degree
from course_management.models import Course, Regulations
from faculty_management.models import general_information, Assessment_master  # ensure this model is correct


from examination_management.models import CourseOutcome, BloomsLevel, Assessments
from user_accounts.models import Add_Department, Degree
from course_management.models import Course, Regulations
from faculty_management.models import general_information, Assessment_master  # ensure this model is correct
import json

# faculty_management/views_ajax.py
# faculty_management/views_ajax.py
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required

from user_accounts.models import Add_Department, Degree
from faculty_management.models import general_information, AssignSubjectFaculty

def _get_faculty(request):
    """Resolve the general_information record for the logged-in user, if present."""
    user = getattr(request, "user", None)
    emp_id = getattr(user, "Employee_id", None) or getattr(user, "employee_id", None) if user else None
    if not emp_id:
        return None
    return general_information.objects.filter(faculty_id=emp_id).first()

def _degree_or_400(deg_id):
    if not deg_id:
        return None, JsonResponse({"ok": False, "error": "Degree ID missing."}, status=400)
    try:
        degree = Degree.objects.get(pk=int(deg_id))
    except (Degree.DoesNotExist, ValueError, TypeError):
        return None, JsonResponse({"ok": False, "error": "Invalid degree."}, status=400)
    return degree, None


def departments_for_degree(request):
    gi_obj = _get_faculty(request)
    if gi_obj is None:
        return JsonResponse({"ok": False, "error": "Faculty not resolved."}, status=403)

    deg_id = request.POST.get("degree_id")
    degree, err = _degree_or_400(deg_id)
    if err:
        return err

    # Restrict to departments assigned to this faculty under this degree
    deps = (
        Add_Department.objects
        .filter(degree=degree, is_active=True, assignsubjectfaculty__faculty=gi_obj, assignsubjectfaculty__is_active=True)
        .order_by("Department")
        .distinct()
        .only("id", "Department", "Department_code", "department_label")
    )

    return JsonResponse({
        "ok": True,
        "departments": [
            {
                "id": d.id,
                "label": d.Department or "Unnamed Department",
                "code": d.Department_code or "",
                "label_alt": d.department_label or "",
            }
            for d in deps
        ],
    })



def semesters_for_degree(request):
    # semesters still depend only on degree.duration
    deg_id = request.POST.get("degree_id")
    degree, err = _degree_or_400(deg_id)
    if err:
        return err

    try:
        duration_years = int(getattr(degree, "effective_duration", None) or getattr(degree, "duration", 0) or 0)
    except (TypeError, ValueError):
        duration_years = 0

    total = max(0, duration_years * 2)
    return JsonResponse({
        "ok": True,
        "semesters": [{"value": i, "label": str(i)} for i in range(1, total + 1)],  # plain numbers
    })








import logging
from decimal import Decimal, InvalidOperation

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpRequest
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.db import models as dj_models

# Import your models (adjust imports to your project's layout)
from examination_management.models import  AssessmentWeightage
from django.db.models import Q, F



logger = logging.getLogger(__name__)


@require_POST
def ajax_assessment_weight(request: HttpRequest):
    """
    AJAX endpoint: expects POST { degree_id, regulation_id }.
    Returns {"ok": True, "percentage": "30"} (string formatted without trailing .0)
    or {"ok": True, "percentage": None} when no record found.
    """
    deg_id = request.POST.get("degree_id")
    reg_id = request.POST.get("regulation_id")

    # Debug logging
    logger.debug("AJAX: ajax_assessment_weight called with degree_id=%r regulation_id=%r", deg_id, reg_id)
    # print("AJAX: ajax_assessment_weight called with degree_id=", deg_id, " regulation_id=", reg_id)

    if not deg_id or not reg_id:
        logger.debug("AJAX: Missing parameters for ajax_assessment_weight")
        return JsonResponse({"ok": False, "error": "Missing parameters"}, status=400)

    try:
        degree = Degree.objects.get(pk=int(deg_id))
        regulation = Regulations.objects.get(pk=int(reg_id))
    except (Degree.DoesNotExist, Regulations.DoesNotExist, ValueError) as e:
        logger.debug("AJAX: invalid degree/regulation: %s", e)
        # print("AJAX: invalid degree/regulation:", repr(e))
        return JsonResponse({"ok": False, "error": "Invalid degree or regulation"}, status=400)

    aw = AssessmentWeightage.objects.filter(degree=degree, regulation=regulation).order_by('-created_at').first()

    logger.debug("AJAX: AssessmentWeightage lookup result: %r", aw)
    # print("AJAX: AssessmentWeightage lookup result:", repr(aw))

    if not aw:
        return JsonResponse({"ok": True, "percentage": None})

    # Format so 30.0 -> "30" and 30.5 -> "30.5"
    try:
        val = float(aw.selected_assessment_percentage)
    except Exception:
        # Fallback: return raw value as string
        formatted = str(aw.selected_assessment_percentage)
        return JsonResponse({"ok": True, "percentage": formatted})

    formatted = "%g" % val
    logger.debug("AJAX: returning formatted percentage: %s", formatted)
    return JsonResponse({"ok": True, "percentage": formatted})




@require_POST
def ajax_cos_for_regulation(request):
    regulation_id = request.POST.get("regulation_id")
    if not regulation_id:
        return JsonResponse({"ok": False, "error": "Missing regulation_id"}, status=400)

    try:
        regulation = Regulations.objects.get(pk=int(regulation_id))
    except (Regulations.DoesNotExist, ValueError, TypeError):
        return JsonResponse({"ok": False, "error": "Invalid regulation"}, status=400)

    regulation_year = str(
        getattr(regulation, "Regulation", "") or
        getattr(regulation, "regulation", "") or
        getattr(regulation, "year", "") or
        ""
    ).strip()

    cos = CourseOutcome.objects.filter(
        regulation__iexact=regulation_year
    ).exclude(
        co_code__isnull=True
    ).exclude(
        co_code__exact=""
    ).order_by("co_code")

    data = [
        {
            "id": co.id,
            "code": co.co_code,
            "name": co.co_name or "",
            "label": f"{co.co_code} - {co.co_name}" if co.co_name else co.co_code,
        }
        for co in cos
    ]

    print("AJAX CO DEBUG => regulation_id =", regulation_id)
    print("AJAX CO DEBUG => regulation_year =", regulation_year)
    print("AJAX CO DEBUG => matched CO count =", len(data))
    print("AJAX CO DEBUG => matched CO data =", data)

    return JsonResponse({"ok": True, "items": data})

@require_POST
def ajax_blooms_levels(request):
    levels = BloomsLevel.objects.exclude(
        level_code__isnull=True
    ).exclude(
        level_code__exact=""
    ).order_by("level_code")

    data = [
        {
            "id": lv.id,
            "code": lv.level_code,
            "name": lv.description or "",
            "label": f"{lv.level_code} - {lv.description}" if lv.description else lv.level_code,
        }
        for lv in levels
    ]
    return JsonResponse({"ok": True, "items": data})


@require_POST
def ajax_batch_sections_for_course(request):
    gi_obj = _get_faculty(request)
    if gi_obj is None:
        return JsonResponse({"ok": False, "error": "Faculty not resolved."}, status=403)

    department_id = request.POST.get("department_id")
    regulation_id = request.POST.get("regulation_id")
    course_id = request.POST.get("course_id")

    if not department_id or not regulation_id or not course_id:
        return JsonResponse({"ok": False, "error": "Missing required parameters."}, status=400)

    try:
        department = Add_Department.objects.get(pk=int(department_id), is_active=True)
        regulation = Regulations.objects.get(pk=int(regulation_id))
        course = Course.objects.get(pk=int(course_id))
    except (Add_Department.DoesNotExist, Regulations.DoesNotExist, Course.DoesNotExist, ValueError, TypeError):
        return JsonResponse({"ok": False, "error": "Invalid selection."}, status=400)

    assignments = (
        AssignSubjectFaculty.objects
        .filter(
            faculty=gi_obj,
            department=department,
            regulation=regulation,
            course=course,
            is_active=True
        )
        .order_by("batch", "section", "id")
    )

    items = []
    for row in assignments:
        batch_val = (row.batch or "").strip()
        section_val = (row.section or "").strip()

        if batch_val and section_val:
            label = f"{batch_val} - {section_val}"
        elif batch_val:
            label = batch_val
        elif section_val:
            label = section_val
        else:
            label = "No Batch / No Section"

        items.append({
            "id": row.id,
            "batch": batch_val,
            "section": section_val,
            "label": label,
        })

    return JsonResponse({"ok": True, "items": items})
@check_permission("add_assessments")
def add_assessments(request: HttpRequest):
    action = (request.POST.get("action") or request.GET.get("action") or "").lower()
    pk = request.POST.get("id") or request.GET.get("id")
    edit = None

    faculty_id_value = ""
    gi_obj = None
    user = getattr(request, "user", None)
    emp_id = (getattr(user, "Employee_id", None) or getattr(user, "employee_id", None)) if user else None
    if emp_id:
        gi_obj = general_information.objects.filter(faculty_id=emp_id).first()
        faculty_id_value = str(getattr(gi_obj, "faculty_id", "") or emp_id)

    if gi_obj:
        allowed_degrees_qs = (
            Degree.objects.filter(
                add_department__assignsubjectfaculty__faculty=gi_obj,
                is_active=True
            )
            .distinct()
            .order_by("degree")
        )
    else:
        allowed_degrees_qs = Degree.objects.none()

    if request.method == "POST":
        # ---------------- SYNC FROM MASTER ----------------
        if action == "sync_assessment_name":
            if not pk:
                messages.error(request, "Missing id for sync.")
                return redirect("add_assessments")

            am = get_object_or_404(Assessment_master, pk=pk)

            if gi_obj and am.degree and not allowed_degrees_qs.filter(pk=am.degree.pk).exists():
                messages.error(request, "You are not allowed to sync this assessment.")
                return redirect("add_assessments")

            if not am.assessment:
                messages.error(request, "No linked predefined Assessment to sync from.")
                return redirect("add_assessments")

            needs_update = False
            update_fields = []

            source_name = (am.assessment.assessment_name or "").strip() or None
            if am.Assessmentname != source_name:
                am.Assessmentname = source_name
                update_fields.append("Assessmentname")
                needs_update = True

            if am.internal_assessment_id != am.assessment.internal_assessment_id:
                am.internal_assessment = am.assessment.internal_assessment
                update_fields.append("internal_assessment")
                needs_update = True

            if not needs_update:
                messages.info(request, "Assessment is already up to date.")
            else:
                am.save(update_fields=update_fields)
                messages.success(request, "Assessment name and IAT updated from master.")

            return redirect("add_assessments")

        # ---------------- DELETE ----------------
        if action == "delete":
            if not pk:
                messages.error(request, "Missing id for deletion.")
                return redirect("add_assessments")
            get_object_or_404(Assessment_master, pk=pk).delete()
            messages.success(request, "Assessment deleted.")
            return redirect("add_assessments")

        # ---------------- CREATE / EDIT ----------------
        try:
            degree_id = int(request.POST.get("degree") or 0)
            department_id = int(request.POST.get("department") or 0)
            regulation_id = int(request.POST.get("regulation") or 0)
            course_id = int(request.POST.get("course") or 0)
        except (TypeError, ValueError):
            messages.error(request, "Please select valid options for all required fields.")
            return redirect("add_assessments")

        degree = get_object_or_404(Degree, pk=degree_id)
        if not allowed_degrees_qs.filter(pk=degree.pk).exists():
            messages.error(request, "Selected degree is not assigned to you.")
            return redirect("add_assessments")

        department = get_object_or_404(Add_Department, pk=department_id, degree=degree, is_active=True)
        if not gi_obj or not AssignSubjectFaculty.objects.filter(
            Q(faculty=gi_obj) | Q(skilled_faculty=gi_obj),
            department=department,
            is_active=True
        ).exists():
            messages.error(request, "Selected department is not assigned to you.")
            return redirect("add_assessments")

        regulation = get_object_or_404(Regulations, pk=regulation_id)
        if not AssignSubjectFaculty.objects.filter(
            Q(faculty=gi_obj) | Q(skilled_faculty=gi_obj),
            department=department,
            regulation=regulation,
            is_active=True
        ).exists():
            messages.error(request, "Selected regulation is not assigned to you for this department.")
            return redirect("add_assessments")

        course = get_object_or_404(Course, pk=course_id)

        assignment_row_id = (request.POST.get("assignment_id") or "").strip()
        if not assignment_row_id:
            messages.error(request, "Please select Batch - Section.")
            return redirect("add_assessments")

        try:
            assignment_obj = AssignSubjectFaculty.objects.get(
                pk=int(assignment_row_id),
                department=department,
                regulation=regulation,
                course=course,
                is_active=True
            )
            if assignment_obj.faculty_id != gi_obj.id and assignment_obj.skilled_faculty_id != gi_obj.id:
                raise AssignSubjectFaculty.DoesNotExist
        except (AssignSubjectFaculty.DoesNotExist, ValueError, TypeError):
            messages.error(request, "Selected Batch - Section is invalid for this course.")
            return redirect("add_assessments")

        try:
            sem_int = int((request.POST.get("semester") or "").strip())
        except (TypeError, ValueError):
            sem_int = 0

        try:
            max_sem = int(getattr(degree, "effective_duration", None) or getattr(degree, "duration", 0) or 0) * 2
        except (TypeError, ValueError):
            max_sem = 0

        if sem_int < 1 or sem_int > max_sem:
            messages.error(request, f"Invalid semester. Degree has {max_sem} semesters.")
            return redirect("add_assessments")

        semester = str(sem_int)

        # MULTIPLE MODULES
        module_values = request.POST.getlist("module")
        module_values = [str(m).strip() for m in module_values if str(m).strip()]

        if not module_values:
            messages.error(request, "Select at least one Unit / Module.")
            return redirect("add_assessments")

        # keep numeric sorted if possible
        try:
            module_values = sorted(set(module_values), key=lambda x: int(x))
        except Exception:
            module_values = list(dict.fromkeys(module_values))

        module = ",".join(module_values)

        # MULTIPLE COs
        co_ids = request.POST.getlist("co_codes")

        # MULTIPLE BLOOMS
        bloom_ids = request.POST.getlist("level_codes")

        if not co_ids:
            messages.error(request, "Select at least one CO.")
            return redirect("add_assessments")

        if not bloom_ids:
            messages.error(request, "Select at least one Bloom Level.")
            return redirect("add_assessments")

        regulation_year = str(
            getattr(regulation, "Regulation", "") or
            getattr(regulation, "regulation", "") or
            getattr(regulation, "year", "") or
            ""
        ).strip()

        print("SAVE DEBUG => co_ids =", co_ids)
        print("SAVE DEBUG => regulation_year =", regulation_year)

        co_qs = CourseOutcome.objects.filter(
            id__in=co_ids,
            regulation__iexact=regulation_year
        )

        print("SAVE DEBUG => co_qs count =", co_qs.count())
        print("SAVE DEBUG => co_qs ids =", list(co_qs.values_list("id", flat=True)))

        if co_qs.count() != len(co_ids):
            messages.error(request, "Some selected COs do not belong to the chosen regulation year.")
            return redirect("add_assessments")

        blooms_qs = BloomsLevel.objects.filter(id__in=bloom_ids)
        if blooms_qs.count() != len(bloom_ids):
            messages.error(request, "Some selected Bloom Levels are invalid.")
            return redirect("add_assessments")

        iat_raw = (request.POST.get("internal_assessment") or "").strip()
        iat_obj = None
        if iat_raw:
            try:
                iat_obj = get_object_or_404(InternalAssessment, pk=int(iat_raw))
            except (TypeError, ValueError):
                messages.error(request, "Invalid Internal Assessment.")
                return redirect("add_assessments")

            if iat_obj.degree_id != degree.id:
                messages.error(request, "Selected Internal Assessment does not belong to the chosen Degree.")
                return redirect("add_assessments")

        a_id = (request.POST.get("assessment_id") or "").strip()
        assessment_obj = None
        custom_assessment_text = None
        assessment_display_name = None

        if a_id == "__custom__":
            custom_assessment_text = (request.POST.get("assessment_name") or "").strip()
            if not custom_assessment_text:
                messages.error(request, "Enter a custom assessment name.")
                return redirect("add_assessments")
            assessment_obj = None
            assessment_display_name = None

        elif a_id == "__iat__":
            if not iat_obj or not (iat_obj.iat or "").strip():
                messages.error(request, "Select a valid Internal Assessment to use its name.")
                return redirect("add_assessments")
            assessment_obj = None
            assessment_display_name = iat_obj.iat.strip()

        else:
            if a_id:
                try:
                    assessment_obj = get_object_or_404(Assessments, pk=int(a_id))
                except (TypeError, ValueError):
                    messages.error(request, "Invalid assessment selection.")
                    return redirect("add_assessments")

                assessment_display_name = (assessment_obj.assessment_name or "").strip() or None

        if assessment_obj and assessment_obj.internal_assessment:
            iat_obj = assessment_obj.internal_assessment

        posted_weightage = (request.POST.get("weightage") or "").strip()
        posted_faculty_id = (request.POST.get("faculty_id") or "").strip()
        faculty_id_final = faculty_id_value or posted_faculty_id

        if a_id == "__iat__":
            maxmarks_val = 100
            aw = AssessmentWeightage.objects.filter(
                degree=degree,
                regulation=regulation
            ).order_by("-created_at").first()

            if aw:
                try:
                    pct = float(aw.selected_assessment_percentage)
                    weightage_val = "%g" % pct
                except Exception:
                    weightage_val = str(aw.selected_assessment_percentage)
            else:
                weightage_val = ""
        else:
            try:
                maxmarks_val = int(request.POST.get("Maxmarks") or 0)
            except (TypeError, ValueError):
                messages.error(request, "Max Marks must be a number.")
                return redirect("add_assessments")

            if maxmarks_val <= 0:
                messages.error(request, "Max Marks must be greater than zero.")
                return redirect("add_assessments")

            weightage_val = posted_weightage

        obj = get_object_or_404(Assessment_master, pk=pk) if (action == "edit" and pk) else Assessment_master()
        obj.degree = degree
        obj.department = department
        obj.regulation = regulation
        obj.semester = semester
        obj.course = course
        obj.module = module
        obj.faculty_id = faculty_id_final
        obj.internal_assessment = iat_obj
        obj.assessment = assessment_obj
        obj.customAssessmentname = (custom_assessment_text or None)
        obj.Assessmentname = (assessment_display_name or None)
        obj.Maxmarks = maxmarks_val
        obj.weightage = weightage_val
        obj.batch = (assignment_obj.batch or "").strip() or None
        obj.section = (assignment_obj.section or "").strip() or None
        obj.save()

        obj.co_codes.set(co_qs)
        obj.bloom_levels.set(blooms_qs)

        messages.success(request, f"Assessment {'updated' if action == 'edit' and pk else 'created'} successfully.")
        return redirect("add_assessments")

    if action == "edit" and pk:
        edit = get_object_or_404(Assessment_master, pk=pk)

    assigned_departments = Add_Department.objects.filter(
        assignsubjectfaculty__faculty=gi_obj,
        assignsubjectfaculty__is_active=True
    ).distinct() if gi_obj else Add_Department.objects.none()

    items = (
        Assessment_master.objects
        .select_related(
            "degree", "department", "regulation", "course",
            "assessment", "internal_assessment",
        )
        .prefetch_related("co_codes", "bloom_levels")
        .filter(
            department__in=assigned_departments,
            faculty_id=faculty_id_value,
        )
        .annotate(
            has_name_mismatch=models.Case(
                models.When(
                    Q(assessment__isnull=False) &
                    ~Q(Assessmentname=F("assessment__assessment_name")),
                    then=models.Value(True)
                ),
                default=models.Value(False),
                output_field=models.BooleanField()
            ),
            has_iat_mismatch=models.Case(
                models.When(
                    Q(assessment__isnull=False) &
                    ~Q(internal_assessment_id=F("assessment__internal_assessment_id")),
                    then=models.Value(True)
                ),
                default=models.Value(False),
                output_field=models.BooleanField()
            ),
            needs_master_update=models.Case(
                models.When(
                    (
                        Q(assessment__isnull=False) &
                        ~Q(Assessmentname=F("assessment__assessment_name"))
                    ) |
                    (
                        Q(assessment__isnull=False) &
                        ~Q(internal_assessment_id=F("assessment__internal_assessment_id"))
                    ),
                    then=models.Value(True)
                ),
                default=models.Value(False),
                output_field=models.BooleanField()
            )
        )
        .order_by("-id")
    )

    selected_modules = []
    if edit and edit.module:
        selected_modules = [m.strip() for m in str(edit.module).split(",") if m.strip()]

    for item in items:
        raw_modules = [m.strip() for m in str(item.module or "").split(",") if m.strip()]
        item.module_display = ", ".join([f"Unit {m} / Module {m}" for m in raw_modules]) if raw_modules else "-"

    context = {
        "is_edit": bool(edit),
        "edit": edit,
        "degrees": allowed_degrees_qs,
        "departments": Add_Department.objects.none(),
        "regulations": Regulations.objects.none(),
        "courses": Course.objects.none(),
        "levels": BloomsLevel.objects.all().order_by("id"),
        "assessments": Assessments.objects.all().order_by("assessment_name"),
        "faculty_id": faculty_id_value,
        "selected_modules": selected_modules,
        "max_semesters": (
            edit.degree.effective_duration * 2
        ) if edit and hasattr(edit.degree, "effective_duration") else (
            (edit.degree.duration * 2) if edit and getattr(edit.degree, "duration", None) else None
        ),
        "items": items,
        "module_range": [str(i) for i in range(1, 11)],
    }
    return render(request, "faculty_management/exams/add_assessments.html", context)







from django.db import transaction
from django.db.models import Q
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404

from course_management.models import AssignSubjectFaculty
from examination_management.models import Assessments
from faculty_management.models import Assessment_master, AssessmentMark, general_information
from student_management.models import StudentDetails  # adjust if different

def _resolve_faculty_id(user):
    """
    Prefer mapped faculty_id from general_information; fallback to user.Employee_id.
    This must match what's stored in AssignSubjectFaculty.faculty_id.
    """
    emp = str(getattr(user, "Employee_id", None))
    gi = None
    if emp:
        gi = general_information.objects.filter(faculty_id=emp).only("faculty_id").first()
        if not gi and hasattr(general_information, "Employee_id"):
            gi = general_information.objects.filter(Employee_id=emp).only("faculty_id").first()
    return str(getattr(gi, "faculty_id", None) or emp)





from django.db import transaction
from django.db.models import Q
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, getcontext
from types import SimpleNamespace

from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Sum

from course_management.models import AssignSubjectFaculty
from examination_management.models import Assessments
from faculty_management.models import Assessment_master, AssessmentMark, general_information
from student_management.models import StudentDetails 
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation, DivisionByZero, getcontext# adjust if different
from django.db import transaction
from django.db.models import Q, Sum
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, DivisionByZero, getcontext
from types import SimpleNamespace

getcontext().prec = 28

def _resolve_faculty_id(user):
    emp = getattr(user, "Employee_id", None) or getattr(user, "employee_id", None)
    if not emp:
        return ""
    emp_str = str(emp)
    gi = general_information.objects.filter(faculty_id=emp_str).only("faculty_id").first()
    if not gi and hasattr(general_information, "Employee_id"):
        gi = general_information.objects.filter(Employee_id=emp_str).only("faculty_id").first()
    return str(getattr(gi, "faculty_id", None) or emp_str)


def _assessment_label(am):
    """
    Return label to use for display and matching StudentInternalMark.exam_name.
    Example: 'IAT1'
    """
    return (
        getattr(am, "customAssessmentname", None)
        or getattr(am, "Assessmentname", None)
        or (am.assessment.assessment_name if getattr(am, "assessment", None) else None)
        or ""
    ).strip()


def _is_internal_prefill_allowed(selected_assessment) -> bool:
    """
    Prefill ONLY if selected assessment is internal-type assessment.
    """
    if not selected_assessment:
        return False

    ia = getattr(selected_assessment, "internal_assessment", None)
    if ia and getattr(ia, "iat", None):
        return True

    label = _assessment_label(selected_assessment).upper().replace(" ", "")
    if label.startswith("IAT"):
        return True

    return False


def _calculate_weighted_marks(marks_raw, max_marks, weightage):
    """
    Calculate weighted marks:
        (marks_raw / max_marks) * weightage
    """
    try:
        marks_raw = Decimal(str(marks_raw))
        max_marks = Decimal(str(max_marks))
        weightage = Decimal(str(weightage))

        if max_marks <= 0:
            return Decimal("0.00")

        weighted = (marks_raw / max_marks) * weightage
        return weighted.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    except (InvalidOperation, DivisionByZero, TypeError, ValueError):
        return Decimal("0.00")


# ------------------------------------------------------------
# MAIN VIEW
# ------------------------------------------------------------
@check_permission("add_assessment_marks")
def add_assessment_marks(request):
    faculty_id = _resolve_faculty_id(request.user)
    faculty = general_information.objects.filter(faculty_id=faculty_id).first()

    # ------------------------------------------------------------
    # Faculty assignments
    # ------------------------------------------------------------
    assignments = (
        AssignSubjectFaculty.objects
        .filter(is_active=True, faculty=faculty)
        .select_related("course", "department")
        .order_by("-id")
    )

    sel_asf_id = request.GET.get("asf") or request.POST.get("asf")
    sel_assessment_id = request.GET.get("assessment") or request.POST.get("assessment")

    selected_asf = None
    selected_assessment = None
    semester = None
    max_marks = None
    assessment_weightage = None

    # ------------------------------------------------------------
    # Selected ASF
    # ------------------------------------------------------------
    if sel_asf_id:
        selected_asf = get_object_or_404(
            AssignSubjectFaculty,
            pk=sel_asf_id,
            faculty=faculty,
            is_active=True
        )
        semester = str(getattr(selected_asf.course, "semester", "") or "")

    # ------------------------------------------------------------
    # Assessments list
    # ------------------------------------------------------------
    assessments_qs = Assessment_master.objects.none()

    if selected_asf:
        assessments_qs = (
            Assessment_master.objects
            .filter(
                faculty_id=faculty_id,
                course_id=selected_asf.course_id,
                batch=(selected_asf.batch or "").strip() or None,
                section=(selected_asf.section or "").strip() or None,
            )
            .select_related("assessment", "course", "internal_assessment")
            .order_by("internal_assessment__iat", "module", "id")
        )

    # ------------------------------------------------------------
    # Selected assessment
    # ------------------------------------------------------------
    if selected_asf and sel_assessment_id:
        selected_assessment = get_object_or_404(
            Assessment_master,
            pk=sel_assessment_id,
            faculty_id=faculty_id,
            course_id=selected_asf.course_id,
            batch=(selected_asf.batch or "").strip() or None,
            section=(selected_asf.section or "").strip() or None,
        )

        max_marks = (
            getattr(selected_assessment, "Maxmarks", None)
            or getattr(selected_assessment, "max_marks", None)
        )

        assessment_weightage = getattr(selected_assessment, "weightage", None)

    # ------------------------------------------------------------
    # STUDENT FILTER
    # ------------------------------------------------------------
    students = []

    if selected_asf:
       
        all_course_enrollments = CourseEnrollment.objects.filter(course_id=selected_asf.course_id)
        

        enrolled_only = all_course_enrollments.filter(enroll=True)
       

        enrollments = CourseEnrollment.objects.filter(
            course_id=selected_asf.course_id,
            enroll=True
        ).select_related("student", "course", "department")

       

        if selected_asf.department:
            enrollments = enrollments.filter(department=selected_asf.department)
           

        if selected_asf.batch:
            enrollments = enrollments.filter(batch=selected_asf.batch)
           

        if selected_asf.section:
            enrollments = enrollments.filter(section=selected_asf.section)
            

      
       

        enrolled_student_ids = list(
            enrollments
            .filter(student__isnull=False)
            .values_list("student_id", flat=True)
            .distinct()
        )

      
        students = list(
            StudentDetails.objects
            .filter(id__in=enrolled_student_ids, is_active=True)
            .order_by("reg_no")
        )

      

    # ------------------------------------------------------------
    # Existing marks
    # ------------------------------------------------------------
    marks_map = {}

    if selected_asf and selected_assessment and students:
        existing_marks = AssessmentMark.objects.filter(
            assignment=selected_asf,
            assessment=selected_assessment,
            student__in=students
        )
        marks_map = {m.student_id: m for m in existing_marks}

    # ------------------------------------------------------------
    # AUTO PREFILL FROM StudentInternalMark (GET only)
    # ------------------------------------------------------------
    internal_sum_map = {}

    if (
        request.method == "GET"
        and selected_asf
        and selected_assessment
        and students
        and _is_internal_prefill_allowed(selected_assessment)
    ):
        exam_label = _assessment_label(selected_assessment).strip()

        reg_nos = [
            s.reg_no.strip()
            for s in students
            if s.reg_no
        ]

        im_qs = StudentInternalMark.objects.filter(
            course_id=selected_asf.course_id,
            exam_name__iexact=exam_label
        )

        if selected_asf.batch:
            im_qs = im_qs.filter(batch=selected_asf.batch)

        if selected_asf.section:
            im_qs = im_qs.filter(section=selected_asf.section)

        im_qs = im_qs.filter(
            Q(student__reg_no__in=reg_nos) |
            Q(reg_no__in=reg_nos)
        )

        rows = (
            im_qs
            .values("student__reg_no", "reg_no")
            .annotate(
                raw_sum=Sum("marks_obtained"),
                max_sum=Sum("max_marks")
            )
        )

        for r in rows:
            rn = r.get("student__reg_no") or r.get("reg_no")
            if not rn:
                continue

            internal_sum_map[rn] = {
                "raw": Decimal(str(r["raw_sum"] or 0)),
                "max": Decimal(str(r["max_sum"] or 0)),
            }

        for s in students:
            if s.id in marks_map:
                continue

            rn = (s.reg_no or "").strip()
            if rn not in internal_sum_map:
                continue

            raw = internal_sum_map[rn]["raw"]

            weighted = None
            if max_marks and selected_assessment.weightage:
                weighted = _calculate_weighted_marks(
                    raw,
                    max_marks,
                    selected_assessment.weightage
                )

            marks_map[s.id] = SimpleNamespace(
                marks_raw=raw.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
                marks_weighted=weighted,
                remarks="",
                prefilled=True,
            )

    # ------------------------------------------------------------
    # POST: SAVE MARKS
    # ------------------------------------------------------------
    if request.method == "POST" and selected_asf and selected_assessment:
        save_count = 0

        try:
            max_m = Decimal(str(max_marks)) if max_marks is not None else None
        except Exception:
            max_m = None

        try:
            weightage = Decimal(str(selected_assessment.weightage or 0))
        except Exception:
            weightage = Decimal("0")

        with transaction.atomic():
            for s in students:
                mk = (request.POST.get(f"mark_{s.id}", "") or "").strip()
                rk = (request.POST.get(f"remark_{s.id}", "") or "").strip()

                obj = marks_map.get(s.id)
                if obj and hasattr(obj, "pk"):
                    mark_obj = obj
                else:
                    mark_obj = AssessmentMark(
                        assignment=selected_asf,
                        assessment=selected_assessment,
                        student=s
                    )

                if mk == "":
                    mark_obj.marks_raw = None
                    mark_obj.marks_weighted = None
                    mark_obj.remarks = rk
                    mark_obj.save()
                    continue

                try:
                    marks_raw = Decimal(str(mk))
                except InvalidOperation:
                    messages.error(request, f"Invalid marks for {s.reg_no}")
                    return redirect(
                        f"{request.path}?asf={selected_asf.id}&assessment={selected_assessment.id}"
                    )

                if marks_raw < 0:
                    messages.error(request, f"Marks cannot be negative for {s.reg_no}")
                    return redirect(
                        f"{request.path}?asf={selected_asf.id}&assessment={selected_assessment.id}"
                    )

                if max_m is not None and marks_raw > max_m:
                    messages.error(request, f"Marks exceed max for {s.reg_no}")
                    return redirect(
                        f"{request.path}?asf={selected_asf.id}&assessment={selected_assessment.id}"
                    )

                marks_weighted = _calculate_weighted_marks(
                    marks_raw=marks_raw,
                    max_marks=max_m if max_m is not None else 0,
                    weightage=weightage
                )

                mark_obj.marks_raw = marks_raw.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                mark_obj.marks_weighted = marks_weighted
                mark_obj.remarks = rk
                mark_obj.save()

                save_count += 1

        messages.success(request, f"Saved marks for {save_count} students.")
        return redirect(f"{request.path}?asf={selected_asf.id}&assessment={selected_assessment.id}")

    # ------------------------------------------------------------
    # CONTEXT
    # ------------------------------------------------------------
    context = {
        "faculty_id": faculty_id,
        "assignments": assignments,
        "selected_asf": selected_asf,
        "assessments": assessments_qs,
        "selected_assessment": selected_assessment,
        "students": students,
        "marks_map": marks_map,
        "max_marks": max_marks,
        "semester": semester,
        "assessment_weightage": assessment_weightage,
    }

    return render(
        request,
        "faculty_management/exams/add_assessmesnt_marks.html",
        context
    )
# v# views.py
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import os

from django.http import HttpResponse, Http404
from django.conf import settings
from django.contrib.staticfiles import finders

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.utils import ImageReader

from course_management.models import AssignSubjectFaculty
from faculty_management.models import Assessment_master
from user_accounts.models import StudentDetails

# make sure this import exists in your file
from faculty_management.models import AssessmentMark


def _get_assessment_name_and_max(a: Assessment_master):
    name = None
    max_marks = None

    for cand in [
        getattr(a, "customAssessmentname", None),
        getattr(a, "Assessmentname", None),
        getattr(getattr(a, "assessment", None), "assessment_name", None),
        getattr(a, "assessment_name", None),
    ]:
        if cand:
            name = cand
            break

    for cand in [
        getattr(a, "Maxmarks", None),
        getattr(a, "max_marks", None),
        getattr(getattr(a, "assessment", None), "max_marks", None),
        getattr(a, "maximum_marks", None),
    ]:
        if cand is not None:
            max_marks = cand
            break

    return name or "(unnamed)", max_marks


from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict
import os
import re

from django.http import Http404, HttpResponse
from django.conf import settings
from django.contrib.staticfiles import finders

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib.utils import ImageReader


from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict
import os
import re

from django.http import Http404, HttpResponse
from django.conf import settings
from django.contrib.staticfiles import finders

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)

# make sure these models are imported in your file:
# AssignSubjectFaculty, Assessment_master, AssessmentMark, StudentDetails, StudentInternalMark


from io import BytesIO
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict
import os
import re

from django.http import Http404, HttpResponse
from django.conf import settings
from django.contrib.staticfiles import finders

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)


# def _get_assessment_name_and_max(a: Assessment_master):
#     name = None
#     max_marks = None

#     for cand in [
#         getattr(a, "customAssessmentname", None),
#         getattr(a, "Assessmentname", None),
#         getattr(getattr(a, "assessment", None), "assessment_name", None),
#         getattr(a, "assessment_name", None),
#     ]:
#         if cand:
#             name = cand
#             break

#     for cand in [
#         getattr(a, "Maxmarks", None),
#         getattr(a, "max_marks", None),
#         getattr(getattr(a, "assessment", None), "max_marks", None),
#         getattr(a, "maximum_marks", None),
#     ]:
#         if cand is not None:
#             max_marks = cand
#             break

#     return name or "(unnamed)", max_marks


# def _convert_iat_to_60(iat_obtained):
#     try:
#         obtained = Decimal(str(iat_obtained or 0))
#         return ((obtained / Decimal("100")) * Decimal("60")).quantize(
#             Decimal("1"), rounding=ROUND_HALF_UP
#         )
#     except Exception:
#         return Decimal("0")


# def marks_pdf(request):
#     """
#     Layout:
#     Student | Reg No | IAT Obtained | Convert 60 |
#     Assignment1-CO1 | Assignment1-CO2 | Assignment2-CO1 | Assignment2-CO2 | ...
#     Assess Total | Convert 40 | Grand Total

#     Rules:
#     - IAT total comes from StudentInternalMark
#     - Convert 60 = (IAT obtained / 100) * 60
#     - Assessment columns are shown CO-wise
#     - Since AssessmentMark has no co_code field, same assessment mark is repeated
#       for each CO of that Assessment_master
#     - Assess Total = sum of all displayed assessment CO-wise marks
#     - Assess Max Total = sum of all repeated max marks
#     - Convert 40 = (Assess Total / Assess Max Total) * 40
#     - Grand Total = Convert 60 + Convert 40

#     Summary shown at bottom:
#     - Overall Percentage
#     - >= 90
#     - >= 80 and < 90
#     - >= 70 and < 80
#     - >= 60 and < 70
#     - >= 50 and < 60
#     - < 50 (Fail)
#     """

#     asf_id = (request.GET.get("asf") or "").strip()
#     assessment_id = (request.GET.get("assessment") or "").strip()

#     if not (asf_id and assessment_id):
#         raise Http404("Missing required parameters.")

#     try:
#         assignment = (
#             AssignSubjectFaculty.objects
#             .select_related("course", "department")
#             .get(id=asf_id)
#         )

#         selected_assessment = (
#             Assessment_master.objects
#             .select_related(
#                 "assessment", "course", "department", "regulation", "internal_assessment"
#             )
#             .prefetch_related("co_codes")
#             .get(id=assessment_id)
#         )
#     except (AssignSubjectFaculty.DoesNotExist, Assessment_master.DoesNotExist):
#         raise Http404("Selection not found.")

#     # ------------------------------------------------------------------
#     # Helpers
#     # ------------------------------------------------------------------
#     def _fmt_marks(val, places="0.01"):
#         if val in (None, ""):
#             return "-"
#         try:
#             d = Decimal(str(val))
#             q = Decimal(places)
#             d = d.quantize(q, rounding=ROUND_HALF_UP)
#             if d == d.to_integral_value():
#                 return str(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
#             return str(d)
#         except Exception:
#             return "-"

#     def _to_decimal(val, default="0"):
#         try:
#             if val in (None, ""):
#                 return Decimal(default)
#             return Decimal(str(val))
#         except Exception:
#             return Decimal(default)

#     def _pick_mark_value(mark_obj):
#         if not mark_obj:
#             return None
#         for field in ["marks_raw", "marks", "marks_weighted"]:
#             v = getattr(mark_obj, field, None)
#             if v not in (None, ""):
#                 return v
#         return None

#     def _safe_text(v):
#         return str(v).strip() if v not in (None, "") else "-"

#     def _resolve_exam_name_candidates(assessment_obj):
#         candidates = []
#         for cand in [
#             getattr(getattr(assessment_obj, "internal_assessment", None), "name", None),
#             getattr(getattr(assessment_obj, "internal_assessment", None), "internal_name", None),
#             getattr(assessment_obj, "internal_name", None),
#             getattr(assessment_obj, "exam_name", None),
#             getattr(assessment_obj, "Assessmentname", None),
#             getattr(assessment_obj, "customAssessmentname", None),
#             getattr(getattr(assessment_obj, "assessment", None), "assessment_name", None),
#         ]:
#             if cand:
#                 txt = str(cand).strip()
#                 if txt and txt not in candidates:
#                     candidates.append(txt)
#         return candidates

#     def _convert_to_40(total_obtained, total_max):
#         try:
#             total_obtained = Decimal(str(total_obtained or 0))
#             total_max = Decimal(str(total_max or 0))
#             if total_max <= 0:
#                 return Decimal("0")
#             return ((total_obtained / total_max) * Decimal("40")).quantize(
#                 Decimal("1"), rounding=ROUND_HALF_UP
#             )
#         except Exception:
#             return Decimal("0")

#     def _co_sort_key(name):
#         m = re.match(r"^CO\s*[-_]?(\d+)$", str(name).strip(), re.IGNORECASE)
#         if m:
#             return (0, int(m.group(1)))
#         return (1, str(name).lower())

#     # ------------------------------------------------------------------
#     # Header info
#     # ------------------------------------------------------------------
#     course = assignment.course
#     department = assignment.department or getattr(course, "department", None)

#     department_code = (
#         getattr(department, "Department_code", None)
#         or getattr(department, "dept_code", None)
#         or ""
#     )
#     department_name = (
#         getattr(department, "Department", None)
#         or getattr(department, "name", None)
#         or ""
#     )

#     course_code = getattr(course, "course_code", "") or ""
#     course_title = (
#         getattr(course, "title", None)
#         or getattr(course, "course_title", None)
#         or str(course)
#     )

#     semester = getattr(course, "semester", "") or getattr(selected_assessment, "semester", "")
#     section = getattr(assignment, "section", "") or getattr(selected_assessment, "section", "")
#     batch = getattr(assignment, "batch", "") or getattr(selected_assessment, "batch", "")
#     department_id = getattr(department, "id", None)

#     # ------------------------------------------------------------------
#     # Related assessments
#     # ------------------------------------------------------------------
#     sibling_qs = Assessment_master.objects.select_related(
#         "assessment", "course", "department", "regulation", "internal_assessment"
#     ).prefetch_related("co_codes").filter(course=selected_assessment.course)

#     if selected_assessment.department_id:
#         sibling_qs = sibling_qs.filter(department=selected_assessment.department)

#     if batch and hasattr(Assessment_master, "batch"):
#         sibling_qs = sibling_qs.filter(batch=batch)

#     if section and hasattr(Assessment_master, "section"):
#         sibling_qs = sibling_qs.filter(section=section)

#     if getattr(selected_assessment, "internal_assessment_id", None):
#         sibling_qs = sibling_qs.filter(
#             internal_assessment_id=selected_assessment.internal_assessment_id
#         )
#     else:
#         selected_name, _ = _get_assessment_name_and_max(selected_assessment)
#         sibling_qs = [
#             a for a in sibling_qs
#             if _get_assessment_name_and_max(a)[0] == selected_name
#         ]

#     if not isinstance(sibling_qs, list):
#         sibling_qs = list(sibling_qs.order_by("id"))

#     if not sibling_qs:
#         sibling_qs = [selected_assessment]

#     sibling_ids = [a.id for a in sibling_qs]

#     # ------------------------------------------------------------------
#     # Assessment marks
#     # ------------------------------------------------------------------
#     marks_qs = (
#         AssessmentMark.objects
#         .select_related("student", "assessment")
#         .filter(assignment=assignment, assessment_id__in=sibling_ids)
#         .order_by("student__reg_no", "student__name", "assessment_id", "id")
#     )

#     student_ids = list(marks_qs.values_list("student_id", flat=True).distinct())

#     # key: (student_id, assessment_id) -> {"obt": Decimal, "max": Decimal}
#     assess_mark_map = defaultdict(lambda: {"obt": Decimal("0"), "max": Decimal("0")})

#     for m in marks_qs:
#         raw_val = _pick_mark_value(m)

#         max_val = (
#             getattr(m, "max_marks", None)
#             or getattr(getattr(m, "assessment", None), "Maxmarks", None)
#             or getattr(getattr(m, "assessment", None), "max_marks", None)
#             or getattr(getattr(getattr(m, "assessment", None), "assessment", None), "max_marks", None)
#             or 0
#         )

#         k = (m.student_id, m.assessment_id)
#         assess_mark_map[k]["obt"] += _to_decimal(raw_val)
#         assess_mark_map[k]["max"] += _to_decimal(max_val)

#     # ------------------------------------------------------------------
#     # Assessment CO-wise columns from Assessment_master.co_codes
#     # ------------------------------------------------------------------
#     assessment_columns = []

#     for a in sibling_qs:
#         a_name, a_max = _get_assessment_name_and_max(a)
#         a_max = _to_decimal(a_max)

#         co_list = list(a.co_codes.all().order_by("co_code")) if hasattr(a, "co_codes") else []

#         if co_list:
#             for co in co_list:
#                 co_label = getattr(co, "co_code", None) or str(co)
#                 assessment_columns.append({
#                     "assessment_id": a.id,
#                     "assessment_name": a_name,
#                     "co_label": co_label,
#                     "max": a_max,
#                     "label": f"{a_name}\n{co_label}",
#                 })
#         else:
#             fallback_co = getattr(a, "module", None) or "CO"
#             assessment_columns.append({
#                 "assessment_id": a.id,
#                 "assessment_name": a_name,
#                 "co_label": str(fallback_co),
#                 "max": a_max,
#                 "label": f"{a_name}\n{fallback_co}",
#             })

#     assessment_columns = sorted(
#         assessment_columns,
#         key=lambda x: (
#             sibling_ids.index(x["assessment_id"]) if x["assessment_id"] in sibling_ids else 9999,
#             _co_sort_key(x["co_label"]),
#         )
#     )

#     # ------------------------------------------------------------------
#     # IAT marks from StudentInternalMark
#     # ------------------------------------------------------------------
#     exam_name_candidates = _resolve_exam_name_candidates(selected_assessment)

#     sim_base_qs = StudentInternalMark.objects.select_related(
#         "student",
#         "enrollment__course",
#         "co_code",
#         "level_code",
#         "student__department",
#     ).filter(course_code=course_code).order_by("student__reg_no", "created_at")

#     if batch:
#         sim_base_qs = sim_base_qs.filter(batch=batch)

#     if department_id:
#         sim_base_qs = sim_base_qs.filter(student__department_id=department_id)

#     sim_rows = []
#     used_exam_name = ""
#     effective_section = section

#     for exam_name in exam_name_candidates or [""]:
#         tmp_qs = sim_base_qs
#         if exam_name:
#             tmp_qs = tmp_qs.filter(exam_name=exam_name)

#         section_filtered_qs = tmp_qs.filter(section=section) if section else tmp_qs
#         rows = list(section_filtered_qs)

#         tmp_effective_section = section
#         if not rows and section:
#             available_sections = [
#                 (sec or "").strip()
#                 for sec in tmp_qs.values_list("section", flat=True).distinct()
#                 if (sec or "").strip()
#             ]
#             if len(available_sections) == 1:
#                 tmp_effective_section = available_sections[0]
#                 rows = list(tmp_qs.filter(section=tmp_effective_section))

#         if rows:
#             sim_rows = rows
#             used_exam_name = exam_name
#             effective_section = tmp_effective_section
#             break

#     if not sim_rows:
#         tmp_qs = sim_base_qs
#         section_filtered_qs = tmp_qs.filter(section=section) if section else tmp_qs
#         rows = list(section_filtered_qs)

#         tmp_effective_section = section
#         if not rows and section:
#             available_sections = [
#                 (sec or "").strip()
#                 for sec in tmp_qs.values_list("section", flat=True).distinct()
#                 if (sec or "").strip()
#             ]
#             if len(available_sections) == 1:
#                 tmp_effective_section = available_sections[0]
#                 rows = list(tmp_qs.filter(section=tmp_effective_section))

#         if rows:
#             sim_rows = rows
#             effective_section = tmp_effective_section

#     sim_by_student = defaultdict(list)
#     for r in sim_rows:
#         sim_by_student[r.student_id].append(r)

#     iat_totals = {}

#     for sid, srows in sim_by_student.items():
#         question_map = defaultdict(lambda: defaultdict(lambda: {"obt": 0, "rows": []}))

#         for m in srows:
#             qkey = (m.part_name or "", m.question_number or "")
#             okey = m.option_letter or ""
#             question_map[qkey][okey]["obt"] += int(m.marks_obtained or 0)
#             question_map[qkey][okey]["rows"].append(m)

#         chosen_rows = []
#         for (_part, _qnum), options in question_map.items():
#             if "" in options and len(options) == 1:
#                 chosen_rows.extend(options[""]["rows"])
#             else:
#                 key = max(options.keys(), key=lambda k: options[k]["obt"])
#                 chosen_rows.extend(options[key]["rows"])

#         co_totals = defaultdict(lambda: {"obt": 0, "max": 0})
#         for r in chosen_rows:
#             co_code = getattr(r.co_code, "co_code", "") if getattr(r, "co_code_id", None) else "—"
#             if not co_code or str(co_code).strip() == "—":
#                 continue
#             co_totals[co_code]["obt"] += int(r.marks_obtained or 0)
#             co_totals[co_code]["max"] += int(r.max_marks or 0)

#         total_obt = sum(v["obt"] for v in co_totals.values())
#         total_max = sum(v["max"] for v in co_totals.values())

#         iat_totals[sid] = {
#             "obt": Decimal(str(total_obt)),
#             "max": Decimal(str(total_max)),
#         }

#     sim_student_ids = list(iat_totals.keys())
#     all_student_ids = sorted(set(student_ids) | set(sim_student_ids))
#     roster_qs = StudentDetails.objects.filter(id__in=all_student_ids).order_by("reg_no", "name")

#     # ------------------------------------------------------------------
#     # PDF layout
#     # ------------------------------------------------------------------
#     fixed_cols = 5
#     total_cols = fixed_cols + len(assessment_columns) + 3
#     use_landscape = total_cols > 9
#     page_size = landscape(A4) if use_landscape else A4

#     buf = BytesIO()
#     doc = SimpleDocTemplate(
#         buf,
#         pagesize=page_size,
#         leftMargin=10 * mm,
#         rightMargin=10 * mm,
#         topMargin=44 * mm,
#         bottomMargin=16 * mm,
#         title="Mark Summary",
#     )

#     styles = getSampleStyleSheet()
#     styles.add(ParagraphStyle(name="MetaLabel", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9))
#     styles.add(ParagraphStyle(name="MetaValue", parent=styles["Normal"], fontName="Helvetica", fontSize=9))
#     styles.add(ParagraphStyle(name="HeadCenter", parent=styles["Normal"], alignment=1, fontName="Helvetica-Bold", fontSize=8, leading=10))
#     styles.add(ParagraphStyle(name="BodyLeft", parent=styles["Normal"], alignment=0, fontSize=8))
#     styles.add(ParagraphStyle(name="BodyRight", parent=styles["Normal"], alignment=2, fontSize=8))

#     story = []

#     row_data = []

# # Internal Exam
#     row_data.append(
#         Paragraph(f"<b>Internal Exam:</b> {_safe_text(used_exam_name or (exam_name_candidates[0] if exam_name_candidates else ''))}", styles["MetaValue"])
#     )

#     # Semester
#     if semester:
#         row_data.append(
#             Paragraph(f"<b>Semester:</b> {_safe_text(semester)}", styles["MetaValue"])
#         )

#     # Batch
#     if batch:
#         row_data.append(
#             Paragraph(f"<b>Batch:</b> {_safe_text(batch)}", styles["MetaValue"])
#         )

#     # Section
#     if effective_section:
#         row_data.append(
#             Paragraph(f"<b>Section:</b> {_safe_text(effective_section)}", styles["MetaValue"])
#         )

#     # Create single row table
#     details_table = Table([row_data], hAlign="CENTER")

#     details_table.setStyle(TableStyle([
#         ("ALIGN", (0, 0), (-1, -1), "CENTER"),
#         ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
#         ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
#         ("TOPPADDING", (0, 0), (-1, -1), 4),
#     ]))

#     story.append(details_table)
#     story.append(Spacer(1, 6))

#     header = [
#         Paragraph("#", styles["HeadCenter"]),
#         Paragraph("Reg No", styles["HeadCenter"]),
#         Paragraph("Student", styles["HeadCenter"]),
#         Paragraph("IAT<br/>Obtained", styles["HeadCenter"]),
#         Paragraph("Convert<br/>60", styles["HeadCenter"]),
#     ]

#     for col in assessment_columns:
#         header.append(Paragraph(col["label"].replace("\n", "<br/>"), styles["HeadCenter"]))

#     header.extend([
#         Paragraph("Assess<br/>Total", styles["HeadCenter"]),
#         Paragraph("Convert<br/>40", styles["HeadCenter"]),
#         Paragraph("Grand<br/>Total", styles["HeadCenter"]),
#     ])

#     data = [header]

#     total_iat_raw = Decimal("0")
#     total_iat_60 = Decimal("0")
#     total_assess_raw = Decimal("0")
#     total_assess_40 = Decimal("0")
#     total_grand = Decimal("0")
#     student_count = 0

#     count_90_and_above = 0
#     count_80_to_89 = 0
#     count_70_to_79 = 0
#     count_60_to_69 = 0
#     count_50_to_59 = 0
#     count_below_50 = 0

#     for idx, student in enumerate(roster_qs, start=1):
#         student_count += 1

#         student_name = (
#             getattr(student, "full_name", None)
#             or getattr(student, "name", None)
#             or str(student)
#         )
#         reg_no = getattr(student, "reg_no", "") or ""

#         iat_raw = iat_totals.get(student.id, {}).get("obt", Decimal("0"))
#         iat_conv60 = _convert_iat_to_60(iat_raw)

#         assess_raw_total = Decimal("0")
#         assess_max_total = Decimal("0")

#         row = [
#             str(idx),
#             reg_no,
#             student_name,
#             _fmt_marks(iat_raw),
#             _fmt_marks(iat_conv60),
#         ]

#         for col in assessment_columns:
#             base_key = (student.id, col["assessment_id"])
#             base_obt = assess_mark_map[base_key]["obt"]
#             base_max = assess_mark_map[base_key]["max"] or col["max"]

#             # same assessment mark repeated for each CO column
#             row.append(_fmt_marks(base_obt))
#             assess_raw_total += base_obt
#             assess_max_total += base_max

#         assess_conv40 = _convert_to_40(assess_raw_total, assess_max_total)
#         grand_total = (iat_conv60 + assess_conv40).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

#         grand_num = Decimal(str(grand_total))
#         if grand_num >= 90:
#             count_90_and_above += 1
#         elif grand_num >= 80:
#             count_80_to_89 += 1
#         elif grand_num >= 70:
#             count_70_to_79 += 1
#         elif grand_num >= 60:
#             count_60_to_69 += 1
#         elif grand_num >= 50:
#             count_50_to_59 += 1
#         else:
#             count_below_50 += 1

#         row.extend([
#             _fmt_marks(assess_raw_total),
#             _fmt_marks(assess_conv40),
#             _fmt_marks(grand_total),
#         ])
#         data.append(row)

#         total_iat_raw += iat_raw
#         total_iat_60 += iat_conv60
#         total_assess_raw += assess_raw_total
#         total_assess_40 += assess_conv40
#         total_grand += grand_total

#     if student_count == 0:
#         raise Http404("No students found for the given selection.")

#     page_w = page_size[0] - doc.leftMargin - doc.rightMargin
#     col_widths = [10 * mm, 22 * mm, 42 * mm, 18 * mm, 18 * mm]
#     col_widths += [18 * mm] * len(assessment_columns)
#     col_widths += [20 * mm, 18 * mm, 18 * mm]

#     total_width = sum(col_widths)
#     if total_width > page_w:
#         scale = page_w / float(total_width)
#         col_widths = [w * scale for w in col_widths]
#         body_font = 7
#         head_font = 7
#         top_pad = 2
#         bottom_pad = 2
#     else:
#         body_font = 8
#         head_font = 8
#         top_pad = 3
#         bottom_pad = 3

#     table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
#     table.setStyle(TableStyle([
#         ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
#         ("FONTSIZE", (0, 0), (-1, 0), head_font),
#         ("FONTSIZE", (0, 1), (-1, -1), body_font),
#         ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
#         ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
#         ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
#         ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9f9f9")]),
#         ("ALIGN", (0, 0), (1, -1), "CENTER"),
#         ("ALIGN", (2, 1), (2, -1), "LEFT"),
#         ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
#         ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
#         ("TOPPADDING", (0, 0), (-1, -1), top_pad),
#         ("BOTTOMPADDING", (0, 0), (-1, -1), bottom_pad),
#         ("LEFTPADDING", (0, 0), (-1, -1), 3),
#         ("RIGHTPADDING", (0, 0), (-1, -1), 3),
#     ]))
#     story.append(table)
#     story.append(Spacer(1, 10))

#     overall_percentage = Decimal("0")
#     if student_count > 0:
#         overall_percentage = (total_grand / Decimal(student_count)).quantize(
#             Decimal("0.01"), rounding=ROUND_HALF_UP
#         )

#     summary_rows = [
#         ["Summary", "Value"],
#         ["Overall Percentage", f"{_fmt_marks(overall_percentage)}%"],
#         [">= 90", str(count_90_and_above)],
#         [">= 80 and < 90", str(count_80_to_89)],
#         [">= 70 and < 80", str(count_70_to_79)],
#         [">= 60 and < 70", str(count_60_to_69)],
#         [">= 50 and < 60", str(count_50_to_59)],
#         ["< 50 (Fail)", str(count_below_50)],
#     ]

#     summary_table = Table(summary_rows, colWidths=[55 * mm, 35 * mm], hAlign="RIGHT")
#     summary_table.setStyle(TableStyle([
#         ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
#         ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
#         ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
#         ("ALIGN", (0, 0), (0, -1), "LEFT"),
#         ("ALIGN", (1, 1), (1, -1), "RIGHT"),
#         ("TOPPADDING", (0, 0), (-1, -1), 3),
#         ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
#         ("LEFTPADDING", (0, 0), (-1, -1), 4),
#         ("RIGHTPADDING", (0, 0), (-1, -1), 4),
#     ]))
#     story.append(summary_table)

#     # ------------------------------------------------------------------
#     # Header / footer
#     # ------------------------------------------------------------------
#     def _on_page(c, _doc):
#         c.saveState()
#         pw, ph = page_size
#         left = doc.leftMargin
#         right = doc.rightMargin

#         logo_rel = "images/ritlogo.png"
#         logo_path = finders.find(logo_rel)

#         if not logo_path:
#             static_root = getattr(settings, "STATIC_ROOT", "")
#             if static_root:
#                 cand = os.path.join(static_root, logo_rel)
#                 if os.path.exists(cand):
#                     logo_path = cand

#         if not logo_path:
#             for d in getattr(settings, "STATICFILES_DIRS", []):
#                 cand = os.path.join(d, logo_rel)
#                 if os.path.exists(cand):
#                     logo_path = cand
#                     break

#         if logo_path and os.path.exists(logo_path):
#             try:
#                 img = ImageReader(logo_path)
#                 iw, ih = img.getSize()
#                 target_h = 18 * mm
#                 target_w = target_h * (iw / float(ih))
#                 c.drawImage(
#                     img,
#                     left,
#                     ph - (target_h + 8 * mm),
#                     width=target_w,
#                     height=target_h,
#                     preserveAspectRatio=True,
#                     mask="auto",
#                 )
#             except Exception:
#                 pass

#         c.setFillColor(colors.black)
#         c.setFont("Helvetica-Bold", 16)
#         c.drawCentredString(pw / 2.0, ph - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

#         c.setFont("Helvetica", 10)
#         c.drawCentredString(pw / 2.0, ph - 15 * mm, "Rajapalayam - 626117")
#         c.drawCentredString(pw / 2.0, ph - 19 * mm, "An Autonomous Institution")

#         c.setFont("Helvetica-Bold", 13)
#         c.drawCentredString(pw / 2.0, ph - 26 * mm, "Internal / Assessment Mark Summary")

#         dept_line = f"{department_code} - {department_name}".strip(" -")
#         course_line = f"{course_code} - {course_title}".strip(" -")

#         c.setFont("Helvetica", 10)
#         c.drawCentredString(pw / 2.0, ph - 31 * mm, dept_line)
#         c.drawCentredString(pw / 2.0, ph - 35 * mm, course_line)

#         rule_y = ph - 40 * mm
#         c.setStrokeColor(colors.black)
#         c.setLineWidth(0.5)
#         c.line(left, rule_y, pw - right, rule_y)

#         c.setFont("Helvetica", 8)
#         c.setFillColor(colors.grey)
#         c.drawRightString(pw - right, 10 * mm, f"Page {c.getPageNumber()}")

#         c.restoreState()

#     doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

#     pdf = buf.getvalue()
#     buf.close()

#     filename = f"MarkSummary_{assignment.id}_{selected_assessment.id}.pdf"
#     response = HttpResponse(pdf, content_type="application/pdf")
#     response["Content-Disposition"] = f'inline; filename="{filename}"'
#     return response


from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import os

from django.http import HttpResponse, Http404
from django.conf import settings
from django.contrib.staticfiles import finders

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.utils import ImageReader

from course_management.models import AssignSubjectFaculty
from faculty_management.models import Assessment_master
from user_accounts.models import StudentDetails

from faculty_management.models import AssessmentMark


def _get_assessment_name_and_max(a: Assessment_master):
    name = None
    max_marks = None

    for cand in [
        getattr(a, "customAssessmentname", None),
        getattr(a, "Assessmentname", None),
        getattr(getattr(a, "assessment", None), "assessment_name", None),
        getattr(a, "assessment_name", None),
    ]:
        if cand:
            name = cand
            break

    for cand in [
        getattr(a, "Maxmarks", None),
        getattr(a, "max_marks", None),
        getattr(getattr(a, "assessment", None), "max_marks", None),
        getattr(a, "maximum_marks", None),
    ]:
        if cand is not None:
            max_marks = cand
            break

    return name or "(unnamed)", max_marks


def marks_pdf(request):
    # ------------------------------------------------------------
    # GET params
    # ------------------------------------------------------------
    asf_id = (request.GET.get("asf") or "").strip()
    assessment_id = (request.GET.get("assessment") or "").strip()

    if not (asf_id and assessment_id):
        raise Http404("Missing required parameters.")

    # ------------------------------------------------------------
    # Fetch objects
    # ------------------------------------------------------------
    try:
        assignment = (
            AssignSubjectFaculty.objects
            .select_related("course", "department")
            .get(id=asf_id)
        )

        assessment = (
            Assessment_master.objects
            .select_related("assessment", "course", "department", "regulation", "internal_assessment")
            .prefetch_related("co_codes")
            .get(id=assessment_id)
        )
    except (AssignSubjectFaculty.DoesNotExist, Assessment_master.DoesNotExist):
        raise Http404("Selection not found.")

    # ------------------------------------------------------------
    # Fetch marks
    # ------------------------------------------------------------
    marks_qs = (
        AssessmentMark.objects
        .select_related("student")
        .filter(assignment=assignment, assessment=assessment)
        .order_by("student__reg_no", "student__name")
    )

    student_ids = list(marks_qs.values_list("student_id", flat=True))

    roster_qs = (
        StudentDetails.objects
        .filter(id__in=student_ids)
        .order_by("reg_no", "name")
    )

    mark_by_student = {m.student_id: m for m in marks_qs}

    assessment_name, max_marks_value = _get_assessment_name_and_max(assessment)

    # ------------------------------------------------------------
    # Header context
    # ------------------------------------------------------------
    course = assignment.course
    department = assignment.department or getattr(course, "department", None)

    department_code = (
        getattr(department, "Department_code", None)
        or getattr(department, "dept_code", None)
        or ""
    )
    department_name = (
        getattr(department, "Department", None)
        or getattr(department, "name", None)
        or ""
    )

    course_code = getattr(course, "course_code", "") or ""
    course_title = (
        getattr(course, "title", None)
        or getattr(course, "course_title", None)
        or str(course)
    )

    semester = getattr(course, "semester", "") or getattr(assessment, "semester", "")
    section = getattr(assignment, "section", "") or getattr(assessment, "section", "")
    batch = getattr(assignment, "batch", "") or getattr(assessment, "batch", "")
    weightage = getattr(assessment, "weightage", "") or ""
    module = getattr(assessment, "module", "") or ""

    # ------------------------------------------------------------
    # CO details
    # ------------------------------------------------------------
    co_qs = assessment.co_codes.all()
    co_labels = []
    for co in co_qs:
        label = (
            getattr(co, "co_code", None)
            or getattr(co, "code", None)
            or getattr(co, "name", None)
            or str(co)
        )
        if label:
            co_labels.append(str(label))

    co_count = len(co_labels)
    co_text = ", ".join(co_labels) if co_labels else "-"

    # Same max mark for every CO
    per_co_max = None
    if max_marks_value is not None and co_count > 0:
        try:
            per_co_max = Decimal(str(max_marks_value)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
        except Exception:
            per_co_max = None

    # ------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------
    def _pick_marks_value(mark_obj):
        if not mark_obj:
            return None
        v = getattr(mark_obj, "marks_raw", None)
        if v is None:
            v = getattr(mark_obj, "marks", None)
        if v is None:
            v = getattr(mark_obj, "marks_weighted", None)
        return v

    def _pick_weighted_value(mark_obj):
        if not mark_obj:
            return None
        return getattr(mark_obj, "marks_weighted", None)

    def _fmt_marks(val):
        if val is None or val == "":
            return "-"
        try:
            d = Decimal(str(val))
            if d == d.to_integral_value():
                return str(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
            return str(d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
        except Exception:
            return "-"

    def _fmt_text(val):
        return str(val).strip() if val not in (None, "") else "-"

    # ------------------------------------------------------------
    # Build PDF
    # ------------------------------------------------------------
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=44 * mm,
        bottomMargin=18 * mm,
        title="Mark Statement",
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Totals", parent=styles["Heading3"], spaceBefore=6))
    styles.add(
        ParagraphStyle(
            name="MetaLabel",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10,
        )
    )
    styles.add(
        ParagraphStyle(
            name="MetaValue",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
        )
    )

    story = []

    # ------------------------------------------------------------
    # Details section
    # ------------------------------------------------------------
    details_data = [
        [
            Paragraph("Assessment", styles["MetaLabel"]),
            Paragraph(_fmt_text(assessment_name), styles["MetaValue"]),
        ],
        [
            Paragraph("Course", styles["MetaLabel"]),
            Paragraph(_fmt_text(f"{course_code} - {course_title}".strip(" -")), styles["MetaValue"]),
        ],
    ]

    if semester:
        details_data.append(
            [
                Paragraph("Semester", styles["MetaLabel"]),
                Paragraph(_fmt_text(semester), styles["MetaValue"]),
            ]
        )

    if batch and section:
        details_data.append(
            [
                Paragraph("Batch - Section", styles["MetaLabel"]),
                Paragraph(f"{batch} - {section}", styles["MetaValue"]),
            ]
        )
    elif batch:
        details_data.append(
            [
                Paragraph("Batch", styles["MetaLabel"]),
                Paragraph(_fmt_text(batch), styles["MetaValue"]),
            ]
        )
    elif section:
        details_data.append(
            [
                Paragraph("Section", styles["MetaLabel"]),
                Paragraph(_fmt_text(section), styles["MetaValue"]),
            ]
        )

    if module:
        details_data.append(
            [
                Paragraph("Module", styles["MetaLabel"]),
                Paragraph(_fmt_text(module), styles["MetaValue"]),
            ]
        )

    details_data.append(
        [
            Paragraph("CO(s)", styles["MetaLabel"]),
            Paragraph(f" {co_text}" if co_count else "-", styles["MetaValue"]),
        ]
    )

    if max_marks_value is not None:
        details_data.append(
            [
                Paragraph("Max Marks", styles["MetaLabel"]),
                Paragraph(_fmt_text(max_marks_value), styles["MetaValue"]),
            ]
        )

    if weightage:
        details_data.append(
            [
                Paragraph("Weightage", styles["MetaLabel"]),
                Paragraph(_fmt_text(weightage), styles["MetaValue"]),
            ]
        )

    details_table = Table(details_data, colWidths=[32 * mm, None], hAlign="LEFT")
    details_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(details_table)
    story.append(Spacer(1, 6))

    # ------------------------------------------------------------
    # Main marks table
    # ------------------------------------------------------------
    data = [["#", "Reg. No", "Student", "Raw Marks", "Weighted", "Remarks"]]

    total_raw = Decimal("0")
    total_weighted = Decimal("0")
    count_with_marks = 0

    if roster_qs.exists():
        for idx, student in enumerate(roster_qs, 1):
            m = mark_by_student.get(student.id)

            raw_val = _pick_marks_value(m)
            weighted_val = _pick_weighted_value(m)

            try:
                if raw_val is not None and str(raw_val) != "":
                    total_raw += Decimal(str(raw_val))
                    count_with_marks += 1
            except Exception:
                pass

            try:
                if weighted_val is not None and str(weighted_val) != "":
                    total_weighted += Decimal(str(weighted_val))
            except Exception:
                pass

            student_name = (
                getattr(student, "full_name", None)
                or getattr(student, "name", None)
                or str(student)
            )

            data.append([
                str(idx),
                getattr(student, "reg_no", "") or "",
                student_name,
                _fmt_marks(raw_val),
                _fmt_marks(weighted_val),
                getattr(m, "remarks", "") if m else "",
            ])
    else:
        data.append(["-", "-", "No students found", "-", "-", "-"])

    col_widths = [12 * mm, 28 * mm, 58 * mm, 22 * mm, 22 * mm, 38 * mm]
    main_table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    main_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (1, -1), "LEFT"),
                ("ALIGN", (3, 1), (4, -1), "RIGHT"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9f9f9")]),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(main_table)
    story.append(Spacer(1, 8))

    # ------------------------------------------------------------
    # Totals section
    # ------------------------------------------------------------
    if roster_qs.exists():
        avg_weighted = Decimal("0")
        if count_with_marks > 0:
            avg_weighted = (total_weighted / Decimal(str(count_with_marks))).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )

      

        summary_rows = [
            ["Summary", "Value"],
            ["Assessment", _fmt_text(assessment_name)],
            ["Max Marks", _fmt_text(max_marks_value)],
            ["Weightage", _fmt_text(weightage)],
        ]

        for co_label in co_labels:
            summary_rows.append([f"{co_label} Max Mark", _fmt_marks(per_co_max)])

        summary_rows.append(["Average Weighted", _fmt_marks(avg_weighted)])

        summary_table = Table(summary_rows, colWidths=[45 * mm, 35 * mm], hAlign="RIGHT")
        summary_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ]
            )
        )
        story.append(Spacer(1, 6))
        story.append(summary_table)

    # ------------------------------------------------------------
    # Header design
    # ------------------------------------------------------------
    def _on_page(c, _doc):
        c.saveState()
        page_w, page_h = A4
        left = 18 * mm
        right = 18 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)

        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if not logo_path:
            for d in getattr(settings, "STATICFILES_DIRS", []):
                cand = os.path.join(d, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand
                    break

        if logo_path and os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                target_h = 18 * mm
                target_w = target_h * (iw / float(ih))
                c.drawImage(
                    img,
                    left,
                    page_h - (target_h + 8 * mm),
                    width=target_w,
                    height=target_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass

        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(page_w / 2.0, page_h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        c.setFont("Helvetica", 10)
        c.drawCentredString(page_w / 2.0, page_h - 15 * mm, "Rajapalayam - 626117")
        c.drawCentredString(page_w / 2.0, page_h - 19 * mm, "Affiliated to Anna University, Chennai")

        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(page_w / 2.0, page_h - 26 * mm, "Mark Statement")

        dept_line = f"{department_code} - {department_name}".strip(" -")
        course_line = f"{course_code} - {course_title}".strip(" -")

        c.setFont("Helvetica", 10)
        c.drawCentredString(page_w / 2.0, page_h - 31 * mm, dept_line)
        c.drawCentredString(page_w / 2.0, page_h - 35 * mm, course_line)

        rule_y = page_h - 40 * mm
        c.setStrokeColor(colors.black)
        c.setLineWidth(0.5)
        c.line(left, rule_y, page_w - right, rule_y)

        c.setFont("Helvetica", 8)
        c.setFillColor(colors.grey)
        c.drawRightString(page_w - right, 12 * mm, f"Page {c.getPageNumber()}")

        c.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buf.getvalue()
    buf.close()

    filename = f"MarkStatement_{assignment.id}_{assessment.id}.pdf"
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response






@check_permission("assign_assessment_student")
def assign_assessment_student(request):
    """
    Assign an assessment (Assessment_master) to students under a faculty's assigned subject,
    matching Course semester, batch, department, and section.
    Selected students get an AssessmentMark row created (marks left blank).
    Unselected students have any existing row removed.
    """
    faculty_id = _resolve_faculty_id(request.user)
    faculty = general_information.objects.filter(faculty_id=faculty_id).first()
    # Assignments for this faculty
    assignments = (
        AssignSubjectFaculty.objects
        .filter(
            Q(is_active=True) if hasattr(AssignSubjectFaculty, "is_active") else Q(),
            Q(faculty=faculty) | Q(skilled_faculty=faculty)
        )
        .select_related("course", "course__department")
        .order_by("-id")
    )
    # print("Assignments for faculty_id", faculty_id, ":", assignments)
    sel_asf_id = request.GET.get("asf") or request.POST.get("asf")
    sel_assessment_id = request.GET.get("assessment") or request.POST.get("assessment")

    selected_asf = None
    selected_assessment = None
    assessments_qs = Assessment_master.objects.none()
    students = []
    assigned_map = {}

    if sel_asf_id:
        selected_asf = get_object_or_404(
            AssignSubjectFaculty,
            Q(faculty=faculty) | Q(skilled_faculty=faculty),
            pk=sel_asf_id
        )

        # Available assessments for this course/faculty
        assessments_qs = (
            Assessment_master.objects
            .select_related("assessment", "course")
            .filter(faculty_id=faculty_id, course_id=selected_asf.course_id)
            .order_by("id")
        )

        # Build student filter: match by department, batch, section, and course semester
        sf = Q()
        course = selected_asf.course
        if course and course.department_id:
            sf &= Q(department_id=course.department_id)
        if selected_asf.batch:
            sf &= Q(batch=selected_asf.batch)
        if selected_asf.section:
            sf &= Q(section=selected_asf.section)
        if course and course.semester:
            sf &= Q(semester=str(course.semester))

        students = list(StudentDetails.objects.filter(sf).order_by("id"))

    if sel_assessment_id and selected_asf:
        selected_assessment = get_object_or_404(
            Assessment_master,
            pk=sel_assessment_id,
            faculty_id=faculty_id,
            course_id=selected_asf.course_id,
        )

        # Build assigned map for display
        if students:
            existing = AssessmentMark.objects.filter(
                assignment=selected_asf,
                assessment=selected_assessment,
                student__in=students,
            )
            assigned_map = {m.student_id: True for m in existing}

    # Handle POST to assign/unassign
    if request.method == "POST" and selected_asf and selected_assessment:
        selected_ids = request.POST.getlist("students")  # list of student IDs as strings
        selected_ids_set = set(int(i) for i in selected_ids if i.isdigit())
        student_ids = [s.id for s in students]

        with transaction.atomic():
            # Create rows for newly selected
            to_create = [sid for sid in student_ids if sid in selected_ids_set and sid not in assigned_map]
            objs = [AssessmentMark(assignment=selected_asf, assessment=selected_assessment, student_id=sid) for sid in to_create]
            if objs:
                AssessmentMark.objects.bulk_create(objs, ignore_conflicts=True)

            # Remove rows for deselected
            to_delete = [sid for sid in student_ids if sid not in selected_ids_set and sid in assigned_map]
            if to_delete:
                AssessmentMark.objects.filter(
                    assignment=selected_asf,
                    assessment=selected_assessment,
                    student_id__in=to_delete,
                ).delete()

        messages.success(request, "Assignment updated successfully.")
        return redirect(f"{request.path}?asf={selected_asf.id}&assessment={selected_assessment.id}")

    context = {
        "faculty_id": faculty_id,
        "assignments": assignments,
        "selected_asf": selected_asf,
        "assessments": assessments_qs,
        "selected_assessment": selected_assessment,
        "students": students,
        "assigned_map": assigned_map,
    }
    return render(request, "faculty_management/exams/assign_assessment_student.html", context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

from faculty_management.models import Open_Elective_Offer, Open_Elective_OfferToDept


def get_offered_to_list(request):
    offer_id = request.GET.get("offer")
    ids = Open_Elective_OfferToDept.objects.filter(
        offer_id=offer_id
    ).values_list("offered_to_dept_id", flat=True)
    return JsonResponse(list(ids), safe=False)

def get_courses_by_reg_sem(request):
    regulation = request.GET.get("regulation")
    semester = request.GET.get("semester")

    courses = Course.objects.filter(
        regulation_id=regulation,
        semester=semester
    ).values("id", "course_code", "title")

    return JsonResponse(list(courses), safe=False)


@check_permission("open_elective_offer")
def open_elective_offer(request):
    faculty = general_information.objects.get(faculty_id=request.user.Employee_id)
    try:
        department = faculty.department
    except Add_Department.DoesNotExist:
        messages.error(request, "Your department is not found or inactive.")
        return redirect(request.path)

    faculty = general_information.objects.filter(
        faculty_id=request.user.Employee_id,
        department=department
    ).first()

    if faculty is None:
        messages.error(request, "Faculty record not mapped to user.")
        return redirect("open_elective_offer")

    # ✅ Dynamic Semesters from Degree Duration
    total_semesters = department.degree.duration * 2
    semester_list = list(range(1, total_semesters + 1))

    if request.method == "POST":
        action = request.POST.get("action")
        offer_id = request.POST.get("offer_id")

        try:
            with transaction.atomic():
                if action == "save":
                    # Only allow creating/updating inside the user's department
                    offer, created = Open_Elective_Offer.objects.update_or_create(
                        id=offer_id if offer_id else None,
                        defaults={
                            "course_id": request.POST.get("course"),
                            "regulation_id": request.POST.get("regulation"),
                            "faculty_id": request.POST.get("faculty"),
                            # Force to user's department (do NOT trust POST)
                            "department_id": faculty.department_id,
                            "offered_from_dept_id": faculty.department_id,
                            "slots": request.POST.get("slots"),
                            "academic_year": request.POST.get("academic_year"),
                            "batch": request.POST.get("batch"),
                            "created_by": faculty,
                            "updated_by": faculty,
                        },
                    )

                    # Refresh Offered-To mapping
                    Open_Elective_OfferToDept.objects.filter(offer=offer).delete()
                    for d in request.POST.getlist("offered_to_dept"):
                        Open_Elective_OfferToDept.objects.create(
                            offer=offer,
                            offered_to_dept_id=d
                        )

                    messages.success(request, "Offer saved successfully!")

                elif action == "delete":
                    # Can delete only offers from user's department
                    offer = get_object_or_404(
                        Open_Elective_Offer,
                        pk=offer_id,
                        department_id=faculty.department_id
                    )
                    offer.delete()
                    messages.success(request, "Offer deleted successfully!")

        except Exception as e:
            messages.error(request, f"Error: {e}")

        return redirect("open_elective_offer")

    # ----- READ (only show this department's offers) -----
    batches = (
        StudentDetails.objects.values_list("batch", flat=True)
        .distinct().order_by("-batch")
    )

    offers_qs = (
        Open_Elective_Offer.objects.filter(department_id=faculty.department_id)
        .prefetch_related("to_departments")
        .select_related("course", "faculty", "regulation", "offered_from_dept")
        .order_by("-created_at")
    )

    # ----- PDF with Professional Design -----
    if request.GET.get("download") == "pdf":
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="open_elective_offers.pdf"'

        doc = BaseDocTemplate(response, pagesize=A4)
        frame = Frame(18 * mm, 25 * mm, A4[0] - 36 * mm, A4[1] - 60 * mm, id='normal')

        def _header(canvas, _doc):
            canvas.saveState()
            page_w, page_h = A4
            left = 18 * mm
            right = 18 * mm
            
            # Professional Logo Styling
            try:
                logo_rel = 'images/ritlogo.png'
                logo_path = finders.find(logo_rel)
                if logo_path and os.path.exists(logo_path):
                    img = ImageReader(logo_path)
                    iw, ih = img.getSize()
                    target_h = 16 * mm
                    target_w = target_h * (iw / float(ih))
                    canvas.drawImage(img, left, page_h - (target_h + 8 * mm),
                                     width=target_w, height=target_h, 
                                     preserveAspectRatio=True, mask='auto')
            except Exception:
                pass

            # Header Text with Professional Styling
            canvas.setFillColor(colors.HexColor("#2C3E50"))
            canvas.setFont('Helvetica-Bold', 16)
            canvas.drawCentredString(page_w / 2.0, page_h - 14 * mm, 'RAMCO INSTITUTE OF TECHNOLOGY')
            
            # canvas.setFont('Helvetica-Bold', 11)
            # canvas.setFillColor(colors.HexColor("#E74C3C"))
            # canvas.drawCentredString(page_w / 2.0, page_h - 21 * mm, 'An Autonomous Institution')
            
            canvas.setFont('Helvetica', 10)
            canvas.setFillColor(colors.HexColor("#2C3E50"))
            canvas.drawCentredString(page_w / 2.0, page_h - 27 * mm, 'RAJAPALAYAM - 626117 | Tamil Nadu')
            
            # Report Title
            canvas.setFont('Helvetica-Bold', 12)
            canvas.setFillColor(colors.HexColor("#3498DB"))
            canvas.drawCentredString(page_w / 2.0, page_h - 34 * mm, 'Open Elective Offers Report')

            # Footer with Decorative Line
            canvas.setStrokeColor(colors.HexColor("#3498DB"))
            canvas.setLineWidth(0.5)
            canvas.line(18 * mm, 15 * mm, page_w - 18 * mm, 15 * mm)
            
            canvas.setFont('Helvetica', 9)
            canvas.setFillColor(colors.HexColor("#7F8C8D"))
            canvas.drawRightString(page_w - right, 10 * mm, f"Page {canvas.getPageNumber()}")
            
            # Footer Text
            from datetime import datetime as _dt
            footer_text = f"Generated on: {_dt.now().strftime('%d-%m-%Y at %H:%M')}"
            canvas.drawString(left, 10 * mm, footer_text)

            canvas.restoreState()

        doc.addPageTemplates([PageTemplate(id='offers', frames=[frame], onPage=_header)])

        styles = getSampleStyleSheet()
        
        # Professional Style Definitions
        header_style = ParagraphStyle(
            'header_style', 
            parent=styles['Normal'], 
            alignment=1,
            fontName='Helvetica-Bold', 
            fontSize=9, 
            textColor=colors.white,
            leading=10
        )
        
        wrap_style = ParagraphStyle(
            'wrap_style', 
            parent=styles['Normal'], 
            fontSize=8,
            leading=10,
            textColor=colors.HexColor("#2C3E50")
        )
        
        title_style = ParagraphStyle(
            'title_style',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor("#2C3E50"),
            alignment=1,
            spaceAfter=12,
            fontName='Helvetica-Bold'
        )
        
        summary_style = ParagraphStyle(
            'summary_style',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor("#27AE60"),
            alignment=1,
            spaceAfter=10
        )
        
        meta_style = ParagraphStyle(
            'meta_style',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor("#34495E"),
            alignment=1,
            spaceAfter=8
        )

        elements = [Spacer(1, 35)]

        # Report Title
        elements.append(Paragraph("OPEN ELECTIVE OFFERS REPORT", title_style))
        
        # Decorative Line
        elements.append(Paragraph("<hr width='70%' color='#3498DB'/>", styles['Normal']))
        elements.append(Spacer(1, 8))

        # Department Information
        elements.append(Paragraph(
            f"<b>Offering Department:</b> {department.Department}",
            ParagraphStyle(
                'dept_style',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor("#34495E"),
                alignment=1,
                spaceAfter=5
            )
        ))

        # Meta Information
        offered_from_req = (request.GET.get('offered_from') or department.Department or '').strip()
        ay_from_req = (request.GET.get('academic_year') or '').strip()
        if not ay_from_req:
            first_ay = next((o.academic_year for o in offers_qs if getattr(o, 'academic_year', None)), '')
            ay_from_req = first_ay or ''

        meta_lines = []
        if offered_from_req:
            meta_lines.append(f"Offered From: {offered_from_req}")
        if ay_from_req:
            meta_lines.append(f"Academic Year: {ay_from_req}")

        if meta_lines:
            elements.append(Paragraph(" • ".join(meta_lines), meta_style))
        
        # Summary Statistics
        total_offers = len(offers_qs)
        elements.append(Paragraph(
            f"<b>📊 SUMMARY:</b> Total {total_offers} open elective offer(s) available",
            summary_style
        ))
        elements.append(Spacer(1, 15))

        # Table Data with Professional Headers - Added Offered From Department
        data = [[
            Paragraph("SL<br/>No", header_style),
            Paragraph("Course<br/>Details", header_style),
            Paragraph("Offered From<br/>Department", header_style),
            Paragraph("Assigned<br/>Faculty", header_style),
            Paragraph("Offered To<br/>Departments", header_style),
            Paragraph("Available<br/>Slots", header_style),
            Paragraph("Batch /<br/>Semester", header_style),
        ]]

        # Populate table rows
        for idx, offer in enumerate(offers_qs, start=1):
            # Course details with title
            course_line = (getattr(offer.course, 'title', '') or '')
            
            # Offered From Department
            offered_from_dept = getattr(offer.offered_from_dept, 'Department', '') or getattr(offer.department, 'Department', '')
            
            # Faculty name
            fac = getattr(offer.faculty, 'name', '') or 'Not Assigned'
            
            # Offered to departments (truncate if too long)
            offered_to_list = [
                getattr(od.offered_to_dept, 'Department', '') or ''
                for od in offer.to_departments.all()
            ]
            offered_to_names = '<br/>'.join(offered_to_list[:3]) if offered_to_list else ''
            if len(offered_to_list) > 3:
                offered_to_names += f"<br/>+{len(offered_to_list) - 3} more"
            
            # Slots and batch/semester
            slots = offer.slots or ''
            batch = offer.batch or ''
            sem = getattr(getattr(offer, 'course', None), 'semester', '') or ''
            batch_sem = f"{batch}" if not sem else f"{batch}<br/>Sem {sem}"

            data.append([
                Paragraph(str(idx), wrap_style),
                Paragraph(course_line, wrap_style),
                Paragraph(offered_from_dept, wrap_style),
                Paragraph(fac, wrap_style),
                Paragraph(offered_to_names, wrap_style),
                Paragraph(str(slots), wrap_style),
                Paragraph(batch_sem, wrap_style),
            ])

        # Adjusted column widths to accommodate the new column
        colWidths = [12 * mm, 45 * mm, 30 * mm, 28 * mm, 35 * mm, 18 * mm, 22 * mm]

        table = Table(data, colWidths=colWidths, repeatRows=1, hAlign='LEFT')
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            
            # Grid and borders
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),  # Center align slots column
            
            # Font styling for content
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),  # Bold SL No
            ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),  # Bold Offered From Department
            
            # Padding
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            
            # Alternate row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ]))

        elements.append(table)
        
        # Empty state message
        if not offers_qs:
            elements.append(Spacer(1, 30))
            elements.append(Paragraph(
                "No open elective offers found for your department.",
                ParagraphStyle(
                    'empty_style',
                    parent=styles['Heading3'],
                    fontSize=12,
                    textColor=colors.HexColor("#95A5A6"),
                    alignment=1
                )
            ))

        elements.append(Spacer(1, 10))
        doc.build(elements)
        return response

    return render(request, "faculty_management/exams/open_elective_offer.html", {
        "offers": offers_qs,
        "regulations": Regulations.objects.all(),
        "faculties": general_information.objects.filter(department_id=faculty.department_id),
        "departments": Add_Department.objects.all(),
        "batches": batches,
        "semesters": semester_list,
        "offer_from_department" : faculty.department
    })


@check_permission("open_elective_assignments")
def open_elective_assignments_view(request):

    # 🔹 Get Logged-in Faculty
    try:
        faculty = general_information.objects.get(
            faculty_id=request.user.Employee_id
        )
    except general_information.DoesNotExist:
        messages.error(request, "Faculty record not mapped to your login.")
        return redirect("open_elective_offer")

    department = faculty.department

    if not department:
        messages.error(request, "Your department is not mapped.")
        return redirect("open_elective_offer")

    # 🔥 IMPORTANT: Filter by logged-in department
    offers_qs = Open_Elective_Offer.objects.select_related(
        "course",
        "faculty",
        "regulation",
        "offered_from_dept"
    ).prefetch_related(
        "to_departments"
    ).filter(
        offered_from_dept=department  # ✅ Only this department's assignments
    ).order_by("-created_at")

    # ---------------- FILTERS ----------------
    academic_year = request.GET.get("academic_year")
    semester = request.GET.get("semester")
    search = request.GET.get("search")

    if academic_year:
        offers_qs = offers_qs.filter(academic_year=academic_year)

    if semester:
        offers_qs = offers_qs.filter(course__semester=semester)

    if search:
        offers_qs = offers_qs.filter(
            Q(course__title__icontains=search) |
            Q(faculty__name__icontains=search) |
            Q(course__course_code__icontains=search)
        )

    # ---------------- PDF DOWNLOAD ----------------
    if request.GET.get("download") == "pdf":

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = \
            'inline; filename="open_elective_assignments.pdf"'

        doc = BaseDocTemplate(response, pagesize=A4)
        frame = Frame(
            18 * mm, 25 * mm,
            A4[0] - 36 * mm,
            A4[1] - 60 * mm,
            id='normal'
        )

        def header(canvas, _doc):
            canvas.saveState()
            page_w, page_h = A4

            canvas.setFont('Helvetica-Bold', 14)
            canvas.drawCentredString(
                page_w / 2,
                page_h - 20 * mm,
                'OPEN ELECTIVE ASSIGNMENTS REPORT'
            )

            canvas.setFont('Helvetica', 9)
            canvas.drawRightString(
                page_w - 20 * mm,
                10 * mm,
                f"Page {canvas.getPageNumber()}"
            )

            canvas.restoreState()

        doc.addPageTemplates([
            PageTemplate(id='assignments',
                         frames=[frame],
                         onPage=header)
        ])

        styles = getSampleStyleSheet()

        elements = []
        elements.append(Spacer(1, 30))

        elements.append(
            Paragraph(
                f"<b>Department:</b> {department.Department}",
                styles["Normal"]
            )
        )

        elements.append(
            Paragraph(
                f"<b>Total Offers:</b> {offers_qs.count()}",
                styles["Normal"]
            )
        )

        elements.append(Spacer(1, 15))

        # -------- TABLE DATA --------
        data = [[
            "Sl No",
            "Course",
            "Faculty",
            "Slots",
            "Batch",
            "Semester"
        ]]

        for idx, offer in enumerate(offers_qs, start=1):
            data.append([
                str(idx),
                f"{offer.course.course_code} - {offer.course.title}",
                offer.faculty.name if offer.faculty else "Not Assigned",
                str(offer.slots),
                offer.batch,
                str(offer.course.semester),
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))

        elements.append(table)

        doc.build(elements)
        return response

    # -------- FILTER DROPDOWN DATA --------
    academic_years = offers_qs.values_list(
        "academic_year",
        flat=True
    ).distinct().order_by("-academic_year")

    semesters = offers_qs.values_list(
        "course__semester",
        flat=True
    ).distinct().order_by("course__semester")

    return render(
        request,
        "faculty_management/exams/open_elective_assignments_view.html",
        {
            "offers": offers_qs,
            "academic_years": academic_years,
            "semesters": semesters,
            "selected_ay": academic_year,
            "selected_sem": semester,
            "search_query": search,
        }
    )

from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required

from user_accounts.models import Add_Department, Degree
from course_management.models import Regulations, Course
from faculty_management.models import general_information, AssignSubjectFaculty
from examination_management.models import InternalAssessment


def _get_faculty(request):
    user = getattr(request, "user", None)
    emp_id = getattr(user, "Employee_id", None) or getattr(user, "employee_id", None) if user else None
    if not emp_id:
        return None
    return general_information.objects.filter(faculty_id=emp_id).first()

def _degree_or_400(deg_id):
    if not deg_id:
        return None, JsonResponse({"ok": False, "error": "Degree ID missing."}, status=400)
    try:
        degree = Degree.objects.get(pk=int(deg_id))
    except (Degree.DoesNotExist, ValueError, TypeError):
        return None, JsonResponse({"ok": False, "error": "Invalid degree."}, status=400)
    return degree, None

def _dept_or_400(dept_id):
    if not dept_id:
        return None, JsonResponse({"ok": False, "error": "Department ID missing."}, status=400)
    try:
        dept = Add_Department.objects.get(pk=int(dept_id))
    except (Add_Department.DoesNotExist, ValueError, TypeError):
        return None, JsonResponse({"ok": False, "error": "Invalid department."}, status=400)
    return dept, None



def iats_for_degree(request):
    deg_id = request.POST.get("degree_id")
    degree, err = _degree_or_400(deg_id)
    if err:
        return err

    iats = InternalAssessment.objects.filter(degree=degree).order_by("iat")
    return JsonResponse({
        "ok": True,
        "iats": [{"id": ia.id, "label": ia.iat or ""} for ia in iats],
    })

def assessments_for_degree_iat(request):
    deg_id = request.POST.get("degree_id")
    iat_id = request.POST.get("iat_id")

    degree, err = _degree_or_400(deg_id)
    if err:
        return err

    if not iat_id:
        return JsonResponse({"ok": False, "error": "IAT ID missing."}, status=400)
    try:
        iat = InternalAssessment.objects.get(pk=int(iat_id), degree=degree)
    except (InternalAssessment.DoesNotExist, ValueError, TypeError):
        return JsonResponse({"ok": False, "error": "Invalid IAT for this degree."}, status=400)

    qs = (Assessments.objects
          .filter(degree=degree, internal_assessment=iat)
          .order_by("assessment_name", "id"))

    assessments = [{"id": a.id, "label": a.assessment_name or ""} for a in qs]

    # Prepend the selected IAT as a regular option (id="iat:<id>")
    assessments.insert(0, {"id": f"iat:{iat.id}", "label": iat.iat or ""})

    return JsonResponse({"ok": True, "assessments": assessments})






def regulations_for_department(request):
    gi_obj = _get_faculty(request)
    if gi_obj is None:
        return JsonResponse({"ok": False, "error": "Faculty not resolved."}, status=403)

    dept_id = request.POST.get("department_id")
    department, err = _dept_or_400(dept_id)
    if err:
        return err

    regs = (
        Regulations.objects
        .filter(assignsubjectfaculty__faculty=gi_obj,
                assignsubjectfaculty__department=department,
                assignsubjectfaculty__is_active=True)
        .distinct()
        .order_by("id")
    )

    return JsonResponse({
        "ok": True,
        "regulations": [{"id": r.id, "label": str(r)} for r in regs],
    })



def courses_for_selection(request):
    """
    Return courses assigned to the logged-in faculty, filtered by:
    - department (required)
    - regulation (required)
    - degree (optional but recommended)
    - semester (optional; if provided, matches Course.semester)

    Uses reverse relation name 'course_assign' from AssignSubjectFaculty.course.
    """
    gi_obj = _get_faculty(request)
    if gi_obj is None:
        return JsonResponse({"ok": False, "error": "Faculty not resolved."}, status=403)

    dept_id = request.POST.get("department_id")
    reg_id = request.POST.get("regulation_id")
    degree_id = request.POST.get("degree_id")      # new
    semester = request.POST.get("semester")        # new

    department, err = _dept_or_400(dept_id)
    if err:
        return err

    if not reg_id:
        return JsonResponse({"ok": False, "error": "Regulation ID missing."}, status=400)
    try:
        regulation = Regulations.objects.get(pk=int(reg_id))
    except (Regulations.DoesNotExist, ValueError, TypeError):
        return JsonResponse({"ok": False, "error": "Invalid regulation."}, status=400)

    # Optional: degree filter (and sanity check that dept belongs to degree, if sent)
    degree = None
    if degree_id:
        try:
            degree = Degree.objects.get(pk=int(degree_id))
        except (Degree.DoesNotExist, ValueError, TypeError):
            return JsonResponse({"ok": False, "error": "Invalid degree."}, status=400)
        # If a degree was supplied but mismatches the department, fail fast
        if department.degree_id != degree.id:
            return JsonResponse({"ok": False, "error": "Department does not belong to selected degree."}, status=400)

    qs = (
        Course.objects.filter(
            is_active=True,
            department=department,
            regulation=regulation,
            # traverse AssignSubjectFaculty via reverse name 'course_assign'
            course_assign__faculty=gi_obj,
            course_assign__department=department,
            course_assign__regulation=regulation,
            course_assign__is_active=True,
        )
        .distinct()
    )

    if degree is not None:
        qs = qs.filter(department__degree=degree)

    # Course.semester is a CharField; compare as string if provided
    if semester:
        qs = qs.filter(semester=str(semester).strip())

    qs = qs.order_by("course_code", "title", "id")

    return JsonResponse({
        "ok": True,
        "courses": [{"id": c.id, "label": str(c)} for c in qs],
    })
    



# ============================================================================
# SEMINAR HALL BOOKING VIEWS
# ============================================================================

@no_cache
@check_permission("shb_hub")
def shb_hub(request):
    """Seminar Hall Booking Hub - Main dashboard"""
    from faculty_management.models import SeminarHallBooking, SHBApplicationApproval
    
    # Get user's role
    user_role_id = request.user.role.id if hasattr(request.user, 'role') and request.user.role else None
    
    # Count pending approvals for this user
    pending_count = 0
    if user_role_id:
        pending_count = SHBApplicationApproval.objects.filter(
            approval_step__approver_role_id=user_role_id,
            status='pending'
        ).count()
    
    # Count total applications
    total_count = SeminarHallBooking.objects.count()
    
    context = {
        'pending_count': pending_count,
        'total_count': total_count,
    }
    
    return render(request, 'faculty_management/shb_hub.html', context)


@no_cache
@check_permission("seminar_hall_booking")
def seminar_hall_booking(request):
    """Seminar Hall Booking Form"""
    from faculty_management.models import SeminarHallBooking, SeminarHall, SHBApprovalWorkflow, SHBApprovalStep, SHBApplicationApproval
    from django.utils import timezone
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Set current page for sidebar
    request.session['current_page'] = 'seminar_hall_booking'
    
    # Get faculty details
    faculty = general_information.objects.get(faculty_id=request.user.Employee_id)
    
    if request.method == 'POST':
        try:
            # Create booking
            booking = SeminarHallBooking()
            booking.faculty = faculty
            booking.faculty_name = faculty.name
            booking.faculty_email = request.user.email
            booking.faculty_phone = request.POST.get('faculty_phone')
            booking.department = faculty.department
            
            # Event details
            booking.event_name = request.POST.get('event_name')
            booking.event_type = request.POST.get('event_type')
            booking.event_description = request.POST.get('event_description')
            
            # Guest speaker
            booking.has_guest_speaker = request.POST.get('has_guest_speaker') == 'yes'
            if booking.has_guest_speaker:
                booking.guest_name = request.POST.get('guest_name')
                booking.guest_designation = request.POST.get('guest_designation')
                booking.guest_organization = request.POST.get('guest_organization')
                booking.guest_email = request.POST.get('guest_email')
                booking.guest_phone = request.POST.get('guest_phone')
            
            # Schedule
            from datetime import datetime as dt
            start_datetime = dt.strptime(f"{request.POST.get('booking_date')} {request.POST.get('start_time')}", '%Y-%m-%d %H:%M')
            end_datetime = dt.strptime(f"{request.POST.get('booking_date')} {request.POST.get('end_time')}", '%Y-%m-%d %H:%M')
            
            booking.booking_date = start_datetime.date()
            booking.start_time = start_datetime.time()
            booking.end_time = end_datetime.time()
            
            # Hall
            booking.preferred_hall = request.POST.get('preferred_hall')
            booking.expected_attendees = request.POST.get('expected_attendees', 0) or 0
            
            # Requirements
            booking.special_requirements = request.POST.get('special_requirements')
            booking.needs_projector = request.POST.get('needs_projector') == 'on'
            booking.needs_microphone = request.POST.get('needs_microphone') == 'on'
            booking.needs_sound_system = request.POST.get('needs_sound_system') == 'on'
            booking.needs_video_conferencing = request.POST.get('needs_video_conferencing') == 'on'
            
            booking.save()
            
            # Create approval workflow
            creator_role_id = request.user.role.id if hasattr(request.user, 'role') and request.user.role else None
            
            if creator_role_id:
                workflow = SHBApprovalWorkflow.objects.filter(creator_role_id=creator_role_id, is_active=True).first()
                
                if workflow:
                    steps = SHBApprovalStep.objects.filter(workflow=workflow, is_active=True).order_by('approval_level')
                    
                    for step in steps:
                        approval = SHBApplicationApproval()
                        approval.application = booking
                        approval.approval_step = step
                        approval.status = 'pending'
                        
                        # Set department for filtering
                        if not step.is_cross_department:
                            approval.approver_department_id = booking.department_id
                        
                        approval.save()
            
            messages.success(request, f'Booking submitted successfully! Booking ID: {booking.booking_id}')
            return redirect('seminar_hall_booking')
            
        except Exception as e:
            logger.error(f"Error creating booking: {e}")
            messages.error(request, f'Error submitting booking: {str(e)}')
    
    # Get previous bookings
    previous_bookings = SeminarHallBooking.objects.filter(faculty=faculty).order_by('-created_at')[:5]
    
    # Get active halls
    seminar_halls = SeminarHall.objects.filter(is_active=True).order_by('hall_number')
    
    from datetime import datetime as dt
    context = {
        'faculty': faculty,
        'previous_bookings': previous_bookings,
        'seminar_halls': seminar_halls,
        'today': dt.now().date()
    }
    
    return render(request, 'faculty_management/seminar_hall_booking.html', context)


@no_cache
@check_permission("shb_my_approvals")
def shb_my_approvals(request):
    """View for approving seminar hall bookings"""
    from faculty_management.models import SHBApplicationApproval
    from django.utils import timezone
    
    # Set current page
    request.session['current_page'] = 'shb_my_approvals'
    
    # Get user's role and department
    user_role_id = request.user.role.id if hasattr(request.user, 'role') and request.user.role else None
    user_dept_id = request.user.Department.id if hasattr(request.user, 'Department') and request.user.Department else None
    
    if request.method == 'POST':
        action = request.POST.get('action')
        approval_id = request.POST.get('approval_id')
        comments = request.POST.get('comments', '')
        
        try:
            approval = SHBApplicationApproval.objects.get(id=approval_id, status='pending')
            
            if action == 'approve':
                approval.status = 'approved'
                approval.approver_id = request.user.id
                approval.comments = comments
                approval.approved_at = timezone.now()
                approval.save()
                
                # Check if all approvals are complete
                remaining = SHBApplicationApproval.objects.filter(
                    application=approval.application,
                    status='pending'
                ).exists()
                
                if not remaining:
                    approval.application.status = 'approved'
                    approval.application.approved_by_id = request.user.id
                    approval.application.approval_date = timezone.now()
                    approval.application.save()
                    messages.success(request, 'Application fully approved!')
                else:
                    messages.success(request, 'Application approved at your level.')
                    
            elif action == 'reject':
                approval.status = 'rejected'
                approval.approver_id = request.user.id
                approval.comments = comments
                approval.approved_at = timezone.now()
                approval.save()
                
                approval.application.status = 'rejected'
                approval.application.rejection_reason = comments
                approval.application.save()
                
                messages.success(request, 'Application rejected.')
                
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
        
        return redirect('shb_my_approvals')
    
    # Get pending approvals
    approvals = SHBApplicationApproval.objects.filter(
        approval_step__approver_role_id=user_role_id,
        status='pending'
    ).select_related('application', 'application__department', 'approval_step').order_by('-created_at')
    
    # Filter by department if not cross-department
    if user_dept_id:
        approvals = approvals.filter(
            models.Q(approval_step__is_cross_department=True) |
            models.Q(approver_department_id=user_dept_id)
        )
    
    context = {
        'approvals': approvals,
    }
    
    return render(request, 'faculty_management/shb_my_approvals.html', context)


@no_cache
@check_permission("shb_all_applications")
def shb_all_applications(request):
    """View all seminar hall booking applications"""
    from faculty_management.models import SeminarHallBooking
    from django.db.models import Q
    
    # Set current page
    request.session['current_page'] = 'shb_all_applications'
    
    # Get filters
    status_filter = request.GET.get('status', '')
    department_filter = request.GET.get('department', '')
    event_type_filter = request.GET.get('event_type', '')
    search_query = request.GET.get('search', '')
    
    # Base queryset
    applications = SeminarHallBooking.objects.all().select_related('faculty', 'department').order_by('-created_at')
    
    # Apply filters
    if status_filter:
        applications = applications.filter(status=status_filter)
    
    if department_filter:
        applications = applications.filter(department_id=department_filter)
    
    if event_type_filter:
        applications = applications.filter(event_type=event_type_filter)
    
    if search_query:
        applications = applications.filter(
            Q(booking_id__icontains=search_query) |
            Q(event_name__icontains=search_query) |
            Q(faculty_name__icontains=search_query)
        )
    
    # Get counts
    total_count = SeminarHallBooking.objects.count()
    pending_count = SeminarHallBooking.objects.filter(status='pending').count()
    approved_count = SeminarHallBooking.objects.filter(status='approved').count()
    rejected_count = SeminarHallBooking.objects.filter(status='rejected').count()
    
    # Get departments for filter
    from user_accounts.models import Add_Department
    departments = Add_Department.objects.all().order_by('Department')
    
    context = {
        'applications': applications,
        'departments': departments,
        'total_count': total_count,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'status_filter': status_filter,
        'department_filter': department_filter,
        'event_type_filter': event_type_filter,
        'search_query': search_query,
    }
    
    return render(request, 'faculty_management/shb_all_applications.html', context)


from faculty_management.models import Faculty_Data_Permission
from django.views.decorators.http import require_GET
from django.core.paginator import Paginator

def assign_faculty_data_permission(request):
    edit_permission = None

    # ---------- DELETE ----------
    if request.method == "POST" and request.POST.get("action") == "delete":
        perm_id = request.POST.get("perm_id")
        Faculty_Data_Permission.objects.filter(id=perm_id).delete()
        messages.success(request, "Permission deleted")
        return redirect("assign_faculty_data_permission")

    # ---------- EDIT LOAD ----------
    if request.method == "GET" and request.GET.get("edit"):
        edit_permission = get_object_or_404(
            Faculty_Data_Permission, id=request.GET.get("edit")
        )

    # ---------- CREATE / UPDATE ----------
    if request.method == "POST" and request.POST.get("action") == "save":
        role_ids = request.POST.getlist("roles[]")
        can_view_all = request.POST.get("can_view_all_faculty_data") == "on"
        can_view_dept = request.POST.get("can_view_department_faculty_data") == "on"
        perm_id = request.POST.get("perm_id")

        if perm_id:
            # ---- EDIT (single row) ----
            Faculty_Data_Permission.objects.filter(id=perm_id).update(
                can_view_all_faculty_data=can_view_all,
                can_view_department_faculty_data=can_view_dept,
            )
            messages.success(request, "Permission updated successfully")
            return redirect("assign_faculty_data_permission")

        # ---- CREATE (bulk roles) ----
        if not role_ids:
            messages.error(request, "At least one role is required")
            return redirect("assign_faculty_data_permission")

        for role_id in role_ids:
            Faculty_Data_Permission.objects.update_or_create(
                role_id=role_id,
                defaults={
                    "can_view_all_faculty_data": can_view_all,
                    "can_view_department_faculty_data": can_view_dept,
                }
            )

        messages.success(request, "Permission saved successfully")
        return redirect("assign_faculty_data_permission")

    # ---------- PAGE LOAD ----------
    roles = Role.objects.using("rit_approval_system").all()

    context = {
        "roles": roles,
        "edit_permission": edit_permission,
    }
    return render(request, "faculty_management/faculty_data_permission.html", context)
 

@require_GET
def faculty_data_permission_api(request):
    search = (request.GET.get("search") or "").strip()
    page = int(request.GET.get("page", 1))

    permissions = Faculty_Data_Permission.objects.all().order_by("id")

    # ---- FETCH ROLES FROM OTHER DB ----
    roles_qs = Role.objects.using("rit_approval_system").all()
    role_map = {r.id: r.role for r in roles_qs}

    # ---- SEARCH BY ROLE NAME (IN OTHER DB) ----
    if search:
        role_ids = list(
            roles_qs.filter(role__icontains=search).values_list("id", flat=True)
        )
        if not role_ids:
            permissions = Faculty_Data_Permission.objects.none()
        else:
            permissions = permissions.filter(role_id__in=role_ids)

    page_size = 25
    paginator = Paginator(permissions, page_size)
    page_obj = paginator.get_page(page)

    data = [
        {
            "id": perm.id,
            "role": role_map.get(perm.role_id, "Unknown"),
            "can_view_all": perm.can_view_all_faculty_data,
            "can_view_dept": perm.can_view_department_faculty_data,
        }
        for perm in page_obj
    ]

    return JsonResponse({
        "results": data,
        "page": page_obj.number,
        "total_pages": paginator.num_pages,
        "has_next": page_obj.has_next(),
        "has_prev": page_obj.has_previous(),
        "total_count": paginator.count,
        "page_size": page_size,
    })


def get_logged_in_student(request):
    """Return the StudentDetails record for the authenticated student."""
    user = getattr(request, "user", None)
    if not user or not getattr(user, "is_authenticated", False):
        return None

    if not getattr(user, "is_student", False):
        return None

    employee_id = getattr(user, "Employee_id", None)
    if not employee_id:
        return None

    return StudentDetails.objects.filter(reg_no=employee_id).first()


def get_logged_in_faculty(request):
    """Return the faculty record for the authenticated non-student user."""
    user = getattr(request, "user", None)
    if not user or not getattr(user, "is_authenticated", False):
        return None

    if getattr(user, "is_student", False) or getattr(user, "is_parent", False):
        return None

    employee_id = getattr(user, "Employee_id", None)
    if not employee_id:
        return None

    try:
        return general_information.objects.filter(faculty_id=employee_id).first()
    except (TypeError, ValueError):
        return None

from faculty_management.models import Ticket

def ticket_raise(request, ticket_id=None):
    
    logged_student = get_logged_in_student(request)
    logged_faculty = get_logged_in_faculty(request)

    logged_user_type = None

    if logged_student:
        logged_user_type = "Student"
    elif logged_faculty:
        logged_user_type = "Faculty"

    update_ticket = None

    if ticket_id:
        update_ticket = get_object_or_404(Ticket, id=ticket_id)

    if request.method == "POST":

        if ticket_id:
            ticket = update_ticket

            status = request.POST.get("status")

            is_ticket_owner = (
                (ticket.student and logged_student and ticket.student == logged_student)
                or
                (ticket.faculty and logged_faculty and ticket.faculty == logged_faculty)
            )

            allowed_status = ["Pending", "Progressing", "Solved"]

            if is_ticket_owner and ticket.status == "Solved":
                allowed_status.append("Accomplished")

            if status in allowed_status:
                ticket.status = status
                ticket.save()
                messages.success(request, "Ticket Status Updated Successfully")
            else:
                messages.error(request, "You are not allowed to update this status.")

            return redirect("ticket_raise")

        issue = request.POST.get("issue")
        location = request.POST.get("location")
        description = request.POST.get("description")

        if not logged_student and not logged_faculty:
            messages.error(request, "Logged-in user details not found.")
            return redirect("ticket_raise")

        ticket = Ticket(
            student=logged_student,
            faculty=logged_faculty if not logged_student else None,
            issue=issue,
            location=location,
            description=description,
            status="Pending"
        )

        try:
            ticket.full_clean()
            ticket.save()
            messages.success(request, "Ticket Raised Successfully")
            return redirect("ticket_raise")

        except Exception as e:
            messages.error(request, e)

    tickets = Ticket.objects.select_related(
        "student",
        "faculty"
    ).order_by("-created_at")

    context = {
        "logged_student": logged_student,
        "logged_faculty": logged_faculty,
        "logged_user_type": logged_user_type,
        "tickets": tickets,
        "update_ticket": update_ticket,
    }

    return render(
        request,
        "faculty_management/power_house/ticket_raise.html",
        context
    )

from faculty_management.models import MaterialRequestApprovers, MaterialRequestApproversData

def material_approval_system(request):
    if request.method == "GET":
        creator_role_id = request.GET.get("creator_role_id")

        # AJAX: load existing hierarchy for selected creator role
        if creator_role_id:
            approvers = MaterialRequestApprovers.objects.filter(
                creator_role_id=creator_role_id
            ).order_by("approver_level")

            hierarchy = []

            for a in approvers:
                hierarchy.append({
                    "id": str(a.approver_role_id),
                    "isCrossDepartment": a.is_cross_department_approver == "YES",
                    "departmentId": str(a.approver_department.id) if a.approver_department else None,
                })

            return JsonResponse({
                "hierarchy": hierarchy
            })

        roles = Role.objects.using(APPROVAL_DB).all()
        departments = Add_Department.objects.all().order_by("Department")

        return render(request, "faculty_management/admin/material_approval_system.html", {
            "roles": roles,
            "departments": departments,
        })

    if request.method == "POST":
        try:
            raw = (request.body or b"").decode("utf-8").strip()

            if not raw:
                return JsonResponse({"error": "Empty body"}, status=400)

            data = json.loads(raw)

            creator_role_id = str(data.get("creatorRole") or "").strip()
            role_hierarchy = data.get("roleHierarchy", [])

            if not creator_role_id:
                return JsonResponse({"error": "Creator role is required."}, status=400)

            if not isinstance(role_hierarchy, list):
                return JsonResponse({"error": "roleHierarchy must be a list."}, status=400)

            MaterialRequestApprovers.objects.filter(
                creator_role_id=creator_role_id
            ).delete()

            for index, role_data in enumerate(role_hierarchy):
                approver_role_id = str(role_data.get("id") or "").strip()
                is_cross_department = role_data.get("isCrossDepartment", False)
                dept_id = role_data.get("departmentId")

                if not approver_role_id:
                    return JsonResponse({"error": "Approver role is required."}, status=400)

                department_obj = None

                if is_cross_department:
                    if not dept_id:
                        return JsonResponse({
                            "error": "Department is required for cross-department approver."
                        }, status=400)

                    department_obj = Add_Department.objects.filter(id=dept_id).first()

                    if not department_obj:
                        return JsonResponse({"error": "Department not found."}, status=404)

                MaterialRequestApprovers.objects.create(
                    creator_role_id=creator_role_id,
                    approver_role_id=approver_role_id,
                    approver_level=index + 1,
                    is_cross_department_approver="YES" if is_cross_department else "NO",
                    approver_department=department_obj,
                )

            return JsonResponse({
                "message": "Material approval hierarchy saved successfully."
            })

        except Exception as e:
            logger.exception("material_approval_system failed")
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request method."}, status=405)




def material_request_approval(request):
    logged_in_faculty = get_logged_in_faculty(request)
    logged_in_role = get_logged_in_role(request)

    if not logged_in_faculty:
        messages.error(request, "Approver faculty record not found.")
        return redirect("dashboard")

    if not logged_in_role:
        messages.error(request, "Approver role not found.")
        return redirect("dashboard")

    if request.method == "POST":
        entry_id = request.POST.get("entry_id")
        action = request.POST.get("action")
        remarks = request.POST.get("remarks") or ""

        if action not in ["approve", "reject"]:
            messages.error(request, "Invalid action.")
            return redirect("material_request_approval")

        try:
            with transaction.atomic():
                entry = get_object_or_404(
                    MaterialRequestApproversData.objects
                    .select_for_update()
                    .select_related(
                        "request",
                        "request__department",
                        "request__ticket",
                        "request__ticket__student",
                        "request__ticket__student__department",
                        "request__ticket__faculty",
                        "request__ticket__faculty__department",
                    ),
                    id=entry_id,
                )

                mat_request = entry.request

                if mat_request.status in ["APPROVED", "REJECTED", "ISSUED"]:
                    messages.error(request, "This request is already closed.")
                    return redirect("material_request_approval")

                if not can_approve_material_entry(entry, logged_in_faculty, logged_in_role):
                    messages.error(request, "You are not authorized to approve this step.")
                    return redirect("material_request_approval")

                entry.approver = logged_in_faculty
                entry.acted_on = timezone.now()

                if action == "approve":
                    entry.status = MaterialRequestApproversData.Status.APPROVED
                    entry.reason = remarks or f"Approved by {logged_in_faculty.name}"
                    entry.save()

                    has_pending = mat_request.approval_entries.filter(
                        status=MaterialRequestApproversData.Status.PENDING
                    ).exists()

                    mat_request.status = "PENDING" if has_pending else "APPROVED"
                    mat_request.save(update_fields=["status"])

                    messages.success(request, f"{mat_request.request_no} approved successfully.")

                else:
                    if not remarks.strip():
                        messages.error(request, "Rejection reason is required.")
                        return redirect("material_request_approval")

                    entry.status = MaterialRequestApproversData.Status.REJECTED
                    entry.reason = remarks
                    entry.save()

                    mat_request.approval_entries.filter(
                        status=MaterialRequestApproversData.Status.PENDING
                    ).exclude(id=entry.id).update(
                        status=MaterialRequestApproversData.Status.REJECTED,
                        reason=f"Auto rejected after level {entry.approver_level}",
                        acted_on=timezone.now(),
                    )

                    mat_request.status = "REJECTED"
                    mat_request.save(update_fields=["status"])

                    messages.success(request, f"{mat_request.request_no} rejected successfully.")

        except Exception as e:
            messages.error(request, f"Approval failed: {str(e)}")

        return redirect("material_request_approval")

    pending_entries = (
        MaterialRequestApproversData.objects
        .select_related(
            "request",
            "request__department",
            "request__ticket",
            "request__ticket__student",
            "request__ticket__student__department",
            "request__ticket__faculty",
            "request__ticket__faculty__department",
        )
        .prefetch_related("request__request_items__item")
        .filter(status=MaterialRequestApproversData.Status.PENDING)
        .order_by("request_id", "approver_level", "id")
    )

    actionable_entries = [
        entry for entry in pending_entries
        if can_approve_material_entry(entry, logged_in_faculty, logged_in_role)
    ]

    my_actions = (
        MaterialRequestApproversData.objects
        .select_related(
            "request",
            "request__department",
            "request__ticket",
            "request__ticket__student",
            "request__ticket__faculty",
        )
        .filter(approver=logged_in_faculty)
        .exclude(status=MaterialRequestApproversData.Status.PENDING)
        .order_by("-acted_on", "-id")[:20]
    )

    return render(request, "faculty_management/power_house/material_request_approval.html", {
        "actionable_entries": actionable_entries,
        "my_actions": my_actions,
    })


    def material_issue(request, request_id=None):
        
        mat_request = None
        approved_requests = None

        if request_id is None:
            approved_requests = (
                MaterialRequest.objects
                .select_related(
                    "department",
                    "ticket",
                    "ticket__student",
                    "ticket__student__department",
                    "ticket__faculty",
                    "ticket__faculty__department",
                )
                .filter(status="APPROVED")
                .order_by("-id")
            )

            return render(request, "faculty_management/power_house/material_issue.html", {
                "approved_requests": approved_requests,
                "mat_request": mat_request,
            })

        mat_request = get_object_or_404(
            MaterialRequest.objects
            .select_related(
                "department",
                "ticket",
                "ticket__student",
                "ticket__student__department",
                "ticket__faculty",
                "ticket__faculty__department",
            )
            .prefetch_related("request_items__item"),
            id=request_id
        )

        if mat_request.status != "APPROVED":
            messages.error(request, "Only approved requests can be issued.")
            return redirect("material_issue")

        if request.method == "POST":
            try:
                with transaction.atomic():
                    issued_any = False

                    for req_item in mat_request.request_items.select_related("item"):
                        issue_qty = float(request.POST.get(f"issue_qty_{req_item.id}") or 0)
                        item = req_item.item

                        if issue_qty <= 0:
                            continue

                        if issue_qty > float(item.stock_qty):
                            messages.error(request, f"Insufficient stock for {item.item_name}.")
                            return redirect("material_issue_detail", request_id=request_id)

                        item.stock_qty = float(item.stock_qty) - issue_qty
                        item.save(update_fields=["stock_qty"])

                        req_item.issued_qty = issue_qty
                        req_item.save(update_fields=["issued_qty"])

                        StockLedger.objects.create(
                            item=item,
                            transaction_type="ISSUE",
                            qty_in=0,
                            qty_out=issue_qty,
                            balance_qty=item.stock_qty,
                            reference_no=mat_request.request_no,
                            handled_by_employee_id=getattr(request.user, "Employee_id", None),
                            handled_by_name=get_user_display_name(request),
                        )

                        issued_any = True

                    if not issued_any:
                        messages.error(request, "Please enter at least one issue quantity.")
                        return redirect("material_issue_detail", request_id=request_id)

                    mat_request.status = "ISSUED"
                    mat_request.save(update_fields=["status"])

                    if mat_request.ticket:
                        mat_request.ticket.status = "Solved"
                        mat_request.ticket.save(update_fields=["status"])

                messages.success(request, "Material issued successfully. Stock reduced.")
                return redirect("material_issue")

            except Exception as e:
                messages.error(request, f"Issue failed: {str(e)}")

        return render(request, "faculty_management/power_house/material_issue.html", {
            "mat_request": mat_request,
            "approved_requests": approved_requests,
        })
    
 
def material_issue(request, request_id=None):
    
    mat_request = None
    approved_requests = None

    if request_id is None:
        approved_requests = (
            MaterialRequest.objects
            .select_related(
                "department",
                "ticket",
                "ticket__student",
                "ticket__student__department",
                "ticket__faculty",
                "ticket__faculty__department",
            )
            .filter(status="APPROVED")
            .order_by("-id")
        )

        return render(request, "faculty_management/power_house/material_issue.html", {
            "approved_requests": approved_requests,
            "mat_request": mat_request,
        })

    mat_request = get_object_or_404(
        MaterialRequest.objects
        .select_related(
            "department",
            "ticket",
            "ticket__student",
            "ticket__student__department",
            "ticket__faculty",
            "ticket__faculty__department",
        )
        .prefetch_related("request_items__item"),
        id=request_id
    )

    if mat_request.status != "APPROVED":
        messages.error(request, "Only approved requests can be issued.")
        return redirect("material_issue")

    if request.method == "POST":
        try:
            with transaction.atomic():
                issued_any = False

                for req_item in mat_request.request_items.select_related("item"):
                    issue_qty = float(request.POST.get(f"issue_qty_{req_item.id}") or 0)
                    item = req_item.item

                    if issue_qty <= 0:
                        continue

                    if issue_qty > float(item.stock_qty):
                        messages.error(request, f"Insufficient stock for {item.item_name}.")
                        return redirect("material_issue_detail", request_id=request_id)

                    req_item.issued_qty = issue_qty
                    req_item.save(update_fields=["issued_qty"])

                    issued_any = True

                if not issued_any:
                    messages.error(request, "Please enter at least one issue quantity.")
                    return redirect("material_issue_detail", request_id=request_id)

                mat_request.status = "ISSUED"
                mat_request.save(update_fields=["status"])

            messages.success(
                request,
                "Material issue recorded. Stock will be reduced only after proof approval."
            )
            return redirect("material_issue")

        except Exception as e:
            messages.error(request, f"Issue failed: {str(e)}")

    return render(request, "faculty_management/power_house/material_issue.html", {
        "mat_request": mat_request,
        "approved_requests": approved_requests,
    })



def inventory_category_entry(request):
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "add":
            names = request.POST.getlist("name[]")
            count = 0

            for name in names:
                name = name.strip()

                if name:
                    InventoryCategory.objects.create(
                        name=name,
                        description="",
                        active=True,
                    )
                    count += 1

            messages.success(request, f"{count} category(s) added successfully.")

        elif action == "edit":
            category_id = request.POST.get("category_id")
            category = get_object_or_404(InventoryCategory, id=category_id)

            category.name = request.POST.get("name")
            category.save()

            messages.success(request, "Category updated successfully.")

        elif action == "delete":
            category_id = request.POST.get("category_id")
            category = get_object_or_404(InventoryCategory, id=category_id)
            category.delete()

            messages.success(request, "Category deleted successfully.")

        return redirect("inventory_category_entry")

    categories = InventoryCategory.objects.all().order_by("name")

    return render(request, "faculty_management/power_house/inventory_category_entry.html", {
        "categories": categories,
    })



def inventory_item_entry(request):
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "add":
            item_codes = request.POST.getlist("item_code[]")
            item_names = request.POST.getlist("item_name[]")
            item_types = request.POST.getlist("item_type[]")
            categories = request.POST.getlist("category[]")
            units = request.POST.getlist("unit[]")
            stock_qtys = request.POST.getlist("stock_qty[]")
            minimum_qtys = request.POST.getlist("minimum_qty[]")
            unit_rates = request.POST.getlist("unit_rate[]")

            count = 0

            try:
                with transaction.atomic():
                    for i, item_code in enumerate(item_codes):
                        item_code = item_code.strip()

                        if not item_code:
                            continue

                        item_name = item_names[i].strip() if i < len(item_names) else ""
                        unit = units[i].strip() if i < len(units) else ""

                        if not item_name or not unit:
                            continue

                        stock_qty = Decimal(stock_qtys[i] or "0") if i < len(stock_qtys) else Decimal("0")
                        minimum_qty = Decimal(minimum_qtys[i] or "0") if i < len(minimum_qtys) else Decimal("0")
                        unit_rate = Decimal(unit_rates[i] or "0") if i < len(unit_rates) else Decimal("0")

                        item_type_id = item_types[i] if i < len(item_types) and item_types[i] else None
                        category_id = categories[i] if i < len(categories) and categories[i] else None

                        if InventoryItem.objects.filter(item_code=item_code).exists():
                            messages.warning(request, f"Item code {item_code} already exists. Skipped.")
                            continue

                        item = InventoryItem.objects.create(
                            item_code=item_code,
                            item_name=item_name,
                            item_type_id=item_type_id,
                            category_id=category_id,
                            unit=unit,
                            stock_qty=stock_qty,
                            minimum_qty=minimum_qty,
                            unit_rate=unit_rate,
                            active=True,
                        )

                        StockLedger.objects.create(
                            item=item,
                            transaction_type="OPENING",
                            qty_in=stock_qty,
                            qty_out=0,
                            balance_qty=stock_qty,
                            reference_no=f"OPENING-{item.item_code}",
                            handled_by_employee_id=getattr(request.user, "Employee_id", None),
                            handled_by_name=get_user_display_name(request),
                        )

                        count += 1

                messages.success(request, f"{count} inventory item(s) added successfully.")

            except Exception as e:
                messages.error(request, f"Error while adding inventory items: {str(e)}")

        elif action == "edit":
            item_id = request.POST.get("item_id")
            item = get_object_or_404(InventoryItem, id=item_id)

            old_stock = item.stock_qty

            try:
                with transaction.atomic():
                    item.item_code = request.POST.get("item_code", "").strip()
                    item.item_name = request.POST.get("item_name", "").strip()
                    item.item_type_id = request.POST.get("item_type") or None
                    item.category_id = request.POST.get("category") or None
                    item.unit = request.POST.get("unit", "").strip()
                    item.stock_qty = Decimal(request.POST.get("stock_qty") or "0")
                    item.minimum_qty = Decimal(request.POST.get("minimum_qty") or "0")
                    item.unit_rate = Decimal(request.POST.get("unit_rate") or "0")
                    item.active = True if request.POST.get("active") == "on" else False
                    item.save()

                    stock_difference = item.stock_qty - old_stock

                    if stock_difference != 0:
                        StockLedger.objects.create(
                            item=item,
                            transaction_type="ADJUSTMENT",
                            qty_in=stock_difference if stock_difference > 0 else 0,
                            qty_out=abs(stock_difference) if stock_difference < 0 else 0,
                            balance_qty=item.stock_qty,
                            reference_no=f"ADJUST-{item.item_code}",
                            handled_by_employee_id=getattr(request.user, "Employee_id", None),
                            handled_by_name=get_user_display_name(request),
                        )

                messages.success(request, "Inventory item updated successfully.")

            except Exception as e:
                messages.error(request, f"Error while updating inventory item: {str(e)}")

        elif action == "delete":
            item_id = request.POST.get("item_id")
            item = get_object_or_404(InventoryItem, id=item_id)

            try:
                item.delete()
                messages.success(request, "Inventory item deleted successfully.")
            except Exception as e:
                messages.error(request, f"Cannot delete this item: {str(e)}")

        return redirect("inventory_item_entry")

    item_types = InventoryItemType.objects.filter(active=True).order_by("type_name")
    categories = InventoryCategory.objects.filter(active=True).order_by("name")

    items = (
        InventoryItem.objects
        .select_related("item_type", "category")
        .all()
        .order_by("item_name")
    )

    return render(request, "faculty_management/power_house/inventory_item_entry.html", {
        "item_types": item_types,
        "categories": categories,
        "items": items,
    })
 


def material_request_entry(request):
    
    creator_role = get_logged_in_role(request)

    if not creator_role:
        messages.error(request, "User role not found.")
        return redirect("dashboard")

    tickets = (
        Ticket.objects
        .select_related("student", "student__department", "faculty", "faculty__department")
        .exclude(status="Solved")
        .order_by("-created_at")
    )

    inventory_items = (
        InventoryItem.objects
        .select_related("item_type", "category")
        .filter(active=True)
        .order_by("item_name")
    )

    if request.method == "POST":

        ticket_id = request.POST.get("ticket")
        purpose = request.POST.get("purpose")
        location = request.POST.get("location")

        item_ids = request.POST.getlist("item_id[]")
        quantities = request.POST.getlist("quantity[]")

        if not ticket_id:
            messages.error(request, "Please select a ticket.")
            return redirect("material_request_entry")

        if not any(item_ids):
            messages.error(request, "Please select at least one item.")
            return redirect("material_request_entry")

        ticket = Ticket.objects.select_related(
            "student",
            "student__department",
            "faculty",
            "faculty__department"
        ).filter(id=ticket_id).first()

        if not ticket:
            messages.error(request, "Selected ticket not found.")
            return redirect("material_request_entry")

        department = None

        if ticket.student and ticket.student.department:
            department = ticket.student.department

        elif ticket.faculty and ticket.faculty.department:
            department = ticket.faculty.department

        try:
            with transaction.atomic():

                mat_request = MaterialRequest.objects.create(
                    creator_role_id=str(creator_role.id),
                    ticket=ticket,
                    department=department,
                    purpose=purpose,
                    location=location or ticket.location,
                    status="PENDING",
                )

                for item_id, qty in zip(item_ids, quantities):
                    if item_id and qty:
                        MaterialRequestItem.objects.create(
                            request=mat_request,
                            item_id=item_id,
                            requested_qty=qty,
                        )

                created_rows = create_material_approver_chain(
                    mat_request=mat_request,
                    creator_role_id=str(creator_role.id),
                )

                if created_rows == 0:
                    mat_request.status = "APPROVED"
                    mat_request.save(update_fields=["status"])
                    messages.warning(
                        request,
                        "Request auto-approved. No approval hierarchy found."
                    )
                else:
                    messages.success(
                        request,
                        "Material request submitted successfully."
                    )

            return redirect("material_request_entry")

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            return redirect("material_request_entry")

    my_requests = (
        MaterialRequest.objects
        .select_related(
            "ticket",
            "ticket__student",
            "ticket__faculty",
            "department"
        )
        .prefetch_related("request_items__item")
        .all()
        .order_by("-created_at")
    )

    return render(
        request,
        "faculty_management/power_house/material_request_entry.html",
        {
            "tickets": tickets,
            "inventory_items": inventory_items,
            "my_requests": my_requests,
        }
    )



def material_approval_system(request):
    if request.method == "GET":
        creator_role_id = request.GET.get("creator_role_id")

        # AJAX: load existing hierarchy for selected creator role
        if creator_role_id:
            approvers = MaterialRequestApprovers.objects.filter(
                creator_role_id=creator_role_id
            ).order_by("approver_level")

            hierarchy = []

            for a in approvers:
                hierarchy.append({
                    "id": str(a.approver_role_id),
                    "isCrossDepartment": a.is_cross_department_approver == "YES",
                    "departmentId": str(a.approver_department.id) if a.approver_department else None,
                })

            return JsonResponse({
                "hierarchy": hierarchy
            })

        roles = Role.objects.using(APPROVAL_DB).all()
        departments = Add_Department.objects.all().order_by("Department")

        return render(request, "faculty_management/admin/material_approval_system.html", {
            "roles": roles,
            "departments": departments,
        })

    if request.method == "POST":
        try:
            raw = (request.body or b"").decode("utf-8").strip()

            if not raw:
                return JsonResponse({"error": "Empty body"}, status=400)

            data = json.loads(raw)

            creator_role_id = str(data.get("creatorRole") or "").strip()
            role_hierarchy = data.get("roleHierarchy", [])

            if not creator_role_id:
                return JsonResponse({"error": "Creator role is required."}, status=400)

            if not isinstance(role_hierarchy, list):
                return JsonResponse({"error": "roleHierarchy must be a list."}, status=400)

            MaterialRequestApprovers.objects.filter(
                creator_role_id=creator_role_id
            ).delete()

            for index, role_data in enumerate(role_hierarchy):
                approver_role_id = str(role_data.get("id") or "").strip()
                is_cross_department = role_data.get("isCrossDepartment", False)
                dept_id = role_data.get("departmentId")

                if not approver_role_id:
                    return JsonResponse({"error": "Approver role is required."}, status=400)

                department_obj = None

                if is_cross_department:
                    if not dept_id:
                        return JsonResponse({
                            "error": "Department is required for cross-department approver."
                        }, status=400)

                    department_obj = Add_Department.objects.filter(id=dept_id).first()

                    if not department_obj:
                        return JsonResponse({"error": "Department not found."}, status=404)

                MaterialRequestApprovers.objects.create(
                    creator_role_id=creator_role_id,
                    approver_role_id=approver_role_id,
                    approver_level=index + 1,
                    is_cross_department_approver="YES" if is_cross_department else "NO",
                    approver_department=department_obj,
                )

            return JsonResponse({
                "message": "Material approval hierarchy saved successfully."
            })

        except Exception as e:
            logger.exception("material_approval_system failed")
            return JsonResponse({"error": str(e)}, status=500)

    return JsonResponse({"error": "Invalid request method."}, status=405)




def material_request_approval(request):
    logged_in_faculty = get_logged_in_faculty(request)
    logged_in_role = get_logged_in_role(request)

    if not logged_in_faculty:
        messages.error(request, "Approver faculty record not found.")
        return redirect("dashboard")

    if not logged_in_role:
        messages.error(request, "Approver role not found.")
        return redirect("dashboard")

    if request.method == "POST":
        entry_id = request.POST.get("entry_id")
        action = request.POST.get("action")
        remarks = request.POST.get("remarks") or ""

        if action not in ["approve", "reject"]:
            messages.error(request, "Invalid action.")
            return redirect("material_request_approval")

        try:
            with transaction.atomic():
                entry = get_object_or_404(
                    MaterialRequestApproversData.objects
                    .select_for_update()
                    .select_related(
                        "request",
                        "request__department",
                        "request__ticket",
                        "request__ticket__student",
                        "request__ticket__student__department",
                        "request__ticket__faculty",
                        "request__ticket__faculty__department",
                    ),
                    id=entry_id,
                )

                mat_request = entry.request

                if mat_request.status in ["APPROVED", "REJECTED", "ISSUED"]:
                    messages.error(request, "This request is already closed.")
                    return redirect("material_request_approval")

                if not can_approve_material_entry(entry, logged_in_faculty, logged_in_role):
                    messages.error(request, "You are not authorized to approve this step.")
                    return redirect("material_request_approval")

                entry.approver = logged_in_faculty
                entry.acted_on = timezone.now()

                if action == "approve":
                    entry.status = MaterialRequestApproversData.Status.APPROVED
                    entry.reason = remarks or f"Approved by {logged_in_faculty.name}"
                    entry.save()

                    has_pending = mat_request.approval_entries.filter(
                        status=MaterialRequestApproversData.Status.PENDING
                    ).exists()

                    mat_request.status = "PENDING" if has_pending else "APPROVED"
                    mat_request.save(update_fields=["status"])

                    messages.success(request, f"{mat_request.request_no} approved successfully.")

                else:
                    if not remarks.strip():
                        messages.error(request, "Rejection reason is required.")
                        return redirect("material_request_approval")

                    entry.status = MaterialRequestApproversData.Status.REJECTED
                    entry.reason = remarks
                    entry.save()

                    mat_request.approval_entries.filter(
                        status=MaterialRequestApproversData.Status.PENDING
                    ).exclude(id=entry.id).update(
                        status=MaterialRequestApproversData.Status.REJECTED,
                        reason=f"Auto rejected after level {entry.approver_level}",
                        acted_on=timezone.now(),
                    )

                    mat_request.status = "REJECTED"
                    mat_request.save(update_fields=["status"])

                    messages.success(request, f"{mat_request.request_no} rejected successfully.")

        except Exception as e:
            messages.error(request, f"Approval failed: {str(e)}")

        return redirect("material_request_approval")

    pending_entries = (
        MaterialRequestApproversData.objects
        .select_related(
            "request",
            "request__department",
            "request__ticket",
            "request__ticket__student",
            "request__ticket__student__department",
            "request__ticket__faculty",
            "request__ticket__faculty__department",
        )
        .prefetch_related("request__request_items__item")
        .filter(status=MaterialRequestApproversData.Status.PENDING)
        .order_by("request_id", "approver_level", "id")
    )

    actionable_entries = [
        entry for entry in pending_entries
        if can_approve_material_entry(entry, logged_in_faculty, logged_in_role)
    ]

    my_actions = (
        MaterialRequestApproversData.objects
        .select_related(
            "request",
            "request__department",
            "request__ticket",
            "request__ticket__student",
            "request__ticket__faculty",
        )
        .filter(approver=logged_in_faculty)
        .exclude(status=MaterialRequestApproversData.Status.PENDING)
        .order_by("-acted_on", "-id")[:20]
    )

    return render(request, "faculty_management/power_house/material_request_approval.html", {
        "actionable_entries": actionable_entries,
        "my_actions": my_actions,
    })


    def material_issue(request, request_id=None):
        
        mat_request = None
        approved_requests = None

        if request_id is None:
            approved_requests = (
                MaterialRequest.objects
                .select_related(
                    "department",
                    "ticket",
                    "ticket__student",
                    "ticket__student__department",
                    "ticket__faculty",
                    "ticket__faculty__department",
                )
                .filter(status="APPROVED")
                .order_by("-id")
            )

            return render(request, "faculty_management/power_house/material_issue.html", {
                "approved_requests": approved_requests,
                "mat_request": mat_request,
            })

        mat_request = get_object_or_404(
            MaterialRequest.objects
            .select_related(
                "department",
                "ticket",
                "ticket__student",
                "ticket__student__department",
                "ticket__faculty",
                "ticket__faculty__department",
            )
            .prefetch_related("request_items__item"),
            id=request_id
        )

        if mat_request.status != "APPROVED":
            messages.error(request, "Only approved requests can be issued.")
            return redirect("material_issue")

        if request.method == "POST":
            try:
                with transaction.atomic():
                    issued_any = False

                    for req_item in mat_request.request_items.select_related("item"):
                        issue_qty = float(request.POST.get(f"issue_qty_{req_item.id}") or 0)
                        item = req_item.item

                        if issue_qty <= 0:
                            continue

                        if issue_qty > float(item.stock_qty):
                            messages.error(request, f"Insufficient stock for {item.item_name}.")
                            return redirect("material_issue_detail", request_id=request_id)

                        item.stock_qty = float(item.stock_qty) - issue_qty
                        item.save(update_fields=["stock_qty"])

                        req_item.issued_qty = issue_qty
                        req_item.save(update_fields=["issued_qty"])

                        StockLedger.objects.create(
                            item=item,
                            transaction_type="ISSUE",
                            qty_in=0,
                            qty_out=issue_qty,
                            balance_qty=item.stock_qty,
                            reference_no=mat_request.request_no,
                            handled_by_employee_id=getattr(request.user, "Employee_id", None),
                            handled_by_name=get_user_display_name(request),
                        )

                        issued_any = True

                    if not issued_any:
                        messages.error(request, "Please enter at least one issue quantity.")
                        return redirect("material_issue_detail", request_id=request_id)

                    mat_request.status = "ISSUED"
                    mat_request.save(update_fields=["status"])

                    if mat_request.ticket:
                        mat_request.ticket.status = "Solved"
                        mat_request.ticket.save(update_fields=["status"])

                messages.success(request, "Material issued successfully. Stock reduced.")
                return redirect("material_issue")

            except Exception as e:
                messages.error(request, f"Issue failed: {str(e)}")

        return render(request, "faculty_management/power_house/material_issue.html", {
            "mat_request": mat_request,
            "approved_requests": approved_requests,
        })
    
 
def material_issue(request, request_id=None):
    
    mat_request = None
    approved_requests = None

    if request_id is None:
        approved_requests = (
            MaterialRequest.objects
            .select_related(
                "department",
                "ticket",
                "ticket__student",
                "ticket__student__department",
                "ticket__faculty",
                "ticket__faculty__department",
            )
            .filter(status="APPROVED")
            .order_by("-id")
        )

        return render(request, "faculty_management/power_house/material_issue.html", {
            "approved_requests": approved_requests,
            "mat_request": mat_request,
        })

    mat_request = get_object_or_404(
        MaterialRequest.objects
        .select_related(
            "department",
            "ticket",
            "ticket__student",
            "ticket__student__department",
            "ticket__faculty",
            "ticket__faculty__department",
        )
        .prefetch_related("request_items__item"),
        id=request_id
    )

    if mat_request.status != "APPROVED":
        messages.error(request, "Only approved requests can be issued.")
        return redirect("material_issue")

    if request.method == "POST":
        try:
            with transaction.atomic():
                issued_any = False

                for req_item in mat_request.request_items.select_related("item"):
                    issue_qty = float(request.POST.get(f"issue_qty_{req_item.id}") or 0)
                    item = req_item.item

                    if issue_qty <= 0:
                        continue

                    if issue_qty > float(item.stock_qty):
                        messages.error(request, f"Insufficient stock for {item.item_name}.")
                        return redirect("material_issue_detail", request_id=request_id)

                    req_item.issued_qty = issue_qty
                    req_item.save(update_fields=["issued_qty"])

                    issued_any = True

                if not issued_any:
                    messages.error(request, "Please enter at least one issue quantity.")
                    return redirect("material_issue_detail", request_id=request_id)

                mat_request.status = "ISSUED"
                mat_request.save(update_fields=["status"])

            messages.success(
                request,
                "Material issue recorded. Stock will be reduced only after proof approval."
            )
            return redirect("material_issue")

        except Exception as e:
            messages.error(request, f"Issue failed: {str(e)}")

    return render(request, "faculty_management/power_house/material_issue.html", {
        "mat_request": mat_request,
        "approved_requests": approved_requests,
    })


def update_stock_after_proof_verification(proof, request):
    mat_request = proof.material_request

    if proof.stock_updated:
        return

    for req_item in mat_request.request_items.select_related("item"):
        item = req_item.item
        issue_qty = float(req_item.issued_qty or 0)

        if issue_qty <= 0:
            continue

        if issue_qty > float(item.stock_qty):
            raise ValueError(f"Insufficient stock for {item.item_name}.")

        item.stock_qty = float(item.stock_qty) - issue_qty
        item.save(update_fields=["stock_qty"])

        StockLedger.objects.create(
            item=item,
            transaction_type="ISSUE",
            qty_in=0,
            qty_out=issue_qty,
            balance_qty=item.stock_qty,
            reference_no=mat_request.request_no,
            handled_by_employee_id=getattr(request.user, "Employee_id", None),
            handled_by_name=get_user_display_name(request),
        )

    proof.stock_updated = True
    proof.save(update_fields=["stock_updated"])

    if mat_request.ticket:
        mat_request.ticket.status = "Solved"
        mat_request.ticket.save(update_fields=["status"])

def low_stock_dashboard(request):
    low_stock_items = (
        InventoryItem.objects
        .select_related("item_type", "category")
        .filter(active=True, stock_qty__lte=F("minimum_qty"))
        .order_by("item_name")
    )

    return render(request, "faculty_management/power_house/low_stock_dashboard.html", {
        "low_stock_items": low_stock_items,
    })



def stock_ledger(request):
    ledger_entries = (
        StockLedger.objects
        .select_related("item")
        .all()
        .order_by("-created_at")
    )

    return render(request, "faculty_management/power_house/stock_ledger.html", {
        "ledger_entries": ledger_entries,
    })


from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from faculty_management.models import (
    MaterialRequest,
    MaterialIssueProof,
)

from faculty_management.models import (
    MaterialRequest,
    MaterialIssueProof,
    MaterialIssueProofApprovalData,
    MaterialScrapItem,
    StockLedger,
)


def create_material_proof_approver_chain(proof, creator_role_id):
    routes = MaterialRequestApprovers.objects.filter(
        creator_role_id=str(creator_role_id)
    ).order_by("approver_level")

    count = 0

    for route in routes:
        MaterialIssueProofApprovalData.objects.create(
            proof=proof,
            creator_role_id=str(creator_role_id),
            approver_role_id=str(route.approver_role_id),
            approver_level=route.approver_level,
            status=MaterialIssueProofApprovalData.Status.PENDING,
        )
        count += 1

    return count


def get_current_pending_proof_entry(proof):
    return (
        proof.approval_entries
        .filter(status=MaterialIssueProofApprovalData.Status.PENDING)
        .order_by("approver_level", "id")
        .first()
    )


def can_approve_material_proof_entry(entry, faculty, logged_in_role):
    if not entry or not faculty or not logged_in_role:
        return False

    if entry.status != MaterialIssueProofApprovalData.Status.PENDING:
        return False

    current = get_current_pending_proof_entry(entry.proof)

    if not current or current.id != entry.id:
        return False

    if str(logged_in_role.id) != str(entry.approver_role_id):
        return False

    if is_global_user(faculty):
        return True

    route = MaterialRequestApprovers.objects.filter(
        creator_role_id=str(entry.creator_role_id),
        approver_role_id=str(entry.approver_role_id),
        approver_level=entry.approver_level,
    ).first()

    if not route:
        return True

    if route.is_cross_department_approver == "YES":
        if route.approver_department_id:
            return faculty.department_id == route.approver_department_id
        return True

    request_department_id = entry.proof.material_request.department_id

    if request_department_id:
        return faculty.department_id == request_department_id

    return True


def material_proof_upload(request):
    logged_in_faculty = get_logged_in_faculty(request)
    creator_role = get_logged_in_role(request)

    if not logged_in_faculty:
        messages.error(request, "Faculty record not found.")
        return redirect("dashboard")

    if not creator_role:
        messages.error(request, "User role not found.")
        return redirect("dashboard")

    if request.method == "POST":
        request_id = request.POST.get("request_id")
        proof_file = request.FILES.get("proof_file")
        description = request.POST.get("description") or ""
        remarks = request.POST.get("remarks") or ""

        scrap_types = request.POST.getlist("scrap_type[]")
        scrap_names = request.POST.getlist("scrap_name[]")
        scrap_counts = request.POST.getlist("scrap_count[]")

        if not request_id:
            messages.error(request, "Please select a material request.")
            return redirect("material_proof_upload")

        if not proof_file:
            messages.error(request, "Please upload a proof document.")
            return redirect("material_proof_upload")

        mat_request = get_object_or_404(
            MaterialRequest.objects.select_related(
                "department",
                "ticket",
                "ticket__student",
                "ticket__student__department",
                "ticket__faculty",
                "ticket__faculty__department",
            ),
            id=request_id,
            status__iexact="ISSUED"
        )

        if MaterialIssueProof.objects.filter(material_request_id=mat_request.id).exists():
            messages.error(request, "Proof already uploaded for this request.")
            return redirect("material_proof_upload")

        try:
            with transaction.atomic():
                proof = MaterialIssueProof.objects.create(
                    material_request=mat_request,
                    proof_file=proof_file,
                    description=description,
                    remarks=remarks,
                    status="PENDING",
                    uploaded_by=logged_in_faculty,
                )

                for scrap_type, name, count in zip(scrap_types, scrap_names, scrap_counts):
                    if scrap_type and name and count:
                        MaterialScrapItem.objects.create(
                            proof=proof,
                            scrap_type=scrap_type,
                            name=name,
                            count=count
                        )

                created_rows = create_material_proof_approver_chain(
                    proof=proof,
                    creator_role_id=str(creator_role.id),
                )

                if created_rows == 0:
                    proof.status = "VERIFIED"
                    proof.verified_by = logged_in_faculty
                    proof.verified_at = timezone.now()
                    proof.verified_remarks = "Auto verified. No approval hierarchy found."
                    proof.save(update_fields=[
                        "status",
                        "verified_by",
                        "verified_at",
                        "verified_remarks",
                    ])

                    messages.warning(request, "Proof uploaded and auto-verified. No approval hierarchy found.")
                else:
                    messages.success(request, "Proof uploaded successfully and sent for verification.")

        except Exception as e:
            messages.error(request, f"Proof upload failed: {str(e)}")

        return redirect("material_proof_upload")

    issued_requests = (
        MaterialRequest.objects
        .select_related(
            "department",
            "ticket",
            "ticket__student",
            "ticket__student__department",
            "ticket__faculty",
            "ticket__faculty__department",
        )
        .filter(status__iexact="ISSUED")
        .exclude(issue_proof__isnull=False)
        .order_by("-created_at")
    )

    uploaded_proofs = (
        MaterialIssueProof.objects
        .select_related(
            "material_request",
            "material_request__department",
            "material_request__ticket",
            "material_request__ticket__student",
            "material_request__ticket__faculty",
            "uploaded_by",
            "verified_by"
        )
        .prefetch_related("scrap_items")
        .filter(uploaded_by=logged_in_faculty)
        .order_by("-uploaded_at")
    )

    return render(request, "faculty_management/power_house/material_proof_upload.html", {
        "issued_requests": issued_requests,
        "uploaded_proofs": uploaded_proofs,
    })



def material_proof_approval(request):
    logged_in_faculty = get_logged_in_faculty(request)
    logged_in_role = get_logged_in_role(request)

    if not logged_in_faculty:
        messages.error(request, "Approver faculty record not found.")
        return redirect("dashboard")

    if not logged_in_role:
        messages.error(request, "Approver role not found.")
        return redirect("dashboard")

    if request.method == "POST":
        entry_id = request.POST.get("entry_id")
        action = request.POST.get("action")
        remarks = request.POST.get("remarks") or ""

        if action not in ["verify", "reject"]:
            messages.error(request, "Invalid action.")
            return redirect("material_proof_approval")

        try:
            with transaction.atomic():
                entry = get_object_or_404(
                    MaterialIssueProofApprovalData.objects
                    .select_for_update()
                    .select_related(
                        "proof",
                        "proof__material_request",
                        "proof__material_request__department",
                        "proof__material_request__ticket",
                        "proof__material_request__ticket__student",
                        "proof__material_request__ticket__faculty",
                        "proof__uploaded_by",
                    ),
                    id=entry_id,
                )

                proof = entry.proof

                if proof.status in ["VERIFIED", "REJECTED"]:
                    messages.error(request, "This proof is already closed.")
                    return redirect("material_proof_approval")

                if not can_approve_material_proof_entry(entry, logged_in_faculty, logged_in_role):
                    messages.error(request, "You are not authorized to verify this proof level.")
                    return redirect("material_proof_approval")

                entry.approver = logged_in_faculty
                entry.acted_on = timezone.now()

                if action == "verify":
                    entry.status = MaterialIssueProofApprovalData.Status.APPROVED
                    entry.reason = remarks or f"Verified by {logged_in_faculty.name}"
                    entry.save()

                    has_pending = proof.approval_entries.filter(
                        status=MaterialIssueProofApprovalData.Status.PENDING
                    ).exists()

                    if has_pending:
                        proof.status = "PENDING"
                        proof.save(update_fields=["status"])

                        messages.success(
                            request,
                            f"{proof.material_request.request_no} verified at level {entry.approver_level}."
                        )

                    else:
                        proof.status = "VERIFIED"
                        proof.verified_by = logged_in_faculty
                        proof.verified_at = timezone.now()
                        proof.verified_remarks = remarks or f"Final verification by {logged_in_faculty.name}"
                        proof.save(update_fields=[
                            "status",
                            "verified_by",
                            "verified_at",
                            "verified_remarks",
                        ])

                        update_stock_after_proof_verification(proof, request)

                        messages.success(
                            request,
                            f"{proof.material_request.request_no} proof fully verified. Stock ledger updated."
                        )

                else:
                    if not remarks.strip():
                        messages.error(request, "Rejection reason is required.")
                        return redirect("material_proof_approval")

                    entry.status = MaterialIssueProofApprovalData.Status.REJECTED
                    entry.reason = remarks
                    entry.save()

                    proof.approval_entries.filter(
                        status=MaterialIssueProofApprovalData.Status.PENDING
                    ).exclude(id=entry.id).update(
                        status=MaterialIssueProofApprovalData.Status.REJECTED,
                        reason=f"Auto rejected after level {entry.approver_level}",
                        acted_on=timezone.now(),
                    )

                    proof.status = "REJECTED"
                    proof.verified_by = logged_in_faculty
                    proof.verified_at = timezone.now()
                    proof.verified_remarks = remarks
                    proof.save(update_fields=[
                        "status",
                        "verified_by",
                        "verified_at",
                        "verified_remarks",
                    ])

                    messages.success(
                        request,
                        f"{proof.material_request.request_no} proof rejected successfully."
                    )

        except Exception as e:
            messages.error(request, f"Proof verification failed: {str(e)}")

        return redirect("material_proof_approval")

    pending_entries = (
        MaterialIssueProofApprovalData.objects
        .select_related(
            "proof",
            "proof__material_request",
            "proof__material_request__department",
            "proof__material_request__ticket",
            "proof__material_request__ticket__student",
            "proof__material_request__ticket__faculty",
            "proof__uploaded_by",
        )
        .prefetch_related(
            "proof__material_request__request_items__item",
            "proof__scrap_items",
        )
        .filter(status=MaterialIssueProofApprovalData.Status.PENDING)
        .order_by("proof_id", "approver_level", "id")
    )

    actionable_entries = [
        entry for entry in pending_entries
        if can_approve_material_proof_entry(entry, logged_in_faculty, logged_in_role)
    ]

    my_actions = (
        MaterialIssueProofApprovalData.objects
        .select_related(
            "proof",
            "proof__material_request",
            "proof__material_request__department",
            "proof__material_request__ticket",
            "proof__material_request__ticket__student",
            "proof__material_request__ticket__faculty",
            "proof__uploaded_by",
            "approver",
        )
        .prefetch_related(
            "proof__material_request__request_items__item",
            "proof__scrap_items",
        )
        .filter(approver=logged_in_faculty)
        .exclude(status=MaterialIssueProofApprovalData.Status.PENDING)
        .order_by("-acted_on", "-id")[:30]
    )

    return render(request, "faculty_management/power_house/material_proof_approval.html", {
        "actionable_entries": actionable_entries,
        "my_actions": my_actions,
    })
  

def inventory_item_type_entry(request):
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "add":
            type_names = request.POST.getlist("type_name[]")
            descriptions = request.POST.getlist("description[]")

            count = 0

            for i, type_name in enumerate(type_names):
                type_name = type_name.strip()
                description = descriptions[i].strip() if i < len(descriptions) else ""

                if type_name:
                    InventoryItemType.objects.create(
                        type_name=type_name,
                        description=description,
                        active=True,
                    )
                    count += 1

            messages.success(request, f"{count} item type(s) added successfully.")

        elif action == "edit":
            item_id = request.POST.get("item_id")
            item_type = get_object_or_404(InventoryItemType, id=item_id)

            item_type.type_name = request.POST.get("type_name")
            item_type.description = request.POST.get("description")
            item_type.active = True if request.POST.get("active") == "on" else False
            item_type.save()

            messages.success(request, "Item type updated successfully.")

        elif action == "delete":
            item_id = request.POST.get("item_id")
            item_type = get_object_or_404(InventoryItemType, id=item_id)
            item_type.delete()

            messages.success(request, "Item type deleted successfully.")

        return redirect("inventory_item_type_entry")

    item_types = InventoryItemType.objects.all().order_by("type_name")

    return render(request, "faculty_management/power_house/inventory_item_type_entry.html", {
        "item_types": item_types,
    })







def get_academic_year():
    """
    Dynamically returns academic year string.
    Example:
      If current month >= June → '2025-2026'
      Else (Jan–May) → '2024-2025'
    """
    today = date.today()
    current_year = today.year
    if today.month >= 6:  # June or later
        return f"{current_year}-{current_year + 1}"
    else:  # Before June → part of previous cycle
        return f"{current_year - 1}-{current_year}"

from decimal import Decimal, InvalidOperation
import json
import logging

from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Max, Count
from django.http import JsonResponse, HttpResponseForbidden
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from user_accounts.decorators import check_permission
from user_accounts.models import StudentDetails
from faculty_management.models import (
    ProgramOrganizationRecord,
    ProgramOrganizationStudentMark,
    general_information,
    ProgramOutcomeMapping,
)

logger = logging.getLogger(__name__)
from datetime import datetime, timedelta


@check_permission("program_org_dashboard")
def program_org_dashboard(request):
    faculty_id = getattr(request.user, "Employee_id", None)
    faculty = general_information.objects.filter(faculty_id=faculty_id).first()

    if not faculty:
        messages.error(request, "Faculty profile not found. Please contact administrator.")
        return redirect("home")

    faculty_department = faculty.department

    DepartmentModel = None
    if faculty_department:
        DepartmentModel = faculty_department.__class__
    else:
        try:
            DepartmentModel = StudentDetails._meta.get_field("department").remote_field.model
        except Exception:
            DepartmentModel = None

    def model_has_field(model, field_name):
        try:
            model._meta.get_field(field_name)
            return True
        except Exception:
            return False

    selected_department_id = (
        request.POST.get("department")
        or request.POST.get("department_id")
        or request.POST.get("selected_department")
        or request.GET.get("department")
        or request.GET.get("department_id")
        or request.GET.get("selected_department")
        or ""
    )
    selected_department_id = str(selected_department_id).strip()

    all_departments = []
    department = faculty_department

    if DepartmentModel:
        dept_order_field = "Department" if model_has_field(DepartmentModel, "Department") else "id"
        all_departments = list(DepartmentModel.objects.all().order_by(dept_order_field))

        if selected_department_id:
            selected_dept = DepartmentModel.objects.filter(pk=selected_department_id).first()
            if selected_dept:
                department = selected_dept
            else:
                messages.warning(request, "Selected department is invalid. Showing your department data.")

    degree = department.degree if department else None
    has_program_department_field = model_has_field(ProgramOrganizationRecord, "department")

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":

        if request.GET.get("get_semesters"):
            year = request.GET.get("year", "").strip()
            semesters = []

            try:
                duration_years = int(
                    getattr(degree, "effective_duration", None)
                    or getattr(degree, "duration", 0)
                    or 0
                )

                if year and degree:
                    year_int = int(year)
                    if year_int <= duration_years:
                        start_sem = (year_int - 1) * 2 + 1
                        end_sem = year_int * 2
                        semesters = list(range(start_sem, end_sem + 1))
                else:
                    semesters = list(range(1, duration_years * 2 + 1))

            except Exception:
                semesters = []

            return JsonResponse({"semesters": semesters})

        if request.GET.get("get_students_for_selection"):
            year = request.GET.get("year", "").strip()
            section = request.GET.get("section", "").strip()
            semester = request.GET.get("semester", "").strip()

            if not year or not section or not semester:
                return JsonResponse(
                    {"error": "Year, section, and semester are required"},
                    status=400
                )

            students_qs = StudentDetails.objects.filter(
                department=department,
                year=year,
                section=section
            ).order_by("reg_no")

            students_data = []

            for student in students_qs:
                students_data.append({
                    "id": student.id,
                    "reg_no": student.reg_no or "",
                    "name": student.name or "",
                    "gender": student.gender or "",
                    "regulation": student.regulation or "",
                })

            return JsonResponse({"students": students_data})

        if request.GET.get("ajax_get_marks"):
            program_id = request.GET.get("program_id", "").strip()
            record_id = request.GET.get("record_id", "").strip()

            if not program_id:
                return JsonResponse({"error": "Program ID is required"}, status=400)

            if record_id:
                marks_qs = ProgramOrganizationStudentMark.objects.filter(
                    program_id=record_id,
                    program__faculty=faculty
                ).select_related(
                    "student",
                    "program",
                    "regulation",
                    "program_outcome_mapping",
                    "program_outcome_mapping__revised_po",
                    "program_outcome_mapping__non_revised_po"
                ).order_by("year", "section", "student__reg_no")
            else:
                marks_qs = ProgramOrganizationStudentMark.objects.filter(
                    program__program_id=program_id,
                    program__faculty=faculty
                ).select_related(
                    "student",
                    "program",
                    "regulation",
                    "program_outcome_mapping",
                    "program_outcome_mapping__revised_po",
                    "program_outcome_mapping__non_revised_po"
                ).order_by("year", "section", "student__reg_no")

            marks_data = []

            for mark in marks_qs:
                po_info = "N/A"

                if mark.program_outcome_mapping:
                    if mark.program_outcome_mapping.revised_po:
                        po = mark.program_outcome_mapping.revised_po
                        po_info = f"PO{po.program_number}: {po.program_name}"
                    elif mark.program_outcome_mapping.non_revised_po:
                        po = mark.program_outcome_mapping.non_revised_po
                        po_info = f"PO{po.program_number}: {po.program_name}"

                marks_data.append({
                    "reg_no": mark.student.reg_no if mark.student else "",
                    "student_name": mark.student.name if mark.student else "Unknown",
                    "year": mark.year or "",
                    "section": mark.section or "",
                    "semester": mark.semester or "",
                    "regulation": str(mark.regulation) if mark.regulation else "N/A",
                    "marks": float(mark.marks) if mark.marks is not None else None,
                    "program_outcome": po_info,
                    "po_mapping_id": mark.program_outcome_mapping.id if mark.program_outcome_mapping else None,
                })

            return JsonResponse({"marks": marks_data})

        if request.GET.get("ajax_students"):
            program_id = request.GET.get("program_id", "").strip()
            year = request.GET.get("year", "").strip()

            if not program_id:
                return JsonResponse({"error": "Program ID is required"}, status=400)

            program_records = ProgramOrganizationRecord.objects.filter(
                program_id=program_id,
                faculty=faculty
            )

            if has_program_department_field and department:
                program_records = program_records.filter(department=department)

            program_records = program_records.order_by("id")

            if not program_records.exists():
                return JsonResponse({"error": "Program not found"}, status=404)

            students = []
            seen_ids = set()

            existing_marks = {}
            marks_qs = ProgramOrganizationStudentMark.objects.filter(
                program__program_id=program_id
            ).select_related("student")

            for mark in marks_qs:
                if mark.student_id:
                    existing_marks[mark.student_id] = (
                        float(mark.marks) if mark.marks is not None else None
                    )

            for rec in program_records:
                if year and rec.year != year:
                    continue

                student_dept = getattr(rec, "department", None) if has_program_department_field else department
                if not student_dept:
                    student_dept = department

                qs = StudentDetails.objects.filter(
                    department=student_dept,
                    year=rec.year,
                    section=rec.section
                )

                if rec.student_selection_type == "boys_only":
                    qs = qs.filter(gender="Male")
                elif rec.student_selection_type == "girls_only":
                    qs = qs.filter(gender="Female")
                elif rec.student_selection_type == "specific_students":
                    if rec.specific_student_reg_numbers:
                        reg_numbers = [
                            rn.strip()
                            for rn in rec.specific_student_reg_numbers.split(",")
                            if rn.strip()
                        ]
                        qs = qs.filter(reg_no__in=reg_numbers) if reg_numbers else qs.none()
                    else:
                        qs = qs.none()

                qs = qs.order_by("reg_no")

                for student in qs:
                    if student.id in seen_ids:
                        continue

                    seen_ids.add(student.id)

                    students.append({
                        "id": student.id,
                        "reg_no": student.reg_no or "",
                        "name": student.name or "",
                        "year": rec.year or "",
                        "section": rec.section or "",
                        "regulation": str(rec.regulation) if rec.regulation else "N/A",
                        "existing_marks": existing_marks.get(student.id),
                        "gender": student.gender or "",
                        "selection_type": rec.student_selection_type,
                    })

            return JsonResponse({"students": students})

    if request.method == "POST":
        action = request.POST.get("action", "").strip()

        if request.headers.get("X-Requested-With") == "XMLHttpRequest" and "report" in request.FILES:
            try:
                record_id = request.POST.get("record_id", "").strip()
                report_file = request.FILES.get("report")

                if not record_id:
                    return JsonResponse({"success": False, "error": "Record ID is required"}, status=400)

                if not report_file:
                    return JsonResponse({"success": False, "error": "No file uploaded"}, status=400)

                max_size = 10 * 1024 * 1024
                if report_file.size > max_size:
                    return JsonResponse({"success": False, "error": "File size exceeds 10MB limit"}, status=400)

                allowed_extensions = [".pdf", ".doc", ".docx"]
                file_extension = os.path.splitext(report_file.name)[1].lower()

                if file_extension not in allowed_extensions:
                    return JsonResponse({"success": False, "error": "Only PDF, DOC, and DOCX files are allowed"}, status=400)

                record = ProgramOrganizationRecord.objects.filter(
                    id=record_id,
                    faculty=faculty
                ).first()

                if not record:
                    return JsonResponse({"success": False, "error": "Program record not found"}, status=404)

                if record.report:
                    try:
                        if os.path.isfile(record.report.path):
                            os.remove(record.report.path)
                    except Exception:
                        pass

                record.report = report_file
                record.save()

                return JsonResponse({
                    "success": True,
                    "message": "Report uploaded successfully",
                    "report_url": record.report.url if record.report else ""
                })

            except Exception as e:
                logger.exception("Error uploading report")
                return JsonResponse({"success": False, "error": str(e)}, status=500)

        if action == "save_marks":
            try:
                program_id = request.POST.get("program_id", "").strip()

                if not program_id:
                    messages.error(request, "Program ID is required.")
                    return redirect("program_org_dashboard")

                program_records = ProgramOrganizationRecord.objects.filter(
                    program_id=program_id,
                    faculty=faculty
                )

                if has_program_department_field and department:
                    program_records = program_records.filter(department=department)

                program_records = program_records.order_by("id")

                if not program_records.exists():
                    messages.error(request, "Program not found.")
                    return redirect("program_org_dashboard")

                po_mappings = ProgramOutcomeMapping.objects.filter(
                    program_organization__in=program_records
                ).distinct()

                if not po_mappings.exists():
                    messages.error(request, "No Program Outcomes found for this program.")
                    return redirect("program_org_dashboard")

                marks_saved = 0
                marks_updated = 0

                for key, value in request.POST.items():
                    if key.startswith("student_marks[") and key.endswith("]"):
                        student_id = key[len("student_marks["):-1]
                        marks_value = value.strip()

                        if not marks_value:
                            continue

                        try:
                            student_id = int(student_id)
                            marks_value = float(marks_value)

                            if marks_value < 0:
                                messages.error(request, f"Invalid marks {marks_value}. Marks cannot be negative.")
                                continue

                            if marks_value > 100:
                                messages.error(request, f"Invalid marks {marks_value}. Marks cannot exceed 100.")
                                continue

                        except Exception:
                            continue

                        student = StudentDetails.objects.filter(
                            id=student_id,
                            is_active=True
                        ).first()

                        if not student:
                            continue

                        matching_record = None

                        for rec in program_records:
                            if rec.year == student.year and rec.section == student.section:
                                matching_record = rec
                                break

                        if not matching_record:
                            continue

                        record_po_mappings = ProgramOutcomeMapping.objects.filter(
                            program_organization=matching_record
                        )

                        mark_department = (
                            matching_record.department
                            if has_program_department_field and getattr(matching_record, "department", None)
                            else department
                        )

                        for po_mapping in record_po_mappings:
                            mark_obj, created = ProgramOrganizationStudentMark.objects.update_or_create(
                                program=matching_record,
                                student=student,
                                program_outcome_mapping=po_mapping,
                                defaults={
                                    "marks": marks_value,
                                    "department": mark_department,
                                    "regulation": matching_record.regulation,
                                    "year": student.year,
                                    "section": student.section,
                                    "semester": matching_record.semester,
                                }
                            )

                            if created:
                                marks_saved += 1
                            else:
                                marks_updated += 1

                if marks_saved > 0 or marks_updated > 0:
                    messages.success(
                        request,
                        f"Marks saved successfully! Created: {marks_saved}, Updated: {marks_updated}"
                    )
                else:
                    messages.warning(request, "No marks were entered.")

                return redirect("program_org_dashboard")

            except Exception as e:
                logger.exception("Error saving marks")
                messages.error(request, f"Error saving marks: {e}")
                return redirect("program_org_dashboard")

        if action == "populate_po_marks":
            try:
                existing_marks = ProgramOrganizationStudentMark.objects.filter(
                    program_outcome_mapping__isnull=True,
                    program__faculty=faculty
                )

                if not existing_marks.exists():
                    messages.info(request, "No marks found that need PO mapping.")
                    return redirect("program_org_dashboard")

                po_mappings_by_program = {}

                for po_mapping in ProgramOutcomeMapping.objects.all():
                    program_obj_id = po_mapping.program_organization.id
                    po_mappings_by_program.setdefault(program_obj_id, []).append(po_mapping)

                created_count = 0
                updated_count = 0

                for mark in existing_marks:
                    program_obj_id = mark.program.id if mark.program else None

                    if program_obj_id and program_obj_id in po_mappings_by_program:
                        po_mappings = po_mappings_by_program[program_obj_id]

                        if po_mappings:
                            mark.program_outcome_mapping = po_mappings[0]
                            mark.save()
                            updated_count += 1

                            for po_mapping in po_mappings[1:]:
                                ProgramOrganizationStudentMark.objects.create(
                                    program=mark.program,
                                    department=mark.department,
                                    student=mark.student,
                                    regulation=mark.regulation,
                                    year=mark.year,
                                    section=mark.section,
                                    semester=mark.semester,
                                    marks=mark.marks,
                                    program_outcome_mapping=po_mapping
                                )
                                created_count += 1

                if created_count > 0 or updated_count > 0:
                    messages.success(
                        request,
                        f"Successfully populated PO marks! Updated: {updated_count}, Created: {created_count}"
                    )
                else:
                    messages.warning(request, "No marks were processed.")

                return redirect("program_org_dashboard")

            except Exception as e:
                logger.exception("Error populating PO marks")
                messages.error(request, f"Error populating PO marks: {e}")
                return redirect("program_org_dashboard")

        try:
            from course_management.models import Program_outcomes, Regulations

            program_name = (request.POST.get("program_name") or "").strip()
            resource_person = (request.POST.get("resource_person") or "").strip()
            address = (request.POST.get("address") or "").strip()
            professional_society_id = request.POST.get("professional_society", "").strip()
            program_type_id = request.POST.get("program_type", "").strip()
            mode_of_program = request.POST.get("mode_of_program", "").strip()
            from_date = request.POST.get("from_date")
            to_date = request.POST.get("to_date")

            years = request.POST.getlist("year[]")
            sections = request.POST.getlist("section[]")
            semesters = request.POST.getlist("semester[]")
            student_selection_types = request.POST.getlist("student_selection_type[]")
            specific_reg_numbers = request.POST.getlist("specific_reg_numbers[]")

            if not program_name:
                messages.error(request, "Program name is required.")
                return redirect("program_org_dashboard")

            if not resource_person:
                messages.error(request, "Resource person is required.")
                return redirect("program_org_dashboard")

            if not address:
                messages.error(request, "Address is required.")
                return redirect("program_org_dashboard")

            if not from_date or not to_date:
                messages.error(request, "From date and To date are required.")
                return redirect("program_org_dashboard")

            if len(years) != len(sections) or len(sections) != len(semesters):
                messages.error(request, "Mismatch in year, section, and semester selections.")
                return redirect("program_org_dashboard")

            if len(years) != len(student_selection_types):
                messages.error(request, "Mismatch in student selection options.")
                return redirect("program_org_dashboard")

            professional_society = None
            if professional_society_id:
                try:
                    professional_society = ProfessionalSociety.objects.get(id=professional_society_id)
                except ProfessionalSociety.DoesNotExist:
                    messages.error(request, "Selected professional society is invalid.")
                    return redirect("program_org_dashboard")

            program_type = None
            if program_type_id:
                try:
                    program_type = ProgramType.objects.get(
                        id=program_type_id,
                        department=department
                    )
                except ProgramType.DoesNotExist:
                    messages.error(request, "Selected program type is invalid.")
                    return redirect("program_org_dashboard")

            try:
                from_date_obj = datetime.strptime(from_date, "%Y-%m-%d").date()
                to_date_obj = datetime.strptime(to_date, "%Y-%m-%d").date()
            except ValueError as e:
                messages.error(request, f"Invalid date format: {e}")
                return redirect("program_org_dashboard")

            if to_date_obj < from_date_obj:
                messages.error(request, "To date must be after from date.")
                return redirect("program_org_dashboard")

            selected_po_ids = request.POST.getlist("selected_program_outcomes[]")

            if not selected_po_ids:
                messages.error(request, "Please select at least one Program Outcome.")
                return redirect("program_org_dashboard")

            selected_pos = Program_outcomes.objects.filter(
                id__in=selected_po_ids,
                is_active=True
            )

            if selected_pos.count() != len(selected_po_ids):
                messages.error(request, "Some selected Program Outcomes are invalid.")
                return redirect("program_org_dashboard")

            created_records = []
            records_created = 0

            for i in range(len(years)):
                y = (years[i] or "").strip()
                sec = (sections[i] or "").strip()
                sem = (semesters[i] or "").strip()
                selection_type = student_selection_types[i] if i < len(student_selection_types) else "both"
                reg_numbers = specific_reg_numbers[i] if i < len(specific_reg_numbers) else ""

                if not (y and sec and sem):
                    continue

                regulation_obj = None

                sample_student = StudentDetails.objects.filter(
                    department=department,
                    year=y,
                    section=sec
                ).exclude(
                    regulation__isnull=True
                ).exclude(
                    regulation__exact=""
                ).first()

                if sample_student:
                    regulation_obj = Regulations.objects.filter(
                        year=sample_student.regulation
                    ).first()

                rec_kwargs = {
                    "faculty": faculty,
                    "regulation": regulation_obj,
                    "year": y,
                    "section": sec,
                    "semester": sem,
                    "program_name": program_name,
                    "resource_person": resource_person,
                    "address": address,
                    "professional_society": professional_society,
                    "program_type": program_type,
                    "mode_of_program": mode_of_program if mode_of_program else None,
                    "from_date": from_date_obj,
                    "to_date": to_date_obj,
                    "no_of_days": (to_date_obj - from_date_obj).days + 1,
                    "approval": "Pending",
                    "student_selection_type": selection_type,
                    "specific_student_reg_numbers": reg_numbers.strip() if reg_numbers else None,
                }

                if has_program_department_field:
                    rec_kwargs["department"] = department

                rec = ProgramOrganizationRecord(**rec_kwargs)
                rec.full_clean()
                rec.save()

                created_records.append(rec)
                records_created += 1

            if not created_records:
                messages.error(request, "No records created. Please check selections.")
                return redirect("program_org_dashboard")

            max_program_id = ProgramOrganizationRecord.objects.aggregate(
                max_id=Max("program_id")
            )["max_id"] or 0

            next_program_id = max_program_id + 1

            for rec in created_records:
                rec.program_id = next_program_id
                rec.save()

            for rec in created_records:
                ProgramOutcomeMapping.objects.filter(program_organization=rec).delete()

                for po in selected_pos:
                    if po.is_revised:
                        ProgramOutcomeMapping.objects.create(
                            program_organization=rec,
                            revised_po=po,
                            non_revised_po=None
                        )
                    else:
                        ProgramOutcomeMapping.objects.create(
                            program_organization=rec,
                            revised_po=None,
                            non_revised_po=po
                        )

            messages.success(
                request,
                f"Successfully created {records_created} record(s) for '{program_name}' with Program ID: {next_program_id}."
            )

            return redirect("program_org_dashboard")

        except Exception as e:
            logger.exception("Error in program application submit")
            messages.error(request, f"Error processing form: {e}")
            return redirect("program_org_dashboard")

    today = timezone.localdate()
    academic_year = f"{today.year}-{today.year + 1}" if today.month >= 6 else f"{today.year - 1}-{today.year}"
    current_year = today.year

    student_base_qs = StudentDetails.objects.filter(
        department=department,
        is_active=True
    )

    sections = list(
        student_base_qs.values_list("section", flat=True)
        .distinct()
        .exclude(section__isnull=True)
        .exclude(section__exact="")
        .order_by("section")
    )

    available_batches = list(
        student_base_qs.values_list("batch", flat=True)
        .distinct()
        .exclude(batch__isnull=True)
        .exclude(batch__exact="")
        .order_by("batch")
    )

    available_years = list(
        student_base_qs.values_list("year", flat=True)
        .distinct()
        .exclude(year__isnull=True)
        .exclude(year__exact="")
        .order_by("year")
    )

    available_semesters = []

    try:
        duration_years = int(
            getattr(degree, "effective_duration", None)
            or getattr(degree, "duration", 0)
            or 0
        )
        available_semesters = list(range(1, duration_years * 2 + 1))
    except Exception:
        available_semesters = []

    available_regulations = list(
        student_base_qs.values_list("regulation", flat=True)
        .distinct()
        .exclude(regulation__isnull=True)
        .exclude(regulation__exact="")
        .order_by("regulation")
    )

    program_outcomes = []
    revised_program_outcomes = []
    non_revised_program_outcomes = []

    try:
        from course_management.models import Program_outcomes

        pos = Program_outcomes.objects.filter(is_active=True).extra(
            select={"program_number_int": "CAST(program_number AS UNSIGNED)"}
        ).order_by("program_number_int")

        for po in pos:
            d = {
                "id": po.id,
                "program_number": po.program_number or "",
                "program_name": po.program_name or "",
                "program_description": po.program_description or "",
                "is_revised": po.is_revised,
            }

            program_outcomes.append(d)

            if po.is_revised:
                revised_program_outcomes.append(d)
            else:
                non_revised_program_outcomes.append(d)

    except Exception:
        pass

    professional_societies = ProfessionalSociety.objects.all().order_by("society_name")
    program_types = ProgramType.objects.filter(department=department).order_by("program_type_name")

    faculty_programs = ProgramOrganizationRecord.objects.filter(faculty=faculty)

    if has_program_department_field:
        faculty_programs = faculty_programs.select_related(
            "faculty",
            "faculty__department",
            "department",
            "professional_society",
            "program_type"
        )
    else:
        faculty_programs = faculty_programs.select_related(
            "faculty",
            "faculty__department",
            "professional_society",
            "program_type"
        )

    stats = {
        "total": faculty_programs.count(),
        "pending": faculty_programs.filter(
            Q(approval__isnull=True) |
            Q(approval="") |
            Q(approval="Pending") |
            Q(approval="HOD_Approved")
        ).count(),
        "approved": faculty_programs.filter(approval="Approved").count(),
        "rejected": faculty_programs.filter(approval="Rejected").count(),
    }

    programs = faculty_programs.order_by("-created_at")

    status_filter = request.GET.get("status", "")
    year_filter = request.GET.get("year", "")
    semester_filter = request.GET.get("semester", "")
    date_filter = request.GET.get("date_range", "")
    search_query = request.GET.get("search", "")
    sort_by = request.GET.get("sort", "-created_at")

    if status_filter:
        if status_filter == "pending":
            programs = programs.filter(
                Q(approval__isnull=True) |
                Q(approval="") |
                Q(approval="Pending")
            )
        elif status_filter == "hod_approved":
            programs = programs.filter(approval="HOD_Approved")
        elif status_filter == "approved":
            programs = programs.filter(approval="Approved")
        elif status_filter == "rejected":
            programs = programs.filter(approval="Rejected")

    if year_filter:
        programs = programs.filter(year=year_filter)

    if semester_filter:
        programs = programs.filter(semester=semester_filter)

    if date_filter:
        if date_filter == "last_week":
            programs = programs.filter(created_at__date__gte=today - timedelta(days=7))
        elif date_filter == "last_month":
            programs = programs.filter(created_at__date__gte=today - timedelta(days=30))
        elif date_filter == "last_3_months":
            programs = programs.filter(created_at__date__gte=today - timedelta(days=90))

    if search_query:
        programs = programs.filter(
            Q(program_name__icontains=search_query) |
            Q(resource_person__icontains=search_query) |
            Q(address__icontains=search_query)
        )

    programs = programs.order_by(sort_by)

    for p in programs:
        p.has_marks = ProgramOrganizationStudentMark.objects.filter(
            program__program_id=p.program_id,
            program__year=p.year,
            program__faculty=faculty
        ).exists()

    years = faculty_programs.values_list("year", flat=True).distinct().order_by("year")
    semesters = faculty_programs.values_list("semester", flat=True).distinct().order_by("semester")

    student_marks = []
    program_info = None

    approved_programs = faculty_programs.filter(approval="Approved").order_by("-created_at")

    if approved_programs.exists():
        first_approved = approved_programs.first()

        program_info = {
            "program_id": first_approved.program_id,
            "program_name": first_approved.program_name,
            "year": first_approved.year,
        }

        marks_qs = ProgramOrganizationStudentMark.objects.filter(
            program__program_id=first_approved.program_id,
            program__faculty=faculty
        ).select_related(
            "student",
            "program",
            "regulation",
            "program_outcome_mapping",
            "program_outcome_mapping__revised_po",
            "program_outcome_mapping__non_revised_po"
        ).order_by("year", "section", "student__reg_no")

        for mark in marks_qs:
            po_info = "N/A"

            if mark.program_outcome_mapping:
                if mark.program_outcome_mapping.revised_po:
                    po = mark.program_outcome_mapping.revised_po
                    po_info = f"PO{po.program_number}: {po.program_name}"
                elif mark.program_outcome_mapping.non_revised_po:
                    po = mark.program_outcome_mapping.non_revised_po
                    po_info = f"PO{po.program_number}: {po.program_name}"

            student_marks.append({
                "reg_no": mark.student.reg_no if mark.student else "",
                "student_name": mark.student.name if mark.student else "Unknown",
                "year": mark.year or "",
                "section": mark.section or "",
                "semester": mark.semester or "",
                "regulation": str(mark.regulation) if mark.regulation else "N/A",
                "marks": float(mark.marks) if mark.marks is not None else None,
                "program_outcome": po_info,
                "po_mapping_id": mark.program_outcome_mapping.id if mark.program_outcome_mapping else None,
            })

    context = {
        "faculty_name": faculty.name if faculty else "Faculty",
        "faculty_degree": degree,
        "departments": all_departments,
        "selected_department": department,
        "selected_department_id": department.id if department else "",
        "dept_code": department.Department_code if department else "",
        "dept_name": department.Department if department else "",
        "academic_year": academic_year,
        "current_year": current_year,
        "section_choices": sections,
        "available_batches": available_batches,
        "available_years": available_years,
        "available_semesters": available_semesters,
        "available_regulations": available_regulations,
        "program_outcomes": program_outcomes,
        "revised_program_outcomes": revised_program_outcomes,
        "non_revised_program_outcomes": non_revised_program_outcomes,
        "professional_societies": professional_societies,
        "program_types": program_types,
        "stats": stats,
        "programs": programs,
        "years": years,
        "semesters": semesters,
        "student_marks": student_marks,
        "program_info": program_info,
        "current_filters": {
            "status": status_filter,
            "year": year_filter,
            "semester": semester_filter,
            "date_range": date_filter,
            "search": search_query,
            "sort": sort_by,
        }
    }

    return render(
        request,
        "faculty_management/faculty/program_organization/program_org_dashboard.html",
        context
    )





def program_org_excel_export(request):
    """
    Export Program Organization records to Excel format
    Supports both faculty dashboard and HOD review exports
    """
    # Check permissions - allow both dashboard and view permissions
    from django.core.exceptions import PermissionDenied
    
    has_dashboard_permission = False
    has_view_permission = False
    
    try:
        # Check if user has dashboard permission (faculty)
        from faculty_management.decorators import check_permission
        # We'll check permissions manually since we need to allow multiple permissions
        user_permissions = getattr(request.user, 'user_permissions', [])
        role_permissions = getattr(request.user.role, 'permissions', []) if hasattr(request.user, 'role') else []
        
        # Check if user has either permission
        has_dashboard_permission = any(perm.codename == "program_org_dashboard" for perm in user_permissions) or \
                                 any(perm.codename == "program_org_dashboard" for perm in role_permissions)
        has_view_permission = any(perm.codename == "program_org_view" for perm in user_permissions) or \
                            any(perm.codename == "program_org_view" for perm in role_permissions)
        
        if not (has_dashboard_permission or has_view_permission):
            raise PermissionDenied("You don't have permission to export program organization data.")
            
    except Exception as e:
        # If permission checking fails, allow access for now and let the view handle it
        pass
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse("Excel export requires openpyxl library. Please install it using: pip install openpyxl", status=500)
    
    from django.http import HttpResponse
    from datetime import datetime
    
    # Get faculty and department
    faculty_id = getattr(request.user, "Employee_id", None)
    faculty = general_information.objects.filter(faculty_id=faculty_id).first()
    if not faculty:
        return HttpResponse("Faculty profile not found", status=404)
    
    department = faculty.department
    
    # Get query parameters for filtering
    program_id = request.GET.get('program_id', '').strip()
    year = request.GET.get('year', '').strip()
    
    # Check if this is from HOD view (program_org_view) or faculty dashboard
    # HOD should see all programs in department, faculty should see only their own
    if request.user.role.role == 'HOD':
        # HOD can see all programs in their department
        programs = ProgramOrganizationRecord.objects.filter(
            faculty__department=department
        ).order_by("-created_at")
    else:
        # Faculty can see only their own programs
        programs = ProgramOrganizationRecord.objects.filter(faculty=faculty).order_by("-created_at")
    
    # Apply filters if provided
    if program_id:
        programs = programs.filter(program_id=program_id)
    if year:
        programs = programs.filter(year=year)
    
    # Check if we're exporting marks or programs
    export_type = request.GET.get('type', 'programs')  # 'programs' or 'marks'
    
    if export_type == 'marks':
        # Export student marks
        if request.user.role.role == 'HOD':
            marks_qs = ProgramOrganizationStudentMark.objects.filter(
                program__faculty__department=department
            ).select_related("student", "program", "regulation").order_by("year", "section", "student__reg_no")
        else:
            marks_qs = ProgramOrganizationStudentMark.objects.filter(
                program__faculty=faculty
            ).select_related("student", "program", "regulation").order_by("year", "section", "student__reg_no")
        
        if program_id:
            marks_qs = marks_qs.filter(program__program_id=program_id)
        if year:
            marks_qs = marks_qs.filter(year=year)
        
        # Check if there are any marks to export
        if not marks_qs.exists():
            return HttpResponse("No student marks found to export.", status=404)
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Marks"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "S.No", "Register No", "Student Name", "Regulation", 
            "Year", "Section", "Semester", "Marks", "Program Name", "Program ID"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row_num, mark in enumerate(marks_qs, 2):
            data = [
                row_num - 1,  # S.No
                mark.student.reg_no if mark.student else "",
                mark.student.name if mark.student else "Unknown",
                str(mark.regulation) if mark.regulation else "N/A",
                mark.year or "",
                mark.section or "",
                mark.semester or "",
                float(mark.marks) if mark.marks is not None else "",
                mark.program.program_name if mark.program else "",
                mark.program.program_id if mark.program else ""
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                if col == 1:  # S.No column
                    cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        filename = f"program_org_student_marks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
    else:
        # Export program records
        # Check if there are any programs to export
        if not programs.exists():
            return HttpResponse("No program records found to export.", status=404)
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Program Organization"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Program ID", "Program Name", "Faculty Name", "Resource Person", "Professional Society",
            "Year", "Section", "Semester", "Student Selection", "From Date", 
            "To Date", "No. of Days", "Address", "Approval Status", "HOD Remarks", "Submitted Date"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Data rows
        for row_num, program in enumerate(programs, 2):
            # Format student selection
            selection_text = {
                'both': 'Both (Boys & Girls)',
                'boys_only': 'Only Boys',
                'girls_only': 'Only Girls',
                'specific_students': 'Specific Students'
            }.get(program.student_selection_type, 'Both (Default)')
            
            data = [
                program.program_id or "",
                program.program_name or "",
                program.faculty.name if program.faculty else "",
                program.resource_person or "",
                program.professional_society.society_name if program.professional_society else "",
                program.year or "",
                program.section or "",
                program.semester or "",
                selection_text,
                program.from_date.strftime('%Y-%m-%d') if program.from_date else "",
                program.to_date.strftime('%Y-%m-%d') if program.to_date else "",
                program.no_of_days or "",
                program.address or "",
                program.approval or "Pending",
                program.hod_remarks or "",
                program.created_at.strftime('%Y-%m-%d %H:%M') if program.created_at else ""
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                if col == 1:  # Program ID column
                    cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        filename = f"program_organization_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    return response

@check_permission("program_org_view")
def program_org_view(request, cache_bust=None):
    """
    Program Organization Review:
    - HOD sees own department faculty organised programs
    - Principal sees HOD approved / approved / rejected programs
    - HOD approval: Pending -> HOD_Approved
    - Principal approval: HOD_Approved -> Approved
    - Program Dept shows saved department from ProgramOrganizationRecord.department
    """

    user = request.user

    can_view_all_program_org, can_view_dept_program_org, perm = get_program_org_permissions(user)

    if not can_view_all_program_org and not can_view_dept_program_org:
        messages.error(request, "You don't have permission to view program organization data.")
        return redirect("dashboard")

    current_user_faculty = general_information.objects.filter(
        faculty_id=getattr(user, "Employee_id", None)
    ).select_related("department").first()

    if not current_user_faculty or not current_user_faculty.department:
        messages.error(request, "Your department is not mapped.")
        return redirect("dashboard")

    current_user_dept = current_user_faculty.department

    role_name = ""
    if getattr(user, "role", None):
        role_name = getattr(user.role, "role", "") or ""
    role_name = role_name.lower().strip()

    is_principal_level = (
        "principal" in role_name
        or "principle" in role_name
        or can_view_all_program_org
    )

    is_hod_level = (
        "hod" in role_name
        or can_view_dept_program_org
    ) and not is_principal_level

    selected_department_id = request.GET.get("department", "").strip()

    has_program_department_field = any(
        field.name == "department"
        for field in ProgramOrganizationRecord._meta.get_fields()
    )

    # ------------------------------
    # AJAX: Get Student Marks
    # ------------------------------
    if request.headers.get("X-Requested-With") == "XMLHttpRequest" and request.GET.get("ajax_get_marks"):
        program_id = request.GET.get("program_id", "").strip()
        record_id = request.GET.get("record_id", "").strip()

        if not program_id:
            return JsonResponse({"error": "Program ID is required"}, status=400)

        if is_principal_level:
            access_qs = ProgramOrganizationRecord.objects.filter(
                approval__in=["HOD_Approved", "Approved", "Rejected"]
            )
        else:
            access_qs = ProgramOrganizationRecord.objects.filter(
                faculty__department=current_user_dept
            )

        if selected_department_id and is_principal_level:
            access_qs = access_qs.filter(
                faculty__department_id=selected_department_id
            )

        if record_id:
            program_record = access_qs.filter(id=record_id).first()

            if not program_record:
                return JsonResponse({"error": "Access denied to this record"}, status=403)

            marks_qs = ProgramOrganizationStudentMark.objects.filter(
                program_id=record_id
            ).select_related(
                "student",
                "program",
                "regulation"
            )

        else:
            program_record = access_qs.filter(program_id=program_id).first()

            if not program_record:
                return JsonResponse({"error": "Access denied to this program"}, status=403)

            marks_qs = ProgramOrganizationStudentMark.objects.filter(
                program__program_id=program_id
            ).select_related(
                "student",
                "program",
                "regulation"
            )

        marks_qs = marks_qs.order_by("year", "section", "student__reg_no")

        marks_data = []

        for mark in marks_qs:
            marks_data.append({
                "reg_no": mark.student.reg_no if mark.student else "",
                "student_name": mark.student.name if mark.student else "Unknown",
                "year": mark.year or "",
                "section": mark.section or "",
                "semester": mark.semester or "",
                "regulation": str(mark.regulation) if mark.regulation else "N/A",
                "marks": float(mark.marks) if mark.marks is not None else None,
            })

        return JsonResponse({"marks": marks_data})

    # ------------------------------
    # POST: Approve / Reject
    # ------------------------------
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip().lower()
        remarks = (request.POST.get("remarks") or "").strip()
        single_id = request.POST.get("record_id")

        if action not in ["approve", "reject"]:
            return HttpResponseForbidden("Invalid action")

        if is_principal_level:
            base_qs = ProgramOrganizationRecord.objects.filter(
                approval__in=["HOD_Approved", "Approved", "Rejected"]
            )
        else:
            base_qs = ProgramOrganizationRecord.objects.filter(
                faculty__department=current_user_dept
            )

        if single_id:
            record = base_qs.filter(id=single_id).first()
        else:
            messages.error(request, "No record selected.")
            return redirect("program_org_view")

        if not record:
            messages.warning(
                request,
                "Record not found or you don't have permission to access it."
            )
            return redirect("program_org_view")

        current_approval = record.approval or "Pending"

        if action == "approve":
            if current_approval == "Pending":
                if is_hod_level:
                    record.approval = "HOD_Approved"
                    record.hod_remarks = remarks
                    record.hod_approved_at = timezone.now()
                    record.hod_approved_by = current_user_faculty
                    record.save()

                    messages.success(
                        request,
                        "Record approved by HOD. Waiting for Principal approval."
                    )
                else:
                    messages.error(request, "Only HOD can approve pending records.")

            elif current_approval == "HOD_Approved":
                if is_principal_level:
                    record.approval = "Approved"
                    record.principal_remarks = remarks
                    record.principal_approved_at = timezone.now()
                    record.principal_approved_by = current_user_faculty
                    record.save()

                    messages.success(request, "Record finally approved by Principal.")
                else:
                    messages.error(request, "Only Principal can approve after HOD approval.")

            elif current_approval == "Approved":
                messages.warning(request, "Record is already approved.")

            elif current_approval == "Rejected":
                messages.warning(request, "Rejected record cannot be approved.")

            else:
                messages.warning(request, "Record cannot be approved at this stage.")

        elif action == "reject":
            if current_approval == "Approved":
                messages.warning(request, "Approved record cannot be rejected.")
                return redirect("program_org_view")

            if current_approval == "Pending":
                if not is_hod_level:
                    messages.error(request, "Only HOD can reject pending records.")
                    return redirect("program_org_view")

                record.approval = "Rejected"
                record.hod_remarks = remarks

            elif current_approval == "HOD_Approved":
                if not is_principal_level:
                    messages.error(request, "Only Principal can reject after HOD approval.")
                    return redirect("program_org_view")

                record.approval = "Rejected"
                record.principal_remarks = remarks

            else:
                record.approval = "Rejected"

                if is_principal_level:
                    record.principal_remarks = remarks
                else:
                    record.hod_remarks = remarks

            record.save()
            messages.warning(request, "Record rejected.")

        return redirect("program_org_view")

    # ------------------------------
    # GET RECORDS
    # ------------------------------
    if is_principal_level:
        records_qs = ProgramOrganizationRecord.objects.filter(
            approval__in=["HOD_Approved", "Approved", "Rejected"]
        )

        if selected_department_id:
            records_qs = records_qs.filter(
                faculty__department_id=selected_department_id
            )
    else:
        records_qs = ProgramOrganizationRecord.objects.filter(
            faculty__department=current_user_dept
        )

    if has_program_department_field:
        records_qs = records_qs.select_related(
            "faculty",
            "faculty__department",
            "department"
        )
    else:
        records_qs = records_qs.select_related(
            "faculty",
            "faculty__department"
        )

    records_qs = records_qs.order_by("-created_at")

    stats = {
        "total": 0,
        "pending": 0,
        "hod_approved": 0,
        "approved": 0,
        "rejected": 0,
    }

    records = []

    for rec in records_qs:
        faculty_name = rec.faculty.name if rec.faculty else "Unknown Faculty"

        faculty_dept = (
            rec.faculty.department.Department
            if rec.faculty and rec.faculty.department
            else "Unknown Department"
        )

        if has_program_department_field and getattr(rec, "department", None):
            program_dept = rec.department.Department
        else:
            program_dept = "Not Mapped"

        marks_count = ProgramOrganizationStudentMark.objects.filter(
            program__program_id=rec.program_id
        ).count()

        records.append({
            "record_id": rec.id,
            "program_id": rec.program_id,
            "program_name": rec.program_name,
            "resource_person": rec.resource_person,
            "address": rec.address,
            "from_date": rec.from_date,
            "to_date": rec.to_date,
            "no_of_days": rec.no_of_days,
            "approval": rec.approval,
            "hod_remarks": rec.hod_remarks,
            "principal_remarks": rec.principal_remarks,
            "hod_approved_at": rec.hod_approved_at,
            "principal_approved_at": rec.principal_approved_at,
            "hod_approved_by": rec.hod_approved_by,
            "principal_approved_by": rec.principal_approved_by,
            "year": rec.year,
            "section": rec.section,
            "semester": rec.semester,
            "created_at": rec.created_at,
            "faculty_name": faculty_name,
            "faculty_department": faculty_dept,
            "program_department": program_dept,
            "marks_count": marks_count,
            "report": rec.report,
        })

        stats["total"] += 1

        st = (rec.approval or "Pending").lower()

        if st == "approved":
            stats["approved"] += 1
        elif st == "rejected":
            stats["rejected"] += 1
        elif st == "hod_approved":
            stats["hod_approved"] += 1
            stats["pending"] += 1
        else:
            stats["pending"] += 1

    departments = []

    if is_principal_level:
        from user_accounts.models import Add_Department
        departments = Add_Department.objects.all().order_by("Department")

    context = {
        "records": records,
        "stats": stats,
        "current_user_dept": current_user_dept.Department,
        "user": user,
        "can_view_all_program_org": can_view_all_program_org,
        "can_view_dept_program_org": can_view_dept_program_org,
        "permission_scope": "All Departments" if is_principal_level else "Department Only",
        "departments": departments,
        "selected_department_id": selected_department_id,
        "is_hod_level": is_hod_level,
        "is_principal_level": is_principal_level,
    }

    return render(
        request,
        "faculty_management/faculty/program_organization/program_org_view.html",
        context
    )

def program_org_pdf(request):
    """Generate Program Organization PDF Report"""
    from datetime import datetime
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Paragraph, Frame
    from reportlab.lib.styles import ParagraphStyle
    from django.conf import settings
    from django.contrib.staticfiles import finders
    import os
    
    user = request.user

    # Get Current User Department
    current_user_faculty = general_information.objects.filter(
        faculty_id=getattr(user, "Employee_id", None)
    ).select_related("department").first()

    if not current_user_faculty or not current_user_faculty.department:
        messages.error(request, "Your department is not mapped.")
        return redirect("program_org_view")

    current_user_dept = current_user_faculty.department

    # Check if this is a marks export (has program_id parameter) or general program records export
    program_id = request.GET.get('program_id')
    year = request.GET.get('year')
    export_type = "marks" if program_id else "programs"

    if export_type == "marks":
        # Export student marks for specific program
        marks_qs = ProgramOrganizationStudentMark.objects.filter(
            program__faculty__department=current_user_dept
        ).select_related("student", "program", "regulation")
        
        if program_id:
            marks_qs = marks_qs.filter(program__program_id=program_id)
        if year:
            marks_qs = marks_qs.filter(program__year=year)
            
        marks_qs = marks_qs.order_by("year", "section", "student__reg_no")
        
        # Prepare response as PDF
        response = HttpResponse(content_type='application/pdf')
        filename = f'Student_Marks_Report_{current_user_dept.Department}_{datetime.now().strftime("%Y%m%d")}.pdf'
        if program_id:
            filename = f'Student_Marks_Program_{program_id}_{datetime.now().strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = f'inline; filename={filename}'

        c = canvas.Canvas(response, pagesize=A4)
        c.setTitle("Student Marks Report")
        
        page_w, page_h = A4
        left_margin = 25 * mm
        right_margin = page_w - 25 * mm
        top_margin = page_h - 20 * mm
        bottom_margin = 20 * mm

        # Header (Professional RIT Format)
        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            for d in getattr(settings, "STATICFILES_DIRS", []):
                cand = os.path.join(d, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand
                    break

        if logo_path and os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                th = 22 * mm
                tw = th * (iw / float(ih))
                c.drawImage(img, left_margin, top_margin - th, width=tw, height=th, mask='auto')
            except Exception:
                pass

        c.setFillColor(colors.HexColor("#2C3E50"))
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(page_w / 2.0, top_margin - 5 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")
        
        c.setFont("Helvetica", 9.5)
        c.drawCentredString(page_w / 2.0, top_margin - 10 * mm,
                           "(Approved by AICTE - New Delhi and Affiliated to Anna University - Chennai)")
        c.drawCentredString(page_w / 2.0, top_margin - 15 * mm,
                           "NAAC Accredited and An ISO 9001:2015 Certified Institution")
        c.drawCentredString(page_w / 2.0, top_margin - 20 * mm,
                           "NBA Accredited UG Programs: CSE, EEE, ECE, MECH")

        # TUV Logo
        tuv_logo_rel = "images/tuvlogo.png"
        tuv_path = finders.find(tuv_logo_rel)
        if tuv_path and os.path.exists(tuv_path):
            try:
                img2 = ImageReader(tuv_path)
                iw, ih = img2.getSize()
                th = 15 * mm
                tw = th * (iw / float(ih))
                c.drawImage(img2, right_margin - tw, top_margin - th + 2 * mm, width=tw, height=th, mask='auto')
            except Exception:
                pass

        c.setStrokeColor(colors.black)
        c.setLineWidth(0.5)
        c.line(left_margin, top_margin - 24 * mm, right_margin, top_margin - 24 * mm)

        # Title & Date
        c.setFont("Helvetica", 11)
        c.drawRightString(right_margin, top_margin - 33 * mm, f"Date: {datetime.now().strftime('%d/%m/%Y')}")
        
        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(page_w / 2.0, top_margin - 43 * mm, "Student Marks Report")
        
        c.setFont("Helvetica", 11)
        c.drawCentredString(page_w / 2.0, top_margin - 50 * mm, f"Department: {current_user_dept.Department}")
        if program_id:
            c.drawCentredString(page_w / 2.0, top_margin - 55 * mm, f"Program ID: {program_id}")

        # Table Headers for Marks
        y_position = top_margin - 70 * mm
        row_height = 6 * mm
        
        # Column widths for marks table - adjusted for better centering
        col_widths = [15*mm, 25*mm, 40*mm, 15*mm, 15*mm, 15*mm, 20*mm, 15*mm]
        table_width = sum(col_widths)
        
        # Center the table on the page
        table_start_x = (page_w - table_width) / 2
        col_positions = [table_start_x]
        for width in col_widths[:-1]:
            col_positions.append(col_positions[-1] + width)

        # Draw table headers
        c.setFillColor(colors.HexColor("#2C3E50"))
        c.rect(table_start_x, y_position - row_height, table_width, row_height, fill=1)
        
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 8)
        headers = ["S.No", "Reg No", "Student Name", "Year", "Section", "Semester", "Regulation", "Marks"]
        
        for i, header in enumerate(headers):
            c.drawString(col_positions[i] + 2*mm, y_position - row_height + 2*mm, header)

        # Draw marks data
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 7)
        
        row_num = 1
        y_position -= row_height
        
        for mark in marks_qs:
            if y_position < bottom_margin + 50*mm:  # Start new page if needed
                c.showPage()
                y_position = top_margin - 70 * mm
                
                # Redraw headers on new page
                c.setFillColor(colors.HexColor("#2C3E50"))
                c.rect(table_start_x, y_position - row_height, table_width, row_height, fill=1)
                
                c.setFillColor(colors.white)
                c.setFont("Helvetica-Bold", 8)
                for i, header in enumerate(headers):
                    c.drawString(col_positions[i] + 2*mm, y_position - row_height + 2*mm, header)
                
                c.setFillColor(colors.black)
                c.setFont("Helvetica", 7)
                y_position -= row_height

            # Alternate row colors
            if row_num % 2 == 0:
                c.setFillColor(colors.HexColor("#f8f9fa"))
                c.rect(table_start_x, y_position - row_height, table_width, row_height, fill=1)
                c.setFillColor(colors.black)

            # Row data for marks
            student_name = mark.student.name if mark.student else "Unknown"
            reg_no = mark.student.reg_no if mark.student else "N/A"
            
            row_data = [
                str(row_num),
                reg_no[:15] + "..." if len(reg_no) > 15 else reg_no,
                student_name[:25] + "..." if len(student_name) > 25 else student_name,
                str(mark.year or "N/A"),
                str(mark.section or "N/A"),
                str(mark.semester or "N/A"),
                str(mark.regulation or "N/A"),
                str(mark.marks if mark.marks is not None else "N/A")
            ]
            
            for i, data in enumerate(row_data):
                c.drawString(col_positions[i] + 2*mm, y_position - row_height + 2*mm, str(data))
            
            y_position -= row_height
            row_num += 1

    else:
        # Export program records (original functionality)
        records_qs = ProgramOrganizationRecord.objects.filter(
            faculty__department=current_user_dept
        ).select_related("faculty", "faculty__department").order_by("-created_at")

        # Apply filters from URL parameters (for dashboard export)
        if program_id:
            records_qs = records_qs.filter(program_id=program_id)
        if year:
            records_qs = records_qs.filter(year=year)

        # Prepare response as PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename=Program_Organization_Report_{current_user_dept.Department}_{datetime.now().strftime("%Y%m%d")}.pdf'

        c = canvas.Canvas(response, pagesize=A4)
        c.setTitle("Program Organization Report")
        
        page_w, page_h = A4
        left_margin = 25 * mm
        right_margin = page_w - 25 * mm
        top_margin = page_h - 20 * mm
        bottom_margin = 20 * mm

        # Header (Professional RIT Format) - same as above
        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            for d in getattr(settings, "STATICFILES_DIRS", []):
                cand = os.path.join(d, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand
                    break

        if logo_path and os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                th = 22 * mm
                tw = th * (iw / float(ih))
                c.drawImage(img, left_margin, top_margin - th, width=tw, height=th, mask='auto')
            except Exception:
                pass

        c.setFillColor(colors.HexColor("#2C3E50"))
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(page_w / 2.0, top_margin - 5 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")
        
        c.setFont("Helvetica", 9.5)
        c.drawCentredString(page_w / 2.0, top_margin - 10 * mm,
                           "(Approved by AICTE - New Delhi and Affiliated to Anna University - Chennai)")
        c.drawCentredString(page_w / 2.0, top_margin - 15 * mm,
                           "NAAC Accredited and An ISO 9001:2015 Certified Institution")
        c.drawCentredString(page_w / 2.0, top_margin - 20 * mm,
                           "NBA Accredited UG Programs: CSE, EEE, ECE, MECH")

        # TUV Logo
        tuv_logo_rel = "images/tuvlogo.png"
        tuv_path = finders.find(tuv_logo_rel)
        if tuv_path and os.path.exists(tuv_path):
            try:
                img2 = ImageReader(tuv_path)
                iw, ih = img2.getSize()
                th = 15 * mm
                tw = th * (iw / float(ih))
                c.drawImage(img2, right_margin - tw, top_margin - th + 2 * mm, width=tw, height=th, mask='auto')
            except Exception:
                pass

        c.setStrokeColor(colors.black)
        c.setLineWidth(0.5)
        c.line(left_margin, top_margin - 24 * mm, right_margin, top_margin - 24 * mm)

        # Title & Date
        c.setFont("Helvetica", 11)
        c.drawRightString(right_margin, top_margin - 33 * mm, f"Date: {datetime.now().strftime('%d/%m/%Y')}")
        
        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(page_w / 2.0, top_margin - 43 * mm, "Program Organization Report")
        
        c.setFont("Helvetica", 11)
        c.drawCentredString(page_w / 2.0, top_margin - 50 * mm, f"Department: {current_user_dept.Department}")

        # Table Headers for Programs
        y_position = top_margin - 70 * mm
        row_height = 6 * mm
        
        # Column widths for programs table - adjusted for better centering
        col_widths = [15*mm, 40*mm, 30*mm, 25*mm, 25*mm, 20*mm, 25*mm]
        table_width = sum(col_widths)
        
        # Center the table on the page
        table_start_x = (page_w - table_width) / 2
        col_positions = [table_start_x]
        for width in col_widths[:-1]:
            col_positions.append(col_positions[-1] + width)

        # Draw table headers
        c.setFillColor(colors.HexColor("#2C3E50"))
        c.rect(table_start_x, y_position - row_height, table_width, row_height, fill=1)
        
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 8)
        headers = ["S.No", "Program Name", "Faculty", "Dates", "Resource Person", "Status", "Remarks"]
        
        for i, header in enumerate(headers):
            c.drawString(col_positions[i] + 2*mm, y_position - row_height + 2*mm, header)

        # Draw table data
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 7)
        
        row_num = 1
        y_position -= row_height
        
        for rec in records_qs:
            if y_position < bottom_margin + 50*mm:  # Start new page if needed
                c.showPage()
                y_position = top_margin - 70 * mm
                
                # Redraw headers on new page
                c.setFillColor(colors.HexColor("#2C3E50"))
                c.rect(table_start_x, y_position - row_height, table_width, row_height, fill=1)
                
                c.setFillColor(colors.white)
                c.setFont("Helvetica-Bold", 8)
                for i, header in enumerate(headers):
                    c.drawString(col_positions[i] + 2*mm, y_position - row_height + 2*mm, header)
                
                c.setFillColor(colors.black)
                c.setFont("Helvetica", 7)
                y_position -= row_height

            # Alternate row colors
            if row_num % 2 == 0:
                c.setFillColor(colors.HexColor("#f8f9fa"))
                c.rect(table_start_x, y_position - row_height, table_width, row_height, fill=1)
                c.setFillColor(colors.black)

            # Row data for programs
            faculty_name = rec.faculty.name if rec.faculty else "Unknown"
            date_range = f"{rec.from_date.strftime('%d/%m/%Y') if rec.from_date else 'N/A'} - {rec.to_date.strftime('%d/%m/%Y') if rec.to_date else 'N/A'}"
            
            row_data = [
                str(row_num),
                rec.program_name[:25] + "..." if len(rec.program_name or "") > 25 else (rec.program_name or "N/A"),
                faculty_name[:20] + "..." if len(faculty_name) > 20 else faculty_name,
                date_range,
                (rec.resource_person[:15] + "..." if len(rec.resource_person or "") > 15 else (rec.resource_person or "N/A")),
                rec.approval or "Pending",
                (rec.hod_remarks[:20] + "..." if len(rec.hod_remarks or "") > 20 else (rec.hod_remarks or "-"))
            ]
            
            for i, data in enumerate(row_data):
                c.drawString(col_positions[i] + 2*mm, y_position - row_height + 2*mm, str(data))
            
            y_position -= row_height
            row_num += 1

    # Footer (common for both types)
    footer_y = 28 * mm
    c.setStrokeColor(colors.HexColor("#C0392B"))
    c.setLineWidth(0.8)
    c.line(left_margin, footer_y + 2 * mm, right_margin, footer_y + 2 * mm)
    
    c.setFont("Helvetica", 9)
    c.setFillColor(colors.HexColor("#2C3E50"))
    c.drawCentredString(page_w / 2.0, footer_y - 1 * mm,
                       "North Venganallur, Ayyanarkovil Road, Rajapalayam - 626 117, Virudhunagar District, Tamil Nadu.")
    c.drawCentredString(page_w / 2.0, footer_y - 6 * mm,
                       "Tel: 04563 233400 | E-mail: rit@ritrjpm.ac.in | Web: www.ritrjpm.ac.in")

    # Computer-Generated Notice
    c.setFont("Helvetica-Oblique", 10)
    c.setFillColor(colors.black)
    report_type = "student marks report" if export_type == "marks" else "program organization report"
    c.drawCentredString(page_w / 2.0, bottom_margin + 60,
                       f"This is a computer-generated {report_type}.")

    c.showPage()
    c.save()
    return response


def professional_society_view(request):

    # DELETE
    if request.method == "POST" and request.POST.get("delete_id"):
        sid = request.POST.get("delete_id")
        ProfessionalSociety.objects.filter(id=sid).delete()
        return JsonResponse({"status": "success"})

    # CREATE / UPDATE
    if request.method == "POST":
        sid = request.POST.get("id")
        name = request.POST.get("society_name")
        dept_id = request.POST.get("department")

        if not name or not dept_id:
            return JsonResponse({"status": "error", "message": "All fields required"})

        # Duplicate protection
        qs = ProfessionalSociety.objects.filter(
            society_name=name,
            department_id=dept_id
        )
        if sid:
            qs = qs.exclude(id=sid)

        if qs.exists():
            return JsonResponse({"status": "error", "message": "Society already exists for this department"})

        if sid:  # UPDATE
            society = ProfessionalSociety.objects.get(id=sid)
            society.society_name = name
            society.department_id = dept_id
            society.save()
        else:  # CREATE
            society = ProfessionalSociety.objects.create(
                society_name=name,
                department_id=dept_id
            )

        return JsonResponse({
            "status": "success",
            "id": society.id,
            "name": society.society_name,
            "dept": society.department.Department,   # THIS feeds table
            "dept_id": society.department.id        # THIS feeds edit
        })

    # PAGE LOAD
    societies = ProfessionalSociety.objects.select_related('department').all()
    departments = Add_Department.objects.all().order_by('Department')

    return render(request, "faculty_management/professional_society.html", {
        "societies": societies,
        "departments": departments
    })




def program_org_permission(request):
    edit_permission = None

    # ---------- DELETE ----------
    if request.method == "POST" and request.POST.get("action") == "delete":
        perm_id = request.POST.get("perm_id")
        program_org_data_Permission.objects.filter(id=perm_id).delete()
        messages.success(request, "Program organization permission deleted successfully.")
        return redirect("program_org_permission")

    # ---------- EDIT LOAD ----------
    if request.method == "GET" and request.GET.get("edit"):
        edit_permission = get_object_or_404(
            program_org_data_Permission,
            id=request.GET.get("edit")
        )

    # ---------- CREATE / UPDATE ----------
    if request.method == "POST" and request.POST.get("action") == "save":
        role_ids = request.POST.getlist("roles[]")
        can_view_all = request.POST.get("can_view_all_program_org_data") == "on"
        can_view_dept = request.POST.get("can_view_department_program_org_data") == "on"
        perm_id = request.POST.get("perm_id")

        if perm_id:
            program_org_data_Permission.objects.filter(id=perm_id).update(
                can_view_all_program_org_data=can_view_all,
                can_view_department_program_org_data=can_view_dept,
            )
            messages.success(request, "Program organization permission updated successfully.")
            return redirect("program_org_permission")

        if not role_ids:
            messages.error(request, "At least one role is required.")
            return redirect("program_org_permission")

        for role_id in role_ids:
            program_org_data_Permission.objects.update_or_create(
                role_id=role_id,
                defaults={
                    "can_view_all_program_org_data": can_view_all,
                    "can_view_department_program_org_data": can_view_dept,
                }
            )

        messages.success(request, "Program organization permission saved successfully.")
        return redirect("program_org_permission")

    roles = Role.objects.using("rit_approval_system").all()
    context = {
        "roles": roles,
        "edit_permission": edit_permission,
    }
    return render(request, "faculty_management/faculty/program_organization/Program_org_permission.html", context)


@require_GET
def program_org_permission_api(request):
    search = (request.GET.get("search") or "").strip()
    page = int(request.GET.get("page", 1))
    
    permissions = program_org_data_Permission.objects.all().order_by("id")
    roles_qs = Role.objects.using("rit_approval_system").all()
    role_map = {r.id: r.role for r in roles_qs}

    if search:
        role_ids = list(roles_qs.filter(role__icontains=search).values_list("id", flat=True))
        if not role_ids:
            permissions = program_org_data_Permission.objects.none()
        else:
            permissions = permissions.filter(role_id__in=role_ids)

    page_size = 25
    paginator = Paginator(permissions, page_size)
    page_obj = paginator.get_page(page)

    data = [
        {
            "id": perm.id,
            "role": role_map.get(perm.role_id, "Unknown"),
            "can_view_all": perm.can_view_all_program_org_data,
            "can_view_dept": perm.can_view_department_program_org_data,
        }
        for perm in page_obj
    ]

    return JsonResponse({
        "results": data,
        "page": page_obj.number,
        "total_pages": paginator.num_pages,
        "has_next": page_obj.has_next(),
        "has_prev": page_obj.has_previous(),
        "total_count": paginator.count,
        "page_size": page_size,
    })


def type_of_program(request):

    # DELETE
    if request.method == "POST" and request.POST.get("delete_id"):
        pid = request.POST.get("delete_id")
        ProgramType.objects.filter(id=pid).delete()
        return JsonResponse({"status": "success"})

    # CREATE / UPDATE
    if request.method == "POST":
        pid = request.POST.get("id")
        name = request.POST.get("program_type_name")
        dept_id = request.POST.get("department")

        if not name or not dept_id:
            return JsonResponse({"status": "error", "message": "All fields required"})

        # Duplicate protection
        qs = ProgramType.objects.filter(
            program_type_name=name,
            department_id=dept_id
        )
        if pid:
            qs = qs.exclude(id=pid)

        if qs.exists():
            return JsonResponse({"status": "error", "message": "Program type already exists for this department"})

        try:
            if pid:  # UPDATE
                program_type = ProgramType.objects.get(id=pid)
                program_type.program_type_name = name
                program_type.department_id = dept_id
                program_type.save()
            else:  # CREATE
                program_type = ProgramType.objects.create(
                    program_type_name=name,
                    department_id=dept_id
                )

            return JsonResponse({
                "status": "success",
                "id": program_type.id,
                "name": program_type.program_type_name,
                "dept": program_type.department.Department,   # THIS feeds table
                "dept_id": program_type.department.id        # THIS feeds edit
            })
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})

    # PAGE LOAD
    program_types = ProgramType.objects.select_related('department').all()
    departments = Add_Department.objects.all().order_by('Department')

    return render(request, "faculty_management/type_of_program.html", {
        "program_types": program_types,
        "departments": departments
    })


# ============================================================
#  LAB MANAGEMENT SYSTEM
#  Faculty-side module. Permission-gated via FacultyFunction
#  (function name = "lab_management"). Appears in the sidebar
#  only when the logged-in user's role has the permission
#  granted from the Permission Management modal.
#
#  NOTE: Business rules / data model are pending. This is the
#  landing page scaffold — sub-features will be added once the
#  rules are finalised.
# ============================================================
@no_cache
@check_permission("lab_management")
def lab_management(request):
    """Lab Management System landing page (faculty side).

    Gating is handled by @check_permission:
      * superusers always pass
      * other users pass only if session['permissions']['lab_management'] is True
    The same permission flag drives whether this page appears in the
    sidebar (built by the @faculty_management decorator).
    """
    # check_permission already sets request.session['current_page'].
    permissions = request.session.get("permissions", {})
    can_access = bool(request.user.is_superuser) or bool(permissions.get("lab_management"))

    faculty = None
    try:
        faculty = general_information.objects.filter(
            faculty_id=request.user.Employee_id
        ).select_related("department").first()
    except Exception:
        faculty = None

    context = {
        "can_access_lab_management": can_access,
        "faculty": faculty,
        # Placeholder sub-features for the landing dashboard. These map to
        # future pages/permissions; for now they are display-only cards.
        "lab_features": [
            {
                "key": "labs",
                "title": "Labs",
                "description": "Register and manage laboratories and their details.",
                "icon": "flask",
                "url_name": None,
            },
            {
                "key": "lab_equipment",
                "title": "Equipment",
                "description": "Track equipment and instruments available in each lab.",
                "icon": "microscope",
                "url_name": None,
            },
            {
                "key": "lab_booking",
                "title": "Lab Bookings",
                "description": "Request and approve lab usage slots.",
                "icon": "calendar-check",
                "url_name": None,
            },
            {
                "key": "lab_incharge",
                "title": "Lab In-charge",
                "description": "Assign faculty/staff responsible for each lab.",
                "icon": "user-gear",
                "url_name": None,
            },
        ],
    }
    return render(request, "faculty_management/lab_management.html", context)

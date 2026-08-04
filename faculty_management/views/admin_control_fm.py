from django.shortcuts import render
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, OperationalError, transaction
from faculty_management.decorators import faculty_management
from faculty_management.models import FacultyCategory, FacultyFunction, StaffCategoryAssignment, general_information
from user_accounts.models import Add_Department, Role, Department, USER, NewUserAdder
from django.utils.dateparse import parse_date
import re
from user_accounts.decorators import faculty_login_required, no_cache, is_super_user, check_permission

APPROVAL_DB = "rit_approval_system"

# @faculty_login_required
@faculty_management
def fm_home(request):
    # # print("fm home page ")
    request.session['current_page'] = 'fm_home'
    return redirect('home')



@check_permission("fm_hello")
def fm_hello(request):
    return render(request, "fm_hello.html")

@faculty_login_required
@no_cache
@is_super_user('faculty_management')
def fm_assign_permission(request):
    if request.method == 'POST':  
        permissions = request.POST
        for role_name, role_permissions in permissions.items():
            if role_name.startswith('permissions'):
                try:
                    # Extract data from role_name using regex
                    extract_data = list(re.findall(r'\[([^\]]+)\]', role_name))
                    if len(extract_data) < 2:  # Ensure there are at least role and function
                        messages.warning(request,f"Invalid format in role_name: {role_name}. Skipping.")
                        continue
                    
                    extract_data.append(role_permissions)

                    # Retrieve the role (Handle if the role does not exist)
                    try:
                        role = Role.objects.using("rit_approval_system").get(role=extract_data[0])

                    except Role.DoesNotExist:
                        messages.error(request,f"Role {extract_data[0]} does not exist.")
                        messages.error(request, f"Role '{extract_data[0]}' does not exist. Skipping this entry.")
                        continue
                    
                    # Parse permissions - handle the case where role_permissions is a list (unlikely with POST data)
                    if isinstance(role_permissions, list):  # Handle list case
                        role_permissions = role_permissions[0]
                    
                    # Convert permission to boolean (True/False)
                    permission = extract_data[2] == 'true'
       
                    
                    # Find or create ApprovalPermissionFunction object
                    permission_obj = FacultyFunction.objects.filter(
                        role=role, function=extract_data[1]
                    ).first()
                    
                    if permission_obj:
                        permission_obj.permission = permission
                        permission_obj.save()
                    else:
                        # Create a new ApprovalPermissionFunction object
                        FacultyFunction.objects.create(
                            role=role,
                            function=extract_data[1],
                            permission=permission
                        )
                except Exception as e:
                    # Catch unexpected errors and log them
                    messages.error(request,f"Error processing role '{role_name}': {str(e)}")
                    messages.error(request, f"An error occurred while processing '{role_name}': {str(e)}")

    # Redirect to admin dashboard after processing
    messages.success(request,"The permission changes have been successfully applied.")
    return redirect('faculty_management')

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.template.loader import render_to_string
from faculty_management.models import DesignationMaster
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def designation_master(request):
    if request.method == "POST":
        action = request.POST.get("action")
        designation_name = request.POST.get("designation_name", "").strip()
        is_teaching = request.POST.get("is_teaching") == "on"

        if action == "add":
            if not designation_name:
                messages.error(request, "Designation name is required.")
                return redirect("designation_master")

            if DesignationMaster.objects.filter(designation_name__iexact=designation_name).exists():
                messages.error(request, "Designation already exists.")
                return redirect("designation_master")

            DesignationMaster.objects.create(
                designation_name=designation_name,
                is_teaching=is_teaching
            )
            messages.success(request, "Designation added successfully!")
            return redirect("designation_master")

        elif action == "update":
            designation_id = request.POST.get("designation_id")
            designation = get_object_or_404(DesignationMaster, id=designation_id)

            if not designation_name:
                messages.error(request, "Designation name is required.")
                return redirect("designation_master")

            if DesignationMaster.objects.filter(
                designation_name__iexact=designation_name
            ).exclude(id=designation_id).exists():
                messages.error(request, "Another designation with this name already exists.")
                return redirect("designation_master")

            designation.designation_name = designation_name
            designation.is_teaching = is_teaching
            designation.save()

            messages.success(request, "Designation updated successfully!")
            return redirect("designation_master")

        elif action == "delete":
            designation_id = request.POST.get("designation_id")
            designation = get_object_or_404(DesignationMaster, id=designation_id)
            designation.delete()
            messages.success(request, "Designation deleted successfully!")
            return redirect("designation_master")

    # Search + Filter
    search_query = request.GET.get("search", "").strip()
    category_filter = request.GET.get("category", "").strip()

    designations = DesignationMaster.objects.all().order_by("id")

    if search_query:
        search_value = search_query.lower()

        if search_value in ["teaching", "non teaching", "non-teaching", "nonteaching"]:
            if search_value == "teaching":
                designations = designations.filter(is_teaching=True)
            else:
                designations = designations.filter(is_teaching=False)
        else:
            designations = designations.filter(designation_name__icontains=search_query)

    if category_filter == "teaching":
        designations = designations.filter(is_teaching=True)
    elif category_filter == "non_teaching":
        designations = designations.filter(is_teaching=False)

    # Pagination
    paginator = Paginator(designations, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "faculty_management/admin/designation_master.html", {
        "page_obj": page_obj,
        "search_query": search_query,
        "category_filter": category_filter,
    })
 
def export_designation_excel(request):
    search_query = request.GET.get("search", "").strip()

    designations = DesignationMaster.objects.filter(is_teaching=True).order_by("designation_name")

    if search_query:
        search_value = search_query.lower()

        # Since export is only for teaching=True, only keep teaching records
        if search_value in ["teaching"]:
            designations = designations.filter(is_teaching=True)
        elif search_value in ["non teaching", "non-teaching", "nonteaching"]:
            designations = designations.none()
        else:
            designations = designations.filter(designation_name__icontains=search_query)

    wb = Workbook()
    ws = wb.active
    ws.title = "Teaching Designations"

    headers = ["designation_name"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for d in designations:
        ws.append([d.designation_name or ""])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = max_length + 3

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="teaching_designations.xlsx"'
    wb.save(response)
    return response




def _mirror_department_to_local(department_name, department_code):
    if not department_code:
        return None

    local_dept, _ = Add_Department.objects.update_or_create(
        Department_code=department_code,
        defaults={
            "Department": department_name,
            "is_active": True,
        },
    )
    return local_dept




def add_approval_user(request):
    if request.method != "POST":
        return redirect("add_employee")

    employee_id = (request.POST.get("employee_id") or "").strip()
    role_id = (request.POST.get("role") or "").strip()
    department_id = (request.POST.get("department") or "").strip()
    category_id = (request.POST.get("category") or "").strip()
    doj_raw = (request.POST.get("doj") or "").strip()

    if not all([employee_id, role_id, department_id, category_id]):
        messages.error(request, "Employee ID, role, department, and category are required.")
        return redirect("add_employee")

    # Date of joining is optional here but, once set, is fixed and shown
    # read-only when the staff member completes their general information.
    doj = None
    if doj_raw:
        doj = parse_date(doj_raw)
        if doj is None:
            messages.error(request, "Date of joining is invalid. Use the date picker (YYYY-MM-DD).")
            return redirect("add_employee")

    try:
        role = Role.objects.using(APPROVAL_DB).get(id=role_id)
        department = Department.objects.using(APPROVAL_DB).get(id=department_id)

        category = FacultyCategory.objects.filter(id=category_id, is_active=True).first()
        if not category:
            messages.error(request, "Selected category is invalid or inactive.")
            return redirect("add_employee")

        if NewUserAdder.objects.using(APPROVAL_DB).filter(
            Employee_id=employee_id,
            role=role,
            Department=department,
        ).exists():
            messages.error(request, f"Employee {employee_id} is already pre-authorized for this role and department.")
            return redirect("add_employee")

        NewUserAdder.objects.using(APPROVAL_DB).create(
            Employee_id=employee_id,
            role=role,
            Department=department,
        )

        # Persist the chosen category so it is fixed (read-only) when the
        # staff member later completes their general information after signup.
        StaffCategoryAssignment.objects.update_or_create(
            employee_id=employee_id,
            defaults={"category": category},
        )

        # Persist the admin-set date of joining into general information now, so
        # it is already stored (and shown read-only) when the staff member later
        # completes their profile. Requires a numeric employee id.
        if doj is not None:
            if employee_id.isdigit():
                info = general_information.objects.filter(faculty_id=int(employee_id)).first()
                if info:
                    info.doj = doj
                    info.save(update_fields=["doj"])
                else:
                    general_information.objects.create(faculty_id=int(employee_id), doj=doj)
            else:
                messages.warning(request, "Date of joining was not saved: employee ID must be numeric.")

        messages.success(request, f"Employee {employee_id} pre-authorized — they can now sign up with this role and department.")
    except OperationalError as e:
        messages.error(request, f"Approval database error: {e}")
    except Exception as e:
        messages.error(request, f"An error occurred: {e}")

    return redirect("add_employee")



def add_approval_role(request):
    if request.method != "POST":
        return redirect("add_role")

    roles = request.POST.getlist("role[]") or [request.POST.get("role", "")]

    for role_name in roles:
        role_name = (role_name or "").strip()
        if not role_name:
            continue

        try:
            if Role.objects.using(APPROVAL_DB).filter(role__iexact=role_name).exists():
                messages.error(request, f"{role_name} already exists.")
                continue

            Role.objects.using(APPROVAL_DB).create(role=role_name)
            messages.success(request, f"Role '{role_name}' added successfully.")
        except IntegrityError as e:
            messages.error(request, f"Integrity error while adding role '{role_name}': {e}")
        except OperationalError as e:
            messages.error(request, f"Approval database error while adding role '{role_name}': {e}")
        except Exception as e:
            messages.error(request, f"An error occurred while adding role '{role_name}': {e}")

    return redirect("add_role")





def assign_approval_role(request):
    if request.method != "POST":
        return redirect("assign_role_to_employee")

    employee_id = (request.POST.get("employeeID") or "").strip()
    role_id = (request.POST.get("role") or "").strip()
    department_id = (request.POST.get("department") or "").strip()

    try:
        role = Role.objects.using(APPROVAL_DB).get(id=role_id)
        department = Department.objects.using(APPROVAL_DB).get(id=department_id)

        if USER.objects.using(APPROVAL_DB).filter(
            Employee_id=employee_id,
            role=role,
            Department=department,
        ).exists():
            messages.error(request, "This employee already has that role in the selected department.")
            return redirect("assign_role_to_employee")

        source_user = (
            USER.objects.using(APPROVAL_DB)
            .filter(Employee_id=employee_id)
            .order_by("id")
            .first()
        )
        if not source_user:
            messages.error(request, f"No user found with Employee ID: {employee_id}")
            return redirect("assign_role_to_employee")

        unique_id = f"{department.Department_code}{employee_id}{role.role}"
        if USER.objects.using(APPROVAL_DB).filter(unique_id=unique_id).exists():
            messages.error(request, f"Unique ID '{unique_id}' already exists.")
            return redirect("assign_role_to_employee")

        with transaction.atomic(using=APPROVAL_DB):
            USER.objects.using(APPROVAL_DB).create(
                Employee_id=source_user.Employee_id,
                username=source_user.username,
                role=role,
                Department=department,
                unique_id=unique_id,
                email=source_user.email,
                password=source_user.password,
                profile_img=source_user.profile_img.name if source_user.profile_img else None,
                is_active=source_user.is_active,
                is_staff=source_user.is_staff,
                is_superuser=source_user.is_superuser,
                is_student=source_user.is_student,
                is_parent=source_user.is_parent,
                last_login=source_user.last_login,
            )

        _mirror_department_to_local(department.Department, department.Department_code)
        messages.success(request, f"Role '{role.role}' assigned to {source_user.username}.")
    except IntegrityError as e:
        messages.error(request, f"Integrity error: {e}")
    except OperationalError as e:
        messages.error(request, f"Approval database error: {e}")
    except Exception as e:
        messages.error(request, f"An error occurred while assigning role: {e}")

    return redirect("assign_role_to_employee")

@check_permission("add_employee")
def approval_add_staff_page(request):
    roles = Role.objects.using(APPROVAL_DB).all().order_by("role")
    departments = Department.objects.using(APPROVAL_DB).all().order_by("Department")
    categories = FacultyCategory.objects.filter(is_active=True).order_by("category_name")
    return render(request, "faculty_management/admin/approval_add_staff.html", {
        "roles": roles,
        "departments": departments,
        "categories": categories,
    })

@check_permission("add_role")
def approval_roles_page(request):
    roles = Role.objects.using(APPROVAL_DB).all().order_by("role")
    return render(request, "faculty_management/admin/approval_roles.html", {
        "roles": roles,
        "role_count": Role.objects.using(APPROVAL_DB).count(),
    })

@check_permission("assign_role_to_employee")
def approval_assign_role_to_employee_page(request):
    roles = Role.objects.using(APPROVAL_DB).all().order_by("role")
    departments = Department.objects.using(APPROVAL_DB).all().order_by("Department")
    return render(request, "faculty_management/admin/approval_assign_role_to_employee.html", {
        "roles": roles,
        "departments": departments,
    })



def faculty_category_master(request):
    if request.method == "POST":
        action = request.POST.get("action")
        category_name = request.POST.get("category_name", "").strip()
        is_active = request.POST.get("is_active") == "on"

        if action == "add":
            if not category_name:
                messages.error(request, "Category name is required.")
                return redirect("faculty_category_master")

            if FacultyCategory.objects.filter(category_name__iexact=category_name).exists():
                messages.error(request, "Category already exists.")
                return redirect("faculty_category_master")

            FacultyCategory.objects.create(category_name=category_name, is_active=is_active)
            messages.success(request, "Category added successfully!")
            return redirect("faculty_category_master")

        if action == "update":
            category_id = request.POST.get("category_id")
            category = get_object_or_404(FacultyCategory, id=category_id)

            if not category_name:
                messages.error(request, "Category name is required.")
                return redirect("faculty_category_master")

            if FacultyCategory.objects.filter(category_name__iexact=category_name).exclude(id=category_id).exists():
                messages.error(request, "Another category with this name already exists.")
                return redirect("faculty_category_master")

            category.category_name = category_name
            category.is_active = is_active
            category.save()
            messages.success(request, "Category updated successfully!")
            return redirect("faculty_category_master")

        if action == "delete":
            category_id = request.POST.get("category_id")
            category = get_object_or_404(FacultyCategory, id=category_id)
            category.delete()
            messages.success(request, "Category deleted successfully!")
            return redirect("faculty_category_master")

    search_query = request.GET.get("search", "").strip()
    categories = FacultyCategory.objects.all().order_by("category_name")
    if search_query:
        categories = categories.filter(category_name__icontains=search_query)

    paginator = Paginator(categories, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "faculty_management/admin/faculty_category_master.html", {
        "page_obj": page_obj,
        "search_query": search_query,
    })

from django.db.models import Count, Q


from django.db.models import *

from faculty_management.models import general_information, DesignationMaster, Faculty_Data_Permission
from user_accounts.models import Add_Department


# @is_super_user('faculty_management')
# @no_cache
# @check_permission("employee_details")
# def admin_faculty_details(request):
#     selected_dept = (request.GET.get('department') or '').strip()
#     selected_desig = (request.GET.get('designation') or '').strip()

#     # ---------- PERMISSION ----------
#     role_id = request.user.role_id

#     # print("Logged in user role_id => ", role_id)
#     perm = Faculty_Data_Permission.objects.filter(role_id=role_id).first()
#     # print("Permission for user => ", perm)
#     base_qs = general_information.objects.all()

#     # detect logged-in user's department id (supports different field names)
#     # user_dept_id = (
#     #     getattr(request.user, "Department_id", None)
#     #     or getattr(request.user, "department_id", None)
#     #     or getattr(getattr(request.user, "Department", None), "id", None)
#     # )
    

#     if request.user.is_superuser:
#         user_dept_id = None  # Superuser can see all departments, so ignore user's department
#     else:
#         faculty = general_information.objects.filter(faculty_id=request.user.Employee_id).first()

#         user_dept_id = faculty.department.id if faculty else None
#     if perm and perm.can_view_all_faculty_data:
#         # no restriction
#         pass

#     elif perm and perm.can_view_department_faculty_data:
#         # restrict to user's department only
#         if user_dept_id:
#             base_qs = base_qs.filter(department_id=user_dept_id)
#         else:
#             base_qs = base_qs.none()

#         # prevent overriding department filter
#         if selected_dept and str(selected_dept) != str(user_dept_id):
#             selected_dept = str(user_dept_id) if user_dept_id else ""

#     else:
#         # no permission -> show nothing
#         base_qs = base_qs.none()
#         selected_dept = ""
#         selected_desig = ""

#     # ---------- APPLY FILTERS ----------
#     faculty_list = base_qs.order_by('name')

#     if selected_dept:
#         faculty_list = faculty_list.filter(department_id=selected_dept)

#     if selected_desig:
#         faculty_list = faculty_list.filter(designation_id=selected_desig)

#     # ---------- SIDE LISTS ----------
#     # departments list should also respect permission
#     departments = Add_Department.objects.all().order_by("Department")
#     if perm and perm.can_view_department_faculty_data and user_dept_id:
#         departments = departments.filter(id=user_dept_id)

#     designations = DesignationMaster.objects.all().annotate(
#         filtered_count=Count('general_information', filter=Q(general_information__in=faculty_list))
#     )

#     total_faculty = faculty_list.count()

#     context = {
#         'faculty_list': faculty_list,
#         'departments': departments,
#         'designations': designations,
#         'selected_dept': selected_dept,
#         'selected_desig': selected_desig,
#         'total_faculty': total_faculty,
#     }

#     # IMPORTANT:
#     # For AJAX, return SAME TEMPLATE (full HTML).
#     # JS will extract #facultyContent and replace.
#     return render(request, 'faculty_management/admin/admin_faculty_details.html', context)


from django.core.paginator import Paginator
from django.db.models import Count, Q

@no_cache
@check_permission("employee_details")
def admin_faculty_details(request):

    selected_dept = (request.GET.get('department') or '').strip()
    selected_desig = (request.GET.get('designation') or '').strip()
    search_query = (request.GET.get('search') or '').strip()
    page_number = request.GET.get('page', 1)

    role_id = request.user.role_id
    perm = Faculty_Data_Permission.objects.filter(role_id=role_id).first()
    base_qs = general_information.objects.select_related(
        'department', 'designation', 'category', 'shift'
    ).all().order_by('faculty_id')

    # ---------- Department Restriction ----------
    if request.user.is_superuser:
        user_dept_id = None
    else:
        faculty = general_information.objects.filter(
            faculty_id=request.user.Employee_id
        ).first()
        user_dept_id = faculty.department.id if faculty else None

    if perm and perm.can_view_all_faculty_data:
        pass

    elif perm and perm.can_view_department_faculty_data:
        if user_dept_id:
            base_qs = base_qs.filter(department_id=user_dept_id)
        else:
            base_qs = base_qs.none()

        if selected_dept and str(selected_dept) != str(user_dept_id):
            selected_dept = str(user_dept_id) if user_dept_id else ""

    else:
        base_qs = base_qs.none()
        selected_dept = ""
        selected_desig = ""
        search_query = ""

    # ---------- Filters ----------
    if selected_dept:
        base_qs = base_qs.filter(department_id=selected_dept)

    if selected_desig:
        base_qs = base_qs.filter(designation_id=selected_desig)

    # ---------- Global Search ----------
    if search_query:
        base_qs = base_qs.filter(
            Q(name__icontains=search_query) |
            Q(faculty_id__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(college_email__icontains=search_query) |
            Q(department__Department__icontains=search_query) |
            Q(designation__designation_name__icontains=search_query)
        )

    faculty_qs = base_qs.order_by('faculty_id')

    # ---------- Pagination ----------
    paginator = Paginator(faculty_qs, 50)  # 10 per page
    page_obj = paginator.get_page(page_number)

    # ---------- Side Lists ----------
    departments = Add_Department.objects.all().order_by("Department")
    if perm and perm.can_view_department_faculty_data and user_dept_id:
        departments = departments.filter(id=user_dept_id)

    designations = DesignationMaster.objects.all().annotate(
        filtered_count=Count(
            'general_information',
            filter=Q(general_information__in=faculty_qs)
        )
    )

    context = {
        'faculty_list': page_obj,
        'page_obj': page_obj,
        'departments': departments,
        'designations': designations,
        'selected_dept': selected_dept,
        'selected_desig': selected_desig,
        'search_query': search_query,
        'total_faculty': faculty_qs.count(),
    }

    return render(
        request,
        'faculty_management/admin/admin_faculty_details.html',
        context
    )


from user_accounts.models import general_information as SourceFaculty
from faculty_management.models import general_information as TargetFaculty
from user_accounts.models import USER

def sync_faculty_data(request):

    if not request.user.is_superuser:
        return redirect("employee_details")

    source_data = SourceFaculty.objects.using("rit_academic_system").all()

    created = 0
    updated = 0
    skipped = 0

    for obj in source_data:

        if not obj.faculty_id:
            skipped += 1
            continue

        # ---------------- Department Mapping ----------------
        department_obj = None
        if obj.department:
            department_obj = Add_Department.objects.filter(
                Department__iexact=obj.department.strip()
            ).first()

        # ---------------- Designation Mapping ----------------
        designation_obj = None
        if obj.designation:
            designation_obj = DesignationMaster.objects.filter(
                designation_name__iexact=obj.designation.strip()
            ).first()

        category_obj = None
        if designation_obj:
            category_obj = FacultyCategory.objects.filter(
                category_name__iexact="Teaching" if designation_obj.is_teaching else "Non Teaching"
            ).first()

        # ---------------- Get Profile Photo ----------------
        user_obj = USER.objects.using("rit_approval_system").filter(Employee_id=str(obj.faculty_id)).first()
        profile_photo = None

        if user_obj and user_obj.profile_img:
            profile_photo = user_obj.profile_img

        # ---------------- Check Existing Faculty ----------------
        faculty = TargetFaculty.objects.filter(
            faculty_id=obj.faculty_id
        ).first()

        # ---------------- CREATE ----------------
        if not faculty:

            TargetFaculty.objects.create(
                faculty_id=obj.faculty_id,
                name=obj.name,
                department=department_obj,
                designation=designation_obj,
                category=category_obj,
                dob=obj.dob,
                address=obj.address,
                personal_email=obj.personal_email,
                college_email=obj.college_email,
                phone=obj.phone,
                blood_group=obj.blood_group,
                community=obj.community,
                caste=obj.caste,
                religion=obj.religion,
                doj=obj.doj,
                apaar_id=obj.apaar_id,
                anu_id=obj.anu_id,
                aicte_id=obj.aicte_id,
                annauniversity_affiliation_id=obj.annauniversity_affiliation_id,
                PAN_number=obj.PAN_number,
                Aadhar_number=obj.Aadhar_number,
                appointment_type=obj.appointment_type,
                basic_pay=obj.basic_pay,
                agp=obj.agp,
                allowances=obj.allowances,
                pay_scale_notes=obj.pay_scale_notes,
                recruitment_mode=obj.recruitment_mode,
                nature_of_duties=obj.nature_of_duties,
                confirmation_date=obj.confirmation_date,
                probation_period_months=obj.probation_period_months,
                probation_confirmation_reference=obj.probation_confirmation_reference,
                approval=obj.approval,
                profile_photo=profile_photo
            )

            created += 1

        # ---------------- UPDATE ONLY MISSING ----------------
        else:

            changed = False

            fields = {
                "name": obj.name,
                "department": department_obj,
                "designation": designation_obj,
                "category": category_obj,
                "dob": obj.dob,
                "address": obj.address,
                "personal_email": obj.personal_email,
                "college_email": obj.college_email,
                "phone": obj.phone,
                "blood_group": obj.blood_group,
                "community": obj.community,
                "caste": obj.caste,
                "religion": obj.religion,
                "doj": obj.doj,
                "apaar_id": obj.apaar_id,
                "anu_id": obj.anu_id,
                "aicte_id": obj.aicte_id,
                "annauniversity_affiliation_id": obj.annauniversity_affiliation_id,
                "PAN_number": obj.PAN_number,
                "Aadhar_number": obj.Aadhar_number,
                "appointment_type": obj.appointment_type,
                "basic_pay": obj.basic_pay,
                "agp": obj.agp,
                "allowances": obj.allowances,
                "pay_scale_notes": obj.pay_scale_notes,
                "recruitment_mode": obj.recruitment_mode,
                "nature_of_duties": obj.nature_of_duties,
                "confirmation_date": obj.confirmation_date,
                "probation_period_months": obj.probation_period_months,
                "probation_confirmation_reference": obj.probation_confirmation_reference,
            }

            for field, value in fields.items():

                current_value = getattr(faculty, field)

                if (current_value is None or current_value == "") and value:
                    setattr(faculty, field, value)
                    changed = True

            # ---------- PROFILE SYNC ----------
            if not faculty.profile_photo and profile_photo:
                faculty.profile_photo = profile_photo
                changed = True

            if changed:
                faculty.save()
                updated += 1
            else:
                skipped += 1

    messages.success(
        request,
        f"Sync Completed → Created: {created}, Updated Missing: {updated}, Skipped: {skipped}"
    )

    return redirect("employee_details")
# def sync_faculty_data(request):

#     if not request.user.is_superuser:
#         return redirect("employee_details")

#     source_data = SourceFaculty.objects.using("rit_academic_system").all()

#     created = 0
#     updated = 0
#     skipped = 0
#     missing_ids = []

#     # print("\n=========== START FACULTY SYNC ===========")

#     for obj in source_data:

#         # print("\n---------------------------------------")
#         # print("Processing Employee ID:", obj.faculty_id)

#         if not obj.faculty_id:
#             # print("❌ Skipped → faculty_id missing")
#             skipped += 1
#             continue

#         # ---------------- Department Mapping ----------------
#         department_obj = None
#         if obj.department:
#             department_obj = Add_Department.objects.filter(
#                 Department__iexact=obj.department.strip()
#             ).first()

#         # print("SOURCE Department:", obj.department)
#         # print("MAPPED Department:", department_obj)

#         # ---------------- Designation Mapping ----------------
#         designation_obj = None
#         if obj.designation:
#             designation_obj = DesignationMaster.objects.filter(
#                 designation_name__iexact=obj.designation.strip()
#             ).first()

#         # print("SOURCE Designation:", obj.designation)
#         # print("MAPPED Designation:", designation_obj)

#         # ---------------- Check Existing Faculty ----------------
#         faculty = TargetFaculty.objects.filter(
#             faculty_id=obj.faculty_id
#         ).first()

#         if faculty:
#             # print("✔ Faculty already exists in TARGET DB:", obj.faculty_id)
#         else:
#             # print("❌ Faculty NOT found in TARGET DB:", obj.faculty_id)
#             missing_ids.append(obj.faculty_id)

#         # ---------------- CREATE ----------------
#         if not faculty:

#             # print("➡ Creating new faculty record")

#             TargetFaculty.objects.create(
#                 faculty_id=obj.faculty_id,
#                 name=obj.name,
#                 department=department_obj,
#                 designation=designation_obj,
#                 dob=obj.dob,
#                 address=obj.address,
#                 personal_email=obj.personal_email,
#                 college_email=obj.college_email,
#                 phone=obj.phone,
#                 blood_group=obj.blood_group,
#                 community=obj.community,
#                 caste=obj.caste,
#                 religion=obj.religion,
#                 doj=obj.doj,
#                 apaar_id=obj.apaar_id,
#                 anu_id=obj.anu_id,
#                 aicte_id=obj.aicte_id,
#                 annauniversity_affiliation_id=obj.annauniversity_affiliation_id,
#                 PAN_number=obj.PAN_number,
#                 Aadhar_number=obj.Aadhar_number,
#                 appointment_type=obj.appointment_type,
#                 basic_pay=obj.basic_pay,
#                 agp=obj.agp,
#                 allowances=obj.allowances,
#                 pay_scale_notes=obj.pay_scale_notes,
#                 recruitment_mode=obj.recruitment_mode,
#                 nature_of_duties=obj.nature_of_duties,
#                 confirmation_date=obj.confirmation_date,
#                 probation_period_months=obj.probation_period_months,
#                 probation_confirmation_reference=obj.probation_confirmation_reference,
#                 approval=obj.approval,
#             )

#             created += 1
#             # print("✅ Faculty created")

#         # ---------------- UPDATE ONLY MISSING ----------------
#         else:

#             # print("➡ Checking for missing fields")

#             changed = False

#             fields = {
#                 "name": obj.name,
#                 "department": department_obj,
#                 "designation": designation_obj,
#                 "dob": obj.dob,
#                 "address": obj.address,
#                 "personal_email": obj.personal_email,
#                 "college_email": obj.college_email,
#                 "phone": obj.phone,
#                 "blood_group": obj.blood_group,
#                 "community": obj.community,
#                 "caste": obj.caste,
#                 "religion": obj.religion,
#                 "doj": obj.doj,
#                 "apaar_id": obj.apaar_id,
#                 "anu_id": obj.anu_id,
#                 "aicte_id": obj.aicte_id,
#                 "annauniversity_affiliation_id": obj.annauniversity_affiliation_id,
#                 "PAN_number": obj.PAN_number,
#                 "Aadhar_number": obj.Aadhar_number,
#                 "appointment_type": obj.appointment_type,
#                 "basic_pay": obj.basic_pay,
#                 "agp": obj.agp,
#                 "allowances": obj.allowances,
#                 "pay_scale_notes": obj.pay_scale_notes,
#                 "recruitment_mode": obj.recruitment_mode,
#                 "nature_of_duties": obj.nature_of_duties,
#                 "confirmation_date": obj.confirmation_date,
#                 "probation_period_months": obj.probation_period_months,
#                 "probation_confirmation_reference": obj.probation_confirmation_reference,
#             }

#             for field, value in fields.items():

#                 current_value = getattr(faculty, field)

#                 if (current_value is None or current_value == "") and value:
#                     # print(f"✔ Updating field {field} → {value}")
#                     setattr(faculty, field, value)
#                     changed = True

#             if changed:
#                 faculty.save()
#                 updated += 1
#                 # print("✅ Faculty updated")
#             else:
#                 skipped += 1
#                 # print("⚠ No update required")

#     # print("\n=========== SYNC SUMMARY ===========")
#     # print("Created:", created)
#     # print("Updated:", updated)
#     # print("Skipped:", skipped)
#     # print("Missing Faculty IDs:", missing_ids)

#     messages.success(
#         request,
#         f"Sync Completed → Created: {created}, Updated Missing: {updated}, Skipped: {skipped}"
#     )

#     return redirect("employee_details")


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from faculty_management.models import FacultyCategory, general_information, DesignationMaster
from faculty_leave_management.models import ShiftMaster
from user_accounts.models import Add_Department

def admin_edit_faculty_details(request, faculty_id):
    faculty = get_object_or_404(general_information, id=faculty_id)
    designations = DesignationMaster.objects.all().order_by("designation_name")
    categories = FacultyCategory.objects.filter(is_active=True).order_by("category_name")
    departments = Add_Department.objects.all().order_by("Department")
    shifts = ShiftMaster.objects.filter(is_active=True).order_by("shift_name")

    if request.method == "POST":
        try:
            faculty.name = request.POST.get("name")
            faculty.department_id = request.POST.get("department")
            faculty.designation_id = request.POST.get("designation")
            faculty.category_id = request.POST.get("category") or None
            faculty.shift_id = request.POST.get("shift") or None
            faculty.personal_email = request.POST.get("personal_email")
            faculty.college_email = request.POST.get("college_email")
            faculty.address = request.POST.get("address")
            faculty.phone = request.POST.get("phone") or None
            faculty.dob = request.POST.get("dob") or None
            faculty.doj = request.POST.get("doj") or None

            # Upload new profile image
            if request.FILES.get("profile_photo"):
                faculty.profile_photo = request.FILES.get("profile_photo")

            # Remove profile image
            if request.POST.get("remove_photo") == "1":
                faculty.profile_photo.delete(save=False)
                faculty.profile_photo = None

            faculty.save()

            messages.success(request, "Faculty details updated successfully!")
            return redirect("employee_details")

        except Exception as e:
            messages.error(request, f"Error updating faculty details: {e}")

    return render(request, "faculty_management/admin/admin_edit_faculty_details.html", {
        "faculty": faculty,
        "designations": designations,
        "categories": categories,
        "departments": departments,
        "shifts": shifts,
    })


def admin_view_faculty_detail(request, faculty_id):
    faculty = get_object_or_404(general_information, id=faculty_id)
    academic_bg = faculty.academic_background_set.all().order_by("degree")
    academic_exp = faculty.academicexperience_set.all().order_by("-from_date")
    industry_exp = faculty.industryexperience_set.all().order_by("-from_date")
    research_exp = faculty.researchexperience_set.all().order_by("-from_date")
    return render(request, "faculty_management/admin/admin_view_faculty_detail.html", {
        "faculty": faculty,
        "academic_bg": academic_bg,
        "academic_exp": academic_exp,
        "industry_exp": industry_exp,
        "research_exp": research_exp,
    })


from django.http import JsonResponse

def faculty_popup_detail(request, faculty_id):
    from django.utils.dateformat import format as dfmt
    faculty = get_object_or_404(general_information, id=faculty_id)

    def fmt_date(d):
        return d.strftime("%d %b %Y") if d else None

    photo_url = None
    if faculty.profile_photo:
        try:
            photo_url = faculty.profile_photo.url
        except Exception:
            photo_url = None

    data = {
        "id":               faculty.id,
        "name":             faculty.name or "",
        "faculty_id":       faculty.faculty_id,
        "department":       faculty.department.Department if faculty.department else None,
        "designation":      faculty.designation.designation_name if faculty.designation else None,
        "gender":           faculty.gender,
        "dob":              fmt_date(faculty.dob),
        "blood_group":      faculty.blood_group,
        "community":        faculty.community,
        "religion":         faculty.religion,
        "phone":            str(faculty.phone) if faculty.phone else None,
        "college_email":    faculty.college_email,
        "personal_email":   faculty.personal_email,
        "address":          faculty.address,
        "doj":              fmt_date(faculty.doj),
        "appointment_type": faculty.appointment_type,
        "recruitment_mode": faculty.recruitment_mode,
        "nature_of_duties": faculty.nature_of_duties,
        "basic_pay":        str(faculty.basic_pay) if faculty.basic_pay else None,
        "approval":         faculty.approval,
        "photo_url":        photo_url,
    }
    return JsonResponse(data)


# from django.views.generic import CreateView, UpdateView, ListView,DeleteView
# from django.urls import reverse_lazy
# import json
# from datetime import date
# from openpyxl import Workbook
# from django.utils.decorators import method_decorator    
# from django.contrib.auth.decorators import login_required
# from faculty_management.models import Designation   
# from faculty_management.forms import DesignationForm
# from control_room.decorators import no_cache,is_super_user
# from django.db.models import Count
# from faculty_management.models import Faculty
# from control_room.models import Department
# from django.db import connection

# import pandas as pd
# from io import BytesIO


# @method_decorator(login_required, name='dispatch')
# @method_decorator(no_cache, name='dispatch')
# @method_decorator(is_super_user('faculty_management'), name='dispatch')
# class DesignationCreateView(CreateView):
#     model=Designation
#     form_class=DesignationForm
#     context_object_name='designations'
#     template_name="faculty_management/admin/designation_form.html"
#     success_url = reverse_lazy('faculty_management')
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['create_or_update'] = True  # Adding extra context variable
#         return context
    
#     def form_valid(self, form):
#         response = super().form_valid(form)
#         messages.success(self.request, "Designation created successfully!")  # Success message
#         return response
#     def form_invalid(self, form):
#         response = super().form_invalid(form)
#         for field, errors in form.errors.items():
#             for error in errors:
#                 messages.error(self.request, f"Error in {field}: {error}")
                
# @method_decorator(login_required, name='dispatch')
# @method_decorator(no_cache, name='dispatch')
# @method_decorator(is_super_user('faculty_management'), name='dispatch')
# class DesignationUpdateView(UpdateView):
#     model=Designation
#     form_class=DesignationForm
#     context_object_name='designations'
#     template_name="faculty_management/admin/designation_form.html"
#     success_url = reverse_lazy('faculty_management')
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['create_or_update'] = True  # Adding extra context variable
#         return context
    
#     def form_valid(self, form):
#         response = super().form_valid(form)
#         messages.success(self.request, "Designation created successfully!")  # Success message
#         return response
#     def form_invalid(self, form):
#         response = super().form_invalid(form)
#         for field, errors in form.errors.items():
#             for error in errors:
#                 messages.error(self.request, f"Error in {field}: {error}")
# @method_decorator(login_required, name='dispatch')
# @method_decorator(no_cache, name='dispatch')
# @method_decorator(is_super_user('faculty_management'), name='dispatch')
# class DesignationListView(ListView):
#     model=Designation
#     form_class=DesignationForm
#     context_object_name='designations'
#     template_name="faculty_management/admin/designation_form.html"
#     success_url = reverse_lazy('faculty_management')
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['list'] = True  # Adding extra context variable
#         return context
# @method_decorator(login_required, name='dispatch')
# @method_decorator(no_cache, name='dispatch')
# @method_decorator(is_super_user('faculty_management'), name='dispatch')
# class DesignationDeleteView(DeleteView):
#     model = Designation
#     context_object_name = 'designation'
#     template_name = "faculty_management/admin/designation_form.html"
#     success_url = reverse_lazy('faculty_management')

#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
#         context['delete'] = True  # Adding extra context variable
#         return context

#     def delete(self, request, *args, **kwargs):
#         messages.success(self.request, "Designation deleted successfully.")
#         return super().delete(request, *args, **kwargs)
    
    
    
    
# from django.forms.models import model_to_dict
# from django.db.models.fields.files import FieldFile


# @login_required
# @no_cache
# @is_super_user('faculty_management')   
# def faculty_filter(request):
#     depts = Department.objects.all()
#     datas = Faculty.objects.all()
#     total_faculty = datas.count()
#     designation_counts = datas.values('present_designation__name').annotate(count=Count('present_designation'))
#     # print("designation count -> ",designation_counts)

#     table_name = "faculty_leave_management_faculty"
#     with connection.cursor() as cursor:
#         cursor.execute(f"SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '{table_name}'")
#         all_columns = [row[0] for row in cursor.fetchall()]

#     filtered_faculty = datas
#     department_name = None
#     filtered_designation_counts = []

#     if request.method == 'POST':
#         department = request.POST.get('department')
#         column_type = request.POST.get('Column_type')
#         selected_columns = request.POST.getlist('custom_columns')

#         filters = {}
#         if department:
#             filters['department'] = department
#             department_name = Department.objects.get(id=department).Department

#         if column_type == "all":
#             selected_columns = all_columns
#         elif column_type == "customised_Column" and not selected_columns:
#             selected_columns = all_columns

#         if selected_columns:
#             filtered_faculty = Faculty.objects.filter(**filters).values('id', *selected_columns)
#         else:
#             filtered_faculty = Faculty.objects.filter(**filters)

#         filtered_faculty_1 = Faculty.objects.filter(department=department)
#         department_name = Department.objects.get(id=department).Department

#         # Convert queryset to serializable dicts
#         faculty_list = []
#         for faculty in Faculty.objects.filter(**filters):
#             item = model_to_dict(faculty, fields=selected_columns if selected_columns else None)
#             for key, value in item.items():
#                 # Convert date fields to string
#                 if isinstance(value, date):
#                     item[key] = value.strftime('%Y-%m-%d')

#                 # Convert file/image fields to URL or path
#                 elif isinstance(value, FieldFile):
#                     item[key] = str(value.url) if value and hasattr(value, 'url') else str(value)

#                 # Handle foreign keys (replace ID with string)
#                 elif isinstance(value, int):
#                     try:
#                         related_field = Faculty._meta.get_field(key)
#                         if related_field.is_relation:
#                             related_model = related_field.related_model
#                             related_object = related_model.objects.filter(id=value).first()
#                             if related_object:
#                                 item[key] = str(related_object)
#                     except Exception:
#                         continue
#             faculty_list.append(item)

#         # Store data in session (must be fully serializable)
#         request.session['filtered_faculty'] = json.dumps(faculty_list)
#         request.session['selected_columns'] = selected_columns

#         filtered_designation_counts = (
#             filtered_faculty_1.values('present_designation__name')
#             .annotate(count=Count('present_designation'))
#         )

#         return render(request, 'faculty_management/admin/faculty_filter.html', {
#             'data': filtered_faculty,
#             'depts': depts,
#             'total_faculty': total_faculty,
#             'designation_counts': designation_counts,
#             'department_name': department_name,
#             'filtered_designation_counts': filtered_designation_counts,
#             'filtered_faculty_1': filtered_faculty_1,
#             'columns': all_columns
#         })

#     return render(request, 'faculty_management/admin/faculty_filter.html', {
#         'data': datas,
#         'depts': depts,
#         'total_faculty': total_faculty,
#         'columns': all_columns,
#         'designation_counts': designation_counts,
#         'filtered_faculty': filtered_faculty,
#     })
    
    
    
    
    
# @login_required
# @no_cache
# @is_super_user('faculty_management')
# def export_faculty_excel(request):
#     # Retrieve filtered data from session
#     filtered_faculty_json = request.session.get('filtered_faculty')
#     selected_columns = request.session.get('selected_columns', [])

#     if not filtered_faculty_json or not selected_columns:
#         return HttpResponse("No data available to export.", status=400)

#     filtered_faculty = json.loads(filtered_faculty_json)

#     # Create a new Excel workbook and sheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Faculty Data"

#     # Write header row
#     ws.append(selected_columns)

#     # Write data rows
#     for row in filtered_faculty:
#         ws.append([row.get(col, '') for col in selected_columns])

#     # Prepare response
#     del request.session['filtered_faculty']
#     del request.session['selected_columns']

#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename="faculty_data.xlsx"'

#     # Save workbook to response
#     wb.save(response)

#     return response

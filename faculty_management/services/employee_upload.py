import re
from datetime import date, datetime

import openpyxl
from django.contrib.auth.hashers import make_password
from django.db import IntegrityError, transaction
from django.http import HttpResponse
from openpyxl.utils.datetime import from_excel

from faculty_management.models import (
    DesignationMaster,
    FacultyCategory,
    StaffCategoryAssignment,
    general_information,
)
from faculty_leave_management.models import ShiftMaster
from user_accounts.models import Add_Department, Department, NewUserAdder, Role, USER


APPROVAL_DB = "rit_approval_system"

REQUIRED_COLUMNS = {
    "employee_id": "Faculty ID / Employee ID",
    "name": "Name / Username",
    "department": "Department",
    "role": "Role",
    "designation": "Designation",
    "category": "Category",
}

OPTIONAL_COLUMNS = {
    "gender": "Gender",
    "dob": "DOB / Date of Birth",
    "address": "Address",
    "personal_email": "Personal Email",
    "college_email": "College Email",
    "email": "Mail / Email (legacy college email)",
    "phone": "Phone Number",
    "blood_group": "Blood Group",
    "community": "Community",
    "caste": "Caste",
    "religion": "Religion",
    "doj": "DOJ / Date of Joining",
    "PAN_number": "PAN Number",
    "Aadhar_number": "Aadhar Number",
    "shift": "Shift",
}

HEADER_ALIASES = {
    "facultyid": "employee_id",
    "faculty_id": "employee_id",
    "employeeid": "employee_id",
    "employee_id": "employee_id",
    "employeid": "employee_id",
    "employe_id": "employee_id",
    "empid": "employee_id",
    "emp_id": "employee_id",
    "id": "employee_id",
    "name": "name",
    "username": "name",
    "user_name": "name",
    "employeename": "name",
    "employee_name": "name",
    "facultyname": "name",
    "faculty_name": "name",
    "department": "department",
    "dept": "department",
    "departmentname": "department",
    "department_name": "department",
    "role": "role",
    "userrole": "role",
    "user_role": "role",
    "designation": "designation",
    "desinination": "designation",
    "desigination": "designation",
    "desgination": "designation",
    "designationname": "designation",
    "designation_name": "designation",
    "category": "category",
    "facultycategory": "category",
    "faculty_category": "category",
    "staffcategory": "category",
    "staff_category": "category",
    "gender": "gender",
    "sex": "gender",
    "mail": "email",
    "email": "email",
    "emailid": "email",
    "email_id": "email",
    "collegeemail": "email",
    "college_email": "email",
    "officialmail": "email",
    "official_mail": "email",
    "officialemail": "email",
    "official_email": "email",
    "personalemail": "personal_email",
    "personal_email": "personal_email",
    "personalmail": "personal_email",
    "personal_mail": "personal_email",
    "college_mail": "college_email",
    "college_email": "college_email",
    "institutionalemail": "college_email",
    "institutional_email": "college_email",
    "institutionalmail": "college_email",
    "institutional_mail": "college_email",
    "shift": "shift",
    "shiftname": "shift",
    "shift_name": "shift",
    "dateofjoining": "doj",
    "date_of_joining": "doj",
    "joiningdate": "doj",
    "joining_date": "doj",
    "doj": "doj",
    "dateofbirth": "dob",
    "date_of_birth": "dob",
    "birthdate": "dob",
    "birth_date": "dob",
    "dob": "dob",
    "address": "address",
    "residentialaddress": "address",
    "residential_address": "address",
    "phone": "phone",
    "phonenumber": "phone",
    "phone_number": "phone",
    "mobile": "phone",
    "mobilenumber": "phone",
    "mobile_number": "phone",
    "contact": "phone",
    "contactnumber": "phone",
    "contact_number": "phone",
    "bloodgroup": "blood_group",
    "blood_group": "blood_group",
    "community": "community",
    "caste": "caste",
    "religion": "religion",
    "pannumber": "PAN_number",
    "pan_number": "PAN_number",
    "pan": "PAN_number",
    "panno": "PAN_number",
    "pan_no": "PAN_number",
    "aadharnumber": "Aadhar_number",
    "aadhaarnumber": "Aadhar_number",
    "aadhar_number": "Aadhar_number",
    "aadhaar_number": "Aadhar_number",
    "aadhar": "Aadhar_number",
    "aadhaar": "Aadhar_number",
    "aadharno": "Aadhar_number",
    "aadhaarno": "Aadhar_number",
    "aadhar_no": "Aadhar_number",
    "aadhaar_no": "Aadhar_number",
}


def upload_context(results=None, summary=None):
    return {
        "required_columns": REQUIRED_COLUMNS.values(),
        "optional_columns": OPTIONAL_COLUMNS.values(),
        "results": results or [],
        "summary": summary,
    }


def build_template_response():
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Employee Upload"
    reference_sheet = workbook.create_sheet("Reference Values")

    worksheet.append([
        "Faculty ID",
        "Name",
        "Department",
        "Role",
        "Designation",
        "Category",
        "Gender",
        "DOB",
        "Address",
        "Personal Email",
        "College Email",
        "Phone Number",
        "Blood Group",
        "Community",
        "Caste",
        "Religion",
        "DOJ",
        "PAN Number",
        "Aadhar Number",
        "Shift",
    ])
    worksheet.append([
        "1001",
        "Sample Faculty",
        "Use exact department from Sheet 2",
        "Use exact role from Sheet 2",
        "Use exact designation from Sheet 2",
        "Use exact category from Sheet 2",
        "Male",
        "1990-01-15",
        "Sample address",
        "sample.personal@example.com",
        "sample@ritrjpm.ac.in",
        "9876543210",
        "O+",
        "BC",
        "Sample caste",
        "Sample religion",
        "2024-06-01",
        "ABCDE1234F",
        "123456789012",
        "Use exact shift from Sheet 2",
    ])
    _style_sheet(worksheet, [16, 28, 36, 18, 28, 18, 16, 16, 34, 30, 30, 18, 16, 18, 18, 18, 16, 18, 20, 24])

    reference_sheet.append(["Departments", "Roles", "Designations", "Categories", "Shifts"])
    for cell in reference_sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    _append_reference_values(reference_sheet)
    for index, width in enumerate([38, 24, 34, 24, 28], start=1):
        reference_sheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="employee_upload_template.xlsx"'
    workbook.save(response)
    return response


def process_upload_file(excel_file):
    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as exc:
        raise ValueError(f"Unable to read Excel file: {exc}") from exc

    return _process_worksheet(workbook.active)


def sync_department_tables():
    summary = {
        "created_academic": 0,
        "created_control": 0,
        "existing_academic": 0,
        "existing_control": 0,
        "skipped": 0,
        "messages": [],
    }

    control_departments = list(Department.objects.using(APPROVAL_DB).all().order_by("Department"))
    local_departments = list(Add_Department.objects.all().order_by("Department"))
    local_names = {_department_key(dept.Department): dept for dept in local_departments if dept.Department}
    control_names = {_department_key(dept.Department): dept for dept in control_departments if dept.Department}

    for control_dept in control_departments:
        key = _department_key(control_dept.Department)
        if not key:
            continue
        if key in local_names:
            summary["existing_academic"] += 1
            _activate_local_department(local_names[key])
            continue
        Add_Department.objects.create(
            Department=control_dept.Department,
            Department_code=control_dept.Department_code,
            is_active=True,
        )
        summary["created_academic"] += 1

    for local_dept in local_departments:
        key = _department_key(local_dept.Department)
        if not key:
            continue
        # Existence is decided by department name only. The code is carried over
        # as-is and never used to skip a record, so matching names are never
        # duplicated regardless of their codes.
        if key in control_names:
            summary["existing_control"] += 1
            continue
        Department.objects.using(APPROVAL_DB).create(
            Department=local_dept.Department,
            Department_code=local_dept.Department_code or "",
        )
        summary["created_control"] += 1

    return summary


def find_missing_control_users():
    """Return the faculty present in ``general_information`` whose Employee ID is
    NOT present in the control-room USER table (rit_approval_system).

    Nothing is created — this is a read-only report."""
    summary = {
        "total": 0,
        "present": 0,
        "missing": 0,
        "missing_rows": [],
    }

    faculties = (
        general_information.objects
        .select_related("department", "designation")
        .exclude(faculty_id__isnull=True)
        .order_by("faculty_id")
    )

    existing_ids = set(
        str(eid) for eid in USER.objects.using(APPROVAL_DB).values_list("Employee_id", flat=True)
    )
    adder_ids = set(
        str(eid) for eid in NewUserAdder.objects.using(APPROVAL_DB).values_list("Employee_id", flat=True)
    )
    control_dept_names = {
        (name or "").strip().lower()
        for name in Department.objects.using(APPROVAL_DB).values_list("Department", flat=True)
    }

    for faculty in faculties:
        summary["total"] += 1
        employee_id = str(faculty.faculty_id)

        if employee_id in existing_ids:
            summary["present"] += 1
            continue

        dept_name = getattr(faculty.department, "Department", "") or ""
        summary["missing"] += 1
        summary["missing_rows"].append({
            "employee_id": employee_id,
            "name": faculty.name or "",
            "department": dept_name,
            "designation": getattr(faculty.designation, "designation_name", "") or "",
            "reason": _missing_user_reason(employee_id, dept_name, control_dept_names, adder_ids),
        })

    return summary


def _missing_user_reason(employee_id, dept_name, control_dept_names, adder_ids):
    """Explain why a faculty in general_information has no control-room USER."""
    if not dept_name:
        return "No department set in General Information - never created by the Excel upload (department is required)."
    if dept_name.strip().lower() not in control_dept_names:
        return f"Department '{dept_name}' not found in control-room departments - upload row would have been skipped."
    if employee_id in adder_ids:
        return "Pre-authorization exists (NewUserAdder) but the control-room USER is missing - user was deleted after upload."
    return "Not created in control-room (no user and no pre-authorization) - upload never reached the user-creation step."


def build_missing_users_report():
    """Excel download listing Employee IDs present in general_information but
    missing from the control-room USER table, with the reason for each."""
    summary = find_missing_control_users()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Missing Users"

    worksheet.append(["S.No", "Employee ID", "Name", "Department", "Designation", "Reason"])
    for cell in worksheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for index, row in enumerate(summary["missing_rows"], start=1):
        worksheet.append([
            index,
            row["employee_id"],
            row["name"],
            row["department"],
            row["designation"],
            row["reason"],
        ])

    _style_sheet(worksheet, [8, 16, 28, 34, 26, 80])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="missing_control_room_users.xlsx"'
    workbook.save(response)
    return response


def _style_sheet(worksheet, widths):
    for cell in worksheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(index)].width = width


def _department_key(value):
    return re.sub(r"\s+", " ", (value or "").strip()).lower()


def _activate_local_department(department):
    if department.is_active is False:
        department.is_active = True
        department.save(update_fields=["is_active"])


def _append_reference_values(reference_sheet):
    departments = list(
        Department.objects.using(APPROVAL_DB)
        .all()
        .order_by("Department")
        .values_list("Department", flat=True)
    )
    roles = list(
        Role.objects.using(APPROVAL_DB).all().order_by("role").values_list("role", flat=True)
    )
    designations = list(
        DesignationMaster.objects
        .all()
        .order_by("designation_name")
        .values_list("designation_name", flat=True)
    )
    categories = list(
        FacultyCategory.objects
        .filter(is_active=True)
        .order_by("category_name")
        .values_list("category_name", flat=True)
    )
    shifts = list(
        ShiftMaster.objects
        .filter(is_active=True)
        .order_by("shift_name")
        .values_list("shift_name", flat=True)
    )

    max_rows = max(len(departments), len(roles), len(designations), len(categories), len(shifts), 1)
    for row_index in range(max_rows):
        reference_sheet.append([
            departments[row_index] if row_index < len(departments) else "",
            roles[row_index] if row_index < len(roles) else "",
            designations[row_index] if row_index < len(designations) else "",
            categories[row_index] if row_index < len(categories) else "",
            shifts[row_index] if row_index < len(shifts) else "",
        ])


def _process_worksheet(worksheet):
    header_row_number, header_map = _find_header(worksheet)
    if not header_row_number:
        raise ValueError("Required columns were not found in the first 20 rows.")

    counts = {
        "created_general": 0,
        "updated_general": 0,
        "created_users": 0,
        "updated_users": 0,
        "created_pre_authorizations": 0,
        "updated_pre_authorizations": 0,
        "skipped": 0,
    }
    results = []

    for row_number in range(header_row_number + 1, worksheet.max_row + 1):
        row_data = _row_data(worksheet[row_number], header_map)
        if not any(row_data.values()):
            continue

        missing = [label for key, label in REQUIRED_COLUMNS.items() if not row_data[key]]
        if missing:
            counts["skipped"] += 1
            results.append(_result(row_number, row_data, "Skipped", f"Missing: {', '.join(missing)}"))
            continue

        try:
            faculty_created, user_created, adder_created = _save_row(row_data)
            counts["created_general" if faculty_created else "updated_general"] += 1
            counts["created_users" if user_created else "updated_users"] += 1
            counts["created_pre_authorizations" if adder_created else "updated_pre_authorizations"] += 1
            message = _saved_message(faculty_created, user_created, adder_created)
            results.append(_result(row_number, row_data, "Saved", message))
        except Exception as exc:
            counts["skipped"] += 1
            results.append(_result(row_number, row_data, "Skipped", str(exc)))

    counts["total"] = counts["created_general"] + counts["updated_general"] + counts["skipped"]
    return results, counts


def _saved_message(faculty_created, user_created, adder_created):
    return "; ".join([
        f"General info {'created' if faculty_created else 'updated'}",
        f"User {'created' if user_created else 'updated/existing'}",
        f"Pre-authorization {'created' if adder_created else 'updated/existing'}",
    ])


def _find_header(worksheet):
    for row_number in range(1, min(worksheet.max_row, 20) + 1):
        header_map = {}
        for index, cell in enumerate(worksheet[row_number]):
            mapped_header = _normalize_header(cell.value)
            if mapped_header and mapped_header not in header_map:
                header_map[mapped_header] = index
        if set(REQUIRED_COLUMNS).issubset(header_map):
            return row_number, header_map
    return None, {}


def _row_data(row, header_map):
    data = {}
    for key in REQUIRED_COLUMNS:
        index = header_map[key]
        data[key] = _clean_value(row[index].value if index < len(row) else "")
    for key in OPTIONAL_COLUMNS:
        index = header_map.get(key)
        data[key] = _clean_value(row[index].value if index is not None and index < len(row) else "")
    return data


def _result(row_number, row_data, status, message):
    return {
        "row": row_number,
        "employee_id": row_data.get("employee_id", ""),
        "name": row_data.get("name", ""),
        "status": status,
        "message": message,
    }


def _clean_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _normalize_header(value):
    cleaned = _clean_value(value).lower()
    cleaned = re.sub(r"\([^)]*\)", "", cleaned)
    cleaned = cleaned.replace("&", "and")
    cleaned = re.sub(r"[^a-z0-9]+", "_", cleaned).strip("_")
    return HEADER_ALIASES.get(cleaned) or HEADER_ALIASES.get(cleaned.replace("_", ""))


def _save_row(row_data):
    employee_id = row_data["employee_id"]
    if not employee_id.isdigit():
        raise ValueError("Employee ID must be numeric to save faculty general information.")

    control_department = _existing_control_department(row_data["department"])
    local_department = _existing_local_department(row_data["department"])
    role = _get_or_create_role(row_data["role"])
    designation = _get_or_create_designation(row_data["designation"], row_data["category"])
    category = _get_or_create_faculty_category(row_data["category"])
    unique_id = _build_unique_id(control_department, employee_id, role)

    existing_unique_user = (
        USER.objects.using(APPROVAL_DB)
        .filter(unique_id=unique_id)
        .exclude(Employee_id=employee_id)
        .first()
    )
    if existing_unique_user:
        raise ValueError(f"Control-room unique id already belongs to employee {existing_unique_user.Employee_id}.")

    with transaction.atomic():
        with transaction.atomic(using=APPROVAL_DB):
            faculty_created = _ensure_general_information(
                row_data,
                local_department,
                designation,
                category,
            )
            StaffCategoryAssignment.objects.get_or_create(
                employee_id=employee_id,
                defaults={"category": category},
            )
            user_created = _ensure_control_user(row_data, control_department, role, unique_id)
            adder_created = _ensure_pre_authorized_employee(employee_id, control_department, role)

    return faculty_created, user_created, adder_created


def _existing_local_department(department_name):
    department = Add_Department.objects.filter(Department__iexact=department_name).first()
    if not department:
        raise ValueError(f"Department '{department_name}' was not found in academic departments.")
    if department.is_active is False:
        department.is_active = True
        department.save()
    return department


def _existing_control_department(department_name):
    department = Department.objects.using(APPROVAL_DB).filter(Department__iexact=department_name).first()
    if not department:
        raise ValueError(f"Department '{department_name}' was not found in control-room departments.")
    return department


def _get_or_create_role(role_name):
    role = Role.objects.using(APPROVAL_DB).filter(role__iexact=role_name).first()
    return role or Role.objects.using(APPROVAL_DB).create(role=role_name)


def _get_or_create_designation(designation_name, category_name):
    designation = DesignationMaster.objects.filter(designation_name__iexact=designation_name).first()
    if designation:
        return designation

    category_key = (category_name or "").strip().lower().replace("-", " ")
    return DesignationMaster.objects.create(
        designation_name=designation_name,
        is_teaching="teaching" in category_key and "non" not in category_key,
    )


def _get_or_create_faculty_category(category_name):
    category = FacultyCategory.objects.filter(category_name__iexact=category_name).first()
    if category:
        if not category.is_active:
            category.is_active = True
            category.save()
        return category
    return FacultyCategory.objects.create(category_name=category_name, is_active=True)


def _build_unique_id(department, employee_id, role):
    return f"{department.Department_code or ''}{employee_id}{role.role or ''}"[:100]


def _ensure_general_information(row_data, local_department, designation, category):
    employee_id = int(row_data["employee_id"])
    values = _general_information_values(row_data, employee_id, local_department, designation, category)
    faculty = general_information.objects.filter(faculty_id=employee_id).first()
    if faculty:
        for field, value in values.items():
            setattr(faculty, field, value)
        faculty.save(update_fields=list(values.keys()))
        return False

    general_information.objects.create(**values)
    return True


def _general_information_values(row_data, employee_id, local_department, designation, category):
    values = {
        "faculty_id": employee_id,
        "name": row_data["name"],
        "department": local_department,
        "designation": designation,
        "category": category,
    }
    if row_data.get("gender"):
        values["gender"] = _parse_gender(row_data["gender"])
    if row_data.get("dob"):
        values["dob"] = _parse_date(row_data["dob"], "DOB")
    if row_data.get("address"):
        values["address"] = row_data["address"]
    if row_data.get("personal_email"):
        values["personal_email"] = row_data["personal_email"].lower()
    college_email = row_data.get("college_email") or row_data.get("email")
    if college_email:
        values["college_email"] = college_email.lower()
    if row_data.get("phone"):
        values["phone"] = _parse_phone(row_data["phone"])
    for field in ("blood_group", "community", "caste", "religion", "PAN_number", "Aadhar_number"):
        if row_data.get(field):
            values[field] = row_data[field]
    if row_data.get("doj"):
        values["doj"] = _parse_date(row_data["doj"], "DOJ")
    if row_data.get("shift"):
        values["shift"] = _existing_shift(row_data["shift"])
    return values


def _parse_gender(value):
    text = _clean_value(value)
    if not text:
        return ""
    choices = {choice.lower(): choice for choice, _ in general_information.GenderChoices.choices}
    normalized = choices.get(text.lower())
    if not normalized:
        valid = ", ".join(choice for choice, _ in general_information.GenderChoices.choices)
        raise ValueError(f"Gender '{text}' is not valid. Use one of: {valid}.")
    return normalized


def _existing_shift(shift_name):
    shift = ShiftMaster.objects.filter(shift_name__iexact=shift_name, is_active=True).first()
    if not shift:
        raise ValueError(f"Shift '{shift_name}' was not found in active shift masters.")
    return shift


def _parse_date(value, label):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return from_excel(value).date()

    text = _clean_value(value)
    if not text:
        return None
    if re.fullmatch(r"\d{5}", text):
        return from_excel(int(text)).date()

    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        pass

    for date_format in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d.%m.%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue
    raise ValueError(f"{label} '{text}' is not a valid date. Use YYYY-MM-DD or DD-MM-YYYY.")


def _parse_phone(value):
    text = _clean_value(value)
    if not text:
        return None
    digits = re.sub(r"\D", "", text)
    if not digits:
        raise ValueError(f"Phone number '{text}' is not valid.")
    return int(digits)


def _ensure_control_user(row_data, department, role, unique_id):
    employee_id = row_data["employee_id"]
    user = (
        USER.objects.using(APPROVAL_DB)
        .filter(Employee_id=employee_id, is_student=False, is_parent=False)
        .order_by("id")
        .first()
    )
    unique_user = USER.objects.using(APPROVAL_DB).filter(unique_id=unique_id).order_by("id").first()

    if unique_user and user and unique_user.pk != user.pk:
        raise ValueError(f"Control-room unique id already belongs to employee {unique_user.Employee_id}.")

    if not user and unique_user:
        user = unique_user

    email = _control_email(
        row_data.get("college_email") or row_data.get("email"),
        employee_id,
        department,
        role,
        current_user_id=getattr(user, "pk", None),
    )

    if user:
        user.username = row_data["name"]
        user.Employee_id = employee_id
        user.Department = department
        user.role = role
        user.unique_id = unique_id
        user.email = email
        user.is_student = False
        user.is_parent = False
        user.is_staff = True
        user.is_active = True
        user.save(
            using=APPROVAL_DB,
            update_fields=[
                "username",
                "Employee_id",
                "Department",
                "role",
                "unique_id",
                "email",
                "is_student",
                "is_parent",
                "is_staff",
                "is_active",
            ],
        )
        return False

    USER.objects.using(APPROVAL_DB).create(
        username=row_data["name"],
        Employee_id=employee_id,
        Department=department,
        role=role,
        unique_id=unique_id,
        email=email,
        password=make_password(employee_id),
        is_student=False,
        is_parent=False,
        is_staff=True,
        is_active=True,
    )
    return True


def _control_email(email, employee_id, department, role, current_user_id=None):
    email = (email or "").strip().lower()
    fallback = f"employee-{employee_id}-{department.id}-{role.id}@employee-upload.local"
    qs = USER.objects.using(APPROVAL_DB)
    if email and not qs.filter(email__iexact=email).exclude(pk=current_user_id).exists():
        return email
    if not qs.filter(email__iexact=fallback).exclude(pk=current_user_id).exists():
        return fallback
    return f"employee-{employee_id}-{department.id}-{role.id}-{datetime.now().strftime('%Y%m%d%H%M%S')}@employee-upload.local"


def _ensure_pre_authorized_employee(employee_id, department, role):
    existing_qs = NewUserAdder.objects.using(APPROVAL_DB).filter(Employee_id=employee_id)
    if existing_qs.exists():
        existing_qs.exclude(Department=department, role=role).update(
            Department=department,
            role=role,
        )
        return False

    try:
        NewUserAdder.objects.using(APPROVAL_DB).create(
            Employee_id=employee_id,
            Department=department,
            role=role,
        )
        return True
    except IntegrityError as exc:
        if "Employee_id" in str(exc):
            return False
        raise

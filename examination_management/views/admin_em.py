from tkinter import RIGHT

from django.shortcuts import render,redirect
from django.http import JsonResponse
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from examination_management.decorators import examination_management

from examination_management.models import ExaminationFunction
from user_accounts.decorators import is_super_user, no_cache, faculty_login_required, check_permission
import re
from user_accounts.models import Role



@check_permission("em_hello")
def em_hello(request):
    return render(request, "em_hello.html")




# @faculty_login_required
@examination_management
def em_home(request):
    # print("cm home page ")
    request.session['current_page'] = 'em_home'
    return redirect('home')




@faculty_login_required
@no_cache
@is_super_user('examination_management')
def em_assign_permission(request):
    if request.method == 'POST':  
        permissions = request.POST
        for role_name, role_permissions in permissions.items():
            if role_name.startswith('permissions'):
                try:
                    # Extract data from role_name using regex
                    extract_data = list(re.findall(r'\[([^\]]+)\]', role_name))
                    if len(extract_data) < 2:  # Ensure there are at least role and function
                        messages.warning(request,f"Invalid format in role_name: {role_name}. Skipping.")
                        continue
                    
                    extract_data.append(role_permissions)

                    # Retrieve the role (Handle if the role does not exist)
                    try:
                        role = Role.objects.using("rit_approval_system").get(role=extract_data[0])

                    except Role.DoesNotExist:
                        messages.error(request,f"Role {extract_data[0]} does not exist.")
                        messages.error(request, f"Role '{extract_data[0]}' does not exist. Skipping this entry.")
                        continue
                    
                    # Parse permissions - handle the case where role_permissions is a list (unlikely with POST data)
                    if isinstance(role_permissions, list):  # Handle list case
                        role_permissions = role_permissions[0]
                    
                    # Convert permission to boolean (True/False)
                    permission = extract_data[2] == 'true'
       
                    
                    # Find or create ApprovalPermissionFunction object
                    permission_obj = ExaminationFunction.objects.filter(
                        role=role, function=extract_data[1]
                    ).first()
                    
                    if permission_obj:
                        permission_obj.permission = permission
                        permission_obj.save()
                    else:
                        # Create a new ApprovalPermissionFunction object
                        ExaminationFunction.objects.create(
                            role=role,
                            function=extract_data[1],
                            permission=permission
                        )
                except Exception as e:
                    # Catch unexpected errors and log them
                    messages.error(request,f"Error processing role '{role_name}': {str(e)}")
                    messages.error(request, f"An error occurred while processing '{role_name}': {str(e)}")

    # Redirect to admin dashboard after processing
    messages.success(request,"The permission changes have been successfully applied.")
    return redirect('examination_management')



from course_management.models import Regulations
from examination_management.models import CourseOutcome

from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.db import IntegrityError

from django.contrib import messages
from django.db import IntegrityError
from django.shortcuts import get_object_or_404, redirect, render

def update_course_outcome(request):
    outcomes = CourseOutcome.objects.all().order_by("id")
    outcome = None

    # ✅ Distinct regulation years
    regulations = Regulations.objects.values_list("year", flat=True).distinct()

    if request.method == "POST":
        if "delete" in request.POST:  # 🔹 Handle Delete
            pk = request.POST.get("delete")
            outcome = get_object_or_404(CourseOutcome, pk=pk)
            outcome.delete()
            messages.success(request, f"Course Outcome '{outcome.co_code}' deleted successfully.")
            return redirect("update_course_outcome")

        pk = request.POST.get("id")
        regulation = request.POST.get("regulation")
        co_code = request.POST.get("co_code")
        co_name = request.POST.get("co_name")

        # 🔹 Validation
        if not regulation or not co_code or not co_name:
            messages.error(request, "All fields are required.")
            return redirect("update_course_outcome")

        try:
            if pk:  # ✅ Update existing
                outcome = get_object_or_404(CourseOutcome, pk=pk)

                # 🔹 Check duplicate CO code (exclude self)
                if CourseOutcome.objects.exclude(pk=pk).filter(co_code=co_code, regulation=regulation).exists():
                    messages.error(request, f"Course Outcome with code '{co_code}' already exists in this regulation.")
                    return redirect("update_course_outcome")

                outcome.regulation = regulation
                outcome.co_code = co_code
                outcome.co_name = co_name
                outcome.save()
                messages.success(request, f"Course Outcome '{co_code}' updated successfully.")
            else:  # ✅ Create new
                if CourseOutcome.objects.filter(co_code=co_code, regulation=regulation).exists():
                    messages.error(request, f"Course Outcome with code '{co_code}' already exists in this regulation.")
                else:
                    CourseOutcome.objects.create(
                        regulation=regulation,
                        co_code=co_code,
                        co_name=co_name,
                    )
                    messages.success(request, f"Course Outcome '{co_code}' created successfully.")
        except IntegrityError as e:
            messages.error(request, f"Database error: {str(e)}")
        except Exception as e:
            messages.error(request, f"Unexpected error: {str(e)}")

        return redirect("update_course_outcome")

    # ✅ Prefill outcome in edit mode
    pk = request.GET.get("edit")
    if pk:
        outcome = get_object_or_404(CourseOutcome, pk=pk)

    return render(request, "examination_management/admin/update_course_outcome.html", {
        "outcomes": outcomes,
        "outcome": outcome,
        "regulations": regulations,
    })




from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.db import IntegrityError

from django.contrib import messages
from django.db import IntegrityError
from django.shortcuts import get_object_or_404, redirect, render

def update_blooms_level(request):
    levels = BloomsLevel.objects.all().order_by("id")
    level = None

    if request.method == "POST":
        if "delete" in request.POST:  # 🔹 Handle Delete
            pk = request.POST.get("delete")
            level = get_object_or_404(BloomsLevel, pk=pk)
            level.delete()
            messages.success(request, f"Bloom's Level '{level.level_code}' deleted successfully.")
            return redirect("update_blooms_level")

        pk = request.POST.get("id")
        level_code = request.POST.get("level_code")
        description = request.POST.get("description")

        # ✅ Validation
        if not level_code or not description:
            messages.error(request, "Both Level Code and Description are required.")
            return redirect("update_blooms_level")

        try:
            if pk:  # ✅ Update existing
                level = get_object_or_404(BloomsLevel, pk=pk)

                # 🔹 Prevent duplicate codes (exclude self)
                if BloomsLevel.objects.exclude(pk=pk).filter(level_code=level_code).exists():
                    messages.error(request, f"Bloom's Level '{level_code}' already exists.")
                    return redirect("update_blooms_level")

                level.level_code = level_code
                level.description = description
                level.save()
                messages.success(request, f"Bloom's Level '{level_code}' updated successfully.")
            else:  # ✅ Create new
                if BloomsLevel.objects.filter(level_code=level_code).exists():
                    messages.error(request, f"Bloom's Level '{level_code}' already exists.")
                else:
                    BloomsLevel.objects.create(
                        level_code=level_code,
                        description=description
                    )
                    messages.success(request, f"Bloom's Level '{level_code}' created successfully.")
        except IntegrityError as e:
            messages.error(request, f"Database error: {str(e)}")
        except Exception as e:
            messages.error(request, f"Unexpected error: {str(e)}")

        return redirect("update_blooms_level")

    # ✅ Prefill form in edit mode
    pk = request.GET.get("edit")
    if pk:
        level = get_object_or_404(BloomsLevel, pk=pk)

    return render(request, "examination_management/admin/update_blooms_level.html", {
        "levels": levels,
        "level": level,
    })



# def grade_master(request):
#     if request.method == "POST":
#         degree = request.POST.get("degree")
#         regulation = request.POST.get("regulation")
#         mark_range = request.POST.get("mark_range")
#         grade_range = request.POST.get("grade_range")
#         class_category = request.POST.get("class_category")
#         grade = request.POST.get("grade")

#         try:
#             # Parse mark range (like "80-89")
#             mark_from, mark_to = None, None
#             if mark_range and "-" in mark_range:
#                 mark_from, mark_to = map(int, mark_range.split("-"))

#             # Parse grade range (like "3.5-4.0")
#             grade_from, grade_to = None, None
#             if grade_range and "-" in grade_range:
#                 grade_from, grade_to = map(float, grade_range.split("-"))

#             # Save / update
#             obj, created = GradeMaster.objects.update_or_create(
#                 degree=degree,
#                 class_category=class_category,
                
#                 defaults={
#                     "regulation": regulation,
#                     "mark_from": mark_from,
#                     "mark_to": mark_to,
#                     "grade_from": grade_from,
#                     "grade_to": grade_to,
#                     "grade": grade,
#                 }
#             )

#             if created:
#                 messages.success(request, f"✅ Grade entry added successfully for {degree} ({class_category}).")
#             else:
#                 messages.success(request, f"✏️ Grade entry updated successfully for {degree} ({class_category}).")

#         except Exception as e:
#             messages.error(request, f"❌ Error saving grade entry: {e}")

#         return redirect("examination_management")  # refresh after save

#     return redirect("examination_management")


# def grade_master_list(request):
#     grade_list = GradeMaster.objects.all()
#     context = {"grade_list": grade_list}
#     return render(request, "examination_management/admin/grade_master_list.html", context)


from examination_management.models import Class_Category
from user_accounts.models import Add_Department, Degree, general_information
from course_management.models import Regulations
def class_category(request):
    categories = Class_Category.objects.all()
    return render(request, "examination_management/admin/class_category.html", {"categories": categories})

# Create & Update in the same function (but keep your view name)
def edit_grade_master(request, grade_id=None):
    if grade_id:
        grade = get_object_or_404(GradeMaster, id=grade_id)
        success_message = "Grade updated successfully!"
    else:
        grade = GradeMaster()
        success_message = "Grade added successfully!"

    if request.method == "POST":
        grade.degree = request.POST.get("degree")
        grade.regulation = request.POST.get("regulation")
        grade.grade_from = request.POST.get("grade_from") or None
        grade.grade_to = request.POST.get("grade_to") or None
        grade.class_category = request.POST.get("class_category")
        grade.mark_from = request.POST.get("mark_from") or None
        grade.mark_to = request.POST.get("mark_to") or None
        grade.grade = request.POST.get("grade")

        grade.save()
        messages.success(request, success_message)
        return redirect("grade_master_list")

    return render(request, "examination_management/admin/edit_grade_master.html", {"grade": grade})


# Delete
def delete_grade_master(request, grade_id):
    grade = get_object_or_404(GradeMaster, id=grade_id)
    grade.delete()
    messages.success(request, "Grade deleted successfully!")
    return redirect("grade_master_list")





from django.shortcuts import redirect
from django.http import JsonResponse
from django.contrib import messages
from django.db import transaction
from django.core.exceptions import MultipleObjectsReturned

from examination_management.models import Degree, Assessments, ModelLab, InternalAssessment


def _sync_exact_iats(degree_obj, count):
    """Ensure EXACTLY IAT1..IAT{count} exist for this degree."""
    desired = {f"IAT{i}" for i in range(1, count + 1)}
    existing_qs = InternalAssessment.objects.filter(degree=degree_obj)
    existing_names = set(existing_qs.values_list("iat", flat=True))

    extras = existing_names - desired
    if extras:
        InternalAssessment.objects.filter(degree=degree_obj, iat__in=extras).delete()

    missing = desired - (existing_names - extras)
    created = 0
    for name in sorted(missing, key=lambda s: int(s.replace("IAT", ""))):
        InternalAssessment.objects.get_or_create(degree=degree_obj, iat=name)
        created += 1
    return created, len(extras)


def _sync_exact_modellabs(degree_obj, count):
    """Ensure EXACTLY Model Lab 1..Model Lab {count} exist for this degree."""
    desired = {f"Model Lab {i}" for i in range(1, count + 1)}
    existing_qs = ModelLab.objects.filter(degree=degree_obj)
    existing_names = set(existing_qs.values_list("model_lab_name", flat=True))

    extras = existing_names - desired
    if extras:
        ModelLab.objects.filter(degree=degree_obj, model_lab_name__in=extras).delete()

    missing = desired - (existing_names - extras)
    created = 0
    for name in sorted(missing, key=lambda s: int(s.replace("Model Lab ", ""))):
        ModelLab.objects.get_or_create(degree=degree_obj, model_lab_name=name)
        created += 1
    return created, len(extras)


def assessments(request):
    """
    AJAX:
      - degree_preview
      - apply_filter
      - existing_for_degree
      - map_assessment_iat
      - map_modellab_iat
      - update_assessment   <-- NEW
      - delete_assessment   <-- NEW

    Normal POST:
      - Optional: create one assessment (needs QPR if name provided)
      - Optional: sync EXACT counts for IATs and Model Labs
    """

    # --- AJAX: degree preview ---
    if request.method == "POST" and request.POST.get("ajax") == "degree_preview":
        degree_id = request.POST.get("degree_id")
        if not degree_id:
            return JsonResponse({"ok": False, "error": "Degree ID missing."}, status=400)
        try:
            d = Degree.objects.get(pk=int(degree_id))
            return JsonResponse({"ok": True, "degree_name": d.degree, "degree_code": d.degree_code})
        except (Degree.DoesNotExist, TypeError, ValueError):
            return JsonResponse({"ok": False, "error": "Invalid degree selected."}, status=400)

    # --- AJAX: apply_filter (lock degree in session) ---
    if request.method == "POST" and request.POST.get("ajax") == "apply_filter":
        degree_id = request.POST.get("degree_id")
        if not degree_id:
            return JsonResponse({"ok": False, "error": "Degree ID missing."}, status=400)
        try:
            d = Degree.objects.get(pk=int(degree_id))
            request.session['filtered_degree_id'] = d.pk
            request.session['filtered_degree_label'] = f"{d.degree_code} - {d.degree}"
            return JsonResponse({"ok": True, "degree_name": d.degree, "degree_code": d.degree_code})
        except (Degree.DoesNotExist, TypeError, ValueError):
            return JsonResponse({"ok": False, "error": "Invalid degree selected."}, status=400)

    # --- AJAX: existing_for_degree ---
    if request.method == "POST" and request.POST.get("ajax") == "existing_for_degree":
        degree_id = request.POST.get("degree_id")
        if not degree_id:
            return JsonResponse({"ok": False, "error": "Degree ID missing."}, status=400)
        try:
            d = Degree.objects.get(pk=int(degree_id))
        except (Degree.DoesNotExist, TypeError, ValueError):
            return JsonResponse({"ok": False, "error": "Invalid degree selected."}, status=400)

        assessments_qs = Assessments.objects.filter(degree=d).select_related('internal_assessment').order_by('id')
        iats_qs = InternalAssessment.objects.filter(degree=d).order_by('iat')
        modellabs_qs = ModelLab.objects.filter(degree=d).select_related('internal_assessment').order_by('model_lab_name')

        payload = {
            "ok": True,
            "assessments": [
                {
                    "id": a.id,
                    "assessment_name": a.assessment_name,
                    "question_paper_required": bool(a.question_paper_required),
                    "internal_assessment_id": (a.internal_assessment_id or None),
                }
                for a in assessments_qs
            ],
            "iats_detail": [{"id": ia.id, "iat": ia.iat} for ia in iats_qs],
            "model_labs": [
                {
                    "id": ml.id,
                    "model_lab_name": ml.model_lab_name,
                    "internal_assessment_id": (ml.internal_assessment_id or None),
                }
                for ml in modellabs_qs
            ],
        }
        return JsonResponse(payload)

    # --- AJAX: map_assessment_iat ---
    if request.method == "POST" and request.POST.get("ajax") == "map_assessment_iat":
        assessment_id = request.POST.get("assessment_id")
        internal_id = request.POST.get("internal_id") or None
        if not assessment_id:
            return JsonResponse({"ok": False, "error": "Assessment ID missing."}, status=400)
        try:
            a = Assessments.objects.get(pk=int(assessment_id))
        except (Assessments.DoesNotExist, ValueError, TypeError):
            return JsonResponse({"ok": False, "error": "Assessment not found."}, status=404)

        if internal_id:
            try:
                ia = InternalAssessment.objects.get(pk=int(internal_id), degree=a.degree)
            except (InternalAssessment.DoesNotExist, ValueError, TypeError):
                return JsonResponse({"ok": False, "error": "Invalid IAT for this degree."}, status=400)
            a.internal_assessment = ia
        else:
            a.internal_assessment = None
        a.save(update_fields=["internal_assessment"])
        return JsonResponse({"ok": True})

    # --- AJAX: map_modellab_iat ---
    if request.method == "POST" and request.POST.get("ajax") == "map_modellab_iat":
        modellab_id = request.POST.get("modellab_id")
        internal_id = request.POST.get("internal_id") or None
        if not modellab_id:
            return JsonResponse({"ok": False, "error": "ModelLab ID missing."}, status=400)
        try:
            ml = ModelLab.objects.get(pk=int(modellab_id))
        except (ModelLab.DoesNotExist, ValueError, TypeError):
            return JsonResponse({"ok": False, "error": "ModelLab not found."}, status=404)

        if internal_id:
            try:
                ia = InternalAssessment.objects.get(pk=int(internal_id), degree=ml.degree)
            except (InternalAssessment.DoesNotExist, ValueError, TypeError):
                return JsonResponse({"ok": False, "error": "Invalid IAT for this degree."}, status=400)
            ml.internal_assessment = ia
        else:
            ml.internal_assessment = None
        ml.save(update_fields=["internal_assessment"])
        return JsonResponse({"ok": True})

    # --- AJAX: update_assessment (rename and/or QPR) ---
    if request.method == "POST" and request.POST.get("ajax") == "update_assessment":
        assessment_id = request.POST.get("assessment_id")
        new_name = (request.POST.get("assessment_name") or "").strip()
        qpr_raw = (request.POST.get("question_paper_required") or "").strip()  # "yes"/"no"

        if not assessment_id:
            return JsonResponse({"ok": False, "error": "Assessment ID missing."}, status=400)
        try:
            a = Assessments.objects.select_related("degree").get(pk=int(assessment_id))
        except (Assessments.DoesNotExist, ValueError, TypeError):
            return JsonResponse({"ok": False, "error": "Assessment not found."}, status=404)

        dirty = False
        if new_name:
            a.assessment_name = new_name
            dirty = True
        if qpr_raw in ("yes", "no"):
            a.question_paper_required = (qpr_raw == "yes")
            dirty = True

        if dirty:
            a.save()
        return JsonResponse({"ok": True})

    # --- AJAX: delete_assessment ---
    if request.method == "POST" and request.POST.get("ajax") == "delete_assessment":
        assessment_id = request.POST.get("assessment_id")
        if not assessment_id:
            return JsonResponse({"ok": False, "error": "Assessment ID missing."}, status=400)
        try:
            a = Assessments.objects.get(pk=int(assessment_id))
        except (Assessments.DoesNotExist, ValueError, TypeError):
            return JsonResponse({"ok": False, "error": "Assessment not found."}, status=404)

        a.delete()
        return JsonResponse({"ok": True})

    # --- Normal form submission branch ---
    if request.method != "POST":
        return redirect("examination_management")

    degree_id = request.POST.get("degree_id")
    assessment_name = (request.POST.get("assessment_name") or "").strip()
    no_of_model_labs = (request.POST.get("no_of_model_labs") or "").strip()
    no_of_internal_assessments = (request.POST.get("no_of_internal_assessments") or "").strip()
    qpr_raw = (request.POST.get("question_paper_required") or "").strip()
    question_paper_required = (qpr_raw == "yes")

    # Enforce Filter lock
    filtered_degree_id = request.session.get("filtered_degree_id")
    if not filtered_degree_id or str(filtered_degree_id) != str(degree_id):
        messages.error(request, "Please select a degree and click Filter before saving.")
        return redirect("examination_management")

    if not degree_id:
        messages.error(request, "Degree is required.")
        return redirect("examination_management")

    # QPR only required if creating/updating a NEW assessment (assessment_name provided)
    if assessment_name and qpr_raw not in ("yes", "no"):
        messages.error(request, "Please choose Question Paper Required for the new assessment.")
        return redirect("examination_management")

    # If nothing to create/update (inline edits already saved via AJAX)
    if not assessment_name and not no_of_model_labs and not no_of_internal_assessments:
        messages.info(request, "Nothing to save. Inline edits and mappings are saved automatically.")
        return redirect("examination_management")

    # Fetch degree
    try:
        degree_obj = Degree.objects.get(pk=int(degree_id))
    except (Degree.DoesNotExist, TypeError, ValueError):
        messages.error(request, "Invalid degree selected.")
        return redirect("examination_management")

    created_labs = deleted_labs = created_iats = deleted_iats = 0
    created_assessment = updated_assessment = False

    with transaction.atomic():
        # Sync EXACT Internal Assessments
        if no_of_internal_assessments:
            try:
                iat_count = int(no_of_internal_assessments)
                assert iat_count > 0
            except Exception:
                messages.error(request, "Number of internal assessments must be a positive integer if provided.")
                return redirect("examination_management")
            c, d = _sync_exact_iats(degree_obj, iat_count)
            created_iats += c
            deleted_iats += d

        # Sync EXACT Model Labs
        if no_of_model_labs:
            try:
                lab_count = int(no_of_model_labs)
                assert lab_count > 0
            except Exception:
                messages.error(request, "Number of model labs must be a positive integer if provided.")
                return redirect("examination_management")
            c, d = _sync_exact_modellabs(degree_obj, lab_count)
            created_labs += c
            deleted_labs += d

        # Optional: create/update ONE assessment by name (apply QPR)
        if assessment_name:
            try:
                obj, created = Assessments.objects.update_or_create(
                    degree=degree_obj,
                    assessment_name=assessment_name,
                    defaults={"question_paper_required": question_paper_required},
                )
                if created:
                    created_assessment = True
                else:
                    if obj.question_paper_required != question_paper_required:
                        obj.question_paper_required = question_paper_required
                        obj.save(update_fields=["question_paper_required"])
                        updated_assessment = True
            except MultipleObjectsReturned:
                dup_qs = Assessments.objects.filter(degree=degree_obj, assessment_name=assessment_name).order_by("id")
                keep = dup_qs.first()
                dup_qs.exclude(pk=keep.pk).delete()
                if keep.question_paper_required != question_paper_required:
                    keep.question_paper_required = question_paper_required
                    keep.save(update_fields=["question_paper_required"])
                updated_assessment = True

    # Messages
    if created_iats or deleted_iats:
        messages.success(request, f"IATs synced: +{created_iats}, -{deleted_iats}.")
    if created_labs or deleted_labs:
        messages.success(request, f"Model Labs synced: +{created_labs}, -{deleted_labs}.")
    if created_assessment:
        messages.success(request, "Assessment created successfully.")
    if updated_assessment:
        messages.info(request, "Assessment updated successfully.")
    if not (created_labs or deleted_labs or created_iats or deleted_iats or created_assessment or updated_assessment):
        messages.info(request, "Nothing new to save. Inline edits and mappings are saved automatically.")

    # Optional: clear filter lock after success
    request.session.pop("filtered_degree_id", None)
    request.session.pop("filtered_degree_label", None)

    return redirect("examination_management")




def assessments_list(request):
    assessments_list = Assessments.objects.all()
    context = {"assessments_list": assessments_list}
    return render(request, "examination_management/admin/assessments_list.html", context)

def edit_assessments(request, assessment_id):
    assessment = get_object_or_404(Assessments, id=assessment_id)
    if request.method == "POST":
        try:
            assessment.degree = request.POST.get("degree")
            assessment.assessment_name = request.POST.get("assessment_name")
            # assessment.max_marks = request.POST.get("max_marks")
            assessment.save()
            messages.success(request, "Assessment updated successfully!")
            return redirect("assessments_list")
        except Exception as e:
            messages.error(request, f"Error updating assessment: {str(e)}")
    context = {"assessment": assessment}
    return render(request, "examination_management/admin/edit_assessment.html", context)

def delete_assessments(request, assessment_id):
    assessment = get_object_or_404(Assessments, id=assessment_id)
    try:
        assessment.delete()
        messages.success(request, "Assessment deleted successfully!")
    except Exception as e:
        messages.error(request, f"Error deleting assessment: {str(e)}")
    return redirect("assessments_list")




from django.http import JsonResponse
from django.contrib import messages
from django.shortcuts import render, redirect
from examination_management.models import SquadMember
from faculty_management.models import general_information


@check_permission("assign_squad_member_list")
def assign_squad_member_list(request):

    # ================= FACULTY AJAX LOOKUP =================
    faculty_id = request.GET.get("facultyId")
    if faculty_id:
        data = {"name": "", "department": "", "designation": "", "error": ""}

        faculty_obj = general_information.objects.select_related(
            "department", "designation"
        ).filter(faculty_id=faculty_id).first()

        if not faculty_obj:
            data["error"] = "Faculty not found"
        else:
            data["name"] = faculty_obj.name
            data["department"] = (
                faculty_obj.department.Department
                if faculty_obj.department else "-"
            )
            data["designation"] = (
                faculty_obj.designation.designation_name
                if faculty_obj.designation else "-"
            )

        return JsonResponse(data)

    # ================= EDIT FETCH =================
    edit_ref = request.GET.get("edit")
    if edit_ref:
        member = SquadMember.objects.filter(appointment_ref=edit_ref).select_related("faculty_id").first()

        if not member:
            return JsonResponse({"error": "Record not found"})

        return JsonResponse({
            "appointment_ref": member.appointment_ref,
            "facultyId": member.faculty_id.faculty_id if member.faculty_id else "",
            "date": member.date.strftime("%Y-%m-%d") if member.date else "",
            "semester": member.semester,
            "iat": member.iat,
            "no_of_hall": member.no_of_hall,
            "hall_numbers": member.hall_numbers,
            "duration": member.duration,
        })

    # ================= DELETE =================
    if request.method == "POST" and request.POST.get("action") == "delete":
        ref = request.POST.get("appointment_ref")

        try:
            SquadMember.objects.filter(appointment_ref=ref).delete()
            return JsonResponse({"ok": True})
        except Exception as e:
            return JsonResponse({"ok": False, "error": str(e)})

    # ================= SAVE =================
    if request.method == "POST":
        try:
            refs = request.POST.getlist("appointment_ref[]")
            fids = request.POST.getlist("facultyId[]")
            dates = request.POST.getlist("date[]")
            semesters = request.POST.getlist("semester[]")
            iats = request.POST.getlist("iat[]")
            halls = request.POST.getlist("no_of_hall[]")
            hall_numbers = request.POST.getlist("hall_numbers[]")
            durations = request.POST.getlist("duration[]")

            for i, ref in enumerate(refs):
                if not ref:
                    continue

                faculty_obj = general_information.objects.filter(
                    faculty_id=fids[i]
                ).first()

                if not faculty_obj:
                    continue

                SquadMember.objects.update_or_create(
                    appointment_ref=ref,
                    defaults={
                        "faculty_id": faculty_obj,
                        "date": dates[i] or None,
                        "semester": semesters[i] or None,
                        "iat": iats[i] or None,
                        "no_of_hall": halls[i] or None,
                        "hall_numbers": hall_numbers[i],
                        "duration": durations[i] or None,
                    },
                )

            messages.success(request, "Squad Members Saved Successfully")
            return redirect("assign_squad_member_list")

        except Exception as e:
            messages.error(request, f"Error: {e}")
            return redirect("assign_squad_member_list")

    # ================= LIST =================
    squad_members = SquadMember.objects.select_related(
        "faculty_id",
        "faculty_id__department",
        "faculty_id__designation"
    ).all().order_by("date")

    return render(
        request,
        "examination_management/admin/assign_squad_member_list.html",
        {"squad_members": squad_members},
    )
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from django.contrib import messages
from io import BytesIO

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors


def squad_member_report(request):

    # ✅ Get logged-in faculty object
    faculty = get_object_or_404(
        general_information,
        faculty_id=request.user.Employee_id
    )

    faculty_id_display = faculty.faculty_id

    # ✅ Always show 8 semesters
    available_semesters = [str(i) for i in range(1, 9)]
    selected_semester = request.GET.get("semester", "")

    # ================= PENDING =================
    assigned_qs = SquadMember.objects.filter(
        faculty_id=faculty,
        reported="Pending"
    ).order_by("date", "appointment_ref")

    # ================= REPORTED =================
    reported_qs = SquadMember.objects.filter(
        faculty_id=faculty,
        reported="Reported"
    )

    if selected_semester:
        reported_qs = reported_qs.filter(semester=selected_semester)

    reported_qs = reported_qs.order_by("-date")

    selected = None
    report = None

    # ================= REPORT SUBMISSION =================
    if assigned_qs.exists():
        selected = assigned_qs.first()

        report, _ = SquadMemberReport.objects.get_or_create(
            squad_member=selected
        )

        if request.method == "POST":

            yn = lambda name: request.POST.get(name) == "Yes"

            report.seating_appropriate = yn("q1")
            report.classrooms_clean = yn("q2")
            report.seating_as_arrangement = yn("q4")
            report.materials_distributed = yn("q5")
            report.only_permitted_materials = yn("q6")
            report.register_no_written = yn("q7")
            report.no_markings_on_paper = yn("q8")
            report.id_worn = yn("q9")
            report.unruly_behaviour = yn("q10")
            report.followed_rules = yn("q11")
            report.faculty_present = yn("q12")
            report.faculty_misconduct = yn("q13")
            report.feedback = request.POST.get("feedback")

            report.save()

            selected.reported = "Reported"
            selected.save(update_fields=["reported"])

            messages.success(request, "Report submitted successfully.")
            return redirect("squad_member_report")

    # ================= PDF GENERATION =================
    view_ref = request.GET.get("view_ref")

    if view_ref:

        selected_squad = get_object_or_404(
            SquadMember,
            appointment_ref=view_ref,
            faculty_id=faculty,
            reported="Reported"
        )

        selected_report = SquadMemberReport.objects.filter(
            squad_member=selected_squad
        ).first()

        selected_report = SquadMemberReport.objects.filter(
            squad_member=selected_squad
        ).first()

        # Title
        p.setFont("Helvetica-Bold", 16)
        p.drawString(180, y, "Squad Member Report Summary")
        y -= 40

        # Squad Info Table
        squad_data = [
            ["Appointment Ref", selected_squad.appointment_ref],
            ["Name", selected_squad.faculty_id.name],
            ["Designation", selected_squad.faculty_id.designation.designation_name if selected_squad.faculty_id.designation else "-"],
            ["Department", selected_squad.faculty_id.department.Department if selected_squad.faculty_id.department else "-"],
            ["Semester", selected_squad.semester],
            ["Date", selected_squad.date.strftime('%d-%m-%Y')],
            ["IAT", selected_squad.iat],
            ["No. of Halls", selected_squad.no_of_hall],
            ["Hall Numbers", selected_squad.hall_numbers],
            ["Duration", selected_squad.duration],
        ]

        table = Table(squad_data, colWidths=[160, 300])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
        ]))
        table.wrapOn(p, width, height)
        table.drawOn(p, 50, y - len(squad_data) * 18)
        y -= len(squad_data) * 18 + 40

        # Observation Table
        if selected_report:
            observations = [
                ["Observation", "Yes/No"],
                ["Seating Appropriate", "Yes" if selected_report.seating_appropriate else "No"],
                ["Classrooms Clean", "Yes" if selected_report.classrooms_clean else "No"],
                ["Rules Followed", "Yes" if selected_report.followed_rules else "No"],
                ["Faculty Present", "Yes" if selected_report.faculty_present else "No"],
            ]

            obs_table = Table(observations, colWidths=[350, 100])
            obs_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONT', (0, 0), (-1, -1), 'Helvetica', 10),
            ]))
            obs_table.wrapOn(p, width, height)
            obs_table.drawOn(p, 50, y - len(observations) * 18)

        p.showPage()
        p.save()

        pdf = buffer.getvalue()
        buffer.close()

        filename = f"{selected_squad.appointment_ref}_report.pdf"

        if pdf_action == "view":
            response["Content-Disposition"] = f'inline; filename="{filename}"'
        else:
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

        response.write(pdf)
        return response

    # ================= RENDER =================
    return render(
        request,
        "examination_management/admin/squad_member_report.html",
        {
            "faculty_id": faculty_id_display,
            "pending_squad": selected,
            "report": report,
            "pending_exists": assigned_qs.exists(),
            "reported_squads": reported_qs,
            "available_semesters": available_semesters,
            "selected_semester": selected_semester,
        },
    )
from django.shortcuts import render, get_object_or_404
from django.http import FileResponse
from examination_management.models import (
    SquadMember,
    SquadMemberReport,
    SquadQuestionAnswer,
    SquadQuestions
)
import io

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse

from user_accounts.models import Add_Department
from examination_management.models import SquadMember, SquadMemberReport

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


@check_permission("squad_member_report_view")
def squad_member_report_view(request):
    # ===========================
    # FILTERED REPORTED SQUADS
    # ===========================
    reported_squads = SquadMember.objects.filter(
        reported="Reported"
    ).order_by("-date")

    semester = request.GET.get("semester", "")
    department = request.GET.get("department", "")
    search_name = request.GET.get("search", "")
    view_ref = request.GET.get("view_ref")
    pdf_action = request.GET.get("pdf")

    if semester:
        reported_squads = reported_squads.filter(semester=semester)

    if department:
        reported_squads = reported_squads.filter(
            faculty_id__department__icontains=department
        )

    if search_name:
        reported_squads = reported_squads.filter(
            faculty_id__name__icontains=search_name
        )

    # ===========================
    # DROPDOWNS
    # ===========================
    departments = (
        Add_Department.objects.filter(is_active=True)
        .values_list("Department", flat=True)
        .order_by("Department")
        .distinct()
    )

    available_semesters = [str(i) for i in range(1, 9)]

    selected_squad = None
    selected_report = None

    # ===========================
    # SELECT REPORT
    # ===========================
    if view_ref:
        selected_squad = get_object_or_404(
            SquadMember,
            appointment_ref=view_ref,
            reported="Reported"
        )

        selected_report = SquadMemberReport.objects.filter(
            squad_member=selected_squad
        ).first()

    # ===========================
    # PDF GENERATION
    # ===========================
    if selected_report and pdf_action and selected_squad:
        buffer = io.BytesIO()

        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()

        PRIMARY_BLUE = colors.HexColor("#0f2f57")
        SECONDARY_BLUE = colors.HexColor("#1a4b8c")
        DARK_GRAY = colors.HexColor("#111827")
        BORDER_GRAY = colors.HexColor("#e5e7eb")
        BG_GRAY = colors.HexColor("#f8fafc")

        title_style = ParagraphStyle(
            "title_style",
            parent=styles["Heading1"],
            fontSize=15,
            textColor=PRIMARY_BLUE,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            spaceAfter=8,
        )

        section_style = ParagraphStyle(
            "section_style",
            parent=styles["Heading2"],
            fontSize=11,
            textColor=SECONDARY_BLUE,
            alignment=TA_LEFT,
            fontName="Helvetica-Bold",
            spaceAfter=6,
        )

        cell_style = ParagraphStyle(
            "cell_style",
            parent=styles["Normal"],
            fontSize=8.5,
            textColor=DARK_GRAY,
            fontName="Helvetica",
        )

        elements = []

        # Title
        elements.append(Paragraph("RAMCO INSTITUTE OF TECHNOLOGY, RAJAPALAYAM", title_style))
        elements.append(Paragraph("Squad Member Report Summary", title_style))
        elements.append(Spacer(1, 12))

        # Squad details
        elements.append(Paragraph("Squad Details", section_style))

        squad_data = [
            ["Appointment Ref", selected_squad.appointment_ref],
            ["Name", selected_squad.faculty_id.name],
            ["Designation", selected_squad.faculty_id.designation],
            ["Department", selected_squad.faculty_id.department],
            ["Semester", selected_squad.semester],
            ["Date", selected_squad.date.strftime("%d-%m-%Y") if selected_squad.date else ""],
            ["IAT", selected_squad.iat],
            ["No. of Halls", selected_squad.no_of_hall],
            ["Hall Numbers", selected_squad.hall_numbers],
            ["Duration (hrs)", selected_squad.duration],
        ]

        squad_table = Table(squad_data, colWidths=[150, 320])
        squad_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), BG_GRAY),
            ("TEXTCOLOR", (0, 0), (-1, -1), DARK_GRAY),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(squad_table)
        elements.append(Spacer(1, 12))

        # Example report details
        elements.append(Paragraph("Report Details", section_style))

        report_data = [
            ["Report ID", str(selected_report.id)],
            ["Created At", selected_report.created_at.strftime("%d-%m-%Y %H:%M") if hasattr(selected_report, "created_at") and selected_report.created_at else ""],
        ]

        report_table = Table(report_data, colWidths=[150, 320])
        report_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), BG_GRAY),
            ("TEXTCOLOR", (0, 0), (-1, -1), DARK_GRAY),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("PADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(report_table)

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="squad_member_report_{selected_squad.appointment_ref}.pdf"'
        response.write(pdf)
        return response

    # ===========================
    # RENDER PAGE
    # ===========================
    return render(
        request,
        "examination_management/admin/squad_member_report_view.html",
        {
            "reported_squads": reported_squads,
            "departments": departments,
            "available_semesters": available_semesters,
            "selected_semester": semester,
            "selected_department": department,
            "search_name": search_name,
            "selected_squad": selected_squad,
            "selected_report": selected_report,
        },
    )
def api_degrees(request):
    # Only active degrees; order by name for stable UX
    data = list(
        Degree.objects.filter(is_active=True)
        .order_by("degree")
        .values("id", "degree", "degree_code", "duration")
    )
    return JsonResponse({"results": data})

def api_regulations(request):
    data = list(
        Regulations.objects.all()
        .order_by("-year")  # newest first; tweak if needed
        .values("id", "year")
    )
    return JsonResponse({"results": data})

def api_assessments(request):
    """
    Expects ?degree_id=<id>
    Returns IATs for that degree. Regulation is not part of this model in your schema,
    but we keep the selection order Degree -> Regulation -> Assessment.
    """
    degree_id = request.GET.get("degree_id")
    qs = InternalAssessment.objects.all()
    if degree_id:
        qs = qs.filter(degree_id=degree_id)
    data = list(qs.order_by("iat").values("id", "iat"))
    return JsonResponse({"results": data})   
    
    
    
    
    
    
    
import json
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt  # remove if you pass a valid CSRF token
from examination_management.models import AssessmentWeightage

    # If you are sending a valid CSRF token from the template, you can drop @csrf_exempt.
@require_http_methods(["GET", "POST"])
@csrf_exempt
def assessments_weightage(request):
    if request.method == "GET":
        # Render your existing template
        return render(request, "examination_management/admin/assessments_weightage.html")

    # POST: save logic
    try:
        # Expecting JSON body from fetch()
        data = json.loads(request.body.decode("utf-8"))

        degree_id = data.get("degree_id")
        regulation_id = data.get("regulation_id")
        selected_pct_raw = data.get("selected_assessment_percentage")
        activity_pct_raw = data.get("activity_percentage")

        # Basic presence checks
        if not degree_id or not regulation_id:
            return JsonResponse({"success": False, "error": "Degree and Regulation are required."}, status=400)
        if selected_pct_raw is None or activity_pct_raw is None:
            return JsonResponse({"success": False, "error": "Both percentage fields are required."}, status=400)

        # Coerce to float safely
        try:
            selected_pct = float(selected_pct_raw)
            activity_pct = float(activity_pct_raw)
        except (TypeError, ValueError):
            return JsonResponse({"success": False, "error": "Percentages must be numbers."}, status=400)

        # Range checks
        if not (0 <= selected_pct <= 100) or not (0 <= activity_pct <= 100):
            return JsonResponse({"success": False, "error": "Percentages must be between 0 and 100."}, status=400)

        # Sum check
        if abs((selected_pct + activity_pct) - 100.0) > 0.001:
            return JsonResponse({"success": False, "error": "Percentages must total 100."}, status=400)

        # Resolve FKs
        degree = Degree.objects.filter(id=degree_id).first()
        regulation = Regulations.objects.filter(id=regulation_id).first()
        if not degree or not regulation:
            return JsonResponse({"success": False, "error": "Invalid Degree or Regulation."}, status=400)

        # Upsert per (degree, regulation)
        obj, created = AssessmentWeightage.objects.update_or_create(
            degree=degree,
            regulation=regulation,
            defaults={
                "selected_assessment_percentage": selected_pct,
                "activity_percentage": activity_pct,
            },
        )

        return JsonResponse({
            "success": True,
            "message": "Created" if created else "Updated",
            "id": obj.id,
        })

    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON payload."}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)
        return render(request, 'examination_management/admin/assessments_weightage.html')

        
        
from django.shortcuts import render, redirect
from django.contrib import messages
from examination_management.models import Degree, Regulations, CourseHourConfig
from examination_management.models import CourseHourConfig

from examination_management.models import InternalAssessmentMasterTemplate

from django.contrib import messages
from django.shortcuts import render, redirect

# def ltp(request):
#     degrees = Degree.objects.filter(is_active=True)
#     regulations = Regulations.objects.all()

#     selected_degree_id = request.GET.get("degree")
#     selected_regulation_id = request.GET.get("regulation")
#     selected_config_id = request.GET.get("selected_config")

#     configs = CourseHourConfig.objects.none()
#     selected_ltp = None
#     internal_assessments = InternalAssessment.objects.none()
#     config_rows = []
#     max_iat_columns = 0

#     theory_master = None
#     practical_master = None
#     theory_lab_master = None

#     if selected_degree_id and selected_regulation_id:
#         configs = CourseHourConfig.objects.filter(
#             degree_id=selected_degree_id,
#             regulation_id=selected_regulation_id
#         ).order_by("id")

#         internal_assessments = list(
#             InternalAssessment.objects.filter(
#                 degree_id=selected_degree_id
#             ).order_by("iat")
#         )

#         if selected_config_id:
#             selected_ltp = configs.filter(id=selected_config_id).first()

#         if not selected_ltp:
#             selected_ltp = configs.first()

#         for cfg in configs:
#             base_iats = list(internal_assessments)

#             if int(cfg.lecture_hours or 0) != 0 and int(cfg.laboratory_hours or 0) != 0:
#                 iat_columns = base_iats + base_iats
#             else:
#                 iat_columns = base_iats

#             max_iat_columns = max(max_iat_columns, len(iat_columns))

#             config_rows.append({
#                 "cfg": cfg,
#                 "iat_columns": iat_columns,
#             })

#         for row in config_rows:
#             missing = max_iat_columns - len(row["iat_columns"])
#             if missing > 0:
#                 row["iat_columns"].extend([None] * missing)

#         theory_master = InternalAssessmentMasterTemplate.objects.filter(
#             degree_id=selected_degree_id,
#             regulation_id=selected_regulation_id,
#             course_type="theory"
#         ).first()

#         practical_master = InternalAssessmentMasterTemplate.objects.filter(
#             degree_id=selected_degree_id,
#             regulation_id=selected_regulation_id,
#             course_type="practical"
#         ).first()

#         theory_lab_master = InternalAssessmentMasterTemplate.objects.filter(
#             degree_id=selected_degree_id,
#             regulation_id=selected_regulation_id,
#             course_type="theory_lab"
#         ).first()

#     if request.method == "POST":
#         action = request.POST.get("action")

#         lecture = request.POST.get("lecture_hours", 0)
#         tutorial = request.POST.get("tutorial_hours", 0)
#         laboratory = request.POST.get("laboratory_hours", 0)

#         theory_pct = request.POST.get("theory_percentage", 0)
#         practical_pct = request.POST.get("practical_percentage", 0)
#         activity_pct = request.POST.get("activity_percentage", 0)

#         degree_id = request.POST.get("degree_id") or selected_degree_id
#         regulation_id = request.POST.get("regulation_id") or selected_regulation_id

#         if action == "create":
#             CourseHourConfig.objects.create(
#                 degree_id=degree_id,
#                 regulation_id=regulation_id,
#                 lecture_hours=lecture,
#                 tutorial_hours=tutorial,
#                 laboratory_hours=laboratory,
#                 theory_percentage=theory_pct,
#                 practical_percentage=practical_pct,
#                 activity_percentage=activity_pct,
#             )
#             messages.success(request, "LTP entry added successfully")

#         elif action == "update":
#             CourseHourConfig.objects.filter(
#                 id=request.POST.get("config_id")
#             ).update(
#                 lecture_hours=lecture,
#                 tutorial_hours=tutorial,
#                 laboratory_hours=laboratory,
#                 theory_percentage=theory_pct,
#                 practical_percentage=practical_pct,
#                 activity_percentage=activity_pct,
#             )
#             messages.success(request, "LTP entry updated successfully")

#         elif action == "delete":
#             CourseHourConfig.objects.filter(
#                 id=request.POST.get("config_id")
#             ).delete()
#             messages.success(request, "LTP entry deleted successfully")

#         elif action == "save_theory_course_master":
#             obj, created = InternalAssessmentMasterTemplate.objects.update_or_create(
#                 degree_id=degree_id,
#                 regulation_id=regulation_id,
#                 course_type="theory",
#                 defaults={
#                     "assessment1_assignment": request.POST.get("theory_assessment1_assignment") or 0,
#                     "assessment1_test": request.POST.get("theory_assessment1_test") or 0,
#                     "assessment2_assignment": request.POST.get("theory_assessment2_assignment") or 0,
#                     "assessment2_test": request.POST.get("theory_assessment2_test") or 0,
#                     "total_internal": request.POST.get("theory_total_internal") or 0,
#                 }
#             )
#             messages.success(request, "Theory Course master saved successfully")

#         elif action == "save_practical_course_master":
#             obj, created = InternalAssessmentMasterTemplate.objects.update_or_create(
#                 degree_id=degree_id,
#                 regulation_id=regulation_id,
#                 course_type="practical",
#                 defaults={
#                     "assessment1_assignment": request.POST.get("practical_observation_record") or 0,
#                     "assessment1_test": request.POST.get("practical_test") or 0,
#                     "assessment2_assignment": 0,
#                     "assessment2_test": 0,
#                     "total_internal": 100,
#                 }
#             )
#             messages.success(request, "Practical Course master saved successfully")

#         elif action == "save_practical_course_master":
#             record_val = int(request.POST.get("practical_observation_record") or 0)
#             test_val = int(request.POST.get("practical_test") or 0)

#             obj, created = InternalAssessmentMasterTemplate.objects.update_or_create(
#                 degree_id=degree_id,
#                 regulation_id=regulation_id,
#                 course_type="practical",
#                 defaults={
#                     "assessment1_assignment": record_val,
#                     "assessment1_test": test_val,
#                     "assessment2_assignment": 0,
#                     "assessment2_test": 0,
#                     "total_internal": record_val + test_val,
#                 }
#             )
#             messages.success(request, "Theory with Laboratory Course master saved successfully")

#         return redirect(
#             f"{request.path}?degree={degree_id}&regulation={regulation_id}"
#         )

#     return render(request, "examination_management/admin/LTP.html", {
#         "degrees": degrees,
#         "regulations": regulations,
#         "selected_degree_id": selected_degree_id,
#         "selected_regulation_id": selected_regulation_id,
#         "configs": configs,
#         "selected_ltp": selected_ltp,
#         "internal_assessments": internal_assessments,
#         "config_rows": config_rows,
#         "max_iat_columns": max_iat_columns,

#         "theory_assessment1_assignment": getattr(theory_master, "assessment1_assignment", 40),
#         "theory_assessment1_test": getattr(theory_master, "assessment1_test", 60),
#         "theory_assessment2_assignment": getattr(theory_master, "assessment2_assignment", 40),
#         "theory_assessment2_test": getattr(theory_master, "assessment2_test", 60),
#         "theory_total_internal": getattr(theory_master, "total_internal", 200),

#         "practical_observation_record": getattr(practical_master, "assessment1_assignment", 75),
#         "practical_test": getattr(practical_master, "assessment1_test", 25),

#         "theory_lab_assessment1_assignment": getattr(theory_lab_master, "assessment1_assignment", 40),
#         "theory_lab_assessment1_test": getattr(theory_lab_master, "assessment1_test", 60),
#         "theory_lab_assessment2_record": getattr(theory_lab_master, "assessment2_assignment", 75),
#         "theory_lab_assessment2_test": getattr(theory_lab_master, "assessment2_test", 25),
#         "theory_lab_total_internal": getattr(theory_lab_master, "total_internal", 200),
#     })


from django.contrib import messages
from django.shortcuts import render, redirect

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def export_ltp_excel(request):


    degree_id = request.GET.get("degree")
    regulation_id = request.GET.get("regulation")

    queryset = CourseHourConfig.objects.select_related(
        "degree", "regulation"
    ).filter(
        degree_id=degree_id,
        regulation_id=regulation_id
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "CourseHourConfig"

    # ✅ DB column names (customized)
    fields = [
        
        "degree",
        "regulation",
        "lecture_hours",
        "tutorial_hours",
        "laboratory_hours",
        "theory_percentage",
        "practical_percentage",
        "activity_percentage",
    ]

    # Header
    ws.append(fields)

    # Data rows
    for obj in queryset:
        row = [
            
            str(obj.degree) if obj.degree else "",
            str(obj.regulation.year) if obj.regulation else "",
            obj.lecture_hours,
            obj.tutorial_hours,
            obj.laboratory_hours,
            obj.theory_percentage,
            obj.practical_percentage,
            obj.activity_percentage,
        ]
        ws.append(row)

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="course_hour_config.xlsx"'

    wb.save(response)
    return response



def ltp(request):
    degrees = Degree.objects.filter(is_active=True)
    regulations = Regulations.objects.all()

    selected_degree_id = request.GET.get("degree")
    selected_regulation_id = request.GET.get("regulation")
    selected_config_id = request.GET.get("selected_config")

    configs = CourseHourConfig.objects.none()
    selected_ltp = None
    internal_assessments = InternalAssessment.objects.none()
    config_rows = []
    max_iat_columns = 0

    theory_master = None
    practical_master = None
    theory_lab_master = None

    if selected_degree_id and selected_regulation_id:
        configs = CourseHourConfig.objects.filter(
            degree_id=selected_degree_id,
            regulation_id=selected_regulation_id
        ).order_by("id")

        internal_assessments = list(
            InternalAssessment.objects.filter(
                degree_id=selected_degree_id
            ).order_by("iat")
        )

        if selected_config_id:
            selected_ltp = configs.filter(id=selected_config_id).first()

        if not selected_ltp:
            selected_ltp = configs.first()

        for cfg in configs:
            base_iats = list(internal_assessments)

            if int(cfg.lecture_hours or 0) != 0 and int(cfg.laboratory_hours or 0) != 0:
                iat_columns = base_iats + base_iats
            else:
                iat_columns = base_iats

            max_iat_columns = max(max_iat_columns, len(iat_columns))

            config_rows.append({
                "cfg": cfg,
                "iat_columns": iat_columns,
            })

        for row in config_rows:
            missing = max_iat_columns - len(row["iat_columns"])
            if missing > 0:
                row["iat_columns"].extend([None] * missing)

        theory_master = InternalAssessmentMasterTemplate.objects.filter(
            degree_id=selected_degree_id,
            regulation_id=selected_regulation_id,
            course_type="theory"
        ).first()

        practical_master = InternalAssessmentMasterTemplate.objects.filter(
            degree_id=selected_degree_id,
            regulation_id=selected_regulation_id,
            course_type="practical"
        ).first()

        theory_lab_master = InternalAssessmentMasterTemplate.objects.filter(
            degree_id=selected_degree_id,
            regulation_id=selected_regulation_id,
            course_type="theory_lab"
        ).first()

    if request.method == "POST":
        action = request.POST.get("action")

        lecture = request.POST.get("lecture_hours", 0)
        tutorial = request.POST.get("tutorial_hours", 0)
        laboratory = request.POST.get("laboratory_hours", 0)

        theory_pct = request.POST.get("theory_percentage", 0)
        practical_pct = request.POST.get("practical_percentage", 0)
        activity_pct = request.POST.get("activity_percentage", 0)

        degree_id = request.POST.get("degree_id") or selected_degree_id
        regulation_id = request.POST.get("regulation_id") or selected_regulation_id

        if action == "create":
            CourseHourConfig.objects.create(
                degree_id=degree_id,
                regulation_id=regulation_id,
                lecture_hours=lecture,
                tutorial_hours=tutorial,
                laboratory_hours=laboratory,
                theory_percentage=theory_pct,
                practical_percentage=practical_pct,
                activity_percentage=activity_pct,
            )
            messages.success(request, "LTP entry added successfully")

        elif action == "update":
            CourseHourConfig.objects.filter(
                id=request.POST.get("config_id")
            ).update(
                lecture_hours=lecture,
                tutorial_hours=tutorial,
                laboratory_hours=laboratory,
                theory_percentage=theory_pct,
                practical_percentage=practical_pct,
                activity_percentage=activity_pct,
            )
            messages.success(request, "LTP entry updated successfully")

        elif action == "delete":
            CourseHourConfig.objects.filter(
                id=request.POST.get("config_id")
            ).delete()
            messages.success(request, "LTP entry deleted successfully")

        elif action == "save_theory_course_master":
            obj, created = InternalAssessmentMasterTemplate.objects.update_or_create(
                degree_id=degree_id,
                regulation_id=regulation_id,
                course_type="theory",
                defaults={
                    "assessment1_assignment": request.POST.get("theory_assessment1_assignment") or 0,
                    "assessment1_test": request.POST.get("theory_assessment1_test") or 0,
                    "assessment2_assignment": request.POST.get("theory_assessment2_assignment") or 0,
                    "assessment2_test": request.POST.get("theory_assessment2_test") or 0,
                    "total_internal": request.POST.get("theory_total_internal") or 0,
                }
            )
            messages.success(request, "Theory Course master saved successfully")

        elif action == "save_practical_course_master":
            obj, created = InternalAssessmentMasterTemplate.objects.update_or_create(
                degree_id=degree_id,
                regulation_id=regulation_id,
                course_type="practical",
                defaults={
                    "assessment1_assignment": request.POST.get("practical_observation_record") or 0,
                    "assessment1_test": request.POST.get("practical_test") or 0,
                    "assessment2_assignment": 0,
                    "assessment2_test": 0,
                    "total_internal": 100,
                }
            )
            messages.success(request, "Practical Course master saved successfully")

        elif action == "save_theory_lab_course_master":
            a1_assignment = int(request.POST.get("theory_lab_assessment1_assignment") or 0)
            a1_test = int(request.POST.get("theory_lab_assessment1_test") or 0)
            a2_record = int(request.POST.get("theory_lab_assessment2_record") or 0)
            a2_test = int(request.POST.get("theory_lab_assessment2_test") or 0)

            obj, created = InternalAssessmentMasterTemplate.objects.update_or_create(
                degree_id=degree_id,
                regulation_id=regulation_id,
                course_type="theory_lab",
                defaults={
                    "assessment1_assignment": a1_assignment,
                    "assessment1_test": a1_test,
                    "assessment2_assignment": a2_record,
                    "assessment2_test": a2_test,
                    "total_internal": a1_assignment + a1_test + a2_record + a2_test,
                }
            )
            messages.success(request, "Theory with Laboratory Course master saved successfully")

        return redirect(
            f"{request.path}?degree={degree_id}&regulation={regulation_id}"
        )

    return render(request, "examination_management/admin/LTP.html", {
        "degrees": degrees,
        "regulations": regulations,
        "selected_degree_id": selected_degree_id,
        "selected_regulation_id": selected_regulation_id,
        "configs": configs,
        "selected_ltp": selected_ltp,
        "internal_assessments": internal_assessments,
        "config_rows": config_rows,
        "max_iat_columns": max_iat_columns,

        "theory_assessment1_assignment": getattr(theory_master, "assessment1_assignment", 40),
        "theory_assessment1_test": getattr(theory_master, "assessment1_test", 60),
        "theory_assessment2_assignment": getattr(theory_master, "assessment2_assignment", 40),
        "theory_assessment2_test": getattr(theory_master, "assessment2_test", 60),
        "theory_total_internal": getattr(theory_master, "total_internal", 200),

        "practical_observation_record": getattr(practical_master, "assessment1_assignment", 75),
        "practical_test": getattr(practical_master, "assessment1_test", 25),

        "theory_lab_assessment1_assignment": getattr(theory_lab_master, "assessment1_assignment", 40),
        "theory_lab_assessment1_test": getattr(theory_lab_master, "assessment1_test", 60),
        "theory_lab_assessment2_record": getattr(theory_lab_master, "assessment2_assignment", 75),
        "theory_lab_assessment2_test": getattr(theory_lab_master, "assessment2_test", 25),
        "theory_lab_total_internal": getattr(theory_lab_master, "total_internal", 200),
    })


    


from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.views.decorators.http import require_POST

from examination_management.models import  Regular_Course_Grade_Master


# views.py
from django.shortcuts import render, redirect, get_object_or_404
from examination_management.models import Regular_Course_Grade_Master, Degree, Self_Learning_Course_Grade_Master

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q

from django.core.paginator import Paginator
from django.db.models import Q

@check_permission("regular_course_grade_master")
def regular_course_grade_master(request):

    if request.method == "POST":

        action = request.POST.get("action")

        if action == "save":

            grade_id = request.POST.get("grade_id")
            degree_id = request.POST.get("degree")
            regulation_id = request.POST.get("regulation")
            letter = request.POST.get("letter_grade").strip().upper()
            desc = request.POST.get("description")
            points = request.POST.get("grade_points")
            mark_from = request.POST.get("mark_from")
            mark_to = request.POST.get("mark_to")

            is_fail_grade = request.POST.get("is_fail_grade") == "on"
            is_active = request.POST.get("is_active") == "on"

            degree = Degree.objects.get(id=degree_id)
            regulation = Regulations.objects.get(id=regulation_id)

            if grade_id:

                grade = Regular_Course_Grade_Master.objects.get(id=grade_id)

                grade.degree = degree
                grade.regulation = regulation
                grade.letter_grade = letter
                grade.description = desc
                grade.grade_points = points
                grade.mark_from = mark_from
                grade.mark_to = mark_to
                grade.is_fail_grade = is_fail_grade
                grade.is_active = is_active

                grade.save()

                messages.success(request, "Grade Updated Successfully")

            else:

                Regular_Course_Grade_Master.objects.create(
                    degree=degree,
                    regulation=regulation,
                    letter_grade=letter,
                    description=desc,
                    grade_points=points,
                    mark_from=mark_from,
                    mark_to=mark_to,
                    is_fail_grade=is_fail_grade,
                    is_active=is_active
                )

                messages.success(request, "Grade Created Successfully")


        elif action == "delete":

            grade_id = request.POST.get("grade_id")

            Regular_Course_Grade_Master.objects.filter(id=grade_id).delete()

            messages.success(request, "Grade Deleted Successfully")


        return redirect("regular_course_grade_master")


    search = request.GET.get("search")

    selected_degree = request.GET.get("degree")
    selected_regulation = request.GET.get("regulation")
    selected_grade = request.GET.get("grade")

    grades = Regular_Course_Grade_Master.objects.select_related(
        "degree", "regulation"
    ).order_by("letter_grade")
    if search:
        grades = grades.filter(

            Q(letter_grade__icontains=search) |
            Q(description__icontains=search) |
            Q(degree__degree__icontains=search) |
            Q(regulation__year__icontains=search)

        )

    if selected_degree:
        grades = grades.filter(degree_id=selected_degree)

    if selected_regulation:
        grades = grades.filter(regulation_id=selected_regulation)

    if selected_grade:
        grades = grades.filter(letter_grade=selected_grade)

    grades = grades.order_by("-id")

    paginator = Paginator(grades,10)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    grade_letters = (
        Regular_Course_Grade_Master.objects
        .values_list("letter_grade",flat=True)
        .distinct()
    )

    return render(

        request,
        "examination_management/regular_course_grade_master.html",
        {

            "grades": page_obj,

            "degrees": Degree.objects.filter(is_active=True),

            "regulations": Regulations.objects.all(),

            "selected_degree": selected_degree,
            "selected_regulation": selected_regulation,
            "selected_grade": selected_grade,

            "grade_letters": grade_letters,

            "page_obj": page_obj
        }
    )



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from examination_management.models import Self_Learning_Course_Grade_Master


@check_permission("self_learning_course_grade_master")
def self_learning_course_grade_master(request):
    degrees = Degree.objects.all()
    regulations = Regulations.objects.all()
    grades = Self_Learning_Course_Grade_Master.objects.select_related("degree", "regulation")

    if request.method == "POST":
        action = request.POST.get("action")

        try:
            if action == "save":
                grade_id = request.POST.get("grade_id")
                degree_id = request.POST.get("degree")
                regulation_id = request.POST.get("regulation")
                letter_grade = request.POST.get("letter_grade", "").strip()
                mark_from = request.POST.get("mark_from")
                mark_to = request.POST.get("mark_to")

                # Validation
                if not all([degree_id, regulation_id, letter_grade, mark_from, mark_to]):
                    messages.error(request, "All fields are required.")
                    return redirect("self_learning_course_grade_master")

                degree = get_object_or_404(Degree, id=degree_id)
                regulation = get_object_or_404(Regulations, id=regulation_id)

                if grade_id:
                    # UPDATE
                    grade = get_object_or_404(Self_Learning_Course_Grade_Master, id=grade_id)
                    grade.degree = degree
                    grade.regulation = regulation
                    grade.letter_grade = letter_grade
                    grade.mark_from = mark_from
                    grade.mark_to = mark_to
                    grade.save()
                    messages.success(request, "Grade updated successfully.")
                else:
                    # CREATE
                    Self_Learning_Course_Grade_Master.objects.create(
                        degree=degree,
                        regulation=regulation,
                        letter_grade=letter_grade,
                        mark_from=mark_from,
                        mark_to=mark_to
                    )
                    messages.success(request, "Grade added successfully.")

            elif action == "delete":
                grade_id = request.POST.get("grade_id")
                grade = get_object_or_404(Self_Learning_Course_Grade_Master, id=grade_id)
                grade.delete()
                messages.success(request, "Grade deleted successfully.")

        except Exception as e:
            messages.error(request, f"Error: {str(e)}")

        return redirect("self_learning_course_grade_master")

    return render(
        request,
        "examination_management/self_learning_course_grade_master.html",
        {
            "degrees": degrees,
            "regulations": regulations,
            "grades": grades,
        }
    )



from faculty_management.models import general_information
from course_management.models import  Program_outcomes
@check_permission("program_outcome")
def program_outcome(request):
    # If an outcome is being edited, populate the form with the current data
    outcome_to_edit = None
    if 'edit' in request.GET:
        outcome_id = request.GET.get('edit')
        outcome_to_edit = get_object_or_404(Program_outcomes, id=outcome_id)

    # Handle POST request for creating a new Program Outcome or updating an existing one
    if request.method == 'POST':
        if 'create' in request.POST:
            program_number = request.POST.get('program_number')
            program_name = request.POST.get('program_name')
            program_description = request.POST.get('program_description')
            is_revised = 'is_revised' in request.POST
            is_active = 'is_active' in request.POST

            # Ensure all required data is provided
            if program_number and program_name and program_description:
                Program_outcomes.objects.create(
                    program_number=program_number,
                    program_name=program_name,
                    program_description=program_description,
                    is_revised=is_revised,
                    is_active=is_active
                )
                messages.success(request, "Program Outcome created successfully.")
                return redirect('program_outcome')

        if 'edit' in request.POST:
            outcome_id = request.POST.get('edit')
            program_number = request.POST.get('program_number')
            program_name = request.POST.get('program_name')
            program_description = request.POST.get('program_description')
            is_revised = 'is_revised' in request.POST
            is_active = 'is_active' in request.POST

            outcome = get_object_or_404(Program_outcomes, id=outcome_id)
            outcome.program_number = program_number
            outcome.program_name = program_name
            outcome.program_description = program_description
            outcome.is_revised = is_revised
            outcome.is_active = is_active
            outcome.save()

            messages.success(request, "Program Outcome updated successfully.")
            return redirect('program_outcome')

    if request.method == 'GET' and 'delete' in request.GET:
        outcome_id = request.GET.get('delete')
        outcome = get_object_or_404(Program_outcomes, id=outcome_id)
        outcome.delete()
        messages.success(request, "Program Outcome deleted successfully.")
        return redirect('program_outcome')

    # Fetch all program outcomes
    outcomes = Program_outcomes.objects.all()

    return render(request, "course_management/faculty/program_outcome.html", {
        'outcomes': outcomes,
        'outcome_to_edit': outcome_to_edit,  # Pass the data to the template if in edit mode
    })
    
    
    
    
    
    
from django.shortcuts import render, redirect, get_object_or_404
from django.conf import settings
from django.db.models import Min

from examination_management.models import Hall
from course_management.models import Hall_Allotment

from faculty_management.models import Add_Department
from student_management.models import StudentDetails

MAX_DEPTS_PER_HALL = 4  # ✅ Maximum departments allowed in one hall (same exam)


# ==========================================================
# HELPERS
# ==========================================================
def _all_db_aliases():
    return list(getattr(settings, "DATABASES", {}).keys()) or ["default"]


def _pick_db_alias_for_models():
    """
    Used only for dropdown defaults.
    (Actual student load uses a stronger cross-db finder if you use it later.)
    """
    for alias in _all_db_aliases():
        try:
            if StudentDetails.objects.using(alias).exists() or Add_Department.objects.using(alias).exists():
                return alias
        except Exception:
            continue
    return "default"


def _clean_list(values):
    out, seen = [], set()
    for v in values:
        if v is None:
            continue
        vv = str(v).strip()
        if vv and vv not in seen:
            seen.add(vv)
            out.append(vv)
    return sorted(out)


# ==========================================================
# ZIG-ZAG SEAT NUMBERING
# col1: 1..5, col2: 10..6, col3: 11..15, col4: 20..16, col5: 21..25
# ==========================================================
def _seat_no_pattern(r, c):
    base = (c - 1) * 5 + 1
    if c % 2 == 1:
        return base + (r - 1)
    return base + (5 - r)


# ==========================================================
# HALL ENTRY
# ==========================================================
def hall_entry(request):
    if request.method == "POST":
        hall_name = (request.POST.get("hall_name") or "").strip()
        benches = (request.POST.get("benches") or "").strip()

        if hall_name:
            try:
                benches_val = int(benches)
            except Exception:
                benches_val = 25

            if benches_val == 25:
                try:
                    Hall.objects.create(hall_name=hall_name, benches=25)
                except Exception:
                    pass

        return redirect("hall_entry")

    halls = Hall.objects.all().order_by("hall_name")
    return render(request, "examination_management/admin/hall_entry.html", {"halls": halls})


def delete_hall(request, id):
    hall = get_object_or_404(Hall, id=id)
    hall.delete()
    return redirect("hall_entry")

def hall_allotment(request):
    DB = (request.GET.get("db_alias") or "").strip() or _pick_db_alias_for_models()

    selected = {
        "regulation": (request.GET.get("regulation") or "").strip(),
        "degree_id": (request.GET.get("degree_id") or "").strip(),
        "department_id": (request.GET.get("department_id") or "").strip(),
        "batch": (request.GET.get("batch") or "").strip(),
        "year": (request.GET.get("year") or "").strip(),
        "semester": (request.GET.get("semester") or "").strip(),
        "exam_type": (request.GET.get("exam_type") or "SEMESTER").strip().upper(),
    }

    hall_id = (request.GET.get("hall_id") or "").strip()
    load = (request.GET.get("load") or "").strip() == "1"
    show_seats_flag = (request.GET.get("show_seats") or "").strip() == "1"

    regulations = list(StudentDetails.objects.using(DB).values_list("regulation", flat=True).distinct())
    regulations = sorted([r for r in regulations if r])

    batches = list(StudentDetails.objects.using(DB).values_list("batch", flat=True).distinct())
    batches = sorted([b for b in batches if b])

    years = list(StudentDetails.objects.using(DB).values_list("year", flat=True).distinct())
    years = sorted([y for y in years if y])

    semesters = list(StudentDetails.objects.using(DB).values_list("semester", flat=True).distinct())
    semesters = sorted([s for s in semesters if s])

    degrees = list(Degree.objects.using(DB).all().order_by("id"))

    dept_qs = Add_Department.objects.using(DB).all().order_by("Department")
    if selected["degree_id"]:
        try:
            dept_qs = dept_qs.filter(degree_id=int(selected["degree_id"]))
        except Exception:
            pass
    departments = list(dept_qs)

    halls = Hall.objects.all().order_by("hall_name")

    selected_hall = None
    if hall_id:
        try:
            selected_hall = Hall.objects.get(id=int(hall_id))
        except Exception:
            selected_hall = None

    required_ok = bool(
        selected["regulation"] and
        selected["batch"] and
        selected["year"] and
        selected["semester"]
    )

    exam_schedule = None
    selected_exam_date = None
    selected_session = ""
    selected_session_display = "-"

    if required_ok:
        exam_schedule = _get_exam_schedule(
            DB=DB,
            regulation=selected["regulation"],
            degree_id=selected["degree_id"],
            department_id=selected["department_id"],
            batch=selected["batch"],
            semester=selected["semester"],
        )

        if exam_schedule:
            selected_exam_date = exam_schedule.exam_date
            selected_session = exam_schedule.session or ""
            selected_session_display = exam_schedule.get_session_display() if exam_schedule.session else "-"

    allocated_students = []
    seat_map = {}
    allocated_count = 0
    hall_capacity = 25
    remaining_seats = 25
    allotted_ids = []

    if selected_hall and required_ok:
        alloc_qs = HallAllotment.objects.filter(
            hall=selected_hall,
            regulation__iexact=selected["regulation"],
            batch__iexact=selected["batch"],
            year__iexact=selected["year"],
            semester__iexact=selected["semester"],
            exam_type__iexact=selected["exam_type"],
        ).order_by("seat_no", "id")

        allocated_students = list(alloc_qs)
        allocated_count = alloc_qs.exclude(student_id__isnull=True).count()

        allotted_ids = list(
            alloc_qs.exclude(student_id__isnull=True).values_list("student_id", flat=True)
        )

        hall_capacity = int(getattr(selected_hall, "benches", 25) or 25)
        if hall_capacity <= 0:
            hall_capacity = 25

        remaining_seats = max(0, hall_capacity - allocated_count)

        for a in allocated_students:
            if getattr(a, "row_no", None) and getattr(a, "col_no", None):
                seat_map[(a.row_no, a.col_no)] = a

    student_rows = []
    if load:
        qs = StudentDetails.objects.using(DB).select_related("department").all()

        if selected["regulation"]:
            qs = qs.filter(regulation__iexact=selected["regulation"])
        if selected["batch"]:
            qs = qs.filter(batch__iexact=selected["batch"])
        if selected["year"]:
            qs = qs.filter(year__iexact=selected["year"])
        if selected["semester"]:
            qs = qs.filter(semester__iexact=selected["semester"])

        if selected["department_id"]:
            try:
                qs = qs.filter(department_id=selected["department_id"])
            except Exception:
                pass

        if selected["degree_id"] and not selected["department_id"]:
            try:
                qs = qs.filter(department__degree_id=int(selected["degree_id"]))
            except Exception:
                pass

        if allotted_ids:
            qs = qs.exclude(id__in=allotted_ids)

        qs = qs.order_by("reg_no").values("id", "reg_no", "name", "department__Department")[:3000]

        for row in qs:
            student_rows.append({
                "id": row["id"],
                "reg_no": (row.get("reg_no") or "").strip(),
                "name": (row.get("name") or "").strip(),
                "dept_name": (row.get("department__Department") or "-").strip(),
                "exam_date": selected_exam_date,
                "session": selected_session,
                "session_display": selected_session_display,
            })

        logger.warning(
            "HALL_ALLOT load=1 DB=%s filters=%s result_count=%s excluded_allotted=%s",
            DB, selected, len(student_rows), len(allotted_ids)
        )

    context = {
        "selected_db_alias": DB,
        "selected": selected,
        "regulations": regulations,
        "degrees": degrees,
        "departments": departments,
        "batches": batches,
        "years": years,
        "semesters": semesters,
        "halls": halls,
        "selected_hall": selected_hall,
        "student_rows": student_rows,
        "allocated_students": allocated_students,
        "seat_map": seat_map,
        "allocated_count": allocated_count,
        "remaining_seats": remaining_seats,
        "hall_capacity": hall_capacity,
        "show_seats_flag": show_seats_flag,
        "required_ok": required_ok,
        "exam_schedule": exam_schedule,
        "selected_exam_date": selected_exam_date,
        "selected_session": selected_session,
        "selected_session_display": selected_session_display,
    }
    return render(request, "examination_management/admin/hall_allotment.html", context)


from django.db import transaction
from django.shortcuts import redirect

def save_hall_allotment(request):
    if request.method != "POST":
        return redirect("hall_allotment")

    DB = (request.POST.get("db_alias") or "").strip() or _pick_db_alias_for_models()

    regulation = (request.POST.get("regulation") or "").strip()
    department_id = (request.POST.get("department_id") or "").strip()
    degree_id = (request.POST.get("degree_id") or "").strip()
    batch = (request.POST.get("batch") or "").strip()
    year = (request.POST.get("year") or "").strip()
    semester = (request.POST.get("semester") or "").strip()
    exam_type = (request.POST.get("exam_type") or "SEMESTER").strip().upper()
    hall_id = (request.POST.get("hall_id") or "").strip()

    exam_date = request.POST.get("exam_date")
    session = (request.POST.get("session") or "").strip()

    student_ids = request.POST.getlist("student_ids")

    if not (regulation and batch and year and semester and hall_id):
        return redirect("hall_allotment")

    if not student_ids:
        return redirect(
            f"/examination_management/hall-allotment/?regulation={regulation}"
            f"&degree_id={degree_id}&department_id={department_id}"
            f"&batch={batch}&year={year}&semester={semester}"
            f"&exam_type={exam_type}&hall_id={hall_id}&load=1&db_alias={DB}"
        )

    try:
        hall_obj = Hall.objects.get(id=int(hall_id))
    except Exception:
        return redirect("hall_allotment")

    hall_capacity = 25
    try:
        if getattr(hall_obj, "capacity", None):
            hall_capacity = int(hall_obj.capacity)
    except Exception:
        hall_capacity = 25

    existing_qs = HallAllotment.objects.filter(
        hall=hall_obj,
        regulation__iexact=regulation,
        batch__iexact=batch,
        year__iexact=year,
        semester__iexact=semester,
        exam_type__iexact=exam_type,
    ).exclude(row_no__isnull=True).exclude(col_no__isnull=True)

    existing_items = []
    existing_ids = set()

    for a in existing_qs.values(
        "student_id",
        "reg_no",
        "student_name",
        "degree",
        "department_id",
        "department_name",
        "row_no",
        "col_no",
    ):
        try:
            sid = int(a["student_id"])
        except Exception:
            continue

        existing_ids.add(sid)
        existing_items.append({
            "student_id": sid,
            "reg_no": (a.get("reg_no") or "").strip(),
            "student_name": (a.get("student_name") or "").strip(),
            "degree": (a.get("degree") or "").strip(),
            "department_id": a.get("department_id"),
            "department_name": (a.get("department_name") or "").strip(),
            "row_no": a.get("row_no"),
            "col_no": a.get("col_no"),
        })

    clean_selected_ids = []
    for s in student_ids:
        try:
            sid = int(s)
        except Exception:
            continue
        if sid not in existing_ids:
            clean_selected_ids.append(sid)

    selected_students = list(
        StudentDetails.objects.using(DB)
        .select_related("department")
        .filter(id__in=clean_selected_ids)
        .order_by("reg_no")
    )

    def get_student_dept_name(st):
        try:
            if st.department and st.department.Department:
                return (st.department.Department or "").strip()
        except Exception:
            pass
        return "Unknown"

    new_items = []
    for st in selected_students:
        new_items.append({
            "student_id": int(st.id),
            "reg_no": (getattr(st, "reg_no", "") or "").strip(),
            "student_name": (getattr(st, "name", "") or "").strip(),
            "degree": "",
            "department_id": getattr(st, "department_id", None),
            "department_name": get_student_dept_name(st),
        })

    remaining_slots = max(0, hall_capacity - len(existing_items))
    if remaining_slots <= 0:
        return redirect(
            f"/examination_management/hall-allotment/?regulation={regulation}"
            f"&degree_id={degree_id}&department_id={department_id}"
            f"&batch={batch}&year={year}&semester={semester}"
            f"&exam_type={exam_type}&hall_id={hall_id}&load=1&show_seats=1&db_alias={DB}"
        )

    if len(new_items) > remaining_slots:
        new_items = new_items[:remaining_slots]

    combined = existing_items + new_items

    dept_groups = {}
    for it in combined:
        dk = _norm_dept(it.get("department_name") or "") or "unknown"
        dept_groups.setdefault(dk, []).append(it)

    for dk in dept_groups:
        dept_groups[dk].sort(key=lambda x: (x.get("reg_no") or "", x.get("student_id") or 0))

    ordered = []
    while True:
        progressed = False
        for dk in list(dept_groups.keys()):
            if dept_groups[dk]:
                ordered.append(dept_groups[dk].pop(0))
                progressed = True
        if not progressed:
            break

    positions = _zigzag_positions_5x5()
    if hall_capacity < len(positions):
        positions = positions[:hall_capacity]

    if len(ordered) > len(positions):
        ordered = ordered[:len(positions)]

    def is_valid_position(r, c, dept_key, pos_map):
        neighbors = [(r, c - 1), (r, c + 1), (r - 1, c), (r + 1, c)]
        for nr, nc in neighbors:
            if (nr, nc) in pos_map and pos_map.get((nr, nc)) == dept_key:
                return False
        return True

    free_positions = list(positions)
    pos_map = {}
    assigned = []
    pending = []

    for it in ordered:
        dept_key = _norm_dept(it.get("department_name") or "") or "unknown"
        placed = False

        for idx, (r, c) in enumerate(free_positions):
            if is_valid_position(r, c, dept_key, pos_map):
                assigned.append((it, r, c))
                pos_map[(r, c)] = dept_key
                free_positions.pop(idx)
                placed = True
                break

        if not placed:
            pending.append(it)

    for it in pending:
        if not free_positions:
            break
        r, c = free_positions.pop(0)
        dept_key = _norm_dept(it.get("department_name") or "") or "unknown"
        assigned.append((it, r, c))
        pos_map[(r, c)] = dept_key

    with transaction.atomic():
        HallAllotment.objects.filter(
            hall=hall_obj,
            regulation__iexact=regulation,
            batch__iexact=batch,
            year__iexact=year,
            semester__iexact=semester,
            exam_type__iexact=exam_type,
        ).delete()

        to_create = []
        for it, r, c in assigned:
            seat_no = _seat_no_pattern(r, c)
            to_create.append(
                HallAllotment(
                    hall=hall_obj,
                    student_id=int(it["student_id"]),
                    reg_no=it.get("reg_no", "") or "",
                    student_name=it.get("student_name", "") or "",
                    degree=it.get("degree", "") or "",
                    department_id=it.get("department_id"),
                    department_name=it.get("department_name", "") or "",
                    regulation=regulation,
                    batch=batch,
                    year=year,
                    semester=semester,
                    exam_type=exam_type,
                    exam_date=exam_date,
                    session=session,
                    seat_no=seat_no,
                    row_no=r,
                    col_no=c,
                )
            )

        HallAllotment.objects.bulk_create(to_create, batch_size=200)

    return redirect(
        f"/examination_management/hall-allotment/?regulation={regulation}"
        f"&degree_id={degree_id}&department_id={department_id}"
        f"&batch={batch}&year={year}&semester={semester}"
        f"&exam_type={exam_type}&hall_id={hall_id}&load=1&show_seats=1&db_alias={DB}"
    )




# ==========================================================
# REMOVE ALLOTMENTS (POST)
# ==========================================================
def remove_hall_allotments(request):
    if request.method != "POST":
        return redirect("hall_allotment")

    regulation = (request.POST.get("regulation") or "").strip()
    batch = (request.POST.get("batch") or "").strip()
    year = (request.POST.get("year") or "").strip()
    semester = (request.POST.get("semester") or "").strip()
    exam_type = (request.POST.get("exam_type") or "SEMESTER").strip().upper()
    hall_id = (request.POST.get("hall_id") or "").strip()
    department_id = (request.POST.get("department_id") or "").strip()
    degree_id = (request.POST.get("degree_id") or "").strip()
    DB = (request.POST.get("db_alias") or "").strip() or _pick_db_alias_for_models()

    if not hall_id:
        return redirect("hall_allotment")

    hall_obj = get_object_or_404(Hall, id=int(hall_id))

    HallAllotment.objects.filter(
        hall=hall_obj,
        regulation__iexact=regulation,
        batch__iexact=batch,
        year__iexact=year,
        semester__iexact=semester,
        exam_type__iexact=exam_type,
    ).delete()

    return redirect(
        f"/examination_management/hall-allotment/?regulation={regulation}"
        f"&department_id={department_id}&batch={batch}&year={year}&semester={semester}"
        f"&exam_type={exam_type}&hall_id={hall_id}&load=1"
    )
    
import io
import os
from datetime import datetime
from collections import Counter

from django.conf import settings
from django.http import FileResponse, HttpResponse
from django.contrib.staticfiles import finders
from django.shortcuts import get_object_or_404

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame,
    Table, TableStyle, Paragraph, Spacer, KeepInFrame, PageBreak
)


def hall_entry_sem(request):
    if request.method == "POST":
        hall_name = (request.POST.get("hall_name") or "").strip()
        benches = (request.POST.get("benches") or "").strip()

        if hall_name:
            try:
                benches_val = int(benches)
            except Exception:
                benches_val = 25

            if benches_val <= 0:
                benches_val = 25

            try:
                Hall.objects.create(hall_name=hall_name, benches=benches_val)
            except Exception:
                pass

        return redirect("hall_entry_sem")

    halls = Hall.objects.all().order_by("hall_name")
    return render(request, "examination_management/admin/hall_entry.html", {"halls": halls})


def delete_hall(request, id):
    hall = get_object_or_404(Hall, id=id)
    hall.delete()
    return redirect("hall_entry_sem")


def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


def seat_no_pattern(row_idx: int, col_idx: int) -> int:
    base = (col_idx * 5) + 1
    is_even_col = (col_idx % 2 == 0)
    return (base + row_idx) if is_even_col else (base + (4 - row_idx))


# ==========================================================
# PDF: SEATING ARRANGEMENT (YOUR SAME CODE - unchanged)
# ==========================================================
def seating_arrangement_pdf(request):
    hall_id = (request.GET.get("hall_id") or "").strip()
    regulation = (request.GET.get("regulation") or "").strip()
    batch = (request.GET.get("batch") or "").strip()
    year = (request.GET.get("year") or "").strip()
    semester = (request.GET.get("semester") or "").strip()
    exam_type = (request.GET.get("exam_type") or "").strip().upper()

    if not hall_id:
        return HttpResponse("hall_id is required", status=400)

    hall = get_object_or_404(Hall, id=hall_id)

    qs = HallAllotment.objects.select_related("hall").filter(hall_id=hall.id)
    if regulation:
        qs = qs.filter(regulation=regulation)
    if batch:
        qs = qs.filter(batch=batch)
    if year:
        qs = qs.filter(year=year)
    if semester:
        qs = qs.filter(semester=semester)
    if exam_type:
        qs = qs.filter(exam_type=exam_type)

    qs = qs.order_by("seat_no", "id")
    allotments = list(qs)

    seat_map = {}
    dept_counter = Counter()

    for a in allotments:
        if a.seat_no:
            seat_map[int(a.seat_no)] = a

        dept_name = safe_str(getattr(a, "department_name", "")) or "-"
        if dept_name != "-":
            dept_counter[dept_name] += 1

    total_students = len(allotments)

    styles = getSampleStyleSheet()

    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1a4b8c")
    DARK_GRAY = colors.HexColor("#111827")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BORDER_BLUE = colors.HexColor("#bcd3ff")
    ACCENT_RED = colors.HexColor("#b91c1c")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle(
        "title_style", parent=styles["Normal"],
        fontSize=14, textColor=PRIMARY_BLUE, alignment=TA_CENTER,
        fontName="Helvetica-Bold", spaceAfter=2
    )
    hall_style = ParagraphStyle(
        "hall_style", parent=styles["Normal"],
        fontSize=10.2, textColor=SECONDARY_BLUE, alignment=TA_CENTER,
        fontName="Helvetica-Bold", spaceAfter=6
    )
    sub_style = ParagraphStyle(
        "sub_style", parent=styles["Normal"],
        fontSize=9.2, textColor=MEDIUM_GRAY, alignment=TA_CENTER,
        spaceAfter=8
    )

    seat_title_style = ParagraphStyle(
        "seat_title_style", parent=styles["Normal"],
        fontSize=9.0, textColor=SECONDARY_BLUE,
        alignment=TA_LEFT, fontName="Helvetica-Bold", leading=10
    )
    reg_style = ParagraphStyle(
        "reg_style", parent=styles["Normal"],
        fontSize=8.6, textColor=DARK_GRAY,
        alignment=TA_LEFT, fontName="Helvetica-Bold", leading=9
    )
    name_style = ParagraphStyle(
        "name_style", parent=styles["Normal"],
        fontSize=8.0, textColor=DARK_GRAY,
        alignment=TA_LEFT, leading=8.5
    )
    dept_style = ParagraphStyle(
        "dept_style", parent=styles["Normal"],
        fontSize=7.6, textColor=MEDIUM_GRAY,
        alignment=TA_LEFT, leading=8
    )

    dept_head_style = ParagraphStyle(
        "dept_head_style", parent=styles["Normal"],
        fontSize=10, textColor=PRIMARY_BLUE, alignment=TA_LEFT,
        fontName="Helvetica-Bold", spaceAfter=4
    )
    dept_item_style = ParagraphStyle(
        "dept_item_style", parent=styles["Normal"],
        fontSize=9, textColor=DARK_GRAY, alignment=TA_LEFT,
        leading=11
    )

    HEADER_HEIGHT = 44 * mm

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = A4
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 8 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(
                ImageReader(logo_path),
                left, top_y - 20 * mm,
                width=30 * mm, height=18 * mm,
                preserveAspectRatio=True, mask="auto"
            )

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawCentredString(center_x, top_y - 6 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "An Autonomous Institution")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 8.2)
        canvas.drawCentredString(center_x, top_y - 18.5 * mm, "Approved by AICTE, New Delhi")
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(center_x, top_y - 23 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
        canvas.drawCentredString(center_x, top_y - 27.5 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

        footer_y = 18 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.8)
        canvas.line(left, footer_y + 7 * mm, right, footer_y + 7 * mm)

        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 8)
        gen_time = datetime.now().strftime("%d %b %Y, %I:%M %p")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawCentredString(center_x, footer_y, f"Hall: {safe_str(hall.hall_name)}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")

        canvas.restoreState()

    buffer = io.BytesIO()
    doc = BaseDocTemplate(
        buffer,
        pagesize=A4,
        title=f"Seating Arrangement - {safe_str(hall.hall_name)}",
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        showBoundary=0
    )

    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin + 6 * mm,
        doc.width,
        doc.height - HEADER_HEIGHT + 8 * mm,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        id="normal"
    )
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

    elements = []
    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph("SEATING ARRANGEMENT (5 x 5)", title_style))
    elements.append(Paragraph(f"Hall: {safe_str(hall.hall_name)}", hall_style))

    subtitle_parts = []
    if regulation:
        subtitle_parts.append(f"Regulation: <b>{regulation}</b>")
    if batch:
        subtitle_parts.append(f"Batch: <b>{batch}</b>")
    if year:
        subtitle_parts.append(f"Year: <b>{year}</b>")
    if semester:
        subtitle_parts.append(f"Semester: <b>{semester}</b>")
    if exam_type:
        subtitle_parts.append(f"Exam: <b>{exam_type}</b>")

    if subtitle_parts:
        elements.append(Paragraph(" | ".join(subtitle_parts), sub_style))
    else:
        elements.append(Spacer(1, 2 * mm))

    elements.append(Paragraph("Department Count", dept_head_style))

    dept_table_data = [[
        Paragraph("<b>Department</b>", dept_item_style),
        Paragraph("<b>Count</b>", dept_item_style),
    ]]

    if dept_counter:
        dept_rows = sorted(dept_counter.items(), key=lambda x: (-x[1], x[0].lower()))
        for dname, cnt in dept_rows:
            dept_table_data.append([
                Paragraph(safe_str(dname), dept_item_style),
                Paragraph(str(cnt), dept_item_style),
            ])

    dept_table_data.append([
        Paragraph("<b>Total</b>", dept_item_style),
        Paragraph(f"<b>{total_students}</b>", dept_item_style),
    ])

    dept_tbl = Table(dept_table_data, colWidths=[doc.width - 30 * mm, 30 * mm])
    dept_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER_GRAY),
        ("BACKGROUND", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 1), (1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(dept_tbl)
    elements.append(Spacer(1, 5 * mm))

    CARD_W = doc.width / 5.0
    CARD_H = 28 * mm
    TOP_H = 5.8 * mm
    REG_H = 5.8 * mm
    NAME_H = 7.2 * mm
    DEPT_H = CARD_H - (TOP_H + REG_H + NAME_H)

    def seat_cell(seat_no: int):
        a = seat_map.get(seat_no)
        regno = safe_str(getattr(a, "reg_no", "")) if a else "-"
        student_name = safe_str(getattr(a, "student_name", "")) if a else "-"
        dept = safe_str(getattr(a, "department_name", "")) if a else "-"

        name_para = Paragraph(student_name if student_name else "-", name_style)
        name_box = KeepInFrame(
            CARD_W - 10, NAME_H,
            [name_para],
            mode="shrink",
            hAlign="LEFT",
            vAlign="TOP"
        )

        dept_para = Paragraph(dept if dept else "-", dept_style)
        dept_box = KeepInFrame(
            CARD_W - 10, DEPT_H,
            [dept_para],
            mode="shrink",
            hAlign="LEFT",
            vAlign="TOP"
        )

        inner = Table(
            [
                [Paragraph(f"Seat {seat_no}", seat_title_style)],
                [Paragraph(regno if regno else "-", reg_style)],
                [name_box],
                [dept_box],
            ],
            colWidths=[CARD_W - 6],
            rowHeights=[TOP_H, REG_H, NAME_H, DEPT_H],
        )
        inner.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))

        outer = Table([[inner]], colWidths=[CARD_W], rowHeights=[CARD_H])
        outer.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.8, BORDER_BLUE),
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return outer

    grid_data = []
    for r in range(5):
        row_cells = []
        for c in range(5):
            sn = seat_no_pattern(r, c)
            row_cells.append(seat_cell(sn))
        grid_data.append(row_cells)

    grid = Table(grid_data, colWidths=[CARD_W] * 5, rowHeights=[CARD_H] * 5)
    grid.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    elements.append(grid)

    doc.build(elements)
    buffer.seek(0)

    filename = f"Seating_{safe_str(hall.hall_name).replace(' ', '')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    resp = FileResponse(buffer, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp
# ==========================================================
# PDF: ABSENTEES STATEMENT  (YOUR SAME CODE - unchanged)
# ==========================================================
def absentees_statement_pdf(request):
    regulation = (request.GET.get("regulation") or "").strip()
    batch = (request.GET.get("batch") or "").strip()
    year = (request.GET.get("year") or "").strip()
    semester = (request.GET.get("semester") or "").strip()
    exam_type = (request.GET.get("exam_type") or "SEMESTER").strip().upper()

    date_of_exam = (request.GET.get("date_of_exam") or "").strip()  # optional
    session = (request.GET.get("session") or "").strip()            # optional

    qs = HallAllotment.objects.select_related("hall").all()
    if regulation:
        qs = qs.filter(regulation__iexact=regulation)
    if batch:
        qs = qs.filter(batch__iexact=batch)
    if year:
        qs = qs.filter(year__iexact=year)
    if semester:
        qs = qs.filter(semester__iexact=semester)
    if exam_type:
        qs = qs.filter(exam_type__iexact=exam_type)

    allotments = list(qs.order_by("hall__hall_name", "department_name", "seat_no", "id"))

    hall_map = defaultdict(lambda: defaultdict(list))
    for a in allotments:
        hall_name = safe_str(getattr(getattr(a, "hall", None), "hall_name", "")) or "-"
        dept_name = safe_str(getattr(a, "department_name", "")) or "-"
        hall_map[hall_name][dept_name].append(a)

    styles = getSampleStyleSheet()
    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    ACCENT_RED = colors.HexColor("#b91c1c")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle(
        "title_style", parent=styles["Normal"],
        fontSize=10.5, textColor=colors.black,
        alignment=TA_CENTER, fontName="Helvetica-Bold", spaceAfter=2
    )
    meta_style = ParagraphStyle(
        "meta_style", parent=styles["Normal"],
        fontSize=8.8, textColor=colors.black,
        alignment=TA_LEFT, leading=11
    )

    th_style = ParagraphStyle(
        "th_style", parent=styles["Normal"],
        fontSize=8.0, textColor=colors.black,
        alignment=TA_CENTER, fontName="Helvetica-Bold", leading=10
    )
    td_style = ParagraphStyle(
        "td_style", parent=styles["Normal"],
        fontSize=8.2, textColor=colors.black,
        alignment=TA_LEFT, leading=10, wordWrap="CJK"
    )
    td_center = ParagraphStyle("td_center", parent=td_style, alignment=TA_CENTER)

    HEADER_HEIGHT = 44 * mm

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = A4
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 8 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(
                ImageReader(logo_path),
                left, top_y - 20 * mm,
                width=30 * mm, height=18 * mm,
                preserveAspectRatio=True, mask="auto"
            )

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawCentredString(center_x, top_y - 6 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "An Autonomous Institution")

        canvas.setFillColor(colors.black)
        canvas.setFont("Helvetica-Bold", 9.5)
        canvas.drawCentredString(center_x, top_y - 17.5 * mm, "EXAMINATION CONTROL OFFICE")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 8.2)
        canvas.drawCentredString(center_x, top_y - 22.5 * mm, "Approved by AICTE, New Delhi")
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(center_x, top_y - 27 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
        canvas.drawCentredString(center_x, top_y - 31.5 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

        footer_y = 18 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.8)
        canvas.line(left, footer_y + 7 * mm, right, footer_y + 7 * mm)

        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 8)
        gen_time = datetime.now().strftime("%d %b %Y, %I:%M %p")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")

        canvas.restoreState()

    buffer = io.BytesIO()
    doc = BaseDocTemplate(
        buffer,
        pagesize=A4,
        title="Absentees Statement",
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        showBoundary=0
    )

    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin + 6 * mm,
        doc.width,
        doc.height - HEADER_HEIGHT + 8 * mm,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        id="normal"
    )
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

    elements = []
    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph("ABSENTEES STATEMENT FOR SEMESTER EXAMINATION", title_style))
    elements.append(Spacer(1, 2 * mm))

    meta_parts = []
    if date_of_exam:
        meta_parts.append(f"Date of Examination: <b>{date_of_exam}</b>")
    if regulation:
        meta_parts.append(f"Regulation: <b>{regulation}</b>")
    if batch:
        meta_parts.append(f"Batch: <b>{batch}</b>")
    if year:
        meta_parts.append(f"Year: <b>{year}</b>")
    if semester:
        meta_parts.append(f"Semester: <b>{semester}</b>")
    if exam_type:
        meta_parts.append(f"Exam Type: <b>{exam_type}</b>")
    if session:
        meta_parts.append(f"Session: <b>{session}</b>")

    if meta_parts:
        elements.append(Paragraph(" | ".join(meta_parts), meta_style))
        elements.append(Spacer(1, 3 * mm))

    data = [[
        Paragraph("Hall No", th_style),
        Paragraph("Semester & Branch", th_style),
        Paragraph("Total No. of Students", th_style),
        Paragraph("No. of Students Present", th_style),
        Paragraph("No. of Students Absent", th_style),
        Paragraph("Register Numbers of Absentees", th_style),
    ]]

    spans = []
    row_idx = 1

    if not hall_map:
        data.append([Paragraph("-", td_center)] * 6)
    else:
        for hall_name in sorted(hall_map.keys(), key=lambda x: (x or "").lower()):
            dept_dict = hall_map[hall_name]
            dept_names = sorted(dept_dict.keys(), key=lambda x: (x or "").lower())

            dept_count = len([d for d in dept_names if (d or "").strip() and d != "-"])
            if dept_count == 0:
                dept_count = len(dept_names)

            start = row_idx
            for dept_name in dept_names:
                total = len(dept_dict[dept_name])

                sem_branch = f"{safe_str(semester)} Sem. {dept_name}".strip()
                if sem_branch.startswith("Sem.") or sem_branch == "Sem.":
                    sem_branch = dept_name or "-"

                hall_cell = f"{hall_name} ({dept_count} Depts)"
                data.append([
                    Paragraph(hall_cell, td_center),
                    Paragraph(sem_branch or "-", td_style),
                    Paragraph(str(total), td_center),
                    Paragraph("", td_center),
                    Paragraph("", td_center),
                    Paragraph("", td_style),
                ])
                row_idx += 1

            end = row_idx - 1
            if end > start:
                spans.append(("SPAN", (0, start), (0, end)))

    col_widths = [
        28 * mm,
        54 * mm,
        24 * mm,
        24 * mm,
        24 * mm,
        doc.width - (28 + 54 + 24 + 24 + 24) * mm,
    ]

    tbl = Table(data, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ("GRID", (0, 0), (-1, -1), 0.9, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 1), (4, -1), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (5, 1), (5, -1), 10),
        ("BOTTOMPADDING", (5, 1), (5, -1), 10),
    ]
    style_cmds.extend(spans)

    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)

    doc.build(elements)
    buffer.seek(0)

    filename = f"Absentees_AllHalls_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    resp = FileResponse(buffer, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


# ==========================================================
# PDF: ABSENTEES SUMMARY (branch-wise)  (keep your existing)
# ==========================================================
def absentees_statement_summary_pdf(request):
    regulation = (request.GET.get("regulation") or "").strip()
    batch = (request.GET.get("batch") or "").strip()
    year = (request.GET.get("year") or "").strip()
    semester = (request.GET.get("semester") or "").strip()
    exam_type = (request.GET.get("exam_type") or "SEMESTER").strip().upper()

    date_of_exam = (request.GET.get("date_of_exam") or "").strip()

    absent_reg_raw = (request.GET.get("absent_reg") or "").strip()
    absent_set = set()
    if absent_reg_raw:
        absent_set = {x.strip() for x in absent_reg_raw.split(",") if x.strip()}

    qs = HallAllotment.objects.select_related("hall").all()
    if regulation:
        qs = qs.filter(regulation__iexact=regulation)
    if batch:
        qs = qs.filter(batch__iexact=batch)
    if year:
        qs = qs.filter(year__iexact=year)
    if semester:
        qs = qs.filter(semester__iexact=semester)
    if exam_type:
        qs = qs.filter(exam_type__iexact=exam_type)

    allotments = list(qs.order_by("department_name", "reg_no", "id"))

    def short_branch(name: str) -> str:
        n = (name or "").strip()
        if not n:
            return "-"
        low = n.lower()

        compact = n.replace("&", "").replace(" ", "")
        if len(n) <= 8 and compact.isalpha():
            return n.upper() if len(n) <= 4 else n.title()

        if "civil" in low:
            return "Civil"
        if "mechan" in low:
            return "Mech"
        if "information technology" in low or low.strip() == "it":
            return "IT"
        if "computer science and business systems" in low or "csbs" in low:
            return "CSBS"
        if "computer science" in low or "cse" in low:
            return "CSE"
        if "electronics and communication" in low or "ece" in low:
            return "ECE"
        if "electrical and electronics" in low or "eee" in low:
            return "EEE"
        if "artificial intelligence" in low and "data" in low:
            return "AD"
        if "artificial intelligence" in low:
            return "AI"
        if "data science" in low:
            return "DS"
        return (n.split()[0] or n).title()

    total_by_branch = Counter()
    abs_regs_by_branch = defaultdict(list)

    for a in allotments:
        dept_name = safe_str(getattr(a, "department_name", "")) or "-"
        br = short_branch(dept_name)
        total_by_branch[br] += 1

        rno = safe_str(getattr(a, "reg_no", "")) or ""
        if rno and rno in absent_set:
            abs_regs_by_branch[br].append(rno)

    for br in list(abs_regs_by_branch.keys()):
        abs_regs_by_branch[br] = sorted(set(abs_regs_by_branch[br]))

    styles = getSampleStyleSheet()
    BG_GRAY = colors.HexColor("#f8fafc")

    title_style = ParagraphStyle(
        "title_style", parent=styles["Normal"],
        fontSize=10.8, textColor=colors.black,
        alignment=TA_CENTER, fontName="Helvetica-Bold", spaceAfter=4
    )
    meta_style = ParagraphStyle(
        "meta_style", parent=styles["Normal"],
        fontSize=8.8, textColor=colors.black,
        alignment=TA_LEFT, leading=11
    )
    th_style = ParagraphStyle(
        "th_style", parent=styles["Normal"],
        fontSize=8.2, textColor=colors.black,
        alignment=TA_CENTER, fontName="Helvetica-Bold", leading=10
    )
    td_style = ParagraphStyle(
        "td_style", parent=styles["Normal"],
        fontSize=8.4, textColor=colors.black,
        alignment=TA_LEFT, leading=10, wordWrap="CJK"
    )
    td_center = ParagraphStyle("td_center", parent=td_style, alignment=TA_CENTER)

    HEADER_HEIGHT = 44 * mm

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = A4
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 8 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(
                ImageReader(logo_path),
                left, top_y - 20 * mm,
                width=30 * mm, height=18 * mm,
                preserveAspectRatio=True, mask="auto"
            )

        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawCentredString(center_x, top_y - 6 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "An Autonomous Institution")

        canvas.setFont("Helvetica", 8.2)
        canvas.drawCentredString(center_x, top_y - 18.5 * mm, "Approved by AICTE, New Delhi")
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(center_x, top_y - 23 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
        canvas.drawCentredString(center_x, top_y - 27.5 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

        footer_y = 18 * mm
        canvas.setFont("Helvetica", 8)
        gen_time = datetime.now().strftime("%d %b %Y, %I:%M %p")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")

        canvas.restoreState()

    buffer = io.BytesIO()
    doc = BaseDocTemplate(
        buffer,
        pagesize=A4,
        title="Absentees Statement Summary",
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        showBoundary=0
    )

    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin + 6 * mm,
        doc.width,
        doc.height - HEADER_HEIGHT + 8 * mm,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        id="normal"
    )
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

    elements = []
    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph("ABSENTEES STATEMENT FOR INTERNAL ASSESSMENT", title_style))

    meta_parts = []
    if date_of_exam:
        meta_parts.append(f"Date of Examination : <b>{date_of_exam}</b>")
    if regulation:
        meta_parts.append(f"Regulation : <b>{regulation}</b>")
    if batch:
        meta_parts.append(f"Batch : <b>{batch}</b>")
    if year:
        meta_parts.append(f"Year : <b>{year}</b>")
    if semester:
        meta_parts.append(f"Semester : <b>{semester}</b>")
    if exam_type:
        meta_parts.append(f"Exam Type : <b>{exam_type}</b>")

    if meta_parts:
        elements.append(Paragraph(" | ".join(meta_parts), meta_style))
        elements.append(Spacer(1, 3 * mm))

    data = [[
        Paragraph("Sl. No.", th_style),
        Paragraph("Branch", th_style),
        Paragraph("Total No. of Students", th_style),
        Paragraph("No. of Students Present", th_style),
        Paragraph("No. of Students Absent", th_style),
        Paragraph("Register Numbers of Absentees", th_style),
    ]]

    total_all = 0
    branches = sorted(total_by_branch.keys(), key=lambda x: (x or "").lower())

    if not branches:
        data.append([Paragraph("-", td_center)] * 6)
    else:
        for i, br in enumerate(branches, start=1):
            total = int(total_by_branch[br])
            total_all += total

            regs = abs_regs_by_branch.get(br, [])
            regs_text = "" if not regs else ", ".join(regs)

            data.append([
                Paragraph(str(i), td_center),
                Paragraph(br, td_center),
                Paragraph(str(total), td_center),
                Paragraph("", td_center),
                Paragraph("", td_center),
                Paragraph(regs_text, td_style),
            ])

        data.append([
            Paragraph("", td_center),
            Paragraph("Total No. of Students", td_style),
            Paragraph(str(total_all), td_center),
            Paragraph("", td_center),
            Paragraph("", td_center),
            Paragraph("", td_style),
        ])

    col_widths = [
        14 * mm,
        22 * mm,
        30 * mm,
        32 * mm,
        32 * mm,
        doc.width - (14 + 22 + 30 + 32 + 32) * mm,
    ]

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.9, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (0, 1), (4, -2), "CENTER"),
        ("ALIGN", (5, 1), (5, -2), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, BG_GRAY]),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (5, 1), (5, -2), 10),
        ("BOTTOMPADDING", (5, 1), (5, -2), 10),
    ]))

    elements.append(tbl)
    doc.build(elements)
    buffer.seek(0)

    filename = f"Absentees_Summary_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    resp = FileResponse(buffer, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp



from django.shortcuts import render, redirect
from django.contrib import messages
from examination_management.models import SquadQuestions
from examination_management.models import Regulations
import datetime

def generate_academic_years():
    """
    Generate academic years as choices (current year + next 5 years).
    """
    current_year = datetime.datetime.now().year
    academic_years = [
        (f"{start_year}-{start_year + 1}")
        for start_year in range(current_year - 3, current_year + 3)
    ]
    return academic_years


from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Q
from course_management.models import Regulations
from examination_management.models import SquadQuestions



def generate_academic_years():
    current_year = datetime.datetime.now().year
    return [
        f"{year}-{year+1}"
        for year in range(current_year, current_year + 6)
    ]

def add_squad_questions(request):

    academic_years = generate_academic_years()
    regulations = Regulations.objects.all()

    # ================= DELETE =================

    delete_id = request.GET.get("delete")

    if delete_id:
        SquadQuestions.objects.filter(id=delete_id).delete()
        messages.success(request, "Question deleted successfully")
        return redirect("add_squad_questions")

    # ================= TOGGLE ACTIVE =================

    toggle_id = request.GET.get("toggle")

    if toggle_id:

        q = SquadQuestions.objects.filter(id=toggle_id).first()

        if q:
            q.is_active = not q.is_active
            q.save()

        return redirect("add_squad_questions")

    # ================= EDIT =================

    edit_id = request.GET.get("edit")
    edit_question = None

    if edit_id:
        edit_question = SquadQuestions.objects.filter(id=edit_id).first()

    # ================= UPDATE =================

    if request.method == "POST" and request.POST.get("update_question_id"):

        qid = request.POST.get("update_question_id")
        text = request.POST.get("update_question")

        SquadQuestions.objects.filter(id=qid).update(
            question=text
        )

        messages.success(request, "Question updated successfully")

        return redirect("add_squad_questions")

    # ================= ADD QUESTIONS =================

    if request.method == "POST" and request.POST.getlist("question[]"):

        academic_year = request.POST.get("academic_year")
        regulation_id = request.POST.get("regulation")

        questions = request.POST.getlist("question[]")

        regulation = Regulations.objects.get(id=regulation_id)

        for q in questions:

            if q.strip():

                SquadQuestions.objects.create(
                    academic_year=academic_year,
                    regulation=regulation,
                    question=q
                )

        messages.success(request, "Questions added successfully")

        return redirect("add_squad_questions")

    # ================= FILTER =================

    filter_year = request.GET.get("academic_year")
    filter_reg = request.GET.get("regulation")
    search = request.GET.get("search")

    questions = SquadQuestions.objects.select_related("regulation")

    if filter_year:
        questions = questions.filter(academic_year=filter_year)

    if filter_reg:
        questions = questions.filter(regulation_id=filter_reg)

    if search:
        questions = questions.filter(
            question__icontains=search
        )

    # ================= GROUP QUESTIONS =================

    grouped_questions = {}

    for q in questions:

        key = (q.academic_year, q.regulation)

        if key not in grouped_questions:
            grouped_questions[key] = []

        grouped_questions[key].append(q)

    context = {

        "academic_years": academic_years,
        "regulations": regulations,
        "grouped_questions": grouped_questions,

        "filter_year": filter_year,
        "filter_reg": filter_reg,
        "search": search,

        "edit_question": edit_question
    }

    return render(
        request,
        "examination_management/admin/add_squad_questions.html",
        context
    )
    
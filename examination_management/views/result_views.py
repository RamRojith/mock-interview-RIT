from django.shortcuts import render,redirect
from django.http import JsonResponse
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from examination_management.decorators import examination_management
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.views.decorators.http import require_GET
from examination_management.models import ExaminationFunction, Result_Permission, Result, Regular_Course_Grade_Master, Self_Learning_Course_Grade_Master
from user_accounts.decorators import is_super_user, no_cache, faculty_login_required, check_permission
import re
from user_accounts.models import Role, Add_Department, StudentDetails
from faculty_management.models import general_information

def assign_result_permission(request):
    edit_permission = None

    # ---------- DELETE ----------
    if request.method == "POST" and request.POST.get("action") == "delete":
        perm_id = request.POST.get("perm_id")
        Result_Permission.objects.filter(id=perm_id).delete()
        messages.success(request, "Permission deleted")
        return redirect("assign_result_permission")

    # ---------- EDIT LOAD ----------
    if request.method == "GET" and request.GET.get("edit"):
        edit_permission = get_object_or_404(
            Result_Permission, id=request.GET.get("edit")
        )

    # ---------- CREATE / UPDATE ----------
    # ---------- CREATE / UPDATE ----------
    if request.method == "POST" and request.POST.get("action") == "save":
        role_ids = request.POST.getlist("roles[]")
        can_view_all = request.POST.get("can_view_all_results") == "on"
        can_view_dept = request.POST.get("can_view_department_results") == "on"
        perm_id = request.POST.get("perm_id")

        if not role_ids:
            messages.error(request, "At least one role is required")
            return redirect("assign_result_permission")

        # ---- EDIT (single row) ----
        if perm_id:
            Result_Permission.objects.filter(id=perm_id).update(
                can_view_all_results=can_view_all,
                can_view_department_results=can_view_dept,
            )
        else:
            # ---- CREATE (bulk roles) ----
            for role_id in role_ids:
                Result_Permission.objects.update_or_create(
                    role_id=role_id,
                    defaults={
                        "can_view_all_results": can_view_all,
                        "can_view_department_results": can_view_dept,
                    }
                )

        messages.success(request, "Permission saved successfully")
        return redirect("assign_result_permission")

    # ---------- PAGE LOAD ----------
    roles = Role.objects.using("rit_approval_system").all()

    context = {
        "roles": roles,
        "edit_permission": edit_permission,
    }
    return render(
        request,
        "examination_management/result/result_permission.html",
        context,
    )
 


 
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from django.shortcuts import get_object_or_404, render
# ---------- AJAX API ----------




@require_GET
def result_permission_api(request):
    search = request.GET.get("search", "").strip()
    page = int(request.GET.get("page", 1))

    permissions = Result_Permission.objects.all().order_by("id")

    # ---- FETCH ROLES FROM OTHER DB ----
    roles_qs = Role.objects.using("rit_approval_system").all()

    role_map = {r.id: r.role for r in roles_qs}

    # ---- IMPORTANT FIX HERE ----
    if search:
        role_ids = list(
            roles_qs.filter(role__icontains=search)
                    .values_list("id", flat=True)
        )

        # If no matching roles, return empty queryset fast
        if not role_ids:
            permissions = Result_Permission.objects.none()
        else:
            permissions = permissions.filter(role_id__in=role_ids)

    paginator = Paginator(permissions, 25)
    page_obj = paginator.get_page(page)

    data = [
        {
            "id": perm.id,
            "role": role_map.get(perm.role_id, "Unknown"),
            "can_view_all": perm.can_view_all_results,
            "can_view_dept": perm.can_view_department_results,
        }
        for perm in page_obj
    ]

    return JsonResponse({
        "results": data,
        "page": page_obj.number,
        "total_pages": paginator.num_pages,
        "has_next": page_obj.has_next(),
        "has_prev": page_obj.has_previous(),
    })




def calculate_sem_and_cum_gpa(student, upto_semester=None):
    """
    Returns:
    - semester_wise_data (dict)
    - cumulative_gpa_upto_sem (float)
    """

    results = Result.objects.filter(student=student)

    sem_data = {}
    cum_grade_total = 0.0
    cum_credit = 0.0

    # Decide semester range
    if upto_semester:
        semester_range = range(1, upto_semester + 1)
    else:
        semester_range = sorted(
            set(int(r.semester) for r in results if r.semester)
        )

    for sem in semester_range:
        sem_results = results.filter(semester=str(sem))

        sem_grade_total = 0.0
        sem_credit = 0.0
        enriched_results = []

        for result in sem_results:
            # Grade points from master
            grade_points = None
            if result.grade:
                try:
                    grade_master = Regular_Course_Grade_Master.objects.get(
                        letter_grade=result.grade.strip().upper()
                    )
                    grade_points = grade_master.grade_points
                except Regular_Course_Grade_Master.DoesNotExist:
                    grade_points = None

            # Credit
            try:
                credit_val = float(result.credit) if result.credit else 0.0
            except ValueError:
                credit_val = 0.0

            grade_total_val = result.grade_total if result.grade_total else 0.0

            enriched_results.append({
                "course": result.course,
                "credit": credit_val,
                "grade": result.grade,
                "grade_total": result.grade_total,
                "grade_points": grade_points,
            })

            sem_credit += credit_val
            sem_grade_total += grade_total_val

        # Semester GPA
        sem_gpa = sem_grade_total / sem_credit if sem_credit > 0 else None

        # Cumulative
        cum_credit += sem_credit
        cum_grade_total += sem_grade_total
        cum_gpa = cum_grade_total / cum_credit if cum_credit > 0 else None

        sem_data[sem] = {
            "results": enriched_results,
            "gpa": sem_gpa,
            "cum_gpa": cum_gpa,
            "sem_credit": sem_credit,
            "cum_credit": cum_credit,
        }

    return sem_data, cum_gpa



from course_management.models import Regulations

@check_permission("results")
def results(request):
    degrees = Degree.objects.all()
    departments = Add_Department.objects.values_list("degree", flat=True).distinct()
    regulations = Regulations.objects.all()
    batches = StudentDetails.objects.values_list("batch", flat=True).distinct().order_by("batch")
    sections = StudentDetails.objects.values_list("section", flat=True).distinct().order_by("section")
    
    context = {
        "degrees": degrees,
        "departments": departments,
        "regulations": regulations,
        "batches": batches,
        "sections": sections,
    }
    return render(request, "examination_management/result/results.html", context)

from django.core.paginator import Paginator
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from collections import defaultdict
from django.db.models import Q
import math

@login_required
@check_permission("results")
def results_data(request):

    degree     = request.GET.get("degree")
    department = request.GET.get("department")
    regulation = request.GET.get("regulation")
    year       = request.GET.get("year")
    semester   = request.GET.get("semester")
    section    = request.GET.get("section")
    batch      = request.GET.get("batch")
    search = request.GET.get("search")

    page_number = int(request.GET.get("page", 1))
    page_size   = int(request.GET.get("page_size", 25))

    user = request.user

    role = Role.objects.using("rit_approval_system").filter(id=user.role.id).first()
    if not role:
        return JsonResponse({"data": [], "total": 0, "pages": 1, "current_page": 1})

    try:
        faculty = general_information.objects.get(faculty_id=user.Employee_id)
    except general_information.DoesNotExist:
        return JsonResponse({"data": [], "total": 0, "pages": 1, "current_page": 1})

    permission = Result_Permission.objects.filter(role_id=role.id).first()
    
    students_qs = StudentDetails.objects.select_related(
        "department",
        "department__degree"
    ).order_by("reg_no")
    regulation_qs = Regulations.objects.get(id=regulation) if regulation else None
    if permission and permission.can_view_department_results:
        students_qs = students_qs.filter(department=faculty.department)

    # Filters
    if degree:
        students_qs = students_qs.filter(department__degree_id=degree)

    if department:
        students_qs = students_qs.filter(department_id=department)

    if regulation:
        students_qs = students_qs.filter(regulation=regulation_qs.year)

    if year:
        students_qs = students_qs.filter(year=year)

    if semester:
        students_qs = students_qs.filter(semester=semester)

    if section:
        students_qs = students_qs.filter(section=section)

    if batch:
        students_qs = students_qs.filter(batch=batch)
    if search:
        students_qs = students_qs.filter(
        Q(reg_no__icontains=search) |
        Q(name__icontains=search) |
        Q(department__Department__icontains=search) |
        Q(batch__icontains=search) |
        Q(section__icontains=search) |
        Q(regulation__icontains=search)
    )
    # PAGINATION FIRST
    paginator = Paginator(students_qs, page_size)
    page_obj = paginator.get_page(page_number)

    students = list(page_obj.object_list)

    student_ids = [s.id for s in students]

    # SINGLE QUERY FOR RESULTS
    results = Result.objects.filter(student_id__in=student_ids)

    results_map = defaultdict(list)

    for r in results:
        results_map[r.student_id].append(r)

    students_data = []

    for student in students:

        student_results = results_map.get(student.id, [])

        sem_credit_total = 0
        sem_grade_total = 0

        cum_credit = 0
        cum_grade = 0

        sem_gpa = None
        cgpa = None

        current_sem = int(student.semester or 0)
        current_year = math.ceil(current_sem / 2)

        for r in student_results:

            credit = float(r.credit or 0)
            grade_total = float(r.grade_total or 0)

            cum_credit += credit
            cum_grade += grade_total

            if int(r.semester or 0) == current_sem:
                sem_credit_total += credit
                sem_grade_total += grade_total

        if sem_credit_total:
            sem_gpa = round(sem_grade_total / sem_credit_total, 2)

        if cum_credit:
            cgpa = round(cum_grade / cum_credit, 2)

        semesters = []

        for i in range(1, current_sem + 1):

            year = math.ceil(i / 2)

            semesters.append({
                "sem": i,
                "year": year,
                "url": f"/examination_management/result/result_detail/{student.id}/{year}/{i}/"
            })

        students_data.append({
            "id": student.id,
            "reg_no": student.reg_no,
            "name": student.name,
            "department": f"{student.department.degree.degree_code} - {student.department}",
            "batch": student.batch,
            "year": student.year,
            "regulation": student.regulation,
            "semester": student.semester,
            "section": student.section,
            "sgpa": sem_gpa if sem_gpa else "-",
            "cgpa": cgpa if cgpa else "-",
            "semesters": semesters
        })

    return JsonResponse({
        "data": students_data,
        "total": paginator.count,
        "pages": paginator.num_pages,
        "current_page": page_obj.number,
        "has_next": page_obj.has_next(),
        "has_previous": page_obj.has_previous(),
    })



from django.http import HttpResponse
from openpyxl import Workbook
from collections import defaultdict
from django.contrib.auth.decorators import login_required
from course_management.models import Course

from django.http import HttpResponse
from openpyxl import Workbook
from collections import defaultdict
from django.contrib.auth.decorators import login_required


from django.http import HttpResponse
from openpyxl import Workbook
from collections import defaultdict
from django.contrib.auth.decorators import login_required


from django.http import HttpResponse
from openpyxl import Workbook
from collections import defaultdict
from django.contrib.auth.decorators import login_required
from course_management.models import Course, CourseEnrollment

from django.http import HttpResponse
from openpyxl import Workbook
from collections import defaultdict
from django.contrib.auth.decorators import login_required


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from collections import defaultdict
from django.contrib.auth.decorators import login_required


@login_required
@check_permission("results")
def export_result_analysis(request):

    degree     = request.GET.get("degree")
    department = request.GET.get("department")
    regulation = request.GET.get("regulation")
    year       = request.GET.get("year")
    semester   = request.GET.get("semester")
    section    = request.GET.get("section")
    batch      = request.GET.get("batch")
    department_obj = Add_Department.objects.filter(id=department).first() if department else None
    degree_obj = Degree.objects.filter(id=degree).first() if degree else None
    regulation_obj = Regulations.objects.filter(id=regulation).first() if regulation else None
    results = Result.objects.select_related("student", "course")

    if degree:
        results = results.filter(degree_id=degree)

    if department:
        results = results.filter(department_id=department)

    if regulation:
        results = results.filter(regulation=regulation)

    if year:
        results = results.filter(year=year)

    if semester:
        results = results.filter(semester=semester)

    if batch:
        results = results.filter(batch=batch)

    if section:
        results = results.filter(student__section=section)

    students = StudentDetails.objects.filter(
        id__in=results.values_list("student_id", flat=True).distinct()
    ).order_by("reg_no")

    # COURSES
    course_ids = CourseEnrollment.objects.filter(
        student_id__in=students.values_list("id", flat=True)
    ).values_list("course_id", flat=True).distinct()

    courses = Course.objects.filter(id__in=course_ids).order_by("course_code")

    # GRADES
    grades = Regular_Course_Grade_Master.objects.filter(is_active=True)
    grade_map = {g.letter_grade: g for g in grades}

    # RESULT MAP
    result_map = {}
    for r in results:
        result_map[(r.student_id, r.course_id)] = r

    # ANALYSIS STRUCTURE
    analysis = {}

    for c in courses:
        analysis[c.id] = {
            "code": f"{c.course_code} - {c.title}",
            "total": 0,
            "pass": 0,
            "fail": 0,
            "grades": defaultdict(int)
        }

    # EXCEL
    wb = Workbook()
    ws = wb.active
    ws.title = "Result Analysis"

    col_count = 3 + (len(courses)*2) + 3

    # COLLEGE NAME
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=col_count)
    cell = ws.cell(row=1,column=1)
    cell.value = "RAMCO INSTITUTE OF TECHNOLOGY, RAJAPALAYAM"
    cell.font = Font(bold=True,size=14)
    cell.alignment = Alignment(horizontal="center")

    # FILTER INFO
    filter_text = f"Degree:{degree_obj.degree_code if degree_obj else 'All'} | Department:{department_obj.Department if department_obj else 'All'} | Regulation:{regulation_obj.year if regulation_obj else 'All'} | Semester:{semester or 'All'} | Batch:{batch or 'All'}"

    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=col_count)
    cell = ws.cell(row=2,column=1)
    cell.value = filter_text
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

    # =============================
    # STUDENT RESULT TABLE
    # =============================

    row_start = 4

    header = ["Sl No","Reg No","Student Name"]

    for c in courses:
        header.append(f"{c.course_code} - {c.title}")
        header.append("Grade Points")

    header += ["Failed","Total Credit","SGPA"]

    ws.append(header)

    for col in range(1,len(header)+1):
        ws.cell(row=row_start,column=col).font = Font(bold=True)
        ws.cell(row=row_start,column=col).alignment = Alignment(horizontal="center")

    current_row = row_start+1

    for i,student in enumerate(students,start=1):

        row=[i,student.reg_no,student.name]

        total_credit=0
        total_points=0
        failed=0

        for course in courses:

            res=result_map.get((student.id,course.id))

            if res:

                grade=res.grade
                credit=float(res.credit or 0)

                grade_obj=grade_map.get(grade)

                gp=credit*grade_obj.grade_points if grade_obj else 0

                row.append(grade)
                row.append(gp)

                if grade_obj and grade_obj.is_fail_grade:
                    failed+=1

                total_credit+=credit
                total_points+=gp

                # ANALYSIS UPDATE
                analysis[course.id]["total"]+=1
                analysis[course.id]["grades"][grade]+=1

                if grade_obj and grade_obj.is_fail_grade:
                    analysis[course.id]["fail"]+=1
                else:
                    analysis[course.id]["pass"]+=1

            else:
                row+=["-","-"]

        sgpa=round(total_points/total_credit,2) if total_credit else 0

        row+=[failed,total_credit,sgpa]

        ws.append(row)
        current_row+=1

    # =============================
    # RESULT ANALYSIS
    # =============================

    ws.append([])
    ws.append([])
    ws.append(["RESULT ANALYSIS"])

    header=["Subject"]

    for a in analysis.values():
        header.append(a["code"])

    ws.append(header)

    # TOTAL
    row=["Total Candidates"]
    for a in analysis.values():
        row.append(a["total"])
    ws.append(row)

    # PASS
    row=["Students Passed"]
    for a in analysis.values():
        row.append(a["pass"])
    ws.append(row)

    # FAIL
    row=["Students Failed"]
    for a in analysis.values():
        row.append(a["fail"])
    ws.append(row)

    # PASS %
    row=["Pass %"]

    for a in analysis.values():
        percent=round((a["pass"]/a["total"])*100,2) if a["total"] else 0
        row.append(percent)

    ws.append(row)

    # =============================
    # GRADE DISTRIBUTION
    # =============================

    ws.append([])
    ws.append(["GRADE DISTRIBUTION"])

    grade_letters=[g.letter_grade for g in grades]

    header=["Grade"]+[a["code"] for a in analysis.values()]
    ws.append(header)

    for g in grade_letters:

        row=[g]

        for a in analysis.values():
            row.append(a["grades"].get(g,0))

        ws.append(row)

    # DOWNLOAD
    response=HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"]="attachment; filename=result_analysis.xlsx"

    wb.save(response)

    return response


from django.db.models import Count, Sum
from collections import defaultdict
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required


from user_accounts.models import Degree
def get_result_departments(request):

    degree_id = request.GET.get("degree")
    # print("Degree ID:", degree_id)  # Debugging line
    departments = Add_Department.objects.filter(
        degree_id=degree_id
    ).values("id", "Department")

    return JsonResponse(list(departments), safe=False)

def get_result_degree_structure(request):

    degree_id = request.GET.get("degree")

    degree = Degree.objects.filter(id=degree_id).first()

    years = []
    semesters = []

    if degree:

        for y in range(1, degree.duration + 1):
            years.append(y)

        for s in range(1, degree.duration * 2 + 1):
            semesters.append(s)

    return JsonResponse({
        "years": years,
        "semesters": semesters
    })



@check_permission("results")
def result_detail(request, student_id, year, semester):

    user = request.user

    try:
        faculty = general_information.objects.get(faculty_id=user.Employee_id)
    except general_information.DoesNotExist:
        return HttpResponseForbidden("Access denied")

    role = Role.objects.using("rit_approval_system").filter(id=user.role.id).first()
    permission = Result_Permission.objects.filter(role_id=role.id).first()

    student = get_object_or_404(StudentDetails, id=student_id)

    allowed = False
    if permission:
        if permission.can_view_all_results:
            allowed = True
        elif permission.can_view_department_results:
            allowed = student.department == faculty.department

    if not allowed:
        return HttpResponseForbidden("Not allowed")

    semester = int(semester)
    year = int(year)

    semester_data, cgpa_upto_sem = calculate_sem_and_cum_gpa(
        student=student,
        upto_semester=semester
    )

    current_sem_data = semester_data.get(semester, {})

    sgpa = current_sem_data.get("gpa")
    cgpa = current_sem_data.get("cum_gpa")

    return render(request, "examination_management/result/result_detail.html", {
        "student": student,
        "semester": semester,
        "year": year,
        "results": current_sem_data.get("results", []),
        "sgpa": round(sgpa, 2) if sgpa is not None else None,
        "cgpa": round(cgpa, 2) if cgpa is not None else None,
        "cum_credit": current_sem_data.get("cum_credit"),
    })


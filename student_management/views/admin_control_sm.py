from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.db.models import Count
from django.db import connection
from datetime import date
import pandas as pd
from io import BytesIO
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse

from django.urls import reverse
from django.contrib import messages
from django.shortcuts import redirect
from django.contrib.auth.decorators import login_required
from user_accounts.decorators import no_cache, is_super_user
from student_management.models import *
from django.shortcuts import render
from django.http import JsonResponse
import json
from django.views.decorators.csrf import csrf_exempt
from django.core.serializers.json import DjangoJSONEncoder
import re

from student_management.decorators import student_management

from user_accounts.models import Role, USER
from student_management.models import StudentManagementPermissions
from user_accounts.decorators import faculty_login_required, check_permission, no_cache
from course_management.models import Discontinued_Student




# @faculty_login_required
@student_management
def sm_home(request):
    # print("sm home page ")
    request.session['current_page'] = 'em_home'
    return redirect('home')

@check_permission('sm_hello')
def sm_hello(request):
    return render(request, "sm_hello.html")



@faculty_login_required
@no_cache
@is_super_user('student_management')
def sm_assign_permission(request):
    if request.method == 'POST':  
        permissions = request.POST
        for role_name, role_permissions in permissions.items():
            if role_name.startswith('permissions'):
                try:
                    # Extract data from role_name using regex
                    extract_data = list(re.findall(r'\[([^\]]+)\]', role_name))
                    if len(extract_data) < 2:  # Ensure there are at least role and function
                        messages.warning(request,f"Invalid format in role_name: {role_name}. Skipping.")
                        continue
                    
                    extract_data.append(role_permissions)

                    # Retrieve the role (Handle if the role does not exist)
                    try:
                        role = Role.objects.using("rit_approval_system").get(role=extract_data[0])

                    except Role.DoesNotExist:
                        messages.error(request,f"Role {extract_data[0]} does not exist.")
                        messages.error(request, f"Role '{extract_data[0]}' does not exist. Skipping this entry.")
                        continue
                    
                    # Parse permissions - handle the case where role_permissions is a list (unlikely with POST data)
                    if isinstance(role_permissions, list):  # Handle list case
                        role_permissions = role_permissions[0]
                    
                    # Convert permission to boolean (True/False)
                    permission = extract_data[2] == 'true'
       
                    
                    # Find or create ApprovalPermissionFunction object
                    permission_obj = StudentManagementPermissions.objects.filter(
                        role=role, function=extract_data[1]
                    ).first()
                    
                    if permission_obj:
                        permission_obj.permission = permission
                        permission_obj.save()
                    else:
                        # Create a new ApprovalPermissionFunction object
                        StudentManagementPermissions.objects.create(
                            role=role,
                            function=extract_data[1],
                            permission=permission
                        )
                except Exception as e:
                    # Catch unexpected errors and log them
                    messages.error(request,f"Error processing role '{role_name}': {str(e)}")
                    messages.error(request, f"An error occurred while processing '{role_name}': {str(e)}")

    # Redirect to admin dashboard after processing
    messages.success(request,"The permission changes have been successfully applied.")
    return redirect('student_management')


from user_accounts.models import StudentDetails
from course_management.models import *



from django.shortcuts import render
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Count, Q


import os
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.shortcuts import render
from django.templatetags.static import static


def student_details(request):
    batches = StudentDetails.objects.values_list("batch", flat=True).distinct().order_by("batch")
    departments = Add_Department.objects.filter(is_active=True).order_by("-degree__degree_code")
    years = StudentDetails.objects.values_list("year", flat=True).distinct().order_by("year")
    semesters = StudentDetails.objects.values_list("semester", flat=True).distinct().order_by("semester")
    sections = (
        StudentDetails.objects
        .exclude(section__isnull=True)
        .exclude(section__exact="")
        .values_list("section", flat=True)
        .distinct()
        .order_by("section")
    )

    context = {
        "batches": batches,
        "departments": departments,
        "years": years,
        "semesters": semesters,
        "sections": sections,
    }
    return render(request, "student_management/admin/student_details.html", context)


def student_details_ajax(request):
    search = request.GET.get("search", "").strip()
    batch = request.GET.get("batch", "").strip()
    department = request.GET.get("department", "").strip()
    year = request.GET.get("year", "").strip()
    semester = request.GET.get("semester", "").strip()
    section = request.GET.get("section", "").strip()
    status = request.GET.get("status", "").strip()
    page = int(request.GET.get("page", 1))
    per_page = int(request.GET.get("per_page", 100))

    qs = StudentDetails.objects.select_related("department", "department__degree").order_by("semester", "year")

    if batch:
        qs = qs.filter(batch=batch)

    if department:
        qs = qs.filter(department_id=department)

    if year:
        qs = qs.filter(year=year)

    if semester:
        qs = qs.filter(semester=semester)

    if section:
        qs = qs.filter(section=section)

    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "inactive":
        qs = qs.filter(is_active=False)

    if search:
        qs = qs.filter(
            Q(reg_no__icontains=search) |
            Q(name__icontains=search) |
            Q(batch__icontains=search) |
            Q(regulation__icontains=search) |
            Q(year__icontains=search) |
            Q(semester__icontains=search) |
            Q(section__icontains=search) |
            Q(mode__icontains=search) |
            Q(department__Department__icontains=search) |
            Q(department__degree__degree_code__icontains=search)
        )

    qs = qs.order_by("semester", "year")

    total_count = StudentDetails.objects.count()
    filtered_count = qs.count()
    active_count = qs.filter(is_active=True).count()
    inactive_count = qs.filter(is_active=False).count()

    dept_counts = list(
        qs.values("department__id", "department__Department")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    year_counts = list(
        qs.values("year")
        .annotate(count=Count("id"))
        .order_by("year")
    )

    semester_counts = list(
        qs.values("semester")
        .annotate(count=Count("id"))
        .order_by("semester")
    )

    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(page)

    default_profile = static("images/profile.jpg")

    students = []
    for i, student in enumerate(page_obj.object_list, start=((page_obj.number - 1) * per_page) + 1):
        profile_img_url = default_profile

        if student.profile_img:
            try:
                if student.profile_img.name and student.profile_img.storage.exists(student.profile_img.name):
                    profile_img_url = student.profile_img.url
            except Exception:
                profile_img_url = default_profile

        students.append({
            "sl_no": i,
            "id": student.id,
            "reg_no": student.reg_no or "-",
            "name": student.name or "-",
            "batch": student.batch or "-",
            "regulation": student.regulation or "-",
            "department": (
                f"{student.department.degree.degree_code} - {student.department.Department}"
                if student.department and student.department.degree else
                (student.department.Department if student.department else "-")
            ),
            "year": student.year or "-",
            "semester": student.semester or "-",
            "section": student.section or "-",
            "mode": student.mode or "-",
            "is_active": student.is_active,
            "is_discontinued": student.is_discontinued or False,
            "status_label": "Active" if student.is_active else "Inactive",
            "edit_url": f"/student_management/student_management_admin/students/edit/{student.id}/",
            "profile_img": profile_img_url,
        })

    return JsonResponse({
        "students": students,
        "pagination": {
            "current_page": page_obj.number,
            "total_pages": paginator.num_pages,
            "has_previous": page_obj.has_previous(),
            "has_next": page_obj.has_next(),
            "previous_page_number": page_obj.previous_page_number() if page_obj.has_previous() else None,
            "next_page_number": page_obj.next_page_number() if page_obj.has_next() else None,
            "start_index": page_obj.start_index() if paginator.count else 0,
            "end_index": page_obj.end_index() if paginator.count else 0,
        },
        "counts": {
            "total_count": total_count,
            "filtered_count": filtered_count,
            "active_count": active_count,
            "inactive_count": inactive_count,
        },
        "analytics": {
            "dept_counts": dept_counts,
            "year_counts": year_counts,
            "semester_counts": semester_counts,
        }
    })


import traceback
from datetime import date
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction

@csrf_exempt
def discontinue_student_ajax(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid request method."}, status=405)

    try:
        student_id = request.POST.get("student_id")
        discontinued_date_str = request.POST.get("discontinued_date")
        year_str = request.POST.get("year_of_discontinuation")
        reason = request.POST.get("reason_for_discontinuation", "").strip()

        if not student_id or not discontinued_date_str or not year_str:
            messages.error(request, "Discontinued Date and Year of Discontinuation are required.")
            return JsonResponse({"success": False})

        discontinued_date_obj = date.fromisoformat(discontinued_date_str)
        year_int = int(year_str)

        with transaction.atomic():
            student = StudentDetails.objects.select_for_update().get(id=student_id)

            if student.is_discontinued:
                messages.warning(request, "This student has already been discontinued.")
                return JsonResponse({"success": False})

            Discontinued_Student.objects.create(
                student=student,
                department=student.department,
                discontinued_date=discontinued_date_obj,
                year_of_discontinuation=year_int,
                reason_for_discontinuation=reason or None,
            )

            student.is_discontinued = True
            student.is_active = False
            student.save(update_fields=["is_discontinued", "is_active"])

        messages.success(request, f"Student '{student.name}' discontinued successfully.")
        return JsonResponse({"success": True})

    except Exception as e:
        traceback.print_exc()
        messages.error(request, "Unable to discontinue student.")
        return JsonResponse({"success": False}, status=400)




def sync_student_reg_numbers(request):

    students = StudentDetails.objects.exclude(
        aadhar_number__isnull=True
    ).exclude(
        aadhar_number=""
    )

    updated = 0
    not_found = 0

    for student in students:

        personal = PersonalDetails.objects.using("admissionform1").filter(
            Aadhaar_Number=student.aadhar_number
        ).first()

        if personal:
            personal.registration_no = student.reg_no
            personal.save(update_fields=["registration_no"])
            updated += 1
        else:
            not_found += 1

    messages.success(
        request,
        f"Sync completed. Updated: {updated}, Not Found: {not_found}"
    )

    return redirect("student_details")

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

def export_student_basic_excel(request):
    search = request.GET.get("search", "").strip()
    batch = request.GET.get("batch", "").strip()
    department = request.GET.get("department", "").strip()
    year = request.GET.get("year", "").strip()
    semester = request.GET.get("semester", "").strip()
    section = request.GET.get("section", "").strip()
    status = request.GET.get("status", "").strip()

    qs = StudentDetails.objects.select_related("department", "department__degree").all()

    if batch:
        qs = qs.filter(batch=batch)

    if department:
        qs = qs.filter(department_id=department)

    if year:
        qs = qs.filter(year=year)

    if semester:
        qs = qs.filter(semester=semester)

    if section:
        qs = qs.filter(section=section)

    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "inactive":
        qs = qs.filter(is_active=False)

    if search:
        qs = qs.filter(
            Q(reg_no__icontains=search) |
            Q(name__icontains=search) |
            Q(batch__icontains=search) |
            Q(regulation__icontains=search) |
            Q(year__icontains=search) |
            Q(semester__icontains=search) |
            Q(section__icontains=search) |
            Q(mode__icontains=search) |
            Q(aadhar_number__icontains=search) |
            Q(department__Department__icontains=search) |
            Q(department__degree__degree_code__icontains=search)
        )

    qs = qs.order_by("semester", "year", "name")

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Export 2"

    headers = [
        "aadhar_number",
        "reg_no",
        "name",
        "department_code",
        "year",
        "semester",
        "section",
        "mode",
        "batch",
        "regulation", "date_of_birth",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for student in qs:
        department_code = ""
        if student.department and student.department.degree:
            department_code = student.department.Department_code
        dob = student.date_of_birth
        formatted_dob = dob.strftime("%d-%m-%Y") if hasattr(dob, "strftime") else ""
        ws.append([
            student.aadhar_number or "",
            student.reg_no or "",
            student.name or "",
            department_code,
            student.year or "",
            student.semester or "",
            student.section or "",
            student.mode or "",
            student.batch or "",
            student.regulation or "",
            formatted_dob,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="students_export_2.xlsx"'

    wb.save(response)
    return response




from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from user_accounts.models import StudentDetails, Add_Department
from course_management.models import SectionMaster

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

from user_accounts.models import StudentDetails, Add_Department
from course_management.models import SectionMaster

from django.contrib import messages
from django.core.files.base import ContentFile
from django.shortcuts import get_object_or_404, redirect, render
from django.templatetags.static import static

def edit_student_details(request, id):
    student = get_object_or_404(StudentDetails, id=id)
    departments = Add_Department.objects.filter(is_active=True)
    sections = SectionMaster.objects.all()

    years = ["1", "2", "3", "4"]
    semesters = ["1", "2", "3", "4", "5", "6", "7", "8"]

    if request.method == "POST":
        old_reg_no = student.reg_no
        old_email = student.email

        student.aadhar_number = request.POST.get("aadhar_number")
        student.name = request.POST.get("name")
        student.reg_no = request.POST.get("reg_no")
        student.batch = request.POST.get("batch")
        student.department_id = request.POST.get("department") or None
        student.is_active = request.POST.get("is_active") == "True"

        student.year = request.POST.get("year")
        student.semester = request.POST.get("semester")
        student.regulation = request.POST.get("regulation")

        student.year_of_admission = request.POST.get("year_of_admission") or None
        student.semester_of_admission = request.POST.get("semester_of_admission") or None

        student.section = request.POST.get("section")
        student.mode = request.POST.get("mode") or None
        student.email = request.POST.get("email")
        student.mobile_no = request.POST.get("mobile_no")

        dob = request.POST.get("date_of_birth")
        student.date_of_birth = dob if dob else None

        age = request.POST.get("age")
        student.age = int(age) if age else None

        student.gender = request.POST.get("gender")
        student.is_active = request.POST.get("is_active") == "True"

        profile_img = request.FILES.get("profile_img")
        remove_profile_img = request.POST.get("remove_profile_img") == "1"

        uploaded_file_name = None
        uploaded_file_bytes = None

        if profile_img:
            uploaded_file_name = profile_img.name
            uploaded_file_bytes = profile_img.read()
            student.profile_img.save(
                uploaded_file_name,
                ContentFile(uploaded_file_bytes),
                save=False
            )
        elif remove_profile_img:
            if student.profile_img:
                student.profile_img.delete(save=False)
            student.profile_img = None

        student.save()

        # Sync USER model in external DB
        user = None

        # Try matching with old values first, then new values
        if old_email:
            user = USER.objects.using("rit_approval_system").filter(email=old_email).first()

        if not user and old_reg_no:
            user = USER.objects.using("rit_approval_system").filter(username=old_reg_no).first()

        if not user and student.email:
            user = USER.objects.using("rit_approval_system").filter(email=student.email).first()

        if not user and student.reg_no:
            user = USER.objects.using("rit_approval_system").filter(username=student.reg_no).first()

        if user:
            user.username = user.username or student.reg_no
            user.email = student.email or user.email
            user.is_active = student.is_active

            if uploaded_file_name and uploaded_file_bytes:
                user.profile_img.save(
                    uploaded_file_name,
                    ContentFile(uploaded_file_bytes),
                    save=False
                )
            elif remove_profile_img:
                if user.profile_img:
                    user.profile_img.delete(save=False)
                user.profile_img = None

            user.save(using="rit_approval_system")
            messages.success(request, "✅ Student details and USER profile updated successfully.")
        else:
            messages.warning(request, "✅ Student details updated, but matching USER account was not found.")

        return redirect("student_details")

    profile_preview = static("images/profile.jpg")
    if student.profile_img:
        try:
            if student.profile_img.name and student.profile_img.storage.exists(student.profile_img.name):
                profile_preview = student.profile_img.url
        except Exception:
            profile_preview = static("images/profile.jpg")

    context = {
        "student": student,
        "departments": departments,
        "sections": sections,
        "years": years,
        "semesters": semesters,
        "profile_preview": profile_preview,
    }
    return render(request, "student_management/admin/edit_student_details.html", context)
from student_management.models import StudentDetails
from user_accounts.models import PersonalDetails, USER, AdmissionRecords, Department, Add_Department, AcademicDetails

from django.shortcuts import render
from openpyxl import Workbook, load_workbook


from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook, load_workbook
from datetime import datetime

from student_management.models import StudentDetails
from user_accounts.models import (
    PersonalDetails, USER, AdmissionRecords,
    Department, Add_Department, AcademicDetails
)
from django.db.models import Q

from django.contrib.auth.hashers import make_password

from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Q
from openpyxl import load_workbook, Workbook
from django.contrib.auth.hashers import make_password
from datetime import datetime
from django.utils import timezone

from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Q
from django.utils import timezone
from django.contrib.auth.hashers import make_password
from openpyxl import load_workbook, Workbook
from datetime import datetime, date
from openpyxl.utils.datetime import from_excel


import re
from decimal import Decimal, InvalidOperation
from datetime import datetime, date

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel

from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.hashers import make_password
from django.utils import timezone


def clean(value):
    if value in ["", None, "-"]:
        return None

    value = str(value).strip()

    if value in ["", "-"]:
        return None

    return value


def clean_number(value):
    """
    Supports:
    953622205001
    9.53622E+11
    1234 5678 9012
    1234-5678-9012
    """
    if value in ["", None, "-"]:
        return None

    value = str(value).strip()

    try:
        if "E" in value.upper():
            value = str(int(Decimal(value)))
    except (InvalidOperation, ValueError):
        pass

    value = re.sub(r"\D", "", value)

    return value or None


def clean_aadhar(value):
    value = clean_number(value)

    if not value:
        return None

    if len(value) != 12:
        return None

    return value


def parse_excel_date(value):
    if value in ["", None, "-"]:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except Exception:
            return None

    value = str(value).strip()

    formats = [
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%d.%m.%y",
        "%d/%m/%y",
        "%d-%m-%y",
        "%Y-%m-%d %H:%M:%S",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue

    return None


def update_if_value(obj, field, value):
    """
    Update only when uploaded value is available.
    Prevents existing DB data from becoming None / blank.
    """
    if value not in ["", None, "-"]:
        setattr(obj, field, value)


def upload_student_details(request):
    TEMPLATE = "student_management/student/upload_student_details.html"

    if request.method != "POST":
        return render(request, TEMPLATE)

    excel_file = request.FILES.get("excel_file")

    if not excel_file:
        return render(request, TEMPLATE, {
            "error": "No Excel file uploaded."
        })

    wb = load_workbook(excel_file, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return render(request, TEMPLATE, {
            "error": "Excel file is empty."
        })

    headers = [str(h).strip() for h in rows[0] if h]

    required = {"reg_no", "aadhar_number"}

    if not required.issubset(headers):
        return render(request, TEMPLATE, {
            "error": "Excel must contain reg_no and aadhar_number columns."
        })

    result_wb = Workbook()
    result_ws = result_wb.active

    result_headers = headers + [
        "student_status",
        "user_status",
        "error_reason",
    ]
    result_ws.append(result_headers)

    for excel_row in rows[1:]:
        row = dict(zip(headers, excel_row))

        student_status = ""
        user_status = ""
        error_reason = ""

        try:
            reg_no = clean_number(row.get("reg_no"))
            aadhar_no = clean_aadhar(row.get("aadhar_number"))

            if not reg_no or not aadhar_no:
                raise Exception("valid reg_no and 12-digit aadhar_number are required")

            name = clean(row.get("name"))
            dept_name = clean(row.get("department"))
            batch = clean(row.get("batch"))
            dob = parse_excel_date(row.get("DOB") or row.get("date_of_birth"))

            email = clean(row.get("email")) or f"{reg_no}@ritrjpm.ac.in"
            mobile_no = clean_number(row.get("mobile_no"))
            gender = clean(row.get("gender"))
            year = clean(row.get("year")) or "1"
            semester = clean(row.get("semester")) or "1"

            student_dept = None
            user_dept = None

            if dept_name:
                student_dept = Add_Department.objects.filter(
                    Department__iexact=dept_name
                ).first()

                user_dept = Department.objects.using(
                    "rit_approval_system"
                ).filter(
                    Department__iexact=dept_name
                ).first()

            student = StudentDetails.objects.filter(
                Q(reg_no=reg_no) | Q(aadhar_number=aadhar_no)
            ).first()

            student_data = {
                "aadhar_number": aadhar_no,
                "reg_no": reg_no,
                "name": name,
                "department": student_dept,
                "email": email,
                "batch": batch,
                "year": year,
                "semester": semester,
                "date_of_birth": dob,
                "mobile_no": mobile_no,
                "gender": gender,
            }

            if student:
                for field, value in student_data.items():
                    update_if_value(student, field, value)

                student.save()
                student_status = "Updated"
            else:
                StudentDetails.objects.create(**student_data)
                student_status = "Created"

            role = Role.objects.using("rit_approval_system").filter(
                role="Student"
            ).first()

            if not user_dept:
                raise Exception(f"USER Department not found: {dept_name}")

            if not role:
                raise Exception("Student role not found in rit_approval_system")

            user = USER.objects.using("rit_approval_system").filter(
                Q(Employee_id=reg_no) | Q(unique_id=aadhar_no)
            ).first()

            if user:
                update_if_value(user, "username", name)
                update_if_value(user, "Employee_id", reg_no)
                update_if_value(user, "unique_id", aadhar_no)
                update_if_value(user, "email", email)
                update_if_value(user, "Department", user_dept)
                update_if_value(user, "role", role)

                user.is_student = True
                user.save(using="rit_approval_system")
                user_status = "Updated"
            else:
                USER.objects.using("rit_approval_system").create(
                    username=name,
                    Employee_id=reg_no,
                    unique_id=aadhar_no,
                    email=email,
                    Department=user_dept,
                    role=role,
                    is_student=True,
                    password=make_password("123"),
                    date_joined=timezone.now(),
                    last_login=timezone.now(),
                )
                user_status = "Created"

        except Exception as e:
            error_reason = f"{type(e).__name__}: {e}"

        result_ws.append([
            row.get(h, "") for h in headers
        ] + [
            student_status,
            user_status,
            error_reason,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="student_upload_result.xlsx"'
    )

    result_wb.save(response)
    return response

import pandas as pd
from django.shortcuts import render
from django.contrib import messages

def sync_user_student_aadhar_to_reg_numbers(request):

    updated = 0
    not_found = 0
    invalid = 0

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "Please upload an Excel file.")
            return render(request, "student_management/admin/sync_user_student_aadhar_to_reg_numbers.html")

        try:
            df = pd.read_excel(excel_file, dtype=str)
            df = df.fillna("")

            for _, row in df.iterrows():

                reg_no = row.get("reg_no", "").strip()
                aadhar = row.get("aadhar_number", "").strip()

                # remove decimal if excel converted it
                aadhar = aadhar.split(".")[0]

                if not reg_no or not aadhar:
                    invalid += 1
                    continue

                student = StudentDetails.objects.filter(
                    aadhar_number=aadhar
                ).first()

                if not student:
                    not_found += 1
                    continue

                user = USER.objects.using("rit_approval_system").filter(
                    unique_id=aadhar
                ).first()

                if user:
                    user.Employee_id = reg_no
                    user.save(update_fields=["Employee_id"])
                    updated += 1
                else:
                    not_found += 1

            messages.success(
                request,
                f"Sync Completed — Updated: {updated}, Not Found: {not_found}, Invalid Rows: {invalid}"
            )

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")

    return render(
        request,
        "student_management/admin/sync_user_student_aadhar_to_reg_numbers.html",
        {
            "updated": updated,
            "not_found": not_found,
            "invalid": invalid
        }
    )




def sync_user_to_admission_student_details(request):

    users = USER.objects.using("rit_approval_system").exclude(
        unique_id__isnull=True
    ).exclude(
        unique_id=""
    )

    updated = 0
    not_found = 0

    for user in users:

        aadhar = user.unique_id
        reg_no = user.Employee_id

        student_data = StudentDetails.objects.filter(
            aadhar_number=aadhar
        ).first()

        personal = PersonalDetails.objects.using("admissionform1").filter(
            Aadhaar_Number=aadhar
        ).first()

        # Update only if both records exist
        if student_data and personal:

            student_data.reg_no = reg_no
            personal.registration_no = reg_no

            student_data.save(update_fields=["reg_no"])
            personal.save(update_fields=["registration_no"])

            updated += 1

        else:
            not_found += 1

    messages.success(
        request,
        f"Sync completed. Updated: {updated}, Not Found: {not_found}"
    )

    return redirect("student_details")



def sync_student_to_user_details(request):

    students = StudentDetails.objects.exclude(
        aadhar_number__isnull=True
    ).exclude(
        aadhar_number=""
    )

    updated = 0
    not_found = 0

    for student in students:

        aadhar = student.aadhar_number
        reg_no = student.reg_no

        student_data = StudentDetails.objects.filter(
            aadhar_number=aadhar
        ).first()

        user = USER.objects.using("rit_approval_system").filter(
            unique_id=aadhar
        ).first()

        # Update only if both records exist
        if student_data and user:

            # student_data.reg_no = reg_no
            user.Employee_id = reg_no

            # student_data.save(update_fields=["reg_no"])
            user.save(update_fields=["Employee_id"])

            updated += 1

        else:
            not_found += 1

    messages.success(
        request,
        f"Sync completed. Updated: {updated}, Not Found: {not_found}"
    )

    return redirect("student_details")


def _fetch_admission_mode(student):
    """
    Fetch the admission Mode for a student from the external college
    admission DB (`admissionform1`).

    Match order:
      1) StudentDetails.reg_no  -> PersonalDetails.registration_no
      2) StudentDetails.aadhar  -> PersonalDetails.Aadhaar_Number
    Then PersonalDetails -> AdmissionRecords.Mode
    Returns the raw Mode string (e.g. "HOSTEL" / "TRANSPORT" / "DAY SCHOLAR"),
    or None if no admission record is found.
    """
    personal = None

    reg_no = str(getattr(student, "reg_no", "") or "").strip()
    if reg_no:
        personal = (
            PersonalDetails.objects.using("admissionform1")
            .filter(registration_no=reg_no)
            .first()
        )

    aadhar = str(getattr(student, "aadhar_number", "") or "").strip()
    if not personal and aadhar:
        personal = (
            PersonalDetails.objects.using("admissionform1")
            .filter(Aadhaar_Number=aadhar)
            .first()
        )

    if not personal:
        return None

    admission = (
        AdmissionRecords.objects.using("admissionform1")
        .filter(PersonalDetailsId=personal)
        .first()
    )

    if not admission:
        return None

    return (admission.Mode or "").strip() or None


def sync_student_mode_from_admission(request):
    """
    Pull the admission Mode (HOSTEL / TRANSPORT / DAY SCHOLAR ...) from the
    external college admission DB (`admissionform1`) and store it on
    StudentDetails.mode so the value is available locally (and used by the
    student Leave/OD approval flow without hitting the external DB).
    """
    students = StudentDetails.objects.all()

    updated = 0
    unchanged = 0
    not_found = 0

    for student in students:
        mode = _fetch_admission_mode(student)

        if mode is None:
            not_found += 1
            continue

        if (student.mode or "").strip() == mode:
            unchanged += 1
            continue

        student.mode = mode
        student.save(update_fields=["mode"])
        updated += 1

    messages.success(
        request,
        f"Mode sync completed. Updated: {updated}, "
        f"Already up-to-date: {unchanged}, Not Found: {not_found}"
    )

    return redirect("student_details")



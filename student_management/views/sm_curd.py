import math
import re
from decimal import Decimal, ROUND_HALF_UP
from course_management.models import *
from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from course_management.decorators import course_management
from django.core.exceptions import ObjectDoesNotExist
from django.utils import timezone
from django.utils.dateparse import parse_datetime

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect
import user_accounts
from user_accounts.models import USER, Role, Department, AdmissionRecords, PersonalDetails
from user_accounts.decorators import check_permission, no_cache
from student_management.decorators import student_management


from django.shortcuts import render
from student_management.models import *
import json
from faculty_management.models import *
from user_accounts.views.dashboards import examination_management


def _parse_leave_datetime(value):
    parsed = parse_datetime(str(value or "").strip())
    if not parsed:
        return None
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


def _calculate_leave_days(from_datetime, to_datetime):
    delta = to_datetime - from_datetime
    seconds = Decimal(str(delta.total_seconds()))
    days = (seconds / Decimal("86400")).quantize(
        Decimal("0.01"),
        rounding=ROUND_HALF_UP,
    )
    return days if days > 0 else Decimal("0.01")


# @student_management
# @check_permission("dashboard")
# def dashboard(request):
#     register_no = request.user.Employee_id  # Student Register Number

#     if not register_no:
#         return render(request, 'student_management/student/student_overview.html', {
#             'error': 'Register number not found.'
#         })

#     # ---------- Personal Details (Optional) ----------
#     cgpa_value = "N/A"
#     gpa_value = "N/A"
#     current_sem = None

#     try:
#         student_personal = PersonalDetails.objects.using('admissionform1').get(registration_no=register_no)
#     except PersonalDetails.DoesNotExist:
#         student_personal = None

#     # ---------- Co/Extra-Curricular Activities ----------
#     co_ex_curricular_all = StudentCO_EX_Curricular.objects.filter(student__reg_no=register_no)

#     co_ex_curricular_approved_count = co_ex_curricular_all.filter(status='APPROVED').count()
#     co_ex_curricular_pending_count = co_ex_curricular_all.filter(status='PENDING').count()
#     co_ex_curricular_rejected_count = co_ex_curricular_all.filter(status='REJECTED').count()

#     # ---------- Student Projects ----------
#     student_projects_all = StudentName.objects.filter(register_no=register_no)

#     projects_pending_count = student_projects_all.filter(project__status='pending').count()
#     projects_ongoing_count = student_projects_all.filter(project__status='on going').count()
#     projects_completed_count = student_projects_all.filter(project__status='completed').count()
#     projects_count = student_projects_all.count()

#     # ---------- Achievements ----------
#     achievements_all = StudentAchievements.objects.filter(register_no=register_no)

#     achievements_pending_count = achievements_all.filter(status='PENDING').count()
#     achievements_approved_count = achievements_all.filter(status='APPROVED').count()
#     achievements_rejected_count = achievements_all.filter(status='REJECTED').count()

#     # ---------- Publications ----------
#     publications_all = StudentPublication.objects.filter(register_no=register_no)

#     publications_approved_count = publications_all.filter(status='APPROVED').count()
#     publications_pending_count = publications_all.filter(status='PENDING').count()
#     publications_presented_count = publications_all.filter(status='APPROVED', presented='Presented').count()
#     publications_not_presented_count = publications_all.filter(status='APPROVED', presented='Not Presented').count()

#     # ---------- Professional Membership ----------
#     latest_professional = StudentProfessionl.objects.filter(
#         register_no=register_no,
#         status='APPROVED'
#     ).order_by('-id').first()

#     # ---------- Chart Data (for Dashboard Graphs) ----------
#     chart_data = {
#         'co_ex_curricular': co_ex_curricular_approved_count,
#         'projects': projects_completed_count,
#         'achievements': achievements_approved_count,
#         'publications': publications_approved_count,
#     }

#     # ---------- Context to Template ----------
#     context = {
#         'cgpa_value': cgpa_value,
#         'gpa_value': gpa_value,
#         'current_sem': current_sem,

#         'co_ex_curricular_count': co_ex_curricular_all.count(),
#         'co_ex_curricular_approved_count': co_ex_curricular_approved_count,
#         'co_ex_curricular_pending_count': co_ex_curricular_pending_count,
#         'co_ex_curricular_rejected_count': co_ex_curricular_rejected_count,

#         'projects_total_count': projects_count,
#         'projects_pending_count': projects_pending_count,
#         'projects_ongoing_count': projects_ongoing_count,
#         'projects_completed_count': projects_completed_count,

#         'achievements_approved_count': achievements_approved_count,
#         'achievements_pending_count': achievements_pending_count,
#         'achievements_rejected_count': achievements_rejected_count,

#         'publications_approved_count': publications_approved_count,
#         'publications_pending_count': publications_pending_count,
#         'publications_presented_count': publications_presented_count,
#         'publications_not_presented_count': publications_not_presented_count,

#         'latest_professional': latest_professional,
#         'chart_data': chart_data,
#     }


#     return render(request, 'student_management/student/student_overview.html', context)

from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from user_accounts.decorators import check_permission  # keep your existing import

from user_accounts.models import StudentDetails
from student_management.models import (
    StudentCO_EX_Curricular,
    StudentAchievements,
    StudentPublication,
    StudentProfessionl,
    StudentProjects,
)



from django.shortcuts import render
from django.db.models import Q

from user_accounts.models import StudentDetails
from student_management.models import (
    StudentCO_EX_Curricular,
    StudentProjects,
    StudentAchievements,
    StudentPublication,
    StudentProfessionl,
)

@check_permission("dashboard")
def dashboard(request):
    """
    Single view:
    - Builds all counts for Student Overview cards
    - Sends chart_payload (JSON) for Chart.js
    - Sends latest_professional record
    """

    TEMPLATE = "student_management/student/student_overview.html"

    # -----------------------------
    # Get StudentDetails mapped to request.user
    # Your mapping: request.user.Employee_id -> reg_no
    # -----------------------------
    reg_no = getattr(request.user, "Employee_id", None)
    student_obj = StudentDetails.objects.filter(reg_no=reg_no).first()

    # if student not found -> render with all zeros
    if not student_obj:
        context = {
            "coex_total": 0, "coex_pending": 0, "coex_approved": 0, "coex_rejected": 0,
            "projects_total": 0, "projects_pending": 0, "projects_ongoing": 0, "projects_completed": 0,
            "projects_approval_pending": 0, "projects_approval_approved": 0, "projects_approval_rejected": 0,
            "achievements_total": 0, "achievements_pending": 0, "achievements_approved": 0, "achievements_rejected": 0,
            "publications_total": 0, "publications_pending": 0, "publications_approved": 0, "publications_rejected": 0,
            "publications_presented": 0, "publications_not_presented": 0,
            "latest_professional": None,
            "chart_payload": {
                "coex": {"pending": 0, "approved": 0, "rejected": 0},
                "projects": {"pending": 0, "ongoing": 0, "completed": 0},
                "achievements": {"pending": 0, "approved": 0, "rejected": 0},
                "publications": {"pending": 0, "approved": 0, "rejected": 0},
            },
        }
        return render(request, TEMPLATE, context)

    # helper for case-insensitive counting
    def c(qs, field, value):
        return qs.filter(**{f"{field}__iexact": value}).count()

    # -----------------------------
    # CO/EX curricular
    # status: pending/approved/rejected (lowercase)
    # -----------------------------
    coex_qs = StudentCO_EX_Curricular.objects.filter(student=student_obj)
    coex_pending = c(coex_qs, "status", "pending")
    coex_approved = c(coex_qs, "status", "approved")
    coex_rejected = c(coex_qs, "status", "rejected")
    coex_total = coex_qs.count()

    # -----------------------------
    # Projects:
    # status: pending / on going / completed (lowercase, "on going" has a space)
    # approval_status: Pending / Approved / Rejected (capitalized)
    # -----------------------------
    projects_qs = StudentProjects.objects.filter(student=student_obj)

    projects_pending = c(projects_qs, "status", "pending")
    projects_ongoing = c(projects_qs, "status", "on going")
    projects_completed = c(projects_qs, "status", "completed")
    projects_total = projects_qs.count()

    projects_approval_pending = c(projects_qs, "approval_status", "Pending")
    projects_approval_approved = c(projects_qs, "approval_status", "Approved")
    projects_approval_rejected = c(projects_qs, "approval_status", "Rejected")

    # -----------------------------
    # Achievements
    # status: Pending/Approved/Rejected (capitalized)
    # -----------------------------
    achievements_qs = StudentAchievements.objects.filter(student=student_obj)
    achievements_pending = c(achievements_qs, "status", "Pending")
    achievements_approved = c(achievements_qs, "status", "Approved")
    achievements_rejected = c(achievements_qs, "status", "Rejected")
    achievements_total = achievements_qs.count()

    # -----------------------------
    # Publications
    # status: Pending/Approved/Rejected
    # presented: Presented / Not Presented
    # -----------------------------
    publications_qs = StudentPublication.objects.filter(student=student_obj)
    publications_pending = c(publications_qs, "status", "Pending")
    publications_approved = c(publications_qs, "status", "Approved")
    publications_rejected = c(publications_qs, "status", "Rejected")
    publications_total = publications_qs.count()

    publications_presented = c(publications_qs, "presented", "Presented")
    publications_not_presented = c(publications_qs, "presented", "Not Presented")

    # -----------------------------
    # Professional membership (latest record)
    # status: Pending/Approved/Rejected (capitalized)
    # -----------------------------
    latest_professional = (
        StudentProfessionl.objects
        .filter(student=student_obj)
        .order_by("-created_at")
        .first()
    )

    # -----------------------------
    # Chart payload
    # -----------------------------
    chart_payload = {
        "coex": {"pending": coex_pending, "approved": coex_approved, "rejected": coex_rejected},
        "projects": {"pending": projects_pending, "ongoing": projects_ongoing, "completed": projects_completed},
        "achievements": {"pending": achievements_pending, "approved": achievements_approved, "rejected": achievements_rejected},
        "publications": {"pending": publications_pending, "approved": publications_approved, "rejected": publications_rejected},
    }

    context = {
        # COEX
        "coex_total": coex_total,
        "coex_pending": coex_pending,
        "coex_approved": coex_approved,
        "coex_rejected": coex_rejected,

        # Projects
        "projects_total": projects_total,
        "projects_pending": projects_pending,
        "projects_ongoing": projects_ongoing,
        "projects_completed": projects_completed,

        # Projects Approval
        "projects_approval_pending": projects_approval_pending,
        "projects_approval_approved": projects_approval_approved,
        "projects_approval_rejected": projects_approval_rejected,

        # Achievements
        "achievements_total": achievements_total,
        "achievements_pending": achievements_pending,
        "achievements_approved": achievements_approved,
        "achievements_rejected": achievements_rejected,

        # Publications
        "publications_total": publications_total,
        "publications_pending": publications_pending,
        "publications_approved": publications_approved,
        "publications_rejected": publications_rejected,
        "publications_presented": publications_presented,
        "publications_not_presented": publications_not_presented,

        # Professional
        "latest_professional": latest_professional,

        # Charts
        "chart_payload": chart_payload,
    }

    return render(request, TEMPLATE, context)

 

# @student_management
# @check_permission("apply_leave_od")
# def student_apply_leave_od(request):
#     """
#     View for students to apply Leave / OD.
#     Department is mapped from StudentDetails -> Add_Department -> Department using iexact.
#     """
#     # ✅ Get logged-in student's register number
#     register_no = getattr(request.user, "Employee_id", None)
#     if not register_no:
#         messages.error(request, "No register number found for the student.")
#         return redirect("home")

#     # ✅ Fetch student details
#     try:
#         student_details = StudentDetails.objects.select_related("department").get(
#             reg_no=register_no
#         )
#     except StudentDetails.DoesNotExist:
#         messages.error(request, "Your student record was not found.")
#         return redirect("home")

#     # ✅ Map Add_Department to Department table using iexact
#     # department_obj = Add_Department.objects.filter(
#     #     Department__iexact=student_details.department.Department
#     # ).first()

#     # if not department_obj:
#     #     messages.error(request, "Department mapping not found in Department table.")
#     #     return redirect("home")

#     if request.method == "POST":
#         # Form data
#         application_type = request.POST.get("application_type")
#         from_date_str = request.POST.get("from_date")
#         to_date_str = request.POST.get("to_date")
#         total_days = request.POST.get("total_days")
#         reason = request.POST.get("reason")
#         proof_file = request.FILES.get("proof_file")

#         # Validate proof if OD
#         if application_type == "OD" and not proof_file:
#             messages.error(request, "Proof document is required for OD applications.")
#             return redirect("apply_leave_od")

#         # Convert dates
#         try:
#             from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
#             to_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()
#         except (ValueError, TypeError):
#             messages.error(request, "Invalid dates provided.")
#             return redirect("apply_leave_od")

#         if from_date > to_date:
#             messages.error(request, "'From Date' cannot be after 'To Date'.")
#             return redirect("apply_leave_od")

#         # Validate total_days
#         try:
#             total_days = int(total_days)
#             if total_days <= 0:
#                 raise ValueError
#         except (ValueError, TypeError):
#             messages.error(request, "Total days must be a positive number.")
#             return redirect("apply_leave_od")

#         # ✅ Save Leave/OD application
#         StudentLeaveOdApplication.objects.create(
#             student=student_details,  # USER instance
#             mentor=None,
#             ca=None,
#             application_type=application_type,
#             from_date=from_date,
#             to_date=to_date,
#             total_days=total_days,
#             reason=reason,
#             proof_file=proof_file,
#             status=StudentLeaveOdApplication.Status.PENDING,
#             department=student_details.department,  # Department instance from iexact mapping
#             study_year=getattr(student_details, "study_year", None),
#         )

#         messages.success(
#             request, f"Your {application_type} request has been submitted successfully."
#         )
#         return redirect("apply_leave_od")

#     # Fetch existing applications
#     leave_od_requests = StudentLeaveOdApplication.objects.filter(
#         student=student_details
#     ).order_by("-created_at")

#     context = {
#         "student": request.user,
#         "user_role": getattr(request.user, "role", None),
#         "leave_od_requests": leave_od_requests,
#         "student": student_details,
#     }

#     return render(
#         request,
#         "student_management/leave_od_templates/student_leave_od_form.html",
#         context,
#     )

@student_management
@check_permission("apply_leave_od")
def student_apply_leave_od(request):


    register_no = getattr(request.user, "Employee_id", None)

   
    if not register_no:
        messages.error(request, "No register number found.")
        return redirect("home")

    try:
        student_details = (
            StudentDetails.objects
            .select_related("department", "mentor", "ca")
            .get(reg_no=register_no)
        )

       

    except StudentDetails.DoesNotExist:
        messages.error(request, "Student record not found.")
        return redirect("home")

    if request.method == "POST":

        

        application_type = request.POST.get("application_type")
        from_date_str = request.POST.get("from_date")
        to_date_str = request.POST.get("to_date")
        reason = request.POST.get("reason")
        proof_file = request.FILES.get("proof_file")

        if application_type == "OD" and not proof_file:
            messages.error(request, "Proof required for OD.")
            return redirect("apply_leave_od")

        from_date = _parse_leave_datetime(from_date_str)
        to_date = _parse_leave_datetime(to_date_str)

        if not from_date or not to_date:
            messages.error(request, "Invalid date and time.")
            return redirect("apply_leave_od")

        if from_date >= to_date:
            messages.error(request, "To date and time must be after from date and time.")
            return redirect("apply_leave_od")

        total_days = _calculate_leave_days(from_date, to_date)

        overlap_qs = (
            StudentLeaveOdApplication.objects
            .filter(student=student_details)
            .filter(
                Q(from_date__lte=to_date)
                &
                Q(to_date__gte=from_date)
            )
        )

      

        if overlap_qs.exists():
            messages.error(request, "Already applied for this date.")
            return redirect("apply_leave_od")

        try:
            with transaction.atomic():

              

                new_application = StudentLeaveOdApplication.objects.create(
                    student=student_details,
                    mentor=None,
                    ca=None,
                    application_type=application_type,
                    from_date=from_date,
                    to_date=to_date,
                    total_days=total_days,
                    reason=reason,
                    proof_file=proof_file,
                    status=StudentLeaveOdApplication.Status.PENDING,
                    department=student_details.department,
                    study_year=student_details.year,
                )
              

                created_rows = _create_approver_chain_for_student_leave(
                    leave_application=new_application,
                    creator_student=student_details
                )

             
                if created_rows == 0:
                    raise ValueError("No approval rows were created for this application.")

                messages.success(
                    request,
                    f"{application_type} applied successfully."
                )

                return redirect("apply_leave_od")

        except Exception as e:
            import traceback
            traceback.print_exc()

      
            messages.error(request, f"Error: {str(e)}")
            return redirect("apply_leave_od")

    leave_od_requests = (
        StudentLeaveOdApplication.objects
        .filter(student=student_details)
        .order_by("-created_at")
    )

    request_ids = [request_obj.id for request_obj in leave_od_requests]
    approval_rows = list(
        Student_LeaveApproversData.objects
        .filter(leave_application_id__in=request_ids)
        .order_by("approver_level", "id")
    )
    approver_user_ids = {
        row.approver_id_id
        for row in approval_rows
        if row.approver_id_id
    }
    approver_user_map = {
        user.id: user
        for user in USER.objects.using("rit_approval_system").filter(id__in=approver_user_ids)
    }
    approval_status_map = {}
    for row in approval_rows:
        approver_user = approver_user_map.get(row.approver_id_id)
        approval_status_map.setdefault(row.leave_application_id, []).append({
            "level": row.approver_level,
            "status": row.status,
            "name": getattr(approver_user, "username", "") or "-",
            "employee_id": getattr(approver_user, "Employee_id", "") or "-",
        })

    for request_obj in leave_od_requests:
        request_obj.approval_status_rows = approval_status_map.get(request_obj.id, [])

    applicant_mode, applicant_gender = _get_student_leave_condition(student_details)

    context = {
        "student": student_details,
        "leave_od_requests": leave_od_requests,
        "student_leave_flow_label": _student_leave_flow_label(applicant_mode, applicant_gender),
    }

    return render(
        request,
        "student_management/leave_od_templates/student_leave_od_form.html",
        context,
    )


def _normalize_student_leave_mode(value):
    value = str(value or "").strip().upper()
    if "HOSTEL" in value:
        return "HOSTEL"
    if "TRANSPORT" in value:
        return "TRANSPORT"
    return "DEFAULT"


def _normalize_student_leave_gender(value):
    value = str(value or "").strip().upper()
    if value in {"M", "MALE", "BOY", "BOYS"}:
        return "MALE"
    if value in {"F", "FEMALE", "GIRL", "GIRLS"}:
        return "FEMALE"
    return "ANY"


def _get_student_leave_condition(creator_student):
    # Prefer the Mode already synced onto StudentDetails (via the
    # "Sync Mode from Admission" button on the student details page).
    # Only fall back to a live lookup in the college admission DB
    # (admissionform1) when the local value is missing.
    local_mode = str(getattr(creator_student, "mode", "") or "").strip()

    if local_mode:
        mode = _normalize_student_leave_mode(local_mode)
    else:
        register_no = str(getattr(creator_student, "reg_no", "") or "").strip()
        admission = None

        if register_no:
            personal = (
                PersonalDetails.objects.using("admissionform1")
                .filter(registration_no=register_no)
                .first()
            )
            if personal:
                admission = (
                    AdmissionRecords.objects.using("admissionform1")
                    .filter(PersonalDetailsId=personal)
                    .first()
                )

        if not admission and getattr(creator_student, "aadhar_number", None):
            personal = (
                PersonalDetails.objects.using("admissionform1")
                .filter(Aadhaar_Number=creator_student.aadhar_number)
                .first()
            )
            if personal:
                admission = (
                    AdmissionRecords.objects.using("admissionform1")
                    .filter(PersonalDetailsId=personal)
                    .first()
                )

        mode = _normalize_student_leave_mode(getattr(admission, "Mode", None))

    gender = _normalize_student_leave_gender(getattr(creator_student, "gender", None))

    if mode != "HOSTEL":
        gender = "ANY"

    return mode, gender


def _student_leave_flow_label(mode, gender):
    if mode == "TRANSPORT":
        return "Transport Students: Faculty -> HOD"
    if mode == "HOSTEL" and gender == "MALE":
        return "Hostel Boys: Faculty -> HOD -> Hostel Warden(BOYS)"
    if mode == "HOSTEL" and gender == "FEMALE":
        return "Hostel Girls: Faculty -> HOD -> Hostel Warden(GIRLS)"
    return "Student Leave: Faculty -> HOD"


def _create_approver_chain_for_student_leave(
    leave_application,
    creator_student,
):

   

    created_rows = 0
    used_employee_ids = set()

    student_user = (
        USER.objects
        .using("rit_approval_system")
        .filter(
            Employee_id=str(creator_student.reg_no),
            is_active=True
        )
        .first()
    )


    if not student_user:
        
        return 0

    creator_role_id = student_user.role_id


    applicant_mode, applicant_gender = _get_student_leave_condition(creator_student)


    matching_conditions = [
        (applicant_mode, applicant_gender),
    ]

    approvers_qs = Student_LeaveApprovers.objects.none()

    for mode, gender in matching_conditions:
        candidate_qs = (
            Student_LeaveApprovers.objects
            .filter(
                creator_role_id=creator_role_id,
                applicant_mode=mode,
                applicant_gender=gender,
            )
            .order_by("approver_level")
        )
        if candidate_qs.exists():
            approvers_qs = candidate_qs
            print("Using Approver Condition:", mode, gender)
            break

    if not approvers_qs.exists():
        raise ValueError(
            "Student leave approval hierarchy is not configured for this student mode/gender."
        )

    if not approvers_qs.exists():
       
        return 0

    student_department = creator_student.department

    student_department_name = getattr(
        student_department,
        "Department",
        None
    )


    student_ext_dept_id = (
        Department.objects
        .using("rit_approval_system")
        .filter(Department__iexact=student_department_name)
        .values_list("id", flat=True)
        .first()
    )


    mentor_user = None
    mentor_faculty = getattr(creator_student, "mentor", None)

    if mentor_faculty:
        mentor_user = (
            USER.objects
            .using("rit_approval_system")
            .filter(
                Employee_id=str(mentor_faculty.faculty_id),
                is_active=True
            )
            .first()
        )


    ca_user = None
    ca_faculty = getattr(creator_student, "ca", None)

    if ca_faculty:
        ca_user = (
            USER.objects
            .using("rit_approval_system")
            .filter(
                Employee_id=str(ca_faculty.faculty_id),
                is_active=True
            )
            .first()
        )

    for approver in approvers_qs:


        level = approver.approver_level
        role_id = approver.approver_role_id
        local_dept_id = approver.approver_department_id

        is_cross = (
            (approver.is_cross_department_approver or "NO").upper() == "YES"
        )


        approver_users = []

        mentor_role_match = (
            mentor_user
            and mentor_user.role_id == role_id
        )

        ca_role_match = (
            ca_user
            and ca_user.role_id == role_id
        )

        same_faculty_user = (
            mentor_user
            and ca_user
            and str(mentor_user.Employee_id) == str(ca_user.Employee_id)
        )

        if mentor_role_match:

            mentor_employee_id = str(mentor_user.Employee_id)

            if mentor_employee_id not in used_employee_ids:

                

                approver_users.append(mentor_user)
                used_employee_ids.add(mentor_employee_id)

        if ca_role_match and not same_faculty_user:

            ca_employee_id = str(ca_user.Employee_id)

            if ca_employee_id not in used_employee_ids:

                

                approver_users.append(ca_user)
                used_employee_ids.add(ca_employee_id)

        approver_filter = {
            "role_id": role_id,
            "is_active": True,
        }

        if is_cross and local_dept_id:


            local_dept = (
                Add_Department.objects
                .filter(id=local_dept_id)
                .first()
            )

            local_department_name = getattr(
                local_dept,
                "Department",
                None
            )

            ext_dept_id = (
                Department.objects
                .using("rit_approval_system")
                .filter(Department__iexact=local_department_name)
                .values_list("id", flat=True)
                .first()
            )

            if ext_dept_id:
                approver_filter["Department_id"] = ext_dept_id

        else:

            if student_ext_dept_id:
                approver_filter["Department_id"] = student_ext_dept_id


        # IMPORTANT FIX:
        # Dynamic user should be added only if Mentor/CA was not added.
        if not approver_users:

            dynamic_user = (
                USER.objects
                .using("rit_approval_system")
                .filter(**approver_filter)
                .exclude(Employee_id__in=used_employee_ids)
                .first()
            )

            if dynamic_user:

               

                approver_users.append(dynamic_user)
                used_employee_ids.add(str(dynamic_user.Employee_id))

        if not approver_users:
            raise ValueError(
                f"No active approver found for level {level}. Please check the student leave hierarchy."
            )

        for user_obj in approver_users:

            try:
                created_obj = Student_LeaveApproversData.objects.create(
                    leave_application=leave_application,
                    creator_id=creator_student,
                    approver_level=level,
                    status=Student_LeaveApproversData.Status.PENDING,
                    reason=f"Approver: {user_obj.Employee_id}",
                )

                Student_LeaveApproversData.objects.filter(
                    id=created_obj.id
                ).update(
                    approver_id_id=user_obj.id
                )

               

                created_rows += 1

            except Exception as create_error:
                import traceback

                
                print(str(create_error))

                traceback.print_exc()


    return created_rows

@check_permission("profile")
@student_management
def profile(request):
    return render(request, "student_management/student/student_profile.html")


from django.shortcuts import render, redirect, get_object_or_404
from student_management.models import StudentAchievements
from user_accounts.models import Add_Department ,StudentDetails

@student_management
def student_achievements_view(request):

    # 🔹 Student object
    student_id = request.user.Employee_id
    student = StudentDetails.objects.filter(reg_no=student_id).first()

    # 🔹 Department
    department_obj = None
    department = ""
    if student and student.department:
        department_obj = student.department
        department = department_obj.Department

    # 🔹 Semester / Batch / Section
    sem_obj = (
        StudentDetails.objects
        .filter(reg_no=student.reg_no)
        .order_by("-last_updated")
        .first()
        if student else None
    )

    semester = sem_obj.semester if sem_obj else "-"
    batch = sem_obj.batch if sem_obj else "-"
    section = sem_obj.section if sem_obj else "-"

    # 🔹 POST (Edit Achievement)
    if request.method == "POST":
        ach_id = request.POST.get("achievement_id")

        achievement = get_object_or_404(
            StudentAchievements,
            id=ach_id,
            student=student
        )

        achievement.award_name = request.POST.get("award_name", "").strip()
        achievement.contest = request.POST.get("contest", "").strip()
        achievement.given_by = request.POST.get("given_by", "").strip()

        if request.POST.get("date"):
            achievement.date = request.POST.get("date")

        if request.FILES.get("certificate"):
            achievement.certificate = request.FILES["certificate"]

        achievement.save()
        return redirect("student_achievements_view")

    # 🔹 Fetch Achievements
    achievements = StudentAchievements.objects.filter(
        student=student,
        event_type="achievements"
    ).order_by("-created_at")

    context = {
        "student": student,
        "department_obj": department_obj,
        "department": department,
        "semester": semester,
        "batch": batch,
        "section": section,
        "achievements": achievements,
    }

    return render(
        request,
        "student_management/student/student_achivement_view.html",
        context
    )
  
 

from django.shortcuts import render, redirect, get_object_or_404
from student_management.models import StudentCO_EX_Curricular
from user_accounts.models import StudentDetails


from user_accounts.models import Add_Department, StudentDetails

@student_management
def student_co_ex_curricular_view(request):
    # 🔹 Student object
    student_id = request.user.Employee_id
    student = StudentDetails.objects.filter(reg_no=student_id).first()

    if not student:
        return redirect("profile")

    # 🔹 Department
    department_obj = student.department if student.department else None
    department = ""
    if department_obj:
        department = getattr(department_obj, "Department", str(department_obj))

    # ✅ Semester / Batch / Section FROM StudentDetails
    semester = student.semester if student.semester else "-"
    batch = student.batch if student.batch else "-"
    section = student.section if student.section else "-"

    # 🔹 Semester / Batch / Section
    sem_obj = (
        StudentDetails.objects
        .filter(reg_no=student.reg_no)
        .order_by("-last_updated")
        .first()
    )

    semester = sem_obj.semester if sem_obj else None
    batch = sem_obj.batch if sem_obj else None
    section = sem_obj.section if sem_obj else None


    # 🔹 POST (Edit Entry)
    if request.method == "POST":
        entry_id = request.POST.get("entry_id")

        entry = get_object_or_404(
            StudentCO_EX_Curricular,
            id=entry_id,
            student=student
        )

        entry.batch = request.POST.get("batch", "").strip() or entry.batch

        activity_type = request.POST.get("activity_type")
        if activity_type:
            entry.activity_type = activity_type

        entry.event_name = request.POST.get("event_name", "").strip()

        level = request.POST.get("level")
        if level:
            entry.level = level

        from_date = request.POST.get("from_date")
        to_date = request.POST.get("to_date")
        entry.from_date = from_date if from_date else None
        entry.to_date = to_date if to_date else None

        total_days_val = request.POST.get("total_days")
        entry.total_days = int(total_days_val) if total_days_val else None

        if request.FILES.get("certificate"):
            entry.certificate = request.FILES["certificate"]

        # ✅ Auto sync academic info (same logic)
        entry.department = department_obj
        entry.semester = None if semester == "-" else semester
        entry.section = None if section == "-" else section

        # ✅ If you want entry.batch always equal to StudentDetails batch, force it:
        # entry.batch = None if batch == "-" else batch

        entry.save()
        return redirect("student_co_ex_curricular_view")

    # 🔹 Fetch entries
    co_curriculars = StudentCO_EX_Curricular.objects.filter(
        student=student
    ).order_by("-created_at")

    # 🔹 Detect file type
    for entry in co_curriculars:
        if entry.certificate:
            url = entry.certificate.url.lower()
            if url.endswith(".pdf"):
                entry.file_type = "pdf"
            elif url.endswith((".jpg", ".jpeg", ".png", ".gif")):
                entry.file_type = "image"
            else:
                entry.file_type = "other"
        else:
            entry.file_type = None

    context = {
        "student": student,
        "department_obj": department_obj,
        "department": department,
        "semester": semester,
        "batch": batch,
        "section": section,
        "co_curriculars": co_curriculars,
        "activity_choices": StudentCO_EX_Curricular.ACTIVITY_CHOICES,
        "level_choices": StudentCO_EX_Curricular.LEVEL_CHOICES,
    }

    return render(
        request,
        "student_management/student/student_co_ex_curricular_view.html",
        context
    )
 

from django.shortcuts import render, redirect, get_object_or_404
from student_management.models import StudentProfessionl
from user_accounts.models import StudentDetails, Add_Department

@student_management
def student_professional_view(request):
    # ✅ StudentDetails from Employee_id -> reg_no
    student_id = request.user.Employee_id
    student = StudentDetails.objects.filter(reg_no=student_id).first()

    if not student:
        return redirect("profile")

    # ✅ Department from StudentDetails (most reliable)
    department_obj = student.department if student.department else None

    # ✅ Academic info from StudentDetails (as you requested)
    semester = student.semester if student.semester else "-"
    batch = student.batch if student.batch else "-"
    section = student.section if student.section else "-"
    year = student.year if student.year else "-"

    # 🔹 POST (Edit Entry)
    if request.method == "POST":
        entry_id = request.POST.get("entry_id")

        prof = get_object_or_404(
            StudentProfessionl,
            id=entry_id,
            student=student
        )

        prof.bodyName = request.POST.get("bodyName", "").strip()
        prof.validity = request.POST.get("validity", "").strip()

        # ✅ store academic info in model (optional but consistent with your other pages)
        prof.department = department_obj
        prof.year = year if year != "-" else prof.year
        prof.semester = semester if semester != "-" else prof.semester
        prof.section = section if section != "-" else prof.section
        prof.academic_year = request.POST.get("academic_year", "").strip() or prof.academic_year

        # ✅ Don't touch is_verified
        prof.save()
        return redirect("student_professional_view")

    # ✅ Correct filtering: model has student FK, not register_no
    professionals = StudentProfessionl.objects.filter(student=student).order_by("-id")

    context = {
        "student": student,
        "professionals": professionals,
        "department_obj": department_obj,
        "semester": semester,
        "batch": batch,
        "section": section,
        "year": year,
    }

    return render(
        request,
        "student_management/student/student_professional_body_view.html",
        context,
    )
 

import datetime
from django.shortcuts import render, redirect, get_object_or_404
from user_accounts.models import StudentDetails
from student_management.models import StudentProjects  # keep as you already use

@student_management
def student_projects_view(request):
    student_id = getattr(request.user, "Employee_id", None)
    student = StudentDetails.objects.filter(reg_no=student_id).select_related("department").first()

    if not student:
        return redirect("profile")

    department_obj = student.department if student.department else None
    department_name = getattr(department_obj, "Department", str(department_obj)) if department_obj else ""

    semester = student.semester if student.semester else "-"
    batch = student.batch if student.batch else "-"
    section = student.section if student.section else "-"
    year = student.year if student.year else "-"

    # ✅ NEW: direct M2M filter
    projects = (
        StudentProjects.objects.filter(students__reg_no=student_id)
        .distinct()
        .order_by("-id")
        .prefetch_related("students")
    )

    # -----------------------------
    # POST: edit project (only if this student belongs to that project)
    # -----------------------------
    if request.method == "POST":
        project_id = request.POST.get("project_id")
        project = get_object_or_404(StudentProjects, id=project_id, students__reg_no=student_id)

        project.title = (request.POST.get("title") or "").strip()
        project.domain = (request.POST.get("domain") or "").strip()

        status_val = (request.POST.get("status") or "").strip()
        if status_val in dict(StudentProjects.STATUS_CHOICES):
            project.status = status_val

        activity_val = (request.POST.get("activity_name") or "").strip()
        if activity_val in dict(StudentProjects.ACTIVITY_CHOICES):
            project.activity_name = activity_val

        project.organisation = (request.POST.get("organisation") or "").strip()
        project.place = (request.POST.get("place") or "").strip()

        date_val = (request.POST.get("date") or "").strip()
        if date_val:
            try:
                project.date = datetime.datetime.strptime(date_val, "%Y-%m-%d").date()
            except ValueError:
                pass

        # keep approval_status unchanged by student
        project.save()
        return redirect("student_projects_view")

    context = {
        "student": student,
        "projects": projects,
        "STATUS_CHOICES": StudentProjects.STATUS_CHOICES,
        "ACTIVITY_CHOICES": StudentProjects.ACTIVITY_CHOICES,
        "department_obj": department_obj,
        "department": department_name,
        "semester": semester,
        "batch": batch,
        "section": section,
        "year": year,
    }
    return render(request, "student_management/student/student_projects_view.html", context)
 


import datetime
from django.shortcuts import render, redirect, get_object_or_404
from student_management.models import StudentPublication
from user_accounts.models import StudentDetails

@student_management
def student_publications_view(request):
    student_id = request.user.Employee_id
    student = StudentDetails.objects.filter(reg_no=student_id).first()

    # ✅ StudentDetails (reg_no)
    if not student:
        return redirect("profile")

    # ✅ Department from StudentDetails
    department_obj = student.department if student.department else None
    department_name = getattr(department_obj, "Department", str(department_obj)) if department_obj else ""

    # ✅ Academic info from StudentDetails
    semester = student.semester if student.semester else "-"
    batch = student.batch if student.batch else "-"
    section = student.section if student.section else "-"
    year = student.year if student.year else "-"

    # ✅ Fetch publications for this logged-in student (MODEL HAS student FK, not register_no)
    publications = StudentPublication.objects.filter(student=student).order_by("-id")

    # 🔹 POST (Edit)
    if request.method == "POST":
        pub_id = request.POST.get("pub_id")

        publication = get_object_or_404(
            StudentPublication,
            id=pub_id,
            student=student
        )

        publication.authors = request.POST.get("authors", "").strip()
        publication.title = request.POST.get("title", "").strip()
        publication.program_name = request.POST.get("program_name", "").strip()
        publication.volume = request.POST.get("volume", "").strip() or None

        presented_val = request.POST.get("presented", "").strip()
        if presented_val:
            publication.presented = presented_val

        # ✅ Auto-sync academic info + dept (same pattern as your other pages)
        publication.department = department_obj
        publication.batch = None if batch == "-" else batch
        publication.semester = None if semester == "-" else semester
        publication.section = None if section == "-" else section
        publication.year = None if year == "-" else year
        # publication.academic_year = ... (only if you want, otherwise leave)

        # ✅ Date
        pub_date = request.POST.get("publication_date", "").strip()
        if pub_date:
            try:
                publication.publication_date = datetime.datetime.strptime(pub_date, "%Y-%m-%d").date()
            except ValueError:
                pass

        # ❌ Do NOT update is_verified
        publication.save()
        return redirect("student_publications_view")

    context = {
        "student": student,
        "publications": publications,
        "PRESENTATION_CHOICES": StudentPublication.PRESENTATION_CHOICES,
        "department_obj": department_obj,
        "department": department_name,
        "semester": semester,
        "batch": batch,
        "section": section,
        "year": year,
    }

    return render(request, "student_management/student/student_publications_view.html", context)
 

@check_permission("student_activity_upload")
@student_management
def student_activity_upload(request):
    return render(request, "student_management/student/student_activity_upload.html")


from django.shortcuts import render, redirect
from django.contrib import messages
from user_accounts.models import StudentDetails
from student_management.models import StudentCO_EX_Curricular

@student_management
def student_co_ex_curricular_upload(request):
    register_no = getattr(request.user, "Employee_id", None)
    if not register_no:
        messages.error(request, "⚠ No Register Number found. Please log in.")
        return redirect("student_overview")

    try:
        student = StudentDetails.objects.get(reg_no=register_no)
    except StudentDetails.DoesNotExist:
        messages.error(request, "Student details not found.")
        return redirect("student_overview")

    if request.method == "POST" and request.FILES.get("certificate"):
        StudentCO_EX_Curricular.objects.create(
            student=student,
            batch=student.batch,
            department=student.department,
            mentor=student.mentor,
            year=student.year,
            semester=student.semester,
            section=student.section,
            academic_year=request.POST.get("academic_year"),
            activity_type=request.POST.get("activity_type"),
            event_name=request.POST.get("event_name"),
            level=request.POST.get("level"),
            from_date=request.POST.get("from_date"),
            to_date=request.POST.get("to_date"),
            total_days=request.POST.get("total_days"),
            certificate=request.FILES["certificate"],
            status="Pending",  # ✅ default
        )
        messages.success(request, "✅ Co/Extra-Curricular certificate uploaded successfully!")
        return redirect("student_co_ex_curricular_upload")

    uploaded_records = StudentCO_EX_Curricular.objects.filter(
        student__reg_no=register_no
    ).order_by("-from_date")

    return render(
        request,
        "student_management/student/upload/student_co_ex_curricular_upolad.html",
        {
            "uploaded_records": uploaded_records,
            "student_name": student.name,
        },
    )


# -----------------------
# Publications Upload
# -----------------------
from django.shortcuts import render, redirect
from django.contrib import messages
from user_accounts.models import StudentDetails
from student_management.models import StudentPublication

@student_management
def student_publications_upload(request):
    register_no = getattr(request.user, "Employee_id", None)
    if not register_no:
        messages.error(request, "⚠ No Register Number found. Please log in.")
        return redirect("student_dashboard")

    try:
        student = StudentDetails.objects.select_related("department").get(reg_no=register_no)
    except StudentDetails.DoesNotExist:
        messages.error(request, "Student details not found.")
        return redirect("student_dashboard")

    if request.method == "POST":
        StudentPublication.objects.create(
            student=student,
            mentor=student.mentor,
            department=student.department,
            batch=student.batch,
            year=student.year,
            semester=student.semester,
            section=student.section,
            academic_year=request.POST.get("academic_year"),

            authors=request.POST.get("authors"),
            title=request.POST.get("title"),
            program_name=request.POST.get("program_name"),
            publication_date=request.POST.get("publication_date"),
            volume=request.POST.get("volume"),
            presented=request.POST.get("presented"),

            status="Pending",  # ✅
        )
        messages.success(request, " Publication submitted successfully!")
        return redirect("student_publications_upload")

    uploaded_publications = StudentPublication.objects.filter(
        student__reg_no=register_no
    ).order_by("-publication_date")

    return render(
        request,
        "student_management/student/upload/student_publications_upload.html",
        {
            "uploaded_publications": uploaded_publications,
            "student_name": student.name,
            "student": student,
        },
    )


# -----------------------
# Achievements Upload
# -----------------------
from django.shortcuts import render, redirect
from django.contrib import messages
from user_accounts.models import StudentDetails
from student_management.models import StudentAchievements

@student_management
def student_achievements_upload(request):
    register_no = getattr(request.user, "Employee_id", None)
    if not register_no:
        messages.error(request, "⚠ No Register Number found. Please log in.")
        return redirect("student_dashboard")

    try:
        student = StudentDetails.objects.select_related("department").get(reg_no=register_no)
    except StudentDetails.DoesNotExist:
        messages.error(request, "Student details not found.")
        return redirect("student_dashboard")

    if request.method == "POST" and request.FILES.get("certificate"):
        StudentAchievements.objects.create(
            student=student,
            mentor=student.mentor,
            department=student.department,
            batch=student.batch,
            section=student.section,
            semester=student.semester,
            year=student.year,
            academic_year=request.POST.get("academic_year"),

            date=request.POST.get("date"),
            award_name=request.POST.get("award_name"),
            contest=request.POST.get("contest"),
            given_by=request.POST.get("given_by"),
            certificate=request.FILES["certificate"],

            status="Pending",  # ✅ default
        )
        messages.success(request, " Achievement uploaded successfully!")
        return redirect("student_achievements_upload")

    uploaded_achievements = StudentAchievements.objects.filter(
        student__reg_no=register_no
    ).order_by("-date")

    return render(
        request,
        "student_management/student/upload/student_achievements_upload.html",
        {
            "uploaded_achievements": uploaded_achievements,
            "student_name": student.name,
            "register_no": register_no,
            "department": student.department,
            "batch": student.batch,
            "section": student.section,
            "semester": student.semester,
        },
    )

# -----------------------
# Projects Upload
# -----------------------
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.contrib import messages

from user_accounts.models import StudentDetails
from student_management.models import StudentProjects
from student_management.models import StudentProjects



from django.contrib import messages
from django.http import JsonResponse
from django.shortcuts import redirect, render

from user_accounts.models import StudentDetails
from student_management.models import StudentProjects


@student_management
def student_project_upload(request):
    # -----------------------------
    # AJAX: verify students by reg no
    # -----------------------------
    if request.GET.get("verify_students") == "1":
        reg_list = request.GET.getlist("regs[]")
        response_data = []

        # empty check
        if not reg_list:
            return JsonResponse({"students": []})

        # duplicate check
        if len(reg_list) != len(set(reg_list)):
            return JsonResponse({"error": "Duplicate register numbers found"}, status=400)

        for reg_no in reg_list:
            reg_no = (reg_no or "").strip()
            if not reg_no:
                response_data.append({"reg_no": reg_no, "error": "Empty register number"})
                continue

            try:
                student = StudentDetails.objects.select_related("department").get(reg_no=reg_no)
                response_data.append(
                    {
                        "reg_no": student.reg_no,
                        "name": getattr(student, "name", ""),
                        "department": student.department.Department if getattr(student, "department", None) else "",
                        "batch": getattr(student, "batch", ""),
                        "section": getattr(student, "section", ""),
                        "semester": getattr(student, "semester", ""),
                        "year": getattr(student, "year", ""),
                    }
                )
            except StudentDetails.DoesNotExist:
                response_data.append({"reg_no": reg_no, "error": "Register number not found"})

        return JsonResponse({"students": response_data})

    # -----------------------------
    # FORM SUBMIT
    # -----------------------------
    if request.method == "POST":
        # mentor id logic (your pattern - kept)
        mentor_id = (
            StudentDetails.objects.filter(reg_no=getattr(request.user, "Employee_id", None))
            .values_list("mentor_id", flat=True)
            .first()
        )

        title = (request.POST.get("title") or "").strip()
        domain = (request.POST.get("domain") or "").strip()
        status = (request.POST.get("status") or "pending").strip()  # pending/on going/completed
        activity_name = (request.POST.get("activity_name") or "").strip()
        organisation = (request.POST.get("organisation") or "").strip()
        place = (request.POST.get("place") or "").strip()
        date = request.POST.get("date") or None
        academic_year = (request.POST.get("academic_year") or "").strip()

        # received regs from frontend
        student_regs = request.POST.getlist("student_register_no[]")
        student_regs = [(r or "").strip() for r in student_regs if (r or "").strip()]

        if not student_regs:
            messages.error(request, "Please add at least one student register number.")
            return redirect("student_project_upload")

        # prevent duplicates in POST
        if len(student_regs) != len(set(student_regs)):
            messages.error(request, "Duplicate register numbers found. Please remove duplicates and try again.")
            return redirect("student_project_upload")

        # fetch students
        students_qs = StudentDetails.objects.select_related("department").filter(reg_no__in=student_regs)
        students_map = {s.reg_no: s for s in students_qs}

        # ensure all exist
        missing = [r for r in student_regs if r not in students_map]
        if missing:
            messages.error(request, f"These register numbers are not found: {', '.join(missing)}")
            return redirect("student_project_upload")

        created_count = 0
        created_regs = []

        # ✅ create ONE StudentProjects row per student (because model has FK student)
        for reg_no in student_regs:
            stu = students_map[reg_no]

            StudentProjects.objects.create(
                student=stu,                               # ✅ FK
                department=getattr(stu, "department", None),
                semester=getattr(stu, "semester", None),
                section=getattr(stu, "section", None),
                year=getattr(stu, "year", None),
                batch=getattr(stu, "batch", None),

                title=title,
                mentor_id=mentor_id,
                domain=domain,
                status=status,
                activity_name=activity_name,
                organisation=organisation,
                place=place,
                date=date,
                academic_year=academic_year,
                approval_status="Pending",
            )

            created_count += 1
            created_regs.append(reg_no)

        messages.success(
            request,
            f"Project '{title}' saved successfully for {created_count} student(s): {', '.join(created_regs)}"
        )
        return redirect("student_project_upload")

    return render(request, "student_management/student/upload/student_projects_upload.html")

# -----------------------
# Professional Body Upload
# -----------------------
from django.shortcuts import render, redirect
from django.contrib import messages
from user_accounts.models import StudentDetails
from student_management.models import StudentProfessionl

@student_management
def student_professional_body_upload(request):
    register_no = getattr(request.user, "Employee_id", None)
    if not register_no:
        messages.error(request, "User register number not found.")
        return redirect("student_dashboard")

    try:
        student = StudentDetails.objects.select_related("department").get(reg_no=register_no)
    except StudentDetails.DoesNotExist:
        messages.error(request, "Student details not found.")
        return redirect("student_dashboard")

    if request.method == "POST":
        body_name = request.POST.get("bodyName")
        validity = request.POST.get("validity")
        academic_year = request.POST.get("academic_year")

        if not all([body_name, validity, academic_year]):
            messages.error(request, "Please fill all required fields.")
            return redirect("student_professional_body_upload")

        StudentProfessionl.objects.create(
            mentor=student.mentor,
            student=student,
            department=student.department,
            semester=student.semester,
            section=student.section,
            year=student.year,
            academic_year=academic_year,
            bodyName=body_name,
            validity=validity,
            status="Pending",  # ✅
        )

        messages.success(request, "Professional body entry submitted successfully!")
        return redirect("student_professional_body_upload")

    context = {
        "name": student.name,
        "reg_no": student.reg_no,
        "department": student.department.Department if student.department else "",
        "section": student.section,
        "semester": student.semester,
        "year": student.year,
    }

    return render(
        request,
        "student_management/student/upload/student_professional_body_upload.html",
        context,
    )


from django.shortcuts import render, get_object_or_404
from django.db.models import Count, Q
from student_management.models import HourAttendance, Daily_Attendance
from course_management.models import CourseHours


@check_permission("student_attendance")
@student_management
def student_attendance(request):
    """
    Student attendance dashboard:
    - Hour-wise attendance from HourAttendance
    - Daily attendance from Daily_Attendance (half-day logic handled)
    """

    # 🧩 Get student details
    try:
        student = StudentDetails.objects.get(reg_no=request.user.Employee_id)
    except StudentDetails.DoesNotExist:
        return render(
            request,
            "student_management/student/student_attendance.html",
            {"error": "⚠ Student record not found. Please contact admin."},
        )

    reg_no = student.reg_no

    try:
        current_sem = int(student.semester or 1)
    except (TypeError, ValueError):
        current_sem = 1

    # Initialize structure for each semester
    semesters = {
        sem: {
            "hourwise": {"courses": [], "total": 0, "attended": 0, "percentage": 0.0},
            "daily": {"total_days": 0, "present_days": 0.0, "percentage": 0.0},
        }
        for sem in range(1, current_sem + 1)
    }

    # ✅ Enrolled courses (only up to current semester)
    enrolled_courses = list(
        CourseEnrollment.objects.filter(
            student=student,
            enroll=True,
            course__semester__lte=current_sem,
        )
        .select_related("course")
        .values(
            "course_id",
            "course__course_code",
            "course__title",
            "course__semester",
        )
    )

    course_ids = [c["course_id"] for c in enrolled_courses]

    # ✅ Hour-wise Attendance (single query aggregation)
    hour_stats = (
        HourAttendance.objects.filter(student=student, course_id__in=course_ids)
        .values("course_id")
        .annotate(
            total=Count("id"),
            attended=Count("id", filter=Q(status="Present")),
        )
    )
    hour_map = {h["course_id"]: h for h in hour_stats}

    for row in enrolled_courses:
        course_id = row["course_id"]
        sem = int(row["course__semester"] or 1)

        total_hours = hour_map.get(course_id, {}).get("total", 0)
        attended_hours = hour_map.get(course_id, {}).get("attended", 0)

        percentage = round((attended_hours / total_hours) * 100, 2) if total_hours else 0.0

        semesters[sem]["hourwise"]["courses"].append(
            {
                "code": row["course__course_code"],
                "name": row["course__title"],
                "total": total_hours,
                "attended": attended_hours,
                "percentage": percentage,
            }
        )

        semesters[sem]["hourwise"]["total"] += total_hours
        semesters[sem]["hourwise"]["attended"] += attended_hours

    # ✅ Compute hour-wise semester totals
    for sem, data in semesters.items():
        total = data["hourwise"]["total"]
        attended = data["hourwise"]["attended"]
        data["hourwise"]["percentage"] = round((attended / total) * 100, 2) if total else 0.0
        data["hourwise"]["courses"].sort(key=lambda c: (c["code"] or ""))

    # ✅ Daily Attendance (keep records once; compute per sem)
    daily_records = Daily_Attendance.objects.filter(student=student).only(
        "semester", "morning_status", "afternoon_status"
    )

    # Group daily by semester in python (faster than repeated DB filters)
    daily_by_sem = defaultdict(list)
    for d in daily_records:
        try:
            sem = int(d.semester or 0)
        except (TypeError, ValueError):
            continue
        if 1 <= sem <= current_sem:
            daily_by_sem[sem].append(d)

    for sem in range(1, current_sem + 1):
        records = daily_by_sem.get(sem, [])
        total_days = len(records)

        present_days = 0.0
        for d in records:
            morning = d.morning_status
            afternoon = d.afternoon_status

            # Full day present / on duty
            if morning in ("Presentresent","Present", "On Duty") and afternoon in ("Present", "On Duty"):
                present_days += 1.0
            # Half day
            elif (morning in ("Present", "On Duty") and afternoon == "Absent") or (
                morning == "Absent" and afternoon in ("Present", "On Duty")
            ):
                present_days += 0.5

        percentage = round((present_days / total_days) * 100, 2) if total_days else 0.0

        semesters[sem]["daily"] = {
            "total_days": total_days,
            "present_days": present_days,
            "percentage": percentage,
        }

    semesters_list = sorted(semesters.items(), key=lambda x: x[0])

    return render(
        request,
        "student_management/student/student_attendance.html",
        {
            "student_name": student.name or "Student",
            "student_reg_no": reg_no,
            "current_sem": current_sem,
            "semesters": semesters,
            "semesters_list": semesters_list,
        },
    )

import io
import os
from datetime import datetime
from collections import defaultdict

from django.http import FileResponse, HttpResponse
from django.shortcuts import render
from django.conf import settings
from django.contrib.staticfiles import finders
from django.db.models import Count, Q

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame,
    Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
)

from student_management.models import HourAttendance, Daily_Attendance
from user_accounts.models import StudentDetails
from course_management.models import CourseEnrollment


def _safe(v):
    return "" if v is None else str(v).strip()


@check_permission("student_attendance")
@student_management
def student_attendance_pdf(request):
    # ----------------------------
    # Get Student
    # ----------------------------
    try:
        student = StudentDetails.objects.get(reg_no=request.user.Employee_id)
    except StudentDetails.DoesNotExist:
        return HttpResponse("Student record not found.", status=404)

    reg_no = student.reg_no
    student_name = _safe(student.name) or "Student"

    try:
        current_sem = int(student.semester or 1)
    except (TypeError, ValueError):
        current_sem = 1

    # ----------------------------
    # Build same data structure as dashboard
    # ----------------------------
    semesters = {
        sem: {
            "hourwise": {"courses": [], "total": 0, "attended": 0, "percentage": 0.0},
            "daily": {"total_days": 0, "present_days": 0.0, "percentage": 0.0},
        }
        for sem in range(1, current_sem + 1)
    }

    enrolled_courses = list(
        CourseEnrollment.objects.filter(
            student=student,
            enroll=True,
            course__semester__lte=current_sem,
        )
        .select_related("course")
        .values(
            "course_id",
            "course__course_code",
            "course__title",
            "course__semester",
        )
    )

    course_ids = [c["course_id"] for c in enrolled_courses]

    # Hour-wise aggregation
    hour_stats = (
        HourAttendance.objects.filter(student=student, course_id__in=course_ids)
        .values("course_id")
        .annotate(
            total=Count("id"),
            attended=Count("id", filter=Q(status="Present")),
        )
    )
    hour_map = {h["course_id"]: h for h in hour_stats}

    for row in enrolled_courses:
        course_id = row["course_id"]
        sem = int(row["course__semester"] or 1)

        total_hours = hour_map.get(course_id, {}).get("total", 0)
        attended_hours = hour_map.get(course_id, {}).get("attended", 0)
        percentage = round((attended_hours / total_hours) * 100, 2) if total_hours else 0.0

        semesters[sem]["hourwise"]["courses"].append(
            {
                "code": row["course__course_code"],
                "name": row["course__title"],
                "total": total_hours,
                "attended": attended_hours,
                "percentage": percentage,
            }
        )

        semesters[sem]["hourwise"]["total"] += total_hours
        semesters[sem]["hourwise"]["attended"] += attended_hours

    for sem, data in semesters.items():
        total = data["hourwise"]["total"]
        attended = data["hourwise"]["attended"]
        data["hourwise"]["percentage"] = round((attended / total) * 100, 2) if total else 0.0
        data["hourwise"]["courses"].sort(key=lambda c: (c["code"] or ""))

    # Daily attendance
    daily_records = Daily_Attendance.objects.filter(student=student).only(
        "semester", "morning_status", "afternoon_status"
    )

    daily_by_sem = defaultdict(list)
    for d in daily_records:
        try:
            sem = int(d.semester or 0)
        except (TypeError, ValueError):
            continue
        if 1 <= sem <= current_sem:
            daily_by_sem[sem].append(d)

    for sem in range(1, current_sem + 1):
        records = daily_by_sem.get(sem, [])
        total_days = len(records)

        present_days = 0.0
        for d in records:
            morning = d.morning_status
            afternoon = d.afternoon_status

            if morning in ("Present", "On Duty") and afternoon in ("Present", "On Duty"):
                present_days += 1.0
            elif (morning in ("Present", "On Duty") and afternoon == "Absent") or (
                morning == "Absent" and afternoon in ("Present", "On Duty")
            ):
                present_days += 0.5

        percentage = round((present_days / total_days) * 100, 2) if total_days else 0.0

        semesters[sem]["daily"] = {
            "total_days": total_days,
            "present_days": present_days,
            "percentage": percentage,
        }

    # ----------------------------
    # PDF Styles
    # ----------------------------
    styles = getSampleStyleSheet()

    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1a4b8c")
    ACCENT_RED = colors.HexColor("#b91c1c")
    DARK_GRAY = colors.HexColor("#111827")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BG_GRAY = colors.HexColor("#f8fafc")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle(
        "title_style",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=PRIMARY_BLUE,
        alignment=TA_CENTER,
        spaceAfter=6,
        fontName="Helvetica-Bold",
    )

    sub_style = ParagraphStyle(
        "sub_style",
        parent=styles["Normal"],
        fontSize=10,
        textColor=MEDIUM_GRAY,
        alignment=TA_CENTER,
        spaceAfter=10,
    )

    section_style = ParagraphStyle(
        "section_style",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=PRIMARY_BLUE,
        alignment=TA_LEFT,
        spaceBefore=8,
        spaceAfter=6,
        fontName="Helvetica-Bold",
    )

    table_header = ParagraphStyle(
        "table_header",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.white,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
    )

    cell_left = ParagraphStyle(
        "cell_left",
        parent=styles["Normal"],
        fontSize=9,
        textColor=DARK_GRAY,
        alignment=TA_LEFT,
        leading=11,
        wordWrap="CJK",
    )

    cell_center = ParagraphStyle(
        "cell_center",
        parent=cell_left,
        alignment=TA_CENTER,
    )

    def make_table(data, col_widths, header_bg=SECONDARY_BLUE, zebra=True):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        ts = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ])
        if zebra and len(data) > 1:
            ts.add("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY])
        t.setStyle(ts)
        return t

    # ----------------------------
    # Document setup
    # ----------------------------
    buffer = io.BytesIO()
    doc = BaseDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Attendance - {reg_no}",
        showBoundary=0
    )

    HEADER_HEIGHT = 36 * mm

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = A4
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 8 * mm

        # Logo
        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(
                ImageReader(logo_path),
                left, top_y - 20 * mm,
                width=30 * mm, height=18 * mm,
                preserveAspectRatio=True, mask="auto"
            )

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawCentredString(center_x, top_y - 6 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "An Autonomous Institution")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 8.2)
        canvas.drawCentredString(center_x, top_y - 18.5 * mm, "Approved by AICTE, New Delhi")
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(center_x, top_y - 23 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
        canvas.drawCentredString(center_x, top_y - 27.5 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

        # Footer line
        footer_y = 18 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.8)
        canvas.line(left, footer_y + 7 * mm, right, footer_y + 7 * mm)

        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 8)
        gen_time = datetime.now().strftime("%d %b %Y, %I:%M %p")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawCentredString(center_x, footer_y, f"Reg No: {reg_no}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")

        canvas.restoreState()

    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin + 6 * mm,
        doc.width,
        doc.height - HEADER_HEIGHT + 8 * mm,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        id="normal"
    )
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

    # ----------------------------
    # Build PDF content
    # ----------------------------
    elements = []
    elements.append(Spacer(1, 5 * mm))
    elements.append(Paragraph("STUDENT ATTENDANCE REPORT", title_style))
    elements.append(Paragraph(f"{student_name} ({reg_no})", sub_style))

    # Semester sections
    for sem in range(1, current_sem + 1):
        data = semesters.get(sem, {})
        hourwise = data.get("hourwise", {})
        daily = data.get("daily", {})

        elements.append(Paragraph(f"Semester {sem}", section_style))
        elements.append(Spacer(1, 2 * mm))

        # Hour-wise table
        elements.append(Paragraph("Hour-wise Attendance", ParagraphStyle(
            "mini", parent=styles["Normal"], fontSize=11, textColor=SECONDARY_BLUE,
            fontName="Helvetica-Bold", spaceAfter=4
        )))

        hw_rows = [[
            Paragraph("S.No", table_header),
            Paragraph("Course Code", table_header),
            Paragraph("Course Name", table_header),
            Paragraph("Total", table_header),
            Paragraph("Attended", table_header),
            Paragraph("%", table_header),
            Paragraph("Status", table_header),
        ]]

        courses = hourwise.get("courses", []) or []
        if courses:
            for i, c in enumerate(courses, start=1):
                pct = float(c.get("percentage", 0) or 0)
                status = "Good" if pct >= 75 else ("Moderate" if pct >= 50 else "Low")

                hw_rows.append([
                    Paragraph(str(i), cell_center),
                    Paragraph(_safe(c.get("code")) or "—", cell_center),
                    Paragraph(_safe(c.get("name")) or "—", cell_left),
                    Paragraph(str(c.get("total", 0) or 0), cell_center),
                    Paragraph(str(c.get("attended", 0) or 0), cell_center),
                    Paragraph(f"{pct:.2f}%", cell_center),
                    Paragraph(status, cell_center),
                ])

            # totals row
            hw_rows.append([
                Paragraph("<b>Total</b>", cell_center),
                Paragraph("-", cell_center),
                Paragraph("-", cell_center),
                Paragraph(f"<b>{hourwise.get('total', 0) or 0}</b>", cell_center),
                Paragraph(f"<b>{hourwise.get('attended', 0) or 0}</b>", cell_center),
                Paragraph(f"<b>{hourwise.get('percentage', 0) or 0:.2f}%</b>", cell_center),
                Paragraph("-", cell_center),
            ])
        else:
            hw_rows.append([Paragraph("No hour-wise records found.", cell_left)] + [""] * 6)

        elements.append(make_table(
            hw_rows,
            col_widths=[12*mm, 26*mm, doc.width - (12+26+16+18+18+18)*mm, 16*mm, 18*mm, 18*mm, 18*mm],
            header_bg=SECONDARY_BLUE
        ))
        elements.append(Spacer(1, 6 * mm))

        # Daily summary
        elements.append(Paragraph("Daily Attendance (Summary)", ParagraphStyle(
            "mini2", parent=styles["Normal"], fontSize=11, textColor=SECONDARY_BLUE,
            fontName="Helvetica-Bold", spaceAfter=4
        )))

        daily_rows = [[
            Paragraph("Total Days", table_header),
            Paragraph("Present Days", table_header),
            Paragraph("Percentage", table_header),
        ]]

        total_days = daily.get("total_days", 0) or 0
        present_days = daily.get("present_days", 0) or 0
        daily_pct = daily.get("percentage", 0) or 0

        daily_rows.append([
            Paragraph(str(total_days), cell_center),
            Paragraph(str(present_days), cell_center),
            Paragraph(f"{daily_pct:.2f}%", cell_center),
        ])

        elements.append(make_table(
            daily_rows,
            col_widths=[doc.width/3, doc.width/3, doc.width/3],
            header_bg=SECONDARY_BLUE,
            zebra=False
        ))

        # page break between semesters (not after last)
        if sem != current_sem:
            elements.append(PageBreak())

    # Build & return
    try:
        doc.build(elements)
    except Exception as e:
        return HttpResponse(f"PDF generation failed: {e}", status=500)

    buffer.seek(0)
    filename = f"Attendance_{reg_no}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename)




from examination_management.models import Part, ExamPattern, StudentInternalMark, ExamPatternSetting
from course_management.models import CourseHours
from collections import defaultdict
from decimal import Decimal

from django.db.models import Sum, F, Value
from django.db.models.functions import Coalesce
from django.shortcuts import render

from collections import defaultdict
from django.shortcuts import render

from collections import defaultdict
from django.shortcuts import render

from collections import defaultdict
from django.shortcuts import render

# make sure these are imported
# from .models import StudentInternalMark, ExamPattern, Part, Question, OptionMarks, CourseHours
# from student_management.models import StudentDetails


def _safe_int(value, default=0):
    try:
        if value is None:
            return default
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def _text(value):
    return str(value or "").strip()


def _option_key(value):
    return str(value or "").strip().lower()


def _sub_key(value):
    return str(value or "").strip().lower()


def _build_course_credit_map(course_ids):
    credit_map = {}

    if not course_ids:
        return credit_map

    course_hours_qs = (
        CourseHours.objects
        .filter(course_id__in=course_ids)
        .order_by("course_id", "-id")
    )

    for ch in course_hours_qs:
        if ch.course_id in credit_map:
            continue
        credit_map[ch.course_id] = _text(ch.credits) or "0"

    return credit_map


def _build_pattern_lookup(pattern_ids):
    """
    pattern_lookup = {
        pattern_id: {
            "parts": {
                "A": {
                    "questions": {
                        1: {
                            "question_total": 2,
                            "options": {}
                        },
                        13: {
                            "question_total": 0,
                            "options": {
                                "a": {"i": 7, "ii": 6, "total": 13},
                                "b": {"i": 7, "ii": 6, "total": 13},
                            }
                        }
                    }
                }
            },
            "pattern_total": 100
        }
    }
    """
    pattern_lookup = {}

    if not pattern_ids:
        return pattern_lookup

    parts_qs = (
        Part.objects
        .filter(exam_pattern_id__in=pattern_ids)
        .prefetch_related("questions__options")
        .order_by("exam_pattern_id", "name", "id")
    )

    for part in parts_qs:
        pattern_id = part.exam_pattern_id
        part_name = _text(part.name)

        if pattern_id not in pattern_lookup:
            pattern_lookup[pattern_id] = {
                "parts": {},
                "pattern_total": 0,
            }

        if part_name not in pattern_lookup[pattern_id]["parts"]:
            pattern_lookup[pattern_id]["parts"][part_name] = {
                "questions": {}
            }

        for question in part.questions.all():
            q_total = question.total_marks or 0
            options_map = {}

            for option in question.options.all():
                opt_key = _option_key(option.option_letter)
                options_map[opt_key] = {
                    "i": option.marks_i or 0,
                    "ii": option.marks_ii or 0,
                    "total": (option.marks_i or 0) + (option.marks_ii or 0),
                }

            pattern_lookup[pattern_id]["parts"][part_name]["questions"][question.number] = {
                "question_total": q_total,
                "options": options_map,
            }

    # compute full paper total for each pattern
    for pattern_id, pdata in pattern_lookup.items():
        total = 0

        for part_name, part_data in pdata["parts"].items():
            for q_no, q_data in part_data["questions"].items():
                options = q_data["options"]

                if options:
                    # for optional questions, total should be counted ONCE per question
                    # use max option total (normally all options should have same total)
                    option_totals = [opt["total"] for opt in options.values()]
                    total += max(option_totals) if option_totals else 0
                else:
                    total += q_data["question_total"] or 0

        pdata["pattern_total"] = total

    return pattern_lookup


def _get_pattern_total(pattern_id, pattern_lookup):
    if not pattern_id:
        return 0
    return pattern_lookup.get(pattern_id, {}).get("pattern_total", 0)


from collections import defaultdict
from django.shortcuts import render

# make sure these imports already exist
# from .models import StudentInternalMark, CourseHours
# from student_management.models import StudentDetails


def _safe_text(value):
    return str(value or "").strip()


def _safe_int(value, default=0):
    try:
        if value is None:
            return default
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def _build_course_credit_map(course_ids):
    credit_map = {}

    if not course_ids:
        return credit_map

    course_hours_qs = (
        CourseHours.objects
        .filter(course_id__in=course_ids)
        .order_by("course_id", "-id")
    )

    for ch in course_hours_qs:
        if ch.course_id in credit_map:
            continue
        credit_map[ch.course_id] = _safe_text(ch.credits) or "0"

    return credit_map


def _get_effective_rows(rows):
    """
    Same logic as student_mark_pdf:
    Group by (part, question), then choose the option with highest obtained marks.
    For non-choice questions, keep the normal rows.
    """
    question_map = defaultdict(
        lambda: defaultdict(lambda: {"max": 0, "obt": 0, "rows": []})
    )

    for m in rows:
        part = _safe_text(m.part_name)
        qnum = _safe_text(m.question_number)
        opt = _safe_text(m.option_letter)  # '' => non-choice question

        question_map[(part, qnum)][opt]["max"] += _safe_int(m.max_marks)
        question_map[(part, qnum)][opt]["obt"] += _safe_int(m.marks_obtained)
        question_map[(part, qnum)][opt]["rows"].append(m)

    chosen_rows = []

    for (_part, _qnum), options in question_map.items():
        if "" in options and len(options) == 1:
            chosen_rows.extend(options[""]["rows"])
        else:
            chosen_key = max(options.keys(), key=lambda k: options[k]["obt"])
            chosen_rows.extend(options[chosen_key]["rows"])

    return chosen_rows


from django.shortcuts import render

from collections import defaultdict

def _exam_sort_key(name):
    """
    Sort exam names like:
    IAT1, IAT2, MODEL, SEMESTER, etc.
    """
    name = (name or "").strip().upper()

    m = re.match(r"([A-Z]+)\s*(\d+)$", name)
    if m:
        return (m.group(1), int(m.group(2)))

    m = re.match(r"([A-Z]+)(\d+)$", name)
    if m:
        return (m.group(1), int(m.group(2)))

    return (name, 999)


def get_internal_mark_overview(student, enrollment, exam_name=None):
    """
    Returns overview mark data for one student + one enrolled course.

    Uses the same logic as PDF:
    - groups by (part, question)
    - if A/B options exist, picks the option with highest obtained marks
    - totals are calculated only for rows that have CO mapping
    """

    base_qs = StudentInternalMark.objects.select_related(
        "co_code",
        "level_code",
        "enrollment__course",
    ).filter(
        student=student,
        enrollment=enrollment,
    )

    if exam_name:
        rows = list(base_qs.filter(exam_name=exam_name).order_by("created_at"))
        resolved_exam_name = exam_name
    else:
        resolved_exam_name = base_qs.order_by("-created_at").values_list(
            "exam_name", flat=True
        ).first()

        if resolved_exam_name:
            rows = list(
                base_qs.filter(exam_name=resolved_exam_name).order_by("created_at")
            )
        else:
            rows = []

    if not rows:
        return {
            "has_mark": False,
            "exam_name": resolved_exam_name or "",
            "total_max": 0,
            "total_obt": 0,
        }

    question_map = defaultdict(
        lambda: defaultdict(lambda: {"max": 0, "obt": 0, "rows": []})
    )

    for m in rows:
        part = m.part_name or ""
        qnum = m.question_number or ""
        opt = m.option_letter or ""

        question_map[(part, qnum)][opt]["max"] += int(m.max_marks or 0)
        question_map[(part, qnum)][opt]["obt"] += int(m.marks_obtained or 0)
        question_map[(part, qnum)][opt]["rows"].append(m)

    chosen_rows = []

    for (_part, _qnum), options in question_map.items():
        if "" in options and len(options) == 1:
            chosen = options[""]
        else:
            chosen_key = max(options.keys(), key=lambda k: options[k]["obt"])
            chosen = options[chosen_key]

        chosen_rows.extend(chosen["rows"])

    # Build CO totals exactly like PDF logic
    co_totals = defaultdict(lambda: {"max": 0, "obt": 0})
    for r in chosen_rows:
        co_code = getattr(r.co_code, "co_code", "") if r.co_code_id else "—"
        co_totals[co_code]["max"] += int(r.max_marks or 0)
        co_totals[co_code]["obt"] += int(r.marks_obtained or 0)

    # Same as PDF total: only valid CO rows
    total_max = sum(v["max"] for k, v in co_totals.items() if k != "—")
    total_obt = sum(v["obt"] for k, v in co_totals.items() if k != "—")

    return {
        "has_mark": True,
        "exam_name": resolved_exam_name or "",
        "total_max": total_max,
        "total_obt": total_obt,
    }


from collections import OrderedDict

from django.http import Http404
from django.shortcuts import render

# import your models properly
# from .models import StudentDetails, CourseEnrollment, StudentInternalMark
# from .decorators import check_permission, student_management


from collections import OrderedDict

from django.http import Http404
from django.shortcuts import render

# from .models import StudentDetails, CourseEnrollment, StudentInternalMark
# from .decorators import check_permission, student_management


def get_internal_mark_all_exams(student, enrollment):
    """
    Returns all exam-wise mark summaries for one enrolled course.
    Example:
    {
        "IAT1": {"has_mark": True, "total_obt": 64, "total_max": 100},
        "IAT2": {"has_mark": True, "total_obt": 58, "total_max": 100},
    }
    """
    exam_names = (
        StudentInternalMark.objects.filter(
            student=student,
            enrollment=enrollment,
        )
        .exclude(exam_name__isnull=True)
        .exclude(exam_name__exact="")
        .values_list("exam_name", flat=True)
        .distinct()
    )

    exam_names = sorted(exam_names, key=_exam_sort_key)

    exam_data = OrderedDict()
    for exam in exam_names:
        exam_data[exam] = get_internal_mark_overview(
            student=student,
            enrollment=enrollment,
            exam_name=exam,
        )

    return exam_data


@check_permission("student_mark")
@student_management
def student_mark(request):
    reg_no = request.user.Employee_id

    student = StudentDetails.objects.filter(reg_no=reg_no).select_related(
        "department__degree"
    ).first()

    if not student:
        raise Http404("Student not found")

    if not student.department or not student.department.degree:
        raise Http404("Student degree details not found")

    duration = int(student.department.degree.duration)
    total_semesters = duration * 2

    semester_cards = []
    semester_details = OrderedDict()

    active_semester = getattr(student, "semester", None)

    for sem in range(1, total_semesters + 1):
        enrollments = (
            CourseEnrollment.objects.filter(
                student=student,
                enroll=True,
                course__semester=sem,
            )
            .select_related("course")
            .prefetch_related("course__semesters")
            .order_by("course__course_code")
        )

        course_rows = []
        semester_exam_set = set()

        for enrollment in enrollments:
            course = enrollment.course
            course_hour = course.semesters.first()

            exam_wise_marks = get_internal_mark_all_exams(
                student=student,
                enrollment=enrollment,
            )

            semester_exam_set.update(exam_wise_marks.keys())

            available_marks = [
                value["total_obt"]
                for value in exam_wise_marks.values()
                if value.get("has_mark")
            ]
            available_totals = [
                value["total_max"]
                for value in exam_wise_marks.values()
                if value.get("has_mark")
            ]

            overall_obt = (
                round(sum(available_marks) / len(available_marks))
                if available_marks else None
            )
            overall_total = (
                round(sum(available_totals) / len(available_totals))
                if available_totals else None
            )

            latest_exam_name = None
            if exam_wise_marks:
                latest_exam_name = list(exam_wise_marks.keys())[-1]

            course_rows.append({
                "enrollment": enrollment,
                "course": course,
                "credits": course_hour.credits if course_hour else "-",
                "exam_wise_marks": exam_wise_marks,
                "overall_obt": overall_obt,
                "overall_total": overall_total,
                "latest_exam_name": latest_exam_name,
            })

        semester_exam_names = sorted(semester_exam_set, key=_exam_sort_key)

        semester_cards.append({
            "semester": sem,
            "course_count": len(course_rows),
        })

        semester_details[sem] = {
            "exam_names": semester_exam_names,
            "courses": course_rows,
        }

    if not active_semester or int(active_semester) not in semester_details:
        for sem, data in semester_details.items():
            if data["courses"]:
                active_semester = sem
                break

    if not active_semester:
        active_semester = 1

    context = {
        "student": student,
        "semester_cards": semester_cards,
        "semester_details": semester_details,
        "active_semester": int(active_semester),
    }

    return render(
        request,
        "student_management/student/student_mark_view.html",
        context,
    )






    duration = student.department.degree.duration
    total_semesters = duration * 2

    semester_courses = {}

    for sem in range(1, total_semesters + 1):
        courses = CourseEnrollment.objects.filter(
            student=student,
            enroll=True,
            course__semester=sem
        ).select_related("course").prefetch_related("course__semesters")

        semester_courses[sem] = courses

    context = {
        "student": student,
        "semester_courses": semester_courses,
    }

    return render(
        request,
        "student_management/student/student_mark_view.html",
        context
    )

@check_permission("student_graduation_details")
@student_management
def student_graduation_details(request):
    return render(request, "student_management/student/student_graduation_details.html")




from django.shortcuts import render
from django.contrib import messages
from course_management.models import PeriodAllocation
from user_accounts.models import StudentDetails, Add_Department

from datetime import datetime


@check_permission("student_timetable")
@student_management
def student_timetable(request):
    try:
        # ✅ Basic details (as you wrote)
        register_no = getattr(request.user, "Employee_id", None)
        department_name = getattr(getattr(request.user, "Department", None), "Department", None)

        # ✅ Student details
        student_obj = StudentDetails.objects.filter(reg_no=register_no).first()
        if not student_obj:
            return render(
                request,
                "student_management/student/student_timetable_view.html",
                {"allocations": [], "error_message": "⚠️ Student details not found."},
            )

        section = student_obj.section
        student_semester = student_obj.semester
        student_batch = student_obj.batch

        if not (register_no and department_name and section):
            return render(
                request,
                "student_management/student/student_timetable_view.html",
                {"allocations": [], "error_message": "⚠️ Missing student details (section/department)."},
            )

        # ✅ Department lookup
        department_obj = Add_Department.objects.filter(
            Department__iexact=department_name,
            is_active=True
        ).first()

        if not department_obj:
            return render(
                request,
                "student_management/student/student_timetable_view.html",
                {"allocations": [], "error_message": f"⚠️ Department not found: {department_name}"},
            )

        if not student_semester:
            return render(
                request,
                "student_management/student/student_timetable_view.html",
                {"allocations": [], "error_message": "⚠️ Semester not assigned yet."},
            )

        # ✅ Fetch timetable records
        allocations_qs = PeriodAllocation.objects.filter(
            department=department_obj,
            semester=student_semester,
            section=section,
        )

        weekday_order = {
            "Monday": 1, "Tuesday": 2, "Wednesday": 3, "Thursday": 4,
            "Friday": 5, "Saturday": 6, "Sunday": 7,
        }
        allocations = sorted(allocations_qs, key=lambda x: weekday_order.get(getattr(x, "day", ""), 99))

        all_periods = [
            "first_period", "second_period", "third_period", "fourth_period",
            "fifth_period", "sixth_period", "seventh_period", "eighth_period",
            "nineth_period", "tenth_period",
        ]

        # ✅ Only show columns having at least one value
        visible_periods = [p for p in all_periods if any(getattr(a, p, None) for a in allocations)]

        filtered_allocations = []
        used_course_codes = set()

        # ✅ Build period_display + collect used course codes
        for alloc in allocations:
            period_display = {}
            for p in visible_periods:
                code = (getattr(alloc, p, "") or "").strip()
                if code:
                    used_course_codes.add(code)
                period_display[p] = code
            alloc.period_display = period_display
            filtered_allocations.append(alloc)

        if not used_course_codes:
            # timetable exists but empty periods
            context = {
                "allocations": filtered_allocations,
                "student_department": department_obj.Department,
                "student_semester": student_semester,
                "student_section": section,
                "visible_periods": visible_periods,
                "student_batch": student_batch,
                "current_day": datetime.today().date(),
                "course_faculties": [],
                "effective_date": None,
                "total_credits": "0",
                "code_map": {},
            }
            return render(request, "student_management/student/student_timetable_view.html", context)

        # ✅ Faculty mapping only for used course codes
        course_faculties_qs = (
            AssignSubjectFaculty.objects.filter(
                department=department_obj,
                batch=student_batch,
                section=section,
                is_active=True,
                course__course_code__in=used_course_codes,
            )
            .select_related("course", "faculty", "skilled_faculty")
        )

        # ---------- helpers ----------
        def safe_float(v, default=0.0):
            if v is None:
                return default
            try:
                s = str(v).strip()
                if s == "":
                    return default
                return float(s)
            except Exception:
                return default

        # ✅ Fetch all CourseHours in ONE query (no duplicates per course)
        course_ids = [x.course_id for x in course_faculties_qs if x.course_id]
        hours_map = {
            ch.course_id: ch
            for ch in CourseHours.objects.filter(course_id__in=course_ids).only(
                "course_id", "leture_npwk", "tutorial_npwk", "laboratory_npwk", "total_hours", "credits"
            )
        }

        # ✅ Build course_data, total_credits, and code_map
        course_data = []
        total_credits = 0.0
        code_map = {}  # course_code -> course_code (or you can show title too)

        for item in course_faculties_qs:
            course = item.course
            cc = (course.course_code or "").strip() if course else ""
            title = (getattr(course, "title", "") or "").strip() if course else ""

            ch = hours_map.get(item.course_id)

            # IMPORTANT FIX: your model field is leture_npwk (not leture_hpwk)
            lecture = getattr(ch, "leture_npwk", None) if ch else None
            credits = getattr(ch, "credits", None) if ch else None

            total_credits += safe_float(credits, 0.0)

            if cc:
                # if you want CODE — TITLE in legend, use this:
                # code_map[cc] = f"{cc} — {title}" if title else cc
                code_map[cc] = cc

            course_data.append(
                {
                    "course": course,
                    "faculty": item.faculty,
                    "skilled_faculty": item.skilled_faculty,
                    "total_hours": lecture or "N/A",
                    "Credits": credits or "N/A",
                }
            )

        # ✅ Determine effective_date (latest non-null)
        effective_date = None
        dates = [a.effective_date for a in filtered_allocations if getattr(a, "effective_date", None)]
        if dates:
            effective_date = max(dates)

        # ✅ Total credits display
        try:
            total_credits_display = str(int(total_credits)) if float(total_credits).is_integer() else str(total_credits)
        except Exception:
            total_credits_display = str(total_credits)

        context = {
            "allocations": filtered_allocations,
            "student_department": department_obj.Department,
            "student_semester": student_semester,
            "student_section": section,
            "visible_periods": visible_periods,
            "student_batch": student_batch,
            "current_day": datetime.today().date(),
            "course_faculties": course_data,
            "effective_date": effective_date,
            "total_credits": total_credits_display,
            "code_map": code_map,
        }

        return render(request, "student_management/student/student_timetable_view.html", context)

    except Exception as e:
        messages.error(request, f"Error fetching timetable: {str(e)}", extra_tags="danger")
        return render(
            request,
            "student_management/student/student_timetable_view.html",
            {"allocations": [], "error_message": f"⚠️ Error fetching timetable: {str(e)}"},
        )


import io
import os
from datetime import datetime
from collections import defaultdict

from django.http import FileResponse, HttpResponse
from django.conf import settings
from django.contrib.staticfiles import finders
from django.shortcuts import render
from django.db.models import Q

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame,
    Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
)

from user_accounts.models import StudentDetails, Add_Department
from course_management.models import PeriodAllocation, AssignSubjectFaculty, CourseHours


def _safe(v):
    return "" if v is None else str(v).strip()


@check_permission("student_timetable")
@student_management
def student_timetable_pdf(request):
    """
    Generate PDF for student timetable in vertical format, same style as attendance PDF
    """
    try:
        # ----------------------------
        # Get Student Details
        # ----------------------------
        register_no = getattr(request.user, "Employee_id", None)
        student_obj = StudentDetails.objects.filter(reg_no=register_no).first()

        if not student_obj:
            return HttpResponse("Student record not found.", status=404)

        department_name = getattr(getattr(request.user, "Department", None), "Department", None)
        section = student_obj.section
        student_semester = student_obj.semester
        student_batch = student_obj.batch
        student_name = _safe(student_obj.name) or "Student"

        if not (register_no and department_name and section):
            return HttpResponse("Missing student details (section/department).", status=404)

        # ----------------------------
        # Department Lookup
        # ----------------------------
        department_obj = Add_Department.objects.filter(
            Department__iexact=department_name,
            is_active=True
        ).first()

        if not department_obj:
            return HttpResponse(f"Department not found: {department_name}", status=404)

        if not student_semester:
            return HttpResponse("Semester not assigned yet.", status=404)

        # ----------------------------
        # Fetch Timetable Records
        # ----------------------------
        allocations_qs = PeriodAllocation.objects.filter(
            department=department_obj,
            semester=student_semester,
            section=section,
        ).order_by('day')

        weekday_order = {
            "Monday": 1, "Tuesday": 2, "Wednesday": 3, "Thursday": 4,
            "Friday": 5, "Saturday": 6, "Sunday": 7,
        }
        allocations = sorted(allocations_qs, key=lambda x: weekday_order.get(getattr(x, "day", ""), 99))

        all_periods = [
            "first_period", "second_period", "third_period", "fourth_period",
            "fifth_period", "sixth_period", "seventh_period", "eighth_period",
            "nineth_period", "tenth_period",
        ]

        # Only show columns having at least one value
        visible_periods = [p for p in all_periods if any(getattr(a, p, None) for a in allocations)]

        # If no visible periods, still show at least one column for structure
        if not visible_periods:
            visible_periods = ["first_period"]

        # Format period names for display
        period_display_names = {
            "first_period": "1st",
            "second_period": "2nd",
            "third_period": "3rd",
            "fourth_period": "4th",
            "fifth_period": "5th",
            "sixth_period": "6th",
            "seventh_period": "7th",
            "eighth_period": "8th",
            "nineth_period": "9th",
            "tenth_period": "10th",
        }

        # Build period display and collect course codes
        filtered_allocations = []
        used_course_codes = set()

        for alloc in allocations:
            period_display = {}
            for p in visible_periods:
                code = (getattr(alloc, p, "") or "").strip()
                if code:
                    used_course_codes.add(code)
                period_display[p] = code
            alloc.period_display = period_display
            filtered_allocations.append(alloc)

        # ----------------------------
        # Faculty & Course Info
        # ----------------------------
        course_data = []
        total_credits = 0.0

        if used_course_codes:
            course_faculties_qs = (
                AssignSubjectFaculty.objects.filter(
                    department=department_obj,
                    batch=student_batch,
                    section=section,
                    is_active=True,
                    course__course_code__in=used_course_codes,
                )
                .select_related("course", "faculty", "skilled_faculty")
            )

            # Fetch CourseHours
            course_ids = [x.course_id for x in course_faculties_qs if x.course_id]
            hours_map = {
                ch.course_id: ch
                for ch in CourseHours.objects.filter(course_id__in=course_ids).only(
                    "course_id", "leture_npwk", "tutorial_npwk", "laboratory_npwk", "total_hours", "credits"
                )
            }

            def safe_float(v, default=0.0):
                if v is None:
                    return default
                try:
                    s = str(v).strip()
                    return float(s) if s else default
                except Exception:
                    return default

            for item in course_faculties_qs:
                course = item.course
                faculty = item.faculty
                skilled_faculty = item.skilled_faculty
                cc = (course.course_code or "").strip() if course else ""
                title = (getattr(course, "title", "") or "").strip() if course else ""

                ch = hours_map.get(item.course_id)
                credits = getattr(ch, "credits", None) if ch else None
                total_credits += safe_float(credits, 0.0)

                course_data.append({
                    "course_code": cc,
                    "course_title": title,
                    "faculty_name": _safe(getattr(faculty, "name", "")) if faculty else "Not Assigned",
                    "faculty_code": _safe(getattr(faculty, "faculty_id", "")) if faculty else "",
                    "skilled_faculty_name": _safe(getattr(skilled_faculty, "name", "")) if skilled_faculty else "-",
                    "skilled_faculty_code": _safe(getattr(skilled_faculty, "faculty_id", "")) if skilled_faculty else "",
                    "credits": credits or "N/A",
                })

        # Format total credits
        try:
            total_credits_display = str(int(total_credits)) if float(total_credits).is_integer() else f"{total_credits:.1f}"
        except Exception:
            total_credits_display = str(total_credits)

        # Effective date
        effective_date = None
        dates = [a.effective_date for a in filtered_allocations if getattr(a, "effective_date", None)]
        if dates:
            effective_date = max(dates)

        # ----------------------------
        # PDF Styles (SAME AS ATTENDANCE PDF)
        # ----------------------------
        styles = getSampleStyleSheet()

        PRIMARY_BLUE = colors.HexColor("#0f2f57")
        SECONDARY_BLUE = colors.HexColor("#1a4b8c")
        ACCENT_RED = colors.HexColor("#b91c1c")
        DARK_GRAY = colors.HexColor("#111827")
        MEDIUM_GRAY = colors.HexColor("#4b5563")
        LIGHT_GRAY = colors.HexColor("#9ca3af")
        BG_GRAY = colors.HexColor("#f8fafc")
        BORDER_GRAY = colors.HexColor("#e5e7eb")

        title_style = ParagraphStyle(
            "title_style",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=PRIMARY_BLUE,
            alignment=TA_CENTER,
            spaceAfter=6,
            fontName="Helvetica-Bold",
        )

        sub_style = ParagraphStyle(
            "sub_style",
            parent=styles["Normal"],
            fontSize=10,
            textColor=MEDIUM_GRAY,
            alignment=TA_CENTER,
            spaceAfter=10,
        )

        section_style = ParagraphStyle(
            "section_style",
            parent=styles["Heading2"],
            fontSize=12,
            textColor=PRIMARY_BLUE,
            alignment=TA_LEFT,
            spaceBefore=8,
            spaceAfter=6,
            fontName="Helvetica-Bold",
        )

        info_style = ParagraphStyle(
            "info_style",
            parent=styles["Normal"],
            fontSize=10,
            textColor=DARK_GRAY,
            alignment=TA_LEFT,
            leading=14,
            fontName="Helvetica",
        )

        table_header = ParagraphStyle(
            "table_header",
            parent=styles["Normal"],
            fontSize=9,
            textColor=colors.white,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
        )

        cell_left = ParagraphStyle(
            "cell_left",
            parent=styles["Normal"],
            fontSize=9,
            textColor=DARK_GRAY,
            alignment=TA_LEFT,
            leading=11,
            wordWrap="CJK",
        )

        cell_center = ParagraphStyle(
            "cell_center",
            parent=cell_left,
            alignment=TA_CENTER,
        )

        def make_table(data, col_widths, header_bg=SECONDARY_BLUE, zebra=True):
            t = Table(data, colWidths=col_widths, repeatRows=1)
            ts = TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), header_bg),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ])
            if zebra and len(data) > 1:
                ts.add("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY])
            t.setStyle(ts)
            return t

        # ----------------------------
        # Document setup (SAME AS ATTENDANCE PDF)
        # ----------------------------
        buffer = io.BytesIO()
        doc = BaseDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=16 * mm,
            rightMargin=16 * mm,
            topMargin=18 * mm,
            bottomMargin=18 * mm,
            title=f"Timetable - {register_no}",
            showBoundary=0
        )

        HEADER_HEIGHT = 36 * mm

        def draw_header_footer(canvas, doc_):
            canvas.saveState()
            page_w, page_h = A4
            left = doc_.leftMargin
            right = page_w - doc_.rightMargin
            center_x = (left + right) / 2
            top_y = page_h - 8 * mm

            # Logo
            logo_rel = "images/ritlogo.png"
            logo_path = finders.find(logo_rel)
            if not logo_path:
                static_root = getattr(settings, "STATIC_ROOT", "")
                if static_root:
                    cand = os.path.join(static_root, logo_rel)
                    if os.path.exists(cand):
                        logo_path = cand

            if logo_path and os.path.exists(logo_path):
                canvas.drawImage(
                    ImageReader(logo_path),
                    left, top_y - 20 * mm,
                    width=30 * mm, height=18 * mm,
                    preserveAspectRatio=True, mask="auto"
                )

            canvas.setFillColor(PRIMARY_BLUE)
            canvas.setFont("Helvetica-Bold", 14)
            canvas.drawCentredString(center_x, top_y - 6 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

            canvas.setFillColor(ACCENT_RED)
            canvas.setFont("Helvetica-Bold", 10)
            canvas.drawCentredString(center_x, top_y - 13 * mm, "An Autonomous Institution")

            canvas.setFillColor(MEDIUM_GRAY)
            canvas.setFont("Helvetica", 8.2)
            canvas.drawCentredString(center_x, top_y - 18.5 * mm, "Approved by AICTE, New Delhi")
            canvas.setFont("Helvetica", 8)
            canvas.drawCentredString(center_x, top_y - 23 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
            canvas.drawCentredString(center_x, top_y - 27.5 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

            # Footer line
            footer_y = 18 * mm
            canvas.setStrokeColor(BORDER_GRAY)
            canvas.setLineWidth(0.8)
            canvas.line(left, footer_y + 7 * mm, right, footer_y + 7 * mm)

            canvas.setFillColor(LIGHT_GRAY)
            canvas.setFont("Helvetica", 8)
            gen_time = datetime.now().strftime("%d %b %Y, %I:%M %p")
            canvas.drawString(left, footer_y, f"Generated: {gen_time}")
            canvas.drawCentredString(center_x, footer_y, f"Reg No: {register_no}")
            canvas.drawRightString(right, footer_y, f"Page {doc_.page}")

            canvas.restoreState()

        frame = Frame(
            doc.leftMargin,
            doc.bottomMargin + 6 * mm,
            doc.width,
            doc.height - HEADER_HEIGHT + 8 * mm,
            leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
            id="normal"
        )
        doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

        # ----------------------------
        # Build PDF Content
        # ----------------------------
        elements = []
        elements.append(Spacer(1, 5 * mm))

        # Title (same look)
        elements.append(Paragraph("STUDENT TIMETABLE", title_style))
        elements.append(Paragraph(f"{student_name} ({register_no})", sub_style))

        # Optional small info line (keeps attendance feel; safe)
        if effective_date:
            eff = effective_date.strftime("%d %b %Y")
            elements.append(Paragraph(f"Effective Date: {eff}", ParagraphStyle(
                "mini_info",
                parent=styles["Normal"],
                fontSize=9,
                textColor=MEDIUM_GRAY,
                alignment=TA_CENTER,
                spaceAfter=8
            )))

        # ----------------------------
        # Timetable Section
        # ----------------------------
        elements.append(Paragraph("WEEKLY TIMETABLE", section_style))

        if not filtered_allocations:
            elements.append(Paragraph("No timetable data available for the selected semester.", info_style))
        else:
            day_col_width = 22 * mm  # slightly wider like attendance tables
            period_col_width = (doc.width - day_col_width) / len(visible_periods)

            header_row = [Paragraph("Day / Period", table_header)]
            for p in visible_periods:
                display_name = period_display_names.get(p, p.replace("_", " ").title())
                header_row.append(Paragraph(display_name, table_header))

            timetable_rows = [header_row]

            # Add timetable data
            for alloc in filtered_allocations:
                day_name = getattr(alloc, "day", "—")
                row = [Paragraph(day_name, cell_center)]

                for p in visible_periods:
                    course_code = alloc.period_display.get(p, "")
                    if course_code:
                        faculty_name = ""
                        skilled_faculty_name = ""
                        for course in course_data:
                            if course["course_code"] == course_code:
                                faculty_name = course["faculty_name"]
                                skilled_faculty_name = course["skilled_faculty_name"]
                                break

                        display_text = f"<b>{course_code}</b>"
                        if faculty_name and faculty_name != "Not Assigned":
                            initials = "".join([w[0].upper() for w in faculty_name.split()[:2]])
                            display_text += f"<br/><font size=7>M: {initials}</font>"
                        if skilled_faculty_name and skilled_faculty_name not in {"-", "Not Assigned"}:
                            skilled_initials = "".join([w[0].upper() for w in skilled_faculty_name.split()[:2]])
                            display_text += f"<br/><font size=7>S: {skilled_initials}</font>"

                        row.append(Paragraph(display_text, cell_center))
                    else:
                        row.append(Paragraph("—", cell_center))

                timetable_rows.append(row)

            col_widths = [day_col_width] + [period_col_width] * len(visible_periods)
            elements.append(make_table(
                timetable_rows,
                col_widths=col_widths,
                header_bg=SECONDARY_BLUE,
                zebra=True
            ))
            elements.append(Spacer(1, 6 * mm))

        # ----------------------------
        # Course Details & Faculty Section
        # ----------------------------
        if course_data:
            elements.append(Paragraph("COURSE DETAILS & FACULTY", section_style))

            legend_rows = [[
                Paragraph("S.No", table_header),
                Paragraph("Course Code", table_header),
                Paragraph("Course Title", table_header),
                Paragraph("Main Faculty", table_header),
                Paragraph("Skilled Faculty", table_header),
                Paragraph("Credits", table_header),
            ]]

            for i, course in enumerate(sorted(course_data, key=lambda x: x["course_code"]), start=1):
                faculty_display = course["faculty_name"]
                if course["faculty_code"]:
                    faculty_display += f" ({course['faculty_code']})"
                skilled_faculty_display = course["skilled_faculty_name"]
                if course["skilled_faculty_code"]:
                    skilled_faculty_display += f" ({course['skilled_faculty_code']})"

                legend_rows.append([
                    Paragraph(str(i), cell_center),
                    Paragraph(course["course_code"] or "—", cell_center),
                    Paragraph(course["course_title"] or "—", cell_left),
                    Paragraph(faculty_display or "—", cell_left),
                    Paragraph(skilled_faculty_display or "-", cell_left),
                    Paragraph(str(course["credits"]), cell_center),
                ])

            # total credits row
            legend_rows.append([
                Paragraph("<b>Total</b>", cell_center),
                Paragraph("-", cell_center),
                Paragraph("-", cell_center),
                Paragraph("-", cell_center),
                Paragraph("-", cell_center),
                Paragraph(f"<b>{total_credits_display}</b>", cell_center),
            ])

            # widths tuned to attendance feel
            elements.append(make_table(
                legend_rows,
                col_widths=[
                    10 * mm,
                    24 * mm,
                    doc.width - (10 + 24 + 34 + 34 + 18) * mm,
                    34 * mm,
                    34 * mm,
                    18 * mm,
                ],
                header_bg=SECONDARY_BLUE,
                zebra=True
            ))

        # ----------------------------
        # Build PDF
        # ----------------------------
        try:
            doc.build(elements)
        except Exception as e:
            return HttpResponse(f"PDF generation failed: {e}", status=500)

        # ----------------------------
        # Return PDF Response
        # ----------------------------
        buffer.seek(0)
        filename = f"Timetable_{register_no}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        return FileResponse(buffer, as_attachment=True, filename=filename)

    except Exception as e:
        return HttpResponse(f"Error generating timetable PDF: {str(e)}", status=500)


from django.shortcuts import render, get_object_or_404
from course_management.models import Course, CourseHours
from user_accounts.models import USER, Department
from student_management.models import Subject_Assignment
from fee_management.decorators import fee_management

from student_management.models import FeeReceipt


@check_permission("fee_receipt_upload")
@fee_management
def fee_receipt_upload(request):
    # Simulate current logged-in student (replace with actual session user if integrated)
    # Example: current_reg_no = request.user.studentdetails.reg_no
    current_reg_no = (
        request.user.Employee_id
    )  # For now, we can test using ?reg_no=<number> in URL

    student = None
    if current_reg_no:
        try:
            student = StudentDetails.objects.select_related("department").get(
                reg_no=current_reg_no
            )
        except StudentDetails.DoesNotExist:
            messages.error(request, "No student found with this register number.")

    if request.method == "POST":
        if not student:
            messages.error(request, "Student details not found.")
            return redirect("fee_receipt_upload")

        receipt_id = (request.POST.get("receipt_id") or "").strip()

        if receipt_id:
            # ----- Replace the file on an existing PENDING receipt -----
            receipt = FeeReceipt.objects.filter(id=receipt_id, student=student).first()

            if not receipt:
                messages.error(request, "Receipt not found.")
                return redirect("fee_receipt_upload")

            if receipt.status != FeeReceipt.Status.PENDING:
                messages.error(request, "This receipt has already been verified and can no longer be edited.")
                return redirect("fee_receipt_upload")

            new_file = request.FILES.get("fee_receipt")
            if not new_file:
                messages.error(request, "Please choose a replacement file before submitting.")
                return redirect("fee_receipt_upload")

            receipt.fee_receipt.delete(save=False)
            receipt.fee_receipt = new_file
            receipt.save(update_fields=["fee_receipt"])

            messages.success(request, "Receipt updated successfully!")
            return redirect("fee_receipt_upload")

        fee_file = request.FILES.get("fee_receipt")

        if fee_file:
            # Create FeeReceipt record using student details
            FeeReceipt.objects.create(
                student=student,
                department=student.department,
                batch=student.batch,
                section=student.section,
                semester=student.semester,
                fee_receipt=fee_file,
            )

            messages.success(
                request, f"Receipt for {student.name} uploaded successfully!"
            )
            return redirect("fee_receipt_upload")
        else:
            messages.error(request, "Please upload a receipt file before submitting.")

    receipts = (
        FeeReceipt.objects.filter(student=student)
        .select_related("fee_entry")
        .order_by("-uploaded_at")
        if student else FeeReceipt.objects.none()
    )

    return render(
        request,
        "student_management/fee/fee_receipt_upload.html",
        {"student": student, "receipts": receipts},
    )


from decimal import Decimal, InvalidOperation
from datetime import date

from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.db import IntegrityError

from user_accounts.decorators import check_permission
from faculty_management.models import general_information
from student_management.models import FeeReceipt, ManualFeeEntry


def get_academic_year():
    today = date.today()
    y = today.year
    if today.month >= 6:
        return f"{y}-{y+1}"
    return f"{y-1}-{y}"


@check_permission("manual_fee_entry")
def manual_fee_entry(request):
    employee_id = request.user.Employee_id

    try:
        staff = general_information.objects.get(faculty_id=employee_id)
        department = staff.department
    except general_information.DoesNotExist:
        messages.error(request, "Employee not found.")
        return redirect("home")

    academic_year = get_academic_year()

    pending_receipts = (
        FeeReceipt.objects.filter(status="PENDING", department=department)
        .exclude(id__in=ManualFeeEntry.objects.values_list("fee_receipt_id", flat=True))
        .select_related("student")
        .order_by("-id")
    )

    verified_entries = (
        ManualFeeEntry.objects.select_related("fee_receipt", "fee_receipt__student")
        .filter(fee_receipt__department=department)
        .order_by("-updated_at")
    )

    # ================= AJAX =================
    if request.headers.get("x-requested-with") == "XMLHttpRequest" and request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        try:
            if action == "save_new":
                receipt_id = (request.POST.get("receipt_id") or "").strip()
                fee_value = (request.POST.get("fee_value") or "").strip()
                txn_id = (request.POST.get("transaction_id") or "").strip()

                if not receipt_id:
                    return JsonResponse({"ok": False, "error": "Receipt ID missing."})

                receipt = FeeReceipt.objects.select_related("student").get(id=receipt_id)

                if receipt.department != department:
                    return JsonResponse({"ok": False, "error": "Unauthorized receipt access."})

                if not fee_value:
                    return JsonResponse({"ok": False, "error": "Enter fee amount."})
                try:
                    fee_dec = Decimal(fee_value)
                    if fee_dec < 0:
                        return JsonResponse({"ok": False, "error": "Fee cannot be negative."})
                except (InvalidOperation, ValueError):
                    return JsonResponse({"ok": False, "error": "Invalid fee amount."})

                if not txn_id:
                    return JsonResponse({"ok": False, "error": "Transaction ID is required."})

                if ManualFeeEntry.objects.filter(transaction_id__iexact=txn_id).exists():
                    return JsonResponse({"ok": False, "error": f"Transaction ID '{txn_id}' already exists."})

                try:
                    entry = ManualFeeEntry.objects.create(
                        fee_receipt=receipt,
                        entered_fee=fee_dec,
                        transaction_id=txn_id,
                        entered_by=str(staff),
                    )
                except IntegrityError:
                    return JsonResponse({"ok": False, "error": f"Transaction ID '{txn_id}' already exists."})

                receipt.status = FeeReceipt.Status.VERIFIED if hasattr(FeeReceipt, "Status") else "VERIFIED"
                receipt.save()

                return JsonResponse({
                    "ok": True,
                    "message": "Verified successfully.",
                    "entry": {
                        "id": entry.id,
                        "entered_fee": str(entry.entered_fee),
                        "transaction_id": entry.transaction_id,
                        "updated_at": entry.updated_at.strftime("%d %b %Y, %I:%M %p"),
                    }
                })

            elif action == "edit_existing":
                entry_id = (request.POST.get("entry_id") or "").strip()
                fee_value = (request.POST.get("fee_value") or "").strip()
                new_txn = (request.POST.get("transaction_id") or "").strip()

                if not entry_id:
                    return JsonResponse({"ok": False, "error": "Entry ID missing."})

                entry = ManualFeeEntry.objects.select_related("fee_receipt", "fee_receipt__student").get(id=entry_id)

                if entry.fee_receipt.department != department:
                    return JsonResponse({"ok": False, "error": "Unauthorized entry access."})

                if not fee_value:
                    return JsonResponse({"ok": False, "error": "Enter fee amount."})
                try:
                    fee_dec = Decimal(fee_value)
                    if fee_dec < 0:
                        return JsonResponse({"ok": False, "error": "Fee cannot be negative."})
                except (InvalidOperation, ValueError):
                    return JsonResponse({"ok": False, "error": "Invalid fee amount."})

                if not new_txn:
                    return JsonResponse({"ok": False, "error": "Transaction ID is required."})

                if ManualFeeEntry.objects.exclude(id=entry.id).filter(transaction_id__iexact=new_txn).exists():
                    return JsonResponse({"ok": False, "error": f"Transaction ID '{new_txn}' already exists."})

                entry.entered_fee = fee_dec
                entry.transaction_id = new_txn
                entry.entered_by = str(staff)
                entry.save()

                return JsonResponse({
                    "ok": True,
                    "message": "Updated successfully.",
                    "entry": {
                        "id": entry.id,
                        "entered_fee": str(entry.entered_fee),
                        "transaction_id": entry.transaction_id,
                        "updated_at": entry.updated_at.strftime("%d %b %Y, %I:%M %p"),
                    }
                })

            return JsonResponse({"ok": False, "error": "Invalid action."})

        except Exception as e:
            return JsonResponse({"ok": False, "error": str(e)})

    return render(
        request,
        "student_management/fee/manual_fee_entry.html",
        {
            "department": department,
            "academic_year": academic_year,
            "pending_receipts": pending_receipts,
            "verified_entries": verified_entries,
        },
    )


from user_accounts.models import *


from django.shortcuts import render


@student_management
def student_admission_details(request):
    # ==== DEBUG: entry ====
    # # print("[ADM VIEW] Enter student_admission_details")

    # 1) Register number (robust)
    register_no = getattr(request.user, "Employee_id", None)
    # # print(f"[ADM VIEW] register_no: {register_no!r}")
    # print("reg no => ", register_no)
    # 2) Department (resolve to string + keep obj for legacy use)
    user_department_name = getattr(request.user, "Department", None)
    # # print(f"[ADM VIEW] raw user.Department: {user_department_name!r}")

    department_obj = None
    department_name = None
    if user_department_name:
        dept_val = getattr(user_department_name, "Department", None) or str(
            user_department_name
        )
        # # print(f"[ADM VIEW] dept_val (normalized): {dept_val!r}")
        department_obj = Add_Department.objects.filter(
            Department__iexact=dept_val
        ).first()
        department_name = getattr(department_obj, "Department", None) or dept_val
        # # print(f"[ADM VIEW] department_obj: {department_obj!r}, department_name: {department_name!r}")

    if not register_no:
        # # print("[ADM VIEW] No register number; returning early.")
        return render(
            request,
            "student_management/student/student_admission_details.html",
            {
                "error": "Register number not found in session.",
                "reg_no": None,
                "department": department_name,
                "batch": None,
                "admission_data": None,
                "sslc_data": None,
                "hsc_data": None,
                "dip_data": None,
                "transport_data": None,
                "department_obj": None,
            },
        )

    try:
        # 3) Student core record
        student = (
            StudentDetails.objects.filter(reg_no=register_no)
            .select_related("department")
            .first()
        )
        # # print(f"[ADM VIEW] student: {student!r}")
        if not student:
            # # print("[ADM VIEW] Student not found; returning.")
            return render(
                request,
                "student_management/student/student_admission_details.html",
                {
                    "error": "Student record not found.",
                    "reg_no": register_no,
                    "department": department_name,
                    "batch": None,
                    "admission_data": None,
                    "sslc_data": None,
                    "hsc_data": None,
                    "dip_data": None,
                    "transport_data": None,
                    "department_obj": department_obj,
                },
            )

        # Prefer department from student if present
        if student.department and getattr(student.department, "Department", None):
            department_name = student.department.Department
            # # print(f"[ADM VIEW] department from student.department: {department_name!r}")

        context = {
            "reg_no": register_no,
            "department": department_name,  # string for template
            "batch": student.batch,
            "admission_data": None,
            "sslc_data": None,
            "hsc_data": None,
            "dip_data": None,
            "transport_data": None,
            "department_obj": department_obj,  # keep around if template uses it elsewhere
        }

        # 4) PersonalDetails — prioritize registration_no (name match is brittle)
        personal = (
            PersonalDetails.objects.using("admissionform1")
            .filter(registration_no=register_no)
            .first()
        )
        # # print(f"[ADM VIEW] personal by registration_no: {personal!r}")

        if not personal:
            personal = (
                PersonalDetails.objects.using("admissionform1")
                .filter(registration_no=register_no, name__iexact=student.name)
                .first()
            )
            # # print(f"[ADM VIEW] personal by reg+name fallback: {personal!r}")

        if not personal:
            # # print("[ADM VIEW] No PersonalDetails; return with basic context.")
            return render(
                request,
                "student_management/student/student_admission_details.html",
                context,
            )

        # 5) AdmissionRecords linked to PersonalDetails
        admission_record = (
            AdmissionRecords.objects.using("admissionform1")
            .filter(PersonalDetailsId=personal)
            .first()
        )
        # # print(f"[ADM VIEW] admission_record: {admission_record!r}")

        if not admission_record:
            # # print("[ADM VIEW] No AdmissionRecords; return with partial context.")
            return render(
                request,
                "student_management/student/student_admission_details.html",
                context,
            )

        # 6) Build admission_data
        context["admission_data"] = {
            "Name": student.name,
            "registration_no": register_no,
            "Date_of_Birth": personal.date_of_birth,
            "Gender": personal.gender,
            "Nationality": personal.nationality,
            "Aadhaar_Number": personal.Aadhaar_Number,
            "Self_Mobile_Number": personal.personal_mobile_no,
            "Self_Email_ID": personal.personal_email_id,
            "Guardian_name": personal.guardian_name,
            "Guardian_Mobile_Number": personal.guardian_mobile_no,
            "Father_name": personal.father_name,
            "Father_Mobile_Number": personal.father_mobile_no,
            "Mother_name": personal.mother_name,
            "Mother_Mobile_Number": personal.mother_mobile_no,
            "Permanent_Address_Line1": f"{personal.Permanent_Address_Door_No} {personal.Permanent_Address_Street_Name} {personal.Permanent_Address_Location} {personal.Permanent_Address_Taluk}".strip(),
            "Permanent_Address_Line2": None,
            "Permanent_Address_City_Town": personal.Permanent_Address_Location,
            "Permanent_Address_District": personal.Permanent_Address_District,
            "Permanent_Address_State": personal.Permanent_Address_State,
            "Permanent_Address_Pincode": personal.Permanent_Address_Pincode,
            "Communication_Address_Line1": f"{personal.Communication_Address_Door_No} {personal.Communication_Address_Street_Name} {personal.Communication_Address_Location} {personal.Communication_Address_Taluk}".strip(),
            "Communication_Address_Line2": None,
            "Communication_Address_City_Town": personal.Communication_Address_Location,
            "Communication_Address_District": personal.Communication_Address_District,
            "Communication_Address_State": personal.Communication_Address_State,
            "Communication_Address_Pincode": personal.Communication_Address_Pincode,
            "admission_no": admission_record.admissionNo,
            "admissionFor": admission_record.admissionFor,
            "Quota": admission_record.Quota,
            "Department": admission_record.Department,
            "Mode": admission_record.Mode,
            "academic_Category": admission_record.academic_Category,
            "certificate_status": admission_record.certificate_status,
            "certification_validation_date": admission_record.certification_valiation_date,
            "group_code": admission_record.group_code,
            "round": admission_record.round,
            "caste": personal.caste,
            "mother_tounge": personal.mother_tounge,
            "religion": personal.religion,
            "EMIS_ID": personal.EMIS_ID,
        }
        # # print("[ADM VIEW] admission_data populated.")

        # 7) SchoolDetailsId is integer on AdmissionRecords
        school_id = admission_record.SchoolDetailsId
        # # print(f"[ADM VIEW] school_id (int): {school_id!r}")

        # 8) Academic records using correct FK names per model
        sslc_qs = SSLCDetails.objects.using("admissionform1").filter(
            SchoolDetailsId=school_id
        )
        hsc_qs = HSCDetails.objects.using("admissionform1").filter(
            school_details_id=school_id
        )
        dip_qs = DiplomoDetails.objects.using("admissionform1").filter(
            school_details_id=school_id
        )

        # # print(f"[ADM VIEW] sslc count: {sslc_qs.count()}, hsc count: {hsc_qs.count()}, dip count: {dip_qs.count()}")

        context["sslc_data"] = sslc_qs.first()
        context["hsc_data"] = hsc_qs.first()
        context["dip_data"] = dip_qs.first()

        # 9) Transport: direct FK on AdmissionRecords
        try:
            transport_obj = (
                admission_record.TransportDetailsId
            )  # FK object (or None if nullable)
            # # print(f"[ADM VIEW] transport via AdmissionRecords.TransportDetailsId: {transport_obj!r}")
        except Exception as _:
            transport_obj = None
            # # print("[ADM VIEW] transport FK access failed; set None.")

        # If you **also** want to try legacy path (TransportDetails linked to PersonalDetails):
        if not transport_obj:
            legacy_transport = (
                TransportDetails.objects.using("admissionform1")
                .filter(admission_records_id=personal)
                .first()
            )
            # # print(f"[ADM VIEW] legacy transport by personal FK: {legacy_transport!r}")
            transport_obj = legacy_transport

        context["transport_data"] = transport_obj

        # # print("[ADM VIEW] Rendering template with populated context.")
        return render(
            request,
            "student_management/student/student_admission_details.html",
            context,
        )

    except Exception as e:
        import traceback

        # # print("[ADM VIEW] EXCEPTION:", repr(e))
        traceback.print_exc()
        return render(
            request,
            "student_management/student/student_admission_details.html",
            {
                "error": "Unexpected error: " + str(e),
                "reg_no": register_no,
                "department": department_name,
                "batch": None,
                "admission_data": None,
                "sslc_data": None,
                "hsc_data": None,
                "dip_data": None,
                "transport_data": None,
                "department_obj": department_obj,
            },
        )


def edit_address(request):
    registration_no = getattr(
        request.user, "Employee_id", None
    )  # map to admissionNo in DB

    # Use correct field name for lookup
    admission_data = get_object_or_404(
        PersonalDetails.objects.using("admissionform1"), registration_no=registration_no
    )

    if request.method == "POST":
        # Use .update() to directly update the row
        PersonalDetails.objects.using("admissionform1").filter(
            registration_no=registration_no
        ).update(
            Communication_Address_Street_Name=request.POST.get(
                "Communication_Address_Street_Name", ""
            ),
            Communication_Address_District=request.POST.get(
                "Communication_Address_District", ""
            ),
            Communication_Address_State=request.POST.get(
                "Communication_Address_State", ""
            ),
            Communication_Address_Pincode=request.POST.get(
                "Communication_Address_Pincode", ""
            ),
        )
        messages.success(request, "Communication Address updated successfully.")
        return redirect("edit_address")

    return render(
        request,
        "student_management/student/Edit_details.html",
        {"admission_data": admission_data},
    )


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from datetime import datetime
from user_accounts.models import StudentDetails


def student_academic_calendar(request):
    student = request.user  # Logged-in student

    # Get StudentDetails using the user's Employee_id (assuming reg_no matches Employee_id)
    try:
        student_details = StudentDetails.objects.get(reg_no=student.Employee_id)
    except StudentDetails.DoesNotExist:
        student_details = None

    batch = None
    semester = None

    if student_details:
        batch = student_details.batch
        semester = student_details.semester

    # Fetch academic calendar only if batch is available
    calendar = AcademicCalendar.objects.none()
    if batch:
        calendar = AcademicCalendar.objects.filter(batch=batch, semester=semester)
        if semester:
            calendar = calendar.filter(semester=semester)

    # print("calendar => ", calendar)

    context = {
        "calendars": calendar,
        "student_info": student_details,
        "student": student,
    }

    return render(
        request, "student_management/student/student_academic_calendar.html", context
    )


# ...existing code...

from datetime import date, datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from course_management.models import AssignSubjectFaculty, PeriodAllocation
from user_accounts.models import USER
from student_management.models import Student_cgpa, Daily_Attendance, HourAttendance
from django.db.models import Q
import calendar
from django.utils.dateparse import parse_date
from django.shortcuts import render, redirect

# @check_permission("hour_attendence")
# @student_management
# def hour_attendence(request):
#     faculty_id = getattr(request.user, "id", None)  # safer way
#     if not faculty_id:
#         messages.error(request, "No faculty logged in.")
#         return redirect("some_error_page")  # handle appropriately

#     # Get all subject assignments for this faculty
#     assigned_subjects = AssignSubjectFaculty.objects.filter(faculty_id=faculty_id)
#     # print("assigned_subjects => ", assigned_subjects)
#     # Get all courses assigned to this faculty
#     assigned_course_ids = assigned_subjects.values_list('course_id', flat=True)
#     courses = Course.objects.filter(id__in=assigned_course_ids)
#     # print("courses => ", courses)
#     return render(request, "faculty_management/hour_attendence.html", {
#         "courses": courses,
#         "assigned_subjects": assigned_subjects,
#     })

import calendar
import math
from datetime import date

from django.contrib import messages
from django.db.models import Case, When, IntegerField, Value
from django.db.models.functions import Cast, Mod
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date

@check_permission("hour_attendence")
@student_management
def hour_attendence(request):
    faculty_id = getattr(request.user, "Employee_id", None)
    if not faculty_id:
        messages.error(request, "No faculty logged in.")
        return redirect("home")

    try:
        faculty = general_information.objects.get(faculty_id=faculty_id)
    except general_information.DoesNotExist:
        messages.error(request, "Faculty not found.")
        return redirect("home")

    assigned_subjects = (
        AssignSubjectFaculty.objects
        .filter(faculty=faculty, is_active=True)
        .select_related("course", "course__department", "course__regulation", "course__elective")
        .annotate(
            semester_int=Cast("course__semester", IntegerField()),
            semester_mod=Mod(Cast("course__semester", IntegerField()), Value(2)),
            semester_order=Case(
                When(semester_mod=0, then=Value(0)),   # EVEN first
                default=Value(1),                      # ODD next
                output_field=IntegerField(),
            )
        )
        .order_by("-academic_year", "semester_order", "course__semester", "batch", "section")
    )

    return render(
        request,
        "faculty_management/hour_attendence.html",
        {"assigned_subjects": assigned_subjects},
    )
 

def get_academic_year():
    today = date.today()
    if today.month >= 6:
        return f"{today.year}-{today.year + 1}"
    return f"{today.year - 1}-{today.year}"


def split_students_register_wise(student_list, split_count):
    """
    Split students in register-number order into balanced groups.
    Example:
      60 students, split_count=2 -> 30 + 30
      61 students, split_count=2 -> 31 + 30
      60 students, split_count=4 -> 15 + 15 + 15 + 15
    """
    total = len(student_list)
    if total == 0:
        return []

    split_count = max(1, int(split_count))
    chunk_size = math.ceil(total / split_count)

    groups = []
    start = 0
    for i in range(split_count):
        end = start + chunk_size
        groups.append(student_list[start:end])
        start = end

    # remove empty groups
    groups = [g for g in groups if g]
    return groups


@check_permission("hour_attendence")
@student_management
def course_attendance_detail(request):
    # ============================================================
    # GET PARAMS
    # ============================================================
    course_code = (request.GET.get("course_code") or "").strip()
    title = (request.GET.get("title") or "").strip()
    department_id = (request.GET.get("department") or "").strip()
    year = (request.GET.get("year") or "").strip()
    semester = (request.GET.get("semester") or "").strip()
    batch = (request.GET.get("batch") or "").strip()
    section = (request.GET.get("section") or "").strip()
    regulation = (request.GET.get("regulation") or "").strip()
    elective = (request.GET.get("elective") or "").strip()

    # ============================================================
    # FACULTY
    # ============================================================
    faculty_emp_id = getattr(request.user, "Employee_id", None)
    faculty = general_information.objects.filter(faculty_id=faculty_emp_id).first()
    if not faculty:
        messages.error(request, "Faculty not found.")
        return redirect("home")

    if not department_id:
        messages.error(request, "Department is missing.")
        return redirect("hour_attendence")

    department = get_object_or_404(Add_Department, id=department_id)

    # ============================================================
    # DATE
    # ============================================================
    selected_date = request.GET.get("date")
    selected_date = parse_date(selected_date) if selected_date else date.today()
    if not selected_date:
        selected_date = date.today()
    selected_date = min(selected_date, date.today())

    weekday_name = calendar.day_name[selected_date.weekday()]
    academic_year = get_academic_year()

    # ============================================================
    # PERIOD
    # ============================================================
    all_periods = list(range(1, 11))
    selected_period = request.GET.get("period")

    try:
        selected_period = int(selected_period) if selected_period else None
    except (TypeError, ValueError):
        selected_period = None

    if selected_period not in all_periods:
        selected_period = None

    # ============================================================
    # ATTENDANCE SPLIT TYPE
    # 1 = Full class
    # 2 = Half class
    # 3 = One-third class
    # 4 = One-fourth class
    # ============================================================
    split_type = request.GET.get("split_type", "1")
    try:
        split_type = int(split_type)
    except (TypeError, ValueError):
        split_type = 1

    if split_type not in [1, 2, 3, 4]:
        split_type = 1

    selected_group_no = request.GET.get("group_no", "1")
    try:
        selected_group_no = int(selected_group_no)
    except (TypeError, ValueError):
        selected_group_no = 1

    # ============================================================
    # FETCH ENROLLED STUDENTS
    # ============================================================
    enrollment_filter = {
        "course__course_code": course_code,
        "batch": batch,
        "regulation_id": regulation,
        "enroll": True,
        "section": section,
    }

    is_open_elective = CourseEnrollment.objects.filter(
        course__course_code=course_code,
        is_open_elective=True,
        enroll=True,
    ).exists()

    if not is_open_elective:
        enrollment_filter.update({
            "department_id": department_id,
            "section": section,
        })

    enrolled_student_ids = (
        CourseEnrollment.objects.filter(**enrollment_filter)
        .values_list("student_id", flat=True)
    )

    students_qs = StudentDetails.objects.filter(
        id__in=enrolled_student_ids, is_active=True
    ).order_by("reg_no", "name")

    student_list = list(students_qs)

    # ============================================================
    # COURSE OBJECT
    # ============================================================
    course_obj = Course.objects.filter(
        course_code=course_code,
        department=department,
        regulation=regulation,
    ).first()

    if not course_obj:
        messages.error(request, "Course not found.")
        return redirect("hour_attendence")

    # ============================================================
    # COURSE PLANS (only those created by THIS faculty for THIS course)
    # Matched by faculty_id + course code (the stored Course PK can differ
    # by department/regulation between the assignment row and the plan row).
    # Scoped to the correct academic year for this course.
    # ============================================================
    selected_unit = (request.GET.get("unit_module_no") or "").strip()

    # All course plans for THIS course (by course code)
    plan_base = CoursePlan.objects.filter(course__course_code=course_code)

    # Prefer plans created by THIS faculty; if this faculty has none for the
    # course, fall back to every plan of the course so the units still show.
    if plan_base.filter(faculty=faculty).exists():
        plan_base = plan_base.filter(faculty=faculty)

    # Academic years that actually have plans for this course
    available_plan_years = list(
        plan_base.exclude(academic_year__isnull=True)
        .exclude(academic_year__exact="")
        .values_list("academic_year", flat=True)
        .distinct()
        .order_by("-academic_year")
    )

    # Use the current academic year if it has plans, otherwise fall back to
    # the latest academic year that actually has plans for this course.
    if academic_year in available_plan_years:
        plan_academic_year = academic_year
    elif available_plan_years:
        plan_academic_year = available_plan_years[0]
    else:
        plan_academic_year = academic_year

    course_plan_base = plan_base.filter(academic_year=plan_academic_year)

    # Distinct units available for THIS course in THIS academic year (with count)
    unit_counts = {}
    for um in (
        course_plan_base
        .exclude(unit_module_no__isnull=True)
        .exclude(unit_module_no__exact="")
        .values_list("unit_module_no", flat=True)
    ):
        unit_counts[um] = unit_counts.get(um, 0) + 1
    unit_options = [
        {"unit_module_no": u, "plan_count": c}
        for u, c in sorted(unit_counts.items())
    ]

    # Apply the unit filter (if a specific unit is selected via the filter)
    course_plans = course_plan_base
    if selected_unit:
        course_plans = course_plans.filter(unit_module_no=selected_unit)
    course_plans = course_plans.order_by("unit_module_no", "period_no", "id")

    # ============================================================
    # SPLIT STUDENTS REG-WISE
    # ============================================================
    grouped_students = split_students_register_wise(student_list, split_type)

    total_groups = len(grouped_students) if grouped_students else 1

    if selected_group_no < 1:
        selected_group_no = 1
    if selected_group_no > total_groups:
        selected_group_no = total_groups

    current_group_students = grouped_students[selected_group_no - 1] if grouped_students else []
    current_group_ids = [s.id for s in current_group_students]

    # ============================================================
    # EXISTING HOUR ATTENDANCE
    # ============================================================
    existing_map = {}
    existing_all_student_ids = set()
    selected_course_plan_id = None

    if selected_period:
        existing_qs = HourAttendance.objects.filter(
            course=course_obj,
            date=selected_date,
            period=selected_period,
            student_id__in=[s.id for s in student_list],
        ).values("student_id", "status", "course_plan")

        existing_map = {row["student_id"]: row["status"] for row in existing_qs}
        existing_all_student_ids = set(existing_map.keys())

        # Pre-select the course plan already saved for this period (if any)
        for row in existing_qs:
            if row.get("course_plan"):
                selected_course_plan_id = row["course_plan"]
                break

    # ============================================================
    # DAILY ATTENDANCE (PREFILL ONLY)
    # ============================================================
    daily_qs = Daily_Attendance.objects.filter(
        date=selected_date,
        student_id__in=[s.id for s in student_list],
        academic_year=academic_year,
        year=year,
        semester=semester,
        section=section,
    ).values("student_id", "morning_status", "afternoon_status", "full_day_status")

    daily_map = {}
    for d in daily_qs:
        full = (d.get("full_day_status") or "").strip()
        ms = (d.get("morning_status") or "").strip()
        af = (d.get("afternoon_status") or "").strip()

        if full in ["Present", "Absent", "On Duty"]:
            final_status = full
        elif ms and af:
            if ms == af:
                final_status = ms
            elif ms == "On Duty" or af == "On Duty":
                final_status = "On Duty"
            else:
                final_status = ""
        else:
            final_status = ms or af or ""

        daily_map[d["student_id"]] = final_status

    # ============================================================
    # BUILD CURRENT GROUP DATA
    # ============================================================
    students_with_attendance = []

    for student in current_group_students:
        hour_status = existing_map.get(student.id)
        daily_status = daily_map.get(student.id, "")

        if hour_status:
            final_status = hour_status
            locked = True
            source = "HOUR"
        else:
            final_status = daily_status or "Present"
            locked = False
            source = "DAILY" if daily_status else "DEFAULT"

        students_with_attendance.append({
            "student": student,
            "status": final_status,
            "locked": locked,
            "source": source,
        })

    # ============================================================
    # GROUP SUMMARY
    # ============================================================
    group_summary = []
    for idx, group in enumerate(grouped_students, start=1):
        group_ids = [s.id for s in group]
        already_marked_count = len([sid for sid in group_ids if sid in existing_all_student_ids])

        group_summary.append({
            "group_no": idx,
            "count": len(group),
            "start_reg": group[0].reg_no if group else "",
            "end_reg": group[-1].reg_no if group else "",
            "already_marked_count": already_marked_count,
            "is_completed": already_marked_count == len(group) and len(group) > 0,
            "is_selected": idx == selected_group_no,
        })

    total_students = len(student_list)
    marked_total = len(existing_all_student_ids)
    remaining_total = total_students - marked_total

    # ============================================================
    # SAVE
    # ============================================================
    if request.method == "POST":
        post_period = request.POST.get("selected_period")
        post_date = request.POST.get("selected_date")
        post_split_type = request.POST.get("split_type", "1")
        post_group_no = request.POST.get("group_no", "1")

        try:
            post_period = int(post_period)
        except (TypeError, ValueError):
            post_period = None

        try:
            post_split_type = int(post_split_type)
        except (TypeError, ValueError):
            post_split_type = 1

        try:
            post_group_no = int(post_group_no)
        except (TypeError, ValueError):
            post_group_no = 1

        parsed_post_date = parse_date(post_date) if post_date else selected_date
        if not parsed_post_date:
            parsed_post_date = selected_date

        if not post_period or post_period not in all_periods:
            messages.error(request, "Please select a valid period.")
            return redirect(request.get_full_path())

        if post_split_type not in [1, 2, 3, 4]:
            post_split_type = 1

        # ----- Course Plan (compulsory) -----
        post_course_plan_id = request.POST.get("course_plan")
        if not post_course_plan_id:
            messages.error(request, "Please select a course plan before saving attendance.")
            return redirect(request.get_full_path())

        course_plan_obj = CoursePlan.objects.filter(
            id=post_course_plan_id,
            course__course_code=course_code,
        ).first()

        if not course_plan_obj:
            messages.error(request, "Invalid course plan selected.")
            return redirect(request.get_full_path())

        created_count = 0
        updated_count = 0

        for entry in students_with_attendance:
            student = entry["student"]

            field_name = f"attendance_{student.id}"
            status = request.POST.get(field_name)

            if status not in ["Present", "Absent", "On Duty"]:
                continue

            existing_attendance = HourAttendance.objects.filter(
                student=student,
                course=course_obj,
                date=parsed_post_date,
                period=post_period,
            ).first()

            if existing_attendance:
                existing_attendance.status = status
                existing_attendance.faculty = faculty
                existing_attendance.department = department
                existing_attendance.batch = batch
                existing_attendance.section = section
                existing_attendance.academic_year = academic_year
                existing_attendance.semester = semester
                existing_attendance.year = year
                existing_attendance.course_plan = course_plan_obj
                existing_attendance.marked_at = timezone.now()
                existing_attendance.save(
                    update_fields=[
                        "status",
                        "faculty",
                        "department",
                        "batch",
                        "section",
                        "academic_year",
                        "semester",
                        "year",
                        "course_plan",
                        "marked_at",
                    ]
                )
                updated_count += 1
            else:
                HourAttendance.objects.create(
                    student=student,
                    faculty=faculty,
                    department=department,
                    batch=batch,
                    section=section,
                    academic_year=academic_year,
                    semester=semester,
                    year=year,
                    course=course_obj,
                    course_plan=course_plan_obj,
                    period=post_period,
                    date=parsed_post_date,
                    status=status,
                    marked_at=timezone.now(),
                )
                created_count += 1

        messages.success(
            request,
            f"Attendance saved successfully for Group {post_group_no}. "
            f"({created_count} records created, {updated_count} records updated)"
        )

        return redirect(
            f"{request.path}"
            f"?course_code={course_code}"
            f"&title={title}"
            f"&department={department_id}"
            f"&year={year}"
            f"&semester={semester}"
            f"&batch={batch}"
            f"&section={section}"
            f"&regulation={regulation}"
            f"&elective={elective}"
            f"&date={parsed_post_date.strftime('%Y-%m-%d')}"
            f"&period={post_period}"
            f"&split_type={post_split_type}"
            f"&group_no={post_group_no}"
        )

    return render(
        request,
        "course_management/hour_attendance/course_attendance_detail.html",
        {
            "course_code": course_code,
            "title": title,
            "department_name": department.Department if department else "",
            "department_id": department_id,
            "year": year,
            "semester": semester,
            "batch": batch,
            "section": section,
            "regulation": regulation,
            "elective": elective,
            "students_with_attendance": students_with_attendance,
            "selected_date": selected_date,
            "weekday_name": weekday_name,
            "academic_year": academic_year,
            "today": date.today(),
            "all_periods": all_periods,
            "selected_period": selected_period,
            "split_type": split_type,
            "selected_group_no": selected_group_no,
            "group_summary": group_summary,
            "total_students": total_students,
            "marked_total": marked_total,
            "remaining_total": remaining_total,
            "total_groups": total_groups,
            "course_plans": course_plans,
            "selected_course_plan_id": selected_course_plan_id,
            "unit_options": unit_options,
            "selected_unit": selected_unit,
            "plan_academic_year": plan_academic_year,
        },
    )


# ============================================================
# COURSE PLAN COMPLETION  /  RECORD OF CLASS WORK
# ------------------------------------------------------------
# For a faculty's course + the SECTION they are assigned to, the
# course plan is laid out as a physical "Record of Class Work"
# register:
#     Date | Period | Cumulative Hours | Topics Covered |
#     Instructional Methods | Initial
# The plan structure (units, topics, methods, cumulative hours)
# comes from CoursePlan; the actual delivery Date + class Period
# for each topic comes from the hour-attendance table
# (student_management_hourattendance) for that exact section.
# ============================================================
def _faculty_course_sections(faculty, course_code):
    """
    Distinct sections (batch / section / academic year) the faculty is
    assigned to for this course code — used to populate the Section filter.
    Source table: AssignSubjectFaculty.
    """
    sections = []
    seen = set()
    if not course_code:
        return sections

    qs = (
        AssignSubjectFaculty.objects
        .filter(faculty=faculty, is_active=True, course__course_code=course_code)
        .order_by("-academic_year", "batch", "section")
    )
    for a in qs:
        batch = (a.batch or "").strip()
        section = (a.section or "").strip()
        ay = (a.academic_year or "").strip()
        key = (batch, section, ay)
        if key in seen:
            continue
        seen.add(key)

        label = f"{batch} / {section}".strip(" /")
        if ay:
            label = f"{label}  ({ay})" if label else ay
        sections.append({
            "batch": batch,
            "section": section,
            "academic_year": ay,
            "key": f"{batch}|{section}|{ay}",
            "label": label or "—",
        })
    return sections


def _resolve_section_key(section_key, sections):
    """
    Parse a 'batch|section|academic_year' key. Falls back to the first
    available section when nothing valid is selected.
    Returns (batch, section, academic_year, normalised_key).
    """
    batch = section = ay = ""
    if section_key:
        parts = section_key.split("|")
        batch = parts[0] if len(parts) > 0 else ""
        section = parts[1] if len(parts) > 1 else ""
        ay = parts[2] if len(parts) > 2 else ""

    valid_keys = {s["key"] for s in sections}
    if section_key not in valid_keys and sections:
        first = sections[0]
        return first["batch"], first["section"], first["academic_year"], first["key"]
    return batch, section, ay, section_key


def _build_class_work_record(faculty, course_code, batch, section, academic_year=None, unit=None):
    """
    Build the 'Record of Class Work' for ONE faculty + course + section.

    Plan structure comes from CoursePlan; the actual delivery Date + class
    Period for each topic comes from student_management_hourattendance for
    that exact section. Returns a dict consumed by the page, the AJAX partial
    and the PDF export.

    rows is a flat list of two kinds of entries:
        {"type": "unit",  "unit": "1"}
        {"type": "topic", "cum_hours": 1, "topic": ..., "method": ...,
         "date": <date|None>, "period": <int|None>, "covered": bool, ...}
    """
    from django.db.models import Count, Q

    info = {
        "course_title": "",
        "plan_academic_year": (academic_year or "").strip(),
        "units": [],
        "selected_unit": (unit or "").strip(),
        "rows": [],
        "summary": {
            "total_topics": 0,
            "completed_topics": 0,
            "pending_topics": 0,
            "completion_pct": 0,
        },
    }
    if not course_code:
        return info

    # ---- Course plan (prefer this faculty's own; fall back to any) ----
    plan_base = CoursePlan.objects.filter(course__course_code=course_code)
    if plan_base.filter(faculty=faculty).exists():
        plan_base = plan_base.filter(faculty=faculty)

    available_years = list(
        plan_base.exclude(academic_year__isnull=True)
        .exclude(academic_year__exact="")
        .values_list("academic_year", flat=True)
        .distinct()
        .order_by("-academic_year")
    )
    req_year = (academic_year or "").strip()
    if req_year and req_year in available_years:
        plan_year = req_year
    elif available_years:
        plan_year = available_years[0]
    else:
        plan_year = req_year
    info["plan_academic_year"] = plan_year

    plans = plan_base.filter(academic_year=plan_year) if plan_year else plan_base
    plan_rows = list(plans.order_by("unit_module_no", "period_no", "id"))
    if not plan_rows:
        return info

    course_obj = Course.objects.filter(course_code=course_code).first()
    info["course_title"] = getattr(course_obj, "title", "") or ""

    # ---- Actual delivery (earliest date+period) per plan, for THIS section ----
    plan_ids = [p.id for p in plan_rows]
    ha = HourAttendance.objects.filter(faculty=faculty, course_plan_id__in=plan_ids)
    if batch:
        ha = ha.filter(batch=batch)
    if section:
        ha = ha.filter(section=section)

    sessions = (
        ha.values("course_plan_id", "date", "period")
        .annotate(
            present=Count("id", filter=Q(status="Present")),
            total=Count("id"),
        )
        .order_by("date", "period")
    )
    delivered = {}
    for s in sessions:
        pid = s["course_plan_id"]
        if pid not in delivered:   # first by (date, period) ordering = earliest
            delivered[pid] = s

    # ---- Build register rows for the FULL plan first ----
    # Cumulative Hours are numbered continuously across every unit, so a unit
    # keeps its true position (e.g. Unit 2 starts at hour 8) even when the
    # report is later filtered down to a single unit.
    all_rows = []
    cum = 0
    unit_counts = {}
    unit_order = []
    last_unit = "\x00"  # sentinel no real unit equals
    for plan in plan_rows:
        u = (plan.unit_module_no or "").strip()
        if u != last_unit:
            all_rows.append({"type": "unit", "unit": u})
            last_unit = u
        if u not in unit_counts:
            unit_counts[u] = 0
            unit_order.append(u)
        unit_counts[u] += 1

        cum += 1
        sess = delivered.get(plan.id)
        covered = sess is not None
        all_rows.append({
            "type": "topic",
            "unit": u,
            "cum_hours": cum,
            "co_no": plan.co_no or "",
            "topic": plan.topic or "",
            "method": plan.delivery_method or "",
            "date": sess["date"] if sess else None,
            "period": sess["period"] if sess else None,
            "present": sess["present"] if sess else None,
            "present_total": sess["total"] if sess else None,
            "covered": covered,
        })

    # ---- Units available for the dropdown ----
    info["units"] = [{"unit_module_no": u, "plan_count": unit_counts[u]} for u in unit_order]

    # ---- Apply the unit filter (blank = All Units) ----
    selected_unit = (unit or "").strip()
    if selected_unit and selected_unit not in unit_counts:
        selected_unit = ""   # invalid selection -> show all
    info["selected_unit"] = selected_unit

    if selected_unit:
        rows = [r for r in all_rows if r.get("unit") == selected_unit]
    else:
        rows = all_rows

    # ---- Summary computed from the DISPLAYED topic rows ----
    topic_rows = [r for r in rows if r["type"] == "topic"]
    total_topics = len(topic_rows)
    completed = sum(1 for r in topic_rows if r["covered"])
    pending = total_topics - completed
    pct = round((completed / total_topics) * 100) if total_topics else 0

    info["rows"] = rows
    info["summary"] = {
        "total_topics": total_topics,
        "completed_topics": completed,
        "pending_topics": pending,
        "completion_pct": pct,
    }
    return info


@check_permission("course_plan_completion_report")
@student_management
def course_plan_completion_report(request):
    """
    Faculty-facing 'Record of Class Work' for a course + section, derived from
    the course plan and the hour-attendance table. Renders the full page on a
    normal request and only the results partial on an AJAX request (so the
    course / section filters swap the table without a full reload).
    """
    # ---- Session / faculty recognition ----
    faculty_emp_id = getattr(request.user, "Employee_id", None)
    faculty = general_information.objects.filter(faculty_id=faculty_emp_id).first()
    if not faculty:
        messages.error(request, "Faculty not found.")
        return redirect("home")

    # ---- Courses assigned to this faculty (distinct by course code) ----
    assigned_courses = []
    seen_codes = set()
    assignments = (
        AssignSubjectFaculty.objects
        .filter(faculty=faculty, is_active=True)
        .select_related("course")
        .order_by("course__course_code")
    )
    for a in assignments:
        course = a.course
        if not course or course.course_code in seen_codes:
            continue
        seen_codes.add(course.course_code)
        assigned_courses.append({
            "course_code": course.course_code,
            "title": course.title,
        })

    course_code = (request.GET.get("course_code") or "").strip()
    section_key = (request.GET.get("section_key") or "").strip()
    unit = (request.GET.get("unit_module_no") or "").strip()

    sections = _faculty_course_sections(faculty, course_code)
    batch, section, academic_year, section_key = _resolve_section_key(section_key, sections)

    data = _build_class_work_record(faculty, course_code, batch, section, academic_year, unit)

    context = {
        "assigned_courses": assigned_courses,
        "selected_course_code": course_code,
        "sections": sections,
        "selected_section_key": section_key,
        "selected_batch": batch,
        "selected_section": section,
        **data,
    }

    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"
    if is_ajax:
        return render(
            request,
            "student_management/_course_plan_completion_result.html",
            context,
        )

    return render(
        request,
        "student_management/course_plan_completion_report.html",
        context,
    )


def course_plan_completion_report_pdf(request):
    """
    PDF export of the 'Record of Class Work' for a course + section,
    laid out like the physical class-work register.
    """
    faculty_emp_id = getattr(request.user, "Employee_id", None)
    faculty = general_information.objects.filter(faculty_id=faculty_emp_id).first()
    if not faculty:
        return HttpResponse("Faculty not found.", status=404)

    course_code = (request.GET.get("course_code") or "").strip()
    section_key = (request.GET.get("section_key") or "").strip()
    unit = (request.GET.get("unit_module_no") or "").strip()
    if not course_code:
        return HttpResponse("Course is required.", status=400)

    sections = _faculty_course_sections(faculty, course_code)
    batch, section, academic_year, section_key = _resolve_section_key(section_key, sections)

    data = _build_class_work_record(faculty, course_code, batch, section, academic_year, unit)
    if not data["rows"]:
        return HttpResponse("No course plan found for this course / section.", status=404)

    return _render_class_work_pdf(faculty, course_code, data, batch, section)


def _render_class_work_pdf(faculty, course_code, data, batch, section):
    """Render the Record of Class Work register as a PDF (portrait A4)."""
    styles = getSampleStyleSheet()
    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1a4b8c")
    ACCENT_RED = colors.HexColor("#b91c1c")
    DARK_GRAY = colors.HexColor("#111827")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    UNIT_BG = colors.HexColor("#eaf1fb")
    GRID_BLUE = colors.HexColor("#1a4b8c")
    GRID_LIGHT = colors.HexColor("#aebfd6")

    title_style = ParagraphStyle("cw_title", parent=styles["Heading1"], fontSize=14,
                                 textColor=PRIMARY_BLUE, alignment=TA_CENTER, spaceAfter=2,
                                 fontName="Helvetica-Bold")
    record_style = ParagraphStyle("cw_record", parent=styles["Heading2"], fontSize=12.5,
                                  textColor=SECONDARY_BLUE, alignment=TA_CENTER, spaceAfter=3,
                                  fontName="Helvetica-Bold")
    info_style = ParagraphStyle("cw_info", parent=styles["Normal"], fontSize=9.5,
                                textColor=DARK_GRAY, alignment=TA_LEFT, leading=13)
    header_cell = ParagraphStyle("cw_hcell", parent=styles["Normal"], fontSize=9.5,
                                 textColor=SECONDARY_BLUE, alignment=TA_CENTER,
                                 fontName="Helvetica-Bold", leading=11, wordWrap="CJK")
    topic_cell = ParagraphStyle("cw_topic", parent=styles["Normal"], fontSize=9.5,
                                textColor=DARK_GRAY, alignment=TA_LEFT, leading=12, wordWrap="CJK")
    center_cell = ParagraphStyle("cw_center", parent=topic_cell, alignment=TA_CENTER)
    unit_cell = ParagraphStyle("cw_unit", parent=styles["Normal"], fontSize=10,
                               textColor=PRIMARY_BLUE, alignment=TA_LEFT,
                               fontName="Helvetica-Bold", leading=12)

    faculty_name = _safe(getattr(faculty, "name", ""))
    faculty_code = _safe(getattr(faculty, "faculty_id", ""))

    def fmt_date(d):
        return d.strftime("%d/%m/%y") if d else ""

    # ---- Table data ----
    table_data = [[
        Paragraph("Date", header_cell),
        Paragraph("Period", header_cell),
        Paragraph("Cumulative Hours", header_cell),
        Paragraph("Topics Covered", header_cell),
        Paragraph("Instructional Methods", header_cell),
    ]]
    unit_row_indices = []
    r = 1
    for row in data["rows"]:
        if row["type"] == "unit":
            unit_no = row.get("unit") or ""
            label = f"UNIT - {unit_no}" if unit_no else "OTHER TOPICS"
            table_data.append(["", "", "", Paragraph(label, unit_cell), ""])
            unit_row_indices.append(r)
        else:
            table_data.append([
                Paragraph(fmt_date(row["date"]), center_cell),
                Paragraph(str(row["period"]) if row["period"] else "", center_cell),
                Paragraph(str(row["cum_hours"]), center_cell),
                Paragraph(row["topic"] or "", topic_cell),
                Paragraph(row["method"] or "", topic_cell),
            ])
        r += 1

    col_widths = [24 * mm, 14 * mm, 22 * mm, 92 * mm, 34 * mm]  # ~186mm
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    ts = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.7, GRID_LIGHT),
        ("BOX", (0, 0), (-1, -1), 1.1, GRID_BLUE),
        ("LINEBELOW", (0, 0), (-1, 0), 1.1, GRID_BLUE),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("ALIGN", (3, 0), (4, -1), "LEFT"),
    ]
    # Unit-header rows: keep the full grid (Date/Period/Cum/Method cells stay
    # empty) with the unit title sitting only in the Topics column, exactly
    # like the physical register.
    for ri in unit_row_indices:
        ts.append(("TOPPADDING", (0, ri), (-1, ri), 7))
        ts.append(("BOTTOMPADDING", (0, ri), (-1, ri), 7))
    table.setStyle(TableStyle(ts))

    # ---- Document ----
    buffer = io.BytesIO()
    PAGE = A4  # portrait
    doc = BaseDocTemplate(
        buffer, pagesize=PAGE,
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=12 * mm,
        title="Record of Class Work", showBoundary=0,
    )
    HEADER_HEIGHT = 26 * mm

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = PAGE
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 4 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand
        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(ImageReader(logo_path), left, top_y - 15 * mm,
                             width=24 * mm, height=14 * mm, preserveAspectRatio=True, mask="auto")

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawCentredString(center_x, top_y - 4 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")
        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawCentredString(center_x, top_y - 9 * mm, "An Autonomous Institution")
        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 7.5)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "COURSE PLAN COMPLETION — RECORD OF CLASS WORK")

        footer_y = 8 * mm
        canvas.setStrokeColor(GRID_LIGHT)
        canvas.setLineWidth(0.6)
        canvas.line(left, footer_y + 4 * mm, right, footer_y + 4 * mm)
        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 7)
        gen_time = datetime.now().strftime("%d %b %Y, %H:%M")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawCentredString(center_x, footer_y, f"Faculty: {faculty_code}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")
        canvas.restoreState()

    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height - HEADER_HEIGHT, id="normal")
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

    course_title = data.get("course_title", "")
    sec_label = f"{batch} / {section}".strip(" /") or "—"
    summary = data["summary"]

    elements = []
    elements.append(Spacer(1, 2 * mm))
    elements.append(Paragraph(f"{course_code} — {course_title}", title_style))

    meta_rows = [
        ["Faculty:", faculty_name or faculty_code, "Section:", sec_label],
        ["Academic Year:", data.get("plan_academic_year", "") or "—",
         "Completion:", f"{summary['completed_topics']}/{summary['total_topics']} ({summary['completion_pct']}%)"],
    ]
    info_tbl = Table(
        [[Paragraph(f"<b>{a}</b>", info_style), Paragraph(str(b), info_style),
          Paragraph(f"<b>{c}</b>", info_style), Paragraph(str(d), info_style)]
         for a, b, c, d in meta_rows],
        colWidths=[26 * mm, doc.width / 2 - 26 * mm, 30 * mm, doc.width / 2 - 30 * mm],
    )
    info_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_tbl)
    elements.append(Spacer(1, 4 * mm))
    sel_unit = (data.get("selected_unit") or "").strip()
    record_title = f"RECORD OF CLASS WORK — UNIT {sel_unit}" if sel_unit else "RECORD OF CLASS WORK"
    elements.append(Paragraph(record_title, record_style))
    elements.append(Spacer(1, 2 * mm))
    elements.append(table)

    try:
        doc.build(elements)
    except Exception as e:
        return HttpResponse(f"PDF generation failed: {e}", status=500)

    buffer.seek(0)
    safe_sec = (sec_label or "section").replace("/", "-").replace(" ", "")
    filename = f"Record_of_Class_Work_{course_code}_{safe_sec}.pdf"
    return FileResponse(buffer, as_attachment=False, filename=filename)



from datetime import date
from user_accounts.models import USER


def get_academic_year():
    """
    Dynamically returns academic year string.
    Example:
      If current month >= June → '2025-2026'
      Else (Jan–May) → '2024-2025'
    """
    today = date.today()
    current_year = today.year
    if today.month >= 6:  # June or later
        return f"{current_year}-{current_year + 1}"
    else:  # Before June → part of previous cycle
        return f"{current_year - 1}-{current_year}"




import io
import os
from collections import defaultdict
from datetime import date, datetime

from django.conf import settings
from django.contrib import messages
from django.contrib.staticfiles import finders
from django.http import HttpResponse, FileResponse
from django.shortcuts import get_object_or_404, redirect, render

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import BaseDocTemplate, PageTemplate, Frame, Table, TableStyle, Paragraph, Spacer, PageBreak
from datetime import date
from collections import defaultdict

from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect, render

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame,
    Table, TableStyle, Paragraph, Spacer
)
from reportlab.pdfbase.pdfmetrics import stringWidth



def _safe(v):
    return "" if v is None else str(v).strip()


def _parse_from_to_date(request):
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    parsed_from = None
    parsed_to = None

    try:
        if date_from:
            parsed_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        if date_to:
            parsed_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return None, None, None, None, HttpResponse("Invalid date format. Use YYYY-MM-DD.", status=400)

    if parsed_from and parsed_to and parsed_from > parsed_to:
        return None, None, None, None, HttpResponse("'From Date' cannot be greater than 'To Date'.", status=400)

    return date_from, date_to, parsed_from, parsed_to, None


def _get_class_advisor_base_data(request, selected_semester=None):
    faculty_id = getattr(request.user, "Employee_id", None)
    faculty = get_object_or_404(general_information, faculty_id=faculty_id)

    assigned_students = (
        StudentDetails.objects
        .filter(ca=faculty, is_active=True)
        .order_by("reg_no")
    )

    if not assigned_students.exists():
        return faculty, [], "", "", "", [], ""

    academic_year = get_academic_year()

    sems_from_students = set(assigned_students.values_list("semester", flat=True))
    sems_from_attendance = set(
        Daily_Attendance.objects.filter(
            student__in=assigned_students,
            academic_year=academic_year
        ).values_list("semester", flat=True)
    )

    available_semesters = sorted(
        {str(s).strip() for s in (sems_from_students | sems_from_attendance) if str(s).strip()},
        key=lambda x: int(x) if x.isdigit() else x
    )

    semester = str(selected_semester).strip() if selected_semester else ""
    if not semester:
        first_student = assigned_students.first()
        semester = _safe(getattr(first_student, "semester", ""))

    filtered_students = assigned_students
    if semester:
        filtered_students = filtered_students.filter(semester=str(semester))

    first_student = filtered_students.first()
    year = _safe(getattr(first_student, "year", "")) if first_student else ""
    section = _safe(getattr(first_student, "section", "")) if first_student else ""

    return faculty, list(filtered_students), semester, year, section, available_semesters, academic_year


@check_permission("class_attendence")
@student_management
def class_attendence(request):
    faculty_id = request.user.Employee_id
    today = date.today()

    selected_date_str = request.GET.get("date") or request.POST.get("attendance_date") or today.strftime("%Y-%m-%d")

    try:
        selected_date = date.fromisoformat(selected_date_str)
    except ValueError:
        selected_date = today
        messages.warning(request, "Invalid date selected. Using today's date.")

    if selected_date > today:
        messages.warning(request, "Future dates are not allowed. Using today's date.")
        selected_date = today

    faculty, assigned_students, semester, year, section, available_semesters, academic_year = _get_class_advisor_base_data(request)

    if not assigned_students:
        messages.info(request, "You are not assigned as a Class Advisor to any students.")
        return render(request, "faculty_management/class_attendence.html", {
            "students": [],
            "selected_date": selected_date,
            "today": today,
            "academic_year": academic_year,
            "available_semesters": [],
            "year": "",
            "semester": "",
            "section": "",
            "batch": "",
            "course_id": 0,
            "regulation_id": 0,
        })

    attendance_records = Daily_Attendance.objects.filter(
        student__in=assigned_students,
        date=selected_date
    )

    attendance_dict = {
        att.student_id: {
            "morning": att.morning_status,
            "afternoon": att.afternoon_status,
            "full": att.full_day_status,
        }
        for att in attendance_records
    }

    for student in assigned_students:
        att = attendance_dict.get(student.id, {})
        student.morning_status = att.get("morning", "Present")
        student.afternoon_status = att.get("afternoon", "Present")
        student.full_day_status = att.get("full", "Present")
        student.attendance_exists = student.id in attendance_dict

    if request.method == "POST":
        attendance_saved = 0

        for student in assigned_students:
            morning_status = request.POST.get(f"morning_{student.id}", "Present")
            afternoon_status = request.POST.get(f"afternoon_{student.id}", "Present")

            if morning_status == afternoon_status:
                full_day_status = morning_status
            else:
                full_day_status = "Half Day"

            Daily_Attendance.objects.update_or_create(
                student=student,
                date=selected_date,
                defaults={
                    "faculty": faculty,
                    "marked_by": faculty,
                    "updated_by": faculty,
                    "morning_status": morning_status,
                    "afternoon_status": afternoon_status,
                    "full_day_status": full_day_status,
                    "year": student.year,
                    "semester": student.semester,
                    "section": student.section,
                    "academic_year": academic_year,
                },
            )
            attendance_saved += 1

        messages.success(
            request,
            f"Attendance saved successfully for {selected_date.strftime('%d-%m-%Y')}. Marked {attendance_saved} students."
        )
        return redirect(f"{request.path}?date={selected_date}")

    context = {
        "students": assigned_students,
        "selected_date": selected_date,
        "today": today,
        "academic_year": academic_year,
        "available_semesters": available_semesters,
        "year": year,
        "semester": semester,
        "section": section,
        "batch": year,          # placeholder for your existing URL pattern
        "course_id": 0,         # placeholder
        "regulation_id": 0,     # placeholder
    }
    return render(request, "faculty_management/class_attendence.html", context)


def class_attendance_detail_view(request, year, semester, course_id, batch, section, regulation_id):
    date_from, date_to, parsed_from, parsed_to, error_response = _parse_from_to_date(request)
    if error_response:
        return error_response

    faculty, assigned_students, actual_semester, actual_year, actual_section, available_semesters, academic_year = _get_class_advisor_base_data(
        request,
        selected_semester=semester
    )

    if not assigned_students:
        return HttpResponse("No students found for this class advisor.", status=404)

    att_qs = Daily_Attendance.objects.filter(
    student__in=assigned_students,
    
).select_related("student").order_by("date", "student__reg_no")

    if parsed_from:
        att_qs = att_qs.filter(date__gte=parsed_from)
    if parsed_to:
        att_qs = att_qs.filter(date__lte=parsed_to)

    grouped = defaultdict(dict)
    all_dates = set()

    for row in att_qs:
        if not row.student_id or not row.date:
            continue
        grouped[row.student_id][row.date] = {
            "morning": _safe(row.morning_status) or "-",
            "afternoon": _safe(row.afternoon_status) or "-",
            "full_day": _safe(row.full_day_status) or "-",
        }
        all_dates.add(row.date)

    sorted_dates = sorted(all_dates)
    students_sorted = sorted(assigned_students, key=lambda s: _safe(getattr(s, "reg_no", "")))

    date_columns = []
    for dt in sorted_dates:
        date_columns.append({
            "date": dt,
            "date_str": dt.strftime("%d-%m-%Y"),
            "subcols": ["Morning", "Afternoon", "Full Day"],
            "subcol_count": 3,
        })

    student_rows = []
    for idx, student in enumerate(students_sorted, start=1):
        values = []
        for dt in sorted_dates:
            rec = grouped[student.id].get(dt, {})
            for key in ["morning", "afternoon", "full_day"]:
                value = rec.get(key, "-")
                values.append({
                    "value": value,
                    "is_absent": str(value).strip().lower() == "absent",
                })

        student_rows.append({
            "sno": idx,
            "reg_no": _safe(getattr(student, "reg_no", "")),
            "name": _safe(getattr(student, "name", "")),
            "values": values,
        })

    return render(
        request,
        "faculty_management/daily_attendance_detail_view.html",
        {
            "faculty": faculty,
            "year": actual_year,
            "semester": actual_semester,
            "batch": year,
            "section": actual_section,
            "date_from": date_from,
            "date_to": date_to,
            "date_columns": date_columns,
            "student_rows": student_rows,
        }
    )


@check_permission("class_attendence")
@student_management
def class_attendance_pdf(request, year, semester, course_id, batch, section, regulation_id):
    date_from, date_to, parsed_from, parsed_to, error_response = _parse_from_to_date(request)
    if error_response:
        return error_response

    faculty, assigned_students, actual_semester, actual_year, actual_section, available_semesters, academic_year = _get_class_advisor_base_data(
        request,
        selected_semester=semester
    )

    if not assigned_students:
        return HttpResponse("No students found for this class advisor.", status=404)

    att_qs = Daily_Attendance.objects.filter(
        student__in=assigned_students,
        
        
    ).select_related("student").order_by("date", "student__reg_no")

    if parsed_from:
        att_qs = att_qs.filter(date__gte=parsed_from)
    if parsed_to:
        att_qs = att_qs.filter(date__lte=parsed_to)

    if not att_qs.exists():
        return HttpResponse("No attendance records found for the selected criteria.", status=404)

    summary = defaultdict(lambda: {"present": 0, "absent": 0, "od": 0, "half": 0, "total": 0})

    for row in att_qs:
        summary[row.student_id]["total"] += 1
        status = _safe(row.full_day_status)
        if status == "Present":
            summary[row.student_id]["present"] += 1
        elif status == "Absent":
            summary[row.student_id]["absent"] += 1
        elif status == "On Duty":
            summary[row.student_id]["od"] += 1
        elif status == "Half Day":
            summary[row.student_id]["half"] += 1

    students_sorted = sorted(assigned_students, key=lambda s: _safe(getattr(s, "reg_no", "")))
    faculty_name = _safe(getattr(faculty, "name", "")) or "Faculty"
    faculty_code = _safe(getattr(faculty, "faculty_id", ""))

    styles = getSampleStyleSheet()
    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1a4b8c")
    ACCENT_RED = colors.HexColor("#b91c1c")
    DARK_GRAY = colors.HexColor("#111827")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BG_GRAY = colors.HexColor("#f8fafc")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle("title_style", parent=styles["Heading1"], fontSize=16, textColor=PRIMARY_BLUE, alignment=TA_CENTER, spaceAfter=4, fontName="Helvetica-Bold")
    sub_style = ParagraphStyle("sub_style", parent=styles["Normal"], fontSize=10, textColor=MEDIUM_GRAY, alignment=TA_CENTER, spaceAfter=8)
    info_style = ParagraphStyle("info_style", parent=styles["Normal"], fontSize=10, textColor=DARK_GRAY, alignment=TA_LEFT, leading=13, fontName="Helvetica")
    section_style = ParagraphStyle("section_style", parent=styles["Heading2"], fontSize=12, textColor=PRIMARY_BLUE, alignment=TA_LEFT, spaceBefore=8, spaceAfter=5, fontName="Helvetica-Bold")
    table_header = ParagraphStyle("table_header", parent=styles["Normal"], fontSize=10, textColor=colors.white, alignment=TA_CENTER, fontName="Helvetica-Bold", leading=12, wordWrap="CJK")
    cell_left = ParagraphStyle("cell_left", parent=styles["Normal"], fontSize=10, textColor=DARK_GRAY, alignment=TA_LEFT, leading=12, wordWrap="CJK")
    cell_center = ParagraphStyle("cell_center", parent=cell_left, alignment=TA_CENTER)

    def p_left(txt):
        return Paragraph(_safe(txt), cell_left)

    def p_center(txt):
        return Paragraph(_safe(txt), cell_center)

    buffer = io.BytesIO()
    PAGE = landscape(A4)

    doc = BaseDocTemplate(
        buffer,
        pagesize=PAGE,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title="Class Attendance Report",
        showBoundary=0
    )

    HEADER_HEIGHT = 26 * mm

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = PAGE
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 4 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(ImageReader(logo_path), left, top_y - 15 * mm, width=24 * mm, height=14 * mm, preserveAspectRatio=True, mask="auto")

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(center_x, top_y - 4 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawCentredString(center_x, top_y - 9 * mm, "An Autonomous Institution")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 7.5)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "CLASS ADVISOR ATTENDANCE REPORT")

        footer_y = 9 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.6)
        canvas.line(left, footer_y + 5 * mm, right, footer_y + 5 * mm)

        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 7)
        gen_time = datetime.now().strftime("%d %b %Y, %H:%M")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawCentredString(center_x, footer_y, f"Faculty: {faculty_code}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")
        canvas.restoreState()

    frame = Frame(doc.leftMargin, doc.bottomMargin + 3 * mm, doc.width, doc.height - HEADER_HEIGHT + 4 * mm, leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0, id="normal")
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

    elements = []
    elements.append(Spacer(1, 2 * mm))
    elements.append(Paragraph("OVERALL CLASS ATTENDANCE REPORT", title_style))
    elements.append(Paragraph(f"{faculty_name} ({faculty_code})", sub_style))

    info_data = [
        [Paragraph("<b>Year:</b>", info_style), Paragraph(str(actual_year), info_style),
         Paragraph("<b>Semester:</b>", info_style), Paragraph(str(actual_semester), info_style)],
        [Paragraph("<b>Section:</b>", info_style), Paragraph(str(actual_section), info_style),
         Paragraph("<b>Academic Year:</b>", info_style), Paragraph(str(academic_year), info_style)],
        [Paragraph("<b>Date Range:</b>", info_style), Paragraph(f"{date_from or '...'} to {date_to or '...'}", info_style),
         Paragraph("<b>Total Students:</b>", info_style), Paragraph(str(len(students_sorted)), info_style)],
    ]

    info_table = Table(info_data, colWidths=[28 * mm, doc.width / 2 - 28 * mm, 30 * mm, doc.width / 2 - 30 * mm])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 5 * mm))
    elements.append(Paragraph("STUDENT SUMMARY", section_style))

    table_data = [[
        Paragraph("S.No", table_header),
        Paragraph("Reg No", table_header),
        Paragraph("Student Name", table_header),
        Paragraph("Present", table_header),
        Paragraph("Absent", table_header),
        Paragraph("On Duty", table_header),
        Paragraph("Half Day", table_header),
        Paragraph("Total", table_header),
        Paragraph("%", table_header),
    ]]

    for idx, student in enumerate(students_sorted, start=1):
        s = summary[student.id]
        total = s["total"]
        attended = s["present"] + s["od"] + (s["half"] * 0.5)
        percentage = (attended / total * 100.0) if total else 0.0

        table_data.append([
            p_center(idx),
            p_center(_safe(getattr(student, "reg_no", ""))),
            p_left(_safe(getattr(student, "name", ""))),
            p_center(s["present"]),
            p_center(s["absent"]),
            p_center(s["od"]),
            p_center(s["half"]),
            p_center(total),
            p_center(f"{percentage:.2f}%"),
        ])

    col_widths = [14 * mm, 28 * mm, 70 * mm, 22 * mm, 22 * mm, 22 * mm, 22 * mm, 22 * mm, 24 * mm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.7, BORDER_GRAY),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("ALIGN", (2, 0), (2, -1), "LEFT"),
        ("ALIGN", (3, 0), (-1, -1), "CENTER"),
    ]))
    elements.append(table)

    try:
        doc.build(elements)
    except Exception as e:
        return HttpResponse(f"PDF generation failed: {e}", status=500)

    buffer.seek(0)
    filename = f"Class_Attendance_Overall_{actual_year}_{actual_semester}_{actual_section}.pdf"
    return FileResponse(buffer, as_attachment=False, filename=filename)


@check_permission("class_attendence")
@student_management
def class_attendance_datewise_pdf(request, year, semester, course_id, batch, section, regulation_id):
    date_from, date_to, parsed_from, parsed_to, error_response = _parse_from_to_date(request)
    if error_response:
        return error_response

    faculty, assigned_students, actual_semester, actual_year, actual_section, available_semesters, academic_year = _get_class_advisor_base_data(
        request,
        selected_semester=semester
    )

    if not assigned_students:
        return HttpResponse("No students found for this class advisor.", status=404)

    att_qs = Daily_Attendance.objects.filter(
        student__in=assigned_students,
        
        
    ).select_related("student").order_by("date", "student__reg_no")

    if parsed_from:
        att_qs = att_qs.filter(date__gte=parsed_from)
    if parsed_to:
        att_qs = att_qs.filter(date__lte=parsed_to)

    if not att_qs.exists():
        return HttpResponse("No attendance records found for the selected criteria.", status=404)

    grouped = defaultdict(dict)
    all_dates = set()

    for row in att_qs:
        if not row.student_id or not row.date:
            continue
        grouped[row.date][row.student_id] = {
            "morning": _safe(row.morning_status) or "-",
            "afternoon": _safe(row.afternoon_status) or "-",
            "full_day": _safe(row.full_day_status) or "-",
        }
        all_dates.add(row.date)

    sorted_dates = sorted(all_dates)
    students_sorted = sorted(assigned_students, key=lambda s: _safe(getattr(s, "reg_no", "")))
    faculty_name = _safe(getattr(faculty, "name", "")) or "Faculty"
    faculty_code = _safe(getattr(faculty, "faculty_id", ""))

    styles = getSampleStyleSheet()
    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1a4b8c")
    ACCENT_RED = colors.HexColor("#b91c1c")
    DARK_GRAY = colors.HexColor("#111827")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BG_GRAY = colors.HexColor("#f8fafc")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle("title_style", parent=styles["Heading1"], fontSize=16, textColor=PRIMARY_BLUE, alignment=TA_CENTER, spaceAfter=4, fontName="Helvetica-Bold")
    sub_style = ParagraphStyle("sub_style", parent=styles["Normal"], fontSize=10, textColor=MEDIUM_GRAY, alignment=TA_CENTER, spaceAfter=8)
    info_style = ParagraphStyle("info_style", parent=styles["Normal"], fontSize=10, textColor=DARK_GRAY, alignment=TA_LEFT, leading=13, fontName="Helvetica")
    table_header = ParagraphStyle("table_header", parent=styles["Normal"], fontSize=9, textColor=colors.white, alignment=TA_CENTER, fontName="Helvetica-Bold", leading=11, wordWrap="CJK")
    cell_left = ParagraphStyle("cell_left", parent=styles["Normal"], fontSize=8.5, textColor=DARK_GRAY, alignment=TA_LEFT, leading=10, wordWrap="CJK")
    cell_center = ParagraphStyle("cell_center", parent=cell_left, alignment=TA_CENTER)

    def p_left(txt):
        return Paragraph(_safe(txt), cell_left)

    def p_center(txt):
        return Paragraph(_safe(txt), cell_center)

    buffer = io.BytesIO()
    PAGE = landscape(A4)
    doc = BaseDocTemplate(buffer, pagesize=PAGE, leftMargin=10 * mm, rightMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm, title="Datewise Class Attendance Report", showBoundary=0)
    HEADER_HEIGHT = 26 * mm

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = PAGE
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 4 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(ImageReader(logo_path), left, top_y - 15 * mm, width=24 * mm, height=14 * mm, preserveAspectRatio=True, mask="auto")

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(center_x, top_y - 4 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawCentredString(center_x, top_y - 9 * mm, "An Autonomous Institution")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 7.5)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "DATE-WISE DAILY ATTENDANCE REPORT")

        footer_y = 9 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.6)
        canvas.line(left, footer_y + 5 * mm, right, footer_y + 5 * mm)

        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 7)
        gen_time = datetime.now().strftime("%d %b %Y, %H:%M")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawCentredString(center_x, footer_y, f"Faculty: {faculty_code}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")
        canvas.restoreState()

    frame = Frame(doc.leftMargin, doc.bottomMargin + 3 * mm, doc.width, doc.height - HEADER_HEIGHT + 4 * mm, leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0, id="normal")
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

    elements = []
    elements.append(Spacer(1, 2 * mm))
    elements.append(Paragraph("DATE-WISE DAILY ATTENDANCE REPORT", title_style))
    elements.append(Paragraph(f"{faculty_name} ({faculty_code})", sub_style))

    info_data = [
        [Paragraph("<b>Year:</b>", info_style), Paragraph(str(actual_year), info_style),
         Paragraph("<b>Semester:</b>", info_style), Paragraph(str(actual_semester), info_style)],
        [Paragraph("<b>Section:</b>", info_style), Paragraph(str(actual_section), info_style),
         Paragraph("<b>Academic Year:</b>", info_style), Paragraph(str(academic_year), info_style)],
        [Paragraph("<b>Date Range:</b>", info_style), Paragraph(f"{date_from or '...'} to {date_to or '...'}", info_style),
         Paragraph("<b>Total Students:</b>", info_style), Paragraph(str(len(students_sorted)), info_style)],
    ]

    info_table = Table(info_data, colWidths=[28 * mm, doc.width / 2 - 28 * mm, 30 * mm, doc.width / 2 - 30 * mm])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 5 * mm))

    for idx, dt in enumerate(sorted_dates, start=1):
        elements.append(Paragraph(f"<b>Attendance Date:</b> {dt.strftime('%d-%m-%Y')}", info_style))
        elements.append(Spacer(1, 2 * mm))

        rows = [[
            Paragraph("S.No", table_header),
            Paragraph("Reg No", table_header),
            Paragraph("Student Name", table_header),
            Paragraph("Morning", table_header),
            Paragraph("Afternoon", table_header),
            Paragraph("Full Day", table_header),
        ]]

        for sno, student in enumerate(students_sorted, start=1):
            rec = grouped[dt].get(student.id, {})
            rows.append([
                p_center(sno),
                p_center(_safe(getattr(student, "reg_no", ""))),
                p_left(_safe(getattr(student, "name", ""))),
                p_center(rec.get("morning", "-")),
                p_center(rec.get("afternoon", "-")),
                p_center(rec.get("full_day", "-")),
            ])

        col_widths = [14 * mm, 30 * mm, 80 * mm, 30 * mm, 30 * mm, 30 * mm]
        table = Table(rows, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.6, BORDER_GRAY),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
            ("ALIGN", (0, 0), (1, -1), "CENTER"),
            ("ALIGN", (3, 1), (-1, -1), "CENTER"),
        ]))
        elements.append(table)

        if idx < len(sorted_dates):
            elements.append(PageBreak())

    try:
        doc.build(elements)
    except Exception as e:
        return HttpResponse(f"PDF generation failed: {e}", status=500)

    buffer.seek(0)
    filename = f"Class_Attendance_Datewise_{actual_year}_{actual_semester}_{actual_section}.pdf"
    return FileResponse(buffer, as_attachment=False, filename=filename)


@check_permission("class_attendence")
@student_management
def class_attendance_datewise_excel(request, year, semester, course_id, batch, section, regulation_id):
    date_from, date_to, parsed_from, parsed_to, error_response = _parse_from_to_date(request)
    if error_response:
        return error_response

    faculty, assigned_students, actual_semester, actual_year, actual_section, available_semesters, academic_year = _get_class_advisor_base_data(
        request,
        selected_semester=semester
    )

    if not assigned_students:
        return HttpResponse("No students found for this class advisor.", status=404)

    att_qs = Daily_Attendance.objects.filter(
        student__in=assigned_students,
        
        
    ).select_related("student").order_by("date", "student__reg_no")

    if parsed_from:
        att_qs = att_qs.filter(date__gte=parsed_from)
    if parsed_to:
        att_qs = att_qs.filter(date__lte=parsed_to)

    if not att_qs.exists():
        return HttpResponse("No attendance records found for the selected criteria.", status=404)

    grouped = defaultdict(dict)
    all_dates = set()

    for row in att_qs:
        if not row.student_id or not row.date:
            continue
        grouped[row.student_id][row.date] = {
            "morning": _safe(row.morning_status) or "-",
            "afternoon": _safe(row.afternoon_status) or "-",
            "full_day": _safe(row.full_day_status) or "-",
        }
        all_dates.add(row.date)

    sorted_dates = sorted(all_dates)
    students_sorted = sorted(assigned_students, key=lambda s: _safe(getattr(s, "reg_no", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    absent_fill = PatternFill(fill_type="solid", fgColor="FDE2E1")

    ws.merge_cells("A1:D1")
    ws["A1"] = "DATE-WISE DAILY ATTENDANCE REPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Faculty"
    ws["B2"] = f"{_safe(getattr(faculty, 'name', ''))} ({_safe(getattr(faculty, 'faculty_id', ''))})"
    ws["A3"] = "Year / Semester"
    ws["B3"] = f"{actual_year} / {actual_semester}"
    ws["A4"] = "Section"
    ws["B4"] = f"{actual_section}"
    ws["C2"] = "Academic Year"
    ws["D2"] = f"{academic_year}"
    ws["C3"] = "Date Range"
    ws["D3"] = f"{date_from or '...'} to {date_to or '...'}"
    ws["C4"] = "Students"
    ws["D4"] = str(len(students_sorted))

    for cell in ["A2", "A3", "A4", "C2", "C3", "C4"]:
        ws[cell].font = bold_font

    header_row_1 = 6
    header_row_2 = 7

    ws.merge_cells(start_row=header_row_1, start_column=1, end_row=header_row_2, end_column=1)
    ws.merge_cells(start_row=header_row_1, start_column=2, end_row=header_row_2, end_column=2)
    ws.merge_cells(start_row=header_row_1, start_column=3, end_row=header_row_2, end_column=3)

    ws.cell(header_row_1, 1).value = "S.No"
    ws.cell(header_row_1, 2).value = "Reg No"
    ws.cell(header_row_1, 3).value = "Student Name"

    current_col = 4
    date_column_map = []

    for dt in sorted_dates:
        start_col = current_col

        ws.cell(header_row_2, current_col).value = "Morning"
        current_col += 1
        ws.cell(header_row_2, current_col).value = "Afternoon"
        current_col += 1
        ws.cell(header_row_2, current_col).value = "Full Day"
        current_col += 1

        end_col = current_col - 1

        ws.merge_cells(start_row=header_row_1, start_column=start_col, end_row=header_row_1, end_column=end_col)
        ws.cell(header_row_1, start_col).value = dt.strftime("%d-%m-%Y")
        date_column_map.append(dt)

    for row_num in [header_row_1, header_row_2]:
        for col in range(1, current_col):
            cell = ws.cell(row_num, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    data_row = 8
    for idx, student in enumerate(students_sorted, start=1):
        ws.cell(data_row, 1).value = idx
        ws.cell(data_row, 2).value = _safe(getattr(student, "reg_no", ""))
        ws.cell(data_row, 3).value = _safe(getattr(student, "name", ""))

        ws.cell(data_row, 1).alignment = center_align
        ws.cell(data_row, 2).alignment = center_align
        ws.cell(data_row, 3).alignment = left_align

        for fixed_col in [1, 2, 3]:
            ws.cell(data_row, fixed_col).border = thin_border

        col_ptr = 4
        for dt in date_column_map:
            rec = grouped[student.id].get(dt, {})

            for key in ["morning", "afternoon", "full_day"]:
                value = rec.get(key, "-")
                cell = ws.cell(data_row, col_ptr)
                cell.value = _safe(value)
                cell.alignment = center_align
                cell.border = thin_border

                if str(value).strip().lower() == "absent":
                    cell.fill = absent_fill

                col_ptr += 1

        data_row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28

    for col in range(4, current_col):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = "D8"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Class_Attendance_Datewise_{actual_year}_{actual_semester}_{actual_section}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response




from django.db.models import Sum


from faculty_management.models import general_information
from examination_management.models import *


from faculty_management.models import general_information


# import your models
import logging


logger = logging.getLogger(__name__)
    


# @check_permission("Internalmarks")
# def internalmarks(request):
#     # ---------- defaults ----------
#     students = None
#     selected_title = ''
#     selected_course_code = ''
#     selected_department = ''
#     selected_department_id = None
#     selected_batch = ''
#     selected_section = ''
#     selected_iat = ''
#     selected_model_lab = ''

#     selected_degree_code = ''
#     selected_degree_name = ''
#     selected_degree_obj = None
#     assessments_qs = Assessments.objects.none()

#     # ---------- Assigned subjects ----------
#     faculty_id = request.user.Employee_id
#     faculty = general_information.objects.filter(faculty_id=faculty_id).first()

#     assigned_subjects = (
#         AssignSubjectFaculty.objects
#         .filter(faculty=faculty, is_active=True)
#         .select_related("course", "department__degree")
#     )

#     # Attach department + degree details for template data-* and compute existing experiments
#     for subject in assigned_subjects:
#         dept = getattr(subject, "department", None)
#         if dept:
#             subject.department_name = getattr(dept, "Department", "") or ""
#             subject.department_code = getattr(dept, "Department_code", "") or ""
#             deg = getattr(dept, "degree", None)
#             subject.degree_id = deg.id if deg else None
#             subject.degree_code = getattr(deg, "degree_code", "") or ""
#             subject.degree_name = getattr(deg, "degree", "") or ""
#         else:
#             subject.department_name = ''
#             subject.department_code = ''
#             subject.degree_id = None
#             subject.degree_code = ''
#             subject.degree_name = ''

#         experiments = []
#         try:
#             course_code = getattr(subject.course, "course_code", "") or ""
#             batch = getattr(subject, "batch", "") or ""
#             section = getattr(subject, "section", "") or ""
#             department = dept

#             course_obj = None
#             if course_code:
#                 course_obj = (
#                     Course.objects.filter(course_code=course_code, department=department)
#                     .order_by("-id").first()
#                     or Course.objects.filter(course_code=course_code).order_by("-id").first()
#                 )

#                 course_hours = (
#                     CourseHours.objects
#                     .filter(course=course_obj)
#                     .select_related("hour_config")
#                     .order_by("-id")
#                     .first()
#                 )
#             else:
#                 course_hours = None

#             subject.has_theory = False
#             subject.has_practical = False
#             lecture_hpwk = 0
#             tutorial_hpwk = 0
#             lab_hpwk = 0

#             if course_hours and course_hours.hour_config:
#                 lecture_hpwk = int(course_hours.hour_config.lecture_hours or 0)
#                 tutorial_hpwk = int(course_hours.hour_config.tutorial_hours or 0)
#                 lab_hpwk = int(course_hours.hour_config.laboratory_hours or 0)

#             if lecture_hpwk == 0 and tutorial_hpwk == 0 and lab_hpwk == 0 and course_hours:
#                 try:
#                     lecture_hpwk = int(str(getattr(course_hours, "leture_npwk", 0) or 0).strip() or 0)
#                 except (TypeError, ValueError):
#                     lecture_hpwk = 0
#                 try:
#                     tutorial_hpwk = int(str(getattr(course_hours, "tutorial_npwk", 0) or 0).strip() or 0)
#                 except (TypeError, ValueError):
#                     tutorial_hpwk = 0
#                 try:
#                     lab_hpwk = int(str(getattr(course_hours, "laboratory_npwk", 0) or 0).strip() or 0)
#                 except (TypeError, ValueError):
#                     lab_hpwk = 0

#             subject.has_theory = (lecture_hpwk > 0) or (tutorial_hpwk > 0)
#             subject.has_practical = (lab_hpwk > 0)

#             if not subject.has_theory and not subject.has_practical and course_obj:
#                 subject.has_theory = True

#             if str(faculty_id) == "1624" and not subject.has_theory and not subject.has_practical:
#                 subject.has_theory = True

#             logger.debug(
#                 "internalmarks: subject=%s course_code=%s resolved_course=%s",
#                 getattr(subject, "id", "<no-id>"),
#                 course_code,
#                 getattr(course_obj, "id", None)
#             )

#             if course_obj:
#                 enrollments_qs = CourseEnrollment.objects.filter(course_id=course_obj.id)

#                 if batch:
#                     enrollments_qs = enrollments_qs.filter(batch=batch)
#                 if section:
#                     enrollments_qs = enrollments_qs.filter(section=section)

#                 enrollment_ids = list(enrollments_qs.values_list("id", flat=True))

#                 logger.debug(
#                     "internalmarks: enrollment_ids=%s for course_id=%s",
#                     enrollment_ids,
#                     course_obj.id
#                 )

#                 if enrollment_ids:
#                     raw_eno_list = list(
#                         experiment_marks.objects
#                         .filter(courses_id__in=enrollment_ids)
#                         .exclude(experiment_no__isnull=True)
#                         .values_list("experiment_no", flat=True)
#                     )

#                     logger.debug("internalmarks: raw_experiment_numbers=%s", raw_eno_list)

#                     eno_set = set()
#                     for eno in raw_eno_list:
#                         if eno is None:
#                             continue
#                         try:
#                             eno_int = int(str(eno).strip())
#                             if eno_int > 0:
#                                 eno_set.add(eno_int)
#                         except Exception:
#                             logger.debug(
#                                 "internalmarks: skipping_non_numeric_experiment_no=%s",
#                                 eno
#                             )

#                     eno_list = sorted(eno_set)
#                     logger.debug("internalmarks: final_experiment_numbers=%s", eno_list)

#                     subject.experiments_json = json.dumps(eno_list)
#                 else:
#                     subject.experiments_json = json.dumps([])
#             else:
#                 subject.experiments_json = json.dumps([])

#         except Exception:
#             logger.exception(
#                 "Error computing experiments for subject %s",
#                 getattr(subject, "id", None)
#             )
#             subject.experiments_json = json.dumps([])

#     # ---------- Build degree -> internal_iats + assessments + mlabs map ----------
#     degree_ids = []
#     degree_map = {}

#     for s in assigned_subjects:
#         dept = getattr(s, "department", None)
#         deg = getattr(dept, "degree", None) if dept else None
#         if deg:
#             degree_map[deg.id] = deg
#             if deg.id not in degree_ids:
#                 degree_ids.append(deg.id)

#     assessments_rows = (
#         Assessments.objects
#         .filter(Q(degree_id__in=degree_ids) | Q(degree__isnull=True))
#         .values('id', 'assessment_name', 'question_paper_required', 'degree_id')
#     )

#     modellab_rows = (
#         ModelLab.objects
#         .filter(Q(degree_id__in=degree_ids) | Q(degree__isnull=True))
#         .values('id', 'model_lab_name', 'degree_id', 'internal_assessment_id')
#     )

#     internal_iat_rows = (
#         InternalAssessment.objects
#         .filter(Q(degree_id__in=degree_ids) | Q(degree__isnull=True))
#         .values('id', 'iat', 'degree_id')
#     )

#     assessments_by_degree = {}

#     for row in assessments_rows:
#         key = row['degree_id'] if row['degree_id'] is not None else "global"
#         assessments_by_degree.setdefault(key, {"internal_iats": [], "assessments": [], "mlabs": []})
#         assessments_by_degree[key]["assessments"].append({
#             "id": row['id'],
#             "name": row['assessment_name'] or "",
#             "qp": bool(row['question_paper_required']),
#         })

#     for row in modellab_rows:
#         key = row['degree_id'] if row['degree_id'] is not None else "global"
#         assessments_by_degree.setdefault(key, {"internal_iats": [], "assessments": [], "mlabs": []})
#         name = (row['model_lab_name'] or "").strip()
#         if name:
#             assessments_by_degree[key]["mlabs"].append({
#                 "id": row['id'],
#                 "name": name,
#                 "internal_iat_id": row['internal_assessment_id'],
#             })

#     for row in internal_iat_rows:
#         key = row['degree_id'] if row['degree_id'] is not None else "global"
#         assessments_by_degree.setdefault(key, {"internal_iats": [], "assessments": [], "mlabs": []})
#         iat_name = (row['iat'] or "").strip()
#         if iat_name:
#             assessments_by_degree[key]["internal_iats"].append({
#                 "id": row['id'],
#                 "name": iat_name,
#             })

#     for key, payload in assessments_by_degree.items():
#         mlabs = payload.get("mlabs", [])
#         seen_ml = {}
#         deduped_ml = []
#         for m in mlabs:
#             nm = (m.get("name") or "").strip()
#             keypair = (nm.lower(), m.get("internal_iat_id"))
#             if nm and keypair not in seen_ml:
#                 seen_ml[keypair] = True
#                 deduped_ml.append({
#                     "id": m.get("id"),
#                     "name": nm,
#                     "internal_iat_id": m.get("internal_iat_id"),
#                 })
#         payload["mlabs"] = sorted(
#             deduped_ml,
#             key=lambda x: (x["name"].lower(), str(x["internal_iat_id"] or ""))
#         )

#         iiats = payload.get("internal_iats", [])
#         seen_i = {}
#         deduped_i = []
#         for it in iiats:
#             nm = (it.get("name") or "").strip()
#             keypair = (nm.lower(), it.get("id"))
#             if nm and keypair not in seen_i:
#                 seen_i[keypair] = True
#                 deduped_i.append({"id": it.get("id"), "name": nm})
#         payload["internal_iats"] = sorted(deduped_i, key=lambda x: x["name"].lower())

#     assessments_by_degree_json = json.dumps(assessments_by_degree)

#     # ---------- CO + Bloom's mapping master data ----------
#     # Match CO by degree -> regulation
#         # ---------- CO + Bloom's mapping master data ----------
#     # CO should match ONLY by regulation
#     regulation_values = set()

#     for subject in assigned_subjects:
#         regulation = getattr(subject, "regulation", "") or ""
#         regulation = str(regulation).strip()
#         if regulation:
#             regulation_values.add(regulation)

#     print("CO DEBUG => regulation_values from assigned_subjects:", regulation_values)

#     all_co_regs = list(
#         CourseOutcome.objects
#         .exclude(regulation__isnull=True)
#         .exclude(regulation__exact='')
#         .values_list("regulation", flat=True)
#         .distinct()
#     )
#     print("CO DEBUG => regulations available in CourseOutcome table:", all_co_regs)

#     co_by_regulation = {}

#     for regulation in regulation_values:
#         print("CO DEBUG => checking regulation:", regulation)

#         co_rows_qs = (
#             CourseOutcome.objects
#             .filter(regulation__iexact=regulation)
#             .exclude(co_code__isnull=True)
#             .exclude(co_code__exact='')
#             .values("id", "co_code", "co_name", "regulation")
#             .order_by("co_code")
#         )

#         co_rows = list(co_rows_qs)

#         print("CO DEBUG => matched rows count for regulation", regulation, "=", len(co_rows))
#         print("CO DEBUG => matched rows for regulation", regulation, "=", co_rows)

#         co_by_regulation[regulation] = [
#             {
#                 "id": row["id"],
#                 "code": row["co_code"] or "",
#                 "name": row["co_name"] or "",
#                 "regulation": row["regulation"] or "",
#             }
#             for row in co_rows
#         ]

#     print("CO DEBUG => final co_by_regulation =", co_by_regulation)

#     blooms_levels = list(
#         BloomsLevel.objects
#         .exclude(level_code__isnull=True)
#         .exclude(level_code__exact='')
#         .values('id', 'level_code', 'description')
#         .order_by('level_code')
#     )

#     co_mapping_json = json.dumps(co_by_regulation)

#     blooms_mapping_json = json.dumps([
#         {
#             "id": row["id"],
#             "code": row["level_code"] or "",
#             "name": row["description"] or "",
#         }
#         for row in blooms_levels
#     ])

#     # ---------- POST handling (theory) ----------
#     if request.method == "POST":
#         selected_course_id = request.POST.get("course_id")
#         selected_title = request.POST.get("course_title") or ''
#         selected_course_code = request.POST.get("course_code") or ''
#         selected_department = request.POST.get("department") or ''
#         selected_department_id = request.POST.get("department_id") or None
#         selected_batch = request.POST.get("batch") or ''
#         selected_section = request.POST.get("section") or ''
#         selected_iat = request.POST.get("iat") or ''
#         selected_model_lab = request.POST.get("model_lab") or ''

#         department = None
#         if selected_department_id:
#             department = (
#                 Add_Department.objects
#                 .select_related("degree")
#                 .filter(id=selected_department_id)
#                 .first()
#             )
#         else:
#             try:
#                 _, department_name = selected_department.split(" ", 1)
#             except ValueError:
#                 department_name = selected_department

#             if department_name:
#                 department = (
#                     Add_Department.objects
#                     .select_related("degree")
#                     .filter(Department=department_name)
#                     .first()
#                 )

#         course_obj = None
#         if selected_course_id:
#             course_obj = Course.objects.filter(id=selected_course_id).first()

#         if department and selected_iat and course_obj:
#             enrollments = (
#                 CourseEnrollment.objects
#                 .select_related("student")
#                 .filter(
#                     department=department,
#                     course_id=course_obj.id,
#                     batch=selected_batch,
#                     section=selected_section,
#                     enroll=True
#                 )
#                 .order_by("student__reg_no")
#             )

#             faculty_enrollments = enrollments.filter(faculty=faculty)
#             if faculty_enrollments.exists():
#                 enrollments = faculty_enrollments

#             students = []
#             for e in enrollments:
#                 if e.student:
#                     students.append({
#                         "reg_no": e.student.reg_no,
#                         "name": e.student.name,
#                         "department": e.student.department_id,
#                     })
#         else:
#             students = []

#         if students:
#             dept_ids = [s['department'] for s in students if s.get('department')]
#             dept_map = {
#                 d.id: d for d in Add_Department.objects.select_related("degree").filter(id__in=dept_ids)
#             }
#             for student in students:
#                 dept_obj = dept_map.get(student.get('department'))
#                 if dept_obj:
#                     student['department_code'] = getattr(dept_obj, "Department_code", "") or ''
#                     student['department_name'] = getattr(dept_obj, "Department", "") or ''
#                     deg = getattr(dept_obj, "degree", None)
#                     student['degree_id'] = deg.id if deg else None
#                     student['degree_code'] = getattr(deg, "degree_code", "") if deg else ""
#                     student['degree_name'] = getattr(deg, "degree", "") if deg else ""
#                 else:
#                     student['department_code'] = ''
#                     student['department_name'] = ''
#                     student['degree_id'] = None
#                     student['degree_code'] = ''
#                     student['degree_name'] = ''

#         dept_for_selected = None
#         if selected_department_id:
#             dept_for_selected = (
#                 Add_Department.objects
#                 .select_related("degree")
#                 .filter(id=selected_department_id)
#                 .first()
#             )
#         else:
#             try:
#                 _, department_name = selected_department.split(" ", 1)
#                 dept_for_selected = (
#                     Add_Department.objects
#                     .select_related("degree")
#                     .filter(Department=department_name)
#                     .first()
#                 )
#             except ValueError:
#                 if selected_department:
#                     dept_for_selected = (
#                         Add_Department.objects
#                         .select_related("degree")
#                         .filter(Department=selected_department)
#                         .first()
#                     )

#         if dept_for_selected and getattr(dept_for_selected, "degree", None):
#             selected_degree_obj = dept_for_selected.degree
#             selected_degree_code = getattr(selected_degree_obj, "degree_code", "") or ""
#             selected_degree_name = getattr(selected_degree_obj, "degree", "") or ""

#             assessments_qs = (
#                 Assessments.objects
#                 .filter(
#                     Q(degree=selected_degree_obj) |
#                     Q(degree__isnull=True) |
#                     Q(degree__degree_code=selected_degree_code) |
#                     Q(degree__degree=selected_degree_name),
#                     question_paper_required=True
#                 )
#                 .distinct()
#             )
#         else:
#             assessments_qs = Assessments.objects.none()

#     # ---------- Context ----------
#     context = {
#         "assigned_subjects": assigned_subjects,
#         "students": students,
#         "selected_iat": selected_iat,
#         "selected_course_code": selected_course_code,
#         "selected_course_title": selected_title,
#         "selected_department": selected_department,
#         "selected_batch": selected_batch,
#         "selected_section": selected_section,
#         "selected_degree_code": selected_degree_code,
#         "selected_degree_name": selected_degree_name,
#         "assessments": assessments_qs,
#         "assessments_by_degree_json": assessments_by_degree_json,
#         "selected_model_lab": selected_model_lab,
#         "selected_department_id": selected_department_id,
#         "co_mapping_json": co_mapping_json,
#         "blooms_mapping_json": blooms_mapping_json,
#     }

#     return render(request, "faculty_management/internal_mark_entry.html", context)




# from django.db import transaction
# from django.db.models import Q, Case, When, IntegerField
# from django.http import JsonResponse
# from django.views.decorators.http import require_POST
# from django.views.decorators.csrf import (
#     csrf_exempt,
# )  # only if you cannot send CSRF; otherwise don't use
# from django.core.exceptions import ObjectDoesNotExist
# from django.db import transaction
# from django.db.models import Case, When, IntegerField
# from django.http import JsonResponse
# from django.views.decorators.http import require_POST
# import logging


# logger = logging.getLogger(__name__)


# @require_POST
# def map_experiment_iat(request):
#     """
#     Instantly map a chosen InternalAssessment (by name) to all experiment_marks
#     for the selected class (course_code + dept + batch + section) and experiment_no.
#     If no experiment_marks exist yet for enrolled students, placeholder rows are created.
#     """
#     try:
#         course_code = (request.POST.get("course_code") or "").strip()
#         department_id = request.POST.get("department_id")
#         batch = (request.POST.get("batch") or "").strip()
#         section = (request.POST.get("section") or "").strip()
#         degree_id = request.POST.get(
#             "degree_id"
#         )  # not mandatory for lookup, but passed along
#         exp_no_raw = request.POST.get("experiment_no") or "0"
#         iat_name = (request.POST.get("iat_name") or "").strip()

#         try:
#             exp_no = int(exp_no_raw)
#         except (TypeError, ValueError):
#             exp_no = 0

#         if not (course_code and department_id and exp_no > 0 and iat_name):
#             return JsonResponse(
#                 {"ok": False, "error": "Missing required parameters."}, status=400
#             )

#         # Resolve Department (+ Degree)
#         department = (
#             Add_Department.objects.select_related("degree")
#             .filter(id=department_id)
#             .first()
#         )
#         if not department:
#             return JsonResponse(
#                 {"ok": False, "error": "Department not found."}, status=404
#             )

#         degree = getattr(department, "degree", None)

#         # Resolve Course (prefer department-scoped, then global by code)
#         course_obj = (
#             Course.objects.filter(course_code=course_code, department=department)
#             .order_by("-id")
#             .first()
#             or Course.objects.filter(course_code=course_code).order_by("-id").first()
#         )
#         if not course_obj:
#             return JsonResponse({"ok": False, "error": "Course not found."}, status=404)

#         # Enrollments for this class (+ optional batch/section filters)
#         enrollments_qs = CourseEnrollment.objects.filter(course_id=course_obj.id)
#         if batch:
#             enrollments_qs = enrollments_qs.filter(batch=batch)
#         if section:
#             enrollments_qs = enrollments_qs.filter(section=section)

#         enrollments = list(
#             enrollments_qs.select_related("student").values("id", "student_id")
#         )
#         if not enrollments:
#             return JsonResponse(
#                 {"ok": False, "error": "No enrollments for this class."}, status=404
#             )

#         enrollment_ids = [e["id"] for e in enrollments]
#         # student_ids = [e["student_id"] for e in enrollments if e["student_id"]]  # not strictly needed

#         # Resolve InternalAssessment by name; prefer degree-specific over global
#         qs = (
#             InternalAssessment.objects.filter(iat=iat_name)
#             .annotate(
#                 prio=Case(
#                     When(degree_id=degree.id if degree else None, then=0),
#                     When(degree__isnull=True, then=1),
#                     default=2,
#                     output_field=IntegerField(),
#                 )
#             )
#             .order_by("prio", "id")
#         )

#         ia = qs.first()
#         if not ia:
#             return JsonResponse(
#                 {"ok": False, "error": "InternalAssessment not found."}, status=404
#             )

#         with transaction.atomic():
#             # 1) Update any existing rows for this experiment
#             updated = experiment_marks.objects.filter(
#                 courses_id__in=enrollment_ids, experiment_no=exp_no
#             ).update(assessment_id=ia.id)  # IMPORTANT: _id for bulk update

#             # 2) Create placeholders for missing students so mapping is complete
#             existing_pairs = set(
#                 experiment_marks.objects.filter(
#                     courses_id__in=enrollment_ids, experiment_no=exp_no
#                 ).values_list("student_id", "courses_id")
#             )

#             to_create = []
#             for e in enrollments:
#                 pair = (e["student_id"], e["id"])
#                 if pair not in existing_pairs:
#                     to_create.append(
#                         experiment_marks(
#                             student_id=e["student_id"],
#                             courses_id=e["id"],
#                             co=None,  # set if you track COs later
#                             assessment=ia,  # instance is fine for create()
#                             work_program=0,
#                             observation=0,
#                             record=0,
#                             experiment_no=exp_no,
#                         )
#                     )

#             if to_create:
#                 experiment_marks.objects.bulk_create(to_create, batch_size=500)

#         return JsonResponse(
#             {
#                 "ok": True,
#                 "updated_rows": updated,
#                 "created_rows": len(to_create),
#                 "iat_id": ia.id,
#                 "experiment_no": exp_no,
#             }
#         )

#     except Exception as exc:
#         logger.exception("map_experiment_iat failed: %s", exc)
#         return JsonResponse({"ok": False, "error": "Server error."}, status=500)


# from django.views.decorators.http import require_POST, require_GET

# logger = logging.getLogger(__name__)


# @require_GET
# def get_experiment_iat_mapping(request):
#     """
#     Return the saved IAT mapping per experiment_no for the given class
#     (course_code + dept + batch + section). If multiple IATs exist for an
#     experiment among students, we return the *majority* IAT for that experiment.
#     Response: { ok: True, mapping: { "1": "IAT1", "2": "IAT2", ... } }
#     """
#     try:
#         course_code = (request.GET.get("course_code") or "").strip()
#         department_id = request.GET.get("department_id")
#         batch = (request.GET.get("batch") or "").strip()
#         section = (request.GET.get("section") or "").strip()

#         if not (course_code and department_id):
#             return JsonResponse(
#                 {"ok": False, "error": "Missing required parameters."}, status=400
#             )

#         # Resolve Department (+ Degree)
#         department = (
#             Add_Department.objects.select_related("degree")
#             .filter(id=department_id)
#             .first()
#         )
#         if not department:
#             return JsonResponse(
#                 {"ok": False, "error": "Department not found."}, status=404
#             )

#         # Resolve Course (prefer department-scoped, then fallback global by code)
#         course_obj = (
#             Course.objects.filter(course_code=course_code, department=department)
#             .order_by("-id")
#             .first()
#             or Course.objects.filter(course_code=course_code).order_by("-id").first()
#         )
#         if not course_obj:
#             return JsonResponse({"ok": False, "error": "Course not found."}, status=404)

#         # Enrollments for this class (+ optional batch/section filters)
#         enrollments_qs = CourseEnrollment.objects.filter(course_id=course_obj.id)
#         if batch:
#             enrollments_qs = enrollments_qs.filter(batch=batch)
#         if section:
#             enrollments_qs = enrollments_qs.filter(section=section)

#         enrollment_ids = list(enrollments_qs.values_list("id", flat=True))
#         if not enrollment_ids:
#             return JsonResponse({"ok": False, "mapping": {}})

#         # Count rows per (experiment_no, assessment__iat), pick majority per experiment
#         qs = (
#             experiment_marks.objects.filter(
#                 courses_id__in=enrollment_ids, assessment__isnull=False
#             )
#             .values("experiment_no", "assessment__iat")
#             .annotate(n=Count("id"))
#             .order_by("experiment_no", "-n", "assessment__iat")
#         )

#         mapping = {}
#         for row in qs:
#             exp = int(row["experiment_no"])
#             if exp not in mapping:
#                 mapping[exp] = (
#                     row["assessment__iat"] or ""
#                 )  # first row per exp is majority

#         # stringify keys for easy JSON use in FE
#         return JsonResponse(
#             {"ok": True, "mapping": {str(k): v for k, v in mapping.items()}}
#         )
#     except Exception as exc:
#         logger.exception("get_experiment_iat_mapping failed: %s", exc)
#         return JsonResponse({"ok": False, "error": "Server error."}, status=500)


from faculty_management.models import general_information
from examination_management.models import *


from faculty_management.models import general_information
import re


# import your models
import logging


logger = logging.getLogger(__name__)
    


@check_permission("Internalmarks")
def internalmarks(request):
    # ---------- defaults ----------
    students = None
    selected_title = ''
    selected_course_code = ''
    selected_department = ''
    selected_department_id = None
    selected_batch = ''
    selected_section = ''
    selected_iat = ''
    selected_model_lab = ''

    selected_degree_code = ''
    selected_degree_name = ''
    selected_degree_obj = None
    assessments_qs = Assessments.objects.none()

    # ---------- Assigned subjects ----------
    faculty_id = request.user.Employee_id
    faculty = general_information.objects.filter(faculty_id=faculty_id).first()

    assigned_subjects = (
        AssignSubjectFaculty.objects
        .filter(faculty=faculty, is_active=True)
        .select_related("course", "department__degree")
    )

    # Attach department + degree details for template data-* and compute existing experiments
    for subject in assigned_subjects:
        dept = getattr(subject, "department", None)
        if dept:
            subject.department_name = getattr(dept, "Department", "") or ""
            subject.department_code = getattr(dept, "Department_code", "") or ""
            deg = getattr(dept, "degree", None)
            subject.degree_id = deg.id if deg else None
            subject.degree_code = getattr(deg, "degree_code", "") or ""
            subject.degree_name = getattr(deg, "degree", "") or ""
        else:
            subject.department_name = ''
            subject.department_code = ''
            subject.degree_id = None
            subject.degree_code = ''
            subject.degree_name = ''

        experiments = []
        try:
            course_code = getattr(subject.course, "course_code", "") or ""
            batch = getattr(subject, "batch", "") or ""
            section = getattr(subject, "section", "") or ""
            department = dept

            course_obj = None
            if course_code:
                course_obj = (
                    Course.objects.filter(course_code=course_code, department=department)
                    .order_by("-id").first()
                    or Course.objects.filter(course_code=course_code).order_by("-id").first()
                )

                course_hours = (
                    CourseHours.objects
                    .filter(course=course_obj)
                    .select_related("hour_config")
                    .order_by("-id")
                    .first()
                )
            else:
                course_hours = None

            subject.has_theory = False
            subject.has_practical = False
            lecture_hpwk = 0
            tutorial_hpwk = 0
            lab_hpwk = 0

            if course_hours and course_hours.hour_config:
                lecture_hpwk = int(course_hours.hour_config.lecture_hours or 0)
                tutorial_hpwk = int(course_hours.hour_config.tutorial_hours or 0)
                lab_hpwk = int(course_hours.hour_config.laboratory_hours or 0)

            if lecture_hpwk == 0 and tutorial_hpwk == 0 and lab_hpwk == 0 and course_hours:
                try:
                    lecture_hpwk = int(str(getattr(course_hours, "leture_npwk", 0) or 0).strip() or 0)
                except (TypeError, ValueError):
                    lecture_hpwk = 0
                try:
                    tutorial_hpwk = int(str(getattr(course_hours, "tutorial_npwk", 0) or 0).strip() or 0)
                except (TypeError, ValueError):
                    tutorial_hpwk = 0
                try:
                    lab_hpwk = int(str(getattr(course_hours, "laboratory_npwk", 0) or 0).strip() or 0)
                except (TypeError, ValueError):
                    lab_hpwk = 0

            subject.has_theory = (lecture_hpwk > 0) or (tutorial_hpwk > 0)
            subject.has_practical = (lab_hpwk > 0)

            if not subject.has_theory and not subject.has_practical and course_obj:
                subject.has_theory = True

            if str(faculty_id) == "1624" and not subject.has_theory and not subject.has_practical:
                subject.has_theory = True

            logger.debug(
                "internalmarks: subject=%s course_code=%s resolved_course=%s",
                getattr(subject, "id", "<no-id>"),
                course_code,
                getattr(course_obj, "id", None)
            )

            if course_obj:
                enrollments_qs = CourseEnrollment.objects.filter(course_id=course_obj.id)

                if batch:
                    enrollments_qs = enrollments_qs.filter(batch=batch)
                if section:
                    enrollments_qs = enrollments_qs.filter(section=section)

                enrollment_ids = list(enrollments_qs.values_list("id", flat=True))

                logger.debug(
                    "internalmarks: enrollment_ids=%s for course_id=%s",
                    enrollment_ids,
                    course_obj.id
                )

                if enrollment_ids:
                    raw_eno_list = list(
                        experiment_marks.objects
                        .filter(courses_id__in=enrollment_ids)
                        .exclude(experiment_no__isnull=True)
                        .values_list("experiment_no", flat=True)
                    )

                    logger.debug("internalmarks: raw_experiment_numbers=%s", raw_eno_list)

                    eno_set = set()
                    for eno in raw_eno_list:
                        if eno is None:
                            continue
                        try:
                            eno_int = int(str(eno).strip())
                            if eno_int > 0:
                                eno_set.add(eno_int)
                        except Exception:
                            logger.debug(
                                "internalmarks: skipping_non_numeric_experiment_no=%s",
                                eno
                            )

                    eno_list = sorted(eno_set)
                    logger.debug("internalmarks: final_experiment_numbers=%s", eno_list)

                    subject.experiments_json = json.dumps(eno_list)
                else:
                    subject.experiments_json = json.dumps([])
            else:
                subject.experiments_json = json.dumps([])

        except Exception:
            logger.exception(
                "Error computing experiments for subject %s",
                getattr(subject, "id", None)
            )
            subject.experiments_json = json.dumps([])

    # ---------- Build degree -> internal_iats + assessments + mlabs map ----------
    degree_ids = []
    degree_map = {}

    for s in assigned_subjects:
        dept = getattr(s, "department", None)
        deg = getattr(dept, "degree", None) if dept else None
        if deg:
            degree_map[deg.id] = deg
            if deg.id not in degree_ids:
                degree_ids.append(deg.id)

    assessments_rows = (
        Assessments.objects
        .filter(Q(degree_id__in=degree_ids) | Q(degree__isnull=True))
        .values('id', 'assessment_name', 'question_paper_required', 'degree_id')
    )

    modellab_rows = (
        ModelLab.objects
        .filter(Q(degree_id__in=degree_ids) | Q(degree__isnull=True))
        .values('id', 'model_lab_name', 'degree_id', 'internal_assessment_id')
    )

    internal_iat_rows = (
        InternalAssessment.objects
        .filter(Q(degree_id__in=degree_ids) | Q(degree__isnull=True))
        .values('id', 'iat', 'degree_id')
    )

    assessments_by_degree = {}

    for row in assessments_rows:
        key = row['degree_id'] if row['degree_id'] is not None else "global"
        assessments_by_degree.setdefault(key, {"internal_iats": [], "assessments": [], "mlabs": []})
        assessments_by_degree[key]["assessments"].append({
            "id": row['id'],
            "name": row['assessment_name'] or "",
            "qp": bool(row['question_paper_required']),
        })

    for row in modellab_rows:
        key = row['degree_id'] if row['degree_id'] is not None else "global"
        assessments_by_degree.setdefault(key, {"internal_iats": [], "assessments": [], "mlabs": []})
        name = (row['model_lab_name'] or "").strip()
        if name:
            assessments_by_degree[key]["mlabs"].append({
                "id": row['id'],
                "name": name,
                "internal_iat_id": row['internal_assessment_id'],
            })

    for row in internal_iat_rows:
        key = row['degree_id'] if row['degree_id'] is not None else "global"
        assessments_by_degree.setdefault(key, {"internal_iats": [], "assessments": [], "mlabs": []})
        iat_name = (row['iat'] or "").strip()
        if iat_name:
            assessments_by_degree[key]["internal_iats"].append({
                "id": row['id'],
                "name": iat_name,
            })

    for key, payload in assessments_by_degree.items():
        mlabs = payload.get("mlabs", [])
        seen_ml = {}
        deduped_ml = []
        for m in mlabs:
            nm = (m.get("name") or "").strip()
            keypair = (nm.lower(), m.get("internal_iat_id"))
            if nm and keypair not in seen_ml:
                seen_ml[keypair] = True
                deduped_ml.append({
                    "id": m.get("id"),
                    "name": nm,
                    "internal_iat_id": m.get("internal_iat_id"),
                })
        payload["mlabs"] = sorted(
            deduped_ml,
            key=lambda x: (x["name"].lower(), str(x["internal_iat_id"] or ""))
        )

        iiats = payload.get("internal_iats", [])
        seen_i = {}
        deduped_i = []
        for it in iiats:
            nm = (it.get("name") or "").strip()
            keypair = (nm.lower(), it.get("id"))
            if nm and keypair not in seen_i:
                seen_i[keypair] = True
                deduped_i.append({"id": it.get("id"), "name": nm})
        payload["internal_iats"] = sorted(deduped_i, key=lambda x: x["name"].lower())

    assessments_by_degree_json = json.dumps(assessments_by_degree)

    # ---------- CO + Bloom's mapping master data ----------
    # Match CO by degree -> regulation
        # ---------- CO + Bloom's mapping master data ----------
    # CO should match ONLY by regulation
       

    regulation_values = set()

    for subject in assigned_subjects:
        regulation = getattr(subject, "regulation", "") or ""
        regulation = str(regulation).strip()
        if regulation:
            regulation_values.add(regulation)



    all_co_regs = list(
        CourseOutcome.objects
        .exclude(regulation__isnull=True)
        .exclude(regulation__exact='')
        .values_list("regulation", flat=True)
        .distinct()
    )
    

    def _extract_reg_year(value):
        value = str(value or "").strip()
        match = re.search(r"(20\d{2})", value)
        return match.group(1) if match else value

    co_by_regulation = {}

    for regulation in regulation_values:
        

        regulation_year = _extract_reg_year(regulation)
        

        co_rows_qs = (
            CourseOutcome.objects
            .filter(regulation__iexact=regulation_year)
            .exclude(co_code__isnull=True)
            .exclude(co_code__exact='')
            .values("id", "co_code", "co_name", "regulation")
            .order_by("co_code")
        )

        co_rows = list(co_rows_qs)



        co_by_regulation[regulation] = [
            {
                "id": row["id"],
                "code": row["co_code"] or "",
                "name": row["co_name"] or "",
                "regulation": row["regulation"] or "",
            }
            for row in co_rows
        ]

    
    blooms_levels = list(
        BloomsLevel.objects
        .exclude(level_code__isnull=True)
        .exclude(level_code__exact='')
        .values('id', 'level_code', 'description')
        .order_by('level_code')
    )

    co_mapping_json = json.dumps(co_by_regulation)

    blooms_mapping_json = json.dumps([
        {
            "id": row["id"],
            "code": row["level_code"] or "",
            "name": row["description"] or "",
        }
        for row in blooms_levels
    ])

    # ---------- POST handling (theory) ----------
    if request.method == "POST":
        selected_course_id = request.POST.get("course_id")
        selected_title = request.POST.get("course_title") or ''
        selected_course_code = request.POST.get("course_code") or ''
        selected_department = request.POST.get("department") or ''
        selected_department_id = request.POST.get("department_id") or None
        selected_batch = request.POST.get("batch") or ''
        selected_section = request.POST.get("section") or ''
        selected_iat = request.POST.get("iat") or ''
        selected_model_lab = request.POST.get("model_lab") or ''

        department = None
        if selected_department_id:
            department = (
                Add_Department.objects
                .select_related("degree")
                .filter(id=selected_department_id)
                .first()
            )
        else:
            try:
                _, department_name = selected_department.split(" ", 1)
            except ValueError:
                department_name = selected_department

            if department_name:
                department = (
                    Add_Department.objects
                    .select_related("degree")
                    .filter(Department=department_name)
                    .first()
                )

        course_obj = None
        if selected_course_id:
            course_obj = Course.objects.filter(id=selected_course_id).first()

        if department and selected_iat and course_obj:
            enrollments = (
                CourseEnrollment.objects
                .select_related("student")
                .filter(
                    department=department,
                    course_id=course_obj.id,
                    batch=selected_batch,
                    section=selected_section,
                    enroll=True, student__is_active=True
                )
                .order_by("student__reg_no")
            )

            faculty_enrollments = enrollments.filter(faculty=faculty)
            if faculty_enrollments.exists():
                enrollments = faculty_enrollments

            students = []
            for e in enrollments:
                if e.student:
                    students.append({
                        "reg_no": e.student.reg_no,
                        "name": e.student.name,
                        "department": e.student.department_id,
                    })
        else:
            students = []

        if students:
            dept_ids = [s['department'] for s in students if s.get('department')]
            dept_map = {
                d.id: d for d in Add_Department.objects.select_related("degree").filter(id__in=dept_ids)
            }
            for student in students:
                dept_obj = dept_map.get(student.get('department'))
                if dept_obj:
                    student['department_code'] = getattr(dept_obj, "Department_code", "") or ''
                    student['department_name'] = getattr(dept_obj, "Department", "") or ''
                    deg = getattr(dept_obj, "degree", None)
                    student['degree_id'] = deg.id if deg else None
                    student['degree_code'] = getattr(deg, "degree_code", "") if deg else ""
                    student['degree_name'] = getattr(deg, "degree", "") if deg else ""
                else:
                    student['department_code'] = ''
                    student['department_name'] = ''
                    student['degree_id'] = None
                    student['degree_code'] = ''
                    student['degree_name'] = ''

        dept_for_selected = None
        if selected_department_id:
            dept_for_selected = (
                Add_Department.objects
                .select_related("degree")
                .filter(id=selected_department_id)
                .first()
            )
        else:
            try:
                _, department_name = selected_department.split(" ", 1)
                dept_for_selected = (
                    Add_Department.objects
                    .select_related("degree")
                    .filter(Department=department_name)
                    .first()
                )
            except ValueError:
                if selected_department:
                    dept_for_selected = (
                        Add_Department.objects
                        .select_related("degree")
                        .filter(Department=selected_department)
                        .first()
                    )

        if dept_for_selected and getattr(dept_for_selected, "degree", None):
            selected_degree_obj = dept_for_selected.degree
            selected_degree_code = getattr(selected_degree_obj, "degree_code", "") or ""
            selected_degree_name = getattr(selected_degree_obj, "degree", "") or ""

            assessments_qs = (
                Assessments.objects
                .filter(
                    Q(degree=selected_degree_obj) |
                    Q(degree__isnull=True) |
                    Q(degree__degree_code=selected_degree_code) |
                    Q(degree__degree=selected_degree_name),
                    question_paper_required=True
                )
                .distinct()
            )
        else:
            assessments_qs = Assessments.objects.none()

    # ---------- Context ----------
    context = {
        "assigned_subjects": assigned_subjects,
        "students": students,
        "selected_iat": selected_iat,
        "selected_course_code": selected_course_code,
        "selected_course_title": selected_title,
        "selected_department": selected_department,
        "selected_batch": selected_batch,
        "selected_section": selected_section,
        "selected_degree_code": selected_degree_code,
        "selected_degree_name": selected_degree_name,
        "assessments": assessments_qs,
        "assessments_by_degree_json": assessments_by_degree_json,
        "selected_model_lab": selected_model_lab,
        "selected_department_id": selected_department_id,
        "co_mapping_json": co_mapping_json,
        "blooms_mapping_json": blooms_mapping_json,
    }

    return render(request, "faculty_management/internal_mark_entry.html", context)











from django.db import transaction
from django.db.models import Q, Case, When, IntegerField
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import (
    csrf_exempt,
)  # only if you cannot send CSRF; otherwise don't use
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.db.models import Case, When, IntegerField
from django.http import JsonResponse
from django.views.decorators.http import require_POST
import logging


logger = logging.getLogger(__name__)


@require_POST
def map_experiment_iat(request):
    """
    Map one IAT + multiple COs + multiple Bloom's levels to all experiment_marks
    rows for the selected class and experiment number.

    If rows do not yet exist for enrolled students, placeholder rows are created.
    """
    try:
        course_code = (request.POST.get("course_code") or "").strip()
        department_id = request.POST.get("department_id")
        batch = (request.POST.get("batch") or "").strip()
        section = (request.POST.get("section") or "").strip()
        exp_no_raw = request.POST.get("experiment_no") or "0"
        iat_name = (request.POST.get("iat_name") or "").strip()

        co_ids_raw = request.POST.get("co_ids", "")
        blooms_ids_raw = request.POST.get("blooms_ids", "")

        try:
            exp_no = int(exp_no_raw)
        except (TypeError, ValueError):
            exp_no = 0

        if not (course_code and department_id and exp_no > 0 and iat_name):
            return JsonResponse(
                {"ok": False, "error": "Missing required parameters."},
                status=400
            )

        co_ids = []
        if co_ids_raw:
            for x in co_ids_raw.split(","):
                x = x.strip()
                if x.isdigit():
                    co_ids.append(int(x))

        blooms_ids = []
        if blooms_ids_raw:
            for x in blooms_ids_raw.split(","):
                x = x.strip()
                if x.isdigit():
                    blooms_ids.append(int(x))

        department = (
            Add_Department.objects.select_related("degree")
            .filter(id=department_id)
            .first()
        )
        if not department:
            return JsonResponse(
                {"ok": False, "error": "Department not found."},
                status=404
            )

        degree = getattr(department, "degree", None)

        course_obj = (
            Course.objects.filter(course_code=course_code, department=department)
            .order_by("-id")
            .first()
            or Course.objects.filter(course_code=course_code).order_by("-id").first()
        )
        if not course_obj:
            return JsonResponse(
                {"ok": False, "error": "Course not found."},
                status=404
            )

        enrollments_qs = CourseEnrollment.objects.filter(course_id=course_obj.id)
        if batch:
            enrollments_qs = enrollments_qs.filter(batch=batch)
        if section:
            enrollments_qs = enrollments_qs.filter(section=section)

        enrollments = list(
            enrollments_qs.select_related("student").values("id", "student_id")
        )
        if not enrollments:
            return JsonResponse(
                {"ok": False, "error": "No enrollments for this class."},
                status=404
            )

        enrollment_ids = [e["id"] for e in enrollments]

        qs = (
            InternalAssessment.objects.filter(iat=iat_name)
            .annotate(
                prio=Case(
                    When(degree_id=degree.id if degree else None, then=0),
                    When(degree__isnull=True, then=1),
                    default=2,
                    output_field=IntegerField(),
                )
            )
            .order_by("prio", "id")
        )

        ia = qs.first()
        if not ia:
            return JsonResponse(
                {"ok": False, "error": "InternalAssessment not found."},
                status=404
            )

        co_objs = list(CourseOutcome.objects.filter(id__in=co_ids)) if co_ids else []
        blooms_objs = list(BloomsLevel.objects.filter(id__in=blooms_ids)) if blooms_ids else []

        with transaction.atomic():
            updated = experiment_marks.objects.filter(
                courses_id__in=enrollment_ids,
                experiment_no=exp_no
            ).update(assessment_id=ia.id)

            existing_pairs = set(
                experiment_marks.objects.filter(
                    courses_id__in=enrollment_ids,
                    experiment_no=exp_no
                ).values_list("student_id", "courses_id")
            )

            to_create = []
            for e in enrollments:
                pair = (e["student_id"], e["id"])
                if pair not in existing_pairs:
                    to_create.append(
                        experiment_marks(
                            student_id=e["student_id"],
                            courses_id=e["id"],
                            assessment=ia,
                            work_program=0,
                            observation=0,
                            record=0,
                            total=0,  # IMPORTANT
                            experiment_no=exp_no,
                        )
                    )

            if to_create:
                experiment_marks.objects.bulk_create(to_create, batch_size=500)

            all_marks = list(
                experiment_marks.objects.filter(
                    courses_id__in=enrollment_ids,
                    experiment_no=exp_no
                )
            )

            for mark in all_marks:
                if co_objs:
                    mark.cos.set(co_objs)
                else:
                    mark.cos.clear()

                if blooms_objs:
                    mark.blooms_levels.set(blooms_objs)
                else:
                    mark.blooms_levels.clear()

        return JsonResponse(
            {
                "ok": True,
                "updated_rows": updated,
                "created_rows": len(to_create),
                "iat_id": ia.id,
                "iat_name": ia.iat,
                "experiment_no": exp_no,
                "co_ids": co_ids,
                "blooms_ids": blooms_ids,
            }
        )

    except Exception as exc:
        logger.exception("map_experiment_iat failed: %s", exc)
        return JsonResponse(
            {"ok": False, "error": str(exc)},
            status=500
        )
from django.views.decorators.http import require_POST, require_GET

logger = logging.getLogger(__name__)


@require_GET
def get_experiment_iat_mapping(request):
    """
    Return saved mapping per experiment_no for the selected class.

    Response format:
    {
        "ok": True,
        "mapping": {
            "1": {
                "iat_name": "IAT1",
                "co_ids": [1, 2],
                "blooms_ids": [3, 4]
            },
            "2": {
                "iat_name": "IAT2",
                "co_ids": [],
                "blooms_ids": [1]
            }
        }
    }
    """
    try:
        course_code = (request.GET.get("course_code") or "").strip()
        department_id = request.GET.get("department_id")
        batch = (request.GET.get("batch") or "").strip()
        section = (request.GET.get("section") or "").strip()

        if not (course_code and department_id):
            return JsonResponse(
                {"ok": False, "error": "Missing required parameters."},
                status=400
            )

        department = (
            Add_Department.objects.select_related("degree")
            .filter(id=department_id)
            .first()
        )
        if not department:
            return JsonResponse(
                {"ok": False, "error": "Department not found."},
                status=404
            )

        course_obj = (
            Course.objects.filter(course_code=course_code, department=department)
            .order_by("-id")
            .first()
            or Course.objects.filter(course_code=course_code).order_by("-id").first()
        )
        if not course_obj:
            return JsonResponse(
                {"ok": False, "error": "Course not found."},
                status=404
            )

        enrollments_qs = CourseEnrollment.objects.filter(course_id=course_obj.id)
        if batch:
            enrollments_qs = enrollments_qs.filter(batch=batch)
        if section:
            enrollments_qs = enrollments_qs.filter(section=section)

        enrollment_ids = list(enrollments_qs.values_list("id", flat=True))
        if not enrollment_ids:
            return JsonResponse({"ok": True, "mapping": {}})

        all_marks = (
            experiment_marks.objects.filter(courses_id__in=enrollment_ids)
            .select_related("assessment")
            .prefetch_related("cos", "blooms_levels")
        )

        by_exp = defaultdict(list)
        for mark in all_marks:
            if mark.experiment_no:
                by_exp[int(mark.experiment_no)].append(mark)

        mapping = {}

        for exp_no, marks in by_exp.items():
            iat_counter = defaultdict(int)
            for mark in marks:
                iat_name = getattr(mark.assessment, "iat", "") if mark.assessment else ""
                if iat_name:
                    iat_counter[iat_name] += 1

            majority_iat = ""
            if iat_counter:
                majority_iat = sorted(
                    iat_counter.items(),
                    key=lambda x: (-x[1], x[0])
                )[0][0]

            co_ids = sorted({
                co.id
                for mark in marks
                for co in mark.cos.all()
            })

            blooms_ids = sorted({
                bl.id
                for mark in marks
                for bl in mark.blooms_levels.all()
            })

            mapping[str(exp_no)] = {
                "iat_name": majority_iat,
                "co_ids": co_ids,
                "blooms_ids": blooms_ids,
            }

        return JsonResponse({"ok": True, "mapping": mapping})

    except Exception as exc:
        logger.exception("get_experiment_iat_mapping failed: %s", exc)
        return JsonResponse(
            {"ok": False, "error": "Server error."},
            status=500
        )




# views.py

# views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction

from student_management.models import Add_Department
from user_accounts.models import StudentDetails
from course_management.models import Course, CourseEnrollment
from examination_management.models import experiment_marks, CourseOutcome
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.db.models import Q

from student_management.models import Add_Department
from user_accounts.models import StudentDetails
from course_management.models import Course, CourseEnrollment
from examination_management.models import (
    experiment_marks,
    CourseOutcome,
    InternalAssessment,
)

# from django.shortcuts import render, redirect
# from django.contrib import messages
# from django.db import transaction
# from django.db.models import Q

# @check_permission("Internalmarks")
# @student_management
# def enter_practical_page(request):
#     """
#     Practical marks entry page + saver for experiment_marks.
#     Persists Course via CourseEnrollment.
#     Supports CO selection per student (by regulation).
#     Accepts 'assessment_iat' (IA name) to map Experiment -> IA; saves FK on experiment_marks.assessment.
#     Observation rule:
#       - if entered observation <= 10 => convert to /20 using (mark / 10) * 20
#       - if entered observation > 10 => save as entered
#     """

#     def q_or_p(name, default=""):
#         return request.POST.get(name, request.GET.get(name, default))

#     # ---------- Query / POST params ----------
#     course_id = q_or_p("course_id") or None
#     course_code = q_or_p("course_code", "")
#     course_title = q_or_p("course_title", "")
#     selected_dep = q_or_p("selected_department") or q_or_p("department", "")
#     selected_dep_id = q_or_p("selected_department_id") or q_or_p("department_id") or None
#     batch = q_or_p("selected_batch") or q_or_p("batch", "")
#     section = q_or_p("selected_section") or q_or_p("section", "")
#     degree_id = q_or_p("degree_id")
#     degree_code = q_or_p("selected_degree_code") or q_or_p("degree_code", "")
#     degree_name = q_or_p("selected_degree_name") or q_or_p("degree_name", "")
#     experiment_no = (q_or_p("experiment_no") or "").strip()
#     assessment_iat = (q_or_p("assessment_iat") or "").strip()

#     # ---------- Resolve department ----------
#     department = None
#     if selected_dep_id:
#         department = (
#             Add_Department.objects.select_related("degree")
#             .filter(id=selected_dep_id)
#             .first()
#         )
#     else:
#         try:
#             _, dep_name = selected_dep.split(" ", 1)
#         except ValueError:
#             dep_name = selected_dep
#         if dep_name:
#             department = (
#                 Add_Department.objects.select_related("degree")
#                 .filter(Department=dep_name)
#                 .first()
#             )

#     # ---------- Fetch students with regulation ----------
#     students = []
#     cid = None
#     if course_id:
#         try:
#             cid = int(course_id)
#         except (TypeError, ValueError):
#             cid = None

#     if cid:
#         enrollments = CourseEnrollment.objects.select_related("student").filter(
#             course_id=cid,
#             batch=batch,
#             section=section,
#             enroll=True
#         ).order_by("student__reg_no")

#         for e in enrollments:
#             if not e.student:
#                 continue

#             stud = e.student

#             reg_str = ""
#             if hasattr(stud, "regulation") and stud.regulation:
#                 reg_str = str(getattr(stud.regulation, "year", stud.regulation))
#             elif hasattr(stud, "Regulation") and stud.Regulation:
#                 reg_str = str(stud.Regulation)

#             students.append({
#                 "id": stud.id,
#                 "reg_no": stud.reg_no,
#                 "name": stud.name,
#                 "department": stud.department_id,
#                 "regulation_str": reg_str or "",
#             })

#     # ---------- Decorate with department / degree ----------
#     for s in students:
#         dept_obj = (
#             Add_Department.objects.select_related("degree")
#             .filter(id=s["department"])
#             .first()
#         )
#         if dept_obj:
#             s["department_code"] = getattr(dept_obj, "Department_code", "") or ""
#             s["department_name"] = getattr(dept_obj, "Department", "") or ""
#             deg = getattr(dept_obj, "degree", None)
#             s["degree_id"] = deg.id if deg else None
#             s["degree_code"] = getattr(deg, "degree_code", "") if deg else ""
#             s["degree_name"] = getattr(deg, "degree", "") if deg else ""
#         else:
#             s["department_code"] = ""
#             s["department_name"] = ""
#             s["degree_id"] = None
#             s["degree_code"] = ""
#             s["degree_name"] = ""

#     # ---------- Resolve course ----------
#     course = None
#     if cid:
#         course = Course.objects.filter(id=cid).first()

#     if course is None and course_code:
#         course = (
#             Course.objects.filter(course_code=course_code, department=department)
#             .order_by("-id")
#             .first()
#             or Course.objects.filter(course_code=course_code).order_by("-id").first()
#         )

#     # ---------- Prefill prior marks + CO ----------
#     if course and experiment_no.isdigit() and int(experiment_no) > 0 and students:
#         exp_no_int = int(experiment_no)
#         student_ids = [s["id"] for s in students]

#         enrollments = CourseEnrollment.objects.filter(
#             student_id__in=student_ids,
#             course_id=course.id,
#             batch=batch or "",
#             section=section or "",
#         ).values("id", "student_id")

#         sid_to_enroll = {e["student_id"]: e["id"] for e in enrollments}

#         marks_qs = experiment_marks.objects.filter(
#             student_id__in=student_ids,
#             courses_id__in=sid_to_enroll.values(),
#             experiment_no=exp_no_int,
#         ).values(
#             "student_id",
#             "work_program",
#             "observation",
#             "record",
#             "total",
#             "co_id",
#             "assessment_id",
#         )

#         marks_map = {
#             m["student_id"]: {
#                 "pref_wp": m["work_program"],
#                 "pref_ob": m["observation"],
#                 "pref_rc": m["record"],
#                 "pref_total": m["total"],
#                 "pref_co_id": m["co_id"],
#                 "pref_assessment_id": m["assessment_id"],
#             }
#             for m in marks_qs
#         }

#         if not assessment_iat and marks_map:
#             any_assessment_id = next(
#                 (
#                     v.get("pref_assessment_id")
#                     for v in marks_map.values()
#                     if v.get("pref_assessment_id")
#                 ),
#                 None,
#             )
#             if any_assessment_id:
#                 ia_obj = InternalAssessment.objects.filter(id=any_assessment_id).first()
#                 if ia_obj and ia_obj.iat:
#                     assessment_iat = ia_obj.iat

#         for s in students:
#             s.update(marks_map.get(s["id"], {}))

#     # ---------- Default values if nothing exists ----------
#     for s in students:
#         if "pref_wp" not in s:
#             s["pref_wp"] = 40
#         if "pref_ob" not in s:
#             s["pref_ob"] = 20
#         if "pref_rc" not in s:
#             s["pref_rc"] = 15
#         if "pref_total" not in s:
#             s["pref_total"] = 75
#         if "pref_co_id" not in s:
#             s["pref_co_id"] = None

#     # ---------- Load CO options per regulation ----------
#     regs = {s.get("regulation_str", "") for s in students}
#     regs.discard("")

#     co_cache = {}
#     if regs:
#         for r in regs:
#             options = (
#                 CourseOutcome.objects.filter(regulation=r)
#                 .order_by("co_code")
#                 .values("id", "co_code", "co_name")
#             )
#             co_cache[r] = [
#                 {"id": o["id"], "label": f"{o['co_code']} - {o['co_name']}"}
#                 for o in options
#             ]

#     for s in students:
#         r = s.get("regulation_str", "")
#         s["co_options"] = co_cache.get(r, [])

#     # ---------- POST save ----------
#     if request.method == "POST":
#         if not course_code:
#             messages.error(request, "Missing course code; cannot save enrollments.")
#             return redirect(request.path)

#         if course is None:
#             course = Course.objects.create(
#                 course_code=course_code,
#                 title=course_title or "",
#                 department=department
#             )
#             course_id = course.id

#         # IA resolution
#         ia_obj = None
#         if assessment_iat:
#             ia_qs = InternalAssessment.objects.filter(iat=assessment_iat)
#             if department and getattr(department, "degree_id", None):
#                 ia_qs = ia_qs.filter(
#                     Q(degree_id=department.degree_id) | Q(degree__isnull=True)
#                 )
#             ia_obj = ia_qs.order_by("-degree_id").first()

#         WP_MAX, OB_MAX, RC_MAX, TOTAL_MAX = 40, 20, 15, 75

#         def to_int(val, max_v):
#             try:
#                 v = int(val)
#             except (TypeError, ValueError):
#                 v = 0
#             return 0 if v < 0 else (max_v if v > max_v else v)

#         reg_nos = request.POST.getlist("reg_nos[]")
#         if not reg_nos:
#             messages.warning(request, "No students to save.")
#             return redirect(request.path)

#         saved = updated = skipped = 0

#         with transaction.atomic():
#             for rn in reg_nos:
#                 student = (
#                     StudentDetails.objects.filter(
#                         reg_no=rn,
#                         **({"department": department} if department else {}),
#                         batch=batch,
#                         section=section,
#                     ).first()
#                     or StudentDetails.objects.filter(reg_no=rn).first()
#                 )
#                 if not student:
#                     skipped += 1
#                     continue

#                 enrollment, _ = CourseEnrollment.objects.get_or_create(
#                     student_id=student.id,
#                     course_id=course.id,
#                     batch=batch or "",
#                     section=section or "",
#                     defaults={"department_id": (department.id if department else None)},
#                 )

#                 desired_dept_id = department.id if department else None
#                 if enrollment.department_id != desired_dept_id:
#                     enrollment.department_id = desired_dept_id
#                     enrollment.save(update_fields=["department_id"])

#                 # Work Program: save as entered (/40)
#                 wp = to_int(request.POST.get(f"performance[{rn}]"), WP_MAX)

#                 # Observation rule
#                 raw_ob = request.POST.get(f"record[{rn}]")
#                 try:
#                     raw_ob = float(raw_ob)
#                 except (TypeError, ValueError):
#                     raw_ob = 0

#                 if raw_ob <= 10:
#                     ob = round((raw_ob / 10) * 20)
#                 else:
#                     ob = round(raw_ob)

#                 ob = min(max(ob, 0), OB_MAX)

#                 # Record: save as entered (/15)
#                 rc = to_int(request.POST.get(f"viva[{rn}]"), RC_MAX)

#                 total = min(wp + ob + rc, TOTAL_MAX)

#                 # CO selection
#                 co_val = (request.POST.get(f"co[{rn}]") or "").strip()
#                 co_obj = None
#                 if co_val.isdigit():
#                     co_obj = CourseOutcome.objects.filter(id=int(co_val)).first()

#                 lookup = {"student": student, "courses": enrollment}
#                 if experiment_no.isdigit() and int(experiment_no) > 0:
#                     lookup["experiment_no"] = int(experiment_no)

#                 obj, created = experiment_marks.objects.update_or_create(
#                     **lookup,
#                     defaults={
#                         "work_program": wp,
#                         "observation": ob,
#                         "record": rc,
#                         "total": total,
#                         "co": co_obj,
#                         "assessment": ia_obj,
#                     },
#                 )

#                 if created:
#                     saved += 1
#                 else:
#                     updated += 1

#         messages.success(request, f"Saved: {saved}, Updated: {updated}, Skipped: {skipped}")
#         return redirect(
#             f"{request.path}"
#             f"?course_id={course.id if course else (course_id or '')}"
#             f"&course_code={course_code}"
#             f"&course_title={course_title}"
#             f"&department_id={(department.id if department else '')}"
#             f"&batch={batch}"
#             f"&section={section}"
#             f"&degree_code={degree_code}"
#             f"&degree_name={degree_name}"
#             f"&experiment_no={experiment_no}"
#             f"&assessment_iat={assessment_iat}"
#         )

#     # ---------- Context ----------
#     context = {
#         "course_id": course.id if course else course_id,
#         "course_code": course_code,
#         "course_title": course_title,
#         "selected_department": selected_dep,
#         "selected_department_id": selected_dep_id,
#         "selected_batch": batch,
#         "selected_section": section,
#         "selected_degree_code": degree_code,
#         "selected_degree_name": degree_name,
#         "experiment_no": experiment_no,
#         "assessment_iat": assessment_iat,
#         "students": students,
#     }
#     return render(request, "faculty_management/practical_mark_entry.html", context)


from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction

from student_management.models import Add_Department
from user_accounts.models import StudentDetails
from course_management.models import Course, CourseEnrollment
from examination_management.models import experiment_marks, CourseOutcome
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.db.models import Q

from student_management.models import Add_Department
from user_accounts.models import StudentDetails
from course_management.models import Course, CourseEnrollment
from examination_management.models import (
    experiment_marks,
    CourseOutcome,
    InternalAssessment,
)

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.db.models import Q

@check_permission("Internalmarks")
@student_management
def enter_practical_page(request):
    """
    Practical marks entry page + saver for experiment_marks.
    Persists Course via CourseEnrollment.
    Accepts 'assessment_iat' (IA name) to map Experiment -> IA; saves FK on experiment_marks.assessment.

    Rules:
      Work Program:
        - if entered work_program < 25 => convert to /40 using (mark / 25) * 40
        - if entered work_program >= 25 => save as entered

      Observation:
        - if entered observation <= 10 => convert to /20 using (mark / 10) * 20
        - if entered observation > 10 => save as entered
    """

    def q_or_p(name, default=""):
        return request.POST.get(name, request.GET.get(name, default))

    # ---------- Query / POST params ----------
    course_id = q_or_p("course_id") or None
    course_code = q_or_p("course_code", "")
    course_title = q_or_p("course_title", "")
    selected_dep = q_or_p("selected_department") or q_or_p("department", "")
    selected_dep_id = q_or_p("selected_department_id") or q_or_p("department_id") or None
    batch = q_or_p("selected_batch") or q_or_p("batch", "")
    section = q_or_p("selected_section") or q_or_p("section", "")
    degree_code = q_or_p("selected_degree_code") or q_or_p("degree_code", "")
    degree_name = q_or_p("selected_degree_name") or q_or_p("degree_name", "")
    experiment_no = (q_or_p("experiment_no") or "").strip()
    assessment_iat = (q_or_p("assessment_iat") or q_or_p("iat") or "").strip()

    # ---------- Resolve department ----------
    department = None
    if selected_dep_id:
        department = (
            Add_Department.objects.select_related("degree")
            .filter(id=selected_dep_id)
            .first()
        )
    else:
        try:
            _, dep_name = selected_dep.split(" ", 1)
        except ValueError:
            dep_name = selected_dep
        if dep_name:
            department = (
                Add_Department.objects.select_related("degree")
                .filter(Department=dep_name)
                .first()
            )

    # ---------- Fetch students with regulation ----------
    students = []
    cid = None
    if course_id:
        try:
            cid = int(course_id)
        except (TypeError, ValueError):
            cid = None

    if cid:
        enrollments = CourseEnrollment.objects.select_related("student").filter(
            course_id=cid,
            batch=batch,
            section=section,
            enroll=True
        ).order_by("student__reg_no")

        for e in enrollments:
            if not e.student:
                continue

            stud = e.student

            reg_str = ""
            if hasattr(stud, "regulation") and stud.regulation:
                reg_str = str(getattr(stud.regulation, "year", stud.regulation))
            elif hasattr(stud, "Regulation") and stud.Regulation:
                reg_str = str(stud.Regulation)

            students.append({
                "id": stud.id,
                "reg_no": stud.reg_no,
                "name": stud.name,
                "department": stud.department_id,
                "regulation_str": reg_str or "",
            })

    # ---------- Decorate with department / degree ----------
    for s in students:
        dept_obj = (
            Add_Department.objects.select_related("degree")
            .filter(id=s["department"])
            .first()
        )
        if dept_obj:
            s["department_code"] = getattr(dept_obj, "Department_code", "") or ""
            s["department_name"] = getattr(dept_obj, "Department", "") or ""
            deg = getattr(dept_obj, "degree", None)
            s["degree_id"] = deg.id if deg else None
            s["degree_code"] = getattr(deg, "degree_code", "") if deg else ""
            s["degree_name"] = getattr(deg, "degree", "") if deg else ""
        else:
            s["department_code"] = ""
            s["department_name"] = ""
            s["degree_id"] = None
            s["degree_code"] = ""
            s["degree_name"] = ""

    # ---------- Resolve course ----------
    course = None
    if cid:
        course = Course.objects.filter(id=cid).first()

    if course is None and course_code:
        course = (
            Course.objects.filter(course_code=course_code, department=department)
            .order_by("-id")
            .first()
            or Course.objects.filter(course_code=course_code).order_by("-id").first()
        )

    # ---------- Prefill prior marks ----------
    if course and experiment_no.isdigit() and int(experiment_no) > 0 and students:
        exp_no_int = int(experiment_no)
        student_ids = [s["id"] for s in students]

        enrollments = CourseEnrollment.objects.filter(
            student_id__in=student_ids,
            course_id=course.id,
            batch=batch or "",
            section=section or "",
        ).values("id", "student_id")

        sid_to_enroll = {e["student_id"]: e["id"] for e in enrollments}

        marks_qs = experiment_marks.objects.filter(
            student_id__in=student_ids,
            courses_id__in=sid_to_enroll.values(),
            experiment_no=exp_no_int,
        ).values(
            "student_id",
            "work_program",
            "observation",
            "record",
            "total",
            "assessment_id",
        )

        marks_map = {
            m["student_id"]: {
                "pref_wp": m["work_program"],
                "pref_ob": m["observation"],
                "pref_rc": m["record"],
                "pref_total": m["total"],
                "pref_assessment_id": m["assessment_id"],
            }
            for m in marks_qs
        }

        if not assessment_iat and marks_map:
            any_assessment_id = next(
                (
                    v.get("pref_assessment_id")
                    for v in marks_map.values()
                    if v.get("pref_assessment_id")
                ),
                None,
            )
            if any_assessment_id:
                ia_obj = InternalAssessment.objects.filter(id=any_assessment_id).first()
                if ia_obj and ia_obj.iat:
                    assessment_iat = ia_obj.iat

        for s in students:
            s.update(marks_map.get(s["id"], {}))

    # ---------- Default values ----------
    for s in students:
        if "pref_wp" not in s:
            s["pref_wp"] = 40
        if "pref_ob" not in s:
            s["pref_ob"] = 20
        if "pref_rc" not in s:
            s["pref_rc"] = 15
        if "pref_total" not in s:
            s["pref_total"] = 75

    # ---------- POST save ----------
    if request.method == "POST":
        if not course_code:
            messages.error(request, "Missing course code; cannot save enrollments.")
            return redirect(request.path)

        if course is None:
            course = Course.objects.create(
                course_code=course_code,
                title=course_title or "",
                department=department
            )
            course_id = course.id

        ia_obj = None
        if assessment_iat:
            ia_qs = InternalAssessment.objects.filter(iat=assessment_iat)
            if department and getattr(department, "degree_id", None):
                ia_qs = ia_qs.filter(
                    Q(degree_id=department.degree_id) | Q(degree__isnull=True)
                )
            ia_obj = ia_qs.order_by("-degree_id").first()

        WP_MAX, OB_MAX, RC_MAX, TOTAL_MAX = 40, 20, 15, 75

        def to_float(val, default=0):
            try:
                return float(val)
            except (TypeError, ValueError):
                return default

        def clamp(val, min_v, max_v):
            if val < min_v:
                return min_v
            if val > max_v:
                return max_v
            return val

        reg_nos = request.POST.getlist("reg_nos[]")
        if not reg_nos:
            messages.warning(request, "No students to save.")
            return redirect(request.path)

        saved = updated = skipped = 0

        with transaction.atomic():
            for rn in reg_nos:
                student = (
                    StudentDetails.objects.filter(
                        reg_no=rn,
                        **({"department": department} if department else {}),
                        batch=batch,
                        section=section,
                    ).first()
                    or StudentDetails.objects.filter(reg_no=rn).first()
                )
                if not student:
                    skipped += 1
                    continue

                enrollment, _ = CourseEnrollment.objects.get_or_create(
                    student_id=student.id,
                    course_id=course.id,
                    batch=batch or "",
                    section=section or "",
                    defaults={"department_id": (department.id if department else None)},
                )

                desired_dept_id = department.id if department else None
                if enrollment.department_id != desired_dept_id:
                    enrollment.department_id = desired_dept_id
                    enrollment.save(update_fields=["department_id"])

                # ---------- Work Program conversion rule ----------
                raw_wp = to_float(request.POST.get(f"performance[{rn}]"), 0)
                raw_wp = clamp(raw_wp, 0, WP_MAX)

                if raw_wp <= 25:
                    wp = round((raw_wp / 25) * 40)
                else:
                    wp = round(raw_wp)

                wp = clamp(wp, 0, WP_MAX)

                # ---------- Observation conversion rule ----------
                raw_ob = to_float(request.POST.get(f"record[{rn}]"), 0)
                raw_ob = clamp(raw_ob, 0, OB_MAX)

                if raw_ob <= 10:
                    ob = round((raw_ob / 10) * 20)
                else:
                    ob = round(raw_ob)

                ob = clamp(ob, 0, OB_MAX)

                # ---------- Record ----------
                raw_rc = to_float(request.POST.get(f"viva[{rn}]"), 0)
                rc = round(clamp(raw_rc, 0, RC_MAX))

                total = min(wp + ob + rc, TOTAL_MAX)

                lookup = {"student": student, "courses": enrollment}
                if experiment_no.isdigit() and int(experiment_no) > 0:
                    lookup["experiment_no"] = int(experiment_no)

                obj, created = experiment_marks.objects.update_or_create(
                    **lookup,
                    defaults={
                        "work_program": wp,
                        "observation": ob,
                        "record": rc,
                        "total": total,
                        "assessment": ia_obj,
                    },
                )

                if created:
                    saved += 1
                else:
                    updated += 1

        messages.success(request, f"Saved: {saved}, Updated: {updated}, Skipped: {skipped}")
        return redirect(
            f"{request.path}"
            f"?course_id={course.id if course else (course_id or '')}"
            f"&course_code={course_code}"
            f"&course_title={course_title}"
            f"&department_id={(department.id if department else '')}"
            f"&batch={batch}"
            f"&section={section}"
            f"&degree_code={degree_code}"
            f"&degree_name={degree_name}"
            f"&experiment_no={experiment_no}"
            f"&assessment_iat={assessment_iat}"
        )

    context = {
        "course_id": course.id if course else course_id,
        "course_code": course_code,
        "course_title": course_title,
        "selected_department": selected_dep,
        "selected_department_id": selected_dep_id,
        "selected_batch": batch,
        "selected_section": section,
        "selected_degree_code": degree_code,
        "selected_degree_name": degree_name,
        "experiment_no": experiment_no,
        "assessment_iat": assessment_iat,
        "students": students,
    }
    return render(request, "faculty_management/practical_mark_entry.html", context)




from collections import defaultdict, Counter
from io import BytesIO
import os

from django.conf import settings
from django.contrib.staticfiles import finders
from django.http import HttpResponse
from django.utils.timezone import now

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.utils import ImageReader

# your models (as you already use)
from student_management.models import Add_Department
from user_accounts.models import StudentDetails
from course_management.models import Course, CourseEnrollment
from examination_management.models import experiment_marks


from collections import defaultdict, Counter
from io import BytesIO
import os

from django.conf import settings
from django.contrib.staticfiles import finders
from django.http import HttpResponse

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
from reportlab.lib.utils import ImageReader

from student_management.models import Add_Department
from user_accounts.models import StudentDetails
from course_management.models import Course, CourseEnrollment
from examination_management.models import experiment_marks


def practical_statement_pdf(request):
    """
    PDF: Reg No | Name | Exp1 (COx) | Exp2 (COy) | ...  [totals from experiment_marks]
    Shows a friendly message instead of 404 when no data is available.
    """

    def _no_data(msg="There is no record. Please enter marks to view PDF."):
        return HttpResponse(
            f"""<!doctype html>
<html><head><meta charset="utf-8"><title>No Data</title>
<style>body{{font-family:system-ui,-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;padding:24px;color:#333}}
.alert{{background:#fff3cd;border:1px solid #ffeeba;color:#856404;border-radius:8px;padding:16px;max-width:680px}}
h1{{margin:0 0 8px 0;font-size:18px}}</style></head>
<body>
  <div class="alert">
    <h1>Practical Statement</h1>
    <div>{msg}</div>
  </div>
</body></html>""",
            content_type="text/html",
            status=200,
        )

    course_code = (request.GET.get("course_code") or "").strip()
    batch = (request.GET.get("batch") or "").strip()
    section = (request.GET.get("section") or "").strip()
    dep_id = (request.GET.get("department_id") or "").strip()

    if not course_code or not batch or not section:
        return _no_data("Missing course code, batch, or section.")

    department = Add_Department.objects.filter(id=dep_id).first() if dep_id else None

    course = (
        Course.objects.filter(course_code=course_code, department=department)
        .order_by("-id")
        .first()
        or Course.objects.filter(course_code=course_code).order_by("-id").first()
    )
    if not course:
        return _no_data("There is no record. Please enter marks to view PDF.")

    students = list(
        StudentDetails.objects.filter(
            **({"department": department} if department else {}),
            batch=batch,
            section=section,
        )
        .values("id", "reg_no", "name")
        .order_by("reg_no")
    )
    if not students:
        return _no_data("There is no record. Please enter marks to view PDF.")

    student_ids = [s["id"] for s in students]

    enrollments = list(
        CourseEnrollment.objects.filter(
            student_id__in=student_ids,
            course_id=course.id,
            batch=batch,
            section=section,
        ).values("id", "student_id")
    )
    if not enrollments:
        return _no_data("There is no record. Please enter marks to view PDF.")

    sid_to_enroll_id = {e["student_id"]: e["id"] for e in enrollments}
    enroll_ids = [e["id"] for e in enrollments]

    marks_qs = (
        experiment_marks.objects.filter(
            student_id__in=student_ids,
            courses_id__in=enroll_ids,
        )
        .prefetch_related("cos")
    )

    if not marks_qs.exists():
        return _no_data("There is no record. Please enter marks to view PDF.")

    exp_set = set()
    grid = defaultdict(dict)
    exp_to_cos = defaultdict(list)

    for m in marks_qs:
        en = m.experiment_no
        if en is None:
            continue

        exp_set.add(en)
        grid[m.student_id][en] = m.total

        for co in m.cos.all():
            co_code = (getattr(co, "co_code", "") or "").strip()
            if co_code:
                exp_to_cos[en].append(co_code)

    if not exp_set:
        return _no_data("There is no record. Please enter marks to view PDF.")

    exp_list = sorted(exp_set, key=lambda x: int(x))

    exp_headers = []
    for en in exp_list:
        label = f"Exp{int(en)}"
        cos = exp_to_cos.get(en, [])
        if cos:
            common = Counter(cos).most_common(1)[0][0]
            label = f"{label} ({common})"
        exp_headers.append(label)

    pagesize = A4 if len(exp_list) <= 8 else landscape(A4)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=pagesize,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=44 * mm,
        bottomMargin=18 * mm,
        title="Practical Statement",
    )

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(name="Meta", parent=styles["Normal"], fontName="Helvetica")
    )
    styles.add(ParagraphStyle(name="Title2", parent=styles["Title"], fontSize=18))

    dept_obj = department or course.department
    dept_code = getattr(dept_obj, "Department_code", "") if dept_obj else ""
    dept_name = getattr(dept_obj, "Department", "") if dept_obj else ""
    dept_line = f"{dept_code} — {dept_name}".strip(" —")
    course_line = f"{course.course_code or ''} — {course.title or ''}".strip(" —")
    batch_section_line = f"Batch: {batch}  Section: {section}"

    story = []
    story.append(Spacer(1, 2 * mm))

    header = ["Reg No", "Name"] + exp_headers
    data = [header]

    for s in students:
        sid = s["id"]
        row = [s["reg_no"], s["name"]]
        for en in exp_list:
            row.append(grid.get(sid, {}).get(en, ""))
        data.append(row)

    page_w, _page_h = pagesize
    usable_w = page_w - doc.leftMargin - doc.rightMargin
    fixed_w = 28 * mm + 60 * mm
    exp_cols = max(1, len(exp_list))
    rem_w = max(usable_w - fixed_w, 40 * mm)
    exp_w = rem_w / exp_cols
    col_widths = [28 * mm, 60 * mm] + [exp_w] * len(exp_list)

    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (2, 1), (-1, -1), "CENTER"),
                ("ALIGN", (0, 1), (1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f9f9f9")],
                ),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)

    def _on_page(c, _doc):
        c.saveState()

        page_w, page_h = _doc.pagesize
        left = 18 * mm
        right = 18 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand
            if not logo_path:
                for d in getattr(settings, "STATICFILES_DIRS", []):
                    cand = os.path.join(d, logo_rel)
                    if os.path.exists(cand):
                        logo_path = cand
                        break

        if logo_path and os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                target_h = 18 * mm
                target_w = target_h * (iw / float(ih))
                c.drawImage(
                    img,
                    left,
                    page_h - (target_h + 8 * mm),
                    width=target_w,
                    height=target_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass

        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(
            page_w / 2.0, page_h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY"
        )

        c.setFont("Helvetica", 10)
        c.drawCentredString(page_w / 2.0, page_h - 15 * mm, "Rajapalayam - 626117")
        c.drawCentredString(
            page_w / 2.0, page_h - 19 * mm, "Affiliated to Anna University, Chennai"
        )

        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(page_w / 2.0, page_h - 26 * mm, "Practical Statement")

        c.setFont("Helvetica", 10)
        c.drawCentredString(page_w / 2.0, page_h - 31 * mm, dept_line)
        c.drawCentredString(page_w / 2.0, page_h - 35 * mm, course_line)
        c.drawCentredString(page_w / 2.0, page_h - 39 * mm, batch_section_line)

        rule_y = page_h - 40 * mm
        c.setStrokeColor(colors.black)
        c.setLineWidth(0.5)
        c.line(left, rule_y, page_w - right, rule_y)

        c.setFont("Helvetica", 8)
        c.setFillColor(colors.grey)
        c.drawRightString(page_w - right, 12 * mm, f"Page {c.getPageNumber()}")

        c.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buf.getvalue()
    buf.close()

    filename = f"practical_statement_{course.course_code}_{batch}_{section}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp











from io import BytesIO
import os
import re
from collections import defaultdict

from django.conf import settings
from django.contrib.staticfiles import finders
from django.http import HttpResponse, HttpResponseBadRequest
from django.utils.timezone import now

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics


from io import BytesIO
import os
import re
from collections import defaultdict

from django.conf import settings
from django.contrib.staticfiles import finders
from django.http import HttpResponse, HttpResponseBadRequest
from django.utils.timezone import now

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics


def model_lab_statement_pdf(request):
    """
    Model Lab class statement in wide format (NO Grand Total):
      Reg No | Student Name | <Lab 1> | <Lab 2> | ...

    Accepts GET params:
      - course_code (required)
      - department_id (optional)
      - batch (optional)
      - section (optional)
      - model_lab_id (optional; prefer numeric id)
      - model_lab (optional; fallback by name)
      - internal_iat (optional; filter by ModelLab.internal_assessment__iat)
    """
    course_code = (request.GET.get("course_code") or "").strip()
    department_id = (request.GET.get("department_id") or "").strip()
    batch = (request.GET.get("batch") or "").strip()
    section = (request.GET.get("section") or "").strip()
    model_lab_id_param = (request.GET.get("model_lab_id") or "").strip()
    model_lab_name_param = (request.GET.get("model_lab") or "").strip()
    internal_iat_param = (request.GET.get("internal_iat") or "").strip()

    if not course_code:
        return HttpResponseBadRequest("Missing 'course_code'.")

    department = None
    if department_id:
        department = (
            Add_Department.objects.filter(id=department_id)
            .select_related("degree")
            .first()
        )

    course = (
        Course.objects.filter(course_code=course_code, department=department)
        .order_by("-id")
        .first()
        or Course.objects.filter(course_code=course_code).order_by("-id").first()
    )
    if not course:
        return HttpResponseBadRequest(
            "Course not found for the given course_code/department."
        )

    # Build base queryset — select_related model_lab and student & courses for efficient access
    qs = (
        ModelLabMarks.objects.filter(
            courses__course_id=course.id,
            courses__batch=batch or "",
            courses__section=section or "",
        )
        .select_related("student", "courses", "model_lab")
        .order_by("student__reg_no", "model_lab_id")
    )

    # If specific model_lab_id provided use it (preferred)
    if model_lab_id_param and model_lab_id_param.isdigit():
        qs = qs.filter(model_lab_id=int(model_lab_id_param))
    elif model_lab_name_param:
        # fallback: filter by related ModelLab.model_lab_name (case-insensitive)
        qs = qs.filter(model_lab__model_lab_name__iexact=model_lab_name_param)

    # Optional: filter by internal IAT if provided (ModelLab.internal_assessment__iat)
    if internal_iat_param:
        qs = qs.filter(model_lab__internal_assessment__iat__iexact=internal_iat_param)

    # ---- Build pivot data ----
    def _lab_sort_key(name: str):
        m = re.findall(r"(\d+)", name or "")
        return (int(m[-1]) if m else float("inf"), (name or "").lower())

    lab_names_set = set()
    by_student = defaultdict(lambda: {"name": "", "reg_no": "", "labs": {}})

    # Collect model_lab ids encountered so we can build canonical names (if some rows lack related ModelLab)
    encountered_lab_ids = set()

    for m in qs:
        # lab name: try related object, fallback to m.model_lab (string if your model uses that)
        lab_name = ""
        if getattr(m, "model_lab", None):
            try:
                lab_name = getattr(m.model_lab, "model_lab_name", "") or str(
                    getattr(m, "model_lab", "") or ""
                )
                if hasattr(m.model_lab, "id"):
                    encountered_lab_ids.add(m.model_lab.id)
                elif getattr(m, "model_lab_id", None):
                    encountered_lab_ids.add(getattr(m, "model_lab_id"))
            except Exception:
                lab_name = str(getattr(m, "model_lab", "") or "")
        else:
            lab_name = str(getattr(m, "model_lab", "") or "")

        reg = getattr(m.student, "reg_no", "") or ""
        name = getattr(m.student, "name", "") or ""
        # total field: prefer explicit total; else sum program + viva
        total = m.total if m.total is not None else ((m.program or 0) + (m.viva or 0))

        by_student[reg]["name"] = name
        by_student[reg]["reg_no"] = reg
        if lab_name:
            by_student[reg]["labs"][lab_name] = total
            lab_names_set.add(lab_name)

    # Defensive: fetch lab names for any encountered ids (select_related should cover it)
    if encountered_lab_ids:
        _ = {
            str(l.id): (l.model_lab_name or "").strip()
            for l in ModelLab.objects.filter(id__in=encountered_lab_ids)
        }

    lab_names = sorted(lab_names_set, key=_lab_sort_key)
    # If user explicitly asked `model_lab` by name and it's missing from the set, ensure it's shown
    if model_lab_name_param and model_lab_name_param not in lab_names:
        lab_names = [model_lab_name_param]

    # ---- Page setup: switch to landscape if many labs ----
    lab_count = len(lab_names)
    use_landscape = lab_count >= 6  # threshold: tune as you like

    PAGE_SIZE = landscape(A4) if use_landscape else A4
    LEFT, RIGHT, BOTTOM = 18 * mm, 18 * mm, 18 * mm
    TOP = 44 * mm  # space for header

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=PAGE_SIZE,
        leftMargin=LEFT,
        rightMargin=RIGHT,
        topMargin=TOP,
        bottomMargin=BOTTOM,
    )

    # ---- Styles + dynamic font sizing ----
    base_font = "Helvetica"
    header_font = "Helvetica-Bold"
    if lab_count <= 5:
        table_font_size = 9
    elif lab_count <= 8:
        table_font_size = 8
    else:
        table_font_size = 7

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="MetaLabel",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=10,
        )
    )
    styles.add(ParagraphStyle(name="MetaValue", parent=styles["Normal"], fontSize=10))
    styles.add(
        ParagraphStyle(
            name="NameCell",
            parent=styles["Normal"],
            fontSize=table_font_size,
            leading=table_font_size + 2,
        )
    )

    story = []

    # ---- Details block ----
    details = []
    if batch and section:
        details.append(
            [
                Paragraph("Batch-Section", styles["MetaLabel"]),
                Paragraph(f"{batch} - {section}", styles["MetaValue"]),
            ]
        )
    elif batch:
        details.append(
            [
                Paragraph("Batch", styles["MetaLabel"]),
                Paragraph(f"{batch}", styles["MetaValue"]),
            ]
        )
    elif section:
        details.append(
            [
                Paragraph("Section", styles["MetaLabel"]),
                Paragraph(f"{section}", styles["MetaValue"]),
            ]
        )

    details_table = Table(details, colWidths=[30 * mm, None], hAlign="LEFT")
    details_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(details_table)
    story.append(Spacer(1, 6))

    # ---- Table data ----
    header_cells = [
        Paragraph("Reg No", styles["MetaLabel"]),
        Paragraph("Student Name", styles["MetaLabel"]),
    ]
    if lab_names:
        header_cells.extend([Paragraph(lab, styles["MetaLabel"]) for lab in lab_names])
    else:
        header_cells.append(Paragraph("Model Lab", styles["MetaLabel"]))
    table_data = [header_cells]

    if by_student:
        for reg in sorted(by_student.keys()):
            row = by_student[reg]
            name_para = Paragraph(row["name"], styles["NameCell"])
            per_lab_vals = []
            for lab in lab_names or []:
                # keep blank if not present; show 0 if stored explicitly
                val = row["labs"].get(lab, "")
                per_lab_vals.append(val if (val != 0 or lab in row["labs"]) else "")
            table_data.append([row["reg_no"], name_para, *per_lab_vals])
    else:
        if lab_names:
            table_data.append(
                ["—", "No records for the selected criteria", *["—"] * (len(lab_names))]
            )
        else:
            table_data.append(["—", "No records for the selected criteria"])

    # ---- Width calculation (same logic as you had) ----
    page_w, _ = PAGE_SIZE
    content_w = page_w - (LEFT + RIGHT)

    reg_min = 22 * mm
    reg_max = 28 * mm

    def measure(txt, font=base_font, size=table_font_size):
        return pdfmetrics.stringWidth(str(txt), font, size)

    longest_reg = (
        max([len(r[0]) for r in table_data[1:]] + [len("Reg No")])
        if len(table_data) > 1
        else len("Reg No")
    )
    reg_w_pts = (
        max(measure("8" * longest_reg), measure("Reg No", header_font, table_font_size))
        + 6
    )
    reg_w = max(reg_min, min(reg_max, reg_w_pts))

    name_min = 45 * mm if not use_landscape else 55 * mm
    name_max = 90 * mm if not use_landscape else 110 * mm

    lab_min = 16 * mm if not use_landscape else 18 * mm
    lab_max = 28 * mm if not use_landscape else 30 * mm

    lab_w_estimates = []
    for idx, lab in enumerate(lab_names):
        header_w = measure(lab, header_font, table_font_size) + 6
        sample_vals = []
        for r in table_data[1:]:
            try:
                sample_vals.append(str(r[2 + idx]))
            except Exception:
                sample_vals.append("")
        val_w = (
            max([measure(v, base_font, table_font_size) for v in sample_vals] + [0]) + 6
        )
        w = max(header_w, val_w)
        lab_w_estimates.append(max(lab_min, min(lab_max, w)))

    if lab_count == 0:
        name_w = max(name_min, min(name_max, content_w - reg_w))
        col_widths = [reg_w, name_w]
    else:
        labs_total = sum(lab_w_estimates)
        remaining_for_name = content_w - reg_w - labs_total
        if remaining_for_name < name_min:
            deficit = name_min - remaining_for_name
            if labs_total > 0:
                ratio = max(0, (labs_total - deficit) / labs_total)
                lab_w_estimates = [max(lab_min, w * ratio) for w in lab_w_estimates]
            labs_total = sum(lab_w_estimates)
            remaining_for_name = content_w - reg_w - labs_total
            name_w = max(name_min, min(name_max, remaining_for_name))
        else:
            name_w = max(name_min, min(name_max, remaining_for_name))
        col_widths = [reg_w, name_w] + lab_w_estimates

    # ---- Build table and styles ----
    marks_table = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    marks_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), header_font),
                ("FONTSIZE", (0, 0), (-1, 0), table_font_size),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 1), (-1, -1), base_font),
                ("FONTSIZE", (0, 1), (-1, -1), table_font_size),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 0), (1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f9f9f9")],
                ),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    story.append(marks_table)

    # ---- Header/footer (unchanged) ----
    def _on_page(c, _doc):
        c.saveState()
        page_w, page_h = _doc.pagesize
        left, right = LEFT, RIGHT

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                candidate = os.path.join(static_root, logo_rel)
                if os.path.exists(candidate):
                    logo_path = candidate
            if not logo_path:
                for dir_ in getattr(settings, "STATICFILES_DIRS", []):
                    candidate = os.path.join(dir_, logo_rel)
                    if os.path.exists(candidate):
                        logo_path = candidate
                        break

        if logo_path and os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                target_h = 18 * mm
                target_w = target_h * (iw / float(ih))
                c.drawImage(
                    img,
                    left,
                    page_h - (target_h + 8 * mm),
                    width=target_w,
                    height=target_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass

        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(
            page_w / 2.0, page_h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY"
        )

        c.setFont("Helvetica", 10)
        c.drawCentredString(page_w / 2.0, page_h - 15 * mm, "Rajapalayam - 626117")
        c.drawCentredString(
            page_w / 2.0, page_h - 19 * mm, "Affiliated to Anna University, Chennai"
        )

        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(page_w / 2.0, page_h - 26 * mm, "Model Lab Statement")

        c.setFont("Helvetica", 10)
        dept_obj = department or getattr(course, "department", None)
        dept_code = getattr(dept_obj, "Department_code", "") if dept_obj else ""
        dept_name = getattr(dept_obj, "Department", "") if dept_obj else ""
        dep_line = f"{dept_code} — {dept_name}" if (dept_code or dept_name) else ""
        course_line = f"{course_code} — {getattr(course, 'title', '')}"
        if dep_line:
            c.drawCentredString(page_w / 2.0, page_h - 31 * mm, dep_line)
            c.drawCentredString(page_w / 2.0, page_h - 35 * mm, course_line)
            rule_y = page_h - 40 * mm
        else:
            c.drawCentredString(page_w / 2.0, page_h - 31 * mm, course_line)
            rule_y = page_h - 36 * mm

        c.setStrokeColor(colors.black)
        c.setLineWidth(0.5)
        c.line(left, rule_y, page_w - right, rule_y)

        c.setFont("Helvetica", 8)
        c.setFillColor(colors.grey)
        c.drawRightString(page_w - right, 12 * mm, f"Page {c.getPageNumber()}")
        c.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buf.getvalue()
    buf.close()

    selected_lab_label = (model_lab_name_param or model_lab_id_param or "ALL").replace(
        " ", "_"
    )
    filename = f"ModelLab_{course_code}_{batch}_{section}_{selected_lab_label}_wide.pdf".replace(
        " ", "_"
    )
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


from django.shortcuts import render, redirect
from django.db.models import Q

# Import your decorators and models
# Adjust these imports to match your project structure.

# If you already have a model to persist model-lab marks, import it here.
# from .models import ModelLabMarks

# views.py
from django.shortcuts import render, redirect
from django.db import transaction
from django.contrib import messages
from django.db.models import Q

# Adjust imports to your project layout
# or wherever Add_Department lives
from examination_management.models import ModelLabMarks  # path to the model you shared
from django.shortcuts import render, redirect
from django.db import transaction
from django.contrib import messages
from django.db.models import Q
from django.shortcuts import render, redirect
from django.db import transaction
from django.contrib import messages

@check_permission("Internalmarks")
@student_management
def model_lab_entry_page(request):
    """
    Model Lab entry page.
    - Expects model_lab_id (FK id) in POST/GET as 'model_lab_id'
    - Supports only_total mode:
        * if checked, Program and Viva are ignored
        * Total is entered directly and saved
        * Program and Viva stored as NULL
    """

    def g(key, default=""):
        return request.POST.get(key, request.GET.get(key, default))

    # ---------- incoming context ----------
    course_id = request.POST.get("course_id") or request.GET.get("course_id")
    course_code = g("course_code")
    course_title = g("course_title")
    sel_dep = g("selected_department") or g("department")
    sel_dep_id = g("selected_department_id") or g("department_id")
    batch = g("selected_batch") or g("batch")
    section = g("selected_section") or g("section")
    degree_code = g("selected_degree_code") or g("degree_code")
    degree_name = g("selected_degree_name") or g("degree_name")
    selected_model_lab_id = g("model_lab_id") or ""
    selected_iat_name = g("selected_iat") or g("iat") or ""

    only_total = g("only_total")
    only_total = str(only_total).lower() in ["1", "true", "on", "yes"]

    # ---------- resolve department -> degree ----------
    department = None
    if sel_dep_id:
        department = Add_Department.objects.select_related("degree").filter(id=sel_dep_id).first()
    else:
        try:
            _, dep_name = (sel_dep or "").split(" ", 1)
        except ValueError:
            dep_name = sel_dep or ""

        if dep_name:
            department = Add_Department.objects.select_related("degree").filter(Department=dep_name).first()

    degree_obj = getattr(department, "degree", None) if department else None

    # ---------- resolve InternalAssessment (IAT) ----------
    ia_obj = None
    if selected_iat_name and degree_obj:
        ia_obj = InternalAssessment.objects.filter(degree=degree_obj, iat=selected_iat_name).first()
        if ia_obj is None:
            ia_obj = InternalAssessment.objects.create(degree=degree_obj, iat=selected_iat_name)

    # ---------- lookup ModelLab FK ----------
    selected_lab_obj = None
    if selected_model_lab_id and degree_obj:
        try:
            lid = int(selected_model_lab_id)
            selected_lab_obj = ModelLab.objects.filter(id=lid, degree=degree_obj).first()
        except (ValueError, TypeError):
            selected_lab_obj = None

    # ---------- fetch students ----------
    students = []
    cid = None

    if department and course_id:
        try:
            cid = int(course_id)
        except (TypeError, ValueError):
            cid = None

        if cid:
            enrollments = (
                CourseEnrollment.objects
                .select_related("student")
                .filter(
                    course_id=cid,
                    batch=batch,
                    section=section,
                    enroll=True
                )
                .order_by("student__reg_no")
            )

            for e in enrollments:
                if e.student:
                    students.append({
                        "id": e.student.id,
                        "reg_no": e.student.reg_no,
                        "name": e.student.name,
                    })

    # ---------- resolve course ----------
    course = None
    if cid:
        course = Course.objects.filter(id=cid).first()

    if course is None and course_code:
        course = (
            Course.objects.filter(course_code=course_code, department=department).order_by("-id").first()
            or Course.objects.filter(course_code=course_code).order_by("-id").first()
        )

    # ---------- default values ----------
    for s in students:
        s["pref_program"] = 75
        s["pref_viva"] = 25
        s["pref_total"] = 100
        s["pref_only_total"] = False

    # ---------- prefill existing marks ----------
    db_only_total_found = False

    if course and students:
        student_ids = [s["id"] for s in students]

        enrollments = (
            CourseEnrollment.objects
            .filter(
                student_id__in=student_ids,
                course_id=course.id,
                batch=batch or "",
                section=section or "",
            )
            .values("id", "student_id")
        )
        sid_to_enroll = {e["student_id"]: e["id"] for e in enrollments}

        marks_q = ModelLabMarks.objects.filter(
            student_id__in=student_ids,
            courses_id__in=list(sid_to_enroll.values()) if sid_to_enroll else [],
            batch=batch or "",
            section=section or "",
        )

        if selected_lab_obj:
            marks_q = marks_q.filter(model_lab=selected_lab_obj)

        if ia_obj:
            marks_q = marks_q.filter(internal_assessment=ia_obj)

        marks_q = marks_q.values(
            "student_id",
            "program",
            "viva",
            "total",
            "only_total",
        )

        marks_map = {m["student_id"]: m for m in marks_q}

        for s in students:
            m = marks_map.get(s["id"])
            if m:
                s["pref_program"] = m["program"] if m["program"] is not None else ""
                s["pref_viva"] = m["viva"] if m["viva"] is not None else ""
                s["pref_total"] = m["total"] if m["total"] is not None else ""
                s["pref_only_total"] = bool(m.get("only_total", False))

                if s["pref_only_total"]:
                    db_only_total_found = True

        if request.method != "POST":
            only_total = db_only_total_found

    # ---------- SAVE ----------
    if request.method == "POST" and request.POST.getlist("reg_nos[]"):
        if not course_code:
            messages.error(request, "Missing course code; cannot save Model Lab marks.")
            return redirect(request.path)

        if course is None:
            course = Course.objects.create(
                course_code=course_code,
                title=course_title or "",
                department=department
            )
            course_id = course.id

        PROGRAM_MAX, VIVA_MAX, TOTAL_MAX = 75, 25, 100

        def to_int(val, max_v):
            try:
                v = int(val)
            except (TypeError, ValueError):
                v = 0
            return 0 if v < 0 else (max_v if v > max_v else v)

        # Re-resolve lab from POST
        post_lab_id = request.POST.get("model_lab_id") or request.GET.get("model_lab_id") or ""
        if post_lab_id:
            try:
                lid = int(post_lab_id)
                selected_lab_obj = ModelLab.objects.filter(id=lid, degree=degree_obj).first()
            except (ValueError, TypeError):
                selected_lab_obj = None

        if not selected_lab_obj:
            messages.error(request, "Please select an existing Model Lab (valid). Cannot save without a valid Model Lab.")
            return redirect(
                f"{request.path}"
                f"?course_id={course_id or ''}"
                f"&course_code={course_code}"
                f"&course_title={course_title}"
                f"&selected_department_id={(department.id if department else '')}"
                f"&selected_batch={batch}"
                f"&selected_section={section}"
                f"&selected_degree_code={degree_code}"
                f"&selected_degree_name={degree_name}"
                f"&model_lab_id={post_lab_id or ''}"
                f"&selected_iat={selected_iat_name or ''}"
                f"&only_total={'1' if only_total else '0'}"
            )

        reg_nos = request.POST.getlist("reg_nos[]")
        saved = updated = skipped = 0

        with transaction.atomic():
            for rn in reg_nos:
                student = (
                    StudentDetails.objects.filter(
                        reg_no=rn,
                        **({"department": department} if department else {}),
                        batch=batch,
                        section=section
                    ).first()
                    or StudentDetails.objects.filter(reg_no=rn).first()
                )

                if not student:
                    skipped += 1
                    continue

                enrollment, _ = CourseEnrollment.objects.get_or_create(
                    student_id=student.id,
                    course_id=course.id,
                    batch=batch or "",
                    section=section or "",
                    defaults={"department_id": (department.id if department else None)},
                )

                desired_dept_id = department.id if department else None
                if enrollment.department_id != desired_dept_id:
                    enrollment.department_id = desired_dept_id
                    enrollment.save(update_fields=["department_id"])

                if only_total:
                    prog = None
                    viv = None
                    total = to_int(request.POST.get(f"total[{rn}]"), TOTAL_MAX)
                else:
                    prog = to_int(request.POST.get(f"program[{rn}]"), PROGRAM_MAX)
                    viv = to_int(request.POST.get(f"viva[{rn}]"), VIVA_MAX)
                    total = min(prog + viv, TOTAL_MAX)

                obj, created = ModelLabMarks.objects.update_or_create(
                    student=student,
                    courses=enrollment,
                    model_lab=selected_lab_obj,
                    internal_assessment=ia_obj,
                    batch=batch or "",
                    section=section or "",
                    defaults={
                        "program": prog,
                        "viva": viv,
                        "total": total,
                        "only_total": only_total,
                    }
                )

                if created:
                    saved += 1
                else:
                    updated += 1

        messages.success(request, f"Model Lab saved — New: {saved}, Updated: {updated}, Skipped: {skipped}")

        return redirect(
            f"{request.path}"
            f"?course_id={course.id if course else (course_id or '')}"
            f"&course_code={course_code}"
            f"&course_title={course_title}"
            f"&selected_department_id={(department.id if department else '')}"
            f"&selected_batch={batch}"
            f"&selected_section={section}"
            f"&selected_degree_code={degree_code}"
            f"&selected_degree_name={degree_name}"
            f"&model_lab_id={(selected_lab_obj.id if selected_lab_obj else '')}"
            f"&selected_iat={selected_iat_name or ''}"
            f"&only_total={'1' if only_total else '0'}"
        )

    # ---------- render ----------
    context = {
        "course_id": course.id if course else course_id,
        "course_code": course_code,
        "course_title": course_title,
        "selected_department": sel_dep,
        "selected_department_id": sel_dep_id,
        "db_only_total_found": db_only_total_found,
        "selected_batch": batch,
        "selected_section": section,
        "selected_degree_code": degree_code,
        "selected_degree_name": degree_name,
        "model_lab": selected_lab_obj.model_lab_name if selected_lab_obj else "",
        "selected_model_lab_id": selected_lab_obj.id if selected_lab_obj else selected_model_lab_id,
        "selected_iat": selected_iat_name,
        "students": students,
        "only_total": only_total,
    }

    return render(request, "faculty_management/model_lab_entry.html", context)















from django.shortcuts import render, redirect
from django.contrib import messages
from examination_management.models import (
    ExamPattern,
    ExamPatternSetting,
    StudentExam,
    StudentMark,
    BloomsLevel,
    CourseOutcome,
    Final_Marks,
)


# views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from examination_management.models import (
    ExamPattern,
    ExamPatternSetting,
    StudentExam,
    StudentMark,
    BloomsLevel,
    CourseOutcome,
    Final_Marks,
)


from django.shortcuts import render
from django.contrib import messages
from django.db.models import Prefetch

from examination_management.models import (
    ExamPattern,
    Part,
    Question,
    OptionMarks,
    BloomsLevel,
    CourseOutcome,
    StudentInternalMark,
)
from student_management.models import StudentDetails
from course_management.models import CourseEnrollment, Regulations


def _safe_get(obj, *attrs, default=None):
    for a in attrs:
        try:
            v = getattr(obj, a)
            if v is not None:
                return v
        except Exception:
            pass
    return default


def _resolve_degree(student, enrollment):
    """Find Degree via common paths."""
    try:
        if student and getattr(student, "department", None):
            deg = getattr(student.department, "degree", None)
            if deg:
                return deg
    except Exception:
        pass
    try:
        if enrollment and getattr(enrollment, "degree", None):
            return enrollment.degree
    except Exception:
        pass
    try:
        prog = getattr(enrollment, "program", None)
        if prog and getattr(prog, "degree", None):
            return prog.degree
    except Exception:
        pass
    try:
        course = getattr(enrollment, "course", None)
        if course:
            dept = getattr(course, "department", None)
            if dept and getattr(dept, "degree", None):
                return dept.degree
    except Exception:
        pass
    try:
        course = getattr(enrollment, "course", None)
        if course and getattr(course, "degree", None):
            return course.degree
    except Exception:
        pass
    return None


def _degree_id_and_label(degree_obj):
    """Normalize Degree -> (id, label)"""
    degree_id = None
    degree_label = None
    if degree_obj is not None:
        degree_id = getattr(degree_obj, "pk", None) or getattr(degree_obj, "id", None)
        degree_label = (
            getattr(degree_obj, "name", None)
            or getattr(degree_obj, "degree_name", None)
            or getattr(degree_obj, "degree_code", None)
            or getattr(degree_obj, "code", None)
            or str(degree_obj)
        )
    if degree_id is None and isinstance(degree_obj, int):
        degree_id = degree_obj
    if degree_id is None and isinstance(degree_obj, str) and degree_obj.isdigit():
        degree_id = int(degree_obj)
    return degree_id, degree_label


def _resolve_regulation_obj(student, enrollment):
    """
    Regulations has only: id, year.
    Map StudentDetails.regulation (string) -> Regulations instance by year.
    """
    raw = _safe_get(student, "regulation") or _safe_get(enrollment, "regulation")
    if not raw:
        return None

    raw_str = str(raw).strip()

    # Exact year match first
    qs = Regulations.objects.filter(year__iexact=raw_str).order_by("-id")
    if qs.exists():
        return qs.first()

    # Contains (handles 'Regulation 2021' stored as text)
    qs = Regulations.objects.filter(year__icontains=raw_str).order_by("-id")
    if qs.exists():
        return qs.first()

    # Digits-only fallback: 'Regulation 2021' -> '2021'
    digits = "".join(c for c in raw_str if c.isdigit())
    if digits:
        qs = Regulations.objects.filter(year__icontains=digits).order_by("-id")
        if qs.exists():
            return qs.first()

    return None


from django.shortcuts import render
from django.db.models import Prefetch
from django.contrib import messages

from django.db.models import Prefetch

from django.shortcuts import render
from django.db.models import Prefetch
from django.contrib import messages
from collections import Counter


# def enter_marks_page(request):
#     # --- Query params ---
#     bulk = request.GET.get("bulk") == "1"
#     # print(f"Bulk mode: {bulk}")

#     reg_no = request.GET.get("reg_no")
#     student_name = request.GET.get("student_name")
#     course_code = request.GET.get("course_code")
#     course_title = request.GET.get("course_title")
#     batch = request.GET.get("batch")
#     section = request.GET.get("section")
#     iat = request.GET.get("iat")
#     department_name = request.GET.get("department_name")
#     department_code = request.GET.get("department_code")
#     department_id = request.GET.get("department_id")
#     degree_id_qs = request.GET.get("degree_id")
#     pattern_id = request.GET.get("pattern_id")
#     print("department name => ", department_name)
#     print("department id => ", department_id)
#     print("department_code  => ", department_code)
#     # print(f"Query Params: reg_no={reg_no}, student_name={student_name}, course_code={course_code}, course_title={course_title}, batch={batch}, section={section}, iat={iat}, department_name={department_name}, department_code={department_code}, department_id={department_id}, degree_id_qs={degree_id_qs}, pattern_id={pattern_id}")

#     # ----------------------------
#     # BULK: Load students list
#     # ----------------------------
#     department = Add_Department.objects.get(Department_code=department_code)
#     students_list = []
#     enrollments = []
#     faculty = general_information.objects.filter(faculty_id=request.user.Employee_id).first() if request.user else None
#     if bulk:
#         enr_qs = CourseEnrollment.objects.select_related("student", "course").filter(
#             course__course_code=course_code,
#             section=section,
#             batch=batch,
#             department=department,
#             enroll=True,
#         )

#         if batch:
#             enr_qs = enr_qs.filter(batch=batch)
#         if section:
#             enr_qs = enr_qs.filter(section=section)
#         if department_id:
#             enr_qs = enr_qs.filter(department_id=department_id)

#         faculty_enr_qs = enr_qs.filter(faculty=faculty) if faculty else enr_qs.none()
#         if faculty_enr_qs.exists():
#             enr_qs = faculty_enr_qs

#         enr_qs = enr_qs.order_by("student__reg_no", "id")
#         enrollments = list(enr_qs)

#         # print(f"Enrollments found for bulk: {len(enrollments)}")

#         for e in enrollments:
#             if e.student:
#                 students_list.append(e.student)

#         if not students_list:
#             messages.warning(request, "No students found for this Course/Batch/Section.")
#             # print(f"No students found for Course/Batch/Section: {course_code}, {batch}, {section}")
#             return render(request, "faculty_management/enter_iat_marks_bulk.html", {
#                 "bulk": True,
#                 "students_list": [],
#                 "course_code": course_code,
#                 "course_title": course_title,
#                 "batch": batch,
#                 "section": section,
#                 "iat": iat,
#                 "department_name": department_name,
#                 "department_code": department_code,
#             })

#         student = students_list[0]
#         enrollment = enrollments[0] if enrollments else None
#         # print(f"Selected student: {student.reg_no}, Enrollment: {enrollment}")

#     else:
#         # ----------------------------
#         # SINGLE
#         # ----------------------------
#         student = StudentDetails.objects.filter(reg_no=reg_no).first()
#         enrollment = None
#         if student:
#             enrollment_qs = CourseEnrollment.objects.filter(
#                 student_id=student.id,
#                 course__course_code=course_code,
#             )
#             if department:
#                 enrollment_qs = enrollment_qs.filter(department=department)
#             elif department_id:
#                 enrollment_qs = enrollment_qs.filter(department_id=department_id)
#             if batch:
#                 enrollment_qs = enrollment_qs.filter(batch=batch)
#             if section:
#                 enrollment_qs = enrollment_qs.filter(section=section)
#             enrollment = enrollment_qs.order_by("-enroll", "-id").first()

#         # print(f"Single student found: {student.reg_no if student else 'None'}, Enrollment: {enrollment}")

#     # --- Resolve Degree/Year/Sem/Regulation ---
#     degree_obj = _resolve_degree(student, enrollment)
#     degree_id, degree_label = _degree_id_and_label(degree_obj)
#     # print(f"Degree: {degree_label}, Degree ID: {degree_id}")

#     year_val = (
#         _safe_get(enrollment, "year")
#         or _safe_get(student, "current_year")
#         or _safe_get(student, "year")
#     )
#     semester_val = (
#         _safe_get(enrollment, "semester")
#         or _safe_get(student, "current_semester")
#         or _safe_get(student, "semester")
#     )
#     year_val = str(year_val) if year_val is not None else None
#     semester_val = str(semester_val) if semester_val is not None else None

#     # print(f"Year: {year_val}, Semester: {semester_val}")

#     regulation_obj = _resolve_regulation_obj(student, enrollment)
#     regulation_display = getattr(regulation_obj, "year", None) or _safe_get(student, "regulation")
#     # print(f"Regulation: {regulation_display}")

#     # --- Build pattern options ---
#     pattern_options = ExamPattern.objects.none()
#     if regulation_obj and degree_id and year_val and semester_val and iat:
#         pattern_options = (
#             ExamPattern.objects.select_related("regulation", "degree")
#             .filter(
#                 regulation=regulation_obj,
#                 degree_id=degree_id,
#                 year=year_val,
#                 semester=semester_val,
#                 for_exam=str(iat).strip(),
#             )
#             .order_by("pattern", "id")
#         )
#     filtered_ids = list(pattern_options.values_list("id", flat=True))
#     # print(f"Pattern options: {filtered_ids}")

#     # --- Auto-restore last used pattern if none was provided ---
#     restored = False
#     if (not pattern_id) and enrollment and iat:
#         last_row = (
#             StudentInternalMark.objects.filter(
#                 enrollment=enrollment,
#                 course=enrollment.course,
#                 exam_name=iat,
#                 pattern__isnull=False,
#             )
#             .select_related("pattern")
#             .order_by("-created_at", "-id")
#             .first()
#         )
#         if last_row and last_row.pattern_id:
#             pattern_id = str(last_row.pattern_id)
#             restored = True

#     # print(f"Restored last pattern: {restored}, Pattern ID: {pattern_id}")

#     # --- Load selected pattern ---
#     exam_pattern_obj = None
#     selected_pid = None
#     if pattern_id:
#         try:
#             selected_pid = int(pattern_id)
#         except ValueError:
#             selected_pid = None

#     if selected_pid:
#         try:
#             exam_pattern_obj = (
#                 ExamPattern.objects.select_related("regulation", "degree")
#                 .prefetch_related(
#                     Prefetch(
#                         "parts",
#                         queryset=Part.objects.prefetch_related(
#                             Prefetch(
#                                 "questions",
#                                 queryset=Question.objects.prefetch_related("options").order_by("number"),
#                                 to_attr="prefetched_questions",
#                             )
#                         ).order_by("name"),
#                         to_attr="prefetched_parts",
#                     )
#                 )
#                 .get(id=selected_pid)
#             )
#         except ExamPattern.DoesNotExist:
#             exam_pattern_obj = None

#     pattern_selected = bool(exam_pattern_obj)
#     # print(f"Exam pattern selected: {pattern_selected}, Exam pattern object: {exam_pattern_obj}")

#     # ----------------------------------------
#     # PREFILL: SINGLE vs BULK
#     # ----------------------------------------
#     prefilled_marks = {}
#     prefilled_co = {}
#     prefilled_blooms = {}
#     prefilled_subpart_max = {}

#     prefilled_marks_by_reg = {}
#     prefilled_co_by_reg = {}
#     prefilled_blooms_by_reg = {}

#     part_marks_summary = {}

#     if enrollment and iat and exam_pattern_obj:
#         if bulk:
#             marks_qs = StudentInternalMark.objects.select_related(
#                 "student", "co_code", "level_code"
#             ).filter(
#                 enrollment__course=enrollment.course,
#                 batch=enrollment.batch,
#                 section=enrollment.section,
#                 exam_name=iat,
#                 pattern=exam_pattern_obj,
#             )

#             for row in marks_qs:
#                 rno = getattr(row.student, "reg_no", None)
#                 if not rno:
#                     continue

#                 part = (row.part_name or "").strip().upper()
#                 qnum = str(row.question_number or "").strip()
#                 sub = (row.sub_question or "").strip().lower()
#                 opt = (row.option_letter or "").strip().lower()

#                 if sub and opt:
#                     suffix = f"{part}_{qnum}_{sub}_{opt}"
#                 elif sub:
#                     suffix = f"{part}_{qnum}_{sub}"
#                 elif opt:
#                     suffix = f"{part}_{qnum}_{opt}"
#                 else:
#                     suffix = f"{part}_{qnum}"

#                 prefilled_marks_by_reg.setdefault(rno, {})[f"marks_{suffix}"] = row.marks_obtained

#                 if row.co_code_id:
#                     prefilled_co_by_reg.setdefault(rno, {})[f"co_{suffix}"] = str(row.co_code_id)
#                     prefilled_co[f"co_{suffix}"] = str(row.co_code_id)

#                 if row.level_code_id:
#                     prefilled_blooms_by_reg.setdefault(rno, {})[f"blooms_{suffix}"] = str(row.level_code_id)
#                     prefilled_blooms[f"blooms_{suffix}"] = str(row.level_code_id)

#                 # PREFILL subpart max once per combination
#                 if sub in ("i", "ii") and opt:
#                     sp_key = f"{part}__{qnum}__{opt}__{sub}"
#                     if sp_key not in prefilled_subpart_max and row.max_marks is not None:
#                         prefilled_subpart_max[sp_key] = row.max_marks

#         else:
#             student_marks_qs = StudentInternalMark.objects.select_related(
#                 "co_code", "level_code"
#             ).filter(
#                 student=student,
#                 enrollment=enrollment,
#                 exam_name=iat,
#             )

#             for row in student_marks_qs:
#                 part = (row.part_name or "").strip().upper()
#                 qnum = str(row.question_number or "").strip()
#                 sub = (row.sub_question or "").strip().lower()
#                 opt = (row.option_letter or "").strip().lower()

#                 if sub and opt:
#                     suffix = f"{part}_{qnum}_{sub}_{opt}"
#                 elif sub:
#                     suffix = f"{part}_{qnum}_{sub}"
#                 elif opt:
#                     suffix = f"{part}_{qnum}_{opt}"
#                 else:
#                     suffix = f"{part}_{qnum}"

#                 prefilled_marks[f"marks_{suffix}"] = row.marks_obtained

#                 if row.co_code_id:
#                     prefilled_co[f"co_{suffix}"] = str(row.co_code_id)

#                 if row.level_code_id:
#                     prefilled_blooms[f"blooms_{suffix}"] = str(row.level_code_id)

#                 # PREFILL subpart max once per combination
#                 if sub in ("i", "ii") and opt:
#                     sp_key = f"{part}__{qnum}__{opt}__{sub}"
#                     if sp_key not in prefilled_subpart_max and row.max_marks is not None:
#                         prefilled_subpart_max[sp_key] = row.max_marks

#     if exam_pattern_obj:
#         parts_list = getattr(exam_pattern_obj, "prefetched_parts", exam_pattern_obj.parts.all())

#         for part in parts_list:
#             if str(part.name).lower() == "a":
#                 if part.total_questions and part.max_marks:
#                     part_marks_summary[part.name] = f"{part.total_questions} × {part.max_marks} Marks"
#                 else:
#                     part_marks_summary[part.name] = f"{part.max_marks} Marks"
#             else:
#                 questions_list = getattr(part, "prefetched_questions", part.questions.all())

#                 question_marks = []
#                 for question in questions_list:
#                     q_mark = question.total_marks or part.max_marks or 0
#                     if q_mark > 0:
#                         question_marks.append(q_mark)

#                 if question_marks:
#                     counts = Counter(question_marks)
#                     summary_parts = []

#                     for mark, count in sorted(counts.items(), reverse=True):
#                         summary_parts.append(f"{count} × {mark} Marks")

#                     part_marks_summary[part.name] = ", ".join(summary_parts)
#                 else:
#                     part_marks_summary[part.name] = f"{part.max_marks} Marks"

#     # print(f"Prefilled marks: {prefilled_marks}")
#     # print(f"Prefilled CO: {prefilled_co}")
#     # print(f"Prefilled Blooms: {prefilled_blooms}")
#     # print(f"Prefilled subpart max: {prefilled_subpart_max}")

#     course_outcomes = CourseOutcome.objects.none()
#     blooms_levels = BloomsLevel.objects.all().order_by("level_code")
#     if exam_pattern_obj and exam_pattern_obj.regulation:
#         # print("Loading course outcomes for regulation:", exam_pattern_obj.regulation)
#         course_outcomes = CourseOutcome.objects.filter(
#             regulation=exam_pattern_obj.regulation.year
#         ).order_by("co_code")
#     # print("course outcome => ", course_outcomes)

#     context = {
#         "bulk": bulk,
#         "students_list": students_list,
#         "reg_no": reg_no,
#         "student_name": student_name,
#         "course_code": course_code,
#         "course_title": course_title,
#         "batch": batch,
#         "section": section,
#         "iat": iat,
#         "exam_name": iat,
#         "department_name": department_name,
#         "department_code": department_code,
#         "pattern_options": pattern_options,
#         "selected_pattern_id": selected_pid,
#         "pattern_selected": pattern_selected,
#         "selected_pattern": getattr(exam_pattern_obj, "pattern", None),
#         "selected_regulation": getattr(regulation_obj, "year", None) or regulation_display,
#         "selected_degree": degree_label,
#         "selected_year": year_val,
#         "selected_semester": semester_val,
#         "exam_pattern_obj": exam_pattern_obj,
#         "exam_pattern_id": getattr(exam_pattern_obj, "id", None),
#         "prefilled_marks": prefilled_marks,
#         "prefilled_co": prefilled_co,
#         "prefilled_blooms": prefilled_blooms,
#         "prefilled_marks_by_reg": prefilled_marks_by_reg,
#         "prefilled_co_by_reg": prefilled_co_by_reg,
#         "prefilled_blooms_by_reg": prefilled_blooms_by_reg,
#         "prefilled_subpart_max": prefilled_subpart_max,
#         "part_marks_summary": part_marks_summary,
#         "course_outcomes": course_outcomes,
#         "blooms_levels": blooms_levels,
#     }

#     if bulk:
#         return render(request, "faculty_management/enter_iat_marks.html", context)

#     return render(request, "faculty_management/enter_iat_marks.html", context)

from collections import Counter, defaultdict
from django.contrib import messages
from django.db.models import Prefetch
from django.shortcuts import render

def enter_marks_page(request):
    bulk = request.GET.get("bulk") == "1"

    reg_no = request.GET.get("reg_no")
    student_name = request.GET.get("student_name")
    course_code = request.GET.get("course_code")
    course_title = request.GET.get("course_title")
    batch = request.GET.get("batch")
    section = request.GET.get("section")
    iat = request.GET.get("iat")
    department_name = request.GET.get("department_name")
    department_code = request.GET.get("department_code")
    department_id = request.GET.get("department_id")
    degree_id_qs = request.GET.get("degree_id")
    pattern_id = request.GET.get("pattern_id")

    department = Add_Department.objects.get(Department_code=department_code)

    students_list = []
    enrollments = []
    faculty = general_information.objects.filter(
        faculty_id=request.user.Employee_id
    ).first() if request.user else None

    if bulk:
        enr_qs = CourseEnrollment.objects.select_related("student", "course").filter(
            course__course_code=course_code,
            section=section,
            batch=batch,
            department=department,
            enroll=True,
        )

        if batch:
            enr_qs = enr_qs.filter(batch=batch)
        if section:
            enr_qs = enr_qs.filter(section=section)
        if department_id:
            enr_qs = enr_qs.filter(department_id=department_id)

        faculty_enr_qs = enr_qs.filter(faculty=faculty) if faculty else enr_qs.none()
        if faculty_enr_qs.exists():
            enr_qs = faculty_enr_qs

        enr_qs = enr_qs.order_by("student__reg_no", "id")
        enrollments = list(enr_qs)

        for e in enrollments:
            if e.student:
                students_list.append(e.student)

        if not students_list:
            messages.warning(request, "No students found for this Course/Batch/Section.")
            return render(request, "faculty_management/enter_iat_marks.html", {
                "bulk": True,
                "students_list": [],
                "course_code": course_code,
                "course_title": course_title,
                "batch": batch,
                "section": section,
                "iat": iat,
                "department_name": department_name,
                "department_code": department_code,
            })

        student = students_list[0]
        enrollment = enrollments[0] if enrollments else None

    else:
        student = StudentDetails.objects.filter(reg_no=reg_no).first()
        enrollment = None

        if student:
            enrollment_qs = CourseEnrollment.objects.filter(
                student_id=student.id,
                course__course_code=course_code,
            )

            if department:
                enrollment_qs = enrollment_qs.filter(department=department)
            elif department_id:
                enrollment_qs = enrollment_qs.filter(department_id=department_id)

            if batch:
                enrollment_qs = enrollment_qs.filter(batch=batch)
            if section:
                enrollment_qs = enrollment_qs.filter(section=section)

            enrollment = enrollment_qs.order_by("-enroll", "-id").first()

    degree_obj = _resolve_degree(student, enrollment)
    degree_id, degree_label = _degree_id_and_label(degree_obj)

    year_val = (
        _safe_get(enrollment, "year")
        or _safe_get(student, "current_year")
        or _safe_get(student, "year")
    )
    semester_val = (
        _safe_get(enrollment, "semester")
        or _safe_get(student, "current_semester")
        or _safe_get(student, "semester")
    )

    year_val = str(year_val) if year_val is not None else None
    semester_val = str(semester_val) if semester_val is not None else None

    regulation_obj = _resolve_regulation_obj(student, enrollment)
    regulation_display = getattr(regulation_obj, "year", None) or _safe_get(student, "regulation")

    pattern_options = ExamPattern.objects.none()
    if regulation_obj and degree_id and year_val and semester_val and iat:
        pattern_options = (
            ExamPattern.objects.select_related("regulation", "degree")
            .filter(
                regulation=regulation_obj,
                degree_id=degree_id,
                year=year_val,
                semester=semester_val,
                for_exam=str(iat).strip(),
            )
            .order_by("pattern", "id")
        )

    if (not pattern_id) and enrollment and iat:
        last_row = (
            StudentInternalMark.objects.filter(
                enrollment=enrollment,
                course=enrollment.course,
                exam_name=iat,
                pattern__isnull=False,
            )
            .select_related("pattern")
            .order_by("-created_at", "-id")
            .first()
        )
        if last_row and last_row.pattern_id:
            pattern_id = str(last_row.pattern_id)

    exam_pattern_obj = None
    selected_pid = None

    if pattern_id:
        try:
            selected_pid = int(pattern_id)
        except ValueError:
            selected_pid = None

    if selected_pid:
        try:
            exam_pattern_obj = (
                ExamPattern.objects.select_related("regulation", "degree")
                .prefetch_related(
                    Prefetch(
                        "parts",
                        queryset=Part.objects.prefetch_related(
                            Prefetch(
                                "questions",
                                queryset=Question.objects.prefetch_related("options").order_by("number"),
                                to_attr="prefetched_questions",
                            )
                        ).order_by("name"),
                        to_attr="prefetched_parts",
                    )
                )
                .get(id=selected_pid)
            )
        except ExamPattern.DoesNotExist:
            exam_pattern_obj = None

    pattern_selected = bool(exam_pattern_obj)

    prefilled_marks = {}
    prefilled_co = {}
    prefilled_blooms = {}
    prefilled_subpart_max = {}

    prefilled_marks_by_reg = {}
    prefilled_co_by_reg = {}
    prefilled_blooms_by_reg = {}

    prefilled_absentees = set()
    prefilled_existing_subparts = defaultdict(list)
    part_marks_summary = {}

    if enrollment and iat and exam_pattern_obj:
        if bulk:
            marks_qs = StudentInternalMark.objects.select_related(
                "student", "co_code", "level_code"
            ).filter(
                enrollment__course=enrollment.course,
                batch=enrollment.batch,
                section=enrollment.section,
                exam_name=iat,
                pattern=exam_pattern_obj,
            ).order_by("student__reg_no", "id")

            for row in marks_qs:
                rno = getattr(row.student, "reg_no", None)
                if not rno:
                    continue

                if row.absentee:
                    prefilled_absentees.add(rno)

                part = (row.part_name or "").strip().upper()
                qnum = str(row.question_number or "").strip()
                sub = (row.sub_question or "").strip().lower()
                opt = (row.option_letter or "").strip().lower()

                if sub and opt:
                    suffix = f"{part}_{qnum}_{sub}_{opt}"
                elif sub:
                    suffix = f"{part}_{qnum}_{sub}"
                elif opt:
                    suffix = f"{part}_{qnum}_{opt}"
                else:
                    suffix = f"{part}_{qnum}"

                # Always keep marks student-wise
                prefilled_marks_by_reg.setdefault(rno, {})[f"marks_{suffix}"] = row.marks_obtained

                if row.co_code_id:
                    prefilled_co_by_reg.setdefault(rno, {})[f"co_{suffix}"] = str(row.co_code_id)

                if row.level_code_id:
                    prefilled_blooms_by_reg.setdefault(rno, {})[f"blooms_{suffix}"] = str(row.level_code_id)

                # IMPORTANT:
                # Only non-absent rows should contribute to common prefilled CO/Blooms/Subpart Max/Subpart structure
                if row.absentee:
                    continue

                if row.co_code_id and f"co_{suffix}" not in prefilled_co:
                    prefilled_co[f"co_{suffix}"] = str(row.co_code_id)

                if row.level_code_id and f"blooms_{suffix}" not in prefilled_blooms:
                    prefilled_blooms[f"blooms_{suffix}"] = str(row.level_code_id)

                if sub and opt:
                    group_key = f"{part}__{qnum}__{opt}"
                    sp_key = f"{group_key}__{sub}"

                    if sub not in prefilled_existing_subparts[group_key]:
                        prefilled_existing_subparts[group_key].append(sub)

                    if sp_key not in prefilled_subpart_max and row.max_marks is not None:
                        prefilled_subpart_max[sp_key] = row.max_marks

        else:
            student_marks_qs = StudentInternalMark.objects.select_related(
                "co_code", "level_code"
            ).filter(
                student=student,
                enrollment=enrollment,
                exam_name=iat,
                pattern=exam_pattern_obj,
            )

            for row in student_marks_qs:
                if row.absentee and student:
                    prefilled_absentees.add(student.reg_no)

                part = (row.part_name or "").strip().upper()
                qnum = str(row.question_number or "").strip()
                sub = (row.sub_question or "").strip().lower()
                opt = (row.option_letter or "").strip().lower()

                if sub and opt:
                    suffix = f"{part}_{qnum}_{sub}_{opt}"
                elif sub:
                    suffix = f"{part}_{qnum}_{sub}"
                elif opt:
                    suffix = f"{part}_{qnum}_{opt}"
                else:
                    suffix = f"{part}_{qnum}"

                prefilled_marks[f"marks_{suffix}"] = row.marks_obtained

                if not row.absentee:
                    if row.co_code_id:
                        prefilled_co[f"co_{suffix}"] = str(row.co_code_id)

                    if row.level_code_id:
                        prefilled_blooms[f"blooms_{suffix}"] = str(row.level_code_id)

                    if sub and opt:
                        group_key = f"{part}__{qnum}__{opt}"
                        sp_key = f"{group_key}__{sub}"

                        if sub not in prefilled_existing_subparts[group_key]:
                            prefilled_existing_subparts[group_key].append(sub)

                        if sp_key not in prefilled_subpart_max and row.max_marks is not None:
                            prefilled_subpart_max[sp_key] = row.max_marks

    roman_order = ["i", "ii", "iii", "iv", "v", "vi", "vii", "viii", "ix", "x"]

    def sort_subparts(subparts):
        return sorted(
            subparts,
            key=lambda x: roman_order.index(x) if x in roman_order else 999
        )

    prefilled_existing_subparts = {
        key: sort_subparts(value)
        for key, value in prefilled_existing_subparts.items()
    }

    if exam_pattern_obj:
        parts_list = getattr(exam_pattern_obj, "prefetched_parts", exam_pattern_obj.parts.all())

        for part in parts_list:
            if str(part.name).lower() == "a":
                if part.total_questions and part.max_marks:
                    part_marks_summary[part.name] = f"{part.total_questions} × {part.max_marks} Marks"
                else:
                    part_marks_summary[part.name] = f"{part.max_marks} Marks"
            else:
                questions_list = getattr(part, "prefetched_questions", part.questions.all())

                question_marks = []
                for question in questions_list:
                    q_mark = question.total_marks or part.max_marks or 0
                    if q_mark > 0:
                        question_marks.append(q_mark)

                if question_marks:
                    counts = Counter(question_marks)
                    summary_parts = []

                    for mark, count in sorted(counts.items(), reverse=True):
                        summary_parts.append(f"{count} × {mark} Marks")

                    part_marks_summary[part.name] = ", ".join(summary_parts)
                else:
                    part_marks_summary[part.name] = f"{part.max_marks} Marks"

    course_outcomes = CourseOutcome.objects.none()
    blooms_levels = BloomsLevel.objects.all().order_by("level_code")

    if exam_pattern_obj and exam_pattern_obj.regulation:
        course_outcomes = CourseOutcome.objects.filter(
            regulation=exam_pattern_obj.regulation.year
        ).order_by("co_code")

    context = {
        "bulk": bulk,
        "students_list": students_list,
        "reg_no": reg_no,
        "student_name": student_name,
        "course_code": course_code,
        "course_title": course_title,
        "batch": batch,
        "section": section,
        "iat": iat,
        "exam_name": iat,
        "department_name": department_name,
        "department_code": department_code,
        "pattern_options": pattern_options,
        "selected_pattern_id": selected_pid,
        "pattern_selected": pattern_selected,
        "selected_pattern": getattr(exam_pattern_obj, "pattern", None),
        "selected_regulation": getattr(regulation_obj, "year", None) or regulation_display,
        "selected_degree": degree_label,
        "selected_year": year_val,
        "selected_semester": semester_val,
        "exam_pattern_obj": exam_pattern_obj,
        "exam_pattern_id": getattr(exam_pattern_obj, "id", None),
        "prefilled_marks": prefilled_marks,
        "prefilled_co": prefilled_co,
        "prefilled_blooms": prefilled_blooms,
        "prefilled_marks_by_reg": prefilled_marks_by_reg,
        "prefilled_co_by_reg": prefilled_co_by_reg,
        "prefilled_blooms_by_reg": prefilled_blooms_by_reg,
        "prefilled_subpart_max": prefilled_subpart_max,
        "prefilled_existing_subparts": prefilled_existing_subparts,
        "prefilled_absentees": prefilled_absentees,
        "part_marks_summary": part_marks_summary,
        "course_outcomes": course_outcomes,
        "blooms_levels": blooms_levels,
    }

    return render(request, "faculty_management/enter_iat_marks.html", context)






from io import BytesIO
from collections import defaultdict
import os

from django.http import Http404, HttpResponse
from django.db.models import Prefetch
from django.conf import settings
from django.contrib.staticfiles import finders

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.utils import ImageReader


from io import BytesIO
from collections import defaultdict
import os

from django.http import Http404, HttpResponse
from django.conf import settings
from django.contrib.staticfiles import finders

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.utils import ImageReader

# --- models (adjust app paths if different)
from examination_management.models import (
    StudentInternalMark,
    CourseOutcome,
    BloomsLevel,
)
from student_management.models import StudentDetails
from course_management.models import CourseEnrollment

def student_mark_pdf(request):
    """
    Build a single-student mark statement PDF using StudentInternalMark
    (NOT StudentExam/StudentMark). Respects A/B options by taking the
    option with the highest obtained marks for each question.
    """
    reg_no = request.GET.get("reg_no")
    course_code = request.GET.get("course_code")
    exam_name = request.GET.get("exam_name")

    if not reg_no or not course_code:
        raise Http404("Missing reg_no or course_code")

    # ---- Resolve student + enrollment ----
    student = StudentDetails.objects.filter(reg_no=reg_no).first()
    if not student:
        raise Http404(f"Student with Reg.No {reg_no} not found")

    enroll_qs = CourseEnrollment.objects.filter(
        student_id=student.id,
        course__course_code=course_code,
    )
    # pick most relevant enrollment (enroll=True first, then latest)
    enrollment = enroll_qs.order_by("-enroll", "-id").select_related("course").first()
    if not enrollment:
        raise Http404(f"No enrollment for {reg_no} in {course_code}")

    # ---- Pull all rows for this sitting from StudentInternalMark ----
    sim_filters = {
        "student": student,
        "enrollment": enrollment,
    }
    if exam_name:
        sim_filters["exam_name"] = exam_name

    rows = list(
        StudentInternalMark.objects.select_related(
            "co_code", "level_code", "enrollment__course"
        )
        .filter(**sim_filters)
        .order_by("created_at")
    )
    if not rows:
        raise Http404("No internal mark rows found for the given parameters")

    # Meta for header/details (prefer denorms, fall back to FK objects)
    # Department could be from student.department (custom model) if available
    dept_obj = getattr(student, "department", None)
    department_code = (
        getattr(dept_obj, "Department_code", None)
        or getattr(dept_obj, "dept_code", None)
        or ""
    )
    department_name = (
        getattr(dept_obj, "Department", None) or getattr(dept_obj, "name", None) or ""
    )
    course_obj = getattr(enrollment, "course", None)
    course_title = (
        getattr(course_obj, "title", None) or getattr(course_obj, "title", None) or ""
    )
    batch = rows[0].batch or enrollment.batch or ""
    section = rows[0].section or enrollment.section or ""
    # exam_name might be absent in querystring; infer from first row if missing
    exam_name = exam_name or (rows[0].exam_name or "")

    # ---------- Group by (Part, Question) then by Option ----------
    # question_map[(part, qnum)][opt] -> {'max':, 'obt':, 'rows':[SIM rows]}
    question_map = defaultdict(
        lambda: defaultdict(lambda: {"max": 0, "obt": 0, "rows": []})
    )
    for m in rows:
        part = m.part_name or ""
        qnum = m.question_number or ""
        opt = m.option_letter or ""  # '' => non-choice
        question_map[(part, qnum)][opt]["max"] += int(m.max_marks or 0)
        question_map[(part, qnum)][opt]["obt"] += int(m.marks_obtained or 0)
        question_map[(part, qnum)][opt]["rows"].append(m)

    # ---------- Choose effective rows (respect A/B) & build display ----------
    display_rows = []  # attempted rows only (for the big table)
    chosen_rows = []  # ALL rows from chosen path (for accurate CO max/obt)

    for (_part, _qnum), options in question_map.items():
        if "" in options and len(options) == 1:
            agg = options[""]
            chosen_rows.extend(agg["rows"])
            for row in agg["rows"]:
                if int(row.marks_obtained or 0) > 0:
                    display_rows.append(row)
        else:
            chosen_key = max(options.keys(), key=lambda k: options[k]["obt"])
            chosen = options[chosen_key]
            chosen_rows.extend(chosen["rows"])
            for row in chosen["rows"]:
                if int(row.marks_obtained or 0) > 0:
                    display_rows.append(row)

    # ---------- Build PDF ----------
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=44 * mm,
        bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Totals", parent=styles["Heading3"], spaceBefore=6))
    styles.add(
        ParagraphStyle(
            name="MetaLabel", parent=styles["Normal"], fontName="Helvetica-Bold"
        )
    )
    styles.add(ParagraphStyle(name="MetaValue", parent=styles["Normal"]))

    story = []

    # ---------- Student details ----------
    details_data = [
        [
            Paragraph("Student", styles["MetaLabel"]),
            Paragraph(f"{student.name or ''} ({student.reg_no})", styles["MetaValue"]),
        ],
    ]
    if batch and section:
        details_data.append(
            [
                Paragraph("Batch-Section", styles["MetaLabel"]),
                Paragraph(f"{batch} - {section}", styles["MetaValue"]),
            ]
        )
    elif batch:
        details_data.append(
            [
                Paragraph("Batch", styles["MetaLabel"]),
                Paragraph(f"{batch}", styles["MetaValue"]),
            ]
        )
    elif section:
        details_data.append(
            [
                Paragraph("Section", styles["MetaLabel"]),
                Paragraph(f"{section}", styles["MetaValue"]),
            ]
        )

    if exam_name:
        details_data.append(
            [
                Paragraph("Exam", styles["MetaLabel"]),
                Paragraph(f"{exam_name}", styles["MetaValue"]),
            ]
        )

    details_table = Table(details_data, colWidths=[30 * mm, None], hAlign="LEFT")
    details_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(details_table)
    story.append(Spacer(1, 6))

    # ---------- Main marks table (attempted rows only) ----------
    data = [["Part", "Qn", "Opt", "Sub", "CO", "Bloom", "Max", "Obtained"]]

    def _safe(s):
        return s or ""

    def _safe_int(v, default=9999):
        try:
            return int(v)
        except (TypeError, ValueError):
            return default

    def _sub_order(s):  # i < ii < others
        return {"i": 1, "ii": 2}.get((s or "").lower(), 99)

    def _part_order(p):
        p = (p or "").strip().upper()
        return {"A": 1, "B": 2, "C": 3, "D": 4}.get(p, 99)

    display_rows.sort(
        key=lambda r: (
            _part_order(r.part_name),
            _safe_int(r.question_number),
            _safe(r.option_letter or "").lower(),
            _sub_order(r.sub_question),
        )
    )

    for m in display_rows:
        co = getattr(m.co_code, "co_code", "") if m.co_code_id else ""
        bl = getattr(m.level_code, "level_code", "") if m.level_code_id else ""
        data.append(
            [
                m.part_name or "",
                m.question_number or "",
                m.option_letter or "",
                m.sub_question or "",
                co,
                bl,
                str(m.max_marks),
                str(m.marks_obtained),
            ]
        )

    if len(data) == 1:
        data.append(["—", "—", "—", "—", "—", "—", "—", "0"])

    col_widths = [
        16 * mm,
        14 * mm,
        14 * mm,
        14 * mm,
        26 * mm,
        26 * mm,
        18 * mm,
        22 * mm,
    ]
    main_table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    main_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (6, 1), (7, -1), "RIGHT"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f9f9f9")],
                ),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    story.append(main_table)
    story.append(Spacer(1, 8))

    # ---------- CO Summary (bottom-right): CO | Max | Obtained ----------
    co_totals = defaultdict(lambda: {"obt": 0, "max": 0})
    for r in chosen_rows:
        co_code = getattr(r.co_code, "co_code", "") if r.co_code_id else "—"
        co_totals[co_code]["obt"] += int(r.marks_obtained or 0)
        co_totals[co_code]["max"] += int(r.max_marks or 0)

    co_total_max = sum(v["max"] for k, v in co_totals.items() if k != "—")
    co_total_obt = sum(v["obt"] for k, v in co_totals.items() if k != "—")

    story.append(Paragraph(f"Total: {co_total_obt} / {co_total_max}", styles["Totals"]))

    def _co_sort_key(k: str):
        import re

        m = re.match(r"^CO\s*[-_]?(\d+)$", k.strip(), re.IGNORECASE)
        if m:
            return (0, int(m.group(1)))
        if k == "—":
            return (2, 0)
        return (1, k.lower())

    co_rows = [["CO", "Max", "Obtained"]]
    for co_code in sorted(co_totals.keys(), key=_co_sort_key):
        if co_code == "—":   # skip rows without CO mapping
            continue
        co_rows.append(
            [
                co_code,
                str(co_totals[co_code]["max"]),
                str(co_totals[co_code]["obt"]),
            ]
        )
    if len(co_rows) > 1:
        co_rows.append(["TOTAL", str(co_total_max), str(co_total_obt)])
        co_table = Table(co_rows, colWidths=[24 * mm, 20 * mm, 20 * mm], hAlign="RIGHT")
        co_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                ]
            )
        )
        story.append(Spacer(1, 6))
        story.append(co_table)

    # ---------- Header (logo + college + address + title + Dept/Course) ----------
    def _on_page(c, _doc):
        c.saveState()
        page_w, page_h = A4
        left = 18 * mm
        right = 18 * mm

        # Find logo
        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand
            if not logo_path:
                for d in getattr(settings, "STATICFILES_DIRS", []):
                    cand = os.path.join(d, logo_rel)
                    if os.path.exists(cand):
                        logo_path = cand
                        break

        if logo_path and os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                target_h = 18 * mm
                target_w = target_h * (iw / float(ih))
                c.drawImage(
                    img,
                    left,
                    page_h - (target_h + 8 * mm),
                    width=target_w,
                    height=target_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass

        # College name & address
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(
            page_w / 2.0, page_h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY"
        )

        c.setFont("Helvetica", 10)
        c.drawCentredString(page_w / 2.0, page_h - 15 * mm, "Rajapalayam - 626117")
        c.drawCentredString(
            page_w / 2.0, page_h - 19 * mm, "Affiliated to Anna University, Chennai"
        )

        # Title
        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(page_w / 2.0, page_h - 26 * mm, "Student Mark Statement")

        # Department & Course
        c.setFont("Helvetica", 10)
        dept_line = f"{department_code} — {department_name}".strip(" —")
        course_line = f"{course_code} — {course_title}".strip(" —")
        c.drawCentredString(page_w / 2.0, page_h - 31 * mm, dept_line)
        c.drawCentredString(page_w / 2.0, page_h - 35 * mm, course_line)

        # Rule
        rule_y = page_h - 40 * mm
        c.setStrokeColor(colors.black)
        c.setLineWidth(0.5)
        c.line(left, rule_y, page_w - right, rule_y)

        # Footer page number
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.grey)
        c.drawRightString(page_w - right, 12 * mm, f"Page {c.getPageNumber()}")
        c.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buf.getvalue()
    buf.close()

    fname = f"{student.reg_no}_{course_code}_{exam_name or 'exam'}_marks.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{fname}"'
    return resp





from io import BytesIO
from collections import defaultdict
import os
import re

from django.http import Http404, HttpResponse
from django.db.models import Prefetch
from django.conf import settings
from django.contrib.staticfiles import finders

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.utils import ImageReader

# from .models import StudentExam, StudentMark

def student_mark_pdf_all(request):
    """
    One PDF with:
    1) Main roster table
    2) Summary table below it:
       How many students scored:
       - < 60
       - 60-69
       - 70-89
       - >= 90

    Main table format:
    Name | Reg No | CO1 | CO1(100 stacked) | CO2 | CO2(100 stacked) | ... | TOTAL | TOTAL(100 stacked)

    - A/B choice respected per (Part, Question): pick option with highest obtained.
    - Untagged ('—' / NULL) CO rows are excluded from CO and TOTAL.
    - Auto-fits: switches to landscape if needed, scales columns, reduces font/padding.
    """
    import os
    import re
    from io import BytesIO
    from collections import defaultdict

    from django.http import Http404, HttpResponse
    from django.conf import settings
    from django.contrib.staticfiles import finders

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.utils import ImageReader

    course_code = request.GET.get("course_code")
    exam_name = request.GET.get("exam_name")
    batch = request.GET.get("batch") or None
    section = request.GET.get("section") or None
    department_id = request.GET.get("departmentId") or None

    print("department_id => ", department_id)

    if not course_code or not exam_name:
        raise Http404("Missing course_code or exam_name")

    # ------------ Pull all rows from StudentInternalMark for this sitting ------------
    sim_qs = (
        StudentInternalMark.objects.select_related(
            "student",
            "enrollment__course",
            "co_code",
            "level_code",
            "student__department",
        )
        .filter(course_code=course_code, exam_name=exam_name)
        .order_by("student__reg_no", "created_at")
    )

    if batch:
        sim_qs = sim_qs.filter(batch=batch)
    if department_id:
        sim_qs = sim_qs.filter(student__department_id=department_id)

    effective_section = section
    section_filtered_qs = sim_qs.filter(section=section) if section else sim_qs
    rows = list(section_filtered_qs)

    if not rows and section:
        available_sections = [
            (sec or "").strip()
            for sec in sim_qs.values_list("section", flat=True).distinct()
            if (sec or "").strip()
        ]
        if len(available_sections) == 1:
            effective_section = available_sections[0]
            rows = list(sim_qs.filter(section=effective_section))

    if not rows:
        raise Http404("No internal mark rows found for the given parameters")

    # ------------ Header/footer meta ------------
    first = rows[0]
    student0 = first.student
    dept_obj = getattr(student0, "department", None)

    department_code = (
        getattr(dept_obj, "Department_code", None)
        or getattr(dept_obj, "dept_code", None)
        or getattr(dept_obj, "code", None)
        or ""
    )
    department_name = (
        getattr(dept_obj, "Department", None)
        or getattr(dept_obj, "name", None)
        or ""
    )

    course_title = ""
    if first.enrollment and getattr(first.enrollment, "course", None):
        course_title = getattr(first.enrollment.course, "title", None) or ""

    header_meta = {
        "department_code": department_code,
        "department_name": department_name,
        "course_code": first.course_code or course_code,
        "course_title": course_title,
        "exam_name": exam_name,
        "batch": batch if batch is not None else (first.batch or ""),
        "section": effective_section if effective_section is not None else (first.section or ""),
    }

    # ------------ Helpers ------------
    def _co_sort_key(k: str):
        m = re.match(r"^CO\s*[-_]?(\d+)$", k.strip(), re.IGNORECASE)
        if m:
            return (0, int(m.group(1)))
        return (1, k.lower())

    def _to100(obt: int, mx: int) -> str:
        if not mx:
            return "0"
        return f"{round((obt * 100.0) / mx):.0f}"

    def _to100_int(obt: int, mx: int) -> int:
        if not mx:
            return 0
        return int(round((obt * 100.0) / mx))

    def _band_label(score_100: int) -> str:
        if score_100 < 50:
            return "<50"
        elif 50 <= score_100 <= 59:
            return "50-59"
        elif 60 <= score_100 <= 69:
            return "60-69"
        elif 70 <= score_100 <= 89:
            return "70-89"
        return ">=90"

    # ------------ Per-student processing ------------
    by_student = defaultdict(list)
    for r in rows:
        by_student[r.student_id].append(r)

    all_co_labels = set()
    per_student = []  # [{name, reg_no, cos: {CO1:{obt,max}}, total:(obt,max)}]

    for sid, srows in by_student.items():
        stu = srows[0].student
        stu_name = getattr(stu, "name", "") or ""
        reg_no_val = getattr(stu, "reg_no", "") or (srows[0].reg_no or "")

        # Choose best option per question
        question_map = defaultdict(lambda: defaultdict(lambda: {"obt": 0, "rows": []}))
        for m in srows:
            qkey = (m.part_name or "", m.question_number or "")
            okey = m.option_letter or ""
            question_map[qkey][okey]["obt"] += int(m.marks_obtained or 0)
            question_map[qkey][okey]["rows"].append(m)

        chosen_rows = []
        for (_part, _qnum), options in question_map.items():
            if "" in options and len(options) == 1:
                chosen_rows.extend(options[""]["rows"])
            else:
                key = max(options.keys(), key=lambda k: options[k]["obt"])
                chosen_rows.extend(options[key]["rows"])

        # Roll up CO totals
        co_totals = defaultdict(lambda: {"obt": 0, "max": 0})
        for r in chosen_rows:
            co_code = getattr(r.co_code, "co_code", "") if r.co_code_id else "—"
            if not co_code or co_code.strip() == "—":
                continue
            co_totals[co_code]["obt"] += int(r.marks_obtained or 0)
            co_totals[co_code]["max"] += int(r.max_marks or 0)

        all_co_labels.update(co_totals.keys())
        total_obt = sum(v["obt"] for v in co_totals.values())
        total_max = sum(v["max"] for v in co_totals.values())

        per_student.append(
            {
                "name": stu_name,
                "reg_no": reg_no_val,
                "cos": co_totals,
                "total": (total_obt, total_max),
            }
        )

    co_labels = sorted(all_co_labels, key=_co_sort_key)
    per_student = sorted(per_student, key=lambda x: x["reg_no"])

    # ------------ Build main table widths ------------
    sl_w = 16 * mm
    name_w = 52 * mm
    reg_w = 32 * mm
    co_raw_w = 18 * mm
    co_100_w = 16 * mm
    total_raw_w = 22 * mm
    total_100_w = 18 * mm

    base_widths = [sl_w, name_w, reg_w]
    for _ in co_labels:
        base_widths.extend([co_raw_w, co_100_w])
    base_widths.extend([total_raw_w, total_100_w])

    use_landscape = len(base_widths) > 12
    page_size = landscape(A4) if use_landscape else A4
    left_margin = 12 * mm
    right_margin = 12 * mm
    top_margin = 44 * mm
    bottom_margin = 16 * mm

    avail_width = page_size[0] - left_margin - right_margin
    total_base = sum(base_widths)

    if total_base > avail_width:
        scale = avail_width / float(total_base)
        col_widths = [w * scale for w in base_widths]
    else:
        col_widths = base_widths
        scale = 1.0

    if total_base > avail_width:
        if scale < 0.60:
            header_font = 8
            body_font = 7
            pad_top = 2
            pad_bottom = 2
        elif scale < 0.80:
            header_font = 9
            body_font = 8
            pad_top = 2.5
            pad_bottom = 2.5
        else:
            header_font = 10
            body_font = 9
            pad_top = 3
            pad_bottom = 3
    else:
        header_font = 10
        body_font = 9
        pad_top = 3
        pad_bottom = 3

    # ------------ Build PDF ------------
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
    )

    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="SmallMuted",
            parent=styles["Normal"],
            fontSize=max(7, body_font - 1),
            textColor=colors.grey,
        )
    )
    styles.add(
        ParagraphStyle(
            name="HeaderLeft",
            parent=styles["Normal"],
            alignment=0,
            fontSize=header_font,
            leading=header_font + 2,
        )
    )
    styles.add(
        ParagraphStyle(
            name="HeaderCenter",
            parent=styles["Normal"],
            alignment=1,
            fontSize=header_font,
            leading=header_font + 2,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SectionTitle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=max(9, body_font + 1),
            leading=max(11, body_font + 3),
            spaceAfter=4,
        )
    )

    story = []

    info_bits = []
    if header_meta["exam_name"]:
        info_bits.append(f"Assessment: {header_meta['exam_name']}")
    if header_meta["batch"]:
        info_bits.append(f"Batch: {header_meta['batch']}")
    if header_meta["section"]:
        info_bits.append(f"Section: {header_meta['section']}")
    if info_bits:
        story.append(Paragraph(" | ".join(info_bits), styles["SmallMuted"]))
        story.append(Spacer(1, 4))

    # ------------ Main table header ------------
    header = [
        Paragraph("Sl No", styles["HeaderCenter"]),
        Paragraph("Name", styles["HeaderLeft"]),
        Paragraph("Reg No", styles["HeaderLeft"]),
    ]
    for co in co_labels:
        header.append(Paragraph(f"{co}", styles["HeaderCenter"]))
        header.append(Paragraph(f"{co}<br/>(100)", styles["HeaderCenter"]))
    header.append(Paragraph("TOTAL", styles["HeaderCenter"]))
    header.append(Paragraph("TOTAL<br/>(100)", styles["HeaderCenter"]))

    data = [header]

    # ------------ Main table rows ------------
    for idx, row in enumerate(per_student, start=1):
        name = row["name"] or ""
        regno = row["reg_no"] or ""
        cells = [str(idx), name, regno]
        cos = row["cos"]

        for co in co_labels:
            if co in cos:
                obt = cos[co]["obt"]
                mx = cos[co]["max"]
                cells.append(f"{obt}/{mx}")
                cells.append(_to100(obt, mx))
            else:
                cells.append("0/0")
                cells.append("0")

        tot_obt, tot_max = row["total"]
        cells.append(f"{tot_obt}/{tot_max}")
        cells.append(_to100(tot_obt, tot_max))
        data.append(cells)

    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT", splitByRow=1)
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), header_font),
                ("FONTSIZE", (0, 1), (-1, -1), body_font),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (1, 0), (1, -1), "LEFT"),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9f9f9")]),
                ("TOPPADDING", (0, 0), (-1, -1), pad_top),
                ("BOTTOMPADDING", (0, 0), (-1, -1), pad_bottom),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    story.append(table)

    # ------------ Summary counts table ------------
    # ------------ Summary counts table (TOTAL only) ------------
    band_keys = ["<50", "50-59", "60-69", "70-89", ">=90"]

    summary_counts = {k: 0 for k in band_keys}

    for row in per_student:
        tot_obt, tot_max = row["total"]
        total_pct = _to100_int(tot_obt, tot_max)
        summary_counts[_band_label(total_pct)] += 1

    story.append(Spacer(1, 10))
    story.append(Paragraph("Performance Summary", styles["SectionTitle"]))
    story.append(Spacer(1, 3))

    summary_data = [
        [
            Paragraph("Range", styles["HeaderCenter"]),
            Paragraph("Total Students", styles["HeaderCenter"]),
        ]
    ]

    for band in band_keys:
        summary_data.append(
            [
                Paragraph(band, styles["HeaderCenter"]),
                Paragraph(str(summary_counts[band]), styles["HeaderCenter"]),
            ]
        )

    summary_col_widths = [50 * mm, 50 * mm]

    summary_table = Table(
        summary_data,
        colWidths=summary_col_widths,
        repeatRows=1,
        hAlign="LEFT",
        splitByRow=1,
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), header_font),
                ("FONTSIZE", (0, 1), (-1, -1), body_font),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eeeeee")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f8f8")]),
                ("TOPPADDING", (0, 0), (-1, -1), pad_top),
                ("BOTTOMPADDING", (0, 0), (-1, -1), pad_bottom),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(summary_table)

    # ------------ Header/footer drawing ------------
    def _on_page(c, _doc):
        c.saveState()
        page_w, page_h = page_size
        left = left_margin
        right = right_margin

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                candidate = os.path.join(static_root, logo_rel)
                if os.path.exists(candidate):
                    logo_path = candidate
            if not logo_path:
                for dir_ in getattr(settings, "STATICFILES_DIRS", []):
                    candidate = os.path.join(dir_, logo_rel)
                    if os.path.exists(candidate):
                        logo_path = candidate

        if logo_path and os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                target_h = 18 * mm
                target_w = target_h * (iw / float(ih))
                c.drawImage(
                    img,
                    left,
                    page_h - (target_h + 8 * mm),
                    width=target_w,
                    height=target_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass

        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 16)
        c.drawCentredString(page_w / 2.0, page_h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        c.setFont("Helvetica", 10)
        c.drawCentredString(page_w / 2.0, page_h - 15 * mm, "Rajapalayam - 626117")
        c.drawCentredString(page_w / 2.0, page_h - 19 * mm, "Affiliated to Anna University, Chennai")

        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(page_w / 2.0, page_h - 26 * mm, "Student CO Summary")

        c.setFont("Helvetica", 10)
        dept_line = f"{header_meta['department_code']} — {header_meta['department_name']}"
        course_line = f"{header_meta['course_code']} — {header_meta['course_title']}"
        c.drawCentredString(page_w / 2.0, page_h - 31 * mm, dept_line)
        c.drawCentredString(page_w / 2.0, page_h - 35 * mm, course_line)

        rule_y = page_h - 40 * mm
        c.setStrokeColor(colors.black)
        c.setLineWidth(0.5)
        c.line(left, rule_y, page_w - right, rule_y)

        c.setFont("Helvetica", 8)
        c.setFillColor(colors.grey)
        c.drawRightString(page_w - right, 12 * mm, f"Page {c.getPageNumber()}")
        c.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buf.getvalue()
    buf.close()

    fname = f"{header_meta['course_code']}_{exam_name}_CO_summary.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{fname}"'
    return resp


import json
# --------------------------
# Main view
# --------------------------


@check_permission("student_document_verification")
@student_management
def student_document_verification(request):
    return render(
        request, "student_management/faculty/student_document_verification.html"
    )


def load_years(request):
    regulation = request.GET.get("regulation")
    years = (
        ExamPattern.objects.filter(regulation__year=regulation)
        .values_list("year", flat=True)
        .distinct()
    )
    return JsonResponse(list(years), safe=False)


def load_semesters(request):
    regulation = request.GET.get("regulation")
    year = request.GET.get("year")
    semesters = (
        ExamPattern.objects.filter(regulation__year=regulation, year=year)
        .values_list("semester", flat=True)
        .distinct()
    )
    return JsonResponse(list(semesters), safe=False)


def load_academic_years(request):
    regulation = request.GET.get("regulation")
    year = request.GET.get("year")
    semester = request.GET.get("semester")
    academic_years = (
        ExamPattern.objects.filter(
            regulation__year=regulation, year=year, semester=semester
        )
        .values_list("academic_year", flat=True)
        .distinct()
    )
    return JsonResponse(list(academic_years), safe=False)


def load_patterns(request):
    regulation = request.GET.get("regulation")
    year = request.GET.get("year")
    semester = request.GET.get("semester")
    academic_year = request.GET.get("academic_year")
    patterns = (
        ExamPattern.objects.filter(
            regulation__year=regulation,
            year=year,
            semester=semester,
            academic_year=academic_year,
        )
        .values_list("pattern", flat=True)
        .distinct()
    )
    return JsonResponse(list(patterns), safe=False)


from django.db import transaction


# from django.db import transaction
from django.contrib import messages
from django.shortcuts import redirect, get_object_or_404
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import Q





from django.contrib import messages
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect

# @transaction.atomic
# def save_student_marks(request):
#     if request.method != "POST":
#         messages.error(request, "Invalid request method!")
#         return redirect("/")

#     # -------------------------------------------------
#     # Common meta
#     # -------------------------------------------------
#     department_code = request.POST.get("department_code")
#     department_name = request.POST.get("department_name")
#     course_code = request.POST.get("course_code")
#     course_title = request.POST.get("course_title")
#     batch = request.POST.get("batch")
#     section = request.POST.get("section")

#     exam_name = (
#         request.POST.get("iat")
#         or request.POST.get("exam_name")
#         or request.GET.get("iat")
#     )

#     pattern_id_str = request.POST.get("pattern_id") or request.POST.get("exam_pattern_id")
#     pattern_id = int(pattern_id_str) if (pattern_id_str and pattern_id_str.isdigit()) else None

#     exam_pattern = None
#     if pattern_id:
#         exam_pattern = get_object_or_404(ExamPattern, id=pattern_id)

#     # -------------------------------------------------
#     # GLOBAL CO / BLOOMS MAPPING
#     # -------------------------------------------------
#     co_map = {}
#     blooms_map = {}

#     for key in request.POST.keys():
#         if key.startswith("co_map_"):
#             suffix = key.replace("co_map_", "").replace("[]", "")
#             values = request.POST.getlist(key)
#             co_map[suffix] = [int(v) for v in values if str(v).isdigit()]

#         if key.startswith("blooms_map_"):
#             suffix = key.replace("blooms_map_", "").replace("[]", "")
#             values = request.POST.getlist(key)
#             blooms_map[suffix] = [int(v) for v in values if str(v).isdigit()]

#     # -------------------------------------------------
#     # GLOBAL SUBPART MAX CONFIGURATION
#     # Reads your "Sub-part Maximum Marks Configuration"
#     # input names like:
#     # subpart_max__B__12__a__i = 6
#     # subpart_max__B__12__a__ii = 7
#     # -------------------------------------------------
#     global_subpart_max = {}

#     for key, value in request.POST.items():
#         if key.startswith("subpart_max__"):
#             parts = key.split("__")
#             # expected: subpart_max__B__12__a__i
#             if len(parts) == 5:
#                 _, part_name, question_number, option_letter, sub_question = parts
#                 normalized_key = f"max_{part_name}_{question_number}_{sub_question}_{option_letter}"
#                 global_subpart_max[normalized_key] = value

#     # -------------------------------------------------
#     # Detect BULK vs SINGLE
#     # -------------------------------------------------
#         bulk_students = request.POST.getlist("students[]")
#         single_reg_no = request.POST.get("reg_no")
#         save_reg_no = request.POST.get("save_reg_no")  # clicked student-wise save button

#         if save_reg_no:
#             bulk_students = [save_reg_no]
#             is_bulk = False
#             single_reg_no = save_reg_no
#         else:
#             is_bulk = bool(bulk_students)

#         if not is_bulk and not single_reg_no:
#             messages.error(request, "No student reg_no provided.")
#             return redirect(request.META.get("HTTP_REFERER", "/"))

#     # -------------------------------------------------
#     # Helpers
#     # -------------------------------------------------
#     def safe_int(value, default=0):
#         try:
#             value = str(value).strip()
#             if value == "":
#                 return default
#             return int(float(value))
#         except Exception:
#             return default

#     def parse_marks_key(key: str):
#         """
#         Expected formats after stripping bulk prefix:
#             marks_A_1
#             marks_B_1_i_a
#             marks_B_1_ii_a
#             marks_C_2_a
#         """
#         parts = key.split("_")
#         if len(parts) < 3 or parts[0] != "marks":
#             return None

#         part_name = parts[1]
#         question_number = parts[2]
#         sub_question = parts[3] if len(parts) > 3 else None
#         option_letter = parts[4] if len(parts) > 4 else None

#         return {
#             "part_name": part_name,
#             "question_number": question_number,
#             "sub_question": sub_question,
#             "option_letter": option_letter,
#         }

#     def build_max_key(part_name, question_number, sub_question=None, option_letter=None):
#         if sub_question and option_letter:
#             return f"max_{part_name}_{question_number}_{sub_question}_{option_letter}"
#         if sub_question:
#             return f"max_{part_name}_{question_number}_{sub_question}"
#         if option_letter:
#             return f"max_{part_name}_{question_number}_{option_letter}"
#         return f"max_{part_name}_{question_number}"

#     def build_suffix(part_name, question_number, sub_question=None, option_letter=None):
#         if sub_question and option_letter:
#             return f"{part_name}_{question_number}_{sub_question}_{option_letter}"
#         if sub_question:
#             return f"{part_name}_{question_number}_{sub_question}"
#         if option_letter:
#             return f"{part_name}_{question_number}_{option_letter}"
#         return f"{part_name}_{question_number}"

#     def get_part_default_max(part_name):
#         if not exam_pattern:
#             return 0
#         part_obj = exam_pattern.parts.filter(name=part_name).first()
#         if not part_obj:
#             return 0
#         return safe_int(getattr(part_obj, "max_marks", 0), 0)

#     def resolve_max_marks(part_name, question_number, sub_question=None, option_letter=None, max_data=None):
#         """
#         Priority:
#         1. Sub-part Maximum Marks Configuration (global_subpart_max)
#         2. Hidden posted max_data (student-specific hidden fields)
#         3. option/question level max
#         4. exam pattern part max
#         """
#         max_data = max_data or {}

#         # 1) exact subpart override from config section
#         if sub_question and option_letter:
#             k = build_max_key(part_name, question_number, sub_question, option_letter)
#             val = safe_int(global_subpart_max.get(k), 0)
#             if val > 0:
#                 return val

#         # 2) exact hidden max field from row
#         exact_key = build_max_key(part_name, question_number, sub_question, option_letter)
#         val = safe_int(max_data.get(exact_key), 0)
#         if val > 0:
#             return val

#         # 3) option-level max
#         if option_letter:
#             option_key = build_max_key(part_name, question_number, option_letter=option_letter)
#             val = safe_int(max_data.get(option_key), 0)
#             if val > 0:
#                 return val

#         # 4) sub-question only max
#         if sub_question:
#             sub_key = build_max_key(part_name, question_number, sub_question=sub_question)
#             val = safe_int(max_data.get(sub_key), 0)
#             if val > 0:
#                 return val

#         # 5) question-level max
#         q_key = build_max_key(part_name, question_number)
#         val = safe_int(max_data.get(q_key), 0)
#         if val > 0:
#             return val

#         # 6) fallback to part max
#         return get_part_default_max(part_name)

#     # -------------------------------------------------
#     # Per-student parser + saver
#     # -------------------------------------------------
#     def parse_and_save_one_student(reg_no: str):
#         if not reg_no:
#             return 0

#         student = StudentDetails.objects.filter(reg_no=reg_no).first()
#         if not student:
#             raise ValueError(f"Student not found: {reg_no}")
#         student_semester = getattr(student, "semester", None)
#         student_degree = getattr(student.department, "degree", None) if student.department else None
#         student_department = getattr(student, "department", None)
        

#         enrollment_qs = CourseEnrollment.objects.filter(
#             student_id=student.id,
#             course__course_code=course_code,
#         )

#         if batch:
#             enrollment_qs = enrollment_qs.filter(batch=batch)

#         if section:
#             enrollment_qs = enrollment_qs.filter(section=section)

#         enrollment = enrollment_qs.order_by("-enroll", "-id").first()
#         if not enrollment:
#             raise ValueError(f"Course enrollment not found for {reg_no}")

#         # ---------------------------------------------
#         # Get marks/max data for this student
#         # ---------------------------------------------
#         bulk_prefix = f"m__{reg_no}__"
#         is_bulk_row = any(k.startswith(bulk_prefix + "marks_") for k in request.POST.keys())

#         if is_bulk_row:
#             marks_data = {
#                 k.replace(bulk_prefix, ""): v
#                 for k, v in request.POST.items()
#                 if k.startswith(bulk_prefix + "marks_")
#             }
#             max_data = {
#                 k.replace(bulk_prefix, ""): v
#                 for k, v in request.POST.items()
#                 if k.startswith(bulk_prefix + "max_")
#             }
#         else:
#             marks_data = {k: v for k, v in request.POST.items() if k.startswith("marks_")}
#             max_data = {k: v for k, v in request.POST.items() if k.startswith("max_")}

#         # ---------------------------------------------
#         # Group by question
#         # ---------------------------------------------
#         grouped = {}
#         for key, value in marks_data.items():
#             parsed = parse_marks_key(key)
#             if not parsed:
#                 continue

#             group_key = (parsed["part_name"], parsed["question_number"])
#             grouped.setdefault(group_key, []).append(
#                 {
#                     "raw_key": key,
#                     "raw_value": value,
#                     "part_name": parsed["part_name"],
#                     "question_number": parsed["question_number"],
#                     "sub_question": parsed["sub_question"],
#                     "option_letter": parsed["option_letter"],
#                 }
#             )

#         upsert_count = 0

#         for (part_name, question_number), entries in grouped.items():
#             selected_option = None
#             pair_totals = {}

#             for entry in entries:
#                 sub_question = entry["sub_question"]
#                 option_letter = entry["option_letter"]
#                 marks_obt = safe_int(entry["raw_value"], 0)

#                 if option_letter and marks_obt > 0:
#                     selected_option = option_letter

#                 if sub_question in ("i", "ii") and option_letter:
#                     pair_totals.setdefault(option_letter, 0)
#                     pair_totals[option_letter] += marks_obt

#             non_zero_options = set()
#             for entry in entries:
#                 option_letter = entry["option_letter"]
#                 marks_obt = safe_int(entry["raw_value"], 0)
#                 if option_letter and marks_obt > 0:
#                     non_zero_options.add(option_letter)

#             if len(non_zero_options) > 1:
#                 raise ValueError(
#                     f"{reg_no} - {part_name}{question_number}: Enter marks for only one option."
#                 )

#             # validate I + II <= total allowed max
#             for option_letter, total in pair_totals.items():
#                 max_i = resolve_max_marks(
#                     part_name=part_name,
#                     question_number=question_number,
#                     sub_question="i",
#                     option_letter=option_letter,
#                     max_data=max_data,
#                 )
#                 max_ii = resolve_max_marks(
#                     part_name=part_name,
#                     question_number=question_number,
#                     sub_question="ii",
#                     option_letter=option_letter,
#                     max_data=max_data,
#                 )
#                 allowed_max = max_i + max_ii

#                 if allowed_max > 0 and total > allowed_max:
#                     raise ValueError(
#                         f"{reg_no} - {part_name}{question_number}{option_letter}: "
#                         f"Sub I + Sub II cannot exceed {allowed_max}. "
#                         f"You entered total {total}."
#                     )

#             # ---------------------------------------------
#             # Save entries
#             # ---------------------------------------------
#             for entry in entries:
#                 sub_question = entry["sub_question"]
#                 option_letter = entry["option_letter"]
#                 marks_obt = safe_int(entry["raw_value"], 0)

#                 if selected_option and option_letter and option_letter != selected_option:
#                     marks_obt = 0

#                 max_marks = resolve_max_marks(
#                     part_name=part_name,
#                     question_number=question_number,
#                     sub_question=sub_question,
#                     option_letter=option_letter,
#                     max_data=max_data,
#                 )

#                 if max_marks > 0 and marks_obt > max_marks:
#                     raise ValueError(
#                         f"{reg_no} - {part_name}{question_number}"
#                         f"{option_letter or ''}{('-' + sub_question.upper()) if sub_question else ''}: "
#                         f"Marks cannot exceed {max_marks}."
#                     )

#                 suffix = build_suffix(
#                     part_name=part_name,
#                     question_number=question_number,
#                     sub_question=sub_question,
#                     option_letter=option_letter,
#                 )

#                 co_values = co_map.get(suffix, [])
#                 bl_values = blooms_map.get(suffix, [])

#                 co_id = co_values[0] if co_values else None
#                 bl_id = bl_values[0] if bl_values else None

#                 StudentInternalMark.objects.update_or_create(
#                     student=student,
#                     enrollment=enrollment,
#                     exam_name=exam_name,
#                     course=enrollment.course,
#                     part_name=part_name,
#                     question_number=question_number,
#                     sub_question=sub_question,
#                     option_letter=option_letter,
#                     defaults={
#                         "pattern": exam_pattern,
#                         "semester": student_semester, 
#                         "degree": student_degree,
#                         "department": student_department,
#                         "max_marks": max_marks,
#                         "marks_obtained": marks_obt,
#                         "co_code_id": co_id,
#                         "level_code_id": bl_id,
#                         "reg_no": reg_no,
#                         "course_code": course_code,
#                         "batch": batch or enrollment.batch,
#                         "section": section or enrollment.section,
#                     },
#                 )
#                 upsert_count += 1

#         return upsert_count

#     # -------------------------------------------------
#     # Execute save
#     # -------------------------------------------------
#     try:
#         total_upserts = 0

#         if save_reg_no:
#             total_upserts = parse_and_save_one_student(save_reg_no)
#             messages.success(
#                 request,
#                 f"Marks saved successfully for {save_reg_no}. Rows updated: {total_upserts}"
#             )
#         elif is_bulk:
#             for rno in bulk_students:
#                 total_upserts += parse_and_save_one_student(rno)
#             messages.success(request, f"Bulk marks saved successfully. Rows updated: {total_upserts}")
#         else:
#             total_upserts = parse_and_save_one_student(single_reg_no)
#             messages.success(request, f"Marks saved successfully. Rows updated: {total_upserts}")

#     except ValueError as e:
#         transaction.set_rollback(True)
#         messages.error(request, str(e))
#         return redirect(request.META.get("HTTP_REFERER", "/"))

#     except Exception as e:
#         transaction.set_rollback(True)
#         messages.error(request, f"Unexpected error while saving marks: {str(e)}")
#         return redirect(request.META.get("HTTP_REFERER", "/"))

#     return redirect(request.META.get("HTTP_REFERER", "/"))

from django.contrib import messages
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from datetime import date


from datetime import date
from django.contrib import messages
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect

@transaction.atomic
def save_student_marks(request):
    if request.method != "POST":
        messages.error(request, "Invalid request method!")
        return redirect("/")

    def get_academic_year():
        """
        Dynamically returns academic year string.
        Example:
          If current month >= June -> '2025-2026'
          Else (Jan-May) -> '2024-2025'
        """
        today = date.today()
        current_year = today.year
        if today.month >= 6:
            return f"{current_year}-{current_year + 1}"
        else:
            return f"{current_year - 1}-{current_year}"

    department_code = request.POST.get("department_code")
    department_name = request.POST.get("department_name")
    course_code = request.POST.get("course_code")
    course_title = request.POST.get("course_title")
    batch = request.POST.get("batch")
    section = request.POST.get("section")

    exam_name = (
        request.POST.get("iat")
        or request.POST.get("exam_name")
        or request.GET.get("iat")
    )
    academic_year = get_academic_year()

    pattern_id_str = request.POST.get("pattern_id") or request.POST.get("exam_pattern_id")
    pattern_id = int(pattern_id_str) if (pattern_id_str and pattern_id_str.isdigit()) else None

    exam_pattern = None
    if pattern_id:
        exam_pattern = get_object_or_404(ExamPattern, id=pattern_id)

    co_map = {}
    blooms_map = {}

    for key in request.POST.keys():
        if key.startswith("co_map_"):
            suffix = key.replace("co_map_", "").replace("[]", "")
            values = request.POST.getlist(key)
            co_map[suffix] = [int(v) for v in values if str(v).isdigit()]

        if key.startswith("blooms_map_"):
            suffix = key.replace("blooms_map_", "").replace("[]", "")
            values = request.POST.getlist(key)
            blooms_map[suffix] = [int(v) for v in values if str(v).isdigit()]

    global_subpart_max = {}

    for key, value in request.POST.items():
        if key.startswith("subpart_max__"):
            parts = key.split("__")
            if len(parts) == 5:
                _, part_name, question_number, option_letter, sub_question = parts
                normalized_key = f"max_{part_name}_{question_number}_{sub_question}_{option_letter}"
                global_subpart_max[normalized_key] = value

    bulk_students = request.POST.getlist("students[]")
    single_reg_no = request.POST.get("reg_no")
    save_reg_no = request.POST.get("save_reg_no")

    absentee_students = set(request.POST.getlist("absentee_students[]"))

    if save_reg_no:
        bulk_students = [save_reg_no]
        is_bulk = False
        single_reg_no = save_reg_no
    else:
        is_bulk = bool(bulk_students)

    if not is_bulk and not single_reg_no:
        messages.error(request, "No student reg_no provided.")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    def safe_int(value, default=None):
        try:
            value = str(value).strip()
            if value == "":
                return default
            return int(float(value))
        except Exception:
            return default

    def parse_marks_key(key: str):
        parts = key.split("_")
        if len(parts) < 3 or parts[0] != "marks":
            return None

        return {
            "part_name": parts[1],
            "question_number": parts[2],
            "sub_question": parts[3] if len(parts) > 3 else None,
            "option_letter": parts[4] if len(parts) > 4 else None,
        }

    def build_max_key(part_name, question_number, sub_question=None, option_letter=None):
        if sub_question and option_letter:
            return f"max_{part_name}_{question_number}_{sub_question}_{option_letter}"
        if sub_question:
            return f"max_{part_name}_{question_number}_{sub_question}"
        if option_letter:
            return f"max_{part_name}_{question_number}_{option_letter}"
        return f"max_{part_name}_{question_number}"

    def build_suffix(part_name, question_number, sub_question=None, option_letter=None):
        if sub_question and option_letter:
            return f"{part_name}_{question_number}_{sub_question}_{option_letter}"
        if sub_question:
            return f"{part_name}_{question_number}_{sub_question}"
        if option_letter:
            return f"{part_name}_{question_number}_{option_letter}"
        return f"{part_name}_{question_number}"

    def get_part_default_max(part_name):
        if not exam_pattern:
            return 0
        part_obj = exam_pattern.parts.filter(name=part_name).first()
        if not part_obj:
            return 0
        return safe_int(getattr(part_obj, "max_marks", 0), 0)

    def resolve_max_marks(part_name, question_number, sub_question=None, option_letter=None, max_data=None):
        max_data = max_data or {}

        if sub_question and option_letter:
            k = build_max_key(part_name, question_number, sub_question, option_letter)
            val = safe_int(global_subpart_max.get(k), 0)
            if val > 0:
                return val

        exact_key = build_max_key(part_name, question_number, sub_question, option_letter)
        val = safe_int(max_data.get(exact_key), 0)
        if val > 0:
            return val

        if option_letter:
            option_key = build_max_key(part_name, question_number, option_letter=option_letter)
            val = safe_int(max_data.get(option_key), 0)
            if val > 0:
                return val

        if sub_question:
            sub_key = build_max_key(part_name, question_number, sub_question=sub_question)
            val = safe_int(max_data.get(sub_key), 0)
            if val > 0:
                return val

        q_key = build_max_key(part_name, question_number)
        val = safe_int(max_data.get(q_key), 0)
        if val > 0:
            return val

        return get_part_default_max(part_name)

    def parse_and_save_one_student(reg_no: str):
        if not reg_no:
            return 0

        is_absent = 1 if reg_no in absentee_students else 0

        student = StudentDetails.objects.filter(reg_no=reg_no).first()
        if not student:
            raise ValueError(f"Student not found: {reg_no}")

        student_semester = getattr(student, "semester", None)
        student_department = getattr(student, "department", None)
        student_degree = getattr(student_department, "degree", None) if student_department else None

        enrollment_qs = CourseEnrollment.objects.filter(
            student_id=student.id,
            course__course_code=course_code,
        )

        if batch:
            enrollment_qs = enrollment_qs.filter(batch=batch)

        if section:
            enrollment_qs = enrollment_qs.filter(section=section)

        enrollment = enrollment_qs.order_by("-enroll", "-id").first()
        if not enrollment:
            raise ValueError(f"Course enrollment not found for {reg_no}")

        faculty_assignment = AssignSubjectFaculty.objects.filter(
            course=enrollment.course,
            batch=batch or enrollment.batch,
            section=section or enrollment.section,
            academic_year=academic_year,
            is_active=True
        ).first()

        bulk_prefix = f"m__{reg_no}__"
        is_bulk_row = any(k.startswith(bulk_prefix + "marks_") for k in request.POST.keys())

        if is_bulk_row:
            marks_data = {
                k.replace(bulk_prefix, ""): v
                for k, v in request.POST.items()
                if k.startswith(bulk_prefix + "marks_")
            }
            max_data = {
                k.replace(bulk_prefix, ""): v
                for k, v in request.POST.items()
                if k.startswith(bulk_prefix + "max_")
            }
        else:
            marks_data = {k: v for k, v in request.POST.items() if k.startswith("marks_")}
            max_data = {k: v for k, v in request.POST.items() if k.startswith("max_")}

        grouped = {}
        for key, value in marks_data.items():
            parsed = parse_marks_key(key)
            if not parsed:
                continue

            group_key = (parsed["part_name"], parsed["question_number"])
            grouped.setdefault(group_key, []).append(
                {
                    "raw_key": key,
                    "raw_value": value,
                    "part_name": parsed["part_name"],
                    "question_number": parsed["question_number"],
                    "sub_question": parsed["sub_question"],
                    "option_letter": parsed["option_letter"],
                }
            )

        if not grouped:
            StudentInternalMark.objects.update_or_create(
                student=student,
                enrollment=enrollment,
                exam_name=exam_name,
                course=enrollment.course,
                part_name=None,
                question_number=None,
                sub_question=None,
                option_letter=None,
                defaults={
                    "pattern": exam_pattern,
                    "semester": student_semester,
                    "degree": student_degree,
                    "department": student_department,
                    "max_marks": None,
                    "marks_obtained": None,
                    "co_code_id": None,
                    "level_code_id": None,
                    "reg_no": reg_no,
                    "course_code": course_code,
                    "absentee": is_absent,
                    "batch": batch or enrollment.batch,
                    "section": section or enrollment.section,
                    "academic_year": academic_year,
                    "faculty_assignment": faculty_assignment,
                },
            )
            return 1

        submitted_subparts = {}
        for (part_name, question_number), entries in grouped.items():
            for entry in entries:
                option_letter = entry["option_letter"]
                sub_question = entry["sub_question"]

                if option_letter and sub_question:
                    k = (part_name, question_number, option_letter)
                    submitted_subparts.setdefault(k, set()).add(sub_question)

        for (part_name, question_number, option_letter), submitted_set in submitted_subparts.items():
            StudentInternalMark.objects.filter(
                student=student,
                enrollment=enrollment,
                exam_name=exam_name,
                course=enrollment.course,
                part_name=part_name,
                question_number=question_number,
                option_letter=option_letter,
            ).exclude(sub_question__in=list(submitted_set)).exclude(sub_question__in=["i", "ii"]).delete()

        upsert_count = 0

        for (part_name, question_number), entries in grouped.items():
            selected_option = None
            pair_totals = {}

            for entry in entries:
                sub_question = entry["sub_question"]
                option_letter = entry["option_letter"]
                marks_obt = safe_int(entry["raw_value"], None)

                if option_letter and marks_obt is not None and marks_obt > 0:
                    selected_option = option_letter

                if sub_question and option_letter and marks_obt is not None:
                    pair_totals.setdefault(option_letter, 0)
                    pair_totals[option_letter] += marks_obt

            non_zero_options = set()
            for entry in entries:
                option_letter = entry["option_letter"]
                marks_obt = safe_int(entry["raw_value"], None)
                if option_letter and marks_obt is not None and marks_obt > 0:
                    non_zero_options.add(option_letter)

            if len(non_zero_options) > 1:
                raise ValueError(
                    f"{reg_no} - {part_name}{question_number}: Enter marks for only one option."
                )

            for option_letter, total in pair_totals.items():
                allowed_max = 0

                for entry in entries:
                    if entry["option_letter"] != option_letter:
                        continue

                    sub_question = entry["sub_question"]
                    if not sub_question:
                        continue

                    suffix = build_suffix(
                        part_name=part_name,
                        question_number=question_number,
                        sub_question=sub_question,
                        option_letter=option_letter,
                    )
                    co_values = co_map.get(suffix, [])
                    co_id = co_values[0] if co_values else None
                    if co_id is None:
                        continue

                    resolved = resolve_max_marks(
                        part_name=part_name,
                        question_number=question_number,
                        sub_question=sub_question,
                        option_letter=option_letter,
                        max_data=max_data,
                    )
                    allowed_max += resolved

                if allowed_max > 0 and total > allowed_max:
                    raise ValueError(
                        f"{reg_no} - {part_name}{question_number}{option_letter}: "
                        f"Total of all sub-questions cannot exceed {allowed_max}. "
                        f"You entered total {total}."
                    )

            for entry in entries:
                sub_question = entry["sub_question"]
                option_letter = entry["option_letter"]
                marks_obt = safe_int(entry["raw_value"], None)

                if selected_option and option_letter and option_letter != selected_option:
                    marks_obt = None

                suffix = build_suffix(
                    part_name=part_name,
                    question_number=question_number,
                    sub_question=sub_question,
                    option_letter=option_letter,
                )

                co_values = co_map.get(suffix, [])
                bl_values = blooms_map.get(suffix, [])

                co_id = co_values[0] if co_values else None
                bl_id = bl_values[0] if bl_values else None

                max_marks = resolve_max_marks(
                    part_name=part_name,
                    question_number=question_number,
                    sub_question=sub_question,
                    option_letter=option_letter,
                    max_data=max_data,
                )

                if is_absent == 1:
                    co_id = None
                    bl_id = None
                    max_marks_to_save = None
                    marks_obt_to_save = None
                elif co_id is None:
                    bl_id = None
                    max_marks_to_save = None
                    marks_obt_to_save = None
                else:
                    max_marks_to_save = max_marks if max_marks > 0 else None
                    marks_obt_to_save = marks_obt if marks_obt is not None else None

                if (
                    marks_obt_to_save is not None
                    and max_marks_to_save is not None
                    and marks_obt_to_save > max_marks_to_save
                ):
                    raise ValueError(
                        f"{reg_no} - {part_name}{question_number}"
                        f"{option_letter or ''}{('-' + sub_question.upper()) if sub_question else ''}: "
                        f"Marks cannot exceed {max_marks_to_save}."
                    )

                StudentInternalMark.objects.update_or_create(
                    student=student,
                    enrollment=enrollment,
                    exam_name=exam_name,
                    course=enrollment.course,
                    part_name=part_name,
                    question_number=question_number,
                    sub_question=sub_question,
                    option_letter=option_letter,
                    defaults={
                        "pattern": exam_pattern,
                        "semester": student_semester,
                        "degree": student_degree,
                        "department": student_department,
                        "max_marks": max_marks_to_save,
                        "marks_obtained": marks_obt_to_save,
                        "co_code_id": co_id,
                        "level_code_id": bl_id,
                        "reg_no": reg_no,
                        "course_code": course_code,
                        "absentee": is_absent,
                        "batch": batch or enrollment.batch,
                        "section": section or enrollment.section,
                        "academic_year": academic_year,
                        "faculty_assignment": faculty_assignment,
                    },
                )
                upsert_count += 1

        return upsert_count

    try:
        total_upserts = 0

        if save_reg_no:
            total_upserts = parse_and_save_one_student(save_reg_no)
            messages.success(
                request,
                f"Marks saved successfully for {save_reg_no}. Rows updated: {total_upserts}"
            )
        elif is_bulk:
            for rno in bulk_students:
                total_upserts += parse_and_save_one_student(rno)
            messages.success(
                request,
                f"Bulk marks saved successfully. Rows updated: {total_upserts}"
            )
        else:
            total_upserts = parse_and_save_one_student(single_reg_no)
            messages.success(
                request,
                f"Marks saved successfully. Rows updated: {total_upserts}"
            )

    except ValueError as e:
        transaction.set_rollback(True)
        messages.error(request, str(e))
        return redirect(request.META.get("HTTP_REFERER", "/"))

    except Exception as e:
        transaction.set_rollback(True)
        messages.error(request, f"Unexpected error while saving marks: {str(e)}")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    return redirect(request.META.get("HTTP_REFERER", "/"))






def exam_marks_subjects(request):
    faculty_id = request.user.Employee_id
    faculty = general_information.objects.filter(faculty_id=faculty_id).first()
    assigned_subjects = AssignSubjectFaculty.objects.filter(
        faculty=faculty
    ).select_related("course")
    return render(request, "faculty_management/exam_marks_subjects.html")


def exam_marks_overview(request):
    return render(request, "faculty_management/exam_marks_overview.html")


import json
from collections import defaultdict

from django.db.models import Q
from django.shortcuts import render

import json
from collections import defaultdict

from django.db.models import Q, Sum
from django.shortcuts import render

# from yourapp.decorators import check_permission, student_management
# from yourapp.models import (
#     general_information,
#     AssignSubjectFaculty,
#     Assessments,
#     Assessment_master,
#     AssessmentMark,
#     experiment_marks,   # <-- PRACTICAL MODEL
# )

import json
from collections import defaultdict

from django.db.models import Q, Sum
from django.shortcuts import render
from django.contrib import messages

# Import your decorators and models as needed:
# from .decorators import check_permission, student_management
# from .models import (general_information, AssignSubjectFaculty, Degree, Add_Department,
#                      Assessments, ModelLab, InternalAssessment, Assessment_master,
# # views.py (overall_consolidate)
import json
from collections import defaultdict
from django.db.models import Q, Sum
from django.shortcuts import render

# import your decorators and models here
# from .decorators import check_permission, student_management
# from .models import (general_information, AssignSubjectFaculty, Assessments, ModelLab,
#                      InternalAssessment, Assessment_master, AssessmentMark, experiment_marks, Course)
from course_management.models import CourseHours

# @check_permission("overall_consolidate")
# def overall_consolidate(request):
#     # --- Resolve faculty & faculty_id string ---
#     user = getattr(request, "user", None)
#     emp_id = getattr(user, "Employee_id", None) or getattr(user, "employee_id", None)
#     current_faculty_id = ""
#     if emp_id:
#         gi = general_information.objects.filter(faculty_id=emp_id).only("faculty_id").first()
#         current_faculty_id = str(gi.faculty_id) if gi and gi.faculty_id else str(emp_id)

#     # --- Assigned subjects for cards ---
#     assigned_subjects = (
#         AssignSubjectFaculty.objects
#         .filter(faculty__faculty_id=emp_id)
#         .select_related("course", "department__degree")
#     )

#     degree_ids, dept_ids, course_ids = set(), set(), set()
#     for subject in assigned_subjects:
#         dept = getattr(subject, "department", None)
#         if dept:
#             subject.department_name = getattr(dept, "Department", "") or ''
#             subject.department_code = getattr(dept, "Department_code", "") or ''
#             dept_ids.add(dept.id)
#             deg = getattr(dept, "degree", None)
#             if deg:
#                 subject.degree_id = deg.id
#                 subject.degree_code = getattr(deg, "degree_code", "") or ""
#                 subject.degree_name = getattr(deg, "degree", "") or ""
#                 degree_ids.add(deg.id)
#             else:
#                 subject.degree_id = None
#                 subject.degree_code = ""
#                 subject.degree_name = ""
#         else:
#             subject.department_name = ''
#             subject.department_code = ''
#             subject.degree_id = None
#             subject.degree_code = ''
#             subject.degree_name = ''

#         crs = getattr(subject, "course", None)

#         course_hours = None
#         lecture_hours = 0
#         lab_hours = 0

#         if crs:
#             course_ids.add(crs.id)

#             # ✅ bring hour_config also
#             course_hours = (
#                 CourseHours.objects
#                 .filter(course=crs)
#                 .select_related("hour_config")
#                 .order_by("-id")
#                 .first()
#             )

#         if course_hours and course_hours.hour_config:
#             try:
#                 lecture_hours = int(course_hours.hour_config.lecture_hours or 0) + int(course_hours.hour_config.tutorial_hours or 0)
#             except (ValueError, TypeError):
#                 lecture_hours = 0

#             try:
#                 lab_hours = int(course_hours.hour_config.laboratory_hours or 0)
#             except (ValueError, TypeError):
#                 lab_hours = 0

#         subject.has_theory = lecture_hours > 0
#         subject.has_lab = lab_hours > 0

            
    

#     # --- Build degree -> {assessments, model_labs, model_labs_iat, iats} for the client ---
#     rows_assess = (
#         Assessments.objects
#         .filter(Q(degree_id__in=list(degree_ids)) | Q(degree__isnull=True))
#         .values("id", "assessment_name", "question_paper_required", "degree_id")
#     )

#     rows_ml = (
#         ModelLab.objects
#         .filter(Q(degree_id__in=list(degree_ids)) | Q(degree__isnull=True))
#         .select_related("internal_assessment")
#         .values("id", "model_lab_name", "degree_id", "internal_assessment__iat")
#     )

#     rows_iat = (
#         InternalAssessment.objects
#         .filter(Q(degree_id__in=list(degree_ids)) | Q(degree__isnull=True))
#         .values("degree_id", "id", "iat")
#     )

#     grouped = defaultdict(lambda: {"assessments": [], "model_labs": [], "model_labs_iat": {}, "iats": []})

#     for r in rows_assess:
#         key = r["degree_id"] if r["degree_id"] is not None else "global"
#         name = (r["assessment_name"] or "").strip()
#         qp = bool(r["question_paper_required"])
#         if not name:
#             continue
#         exists = next((a for a in grouped[key]["assessments"] if a["name"] == name), None)
#         if exists:
#             exists["qp"] = exists["qp"] or qp
#         else:
#             grouped[key]["assessments"].append({"id": r["id"], "name": name, "qp": qp})

#     for r in rows_ml:
#         key = r["degree_id"] if r["degree_id"] is not None else "global"
#         lab = (r["model_lab_name"] or "").strip()
#         iat = (r["internal_assessment__iat"] or "").strip()
#         if not lab:
#             continue
#         if iat:
#             grouped[key]["model_labs_iat"].setdefault(iat, []).append(lab)
#         else:
#             grouped[key]["model_labs"].append(lab)

#     for r in rows_iat:
#         key = r["degree_id"] if r["degree_id"] is not None else "global"
#         iat = (r["iat"] or "").strip()
#         if iat:
#             grouped[key]["iats"].append(iat)

#     for key in list(grouped.keys()):
#         grouped[key]["model_labs"] = sorted(list(dict.fromkeys(grouped[key]["model_labs"])), key=lambda x: x.lower())
#         for iat_key, labs in list(grouped[key]["model_labs_iat"].items()):
#             grouped[key]["model_labs_iat"][iat_key] = sorted(list(dict.fromkeys(labs)), key=lambda x: x.lower())
#         grouped[key]["iats"] = sorted(list(dict.fromkeys(grouped[key]["iats"])), key=lambda x: x.lower())

#     assessments_by_degree_json = json.dumps(grouped)

#     # --- Assessment_master payload for UI (INCLUDES IAT + assessment info) ---
#     am_qs = (
#         Assessment_master.objects
#         .filter(
#             faculty_id=current_faculty_id,
#             degree_id__in=list(degree_ids) if degree_ids else [],
#             department_id__in=list(dept_ids) if dept_ids else [],
#             course_id__in=list(course_ids) if course_ids else [],
#         )
#         .select_related("assessment", "internal_assessment")
#         .values(
#             "id",
#             "degree_id", "department_id", "course_id", "faculty_id",
#             "assessment_id",
#             "Assessmentname", "customAssessmentname", "Maxmarks",
#             "internal_assessment_id",
#             "internal_assessment__iat",
#             "assessment__assessment_name",
#             "assessment__question_paper_required",
#         )
#         .order_by("id")
#     )

#     # Debug header: show GET-selected IAT for easier reading in logs
#     selected_iat_from_get = request.GET.get("iat", "")
#     # print("\n========== DEBUG: Assessment Matching Trace ==========")
#     # print(f"Selected IAT (from GET): {selected_iat_from_get}\n")

#     am_payload = []
#     for r in am_qs:
#         assessment_name = (r.get("assessment__assessment_name") or "").strip()
#         cached_name = (r.get("Assessmentname") or "").strip()
#         custom_name = (r.get("customAssessmentname") or "").strip()
#         selected_iat = (r.get("internal_assessment__iat") or "").strip()

#         # Determine which column we're effectively using as display_name
#         if r.get("assessment_id"):
#             display_name = assessment_name or cached_name or custom_name or ""
#             source_col = "assessment__assessment_name"
#         else:
#             display_name = custom_name or cached_name or assessment_name or ""
#             source_col = "customAssessmentname"

#         # PRINT DEBUG INFO FOR EACH ROW (IAT + which column produced the display name)
#         # print(f"IAT: {selected_iat or 'N/A'} | Assessment_master ID: {r['id']}")
#         # print(f"   Source Column: {source_col}")
#         # print(f"   Assessment ID (FK): {r.get('assessment_id') or 'None (Custom)'}")
#         # print(f"   Display Name: {display_name}")
#         # print("------------------------------------------------------")

#         am_payload.append({
#             "id": r["id"],
#             "degree_id": r["degree_id"],
#             "department_id": r["department_id"],
#             "course_id": r["course_id"],
#             "faculty_id": r["faculty_id"],
#             "assessment_id": r["assessment_id"],
#             "Assessmentname": cached_name,
#             "customAssessmentname": custom_name,
#             "display_name": display_name,
#             "Maxmarks": r["Maxmarks"],
#             "iat_id": r["internal_assessment_id"],
#             "iat": selected_iat,
#             "assessment_name": assessment_name,
#             "qp": bool(r["assessment__question_paper_required"]) if r["assessment__question_paper_required"] is not None else False,
#         })

#     # print("======================================================\n")

#     # -- Read GET selection ---
#     raw_assessment_ids = request.GET.getlist("assessment_ids")
#     std_assessment_ids_param = request.GET.getlist("standard_assessment_ids")
#     qp_assessment_ids_param = request.GET.getlist("qp_assessment_ids")
#     custom_assessment_names = request.GET.getlist("custom_assessment_names")
#     custom_row_ids_param = request.GET.getlist("custom_assessment_row_ids")
#     model_labs = request.GET.getlist("model_labs")
#     include_practical = request.GET.get("include_practical") == "1"
#     selected_iat = request.GET.get("iat", "")
#     assessment_iats = request.GET.getlist("assessment_iats")  # New parameter for individual assessment IATs

#     # --- Handle BOTH IATs mode ---
#     selected_am_rows = []
#     if selected_iat == "BOTH":
#         # Filter for IAT1 and IAT2 only
#         selected_am_rows = [
#             row for row in am_payload 
#             if row.get("iat", "").upper() in ["IAT1", "IAT2"]
#         ]
        
#         # Filter based on selected assessment IDs
#         if raw_assessment_ids:
#             selected_am_rows = [row for row in selected_am_rows if str(row.get("id")) in raw_assessment_ids]
        
#     else:
       
#         if raw_assessment_ids:
#             selected_am_rows = [row for row in am_payload if str(row.get("id")) in raw_assessment_ids]
#             if selected_iat:
#                 # Further filter by selected IAT if specified
#                 selected_am_rows = [row for row in selected_am_rows if row.get("iat", "").lower() == selected_iat.lower()]
    
#     # --- Process selected assessments ---
#     standard_assessment_ids = []
#     qp_assessment_ids = []
#     custom_assessment_data = []
    
#     for row in selected_am_rows:
#         if row.get("assessment_id"):
#             std_id = str(row["assessment_id"])
#             standard_assessment_ids.append(std_id)
#             if row.get("qp"):
#                 qp_assessment_ids.append(std_id)
#         else:
#             # Custom assessment (no FK to Assessments table)
#             custom_assessment_data.append({
#                 "row_id": row["id"],
#                 "name": row.get("display_name", ""),
#                 "max_marks": row.get("Maxmarks", 0),
#                 "iat": row.get("iat", "")
#             })

#     # --- Get student marks data ---
#     selected_course_id = request.GET.get("course_id", "")
#     selected_batch = request.GET.get("batch", "")
#     selected_section = request.GET.get("section", "")
    
#     student_marks = []
#     if selected_course_id and standard_assessment_ids:
#         marks_qs = AssessmentMark.objects.filter(
#             assessment_id__in=standard_assessment_ids,
#             course_id=selected_course_id
#         ).select_related("student")
        
#         if selected_batch:
#             marks_qs = marks_qs.filter(student__batch=selected_batch)
#         if selected_section:
#             marks_qs = marks_qs.filter(student__section=selected_section)
        
#         student_marks = list(marks_qs)
    
#     # --- Get custom assessment marks ---
#     custom_marks = []
#     if custom_assessment_data:
#         custom_row_ids = [item["row_id"] for item in custom_assessment_data]
#         custom_marks_qs = AssessmentMark.objects.filter(
#             assessment_master_id__in=custom_row_ids,
#             course_id=selected_course_id
#         ).select_related("student")
        
#         if selected_batch:
#             custom_marks_qs = custom_marks_qs.filter(student__batch=selected_batch)
#         if selected_section:
#             custom_marks_qs = custom_marks_qs.filter(student__section=selected_section)
        
#         custom_marks = list(custom_marks_qs)
    
#     # --- Calculate practical marks if included ---
#     practical_data = []
#     if include_practical and selected_course_id:
#         practical_qs = AssessmentMark.objects.filter(
#             course_id=selected_course_id,
#             assessment__assessment_name__icontains="practical"
#         ).select_related("student", "assessment")
        
#         if selected_batch:
#             practical_qs = practical_qs.filter(student__batch=selected_batch)
#         if selected_section:
#             practical_qs = practical_qs.filter(student__section=selected_section)
        
#         practical_data = list(practical_qs)
    
#     # --- Prepare context for template ---
#     selection = {
#         "course_id": request.GET.get("course_id", ""),
#         "course_code": request.GET.get("course_code", ""),
#         "course_title": request.GET.get("course_title", ""),
#         "department": request.GET.get("department", ""),
#         "department_id": request.GET.get("department_id", ""),
#         "batch": request.GET.get("batch", ""),
#         "section": request.GET.get("section", ""),
#         "degree_id": request.GET.get("degree_id", ""),
#         "degree_code": request.GET.get("degree_code", ""),
#         "degree_name": request.GET.get("degree_name", ""),
#         "assessment_ids": raw_assessment_ids,
#         "standard_assessment_ids": standard_assessment_ids,
#         "qp_assessment_ids": qp_assessment_ids,
#         "custom_assessment_names": custom_assessment_names,
#         "custom_assessment_row_ids": custom_row_ids_param,
#         "model_labs": model_labs,
#         "include_practical": include_practical,
#         "iat": selected_iat,
#         "assessment_iats": assessment_iats,
#     }
    
#     context_data = {
#         "student_marks": student_marks,
#         "custom_marks": custom_marks,
#         "practical_data": practical_data,
#         "selected_assessments": selected_am_rows,
#         "custom_assessments": custom_assessment_data,
#         "model_labs_selected": model_labs,
#         "include_practical": include_practical,
#         "selection": selection,
#     }

#     am_master_json = json.dumps(am_payload)

#     context = {
#         "assigned_subjects": assigned_subjects,
#         "assessments_by_degree_json": assessments_by_degree_json,
#         "am_master_json": am_master_json,
#         "current_faculty_id": current_faculty_id,
#         "selection": selection,
#         "context_data": context_data,
#     }
#     return render(request, "faculty_management/overall_consolidate.html", context)


from collections import defaultdict
from django.db.models import Q
from django.shortcuts import render

from course_management.models import CourseHours

@check_permission("overall_consolidate")
def overall_consolidate(request):
    user = getattr(request, "user", None)
    emp_id = getattr(user, "Employee_id", None) or getattr(user, "employee_id", None)

    current_faculty_id = ""
    if emp_id:
        gi = general_information.objects.filter(faculty_id=emp_id).only("faculty_id").first()
        current_faculty_id = str(gi.faculty_id) if gi and gi.faculty_id else str(emp_id)

    assigned_subjects = (
        AssignSubjectFaculty.objects
        .filter(faculty__faculty_id=emp_id)
        .select_related("course", "department__degree")
    )

    degree_ids, dept_ids, course_ids = set(), set(), set()

    for subject in assigned_subjects:
        dept = getattr(subject, "department", None)
        if dept:
            subject.department_name = getattr(dept, "Department", "") or ''
            subject.department_code = getattr(dept, "Department_code", "") or ''
            dept_ids.add(dept.id)

            deg = getattr(dept, "degree", None)
            if deg:
                subject.degree_id = deg.id
                subject.degree_code = getattr(deg, "degree_code", "") or ""
                subject.degree_name = getattr(deg, "degree", "") or ""
                degree_ids.add(deg.id)
            else:
                subject.degree_id = None
                subject.degree_code = ""
                subject.degree_name = ""
        else:
            subject.department_name = ''
            subject.department_code = ''
            subject.degree_id = None
            subject.degree_code = ''
            subject.degree_name = ''

        crs = getattr(subject, "course", None)

        course_hours = None
        lecture_hours = 0
        lab_hours = 0

        if crs:
            course_ids.add(crs.id)
            course_hours = (
                CourseHours.objects
                .filter(course=crs)
                .select_related("hour_config")
                .order_by("-id")
                .first()
            )

        if course_hours and course_hours.hour_config:
            try:
                lecture_hours = int(course_hours.hour_config.lecture_hours or 0) + int(course_hours.hour_config.tutorial_hours or 0)
            except (ValueError, TypeError):
                lecture_hours = 0

            try:
                lab_hours = int(course_hours.hour_config.laboratory_hours or 0)
            except (ValueError, TypeError):
                lab_hours = 0

        subject.has_theory = lecture_hours > 0
        subject.has_lab = lab_hours > 0

    rows_assess = (
        Assessments.objects
        .filter(Q(degree_id__in=list(degree_ids)) | Q(degree__isnull=True))
        .values("id", "assessment_name", "question_paper_required", "degree_id")
    )

    rows_ml = (
        ModelLab.objects
        .filter(Q(degree_id__in=list(degree_ids)) | Q(degree__isnull=True))
        .select_related("internal_assessment")
        .values("id", "model_lab_name", "degree_id", "internal_assessment__iat")
    )

    rows_iat = (
        InternalAssessment.objects
        .filter(Q(degree_id__in=list(degree_ids)) | Q(degree__isnull=True))
        .values("degree_id", "id", "iat")
    )

    grouped = defaultdict(lambda: {
        "assessments": [],
        "model_labs": [],
        "model_labs_iat": {},
        "iats": []
    })

    for r in rows_assess:
        key = r["degree_id"] if r["degree_id"] is not None else "global"
        name = (r["assessment_name"] or "").strip()
        qp = bool(r["question_paper_required"])
        if not name:
            continue

        exists = next((a for a in grouped[key]["assessments"] if a["name"] == name), None)
        if exists:
            exists["qp"] = exists["qp"] or qp
        else:
            grouped[key]["assessments"].append({
                "id": r["id"],
                "name": name,
                "qp": qp
            })

    for r in rows_ml:
        key = r["degree_id"] if r["degree_id"] is not None else "global"
        lab = (r["model_lab_name"] or "").strip()
        iat = (r["internal_assessment__iat"] or "").strip()

        if not lab:
            continue

        if iat:
            grouped[key]["model_labs_iat"].setdefault(iat, []).append(lab)
        else:
            grouped[key]["model_labs"].append(lab)

    for r in rows_iat:
        key = r["degree_id"] if r["degree_id"] is not None else "global"
        iat = (r["iat"] or "").strip()
        if iat:
            grouped[key]["iats"].append(iat)

    for key in list(grouped.keys()):
        grouped[key]["model_labs"] = sorted(
            list(dict.fromkeys(grouped[key]["model_labs"])),
            key=lambda x: x.lower()
        )
        for iat_key, labs in list(grouped[key]["model_labs_iat"].items()):
            grouped[key]["model_labs_iat"][iat_key] = sorted(
                list(dict.fromkeys(labs)),
                key=lambda x: x.lower()
            )
        grouped[key]["iats"] = sorted(
            list(dict.fromkeys(grouped[key]["iats"])),
            key=lambda x: x.lower()
        )

    assessments_by_degree_json = json.dumps(grouped)

    am_qs = (
        Assessment_master.objects
        .filter(
            faculty_id=current_faculty_id,
            degree_id__in=list(degree_ids) if degree_ids else [],
            department_id__in=list(dept_ids) if dept_ids else [],
            course_id__in=list(course_ids) if course_ids else [],
        )
        .select_related("assessment", "internal_assessment")
        .values(
            "id",
            "degree_id", "department_id", "course_id", "faculty_id",
            "assessment_id",
            "Assessmentname", "customAssessmentname", "Maxmarks",
            "internal_assessment_id",
            "internal_assessment__iat",
            "assessment__assessment_name",
            "assessment__question_paper_required",
        )
        .order_by("id")
    )

    am_payload = []
    for r in am_qs:
        assessment_name = (r.get("assessment__assessment_name") or "").strip()
        cached_name = (r.get("Assessmentname") or "").strip()
        custom_name = (r.get("customAssessmentname") or "").strip()
        selected_iat = (r.get("internal_assessment__iat") or "").strip()

        if r.get("assessment_id"):
            display_name = assessment_name or cached_name or custom_name or ""
            source_col = "assessment__assessment_name"
        else:
            display_name = custom_name or cached_name or assessment_name or ""
            source_col = "customAssessmentname"

        am_payload.append({
            "id": r["id"],
            "degree_id": r["degree_id"],
            "department_id": r["department_id"],
            "course_id": r["course_id"],
            "faculty_id": r["faculty_id"],
            "assessment_id": r["assessment_id"],
            "Assessmentname": cached_name,
            "customAssessmentname": custom_name,
            "display_name": display_name,
            "display_source": source_col,
            "Maxmarks": r["Maxmarks"],
            "iat_id": r["internal_assessment_id"],
            "iat": selected_iat,
            "assessment_name": assessment_name,
            "qp": bool(r["assessment__question_paper_required"]) if r["assessment__question_paper_required"] is not None else False,
            "is_custom_display": not bool(r.get("assessment_id")),
        })

    raw_assessment_ids = request.GET.getlist("assessment_ids")
    std_assessment_ids_param = request.GET.getlist("standard_assessment_ids")
    qp_assessment_ids_param = request.GET.getlist("qp_assessment_ids")
    custom_assessment_names = request.GET.getlist("custom_assessment_names")
    custom_row_ids_param = request.GET.getlist("custom_assessment_row_ids")
    model_labs = request.GET.getlist("model_labs")
    include_practical = request.GET.get("include_practical") == "1"
    selected_iat = request.GET.get("iat", "")
    assessment_iats = request.GET.getlist("assessment_iats")
    selected_iat_groups = request.GET.getlist("selected_iat_groups")

    def _norm_iat(v):
        return (v or "").strip().lower().replace(" ", "")

    selected_am_rows = []

    if raw_assessment_ids:
        selected_am_rows = [
            row for row in am_payload
            if str(row.get("id")) in raw_assessment_ids
            or f"am-{row.get('id')}" in raw_assessment_ids
            or str(row.get("assessment_id")) in raw_assessment_ids
        ]

        if selected_iat and selected_iat.upper() != "ALL":
            selected_am_rows = [
                row for row in selected_am_rows
                if _norm_iat(row.get("iat")) == _norm_iat(selected_iat)
            ]

    if selected_iat_groups:
        iat_group_rows = []

        for row in am_payload:
            row_iat = row.get("iat", "")
            display_name = row.get("display_name", "")

            for grp in selected_iat_groups:
                if _norm_iat(row_iat) == _norm_iat(grp):
                    if _norm_iat(display_name) == _norm_iat(grp):
                        iat_group_rows.append(row)

        if not iat_group_rows:
            for row in am_payload:
                row_iat = row.get("iat", "")
                for grp in selected_iat_groups:
                    if _norm_iat(row_iat) == _norm_iat(grp):
                        iat_group_rows.append(row)

        selected_am_rows.extend(iat_group_rows)

    seen_ids = set()
    deduped_rows = []
    for row in selected_am_rows:
        rid = row.get("id")
        if rid in seen_ids:
            continue
        seen_ids.add(rid)
        deduped_rows.append(row)

    selected_am_rows = deduped_rows

    standard_assessment_ids = []
    qp_assessment_ids = []
    custom_assessment_data = []

    for row in selected_am_rows:
        if row.get("assessment_id"):
            std_id = str(row["assessment_id"])
            standard_assessment_ids.append(std_id)
            if row.get("qp"):
                qp_assessment_ids.append(std_id)
        else:
            custom_assessment_data.append({
                "row_id": row["id"],
                "name": row.get("display_name", ""),
                "max_marks": row.get("Maxmarks", 0),
                "iat": row.get("iat", "")
            })

    selected_course_id = request.GET.get("course_id", "")
    selected_batch = request.GET.get("batch", "")
    selected_section = request.GET.get("section", "")

    student_marks = []
    if selected_course_id and standard_assessment_ids:
        marks_qs = AssessmentMark.objects.filter(
            assessment_id__in=standard_assessment_ids,
            course_id=selected_course_id
        ).select_related("student")

        if selected_batch:
            marks_qs = marks_qs.filter(student__batch=selected_batch)
        if selected_section:
            marks_qs = marks_qs.filter(student__section=selected_section)

        student_marks = list(marks_qs)

    custom_marks = []
    if custom_assessment_data:
        custom_row_ids = [item["row_id"] for item in custom_assessment_data]
        custom_marks_qs = AssessmentMark.objects.filter(
            assessment_master_id__in=custom_row_ids,
            course_id=selected_course_id
        ).select_related("student")

        if selected_batch:
            custom_marks_qs = custom_marks_qs.filter(student__batch=selected_batch)
        if selected_section:
            custom_marks_qs = custom_marks_qs.filter(student__section=selected_section)

        custom_marks = list(custom_marks_qs)

    practical_data = []
    if include_practical and selected_course_id:
        practical_qs = AssessmentMark.objects.filter(
            course_id=selected_course_id,
            assessment__assessment_name__icontains="practical"
        ).select_related("student", "assessment")

        if selected_batch:
            practical_qs = practical_qs.filter(student__batch=selected_batch)
        if selected_section:
            practical_qs = practical_qs.filter(student__section=selected_section)

        practical_data = list(practical_qs)

    selection = {
        "course_id": request.GET.get("course_id", ""),
        "course_code": request.GET.get("course_code", ""),
        "course_title": request.GET.get("course_title", ""),
        "department": request.GET.get("department", ""),
        "department_id": request.GET.get("department_id", ""),
        "batch": request.GET.get("batch", ""),
        "section": request.GET.get("section", ""),
        "degree_id": request.GET.get("degree_id", ""),
        "degree_code": request.GET.get("degree_code", ""),
        "degree_name": request.GET.get("degree_name", ""),
        "assessment_ids": raw_assessment_ids,
        "standard_assessment_ids": standard_assessment_ids,
        "qp_assessment_ids": qp_assessment_ids,
        "custom_assessment_names": custom_assessment_names,
        "custom_assessment_row_ids": custom_row_ids_param,
        "model_labs": model_labs,
        "include_practical": include_practical,
        "iat": selected_iat,
        "assessment_iats": assessment_iats,
        "selected_iat_groups": selected_iat_groups,
    }

    context_data = {
        "student_marks": student_marks,
        "custom_marks": custom_marks,
        "practical_data": practical_data,
        "selected_assessments": selected_am_rows,
        "custom_assessments": custom_assessment_data,
        "model_labs_selected": model_labs,
        "include_practical": include_practical,
        "selection": selection,
    }

    am_master_json = json.dumps(am_payload)

    context = {
        "assigned_subjects": assigned_subjects,
        "assessments_by_degree_json": assessments_by_degree_json,
        "am_master_json": am_master_json,
        "current_faculty_id": current_faculty_id,
        "selection": selection,
        "context_data": context_data,
    }
    return render(request, "faculty_management/overall_consolidate.html", context)





import io
from collections import defaultdict

from django.http import HttpResponse
from django.db.models import Sum, Q

# ReportLab
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Table,
    TableStyle,
    Spacer,
    PageBreak,
)
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.styles import ParagraphStyle

#
# --- Standard library
import io
from collections import defaultdict

# --- Django
from django.http import HttpResponse
from django.db.models import Q, Sum

# --- ReportLab
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# --- Your decorators (unchanged)
# from .decorators import check_permission, student_management

# --- Your models (import according to your app structure)
# Adjust import paths to match your project layout.

# --- Standard library
import io
from collections import defaultdict, OrderedDict

# --- Django
from django.http import HttpResponse
from django.db.models import Q, Sum

# --- ReportLab
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# --- Your decorators
# from .decorators import check_permission, student_management

# --- Your models (adjust import paths to your project)

# --- Standard library
import io
from collections import defaultdict, OrderedDict

# --- Django
from django.http import HttpResponse
from django.db.models import Q, Sum

# --- ReportLab
# --- Stdlib
import io
from collections import defaultdict, OrderedDict
import os

from django.http import HttpResponse, Http404
from django.db.models import Sum
from django.contrib.staticfiles import finders
from django.conf import settings

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.utils import ImageReader

# from your_app.models import AssignSubjectFaculty, Assessment_master, AssessmentMark, experiment_marks, general_information, StudentDetails
# from your_app.decorators import check_permission, student_management

# views.py

import io
import os
from collections import defaultdict, OrderedDict
from django.conf import settings
from django.contrib.staticfiles import finders
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# @check_permission("overall_consolidate")
# @student_management
# @check_permission("overall_consolidate")
# @student_management
import io
import os
from collections import defaultdict, OrderedDict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, getcontext

from django.http import HttpResponse
from django.db.models import Sum
from django.contrib.staticfiles import finders
from django.conf import settings
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
import io, os, logging
from django.http import HttpResponse
from django.contrib.staticfiles import finders
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.utils import ImageReader
from decimal import Decimal, ROUND_HALF_UP
from django.db import transaction

# def overall_consolidate_pdf(request):

#     logger = logging.getLogger(__name__)
    

#     # ---------------- REQUEST PARAMS ----------------
#     selection = {
#         "course_id": request.GET.get("course_id", ""),
#         "course_code": request.GET.get("course_code", ""),
#         "course_title": request.GET.get("course_title", ""),
#         "department": request.GET.get("department", ""),
#         "department_id": request.GET.get("department_id", ""),
#         "batch": request.GET.get("batch", ""),
#         "section": request.GET.get("section", ""),
#         "degree_code": request.GET.get("degree_code", ""),
#         "degree_name": request.GET.get("degree_name", ""),
#         "iat": request.GET.get("iat", ""),
#     }

#     # print(" FULL GET:", dict(request.GET))

#     def _to_int(v, default=0):
#         try:
#             if v is None or str(v).strip() == "":
#                 return default
#             return int(float(str(v).strip()))
#         except Exception:
#             return default

#     hour_config_id = (
#         request.GET.get("hour_config_id")
#         or request.GET.get("coursehour_config_id")
#         or request.GET.get("hour_config")
#         or ""
#     )

    
#     ch_qs = CourseHours.objects.filter(course_id=selection["course_id"])
#     if hour_config_id:
#         ch_qs = ch_qs.filter(hour_config_id=hour_config_id)

#     course_hours = ch_qs.first()


#     if (not hour_config_id) and course_hours and course_hours.hour_config_id:
#         hour_config_id = str(course_hours.hour_config_id)

#     # fetch config (source of truth)
#     cfg = CourseHourConfig.objects.filter(id=hour_config_id).first() if hour_config_id else None

#     # print("\n================ COURSE HOUR CONFIG DEBUG ================")
#     # print(" course_id:", selection["course_id"])
#     # print(" final hour_config_id (GET or inferred):", hour_config_id or "NOT AVAILABLE")
#     # print(" CourseHours row found:", bool(course_hours))
#     # print(" CourseHourConfig found:", bool(cfg))

#     if cfg:
#         lecture_hours = _to_int(cfg.lecture_hours)
#         tutorial_hours = _to_int(cfg.tutorial_hours)
#         lab_hours = _to_int(cfg.laboratory_hours)

#         theory_pct = _to_int(cfg.theory_percentage)
#         activity_pct = _to_int(cfg.activity_percentage)
#         practical_pct = _to_int(cfg.practical_percentage)
        
#         if lecture_hours > 0 and tutorial_hours == 0 and lab_hours == 0:
#             theory_pct = min(100, theory_pct * 2)
#             theory_pct, practical_pct, activity_pct
#     else:
#         print()


#     assessment_ids = request.GET.getlist("assessment_ids")
#     # print(" RAW assessment_ids from request:", assessment_ids)
#     logger.info(f"RAW assessment_ids from request: {assessment_ids}")

#     model_labs = request.GET.getlist("model_labs")
#     include_practical = request.GET.get("include_practical") == "1"


#     # ---------------- FETCH STUDENTS ----------------
#     enrollments = (
#         CourseEnrollment.objects
#         .select_related("student")
#         .filter(
#             course_id=selection["course_id"],
#             batch=selection["batch"],
#             section=selection["section"],
#             enroll=True
#         )
#         .order_by("student__reg_no")
#     )

#     students = []
#     student_ids = []

#     for e in enrollments:
#         if e.student:
#             students.append({
#                 "id": e.student.id,
#                 "reg_no": e.student.reg_no,
#                 "name": e.student.name
#             })
#             student_ids.append(e.student.id)

#     # ---------------- RESOLVE ASSESSMENTS ----------------
#     assessment_objs = []
#     # print("assesment_obj (initial empty):", assessment_objs)


#     for aid in assessment_ids:
#         clean_id = aid.replace("am-", "").strip()

#         am = Assessment_master.objects.filter(
#             id=clean_id,
#             course_id=selection["course_id"]
#         ).first()

#         if not am:
#             am = Assessment_master.objects.filter(
#                 assessment_id=clean_id,
#                 course_id=selection["course_id"]
#             ).first()

#         if am:
#             assessment_objs.append(am)
#             # print("    Found Assessment:")
#             # print("      DB id:", am.id)
#             # print("      assessment_id field:", am.assessment_id)
#             # print("      Assessmentname:", am.Assessmentname)
#             # print("      customAssessmentname:", am.customAssessmentname)
#             # print("      weightage:", am.weightage)
        
       
#     # ---------------- RESOLVE MODEL LABS ----------------
#     model_lab_objs = []

#     if model_labs:
#         model_lab_objs = list(
#             ModelLab.objects.filter(model_lab_name__in=model_labs)
#         )
#     marks_map = {}

#     # ---------------- FETCH MODEL LAB MARKS ----------------
#     # ---------------- FETCH MODEL LAB MARKS (AVERAGE) ----------------
#     MODEL_LAB_COL_ID = "MODEL_LAB"

#     from collections import defaultdict
#     student_lab_totals = defaultdict(list)

#     if model_lab_objs:
#         ml_qs = ModelLabMarks.objects.filter(
#             student_id__in=student_ids,
#             model_lab__in=model_lab_objs,
#             batch=selection["batch"],
#             section=selection["section"],
#         )

#         # Respect selected IAT
#         # Respect selected IAT (skip filter for ALL/BOTH/empty)
#         iat_val = (selection.get("iat") or "").strip().lower()
#         if iat_val and iat_val not in ("both", "all"):
#             ml_qs = ml_qs.filter(internal_assessment__iat__iexact=selection["iat"])


#         # Collect all lab marks per student
#         for ml in ml_qs:
#             student_lab_totals[ml.student_id].append(ml.total)

#         # Calculate INTEGER average per student
#         for student_id, totals in student_lab_totals.items():
#             avg = sum(totals) // len(totals)   # 👈 integer average (NO FLOAT)
#             marks_map[(student_id, MODEL_LAB_COL_ID)] = avg

        
#         # ---------------- ADD MODEL LABS AS ACTIVITY ASSESSMENTS ----------------
   
#          # ---------------- FETCH PRACTICAL MARKS (AVERAGE, 75) ----------------
#     PRACTICAL_COL_ID = "PRACTICAL"

#     from collections import defaultdict
#     student_practical_totals = defaultdict(list)
#     practical_experiments = set()

#     if include_practical:
#         exp_qs = experiment_marks.objects.filter(
#             student_id__in=student_ids,
#             courses__course_id=selection["course_id"],
#         )

#         # Respect selected IAT (skip filter for ALL/BOTH/empty)
#         iat_val = (selection.get("iat") or "").strip().lower()
#         if iat_val and iat_val not in ("both", "all"):
#             exp_qs = exp_qs.filter(assessment__iat__iexact=selection["iat"])

#         # Collect experiment totals per student
#         for exp in exp_qs:
#             student_practical_totals[exp.student_id].append(exp.total)
#             if exp.experiment_no:
#                 practical_experiments.add(exp.experiment_no)

#         # Compute INTEGER average per student
#         for student_id, totals in student_practical_totals.items():
#             avg = sum(totals) // len(totals)
#             marks_map[(student_id, PRACTICAL_COL_ID)] = avg




#     # ---------------- FETCH MARKS (MATCH BY NAME) ----------------
#     marks_qs = AssessmentMark.objects.filter(
#         student_id__in=student_ids,
#         assessment__in=assessment_objs
#     ).select_related("assessment")

#     # 🔑 KEY CHANGE: NAME-BASED MAP
    

#     for m in marks_qs:
#         am_id = m.assessment_id 
#         value = (
#             int(m.marks_weighted)
#             if m.marks_weighted is not None
#             else (m.marks_raw if m.marks_raw is not None else m.marks)
#         )# Assessment_master primary key

#         marks_map[(m.student_id, am_id)] = value


        

        

#     # ---------------- PDF SETUP ----------------
#     buffer = io.BytesIO()
#     doc = SimpleDocTemplate(
#         buffer,
#         pagesize=landscape(A4),
#         leftMargin=15 * mm,
#         rightMargin=15 * mm,
#         topMargin=42 * mm,
#         bottomMargin=15 * mm
#     )

#     styles = getSampleStyleSheet()
#     style_small = ParagraphStyle(
#         "small",
#         parent=styles["Normal"],
#         fontSize=9,
#         textColor=colors.grey
#     )

#     story = []

#     # =================================================
#     # THEORY vs ACTIVITY (UNCHANGED)
#     # =================================================
#     theory_assessments = []
#     activity_assessments = []
#     practical_assessments = []

#     selected_iat = (selection["iat"] or "").strip().lower()   # ex: "iat2" or "BOTH"

#     for am in assessment_objs:
#         name = (am.customAssessmentname or am.Assessmentname or "").strip()
#         if not name:
#             continue

#         name_key = name.strip().lower()   # ex: "iat 2", "IAT2", "Assignment 1"

#         entry = {
#             "id": am.id,
#             "name": name,
#             "weightage": am.weightage or ""
#         }

#         # ✅ THEORY condition:
#         # 1) if user selected IAT1/IAT2 and the assessment name matches it
#         # 2) OR if assessment name contains "iat" (common case)
#         is_theory = False

#         if selected_iat and selected_iat not in ("both", "all"):

#             # match "iat2" with "iat 2" / "iat2"
#             compact_name = name_key.replace(" ", "")
#             if selected_iat.replace(" ", "") in compact_name:
#                 is_theory = True

#         # fallback: any assessment whose name contains "iat" treat as theory
#         if "iat" in name_key:
#             is_theory = True

#         if is_theory:
#             theory_assessments.append(entry)
#         else:
#             activity_assessments.append(entry)
#     if model_lab_objs:
#         activity_assessments.append({
#             "id": MODEL_LAB_COL_ID,
#             "name": "Model Lab",
#             "weightage": "100"
#         })
#     if include_practical:
#         practical_assessments.append({
#             "id": PRACTICAL_COL_ID,
#             "name": "Practical",
#             "weightage": "75"
#         })



#     def unique(items):
#         seen = set()
#         out = []
#         for i in items:
#             if i["id"] not in seen:
#                 seen.add(i["id"])
#                 out.append(i)
#         return out

#     theory_assessments = unique(theory_assessments)
#     activity_assessments = unique(activity_assessments)
#     practical_assessments = unique(practical_assessments)
    
#     # =========================================================
#     # DEBUG: STUDENT TOTALS (THEORY / ACTIVITY / PRACTICAL)
#     # =========================================================
#     def _num(x):
#         try:
#             if x is None or x == "":
#                 return 0
#             return int(float(x))
#         except Exception:
#             return 0

#     theory_ids = [a["id"] for a in theory_assessments]
#     activity_ids = [a["id"] for a in activity_assessments]
#     practical_ids = [a["id"] for a in practical_assessments]

#     # print("\n================ BUCKET TOTAL DEBUG ================")
#     # print(" Theory columns:", [a["name"] for a in theory_assessments])
#     # print("Activity columns:", [a["name"] for a in activity_assessments])
#     # print("Practical columns:", [a["name"] for a in practical_assessments])

#     # % from CourseHourConfig (safe defaults)
#     tp = theory_pct if cfg else 0
#     ap = activity_pct if cfg else 0
#     pp = practical_pct if cfg else 0
    
#     weighted_grand_total = tp + ap + pp
    

#     def _safe_pct(score, total):
#         return (score / total * 100) if total > 0 else 0.0

#     # print("\n================ BUCKET TOTAL + WEIGHTED DEBUG ================")
#     # print(" Config % -> Theory/Activity/Practical:", tp, ap, pp)
#     # print(" Theory columns:", [a["name"] for a in theory_assessments])
#     # print(" Activity columns:", [a["name"] for a in activity_assessments])
#     # print(" Practical columns:", [a["name"] for a in practical_assessments])
#     weighted_grand_map = {}
#     converted_grand_map = {}    
    
    
#     for s in students:
#         sid = s["id"]

#         # raw totals (sum of selected columns)
#         theory_total = sum(_num(marks_map.get((sid, aid))) for aid in theory_ids)
#         activity_total = sum(_num(marks_map.get((sid, aid))) for aid in activity_ids)
#         practical_total = sum(_num(marks_map.get((sid, aid))) for aid in practical_ids)

#         # max totals (based on selected assessments weightage)
#         # NOTE: uses Assessment_master.weightage; Model Lab assumed 100; Practical assumed 75 (as you already set)
#         theory_max = sum(_to_int(a.get("weightage", 0)) for a in theory_assessments)
#         activity_max = sum(_to_int(a.get("weightage", 0)) for a in activity_assessments)
#         practical_max = sum(_to_int(a.get("weightage", 0)) for a in practical_assessments)

#         # bucket % scored by student
#         theory_scored_pct = _safe_pct(theory_total, theory_max)
#         activity_scored_pct = _safe_pct(activity_total, activity_max)
#         practical_scored_pct = _safe_pct(practical_total, practical_max)

#         # ✅ convert bucket performance into configured contribution
#         theory_weighted = (theory_scored_pct * tp) / 100
#         activity_weighted = (activity_scored_pct * ap) / 100
#         practical_weighted = (practical_scored_pct * pp) / 100

#         weighted_grand = theory_weighted + activity_weighted + practical_weighted
#         if lecture_hours > 0 and lab_hours == 0:
#             tot_to_convert = 40
#         elif lecture_hours > 0 and lab_hours > 0:
#             tot_to_convert = 50
#         elif lecture_hours == 0 and lab_hours > 0:
#             tot_to_convert = 60
#         else:
#             tot_to_convert = 100  # safe fallback

#         # print("🎯 TOTAL TO CONVERT:", tot_to_convert)
#         converted_total = (weighted_grand / 100) * tot_to_convert

#         converted_total_to_display = f"{int(round(converted_total))}"
        
#         # print(" converted_total_to_display", converted_total_to_display)
        
#         weighted_display = int(round(weighted_grand))
#         converted_total = (weighted_grand / 100) * tot_to_convert
#         converted_display = int(round(converted_total))

#         weighted_grand_map[sid] = str(weighted_display)
#         converted_grand_map[sid] = str(converted_display)

        

        
#     meta_lines = [
#         f"<b>Batch-Section:</b> {selection['batch']} - {selection['section']}",
#         f"<b>IAT:</b> {selection['iat'] or 'ALL'}",
#     ]

#     if theory_assessments:
#         meta_lines.append("<b>Theory Assessment(s):</b> " + ", ".join(a["name"] for a in theory_assessments))

#     if activity_assessments:
#         meta_lines.append("<b>Activity Assessment(s):</b> " + ", ".join(a["name"] for a in activity_assessments))
#     if practical_assessments:
#         meta_lines.append("<b>practical_assessments(s):</b> " + ", ".join(a["name"] for a in practical_assessments))
#     if model_labs:
#         meta_lines.append(f"<b>Model Labs:</b> {', '.join(model_labs)}")

    
#     if include_practical and practical_experiments:
#         exp_list = ", ".join(f"Exp {e}" for e in sorted(practical_experiments))
#         meta_lines.append(f"<b>Practical Included:</b> {exp_list}")

#     story.append(Paragraph("<br/>".join(meta_lines), style_small))
#     story.append(Spacer(1, 10))

#     # =================================================
#     # TABLE (NAME-BASED MARK MATCHING)
#     # =================================================
#     assessment_columns = theory_assessments + activity_assessments + practical_assessments

#     # Short labels for assessment columns
#     assessment_label_map = {}
#     for idx, a in enumerate(assessment_columns, start=1):
#         assessment_label_map[a["id"]] = f"X{idx}"

#     def D(v):
    
#         try:
#             if v is None or v == "":
#                 return Decimal("0.00")
#             return Decimal(str(v)).quantize(Decimal("0.00"), rounding=ROUND_HALF_UP)
#         except Exception:
#             return Decimal("0.00")
    
#     def D_INT(value):
#         """
#         Round to nearest integer and return Decimal without decimals.
#         """
#         if value is None:
#             return None
#         return Decimal(value).quantize(Decimal("0"), rounding=ROUND_HALF_UP)

    
#     enrollment_by_student_id = {e.student_id: e for e in enrollments if e.student_id}
#     # ---------------------------------------------------------
#     # SAVE CONSOLIDATED RESULT INTO DB (PER STUDENT)
#     # ---------------------------------------------------------
#     theory_names = ", ".join(a["name"] for a in theory_assessments) if theory_assessments else ""
#     activity_names = ", ".join(a["name"] for a in activity_assessments) if activity_assessments else ""
#     practical_names = ", ".join(a["name"] for a in practical_assessments) if practical_assessments else ""

#     theory_max = sum(_to_int(a.get("weightage", 0)) for a in theory_assessments)
#     activity_max = sum(_to_int(a.get("weightage", 0)) for a in activity_assessments)
#     practical_max = sum(_to_int(a.get("weightage", 0)) for a in practical_assessments)

#     tp = theory_pct if cfg else 0
#     ap = activity_pct if cfg else 0
#     pp = practical_pct if cfg else 0
#     weighted_grand_total = tp + ap + pp

#     def _safe_pct(score, total):
#         return (score / total * 100) if total > 0 else 0.0

#     enrollment_by_student_id = {e.student_id: e for e in enrollments if e.student_id}
#     # ✅ Student meta (department/degree/batch/section) - fetch once (fast)
#     student_meta = {
#         st.id: st
#         for st in StudentDetails.objects
#             .filter(id__in=student_ids)
#             .select_related("department__degree")
#             .only("id", "batch", "section", "semester", "department", "department__degree")
#     }

#     def _display_max_str(assessments):
#         # comma string of weightages in the SAME ORDER as PDF columns
#         # ex: "35,35"
#         return ",".join(str(_to_int(a.get("weightage", 0))) for a in assessments) if assessments else ""

#     def _display_actual_str(sid, assessments):
#         # comma string of student marks aligned to the SAME ORDER
#         # ex: "28,31"
#         out = []
#         for a in assessments:
#             v = marks_map.get((sid, a["id"]))
#             out.append("" if v is None else str(_num(v)))
#         return ",".join(out) if assessments else ""


#     rows_to_save = []

#     for s in students:
#         sid = s["id"]

#         theory_total = sum(_num(marks_map.get((sid, aid))) for aid in theory_ids)
#         activity_total = sum(_num(marks_map.get((sid, aid))) for aid in activity_ids)
#         practical_total = sum(_num(marks_map.get((sid, aid))) for aid in practical_ids)

#         # compute weighted numeric (same as your debug)
#         theory_scored_pct = _safe_pct(theory_total, theory_max)
#         activity_scored_pct = _safe_pct(activity_total, activity_max)
#         practical_scored_pct = _safe_pct(practical_total, practical_max)

#         theory_weighted = (theory_scored_pct * tp) / 100
#         activity_weighted = (activity_scored_pct * ap) / 100
#         practical_weighted = (practical_scored_pct * pp) / 100
#         weighted_grand = theory_weighted + activity_weighted + practical_weighted

#         st = student_meta.get(sid)
#         dept = st.department if st else None
#         deg = dept.degree if dept else None

#         rows_to_save.append(
#             ConsolidatedAssessmentResult(
#                 student_id=sid,
#                 course=enrollment_by_student_id.get(sid),     # CourseEnrollment FK
#                 hour_config=cfg,   
#                 department=dept,
#                 degree=deg,
#                 batch=(st.batch if st else None) or selection.get("batch") or None,
#                 section=(st.section if st else None) or selection.get("section") or None,
#                 current_semester=getattr(st, "semester", None),# ⭐ CourseHourConfig FK (stores id used)
#                 theory_assessment_name=theory_names or None,
#                 activity_assessment_name=activity_names or None,
#                 practical_assessment_name=practical_names or None,

#                 theory_max_mark=D_INT(tp),
#                 theory_actual_mark=D_INT(theory_weighted),

#                 activity_max_mark=D_INT(ap),
#                 activity_actual_mark=D_INT(activity_weighted),

#                 practical_max_mark=D_INT(pp),
#                 practical_actual_mark=D_INT(practical_weighted),

#                 theory_display_max_mark=_display_max_str(theory_assessments),
#                 theory_display_actual_mark=_display_actual_str(sid, theory_assessments),

#                 activity_display_max_mark=_display_max_str(activity_assessments),
#                 activity_display_actual_mark=_display_actual_str(sid, activity_assessments),

#                 practical_display_max_mark=_display_max_str(practical_assessments),
#                 practical_display_actual_mark=_display_actual_str(sid, practical_assessments),
#             )
#         )

#     # Save in one shot
#     # ---------------- UPSERT (update existing, create missing) ----------------
#     # "Same details" key = (student, course, hour_config)
#     existing_qs = ConsolidatedAssessmentResult.objects.filter(
#         student_id__in=student_ids,
#         course__in=enrollments,        # course is FK to CourseEnrollment
#         hour_config=cfg
#     )

#     existing_map = {
#         (r.student_id, r.course_id, r.hour_config_id): r
#         for r in existing_qs
#     }

#     to_create = []
#     to_update = []

#     for obj in rows_to_save:
#         key = (obj.student_id, obj.course_id, obj.hour_config_id)
#         old = existing_map.get(key)

#         if old:
#             old.theory_assessment_name = obj.theory_assessment_name
#             old.activity_assessment_name = obj.activity_assessment_name
#             old.practical_assessment_name = obj.practical_assessment_name
#             old.current_semester = obj.current_semester

#             old.theory_max_mark = obj.theory_max_mark
#             old.theory_actual_mark = obj.theory_actual_mark

#             old.activity_max_mark = obj.activity_max_mark
#             old.activity_actual_mark = obj.activity_actual_mark

#             old.practical_max_mark = obj.practical_max_mark
#             old.practical_actual_mark = obj.practical_actual_mark
            
#             old.theory_display_max_mark = obj.theory_display_max_mark
#             old.theory_display_actual_mark = obj.theory_display_actual_mark

#             old.activity_display_max_mark = obj.activity_display_max_mark
#             old.activity_display_actual_mark = obj.activity_display_actual_mark

#             old.practical_display_max_mark = obj.practical_display_max_mark
#             old.practical_display_actual_mark = obj.practical_display_actual_mark

#             old.department = obj.department
#             old.degree = obj.degree
#             old.batch = obj.batch
#             old.section = obj.section

#             # generated_on will NOT change (auto_now_add)
#             to_update.append(old)
#         else:
#             to_create.append(obj)

#     with transaction.atomic():
#         if to_create:
#             ConsolidatedAssessmentResult.objects.bulk_create(to_create, batch_size=500)

#         if to_update:
#             ConsolidatedAssessmentResult.objects.bulk_update(
#                 to_update,
#                 fields=[
#                     "theory_assessment_name",
#                     "activity_assessment_name",
#                     "practical_assessment_name",
#                     "theory_max_mark",
#                     "theory_actual_mark",
#                     "activity_max_mark",
#                     "activity_actual_mark",
#                     "practical_max_mark",
#                     "practical_actual_mark",
#                     "theory_display_max_mark",
#                     "theory_display_actual_mark",
#                     "activity_display_max_mark",
#                     "activity_display_actual_mark",
#                     "practical_display_max_mark",
#                     "practical_display_actual_mark",
#                     "department",
#                     "degree",
#                     "batch",
#                     "section",
#                     "current_semester",
#                 ],
#                 batch_size=500
#             )

#     header = ["Reg No", "Name"]
#     for a in assessment_columns:
#         short_label = assessment_label_map[a["id"]]
#         header.append(f'{short_label} ({a["weightage"]})' if a["weightage"] else short_label)
#     header.append(f"Total ({weighted_grand_total})")
#     header.append(f"Conv ({tot_to_convert})")

#     data = [header]

#     for s in students:
#         row = [s["reg_no"], s["name"]]

#         for a in assessment_columns:
#             mark = marks_map.get((s["id"], a["id"]))
#             row.append("" if mark is None else str(mark))

#         row.append(weighted_grand_map.get(s["id"], ""))
#         row.append(converted_grand_map.get(s["id"], ""))
#         data.append(row)

#     total_width = 297 * mm - 30 * mm

#     # fixed widths
#     w_reg = 30 * mm
#     w_name = 55 * mm
#     w_total = 22 * mm
#     w_conv = 22 * mm

#     col_widths = [w_reg, w_name]

#     # remaining width for assessment columns + total
#     remaining = total_width - (w_reg + w_name + w_total + w_conv)

#     if assessment_columns:
#         each = remaining / len(assessment_columns)
#         col_widths += [each] * len(assessment_columns)
#         col_widths += [w_total, w_conv]

#     table = Table(data, repeatRows=1, colWidths=col_widths)
#     table.setStyle(TableStyle([
#         ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
#         ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
#         ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
#         ("ALIGN", (2, 1), (-1, -1), "CENTER"),
#         ("ALIGN", (0, 1), (1, -1), "LEFT"),
#         ("FONTSIZE", (0, 0), (-1, -1), 9),
#         ("TOPPADDING", (0, 0), (-1, -1), 6),
#         ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
#     ]))

#     story.append(table)

#     # ---------------- ASSESSMENT MAPPING AT END OF PDF ----------------
#     if assessment_columns:
#         story.append(Spacer(1, 10))

#         legend_lines = ["<b>Assessment Mapping:</b>"]
#         for a in assessment_columns:
#             short_label = assessment_label_map[a["id"]]
#             if a.get("weightage"):
#                 legend_lines.append(f"{short_label} = {a['name']} ({a['weightage']})")
#             else:
#                 legend_lines.append(f"{short_label} = {a['name']}")

#         story.append(Paragraph("<br/>".join(legend_lines), style_small))

#     # ---------------- HEADER / FOOTER (UNCHANGED) ----------------
#     def _find_logo():
#         p = finders.find("images/ritlogo.png")
#         return p if p and os.path.exists(p) else None

#     logo_path = _find_logo()

#     def _on_page(canvas, doc):
#         canvas.saveState()
#         w, h = landscape(A4)
#         left = 15 * mm

#         if logo_path:
#             try:
#                 img = ImageReader(logo_path)
#                 canvas.drawImage(img, left, h - 26 * mm,
#                                  height=18 * mm, preserveAspectRatio=True)
#             except:
#                 pass

#         canvas.setFont("Helvetica-Bold", 16)
#         canvas.drawCentredString(w / 2, h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")
#         canvas.setFont("Helvetica", 10)
#         canvas.drawCentredString(w / 2, h - 15 * mm, "Rajapalayam - 626117")
#         canvas.drawCentredString(w / 2, h - 19 * mm, "Affiliated to Anna University, Chennai")
#         canvas.setFont("Helvetica-Bold", 13)
#         canvas.drawCentredString(w / 2, h - 26 * mm, "Overall Consolidated Statement")
#         canvas.setFont("Helvetica", 10)
#         canvas.drawCentredString(w / 2, h - 31 * mm, f"{selection['degree_code']} — {selection['degree_name']}")
#         canvas.drawCentredString(w / 2, h - 35 * mm, selection["department"])
#         canvas.drawCentredString(w / 2, h - 39 * mm, f"{selection['course_code']} — {selection['course_title']}")
#         canvas.line(left, h - 42 * mm, w - 15 * mm, h - 42 * mm)
#         canvas.setFont("Helvetica", 8)
#         canvas.setFillColor(colors.grey)
#         canvas.drawRightString(w - 15 * mm, 12 * mm, f"Page {canvas.getPageNumber()}")
#         canvas.restoreState()

#     doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

#     pdf = buffer.getvalue()
#     buffer.close()

#     response = HttpResponse(pdf, content_type="application/pdf")
#     response["Content-Disposition"] = f'inline; filename="Consolidate_{selection["course_code"]}.pdf"'
#     return response



from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.utils import ImageReader
from django.db.models import Q, Sum, Max
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.utils import ImageReader
from django.db.models import Q, Sum, Max



def overall_consolidate_pdf(request):
    logger = logging.getLogger(__name__)

    from collections import defaultdict
    from decimal import Decimal, ROUND_HALF_UP
    import io
    import os

    from django.db.models import Q, Sum
    from django.http import HttpResponse
    from django.contrib.staticfiles import finders

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.utils import ImageReader

    logger = logging.getLogger(__name__)

    # ---------------- REQUEST PARAMS ----------------
    selection = {
        "course_id": request.GET.get("course_id", ""),
        "course_code": request.GET.get("course_code", ""),
        "course_title": request.GET.get("course_title", ""),
        "department": request.GET.get("department", ""),
        "department_id": request.GET.get("department_id", ""),
        "degree_id": request.GET.get("degree_id", ""),
        "batch": request.GET.get("batch", ""),
        "section": request.GET.get("section", ""),
        "degree_code": request.GET.get("degree_code", ""),
        "degree_name": request.GET.get("degree_name", ""),
        "iat": request.GET.get("iat", ""),
    }

    def _to_int(v, default=0):
        try:
            if v is None or str(v).strip() == "":
                return default
            return int(float(str(v).strip()))
        except Exception:
            return default

    def _num(x):
        try:
            if x is None or x == "":
                return 0
            return int(float(x))
        except Exception:
            return 0

    def _safe_pct(score, total):
        return (score / total * 100) if total > 0 else 0.0

    def _norm_iat(v):
        return (v or "").strip().lower().replace(" ", "")

    def _convert_to_target(obtained, source_total, target_total):
        obtained = _to_int(obtained, 0)
        source_total = _to_int(source_total, 0)
        target_total = _to_int(target_total, 0)
        if source_total <= 0 or target_total <= 0:
            return 0
        return int(round((obtained / source_total) * target_total))

    hour_config_id = (
        request.GET.get("hour_config_id")
        or request.GET.get("coursehour_config_id")
        or request.GET.get("hour_config")
        or ""
    )

    # ---------------- FETCH COURSE HOURS ----------------
    ch_qs = CourseHours.objects.select_related("course__regulation", "hour_config").filter(
        course_id=selection["course_id"]
    )
    if hour_config_id:
        ch_qs = ch_qs.filter(hour_config_id=hour_config_id)

    course_hours = ch_qs.first()

    if (not hour_config_id) and course_hours and course_hours.hour_config_id:
        hour_config_id = str(course_hours.hour_config_id)

    # ---------------- FETCH CONFIG ----------------
    cfg = CourseHourConfig.objects.select_related("regulation").filter(
        id=hour_config_id
    ).first() if hour_config_id else None

    # ---------------- REGULATION + L/T/L HOURS ----------------
    regulation_value = ""
    if cfg and cfg.regulation:
        regulation_value = str(cfg.regulation)
    elif course_hours and course_hours.course and course_hours.course.regulation:
        regulation_value = str(course_hours.course.regulation)

    if cfg:
        lecture_hours = _to_int(cfg.lecture_hours)
        tutorial_hours = _to_int(cfg.tutorial_hours)
        lab_hours = _to_int(cfg.laboratory_hours)

        theory_pct = _to_int(cfg.theory_percentage)
        activity_pct = _to_int(cfg.activity_percentage)
        practical_pct = _to_int(cfg.practical_percentage)

        if lecture_hours > 0 and tutorial_hours == 0 and lab_hours == 0:
            theory_pct = min(100, theory_pct * 2)



    elif course_hours:
        lecture_hours = _to_int(course_hours.leture_npwk)
        tutorial_hours = _to_int(course_hours.tutorial_npwk)
        lab_hours = _to_int(course_hours.laboratory_npwk)

        theory_pct = 0
        activity_pct = 0
        practical_pct = 0

        logger.info(
            f"Regulation: {regulation_value or 'N/A'} | "
            f"Lecture Hours: {lecture_hours} | "
            f"Tutorial Hours: {tutorial_hours} | "
            f"Laboratory Hours: {lab_hours}"
        )
    else:
        lecture_hours = 0
        tutorial_hours = 0
        lab_hours = 0
        theory_pct = 0
        activity_pct = 0
        practical_pct = 0

        logger.info(
            f"Regulation: {regulation_value or 'N/A'} | "
            f"Lecture Hours: 0 | Tutorial Hours: 0 | Laboratory Hours: 0"
        )

    # ---------------- CHECK INTERNAL ASSESSMENT TEMPLATE ----------------
    internal_template = None
    course_type = None

    if lecture_hours > 0 and tutorial_hours == 0 and lab_hours == 0:
        course_type = "theory"
    elif lecture_hours == 0 and tutorial_hours == 0 and lab_hours > 0:
        course_type = "practical"
    else:
        course_type = "theory_lab"

    regulation_obj = None
    degree_obj = None

    if cfg and cfg.regulation:
        regulation_obj = cfg.regulation

    if selection.get("degree_id"):
        degree_obj = Degree.objects.filter(id=selection["degree_id"]).first()
    elif cfg and getattr(cfg, "degree_id", None):
        degree_obj = cfg.degree

    regulation_text = str(regulation_obj).strip() if regulation_obj else (regulation_value or "").strip()

    internal_template = InternalAssessmentMasterTemplate.objects.filter(
        degree=degree_obj,
        regulation=regulation_obj,
        course_type=course_type
    ).first()






    assessment_ids = request.GET.getlist("assessment_ids")
    selected_iat_groups = request.GET.getlist("selected_iat_groups")
    model_labs = request.GET.getlist("model_labs")
    include_practical = request.GET.get("include_practical") == "1"

    logger.info(f"RAW assessment_ids from request: {assessment_ids}")
    logger.info(f"RAW selected_iat_groups from request: {selected_iat_groups}")

    # ---------------- FETCH STUDENTS ----------------
    enrollments = (
        CourseEnrollment.objects
        .select_related("student")
        .filter(
            course_id=selection["course_id"],
            batch=selection["batch"],
            section=selection["section"],
            enroll=True
        )
        .order_by("student__reg_no")
    )

    students = []
    student_ids = []

    for e in enrollments:
        if e.student:
            students.append({
                "id": e.student.id,
                "reg_no": e.student.reg_no,
                "name": e.student.name
            })
            student_ids.append(e.student.id)

    # ---------------- RESOLVE NORMAL ASSESSMENTS ----------------
    assessment_objs = []

    for aid in assessment_ids:
        if not aid or not str(aid).strip():
            continue

        clean_id = str(aid).replace("am-", "").strip()

        am = Assessment_master.objects.filter(
            id=clean_id,
            course_id=selection["course_id"]
        ).first()

        if not am:
            am = Assessment_master.objects.filter(
                assessment_id=clean_id,
                course_id=selection["course_id"]
            ).first()

        if am:
            assessment_objs.append(am)

    # ---------------- RESOLVE MODEL LABS ----------------
    model_lab_objs = []
    if model_labs:
        model_lab_objs = list(
            ModelLab.objects.filter(model_lab_name__in=model_labs)
        )

    marks_map = {}
    iat_group_entries = []
    iat_weightage_map = {}

    # ---------------- FETCH IAT GROUP MARKS FROM StudentInternalMark ----------------
    if selected_iat_groups and student_ids:
        sim_qs = StudentInternalMark.objects.filter(
            student_id__in=student_ids,
            course_id=selection["course_id"],
            batch=selection["batch"],
            section=selection["section"],
        )

        if selection.get("degree_id"):
            sim_qs = sim_qs.filter(degree_id=selection["degree_id"])

        if selection.get("department_id"):
            sim_qs = sim_qs.filter(department_id=selection["department_id"])

        exam_q = Q()
        for grp in selected_iat_groups:
            grp = (grp or "").strip()
            if grp:
                exam_q |= Q(exam_name__iexact=grp)

        if exam_q:
            sim_qs = sim_qs.filter(exam_q)
        else:
            sim_qs = sim_qs.none()

        option_level_rows = (
            sim_qs.values(
                "student_id",
                "exam_name",
                "part_name",
                "question_number",
                "option_letter",
            )
            .annotate(
                option_obtained=Sum("marks_obtained"),
                option_max=Sum("max_marks"),
            )
            .order_by("student_id", "exam_name", "part_name", "question_number", "option_letter")
        )

        question_choice_map = defaultdict(lambda: {"chosen_obtained": 0, "chosen_max": 0})

        for row in option_level_rows:
            key = (
                row.get("student_id"),
                (row.get("exam_name") or "").strip(),
                row.get("part_name"),
                row.get("question_number"),
            )

            option_obtained = _to_int(row.get("option_obtained", 0))
            option_max = _to_int(row.get("option_max", 0))

            if option_max > question_choice_map[key]["chosen_max"]:
                question_choice_map[key]["chosen_max"] = option_max
                question_choice_map[key]["chosen_obtained"] = option_obtained
            elif (
                option_max == question_choice_map[key]["chosen_max"]
                and option_obtained > question_choice_map[key]["chosen_obtained"]
            ):
                question_choice_map[key]["chosen_obtained"] = option_obtained

        exam_totals = defaultdict(lambda: {"total_obtained": 0, "total_max": 0})

        for (student_id, exam_name, part_name, question_number), chosen in question_choice_map.items():
            if not exam_name or not student_id:
                continue

            key = (student_id, exam_name)
            exam_totals[key]["total_obtained"] += _to_int(chosen.get("chosen_obtained", 0))
            exam_totals[key]["total_max"] += _to_int(chosen.get("chosen_max", 0))

        for (student_id, exam_name), totals in exam_totals.items():
            col_id = f"IAT::{_norm_iat(exam_name)}"
            marks_map[(student_id, col_id)] = _to_int(totals.get("total_obtained", 0))

            if col_id not in iat_weightage_map:
                iat_weightage_map[col_id] = _to_int(totals.get("total_max", 0))

        for grp in selected_iat_groups:
            grp = (grp or "").strip()
            if not grp:
                continue
            col_id = f"IAT::{_norm_iat(grp)}"
            iat_group_entries.append({
                "id": col_id,
                "name": grp,
                "weightage": str(iat_weightage_map.get(col_id, 0) or "")
            })

    # ---------------- FETCH MODEL LAB MARKS (AVERAGE) ----------------
    MODEL_LAB_COL_ID = "MODEL_LAB"
    student_lab_totals = defaultdict(list)

    if model_lab_objs:
        ml_qs = ModelLabMarks.objects.filter(
            student_id__in=student_ids,
            model_lab__in=model_lab_objs,
            batch=selection["batch"],
            section=selection["section"],
        )

        iat_val = (selection.get("iat") or "").strip().lower()
        if iat_val and iat_val not in ("both", "all"):
            ml_qs = ml_qs.filter(internal_assessment__iat__iexact=selection["iat"])

        for ml in ml_qs:
            student_lab_totals[ml.student_id].append(ml.total)

        for student_id, totals in student_lab_totals.items():
            avg = sum(totals) // len(totals)
            marks_map[(student_id, MODEL_LAB_COL_ID)] = avg

    # ---------------- FETCH PRACTICAL MARKS (AVERAGE) ----------------
    PRACTICAL_COL_ID = "PRACTICAL"
    student_practical_totals = defaultdict(list)
    practical_experiments = set()

    if include_practical:
        exp_qs = experiment_marks.objects.filter(
            student_id__in=student_ids,
            courses__course_id=selection["course_id"],
        )

        iat_val = (selection.get("iat") or "").strip().lower()
        if iat_val and iat_val not in ("both", "all"):
            exp_qs = exp_qs.filter(assessment__iat__iexact=selection["iat"])

        for exp in exp_qs:
            student_practical_totals[exp.student_id].append(exp.total)
            if exp.experiment_no:
                practical_experiments.add(exp.experiment_no)

        for student_id, totals in student_practical_totals.items():
            avg = sum(totals) // len(totals)
            marks_map[(student_id, PRACTICAL_COL_ID)] = avg

    # ---------------- FETCH NORMAL AssessmentMark MARKS ----------------
    marks_qs = AssessmentMark.objects.filter(
        student_id__in=student_ids,
        assessment__in=assessment_objs
    ).select_related("assessment")

    for m in marks_qs:
        am_id = m.assessment_id
        value = (
            int(m.marks_weighted)
            if m.marks_weighted is not None
            else (m.marks_raw if m.marks_raw is not None else m.marks)
        )
        marks_map[(m.student_id, am_id)] = value

    # ---------------- PDF SETUP ----------------
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=42 * mm,
        bottomMargin=15 * mm
    )

    styles = getSampleStyleSheet()
    style_small = ParagraphStyle(
        "small",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey
    )

    story = []

    # =================================================
    # THEORY / ACTIVITY / PRACTICAL SPLIT
    # =================================================
    theory_assessments = []
    activity_assessments = []
    practical_assessments = []

    selected_iat = (selection["iat"] or "").strip().lower()

    for iat_entry in iat_group_entries:
        theory_assessments.append(iat_entry)

    for am in assessment_objs:
        name = (am.customAssessmentname or am.Assessmentname or "").strip()
        if not name:
            continue

        name_key = name.strip().lower()

        entry = {
            "id": am.id,
            "name": name,
            "weightage": am.weightage or ""
        }

        is_theory = False

        if selected_iat and selected_iat not in ("both", "all"):
            compact_name = name_key.replace(" ", "")
            if selected_iat.replace(" ", "") in compact_name:
                is_theory = True

        if "iat" in name_key:
            is_theory = True

        if is_theory:
            theory_assessments.append(entry)
        else:
            activity_assessments.append(entry)

    if model_lab_objs:
        activity_assessments.append({
            "id": MODEL_LAB_COL_ID,
            "name": "Model Lab",
            "weightage": "100"
        })

    if include_practical:
        practical_assessments.append({
            "id": PRACTICAL_COL_ID,
            "name": "Practical",
            "weightage": "75"
        })

    def unique(items):
        seen = set()
        out = []
        for i in items:
            if i["id"] not in seen:
                seen.add(i["id"])
                out.append(i)
        return out

    theory_assessments = unique(theory_assessments)
    activity_assessments = unique(activity_assessments)
    practical_assessments = unique(practical_assessments)

    theory_ids = [a["id"] for a in theory_assessments]
    activity_ids = [a["id"] for a in activity_assessments]
    practical_ids = [a["id"] for a in practical_assessments]

    tp = theory_pct if cfg else 0
    ap = activity_pct if cfg else 0
    pp = practical_pct if cfg else 0

    meta_lines = [
        f"<b>Batch-Section:</b> {selection['batch']} - {selection['section']}",
        f"<b>IAT:</b> {selection['iat'] or 'ALL'}",
        f"<b>Regulation:</b> {regulation_value or 'N/A'}",
        f"<b>L-T-L:</b> {lecture_hours}-{tutorial_hours}-{lab_hours}",
    ]

    if theory_assessments:
        meta_lines.append("<b>Theory Assessment(s):</b> " + ", ".join(a["name"] for a in theory_assessments))

    if activity_assessments:
        meta_lines.append("<b>Activity Assessment(s):</b> " + ", ".join(a["name"] for a in activity_assessments))

    if practical_assessments:
        meta_lines.append("<b>Practical Assessment(s):</b> " + ", ".join(a["name"] for a in practical_assessments))

    if model_labs:
        meta_lines.append(f"<b>Model Labs:</b> {', '.join(model_labs)}")

    if include_practical and practical_experiments:
        exp_list = ", ".join(f"Exp {e}" for e in sorted(practical_experiments))
        meta_lines.append(f"<b>Practical Included:</b> {exp_list}")

    # ---------------- CONVERSION TARGETS ----------------
    use_template_conversion = bool(
    internal_template and course_type in ("theory_lab", "theory", "practical")
)

    selected_iat_norm = _norm_iat(selection.get("iat"))

    converted_total_max = 0
    theory_target = 0
    activity_target = 0
    model_lab_target = 0
    practical_target = 0

    if use_template_conversion:
        

        if course_type == "theory_lab":
            model_lab_target = 25
            practical_target = 75

            if "1" in selected_iat_norm:
                theory_target = _to_int(internal_template.assessment1_test)
                activity_target = _to_int(internal_template.assessment1_assignment)
                
            elif "2" in selected_iat_norm:
                theory_target = _to_int(internal_template.assessment2_test)
                activity_target = _to_int(internal_template.assessment2_assignment)
                
            else:
                print("IAT not clearly IAT1/IAT2, template conversion skipped")

            converted_total_max = _to_int(internal_template.total_internal)


        elif course_type == "theory":
            if "1" in selected_iat_norm:
                theory_target = _to_int(internal_template.assessment1_test)
                activity_target = _to_int(internal_template.assessment1_assignment)
                
            elif "2" in selected_iat_norm:
                theory_target = _to_int(internal_template.assessment2_test)
                activity_target = _to_int(internal_template.assessment2_assignment)
                


            converted_total_max = theory_target + activity_target



        elif course_type == "practical":
            # For practical course:
            # IAT1 or IAT2 practical -> A1 Assignment
            # Model Lab -> A1 Test
            # Total = 100
            activity_target = _to_int(internal_template.assessment1_assignment)
            model_lab_target = _to_int(internal_template.assessment1_test)
            converted_total_max = _to_int(internal_template.total_internal)




    story.append(Paragraph("<br/>".join(meta_lines), style_small))
    story.append(Spacer(1, 10))

    # =================================================
    # TABLE
    # =================================================
    assessment_columns = theory_assessments + activity_assessments + practical_assessments

    assessment_label_map = {}
    next_x_idx = 1

    for a in assessment_columns:
        if str(a["id"]).startswith("IAT::"):
            assessment_label_map[a["id"]] = a["name"]
        else:
            assessment_label_map[a["id"]] = f"X{next_x_idx}"
            next_x_idx += 1

    def D_INT(value):
        if value is None:
            return None
        return Decimal(value).quantize(Decimal("0"), rounding=ROUND_HALF_UP)

    theory_max = sum(_to_int(a.get("weightage", 0)) for a in theory_assessments)

    # exclude model lab from general activity max
    activity_non_model_ids = [a["id"] for a in activity_assessments if a["id"] != MODEL_LAB_COL_ID]
    activity_non_model_max = sum(
        _to_int(a.get("weightage", 0)) for a in activity_assessments if a["id"] != MODEL_LAB_COL_ID
    )

    model_lab_max = sum(
        _to_int(a.get("weightage", 0)) for a in activity_assessments if a["id"] == MODEL_LAB_COL_ID
    )

    practical_max = sum(_to_int(a.get("weightage", 0)) for a in practical_assessments)

    # ---------------- PDF TABLE DATA ----------------
    header = ["Reg No", "Name"]
    for a in assessment_columns:
        short_label = assessment_label_map[a["id"]]
        if str(a["id"]).startswith("IAT::"):
            header.append(f'{a["name"]} ({a["weightage"]})' if a["weightage"] else a["name"])
        else:
            header.append(f'{short_label} ({a["weightage"]})' if a["weightage"] else short_label)

    if use_template_conversion and converted_total_max > 0:
        header.append(f"Total ({converted_total_max})")

    data = [header]

    for s in students:
        row = [s["reg_no"], s["name"]]

        for a in assessment_columns:
            mark = marks_map.get((s["id"], a["id"]))
            row.append("" if mark is None else str(mark))

        # -------- TEMPLATE CONVERTED TOTAL --------
        if use_template_conversion and converted_total_max > 0:
            student_id = s["id"]

            theory_obt = sum(_to_int(marks_map.get((student_id, aid), 0)) for aid in theory_ids)
            activity_obt = sum(_to_int(marks_map.get((student_id, aid), 0)) for aid in activity_non_model_ids)
            model_lab_obt = _to_int(marks_map.get((student_id, MODEL_LAB_COL_ID), 0))
            practical_obt = _to_int(marks_map.get((student_id, PRACTICAL_COL_ID), 0))

            theory_conv = 0
            activity_conv = 0
            model_lab_conv = 0
            practical_conv = 0

            if course_type == "theory_lab":
                theory_conv = _convert_to_target(theory_obt, theory_max, theory_target)
                activity_conv = _convert_to_target(activity_obt, activity_non_model_max, activity_target)
                model_lab_conv = _convert_to_target(model_lab_obt, model_lab_max, model_lab_target)
                practical_conv = _convert_to_target(practical_obt, practical_max, practical_target)

            elif course_type == "theory":
                theory_conv = _convert_to_target(theory_obt, theory_max, theory_target)
                activity_conv = _convert_to_target(activity_obt, activity_non_model_max, activity_target)

            elif course_type == "practical":
                # Practical marks convert to A1 Assignment
                practical_conv = _convert_to_target(practical_obt, practical_max, activity_target)
                # Model lab converts to A1 Test
                model_lab_conv = _convert_to_target(model_lab_obt, model_lab_max, model_lab_target)

            converted_total = theory_conv + activity_conv + model_lab_conv + practical_conv



            row.append(str(converted_total))

        data.append(row)

    total_width = 297 * mm - 30 * mm
    w_reg = 30 * mm
    w_name = 55 * mm

    col_widths = [w_reg, w_name]
    remaining = total_width - (w_reg + w_name)

    if len(header) > 2:
        each = remaining / (len(header) - 2)
        col_widths += [each] * (len(header) - 2)

    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ALIGN", (2, 1), (-1, -1), "CENTER"),
        ("ALIGN", (0, 1), (1, -1), "LEFT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    story.append(table)

    # ---------------- ASSESSMENT MAPPING ----------------
    if assessment_columns:
        story.append(Spacer(1, 10))

        legend_lines = ["<b>Assessment Mapping:</b>"]
        for a in assessment_columns:
            short_label = assessment_label_map[a["id"]]
            if str(a["id"]).startswith("IAT::"):
                if a.get("weightage"):
                    legend_lines.append(f"{a['name']} ({a['weightage']})")
                else:
                    legend_lines.append(f"{a['name']}")
            else:
                if a.get("weightage"):
                    legend_lines.append(f"{short_label} = {a['name']} ({a['weightage']})")
                else:
                    legend_lines.append(f"{short_label} = {a['name']}")

        if use_template_conversion and converted_total_max > 0:
            legend_lines.append("<br/><b>2021 Conversion:</b>")

            if course_type == "theory_lab":
                legend_lines.append(f"IAT/Test → {theory_target}")
                legend_lines.append(f"Assessment/Assignment → {activity_target}")
                legend_lines.append(f"Model Lab → {model_lab_target}")
                legend_lines.append(f"Practical → {practical_target}")

            elif course_type == "theory":
                legend_lines.append(f"IAT/Test → {theory_target}")
                legend_lines.append(f"Assessment/Assignment → {activity_target}")

            elif course_type == "practical":
                legend_lines.append(f"Model Lab → {model_lab_target}")
                legend_lines.append(f"Practical → {activity_target}")

            legend_lines.append(f"Final Total → {converted_total_max}")

        story.append(Paragraph("<br/>".join(legend_lines), style_small))

    # ---------------- HEADER / FOOTER ----------------
    def _find_logo():
        p = finders.find("images/ritlogo.png")
        return p if p and os.path.exists(p) else None

    logo_path = _find_logo()

    def _on_page(canvas, doc):
        canvas.saveState()
        w, h = landscape(A4)
        left = 15 * mm

        if logo_path:
            try:
                img = ImageReader(logo_path)
                canvas.drawImage(
                    img,
                    left,
                    h - 26 * mm,
                    height=18 * mm,
                    preserveAspectRatio=True
                )
            except Exception:
                pass

        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(w / 2, h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w / 2, h - 15 * mm, "Rajapalayam - 626117")
        canvas.drawCentredString(w / 2, h - 19 * mm, "Affiliated to Anna University, Chennai")
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(w / 2, h - 26 * mm, "Overall Consolidated Statement")
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w / 2, h - 31 * mm, f"{selection['degree_code']} — {selection['degree_name']}")
        canvas.drawCentredString(w / 2, h - 35 * mm, selection["department"])
        canvas.drawCentredString(w / 2, h - 39 * mm, f"{selection['course_code']} — {selection['course_title']}")
        canvas.line(left, h - 42 * mm, w - 15 * mm, h - 42 * mm)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(w - 15 * mm, 12 * mm, f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    filename = f'Consolidate_{selection["course_code"]}.pdf'
    response["Content-Disposition"] = f'inline; filename="{filename}"'
    return response








def consolidated_assessment_pdf(request):
    """
    Consolidated Assessment Report (Selected Assessments Only)
    DESIGN MATCHES overall_consolidate_pdf EXACTLY
    """

    import io, os, logging
    from decimal import Decimal, ROUND_HALF_UP
    from collections import defaultdict
    from django.http import HttpResponse
    from django.contrib.staticfiles import finders
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.utils import ImageReader

    logger = logging.getLogger(__name__)

    # ===================== REQUEST PARAMS =====================
    degree_id = request.GET.get("degree")
    department_id = request.GET.get("department")
    year = request.GET.get("year")
    semester = request.GET.get("semester")
    course_id = request.GET.get("course")
    assessment_names = request.GET.get("assessments", "").split(",")

    # ===================== QUERY =====================
    qs = ConsolidatedAssessmentResult.objects.select_related(
        "student", "course", "course__course"
    ).filter(
        student__department_id=department_id,
        assessment_name__in=assessment_names
    )

    if year:
        qs = qs.filter(student__year=year)
    if semester:
        qs = qs.filter(student__semester=semester)
    if course_id:
        qs = qs.filter(course_id=course_id)

    if not qs.exists():
        return HttpResponse("No data found")

    # ===================== META DETAILS =====================
    degree_obj = Degree.objects.filter(id=degree_id).first() if degree_id else None
    department_obj = Add_Department.objects.filter(id=department_id).first() if department_id else None
    course_obj = Course.objects.filter(id=course_id).first() if course_id else None

    degree_name = degree_obj.degree if degree_obj else "All Degrees"
    department_name = department_obj.Department if department_obj else "All Departments"
    course_display = (
        f"{course_obj.course_code} — {course_obj.title}"
        if course_obj else "All Courses"
    )

    # ===================== GROUP DATA =====================
    student_rows = defaultdict(dict)
    student_info = {}

    assessment_columns = sorted({
        f"{r.assessment_name} ({int(r.max_mark)})"
        if r.max_mark is not None else r.assessment_name
        for r in qs
    })

    for r in qs:
        sid = r.student_id
        student_info[sid] = r.student
        key = (
            f"{r.assessment_name} ({int(r.max_mark)})"
            if r.max_mark is not None else r.assessment_name
        )
        student_rows[sid][key] = r.actual_mark

    def fmt(val):
        if val is None:
            return "—"
        return str(Decimal(val).quantize(Decimal("1"), ROUND_HALF_UP))

    # ===================== BUILD TABLE ROWS (WITH DEBUG PRINTS) =====================
    rows_data = []

    # print("\n========== TOTAL CALCULATION DEBUG ==========")

    for sid, marks in student_rows.items():
        stu = student_info[sid]
        stu.regulation
        total = Decimal("0")

        # print(f"\nStudent: {stu.reg_no} | {stu.name} | Regulation: {stu.regulation}")
        # print("-" * 60)

        row = [stu.reg_no, stu.name]

        for col in assessment_columns:
            raw_mark = marks.get(col, 0)
            mark = Decimal(raw_mark or 0)

            # print(f"{col:<35} → {mark}")

            total += mark
            row.append(fmt(mark))

        # print("-" * 60)
        # print(f"{'TOTAL':<35} → {total}")
        # print("-" * 60)

        row.append(fmt(total))
        rows_data.append(row)

    # print("========== END DEBUG ==========\n")

    # ===================== PDF SETUP =====================
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=42 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    story = []

    # ===================== META BLOCK =====================
    style_small = ParagraphStyle(
        "small",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey
    )

    story.append(Paragraph(
        f"""
        <b>Degree:</b> {degree_name}<br/>
        <b>Department:</b> {department_name}<br/>
        <b>Course:</b> {course_display}
        """,
        style_small
    ))
    story.append(Spacer(1, 8))

    # ===================== TABLE =====================
    header = ["Reg No", "Name"] + assessment_columns + ["Total"]
    data = [header] + rows_data

    col_widths = [32 * mm, 48 * mm] + \
                 [30 * mm] * len(assessment_columns) + [26 * mm]

    table = Table(data, repeatRows=1, colWidths=col_widths)

    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (1, -1), "LEFT"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
            [colors.white, colors.Color(0.98, 0.98, 0.98)]
        ),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    story.append(table)

    # ===================== HEADER / FOOTER =====================
    def _find_logo():
        p = finders.find("images/ritlogo.png")
        return p if p and os.path.exists(p) else None

    logo_path = _find_logo()

    def _on_page(canvas, doc):
        canvas.saveState()
        w, h = landscape(A4)
        left = 15 * mm

        if logo_path:
            try:
                img = ImageReader(logo_path)
                canvas.drawImage(img, left, h - 26 * mm,
                                 height=18 * mm,
                                 preserveAspectRatio=True)
            except:
                pass

        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(w / 2, h - 10 * mm,
                                 "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w / 2, h - 15 * mm,
                                 "Rajapalayam - 626117")
        canvas.drawCentredString(w / 2, h - 19 * mm,
                                 "Affiliated to Anna University, Chennai")

        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(w / 2, h - 26 * mm,
                                 "Consolidated Assessment Statement")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w / 2, h - 31 * mm, degree_name)
        canvas.drawCentredString(w / 2, h - 35 * mm, department_name)
        canvas.drawCentredString(w / 2, h - 39 * mm, course_display)

        canvas.line(left, h - 42 * mm, w - 15 * mm, h - 42 * mm)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(w - 15 * mm, 12 * mm,
                               f"Page {canvas.getPageNumber()}")

        canvas.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="Consolidated_Assessment.pdf"'
    return response




from django.http import JsonResponse
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction

from collections import defaultdict
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from django.db.models import Q, Sum

@require_POST
@check_permission("overall_consolidate")
def save_overall_consolidate_record(request):
    user = getattr(request, "user", None)
    emp_id = getattr(user, "Employee_id", None) or getattr(user, "employee_id", None)

    current_faculty_id = ""
    if emp_id:
        gi = general_information.objects.filter(faculty_id=emp_id).only("faculty_id").first()
        current_faculty_id = str(gi.faculty_id) if gi and gi.faculty_id else str(emp_id)

    course_id = request.POST.get("course_id", "").strip()
    department_id = request.POST.get("department_id", "").strip()
    degree_id = request.POST.get("degree_id", "").strip()
    batch = request.POST.get("batch", "").strip()
    section = request.POST.get("section", "").strip()
    selected_iat = request.POST.get("iat", "").strip()

    assessment_ids = request.POST.getlist("assessment_ids")
    selected_iat_groups = request.POST.getlist("selected_iat_groups")
    custom_assessment_names = request.POST.getlist("custom_assessment_names")
    custom_assessment_row_ids = request.POST.getlist("custom_assessment_row_ids")
    standard_assessment_ids = request.POST.getlist("standard_assessment_ids")
    qp_assessment_ids = request.POST.getlist("qp_assessment_ids")
    assessment_iats = request.POST.getlist("assessment_iats")
    model_labs = request.POST.getlist("model_labs")
    include_practical = request.POST.get("include_practical", "").strip() == "1"

    if not course_id:
        return JsonResponse({
            "status": "error",
            "message": "Course is required."
        }, status=400)

    course_obj = Course.objects.filter(id=course_id).first()
    department_obj = Add_Department.objects.filter(id=department_id).first() if department_id else None
    degree_obj = Degree.objects.filter(id=degree_id).first() if degree_id else None

    if not course_obj:
        return JsonResponse({
            "status": "error",
            "message": "Invalid course."
        }, status=400)

    def unique_clean(values):
        out = []
        seen = set()
        for v in values:
            txt = str(v or "").strip()
            if txt and txt not in seen:
                seen.add(txt)
                out.append(txt)
        return out

    def _to_int(v, default=0):
        try:
            if v is None or str(v).strip() == "":
                return default
            return int(float(str(v).strip()))
        except Exception:
            return default

    def _norm_iat(v):
        return (v or "").strip().lower().replace(" ", "")

    # -------------------------------------------------
    # Fetch all students from enrollment
    # -------------------------------------------------
    enrollments = (
        CourseEnrollment.objects
        .select_related("student")
        .filter(
            course_id=course_id,
            batch=batch,
            section=section,
            enroll=True
        )
        .order_by("student__reg_no")
    )

    students = [e.student for e in enrollments if e.student]
    student_ids = [s.id for s in students if s]

    if not students:
        return JsonResponse({
            "status": "error",
            "message": "No enrolled students found for this batch and section."
        }, status=400)

    # -------------------------------------------------
    # Resolve assessment objects exactly like PDF logic
    # -------------------------------------------------
    assessment_objs = []

    for aid in assessment_ids:
        if not aid or not str(aid).strip():
            continue

        clean_id = str(aid).replace("am-", "").strip()

        am = Assessment_master.objects.filter(
            id=clean_id,
            course_id=course_id
        ).first()

        if not am:
            am = Assessment_master.objects.filter(
                assessment_id=clean_id,
                course_id=course_id
            ).first()

        if am:
            assessment_objs.append(am)

    model_lab_objs = []
    if model_labs:
        model_lab_objs = list(ModelLab.objects.filter(model_lab_name__in=model_labs))

    # -------------------------------------------------
    # Build marks map using same logic as PDF
    # marks_map[(student_id, col_id)] = obtained_mark
    # max_map[col_id] = max_mark
    # -------------------------------------------------
    marks_map = {}
    iat_group_entries = []
    iat_weightage_map = {}

    # ---------- IAT GROUP MARKS ----------
    if selected_iat_groups and student_ids:
        sim_qs = StudentInternalMark.objects.filter(
            student_id__in=student_ids,
            course_id=course_id,
            batch=batch,
            section=section,
        )

        if degree_id:
            sim_qs = sim_qs.filter(degree_id=degree_id)

        if department_id:
            sim_qs = sim_qs.filter(department_id=department_id)

        exam_q = Q()
        for grp in selected_iat_groups:
            grp = (grp or "").strip()
            if grp:
                exam_q |= Q(exam_name__iexact=grp)

        if exam_q:
            sim_qs = sim_qs.filter(exam_q)
        else:
            sim_qs = sim_qs.none()

        option_level_rows = (
            sim_qs.values(
                "student_id",
                "exam_name",
                "part_name",
                "question_number",
                "option_letter",
            )
            .annotate(
                option_obtained=Sum("marks_obtained"),
                option_max=Sum("max_marks"),
            )
            .order_by("student_id", "exam_name", "part_name", "question_number", "option_letter")
        )

        question_choice_map = defaultdict(lambda: {"chosen_obtained": 0, "chosen_max": 0})

        for row in option_level_rows:
            key = (
                row.get("student_id"),
                (row.get("exam_name") or "").strip(),
                row.get("part_name"),
                row.get("question_number"),
            )

            option_obtained = _to_int(row.get("option_obtained", 0))
            option_max = _to_int(row.get("option_max", 0))

            if option_max > question_choice_map[key]["chosen_max"]:
                question_choice_map[key]["chosen_max"] = option_max
                question_choice_map[key]["chosen_obtained"] = option_obtained
            elif (
                option_max == question_choice_map[key]["chosen_max"]
                and option_obtained > question_choice_map[key]["chosen_obtained"]
            ):
                question_choice_map[key]["chosen_obtained"] = option_obtained

        exam_totals = defaultdict(lambda: {"total_obtained": 0, "total_max": 0})

        for (student_id_val, exam_name, part_name, question_number), chosen in question_choice_map.items():
            if not exam_name or not student_id_val:
                continue

            key = (student_id_val, exam_name)
            exam_totals[key]["total_obtained"] += _to_int(chosen.get("chosen_obtained", 0))
            exam_totals[key]["total_max"] += _to_int(chosen.get("chosen_max", 0))

        for (student_id_val, exam_name), totals in exam_totals.items():
            col_id = f"IAT::{_norm_iat(exam_name)}"
            marks_map[(student_id_val, col_id)] = _to_int(totals.get("total_obtained", 0))

            if col_id not in iat_weightage_map:
                iat_weightage_map[col_id] = _to_int(totals.get("total_max", 0))

        for grp in unique_clean(selected_iat_groups):
            grp = (grp or "").strip()
            if not grp:
                continue

            col_id = f"IAT::{_norm_iat(grp)}"
            iat_group_entries.append({
                "id": col_id,
                "name": grp,
                "weightage": str(iat_weightage_map.get(col_id, 0) or "")
            })

    # ---------- MODEL LAB MARKS ----------
    MODEL_LAB_COL_ID = "MODEL_LAB"
    student_lab_totals = defaultdict(list)

    if model_lab_objs:
        ml_qs = ModelLabMarks.objects.filter(
            student_id__in=student_ids,
            model_lab__in=model_lab_objs,
            batch=batch,
            section=section,
        )

        iat_val = (selected_iat or "").strip().lower()
        if iat_val and iat_val not in ("both", "all"):
            ml_qs = ml_qs.filter(internal_assessment__iat__iexact=selected_iat)

        for ml in ml_qs:
            student_lab_totals[ml.student_id].append(_to_int(ml.total))

        for student_id_val, totals in student_lab_totals.items():
            avg = sum(totals) // len(totals) if totals else 0
            marks_map[(student_id_val, MODEL_LAB_COL_ID)] = avg

    # ---------- PRACTICAL MARKS ----------
    PRACTICAL_COL_ID = "PRACTICAL"
    student_practical_totals = defaultdict(list)

    if include_practical:
        exp_qs = experiment_marks.objects.filter(
            student_id__in=student_ids,
            courses__course_id=course_id,
        )

        iat_val = (selected_iat or "").strip().lower()
        if iat_val and iat_val not in ("both", "all"):
            exp_qs = exp_qs.filter(assessment__iat__iexact=selected_iat)

        for exp in exp_qs:
            student_practical_totals[exp.student_id].append(_to_int(exp.total))

        for student_id_val, totals in student_practical_totals.items():
            avg = sum(totals) // len(totals) if totals else 0
            marks_map[(student_id_val, PRACTICAL_COL_ID)] = avg

    # ---------- NORMAL ASSESSMENT MARKS ----------
    marks_qs = AssessmentMark.objects.filter(
        student_id__in=student_ids,
        assessment__in=assessment_objs
    ).select_related("assessment")

    for m in marks_qs:
        am_id = m.assessment_id
        value = (
            int(m.marks_weighted)
            if m.marks_weighted is not None
            else (m.marks_raw if m.marks_raw is not None else m.marks)
        )
        marks_map[(m.student_id, am_id)] = _to_int(value)

    # -------------------------------------------------
    # Split theory/activity/practical exactly like PDF
    # -------------------------------------------------
    theory_assessments = []
    activity_assessments = []
    practical_assessments = []

    selected_iat_lower = (selected_iat or "").strip().lower()

    for iat_entry in iat_group_entries:
        theory_assessments.append(iat_entry)

    for am in assessment_objs:
        name = (am.customAssessmentname or am.Assessmentname or "").strip()
        if not name:
            continue

        entry = {
            "id": am.id,
            "name": name,
            "weightage": str(am.weightage or "")
        }

        is_theory = False
        name_key = name.lower()

        if selected_iat_lower and selected_iat_lower not in ("both", "all"):
            compact_name = name_key.replace(" ", "")
            if selected_iat_lower.replace(" ", "") in compact_name:
                is_theory = True

        if "iat" in name_key:
            is_theory = True

        if is_theory:
            theory_assessments.append(entry)
        else:
            activity_assessments.append(entry)

    if model_lab_objs:
        activity_assessments.append({
            "id": MODEL_LAB_COL_ID,
            "name": "Model Lab",
            "weightage": "100"
        })

    if include_practical:
        practical_assessments.append({
            "id": PRACTICAL_COL_ID,
            "name": "Practical",
            "weightage": "75"
        })

    def unique_items(items):
        seen = set()
        out = []
        for item in items:
            if item["id"] not in seen:
                seen.add(item["id"])
                out.append(item)
        return out

    theory_assessments = unique_items(theory_assessments)
    activity_assessments = unique_items(activity_assessments)
    practical_assessments = unique_items(practical_assessments)

    # -------------------------------------------------
    # Static assessment strings and max strings
    # -------------------------------------------------
    theory_assessment = ", ".join(a["name"] for a in theory_assessments) if theory_assessments else None
    activity_assessment = ", ".join(a["name"] for a in activity_assessments) if activity_assessments else None
    practical_assessment = ", ".join(a["name"] for a in practical_assessments) if practical_assessments else None

    theory_max_mark = ", ".join(str(_to_int(a.get("weightage", 0))) for a in theory_assessments) if theory_assessments else None
    activity_max_mark = ", ".join(str(_to_int(a.get("weightage", 0))) for a in activity_assessments) if activity_assessments else None
    practical_max_mark = ", ".join(str(_to_int(a.get("weightage", 0))) for a in practical_assessments) if practical_assessments else None

    print("========== CONSOLIDATE STATIC VALUES ==========")
    print("Theory Assessments :", theory_assessment)
    print("Theory Max Marks   :", theory_max_mark)
    print("Activity Assessments :", activity_assessment)
    print("Activity Max Marks   :", activity_max_mark)
    print("Practical Assessments:", practical_assessment)
    print("Practical Max Marks  :", practical_max_mark)
    print("==============================================")

    # -------------------------------------------------
    # Save one row per student
    # actual mark strings must match assessment order
    # -------------------------------------------------
    created_ids = []

    with transaction.atomic():
        for student_obj in students:
            theory_actual_list = []
            activity_actual_list = []
            practical_actual_list = []

            for a in theory_assessments:
                theory_actual_list.append(str(_to_int(marks_map.get((student_obj.id, a["id"]), 0))))

            for a in activity_assessments:
                activity_actual_list.append(str(_to_int(marks_map.get((student_obj.id, a["id"]), 0))))

            for a in practical_assessments:
                practical_actual_list.append(str(_to_int(marks_map.get((student_obj.id, a["id"]), 0))))

            theory_actual_mark = ", ".join(theory_actual_list) if theory_actual_list else None
            activity_actual_mark = ", ".join(activity_actual_list) if activity_actual_list else None
            practical_actual_mark = ", ".join(practical_actual_list) if practical_actual_list else None

            print("------------------------------------------------")
            print(f"Student: {student_obj.reg_no} - {student_obj.name}")
            print("Theory Assessment :", theory_assessment)
            print("Theory Max        :", theory_max_mark)
            print("Theory Actual     :", theory_actual_mark)
            print("Activity Assessment :", activity_assessment)
            print("Activity Max        :", activity_max_mark)
            print("Activity Actual     :", activity_actual_mark)
            print("Practical Assessment:", practical_assessment)
            print("Practical Max       :", practical_max_mark)
            print("Practical Actual    :", practical_actual_mark)
            print("------------------------------------------------")

            record, created = OverallConsolidateRecord.objects.update_or_create(
            student=student_obj,
            course=course_obj,
            batch=batch or None,
            section=section or None,
            defaults={
                "faculty_id": current_faculty_id,
                "department": department_obj,
                "degree": degree_obj,

                "theory_assessment": theory_assessment,
                "activity_assessment": activity_assessment,
                "practical_assessment": practical_assessment,

                "theory_max_mark": theory_max_mark,
                "activity_max_mark": activity_max_mark,
                "practical_max_mark": practical_max_mark,

                "theory_actual_mark": theory_actual_mark,
                "activity_actual_mark": activity_actual_mark,
                "practical_actual_mark": practical_actual_mark,
            }
        )

        created_ids.append(record.id)

        print(
            f"{'CREATED' if created else 'UPDATED'} -> "
            f"Student: {student_obj.reg_no} | Record ID: {record.id}"
        )

    return JsonResponse({
        "status": "success",
        "message": f"Overall consolidate selection saved successfully for {len(created_ids)} student(s).",
        "record_ids": created_ids,
        "saved_data": {
            "theory_assessment": theory_assessment,
            "theory_max_mark": theory_max_mark,
            "activity_assessment": activity_assessment,
            "activity_max_mark": activity_max_mark,
            "practical_assessment": practical_assessment,
            "practical_max_mark": practical_max_mark,
            "student_count": len(created_ids),
        }
    })





from django.shortcuts import render, get_object_or_404
from faculty_management.models import general_information

def faculty_internal_timetable(request):
    emp_id = request.user.Employee_id

    faculty_info = get_object_or_404(
        general_information,
        faculty_id=emp_id
    )

    department = faculty_info.department
    degree = department.degree if department else None

    # 🔹 GET FILTERS
    batch = request.GET.get("batch")
    semester = request.GET.get("semester")
    iat_id = request.GET.get("iat")

    timetable = InternalTimeTable.objects.none()

    if batch and semester and iat_id:
        timetable = InternalTimeTable.objects.select_related(
            "course", "internal_assessment"
        ).filter(
            degree=degree,
            department=department,
            batch=batch,
            semester=semester,
            internal_assessment_id=iat_id
        ).order_by("exam_date", "session")

    return render(
        request,
        "faculty_management/exams/faculty_internal_timetable.html",
        {
            "faculty": faculty_info,
            "department": department,
            "degree": degree,
            "timetable": timetable,
        }
    )



def fit_batches(request):
    degree_id = request.GET.get("degree")
    department_id = request.GET.get("department")

    qs = InternalTimeTable.objects.filter(
        degree_id=degree_id,
        department_id=department_id
    ).values_list("batch", flat=True).distinct()

    items = [{"id": b, "text": b} for b in qs if b]
    return JsonResponse({"items": items})



def fit_semesters(request):
    degree_id = request.GET.get("degree")
    department_id = request.GET.get("department")
    batch = request.GET.get("batch")

    qs = InternalTimeTable.objects.filter(
        degree_id=degree_id,
        department_id=department_id,
        batch=batch
    ).values_list("semester", flat=True).distinct()

    items = [{"id": s, "text": f"Semester {s}"} for s in qs if s]
    return JsonResponse({"items": items})



def fit_iats(request):
    degree_id = request.GET.get("degree")
    department_id = request.GET.get("department")
    batch = request.GET.get("batch")
    semester = request.GET.get("semester")

    qs = InternalTimeTable.objects.filter(
        degree_id=degree_id,
        department_id=department_id,
        batch=batch,
        semester=semester
    ).select_related("internal_assessment").values(
        "internal_assessment__id",
        "internal_assessment__iat"
    ).distinct()

    items = [
        {"id": x["internal_assessment__id"], "text": x["internal_assessment__iat"]}
        for x in qs if x["internal_assessment__id"]
    ]


    return JsonResponse({"items": items})




from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
def faculty_internal_timetable_pdf(request):
    import io, os
    from django.http import HttpResponse
    from django.shortcuts import get_object_or_404
    from django.contrib.staticfiles import finders
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.utils import ImageReader

    emp_id = request.user.Employee_id

    faculty = get_object_or_404(
        general_information,
        faculty_id=emp_id
    )

    department = faculty.department
    degree = department.degree if department else None

    batch = request.GET.get("batch")
    semester = request.GET.get("semester")
    iat_id = request.GET.get("iat")

    qs = InternalTimeTable.objects.select_related(
        "course", "internal_assessment"
    ).filter(
        degree=degree,
        department=department,
        batch=batch,
        semester=semester,
        internal_assessment_id=iat_id
    ).order_by("exam_date", "session")

    if not qs.exists():
        return HttpResponse("No timetable found")

    # ✅ GET IAT NAME (IAT-1 / IAT-2 etc.)
    iat_name = qs.first().internal_assessment.iat

    # ===================== PDF SETUP =====================
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=42 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    story = []

    # ===================== META BLOCK =====================
    style_small = ParagraphStyle(
        "small",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey
    )

    story.append(Paragraph(
        f"""
        <b>Degree:</b> {degree}<br/>
        <b>Department:</b> {department}<br/>
        <b>Batch:</b> {batch} &nbsp;&nbsp; | &nbsp;&nbsp;
        <b>Semester:</b> {semester}<br/>
        <b>Internal Assessment:</b> {iat_name}
        """,
        style_small
    ))
    story.append(Spacer(1, 8))

    # ===================== TABLE =====================
    data = [["S.No", "Course Code", "Course Title", "Date", "Session"]]

    for idx, r in enumerate(qs, start=1):
        data.append([
            idx,
            r.course.course_code,
            r.course.title,
            r.exam_date.strftime("%d-%m-%Y"),
            r.session
        ])

    col_widths = [
        16 * mm,   # S.No
        36 * mm,   # Course Code
        90 * mm,   # Course Title
        32 * mm,   # Date
        26 * mm,   # Session
    ]

    table = Table(data, repeatRows=1, colWidths=col_widths)

    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),   # S.No
        ("ALIGN", (-2, 1), (-1, -1), "CENTER"), # Date & Session
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
            [colors.white, colors.Color(0.98, 0.98, 0.98)]
        ),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    story.append(table)

    # ===================== HEADER / FOOTER =====================
    def _find_logo():
        p = finders.find("images/ritlogo.png")
        return p if p and os.path.exists(p) else None

    logo_path = _find_logo()

    def _on_page(canvas, doc):
        canvas.saveState()
        w, h = landscape(A4)
        left = 15 * mm

        if logo_path:
            try:
                img = ImageReader(logo_path)
                canvas.drawImage(img, left, h - 26 * mm,
                                 height=18 * mm,
                                 preserveAspectRatio=True)
            except:
                pass

        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(w / 2, h - 10 * mm,
                                 "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w / 2, h - 15 * mm,
                                 "Rajapalayam - 626117")
        canvas.drawCentredString(w / 2, h - 19 * mm,
                                 "Affiliated to Anna University, Chennai")

        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(w / 2, h - 26 * mm,
                                 "Internal Exam Timetable")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w / 2, h - 31 * mm, str(degree))
        canvas.drawCentredString(w / 2, h - 35 * mm, str(department))
        canvas.drawCentredString(w / 2, h - 39 * mm, iat_name)

        canvas.line(left, h - 42 * mm, w - 15 * mm, h - 42 * mm)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(
            w - 15 * mm,
            12 * mm,
            f"Page {canvas.getPageNumber()}"
        )

        canvas.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="Internal_Exam_Timetable.pdf"'
    return response








def result_analysis(request):
    assigned_subjects = []

    faculty_id = getattr(request.user, "Employee_id", None)
    faculty = general_information.objects.filter(faculty_id=faculty_id).first()

    if faculty:
        assigned_subjects = (
            AssignSubjectFaculty.objects
            .filter(faculty=faculty)
            .select_related("course", "department__degree")
        )

        for subject in assigned_subjects:
            dept = getattr(subject, "department", None)

            if dept:
                subject.department_name = getattr(dept, "Department", "") or ""
                subject.department_code = getattr(dept, "Department_code", "") or ""

                deg = getattr(dept, "degree", None)
                subject.degree_id = deg.id if deg else None
                subject.degree_code = getattr(deg, "degree_code", "") or ""
                subject.degree_name = getattr(deg, "degree", "") or ""
            else:
                subject.department_name = ""
                subject.department_code = ""
                subject.degree_id = None
                subject.degree_code = ""
                subject.degree_name = ""

    context = {
        "assigned_subjects": assigned_subjects,
    }

    return render(request, "faculty_management/result_analysis.html", context)

import io, os
from collections import defaultdict
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.staticfiles import finders
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.utils import ImageReader
from django.db.models import Sum

def result_analysis_pdf(request):
    

    course_id = request.GET.get("course_id")
    batch = request.GET.get("batch")
    section = request.GET.get("section")
    department_id = request.GET.get("department_id")
    exam_name = request.GET.get("exam_name")
    year = request.GET.get("year")

    course = get_object_or_404(Course, id=course_id)
    department = get_object_or_404(Add_Department, id=department_id)

    marks_qs = StudentInternalMark.objects.filter(
        course_id=course_id,
        batch=batch,
        section=section,
        student__department_id=department_id
    ).select_related("student", "co_code")

    if exam_name:
        marks_qs = marks_qs.filter(exam_name=exam_name)

    if not marks_qs.exists():
        return HttpResponse("No internal marks found")

    first_mark = marks_qs.first()
    student = first_mark.student

    regulation = Regulations.objects.filter(year=student.regulation).first()
    if not regulation:
        return HttpResponse("Regulation not found")

    degree = Degree.objects.filter(is_active=True).first()

    pass_value_obj = PassValue.objects.filter(
        degree=degree,
        regulation=regulation
    ).first()

    if not pass_value_obj:
        return HttpResponse("Pass value not configured")

    pass_mark = pass_value_obj.iat_pass_value

    student_totals = (
        marks_qs.values(
            "student__reg_no",
            "student__name"
        )
        .annotate(total_marks=Sum("marks_obtained"))
        .order_by("student__reg_no")
    )

    total_students = student_totals.count()
    passed_students = student_totals.filter(total_marks__gte=pass_mark).count()
    failed_students = student_totals.filter(total_marks__lt=pass_mark).count()

    pass_percentage = round((passed_students / total_students) * 100, 2) if total_students else 0

    # ======================================================
    # BUILD STUDENT-WISE CO MAX AND CO OBTAINED
    # ======================================================

    all_rows = list(
        marks_qs.select_related("student", "co_code").order_by(
            "student__reg_no", "part_name", "question_number", "option_letter", "sub_question"
        )
    )

    student_row_map = defaultdict(list)
    for r in all_rows:
        student_row_map[r.student.reg_no].append(r)

    all_cos = set()
    student_co_summary = {}

    for reg_no, rows in student_row_map.items():

        question_map = defaultdict(
            lambda: defaultdict(lambda: {"max": 0, "obt": 0, "rows": []})
        )

        for m in rows:
            part = m.part_name or ""
            qnum = m.question_number or ""
            opt = m.option_letter or ""

            question_map[(part, qnum)][opt]["max"] += int(m.max_marks or 0)
            question_map[(part, qnum)][opt]["obt"] += int(m.marks_obtained or 0)
            question_map[(part, qnum)][opt]["rows"].append(m)

        chosen_rows = []

        for (_part, _qnum), options in question_map.items():
            if "" in options and len(options) == 1:
                chosen_rows.extend(options[""]["rows"])
            else:
                chosen_key = max(options.keys(), key=lambda k: options[k]["obt"])
                chosen_rows.extend(options[chosen_key]["rows"])

        co_summary = defaultdict(lambda: {"max": 0, "obt": 0})

        for r in chosen_rows:
            if r.co_code_id and r.co_code and r.co_code.co_code:
                co = r.co_code.co_code
                co_summary[co]["max"] += int(r.max_marks or 0)
                co_summary[co]["obt"] += int(r.marks_obtained or 0)
                all_cos.add(co)

        student_co_summary[reg_no] = co_summary

    co_list = sorted(all_cos)

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=42 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    story = []

    style_small = ParagraphStyle(
        "small",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey
    )

    display_exam_name = exam_name or first_mark.exam_name or "-"
    display_year = first_mark.created_at.year if first_mark.created_at else "-"

    story.append(Paragraph(
        f"""
        <b>Course:</b> {course.course_code} - {course.title}<br/>
        <b>Degree:</b> {degree}<br/>
        <b>Department:</b> {department}<br/>
        <b>Batch:</b> {batch} &nbsp;&nbsp; | &nbsp;&nbsp;
        <b>Section:</b> {section}<br/>
        <b>Regulation:</b> {regulation}<br/>
        <b>Exam Name:</b> {display_exam_name}<br/>
        <b>Pass Mark:</b> {pass_mark}
        """,
        style_small
    ))

    story.append(Spacer(1, 12))

    headers = ["S.No", "Register No", "Student Name"] + co_list + ["Total Marks", "Result"]
    data = [headers]

    for idx, s in enumerate(student_totals, start=1):
        reg_no = s["student__reg_no"]
        name = s["student__name"]

        row = [idx, reg_no, name]
        total = 0

        co_summary = student_co_summary.get(reg_no, {})

        for co_code in co_list:
            co_obt = co_summary.get(co_code, {}).get("obt", 0)
            co_max = co_summary.get(co_code, {}).get("max", 0)

            row.append(f"{co_obt}/{co_max}")
            total += co_obt

        result = "PASS" if total >= pass_mark else "FAIL"

        row.extend([total, result])
        data.append(row)

    col_widths = [20*mm, 45*mm, 70*mm] + [25*mm]*len(co_list) + [30*mm, 30*mm]

    student_table = Table(
        data,
        repeatRows=1,
        colWidths=col_widths
    )

    student_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (-2, 1), (-1, -1), "CENTER"),
        ("ALIGN", (3, 1), (-3, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
            [colors.white, colors.Color(0.97, 0.97, 0.97)]
        ),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))

    story.append(student_table)
    story.append(Spacer(1, 20))

    summary_data = [
        ["Total Students", "Passed Students", "Failed Students", "Pass Mark", "Pass %"],
        [total_students, passed_students, failed_students, pass_mark, f"{pass_percentage}%"]
    ]

    summary_table = Table(summary_data, colWidths=[50*mm, 50*mm, 50*mm, 40*mm, 30*mm])

    summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
    ]))

    story.append(summary_table)
    story.append(Spacer(1, 15))

    university_pass_mark = pass_value_obj.university_iat_pass_value

    university_passed_students = student_totals.filter(
        total_marks__gte=university_pass_mark
    ).count()

    university_failed_students = student_totals.filter(
        total_marks__lt=university_pass_mark
    ).count()

    university_pass_percentage = round(
        (university_passed_students / total_students) * 100, 2
    ) if total_students else 0

    university_summary_data = [
        ["Total Students", "Passed Students", "Failed Students", "University Pass Mark", "University Pass %"],
        [total_students, university_passed_students, university_failed_students,
         university_pass_mark, f"{university_pass_percentage}%"]
    ]

    university_summary_table = Table(
        university_summary_data,
        colWidths=[45*mm, 45*mm, 45*mm, 55*mm, 40*mm]
    )

    university_summary_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
    ]))

    story.append(university_summary_table)

    def _find_logo():
        p = finders.find("images/ritlogo.png")
        return p if p and os.path.exists(p) else None

    logo_path = _find_logo()

    def _on_page(canvas, doc):
        canvas.saveState()

        w, h = landscape(A4)
        left = 15 * mm

        if logo_path:
            try:
                img = ImageReader(logo_path)
                canvas.drawImage(
                    img,
                    left,
                    h - 26 * mm,
                    height=18 * mm,
                    preserveAspectRatio=True
                )
            except:
                pass

        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(w/2, h-10*mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w/2, h-15*mm, "Rajapalayam - 626117")
        canvas.drawCentredString(w/2, h-19*mm, "Affiliated to Anna University, Chennai")

        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(
            w/2,
            h-26*mm,
            f"Result Analysis - {display_exam_name} ({display_year})"
        )

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w/2, h-31*mm, str(degree))
        canvas.drawCentredString(w/2, h-35*mm, str(department))
        canvas.drawCentredString(w/2, h-39*mm, course.course_code)

        canvas.line(left, h-42*mm, w-15*mm, h-42*mm)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(
            w - 15*mm,
            12*mm,
            f"Page {canvas.getPageNumber()}"
        )

        canvas.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="Result_Analysis_{course.course_code}.pdf"'

    return response




from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Q

def retest_internalmark(request):
    faculty_id = request.user.Employee_id
    faculty = general_information.objects.filter(faculty_id=faculty_id).first()

    if not faculty:
        return render(request, "faculty_management/retestinternalmark.html", {
            "assigned_subjects": [],
            "students": [],
        })

    marks_qs = StudentInternalMark.objects.filter(
        faculty_assignment__faculty=faculty,
        absentee=1
    ).select_related(
        "student",
        "faculty_assignment__course",
        "faculty_assignment__department__degree",
        "faculty_assignment__department",
        "course",
        "department",
        "degree",
    )

    assigned_subjects = []
    seen = set()

    for mark in marks_qs:
        subject = mark.faculty_assignment
        if subject and subject.id not in seen:
            seen.add(subject.id)
            assigned_subjects.append(subject)

    for subject in assigned_subjects:
        dept = getattr(subject, "department", None)

        if dept:
            subject.department_name = getattr(dept, "Department", "") or ""
            subject.department_code = getattr(dept, "Department_code", "") or ""

            deg = getattr(dept, "degree", None)
            subject.degree_id = deg.id if deg else None
            subject.degree_code = getattr(deg, "degree_code", "") or ""
            subject.degree_name = getattr(deg, "degree", "") or ""
        else:
            subject.department_name = ""
            subject.department_code = ""
            subject.degree_id = None
            subject.degree_code = ""
            subject.degree_name = ""

        subject_exam_qs = StudentInternalMark.objects.filter(
            faculty_assignment=subject,
            absentee=1
        ).exclude(
            exam_name__isnull=True
        ).exclude(
            exam_name__exact=""
        ).values_list("exam_name", flat=True).distinct()

        subject.exam_names = list(subject_exam_qs)

    students = []
    selected_course_id = ""
    selected_course_code = ""
    selected_course_title = ""
    selected_batch = ""
    selected_section = ""
    selected_department = ""
    selected_department_id = ""
    selected_degree_code = ""
    selected_degree_name = ""
    selected_exam_name = ""

    if request.method == "POST":
        course_id = request.POST.get("course_id")
        exam_name = request.POST.get("exam_name")
        batch = request.POST.get("batch")
        section = request.POST.get("section")
        department_id = request.POST.get("department_id")

        student_marks = StudentInternalMark.objects.filter(
            faculty_assignment__faculty=faculty,
            absentee=1,
            course_id=course_id,
            exam_name=exam_name,
        ).select_related("student", "department", "degree", "course")

        if batch:
            student_marks = student_marks.filter(batch=batch)

        if section:
            student_marks = student_marks.filter(section=section)

        if department_id:
            student_marks = student_marks.filter(department_id=department_id)

        unique_students = {}
        for row in student_marks:
            stu = row.student
            if not stu:
                continue

            if stu.id not in unique_students:
                dept = row.department
                deg = row.degree

                unique_students[stu.id] = {
                    "id": stu.id,
                    "reg_no": getattr(stu, "reg_no", row.reg_no) or row.reg_no or "",
                    "name": getattr(stu, "student_name", "") or getattr(stu, "name", "") or str(stu),
                    "department_name": getattr(dept, "Department", "") if dept else "",
                    "department_code": getattr(dept, "Department_code", "") if dept else "",
                    "degree_id": deg.id if deg else "",
                    "degree_code": getattr(deg, "degree_code", "") if deg else "",
                    "degree_name": getattr(deg, "degree", "") if deg else "",
                }

        students = list(unique_students.values())

        first_row = student_marks.first()
        selected_course = first_row.course if first_row else None

        selected_course_id = course_id or ""
        selected_course_code = getattr(selected_course, "course_code", "") if selected_course else request.POST.get("course_code", "")
        selected_course_title = getattr(selected_course, "title", "") if selected_course else request.POST.get("course_title", "")
        selected_batch = batch or ""
        selected_section = section or ""
        selected_department_id = department_id or ""
        selected_department = request.POST.get("department", "")
        selected_degree_code = request.POST.get("degree_code", "")
        selected_degree_name = request.POST.get("degree_name", "")
        selected_exam_name = exam_name or ""

    context = {
        "assigned_subjects": assigned_subjects,
        "students": students,
        "selected_course_id": selected_course_id,
        "selected_course_code": selected_course_code,
        "selected_course_title": selected_course_title,
        "selected_batch": selected_batch,
        "selected_section": selected_section,
        "selected_department": selected_department,
        "selected_department_id": selected_department_id,
        "selected_degree_code": selected_degree_code,
        "selected_degree_name": selected_degree_name,
        "selected_exam_name": selected_exam_name,
    }

    return render(request, "faculty_management/retestinternalmark.html", context)







from collections import Counter, defaultdict
from django.contrib import messages
from django.db.models import Prefetch
from django.shortcuts import render


def retest_enter_mark_page(request):
    reg_no = request.GET.get("reg_no")
    student_name = request.GET.get("student_name")
    course_id = request.GET.get("course_id")
    course_code = request.GET.get("course_code")
    course_title = request.GET.get("course_title")
    batch = request.GET.get("batch")
    section = request.GET.get("section")
    exam_name = request.GET.get("exam_name")
    department_name = request.GET.get("department_name")
    department_code = request.GET.get("department_code")
    department_id = request.GET.get("department_id")
    degree_id_qs = request.GET.get("degree_id")
    degree_code = request.GET.get("degree_code")
    degree_name = request.GET.get("degree_name")
    pattern_id = request.GET.get("pattern_id")

    student = StudentDetails.objects.filter(reg_no=reg_no).first()

    department = None
    if department_code:
        department = Add_Department.objects.filter(Department_code=department_code).first()
    elif department_id:
        department = Add_Department.objects.filter(id=department_id).first()

    enrollment = None
    if student:
        enrollment_qs = CourseEnrollment.objects.filter(
            student_id=student.id,
            course__course_code=course_code,
        )

        if department:
            enrollment_qs = enrollment_qs.filter(department=department)
        elif department_id:
            enrollment_qs = enrollment_qs.filter(department_id=department_id)

        if batch:
            enrollment_qs = enrollment_qs.filter(batch=batch)
        if section:
            enrollment_qs = enrollment_qs.filter(section=section)

        enrollment = enrollment_qs.order_by("-enroll", "-id").first()

    if not student:
        messages.warning(request, "Student not found.")
    if not enrollment:
        messages.warning(request, "Enrollment not found for the selected student/course.")

    degree_obj = _resolve_degree(student, enrollment) if student else None
    degree_id, degree_label = _degree_id_and_label(degree_obj) if degree_obj else (None, "")

    year_val = (
        _safe_get(enrollment, "year")
        or _safe_get(student, "current_year")
        or _safe_get(student, "year")
    )
    semester_val = (
        _safe_get(enrollment, "semester")
        or _safe_get(student, "current_semester")
        or _safe_get(student, "semester")
    )

    year_val = str(year_val) if year_val is not None else None
    semester_val = str(semester_val) if semester_val is not None else None

    regulation_obj = _resolve_regulation_obj(student, enrollment) if student else None
    regulation_display = getattr(regulation_obj, "year", None) or _safe_get(student, "regulation")

    pattern_options = ExamPattern.objects.none()
    if regulation_obj and degree_id and year_val and semester_val and exam_name:
        pattern_options = (
            ExamPattern.objects.select_related("regulation", "degree")
            .filter(
                regulation=regulation_obj,
                degree_id=degree_id,
                year=year_val,
                semester=semester_val,
                for_exam=str(exam_name).strip(),
            )
            .order_by("pattern", "id")
        )

    if (not pattern_id) and enrollment and exam_name:
        last_row = (
            StudentInternalMark.objects.filter(
                student=student,
                enrollment=enrollment,
                course__course_code=course_code,
                exam_name=exam_name,
                pattern__isnull=False,
            )
            .select_related("pattern")
            .order_by("-created_at", "-id")
            .first()
        )
        if last_row and last_row.pattern_id:
            pattern_id = str(last_row.pattern_id)

    exam_pattern_obj = None
    selected_pid = None

    if pattern_id:
        try:
            selected_pid = int(pattern_id)
        except (TypeError, ValueError):
            selected_pid = None

    if selected_pid:
        try:
            exam_pattern_obj = (
                ExamPattern.objects.select_related("regulation", "degree")
                .prefetch_related(
                    Prefetch(
                        "parts",
                        queryset=Part.objects.prefetch_related(
                            Prefetch(
                                "questions",
                                queryset=Question.objects.prefetch_related("options").order_by("number"),
                                to_attr="prefetched_questions",
                            )
                        ).order_by("name"),
                        to_attr="prefetched_parts",
                    )
                )
                .get(id=selected_pid)
            )
        except ExamPattern.DoesNotExist:
            exam_pattern_obj = None

    pattern_selected = bool(exam_pattern_obj)

    prefilled_marks = {}
    prefilled_co = {}
    prefilled_blooms = {}
    prefilled_subpart_max = {}
    prefilled_existing_subparts = defaultdict(list)
    part_marks_summary = {}

    if student and enrollment and exam_name and exam_pattern_obj:
        student_marks_qs = (
            StudentInternalMark.objects.select_related("co_code", "level_code")
            .filter(
                student=student,
                enrollment=enrollment,
                exam_name=exam_name,
                pattern=exam_pattern_obj,
            )
            .order_by("id")
        )

        if course_id:
            student_marks_qs = student_marks_qs.filter(course_id=course_id)
        elif course_code:
            student_marks_qs = student_marks_qs.filter(course__course_code=course_code)

        for row in student_marks_qs:
            part = (row.part_name or "").strip().upper()
            qnum = str(row.question_number or "").strip()
            sub = (row.sub_question or "").strip().lower()
            opt = (row.option_letter or "").strip().lower()

            if sub and opt:
                suffix = f"{part}_{qnum}_{sub}_{opt}"
            elif sub:
                suffix = f"{part}_{qnum}_{sub}"
            elif opt:
                suffix = f"{part}_{qnum}_{opt}"
            else:
                suffix = f"{part}_{qnum}"

            prefilled_marks[f"marks_{suffix}"] = row.marks_obtained

            if row.co_code_id:
                prefilled_co[f"co_{suffix}"] = str(row.co_code_id)

            if row.level_code_id:
                prefilled_blooms[f"blooms_{suffix}"] = str(row.level_code_id)

            if sub and opt:
                group_key = f"{part}__{qnum}__{opt}"
                sp_key = f"{group_key}__{sub}"

                if sub not in prefilled_existing_subparts[group_key]:
                    prefilled_existing_subparts[group_key].append(sub)

                if sp_key not in prefilled_subpart_max and row.max_marks is not None:
                    prefilled_subpart_max[sp_key] = row.max_marks

    roman_order = ["i", "ii", "iii", "iv", "v", "vi", "vii", "viii", "ix", "x"]

    def sort_subparts(subparts):
        return sorted(
            subparts,
            key=lambda x: roman_order.index(x) if x in roman_order else 999
        )

    prefilled_existing_subparts = {
        key: sort_subparts(value)
        for key, value in prefilled_existing_subparts.items()
    }

    if exam_pattern_obj:
        parts_list = getattr(exam_pattern_obj, "prefetched_parts", exam_pattern_obj.parts.all())

        for part in parts_list:
            if str(part.name).lower() == "a":
                if part.total_questions and part.max_marks:
                    part_marks_summary[part.name] = f"{part.total_questions} × {part.max_marks} Marks"
                else:
                    part_marks_summary[part.name] = f"{part.max_marks} Marks"
            else:
                questions_list = getattr(part, "prefetched_questions", part.questions.all())

                question_marks = []
                for question in questions_list:
                    q_mark = question.total_marks or part.max_marks or 0
                    if q_mark > 0:
                        question_marks.append(q_mark)

                if question_marks:
                    counts = Counter(question_marks)
                    summary_parts = []
                    for mark, count in sorted(counts.items(), reverse=True):
                        summary_parts.append(f"{count} × {mark} Marks")
                    part_marks_summary[part.name] = ", ".join(summary_parts)
                else:
                    part_marks_summary[part.name] = f"{part.max_marks} Marks"

    course_outcomes = CourseOutcome.objects.none()
    blooms_levels = BloomsLevel.objects.all().order_by("level_code")

    if exam_pattern_obj and exam_pattern_obj.regulation:
        course_outcomes = CourseOutcome.objects.filter(
            regulation=exam_pattern_obj.regulation.year
        ).order_by("co_code")

    context = {
        "bulk": False,
        "students_list": [student] if student else [],
        "reg_no": reg_no,
        "student_name": student_name or getattr(student, "name", "") or getattr(student, "student_name", ""),
        "course_id": course_id,
        "course_code": course_code,
        "course_title": course_title,
        "batch": batch,
        "section": section,
        "iat": exam_name,
        "exam_name": exam_name,
        "department_name": department_name,
        "department_code": department_code,
        "department_id": department_id,
        "degree_id": degree_id_qs or degree_id,
        "degree_code": degree_code,
        "degree_name": degree_name,
        "pattern_options": pattern_options,
        "selected_pattern_id": selected_pid,
        "pattern_selected": pattern_selected,
        "selected_pattern": getattr(exam_pattern_obj, "pattern", None),
        "selected_regulation": getattr(regulation_obj, "year", None) or regulation_display,
        "selected_degree": degree_label,
        "selected_year": year_val,
        "selected_semester": semester_val,
        "exam_pattern_obj": exam_pattern_obj,
        "exam_pattern_id": getattr(exam_pattern_obj, "id", None),
        "prefilled_marks": prefilled_marks,
        "prefilled_co": prefilled_co,
        "prefilled_blooms": prefilled_blooms,
        "prefilled_subpart_max": prefilled_subpart_max,
        "prefilled_existing_subparts": prefilled_existing_subparts,
        "part_marks_summary": part_marks_summary,
        "course_outcomes": course_outcomes,
        "blooms_levels": blooms_levels,
    }

    return render(request, "faculty_management/retest_enter_mark_page.html", context)







from datetime import date
from django.contrib import messages
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect


def get_academic_year():
    """
    Dynamically returns academic year string.
    Example:
      If current month >= June -> '2025-2026'
      Else (Jan-May) -> '2024-2025'
    """
    today = date.today()
    current_year = today.year
    if today.month >= 6:
        return f"{current_year}-{current_year + 1}"
    else:
        return f"{current_year - 1}-{current_year}"


@transaction.atomic
def save_retest_marks(request):
    if request.method != "POST":
        messages.error(request, "Invalid request method!")
        return redirect("/")

    reg_no = request.POST.get("reg_no")
    if not reg_no:
        student_list = request.POST.getlist("students[]")
        reg_no = student_list[0] if student_list else None

    department_code = request.POST.get("department_code")
    department_name = request.POST.get("department_name")
    course_code = request.POST.get("course_code")
    course_title = request.POST.get("course_title")
    batch = request.POST.get("batch")
    section = request.POST.get("section")

    exam_name = (
        request.POST.get("iat")
        or request.POST.get("exam_name")
        or request.GET.get("iat")
    )

    academic_year = get_academic_year()

    pattern_id_str = request.POST.get("pattern_id") or request.POST.get("exam_pattern_id")
    pattern_id = int(pattern_id_str) if (pattern_id_str and pattern_id_str.isdigit()) else None

    exam_pattern = None
    if pattern_id:
        exam_pattern = get_object_or_404(ExamPattern, id=pattern_id)

    if not reg_no:
        messages.error(request, "No student reg_no provided.")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    co_map = {}
    blooms_map = {}

    for key in request.POST.keys():
        if key.startswith("co_map_"):
            suffix = key.replace("co_map_", "").replace("[]", "")
            values = request.POST.getlist(key)
            co_map[suffix] = [int(v) for v in values if str(v).isdigit()]

        if key.startswith("blooms_map_"):
            suffix = key.replace("blooms_map_", "").replace("[]", "")
            values = request.POST.getlist(key)
            blooms_map[suffix] = [int(v) for v in values if str(v).isdigit()]

    global_subpart_max = {}

    for key, value in request.POST.items():
        if key.startswith("subpart_max__"):
            parts = key.split("__")
            if len(parts) == 5:
                _, part_name, question_number, option_letter, sub_question = parts
                normalized_key = f"max_{part_name}_{question_number}_{sub_question}_{option_letter}"
                global_subpart_max[normalized_key] = value

    def safe_int(value, default=None):
        try:
            value = str(value).strip()
            if value == "":
                return default
            return int(float(value))
        except Exception:
            return default

    def parse_marks_key(key: str):
        parts = key.split("_")
        if len(parts) < 3 or parts[0] != "marks":
            return None

        return {
            "part_name": parts[1],
            "question_number": parts[2],
            "sub_question": parts[3] if len(parts) > 3 else None,
            "option_letter": parts[4] if len(parts) > 4 else None,
        }

    def build_max_key(part_name, question_number, sub_question=None, option_letter=None):
        if sub_question and option_letter:
            return f"max_{part_name}_{question_number}_{sub_question}_{option_letter}"
        if sub_question:
            return f"max_{part_name}_{question_number}_{sub_question}"
        if option_letter:
            return f"max_{part_name}_{question_number}_{option_letter}"
        return f"max_{part_name}_{question_number}"

    def build_suffix(part_name, question_number, sub_question=None, option_letter=None):
        if sub_question and option_letter:
            return f"{part_name}_{question_number}_{sub_question}_{option_letter}"
        if sub_question:
            return f"{part_name}_{question_number}_{sub_question}"
        if option_letter:
            return f"{part_name}_{question_number}_{option_letter}"
        return f"{part_name}_{question_number}"

    def get_part_default_max(part_name):
        if not exam_pattern:
            return 0
        part_obj = exam_pattern.parts.filter(name=part_name).first()
        if not part_obj:
            return 0
        return safe_int(getattr(part_obj, "max_marks", 0), 0)

    def resolve_max_marks(part_name, question_number, sub_question=None, option_letter=None, max_data=None):
        max_data = max_data or {}

        if sub_question and option_letter:
            k = build_max_key(part_name, question_number, sub_question, option_letter)
            val = safe_int(global_subpart_max.get(k), 0)
            if val > 0:
                return val

        exact_key = build_max_key(part_name, question_number, sub_question, option_letter)
        val = safe_int(max_data.get(exact_key), 0)
        if val > 0:
            return val

        if option_letter:
            option_key = build_max_key(part_name, question_number, option_letter=option_letter)
            val = safe_int(max_data.get(option_key), 0)
            if val > 0:
                return val

        if sub_question:
            sub_key = build_max_key(part_name, question_number, sub_question=sub_question)
            val = safe_int(max_data.get(sub_key), 0)
            if val > 0:
                return val

        q_key = build_max_key(part_name, question_number)
        val = safe_int(max_data.get(q_key), 0)
        if val > 0:
            return val

        return get_part_default_max(part_name)

    try:
        student = StudentDetails.objects.filter(reg_no=reg_no).first()
        if not student:
            raise ValueError(f"Student not found: {reg_no}")

        student_semester = getattr(student, "semester", None)
        student_department = getattr(student, "department", None)
        student_degree = getattr(student_department, "degree", None) if student_department else None

        enrollment_qs = CourseEnrollment.objects.filter(
            student_id=student.id,
            course__course_code=course_code,
        )

        if batch:
            enrollment_qs = enrollment_qs.filter(batch=batch)

        if section:
            enrollment_qs = enrollment_qs.filter(section=section)

        enrollment = enrollment_qs.order_by("-enroll", "-id").first()
        if not enrollment:
            raise ValueError(f"Course enrollment not found for {reg_no}")

        faculty_assignment = AssignSubjectFaculty.objects.filter(
            course=enrollment.course,
            batch=batch or enrollment.batch,
            section=section or enrollment.section,
            academic_year=academic_year,
            is_active=True
        ).first()

        existing_absent_row = (
            StudentInternalMark.objects.filter(
                student=student,
                enrollment=enrollment,
                exam_name=exam_name,
                course=enrollment.course,
            )
            .order_by("-id")
            .first()
        )
        existing_absentee = existing_absent_row.absentee if existing_absent_row else 0

        bulk_prefix = f"m__{reg_no}__"
        marks_data = {
            k.replace(bulk_prefix, ""): v
            for k, v in request.POST.items()
            if k.startswith(bulk_prefix + "marks_")
        }
        max_data = {
            k.replace(bulk_prefix, ""): v
            for k, v in request.POST.items()
            if k.startswith(bulk_prefix + "max_")
        }

        grouped = {}
        for key, value in marks_data.items():
            parsed = parse_marks_key(key)
            if not parsed:
                continue

            group_key = (parsed["part_name"], parsed["question_number"])
            grouped.setdefault(group_key, []).append(
                {
                    "raw_key": key,
                    "raw_value": value,
                    "part_name": parsed["part_name"],
                    "question_number": parsed["question_number"],
                    "sub_question": parsed["sub_question"],
                    "option_letter": parsed["option_letter"],
                }
            )

        if not grouped:
            StudentInternalMark.objects.update_or_create(
                student=student,
                enrollment=enrollment,
                exam_name=exam_name,
                course=enrollment.course,
                part_name=None,
                question_number=None,
                sub_question=None,
                option_letter=None,
                defaults={
                    "pattern": exam_pattern,
                    "semester": student_semester,
                    "degree": student_degree,
                    "department": student_department,
                    "max_marks": None,
                    "marks_obtained": None,
                    "co_code_id": None,
                    "level_code_id": None,
                    "reg_no": reg_no,
                    "course_code": course_code,
                    "absentee": existing_absentee,
                    "retest_attempted": 1,
                    "batch": batch or enrollment.batch,
                    "section": section or enrollment.section,
                    "academic_year": academic_year,
                    "faculty_assignment": faculty_assignment,
                },
            )
            print("faculty_assignment:", faculty_assignment)
            messages.success(request, f"Retest saved successfully for {reg_no}.")
            return redirect(request.META.get("HTTP_REFERER", "/"))

        submitted_subparts = {}
        for (part_name, question_number), entries in grouped.items():
            for entry in entries:
                option_letter = entry["option_letter"]
                sub_question = entry["sub_question"]

                if option_letter and sub_question:
                    k = (part_name, question_number, option_letter)
                    submitted_subparts.setdefault(k, set()).add(sub_question)

        for (part_name, question_number, option_letter), submitted_set in submitted_subparts.items():
            StudentInternalMark.objects.filter(
                student=student,
                enrollment=enrollment,
                exam_name=exam_name,
                course=enrollment.course,
                part_name=part_name,
                question_number=question_number,
                option_letter=option_letter,
            ).exclude(sub_question__in=list(submitted_set)).exclude(sub_question__in=["i", "ii"]).delete()

        upsert_count = 0

        for (part_name, question_number), entries in grouped.items():
            selected_option = None
            pair_totals = {}

            for entry in entries:
                sub_question = entry["sub_question"]
                option_letter = entry["option_letter"]
                marks_obt = safe_int(entry["raw_value"], None)

                if option_letter and marks_obt is not None and marks_obt > 0:
                    selected_option = option_letter

                if sub_question and option_letter and marks_obt is not None:
                    pair_totals.setdefault(option_letter, 0)
                    pair_totals[option_letter] += marks_obt

            non_zero_options = set()
            for entry in entries:
                option_letter = entry["option_letter"]
                marks_obt = safe_int(entry["raw_value"], None)
                if option_letter and marks_obt is not None and marks_obt > 0:
                    non_zero_options.add(option_letter)

            if len(non_zero_options) > 1:
                raise ValueError(
                    f"{reg_no} - {part_name}{question_number}: Enter marks for only one option."
                )

            for option_letter, total in pair_totals.items():
                allowed_max = 0

                for entry in entries:
                    if entry["option_letter"] != option_letter:
                        continue

                    sub_question = entry["sub_question"]
                    if not sub_question:
                        continue

                    suffix = build_suffix(
                        part_name=part_name,
                        question_number=question_number,
                        sub_question=sub_question,
                        option_letter=option_letter,
                    )
                    co_values = co_map.get(suffix, [])
                    co_id = co_values[0] if co_values else None
                    if co_id is None:
                        continue

                    resolved = resolve_max_marks(
                        part_name=part_name,
                        question_number=question_number,
                        sub_question=sub_question,
                        option_letter=option_letter,
                        max_data=max_data,
                    )
                    allowed_max += resolved

                if allowed_max > 0 and total > allowed_max:
                    raise ValueError(
                        f"{reg_no} - {part_name}{question_number}{option_letter}: "
                        f"Total of all sub-questions cannot exceed {allowed_max}. "
                        f"You entered total {total}."
                    )

            for entry in entries:
                sub_question = entry["sub_question"]
                option_letter = entry["option_letter"]
                marks_obt = safe_int(entry["raw_value"], None)

                if selected_option and option_letter and option_letter != selected_option:
                    marks_obt = None

                suffix = build_suffix(
                    part_name=part_name,
                    question_number=question_number,
                    sub_question=sub_question,
                    option_letter=option_letter,
                )

                co_values = co_map.get(suffix, [])
                bl_values = blooms_map.get(suffix, [])

                co_id = co_values[0] if co_values else None
                bl_id = bl_values[0] if bl_values else None

                max_marks = resolve_max_marks(
                    part_name=part_name,
                    question_number=question_number,
                    sub_question=sub_question,
                    option_letter=option_letter,
                    max_data=max_data,
                )

                if co_id is None:
                    bl_id = None
                    max_marks_to_save = None
                    marks_obt_to_save = None
                else:
                    max_marks_to_save = max_marks if max_marks > 0 else None
                    marks_obt_to_save = marks_obt if marks_obt is not None else None

                if (
                    marks_obt_to_save is not None
                    and max_marks_to_save is not None
                    and marks_obt_to_save > max_marks_to_save
                ):
                    raise ValueError(
                        f"{reg_no} - {part_name}{question_number}"
                        f"{option_letter or ''}{('-' + sub_question.upper()) if sub_question else ''}: "
                        f"Marks cannot exceed {max_marks_to_save}."
                    )

                StudentInternalMark.objects.update_or_create(
                    student=student,
                    enrollment=enrollment,
                    exam_name=exam_name,
                    course=enrollment.course,
                    part_name=part_name,
                    question_number=question_number,
                    sub_question=sub_question,
                    option_letter=option_letter,
                    defaults={
                        "pattern": exam_pattern,
                        "semester": student_semester,
                        "degree": student_degree,
                        "department": student_department,
                        "max_marks": max_marks_to_save,
                        "marks_obtained": marks_obt_to_save,
                        "co_code_id": co_id,
                        "level_code_id": bl_id,
                        "reg_no": reg_no,
                        "course_code": course_code,
                        "absentee": existing_absentee,
                        "retest_attempted": 1,
                        "batch": batch or enrollment.batch,
                        "section": section or enrollment.section,
                        "academic_year": academic_year,
                        "faculty_assignment": faculty_assignment,
                    },
                )
                upsert_count += 1

        messages.success(
            request,
            f"Retest marks saved successfully for {reg_no}. Rows updated: {upsert_count}"
        )

    except ValueError as e:
        transaction.set_rollback(True)
        messages.error(request, str(e))
        return redirect(request.META.get("HTTP_REFERER", "/"))

    except Exception as e:
        transaction.set_rollback(True)
        messages.error(request, f"Unexpected error while saving retest marks: {str(e)}")
        return redirect(request.META.get("HTTP_REFERER", "/"))

    return redirect(request.META.get("HTTP_REFERER", "/"))



import io
import os
from collections import defaultdict
from datetime import datetime, date

from django.conf import settings
from django.contrib import messages
from django.contrib.staticfiles import finders
from django.http import HttpResponse, FileResponse
from django.shortcuts import get_object_or_404, redirect, render

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame,
    Table, TableStyle, Paragraph, Spacer, PageBreak
)




def _safe(v):
    return "" if v is None else str(v).strip()



def _parse_date_range(request):
    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    parsed_from = None
    parsed_to = None

    try:
        if date_from:
            parsed_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        if date_to:
            parsed_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return None, None, date_from, date_to, HttpResponse(
            "Invalid date format. Use YYYY-MM-DD.", status=400
        )

    if parsed_from and parsed_to and parsed_from > parsed_to:
        return None, None, date_from, date_to, HttpResponse(
            "'From Date' cannot be greater than 'To Date'.", status=400
        )

    return parsed_from, parsed_to, date_from, date_to, None


def _get_daily_attendance_common_data(request, year, semester, course_id, batch, section, regulation_id):
    faculty_id = getattr(request.user, "Employee_id", None)
    faculty = get_object_or_404(general_information, faculty_id=faculty_id)
    course = get_object_or_404(Course, id=course_id)
    regulation = get_object_or_404(Regulations, id=regulation_id)

    course_enrollments = (
        CourseEnrollment.objects
        .filter(
            course_id=course_id,
            batch=batch,
            section=section,
            regulation_id=regulation_id,
            faculty=faculty,
            enroll=True
        )
        .select_related("student", "course", "faculty", "department__degree")
        .order_by("student__reg_no")
    )

    students = [e.student for e in course_enrollments]
    student_ids = [s.id for s in students]
    students_sorted = sorted(students, key=lambda s: _safe(getattr(s, "reg_no", "")))

    enrollment = course_enrollments.first()
    degree_department = ""
    if enrollment and getattr(enrollment, "department", None) and getattr(enrollment.department, "degree", None):
        degree_code = _safe(getattr(enrollment.department.degree, "degree_code", ""))
        department_name = _safe(getattr(enrollment.department, "Department", ""))
        degree_department = f"{degree_code} - {department_name}".strip(" -")

    parsed_from, parsed_to, date_from, date_to, error_response = _parse_date_range(request)
    if error_response:
        return {"error_response": error_response}

    att_qs = Daily_Attendance.objects.filter(
        year=str(year),
        semester=str(semester),
        section=str(section),
    )

    if student_ids:
        att_qs = att_qs.filter(student_id__in=student_ids)
    else:
        att_qs = att_qs.none()

    if parsed_from:
        att_qs = att_qs.filter(date__gte=parsed_from)
    if parsed_to:
        att_qs = att_qs.filter(date__lte=parsed_to)

    att_qs = att_qs.select_related("student").order_by("date", "student__reg_no")

    grouped = defaultdict(lambda: defaultdict(dict))
    all_slots_by_date = defaultdict(set)

    for row in att_qs:
        if not row.student_id or not row.date:
            continue

        grouped[row.student_id][row.date]["M"] = row.morning_status or "-"
        grouped[row.student_id][row.date]["A"] = row.afternoon_status or "-"
        grouped[row.student_id][row.date]["F"] = row.full_day_status or "-"
        all_slots_by_date[row.date].update(["M", "A", "F"])

    sorted_dates = sorted(all_slots_by_date.keys())

    date_columns = []
    for dt in sorted_dates:
        slots = ["M", "A", "F"]
        date_columns.append({
            "date": dt,
            "date_str": dt.strftime("%d-%m-%Y"),
            "periods": slots,
            "period_count": len(slots),
        })

    return {
        "faculty_id": faculty_id,
        "faculty": faculty,
        "course": course,
        "regulation": regulation,
        "course_enrollments": course_enrollments,
        "students": students,
        "students_sorted": students_sorted,
        "student_ids": student_ids,
        "degree_department": degree_department,
        "date_from": date_from,
        "date_to": date_to,
        "parsed_from": parsed_from,
        "parsed_to": parsed_to,
        "att_qs": att_qs,
        "grouped": grouped,
        "sorted_dates": sorted_dates,
        "date_columns": date_columns,
        "year": year,
        "semester": semester,
        "batch": batch,
        "section": section,
        "regulation_id": regulation_id,
        "total_students": len(students_sorted),
        "total_slots": len(date_columns) * 3,
    }


def daily_attendance_view(request, year, semester, course_id, batch, section, regulation_id):
    data = _get_daily_attendance_common_data(
        request, year, semester, course_id, batch, section, regulation_id
    )
    if data.get("error_response"):
        return data["error_response"]

    student_rows = []
    for idx, student in enumerate(data["students_sorted"], start=1):
        values = []

        for dt_info in data["date_columns"]:
            dt = dt_info["date"]
            for slot in dt_info["periods"]:
                status_value = data["grouped"][student.id].get(dt, {}).get(slot, "-")
                normalized = str(status_value).strip().lower()

                values.append({
                    "slot": slot,
                    "value": status_value,
                    "is_absent": normalized == "absent",
                    "is_present": normalized == "present",
                    "is_od": normalized == "on duty",
                    "is_half_day": normalized == "half day",
                })

        student_rows.append({
            "sno": idx,
            "reg_no": _safe(getattr(student, "reg_no", "")),
            "name": _safe(getattr(student, "name", "")),
            "values": values,
        })

    return render(
        request,
        "faculty_management/daily_attendance_view.html",
        {
            "faculty": data["faculty"],
            "course": data["course"],
            "regulation": data["regulation"],
            "year": data["year"],
            "semester": data["semester"],
            "batch": data["batch"],
            "section": data["section"],
            "date_from": data["date_from"],
            "date_to": data["date_to"],
            "date_columns": data["date_columns"],
            "student_rows": student_rows,
            "total_students": data["total_students"],
            "total_slots": data["total_slots"],
        }
    )


def daily_attendance_pdf(request, year, semester, course_id, batch, section, regulation_id):
    data = _get_daily_attendance_common_data(
        request, year, semester, course_id, batch, section, regulation_id
    )
    if data.get("error_response"):
        return data["error_response"]

    faculty = data["faculty"]
    faculty_id = data["faculty_id"]
    course = data["course"]
    regulation = data["regulation"]
    total_students = data["total_students"]
    degree_department = data["degree_department"]
    att_qs = data["att_qs"]
    students_sorted = data["students_sorted"]

    faculty_name = _safe(getattr(faculty, "name", "")) or "Faculty"
    faculty_code = faculty_id or ""

    total_days = att_qs.values("date").distinct().count()

    summary = defaultdict(lambda: {
        "present": 0,
        "absent": 0,
        "od": 0,
        "half": 0,
    })

    for row in att_qs:
        if not row.student_id:
            continue
        status = _safe(row.full_day_status)
        if status == "Present":
            summary[row.student_id]["present"] += 1
        elif status == "Absent":
            summary[row.student_id]["absent"] += 1
        elif status == "On Duty":
            summary[row.student_id]["od"] += 1
        elif status == "Half Day":
            summary[row.student_id]["half"] += 1

    styles = getSampleStyleSheet()

    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1a4b8c")
    ACCENT_RED = colors.HexColor("#b91c1c")
    DARK_GRAY = colors.HexColor("#111827")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BG_GRAY = colors.HexColor("#f8fafc")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle(
        "title_style",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=PRIMARY_BLUE,
        alignment=TA_CENTER,
        spaceAfter=4,
        fontName="Helvetica-Bold",
    )

    sub_style = ParagraphStyle(
        "sub_style",
        parent=styles["Normal"],
        fontSize=10,
        textColor=MEDIUM_GRAY,
        alignment=TA_CENTER,
        spaceAfter=8,
    )

    section_style = ParagraphStyle(
        "section_style",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=PRIMARY_BLUE,
        alignment=TA_LEFT,
        spaceBefore=8,
        spaceAfter=5,
        fontName="Helvetica-Bold",
    )

    info_style = ParagraphStyle(
        "info_style",
        parent=styles["Normal"],
        fontSize=10,
        textColor=DARK_GRAY,
        alignment=TA_LEFT,
        leading=13,
        fontName="Helvetica",
    )

    table_header = ParagraphStyle(
        "table_header",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.white,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        leading=12,
        wordWrap="CJK",
    )

    cell_left = ParagraphStyle(
        "cell_left",
        parent=styles["Normal"],
        fontSize=10,
        textColor=DARK_GRAY,
        alignment=TA_LEFT,
        leading=12,
        wordWrap="CJK",
    )

    cell_center = ParagraphStyle(
        "cell_center",
        parent=cell_left,
        alignment=TA_CENTER,
    )

    def calc_col_widths(page_width):
        fixed = {
            "sno": 14 * mm,
            "reg": 34 * mm,
            "p": 16 * mm,
            "od": 16 * mm,
            "a": 16 * mm,
            "half": 18 * mm,
            "total": 20 * mm,
            "perc": 24 * mm,
            "status": 22 * mm,
        }
        fixed_sum = sum(fixed.values())
        name_w = max(90 * mm, page_width - fixed_sum)

        return [
            fixed["sno"], fixed["reg"], name_w, fixed["p"], fixed["od"],
            fixed["a"], fixed["half"], fixed["total"], fixed["perc"], fixed["status"],
        ]

    def make_table(table_data, col_widths, header_bg=SECONDARY_BLUE, zebra=True):
        t = Table(table_data, colWidths=col_widths, repeatRows=1)

        ts = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.7, BORDER_GRAY),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("ALIGN", (2, 0), (2, -1), "LEFT"),
            ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ])

        if zebra and len(table_data) > 2:
            ts.add("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY])

        t.setStyle(ts)
        return t

    buffer = io.BytesIO()
    PAGE = landscape(A4)

    doc = BaseDocTemplate(
        buffer,
        pagesize=PAGE,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Daily Attendance - {course.course_code}",
        showBoundary=0
    )

    HEADER_HEIGHT = 26 * mm

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = PAGE
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 4 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(
                ImageReader(logo_path),
                left, top_y - 15 * mm,
                width=24 * mm, height=14 * mm,
                preserveAspectRatio=True, mask="auto"
            )

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(center_x, top_y - 4 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawCentredString(center_x, top_y - 9 * mm, "An Autonomous Institution")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 7.5)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "Approved by AICTE, New Delhi")
        canvas.setFont("Helvetica", 7)
        canvas.drawCentredString(center_x, top_y - 16.5 * mm, "Accredited by NAAC & ISO 9001:2015 Certified")
        canvas.drawCentredString(center_x, top_y - 20 * mm, "NBA Accredited: CSE, EEE, ECE, MECH")

        footer_y = 9 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.6)
        canvas.line(left, footer_y + 5 * mm, right, footer_y + 5 * mm)

        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 7)
        gen_time = datetime.now().strftime("%d %b %Y, %H:%M")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawCentredString(center_x, footer_y, f"Faculty: {faculty_code}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")

        canvas.restoreState()

    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin + 3 * mm,
        doc.width,
        doc.height - HEADER_HEIGHT + 4 * mm,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        id="normal"
    )
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

    def p_left(txt):
        return Paragraph(_safe(txt), cell_left)

    def p_center(txt):
        return Paragraph(_safe(txt), cell_center)

    rows = []
    for idx, s in enumerate(students_sorted, start=1):
        r = summary.get(s.id, {})
        present = int(r.get("present") or 0)
        absent = int(r.get("absent") or 0)
        od = int(r.get("od") or 0)
        half = int(r.get("half") or 0)

        attended_equivalent = present + od + (half * 0.5)
        perc = (attended_equivalent / total_days * 100.0) if total_days else 0.0

        if perc >= 75:
            status_text = "Good"
        elif perc >= 50:
            status_text = "Moderate"
        else:
            status_text = "Low"

        rows.append([
            p_center(str(idx)),
            p_center(_safe(getattr(s, "reg_no", ""))),
            p_left(_safe(getattr(s, "name", ""))),
            p_center(str(present)),
            p_center(str(od)),
            p_center(str(absent)),
            p_center(str(half)),
            p_center(str(total_days or 0)),
            p_center(f"{perc:.2f}%"),
            p_center(status_text),
        ])

    elements = []
    elements.append(Spacer(1, 2 * mm))
    elements.append(Paragraph("DAILY ATTENDANCE REPORT", title_style))
    elements.append(Paragraph(f"{faculty_name} ({faculty_code})", sub_style))

    info_data = [
        [Paragraph("<b>Course:</b>", info_style),
         Paragraph(f"{course.course_code} - {course.title}", info_style),
         Paragraph("<b>Regulation:</b>", info_style),
         Paragraph(str(regulation), info_style)],

        [Paragraph("<b>Department:</b>", info_style),
         Paragraph(degree_department, info_style),
         Paragraph("<b>Year / Sem:</b>", info_style),
         Paragraph(f"{data['year']} / {data['semester']}", info_style)],

        [Paragraph("<b>Batch/Sec:</b>", info_style),
         Paragraph(f"{data['batch']} / {data['section']}", info_style),
         Paragraph("<b>Students:</b>", info_style),
         Paragraph(str(total_students), info_style)],
    ]

    if data["parsed_from"] or data["parsed_to"]:
        date_range = f"{data['date_from'] or '...'} to {data['date_to'] or '...'}"
        info_data.append([
            Paragraph("<b>Date Range:</b>", info_style),
            Paragraph(date_range, info_style),
            Paragraph("<b>Total Days:</b>", info_style),
            Paragraph(str(total_days), info_style),
        ])
    else:
        info_data.append([
            Paragraph("<b>Total Days:</b>", info_style),
            Paragraph(str(total_days), info_style),
            Paragraph("<b>Students:</b>", info_style),
            Paragraph(str(total_students), info_style),
        ])

    info_table = Table(info_data, colWidths=[26 * mm, doc.width - 86 * mm, 26 * mm, 34 * mm])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 5 * mm))
    elements.append(Paragraph("STUDENT ATTENDANCE DETAILS", section_style))

    if not rows:
        elements.append(Paragraph("No daily attendance records found for the selected criteria.", info_style))
    else:
        header_row = [
            Paragraph("S.No", table_header),
            Paragraph("Reg No", table_header),
            Paragraph("Student Name", table_header),
            Paragraph("P", table_header),
            Paragraph("OD", table_header),
            Paragraph("A", table_header),
            Paragraph("Half", table_header),
            Paragraph("Total", table_header),
            Paragraph("%", table_header),
            Paragraph("Status", table_header),
        ]

        table_data = [header_row] + rows
        col_widths = calc_col_widths(doc.width)
        attendance_table = make_table(table_data, col_widths, header_bg=SECONDARY_BLUE, zebra=True)
        elements.append(attendance_table)

    try:
        doc.build(elements)
    except Exception as e:
        return HttpResponse(f"PDF generation failed: {e}", status=500)

    buffer.seek(0)
    date_suffix = f"_{data['date_from']}_{data['date_to']}" if (data["date_from"] or data["date_to"]) else ""
    filename = f"Daily_Attendance_{course.course_code}_{data['batch']}_{data['section']}_Y{data['year']}_S{data['semester']}{date_suffix}.pdf"
    return FileResponse(buffer, as_attachment=False, filename=filename)


def daily_attendance_datewise_pdf(request, year, semester, course_id, batch, section, regulation_id):
    class AttendanceDateDocTemplate(BaseDocTemplate):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.current_attendance_date = ""

        def afterFlowable(self, flowable):
            if isinstance(flowable, Paragraph):
                try:
                    txt = flowable.getPlainText().strip()
                except Exception:
                    txt = ""
                if txt.startswith("Attendance Date:"):
                    self.current_attendance_date = txt.replace("Attendance Date:", "").strip()

    data = _get_daily_attendance_common_data(
        request, year, semester, course_id, batch, section, regulation_id
    )
    if data.get("error_response"):
        return data["error_response"]

    if not data["att_qs"].exists():
        return HttpResponse("No daily attendance records found for the selected criteria.", status=404)

    faculty = data["faculty"]
    faculty_id = data["faculty_id"]
    course = data["course"]
    regulation = data["regulation"]
    students_sorted = data["students_sorted"]
    total_students = data["total_students"]

    faculty_name = _safe(getattr(faculty, "name", "")) or "Faculty"
    faculty_code = faculty_id or ""

    styles = getSampleStyleSheet()

    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1a4b8c")
    ACCENT_RED = colors.HexColor("#b91c1c")
    DARK_GRAY = colors.HexColor("#111827")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BG_GRAY = colors.HexColor("#f8fafc")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle("title_style", parent=styles["Heading1"], fontSize=16, textColor=PRIMARY_BLUE, alignment=TA_CENTER, spaceAfter=4, fontName="Helvetica-Bold")
    sub_style = ParagraphStyle("sub_style", parent=styles["Normal"], fontSize=10, textColor=MEDIUM_GRAY, alignment=TA_CENTER, spaceAfter=8)
    date_marker_style = ParagraphStyle("date_marker_style", parent=styles["Normal"], fontSize=1, textColor=colors.white, alignment=TA_LEFT, leading=1, spaceBefore=0, spaceAfter=0)
    info_style = ParagraphStyle("info_style", parent=styles["Normal"], fontSize=10, textColor=DARK_GRAY, alignment=TA_LEFT, leading=13, fontName="Helvetica")
    table_header = ParagraphStyle("table_header", parent=styles["Normal"], fontSize=9, textColor=colors.white, alignment=TA_CENTER, fontName="Helvetica-Bold", leading=11, wordWrap="CJK")
    cell_left = ParagraphStyle("cell_left", parent=styles["Normal"], fontSize=8.5, textColor=DARK_GRAY, alignment=TA_LEFT, leading=10, wordWrap="CJK")
    cell_center = ParagraphStyle("cell_center", parent=cell_left, alignment=TA_CENTER)

    buffer = io.BytesIO()
    PAGE = landscape(A4)

    doc = AttendanceDateDocTemplate(
        buffer,
        pagesize=PAGE,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Datewise Daily Attendance - {course.course_code}",
        showBoundary=0
    )

    HEADER_HEIGHT = 26 * mm

    if data["sorted_dates"]:
        doc.current_attendance_date = data["sorted_dates"][0].strftime("%d-%m-%Y")

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = PAGE
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 4 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(ImageReader(logo_path), left, top_y - 15 * mm, width=24 * mm, height=14 * mm, preserveAspectRatio=True, mask="auto")

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(center_x, top_y - 4 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawCentredString(center_x, top_y - 9 * mm, "An Autonomous Institution")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 7.5)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "Approved by AICTE, New Delhi")
        canvas.setFont("Helvetica", 7)
        canvas.drawCentredString(center_x, top_y - 16.5 * mm, "Accredited by NAAC & ISO 9001:2015 Certified")
        canvas.drawCentredString(center_x, top_y - 20 * mm, "NBA Accredited: CSE, EEE, ECE, MECH")

        current_date = getattr(doc_, "current_attendance_date", "")
        if current_date:
            canvas.setFillColor(ACCENT_RED)
            canvas.setFont("Helvetica-Bold", 10)
            canvas.drawString(left, top_y - 24.5 * mm, f"Date: {current_date}")

        footer_y = 9 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.6)
        canvas.line(left, footer_y + 5 * mm, right, footer_y + 5 * mm)

        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 7)
        gen_time = datetime.now().strftime("%d %b %Y, %H:%M")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawCentredString(center_x, footer_y, f"Faculty: {faculty_code}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")

        canvas.restoreState()

    frame = Frame(doc.leftMargin, doc.bottomMargin + 3 * mm, doc.width, doc.height - HEADER_HEIGHT + 4 * mm, leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0, id="normal")
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

    def p_left(txt):
        return Paragraph(_safe(txt), cell_left)

    def p_center(txt):
        return Paragraph(_safe(txt), cell_center)

    elements = []
    elements.append(Spacer(1, 2 * mm))
    elements.append(Paragraph("DATE-WISE DAILY ATTENDANCE REPORT", title_style))
    elements.append(Paragraph(f"{faculty_name} ({faculty_code})", sub_style))

    info_data = [
        [Paragraph("<b>Course:</b>", info_style), Paragraph(f"{course.course_code} - {course.title}", info_style), Paragraph("<b>Regulation:</b>", info_style), Paragraph(str(regulation), info_style)],
        [Paragraph("<b>Department:</b>", info_style), Paragraph(data["degree_department"], info_style), Paragraph("<b>Year / Sem:</b>", info_style), Paragraph(f"{data['year']} / {data['semester']}", info_style)],
        [Paragraph("<b>Batch/Sec:</b>", info_style), Paragraph(f"{data['batch']} / {data['section']}", info_style), Paragraph("<b>Students:</b>", info_style), Paragraph(str(total_students), info_style)],
        [Paragraph("<b>Date Range:</b>", info_style), Paragraph(f"{data['date_from'] or '...'} to {data['date_to'] or '...'}", info_style), Paragraph("<b>Total Dates:</b>", info_style), Paragraph(str(len(data["sorted_dates"])), info_style)],
    ]

    info_table = Table(info_data, colWidths=[26 * mm, doc.width - 86 * mm, 26 * mm, 34 * mm])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 5 * mm))

    if not data["sorted_dates"]:
        elements.append(Paragraph("No daily attendance records found for the selected criteria.", info_style))
    else:
        slots = ["M", "A", "F"]

        for idx, dt in enumerate(data["sorted_dates"], start=1):
            elements.append(Paragraph(f"Attendance Date: {dt.strftime('%d-%m-%Y')}", date_marker_style))
            elements.append(Spacer(1, 2 * mm))

            header_row = [
                Paragraph("S.No", table_header),
                Paragraph("Reg No", table_header),
                Paragraph("Student Name", table_header),
            ] + [Paragraph(slot, table_header) for slot in slots]

            rows = [header_row]

            for sno, student in enumerate(students_sorted, start=1):
                slot_status_map = data["grouped"][student.id].get(dt, {})

                row = [
                    p_center(str(sno)),
                    p_center(_safe(getattr(student, "reg_no", ""))),
                    p_left(_safe(getattr(student, "name", ""))),
                ]

                for slot in slots:
                    row.append(p_center(_safe(slot_status_map.get(slot, "-"))))

                rows.append(row)

            base_width = doc.width
            sno_w = 12 * mm
            reg_w = 28 * mm
            name_w = 85 * mm
            remaining = base_width - (sno_w + reg_w + name_w)
            slot_w = remaining / max(len(slots), 1)

            col_widths = [sno_w, reg_w, name_w] + [slot_w] * len(slots)

            date_table = Table(rows, colWidths=col_widths, repeatRows=1)
            date_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.6, BORDER_GRAY),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
                ("ALIGN", (0, 0), (1, -1), "CENTER"),
                ("ALIGN", (3, 1), (-1, -1), "CENTER"),
            ]))

            elements.append(date_table)

            if idx < len(data["sorted_dates"]):
                elements.append(PageBreak())

    try:
        doc.build(elements)
    except Exception as e:
        return HttpResponse(f"PDF generation failed: {e}", status=500)

    buffer.seek(0)
    date_suffix = f"_{data['date_from']}_{data['date_to']}" if (data["date_from"] or data["date_to"]) else ""
    filename = f"Daily_Attendance_Datewise_{course.course_code}_{data['batch']}_{data['section']}_Y{data['year']}_S{data['semester']}{date_suffix}.pdf"
    return FileResponse(buffer, as_attachment=False, filename=filename)


def daily_attendance_datewise_excel(request, year, semester, course_id, batch, section, regulation_id):
    data = _get_daily_attendance_common_data(
        request, year, semester, course_id, batch, section, regulation_id
    )
    if data.get("error_response"):
        return data["error_response"]

    if not data["att_qs"].exists():
        return HttpResponse("No daily attendance records found for the selected criteria.", status=404)

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Attendance"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    absent_fill = PatternFill(fill_type="solid", fgColor="FDE2E1")
    present_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    od_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    half_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    ws.merge_cells("A1:D1")
    ws["A1"] = "DATE-WISE DAILY ATTENDANCE REPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Course"
    ws["B2"] = f"{data['course'].course_code} - {data['course'].title}"
    ws["A3"] = "Faculty"
    ws["B3"] = f"{_safe(getattr(data['faculty'], 'name', ''))} ({_safe(data['faculty_id'])})"
    ws["A4"] = "Batch / Section"
    ws["B4"] = f"{data['batch']} / {data['section']}"
    ws["C2"] = "Regulation"
    ws["D2"] = str(data["regulation"])
    ws["C3"] = "Year / Semester"
    ws["D3"] = f"{data['year']} / {data['semester']}"
    ws["C4"] = "Date Range"
    ws["D4"] = f"{data['date_from'] or '...'} to {data['date_to'] or '...'}"

    for cell in ["A2", "A3", "A4", "C2", "C3", "C4"]:
        ws[cell].font = bold_font

    header_row_1 = 6
    header_row_2 = 7

    ws.merge_cells(start_row=header_row_1, start_column=1, end_row=header_row_2, end_column=1)
    ws.merge_cells(start_row=header_row_1, start_column=2, end_row=header_row_2, end_column=2)
    ws.merge_cells(start_row=header_row_1, start_column=3, end_row=header_row_2, end_column=3)

    ws.cell(header_row_1, 1).value = "S.No"
    ws.cell(header_row_1, 2).value = "Reg No"
    ws.cell(header_row_1, 3).value = "Student Name"

    current_col = 4
    date_column_map = []

    for dt in data["sorted_dates"]:
        slots = ["M", "A", "F"]
        start_col = current_col

        for slot in slots:
            ws.cell(header_row_2, current_col).value = slot
            current_col += 1

        end_col = current_col - 1

        if start_col <= end_col:
            ws.merge_cells(
                start_row=header_row_1,
                start_column=start_col,
                end_row=header_row_1,
                end_column=end_col
            )
            ws.cell(header_row_1, start_col).value = dt.strftime("%d-%m-%Y")
            date_column_map.append((dt, slots))

    for row_num in [header_row_1, header_row_2]:
        for col in range(1, current_col):
            cell = ws.cell(row_num, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    data_row = 8
    for idx, student in enumerate(data["students_sorted"], start=1):
        ws.cell(data_row, 1).value = idx
        ws.cell(data_row, 2).value = _safe(getattr(student, "reg_no", ""))
        ws.cell(data_row, 3).value = _safe(getattr(student, "name", ""))

        ws.cell(data_row, 1).alignment = center_align
        ws.cell(data_row, 2).alignment = center_align
        ws.cell(data_row, 3).alignment = left_align

        for fixed_col in [1, 2, 3]:
            ws.cell(data_row, fixed_col).border = thin_border

        col_ptr = 4
        for dt, slots in date_column_map:
            for slot in slots:
                value = data["grouped"][student.id].get(dt, {}).get(slot, "-")
                cell = ws.cell(data_row, col_ptr)
                cell.value = _safe(value)
                cell.alignment = center_align
                cell.border = thin_border

                normalized = str(value).strip().lower()
                if normalized == "absent":
                    cell.fill = absent_fill
                elif normalized == "present":
                    cell.fill = present_fill
                elif normalized == "on duty":
                    cell.fill = od_fill
                elif normalized == "half day":
                    cell.fill = half_fill

                col_ptr += 1

        data_row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28

    for col in range(4, current_col):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = "D8"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Daily_Attendance_Datewise_{data['course'].course_code}_{data['batch']}_{data['section']}_Y{data['year']}_S{data['semester']}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
 



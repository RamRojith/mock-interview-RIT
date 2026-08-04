from collections import defaultdict
import datetime

from user_accounts.decorators import check_permission, no_cache,is_super_user
from django.utils.decorators import method_decorator
from django.shortcuts import render,redirect,get_object_or_404
import re
from django.contrib.auth.decorators import login_required
from django.views.generic import CreateView, UpdateView, ListView,DeleteView
from django.urls import reverse_lazy, reverse
from django.core.paginator import Paginator
from user_accounts.models import Role,Department
from faculty_leave_management.models import LeaveAllotment,LeaveType,LeavePermissionFunction,LeaveApplication,LeaveApprovers, ShiftMaster , ShiftDetail, Employee_Holidays, AttendancePolicy, AttendancePolicyAssignment
from faculty_management.models import general_information, DesignationMaster, FacultyCategory
from django.contrib import messages
from django.db import IntegrityError, connections
from django.db.models import Q
from faculty_leave_management.forms import LeaveTypeForm,LeaveAllotmentForm,LeaveAllotmentFormSet,LeaveAllotmentUpdateForm,LeaveApplicationFormSet,LeaveApplicationForm
from django.http import JsonResponse
import json
from django.views.decorators.csrf import csrf_exempt
import logging
logger = logging.getLogger(__name__)

from user_accounts.decorators import no_cache, is_super_user
from user_accounts.models import *
from faculty_leave_management.decorators import faculty_leave_management


def _filter_holidays_for_category(queryset, category_id):
    if category_id:
        return queryset.filter(Q(category_id=category_id) | Q(category__isnull=True))
    return queryset.filter(category__isnull=True)


def _get_matching_holiday(role_id, holiday_date, category_id=None):
    if not role_id:
        return None
    return (
        _filter_holidays_for_category(
            Employee_Holidays.objects.filter(role_id=role_id, holiday_date=holiday_date),
            category_id,
        )
        .order_by("-category_id")
        .first()
    )


def _holiday_map_for_employee(holiday_map, role_id, category_id=None):
    if not role_id:
        return {}
    role_days = holiday_map.get(role_id, {})
    employee_holidays = {}
    for holiday_date, category_items in role_days.items():
        holiday_obj = category_items.get(category_id) if category_id else None
        if not holiday_obj:
            holiday_obj = category_items.get(None)
        if holiday_obj:
            employee_holidays[holiday_date] = holiday_obj
    return employee_holidays


def _add_to_holiday_map(holiday_map, holiday_obj):
    holiday_map.setdefault(holiday_obj.role_id, {}).setdefault(
        holiday_obj.holiday_date, {}
    )[holiday_obj.category_id] = holiday_obj

@faculty_leave_management
def flm_home(request):
    # print("flm home page ")
    request.session['current_page'] = 'flm_home'
    return redirect('home')

@check_permission('flm_hello')
def flm_hello(request):
    return render(request, 'flm_hello.html')

@no_cache
@is_super_user('faculty_leave_management')
def flm_assign_permission(request):
    if request.method == 'POST':
        permissions = request.POST
        for role_name, role_permissions in permissions.items():
            if role_name.startswith('permissions'):
                try:
                    # Extract data from role_name using regex
                    extract_data = list(re.findall(r'\[([^\]]+)\]', role_name))
                    if len(extract_data) < 2:  # Ensure there are at least role and function
                        messages.warning(request,f"Invalid format in role_name: {role_name}. Skipping.")
                        continue

                    extract_data.append(role_permissions)

                    # Retrieve the role (Handle if the role does not exist)
                    try:
                        role = Role.objects.using("rit_approval_system").get(role=extract_data[0])

                    except Role.DoesNotExist:
                        messages.error(request,f"Role {extract_data[0]} does not exist.")
                        messages.error(request, f"Role '{extract_data[0]}' does not exist. Skipping this entry.")
                        continue

                    # Parse permissions - handle the case where role_permissions is a list (unlikely with POST data)
                    if isinstance(role_permissions, list):  # Handle list case
                        role_permissions = role_permissions[0]

                    # Convert permission to boolean (True/False)
                    permission = extract_data[2] == 'true'


                    # Find or create RolePermissionFunction object
                    permission_obj = LeavePermissionFunction.objects.filter(
                        role=role, function=extract_data[1]
                    ).first()

                    if permission_obj:
                        permission_obj.permission = permission
                        permission_obj.save()
                    else:
                        # Create a new RolePermissionFunction object
                        LeavePermissionFunction.objects.create(
                            role=role,
                            function=extract_data[1],
                            permission=permission
                        )
                except Exception as e:
                    # Catch unexpected errors and log them
                    messages.error(request,f"Error processing role '{role_name}': {str(e)}")
                    messages.error(request, f"An error occurred while processing '{role_name}': {str(e)}")

    # Redirect to admin dashboard after processing
    messages.success(request,"The permission changes have been successfully applied.")
    return redirect('faculty_leave_management')


from faculty_leave_management.models import LeaveAllotment

from django.db import IntegrityError, transaction

def faculty_leave_allotments_list(request):
    leave_allotments = (
        LeaveAllotment.objects
        .select_related('leave_type', 'role', 'category')
        .order_by('-active', '-academic_year', 'role__designation_name')
    )

    leave_types = LeaveType.objects.all().order_by('name')
    designations = DesignationMaster.objects.all().order_by('designation_name')
    categories = FacultyCategory.objects.filter(is_active=True).order_by('category_name')

    # Build academic years for dropdown/filter
    from datetime import date
    current_year = date.today().year
    academic_years = [f"{y}-{y + 1}" for y in range(current_year - 1, current_year + 6)]
    default_academic_year = f"{current_year}-{current_year + 1}"

    if request.method == "POST":
        action = request.POST.get('action')

        # ADD
        if action == "add":
            academic_year = request.POST.get('academic_year')
            leave_type_id = request.POST.get('leave_type')
            target_type = request.POST.get('target_type') or 'role'
            designation_ids = request.POST.getlist('role')
            category_ids = request.POST.getlist('category')
            default_allotment = request.POST.get('default_allotment')
            frequency = request.POST.get('frequency') or 'yearly'
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            active = request.POST.get('active') == 'on'

            selected_ids = category_ids if target_type == 'category' else designation_ids

            if not all([academic_year, leave_type_id, selected_ids, default_allotment, start_date, end_date]):
                messages.error(request, "All fields are required.")
                return redirect('faculty_leave_allotments_list')

            if start_date > end_date:
                messages.error(request, "Start date cannot be greater than end date.")
                return redirect('faculty_leave_allotments_list')

            try:
                leave_type = LeaveType.objects.get(id=leave_type_id)
            except LeaveType.DoesNotExist:
                messages.error(request, "Selected leave type not found.")
                return redirect('faculty_leave_allotments_list')

            added_count = 0
            skipped_count = 0

            for target_id in selected_ids:
                role = None
                category = None
                try:
                    if target_type == 'category':
                        category = FacultyCategory.objects.get(id=target_id)
                        target_label = category.category_name
                    else:
                        role = DesignationMaster.objects.get(id=target_id)
                        target_label = role.designation_name
                except (DesignationMaster.DoesNotExist, FacultyCategory.DoesNotExist):
                    skipped_count += 1
                    continue

                try:
                    with transaction.atomic():
                        allotment, created = LeaveAllotment.objects.get_or_create(
                            academic_year=academic_year,
                            role=role,
                            category=category,
                            leave_type=leave_type,
                            defaults={
                                'default_allotment': default_allotment,
                                'frequency': frequency,
                                'start_date': start_date,
                                'end_date': end_date,
                                'active': active,
                            }
                        )
                        if created:
                            added_count += 1
                        else:
                            skipped_count += 1
                            messages.warning(
                                request,
                                f"Already exists for {target_label} in {academic_year}."
                            )
                except IntegrityError:
                    skipped_count += 1
                    messages.warning(
                        request,
                        f"Duplicate entry detected for {target_label} in {academic_year}."
                    )

            if added_count:
                messages.success(request, f"{added_count} allotment(s) added successfully.")

            if not added_count and skipped_count:
                messages.info(request, "No new allotments were added.")

        # EDIT
        elif action == "edit":
            allotment_id = request.POST.get('id')
            academic_year = request.POST.get('academic_year')
            leave_type_id = request.POST.get('leave_type')
            target_type = request.POST.get('target_type') or 'role'
            role_id = request.POST.get('role')       # first selected checkbox value
            category_id = request.POST.get('category')
            default_allotment = request.POST.get('default_allotment')
            frequency = request.POST.get('frequency') or 'yearly'
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            active = request.POST.get('active') == 'on'

            target_id = category_id if target_type == 'category' else role_id

            if not all([allotment_id, academic_year, leave_type_id, target_id, default_allotment, start_date, end_date]):
                messages.error(request, "All fields are required for editing.")
                return redirect('faculty_leave_allotments_list')

            if start_date > end_date:
                messages.error(request, "Start date cannot be greater than end date.")
                return redirect('faculty_leave_allotments_list')

            try:
                allotment = LeaveAllotment.objects.get(id=allotment_id)
            except LeaveAllotment.DoesNotExist:
                messages.error(request, "Allotment not found.")
                return redirect('faculty_leave_allotments_list')

            role = None
            category = None
            try:
                leave_type = LeaveType.objects.get(id=leave_type_id)
                if target_type == 'category':
                    category = FacultyCategory.objects.get(id=target_id)
                    target_label = category.category_name
                else:
                    role = DesignationMaster.objects.get(id=target_id)
                    target_label = role.designation_name
            except (DesignationMaster.DoesNotExist, FacultyCategory.DoesNotExist, LeaveType.DoesNotExist):
                messages.error(request, "Invalid designation, category or leave type selected.")
                return redirect('faculty_leave_allotments_list')

            duplicate_exists = (
                LeaveAllotment.objects
                .exclude(id=allotment.id)
                .filter(
                    academic_year=academic_year,
                    role=role,
                    category=category,
                    leave_type=leave_type
                )
                .exists()
            )

            if duplicate_exists:
                messages.warning(
                    request,
                    f"Duplicate allotment exists for {target_label} in {academic_year}."
                )
            else:
                allotment.academic_year = academic_year
                allotment.role = role
                allotment.category = category
                allotment.leave_type = leave_type
                allotment.default_allotment = default_allotment
                allotment.frequency = frequency
                allotment.start_date = start_date
                allotment.end_date = end_date
                allotment.active = active
                allotment.save()

                messages.success(request, "Leave allotment updated successfully.")

        # DELETE
        elif action == "delete":
            allotment_id = request.POST.get('id')

            try:
                LeaveAllotment.objects.get(id=allotment_id).delete()
                messages.success(request, "Allotment deleted successfully.")
            except LeaveAllotment.DoesNotExist:
                messages.error(request, "Allotment not found.")

        return redirect('faculty_leave_allotments_list')

    context = {
        'leave_allotments': leave_allotments,
        'leave_types': leave_types,
        'designations': designations,
        'categories': categories,
        'academic_years': academic_years,
        'default_academic_year': default_academic_year,
    }
    return render(
        request,
        'faculty_leave_management/admin/faculty_leave_allotments_list.html',
        context
    )





def faculty_leave_types(request):
    if request.method == "POST":
        action = request.POST.get('action')

        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip().upper()
        is_leave = request.POST.get('is_leave') == 'true'
        is_active = request.POST.get('is_active') == 'true'
        restriction = request.POST.get('restriction') or 'restricted'

        # ADD
        if action == 'add':
            if not name or not code:
                messages.error(request, "Name and Code are required.")
            elif LeaveType.objects.filter(name=name).exists():
                messages.error(request, f"Leave type '{name}' already exists.")
            elif LeaveType.objects.filter(code=code).exists():
                messages.error(request, f"Code '{code}' already exists.")
            else:
                LeaveType.objects.create(
                    name=name,
                    code=code,
                    is_leave=is_leave,
                    is_active=is_active,
                    restriction=restriction,
                )
                messages.success(request, f"'{name}' added successfully.")

        # EDIT
        elif action == 'edit':
            lt_id = request.POST.get('id')

            try:
                lt = LeaveType.objects.get(id=lt_id)

                if LeaveType.objects.exclude(id=lt_id).filter(name=name).exists():
                    messages.error(request, f"Name '{name}' already exists.")
                elif LeaveType.objects.exclude(id=lt_id).filter(code=code).exists():
                    messages.error(request, f"Code '{code}' already exists.")
                else:
                    lt.name = name
                    lt.code = code
                    lt.is_leave = is_leave
                    lt.is_active = is_active
                    lt.restriction = restriction
                    lt.save()

                    messages.success(request, f"'{name}' updated successfully.")

            except LeaveType.DoesNotExist:
                messages.error(request, "Leave type not found.")

        # DELETE
        elif action == 'delete':
            lt_id = request.POST.get('id')

            try:
                lt = LeaveType.objects.get(id=lt_id)
                lt_name = lt.name
                lt.delete()
                messages.success(request, f"'{lt_name}' deleted successfully.")
            except LeaveType.DoesNotExist:
                messages.error(request, "Leave type not found.")

        return redirect('faculty_leave_management')

    return redirect('faculty_leave_management')


@check_permission('leave_approval_management')
def leave_approval_management(request):

    if request.method == 'GET':
        roles = Role.objects.using("rit_approval_system").all()
        departments = Add_Department.objects.all()

        return render(request, "faculty_leave_management/admin/models/leave_management.html", {
            'roles': roles,
            'departments': departments,
        })

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            creator_role_id_raw = data.get('creatorRole')
            role_hierarchy = data.get('roleHierarchy', [])

            try:
                creator_role_id = int(creator_role_id_raw)
            except:
                return JsonResponse({'error': 'Invalid creatorRole'}, status=400)

            # ✅ delete old
            LeaveApprovers.objects.filter(
                creator_role_id=creator_role_id
            ).delete()

            for index, role_data in enumerate(role_hierarchy):

                # role id
                try:
                    approver_role_id = int(role_data.get('id'))
                except:
                    return JsonResponse({'error': 'Invalid role id'}, status=400)

                is_cross_department = bool(role_data.get('isCrossDepartment', False))

                dept_id = role_data.get('departmentId')
                if dept_id in ["", None, "null", "undefined"]:
                    dept_id = None
                else:
                    dept_id = int(dept_id)

                # validation
                if is_cross_department and not dept_id:
                    return JsonResponse({
                        'error': f'Department required for role {approver_role_id}'
                    }, status=400)

                department_obj = Add_Department.objects.filter(id=dept_id).first() if dept_id else None

                LeaveApprovers.objects.create(
                    creator_role_id=creator_role_id,
                    approver_role_id=approver_role_id,
                    approver_level=index + 1,
                    is_cross_department_approver="YES" if is_cross_department else "NO",
                    approver_department=department_obj
                )

            return JsonResponse({'message': 'Roles submitted successfully'}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request'}, status=405)




from django.http import JsonResponse
from django.views.decorators.http import require_GET

@is_super_user('faculty_leave_management')
def api_leave_roles(request, creatorRoleId):
    try:
        # ✅ creatorRoleId from URL is string, cast to int safely
        try:
            creator_role_id = int(creatorRoleId)
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid creatorRoleId"}, status=400)

        ROLE_DB = "rit_approval_system"

        # ✅ Ensure creator role exists in ROLE_DB
        creator_role = (
            Role.objects.using(ROLE_DB)
            .filter(id=creator_role_id)
            .values("id", "role")
            .first()
        )
        if not creator_role:
            return JsonResponse({"error": "Creator role not found"}, status=404)

        # ✅ All roles, including the creator role itself — a role can be an
        # approver of its own members' leave, so it must be available to drag
        # into the hierarchy.
        all_roles_qs = Role.objects.using(ROLE_DB).all()

        # ✅ Approver mappings (from your default DB unless LeaveApprovers is routed)
        # If LeaveApprovers is also in a different DB, add .using("that_db") here.
        approvers_qs = LeaveApprovers.objects.filter(creator_role_id=creator_role_id)

        matched_roles = []
        matched_ids = set()

        # ✅ Collect role_ids first (avoid N queries)
        approver_role_ids = list(
            approvers_qs.values_list("approver_role_id", flat=True)
        )

        # ✅ Map role_id -> role_name from ROLE_DB
        role_name_map = {
            r["id"]: r["role"]
            for r in Role.objects.using(ROLE_DB)
                .filter(id__in=approver_role_ids)
                .values("id", "role")
        }

        for item in approvers_qs:
            rid = item.approver_role_id
            matched_ids.add(rid)

            matched_roles.append({
                "id": rid,
                "role": role_name_map.get(rid, ""),  # ✅ no cross-db FK deref
                "is_cross_department": str(getattr(item, "is_cross_department_approver", "")).upper() == "YES",
                "approver_department_id": item.approver_department_id,
                "approver_department_name": item.approver_department.Department if item.approver_department else None,

            })

        unmatched_roles = [
            {"id": r.id, "role": r.role}
            for r in all_roles_qs.exclude(id__in=list(matched_ids))
        ]

        return JsonResponse({
            "matched_roles": matched_roles,
            "unmatched_roles": unmatched_roles
        })

    except Exception as e:
        # helpful during dev; you can log traceback too
        return JsonResponse({"error": str(e)}, status=500)




from faculty_leave_management.models import DeviceInfo

@check_permission('device_entry')
def device_entry(request):
    edit_id = request.GET.get("edit")
    delete_id = request.GET.get("delete")
    toggle_id = request.GET.get("toggle")

    # DELETE
    if delete_id:
        obj = get_object_or_404(DeviceInfo, id=delete_id)
        obj.delete()
        messages.success(request, "Device deleted successfully")
        return redirect("device_entry")

    # TOGGLE ACTIVE
    if toggle_id:
        obj = get_object_or_404(DeviceInfo, id=toggle_id)
        obj.is_active = not obj.is_active
        obj.save()
        messages.success(request, "Status updated")
        return redirect("device_entry")

    # EDIT LOAD
    edit_obj = None
    if edit_id:
        edit_obj = get_object_or_404(DeviceInfo, id=edit_id)

    # SAVE / UPDATE
    if request.method == "POST":
        device_id = (request.POST.get("deviceid") or "").strip()
        location = (request.POST.get("devicelocation") or "").strip()
        is_active = request.POST.get("is_active") == "on"
        is_attendance = request.POST.get("is_attendance") == "on"
        is_mess = request.POST.get("is_mess") == "on"
        obj_id = (request.POST.get("id") or "").strip()

        if not device_id or not location:
            messages.error(request, "Device ID and Location are required.")
            return redirect("device_entry")

        if obj_id:
            obj = get_object_or_404(DeviceInfo, id=obj_id)
            obj.deviceid = device_id
            obj.devicelocation = location
            obj.is_active = is_active
            obj.is_attendance = is_attendance
            obj.is_mess = is_mess
            obj.save()
            messages.success(request, "Device updated successfully.")
        else:
            DeviceInfo.objects.create(
                deviceid=device_id,
                devicelocation=location,
                is_active=is_active,
                is_attendance=is_attendance,
                is_mess=is_mess,
            )
            messages.success(request, "Device added successfully.")

        return redirect("device_entry")

    # LIST
    devices = DeviceInfo.objects.all().order_by("-id")

    return render(request, "faculty_leave_management/device_entry.html", {
        "devices": devices,
        "edit_obj": edit_obj
    })


import calendar
from datetime import datetime, date, timedelta
from urllib.parse import urlencode

from django.shortcuts import render
from django.db import connections

from user_accounts.decorators import check_permission
from faculty_leave_management.models import Employee_Holidays
from faculty_leave_management.models import DeviceInfo, ShiftDetail
from faculty_management.models import general_information

MANUAL_PUNCH_DEVICE_ID = "18"


@check_permission('punch_attendance')
def punch_attendance(request):
    user_id = request.user.Employee_id
    now = datetime.now()

    # Fetch the user's shift details
    user_info = general_information.objects.filter(faculty_id=user_id).first()
    user_shift = user_info.shift if user_info else None
    user_category_id = user_info.category_id if user_info else None

    shift_start_time = shift_end_time = None
    if user_shift:
        first_shift = ShiftDetail.objects.filter(shift_master=user_shift).order_by('shift_no').first()
        if first_shift:
            shift_start_time = first_shift.start_time
            shift_end_time = first_shift.end_time

    year = int(request.GET.get("year", now.year))
    month = int(request.GET.get("month", now.month))
    weekday = (request.GET.get("weekday") or "").strip()
    device_type = (request.GET.get("device_type") or "").strip()

    # Attendance cycle: 26th of the previous month -> 25th of the selected month.
    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year
    cycle_start = date(prev_year, prev_month, 26)
    cycle_end = date(year, month, 25)

    # Get logged-in user role ID
    user_role_id = None
    candidate_role_ids = set()
    if hasattr(request.user, "role_id") and request.user.role_id:
        try:
            candidate_role_ids.add(int(request.user.role_id))
        except (TypeError, ValueError):
            pass
    if user_id:
        try:
            from user_accounts.models import USER
            rows = (
                USER.objects.using("rit_approval_system")
                .filter(Employee_id=str(user_id))
                .values_list("role_id", flat=True)
            )
            for rid in rows:
                try:
                    if rid is not None:
                        candidate_role_ids.add(int(rid))
                except (TypeError, ValueError):
                    pass
        except Exception:
            pass
    if candidate_role_ids:
        role_holiday_counts = {
            rid: _filter_holidays_for_category(
                Employee_Holidays.objects.filter(
                    role_id=rid,
                    holiday_date__gte=cycle_start,
                    holiday_date__lte=cycle_end,
                ),
                user_category_id,
            ).count()
            for rid in candidate_role_ids
        }
        user_role_id = sorted(
            candidate_role_ids,
            key=lambda rid: (-role_holiday_counts.get(rid, 0), rid)
        )[0]

    # DEVICE MASTER — map device_id -> info dict
    devices_qs = DeviceInfo.objects.filter(is_active=True).order_by("deviceid")
    device_master = {}
    attendance_device_ids = []
    mess_device_ids = []
    for d in devices_qs:
        dev_id = str(d.deviceid or "").strip()
        if not dev_id:
            continue
        if d.is_attendance:
            dev_type = "Attendance"
            attendance_device_ids.append(dev_id)
        elif d.is_mess:
            dev_type = "Mess"
            mess_device_ids.append(dev_id)
        else:
            dev_type = "Other"
        device_master[dev_id] = {
            "location": d.devicelocation or "-",
            "is_attendance": d.is_attendance,
            "is_mess": d.is_mess,
            "type_label": dev_type,
        }

    # Determine device IDs to filter by based on device_type
    filter_device_ids = None
    if device_type == "attendance":
        filter_device_ids = attendance_device_ids
    elif device_type == "mess":
        filter_device_ids = mess_device_ids

    # RAW SQL QUERY against attendance_db. Punches live in per-month tables
    # (DeviceLogs_{month}_{year}), so the 26th->25th cycle spans two tables:
    # the previous month and the selected month.
    from faculty_leave_management.signals import punch_data_fetched

    rows = []
    error_message = None
    for tbl_month, tbl_year in [(prev_month, prev_year), (month, year)]:
        if filter_device_ids:
            placeholders = ",".join(["%s"] * len(filter_device_ids))
            query = f"""
                SELECT DeviceLogId, UserId, LogDate, Direction, DeviceId
                FROM DeviceLogs_{tbl_month}_{tbl_year}
                WHERE UserId = %s AND DeviceId IN ({placeholders})
                ORDER BY LogDate ASC
            """
            params = [user_id] + filter_device_ids
        else:
            query = f"""
                SELECT DeviceLogId, UserId, LogDate, Direction, DeviceId
                FROM DeviceLogs_{tbl_month}_{tbl_year}
                WHERE UserId = %s
                ORDER BY LogDate ASC
            """
            params = [user_id]

        try:
            with connections["attendance_db"].cursor() as cursor:
                cursor.execute(query, params)
                columns = [col[0] for col in cursor.description]
                tbl_rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

            rows.extend(tbl_rows)

            # Fire signal to sync fetched rows into local DB
            punch_data_fetched.send(
                sender=None,
                userid=user_id,
                month=tbl_month,
                year=tbl_year,
                rows=tbl_rows,
            )
        except Exception as e:
            # A missing month table (e.g. no previous-month data) shouldn't
            # abort the whole view — record the message and keep going.
            error_message = str(e)

    # GROUP PUNCHES BY DATE
    grouped = {}
    for row in rows:
        log_dt = row.get("LogDate")
        if not log_dt:
            continue
        dev_id = str(row.get("DeviceId") or "").strip()
        date_key = log_dt.date()
        grouped.setdefault(date_key, []).append({
            "DeviceLogId": row.get("DeviceLogId"),
            "UserId": row.get("UserId"),
            "LogDate": log_dt,
            "Direction": row.get("Direction"),
            "DeviceId": dev_id,
            # store the full info dict so template can use log.DeviceLocation.location
            "DeviceLocation": device_master.get(dev_id, {"location": "-", "is_attendance": False, "is_mess": False, "type_label": "Other"}),
        })

    # Generate all dates in the attendance cycle (26th prev month -> 25th sel month)
    all_dates = []
    _d = cycle_start
    while _d <= cycle_end:
        all_dates.append(_d)
        _d += timedelta(days=1)

    # Applied leaves / permissions / On Duty for this faculty in the selected
    # month, mapped per date so the Attendance column can show them.
    from faculty_leave_management.models import LeaveApplication as _LeaveApplication
    leave_by_date = {}
    if user_info:
        month_start = cycle_start
        month_end = cycle_end
        status_rank = {"Approved": 0, "Pre-approved": 1, "Pending": 2}
        leave_qs = (
            _LeaveApplication.objects
            .filter(
                faculty=user_info,
                from_date__lte=month_end,
                to_date__gte=month_start,
                status__in=["Approved", "Pre-approved", "Pending"],
            )
            .select_related("leave_type")
        )
        for lv in leave_qs:
            lt = lv.leave_type
            name = (getattr(lt, "name", None) or "Leave")
            code = (getattr(lt, "code", "") or "").upper()
            is_od = code in ("OD", "ROD") or "on duty" in name.lower()
            label = name
            if lv.from_time and lv.to_time:
                label = f"{name} ({lv.from_time.strftime('%I:%M %p')}–{lv.to_time.strftime('%I:%M %p')})"
            # Always show the approval status — for both Approved and Pending.
            label = f"{label} · {lv.status}"
            rank = status_rank.get(lv.status, 9)
            # Pending/Pre-approved use an amber badge; approved uses the type badge.
            if (lv.status or "") != "Approved":
                badge = "badge-fn"
            else:
                badge = "badge-present" if is_od else "badge-leave"
            d0 = max(lv.from_date, month_start)
            d1 = min(lv.to_date, month_end)
            while d0 <= d1:
                existing = leave_by_date.get(d0)
                if not existing or rank < existing["rank"]:
                    leave_by_date[d0] = {"label": label, "badge": badge, "rank": rank}
                d0 += timedelta(days=1)

    # PREPARE PUNCH DATA
    punch_data = []
    for att_date in all_dates:
        # Apply weekday filter
        if weekday and calendar.day_name[att_date.weekday()] != weekday:
            continue

        logs = grouped.get(att_date, [])
        attendance_logs = [x for x in logs if x["DeviceLocation"].get("is_attendance")]
        mess_logs = [x for x in logs if x["DeviceLocation"].get("is_mess")]

        first_in = "-"
        last_out = "-"
        first_in_time = None
        last_out_time = None

        att_count = len(attendance_logs)

        if att_count >= 1:
            sorted_att = sorted(attendance_logs, key=lambda x: x["LogDate"])
            first_in_time = sorted_att[0]["LogDate"]
            first_in = first_in_time.strftime("%I:%M:%S %p")
            # Last out only when there are 2+ punches
            if att_count >= 2:
                last_out_time = sorted_att[-1]["LogDate"]
                last_out = last_out_time.strftime("%I:%M:%S %p")

        # Use shift midpoint to classify each punch as IN or OUT
        shift_mid_dt = None
        if shift_start_time and shift_end_time:
            shift_start_dt = datetime.combine(att_date, shift_start_time)
            shift_end_dt = datetime.combine(att_date, shift_end_time)
            shift_mid_dt = shift_start_dt + (shift_end_dt - shift_start_dt) / 2

        for log in logs:
            log_dt = log["LogDate"]
            if log["DeviceLocation"].get("is_attendance"):
                if shift_mid_dt:
                    log["computed_direction"] = "IN" if log_dt <= shift_mid_dt else "OUT"
                else:
                    # fallback: first punch IN, last punch OUT
                    log["computed_direction"] = "IN"
            else:
                log["computed_direction"] = "MESS"

        # Re-assign first/last based on computed direction for attendance logs
        in_logs = [l for l in logs if l.get("computed_direction") == "IN"]
        out_logs = [l for l in logs if l.get("computed_direction") == "OUT"]
        if in_logs:
            first_in_time = min(l["LogDate"] for l in in_logs)
            first_in = first_in_time.strftime("%I:%M:%S %p")
        if out_logs:
            last_out_time = max(l["LogDate"] for l in out_logs)
            last_out = last_out_time.strftime("%I:%M:%S %p")

        delayed_punch = False
        early_punch = False
        if shift_start_time and shift_end_time:
            shift_start_dt = datetime.combine(att_date, shift_start_time)
            shift_end_dt = datetime.combine(att_date, shift_end_time)
            if first_in_time:
                delayed_punch = first_in_time > shift_start_dt
            # early exit only meaningful when there's an actual out punch
            if last_out_time:
                early_punch = last_out_time < shift_end_dt

        # Holiday status (computed first so attendance can respect it)
        holiday_status, holiday_badge, holiday_reason = "Working Day", "badge-working", ""
        if user_role_id:
            holiday_obj = _get_matching_holiday(user_role_id, att_date, user_category_id)
            if holiday_obj:
                holiday_reason = holiday_obj.reason or ""
                if holiday_obj.session_type in ("H", "F"):
                    holiday_status, holiday_badge = "Holiday", "badge-holiday"
                elif holiday_obj.session_type == "FN":
                    holiday_status, holiday_badge = "Forenoon Holiday", "badge-fn"
                elif holiday_obj.session_type == "AN":
                    holiday_status, holiday_badge = "Afternoon Holiday", "badge-an"

        is_holiday = holiday_status != "Working Day"

        # Attendance status
        if att_count == 0:
            if is_holiday:
                attendance_status, attendance_badge = "Holiday", "badge-holiday"
            else:
                attendance_status, attendance_badge = "Absent", "badge-absent"
        elif att_count == 1:
            attendance_status, attendance_badge = "Single Punch", "badge-morning"
        elif delayed_punch and not is_holiday:
            attendance_status, attendance_badge = "Absent", "badge-absent"
        else:
            attendance_status, attendance_badge = "Present", "badge-present"

        # Mess status
        mess_count = len(mess_logs)
        if mess_count == 0:
            mess_status, mess_badge = "No Mess Punch", "badge-mess-none"
        elif mess_count == 1:
            mess_status, mess_badge = "Mess Punched Once", "badge-mess-one"
        else:
            mess_status, mess_badge = f"Mess Punched ({mess_count})", "badge-mess-many"

        # Compute work duration
        work_time = None
        work_time_str = "-"
        if first_in_time and last_out_time and last_out_time > first_in_time:
            delta = last_out_time - first_in_time
            total_minutes = int(delta.total_seconds() // 60)
            hours, minutes = divmod(total_minutes, 60)
            work_time_str = f"{hours}h {minutes:02d}m"
            work_time = delta

        # If a leave / permission / On Duty is applied on this date, surface it
        # in the Attendance column (overriding the punch-derived status).
        leave_info = leave_by_date.get(att_date)
        if leave_info:
            attendance_status = leave_info["label"]
            attendance_badge = leave_info["badge"]

        punch_data.append({
            "date": att_date,
            "day_name": calendar.day_name[att_date.weekday()],
            "first_in": first_in,
            "last_out": last_out,
            "work_time": work_time_str,
            "logs": logs,
            "total_logs": len(logs),
            "attendance_status": attendance_status,
            "attendance_badge": attendance_badge,
            "leave_applied": leave_info["label"] if leave_info else "",
            "mess_status": mess_status,
            "mess_badge": mess_badge,
            "holiday_status": holiday_status,
            "holiday_badge": holiday_badge,
            "holiday_reason": holiday_reason,
            "delayed_punch": delayed_punch,
            "early_punch": early_punch,
        })

    # Build filter options for template
    current_year = now.year
    year_options = list(range(current_year - 2, current_year + 2))
    month_options = [(i, calendar.month_name[i]) for i in range(1, 13)]
    weekday_options = list(calendar.day_name)

    context = {
        "punch_data": punch_data,
        "devices": devices_qs,
        "selected_year": year,
        "selected_month": month,
        "selected_weekday": weekday,
        "selected_device_type": device_type,
        "cycle_start": cycle_start,
        "cycle_end": cycle_end,
        "year_options": year_options,
        "month_options": month_options,
        "weekday_options": weekday_options,
        "error_message": error_message,
        "shift_start_time": shift_start_time,
        "shift_end_time": shift_end_time,
    }

    return render(request, "faculty_leave_management/punch_attendance.html", context)


def _compute_mess_bill(user_id, year, month):
    """Compute a faculty's mess bill for the cycle (26th prev month -> 25th
    selected month). Mess punches are matched to a meal by time using the
    MessDetails master, and the amount is summed from each meal's price.
    Returns a dict of the report data (used by both the personal and admin views)."""
    from collections import defaultdict
    from decimal import Decimal
    from faculty_leave_management.models import MessDetails
    from faculty_leave_management.signals import punch_data_fetched

    # Mess cycle: 26th of the previous month -> 25th of the selected month.
    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year
    cycle_start = date(prev_year, prev_month, 26)
    cycle_end = date(year, month, 25)

    # Mess devices
    mess_device_ids = [
        str(d).strip()
        for d in DeviceInfo.objects.filter(is_active=True, is_mess=True)
        .values_list("deviceid", flat=True)
        if d
    ]

    # Mess timing + price master. Prefer the current academic year; if nothing
    # is configured for it, fall back to all active entries.
    academic_year = get_academic_year()
    meals = list(
        MessDetails.objects.filter(is_active=True, academic_year=academic_year)
        .order_by("from_time")
    )
    if not meals:
        meals = list(MessDetails.objects.filter(is_active=True).order_by("from_time"))

    # Fetch mess punches from both month tables spanning the cycle.
    rows = []
    error_message = None
    if not user_id:
        error_message = None  # no employee selected yet
    elif not mess_device_ids:
        error_message = "No mess devices are configured."
    elif not meals:
        error_message = "No mess details (timing/price) are configured."
    else:
        placeholders = ",".join(["%s"] * len(mess_device_ids))
        for tbl_month, tbl_year in [(prev_month, prev_year), (month, year)]:
            query = f"""
                SELECT DeviceLogId, UserId, LogDate, Direction, DeviceId
                FROM DeviceLogs_{tbl_month}_{tbl_year}
                WHERE UserId = %s AND DeviceId IN ({placeholders})
                ORDER BY LogDate ASC
            """
            params = [user_id] + mess_device_ids
            try:
                with connections["attendance_db"].cursor() as cursor:
                    cursor.execute(query, params)
                    columns = [col[0] for col in cursor.description]
                    tbl_rows = [dict(zip(columns, row)) for row in cursor.fetchall()]
                rows.extend(tbl_rows)
                punch_data_fetched.send(
                    sender=None, userid=user_id,
                    month=tbl_month, year=tbl_year, rows=tbl_rows,
                )
            except Exception as e:
                error_message = str(e)

    # Match a punch time to a meal window.
    def match_meal(t):
        for m in meals:
            if m.from_time and m.to_time and m.from_time <= t <= m.to_time:
                return m
        return None

    # Group punches by date -> meal (one meal counts once per day, earliest punch).
    daily = defaultdict(dict)
    for row in rows:
        log_dt = row.get("LogDate")
        if not log_dt:
            continue
        d = log_dt.date()
        if d < cycle_start or d > cycle_end:
            continue
        meal = match_meal(log_dt.time())
        if not meal:
            continue
        entry = daily[d].get(meal.id)
        if not entry:
            daily[d][meal.id] = log_dt
        elif log_dt < entry:
            daily[d][meal.id] = log_dt

    # Build daily rows + per-meal and grand-total summary.
    meal_summary = {
        m.id: {"meal": m, "count": 0, "amount": Decimal("0")} for m in meals
    }
    bill_rows = []
    grand_total = Decimal("0")
    total_meal_count = 0

    for d in sorted(daily.keys()):
        day_meals = daily[d]
        day_items = []
        day_amount = Decimal("0")
        for m in meals:
            if m.id in day_meals:
                rupees = m.rupees or Decimal("0")
                day_items.append({
                    "name": m.name,
                    "time": day_meals[m.id],
                    "rupees": rupees,
                })
                day_amount += rupees
                meal_summary[m.id]["count"] += 1
                meal_summary[m.id]["amount"] += rupees
                total_meal_count += 1
        grand_total += day_amount
        bill_rows.append({
            "date": d,
            "day_name": calendar.day_name[d.weekday()],
            "items": day_items,
            "amount": day_amount,
        })

    return {
        "bill_rows": bill_rows,
        "meal_summary": [meal_summary[m.id] for m in meals],
        "grand_total": grand_total,
        "total_meal_count": total_meal_count,
        "academic_year": academic_year,
        "selected_year": year,
        "selected_month": month,
        "cycle_start": cycle_start,
        "cycle_end": cycle_end,
        "error_message": error_message,
    }


def mess_bills(request):
    """Personal mess bill — the logged-in user views only their own bill."""
    now = datetime.now()
    user_id = request.user.Employee_id
    user_info = general_information.objects.filter(faculty_id=user_id).first()

    year = int(request.GET.get("year", now.year))
    month = int(request.GET.get("month", now.month))

    context = _compute_mess_bill(user_id, year, month)
    context.update({
        "year_options": list(range(now.year - 2, now.year + 2)),
        "month_options": [(i, calendar.month_name[i]) for i in range(1, 13)],
        "faculty": user_info,
    })
    return render(request, "faculty_leave_management/mess_bills.html", context)


@check_permission("employee_attandance_report")
def emp_mess_bills(request):
    """Admin mess bill — pick any employee (like emp_punch_details) and view
    their mess bill for the selected cycle."""
    now = datetime.now()

    employees = (
        general_information.objects
        .filter(faculty_id__isnull=False)
        .select_related("designation", "department")
        .order_by("faculty_id")
    )

    selected_faculty_id = (request.GET.get("faculty_id") or "").strip()
    year = int(request.GET.get("year", now.year))
    month = int(request.GET.get("month", now.month))

    selected_employee = None
    if selected_faculty_id:
        selected_employee = employees.filter(faculty_id=selected_faculty_id).first()

    context = _compute_mess_bill(selected_faculty_id, year, month)
    context.update({
        "year_options": list(range(now.year - 2, now.year + 2)),
        "month_options": [(i, calendar.month_name[i]) for i in range(1, 13)],
        "employees": employees,
        "selected_faculty_id": selected_faculty_id,
        "selected_employee": selected_employee,
    })
    return render(request, "faculty_leave_management/admin/emp_mess_bills.html", context)



from datetime import datetime
def shift_entry(request):
    edit_id = request.GET.get("edit")
    delete_id = request.GET.get("delete")

    if delete_id:
        master = get_object_or_404(ShiftMaster, id=delete_id)
        master.delete()
        messages.success(request, "Shift master deleted successfully.")
        return redirect("shift_entry")

    edit_master = None
    edit_details = []
    generated_rows = 0
    shift_name = ""
    master_id = ""

    if edit_id:
        edit_master = get_object_or_404(ShiftMaster, id=edit_id)
        edit_details = list(edit_master.shift_details.all())
        generated_rows = edit_master.no_of_shifts
        shift_name = edit_master.shift_name
        master_id = edit_master.id

    if request.method == "POST":
        action = request.POST.get("action")
        shift_name = (request.POST.get("shift_name") or "").strip()
        no_of_shifts = request.POST.get("no_of_shifts")
        master_id = (request.POST.get("master_id") or "").strip()

        try:
            generated_rows = int(no_of_shifts) if no_of_shifts else 0
        except:
            generated_rows = 0

        if master_id:
            edit_master = ShiftMaster.objects.filter(id=master_id).first()
            if edit_master:
                edit_details = list(edit_master.shift_details.all())

        if action == "generate":
            masters = ShiftMaster.objects.all().order_by("-id")
            context = {
                "edit_master": edit_master,
                "edit_details": edit_details,
                "generated_rows": generated_rows,
                "masters": masters,
                "shift_name": shift_name,
                "master_id": master_id,
            }
            return render(request, "faculty_leave_management/admin/shift_entry.html", context)

        elif action == "save":
            if not shift_name:
                messages.error(request, "Shift name is required.")
                return redirect("shift_entry")

            try:
                no_of_shifts = int(no_of_shifts)
                if no_of_shifts <= 0:
                    raise ValueError
            except:
                messages.error(request, "Enter valid number of shifts.")
                return redirect("shift_entry")

            for i in range(1, no_of_shifts + 1):
                start_time = request.POST.get(f"start_time_{i}")
                end_time = request.POST.get(f"end_time_{i}")

                if not start_time or not end_time:
                    messages.error(request, f"Please enter both start time and end time for Shift {i}.")
                    masters = ShiftMaster.objects.all().order_by("-id")
                    context = {
                        "edit_master": edit_master,
                        "edit_details": edit_details,
                        "generated_rows": no_of_shifts,
                        "masters": masters,
                        "shift_name": shift_name,
                        "master_id": master_id,
                    }
                    return render(request, "faculty_leave_management/admin/shift_entry.html", context)

            if master_id:
                master = get_object_or_404(ShiftMaster, id=master_id)
                master.shift_name = shift_name
                master.no_of_shifts = no_of_shifts
                master.save()
                master.shift_details.all().delete()
                msg = "Shift master updated successfully."
            else:
                master = ShiftMaster.objects.create(
                    shift_name=shift_name,
                    no_of_shifts=no_of_shifts
                )
                msg = "Shift master created successfully."

            for i in range(1, no_of_shifts + 1):
                start_time = request.POST.get(f"start_time_{i}")
                end_time = request.POST.get(f"end_time_{i}")
                is_next_day = request.POST.get(f"is_next_day_{i}") == "on"

                ShiftDetail.objects.create(
                    shift_master=master,
                    shift_no=i,
                    start_time=start_time,
                    end_time=end_time,
                    is_next_day=is_next_day
                )

            messages.success(request, msg)
            return redirect("shift_entry")

    masters = ShiftMaster.objects.all().order_by("-id")

    context = {
        "edit_master": edit_master,
        "edit_details": edit_details,
        "generated_rows": generated_rows,
        "masters": masters,
        "shift_name": shift_name,
        "master_id": master_id,
    }
    return render(request, "faculty_leave_management/admin/shift_entry.html", context)

from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from user_accounts.decorators import check_permission
from faculty_leave_management.models import PermissionTimingMaster


from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

def permission_timing_master(request):
    edit_id = request.GET.get("edit")
    delete_id = request.GET.get("delete")

    # ---------------- DELETE ----------------
    if delete_id:
        obj = get_object_or_404(PermissionTimingMaster, id=delete_id)
        obj.delete()
        messages.success(request, "Permission timing deleted successfully.")
        return redirect("permission_timing_master")

    # ---------------- EDIT ----------------
    edit_obj = None
    if edit_id:
        edit_obj = get_object_or_404(PermissionTimingMaster, id=edit_id)

    # ---------------- HELPER: SAFE TIME PARSER ----------------
    def parse_time_safe(value):
        try:
            return datetime.strptime(value, "%H:%M").time()
        except ValueError:
            try:
                return datetime.strptime(value, "%H:%M:%S").time()
            except ValueError:
                return None

    # ---------------- CREATE / UPDATE ----------------
    if request.method == "POST":
        row_id = request.POST.get("row_id")

        session_name = (request.POST.get("session_name") or "").strip()
        start_time = (request.POST.get("start_time") or "").strip()
        end_time = (request.POST.get("end_time") or "").strip()
        value = (request.POST.get("value") or "").strip()
        is_active = request.POST.get("is_active") == "on"

        # -------- BASIC VALIDATION --------
        if not all([session_name, start_time, end_time, value]):
            messages.error(request, "All fields are required.")
            return redirect("permission_timing_master")

        # -------- SAFE TIME PARSE --------
        start_obj = parse_time_safe(start_time)
        end_obj = parse_time_safe(end_time)

        if not start_obj or not end_obj:
            messages.error(request, "Invalid time format.")
            return redirect("permission_timing_master")

        # -------- TIME VALIDATION --------
        if start_obj >= end_obj:
            messages.error(request, "End time must be later than start time.")
            return redirect("permission_timing_master")

        # -------- VALUE VALIDATION --------
        try:
            value = float(value)
        except ValueError:
            messages.error(request, "Value must be numeric (e.g., 0.5, 1.0).")
            return redirect("permission_timing_master")

        if value <= 0:
            messages.error(request, "Value must be greater than 0.")
            return redirect("permission_timing_master")

        # -------- OPTIONAL: DUPLICATE CHECK --------
        duplicate = PermissionTimingMaster.objects.filter(
            start_time=start_obj,
            end_time=end_obj
        )

        if row_id:
            duplicate = duplicate.exclude(id=row_id)

        if duplicate.exists():
            messages.error(request, "This exact time range already exists.")
            return redirect("permission_timing_master")

        # -------- SAVE / UPDATE --------
        if row_id:
            obj = get_object_or_404(PermissionTimingMaster, id=row_id)
            obj.session_name = session_name
            obj.start_time = start_obj
            obj.end_time = end_obj
            obj.value = value
            obj.is_active = is_active
            obj.save()
            messages.success(request, "Permission timing updated successfully.")
        else:
            PermissionTimingMaster.objects.create(
                session_name=session_name,
                start_time=start_obj,
                end_time=end_obj,
                value=value,
                is_active=is_active
            )
            messages.success(request, "Permission timing added successfully.")

        return redirect("permission_timing_master")

    # ---------------- FETCH DATA ----------------
    timings = PermissionTimingMaster.objects.all().order_by("start_time")

    return render(request, "faculty_leave_management/admin/permission_timing_master.html", {
        "timings": timings,
        "edit_obj": edit_obj,
    })


def get_academic_year():
    """
    Dynamically returns academic year string.
    Example:
      If current month >= June -> '2025-2026'
      Else (Jan-May) -> '2024-2025'
    """
    from datetime import date
    today = date.today()
    current_year = today.year
    if today.month >= 6:  # June or later
        return f"{current_year}-{current_year + 1}"
    else:  # Before June -> part of previous cycle
        return f"{current_year - 1}-{current_year}"


def mess_details(request):
    from datetime import datetime
    from decimal import Decimal, InvalidOperation
    from faculty_leave_management.models import MessDetails

    edit_id = request.GET.get("edit")
    delete_id = request.GET.get("delete")

    # ---------------- DELETE ----------------
    if delete_id:
        obj = get_object_or_404(MessDetails, id=delete_id)
        obj.delete()
        messages.success(request, "Mess detail deleted successfully.")
        return redirect("mess_details")

    # ---------------- EDIT ----------------
    edit_obj = None
    if edit_id:
        edit_obj = get_object_or_404(MessDetails, id=edit_id)

    # ---------------- HELPER: SAFE TIME PARSER ----------------
    def parse_time_safe(value):
        try:
            return datetime.strptime(value, "%H:%M").time()
        except ValueError:
            try:
                return datetime.strptime(value, "%H:%M:%S").time()
            except ValueError:
                return None

    # ---------------- CREATE / UPDATE ----------------
    if request.method == "POST":
        row_id = request.POST.get("row_id")

        name = (request.POST.get("name") or "").strip()
        from_time = (request.POST.get("from_time") or "").strip()
        to_time = (request.POST.get("to_time") or "").strip()
        rupees = (request.POST.get("rupees") or "").strip()
        academic_year = (request.POST.get("academic_year") or "").strip()
        is_active = request.POST.get("is_active") == "on"

        # -------- BASIC VALIDATION --------
        if not all([name, from_time, to_time, rupees, academic_year]):
            messages.error(request, "All fields are required.")
            return redirect("mess_details")

        # -------- SAFE TIME PARSE --------
        from_obj = parse_time_safe(from_time)
        to_obj = parse_time_safe(to_time)

        if not from_obj or not to_obj:
            messages.error(request, "Invalid time format.")
            return redirect("mess_details")

        if from_obj >= to_obj:
            messages.error(request, "To time must be later than from time.")
            return redirect("mess_details")

        # -------- PRICE VALIDATION --------
        try:
            rupees_val = Decimal(rupees)
        except (InvalidOperation, ValueError):
            messages.error(request, "Rupees must be numeric (e.g., 40, 55.50).")
            return redirect("mess_details")

        if rupees_val < 0:
            messages.error(request, "Rupees cannot be negative.")
            return redirect("mess_details")

        # -------- SAVE / UPDATE --------
        if row_id:
            obj = get_object_or_404(MessDetails, id=row_id)
            obj.name = name
            obj.from_time = from_obj
            obj.to_time = to_obj
            obj.rupees = rupees_val
            obj.academic_year = academic_year
            obj.is_active = is_active
            obj.save()
            messages.success(request, "Mess detail updated successfully.")
        else:
            MessDetails.objects.create(
                name=name,
                from_time=from_obj,
                to_time=to_obj,
                rupees=rupees_val,
                academic_year=academic_year,
                is_active=is_active,
            )
            messages.success(request, "Mess detail added successfully.")

        return redirect("mess_details")

    # ---------------- FETCH DATA ----------------
    mess_list = MessDetails.objects.all().order_by("academic_year", "from_time")

    # Build the dropdown from the current academic year (get_academic_year)
    # spanning two years before to two years after the present.
    current_academic_year = get_academic_year()
    start_year = int(current_academic_year.split("-")[0])
    academic_years = [f"{y}-{y + 1}" for y in range(start_year - 2, start_year + 3)]

    return render(request, "faculty_leave_management/admin/mess_details.html", {
        "mess_list": mess_list,
        "edit_obj": edit_obj,
        "academic_years": academic_years,
        "current_academic_year": current_academic_year,
    })


def ccl_timing_master(request):
    from decimal import Decimal
    from faculty_leave_management.models import CCLTimingMaster

    edit_id = request.GET.get("edit")
    delete_id = request.GET.get("delete")

    if delete_id:
        obj = get_object_or_404(CCLTimingMaster, id=delete_id)
        obj.delete()
        messages.success(request, "CCL timing deleted successfully.")
        return redirect("ccl_timing_master")

    edit_obj = get_object_or_404(CCLTimingMaster, id=edit_id) if edit_id else None

    if request.method == "POST":
        row_id = request.POST.get("row_id")
        session_name = (request.POST.get("session_name") or "").strip()
        min_hours_raw = (request.POST.get("min_hours") or "").strip()
        max_hours_raw = (request.POST.get("max_hours") or "").strip()
        ccl_days_raw = (request.POST.get("ccl_days") or "").strip()
        is_active = request.POST.get("is_active") == "on"

        if not session_name or not min_hours_raw or not ccl_days_raw:
            messages.error(request, "Session name, minimum hours, and CCL days are required.")
            return redirect("ccl_timing_master")

        try:
            min_hours = Decimal(min_hours_raw)
            max_hours = Decimal(max_hours_raw) if max_hours_raw else None
            ccl_days = Decimal(ccl_days_raw)
        except Exception:
            messages.error(request, "Hours and CCL days must be valid numbers.")
            return redirect("ccl_timing_master")

        if min_hours < 0:
            messages.error(request, "Minimum hours cannot be negative.")
            return redirect("ccl_timing_master")
        if max_hours is not None and max_hours <= min_hours:
            messages.error(request, "Maximum hours must be greater than minimum hours.")
            return redirect("ccl_timing_master")
        if ccl_days <= 0:
            messages.error(request, "CCL days must be greater than 0.")
            return redirect("ccl_timing_master")

        if row_id:
            obj = get_object_or_404(CCLTimingMaster, id=row_id)
            obj.session_name = session_name
            obj.min_hours = min_hours
            obj.max_hours = max_hours
            obj.ccl_days = ccl_days
            obj.is_active = is_active
            obj.save()
            messages.success(request, "CCL timing updated successfully.")
        else:
            CCLTimingMaster.objects.create(
                session_name=session_name,
                min_hours=min_hours,
                max_hours=max_hours,
                ccl_days=ccl_days,
                is_active=is_active,
            )
            messages.success(request, "CCL timing added successfully.")

        return redirect("ccl_timing_master")

    timings = CCLTimingMaster.objects.all().order_by("min_hours", "max_hours", "session_name")
    return render(request, "faculty_leave_management/admin/ccl_timing_master.html", {
        "timings": timings,
        "edit_obj": edit_obj,
    })


def attendance_policy_master(request):
    from decimal import Decimal, InvalidOperation
    from django.core.exceptions import ValidationError

    edit_policy_id = request.GET.get("edit_policy")
    delete_policy_id = request.GET.get("delete_policy")
    delete_assignment_id = request.GET.get("delete_assignment")

    if delete_policy_id:
        policy = get_object_or_404(AttendancePolicy, id=delete_policy_id)
        policy.delete()
        messages.success(request, "Attendance policy deleted successfully.")
        return redirect("attendance_policy_master")

    if delete_assignment_id:
        assignment = get_object_or_404(AttendancePolicyAssignment, id=delete_assignment_id)
        assignment.delete()
        messages.success(request, "Attendance policy assignment deleted successfully.")
        return redirect("attendance_policy_master")

    edit_policy = get_object_or_404(AttendancePolicy, id=edit_policy_id) if edit_policy_id else None

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "save_policy":
            policy_id = request.POST.get("policy_id")
            policy_name = (request.POST.get("policy_name") or "").strip()
            punch_mode = request.POST.get("punch_mode") or "FIRST_LAST"
            odd_punch_handling = request.POST.get("odd_punch_handling") or "MISSING_OUT"
            late_early_check = request.POST.get("late_early_check") == "on"
            shift_required = request.POST.get("shift_required") == "on"
            is_active = request.POST.get("is_active") == "on"

            try:
                minimum_punches_required = int(request.POST.get("minimum_punches_required") or 2)
                minimum_working_hours = Decimal(request.POST.get("minimum_working_hours") or "0")
            except (ValueError, InvalidOperation):
                messages.error(request, "Minimum punches and working hours must be valid numbers.")
                return redirect("attendance_policy_master")

            if not policy_name:
                messages.error(request, "Policy name is required.")
                return redirect("attendance_policy_master")
            if minimum_punches_required < 1:
                messages.error(request, "Minimum punches must be at least 1.")
                return redirect("attendance_policy_master")
            if minimum_working_hours < 0:
                messages.error(request, "Minimum working hours cannot be negative.")
                return redirect("attendance_policy_master")

            policy = get_object_or_404(AttendancePolicy, id=policy_id) if policy_id else AttendancePolicy()
            policy.policy_name = policy_name
            policy.punch_mode = punch_mode
            policy.minimum_punches_required = minimum_punches_required
            policy.minimum_working_hours = minimum_working_hours
            policy.late_early_check = late_early_check
            policy.shift_required = shift_required
            policy.odd_punch_handling = odd_punch_handling
            policy.is_active = is_active
            policy.save()
            messages.success(request, "Attendance policy saved successfully.")
            return redirect("attendance_policy_master")

        if action == "save_assignment":
            policy_id = request.POST.get("policy")
            scope_type = request.POST.get("scope_type")
            notes = (request.POST.get("notes") or "").strip()

            policy = AttendancePolicy.objects.filter(id=policy_id, is_active=True).first()
            if not policy:
                messages.error(request, "Choose an active attendance policy.")
                return redirect("attendance_policy_master")

            defaults = {
                "policy": policy,
                "is_active": request.POST.get("assignment_is_active") == "on",
                "notes": notes,
            }

            lookup = {}
            if scope_type == "employee":
                employee_id = request.POST.get("employee")
                if not employee_id:
                    messages.error(request, "Choose an employee for employee-specific policy.")
                    return redirect("attendance_policy_master")
                lookup = {"employee_id": employee_id}
                defaults.update({"designation": None, "category": None, "is_default": False})
            elif scope_type == "designation":
                designation_id = request.POST.get("designation")
                if not designation_id:
                    messages.error(request, "Choose a designation for designation policy.")
                    return redirect("attendance_policy_master")
                lookup = {"designation_id": designation_id}
                defaults.update({"employee": None, "category": None, "is_default": False})
            elif scope_type == "category":
                category_id = request.POST.get("category")
                if not category_id:
                    messages.error(request, "Choose a category for category policy.")
                    return redirect("attendance_policy_master")
                lookup = {"category_id": category_id}
                defaults.update({"employee": None, "designation": None, "is_default": False})
            elif scope_type == "default":
                lookup = {"is_default": True}
                defaults.update({"employee": None, "designation": None, "category": None})
            else:
                messages.error(request, "Choose a valid assignment scope.")
                return redirect("attendance_policy_master")

            assignment, _created = AttendancePolicyAssignment.objects.update_or_create(
                **lookup,
                defaults=defaults,
            )
            try:
                assignment.full_clean()
                assignment.save()
            except ValidationError as exc:
                assignment.delete()
                messages.error(request, "; ".join(exc.messages))
                return redirect("attendance_policy_master")

            messages.success(request, "Attendance policy assignment saved successfully.")
            return redirect("attendance_policy_master")

    employees = (
        general_information.objects
        .filter(faculty_id__isnull=False)
        .select_related("department", "designation", "category")
        .order_by("faculty_id")
    )
    policies = AttendancePolicy.objects.all().order_by("policy_name")
    active_policies = policies.filter(is_active=True)
    assignments = (
        AttendancePolicyAssignment.objects
        .select_related("policy", "employee", "employee__department", "designation", "category")
        .all()
    )
    designations = DesignationMaster.objects.all().order_by("designation_name")
    categories = FacultyCategory.objects.filter(is_active=True).order_by("category_name")

    return render(request, "faculty_leave_management/admin/attendance_policy_master.html", {
        "policies": policies,
        "active_policies": active_policies,
        "assignments": assignments,
        "employees": employees,
        "designations": designations,
        "categories": categories,
        "edit_policy": edit_policy,
    })

from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

from faculty_leave_management.models import Employee_Holidays
from user_accounts.models import Role

@check_permission('Employee_holidays')
def Employee_holidays(request):
    delete_id = request.GET.get("delete")
    roles = list(Role.objects.using("rit_approval_system").all().order_by("role"))
    categories = list(FacultyCategory.objects.filter(is_active=True).order_by("category_name"))
    category_ids = [category.id for category in categories]

    # Map each category -> the roles (rit_approval_system Role ids) held by the
    # faculty assigned to that category in general_information. The UI uses this
    # to auto-tick a category's roles when the category is selected.
    from user_accounts.models import USER as _USER
    category_role_map = {}
    for _cat in categories:
        _fac_ids = [
            str(fid)
            for fid in general_information.objects
            .filter(category=_cat)
            .values_list("faculty_id", flat=True)
            if fid
        ]
        _role_ids = []
        if _fac_ids:
            _role_ids = list(
                _USER.objects.using("rit_approval_system")
                .filter(Employee_id__in=_fac_ids, is_active=True)
                .values_list("role_id", flat=True)
                .distinct()
            )
        category_role_map[_cat.id] = sorted({int(r) for r in _role_ids if r})

    # DELETE
    if delete_id:
        obj = get_object_or_404(Employee_Holidays, id=delete_id)
        obj.delete()
        messages.success(request, "Holiday assignment deleted successfully.")
        return redirect("Employee_holidays")

    # EDIT LOAD (per-date): pre-fill the form with every role/category assigned to that date
    edit_date_data = None
    edit_date_param = request.GET.get("edit_date")
    if edit_date_param:
        try:
            ed = datetime.strptime(edit_date_param, "%Y-%m-%d").date()
        except ValueError:
            ed = None
        if ed:
            recs = list(
                Employee_Holidays.objects.filter(holiday_date=ed)
                .select_related("category")
                .order_by("role_id", "category__category_name")
            )
            if recs:
                saved_category_ids = [r.category_id for r in recs if r.category_id]
                has_all_category_entry = any(not r.category_id for r in recs)
                edit_date_data = {
                    "holiday_date": ed,
                    "reason": next((r.reason for r in recs if r.reason), ""),
                    "session_type": recs[0].session_type,
                    "role_ids": sorted({int(r.role_id) for r in recs if r.role_id}),
                    "category_ids": category_ids if has_all_category_entry else sorted(set(saved_category_ids)),
                }

    # SAVE / UPDATE
    if request.method == "POST":
        edit_date_str = (request.POST.get("edit_date") or "").strip()
        holiday_date = (request.POST.get("holiday_date") or "").strip()
        session_type = (request.POST.get("session_type") or "").strip()
        reason = (request.POST.get("reason") or "").strip()

        if not holiday_date:
            messages.error(request, "Please select a date.")
            return redirect("Employee_holidays")

        if not session_type:
            messages.error(request, "Please select Holiday / FN / AN.")
            return redirect("Employee_holidays")

        try:
            parsed_date = datetime.strptime(holiday_date, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Invalid date format.")
            return redirect("Employee_holidays")

        role_ids = request.POST.getlist("role")
        if not role_ids:
            messages.error(request, "Please select at least one role.")
            return redirect("Employee_holidays")

        selected_category_ids = request.POST.getlist("category")
        if not selected_category_ids:
            messages.error(request, "Please select at least one category.")
            return redirect("Employee_holidays")

        valid_roles = Role.objects.using("rit_approval_system").filter(id__in=role_ids)
        valid_role_ids = [int(r.id) for r in valid_roles]
        if not valid_role_ids:
            messages.error(request, "Selected roles are invalid.")
            return redirect("Employee_holidays")

        valid_categories = FacultyCategory.objects.filter(
            id__in=selected_category_ids,
            is_active=True,
        ).order_by("category_name")
        valid_category_ids = [category.id for category in valid_categories]
        if not valid_category_ids:
            messages.error(request, "Selected categories are invalid.")
            return redirect("Employee_holidays")

        # DATE EDIT MODE: replace the entire set of roles for the edited date
        if edit_date_str:
            try:
                original_date = datetime.strptime(edit_date_str, "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "Invalid date format.")
                return redirect("Employee_holidays")

            Employee_Holidays.objects.filter(holiday_date=original_date).delete()
            for rid in valid_role_ids:
                for category_id in valid_category_ids:
                    Employee_Holidays.objects.update_or_create(
                        holiday_date=parsed_date,
                        role_id=rid,
                        category_id=category_id,
                        defaults={"session_type": session_type, "reason": reason},
                    )

            messages.success(
                request,
                f"Holiday updated with {len(valid_role_ids) * len(valid_category_ids)} assignment(s)."
            )
            return redirect("Employee_holidays")

        # CREATE MODE
        created_count = 0
        updated_count = 0

        for rid in valid_role_ids:
            for category_id in valid_category_ids:
                obj, created = Employee_Holidays.objects.update_or_create(
                    holiday_date=parsed_date,
                    role_id=rid,
                    category_id=category_id,
                    defaults={
                        "session_type": session_type,
                        "reason": reason,
                    }
                )

                if created:
                    created_count += 1
                else:
                    updated_count += 1

        if created_count > 0 and updated_count > 0:
            messages.success(
                request,
                f"{created_count} assignment(s) added and {updated_count} assignment(s) updated successfully."
            )
        elif created_count > 0:
            messages.success(request, f"{created_count} assignment(s) saved successfully.")
        elif updated_count > 0:
            messages.success(request, f"{updated_count} assignment(s) updated successfully.")

        return redirect("Employee_holidays")

    role_map = {int(role.id): role.role for role in roles}
    category_map = {category.id: category.category_name for category in categories}
    session_map = dict(Employee_Holidays.SESSION_CHOICES)

    holiday_qs = (
        Employee_Holidays.objects
        .select_related("category")
        .all()
        .order_by("-holiday_date", "role_id", "category__category_name")
    )

    grouped_holidays = {}
    for item in holiday_qs:
        date_key = item.holiday_date

        if date_key not in grouped_holidays:
            grouped_holidays[date_key] = {
                "holiday_date": item.holiday_date,
                "reason": item.reason,
                "roles": [],
            }

        grouped_holidays[date_key]["roles"].append({
            "id": item.id,
            "role_id": item.role_id,
            "role_name": role_map.get(int(item.role_id), "-") if item.role_id else "-",
            "category_id": item.category_id,
            "category_name": (
                item.category.category_name
                if item.category
                else category_map.get(item.category_id, "All Categories")
            ),
            "session_type": item.session_type,
            "session_label": session_map.get(
                item.session_type,
                "Full Day" if item.session_type == "H" else "-",
            ),
        })

        if not grouped_holidays[date_key]["reason"] and item.reason:
            grouped_holidays[date_key]["reason"] = item.reason

    grouped_holidays = list(grouped_holidays.values())

    import json
    context = {
        "roles": roles,
        "categories": categories,
        "grouped_holidays": grouped_holidays,
        "edit_date_data": edit_date_data,
        "session_choices": Employee_Holidays.SESSION_CHOICES,
        "category_role_map_json": json.dumps(category_role_map),
    }
    return render(request, "faculty_leave_management/admin/Employee_holidays.html", context)


def _get_day_status(att_date, punches, leave_apps_day, holiday_obj, ccl_dates=None, perm_session=None, attendance_policy=None):
    h_code = (holiday_obj.session_type or "").strip().upper() if holiday_obj else ""

    leave_code = ""
    for la in leave_apps_day:
        if la.from_date <= att_date <= la.to_date:
            leave_code = ((la.leave_type.code if la.leave_type else "L") or "L").strip().upper()
            break
    has_ccl_on_day = bool(ccl_dates and att_date in ccl_dates)
    if not leave_code and has_ccl_on_day:
        leave_code = "CCL"

    att_punches = [p for p in punches if p.get("is_attendance")]
    att_count   = len(att_punches)
    punched     = att_count >= 2

    # Holidays come only from the role-based holiday table (holiday_obj); a Sunday
    # is NOT automatically a holiday unless it is assigned for this role.
    # ── Full holiday ──
    if h_code in ("H", "F"):
        if has_ccl_on_day:
            return "CCL"
        return "H"

    # ── FN holiday ──
    if h_code == "FN":
        has_half_day_presence = att_count >= 1
        # Check for permission in afternoon
        if perm_session == "afternoon":
            return "FN/PER"
        if leave_code:
            return "FN/P" if has_half_day_presence else f"FN/{leave_code}"
        return "FN/P" if has_half_day_presence else "FN/A"

    # ── AN holiday ──
    if h_code == "AN":
        has_half_day_presence = att_count >= 1
        # Check for permission in morning
        if perm_session == "morning":
            return "PER/AN"
        if leave_code:
            return "P/AN" if has_half_day_presence else f"{leave_code}/AN"
        return "P/AN" if has_half_day_presence else "A/AN"

    # ── Normal working day ──
    if _is_flexible_pair_policy(attendance_policy):

        policy_status, _summary = _get_flexible_policy_status(attendance_policy, punches)

        return policy_status


    if att_count == 0:

        punch_status = "A"

    elif att_count == 1:
        punch_status = "SP"
    else:
        punch_status = "P"

    is_present = punch_status == "P"

    # ── Permission (check before leave, regardless of punch status) ──
    if perm_session:
        if perm_session == "morning":
            # Morning permission: PER/P if present, PER/A if absent
            if is_present:
                return "PER/P"
            elif leave_code:
                return f"PER/{leave_code}"
            else:
                return "PER/A"
        elif perm_session == "afternoon":
            # Afternoon permission: P/PER if present, A/PER if absent
            if is_present:
                return "P/PER"
            elif leave_code:
                return f"{leave_code}/PER"
            else:
                return "A/PER"
        else:  # both
            return "PER/PER"

    # ── Leave ──
    if leave_code:
        return f"P/{leave_code}" if is_present else leave_code

    return punch_status


def _safe_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _nearest_shift_start(punch_time, shift_starts):
    """Pick the shift start (a datetime.time) closest to ``punch_time``, handling
    midnight wraparound. Used for rotating shifts so a morning punch is compared
    to the day-shift start and an evening punch to the night-shift start."""
    if not shift_starts:
        return punch_time
    pm = punch_time.hour * 60 + punch_time.minute

    def _dist(s):
        sm = s.hour * 60 + s.minute
        d = abs(pm - sm)
        return min(d, 1440 - d)  # circular distance across midnight

    return min(shift_starts, key=_dist)


FLEXIBLE_PRESENT_STATUS = "P"
FLEXIBLE_SHORT_HOURS_STATUS = "SH"
FLEXIBLE_MISSING_OUT_STATUS = "MO"


def _is_flexible_pair_policy(policy):
    return bool(policy and policy.punch_mode == "PAIR" and not policy.shift_required)


def _policy_skips_time_checks(policy):
    return bool(policy and (not policy.shift_required or not policy.late_early_check))


def _attendance_pair_hours(punches):
    att_punches = sorted(
        [p for p in punches if p.get("is_attendance") and p.get("logdate")],
        key=lambda p: p["logdate"],
    )
    total_seconds = 0
    for index in range(0, len(att_punches) - 1, 2):
        start_dt = att_punches[index]["logdate"]
        end_dt = att_punches[index + 1]["logdate"]
        if end_dt > start_dt:
            total_seconds += (end_dt - start_dt).total_seconds()
    return {
        "punch_count": len(att_punches),
        "total_hours": round(total_seconds / 3600, 2),
        "has_odd_punch": len(att_punches) % 2 == 1,
    }


def _get_flexible_policy_status(policy, punches):
    summary = _attendance_pair_hours(punches)
    punch_count = summary["punch_count"]
    if punch_count == 0:
        return "A", summary
    if summary["has_odd_punch"] and policy.odd_punch_handling == "MISSING_OUT":
        return FLEXIBLE_MISSING_OUT_STATUS, summary
    if punch_count < int(policy.minimum_punches_required or 2):
        return FLEXIBLE_MISSING_OUT_STATUS, summary
    if summary["total_hours"] < float(policy.minimum_working_hours or 0):
        return FLEXIBLE_SHORT_HOURS_STATUS, summary
    return FLEXIBLE_PRESENT_STATUS, summary


def _build_attendance_policy_map(employees):
    employees = list(employees)
    employee_pks = [emp.pk for emp in employees]
    designation_ids = [emp.designation_id for emp in employees if getattr(emp, "designation_id", None)]
    category_ids = [emp.category_id for emp in employees if getattr(emp, "category_id", None)]

    assignments = list(
        AttendancePolicyAssignment.objects
        .select_related("policy", "employee", "designation", "category")
        .filter(is_active=True, policy__is_active=True)
        .filter(
            Q(employee_id__in=employee_pks) |
            Q(designation_id__in=designation_ids) |
            Q(category_id__in=category_ids) |
            Q(is_default=True)
        )
        .order_by("id")
    )

    by_employee = {item.employee_id: item.policy for item in assignments if item.employee_id}
    by_designation = {item.designation_id: item.policy for item in assignments if item.designation_id}
    by_category = {item.category_id: item.policy for item in assignments if item.category_id}
    default_policy = next((item.policy for item in assignments if item.is_default), None)

    policy_map = {}
    for emp in employees:
        policy_map[emp.pk] = (
            by_employee.get(emp.pk)
            or by_designation.get(getattr(emp, "designation_id", None))
            or by_category.get(getattr(emp, "category_id", None))
            or default_policy
        )
    return policy_map

from faculty_leave_management.models import leave_day_fraction


def _overlap_leave_days(leave_app, period_start, period_end):
    if not leave_app.from_date or not leave_app.to_date:
        return 0.0

    overlap_start = max(leave_app.from_date, period_start)
    overlap_end = min(leave_app.to_date, period_end)
    if overlap_start > overlap_end:
        return 0.0

    total = 0.0
    d = overlap_start
    while d <= overlap_end:
        total += leave_day_fraction(
            d,
            leave_app.from_date,
            leave_app.to_date,
            leave_app.session,
            leave_app.to_session,
        )
        d += timedelta(days=1)
    return total


def _holiday_fraction(att_date, holiday_obj):
    # Holidays come only from the role-based holiday table; Sundays are not
    # automatically holidays.
    h_code = (holiday_obj.session_type or "").strip().upper() if holiday_obj else ""
    if h_code in ("H", "F"):
        return 1.0
    if h_code in ("FN", "AN"):
        return 0.5
    return 0.0


def _effective_leave_days(leave_app, period_start, period_end, emp_holidays):
    if not leave_app.from_date or not leave_app.to_date:
        return 0.0

    overlap_start = max(leave_app.from_date, period_start)
    overlap_end = min(leave_app.to_date, period_end)
    if overlap_start > overlap_end:
        return 0.0

    total = 0.0
    d = overlap_start
    while d <= overlap_end:
        holiday_obj = emp_holidays.get(d) if emp_holidays else None
        day_working_fraction = 1.0 - _holiday_fraction(d, holiday_obj)
        leave_fraction = leave_day_fraction(
            d,
            leave_app.from_date,
            leave_app.to_date,
            leave_app.session,
            leave_app.to_session,
        )
        total += leave_fraction * day_working_fraction
        d += timedelta(days=1)
    return total

@check_permission("employee_attandance_report")
def emp_attandance_report(request):
    """
    Two-step attendance report:
    1. If no filters submitted, show landing page with filter form
    2. If filters submitted, show the full report
    """
    from django.db.models import Q, Sum
    from faculty_management.models import FacultyCategory, general_information
    from user_accounts.models import Add_Department
    from faculty_leave_management.models import (
        LeaveApplication, LeaveBalance, LeaveType,
        CCL_Claim, Employee_Holidays, DeviceInfo, ShiftDetail,
    )

    now = datetime.now()

    # Check if this is the initial landing (no filters submitted)
    if request.method == "GET" and not request.GET.get("view_report"):
        # Show landing page with filter form
        departments = Add_Department.objects.all().order_by("Department")
        categories = FacultyCategory.objects.filter(is_active=True).order_by("category_name")

        context = {
            "show_landing": True,
            "departments": departments,
            "categories": categories,
            # Default range: 1st of the current month → today.
            "default_from_date": date(now.year, now.month, 1).strftime("%Y-%m-%d"),
            "default_to_date": now.date().strftime("%Y-%m-%d"),
        }
        return render(request, "faculty_leave_management/admin/emp_attendance_report.html", context)

    # Otherwise, process and show the full report
    dept_id  = request.GET.get("dept_id",  "")
    category_id = request.GET.get("category_id", "")
    search   = (request.GET.get("search") or "").strip()

    # New parameters for entry/exit times and allowed permissions
    try:
        allowed_permissions = float(request.GET.get("allowed_permissions", 2))
    except (ValueError, TypeError):
        allowed_permissions = 2.0

    # Entry/Exit time apply ONLY to employees who have no shift assigned; shift
    # employees are checked against their own shift start/end.
    entry_time_str = request.GET.get("entry_time", "09:00") or "09:00"
    exit_time_str = request.GET.get("exit_time", "16:30") or "16:30"
    try:
        entry_time = datetime.strptime(entry_time_str, "%H:%M").time()
    except (ValueError, TypeError):
        entry_time = datetime.strptime("09:00", "%H:%M").time()
    try:
        exit_time = datetime.strptime(exit_time_str, "%H:%M").time()
    except (ValueError, TypeError):
        exit_time = datetime.strptime("16:30", "%H:%M").time()

    # Report period: user-selected From Date → To Date.
    from_date_str = (request.GET.get("from_date") or "").strip()
    to_date_str   = (request.GET.get("to_date") or "").strip()
    try:
        period_start = datetime.strptime(from_date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        period_start = date(now.year, now.month, 1)
    try:
        period_end = datetime.strptime(to_date_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        period_end = now.date()
    if period_start > period_end:
        period_start, period_end = period_end, period_start

    period_days = []
    cur_d = period_start
    while cur_d <= period_end:
        period_days.append(cur_d)
        cur_d += timedelta(days=1)

    academic_year = _academic_year_for_date(period_end)

    departments  = Add_Department.objects.all().order_by("Department")
    categories = FacultyCategory.objects.filter(is_active=True).order_by("category_name")
    all_leave_types = list(LeaveType.objects.all().order_by("name"))
    leave_type_by_code = {lt.code: lt for lt in all_leave_types if lt.code}
    active_leave_codes = {lt.code for lt in all_leave_types if lt.is_active and lt.code}
    leave_types = []

    # ── Device master — same as punch_attendance ──
    devices_qs = DeviceInfo.objects.filter(is_active=True).order_by("deviceid")
    device_master        = {}
    attendance_device_ids = []
    for d in devices_qs:
        dev_id = str(d.deviceid or "").strip()
        if not dev_id:
            continue
        if d.is_attendance:
            attendance_device_ids.append(dev_id)
        device_master[dev_id] = {
            "is_attendance": d.is_attendance,
            "is_mess":       d.is_mess,
        }
    # If no attendance devices configured, treat all as attendance
    no_device_filter = len(attendance_device_ids) == 0

    report_rows  = []
    error_message = None
    punch_errors  = []
    debug_punch_count = 0

    try:
        employees = (
            general_information.objects
            .select_related("department", "designation", "category", "shift")
            .filter(faculty_id__isnull=False)
            .order_by("faculty_id")
        )
        if dept_id:
            employees = employees.filter(department_id=dept_id)
        if category_id:
            employees = employees.filter(category_id=category_id)
        if search:
            employees = employees.filter(
                Q(name__icontains=search) | Q(faculty_id__icontains=search)
            )

        faculty_ids = []
        for fid in employees.values_list("faculty_id", flat=True):
            try:
                faculty_ids.append(int(fid))
            except (TypeError, ValueError):
                pass  # skip non-integer faculty_ids like names

        if not faculty_ids:
            raise ValueError("No employees found for the selected filters.")

        # ── Query attendance_db ──
        # UserId column in SmartOfficedb is VARCHAR — compare as strings to avoid cast errors
        all_punches = []
        id_list_str = ",".join(f"'{fid}'" for fid in faculty_ids)  # quoted strings
        # Every (month, year) DeviceLogs table spanning the range, starting one day
        # before period_start so a night shift ending on the first day still finds
        # its previous-evening IN punch.
        span_tables = []
        _sm = period_start - timedelta(days=1)
        _sy, _smn = _sm.year, _sm.month
        while (_sy, _smn) <= (period_end.year, period_end.month):
            span_tables.append((_smn, _sy))
            _smn += 1
            if _smn > 12:
                _smn = 1
                _sy += 1
        for tbl_month, tbl_year in span_tables:
            sql = (
                f"SELECT UserId, LogDate, DeviceId "
                f"FROM DeviceLogs_{tbl_month}_{tbl_year} "
                f"WHERE CAST(UserId AS VARCHAR) IN ({id_list_str}) "
                f"ORDER BY LogDate ASC"
            )
            try:
                with connections["attendance_db"].cursor() as cur:
                    cur.execute(sql)
                    cols = [c[0] for c in cur.description]
                    all_punches.extend(dict(zip(cols, row)) for row in cur.fetchall())
            except Exception as ex:
                punch_errors.append(f"DeviceLogs_{tbl_month}_{tbl_year}: {ex}")

        debug_punch_count = len(all_punches)

        from collections import defaultdict as _defaultdict

        # ── Per-employee shift details keyed by faculty_id int ──
        #   shift_by_uid[uid]        = (start, end) of the first shift row
        #   shift_starts_by_uid[uid] = every shift start (day + night) for the late
        #                              check, so the applicable start is chosen per
        #                              day from the punch (see _nearest_shift_start).
        # A rotating shift master (e.g. Security: 07:00-18:00 day + 18:00-07:00
        # night) can't be resolved to one window by time alone, so punches are
        # paired into sessions and credited to the day the shift ends (see below).
        shift_by_uid = {}
        shift_starts_by_uid = {}   # uid -> [start_time, ...] across all shift rows
        night_start_by_uid = {}    # uid -> start_time of the midnight-crossing shift
        shift_ids = {emp.shift_id for emp in employees if emp.shift_id}
        details_by_shift = _defaultdict(list)
        if shift_ids:
            for sd in ShiftDetail.objects.filter(
                shift_master_id__in=shift_ids
            ).order_by("shift_no"):
                if sd.start_time and sd.end_time:
                    details_by_shift[sd.shift_master_id].append((sd.start_time, sd.end_time))
            for emp in employees:
                try:
                    uid_key = int(emp.faculty_id)
                except (TypeError, ValueError):
                    continue
                dets = details_by_shift.get(emp.shift_id) if emp.shift_id else None
                if not dets:
                    continue
                shift_by_uid[uid_key] = dets[0]
                shift_starts_by_uid[uid_key] = [s for s, e in dets]
                # Earliest start among midnight-crossing (night) shift rows.
                night_starts = [s for s, e in dets if e < s]
                if night_starts:
                    night_start_by_uid[uid_key] = min(night_starts)

        # ── Collect raw punches per employee ──
        raw_by_uid = _defaultdict(list)   # uid -> [(log_dt, is_att)]
        for p in all_punches:
            uid    = p.get("UserId")
            log_dt = p.get("LogDate")
            if uid is None or log_dt is None:
                continue
            try:
                uid = int(uid)  # UserId is varchar in SmartOfficedb, cast back to int
            except (ValueError, TypeError):
                continue
            dev_id = str(p.get("DeviceId") or "").strip()
            if no_device_filter:
                is_att = True
            else:
                is_att = device_master.get(dev_id, {}).get("is_attendance", False)
            raw_by_uid[uid].append((log_dt, is_att))

        # ── Build punch_by_emp[uid][date] = [{"logdate":.., "is_attendance":..}] ──
        # Day-only / no-shift employees: plain calendar-day grouping.
        # Night-capable shifts (a row where end < start, e.g. Security 18:00-07:00):
        #   pair each IN->OUT into a session and credit it to the day it ENDS.
        #   A cross-midnight pair counts as a night shift ONLY when the IN is at or
        #   after the night-shift start time — so a day worker's ~17:30 evening OUT
        #   is never mistaken for a night IN. So a night entry on the 25th correctly
        #   counts for the 26th, while day shifts stay on their own calendar day.
        MAX_SESSION = timedelta(hours=15)
        punch_by_emp = {}
        for uid, plist in raw_by_uid.items():
            night_start = night_start_by_uid.get(uid)

            def _add(dkey, log_dt, is_att):
                if dkey < period_start or dkey > period_end:
                    return
                punch_by_emp.setdefault(uid, {}).setdefault(dkey, []).append({
                    "logdate": log_dt, "is_attendance": is_att,
                })

            if night_start is None:
                # Day-only shift or no shift -> calendar grouping.
                for dt, a in plist:
                    _add(dt.date() if hasattr(dt, "date") else dt, dt, a)
                continue

            plist.sort(key=lambda x: x[0])
            att = [dt for dt, a in plist if a]
            non_att = [(dt, a) for dt, a in plist if not a]
            i = 0
            while i < len(att):
                a_in = att[i]
                paired = False
                if i + 1 < len(att):
                    b_out = att[i + 1]
                    gap = b_out - a_in
                    same_day = a_in.date() == b_out.date()
                    if same_day and gap <= MAX_SESSION:
                        # Day-shift session -> stays on its own day.
                        _add(a_in.date(), a_in, True)
                        _add(a_in.date(), b_out, True)
                        paired = True
                    elif (not same_day) and a_in.time() >= night_start and gap <= MAX_SESSION:
                        # Night-shift session -> credit both to the day it ends.
                        d = b_out.date()
                        _add(d, a_in, True)
                        _add(d, b_out, True)
                        paired = True
                if paired:
                    i += 2
                else:
                    _add(a_in.date(), a_in, True)   # lone punch -> its own day
                    i += 1
            for dt, a in non_att:
                _add(dt.date() if hasattr(dt, "date") else dt, dt, a)

        # ── Approved leaves in period ──
        all_leaves = list(
            LeaveApplication.objects
            .select_related("leave_type", "faculty")
            .filter(
                faculty__faculty_id__in=faculty_ids,
                from_date__lte=period_end,
                to_date__gte=period_start,
                status__iexact="Approved",
            )
        )
        leave_by_emp = {}
        for la in all_leaves:
            fid = _safe_int(la.faculty.faculty_id if la.faculty else None)
            if fid:
                leave_by_emp.setdefault(fid, []).append(la)

        # ── Role ID map: faculty_id -> role_id (from USER model, for holiday lookup) ──
        from user_accounts.models import USER
        role_ids_map = {}
        try:
            for u in USER.objects.using("rit_approval_system").filter(
                Employee_id__in=[str(fid) for fid in faculty_ids]
            ).values("Employee_id", "role_id"):
                try:
                    emp_id = int(u["Employee_id"])
                    rid = int(u["role_id"])
                    role_ids_map.setdefault(emp_id, set()).add(rid)
                except (TypeError, ValueError):
                    pass
        except Exception as role_ex:
            punch_errors.append(f"role_id_map: {role_ex}")
        role_id_map = {}
        debug_role_map_size = 0

        # ── Employee_Holidays (Employee_Holidays) — all in period ──
        holiday_map = {}
        all_holiday_role_ids = set()
        for h in Employee_Holidays.objects.filter(
            holiday_date__gte=period_start, holiday_date__lte=period_end
        ).select_related("category"):
            _add_to_holiday_map(holiday_map, h)
            all_holiday_role_ids.add(h.role_id)

        # Resolve one role per employee.
        # If an employee has multiple roles, prefer the role with the most holidays
        # configured in this selected period (prevents Faculty/HOD ambiguity).
        holiday_role_day_count = {rid: len(days) for rid, days in holiday_map.items()}
        for emp_id, role_ids in role_ids_map.items():
            role_ids_sorted = sorted(role_ids)
            if role_ids_sorted:
                role_id_map[emp_id] = sorted(
                    role_ids_sorted,
                    key=lambda rid: (-holiday_role_day_count.get(rid, 0), rid)
                )[0]
        debug_role_map_size = len(role_id_map)

        # ── Leave balances ──
        balance_map = {}
        for b in (
            LeaveBalance.objects
            .filter(faculty__faculty_id__in=faculty_ids, academic_year=academic_year)
            .values("faculty__faculty_id", "leave_type__code")
            .annotate(available=Sum("available"), used=Sum("used"))
        ):
            fid = _safe_int(b["faculty__faculty_id"])
            if fid is None:
                continue
            balance_map.setdefault(fid, {})[b["leave_type__code"]] = {
                "available": b["available"] or 0,
                "used":      b["used"]      or 0,
            }

        # ── CCL balance ──
        ccl_map = {}
        for r in (
            CCL_Claim.objects
            .filter(faculty__faculty_id__in=faculty_ids, academic_year=academic_year)
            .values("faculty__faculty_id")
            .annotate(claimed=Sum("claimed"), used=Sum("used"))
        ):
            fid = _safe_int(r["faculty__faculty_id"])
            if fid is not None:
                ccl_map[fid] = r

        # ── emp_pk_map: general_information.pk → faculty_id (avoids cross-FK JOIN) ──
        emp_pk_map = {}
        for _emp in employees:
            _fid = _safe_int(getattr(_emp, "faculty_id", None))
            if _fid is not None:
                emp_pk_map[_emp.pk] = _fid

        # ── CCL approved dates (Sundays/holidays worked) ──
        # Query by PK directly — avoids JOIN that may fail if models span databases.
        from faculty_leave_management.models import CCL_Application, PermissionRequest
        ccl_dates_map = {}
        ccl_days_map = {}
        for ca in CCL_Application.objects.filter(
            faculty_id__in=list(emp_pk_map.keys()),
            date__gte=period_start,
            date__lte=period_end,
            status__iexact="Approved",
        ):
            fid = emp_pk_map.get(ca.faculty_id)
            if fid and ca.date:
                ccl_dates_map.setdefault(fid, set()).add(ca.date)
                ccl_days_map[fid] = ccl_days_map.get(fid, 0.0) + float(ca.days or 1)

        # ── Permission per employee per date in period ──
        # perm_map[fid][date] = "morning" | "afternoon" | "both"
        perm_map = {}
        perm_count_map = {}  # fid -> permission count
        permission_event_ids = {}  # fid -> unique permission records
        from datetime import time as dt_time
        NOON = dt_time(12, 0)

        permissions_qs = (
            PermissionRequest.objects
            .filter(
                Q(faculty__faculty_id__in=faculty_ids) | Q(faculty__isnull=True, user_id__isnull=False),
                date__gte=period_start,
                date__lte=period_end,
                status__iexact="Approved",
            )
            .select_related("faculty")
        )
        permissions = list(permissions_qs)

        permission_user_ids = {pr.user_id for pr in permissions if not pr.faculty_id and pr.user_id}
        permission_user_emp_map = {}
        if permission_user_ids:
            from user_accounts.models import USER
            # Prefer default DB when available; fallback to rit_approval_system.
            for alias in (None, "rit_approval_system"):
                pending_ids = [uid for uid in permission_user_ids if uid not in permission_user_emp_map]
                if not pending_ids:
                    break
                try:
                    qs = USER.objects.filter(id__in=pending_ids)
                    if alias:
                        qs = USER.objects.using(alias).filter(id__in=pending_ids)
                    for u in qs.values("id", "Employee_id"):
                        uid = _safe_int(u.get("id"))
                        emp_id = _safe_int(u.get("Employee_id"))
                        if uid is not None and emp_id is not None:
                            permission_user_emp_map[uid] = emp_id
                except Exception:
                    continue

        for pr in permissions:
            fid = _safe_int(pr.faculty.faculty_id if pr.faculty else None)
            if fid is None:
                fid = permission_user_emp_map.get(pr.user_id)
            if not fid or not pr.date:
                continue
            permission_event_ids.setdefault(fid, set()).add(f"PR:{pr.id}")
            session = "morning" if (pr.from_time and pr.from_time < NOON) else "afternoon"
            existing = perm_map.setdefault(fid, {}).get(pr.date)
            if existing and existing != session:
                perm_map[fid][pr.date] = "both"
            else:
                perm_map[fid][pr.date] = session

        # ── Leave days used per type per employee in period ──
        # leave_used_map[fid][leave_code] = total days used (decimal)
        leave_used_map = {}
        leave_codes_in_use = set()
        emp_designation_map = {}
        emp_category_map = {}
        for emp in employees:
            fid_emp = _safe_int(getattr(emp, "faculty_id", None))
            desig_emp = _safe_int(getattr(emp, "designation_id", None))
            if fid_emp is not None:
                emp_designation_map[fid_emp] = desig_emp
                emp_category_map[fid_emp] = getattr(emp, "category_id", None)
        for la in all_leaves:
            fid = _safe_int(la.faculty.faculty_id if la.faculty else None)
            if not fid:
                continue
            code = ((la.leave_type.code if la.leave_type else "L") or "L").strip().upper()
            role_id = role_id_map.get(fid) or emp_designation_map.get(fid)
            emp_holidays = _holiday_map_for_employee(
                holiday_map,
                role_id,
                emp_category_map.get(fid),
            )
            days_val = _effective_leave_days(la, period_start, period_end, emp_holidays)
            if days_val <= 0:
                continue

            # Permission leaves (PER) should be counted in Permission column,
            # not as a leave bucket column in this report grid.
            if code == "PER":
                permission_event_ids.setdefault(fid, set()).add(f"LA:{la.id}")
                # Extract session from LeaveApplication.session for permission display
                if la.session and la.session.session_name:
                    session_name = la.session.session_name.strip().upper()
                    # Map session_name to morning/afternoon
                    if session_name in ("FN", "MORNING"):
                        perm_session_val = "morning"
                    elif session_name in ("AN", "AFTERNOON"):
                        perm_session_val = "afternoon"
                    elif session_name in ("FULL DAY", "FULL", "BOTH"):
                        perm_session_val = "both"
                    else:
                        perm_session_val = "morning" if (la.session.start_time and la.session.start_time < NOON) else "afternoon"

                    # Apply to all dates in the leave range
                    d = max(la.from_date, period_start)
                    end_d = min(la.to_date, period_end)
                    while d <= end_d:
                        existing = perm_map.setdefault(fid, {}).get(d)
                        if existing and existing != perm_session_val:
                            perm_map[fid][d] = "both"
                        else:
                            perm_map[fid][d] = perm_session_val
                        d += timedelta(days=1)
                continue

            leave_used_map.setdefault(fid, {})
            leave_used_map[fid][code] = leave_used_map[fid].get(code, 0.0) + days_val
            leave_codes_in_use.add(code)

        for fid, ccl_days in ccl_days_map.items():
            if ccl_days <= 0:
                continue
            leave_used_map.setdefault(fid, {})
            leave_used_map[fid]["CCL"] = leave_used_map[fid].get("CCL", 0.0) + ccl_days
            leave_codes_in_use.add("CCL")

        preferred_leave_order = ["CL", "CCL", "OD", "LOP", "AB", "VL", "ML"]
        preferred_index = {code: idx for idx, code in enumerate(preferred_leave_order)}
        visible_leave_codes = set(preferred_leave_order)

        for fid, events in permission_event_ids.items():
            perm_count_map[fid] = float(len(events))

        leave_types = []
        for code in sorted(visible_leave_codes, key=lambda c: (preferred_index.get(c, 999), c)):
            lt_obj = leave_type_by_code.get(code)
            leave_types.append({
                "code": code,
                "name": lt_obj.name if lt_obj else code,
            })

        # ── Shift map ──
        shift_map = {}
        for emp in employees:
            if emp.shift_id:
                sd = ShiftDetail.objects.filter(
                    shift_master_id=emp.shift_id
                ).order_by("shift_no").first()
                if sd:
                    shift_map[emp.faculty_id] = (sd.start_time, sd.end_time)

        policy_map = _build_attendance_policy_map(employees)

        # ── Build rows ──
        for emp in employees:
            try:
                fid = int(emp.faculty_id)
            except (TypeError, ValueError):
                continue
            role_id = role_id_map.get(fid)           # Role from approval USER table
            if role_id is None and getattr(emp, "designation_id", None) is not None:
                # Fallback for cases where USER-role mapping is unavailable/missing.
                role_id = int(emp.designation_id)
            emp_punches   = punch_by_emp.get(fid, {})
            emp_leaves    = leave_by_emp.get(fid, [])
            emp_holidays  = _holiday_map_for_employee(
                holiday_map,
                role_id,
                getattr(emp, "category_id", None),
            )
            emp_ccl_dates = ccl_dates_map.get(fid, set())
            emp_perm_dates = perm_map.get(fid, {})
            attendance_policy = policy_map.get(emp.pk)
            standard_leave_codes = {"CL", "CCL", "OD", "LOP", "AB", "VL", "ML", "WL"}

            # ── Per-employee shift start times for the late check ──
            # Use the employee's own shift start(s) when a shift is assigned,
            # otherwise fall back to the global Entry Time filter. For a rotating
            # shift the applicable start is chosen per day from the day's first
            # punch (see _nearest_shift_start), so a day shift is checked against
            # its day start and a night shift against its night start.
            emp_shift_starts = shift_starts_by_uid.get(fid) or [entry_time]

            # ── Track late entries for permission limit ──
            late_entry_count = 0
            late_entry_dates = []

            # First pass: identify late entries
            for att_date in period_days:
                punches_today = emp_punches.get(att_date, [])
                holiday_obj   = emp_holidays.get(att_date)
                leaves_today  = [la for la in emp_leaves
                                  if la.from_date <= att_date <= la.to_date]

                # Skip on-leave/holiday/permission days and days whose policy
                # disables time checks.
                if leaves_today or holiday_obj or emp_perm_dates.get(att_date) or _policy_skips_time_checks(attendance_policy):
                    continue

                # Late entry only applies to a full (2-punch) day. A single-punch
                # day stays "SP" — a lone late punch is not a late *arrival*.
                att_punches = [p for p in punches_today if p.get("is_attendance")]
                if len(att_punches) >= 2:
                    first_punch_dt = None
                    for p in att_punches:
                        log_dt = p.get("logdate")
                        if log_dt and hasattr(log_dt, "time"):
                            if first_punch_dt is None or log_dt < first_punch_dt:
                                first_punch_dt = log_dt

                    if first_punch_dt:
                        applicable_start = _nearest_shift_start(
                            first_punch_dt.time(), emp_shift_starts
                        )
                        # 1-minute grace, same as the Punch Details page.
                        start_dt = datetime.combine(first_punch_dt.date(), applicable_start)
                        if first_punch_dt > start_dt + timedelta(minutes=1):
                            late_entry_dates.append(att_date)

            day_cells = []
            present_count = 0.0
            ab_count = 0.0
            holidays_count = 0.0
            working_days = 0.0
            late_permission_count = 0  # Track late permissions used

            for att_date in period_days:
                # Future dates haven't happened yet — show blank, don't count as
                # present/absent/working.
                if att_date > now.date():
                    day_cells.append({"date": att_date, "status": ""})
                    continue

                punches_today = emp_punches.get(att_date, [])
                holiday_obj   = emp_holidays.get(att_date)
                leaves_today  = [la for la in emp_leaves
                                  if la.from_date <= att_date <= la.to_date]

                # Check if this is a late entry that should be marked as permission or absent
                is_late_entry = att_date in late_entry_dates
                if is_late_entry:
                    if late_entry_count < allowed_permissions:
                        # Within allowed permissions - mark as late permission (PER/P format)
                        late_entry_count += 1
                        late_permission_count += 1
                        status = "PER/P"  # Show as permission + present
                    else:
                        # Exceeded allowed permissions - mark as absent
                        status = "A"
                else:
                    status = _get_day_status(att_date, punches_today, leaves_today, holiday_obj, emp_ccl_dates, emp_perm_dates.get(att_date), attendance_policy)

                # Display transform: FN/ -> H/, /AN -> /H
                if status and status.upper().startswith("FN/"):
                    display_status = "H/" + status[3:]
                elif status and status.upper().endswith("/AN"):
                    display_status = status[:-3] + "/H"
                else:
                    display_status = status

                day_cells.append({"date": att_date, "status": display_status})

                s = (status or "").strip().upper()
                h_code = (holiday_obj.session_type or "").strip().upper() if holiday_obj else ""

                # Holiday is driven ONLY by the role-based holiday table (F/H=1,
                # FN/AN=0.5). Sundays are not automatically holidays.
                # CCL means employee worked on this holiday — treat as working day.
                is_ccl_day = (s == "CCL")
                if is_ccl_day:
                    holiday_fraction = 0.0
                elif h_code in ("H", "F"):
                    holiday_fraction = 1.0
                elif h_code in ("FN", "AN"):
                    holiday_fraction = 0.5
                else:
                    holiday_fraction = 0.0

                holidays_count += holiday_fraction
                working_days += (1.0 - holiday_fraction)

                if s == "CCL":
                    present_count += 1
                elif s == "H" or s == "S":
                    pass
                elif s.startswith("FN/"):
                    right = s.split("/", 1)[1].strip().upper()
                    if right == "P":
                        present_count += 0.5
                    elif right in standard_leave_codes:
                        pass
                    else:
                        ab_count += 0.5
                elif s.endswith("/AN"):
                    left = s.split("/", 1)[0].strip().upper()
                    if left == "P":
                        present_count += 0.5
                    elif left in standard_leave_codes:
                        pass
                    else:
                        ab_count += 0.5
                elif s == "A":
                    ab_count += 1
                elif s == "P":
                    present_count += 1
                elif s == "SP":
                    ab_count += 1
                elif s in (FLEXIBLE_MISSING_OUT_STATUS, FLEXIBLE_SHORT_HOURS_STATUS):
                    ab_count += 1
                elif "/" in s:
                    if "P" in s:
                        left, right = s.split("/", 1)
                        left = left.strip().upper()
                        right = right.strip().upper()

                        # Half-day leave + half-day present should contribute only 0.5 present
                        # (e.g., P/CL or CL/P).
                        if (left == "P" and right in standard_leave_codes) or (right == "P" and left in standard_leave_codes):
                            present_count += 0.5
                        else:
                            present_count += 1.0
                    else:
                        left, right = s.split("/", 1)
                        left = left.strip().upper()
                        right = right.strip().upper()
                        if left in standard_leave_codes or right in standard_leave_codes:
                            pass
                        else:
                            ab_count += 1
                else:
                    # Bare leave code (CL, ML, VL...) is leave and not AB.
                    pass

            lu = leave_used_map.get(fid, {})
            lu["AB"] = float(ab_count)

            # Metric definitions for report summary
            cl_days = float(lu.get("CL", 0) or 0)
            ccl_days = float(lu.get("CCL", 0) or 0)
            od_days = float(lu.get("OD", 0) or 0)
            vl_days = float(lu.get("VL", 0) or 0)
            ml_days = float(lu.get("ML", 0) or 0)
            lop_days = float(lu.get("LOP", 0) or 0)
            ab_days = float(lu.get("AB", 0) or 0)
            wl_days = float(lu.get("WL", 0) or 0)

            # Worked = CL + CCL + WL + OD + VL + ML + Present + Holiday
            # AB and LOP are NOT subtracted because:
            # - They are already excluded from Present count
            # - Present count only includes days marked as "P" (present)
            # - Worked = paid leaves + present days + holidays
            # - AB and LOP are tracked separately in their own columns
            worked = cl_days + ccl_days + wl_days + od_days + vl_days + ml_days + present_count + holidays_count
            worked = max(0.0, worked)

            # Add late permissions to total permission count
            total_permissions = perm_count_map.get(fid, 0) + late_permission_count

            report_rows.append({
                "emp":          emp,
                "policy_name":  attendance_policy.policy_name if attendance_policy else "Shift",
                "day_cells":    day_cells,
                "present":      present_count,
                "absent":       wl_days,
                "holidays":     holidays_count,
                "working_days": working_days,
                "worked":       worked,
                "permission":   total_permissions,
                "leave_used":   lu,
                "balances":     balance_map.get(fid, {}),
                "ccl_claimed":  ccl_map.get(fid, {}).get("claimed", 0),
                "ccl_used":     ccl_map.get(fid, {}).get("used", 0),
                "ccl_remaining": (ccl_map.get(fid, {}).get("claimed", 0) or 0) - (ccl_map.get(fid, {}).get("used", 0) or 0),
            })

    except Exception as e:
        import traceback
        error_message = traceback.format_exc()

    context = {
        "show_landing":       False,
        "report_rows":        report_rows,
        "period_days":        period_days,
        "leave_types":        leave_types,
        "departments":        departments,
        "categories":         categories,
        "selected_from_date": period_start.strftime("%Y-%m-%d"),
        "selected_to_date":   period_end.strftime("%Y-%m-%d"),
        "selected_dept":      dept_id,
        "selected_category":  category_id,
        "search":             search,
        "allowed_permissions": allowed_permissions,
        "entry_time":         entry_time_str,
        "exit_time":          exit_time_str,
        "period_start":       period_start,
        "period_end":         period_end,
        "academic_year":      academic_year,
        "error_message":      error_message,
        "punch_errors":       punch_errors,
        "debug_punch_count":  debug_punch_count,
        "debug_role_map_size": len(role_id_map) if 'role_id_map' in dir() else 0,
        "debug_holiday_role_ids": sorted(all_holiday_role_ids) if 'all_holiday_role_ids' in dir() else [],
        "debug_role_id_map":  dict(list(role_id_map.items())[:10]) if 'role_id_map' in dir() else {},
    }
    return render(request, "faculty_leave_management/admin/emp_attendance_report.html", context)

def _get_attendance_device_context():
    attendance_device_ids = []
    device_map = {}

    for device in DeviceInfo.objects.filter(is_active=True, is_attendance=True).order_by("deviceid"):
        device_id = str(device.deviceid or "").strip()
        if not device_id:
            continue
        attendance_device_ids.append(device_id)
        device_map[device_id] = device

    return attendance_device_ids, device_map


def _attendance_table_exists(table_name):
    with connections["attendance_db"].cursor() as cursor:
        cursor.execute(
            "SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = %s",
            [table_name],
        )
        return cursor.fetchone() is not None


def _get_faculty_single_punch_rows(faculty_id, year, month):
    """Return every single-punch day for one faculty within a given month.

    Mirrors `_get_single_punch_rows` but is scoped to a single UserId and a
    whole month, grouping punches by calendar day and keeping only the days
    that have exactly one punch on an attendance device.
    """
    attendance_device_ids, device_map = _get_attendance_device_context()
    if not attendance_device_ids:
        return [], {}, "No active attendance devices are configured."

    table_name = f"DeviceLogs_{month}_{year}"
    if not _attendance_table_exists(table_name):
        return [], {}, f"Attendance table '{table_name}' was not found."

    placeholders = ",".join(["%s"] * len(attendance_device_ids))
    query = f"""
        SELECT *
        FROM {table_name}
        WHERE UserId = %s
          AND DeviceId IN ({placeholders})
        ORDER BY LogDate ASC
    """

    with connections["attendance_db"].cursor() as cursor:
        cursor.execute(query, [str(faculty_id)] + attendance_device_ids)
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    grouped_rows = defaultdict(list)
    for row in rows:
        log_date = row.get("LogDate")
        if not log_date:
            continue
        day_key = log_date.date() if hasattr(log_date, "date") else log_date
        grouped_rows[day_key].append(row)

    single_punch_rows = []
    source_row_map = {}

    for day in sorted(grouped_rows.keys()):
        day_rows = grouped_rows[day]
        if len(day_rows) != 1:
            continue

        row = day_rows[0]
        device_id = str(row.get("DeviceId") or "").strip()
        device = device_map.get(device_id)
        day_iso = day.isoformat()

        single_punch_rows.append({
            "log_date": day,
            "log_date_iso": day_iso,
            "first_punch": row.get("LogDate"),
            "device_id": device_id or "-",
            "device_location": (
                device.devicelocation if device and device.devicelocation else "-"
            ),
        })
        source_row_map[day_iso] = row

    return single_punch_rows, source_row_map, None


@check_permission("punch_entry")
def punch_entry(request):
    today = date.today()

    faculty_id_str = (request.POST.get("faculty_id") or request.GET.get("faculty_id") or "").strip()
    selected_date_str = (request.POST.get("punch_date") or request.GET.get("punch_date") or today.isoformat()).strip()
    try:
        selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
    except ValueError:
        messages.error(request, "Please choose a valid punch date.")
        selected_date = today

    selected_year = selected_date.year
    selected_month = selected_date.month

    faculty_id = None
    if faculty_id_str:
        try:
            faculty_id = int(faculty_id_str)
        except ValueError:
            messages.error(request, "Please select a valid faculty.")
            faculty_id = None

    def _redirect_back():
        params = {"punch_date": selected_date.isoformat()}
        if faculty_id:
            params["faculty_id"] = faculty_id
        return redirect(f"{request.path}?{urlencode(params)}")

    if request.method == "POST":
        if not faculty_id:
            messages.error(request, "Please select a faculty before saving.")
            return _redirect_back()

        if not general_information.objects.filter(faculty_id=faculty_id).exists():
            messages.error(request, "The selected faculty could not be found.")
            return _redirect_back()

        table_name = f"DeviceLogs_{selected_month}_{selected_year}"
        if not _attendance_table_exists(table_name):
            messages.error(request, f"Attendance table '{table_name}' was not found.")
            return _redirect_back()

        attendance_device_ids, _ = _get_attendance_device_context()
        attendance_device_ids = list(dict.fromkeys(attendance_device_ids + [MANUAL_PUNCH_DEVICE_ID]))
        device_id = MANUAL_PUNCH_DEVICE_ID

        punch_time_str = (request.POST.get("punch_time") or "").strip()
        if not punch_time_str:
            messages.error(request, "Please enter the punch time.")
            return _redirect_back()

        try:
            punch_time = datetime.strptime(punch_time_str, "%H:%M").time()
        except ValueError:
            messages.error(request, "Please enter a valid punch time.")
            return _redirect_back()

        punch_dt = datetime.combine(selected_date, punch_time)

        try:
            with connections["attendance_db"].cursor() as cursor:
                saved_count, skipped_count = _save_manual_punch(
                    cursor,
                    table_name,
                    faculty_id,
                    device_id,
                    punch_dt,
                    attendance_device_ids,
                )
        except Exception as exc:
            messages.error(request, f"Unable to save punch entry: {exc}")
            return _redirect_back()

        if saved_count:
            messages.success(request, f"{saved_count} punch record(s) saved successfully.")
        if skipped_count:
            messages.warning(request, f"{skipped_count} punch record(s) already existed and were not duplicated.")
        if not saved_count and not skipped_count:
            messages.info(request, "No new punch records were saved.")

        return _redirect_back()

    faculty_options = list(
        general_information.objects
        .filter(faculty_id__isnull=False)
        .order_by("faculty_id")
        .values("faculty_id", "name")
    )

    _, device_map = _get_attendance_device_context()
    device_options = [
        {
            "device_id": device_id,
            "location": device.devicelocation or device_id,
        }
        for device_id, device in device_map.items()
    ]

    selected_faculty = None
    punch_detail_rows = []
    selected_date_punches = []
    punch_summary = _empty_punch_entry_summary()
    error_message = None

    if faculty_id:
        selected_faculty = (
            general_information.objects
            .filter(faculty_id=faculty_id)
            .select_related("department", "designation", "shift")
            .first()
        )
        if not selected_faculty:
            error_message = "The selected faculty could not be found."
        else:
            punch_detail_rows, selected_date_punches, punch_summary, error_message = _get_employee_punch_entry_details(
                selected_faculty,
                selected_year,
                selected_month,
                selected_date,
            )

    context = {
        "faculty_options": faculty_options,
        "selected_faculty_id": faculty_id,
        "selected_faculty": selected_faculty,
        "selected_date": selected_date,
        "selected_year": selected_year,
        "selected_month": selected_month,
        "month_label": calendar.month_name[selected_month],
        "device_options": device_options,
        "default_device_id": device_options[0]["device_id"] if device_options else "",
        "manual_punch_device_id": MANUAL_PUNCH_DEVICE_ID,
        "punch_detail_rows": punch_detail_rows,
        "selected_date_punches": selected_date_punches,
        "punch_summary": punch_summary,
        "error_message": error_message,
    }
    return render(request, "faculty_leave_management/admin/punch_entry.html", context)


def _save_manual_punch(cursor, table_name, faculty_id, device_id, punch_dt, attendance_device_ids):
    cursor.execute(
        f"""
        SELECT COUNT(*)
        FROM {table_name}
        WHERE UserId = %s
          AND DeviceId = %s
          AND LogDate = %s
        """,
        [str(faculty_id), device_id, punch_dt],
    )
    if cursor.fetchone()[0]:
        return 0, 1

    day_start = datetime.combine(punch_dt.date(), datetime.min.time())
    day_end = day_start + timedelta(days=1)
    placeholders = ",".join(["%s"] * len(attendance_device_ids))
    cursor.execute(
        f"""
        SELECT LogDate
        FROM {table_name}
        WHERE UserId = %s
          AND DeviceId IN ({placeholders})
          AND LogDate >= %s
          AND LogDate < %s
        ORDER BY LogDate ASC
        """,
        [str(faculty_id)] + attendance_device_ids + [day_start, day_end],
    )
    existing_punches = [
        row[0]
        for row in cursor.fetchall()
        if row and row[0]
    ]
    ordered_punches = sorted(existing_punches + [punch_dt])
    direction = "in" if ordered_punches.index(punch_dt) % 2 == 0 else "out"

    cursor.execute(
        f"""
        INSERT INTO {table_name} (
            DownloadDate,
            DeviceId,
            UserId,
            LogDate,
            Direction,
            AttDirection,
            StatusCode,
            Remarks,
            VerificationMode,
            IsApproved,
            AttenndanceMarkingType,
            APIResponseText,
            sIOMode
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        [
            punch_dt,
            device_id,
            str(faculty_id),
            punch_dt,
            direction,
            " ",
            "0",
            "Manual punch entry",
            "1073741824",
            1,
            "Biometric",
            "Manual punch entry",
            "",
        ],
    )
    return 1, 0


def _get_employee_punch_entry_details(selected_employee, year, month, selected_date):
    from faculty_leave_management.models import CCL_Application, Employee_Holidays, LeaveApplication
    from user_accounts.models import USER

    faculty_id = selected_employee.faculty_id
    employee_category_id = getattr(selected_employee, "category_id", None)
    attendance_device_ids, device_map = _get_attendance_device_context()
    attendance_device_ids = list(dict.fromkeys(attendance_device_ids + [MANUAL_PUNCH_DEVICE_ID]))

    table_name = f"DeviceLogs_{month}_{year}"
    if not _attendance_table_exists(table_name):
        return [], [], _empty_punch_entry_summary(), f"Attendance table '{table_name}' was not found."

    placeholders = ",".join(["%s"] * len(attendance_device_ids))
    query = f"""
        SELECT DeviceLogId, UserId, LogDate, Direction, DeviceId
        FROM {table_name}
        WHERE UserId = %s
          AND DeviceId IN ({placeholders})
        ORDER BY LogDate ASC
    """

    with connections["attendance_db"].cursor() as cursor:
        cursor.execute(query, [str(faculty_id)] + attendance_device_ids)
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    grouped = defaultdict(list)
    selected_date_punches = []
    for row in rows:
        log_dt = row.get("LogDate")
        if not log_dt:
            continue
        log_day = log_dt.date() if hasattr(log_dt, "date") else log_dt
        device_id = str(row.get("DeviceId") or "").strip()
        device = device_map.get(device_id)
        item = {
            "device_log_id": row.get("DeviceLogId"),
            "logdate": log_dt,
            "time": log_dt.strftime("%H:%M:%S") if hasattr(log_dt, "strftime") else log_dt,
            "direction": row.get("Direction") or "",
            "device_id": device_id,
            "device_location": device.devicelocation if device and device.devicelocation else device_id,
        }
        grouped[log_day].append(item)
        if log_day == selected_date:
            selected_date_punches.append(item)

    shift_start_time = None
    if selected_employee.shift:
        first_shift = ShiftDetail.objects.filter(
            shift_master=selected_employee.shift
        ).order_by("shift_no").first()
        if first_shift:
            shift_start_time = first_shift.start_time
    if not shift_start_time:
        from datetime import time as dtime
        shift_start_time = dtime(9, 0)

    month_start = date(year, month, 1)
    month_end = date(year, month, calendar.monthrange(year, month)[1])
    leave_dates = {}
    try:
        for leave in LeaveApplication.objects.filter(
            faculty=selected_employee,
            status="Approved",
            from_date__lte=month_end,
            to_date__gte=month_start,
        ).select_related("leave_type"):
            cur_day = leave.from_date
            while cur_day <= leave.to_date:
                if month_start <= cur_day <= month_end:
                    leave_dates[cur_day] = (
                        leave.leave_type.code if leave.leave_type else "LEAVE"
                    ).strip().upper()
                cur_day += timedelta(days=1)
    except Exception:
        pass

    try:
        for ccl in CCL_Application.objects.filter(
            faculty=selected_employee,
            status__iexact="Approved",
            date__gte=month_start,
            date__lte=month_end,
        ):
            if ccl.date and ccl.date not in leave_dates:
                leave_dates[ccl.date] = "CCL"
    except Exception:
        pass

    user_role_id = None
    try:
        role_rows = (
            USER.objects.using("rit_approval_system")
            .filter(Employee_id=str(faculty_id))
            .values_list("role_id", flat=True)
        )
        candidate_role_ids = {int(role_id) for role_id in role_rows if role_id}
        if candidate_role_ids:
            role_counts = {
                role_id: _filter_holidays_for_category(
                    Employee_Holidays.objects.filter(
                        role_id=role_id,
                        holiday_date__year=year,
                        holiday_date__month=month,
                    ),
                    employee_category_id,
                ).count()
                for role_id in candidate_role_ids
            }
            user_role_id = sorted(candidate_role_ids, key=lambda role_id: (-role_counts.get(role_id, 0), role_id))[0]
    except Exception:
        pass

    detail_rows = []
    days_in_month = calendar.monthrange(year, month)[1]
    for day_num in range(1, days_in_month + 1):
        att_date = date(year, month, day_num)
        day_punches = sorted(grouped.get(att_date, []), key=lambda item: item["logdate"])
        in_punch = day_punches[0] if day_punches else None
        out_punch = day_punches[-1] if len(day_punches) >= 2 else None
        duration = ""
        if in_punch and out_punch and out_punch["logdate"] > in_punch["logdate"]:
            delta = out_punch["logdate"] - in_punch["logdate"]
            seconds = int(delta.total_seconds())
            hours, remainder = divmod(seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            duration = f"{hours:02d}:{minutes:02d}:{seconds:02d}"

        if len(day_punches) >= 2:
            status = "Present"
        elif len(day_punches) == 1:
            status = "Single Punch"
        else:
            status = "No Punch"

        is_weekend = att_date.weekday() >= 5
        is_holiday = is_weekend
        if _get_matching_holiday(user_role_id, att_date, employee_category_id):
            is_holiday = True

        leave_code = leave_dates.get(att_date, "")
        is_late = False
        if in_punch:
            shift_start_dt = datetime.combine(att_date, shift_start_time)
            is_late = in_punch["logdate"] > shift_start_dt + timedelta(minutes=1)
            in_remark = "Late" if is_late else "Ok"
        else:
            in_remark = "No Punch"
        out_remark = "Ok" if out_punch else "No Punch"
        status_remark = "Holiday" if is_holiday else (f"{leave_code} Approved" if leave_code else "")

        detail_rows.append({
            "date": att_date,
            "date_str": att_date.strftime("%d-%m-%Y"),
            "day_name": calendar.day_name[att_date.weekday()],
            "in_time": in_punch["time"] if in_punch else "",
            "out_time": out_punch["time"] if out_punch else "",
            "duration": duration,
            "in_remark": in_remark,
            "out_remark": out_remark,
            "status_remark": status_remark,
            "leave_code": leave_code,
            "is_late": is_late,
            "is_holiday": is_holiday,
            "is_weekend": is_weekend,
            "status": status,
            "total_punches": len(day_punches),
            "is_selected": att_date == selected_date,
        })

    summary = {
        "total": len(detail_rows),
        "present": sum(1 for row in detail_rows if row["in_time"] and row["out_time"] and not row["is_holiday"] and not row["leave_code"]),
        "absent": sum(1 for row in detail_rows if not row["in_time"] and not row["is_holiday"] and not row["leave_code"]),
        "late": sum(1 for row in detail_rows if row["is_late"]),
        "holiday": sum(1 for row in detail_rows if row["is_holiday"]),
        "leave": sum(1 for row in detail_rows if row["leave_code"]),
        "single": sum(1 for row in detail_rows if row["status"] == "Single Punch"),
    }
    return detail_rows, selected_date_punches, summary, None


def _empty_punch_entry_summary():
    return {
        "total": 0,
        "present": 0,
        "absent": 0,
        "late": 0,
        "holiday": 0,
        "leave": 0,
        "single": 0,
    }






@login_required
def fix_approved_ccl_approvers(request):
    """
    One-time fix: Update all approvers for CCL/Leave applications that are marked as 'Approved'
    but have approvers still in 'Pending' status.
    """
    from faculty_leave_management.models import (
        CCL_Application, CCL_Approvers_Data,
        LeaveApplication, LeaveApproversData
    )
    from django.utils import timezone
    from django.contrib import messages
    from django.shortcuts import redirect

    if request.method != "POST":
        from django.http import HttpResponseNotAllowed
        return HttpResponseNotAllowed(["POST"])

    ccl_fixed = 0
    leave_fixed = 0

    try:
        # Fix CCL Applications
        approved_ccls = CCL_Application.objects.filter(status="Approved")
        for ccl in approved_ccls:
            pending_approvers = CCL_Approvers_Data.objects.filter(
                ccl_application=ccl,
                status="Pending"
            )
            if pending_approvers.exists():
                for approver in pending_approvers:
                    approver.status = "Approved"
                    approver.action_date = timezone.now()
                    approver.remarks = (approver.remarks or "") + " [Auto-approved - Fix]"
                    approver.save()
                ccl_fixed += 1

        # Fix Leave Applications
        approved_leaves = LeaveApplication.objects.filter(status="Approved")
        for leave in approved_leaves:
            pending_approvers = LeaveApproversData.objects.filter(
                leave_application=leave,
                status="PENDING"
            )
            if pending_approvers.exists():
                for approver in pending_approvers:
                    approver.status = "APPROVED"
                    approver.approved_date = timezone.now()
                    approver.reason = (approver.reason or "") + " [Auto-approved - Fix]"
                    approver.save()
                leave_fixed += 1

        messages.success(
            request,
            f"Fixed approver chains: {ccl_fixed} CCL applications, {leave_fixed} Leave applications"
        )
    except Exception as e:
        messages.error(request, f"Error fixing approvers: {str(e)}")

    return redirect("employee_attandance_report")


def sync_ccl_from_erp(request):
    """
    Syncs from erp_rit DB for faculty already present in general_information.
      1. `faculty_leave` table -> LeaveApplication, including leave_type='CCL'
      2. `employee_leave`      -> LeaveApplication, including leave_type='CCL'
      3. `ccl` table           -> CCL_Application for worked/earned CCL dates
    Dedup by (faculty_pk, date) for CCL and (faculty_pk, from_date, to_date, leave_type) for leaves.
    """
    from faculty_management.models import general_information
    from faculty_leave_management.models import (
        CCL_Application, CCL_Claim, LeaveApplication, LeaveType, LeaveBalance,
    )
    from user_accounts.models import USER
    from datetime import datetime as _dt, date as _date, timedelta
    from django.utils import timezone

    if request.method != "POST":
        from django.http import HttpResponseNotAllowed
        return HttpResponseNotAllowed(["POST"])

    ccl_synced    = 0
    ccl_updated   = 0
    ccl_skipped   = 0
    ccl_existing_skipped = 0
    leave_synced  = 0
    leave_updated = 0
    leave_skipped = 0
    leave_existing_skipped = 0
    invalid_rows = 0
    missing_leave_type = 0
    unmatched_faculty_ids = set()
    errors = []
    diag  = {}

    def _parse_date(raw):
        if not raw:
            return None
        if isinstance(raw, _date):
            return raw
        if hasattr(raw, "date") and callable(raw.date):
            return raw.date()
        text = str(raw).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return _dt.strptime(text[:10], fmt).date()
            except ValueError:
                continue
        return None

    def _to_int(v):
        try:
            return int(str(v).strip())
        except (TypeError, ValueError):
            return None

    def _ay(dt):
        if not dt:
            return None
        return f"{dt.year}-{dt.year+1}" if dt.month >= 6 else f"{dt.year-1}-{dt.year}"

    def _str(v):
        return str(v or "").strip() or None

    def _days(v, fallback=1.0):
        try:
            d = float(v or fallback)
            return d if d > 0 else fallback
        except (TypeError, ValueError):
            return fallback

    def _calculate_to_date(from_date, to_date_raw, days_count):
        """
        Calculate the correct to_date for a leave application.
        If to_date is provided in ERP, use it.
        Otherwise, calculate it from from_date + (days_count - 1).
        """
        if not from_date:
            return None

        # Try to parse the provided to_date
        parsed_to_date = _parse_date(to_date_raw)
        if parsed_to_date:
            return parsed_to_date

        # Calculate to_date from from_date and days_count
        try:
            days_int = int(days_count) if days_count else 1
            if days_int <= 0:
                days_int = 1
            # Subtract 1 because if leave is for 1 day, from_date = to_date
            return from_date + timedelta(days=days_int - 1)
        except (TypeError, ValueError):
            return from_date

    def _row_get(row, *names):
        if not isinstance(row, dict):
            return None
        lower_map = {str(k).lower(): v for k, v in row.items()}
        for name in names:
            key = str(name).lower()
            if key in lower_map:
                return lower_map[key]
        for key, value in lower_map.items():
            if any(key.startswith(str(name).lower()) for name in names):
                return value
        return None

    def _status_text(value):
        return str(value or "").strip().lower()

    def _is_approved_value(value):
        return _status_text(value) in {"approved", "approve", "1", "yes", "y", "true", "a"}

    def _is_rejected_value(value):
        return _status_text(value) in {
            "rejected", "reject", "denied", "cancelled", "canceled", "0", "no", "n", "false"
        }

    def _map_erp_status(erp_status_raw):
        """
        Map ERP approval columns to local status.
        ERP tables use columns such as hod_approval, principal_approval,
        director_approval and office_approval instead of a single status column.
        Local statuses: 'Approved', 'Pending', 'Rejected'
        """
        if isinstance(erp_status_raw, dict):
            row = erp_status_raw
            cancelled = _row_get(row, "cancelled", "cancel")
            if _status_text(cancelled) in {"yes", "y", "true", "1", "cancelled", "canceled"}:
                return "Rejected"

            direct_status = _row_get(row, "status", "leave_status", "approval_status")
            if direct_status is not None:
                direct = _map_erp_status(direct_status)
                if direct != "Pending" or _status_text(direct_status) in {"pending", ""}:
                    return direct

            approval_values = [
                _row_get(row, "hod_approval"),
                _row_get(row, "office_approval"),
                _row_get(row, "principal_approval", "principal_appr"),
                _row_get(row, "director_approval", "director_appr"),
            ]
            approval_values = [value for value in approval_values if _status_text(value)]
            if any(_is_rejected_value(value) and _status_text(value) != "no" for value in approval_values):
                return "Rejected"
            director_approval = _row_get(row, "director_approval", "director_appr")
            principal_approval = _row_get(row, "principal_approval", "principal_appr")
            office_approval = _row_get(row, "office_approval")
            if _is_approved_value(director_approval):
                return "Approved"
            if not _status_text(director_approval) and _is_approved_value(principal_approval):
                return "Approved"
            if not _status_text(director_approval) and not _status_text(principal_approval) and _is_approved_value(office_approval):
                return "Approved"
            if approval_values and all(_is_approved_value(value) for value in approval_values):
                return "Approved"
            if any(_is_approved_value(value) for value in approval_values) and not any(
                _status_text(value) == "pending" for value in approval_values
            ):
                return "Approved"
            return "Pending"

        status_str = str(erp_status_raw or "").strip().lower()
        if status_str in ("approved", "approve", "1", "yes"):
            return "Approved"
        elif status_str in ("rejected", "reject", "denied", "0", "no"):
            return "Rejected"
        else:
            return "Pending"

    # When `all` is set, pull every record from ERP (no year/date window).
    sync_all = str(request.POST.get("all") or request.GET.get("all") or "").strip().lower() in {
        "1", "true", "yes", "all", "on"
    }
    target_year = _to_int(request.POST.get("year") or request.GET.get("year")) or timezone.localdate().year
    sync_start = _date(target_year, 5, 1)
    sync_end = _date(target_year, 12, 31)

    # Only sync ERP rows for faculty that exist locally in general_information.
    local_fac_ids = sorted(
        fid for fid in general_information.objects
        .filter(faculty_id__isnull=False)
        .values_list("faculty_id", flat=True)
        if fid is not None
    )

    # Collect log lines and flush them to a text file under MEDIA_ROOT at the end.
    log_lines = []

    def _log(msg):
        log_lines.append(str(msg))
        print(f"[sync_ccl_from_erp] {msg}")

    def _write_log_file():
        import os
        from django.conf import settings
        stamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
        log_dir = os.path.join(settings.MEDIA_ROOT, "ccl_erp_sync_logs")
        try:
            os.makedirs(log_dir, exist_ok=True)
            path = os.path.join(log_dir, f"ccl_erp_sync_{stamp}.txt")
            header = (
                f"CCL / Leave ERP Sync Log\n"
                f"Run at   : {timezone.localtime().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"Scope    : {'ALL records' if sync_all else f'Year {target_year} ({sync_start} to {sync_end})'}\n"
                f"Local faculty in general_information : {len(local_fac_ids)}\n"
                f"{'=' * 60}\n"
            )
            with open(path, "w", encoding="utf-8") as fh:
                fh.write(header)
                fh.write("\n".join(log_lines))
                fh.write("\n")
            logger.info("sync_ccl_from_erp: wrote log file %s", path)
            return path
        except Exception as exc:
            logger.error("sync_ccl_from_erp: failed to write log file: %s", exc)
            return None
    target_year_labels = {
        str(target_year),
        f"{target_year - 1}-{target_year}",
        f"{target_year}-{target_year + 1}",
    }

    def _normalize_year_label(raw):
        value = _str(raw)
        if not value:
            return None
        return (
            value.replace("/", "-")
            .replace(" ", "")
            .replace("\u2013", "-")
            .replace("\u2014", "-")
        )

    def _first_date(row, *field_names):
        for field_name in field_names:
            parsed = _parse_date(_row_get(row, field_name))
            if parsed:
                return parsed
        return None

    def _row_in_sync_year(row, from_fields, to_fields=()):
        from_dt = _first_date(row, *from_fields)
        to_dt = _first_date(row, *to_fields) if to_fields else None
        if from_dt:
            to_dt = to_dt or from_dt
            return from_dt <= sync_end and to_dt >= sync_start

        year_label = _normalize_year_label(
            _row_get(row, "ay", "xy", "academic_year")
        )
        return year_label in target_year_labels

    def _mysql_date_expr(field_name):
        field_sql = f"`{field_name}`"
        value_sql = f"LEFT(TRIM({field_sql}), 10)"
        return (
            "CASE "
            f"WHEN {field_sql} IS NULL OR TRIM({field_sql}) = '' THEN NULL "
            f"WHEN {value_sql} REGEXP '^[0-9]{{4}}-[0-9]{{2}}-[0-9]{{2}}$' THEN STR_TO_DATE({value_sql}, '%%Y-%%m-%%d') "
            f"WHEN {value_sql} REGEXP '^[0-9]{{4}}/[0-9]{{2}}/[0-9]{{2}}$' THEN STR_TO_DATE({value_sql}, '%%Y/%%m/%%d') "
            f"WHEN {value_sql} REGEXP '^[0-9]{{2}}-[0-9]{{2}}-[0-9]{{4}}$' THEN STR_TO_DATE({value_sql}, '%%d-%%m-%%Y') "
            f"WHEN {value_sql} REGEXP '^[0-9]{{2}}/[0-9]{{2}}/[0-9]{{4}}$' THEN STR_TO_DATE({value_sql}, '%%d/%%m/%%Y') "
            "ELSE NULL END"
        )

    def _fetch_erp_rows(table_name, from_field, to_field=None):
        from_expr = _mysql_date_expr(from_field)
        to_expr = _mysql_date_expr(to_field) if to_field else from_expr
        label_placeholders = ", ".join(["%s"] * len(target_year_labels))
        ay_expr = "REPLACE(REPLACE(REPLACE(TRIM(COALESCE(`ay`, '')), '/', '-'), ' ', ''), '–', '-')"
        # Restrict to faculty that exist locally in general_information.
        clauses = []
        params = []
        if not sync_all:
            clauses.append(
                f"(({from_expr}) <= %s AND COALESCE(({to_expr}), ({from_expr})) >= %s) "
                f"OR (({from_expr}) IS NULL AND {ay_expr} IN ({label_placeholders}))"
            )
            params.extend([sync_end, sync_start, *sorted(target_year_labels)])
        if local_fac_ids:
            fac_placeholders = ", ".join(["%s"] * len(local_fac_ids))
            clauses.append(f"`fac_id` IN ({fac_placeholders})")
            params.extend(local_fac_ids)
        # Wrap each clause so OR conditions inside a clause stay grouped.
        where_sql = " AND ".join(f"({c})" for c in clauses) if clauses else None

        with connections["erp_rit"].cursor() as cur:
            cur.execute(f"SELECT COUNT(*) FROM `{table_name}`")
            total_rows = cur.fetchone()[0]
            cur.execute(f"SHOW COLUMNS FROM `{table_name}`")
            columns = [row[0] for row in cur.fetchall()]
            if where_sql is None:
                cur.execute(f"SELECT * FROM `{table_name}` ORDER BY `id`")
            else:
                cur.execute(
                    f"SELECT * FROM `{table_name}` WHERE {where_sql} ORDER BY `id`",
                    params,
                )
            row_columns = [c[0] for c in cur.description]
            rows = [dict(zip(row_columns, row)) for row in cur.fetchall()]
        return rows, columns, total_rows

    def _print_erp_debug(table_name, total_rows, columns, filtered_rows):
        sample_rows = filtered_rows[:5]
        print(f"[sync_ccl_from_erp] {table_name}: total={total_rows}, selected_year_rows={len(filtered_rows)}")
        print(f"[sync_ccl_from_erp] {table_name} columns: {columns}")
        for index, row in enumerate(sample_rows, start=1):
            print(f"[sync_ccl_from_erp] {table_name} sample {index}: {row}")
        logger.info(
            "sync_ccl_from_erp %s total=%s selected_year_rows=%s columns=%s sample=%s",
            table_name,
            total_rows,
            len(filtered_rows),
            columns,
            sample_rows,
        )

    try:
        if not local_fac_ids:
            msg = "No faculty found in general_information; nothing to sync."
            _log(msg)
            logger.warning("sync_ccl_from_erp: %s", msg)
            messages.warning(request, msg)
            _write_log_file()
            return redirect("employee_attandance_report")

        _log(f"Syncing only for {len(local_fac_ids)} faculty present in general_information.")

        # ── Fetch the three ERP tables (only for local faculty) ──
        fl_rows, fl_cols, fl_total_rows = _fetch_erp_rows(
            "faculty_leave",
            from_field="from_date",
            to_field="to_date",
        )

        ccl_rows, ccl_cols, ccl_total_rows = _fetch_erp_rows(
            "ccl",
            from_field="date",
        )

        leave_rows, lv_cols, leave_total_rows = _fetch_erp_rows(
            "employee_leave",
            from_field="from_date",
            to_field="to_date",
        )

        diag["sync_year"] = target_year
        diag["sync_start"] = sync_start
        diag["sync_end"] = sync_end
        diag["remote_fl_rows_all"] = fl_total_rows
        diag["remote_ccl_rows_all"] = ccl_total_rows
        diag["remote_leave_rows_all"] = leave_total_rows

        if not fl_total_rows and not ccl_total_rows and not leave_total_rows:
            msg = "No rows found in erp_rit tables: faculty_leave, ccl, employee_leave."
            _log(msg)
            logger.warning("sync_ccl_from_erp: %s", msg)
            messages.warning(request, msg)
            _write_log_file()
            return redirect("employee_attandance_report")

        if sync_all:
            _log("Sync scope=ALL records (no date window)")
        else:
            _log(f"Sync year={target_year}, date window={sync_start} to {sync_end}")
        _log(f"faculty_leave: total={fl_total_rows}, selected={len(fl_rows)}")
        _log(f"ccl: total={ccl_total_rows}, selected={len(ccl_rows)}")
        _log(f"employee_leave: total={leave_total_rows}, selected={len(leave_rows)}")
        _print_erp_debug("faculty_leave", fl_total_rows, fl_cols, fl_rows)
        _print_erp_debug("ccl", ccl_total_rows, ccl_cols, ccl_rows)
        _print_erp_debug("employee_leave", leave_total_rows, lv_cols, leave_rows)

        if not fl_rows and not ccl_rows and not leave_rows:
            msg = (
                "ERP data exists, but no faculty_leave/ccl/employee_leave rows were selected "
                "for the local faculty."
                if sync_all else
                f"ERP data exists, but no faculty_leave/ccl/employee_leave rows "
                f"matched {sync_start} to {sync_end} for the local faculty."
            )
            _log(msg)
            logger.warning("sync_ccl_from_erp: %s", msg)
            messages.warning(request, msg)
            _write_log_file()
            return redirect("employee_attandance_report")

        # ── Collect all remote fac_ids ──
        all_fac_int_ids = set()
        for r in fl_rows + ccl_rows + leave_rows:
            v = _to_int(_row_get(r, "fac_id"))
            if v is not None:
                all_fac_int_ids.add(v)

        # ── Faculty map: only those already in general_information ──
        faculty_map = {
            gi.faculty_id: gi
            for gi in general_information.objects
                .select_related("designation")
                .filter(faculty_id__in=all_fac_int_ids)
            if gi.faculty_id is not None
        }
        faculty_pks = [gi.pk for gi in faculty_map.values()]

        diag["remote_fl_rows"]    = len(fl_rows)
        diag["remote_ccl_rows"]   = len(ccl_rows)
        diag["remote_leave_rows"] = len(leave_rows)
        diag["remote_fac_ids"]    = len(all_fac_int_ids)
        diag["matched_faculty"]   = len(faculty_map)

        # ── User pk + role_id + dept map from rit_approval_system ──
        user_map = {}
        faculty_role_map = {}
        faculty_dept_map = {}
        for row_u in USER.objects.using("rit_approval_system") \
                .filter(Employee_id__in=[str(f) for f in all_fac_int_ids]) \
                .values("Employee_id", "id", "role_id", "Department_id"):
            try:
                fac_key = int(str(row_u["Employee_id"]).strip())
                user_map[fac_key] = row_u["id"]
                if row_u.get("role_id"):
                    faculty_role_map[fac_key] = row_u["role_id"]
                if row_u.get("Department_id"):
                    faculty_dept_map[fac_key] = row_u["Department_id"]
            except (TypeError, ValueError):
                pass

        # ── Ensure all needed LeaveTypes exist ──
        DEFAULT_LEAVE_TYPES = {
            "CL": ("Casual Leave", True),
            "CCL": ("Compensatory Casual Leave", True),
            "OD": ("On Duty", True),
            "ROD": ("Restricted On Duty", True),
            "LOP": ("Loss of Pay", True),
            "VL": ("Vacation Leave", True),
            "ML": ("Medical Leave", True),
            "WL": ("Without Leave", False),
        }
        lt_map = {
            lt.code.upper().strip(): lt
            for lt in LeaveType.objects.exclude(code__isnull=True).exclude(code="")
        }

        def _get_leave_type(raw_code):
            code = str(raw_code or "").strip().upper()
            if not code:
                return None
            if code in lt_map:
                return lt_map[code]
            default_name, is_leave = DEFAULT_LEAVE_TYPES.get(code, (code, True))
            lt_obj, _ = LeaveType.objects.get_or_create(
                code=code,
                defaults={
                    "name": default_name,
                    "is_leave": is_leave,
                    "is_active": True,
                },
            )
            lt_map[code] = lt_obj
            return lt_obj

        for code in DEFAULT_LEAVE_TYPES:
            if code not in lt_map:
                _get_leave_type(code)

        # ── Existing dedup sets ──
        ccl_existing = {
            (obj.faculty_id, obj.date): obj
            for obj in CCL_Application.objects.filter(faculty_id__in=faculty_pks)
            if obj.faculty_id and obj.date
        }
        leave_existing = {
            (obj.faculty_id, obj.from_date, obj.to_date, obj.leave_type_id): obj
            for obj in LeaveApplication.objects.filter(faculty_id__in=faculty_pks)
            if obj.faculty_id and obj.from_date and obj.leave_type_id
        }
        faculty_leave_keys = set()

        # ── Approver-chain template cache ──
        # The resolved approver chain depends only on (creator_role_id, creator_dept_id),
        # so resolve it ONCE per unique pair (this is what used to hit the remote
        # rit_approval_system DB thousands of times, one lookup per record).
        from faculty_leave_management.models import (
            LeaveApprovers, CCL_Approvers_Data, LeaveApproversData,
        )
        from user_accounts.models import Add_Department
        try:
            from user_accounts.models import Department as _ExtDepartment
        except Exception:
            _ExtDepartment = None

        _chain_template_cache = {}
        # Cache Add_Department -> external dept id resolution too.
        _ext_dept_cache = {}

        def _ext_dept_id_for_local(local_dept_id):
            if local_dept_id in _ext_dept_cache:
                return _ext_dept_cache[local_dept_id]
            ext_id = None
            local_dept = Add_Department.objects.filter(id=local_dept_id).first()
            dept_code = getattr(local_dept, "Department_code", None)
            if dept_code and _ExtDepartment is not None:
                ext_id = (
                    _ExtDepartment.objects.using("rit_approval_system")
                    .filter(Department_code=dept_code)
                    .values_list("id", flat=True)
                    .first()
                )
            _ext_dept_cache[local_dept_id] = ext_id
            return ext_id

        def _resolve_chain_template(creator_role_id, creator_dept_id):
            """Return a list of dicts describing each approver level for this
            (creator_role_id, creator_dept_id) pair. Cached across records."""
            key = (creator_role_id, creator_dept_id)
            if key in _chain_template_cache:
                return _chain_template_cache[key]

            template = []
            created_levels = set()
            approvers_qs = LeaveApprovers.objects.filter(
                creator_role_id=creator_role_id
            ).order_by("approver_level")
            for approver in approvers_qs:
                level = approver.approver_level
                if level in created_levels:
                    continue
                role_id = approver.approver_role_id
                local_dept_id = approver.approver_department_id
                is_cross = (approver.is_cross_department_approver or "NO").upper() == "YES"

                approver_filter = {"role_id": role_id, "is_active": True}
                if is_cross and local_dept_id:
                    ext_dept_id = _ext_dept_id_for_local(local_dept_id)
                    if ext_dept_id:
                        approver_filter["Department_id"] = ext_dept_id
                elif not is_cross and not local_dept_id and creator_dept_id:
                    approver_filter["Department_id"] = creator_dept_id

                approver_user = USER.objects.using("rit_approval_system").filter(
                    **approver_filter
                ).first()
                if not approver_user:
                    continue
                approver_faculty = general_information.objects.filter(
                    faculty_id=approver_user.Employee_id
                ).select_related("designation").first()
                if not approver_faculty:
                    continue

                template.append({
                    "level": level,
                    "role_id": role_id,
                    "approver_faculty": approver_faculty,
                    "approver_emp_id": approver_user.Employee_id,
                })
                created_levels.add(level)

            _chain_template_cache[key] = template
            return template

        # Approver rows are collected here and bulk-inserted after each part.
        ccl_approver_rows = []
        leave_approver_rows = []

        # ── Session map: ERP session string -> PermissionTimingMaster ──
        from faculty_leave_management.models import PermissionTimingMaster
        _fn_session   = PermissionTimingMaster.objects.filter(session_name="FN",       is_active=True).first()
        _an_session   = PermissionTimingMaster.objects.filter(session_name="AN",       is_active=True).first()
        _full_session = PermissionTimingMaster.objects.filter(session_name="FULL Day", is_active=True).first()

        def _resolve_session(from_s, to_s):
            """Map ERP from_session/to_session to local PermissionTimingMaster."""
            f = str(from_s or "").strip().upper()
            t = str(to_s   or "").strip().upper()
            if f == "FN"   and t == "FN":   return _fn_session
            if f == "AN"   and t == "AN":   return _an_session
            if f == "BOTH" and t == "BOTH": return _full_session
            # Mixed sessions (FN->both, AN->both, both->FN, both->AN, AN->FN) = full day
            return _full_session

        # ════════════════════════════════════════════════════════
        # PART 1 — faculty_leave table (primary, most complete)
        # ════════════════════════════════════════════════════════
        for row in fl_rows:
            fac_int = _to_int(_row_get(row, "fac_id"))
            if fac_int is None:
                leave_skipped += 1
                continue
            gi = faculty_map.get(fac_int)
            if gi is None:
                unmatched_faculty_ids.add(fac_int)
                leave_skipped += 1
                continue

            raw_lt  = str(_row_get(row, "leave_type") or "").strip().upper()
            from_dt = _parse_date(_row_get(row, "from_date"))
            days    = _days(_row_get(row, "days"))
            to_dt   = _calculate_to_date(from_dt, _row_get(row, "to_date"), days)
            ay      = _str(_row_get(row, "ay")) or _ay(from_dt)
            reason  = _str(_row_get(row, "reason"))

            # Read status from ERP
            erp_status = _map_erp_status(row)

            lt_obj = _get_leave_type(raw_lt)
            if lt_obj is None:
                missing_leave_type += 1
                leave_skipped += 1
                continue
            if not from_dt or not to_dt:
                invalid_rows += 1
                leave_skipped += 1
                continue
            leave_key = (gi.pk, from_dt, to_dt, lt_obj.pk)
            faculty_leave_keys.add(leave_key)
            existing_leave = leave_existing.get(leave_key)
            if existing_leave:
                # Already present locally — skip, leave the existing record untouched.
                leave_existing_skipped += 1
                continue
            session_obj = _resolve_session(_row_get(row, "from_session"), _row_get(row, "to_session"))
            new_leave = LeaveApplication.objects.create(
                faculty_id=gi.pk,
                user_id=user_map.get(fac_int),
                designation_id=gi.designation_id,
                academic_year=ay,
                from_date=from_dt,
                to_date=to_dt,
                leave_type=lt_obj,
                reason=reason,
                status=erp_status,
                session=session_obj,
            )
            creator_role_id = faculty_role_map.get(fac_int)
            if creator_role_id:
                approved = erp_status == "Approved"
                for t in _resolve_chain_template(creator_role_id, faculty_dept_map.get(fac_int)):
                    leave_approver_rows.append(LeaveApproversData(
                        leave_application=new_leave,
                        approver_id=t["approver_faculty"],
                        creator_id=gi,
                        approver_level=t["level"],
                        approver_role_id=t["role_id"],
                        creator_role_id=creator_role_id,
                        status=LeaveApproversData.Status.APPROVED if approved else LeaveApproversData.Status.PENDING,
                        reason="Auto-approved from ERP sync" if approved else f"Approver: {t['approver_emp_id']}",
                        approved_date=timezone.now(),
                    ))
            leave_existing[leave_key] = new_leave
            leave_synced += 1

        # ════════════════════════════════════════════════════════
        # PART 2 — ccl table (older records not in faculty_leave)
        # ════════════════════════════════════════════════════════
        for row in ccl_rows:
            fac_int = _to_int(_row_get(row, "fac_id"))
            if fac_int is None:
                ccl_skipped += 1; continue
            gi = faculty_map.get(fac_int)
            if gi is None:
                unmatched_faculty_ids.add(fac_int)
                ccl_skipped += 1; continue

            ccl_date = _parse_date(_row_get(row, "date"))
            if not ccl_date:
                invalid_rows += 1
                ccl_skipped += 1; continue
            days_count = _days(_row_get(row, "value", "days"))
            ay = _str(_row_get(row, "ay", "xy")) or _ay(ccl_date)
            erp_status = _map_erp_status(row)
            ccl_key = (gi.pk, ccl_date)
            existing_ccl = ccl_existing.get(ccl_key)
            if existing_ccl:
                # Already present locally — skip, leave the existing record untouched.
                ccl_existing_skipped += 1; continue

            new_ccl = CCL_Application.objects.create(
                faculty_id=gi.pk,
                user_id=user_map.get(fac_int),
                designation_id=gi.designation_id,
                academic_year=ay,
                date=ccl_date,
                reason=_str(_row_get(row, "reason")),
                status=erp_status,
                days=days_count,
                is_claimed=False,
            )
            creator_role_id = faculty_role_map.get(fac_int)
            if creator_role_id:
                approved = erp_status == "Approved"
                for t in _resolve_chain_template(creator_role_id, faculty_dept_map.get(fac_int)):
                    ccl_approver_rows.append(CCL_Approvers_Data(
                        ccl_application=new_ccl,
                        approver_id=t["approver_faculty"],
                        approver_level=t["level"],
                        approver_role_id=t["role_id"],
                        creator_role_id=creator_role_id,
                        status=CCL_Approvers_Data.Status.APPROVED if approved else CCL_Approvers_Data.Status.PENDING,
                        action_date=timezone.now() if approved else None,
                        remarks="Auto-approved from ERP sync" if approved else f"Approver: {t['approver_emp_id']}",
                    ))
            ccl_existing[ccl_key] = new_ccl
            ccl_synced += 1

        # ════════════════════════════════════════════════════════
        # PART 3 — employee_leave (older records not in faculty_leave)
        # ════════════════════════════════════════════════════════
        for row in leave_rows:
            fac_int = _to_int(_row_get(row, "fac_id"))
            if fac_int is None:
                leave_skipped += 1; continue
            gi = faculty_map.get(fac_int)
            if gi is None:
                unmatched_faculty_ids.add(fac_int)
                leave_skipped += 1; continue

            raw_lt = str(_row_get(row, "leave_type") or "").strip().upper()
            erp_status = _map_erp_status(row)

            lt_obj = _get_leave_type(raw_lt)
            if lt_obj is None:
                missing_leave_type += 1
                leave_skipped += 1; continue

            from_date = _parse_date(_row_get(row, "from_date"))
            days_count = _days(_row_get(row, "days", "value"))
            to_date = _calculate_to_date(from_date, _row_get(row, "to_date"), days_count)
            if not from_date or not to_date:
                invalid_rows += 1
                leave_skipped += 1; continue

            leave_key = (gi.pk, from_date, to_date, lt_obj.pk)
            if leave_key in faculty_leave_keys:
                leave_existing_skipped += 1; continue
            existing_leave = leave_existing.get(leave_key)
            if existing_leave:
                # Already present locally — skip, leave the existing record untouched.
                leave_existing_skipped += 1; continue

            ay = _str(_row_get(row, "ay", "xy")) or _ay(from_date)
            session_obj = _resolve_session(_row_get(row, "from_session"), _row_get(row, "to_session"))

            new_leave = LeaveApplication.objects.create(
                faculty_id=gi.pk,
                user_id=user_map.get(fac_int),
                designation_id=gi.designation_id,
                academic_year=ay,
                from_date=from_date,
                to_date=to_date,
                leave_type=lt_obj,
                reason=_str(_row_get(row, "reason")),
                status=erp_status,
                session=session_obj,
            )
            creator_role_id = faculty_role_map.get(fac_int)
            if creator_role_id:
                approved = erp_status == "Approved"
                for t in _resolve_chain_template(creator_role_id, faculty_dept_map.get(fac_int)):
                    leave_approver_rows.append(LeaveApproversData(
                        leave_application=new_leave,
                        approver_id=t["approver_faculty"],
                        creator_id=gi,
                        approver_level=t["level"],
                        approver_role_id=t["role_id"],
                        creator_role_id=creator_role_id,
                        status=LeaveApproversData.Status.APPROVED if approved else LeaveApproversData.Status.PENDING,
                        reason="Auto-approved from ERP sync" if approved else f"Approver: {t['approver_emp_id']}",
                        approved_date=timezone.now(),
                    ))
            leave_existing[leave_key] = new_leave
            leave_synced += 1

        # ── Bulk-insert all collected approver rows (few big inserts) ──
        if ccl_approver_rows:
            CCL_Approvers_Data.objects.bulk_create(ccl_approver_rows, batch_size=500)
        if leave_approver_rows:
            LeaveApproversData.objects.bulk_create(leave_approver_rows, batch_size=500)
        _log(
            f"Approver rows created: CCL={len(ccl_approver_rows)}, "
            f"Leave={len(leave_approver_rows)}; unique chain templates resolved="
            f"{len(_chain_template_cache)}"
        )

    except Exception:
        import traceback
        errors.append(traceback.format_exc())

    if errors:
        logger.error("sync_ccl_from_erp error:\n%s", errors[0])
        _log("SYNC FAILED with an exception:")
        _log(errors[0])
        messages.error(request, f"Sync failed: {errors[0][:3000]}")
    else:
        unmatched_count = diag.get('remote_fac_ids', 0) - diag.get('matched_faculty', 0)
        diag["invalid_rows"] = invalid_rows
        diag["missing_leave_type"] = missing_leave_type
        diag["unmatched_faculty_ids"] = sorted(unmatched_faculty_ids)[:25]
        scope_label = (
            "ALL records"
            if sync_all
            else f"Year: {diag.get('sync_year')} ({diag.get('sync_start')} to {diag.get('sync_end')})"
        )
        summary = (
            f"Sync complete - "
            f"{scope_label} | "
            f"CCL earned: {ccl_synced} created / {ccl_existing_skipped} already-present skipped / {ccl_skipped} invalid skipped "
            f"(ccl: {diag.get('remote_ccl_rows',0)} of {diag.get('remote_ccl_rows_all',0)} rows) | "
            f"Leaves: {leave_synced} created / {leave_existing_skipped} already-present skipped / {leave_skipped} invalid skipped "
            f"(faculty_leave: {diag.get('remote_fl_rows',0)} of {diag.get('remote_fl_rows_all',0)} rows, "
            f"employee_leave: {diag.get('remote_leave_rows',0)} of {diag.get('remote_leave_rows_all',0)} rows) | "
            f"Faculty matched: {diag.get('matched_faculty',0)} / {diag.get('remote_fac_ids',0)} "
            f"({unmatched_count} not in local DB) | "
            f"Invalid date rows: {invalid_rows} | Missing leave type: {missing_leave_type}"
        )
        _log(summary)
        if unmatched_count > 0:
            _log(f"Unmatched ERP fac_ids (sample): {diag['unmatched_faculty_ids']}")
            logger.warning(
                "sync_ccl_from_erp: %d ERP fac_ids not matched locally. Sample: %s",
                unmatched_count,
                diag["unmatched_faculty_ids"],
            )
        messages.success(request, summary)

    log_path = _write_log_file()
    if log_path:
        import os
        from django.conf import settings
        rel = os.path.relpath(log_path, settings.MEDIA_ROOT).replace("\\", "/")
        messages.info(request, f"Sync log saved: {settings.MEDIA_URL}{rel}")

    return redirect("employee_attandance_report")



def emp_attendance_export(request):
    """
    Exports the full attendance grid (matching the screen report) as Excel or PDF.
    Salary conditions applied:
      - perm_allowed: max permissions before deduction
      - entry_time / exit_time: late entry / early exit flagged as SP -> deduction
    """
    from django.http import HttpResponse
    from datetime import datetime as _dt, date as _date, timedelta, time as _time
    from faculty_management.models import FacultyCategory, general_information
    from user_accounts.models import Add_Department
    from faculty_leave_management.models import (
        LeaveApplication, LeaveBalance, LeaveType,
        CCL_Claim, Employee_Holidays, DeviceInfo, ShiftDetail,
        CCL_Application, PermissionRequest,
    )
    from django.db.models import Q, Sum
    import calendar

    # ── Params ──
    now      = datetime.now()
    year     = int(request.GET.get("year",  now.year))
    month    = int(request.GET.get("month", now.month))
    dept_id  = request.GET.get("dept_id",  "")
    category_id = request.GET.get("category_id", "")
    search   = (request.GET.get("search") or "").strip()
    fmt      = request.GET.get("fmt", "excel")

    from_date_str = (request.GET.get("from_date") or "").strip()
    to_date_str   = (request.GET.get("to_date")   or "").strip()
    try:
        perm_allowed = float(request.GET.get("perm_allowed", 2) or 2)
    except (ValueError, TypeError):
        perm_allowed = 2.0

    entry_time_str = request.GET.get("entry_time", "09:00") or "09:00"
    exit_time_str  = request.GET.get("exit_time",  "16:30") or "16:30"

    try:
        entry_limit = _dt.strptime(entry_time_str, "%H:%M").time()
    except ValueError:
        entry_limit = _time(9, 0)
    try:
        exit_limit = _dt.strptime(exit_time_str, "%H:%M").time()
    except ValueError:
        exit_limit = _time(17, 0)

    # ── Period ──
    if from_date_str:
        try:
            period_start = _dt.strptime(from_date_str, "%Y-%m-%d").date()
        except ValueError:
            period_start = None
    else:
        period_start = None

    if to_date_str:
        try:
            period_end = _dt.strptime(to_date_str, "%Y-%m-%d").date()
        except ValueError:
            period_end = None
    else:
        period_end = None

    if not period_start or not period_end:
        prev_year, prev_month = (year-1, 12) if month == 1 else (year, month-1)
        period_start = period_start or _date(prev_year, prev_month, 26)
        period_end   = period_end   or _date(year, month, 25)

    period_days = []
    cur = period_start
    while cur <= period_end:
        period_days.append(cur)
        cur += timedelta(days=1)

    academic_year = _academic_year_for_date(period_end)

    # ── Employees ──
    employees = (
        general_information.objects
        .select_related("department", "designation", "category", "shift")
        .filter(faculty_id__isnull=False)
        .order_by("name")
    )
    if dept_id:
        employees = employees.filter(department_id=dept_id)
    if category_id:
        employees = employees.filter(category_id=category_id)
    if search:
        employees = employees.filter(
            Q(name__icontains=search) | Q(faculty_id__icontains=search)
        )

    faculty_ids = []
    for fid in employees.values_list("faculty_id", flat=True):
        try:
            faculty_ids.append(int(fid))
        except (TypeError, ValueError):
            pass

    if not faculty_ids:
        return HttpResponse("No employees found.", status=400)

    # Build pk → faculty_id map (avoids cross-FK JOIN for CCL queries)
    emp_pk_map_export = {}
    for _emp in employees:
        _fid = _safe_int(getattr(_emp, "faculty_id", None))
        if _fid is not None:
            emp_pk_map_export[_emp.pk] = _fid

    # ── Punch data ──
    all_punches = []
    id_list_str = ",".join(f"'{fid}'" for fid in faculty_ids)
    # Every (month, year) table spanning the range (one day before period_start so
    # a night shift ending on the first day still finds its previous-evening IN).
    _span = []
    _d0 = period_start - timedelta(days=1)
    _sy, _sm = _d0.year, _d0.month
    while (_sy, _sm) <= (period_end.year, period_end.month):
        _span.append((_sm, _sy))
        _sm += 1
        if _sm > 12:
            _sm = 1
            _sy += 1
    for tbl_month, tbl_year in _span:
        sql = (
            f"SELECT UserId, LogDate, DeviceId "
            f"FROM DeviceLogs_{tbl_month}_{tbl_year} "
            f"WHERE CAST(UserId AS VARCHAR) IN ({id_list_str}) "
            f"ORDER BY LogDate ASC"
        )
        try:
            with connections["attendance_db"].cursor() as cur:
                cur.execute(sql)
                cols = [c[0] for c in cur.description]
                all_punches.extend(dict(zip(cols, row)) for row in cur.fetchall())
        except Exception:
            pass

    # Device filter
    devices_qs = DeviceInfo.objects.filter(is_active=True).order_by("deviceid")
    device_master = {}
    attendance_device_ids = []
    for d in devices_qs:
        dev_id = str(d.deviceid or "").strip()
        if not dev_id:
            continue
        if d.is_attendance:
            attendance_device_ids.append(dev_id)
        device_master[dev_id] = {"is_attendance": d.is_attendance}
    no_device_filter = len(attendance_device_ids) == 0

    punch_by_emp = {}
    for p in all_punches:
        uid = p.get("UserId")
        log_dt = p.get("LogDate")
        if uid is None or log_dt is None:
            continue
        try:
            uid = int(uid)
        except (ValueError, TypeError):
            continue
        dkey = log_dt.date() if hasattr(log_dt, "date") else log_dt
        if dkey < period_start or dkey > period_end:
            continue
        dev_id = str(p.get("DeviceId") or "").strip()
        is_att = True if no_device_filter else device_master.get(dev_id, {}).get("is_attendance", False)
        punch_by_emp.setdefault(uid, {}).setdefault(dkey, []).append({
            "logdate": log_dt, "is_attendance": is_att,
        })

    # ── Leaves ──
    all_leaves = list(
        LeaveApplication.objects
        .select_related("leave_type", "faculty")
        .filter(
            faculty__faculty_id__in=faculty_ids,
            from_date__lte=period_end,
            to_date__gte=period_start,
            status__iexact="Approved",
        )
    )
    leave_by_emp = {}
    for la in all_leaves:
        fid = _safe_int(la.faculty.faculty_id if la.faculty else None)
        if fid:
            leave_by_emp.setdefault(fid, []).append(la)

    # ── Employee_Holidays ──
    holiday_map = {}
    for h in Employee_Holidays.objects.filter(
        holiday_date__gte=period_start, holiday_date__lte=period_end
    ).select_related("category"):
        _add_to_holiday_map(holiday_map, h)

    # ── Role map ──
    from user_accounts.models import USER
    role_id_map = {}
    try:
        for u in USER.objects.using("rit_approval_system").filter(
            Employee_id__in=[str(fid) for fid in faculty_ids]
        ).values("Employee_id", "role_id"):
            try:
                role_id_map.setdefault(int(u["Employee_id"]), set()).add(int(u["role_id"]))
            except (TypeError, ValueError):
                pass
    except Exception:
        pass
    holiday_role_day_count = {rid: len(days) for rid, days in holiday_map.items()}
    role_id_map_single = {}
    for emp_id, role_ids in role_id_map.items():
        role_id_map_single[emp_id] = sorted(
            role_ids, key=lambda rid: (-holiday_role_day_count.get(rid, 0), rid)
        )[0]

    # ── CCL dates — query by PK to avoid cross-FK JOIN ──
    ccl_dates_map = {}
    for ca in CCL_Application.objects.filter(
        faculty_id__in=list(emp_pk_map_export.keys()),
        date__gte=period_start, date__lte=period_end,
        status__iexact="Approved",
    ):
        fid = emp_pk_map_export.get(ca.faculty_id)
        if fid and ca.date:
            ccl_dates_map.setdefault(fid, set()).add(ca.date)

    # ── Permissions ──
    perm_map = {}
    NOON = _time(12, 0)
    for pr in PermissionRequest.objects.filter(
        faculty__faculty_id__in=faculty_ids,
        date__gte=period_start, date__lte=period_end,
        status__iexact="Approved",
    ).select_related("faculty"):
        fid = _safe_int(pr.faculty.faculty_id if pr.faculty else None)
        if not fid or not pr.date:
            continue
        session = "morning" if (pr.from_time and pr.from_time < NOON) else "afternoon"
        existing = perm_map.setdefault(fid, {}).get(pr.date)
        if existing and existing != session:
            perm_map[fid][pr.date] = "both"
        else:
            perm_map[fid][pr.date] = session

    # ── Leave types ──
    all_leave_types = list(LeaveType.objects.all().order_by("name"))
    leave_type_by_code = {lt.code: lt for lt in all_leave_types if lt.code}
    preferred_order = ["CL", "CCL", "OD", "LOP", "AB", "VL", "ML"]
    leave_types = [{"code": c, "name": leave_type_by_code[c].name if c in leave_type_by_code else c}
                   for c in preferred_order]

    # ── Build report rows (same logic as emp_attandance_report) ──
    policy_map = _build_attendance_policy_map(employees)

    report_rows = []
    standard_leave_codes = {"CL", "CCL", "OD", "LOP", "AB", "VL", "ML", "WL"}

    for emp in employees:
        try:
            fid = int(emp.faculty_id)
        except (TypeError, ValueError):
            continue

        role_id = role_id_map_single.get(fid)
        if role_id is None and getattr(emp, "designation_id", None) is not None:
            role_id = int(emp.designation_id)

        emp_punches   = punch_by_emp.get(fid, {})
        emp_leaves    = leave_by_emp.get(fid, [])
        emp_holidays  = _holiday_map_for_employee(
            holiday_map,
            role_id,
            getattr(emp, "category_id", None),
        )
        emp_ccl_dates = ccl_dates_map.get(fid, set())
        emp_perm_dates = perm_map.get(fid, {})
        attendance_policy = policy_map.get(emp.pk)

        # ── Track late entries for permission limit (same as main report) ──
        late_entry_count_for_limit = 0
        late_entry_dates = []

        # First pass: identify late entries
        for att_date in period_days:
            punches_today = emp_punches.get(att_date, [])
            holiday_obj   = emp_holidays.get(att_date)
            leaves_today  = [la for la in emp_leaves
                              if la.from_date <= att_date <= la.to_date]

            # Skip if on leave, holiday, or has approved permission
            if leaves_today or holiday_obj or emp_perm_dates.get(att_date) or _policy_skips_time_checks(attendance_policy):
                continue

            # Check if late entry (first punch after entry_time)
            if punches_today:
                first_punch_time = None
                for p in punches_today:
                    log_dt = p.get("logdate")
                    if log_dt and hasattr(log_dt, "time"):
                        punch_time = log_dt.time()
                        if first_punch_time is None or punch_time < first_punch_time:
                            first_punch_time = punch_time

                if first_punch_time and first_punch_time > entry_limit:
                    late_entry_dates.append(att_date)

        day_cells = []
        present_count = 0.0
        ab_count = 0.0
        holidays_count = 0.0
        working_days = 0.0
        perm_count = 0.0  # Changed to float to support half-day permissions
        late_entry_count = 0
        early_exit_count = 0
        late_permission_count = 0  # Track late permissions used

        for att_date in period_days:
            punches_today = emp_punches.get(att_date, [])
            holiday_obj   = emp_holidays.get(att_date)
            leaves_today  = [la for la in emp_leaves if la.from_date <= att_date <= la.to_date]

            # Check if this is a late entry that should be marked as permission or absent
            is_late_entry = att_date in late_entry_dates
            if is_late_entry:
                if late_entry_count_for_limit < perm_allowed:
                    # Within allowed permissions - mark as late permission (PER/P format)
                    late_entry_count_for_limit += 1
                    late_permission_count += 1
                    late_entry_count += 1  # For salary tracking
                    status = "PER/P"  # Show as permission + present
                else:
                    # Exceeded allowed permissions - mark as absent
                    status = "A"
            else:
                status = _get_day_status(
                    att_date, punches_today, leaves_today, holiday_obj,
                    emp_ccl_dates, emp_perm_dates.get(att_date), attendance_policy
                )

            # ── Early exit check for salary (separate from late entry) ──
            s_raw = (status or "").strip().upper()
            if not _policy_skips_time_checks(attendance_policy) and s_raw in ("P", "PER/P") and punches_today:
                att_punches = [p for p in punches_today if p.get("is_attendance")]
                if att_punches:
                    times = sorted(
                        p["logdate"] for p in att_punches if p.get("logdate")
                    )
                    if times:
                        last_t = times[-1].time() if hasattr(times[-1], "time") else times[-1]
                        if last_t < exit_limit:
                            early_exit_count += 1

            # Display transform
            if status and status.upper().startswith("FN/"):
                display = "H/" + status[3:]
            elif status and status.upper().endswith("/AN"):
                display = status[:-3] + "/H"
            else:
                display = status

            day_cells.append({"date": att_date, "status": status, "display": display})

            s = (status or "").strip().upper()
            h_code = (holiday_obj.session_type or "").strip().upper() if holiday_obj else ""

            # Holidays come only from the role-based holiday table; Sundays are not
            # automatically holidays. CCL = worked on a holiday -> working day.
            is_ccl_day = (s == "CCL")
            if is_ccl_day:
                hf = 0.0
            elif h_code in ("H", "F"):
                hf = 1.0
            elif h_code in ("FN", "AN"):
                hf = 0.5
            else:
                hf = 0.0
            holidays_count += hf
            working_days += (1.0 - hf)

            if s == "CCL":
                present_count += 1
            elif s in ("H", "S"):
                pass
            elif s.startswith("FN/"):
                right = s.split("/", 1)[1].strip().upper()
                if right == "P":
                    present_count += 0.5
                elif right not in standard_leave_codes:
                    ab_count += 0.5
            elif s.endswith("/AN"):
                left = s.split("/", 1)[0].strip().upper()
                if left == "P":
                    present_count += 0.5
                elif left not in standard_leave_codes:
                    ab_count += 0.5
            elif s == "A":
                ab_count += 1
            elif s == "P":
                present_count += 1
            elif s == "SP":
                ab_count += 1
            elif s in (FLEXIBLE_MISSING_OUT_STATUS, FLEXIBLE_SHORT_HOURS_STATUS):
                ab_count += 1
            elif "/" in s:
                left, right = s.split("/", 1)
                left, right = left.strip().upper(), right.strip().upper()
                if (left == "P" and right in standard_leave_codes) or (right == "P" and left in standard_leave_codes):
                    present_count += 0.5
                elif "P" in (left, right):
                    present_count += 1.0
                elif left not in standard_leave_codes and right not in standard_leave_codes:
                    ab_count += 1

            # Permission count (includes PER/P from late entries)
            # Handle full-day and half-day permissions
            if s in ("PER/P", "P/PER"):
                perm_count += 1  # Full day permission
            elif s == "PER/PER":
                perm_count += 1  # Full day permission
            elif "PER" in s and "/" in s:
                # Half-day permission combinations like PER/CL, CL/PER, etc.
                perm_count += 0.5

        leave_used_map = {}
        for la in emp_leaves:
            code = ((la.leave_type.code if la.leave_type else "L") or "L").strip().upper()
            days_val = _effective_leave_days(la, period_start, period_end, emp_holidays)
            if days_val > 0 and code != "PER":
                leave_used_map[code] = leave_used_map.get(code, 0.0) + days_val
        for ccl_d in emp_ccl_dates:
            if period_start <= ccl_d <= period_end:
                leave_used_map["CCL"] = leave_used_map.get("CCL", 0.0) + 1.0
        leave_used_map["AB"] = float(ab_count)

        cl  = float(leave_used_map.get("CL",  0) or 0)
        ccl = float(leave_used_map.get("CCL", 0) or 0)
        od  = float(leave_used_map.get("OD",  0) or 0)
        vl  = float(leave_used_map.get("VL",  0) or 0)
        ml  = float(leave_used_map.get("ML",  0) or 0)
        lop = float(leave_used_map.get("LOP", 0) or 0)
        ab  = float(leave_used_map.get("AB",  0) or 0)
        wl  = float(leave_used_map.get("WL",  0) or 0)

        # Worked = CL + CCL + WL + OD + VL + ML + Present + Holiday
        # AB and LOP are NOT subtracted because:
        # - They are already excluded from Present count
        # - Present count only includes days marked as "P" (present)
        # - Worked = paid leaves + present days + holidays
        # - AB and LOP are tracked separately in their own columns
        worked = cl + ccl + wl + od + vl + ml + present_count + holidays_count
        worked = max(0.0, worked)

        # Salary deduction flags
        excess_perm = max(0, perm_count - perm_allowed)

        report_rows.append({
            "emp":           emp,
            "policy_name":   attendance_policy.policy_name if attendance_policy else "Shift",
            "day_cells":     day_cells,
            "present":       present_count,
            "absent":        ab_count,
            "holidays":      holidays_count,
            "working_days":  working_days,
            "worked":        worked,
            "permission":    perm_count,
            "leave_used":    leave_used_map,
            "late_entry":    late_entry_count,
            "early_exit":    early_exit_count,
            "excess_perm":   excess_perm,
        })

    # ════════════════════════════════════════
    # EXCEL EXPORT ONLY - Professional Design
    # ════════════════════════════════════════
    month_name = calendar.month_name[month]
    category_text = ""
    if category_id:
        try:
            category_obj = FacultyCategory.objects.get(id=category_id)
            category_text = f"{category_obj.category_name} "
        except Exception:
            category_text = ""

    if fmt.lower() == "pdf":
        from io import BytesIO
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        buffer = BytesIO()
        page_size = landscape(A4)
        left_margin = right_margin = 6 * mm
        content_width = page_size[0] - left_margin - right_margin
        doc = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            leftMargin=left_margin,
            rightMargin=right_margin,
            topMargin=7 * mm,
            bottomMargin=7 * mm,
            title="Attendance Report",
        )

        title_style = ParagraphStyle(
            "AttendanceTitle",
            fontName="Helvetica-Bold",
            fontSize=10,
            alignment=TA_CENTER,
            leading=12,
            spaceAfter=2,
        )
        sub_style = ParagraphStyle(
            "AttendanceSubTitle",
            fontName="Helvetica",
            fontSize=7,
            alignment=TA_CENTER,
            leading=9,
            spaceAfter=4,
        )
        cell_style = ParagraphStyle(
            "AttendanceCell",
            fontName="Helvetica",
            fontSize=4.6,
            leading=5.2,
            alignment=TA_CENTER,
        )
        left_cell_style = ParagraphStyle(
            "AttendanceLeftCell",
            parent=cell_style,
            alignment=TA_LEFT,
        )

        def pcell(value, style=cell_style):
            return Paragraph(str(value if value is not None else ""), style)

        def export_value(value):
            if isinstance(value, (int, float)):
                if value == 0:
                    return 0
                if value % 1:
                    return f"{value:.1f}"
                return int(value)
            return value if value is not None else 0

        summary_headers = ["CL", "CCL", "OD", "LOP", "AB", "VL", "ML", "P", "WD", "Worked", "PER", "WL", "H"]
        header_row = (
            ["#", "Emp ID", "Name", "Policy"]
            + [f"{d.strftime('%d')}\n{d.strftime('%a')[:2]}" for d in period_days]
            + summary_headers
        )
        table_data = [[pcell(value) for value in header_row]]

        for index, row in enumerate(report_rows, start=1):
            emp = row["emp"]
            lu = row["leave_used"]
            values = [
                index,
                emp.faculty_id,
                pcell(emp.name or "", left_cell_style),
                pcell(row.get("policy_name") or "", left_cell_style),
            ]
            values.extend((cell.get("display") or cell.get("status") or "") for cell in row["day_cells"])
            values.extend([
                lu.get("CL", 0) or 0,
                lu.get("CCL", 0) or 0,
                lu.get("OD", 0) or 0,
                lu.get("LOP", 0) or 0,
                lu.get("AB", 0) or 0,
                lu.get("VL", 0) or 0,
                lu.get("ML", 0) or 0,
                row["present"],
                row["working_days"],
                row["worked"],
                row["permission"],
                lu.get("WL", 0) or 0,
                row["holidays"],
            ])
            table_data.append([
                item if hasattr(item, "wrap") else pcell(export_value(item))
                for item in values
            ])

        fixed_widths = [15, 32, 68, 40]
        day_width = 10.5
        summary_width = max(
            16,
            (content_width - sum(fixed_widths) - (day_width * len(period_days))) / len(summary_headers),
        )
        col_widths = fixed_widths + ([day_width] * len(period_days)) + ([summary_width] * len(summary_headers))

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 4.6),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.black),
            ("LEFTPADDING", (0, 0), (-1, -1), 1.2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1.2),
            ("TOPPADDING", (0, 0), (-1, -1), 1.2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1.2),
        ]))

        story = [
            Paragraph("RAMCO INSTITUTE OF TECHNOLOGY", title_style),
            Paragraph(f"{category_text}Attendance Report {month_name} Month {year}", sub_style),
            Paragraph(f"{period_start.strftime('%d-%m-%Y')} to {period_end.strftime('%d-%m-%Y')}", sub_style),
            Spacer(1, 2 * mm),
            table,
            Spacer(1, 2 * mm),
            Paragraph(
                "Legend: P=Present | A=Absent | H=Holiday | SP=Single Punch | SH=Short Hours | "
                "MO=Missing OUT | PER=Permission | CCL=Compensatory Leave",
                sub_style,
            ),
        ]
        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        fname = f"attendance_report_{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{fname}"'
        return response

    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Print-friendly black and white styling.
    thin = Side(style="thin", color="000000")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(r, c, val, bg=None, fg="000000", sz=8, bold=True, wrap=True):
        cell = ws.cell(r, c, val)
        cell.font = Font(bold=bold, size=sz, color=fg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        cell.border = bdr
        return cell

    def data_cell(r, c, val, bg=None, sz=8, center=True, wrap=False):
        cell = ws.cell(r, c, val)
        cell.font = Font(size=sz)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center",
            wrap_text=wrap,
        )
        cell.border = bdr
        return cell

    # ══════════════════════════════════════════
    # INSTITUTIONAL HEADER ROWS
    # ══════════════════════════════════════════
    fixed_cols = 4
    summary_col_count = 13
    total_cols = fixed_cols + len(period_days) + summary_col_count

    # Row 1: Institute Name
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    inst_name = ws.cell(1, 1, "RAMCO INSTITUTE OF TECHNOLOGY")
    inst_name.font = Font(bold=True, size=14, color="000000")
    inst_name.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Row 2: Autonomous Institution
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    autonomous = ws.cell(2, 1, "An Autonomous Institution")
    autonomous.font = Font(bold=True, size=10, color="000000")
    autonomous.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Row 3: Location
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=total_cols)
    location = ws.cell(3, 1, "Rajapalayam")
    location.font = Font(bold=True, size=11, color="000000")
    location.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    # Row 4: Report Title with Category
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_cols)
    report_title_text = f"{category_text}Attendance Report {month_name} Month {year}"
    report_title = ws.cell(4, 1, report_title_text)
    report_title.font = Font(bold=True, size=11, color="000000")
    report_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 18

    # Row 5: Empty spacer
    ws.row_dimensions[5].height = 8

    # Row 6: Generated on timestamp
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=total_cols)
    generated_time = _dt.now().strftime("%d-%m-%Y %H:%M")
    gen_cell = ws.cell(6, 1, f"Generated on:{generated_time}")
    gen_cell.font = Font(size=9, color="000000")
    gen_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[6].height = 16

    # Row 7: Empty spacer
    ws.row_dimensions[7].height = 8

    # ══════════════════════════════════════════
    # HEADER ROW 8 - Column Headers
    # ══════════════════════════════════════════
    hdr_cell(8, 1, "S.No", sz=7)
    hdr_cell(8, 2, "Emp ID", sz=7)
    hdr_cell(8, 3, "Employee Name", sz=7)
    hdr_cell(8, 4, "Policy", sz=7)

    # Merge fixed columns for rows 8-9
    ws.merge_cells(start_row=8, start_column=1, end_row=9, end_column=1)
    ws.merge_cells(start_row=8, start_column=2, end_row=9, end_column=2)
    ws.merge_cells(start_row=8, start_column=3, end_row=9, end_column=3)
    ws.merge_cells(start_row=8, start_column=4, end_row=9, end_column=4)

    # Day numbers
    for i, d in enumerate(period_days):
        col = fixed_cols + 1 + i
        hdr_cell(8, col, d.strftime("%d"), sz=6)

    # Summary headers (span 2 rows)
    sum_start = fixed_cols + 1 + len(period_days)
    summary_headers = ["CL", "CCL", "OD", "LOP", "AB", "VL", "ML", "Present", "Working\nDays", "Worked", "Permission", "WL", "Holiday"]

    for i, sh in enumerate(summary_headers):
        c = sum_start + i
        hdr_cell(8, c, sh, sz=6)
        ws.merge_cells(start_row=8, start_column=c, end_row=9, end_column=c)

    # ══════════════════════════════════════════
    # HEADER ROW 9 - Day Names
    # ══════════════════════════════════════════
    for i, d in enumerate(period_days):
        col = fixed_cols + 1 + i
        hdr_cell(9, col, d.strftime("%a")[:2], sz=5)

    ws.row_dimensions[8].height = 18
    ws.row_dimensions[9].height = 14
    # HEADER ROW 9 - Day Names
    # ══════════════════════════════════════════
    for i, d in enumerate(period_days):
        col = 4 + i

    # ══════════════════════════════════════════
    # DATA ROWS
    # ══════════════════════════════════════════
    for ri, row in enumerate(report_rows):
        r = ri + 10  # Start from row 10 (after headers at rows 8-9)
        emp = row["emp"]

        # Fixed columns with auto-numbering (S.No)
        data_cell(r, 1, ri + 1, sz=7)
        data_cell(r, 2, emp.faculty_id, sz=7)
        data_cell(r, 3, emp.name or "", sz=7, center=False, wrap=True)
        data_cell(r, 4, row.get("policy_name") or "", sz=7, center=False, wrap=True)

        # Day status cells
        for i, cell in enumerate(row["day_cells"]):
            col = fixed_cols + 1 + i
            disp = cell.get("display") or cell.get("status") or ""
            c = ws.cell(r, col, disp)
            c.font = Font(size=6, bold=(disp in ("P", "A", "H")))
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = bdr

        # Summary columns
        lu = row["leave_used"]
        summary_values = [
            lu.get("CL", 0) or 0,
            lu.get("CCL", 0) or 0,
            lu.get("OD", 0) or 0,
            lu.get("LOP", 0) or 0,
            lu.get("AB", 0) or 0,
            lu.get("VL", 0) or 0,
            lu.get("ML", 0) or 0,
            row["present"],
            row["working_days"],
            row["worked"],
            row["permission"],
            lu.get("WL", 0) or 0,
            row["holidays"],
        ]

        for i, val in enumerate(summary_values):
            # Format values - always show 0 instead of empty
            if isinstance(val, (int, float)):
                if val == 0:
                    display_val = 0  # Show 0 instead of empty
                elif val % 1 != 0:  # Has decimal
                    display_val = f"{val:.1f}"
                else:  # Whole number
                    display_val = int(val)
            else:
                display_val = val if val is not None else 0

            data_cell(r, sum_start + i, display_val, sz=7)

    # ══════════════════════════════════════════
    # COLUMN WIDTHS
    # ══════════════════════════════════════════
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 11

    # Day columns
    for i in range(len(period_days)):
        ws.column_dimensions[get_column_letter(fixed_cols + 1 + i)].width = 3.4

    # Summary columns
    for i in range(summary_col_count):
        ws.column_dimensions[get_column_letter(sum_start + i)].width = 7

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.25, bottom=0.25, header=0.1, footer=0.1)
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "8:9"
    ws.print_title_cols = "A:D"

    # Freeze panes at E10 (freeze first 4 columns and first 9 rows).
    ws.freeze_panes = "E10"

    # ══════════════════════════════════════════
    # LEGEND ROW (at the bottom)
    # ══════════════════════════════════════════
    last_row = len(report_rows) + 11  # Adjusted for new header rows (data starts at row 10)
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=total_cols)
    legend_cell = ws.cell(last_row, 1,
        "Legend: P=Present | A=Absent | H=Holiday | SP=Single Punch | SH=Short Hours | MO=Missing OUT | PER/P=Permission | "
        "CCL=Compensatory Leave | CL=Casual Leave | OD=On Duty | LOP=Loss of Pay | "
        "VL=Vacation Leave | WL=Without Leave | ML=Medical Leave"
    )
    legend_cell.font = Font(size=8, italic=True, color="000000")
    legend_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[last_row].height = 30

    # ══════════════════════════════════════════
    # SAVE AND RETURN
    # ══════════════════════════════════════════
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    fname = f"attendance_report_{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response


# def bulk_approve_all_ccl(request):
#     """
#     Approve ALL pending CCL applications following the approval hierarchy.
#     This will approve level by level - Level 1 first, then Level 2, then Level 3, etc.
#     Processes in multiple rounds to handle cascading approvals.
#     """
#     from django.contrib import messages
#     from django.shortcuts import redirect
#     from django.utils import timezone
#     from django.db import transaction
#     from faculty_leave_management.models import CCL_Approvers_Data, CCL_Application

#     if request.method != 'POST':
#         messages.error(request, "Invalid request method.")
#         return redirect('emp_attandance_report')

#     total_approved = 0
#     total_completed = 0
#     max_rounds = 10  # Prevent infinite loops

#     try:
#         # Process in multiple rounds to handle hierarchical approvals
#         for round_num in range(max_rounds):
#             round_approved = 0

#             # Get all pending approver records, ordered by level
#             pending_approvers = CCL_Approvers_Data.objects.filter(
#                 status=CCL_Approvers_Data.Status.PENDING
#             ).select_related('ccl_application', 'approver_id').order_by('approver_level', 'id')

#             if not pending_approvers.exists():
#                 break  # No more pending approvals

#             # Process each pending approval
#             for cad in pending_approvers:
#                 try:
#                     with transaction.atomic():
#                         # Refresh from DB to get latest status
#                         cad.refresh_from_db()

#                         # Skip if already approved
#                         if cad.status != CCL_Approvers_Data.Status.PENDING:
#                             continue

#                         # Check if lower levels are approved
#                         lower_pending = CCL_Approvers_Data.objects.filter(
#                             ccl_application=cad.ccl_application,
#                             approver_level__lt=cad.approver_level,
#                         ).exclude(status=CCL_Approvers_Data.Status.APPROVED).exists()

#                         # Only approve if lower levels are done
#                         if not lower_pending:
#                             cad.status = CCL_Approvers_Data.Status.APPROVED
#                             cad.action_date = timezone.now()
#                             cad.remarks = f"Bulk approved by system (Level {cad.approver_level})"
#                             cad.save()

#                             round_approved += 1
#                             total_approved += 1

#                             # Check if all levels are now approved for this CCL
#                             ccl_app = cad.ccl_application
#                             all_approvers = CCL_Approvers_Data.objects.filter(ccl_application=ccl_app)

#                             pending_count = all_approvers.filter(status=CCL_Approvers_Data.Status.PENDING).count()
#                             rejected_count = all_approvers.filter(status=CCL_Approvers_Data.Status.REJECTED).count()

#                             if pending_count == 0 and rejected_count == 0:
#                                 # All levels approved - mark CCL as approved
#                                 ccl_app.status = 'APPROVED'
#                                 ccl_app.save()
#                                 total_completed += 1

#                 except Exception as e:
#                     continue

#             # If no approvals in this round, we're done
#             if round_approved == 0:
#                 break

#         if total_approved > 0:
#             messages.success(
#                 request,
#                 f"Successfully approved {total_approved} approval level(s). {total_completed} CCL application(s) fully approved."
#             )
#         else:
#             messages.info(request, "No CCL applications were ready for approval.")

#     except Exception as e:
#         messages.error(request, f"Error during bulk approval: {str(e)}")

#     return redirect('emp_attandance_report')


# ─────────────────────────────────────────────────────────────
# EMP PUNCH DETAILS — admin view showing per-day IN/OUT/duration
# ─────────────────────────────────────────────────────────────
def emp_punch_details(request):
    from faculty_management.models import general_information
    from faculty_leave_management.models import (
        DeviceInfo, ShiftDetail,
        Employee_Holidays, LeaveApplication, CCL_Application,
    )

    now = datetime.now()

    employees = (
        general_information.objects
        .filter(faculty_id__isnull=False)
        .select_related("designation", "shift")
        .order_by("faculty_id")
    )

    selected_faculty_id = (request.GET.get("faculty_id") or "").strip()
    ym = request.GET.get("ym", f"{now.year}-{now.month:02d}")
    try:
        year, month = [int(x) for x in ym.split("-")]
    except Exception:
        year, month = now.year, now.month

    active_tab = request.GET.get("tab", "history")

    punch_data = []
    missed_punch_data = []
    ccl_data = []
    selected_employee = None
    error_message = None

    if selected_faculty_id:
        try:
            faculty_id_int = int(selected_faculty_id)
        except ValueError:
            faculty_id_int = None

        if faculty_id_int:
            selected_employee = employees.filter(faculty_id=str(faculty_id_int)).first()
            employee_category_id = getattr(selected_employee, "category_id", None)

            # Shift times
            shift_start_time = None
            shift_end_time = None
            if selected_employee and selected_employee.shift:
                first_shift = ShiftDetail.objects.filter(
                    shift_master=selected_employee.shift
                ).order_by("shift_no").first()
                if first_shift:
                    shift_start_time = first_shift.start_time
                    shift_end_time = first_shift.end_time
            if not shift_start_time:
                from datetime import time as dtime
                shift_start_time = dtime(9, 0)
            if not shift_end_time:
                from datetime import time as dtime
                shift_end_time = dtime(16, 30)

            # Device master
            devices_qs = DeviceInfo.objects.filter(is_active=True)
            device_master_map = {}
            attendance_device_ids = []
            for d in devices_qs:
                dev_id = str(d.deviceid or "").strip()
                if not dev_id:
                    continue
                if d.is_attendance:
                    attendance_device_ids.append(dev_id)
                device_master_map[dev_id] = {"is_attendance": d.is_attendance}
            no_device_filter = len(attendance_device_ids) == 0

            # Query attendance DB
            rows_att = []
            try:
                query = (
                    f"SELECT UserId, LogDate, DeviceId "
                    f"FROM DeviceLogs_{month}_{year} "
                    f"WHERE UserId = %s ORDER BY LogDate ASC"
                )
                with connections["attendance_db"].cursor() as cur:
                    cur.execute(query, [faculty_id_int])
                    cols = [c[0] for c in cur.description]
                    rows_att = [dict(zip(cols, r)) for r in cur.fetchall()]
            except Exception as e:
                error_message = f"Attendance data unavailable: {e}"

            # Group by date
            grouped = {}
            for row in rows_att:
                log_dt = row.get("LogDate")
                if not log_dt:
                    continue
                dev_id = str(row.get("DeviceId") or "").strip()
                dk = log_dt.date() if hasattr(log_dt, "date") else log_dt
                is_att = True if no_device_filter else device_master_map.get(dev_id, {}).get("is_attendance", False)
                grouped.setdefault(dk, []).append({"logdate": log_dt, "is_attendance": is_att})

            # Role for holiday lookup
            user_role_id = None
            try:
                from user_accounts.models import USER
                role_rows = (
                    USER.objects.using("rit_approval_system")
                    .filter(Employee_id=str(faculty_id_int))
                    .values_list("role_id", flat=True)
                )
                candidate_role_ids = {int(rid) for rid in role_rows if rid}
                if candidate_role_ids:
                    counts = {
                        rid: _filter_holidays_for_category(
                            Employee_Holidays.objects.filter(
                                role_id=rid,
                                holiday_date__year=year,
                                holiday_date__month=month,
                            ),
                            employee_category_id,
                        ).count()
                        for rid in candidate_role_ids
                    }
                    user_role_id = sorted(candidate_role_ids, key=lambda r: (-counts.get(r, 0), r))[0]
            except Exception:
                pass

            # Approved leaves → date map  {date: leave_code}
            month_start = date(year, month, 1)
            month_end = date(year, month, calendar.monthrange(year, month)[1])
            leave_dates = {}
            try:
                for la in LeaveApplication.objects.filter(
                    faculty=selected_employee,
                    status="Approved",
                    from_date__lte=month_end,
                    to_date__gte=month_start,
                ).select_related("leave_type"):
                    cur_d = la.from_date
                    while cur_d <= la.to_date:
                        if month_start <= cur_d <= month_end:
                            code = (la.leave_type.code if la.leave_type else "LEAVE") or "LEAVE"
                            leave_dates[cur_d] = code.strip().upper()
                        cur_d += timedelta(days=1)
            except Exception:
                pass

            # Approved CCL → date map  {date: "CCL"}
            try:
                for ca in CCL_Application.objects.filter(
                    faculty=selected_employee,
                    status__iexact="Approved",
                    date__gte=month_start,
                    date__lte=month_end,
                ):
                    if ca.date and ca.date not in leave_dates:
                        leave_dates[ca.date] = "CCL"
            except Exception:
                pass

            # Any leave / CCL APPLICATION on a day, regardless of status
            # (Pending / Approved / Rejected). {date: [{code, status, type}]}
            applications_by_date = {}
            try:
                for la in LeaveApplication.objects.filter(
                    faculty=selected_employee,
                    from_date__lte=month_end,
                    to_date__gte=month_start,
                ).select_related("leave_type"):
                    code = ((la.leave_type.code if la.leave_type else "LEAVE") or "LEAVE").strip().upper()
                    st = (la.status or "Pending").strip().title()
                    cur_d = la.from_date
                    while cur_d <= la.to_date:
                        if month_start <= cur_d <= month_end:
                            applications_by_date.setdefault(cur_d, []).append(
                                {"code": code, "status": st, "type": "Leave"}
                            )
                        cur_d += timedelta(days=1)
            except Exception:
                pass
            try:
                for ca in CCL_Application.objects.filter(
                    faculty=selected_employee,
                    date__gte=month_start,
                    date__lte=month_end,
                ):
                    if ca.date:
                        st = (ca.status or "Pending").strip().title()
                        applications_by_date.setdefault(ca.date, []).append(
                            {"code": "CCL", "status": st, "type": "CCL"}
                        )
            except Exception:
                pass

            # Build per-day rows
            days_in_month = calendar.monthrange(year, month)[1]
            for day_num in range(1, days_in_month + 1):
                att_date = date(year, month, day_num)
                day_logs = grouped.get(att_date, [])
                att_sorted = sorted([l for l in day_logs if l["is_attendance"]], key=lambda x: x["logdate"])

                in_time_obj = att_sorted[0]["logdate"] if att_sorted else None
                out_time_obj = att_sorted[-1]["logdate"] if len(att_sorted) >= 2 else None
                in_time_str = in_time_obj.strftime("%H:%M:%S") if in_time_obj else ""
                out_time_str = out_time_obj.strftime("%H:%M:%S") if out_time_obj else ""

                duration_str = ""
                if in_time_obj and out_time_obj and out_time_obj > in_time_obj:
                    delta = out_time_obj - in_time_obj
                    tot_s = int(delta.total_seconds())
                    h, rem = divmod(tot_s, 3600)
                    m, s = divmod(rem, 60)
                    duration_str = f"{h:02d}:{m:02d}:{s:02d}"

                # Holiday only when one is entered for this date in the role-based
                # holiday table — weekends are not automatically holidays.
                hol_obj = _get_matching_holiday(user_role_id, att_date, employee_category_id)
                is_weekend = False
                is_holiday = bool(hol_obj)
                holiday_label = "Holiday" if hol_obj else ""

                leave_code = leave_dates.get(att_date, "")

                # IN remark
                is_late = False
                if in_time_obj:
                    shift_start_dt = datetime.combine(att_date, shift_start_time)
                    is_late = in_time_obj > shift_start_dt + timedelta(minutes=1)
                    in_remark = "Late" if is_late else "Ok"
                else:
                    in_remark = "No Punch"

                out_remark = "Ok" if out_time_obj else "No Punch"
                status_remark = holiday_label or (f"{leave_code} Approved" if leave_code else "")

                is_missed = bool(in_time_obj) ^ bool(out_time_obj)

                row_data = {
                    "date_str": att_date.strftime("%d-%m-%Y"),
                    "day_name": calendar.day_name[att_date.weekday()],
                    "in_time": in_time_str,
                    "out_time": out_time_str,
                    "duration": duration_str,
                    "in_remark": in_remark,
                    "out_remark": out_remark,
                    "status_remark": status_remark,
                    "leave_code": leave_code,
                    "is_late": is_late,
                    "is_holiday": is_holiday,
                    "is_weekend": is_weekend,
                    "is_missed": is_missed,
                    "total_punches": len(att_sorted),
                    "applications": applications_by_date.get(att_date, []),
                }
                punch_data.append(row_data)
                if is_missed:
                    missed_punch_data.append(row_data)

            # CCL data for this employee/month
            try:
                ccl_data = list(
                    CCL_Application.objects.filter(
                        faculty=selected_employee,
                        date__year=year,
                        date__month=month,
                    ).select_related("designation").order_by("date")
                )
            except Exception:
                pass

    summary = {
        "total": len(punch_data),
        "present": sum(1 for r in punch_data if r["in_time"] and r["out_time"] and not r["is_holiday"] and not r["status_remark"]),
        "absent": sum(1 for r in punch_data if not r["in_time"] and not r["is_holiday"] and not r["status_remark"]),
        "late": sum(1 for r in punch_data if r["is_late"]),
        "holiday": sum(1 for r in punch_data if r["is_holiday"]),
        "leave": sum(1 for r in punch_data if r["status_remark"] and "Approved" in r["status_remark"]),
        "single": sum(1 for r in punch_data if r["total_punches"] == 1 and not r["is_holiday"]),
    }

    context = {
        "employees": employees,
        "selected_faculty_id": selected_faculty_id,
        "selected_year": year,
        "selected_month": month,
        "selected_month_name": calendar.month_name[month],
        "ym": f"{year}-{month:02d}",
        "punch_data": punch_data,
        "selected_employee": selected_employee,
        "error_message": error_message,
        "summary": summary,
        "punch_details_url": "/faculty_leave_management/faculty/leave/punch_details/",
        "award_ccl_url": "/faculty_leave_management/faculty/leave/award_ccl_admin/",
    }
    return render(request, "faculty_leave_management/admin/punch_details.html", context)


# ─────────────────────────────────────────────────────────────
# AWARD CCL — admin awards compensatory leave to a faculty
# ─────────────────────────────────────────────────────────────
@check_permission("award_ccl_to_employee")
def award_ccl_admin(request):
    from decimal import Decimal
    from faculty_management.models import general_information
    from faculty_leave_management.models import CCL_Application, CCL_Claim, CCLTimingMaster

    employees = (
        general_information.objects
        .filter(faculty_id__isnull=False)
        .select_related("designation", "department")
        .order_by("faculty_id")
    )
    ccl_timing_rules = CCLTimingMaster.objects.filter(is_active=True).order_by("min_hours", "max_hours")

    selected_faculty_id = (request.GET.get("faculty_id") or "").strip()
    selected_academic_year = (
        request.GET.get("academic_year") or _academic_year_for_date(date.today())
    ).strip()
    existing_years = set(
        CCL_Application.objects.exclude(academic_year__isnull=True)
        .exclude(academic_year="")
        .values_list("academic_year", flat=True)
    )
    current_year = date.today().year
    academic_year_options = sorted(
        existing_years
        | {f"{year}-{year + 1}" for year in range(current_year - 3, current_year + 3)}
    )
    selected_employee = None
    ccl_history = []
    ccl_balance = None

    if selected_faculty_id:
        selected_employee = employees.filter(faculty_id=selected_faculty_id).first()
        if selected_employee:
            ccl_history = list(
                CCL_Application.objects.filter(
                    faculty=selected_employee,
                    academic_year=selected_academic_year,
                )
                .select_related("designation")
                .order_by("-date")[:60]
            )
            ccl_balance = CCL_Claim.objects.filter(
                faculty=selected_employee,
                academic_year=selected_academic_year,
            ).first()

    if request.method == "POST":
        fid = (request.POST.get("faculty_id") or "").strip()
        date_str = (request.POST.get("award_date") or "").strip()
        from_time_str = (request.POST.get("from_time") or "").strip()
        to_time_str = (request.POST.get("to_time") or "").strip()
        reason = (request.POST.get("reason") or "").strip()

        try:
            faculty = general_information.objects.filter(faculty_id=fid).first()
            if not faculty:
                messages.error(request, "Employee not found.")
                return redirect(f"{request.path}?faculty_id={fid}")

            award_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            from_time = datetime.strptime(from_time_str, "%H:%M").time()
            to_time = datetime.strptime(to_time_str, "%H:%M").time()
            worked_hours, days_val, session_name = _calculate_ccl_award_from_times(
                award_date,
                from_time,
                to_time,
            )

            academic_year = _academic_year_for_date(award_date)

            from user_accounts.models import USER
            user_obj = None
            try:
                user_obj = USER.objects.using("rit_approval_system").filter(
                    Employee_id=str(fid)
                ).first()
            except Exception:
                pass

            from faculty_leave_management.models import CCL_Approvers_Data
            from faculty_leave_management.views.flm_crud import _credit_ccl_on_approval

            with transaction.atomic():
                new_ccl = CCL_Application.objects.create(
                    user_id=user_obj.pk if user_obj else None,
                    faculty_id=faculty.pk,
                    designation_id=faculty.designation_id,
                    academic_year=academic_year,
                    date=award_date,
                    reason=reason,
                    days=days_val,
                    from_time=from_time,
                    to_time=to_time,
                    worked_hours=worked_hours,
                    session=session_name,
                    status="Approved",
                    is_claimed=False,
                )

                # Remove any auto-created approver rows — admin award needs no approval chain.
                CCL_Approvers_Data.objects.filter(ccl_application=new_ccl).delete()

                # Credit directly (no approval) — updates both the CCL_Claim ledger
                # and the faculty's CCL LeaveBalance (the balance the leave form spends
                # from / the applicant balance card shows), and marks it claimed.
                _credit_ccl_on_approval(new_ccl)

            messages.success(
                request,
                f"CCL of {days_val} day(s) awarded to {faculty.name} for {worked_hours} hour(s) on {award_date.strftime('%d %b %Y')}."
            )
            return redirect(f"{request.path}?faculty_id={fid}&academic_year={academic_year}")

        except Exception as e:
            messages.error(request, f"Error awarding CCL: {e}")

    context = {
        "employees": employees,
        "selected_faculty_id": selected_faculty_id,
        "selected_academic_year": selected_academic_year,
        "academic_year_options": academic_year_options,
        "selected_employee": selected_employee,
        "ccl_history": ccl_history,
        "ccl_balance": ccl_balance,
        "ccl_timing_rules": ccl_timing_rules,
        "punch_details_url": "/faculty_leave_management/faculty/leave/punch_details/",
        "award_ccl_url": "/faculty_leave_management/faculty/leave/award_ccl_admin/",
    }
    return render(request, "faculty_leave_management/admin/award_ccl_admin.html", context)


def _academic_year_for_date(value):
    return f"{value.year}-{value.year + 1}" if value.month >= 6 else f"{value.year - 1}-{value.year}"


def _calculate_ccl_award_from_times(award_date, from_time, to_time):
    from decimal import Decimal, ROUND_HALF_UP
    from faculty_leave_management.models import CCLTimingMaster

    start_dt = datetime.combine(award_date, from_time)
    end_dt = datetime.combine(award_date, to_time)
    if end_dt <= start_dt:
        raise ValueError("To time must be later than from time.")

    total_hours = Decimal(str((end_dt - start_dt).total_seconds() / 3600)).quantize(
        Decimal("0.01"),
        rounding=ROUND_HALF_UP,
    )
    timing_rules = CCLTimingMaster.objects.filter(is_active=True).order_by("-min_hours")
    matched_rule = None
    for rule in timing_rules:
        if total_hours > rule.min_hours and (rule.max_hours is None or total_hours <= rule.max_hours):
            matched_rule = rule
            break
    if not matched_rule:
        raise ValueError(
            f"No active CCL timing rule found for {total_hours} hour(s). Please configure CCL Timing Master."
        )

    days_val = matched_rule.ccl_days
    session_name = matched_rule.session_name
    return total_hours, days_val, session_name


def edit_leave(request):
    """
    Admin: per-employee leave allotment (balance) management.

    Lists every employee with their leave balance for the selected academic
    year + leave type. Admin edits the TOTAL days granted; remaining is stored
    as LeaveBalance.available = total - used. The leave application form
    hard-blocks anything beyond the remaining balance.
    """
    from datetime import date as _date
    from decimal import Decimal, InvalidOperation
    from django.core.paginator import Paginator
    from django.urls import reverse
    from django.db import transaction
    from django.conf import settings
    from faculty_leave_management.models import LeaveBalance
    from faculty_leave_management.views.flm_crud import _resolve_leave_allotment

    current_year = _date.today().year
    academic_years = [f"{y}-{y + 1}" for y in range(current_year - 1, current_year + 6)]
    default_year = getattr(settings, "ACADEMIC_YEAR", None) or f"{current_year}-{current_year + 1}"

    # CCL (code "CCL") is is_leave=True, so it is included and manageable here.
    leave_types = LeaveType.objects.filter(is_active=True, is_leave=True).order_by("name")

    def _balance_window(fac, ay, lt):
        """Start/end dates for a balance row so the leave/CCL forms (which look
        up balances by date range) can find it. Reuse the allotment window for
        the year; fall back to any allotment for the year, then a derived one."""
        a = _resolve_leave_allotment(fac, ay, lt)
        if a and a.start_date and a.end_date:
            return a.start_date, a.end_date
        a2 = LeaveAllotment.objects.filter(
            academic_year=ay, leave_type=lt,
            start_date__isnull=False, end_date__isnull=False,
        ).first()
        if a2:
            return a2.start_date, a2.end_date
        a3 = LeaveAllotment.objects.filter(
            academic_year=ay,
            start_date__isnull=False, end_date__isnull=False,
        ).first()
        if a3:
            return a3.start_date, a3.end_date
        try:
            y1, y2 = ay.split("-")
            return _date(int(y1), 6, 1), _date(int(y2), 5, 31)
        except Exception:
            return None, None

    # ---------------------------------
    # SAVE (POST)
    # ---------------------------------
    if request.method == "POST":
        action = request.POST.get("action", "save_balance")

        # preserve current filters on redirect
        back = (
            f"{reverse('edit_leave')}"
            f"?year={request.POST.get('academic_year', '')}"
        )
        q_val = (request.POST.get("q") or "").strip()
        page_val = (request.POST.get("page") or "").strip()
        if q_val:
            back += f"&q={q_val}"
        if page_val:
            back += f"&page={page_val}"

        # ---- Save employee profile (designation / category / shift) ----
        if action == "save_profile":
            faculty = get_object_or_404(
                general_information, id=request.POST.get("faculty_id")
            )
            designation_id = (request.POST.get("designation") or "").strip()
            category_id = (request.POST.get("category") or "").strip()
            shift_id = (request.POST.get("shift") or "").strip()
            gender_val = (request.POST.get("gender") or "").strip()
            valid_genders = {c[0] for c in general_information.GenderChoices.choices}

            try:
                with transaction.atomic():
                    faculty.designation = (
                        DesignationMaster.objects.filter(id=designation_id).first()
                        if designation_id else None
                    )
                    faculty.category = (
                        FacultyCategory.objects.filter(id=category_id).first()
                        if category_id else None
                    )
                    faculty.shift = (
                        ShiftMaster.objects.filter(id=shift_id).first()
                        if shift_id else None
                    )
                    faculty.gender = gender_val if gender_val in valid_genders else None
                    faculty.save(update_fields=["designation", "category", "shift", "gender"])

                    # Keep this employee's leave balances findable: designation is
                    # part of the lookup key used by the leave/CCL application forms.
                    LeaveBalance.objects.filter(faculty=faculty).update(
                        designation=faculty.designation
                    )
                messages.success(request, f"Profile updated for {faculty.name}.")
            except Exception as e:
                messages.error(request, f"Failed to update profile: {e}")
            return redirect(back)

        # ---- Save leave balances (one row = all leave types for one employee) ----
        faculty_id = request.POST.get("faculty_id")
        academic_year = (request.POST.get("academic_year") or "").strip()

        if not faculty_id or not academic_year:
            messages.error(request, "Missing data — could not save.")
            return redirect(back)

        faculty = get_object_or_404(general_information, id=faculty_id)

        updated, skipped = [], []
        try:
            with transaction.atomic():
                for lt in leave_types:
                    total_raw = (request.POST.get(f"total_{lt.id}") or "").strip()
                    orig_raw = (request.POST.get(f"orig_{lt.id}") or "").strip()
                    if total_raw == "":
                        continue

                    try:
                        # Allow fractional day allotments (e.g. 0.5); 0 is valid.
                        total_val = Decimal(total_raw)
                    except (InvalidOperation, TypeError, ValueError):
                        skipped.append(f"{lt.name} (not a number)")
                        continue

                    # only touch a leave type whose value the admin actually changed
                    try:
                        if orig_raw != "" and Decimal(orig_raw) == total_val:
                            continue
                    except (InvalidOperation, TypeError, ValueError):
                        pass

                    if total_val < 0:
                        skipped.append(f"{lt.name} (negative)")
                        continue

                    lb = LeaveBalance.objects.select_for_update().filter(
                        faculty=faculty,
                        leave_type=lt,
                        academic_year=academic_year,
                    ).first()

                    start_d, end_d = _balance_window(faculty, academic_year, lt)

                    if lb:
                        used = lb.used or 0
                        if total_val < used:
                            skipped.append(f"{lt.name} (below used {used})")
                            continue
                        lb.available = total_val - used
                        lb.designation = faculty.designation
                        if start_d and end_d:
                            lb.start_date = start_d
                            lb.end_date = end_d
                        lb.save()
                    else:
                        LeaveBalance.objects.create(
                            faculty=faculty,
                            designation=faculty.designation,
                            leave_type=lt,
                            academic_year=academic_year,
                            available=total_val,
                            used=0,
                            start_date=start_d,
                            end_date=end_d,
                        )
                    updated.append(lt.name)

            if updated:
                messages.success(
                    request,
                    f"Updated {', '.join(updated)} for {faculty.name}.",
                )
            if skipped:
                messages.warning(request, f"Skipped: {', '.join(skipped)}.")
            if not updated and not skipped:
                messages.info(request, "No changes to save.")
        except Exception as e:
            messages.error(request, f"Failed to save: {e}")

        return redirect(back)

    # ---------------------------------
    # LIST (GET)
    # ---------------------------------
    sel_year = (request.GET.get("year") or default_year).strip()
    q = (request.GET.get("q") or "").strip()

    lt_list = list(leave_types)

    employees = general_information.objects.select_related(
        "designation", "category", "department", "shift"
    )
    if q:
        flt = Q(name__icontains=q)
        if q.isdigit():
            flt |= Q(faculty_id=int(q))
        employees = employees.filter(flt)
    employees = employees.order_by("name")

    paginator = Paginator(employees, 25)
    page_obj = paginator.get_page(request.GET.get("page"))
    page_faculty = list(page_obj)

    # existing balances for this page, keyed by (faculty_id, leave_type_id)
    balances = {
        (b.faculty_id, b.leave_type_id): b
        for b in LeaveBalance.objects.filter(
            faculty_id__in=[f.id for f in page_faculty],
            leave_type__in=lt_list,
            academic_year=sel_year,
        )
    }

    # preload active allotments for the year into in-memory maps (category / role)
    cat_allot, role_allot = {}, {}
    for a in LeaveAllotment.objects.filter(
        academic_year=sel_year, active=True, leave_type__in=lt_list
    ):
        if a.category_id:
            cat_allot[(a.category_id, a.leave_type_id)] = a
        elif a.role_id:
            role_allot[(a.role_id, a.leave_type_id)] = a

    rows = []
    for f in page_faculty:
        cells = []
        for lt in lt_list:
            b = balances.get((f.id, lt.id))
            if b:
                used = b.used or 0
                remaining = b.available or 0
                total = remaining + used
                source = "custom"
            else:
                a = None
                if f.category_id:
                    a = cat_allot.get((f.category_id, lt.id))
                if not a and f.designation_id:
                    a = role_allot.get((f.designation_id, lt.id))
                total = a.default_allotment if a else 0
                used = 0
                remaining = total
                source = "allotment" if a else "none"
            cells.append({
                "lt": lt,
                "total": total,
                "used": used,
                "remaining": remaining,
                "source": source,
            })
        rows.append({"faculty": f, "cells": cells})

    context = {
        "academic_years": academic_years,
        "leave_types": lt_list,
        "sel_year": sel_year,
        "q": q,
        "rows": rows,
        "page_obj": page_obj,
        "paginator": paginator,
        # dropdown data for the "Edit profile" modal
        "all_categories": FacultyCategory.objects.filter(is_active=True).order_by("category_name"),
        "all_designations": DesignationMaster.objects.all().order_by("designation_name"),
        "all_shifts": ShiftMaster.objects.filter(is_active=True).order_by("shift_name"),
        "gender_choices": general_information.GenderChoices.choices,
    }
    return render(request, "faculty_leave_management/admin/edit_leave.html", context)


def upload_carry_forward(request):
    """Admin: bulk-add carry-forward leave counts from an Excel (.xlsx) upload.

    The Excel has an employee-id column plus one column per leave type
    (matched by code or name — e.g. CCL, VL, CL, OD ...). An empty cell is
    treated as 0. Each value is ADDED to that faculty's LeaveBalance.available
    for the selected academic year and leave type; a balance row is created if
    none exists. Re-uploading the same file adds the counts again.
    """
    from datetime import date as _date
    from decimal import Decimal, InvalidOperation
    from django.conf import settings
    from django.db import transaction
    from django.http import HttpResponse
    from faculty_leave_management.models import LeaveBalance
    from faculty_leave_management.views.flm_crud import _resolve_leave_allotment

    current_year = _date.today().year
    academic_years = [f"{y}-{y + 1}" for y in range(current_year - 1, current_year + 6)]
    default_year = getattr(settings, "ACADEMIC_YEAR", None) or f"{current_year}-{current_year + 1}"

    leave_types = list(
        LeaveType.objects.filter(is_active=True, is_leave=True).order_by("name")
    )

    def _norm(v):
        return str(v).strip().lower().replace(" ", "").replace("_", "") if v is not None else ""

    # ---- Download a blank template (.xlsx) ----
    if request.method == "GET" and request.GET.get("template"):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CarryForward"
        ws.append(["EmployeeID"] + [(lt.code or lt.name) for lt in leave_types])
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="carry_forward_template.xlsx"'
        wb.save(resp)
        return resp

    def _balance_window(fac, ay, lt):
        """Start/end dates for a balance row so the leave/CCL forms (which look
        up balances by date range) can find it."""
        a = _resolve_leave_allotment(fac, ay, lt)
        if a and a.start_date and a.end_date:
            return a.start_date, a.end_date
        a2 = LeaveAllotment.objects.filter(
            academic_year=ay, leave_type=lt,
            start_date__isnull=False, end_date__isnull=False,
        ).first()
        if a2:
            return a2.start_date, a2.end_date
        a3 = LeaveAllotment.objects.filter(
            academic_year=ay,
            start_date__isnull=False, end_date__isnull=False,
        ).first()
        if a3:
            return a3.start_date, a3.end_date
        try:
            y1, y2 = ay.split("-")
            return _date(int(y1), 6, 1), _date(int(y2), 5, 31)
        except Exception:
            return None, None

    results = []
    summary = None
    sel_year = default_year

    if request.method == "POST":
        import openpyxl
        sel_year = (request.POST.get("academic_year") or default_year).strip()
        excel_file = request.FILES.get("excel_file")

        wb = None
        if not excel_file:
            messages.error(request, "Please choose an Excel (.xlsx) file to upload.")
        elif not excel_file.name.lower().endswith(".xlsx"):
            messages.error(request, "Invalid file type. Please upload a .xlsx file.")
        else:
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
            except Exception as exc:
                messages.error(request, f"Unable to read Excel file: {exc}")
                wb = None

        if wb is not None:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)

            if not header:
                messages.error(request, "The uploaded file is empty.")
            else:
                norm_headers = [_norm(h) for h in header]

                emp_aliases = {
                    "employeeid", "empid", "employee", "employeeno", "employeenumber",
                    "facultyid", "empno", "id",
                }
                emp_col = next(
                    (idx for idx, h in enumerate(norm_headers) if h in emp_aliases),
                    None,
                )

                lt_by_key = {}
                for lt in leave_types:
                    if lt.code:
                        lt_by_key[_norm(lt.code)] = lt
                    if lt.name:
                        lt_by_key[_norm(lt.name)] = lt

                col_to_lt = {
                    idx: lt_by_key[h]
                    for idx, h in enumerate(norm_headers)
                    if idx != emp_col and h in lt_by_key
                }

                if emp_col is None:
                    messages.error(request, "Could not find an 'EmployeeID' column in the file.")
                elif not col_to_lt:
                    messages.error(
                        request,
                        "No recognizable leave-type columns were found. Use leave codes "
                        "(e.g. CCL, VL, CL) or names as headers.",
                    )
                else:
                    saved_count = skipped_count = applied_total = 0
                    try:
                        with transaction.atomic():
                            for r_i, row in enumerate(rows_iter, start=2):
                                if row is None or all(
                                    c is None or str(c).strip() == "" for c in row
                                ):
                                    continue

                                emp_raw = row[emp_col] if emp_col < len(row) else None
                                emp_str = str(emp_raw).strip() if emp_raw is not None else ""
                                if emp_str.endswith(".0"):
                                    emp_str = emp_str[:-2]

                                if not emp_str:
                                    results.append({"row": r_i, "employee": "-", "status": "Skipped", "message": "Missing employee id"})
                                    skipped_count += 1
                                    continue
                                if not emp_str.isdigit():
                                    results.append({"row": r_i, "employee": emp_str, "status": "Skipped", "message": "Employee id is not numeric"})
                                    skipped_count += 1
                                    continue

                                faculty = (
                                    general_information.objects
                                    .filter(faculty_id=int(emp_str))
                                    .select_related("designation")
                                    .first()
                                )
                                if not faculty:
                                    results.append({"row": r_i, "employee": emp_str, "status": "Skipped", "message": "No faculty found for this employee id"})
                                    skipped_count += 1
                                    continue

                                applied_parts = []
                                for c_idx, lt in col_to_lt.items():
                                    raw = row[c_idx] if c_idx < len(row) else None
                                    if raw is None or str(raw).strip() == "":
                                        cf = Decimal("0")
                                    else:
                                        try:
                                            cf = Decimal(str(raw).strip()).quantize(Decimal("0.01"))
                                        except (TypeError, ValueError, InvalidOperation):
                                            applied_parts.append(f"{lt.code or lt.name}: invalid (skipped)")
                                            continue
                                    if cf <= 0:
                                        continue  # empty / zero → nothing to add

                                    lb = LeaveBalance.objects.select_for_update().filter(
                                        faculty=faculty,
                                        leave_type=lt,
                                        academic_year=sel_year,
                                    ).first()
                                    start_d, end_d = _balance_window(faculty, sel_year, lt)

                                    if lb:
                                        lb.available = (lb.available or 0) + cf
                                        lb.designation = faculty.designation
                                        if start_d and end_d:
                                            lb.start_date = start_d
                                            lb.end_date = end_d
                                        lb.save()
                                    else:
                                        LeaveBalance.objects.create(
                                            faculty=faculty,
                                            designation=faculty.designation,
                                            leave_type=lt,
                                            academic_year=sel_year,
                                            available=cf,
                                            used=0,
                                            start_date=start_d,
                                            end_date=end_d,
                                        )
                                    applied_parts.append(f"{lt.code or lt.name}: +{cf}")
                                    applied_total += cf

                                # A valid employee row is always a success, even if
                                # every value is 0/empty — "Skipped" is reserved for
                                # genuine errors (bad id / faculty not found).
                                if applied_parts:
                                    results.append({"row": r_i, "employee": f"{faculty.name} ({emp_str})", "status": "Saved", "message": ", ".join(applied_parts)})
                                else:
                                    results.append({"row": r_i, "employee": f"{faculty.name} ({emp_str})", "status": "Saved", "message": "No carry-forward to add (all 0/empty)"})
                                saved_count += 1
                    except Exception as exc:
                        messages.error(request, f"Upload failed and was rolled back: {exc}")
                        results = []
                    else:
                        summary = {
                            "saved": saved_count,
                            "skipped": skipped_count,
                            "applied_total": applied_total,
                            "academic_year": sel_year,
                        }
                        if saved_count:
                            messages.success(
                                request,
                                f"Carry-forward applied to {saved_count} employee(s) for "
                                f"{sel_year} (+{applied_total} day(s) total).",
                            )
                        if skipped_count and not saved_count:
                            messages.warning(request, f"No rows applied. {skipped_count} row(s) skipped.")
                        elif skipped_count:
                            messages.warning(request, f"{skipped_count} row(s) skipped — see the results below.")

    context = {
        "academic_years": academic_years,
        "default_year": default_year,
        "sel_year": sel_year,
        "leave_types": leave_types,
        "results": results,
        "summary": summary,
    }
    return render(request, "faculty_leave_management/admin/upload_carry_forward.html", context)


from faculty_leave_management.models import Faculty_Leave_Page_Permission

FLPP_DB = "rit_approval_system"


def _flpp_user_display_map(user_ids):
    """USER/Department/Role live in the `rit_approval_system` database, which
    is unrelated to Faculty_Leave_Page_Permission's default database — so the
    linked user is always looked up separately by id, never through a FK
    join."""
    user_ids = {uid for uid in user_ids if uid is not None}
    if not user_ids:
        return {}

    rows = (
        USER.objects.using(FLPP_DB)
        .filter(id__in=user_ids)
        .select_related("Department", "role")
    )
    return {
        row.id: {
            "id": row.id,
            "name": row.username,
            "employee_id": row.Employee_id,
            "department_id": row.Department_id,
            "department_label": getattr(row.Department, "Department", None),
            "role_id": row.role_id,
            "role_label": getattr(row.role, "role", None),
        }
        for row in rows
    }
from django.core.paginator import Paginator

@no_cache
@is_super_user('faculty_leave_management')
def faculty_leave_page_permission(request):
    delete_id = request.GET.get("delete")
    if delete_id:
        Faculty_Leave_Page_Permission.objects.filter(id=delete_id).delete()
        messages.success(request, "Permission entry deleted successfully.")
        return redirect("faculty_leave_page_permission")

    if request.method == "POST":
        action = request.POST.get("action") or "save"

        filter_params = {
            key: request.POST.get(key)
            for key in ("q", "department", "role", "role_count")
            if request.POST.get(key)
        }
        redirect_url = reverse("faculty_leave_page_permission")
        if filter_params:
            redirect_url = f"{redirect_url}?{urlencode(filter_params)}"

        if action == "delete":
            row_id = request.POST.get("id")
            Faculty_Leave_Page_Permission.objects.filter(id=row_id).delete()
            messages.success(request, "Permission entry deleted successfully.")
            return redirect(redirect_url)

        user_ids = request.POST.getlist("user_ids")
        is_hidden = request.POST.get("is_hidden") == "true"

        clean_ids = []
        for uid in user_ids:
            try:
                clean_ids.append(int(uid))
            except (TypeError, ValueError):
                continue

        if not clean_ids:
            messages.error(request, "Select at least one user.")
            return redirect(redirect_url)

        for uid in clean_ids:
            Faculty_Leave_Page_Permission.objects.update_or_create(
                user_id=uid,
                defaults={"is_hidden": is_hidden},
            )

        status_label = "hidden" if is_hidden else "visible"
        messages.success(
            request,
            f"Leave page marked {status_label} for {len(clean_ids)} user(s)."
        )
        return redirect(redirect_url)

    # ---------------- FILTERS ----------------
    FLPP_PAGE_SIZE = 50

    search_query = (request.GET.get("q") or "").strip()
    selected_department = request.GET.get("department") or ""
    selected_role = request.GET.get("role") or ""
    # "" = all, "single" = users with exactly one role, "multiple" = users with many roles
    selected_role_count = request.GET.get("role_count") or ""
    edit_id = request.GET.get("edit")

    departments = list(Department.objects.using(FLPP_DB).all().order_by("Department"))
    roles = list(Role.objects.using(FLPP_DB).all().order_by("role"))

    selected_department_obj = next(
        (d for d in departments if str(d.id) == str(selected_department)), None
    ) if selected_department else None
    selected_role_obj = next(
        (r for r in roles if str(r.id) == str(selected_role)), None
    ) if selected_role else None

    # ---------------- USERS AVAILABLE FOR ASSIGNMENT (not paginated) ----------------
    users_qs = (
        USER.objects.using(FLPP_DB)
        .filter(is_student=False, is_active=True)
        .select_related("Department", "role")
        .order_by("username")
    )
    if selected_department:
        users_qs = users_qs.filter(Department_id=selected_department)
    if selected_role:
        users_qs = users_qs.filter(role_id=selected_role)
    if search_query:
        users_qs = users_qs.filter(
            Q(username__icontains=search_query)
            | Q(Employee_id__icontains=search_query)
            | Q(Department__Department__icontains=search_query)
            | Q(role__role__icontains=search_query)
        )

    users = list(users_qs)

    # ---------------- FILTER BY NUMBER OF ROLES PER USER ----------------
    # A person with multiple roles has multiple USER rows sharing one Employee_id.
    # Counts are computed over the whole active non-student population (independent
    # of the department/role filters above) so a user filtered to a single role is
    # still recognised as a "multiple role" user.
    if selected_role_count in ("single", "multiple"):
        from django.db.models import Count

        role_counts = dict(
            USER.objects.using(FLPP_DB)
            .filter(is_student=False, is_active=True)
            .values("Employee_id")
            .annotate(role_total=Count("role_id", distinct=True))
            .values_list("Employee_id", "role_total")
        )
        if selected_role_count == "single":
            users = [u for u in users if role_counts.get(u.Employee_id, 0) <= 1]
        else:  # multiple
            users = [u for u in users if role_counts.get(u.Employee_id, 0) > 1]

    # ---------------- EXISTING PERMISSIONS (searchable + paginated) ----------------
    all_permissions = list(Faculty_Leave_Page_Permission.objects.all().order_by("-id"))
    user_map = _flpp_user_display_map(p.user_id for p in all_permissions)
    for permission in all_permissions:
        permission.user_info = user_map.get(permission.user_id)

    perm_search_query = (request.GET.get("pq") or "").strip()
    if perm_search_query:
        term = perm_search_query.lower()

        def _permission_matches(permission):
            info = permission.user_info or {}
            haystack = " ".join(
                str(info.get(key) or "")
                for key in ("employee_id", "name", "department_label", "role_label")
            )
            return term in haystack.lower()

        filtered_permissions = [p for p in all_permissions if _permission_matches(p)]
    else:
        filtered_permissions = all_permissions

    permissions_paginator = Paginator(filtered_permissions, FLPP_PAGE_SIZE)
    permissions_page = permissions_paginator.get_page(request.GET.get("perm_page"))

    perm_querystring = urlencode({
        k: v for k, v in {
            "q": search_query,
            "department": selected_department,
            "role": selected_role,
            "role_count": selected_role_count,
            "pq": perm_search_query,
        }.items() if v
    })

    edit_obj = None
    edit_selected_user_ids = set()
    if edit_id:
        edit_obj = Faculty_Leave_Page_Permission.objects.filter(id=edit_id).first()
        if edit_obj:
            edit_obj.user_info = _flpp_user_display_map([edit_obj.user_id]).get(edit_obj.user_id)
            edit_selected_user_ids = {edit_obj.user_id}

    context = {
        "users": users,
        "permissions": permissions_page,
        "assigned_user_ids": {p.user_id for p in all_permissions},
        "departments": departments,
        "roles": roles,
        "search_query": search_query,
        "selected_department": selected_department,
        "selected_role": selected_role,
        "selected_role_count": selected_role_count,
        "selected_department_label": getattr(selected_department_obj, "Department", None),
        "selected_role_label": getattr(selected_role_obj, "role", None),
        "perm_search_query": perm_search_query,
        "perm_querystring": perm_querystring,
        "edit_obj": edit_obj,
        "edit_selected_user_ids": edit_selected_user_ids,
    }
    return render(request, "faculty_leave_management/admin/faculty_leave_page_permission.html", context)






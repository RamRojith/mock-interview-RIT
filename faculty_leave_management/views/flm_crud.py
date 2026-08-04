import json
import logging

from requests import request
from user_accounts.decorators import no_cache,is_super_user,check_permission
from django.utils.decorators import method_decorator
from django.utils import timezone
from django.utils import timezone

from django.urls import reverse
from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.generic import CreateView, UpdateView, ListView,DeleteView
from django.urls import reverse_lazy
from faculty_leave_management.models import LeaveAllotment,LeaveType,LeaveApplication,general_information,LeaveApproversData,LeaveBalance
from django.contrib import messages
from django.db import IntegrityError, transaction, connections
from django.db.models import Q, Prefetch, Exists, OuterRef, Value, CharField
from django.db.models.functions import Concat, Cast
from django.http import JsonResponse
from django.core.paginator import Paginator, EmptyPage
from functools import wraps
from faculty_leave_management.forms import LeaveApplicationForm,LeaveApplicationFormSet
from faculty_leave_management.decorators import faculty_leave_management

from user_accounts.models import Department, Add_Department
from user_accounts.decorators import no_cache, faculty_login_required
from django.core.exceptions import ValidationError
from datetime import date, datetime, timedelta
from django.db import transaction
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils import timezone

from faculty_leave_management.models import *
from django.utils.dateparse import parse_date, parse_time

from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.shortcuts import redirect, render, get_object_or_404
from django.core.exceptions import ValidationError


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.core.exceptions import ValidationError
from django.utils import timezone


import os
from collections import defaultdict
from django.conf import settings
from django.http import HttpResponse
from django.contrib.staticfiles import finders
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Spacer,
    Paragraph,
    KeepTogether,
)
from reportlab.lib.utils import ImageReader
from django.utils.timezone import now

# import your model
# from faculty_leave_management.models import CCL_Claim


import os
from collections import defaultdict

from django.conf import settings
from django.contrib.staticfiles import finders
from django.http import HttpResponse
from django.utils import timezone

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Spacer,
    Paragraph,
)
from faculty_leave_management.models import PermissionTimingMaster

from django.conf import settings

# @check_permission("leave_application_form")
# def leave_application_form(request):
#     # print("acamdemic year settings:",settings.ACADEMIC_YEAR)
#     """
#     Create / Update / Delete:
#     - Only Pending leaves are editable
#     - AND only if no approver has already Approved/Rejected (approval process started)
#     """

#     # ---- Get Current User ----
#     ext_user = USER.objects.using("rit_approval_system").filter(Employee_id=request.user.Employee_id, is_active=True).first()
#     faculty = general_information.objects.filter(faculty_id=request.user.Employee_id).select_related("designation").first()

#     if not ext_user:
#         messages.error(request, "User information not found in external system.")
#         return redirect("dashboard")

#     if not faculty:
#         messages.error(request, "Faculty record not found in local database.")
#         return redirect("dashboard")

#     creator_role_id = ext_user.role_id

#     # ----------------------------
#     # ✅ Helper: lock rule
#     # ----------------------------
#     def is_leave_locked(leave: LeaveApplication) -> bool:
#         # not pending -> locked
#         if (leave.status or "").lower() != "pending":
#             return True

#         # if any approver already approved/rejected -> locked
#         return LeaveApproversData.objects.filter(
#             leave_application=leave
#         ).exclude(
#             status=LeaveApproversData.Status.PENDING
#         ).exists()

#     # ----------------------------
#     # ✅ DELETE (ONLY IF NOT LOCKED)
#     # ----------------------------
#     if request.method == "GET" and request.GET.get("delete_id"):
#         leave = get_object_or_404(LeaveApplication, id=request.GET.get("delete_id"), faculty=faculty)

#         if is_leave_locked(leave):
#             messages.error(request, "This leave cannot be deleted (already processed / not pending).")
#             return redirect("leave_application_form")

#         try:
#             with transaction.atomic():
#                 # refund leave balance (if exists)
#                 days = leave.days

#                 lb = LeaveBalance.objects.select_for_update().filter(
#                     faculty=leave.faculty,
#                     designation=leave.designation,
#                     leave_type=leave.leave_type,
#                     academic_year=leave.academic_year,
                    
#                 ).first()

#                 if lb:
#                     lb.available = (lb.available or 0) + days
#                     lb.used = max(0, (lb.used or 0) - days)
#                     lb.save(update_fields=["available", "used"])

#                 # remove approver rows
#                 LeaveApproversData.objects.filter(leave_application=leave).delete()
#                 leave.delete()

#             messages.success(request, "Leave application deleted successfully.")
#         except Exception as e:
#             messages.error(request, f"Failed to delete leave: {str(e)}")

#         return redirect("leave_application_form")

#     # ----------------------------
#     # ✅ POST (CREATE / UPDATE)
#     # ----------------------------
#     if request.method == "POST":
#         action = request.POST.get("action", "create")  # create | update
#         leave_id = request.POST.get("leave_id")

#         leave_type_id = request.POST.get("leave_type")
#         academic_year = request.POST.get("academic_year")
#         from_date = parse_date(request.POST.get("from_date"))
#         to_date = parse_date(request.POST.get("to_date"))
#         reason = request.POST.get("reason")
#         session_id = request.POST.get("session")

#         if not all([leave_type_id, academic_year, from_date, to_date]):
#             messages.error(request, "Please fill in all required fields.")
#             return redirect("leave_application_form")

#         if from_date > to_date:
#             messages.error(request, "From date cannot be after To date.")
#             return redirect("leave_application_form")

#         days = (to_date - from_date).days + 1

#         try:
#             with transaction.atomic():

#                 # =============================
#                 # ✅ UPDATE (ONLY IF NOT LOCKED)
#                 # =============================
#                 if action == "update":
#                     leave = get_object_or_404(LeaveApplication, id=leave_id, faculty=faculty)

#                     if is_leave_locked(leave):
#                         messages.error(request, "This leave cannot be edited (already processed / not pending).")
#                         return redirect("leave_application_form")

#                     # Overlap check excluding itself
#                     overlap_qs = LeaveApplication.objects.filter(
#                         faculty=faculty,
#                         status__in=["Pending", "Approved", "Pre-approved"],
#                     ).exclude(id=leave.id).filter(
#                         Q(from_date__lte=to_date) & Q(to_date__gte=from_date)
#                     )
#                     if overlap_qs.exists():
#                         messages.error(request, "You already have a leave overlapping this period.")
#                         return redirect("leave_application_form")

#                     # Refund old balance
#                     old_days = leave.days
#                     old_lb = LeaveBalance.objects.select_for_update().filter(
#                         faculty=leave.faculty,
#                         designation=leave.designation,
#                         leave_type=leave.leave_type,
#                         academic_year=leave.academic_year,
#                     ).first()
#                     if old_lb:
#                         old_lb.available = (old_lb.available or 0) + old_days
#                         old_lb.used = max(0, (old_lb.used or 0) - old_days)
#                         old_lb.save(update_fields=["available", "used"])

#                     # Update leave with the session field
#                     leave.academic_year = academic_year
#                     leave.leave_type_id = leave_type_id
#                     leave.from_date = from_date
#                     leave.to_date = to_date
#                     leave.reason = reason
#                     leave.designation = faculty.designation
#                     leave.status = "Pending"
#                     leave.session_id = session_id  # Save the session field
#                     leave.save()

#                     # Deduct new balance
#                     new_lb = LeaveBalance.objects.select_for_update().filter(
#                         faculty=leave.faculty,
#                         designation=leave.designation,
#                         leave_type=leave.leave_type,
#                         academic_year=leave.academic_year,
#                         start_date__lte=leave.from_date,
#                         end_date__gte=leave.to_date,
#                     ).first()

#                     if not new_lb:
#                         new_lb = LeaveBalance.objects.create(
#                             faculty=leave.faculty,
#                             designation=leave.designation,
#                             leave_type=leave.leave_type,
#                             academic_year=leave.academic_year,
#                             available=0,
#                             used=0,
#                             start_date=leave.from_date,
#                             end_date=leave.to_date,
#                         )

#                     new_lb.available = max(0, (new_lb.available or 0) - days)
#                     new_lb.used = (new_lb.used or 0) + days
#                     new_lb.save(update_fields=["available", "used"])

#                     # Rebuild approver rows
#                     LeaveApproversData.objects.filter(leave_application=leave).delete()
#                     _create_approver_chain_for_leave(
#                         leave=leave,
#                         creator_role_id=creator_role_id,
#                         creator_faculty=faculty
#                     )

#                     messages.success(request, "Leave updated successfully.")
#                     return redirect("leave_application_form")

#                 # =============================
#                 # ✅ CREATE
#                 # =============================
#                 overlap_qs = LeaveApplication.objects.filter(
#                     faculty=faculty,
#                     status__in=["Pending", "Approved", "Pre-approved"]
#                 ).filter(Q(from_date__lte=to_date) & Q(to_date__gte=from_date))

#                 if overlap_qs.exists():
#                     messages.error(request, "You already have a leave overlapping this period.")
#                     return redirect("leave_application_form")

#                 new_leave = LeaveApplication.objects.create(
#                     user_id=ext_user.id,
#                     faculty=faculty,
#                     designation=faculty.designation,
#                     academic_year=academic_year,
#                     leave_type_id=leave_type_id,
#                     from_date=from_date,
#                     to_date=to_date,
#                     reason=reason,
#                     status="Pending",
#                     session_id=session_id,  # Save the session when creating
#                 )

#                 # deduct balance
#                 lb = LeaveBalance.objects.select_for_update().filter(
#                     faculty=new_leave.faculty,
#                     designation=new_leave.designation,
#                     leave_type=new_leave.leave_type,
#                     academic_year=new_leave.academic_year,
#                     start_date__lte=new_leave.from_date,
#                     end_date__gte=new_leave.to_date,
#                 ).first()
                

#                 if not lb:
#                     allotment = LeaveAllotment.objects.filter(
#                         academic_year=academic_year,
#                         role=faculty.designation,
#                         leave_type_id=leave_type_id,
#                         active=True
#                     ).first()

#                     default_days = allotment.default_allotment if allotment else 0

#                     lb = LeaveBalance.objects.create(
#                         faculty=faculty,
#                         designation=faculty.designation,
#                         leave_type_id=leave_type_id,
#                         academic_year=academic_year,
#                         available=default_days,
#                         used=0,
#                         start_date=allotment.start_date if allotment else from_date,
#                         end_date=allotment.end_date if allotment else to_date,
#                     )

#                 lb.available = max(0, (lb.available or 0) - days)
#                 lb.used = (lb.used or 0) + days
#                 lb.save(update_fields=["available", "used"])

#                 # create approver rows
#                 created_rows = _create_approver_chain_for_leave(
#                     leave=new_leave,
#                     creator_role_id=creator_role_id,
#                     creator_faculty=faculty
#                 )

#                 if created_rows == 0:
#                     new_leave.status = "Pre-approved"
#                     new_leave.save(update_fields=["status"])

#                 messages.success(request, f"Leave applied successfully ({days} day{'s' if days > 1 else ''}).")
#                 return redirect("leave_application_form")

#         except ValidationError as e:
#             messages.error(request, e.messages[0] if e.messages else "Validation error.")
#         except Exception as e:
#             messages.error(request, f"An unexpected error occurred: {str(e)}")

#         return redirect("leave_application_form")

#     # ----------------------------
#     # ✅ FILTERS (GET)
#     # ----------------------------
#     fy = request.GET.get("fy", "").strip()
#     lt = request.GET.get("lt", "").strip()
#     status_f = request.GET.get("st", "").strip()
#     df = parse_date(request.GET.get("df") or "")
#     dt = parse_date(request.GET.get("dt") or "")

#     leaves_qs = LeaveApplication.objects.filter(faculty=faculty).select_related("leave_type", "designation").order_by("-requested_date")

#     if fy:
#         leaves_qs = leaves_qs.filter(academic_year=fy)
#     if lt:
#         leaves_qs = leaves_qs.filter(leave_type_id=lt)
#     if status_f:
#         leaves_qs = leaves_qs.filter(status=status_f)

#     # date range filter (overlap range)
#     if df and dt:
#         leaves_qs = leaves_qs.filter(from_date__lte=dt, to_date__gte=df)
#     elif df:
#         leaves_qs = leaves_qs.filter(to_date__gte=df)
#     elif dt:
#         leaves_qs = leaves_qs.filter(from_date__lte=dt)

#     leaves = list(leaves_qs)

#     # ----------------------------
#     # ✅ lock_map + approver status map
#     # ----------------------------
#     locked_ids = set(
#         LeaveApproversData.objects.filter(leave_application_id__in=[l.id for l in leaves])
#         .exclude(status=LeaveApproversData.Status.PENDING)
#         .values_list("leave_application_id", flat=True)
#     )

#     lock_map = {l.id: ((l.status or "") != "Pending" or l.id in locked_ids) for l in leaves}

#     # Approver status list (name + designation)
#     approvals_qs = (
#         LeaveApproversData.objects.filter(leave_application_id__in=[l.id for l in leaves])
#         .select_related("approver_id", "approver_id__designation")
#         .order_by("approver_level", "id")
#     )

#     approver_status_map = {}
#     for row in approvals_qs:
#         appr = row.approver_id
#         approver_emp = getattr(appr, "faculty_id", "—")
#         approver_name = getattr(appr, "name", "") or getattr(appr, "username", "")
#         approver_desig = getattr(getattr(appr, "designation", None), "designation_name", None) or str(getattr(appr, "designation", "") or "")

#         approver_status_map.setdefault(row.leave_application_id, []).append({
#             "level": row.approver_level,
#             "emp": approver_emp,
#             "name": approver_name,
#             "desig": approver_desig,
#             "status": row.status,
#         })

#     # ----------------------------
#     # ✅ COUNTS
#     # ----------------------------
#     total_count = len(leaves)
#     pending_count = sum(1 for l in leaves if (l.status or "") == "Pending")
#     approved_count = sum(1 for l in leaves if (l.status or "") == "Approved")
#     rejected_count = sum(1 for l in leaves if (l.status or "") == "Rejected")
#     preapproved_count = sum(1 for l in leaves if (l.status or "") == "Pre-approved")
#     locked_count = sum(1 for l in leaves if lock_map.get(l.id))
#     editable_count = total_count - locked_count

#     leave_types = LeaveType.objects.all().order_by("name")
#     academic_years = (
#         LeaveApplication.objects.filter(faculty=faculty)
#         .values_list("academic_year", flat=True)
#         .distinct()
#         .order_by("academic_year")
#     )



#     # ----------------------------
#     # ✅ LEAVE BALANCE SUMMARY
#     # ----------------------------
#     leave_balance_summary = []

#     for lt in LeaveType.objects.all():
#         lb = LeaveBalance.objects.filter(
#             faculty=faculty,
#             leave_type=lt,
#             academic_year=settings.ACADEMIC_YEAR
#         ).first()

#         # fallback to allotment if balance not created yet
#         allotment = LeaveAllotment.objects.filter(
#             academic_year=settings.ACADEMIC_YEAR,
#             role=faculty.designation,
#             leave_type=lt,
#             active=True
#         ).first()

#         total = allotment.default_allotment if allotment else 0
#         used = lb.used if lb else 0
#         available = total - used

#         leave_balance_summary.append({
#             "leave_type": lt.name,
#             "total": total,
#             "used": used,
#             "available": available,
#         })
    
#     return render(
#         request,
#         "faculty_leave_management/leave_application.html",
#         {
#             "leaves": leaves,
#             "leave_types": leave_types,
#             "academic_years": academic_years,

#             # filter selections
#             "fy": fy,
#             "lt": lt,
#             "st": status_f,
#             "df": request.GET.get("df", ""),
#             "dt": request.GET.get("dt", ""),

#             # maps
#             "lock_map": lock_map,
#             "approver_status_map": approver_status_map,
#             "sessions": PermissionTimingMaster.objects.filter(is_active=True),
#             # counts
#             "total_count": total_count,
#             "pending_count": pending_count,
#             "approved_count": approved_count,
#             "rejected_count": rejected_count,
#             "preapproved_count": preapproved_count,
#             "locked_count": locked_count,
#             "editable_count": editable_count,
#             "academic_year": settings.ACADEMIC_YEAR,
#             "leave_balance_summary": leave_balance_summary,
#         },
#     )


from decimal import Decimal, ROUND_HALF_UP

# TEMP DEBUG: traces every LeaveBalance mutation (site tag, row pk, faculty,
# leave type, before/after available+used, lookup path taken). Safe to remove
# once the CCL double-deduction issue is confirmed fixed in production.
balance_debug_logger = logging.getLogger("faculty_leave_management.balance_debug")


def _log_balance_change(site, lb, faculty=None, before=None, extra=""):
    """TEMP DEBUG helper — logs a single LeaveBalance write with enough context
    to tell which code path performed it and whether it reused an existing row
    or created a new one."""
    try:
        fac_id = getattr(faculty or getattr(lb, "faculty", None), "faculty_id", None)
        lt_code = getattr(getattr(lb, "leave_type", None), "code", None)
        before_txt = f"before(available={before[0]}, used={before[1]}) " if before else ""
        balance_debug_logger.info(
            "CCL_BALANCE_DEBUG site=%s lb_id=%s faculty=%s leave_type=%s academic_year=%s "
            "%safter(available=%s, used=%s) %s",
            site, getattr(lb, "id", None), fac_id, lt_code,
            getattr(lb, "academic_year", None), before_txt,
            getattr(lb, "available", None), getattr(lb, "used", None), extra,
        )
    except Exception:
        balance_debug_logger.exception("CCL_BALANCE_DEBUG logging failed at site=%s", site)


# ---------------------------------------------------------------------------
# Shared building blocks for the JSON/API-driven CCL listing endpoints
# (CCL Application self-service page + CCL Approval page).
# ---------------------------------------------------------------------------

def _json_forbidden(view_func):
    """Wrap a view already protected by check_permission/faculty_leave_management
    so a permission failure comes back as JSON instead of an HTML 403 page —
    required for endpoints consumed by fetch()."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        response = view_func(request, *args, **kwargs)
        if response.status_code == 403 and not isinstance(response, JsonResponse):
            return JsonResponse(
                {"error": "You do not have permission to perform this action."},
                status=403,
            )
        return response
    return _wrapped


def _hidden_page_json_or_none(request):
    """JSON equivalent of the inline Faculty_Leave_Page_Permission.is_hidden gate
    used by the CCL Application (self-service) page. Returns a 403 JsonResponse
    or None."""
    if Faculty_Leave_Page_Permission.objects.filter(
        user_id=request.user.id, is_hidden=True
    ).exists():
        return JsonResponse(
            {"error": "This page has been hidden for your role."}, status=403
        )
    return None


def _paginate(request, qs, default_page_size=10, max_page_size=100):
    """Int-parse page/page_size from request.GET, clamp page_size, and return
    (page_obj, paginator, page_size)."""
    try:
        page = int(request.GET.get("page", 1))
    except (TypeError, ValueError):
        page = 1

    try:
        page_size = int(request.GET.get("page_size", default_page_size))
    except (TypeError, ValueError):
        page_size = default_page_size
    page_size = max(1, min(page_size, max_page_size))

    paginator = Paginator(qs, page_size)
    try:
        page_obj = paginator.page(page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages if paginator.num_pages else 1)

    return page_obj, paginator, page_size


def _json_body(request):
    """Parse a JSON request body into a dict; returns {} on empty/invalid body
    (fetch() calls from the new CCL pages send application/json bodies, which
    Django's request.POST does not parse)."""
    if not request.body:
        return {}
    try:
        data = json.loads(request.body.decode("utf-8"))
        return data if isinstance(data, dict) else {}
    except (ValueError, UnicodeDecodeError):
        return {}


def _pagination_meta(page_obj, paginator, page_size):
    return {
        "page": page_obj.number,
        "page_size": page_size,
        "total": paginator.count,
        "total_pages": paginator.num_pages,
        "has_next": page_obj.has_next(),
        "has_prev": page_obj.has_previous(),
    }


def calculate_leave_days(from_date, to_date, from_session_id, to_session_id=None):
    """Session-aware leave-day total.

    ``from_session_id`` is the session on the first day and ``to_session_id``
    the session on the last day, so a range like 06-Jul (AN) -> 07-Jul (AN)
    correctly yields 1.5 days. When ``to_session_id`` is omitted the legacy
    single-session behaviour is preserved.
    """
    if not from_date or not to_date:
        return Decimal("0.0")

    from_session = (
        PermissionTimingMaster.objects.filter(id=from_session_id, is_active=True).first()
        if from_session_id else None
    )
    to_session = (
        PermissionTimingMaster.objects.filter(id=to_session_id, is_active=True).first()
        if to_session_id else None
    )

    days = compute_leave_days(from_date, to_date, from_session, to_session)
    return Decimal(str(days)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
    

# @check_permission("leave_application_form")
# def leave_application_form(request):
#     # print("acamdemic year settings:",settings.ACADEMIC_YEAR)
#     """
#     Create / Update / Delete:
#     - Only Pending leaves are editable
#     - AND only if no approver has already Approved/Rejected (approval process started)
#     """

#     # ---- Get Current User ----
#     ext_user = USER.objects.using("rit_approval_system").filter(Employee_id=request.user.Employee_id, is_active=True).first()
#     faculty = general_information.objects.filter(faculty_id=request.user.Employee_id).select_related("designation").first()

#     if not ext_user:
#         messages.error(request, "User information not found in external system.")
#         return redirect("dashboard")

#     if not faculty:
#         messages.error(request, "Faculty record not found in local database.")
#         return redirect("dashboard")

#     creator_role_id = ext_user.role_id

#     # ----------------------------

#     # ✅ Helper: lock rule
#     def is_leave_locked(leave: LeaveApplication) -> bool:
#         # not pending -> locked
#         if (leave.status or "").lower() != "pending":
#             return True

#         # if any approver already approved/rejected -> locked
#         return LeaveApproversData.objects.filter(
#             leave_application=leave
#         ).exclude(
#             status=LeaveApproversData.Status.PENDING
#         ).exists()

#     # ----------------------------

#     # ✅ DELETE (ONLY IF NOT LOCKED)
#     if request.method == "GET" and request.GET.get("delete_id"):
#         leave = get_object_or_404(LeaveApplication, id=request.GET.get("delete_id"), faculty=faculty)

#         if is_leave_locked(leave):
#             messages.error(request, "This leave cannot be deleted (already processed / not pending).")
#             return redirect("leave_application_form")

#         try:
#             with transaction.atomic():
#                 # refund leave balance (if exists)
#                 days = calculate_leave_days(
#     leave.from_date,
#     leave.to_date,
#     leave.session_id
# )

#                 lb = LeaveBalance.objects.select_for_update().filter(
#                     faculty=leave.faculty,
#                     designation=leave.designation,
#                     leave_type=leave.leave_type,
#                     academic_year=leave.academic_year,
#                 ).first()

#                 if lb:
#                     lb.available = (lb.available or 0) + days
#                     lb.used = max(0, (lb.used or 0) - days)
#                     lb.save(update_fields=["available", "used"])

#                 # remove approver rows
#                 LeaveApproversData.objects.filter(leave_application=leave).delete()
#                 leave.delete()

#             messages.success(request, "Leave application deleted successfully.")
#         except Exception as e:
#             messages.error(request, f"Failed to delete leave: {str(e)}")

#         return redirect("leave_application_form")

#     # ----------------------------

#     # ✅ POST (CREATE / UPDATE)
#     if request.method == "POST":
#         action = request.POST.get("action", "create")  # create | update
#         leave_id = request.POST.get("leave_id")

#         leave_type_id = request.POST.get("leave_type")
#         academic_year = request.POST.get("academic_year")
#         _raw_from = (request.POST.get("from_date") or "").strip()
#         _raw_to = (request.POST.get("to_date") or "").strip()
#         from_date = parse_date(_raw_from) if _raw_from else None
#         to_date = parse_date(_raw_to) if _raw_to else None
#         reason = request.POST.get("reason")
#         session_id = request.POST.get("session")

#         if not all([leave_type_id, academic_year, from_date, to_date, session_id]):
#             messages.error(request, "Please fill in all required fields.")
#             return redirect("leave_application_form")

#         if from_date > to_date:
#             messages.error(request, "From date cannot be after To date.")
#             return redirect("leave_application_form")

#         days = calculate_leave_days(from_date, to_date, session_id)

#         try:
#             with transaction.atomic():

#                 # =============================
#                 # ✅ UPDATE (ONLY IF NOT LOCKED)
#                 # =============================
#                 if action == "update":
#                     leave = get_object_or_404(LeaveApplication, id=leave_id, faculty=faculty)

#                     if is_leave_locked(leave):
#                         messages.error(request, "This leave cannot be edited (already processed / not pending).")
#                         return redirect("leave_application_form")

#                     # Overlap check excluding itself
#                     overlap_qs = LeaveApplication.objects.filter(
#                         faculty=faculty,
#                         status__in=["Pending", "Approved", "Pre-approved"],
#                     ).exclude(id=leave.id).filter(
#                         Q(from_date__lte=to_date) & Q(to_date__gte=from_date)
#                     )
#                     if overlap_qs.exists():
#                         messages.error(request, "You already have a leave overlapping this period.")
#                         return redirect("leave_application_form")

#                     # Refund old balance if the leave type is regular (is_leave=True)
#                     old_days = calculate_leave_days(
#     leave.from_date,
#     leave.to_date,
#     leave.session_id
# )
#                     old_lb = LeaveBalance.objects.select_for_update().filter(
#                         faculty=leave.faculty,
#                         designation=leave.designation,
#                         leave_type=leave.leave_type,
#                         academic_year=leave.academic_year,
#                     ).first()

#                     if old_lb and leave.leave_type.is_leave:
#                         old_lb.available = (old_lb.available or 0) + old_days
#                         old_lb.used = max(0, (old_lb.used or 0) - old_days)
#                         old_lb.save(update_fields=["available", "used"])

#                     # Update leave with the session field
#                     leave.academic_year = academic_year
#                     leave.leave_type_id = leave_type_id
#                     leave.from_date = from_date
#                     leave.to_date = to_date
#                     leave.reason = reason
#                     leave.designation = faculty.designation
#                     leave.status = "Pending"
#                     leave.session_id = session_id  # Save the session field
#                     leave.save()

#                     # Deduct balance if the leave type is regular (is_leave=True)
#                     if leave.leave_type.is_leave:
#                         new_lb = LeaveBalance.objects.select_for_update().filter(
#                             faculty=leave.faculty,
#                             designation=leave.designation,
#                             leave_type=leave.leave_type,
#                             academic_year=leave.academic_year,
#                             start_date__lte=leave.from_date,
#                             end_date__gte=leave.to_date,
#                         ).first()

#                         if not new_lb:
#                             new_lb = LeaveBalance.objects.create(
#                                 faculty=leave.faculty,
#                                 designation=leave.designation,
#                                 leave_type=leave.leave_type,
#                                 academic_year=leave.academic_year,
#                                 available=0,
#                                 used=0,
#                                 start_date=leave.from_date,
#                                 end_date=leave.to_date,
#                             )

#                         new_lb.available = max(0, (new_lb.available or 0) - days)
#                         new_lb.used = (new_lb.used or 0) + days
#                         new_lb.save(update_fields=["available", "used"])

#                     # Rebuild approver rows
#                     LeaveApproversData.objects.filter(leave_application=leave).delete()
#                     _create_approver_chain_for_leave(
#                         leave=leave,
#                         creator_role_id=creator_role_id,
#                         creator_faculty=faculty
#                     )

#                     messages.success(request, "Leave updated successfully.")
#                     return redirect("leave_application_form")

#                 # =============================
#                 # ✅ CREATE
#                 # =============================
#                 overlap_qs = LeaveApplication.objects.filter(
#                     faculty=faculty,
#                     status__in=["Pending", "Approved", "Pre-approved"]
#                 ).filter(Q(from_date__lte=to_date) & Q(to_date__gte=from_date))

#                 if overlap_qs.exists():
#                     messages.error(request, "You already have a leave overlapping this period.")
#                     return redirect("leave_application_form")

#                 new_leave = LeaveApplication.objects.create(
#                     user_id=ext_user.id,
#                     faculty=faculty,
#                     designation=faculty.designation,
#                     academic_year=academic_year,
#                     leave_type_id=leave_type_id,
#                     from_date=from_date,
#                     to_date=to_date,
#                     reason=reason,
#                     status="Pending",
#                     session_id=session_id,  # Save the session when creating
#                 )

#                 # deduct balance if the leave type is regular (is_leave=True)
#                 if new_leave.leave_type.is_leave:
#                     lb = LeaveBalance.objects.select_for_update().filter(
#                         faculty=new_leave.faculty,
#                         designation=new_leave.designation,
#                         leave_type=new_leave.leave_type,
#                         academic_year=new_leave.academic_year,
#                         start_date__lte=new_leave.from_date,
#                         end_date__gte=new_leave.to_date,
#                     ).first()

#                     if not lb:
#                         # Category-based allotment takes precedence over role-based.
#                         allotment = None
#                         if faculty.category_id:
#                             allotment = LeaveAllotment.objects.filter(
#                                 academic_year=academic_year,
#                                 category_id=faculty.category_id,
#                                 leave_type_id=leave_type_id,
#                                 active=True
#                             ).first()
#                         if not allotment:
#                             allotment = LeaveAllotment.objects.filter(
#                                 academic_year=academic_year,
#                                 role=faculty.designation,
#                                 leave_type_id=leave_type_id,
#                                 active=True
#                             ).first()

#                         default_days = allotment.default_allotment if allotment else 0

#                         lb = LeaveBalance.objects.create(
#                             faculty=faculty,
#                             designation=faculty.designation,
#                             leave_type_id=leave_type_id,
#                             academic_year=academic_year,
#                             available=default_days,
#                             used=0,
#                             start_date=allotment.start_date if allotment else from_date,
#                             end_date=allotment.end_date if allotment else to_date,
#                         )

#                     lb.available = max(0, (lb.available or 0) - days)
#                     lb.used = (lb.used or 0) + days
#                     lb.save(update_fields=["available", "used"])

#                 # create approver rows
#                 created_rows = _create_approver_chain_for_leave(
#                     leave=new_leave,
#                     creator_role_id=creator_role_id,
#                     creator_faculty=faculty
#                 )

#                 if created_rows == 0:
#                     new_leave.status = "Pre-approved"
#                     new_leave.save(update_fields=["status"])

#                 messages.success(request, f"Leave applied successfully ({days} day{'s' if days > 1 else ''}).")
#                 return redirect("leave_application_form")

#         except ValidationError as e:
#             messages.error(request, e.messages[0] if e.messages else "Validation error.")
#         except Exception as e:
#             messages.error(request, f"An unexpected error occurred: {str(e)}")

#         return redirect("leave_application_form")

#     # ----------------------------

#     # ✅ FILTERS (GET)
#     fy = request.GET.get("fy", "").strip()
#     lt = request.GET.get("lt", "").strip()
#     status_f = request.GET.get("st", "").strip()
#     _raw_df = (request.GET.get("df") or "").strip()
#     _raw_dt = (request.GET.get("dt") or "").strip()
#     df = parse_date(_raw_df) if _raw_df else None
#     dt = parse_date(_raw_dt) if _raw_dt else None

#     leaves_qs = LeaveApplication.objects.filter(faculty=faculty).select_related("leave_type", "designation").order_by("-requested_date")

#     if fy:
#         leaves_qs = leaves_qs.filter(academic_year=fy)
#     if lt:
#         leaves_qs = leaves_qs.filter(leave_type_id=lt)
#     if status_f:
#         leaves_qs = leaves_qs.filter(status=status_f)

#     # date range filter (overlap range)
#     if df and dt:
#         leaves_qs = leaves_qs.filter(from_date__lte=dt, to_date__gte=df)
#     elif df:
#         leaves_qs = leaves_qs.filter(to_date__gte=df)
#     elif dt:
#         leaves_qs = leaves_qs.filter(from_date__lte=dt)

#     leaves = list(leaves_qs)

#     # ----------------------------

#     # ✅ lock_map + approver status map
#     locked_ids = set(
#         LeaveApproversData.objects.filter(leave_application_id__in=[l.id for l in leaves])
#         .exclude(status=LeaveApproversData.Status.PENDING)
#         .values_list("leave_application_id", flat=True)
#     )

#     lock_map = {l.id: ((l.status or "") != "Pending" or l.id in locked_ids) for l in leaves}

#     # Approver status list (name + designation)
#     approvals_qs = (
#         LeaveApproversData.objects.filter(leave_application_id__in=[l.id for l in leaves])
#         .select_related("approver_id", "approver_id__designation")
#         .order_by("approver_level", "id")
#     )
#     approver_status_map = {}
#     for row in approvals_qs:
#         appr = row.approver_id
#         approver_emp = getattr(appr, "faculty_id", "—")
#         approver_name = getattr(appr, "name", "") or getattr(appr, "username", "")
#         approver_desig = getattr(getattr(appr, "designation", None), "designation_name", None) or str(getattr(appr, "designation", "") or "")

#         approver_status_map.setdefault(row.leave_application_id, []).append({
#             "level": row.approver_level,
#             "emp": approver_emp,
#             "name": approver_name,
#             "desig": approver_desig,
#             "status": row.status,
#         })

#     # ----------------------------

#     # ✅ COUNTS
#     total_count = len(leaves)
#     pending_count = sum(1 for l in leaves if (l.status or "") == "Pending")
#     approved_count = sum(1 for l in leaves if (l.status or "") == "Approved")
#     rejected_count = sum(1 for l in leaves if (l.status or "") == "Rejected")
#     preapproved_count = sum(1 for l in leaves if (l.status or "") == "Pre-approved")
#     locked_count = sum(1 for l in leaves if lock_map.get(l.id))
#     editable_count = total_count - locked_count

#     leave_types = LeaveType.objects.all().order_by("name")
#     academic_years = (
#         LeaveApplication.objects.filter(faculty=faculty)
#         .values_list("academic_year", flat=True)
#         .distinct()
#         .order_by("academic_year")
#     )

#     # ----------------------------

#     # ✅ LEAVE BALANCE SUMMARY
#     from decimal import Decimal

#     leave_balance_summary = []

#     for lt in LeaveType.objects.filter(is_active=True, is_leave=True):

#         # 🔥 Recalculate used dynamically (SESSION-BASED)
#         leaves_records = LeaveApplication.objects.filter(
#             faculty=faculty,
#             leave_type=lt,
#             academic_year=settings.ACADEMIC_YEAR,
#             status__in=["Pending", "Approved", "Pre-approved"]
#         )

#         used = Decimal("0.0")

#         for leave in leaves_records:
#             used += calculate_leave_days(
#                 leave.from_date,
#                 leave.to_date,
#                 leave.session_id
#             )

#         # allotment — category-based allotment takes precedence over role-based.
#         allotment = None
#         if faculty.category_id:
#             allotment = LeaveAllotment.objects.filter(
#                 academic_year=settings.ACADEMIC_YEAR,
#                 category_id=faculty.category_id,
#                 leave_type=lt,
#                 active=True
#             ).first()
#         if not allotment:
#             allotment = LeaveAllotment.objects.filter(
#                 academic_year=settings.ACADEMIC_YEAR,
#                 role=faculty.designation,
#                 leave_type=lt,
#                 active=True
#             ).first()

#         total = Decimal(allotment.default_allotment) if allotment else Decimal("0.0")

#         available = total - used

#         leave_balance_summary.append({
#             "leave_type": lt.name,
#             "total": total,
#             "used": used,
#             "available": available,
#         })
#     return render(
#         request,
#         "faculty_leave_management/leave_application.html",
#         {
#             "leaves": leaves,
#             "leave_types": leave_types,
#             "academic_years": academic_years,

#             # filter selections
#             "fy": fy,
#             "lt": lt,
#             "st": status_f,
#             "df": request.GET.get("df", ""),
#             "dt": request.GET.get("dt", ""),

#             # maps
#             "lock_map": lock_map,
#             "approver_status_map": approver_status_map,
#             "sessions": PermissionTimingMaster.objects.filter(is_active=True),
#             # counts
#             "total_count": total_count,
#             "pending_count": pending_count,
#             "approved_count": approved_count,
#             "rejected_count": rejected_count,
#             "preapproved_count": preapproved_count,
#             "locked_count": locked_count,
#             "editable_count": editable_count,
#             "academic_year": settings.ACADEMIC_YEAR,
#             "leave_balance_summary": leave_balance_summary,
#         },
#     )


def _get_leave_allotment(faculty, academic_year, leave_type_id):
    allotment = None

    if faculty.category_id:
        allotment = LeaveAllotment.objects.filter(
            academic_year=academic_year,
            category_id=faculty.category_id,
            leave_type_id=leave_type_id,
            active=True
        ).first()

    if not allotment:
        allotment = LeaveAllotment.objects.filter(
            academic_year=academic_year,
            role=faculty.designation,
            leave_type_id=leave_type_id,
            active=True
        ).first()

    return allotment


def _is_od_leave_type(leave_type):
    """On Duty / Research On Duty types.

    These require a PDF proof to be uploaded and may be applied even when the
    yearly allotment/balance is zero (they never consume balance).
    """
    if not leave_type:
        return False
    name = (getattr(leave_type, "name", "") or "").strip().lower()
    code = (getattr(leave_type, "code", "") or "").strip().upper()
    return "on duty" in name or code in {"OD", "ROD"}


def _od_awaiting_proof(leave_app):
    """On Duty / Research On Duty may be submitted without proof, but they cannot
    be approved until the PDF proof has been uploaded. True while proof is
    missing on an OD/ROD application."""
    if not leave_app:
        return False
    return _is_od_leave_type(leave_app.leave_type) and not leave_app.proof_file


# Leave types that may ALWAYS be applied — even when the available balance is 0 —
# and therefore never enforce/consume a stored balance (their remaining is never
# shown negative on the form). Everything else (Casual, Vacation, Compensatory
# Casual, Maternity, …) is blocked once its available balance hits 0.
_BALANCE_EXEMPT_CODES = {"LOP", "OD", "ROD", "WL"}


def _is_balance_exempt_leave_type(leave_type):
    """On Duty, Research On Duty, Loss of Pay, Without Leave — appliable at 0."""
    if not leave_type:
        return False
    code = (getattr(leave_type, "code", "") or "").strip().upper()
    return code in _BALANCE_EXEMPT_CODES or _is_od_leave_type(leave_type)


def _leave_consumes_balance(leave_type):
    """True when the leave type must have available balance to be applied and
    consumes it — Casual, Vacation, Compensatory Casual, Maternity, etc.

    Loss of Pay, On Duty, Research On Duty and Without Leave are exempt: they can
    be applied even when the available balance is 0 and never drive it negative.
    """
    if not leave_type:
        return False
    return not _is_balance_exempt_leave_type(leave_type)


def _leave_type_is_restricted(leave_type):
    """A restricted leave type blocks application once the balance is 0 and its
    remaining is a real (possibly deducted) figure. A "Not Restricted" leave type
    can be applied even at 0 and never consumes / goes negative. Configured on the
    Leave Type. A missing leave type defaults to restricted (safe)."""
    if leave_type is None:
        return True
    return (getattr(leave_type, "restriction", "restricted") or "restricted") != "unrestricted"


def _build_balance_summary(faculty, leave_types):
    """Build the Total / Used / Remaining rows shown on the Leave and Permission
    pages, honouring the allotment ``frequency``.

    - ``yearly`` (default): a stored LeaveBalance for the academic year is the
      source of truth; otherwise fall back to the configured allotment and
      compute usage live for the whole year.
    - ``monthly``: the allotment is granted per calendar month, so usage is
      counted for the current month only and the balance resets each month.
    """
    from decimal import Decimal

    today = timezone.now().date()
    summary = []

    for lt_obj in leave_types:
        allotment = _get_leave_allotment(
            faculty=faculty,
            academic_year=settings.ACADEMIC_YEAR,
            leave_type_id=lt_obj.id,
        )
        frequency = (getattr(allotment, "frequency", None) or "yearly")

        if frequency == "monthly":
            total = Decimal(allotment.default_allotment or 0)
            used = Decimal("0.0")
            for leave in LeaveApplication.objects.filter(
                faculty=faculty,
                leave_type=lt_obj,
                status__in=["Pending", "Approved", "Pre-approved"],
                from_date__year=today.year,
                from_date__month=today.month,
            ):
                used += calculate_leave_days(
                    leave.from_date, leave.to_date, leave.session_id, leave.to_session_id
                )
            available = total - used
        else:
            lb = LeaveBalance.objects.filter(
                faculty=faculty,
                leave_type=lt_obj,
                academic_year=settings.ACADEMIC_YEAR,
            ).first()
            if lb:
                used = Decimal(lb.used or 0)
                available = Decimal(lb.available or 0)
                total = available + used
            else:
                used = Decimal("0.0")
                for leave in LeaveApplication.objects.filter(
                    faculty=faculty,
                    leave_type=lt_obj,
                    academic_year=settings.ACADEMIC_YEAR,
                    status__in=["Pending", "Approved", "Pre-approved"],
                ):
                    used += calculate_leave_days(
                        leave.from_date, leave.to_date, leave.session_id, leave.to_session_id
                    )
                total = Decimal(allotment.default_allotment) if allotment else Decimal("0.0")
                available = total - used

        restricted = _leave_type_is_restricted(lt_obj)
        exempt = _is_balance_exempt_leave_type(lt_obj)
        # Leaves that can be applied even at 0 (Loss of Pay / On Duty / Research
        # On Duty / Without Leave, or any type flagged "Not Restricted") must
        # never show a negative remaining. Clamp it to 0 and keep the row
        # consistent (Total = Used + Remaining).
        if (exempt or not restricted) and available < 0:
            available = Decimal("0.0")
            total = used + available

        summary.append({
            "leave_type": lt_obj.name,
            "total": total,
            "used": used,
            "available": available,
            "frequency": frequency,
            "restricted": restricted,
            "exempt": exempt,
        })

    return summary


@check_permission("leave_application_form")
def leave_application_form(request):

    if Faculty_Leave_Page_Permission.objects.filter(
        user_id=request.user.id, is_hidden=True
    ).exists():
        role_obj = Role.objects.using("rit_approval_system").get(id=request.user.role_id)
        role_label = getattr(role_obj, "role", None) or "current"
        return render(
            request,
            "faculty_leave_management/leave_access_restricted.html",
            {"role_label": role_label},
            status=403,
        )

    ext_user = USER.objects.using("rit_approval_system").filter(
        Employee_id=request.user.Employee_id,
        is_active=True
    ).first()

    faculty = general_information.objects.filter(
        faculty_id=request.user.Employee_id
    ).select_related("designation").first()

    if not ext_user:
        messages.error(request, "User information not found in external system.")
        return redirect("dashboard")

    if not faculty:
        messages.error(request, "Faculty record not found in local database.")
        return redirect("dashboard")

    # Use the role of the currently logged-in session user, not ext_user.role_id.
    # ext_user is fetched by Employee_id with .first(), so for an employee that has
    # multiple USER rows (multiple roles) it would pick an arbitrary role. The role
    # the user is actually logged in as is request.user.role.
    creator_role_id = request.user.role_id

    def is_leave_locked(leave: LeaveApplication) -> bool:
        if (leave.status or "").lower() != "pending":
            return True

        return LeaveApproversData.objects.filter(
            leave_application=leave
        ).exclude(
            status=LeaveApproversData.Status.PENDING
        ).exists()

    def _consumes_restricted(leave_type, academic_year=None):
        """A leave both consumes balance AND is balance-restricted: it deducts
        from and is blocked by the balance. "Not Restricted" leave types (and On
        Duty / Research On Duty) return False, so they can be applied at 0 and
        never consume / go negative. Restriction is configured on the Leave Type."""
        if not _leave_consumes_balance(leave_type):
            return False
        return _leave_type_is_restricted(leave_type)

    # DELETE
    if request.method == "GET" and request.GET.get("delete_id"):
        leave = get_object_or_404(
            LeaveApplication,
            id=request.GET.get("delete_id"),
            faculty=faculty
        )

        if is_leave_locked(leave):
            messages.error(request, "This leave cannot be deleted (already processed / not pending).")
            return redirect("leave_application_form")

        try:
            with transaction.atomic():
                days = calculate_leave_days(
                    leave.from_date,
                    leave.to_date,
                    leave.session_id,
                    leave.to_session_id,
                )

                lb = LeaveBalance.objects.select_for_update().filter(
                    faculty=leave.faculty,
                    designation=leave.designation,
                    leave_type=leave.leave_type,
                    academic_year=leave.academic_year,
                ).first()

                if lb and _consumes_restricted(leave.leave_type, leave.academic_year):
                    _before = (lb.available, lb.used)
                    lb.available = (lb.available or 0) + days
                    lb.used = max(0, (lb.used or 0) - days)
                    lb.save(update_fields=["available", "used"])
                    _log_balance_change("leave_application_form:DELETE_refund", lb, before=_before,
                                         extra=f"leave_id={leave.id} days={days}")

                LeaveApproversData.objects.filter(leave_application=leave).delete()
                leave.delete()

            messages.success(request, "Leave application deleted successfully.")

        except Exception as e:
            messages.error(request, f"Failed to delete leave: {str(e)}")

        return redirect("leave_application_form")

    # UPLOAD PROOF (post-submission) — attach the OD/ROD PDF to an existing
    # pending application without re-editing the whole form.
    if request.method == "POST" and request.POST.get("action") == "upload_proof":
        leave = get_object_or_404(
            LeaveApplication, id=request.POST.get("leave_id"), faculty=faculty
        )
        if is_leave_locked(leave):
            messages.error(request, "Proof cannot be uploaded (already processed / not pending).")
            return redirect("leave_application_form")
        if not _is_od_leave_type(leave.leave_type):
            messages.error(request, "Proof upload applies only to On Duty / Research On Duty.")
            return redirect("leave_application_form")
        proof_file = request.FILES.get("proof_file")
        if not proof_file:
            messages.error(request, "Please choose a PDF file to upload.")
            return redirect("leave_application_form")
        if not proof_file.name.lower().endswith(".pdf"):
            messages.error(request, "Proof must be a PDF file.")
            return redirect("leave_application_form")
        leave.proof_file = proof_file
        leave.save(update_fields=["proof_file"])
        messages.success(request, "Proof uploaded successfully.")
        return redirect("leave_application_form")

    # CREATE / UPDATE
    if request.method == "POST":
        action = request.POST.get("action", "create")
        leave_id = request.POST.get("leave_id")

        leave_type_id = request.POST.get("leave_type")
        academic_year = request.POST.get("academic_year")

        _raw_from = (request.POST.get("from_date") or "").strip()
        _raw_to = (request.POST.get("to_date") or "").strip()

        from_date = parse_date(_raw_from) if _raw_from else None
        to_date = parse_date(_raw_to) if _raw_to else None

        reason = request.POST.get("reason")
        session_id = request.POST.get("session")
        to_session_id = request.POST.get("to_session")
        proof_file = request.FILES.get("proof_file")

        if not all([leave_type_id, academic_year, from_date, to_date, session_id, to_session_id]):
            messages.error(request, "Please fill in all required fields.")
            return redirect("leave_application_form")

        if from_date > to_date:
            messages.error(request, "From date cannot be after To date.")
            return redirect("leave_application_form")

        if proof_file and not proof_file.name.lower().endswith(".pdf"):
            messages.error(request, "Proof must be a PDF file.")
            return redirect("leave_application_form")

        days = calculate_leave_days(from_date, to_date, session_id, to_session_id)

        if days <= 0:
            messages.error(
                request,
                "Invalid session selection: the end session falls before the start session."
            )
            return redirect("leave_application_form")

        try:
            with transaction.atomic():

                leave_type = get_object_or_404(LeaveType, id=leave_type_id)

                # On Duty / Research On Duty: the PDF proof is OPTIONAL at
                # submission — the applicant may upload it later (before approval)
                # via the "Upload Proof" action. These can also be applied even
                # when the allotment/balance is zero.
                requires_proof = _is_od_leave_type(leave_type)

                # check allotment for regular leave
                allotment = None
                default_days = 0

                if _consumes_restricted(leave_type, academic_year):
                    # A stored LeaveBalance is the source of truth for what the
                    # employee may apply for (this is what the Leave Balance Summary
                    # displays). Fall back to the configured allotment only when no
                    # balance row exists yet for this academic year.
                    existing_lb = LeaveBalance.objects.filter(
                        faculty=faculty,
                        leave_type_id=leave_type_id,
                        academic_year=academic_year,
                    ).first()

                    allotment = _get_leave_allotment(
                        faculty=faculty,
                        academic_year=academic_year,
                        leave_type_id=leave_type_id
                    )

                    default_days = allotment.default_allotment if allotment else 0

                    if not existing_lb and (not allotment or default_days <= 0):
                        messages.error(
                            request,
                            "Leave cannot be applied because allotment days are 0 or not configured."
                        )
                        return redirect("leave_application_form")

                # UPDATE
                if action == "update":
                    leave = get_object_or_404(
                        LeaveApplication,
                        id=leave_id,
                        faculty=faculty
                    )

                    if is_leave_locked(leave):
                        messages.error(request, "This leave cannot be edited (already processed / not pending).")
                        return redirect("leave_application_form")

                    # Proof stays optional at update time too; it can be added
                    # later via the "Upload Proof" action before approval.

                    overlap_qs = LeaveApplication.objects.filter(
                        faculty=faculty,
                        status__in=["Pending", "Approved", "Pre-approved"],
                    ).exclude(id=leave.id).filter(
                        Q(from_date__lte=to_date) & Q(to_date__gte=from_date)
                    )

                    if overlap_qs.exists():
                        messages.error(request, "You already have a leave overlapping this period.")
                        return redirect("leave_application_form")

                    old_days = calculate_leave_days(
                        leave.from_date,
                        leave.to_date,
                        leave.session_id,
                        leave.to_session_id,
                    )

                    old_lb = LeaveBalance.objects.select_for_update().filter(
                        faculty=leave.faculty,
                        designation=leave.designation,
                        leave_type=leave.leave_type,
                        academic_year=leave.academic_year,
                    ).first()

                    new_lb = None

                    if _consumes_restricted(leave_type, academic_year):
                        new_lb = LeaveBalance.objects.select_for_update().filter(
                            faculty=faculty,
                            designation=faculty.designation,
                            leave_type_id=leave_type_id,
                            academic_year=academic_year,
                            start_date__lte=from_date,
                            end_date__gte=to_date,
                        ).first()

                        if not new_lb:
                            # Fall back to any balance row for this academic year —
                            # the stored balance is authoritative even if its date
                            # window doesn't span the requested period.
                            new_lb = LeaveBalance.objects.select_for_update().filter(
                                faculty=faculty,
                                designation=faculty.designation,
                                leave_type_id=leave_type_id,
                                academic_year=academic_year,
                            ).first()

                        if not new_lb:
                            new_lb = LeaveBalance.objects.create(
                                faculty=faculty,
                                designation=faculty.designation,
                                leave_type_id=leave_type_id,
                                academic_year=academic_year,
                                available=default_days,
                                used=0,
                                start_date=allotment.start_date if allotment else from_date,
                                end_date=allotment.end_date if allotment else to_date,
                            )

                        effective_available = new_lb.available or 0

                        if (
                            old_lb
                            and old_lb.id == new_lb.id
                            and _consumes_restricted(leave.leave_type, leave.academic_year)
                        ):
                            effective_available += old_days

                        if effective_available < days:
                            messages.error(
                                request,
                                f"Insufficient leave balance. Available: {effective_available}, Required: {days}."
                            )
                            return redirect("leave_application_form")

                    # refund old balance
                    if old_lb and _consumes_restricted(leave.leave_type, leave.academic_year):
                        _before = (old_lb.available, old_lb.used)
                        old_lb.available = (old_lb.available or 0) + old_days
                        old_lb.used = max(0, (old_lb.used or 0) - old_days)
                        old_lb.save(update_fields=["available", "used"])
                        _log_balance_change("leave_application_form:UPDATE_refund_old", old_lb, before=_before,
                                             extra=f"leave_id={leave.id} old_days={old_days}")

                    leave.academic_year = academic_year
                    leave.leave_type_id = leave_type_id
                    leave.from_date = from_date
                    leave.to_date = to_date
                    leave.reason = reason
                    leave.designation = faculty.designation
                    leave.status = "Pending"
                    leave.session_id = session_id
                    leave.to_session_id = to_session_id
                    if proof_file:
                        leave.proof_file = proof_file
                    leave.save()

                    # Deduct balance if the leave type is restricted (consumes balance)
                    if _consumes_restricted(leave_type, academic_year):
                        new_lb = LeaveBalance.objects.select_for_update().filter(
                            faculty=leave.faculty,
                            designation=leave.designation,
                            leave_type=leave.leave_type,
                            academic_year=leave.academic_year,
                            start_date__lte=leave.from_date,
                            end_date__gte=leave.to_date,
                        ).first()
                        _lookup_path = "exact"

                        if not new_lb:
                            # Fall back to any balance row for this academic year —
                            # the stored balance is authoritative even if its date
                            # window doesn't span the requested period. Without this
                            # fallback (which every other create/deduct site in this
                            # view already has), a CCL balance row credited with a
                            # narrow earn-date window would never be matched here,
                            # and a duplicate LeaveBalance row would be created and
                            # deducted instead of the row that was just refunded
                            # above — silently doubling the applicant's "used" count.
                            new_lb = LeaveBalance.objects.select_for_update().filter(
                                faculty=leave.faculty,
                                designation=leave.designation,
                                leave_type=leave.leave_type,
                                academic_year=leave.academic_year,
                            ).first()
                            _lookup_path = "fallback_no_date_range"

                        if not new_lb:
                            allotment = _resolve_leave_allotment(
                                faculty, academic_year, leave.leave_type
                            )
                            new_lb = LeaveBalance.objects.create(
                                faculty=leave.faculty,
                                designation=leave.designation,
                                leave_type=leave.leave_type,
                                academic_year=leave.academic_year,
                                available=allotment.default_allotment if allotment else 0,
                                used=0,
                                start_date=allotment.start_date if allotment else leave.from_date,
                                end_date=allotment.end_date if allotment else leave.to_date,
                            )
                            _lookup_path = "created_new_row"

                        # Hard limit: cannot apply for more than the remaining balance
                        if days > (new_lb.available or 0):
                            raise ValidationError(
                                f"Insufficient {leave.leave_type.name} balance: "
                                f"{new_lb.available or 0} day(s) remaining but {days} requested."
                            )

                        _before = (new_lb.available, new_lb.used)
                        new_lb.available = (new_lb.available or 0) - days
                        new_lb.used = (new_lb.used or 0) + days
                        new_lb.save(update_fields=["available", "used"])
                        _log_balance_change("leave_application_form:UPDATE_deduct_new", new_lb, before=_before,
                                             extra=f"leave_id={leave.id} days={days} lookup_path={_lookup_path}")

                    LeaveApproversData.objects.filter(leave_application=leave).delete()
                    created_rows = _create_approver_chain_for_leave(
                        leave=leave,
                        creator_role_id=creator_role_id,
                        creator_faculty=faculty
                    )

                    if created_rows == 0:
                        # No approver hierarchy for this role — abort and roll back
                        # the update (and its balance adjustments).
                        raise ValidationError(
                            "No approver has been assigned for your role. Please contact the administrator."
                        )

                    messages.success(request, "Leave updated successfully.")
                    return redirect("leave_application_form")

                # CREATE
                overlap_qs = LeaveApplication.objects.filter(
                    faculty=faculty,
                    status__in=["Pending", "Approved", "Pre-approved"]
                ).filter(
                    Q(from_date__lte=to_date) & Q(to_date__gte=from_date)
                )

                if overlap_qs.exists():
                    messages.error(request, "You already have a leave overlapping this period.")
                    return redirect("leave_application_form")

                lb = None

                if _consumes_restricted(leave_type, academic_year):
                    lb = LeaveBalance.objects.select_for_update().filter(
                        faculty=faculty,
                        designation=faculty.designation,
                        leave_type_id=leave_type_id,
                        academic_year=academic_year,
                        start_date__lte=from_date,
                        end_date__gte=to_date,
                    ).first()

                    if not lb:
                        # Fall back to any balance row for this academic year —
                        # the stored balance is authoritative even if its date
                        # window doesn't span the requested period.
                        lb = LeaveBalance.objects.select_for_update().filter(
                            faculty=faculty,
                            designation=faculty.designation,
                            leave_type_id=leave_type_id,
                            academic_year=academic_year,
                        ).first()

                    if not lb:
                        lb = LeaveBalance.objects.create(
                            faculty=faculty,
                            designation=faculty.designation,
                            leave_type_id=leave_type_id,
                            academic_year=academic_year,
                            available=default_days,
                            used=0,
                            start_date=allotment.start_date if allotment else from_date,
                            end_date=allotment.end_date if allotment else to_date,
                        )

                    if (lb.available or 0) < days:
                        messages.error(
                            request,
                            f"Insufficient leave balance. Available: {lb.available or 0}, Required: {days}."
                        )
                        return redirect("leave_application_form")

                new_leave = LeaveApplication.objects.create(
                    user_id=ext_user.id,
                    faculty=faculty,
                    designation=faculty.designation,
                    academic_year=academic_year,
                    leave_type_id=leave_type_id,
                    from_date=from_date,
                    to_date=to_date,
                    reason=reason,
                    status="Pending",
                    session_id=session_id,
                    to_session_id=to_session_id,
                    proof_file=proof_file,
                )

                # deduct balance only if the leave type is restricted (consumes balance)
                if _consumes_restricted(new_leave.leave_type, new_leave.academic_year):
                    lb = LeaveBalance.objects.select_for_update().filter(
                        faculty=new_leave.faculty,
                        designation=new_leave.designation,
                        leave_type=new_leave.leave_type,
                        academic_year=new_leave.academic_year,
                        start_date__lte=new_leave.from_date,
                        end_date__gte=new_leave.to_date,
                    ).first()

                    if not lb:
                        # Fall back to any balance row for this academic year before
                        # creating a new one, so we deduct from the stored balance
                        # instead of duplicating it.
                        lb = LeaveBalance.objects.select_for_update().filter(
                            faculty=new_leave.faculty,
                            designation=new_leave.designation,
                            leave_type=new_leave.leave_type,
                            academic_year=new_leave.academic_year,
                        ).first()

                    if not lb:
                        # Category-based allotment takes precedence over role-based.
                        allotment = _resolve_leave_allotment(
                            faculty, academic_year, new_leave.leave_type
                        )

                        default_days = allotment.default_allotment if allotment else 0

                        lb = LeaveBalance.objects.create(
                            faculty=faculty,
                            designation=faculty.designation,
                            leave_type_id=leave_type_id,
                            academic_year=academic_year,
                            available=default_days,
                            used=0,
                            start_date=allotment.start_date if allotment else from_date,
                            end_date=allotment.end_date if allotment else to_date,
                        )

                    # Hard limit: cannot apply for more than the remaining balance
                    if days > (lb.available or 0):
                        raise ValidationError(
                            f"Insufficient {new_leave.leave_type.name} balance: "
                            f"{lb.available or 0} day(s) remaining but {days} requested."
                        )

                    _before = (lb.available, lb.used)
                    lb.available = (lb.available or 0) - days
                    lb.used = (lb.used or 0) + days
                    lb.save(update_fields=["available", "used"])
                    _log_balance_change("leave_application_form:CREATE_deduct", lb, before=_before,
                                         extra=f"leave_id={new_leave.id} days={days}")

                created_rows = _create_approver_chain_for_leave(
                    leave=new_leave,
                    creator_role_id=creator_role_id,
                    creator_faculty=faculty
                )

                if created_rows == 0:
                    # No approver hierarchy could be resolved for this applicant's
                    # role — abort so the leave is not silently auto-approved.
                    # Raising inside transaction.atomic() rolls back the leave and
                    # the balance deduction made above.
                    raise ValidationError(
                        "No approver has been assigned for your role. Please contact the administrator."
                    )

                messages.success(
                    request,
                    f"Leave applied successfully ({days} day{'s' if days > 1 else ''})."
                )
                return redirect("leave_application_form")

        except ValidationError as e:
            messages.error(request, e.messages[0] if e.messages else "Validation error.")

        except Exception as e:
            messages.error(request, f"An unexpected error occurred: {str(e)}")

        return redirect("leave_application_form")

    # FILTERS
    fy = request.GET.get("fy", "").strip()
    lt = request.GET.get("lt", "").strip()
    status_f = request.GET.get("st", "").strip()

    _raw_df = (request.GET.get("df") or "").strip()
    _raw_dt = (request.GET.get("dt") or "").strip()

    df = parse_date(_raw_df) if _raw_df else None
    dt = parse_date(_raw_dt) if _raw_dt else None

    # Permission-type applications (is_leave=False) are managed on the dedicated
    # permission_form page — this page handles regular leaves only.
    leaves_qs = LeaveApplication.objects.filter(
        faculty=faculty,
        leave_type__is_leave=True,
    ).select_related(
        "leave_type",
        "designation"
    ).order_by("-requested_date")

    if fy:
        leaves_qs = leaves_qs.filter(academic_year=fy)

    if lt:
        leaves_qs = leaves_qs.filter(leave_type_id=lt)

    if status_f:
        leaves_qs = leaves_qs.filter(status=status_f)

    if df and dt:
        leaves_qs = leaves_qs.filter(from_date__lte=dt, to_date__gte=df)
    elif df:
        leaves_qs = leaves_qs.filter(to_date__gte=df)
    elif dt:
        leaves_qs = leaves_qs.filter(from_date__lte=dt)

    leaves = list(leaves_qs)

    locked_ids = set(
        LeaveApproversData.objects.filter(
            leave_application_id__in=[l.id for l in leaves]
        ).exclude(
            status=LeaveApproversData.Status.PENDING
        ).values_list("leave_application_id", flat=True)
    )

    lock_map = {
        l.id: ((l.status or "") != "Pending" or l.id in locked_ids)
        for l in leaves
    }

    approvals_qs = (
        LeaveApproversData.objects.filter(
            leave_application_id__in=[l.id for l in leaves]
        )
        .select_related("approver_id", "approver_id__designation")
        .order_by("approver_level", "id")
    )

    approver_status_map = {}

    for row in approvals_qs:
        appr = row.approver_id
        approver_emp = getattr(appr, "faculty_id", "—")
        approver_name = getattr(appr, "name", "") or getattr(appr, "username", "")
        approver_desig = (
            getattr(getattr(appr, "designation", None), "designation_name", None)
            or str(getattr(appr, "designation", "") or "")
        )

        approver_status_map.setdefault(row.leave_application_id, []).append({
            "level": row.approver_level,
            "emp": approver_emp,
            "name": approver_name,
            "desig": approver_desig,
            "status": row.status,
        })

    total_count = len(leaves)
    pending_count = sum(1 for l in leaves if (l.status or "") == "Pending")
    approved_count = sum(1 for l in leaves if (l.status or "") == "Approved")
    rejected_count = sum(1 for l in leaves if (l.status or "") == "Rejected")
    preapproved_count = sum(1 for l in leaves if (l.status or "") == "Pre-approved")
    locked_count = sum(1 for l in leaves if lock_map.get(l.id))
    editable_count = total_count - locked_count

    # Only regular leave types here; permission types have their own page.
    leave_types = LeaveType.objects.filter(is_active=True, is_leave=True).order_by("name")

    academic_years = (
        LeaveApplication.objects.filter(faculty=faculty)
        .values_list("academic_year", flat=True)
        .distinct()
        .order_by("academic_year")
    )

    leave_balance_summary = _build_balance_summary(
        faculty, LeaveType.objects.filter(is_active=True, is_leave=True)
    )

    return render(
        request,
        "faculty_leave_management/leave_application.html",
        {
            "leaves": leaves,
            "leave_types": leave_types,
            "academic_years": academic_years,
            "fy": fy,
            "lt": lt,
            "st": status_f,
            "df": request.GET.get("df", ""),
            "dt": request.GET.get("dt", ""),
            "lock_map": lock_map,
            "approver_status_map": approver_status_map,
            "sessions": PermissionTimingMaster.objects.filter(is_active=True),
            "total_count": total_count,
            "pending_count": pending_count,
            "approved_count": approved_count,
            "rejected_count": rejected_count,
            "preapproved_count": preapproved_count,
            "locked_count": locked_count,
            "editable_count": editable_count,
            "academic_year": settings.ACADEMIC_YEAR,
            "leave_balance_summary": leave_balance_summary,
        },
    )



def _create_approver_chain_for_leave(leave, creator_role_id, creator_faculty):
    """
    Build the approver chain for a LeaveApplication.

    Logic:
    - Reads LeaveApprovers master table ordered by level.
    - For HOD-type approvers (same-dept, cross_dept=NO, no dept override):
        uses the creator's own Department_id from rit_approval_system to find
        the correct HOD — not just .first() across all HODs.
    - For cross-dept approvers: uses the configured approver_department.
    - Principal / Director (single-user roles): matched by role_id only.
    """
    approvers_qs = LeaveApprovers.objects.filter(
        creator_role_id=creator_role_id
    ).order_by("approver_level")

    created_rows = 0
    created_levels = set()

    # Get creator's Department_id from rit_approval_system (for HOD matching)
    creator_ext_user = USER.objects.using("rit_approval_system").filter(
        Employee_id=creator_faculty.faculty_id, is_active=True
    ).first()
    creator_dept_id = getattr(creator_ext_user, "Department_id", None) if creator_ext_user else None

    for approver in approvers_qs:
        level = approver.approver_level
        role_id = approver.approver_role_id
        local_dept_id = approver.approver_department_id
        is_cross = (approver.is_cross_department_approver or "NO").upper() == "YES"

        if level in created_levels:
            continue

        approver_filter = {"role_id": role_id, "is_active": True}

        if is_cross and local_dept_id:
            # Cross-dept: use the configured department
            local_dept = Add_Department.objects.filter(id=local_dept_id).first()
            dept_code = getattr(local_dept, "Department_code", None)
            if dept_code:
                ext_dept_id = (
                    Department.objects.using("rit_approval_system")
                    .filter(Department_code=dept_code)
                    .values_list("id", flat=True)
                    .first()
                )
                if ext_dept_id:
                    approver_filter["Department_id"] = ext_dept_id
        elif not is_cross and not local_dept_id and creator_dept_id:
            # Same-dept approver (e.g. HOD): match by creator's department
            approver_filter["Department_id"] = creator_dept_id

        approver_user = USER.objects.using("rit_approval_system").filter(
            **approver_filter
        ).first()
        if not approver_user:
            continue

        approver_faculty = general_information.objects.filter(
            faculty_id=approver_user.Employee_id
        ).select_related("designation").first()
        if not approver_faculty:
            continue

        LeaveApproversData.objects.create(
            leave_application=leave,
            approver_id=approver_faculty,
            creator_id=creator_faculty,
            approver_level=level,
            approver_role_id=role_id,
            creator_role_id=creator_role_id,
            status=LeaveApproversData.Status.PENDING,
            reason=f"Approver: {approver_user.Employee_id}",
            approved_date=timezone.now(),
        )

        created_rows += 1
        created_levels.add(level)

    return created_rows



@check_permission("permission_form")
def permission_form(request):
    """Apply / manage *permission*-type applications — leave types with
    ``is_leave=False`` (e.g. Loss of Pay, Without Leave, Maternity Leave).

    Mirrors ``leave_application_form`` but:
    - the leave-type dropdown lists only permission types (is_leave=False),
    - the applicant supplies a date range PLUS a from-time / to-time window,
    - there is no leave-balance consumption,
    - it routes through the same multi-level approver chain as leaves.
    """
    if Faculty_Leave_Page_Permission.objects.filter(
        user_id=request.user.id, is_hidden=True
    ).exists():
        role_obj = Role.objects.using("rit_approval_system").get(id=request.user.role_id)
        role_label = getattr(role_obj, "role", None) or "current"
        return render(
            request,
            "faculty_leave_management/leave_access_restricted.html",
            {"role_label": role_label},
            status=403,
        )

    ext_user = USER.objects.using("rit_approval_system").filter(
        Employee_id=request.user.Employee_id,
        is_active=True
    ).first()

    faculty = general_information.objects.filter(
        faculty_id=request.user.Employee_id
    ).select_related("designation").first()

    if not ext_user:
        messages.error(request, "User information not found in external system.")
        return redirect("faculty_dashboard")

    if not faculty:
        messages.error(request, "Faculty record not found in local database.")
        return redirect("faculty_dashboard")

    creator_role_id = request.user.role_id

    def is_locked(leave: LeaveApplication) -> bool:
        if (leave.status or "").lower() != "pending":
            return True
        return LeaveApproversData.objects.filter(
            leave_application=leave
        ).exclude(
            status=LeaveApproversData.Status.PENDING
        ).exists()

    permission_types = LeaveType.objects.filter(
        is_active=True, is_leave=False
    ).order_by("name")

    # DELETE
    if request.method == "GET" and request.GET.get("delete_id"):
        leave = get_object_or_404(
            LeaveApplication,
            id=request.GET.get("delete_id"),
            faculty=faculty,
            leave_type__is_leave=False,
        )
        if is_locked(leave):
            messages.error(request, "This permission cannot be deleted (already processed / not pending).")
            return redirect("permission_form")
        try:
            with transaction.atomic():
                LeaveApproversData.objects.filter(leave_application=leave).delete()
                leave.delete()
            messages.success(request, "Permission deleted successfully.")
        except Exception as e:
            messages.error(request, f"Failed to delete permission: {str(e)}")
        return redirect("permission_form")

    # CREATE / UPDATE
    if request.method == "POST":
        action = request.POST.get("action", "create")
        leave_id = request.POST.get("leave_id")
        leave_type_id = request.POST.get("leave_type")
        academic_year = request.POST.get("academic_year")
        reason = request.POST.get("reason")

        # Permission is for a single date plus a from-time / to-time window.
        _raw_date = (request.POST.get("date") or "").strip()
        perm_date = parse_date(_raw_date) if _raw_date else None
        from_date = to_date = perm_date

        from_time = parse_time((request.POST.get("from_time") or "").strip())
        to_time = parse_time((request.POST.get("to_time") or "").strip())

        if not all([leave_type_id, academic_year, perm_date, from_time, to_time]):
            messages.error(request, "Please fill in all required fields.")
            return redirect("permission_form")

        if from_time >= to_time:
            messages.error(request, "From time must be earlier than To time.")
            return redirect("permission_form")

        try:
            with transaction.atomic():
                # Only permission (is_leave=False) types may be applied here.
                leave_type = get_object_or_404(
                    LeaveType, id=leave_type_id, is_leave=False
                )

                if action == "update":
                    leave = get_object_or_404(
                        LeaveApplication,
                        id=leave_id,
                        faculty=faculty,
                        leave_type__is_leave=False,
                    )
                    if is_locked(leave):
                        messages.error(request, "This permission cannot be edited (already processed / not pending).")
                        return redirect("permission_form")

                    leave.academic_year = academic_year
                    leave.leave_type = leave_type
                    leave.from_date = from_date
                    leave.to_date = to_date
                    leave.from_time = from_time
                    leave.to_time = to_time
                    leave.reason = reason
                    leave.designation = faculty.designation
                    leave.status = "Pending"
                    leave.session = None
                    leave.to_session = None
                    leave.save()

                    LeaveApproversData.objects.filter(leave_application=leave).delete()
                    created_rows = _create_approver_chain_for_leave(
                        leave=leave,
                        creator_role_id=creator_role_id,
                        creator_faculty=faculty,
                    )
                    if created_rows == 0:
                        raise ValidationError(
                            "No approver has been assigned for your role. Please contact the administrator."
                        )

                    messages.success(request, "Permission updated successfully.")
                    return redirect("permission_form")

                # CREATE
                new_leave = LeaveApplication.objects.create(
                    user_id=ext_user.id,
                    faculty=faculty,
                    designation=faculty.designation,
                    academic_year=academic_year,
                    leave_type=leave_type,
                    from_date=from_date,
                    to_date=to_date,
                    from_time=from_time,
                    to_time=to_time,
                    reason=reason,
                    status="Pending",
                )

                created_rows = _create_approver_chain_for_leave(
                    leave=new_leave,
                    creator_role_id=creator_role_id,
                    creator_faculty=faculty,
                )
                if created_rows == 0:
                    raise ValidationError(
                        "No approver has been assigned for your role. Please contact the administrator."
                    )

                messages.success(request, "Permission applied successfully.")
                return redirect("permission_form")

        except ValidationError as e:
            messages.error(request, e.messages[0] if e.messages else "Validation error.")
        except Exception as e:
            messages.error(request, f"An unexpected error occurred: {str(e)}")
        return redirect("permission_form")

    # FILTERS + LIST
    fy = request.GET.get("fy", "").strip()
    lt = request.GET.get("lt", "").strip()
    status_f = request.GET.get("st", "").strip()
    _raw_df = (request.GET.get("df") or "").strip()
    _raw_dt = (request.GET.get("dt") or "").strip()
    df = parse_date(_raw_df) if _raw_df else None
    dt = parse_date(_raw_dt) if _raw_dt else None

    leaves_qs = LeaveApplication.objects.filter(
        faculty=faculty,
        leave_type__is_leave=False,
    ).select_related("leave_type", "designation").order_by("-requested_date")

    if fy:
        leaves_qs = leaves_qs.filter(academic_year=fy)
    if lt:
        leaves_qs = leaves_qs.filter(leave_type_id=lt)
    if status_f:
        leaves_qs = leaves_qs.filter(status=status_f)
    if df and dt:
        leaves_qs = leaves_qs.filter(from_date__lte=dt, to_date__gte=df)
    elif df:
        leaves_qs = leaves_qs.filter(to_date__gte=df)
    elif dt:
        leaves_qs = leaves_qs.filter(from_date__lte=dt)

    leaves = list(leaves_qs)

    locked_ids = set(
        LeaveApproversData.objects.filter(
            leave_application_id__in=[l.id for l in leaves]
        ).exclude(
            status=LeaveApproversData.Status.PENDING
        ).values_list("leave_application_id", flat=True)
    )
    lock_map = {
        l.id: ((l.status or "") != "Pending" or l.id in locked_ids)
        for l in leaves
    }

    approvals_qs = (
        LeaveApproversData.objects.filter(
            leave_application_id__in=[l.id for l in leaves]
        )
        .select_related("approver_id", "approver_id__designation")
        .order_by("approver_level", "id")
    )

    approver_status_map = {}
    for row in approvals_qs:
        appr = row.approver_id
        approver_emp = getattr(appr, "faculty_id", "—")
        approver_name = getattr(appr, "name", "") or getattr(appr, "username", "")
        approver_desig = (
            getattr(getattr(appr, "designation", None), "designation_name", None)
            or str(getattr(appr, "designation", "") or "")
        )
        approver_status_map.setdefault(row.leave_application_id, []).append({
            "level": row.approver_level,
            "emp": approver_emp,
            "name": approver_name,
            "desig": approver_desig,
            "status": row.status,
        })

    total_count = len(leaves)
    pending_count = sum(1 for l in leaves if (l.status or "") == "Pending")
    approved_count = sum(1 for l in leaves if (l.status or "") == "Approved")
    rejected_count = sum(1 for l in leaves if (l.status or "") == "Rejected")

    academic_years = (
        LeaveApplication.objects.filter(faculty=faculty, leave_type__is_leave=False)
        .values_list("academic_year", flat=True)
        .distinct()
        .order_by("academic_year")
    )

    # Balance summary (Total / Used / Remaining) for each permission type,
    # honouring the allotment frequency (yearly vs monthly).
    leave_balance_summary = _build_balance_summary(faculty, permission_types)

    return render(
        request,
        "faculty_leave_management/permission_form.html",
        {
            "leaves": leaves,
            "leave_types": permission_types,
            "academic_years": academic_years,
            "fy": fy,
            "lt": lt,
            "st": status_f,
            "df": request.GET.get("df", ""),
            "dt": request.GET.get("dt", ""),
            "lock_map": lock_map,
            "approver_status_map": approver_status_map,
            "total_count": total_count,
            "pending_count": pending_count,
            "approved_count": approved_count,
            "rejected_count": rejected_count,
            "leave_balance_summary": leave_balance_summary,
            "academic_year": settings.ACADEMIC_YEAR,
        },
    )


# ============================================================
# VIEW — leave approvals page
# ============================================================
# @no_cache
# @check_permission("leave_approvals")
# @faculty_leave_management
# def leave_approvals(request):
#     """
#     Displays hierarchical leave applications pending approval for the current approver.
#     - Each creator_role_id defines a separate hierarchy.
#     - Approvers only see leaves from hierarchies they are part of.
#     - Next level approver sees a leave only after all lower levels are approved.
#     """

#     approver_user = USER.objects.using("rit_approval_system").filter(Employee_id=request.user.Employee_id, is_active=True).first()
#     if not approver_user:
#         messages.error(request, "Approver not found.")
#         return redirect("dashboard")

#     approver_faculty = general_information.objects.filter(faculty_id=request.user.Employee_id).first()
#     if not approver_faculty:
#         messages.error(request, "Approver faculty record not found.")
#         return redirect("dashboard")

#     # print("\n===== LEAVE APPROVAL DEBUG START =====")
#     # print(f"Approver Employee ID: {approver_user.Employee_id}")
#     # print(f"Approver DB ID: {approver_user.id}")
#     # print("======================================\n")

#     # ---- All possible hierarchies ----
#     all_hierarchies = list(
#         LeaveApprovers.objects.all()
#         .values("id", "approver_level", "approver_role_id", "creator_role_id")
#         .order_by("creator_role_id", "approver_level")
#     )

#     # ---- Determine which hierarchies current approver participates in ----
#     related_creator_roles = {
#         h["creator_role_id"] for h in all_hierarchies if h["approver_role_id"] == approver_user.role_id
#     }

#     if not related_creator_roles:
#         # print("⚠️ No hierarchy found for this approver role.")
#         return render(
#             request,
#             "faculty_leave_management/leave_approval.html",
#             {"grouped_approvals": {}, "user_map": {}, "role_map": {}},
#         )

#     # print(f"✅ Approver participates in creator_role_ids: {related_creator_roles}")

#     grouped_approvals = {}
#     eligible_apps = []

#     # ---- For each hierarchy (creator_role_id) ----
#     for creator_role_id in related_creator_roles:
#         print()

#         hierarchy_levels = [
#             h for h in all_hierarchies if h["creator_role_id"] == creator_role_id
#         ]
#         for h in hierarchy_levels:
#             print()

#         # Find current approver’s level in this hierarchy
#         current_level = next(
#             (h["approver_level"] for h in hierarchy_levels if h["approver_role_id"] == approver_user.role_id),
#             None,
#         )
#         if current_level is None:
#             continue

#         # print(f"   Current approver level for this hierarchy: {current_level}")

#         # ---- Find pending applications for this approver ----
#         pending_qs = LeaveApproversData.objects.filter(
#     approver_id=approver_faculty,
#     approver_level=current_level,
#     status=LeaveApproversData.Status.PENDING,
# ).select_related("leave_application", "creator_id")

#         # print(f"   Found pending raw: {pending_qs.count()}")

#         for lad in pending_qs:
#             lower_pending = LeaveApproversData.objects.filter(
#                 leave_application=lad.leave_application,
#                 approver_level__lt=current_level,
#             ).exclude(status=LeaveApproversData.Status.APPROVED).exists()

#             if not lower_pending:
#                 eligible_apps.append(lad)

#     # print(f"\n✅ Total eligible applications across hierarchies: {len(eligible_apps)}")

#     creator_ids = [lad.creator_id_id for lad in eligible_apps]
#     external_users = general_information.objects.filter(id__in=creator_ids)
#     user_map = {u.id: u for u in external_users}

#     # ---- Resolve role info ----
#     creator_emp_ids = [u.faculty_id for u in external_users]  # faculty_id = Employee_id
#     ext_users = USER.objects.using("rit_approval_system").filter(Employee_id__in=creator_emp_ids, is_active=True)

#     ext_user_map = {u.Employee_id: u for u in ext_users}
#     role_ids = [u.role_id for u in ext_users if u.role_id]
#     external_roles = Role.objects.using("rit_approval_system").filter(id__in=role_ids)
#     role_map = {r.id: r for r in external_roles}

#     # ---- Group by Approver Level for Display ----
#     for lad in eligible_apps:
#         grouped_approvals.setdefault(lad.approver_level, []).append(lad)

#     # print("Grouped approvals by level:")
#     for level, items in grouped_approvals.items():
#         print()

#     # ---- POST ACTIONS ----
#     if request.method == "POST":
#         action = request.POST.get("action")
#         selected_ids = request.POST.getlist("selected_ids")

#         if not selected_ids:
#             messages.warning(request, "No leave applications selected.")
#             return redirect(reverse("leave_approvals"))

#         selected_leaves = LeaveApproversData.objects.filter(id__in=selected_ids).select_related("leave_application")

#         if action == "approve":
#             with transaction.atomic():
#                 for lad in selected_leaves:
#                     lad.status = LeaveApproversData.Status.APPROVED
#                     lad.approved_date = timezone.now()
#                     lad.save(update_fields=["status", "approved_date"])

#                     app = lad.leave_application
#                     all_approved = not LeaveApproversData.objects.filter(
#                         leave_application=app
#                     ).exclude(status=LeaveApproversData.Status.APPROVED).exists()

#                     if all_approved and app.status != "Approved":
#                         app.status = "Approved"
#                         app.save(update_fields=["status"])

#             messages.success(request, "Selected leave applications approved successfully.")
#             return redirect(reverse("leave_approvals"))

#         elif action == "reject":
#             reason = request.POST.get("rejection_reason", "").strip()
#             if not reason:
#                 messages.error(request, "Rejection reason is required.")
#                 return redirect(reverse("leave_approvals"))

#             with transaction.atomic():
#                 for lad in selected_leaves:
#                     app = lad.leave_application
#                     days = (app.to_date - app.from_date).days + 1

#                     # ✅ Refund using faculty + leave_type + date window
#                     lb = LeaveBalance.objects.select_for_update().filter(
#                         faculty=app.faculty,
#                         designation=app.designation,  # ✅ DesignationMaster FK
#                         leave_type=app.leave_type,
#                         start_date__lte=app.from_date,
#                         end_date__gte=app.to_date,
#                     ).first()

#                     if lb:
#                         lb.available = (lb.available or 0) + days
#                         lb.used = max(0, (lb.used or 0) - days)
#                         lb.save(update_fields=["available", "used"])

#                     # ✅ Update application status
#                     app.status = "Rejected"
#                     app.save(update_fields=["status"])

#                     # ✅ Update approver row
#                     lad.status = LeaveApproversData.Status.REJECTED
#                     lad.reason = reason
#                     lad.approved_date = timezone.now()
#                     lad.save(update_fields=["status", "reason", "approved_date"])

#             messages.success(request, "Selected applications rejected and balances refunded.")
#             return redirect(reverse("leave_approvals"))





#         else:
#             messages.error(request, "Invalid action.")
#             return redirect(reverse("leave_approvals"))

#     # print("===== LEAVE APPROVAL DEBUG END =====\n")

#     return render(
#         request,
#         "faculty_leave_management/leave_approval.html",
#         {
#         "grouped_approvals": grouped_approvals,
#         "user_map": user_map,         # general_information by id
#         "ext_user_map": ext_user_map, # ✅ external USER by Employee_id
#         "role_map": role_map,         # external Role by role_id
#     },
#     )



from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.urls import reverse
from django.core.paginator import Paginator
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# -------------------------------------------------------------------
# LEAVE APPROVAL VIEW
# -------------------------------------------------------------------
# @no_cache
# @check_permission("leave_approvals")
# @faculty_leave_management
# def leave_approvals(request):
#     """
#     Displays hierarchical leave applications for the current approver.

#     Tabs:
#     - Pending: only currently actionable applications
#     - Approved: applications already approved by current approver
#     - Rejected: applications rejected by current approver

#     Supports:
#     - search
#     - approver level filter
#     - academic year filter
#     - pagination
#     - export (excel / pdf)
#     """

#     approver_user = USER.objects.using("rit_approval_system").filter(
#         Employee_id=request.user.Employee_id,
#         is_active=True
#     ).first()
#     if not approver_user:
#         messages.error(request, "Approver not found.")
#         return redirect("dashboard")

#     approver_faculty = general_information.objects.filter(
#         faculty_id=request.user.Employee_id
#     ).first()
#     if not approver_faculty:
#         messages.error(request, "Approver faculty record not found.")
#         return redirect("dashboard")

#     # -------------------------------
#     # Read search / filter params
#     # -------------------------------
#     q = request.GET.get("q", "").strip()
#     level_filter = request.GET.get("level", "").strip()
#     academic_year_filter = request.GET.get("academic_year", "").strip()
#     tab = request.GET.get("tab", "pending").strip().lower()  # pending / approved / rejected
#     page_number = request.GET.get("page", 1)
#     per_page = 50

#     valid_tabs = ["pending", "approved", "rejected"]
#     if tab not in valid_tabs:
#         tab = "pending"

#     # ---- All possible hierarchies ----
#     all_hierarchies = list(
#         LeaveApprovers.objects.all()
#         .values("id", "approver_level", "approver_role_id", "creator_role_id")
#         .order_by("creator_role_id", "approver_level")
#     )

#     # ---- Determine which hierarchies current approver participates in ----
#     related_creator_roles = {
#         h["creator_role_id"]
#         for h in all_hierarchies
#         if h["approver_role_id"] == approver_user.role_id
#     }

#     if not related_creator_roles:
#         return render(
#             request,
#             "faculty_leave_management/leave_approval.html",
#             {
#                 "page_obj": None,
#                 "grouped_approvals": {},
#                 "user_map": {},
#                 "ext_user_map": {},
#                 "role_map": {},
#                 "q": q,
#                 "level_filter": level_filter,
#                 "academic_year_filter": academic_year_filter,
#                 "available_levels": [],
#                 "available_academic_years": [],
#                 "tab": tab,
#                 "pending_count": 0,
#                 "approved_count": 0,
#                 "rejected_count": 0,
#             },
#         )

#     # --------------------------------------------------
#     # Build eligible pending ids (only actionable rows)
#     # --------------------------------------------------
#     eligible_pending_ids = []

#     for creator_role_id in related_creator_roles:
#         hierarchy_levels = [
#             h for h in all_hierarchies
#             if h["creator_role_id"] == creator_role_id
#         ]

#         current_level = next(
#             (
#                 h["approver_level"]
#                 for h in hierarchy_levels
#                 if h["approver_role_id"] == approver_user.role_id
#             ),
#             None,
#         )

#         if current_level is None:
#             continue

#         pending_qs = (
#             LeaveApproversData.objects.filter(
#                 approver_id=approver_faculty,
#                 approver_level=current_level,
#                 status=LeaveApproversData.Status.PENDING,
#             )
#             .select_related("leave_application", "creator_id")
#         )

#         for lad in pending_qs:
#             lower_pending = LeaveApproversData.objects.filter(
#                 leave_application=lad.leave_application,
#                 approver_level__lt=current_level,
#             ).exclude(status=LeaveApproversData.Status.APPROVED).exists()

#             if not lower_pending:
#                 eligible_pending_ids.append(lad.id)

#     # --------------------------------------------------
#     # Base querysets for each tab
#     # --------------------------------------------------
#     pending_base_qs = LeaveApproversData.objects.filter(
#         id__in=eligible_pending_ids
#     )

#     approved_base_qs = LeaveApproversData.objects.filter(
#         approver_id=approver_faculty,
#         status=LeaveApproversData.Status.APPROVED,
#     )

#     rejected_base_qs = LeaveApproversData.objects.filter(
#         approver_id=approver_faculty,
#         status=LeaveApproversData.Status.REJECTED,
#     )

#     # --------------------------------------------------
#     # Tab counts (before filters)
#     # --------------------------------------------------
#     pending_count = pending_base_qs.count()
#     approved_count = approved_base_qs.count()
#     rejected_count = rejected_base_qs.count()

#     # --------------------------------------------------
#     # Choose queryset based on active tab
#     # --------------------------------------------------
#     if tab == "approved":
#         approvals_qs = approved_base_qs
#     elif tab == "rejected":
#         approvals_qs = rejected_base_qs
#     else:
#         approvals_qs = pending_base_qs

#     approvals_qs = approvals_qs.select_related(
#     "leave_application",
#     "leave_application__faculty",
#     "leave_application__designation",
#     "leave_application__leave_type",
#     "approver_id",
#     "creator_id",
# ).order_by("approver_level", "-leave_application__requested_date", "-id")

#     # -------------------------------
#     # Search
#     # -------------------------------
#     if q:
#         approvals_qs = approvals_qs.filter(
#             Q(leave_application__from_date__icontains=q) |
#             Q(leave_application__to_date__icontains=q) |
#             Q(leave_application__academic_year__icontains=q) |
#             Q(leave_application__reason__icontains=q) |
#             Q(leave_application__status__icontains=q) |
#             Q(leave_application__faculty__faculty_id__icontains=q) |
#             Q(leave_application__faculty__name__icontains=q) |
#             Q(leave_application__leave_type__leave_type__icontains=q)
#         )

#     # -------------------------------
#     # Level filter
#     # -------------------------------
#     if level_filter:
#         try:
#             approvals_qs = approvals_qs.filter(approver_level=int(level_filter))
#         except ValueError:
#             pass

#     # -------------------------------
#     # Academic year filter
#     # -------------------------------
#     if academic_year_filter:
#         approvals_qs = approvals_qs.filter(
#             leave_application__academic_year=academic_year_filter
#         )

#     # -------------------------------
#     # Dropdown data
#     # -------------------------------
#     all_my_rows = LeaveApproversData.objects.filter(
#         approver_id=approver_faculty
#     )

#     available_levels = (
#         all_my_rows.values_list("approver_level", flat=True)
#         .distinct()
#         .order_by("approver_level")
#     )

#     available_academic_years = (
#         all_my_rows.values_list("leave_application__academic_year", flat=True)
#         .distinct()
#         .order_by("leave_application__academic_year")
#     )

#     # -------------------------------
#     # Pagination
#     # -------------------------------
#     paginator = Paginator(approvals_qs, per_page)
#     page_obj = paginator.get_page(page_number)

#     paginated_approvals = list(page_obj.object_list)

#     # -------------------------------
#     # Build related maps
#     # -------------------------------
#     creator_ids = [
#         lad.creator_id_id
#         for lad in paginated_approvals
#         if lad.creator_id_id
#     ]

#     external_users = general_information.objects.filter(id__in=creator_ids)
#     user_map = {u.id: u for u in external_users}

#     creator_emp_ids = [u.faculty_id for u in external_users if u.faculty_id]

#     ext_users = USER.objects.using("rit_approval_system").filter(
#         Employee_id__in=creator_emp_ids,
#         is_active=True
#     )
#     ext_user_map = {u.Employee_id: u for u in ext_users}

#     role_ids = [u.role_id for u in ext_users if u.role_id]
#     external_roles = Role.objects.using("rit_approval_system").filter(id__in=role_ids)
#     role_map = {r.id: r for r in external_roles}

#     # -------------------------------
#     # Group by approver level
#     # -------------------------------
#     grouped_approvals = {}
#     for lad in paginated_approvals:
#         grouped_approvals.setdefault(lad.approver_level, []).append(lad)

#     # -------------------------------
#     # POST ACTIONS
#     # -------------------------------
#     if request.method == "POST":

#         # ---------------- EXPORT ----------------
#         if request.POST.get("export") == "excel":
#             return export_leave_approvals_to_excel(approvals_qs, tab)

#         if request.POST.get("export") == "pdf":
#             return export_leave_approvals_to_pdf(approvals_qs, tab)

#         # ---------------- APPROVE / REJECT ----------------
#         action = request.POST.get("action")
#         selected_ids = request.POST.getlist("selected_ids")

#         if not selected_ids:
#             messages.warning(request, "No leave applications selected.")
#             return redirect(reverse("leave_approvals"))

#         selected_leaves = (
#             LeaveApproversData.objects.filter(
#                 id__in=selected_ids,
#                 approver_id=approver_faculty
#             ).select_related("leave_application")
#         )

#         if action == "approve":
#             with transaction.atomic():
#                 for lad in selected_leaves:
#                     if lad.status != LeaveApproversData.Status.PENDING:
#                         continue

#                     lad.status = LeaveApproversData.Status.APPROVED
#                     lad.approved_date = timezone.now()
#                     lad.save(update_fields=["status", "approved_date"])

#                     app = lad.leave_application
#                     all_approved = not LeaveApproversData.objects.filter(
#                         leave_application=app
#                     ).exclude(status=LeaveApproversData.Status.APPROVED).exists()

#                     if all_approved and app.status != "Approved":
#                         app.status = "Approved"
#                         app.save(update_fields=["status"])

#             messages.success(request, "Selected leave applications approved successfully.")
#             return redirect(f"{reverse('leave_approvals')}?tab=pending")

#         elif action == "reject":
#             reason = request.POST.get("rejection_reason", "").strip()
#             if not reason:
#                 messages.error(request, "Rejection reason is required.")
#                 return redirect(f"{reverse('leave_approvals')}?tab=pending")

#             with transaction.atomic():
#                 for lad in selected_leaves:
#                     if lad.status != LeaveApproversData.Status.PENDING:
#                         continue

#                     app = lad.leave_application
#                     days = (app.to_date - app.from_date).days + 1

#                     # Refund using faculty + leave_type + date window
#                     lb = LeaveBalance.objects.select_for_update().filter(
#                         faculty=app.faculty,
#                         designation=app.designation,
#                         leave_type=app.leave_type,
#                         start_date__lte=app.from_date,
#                         end_date__gte=app.to_date,
#                     ).first()

#                     if lb:
#                         lb.available = (lb.available or 0) + days
#                         lb.used = max(0, (lb.used or 0) - days)
#                         lb.save(update_fields=["available", "used"])

#                     app.status = "Rejected"
#                     app.save(update_fields=["status"])

#                     lad.status = LeaveApproversData.Status.REJECTED
#                     lad.reason = reason
#                     lad.approved_date = timezone.now()
#                     lad.save(update_fields=["status", "reason", "approved_date"])

#             messages.success(request, "Selected applications rejected and balances refunded.")
#             return redirect(f"{reverse('leave_approvals')}?tab=pending")

#         else:
#             messages.error(request, "Invalid action.")
#             return redirect(reverse("leave_approvals"))

#     return render(
#         request,
#         "faculty_leave_management/leave_approval.html",
#         {
#             "page_obj": page_obj,
#             "grouped_approvals": grouped_approvals,
#             "user_map": user_map,
#             "ext_user_map": ext_user_map,
#             "role_map": role_map,
#             "q": q,
#             "level_filter": level_filter,
#             "available_levels": available_levels,
#             "available_academic_years": available_academic_years,
#             "academic_year_filter": academic_year_filter,
#             "tab": tab,
#             "pending_count": pending_count,
#             "approved_count": approved_count,
#             "rejected_count": rejected_count,
#         },
#     )


# def export_leave_approvals_to_excel(approvals_qs, tab="pending"):
#     workbook = Workbook()
#     worksheet = workbook.active
#     worksheet.title = "Leave Report"

#     headers = [
#         "Approver Level",
#         "Faculty ID",
#         "Faculty Name",
#         "Department",
#         "Designation",
#         "Academic Year",
#         "Leave Type",
#         "From Date",
#         "To Date",
#         "Days",
#         "Reason",
#         "Applied On",
#         "Status",
#     ]

#     if tab == "rejected":
#         headers.append("Remarks")

#     # Styling
#     header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
#     header_font = Font(color="FFFFFF", bold=True)

#     for col_num, header in enumerate(headers, 1):
#         cell = worksheet.cell(row=1, column=col_num, value=header)
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = Alignment(horizontal="center", vertical="center")

#     row_num = 2

#     for approval in approvals_qs:
#         app = approval.leave_application
#         faculty = app.faculty if app else None

#         faculty_id = getattr(faculty, "faculty_id", "")
#         faculty_name = getattr(faculty, "name", "")

#         department = ""
#         try:
#             department = str(faculty.department) if faculty and faculty.department else ""
#         except:
#             department = ""

#         designation = ""
#         try:
#             designation = str(app.designation) if app and app.designation else ""
#         except:
#             designation = ""

#         leave_type = ""
#         try:
#             leave_type = str(app.leave_type) if app and app.leave_type else ""
#         except:
#             leave_type = ""

#         days = ""
#         try:
#             if app and app.from_date and app.to_date:
#                 days = (app.to_date - app.from_date).days + 1
#         except:
#             days = ""

#         row = [
#             approval.approver_level,
#             faculty_id,
#             faculty_name,
#             department,
#             designation,
#             app.academic_year if app else "",
#             leave_type,
#             app.from_date.strftime("%d-%m-%Y") if app and app.from_date else "",
#             app.to_date.strftime("%d-%m-%Y") if app and app.to_date else "",
#             days,
#             app.reason if app else "",
#             app.requested_date.strftime("%d-%m-%Y %H:%M") if app and app.requested_date else "",
#             approval.status,
#         ]

#         if tab == "rejected":
#             row.append(approval.reason or "")

#         for col_num, value in enumerate(row, 1):
#             cell = worksheet.cell(row=row_num, column=col_num, value=value)
#             cell.alignment = Alignment(vertical="top", wrap_text=True)

#         row_num += 1

#     # Auto width
#     for column_cells in worksheet.columns:
#         max_length = 0
#         column_letter = column_cells[0].column_letter

#         for cell in column_cells:
#             try:
#                 if cell.value:
#                     max_length = max(max_length, len(str(cell.value)))
#             except:
#                 pass

#         worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

#     filename = f"leave_{tab}_report.xlsx"

#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response["Content-Disposition"] = f'inline; filename="{filename}"'

#     workbook.save(response)
#     return response


# def export_leave_approvals_to_pdf(approvals_qs, tab="pending"):
#     # connect your existing PDF generation logic here
#     # example:
#     # return your_leave_pdf_function(approvals_qs, tab)
#     pass



from collections import defaultdict
import os

from django.contrib import messages
from django.contrib.staticfiles import finders
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# make sure these models are imported correctly in your file
# from faculty_leave_management.models import LeaveApprovers, LeaveApproversData, LeaveBalance
# from faculty_leave_management.models import LeaveApplication
# from user_accounts.models import Add_Department
# from faculty_management.models import general_information
# from your approval db models import USER, Role


def leave_report_header(canvas, doc):
    canvas.saveState()
    canvas.setTitle(getattr(doc, "title", "Leave Report"))

    page_w, page_h = landscape(A4)
    left = doc.leftMargin
    right = page_w - doc.rightMargin
    center_x = page_w / 2

    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    ACCENT_RED = colors.HexColor("#b91c1c")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    logo_rel = "images/ritlogo.png"
    logo_path = finders.find(logo_rel)

    if not logo_path:
        static_root = getattr(settings, "STATIC_ROOT", "")
        if static_root:
            candidate = os.path.join(static_root, logo_rel)
            if os.path.exists(candidate):
                logo_path = candidate

    if logo_path and os.path.exists(logo_path):
        try:
            canvas.drawImage(
                ImageReader(logo_path),
                left,
                page_h - 24 * mm,
                width=30 * mm,
                height=18 * mm,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass

    canvas.setFillColor(PRIMARY_BLUE)
    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawCentredString(center_x, page_h - 8 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

    canvas.setFillColor(ACCENT_RED)
    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawCentredString(center_x, page_h - 14 * mm, "An Autonomous Institution")

    canvas.setFillColor(MEDIUM_GRAY)
    canvas.setFont("Helvetica", 8.2)
    canvas.drawCentredString(center_x, page_h - 19 * mm, "Approved by AICTE, New Delhi")
    canvas.setFont("Helvetica", 8)
    canvas.drawCentredString(center_x, page_h - 23.5 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
    canvas.drawCentredString(center_x, page_h - 28 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

    canvas.setStrokeColor(BORDER_GRAY)
    canvas.setLineWidth(0.8)
    canvas.line(left, page_h - 31.5 * mm, right, page_h - 31.5 * mm)

    footer_y = 8 * mm
    canvas.line(left, footer_y + 7 * mm, right, footer_y + 7 * mm)

    local_time = timezone.localtime()
    canvas.setFillColor(LIGHT_GRAY)
    canvas.setFont("Helvetica", 8)
    canvas.drawString(left, footer_y, f"Generated: {local_time.strftime('%d %b %Y, %I:%M %p')}")
    canvas.drawCentredString(center_x, footer_y, "Leave Approval Report")
    canvas.drawRightString(right, footer_y, f"Page {doc.page}")

    canvas.restoreState()


@no_cache
@check_permission("leave_approvals")
@faculty_leave_management
def leave_approvals(request):
    """
    Displays hierarchical leave applications for the current approver.

    Tabs:
    - Pending: only currently actionable applications
    - Approved: applications already approved by current approver
    - Rejected: applications rejected by current approver

    Supports:
    - search
    - approver level filter
    - academic year filter
    - pagination
    - export (excel / pdf)
    """

    approver_user = USER.objects.using("rit_approval_system").filter(
        Employee_id=request.user.Employee_id,
        is_active=True
    ).first()
    if not approver_user:
        messages.error(request, "Approver not found.")
        return redirect("dashboard")

    approver_faculty = general_information.objects.filter(
        faculty_id=request.user.Employee_id
    ).first()
    if not approver_faculty:
        messages.error(request, "Approver faculty record not found.")
        return redirect("dashboard")

    # -------------------------------
    # Read search / filter params
    # -------------------------------
    request_data = request.POST if request.method == "POST" else request.GET

    q = (request_data.get("q") or "").strip()
    level_filter = (request_data.get("level") or "").strip()
    academic_year_filter = (request_data.get("academic_year") or "").strip()
    tab = (request_data.get("tab") or "pending").strip().lower()
    page_number = request_data.get("page", 1)
    per_page = 50

    valid_tabs = ["pending", "approved", "rejected"]
    if tab not in valid_tabs:
        tab = "pending"

    # -----------------------------------------
    # Build actionable pending ids DIRECTLY from this approver's rows.
    #
    # Keyed on the approver's faculty record and the stored approver_level, so it
    # works even when the employee holds multiple roles (e.g. Assoc. Prof + HOD).
    # The previous approach resolved a single role via USER.first(), which could
    # pick a non-approver role and hide every application from the approver.
    # -----------------------------------------
    eligible_pending_ids = []

    # Scope to the role the approver is logged in as, while still supporting
    # legacy rows created before approver_role_id was populated and cases where
    # the local session role and rit_approval_system role do not line up.
    current_role_ids = {
        rid for rid in (
            getattr(request.user, "role_id", None),
            getattr(approver_user, "role_id", None),
        ) if rid
    }
    if current_role_ids:
        role_scope = Q(approver_role_id__in=current_role_ids) | Q(approver_role_id__isnull=True)
    else:
        role_scope = Q(approver_role_id__isnull=True)

    my_pending = (
        LeaveApproversData.objects.filter(
            role_scope,
            approver_id=approver_faculty,
            status=LeaveApproversData.Status.PENDING,
        )
        .select_related(
            "leave_application",
            "leave_application__faculty",
            "leave_application__designation",
            "leave_application__leave_type",
            "creator_id",
        )
    )

    for lad in my_pending:
        # Actionable only once every lower level has already approved.
        lower_pending = LeaveApproversData.objects.filter(
            leave_application=lad.leave_application,
            approver_level__lt=lad.approver_level,
        ).exclude(status=LeaveApproversData.Status.APPROVED).exists()

        if not lower_pending:
            eligible_pending_ids.append(lad.id)

    # -----------------------------------------
    # Base querysets for tabs
    # -----------------------------------------
    pending_base_qs = LeaveApproversData.objects.filter(
        id__in=eligible_pending_ids
    )

    approved_base_qs = LeaveApproversData.objects.filter(
        role_scope,
        approver_id=approver_faculty,
        status=LeaveApproversData.Status.APPROVED,
    )

    rejected_base_qs = LeaveApproversData.objects.filter(
        role_scope,
        approver_id=approver_faculty,
        status=LeaveApproversData.Status.REJECTED,
    )

    # -----------------------------------------
    # Tab counts (before extra filters)
    # -----------------------------------------
    pending_count = pending_base_qs.count()
    approved_count = approved_base_qs.count()
    rejected_count = rejected_base_qs.count()

    # -----------------------------------------
    # Choose queryset
    # -----------------------------------------
    if tab == "approved":
        approvals_qs = approved_base_qs
    elif tab == "rejected":
        approvals_qs = rejected_base_qs
    else:
        approvals_qs = pending_base_qs

    approvals_qs = approvals_qs.select_related(
        "leave_application",
        "leave_application__faculty",
        "leave_application__designation",
        "leave_application__leave_type",
        "approver_id",
        "creator_id",
    ).order_by("approver_level", "-leave_application__requested_date", "-id")

    # -------------------------------
    # Search
    # -------------------------------
    if q:
        approvals_qs = approvals_qs.filter(
            Q(leave_application__from_date__icontains=q) |
            Q(leave_application__to_date__icontains=q) |
            Q(leave_application__academic_year__icontains=q) |
            Q(leave_application__reason__icontains=q) |
            Q(leave_application__status__icontains=q) |
            Q(leave_application__faculty__faculty_id__icontains=q) |
            Q(leave_application__faculty__name__icontains=q) |
            Q(leave_application__leave_type__leave_type__icontains=q)
        )

    # -------------------------------
    # Level filter
    # -------------------------------
    if level_filter:
        try:
            approvals_qs = approvals_qs.filter(approver_level=int(level_filter))
        except ValueError:
            pass

    # -------------------------------
    # Academic year filter
    # -------------------------------
    if academic_year_filter:
        approvals_qs = approvals_qs.filter(
            leave_application__academic_year=academic_year_filter
        )

    # -------------------------------
    # Dropdown data
    # -------------------------------
    all_my_rows = LeaveApproversData.objects.filter(
        approver_id=approver_faculty
    )

    available_levels = (
        all_my_rows.values_list("approver_level", flat=True)
        .distinct()
        .order_by("approver_level")
    )

    available_academic_years = (
        all_my_rows.values_list("leave_application__academic_year", flat=True)
        .distinct()
        .order_by("leave_application__academic_year")
    )

    # -------------------------------
    # POST ACTIONS
    # -------------------------------
    if request.method == "POST":

        # ---------------- EXPORT ----------------
        if request.POST.get("export") == "excel":
            return export_leave_approvals_to_excel(approvals_qs, tab)

        if request.POST.get("export") == "pdf":
            return export_leave_approvals_to_pdf(approvals_qs, tab)

        # ---------------- APPROVE / REJECT ----------------
        action = request.POST.get("action")
        selected_ids = request.POST.getlist("selected_ids")

        if action == "approve_all":
            # Approve every eligible pending application (all pages, ignores checkboxes)
            all_pending = (
                LeaveApproversData.objects.filter(
                    id__in=eligible_pending_ids,
                    approver_id=approver_faculty,
                    status=LeaveApproversData.Status.PENDING,
                ).select_related(
                    "leave_application",
                    "leave_application__faculty",
                    "leave_application__designation",
                    "leave_application__leave_type",
                )
            )
            approved_count_all = 0
            skipped_proof_all = 0
            with transaction.atomic():
                for lad in all_pending:
                    lower_pending = LeaveApproversData.objects.filter(
                        leave_application=lad.leave_application,
                        approver_level__lt=lad.approver_level,
                    ).exclude(status=LeaveApproversData.Status.APPROVED).exists()
                    if lower_pending:
                        continue
                    # On Duty / Research On Duty cannot be approved until proof is uploaded.
                    if _od_awaiting_proof(lad.leave_application):
                        skipped_proof_all += 1
                        continue
                    lad.status = LeaveApproversData.Status.APPROVED
                    lad.approved_date = timezone.now()
                    lad.save(update_fields=["status", "approved_date"])
                    approved_count_all += 1
                    app = lad.leave_application
                    all_approved = not LeaveApproversData.objects.filter(
                        leave_application=app
                    ).exclude(status=LeaveApproversData.Status.APPROVED).exists()
                    if all_approved and app.status != "Approved":
                        app.status = "Approved"
                        app.save(update_fields=["status"])
            if approved_count_all:
                messages.success(request, f"{approved_count_all} leave application(s) approved successfully.")
            if skipped_proof_all:
                messages.warning(
                    request,
                    f"{skipped_proof_all} On Duty / Research On Duty application(s) skipped — "
                    f"proof PDF not yet uploaded."
                )
            return redirect(f"{reverse('leave_approvals')}?tab=pending")

        if not selected_ids:
            messages.warning(request, "No leave applications selected.")
            return redirect(reverse("leave_approvals"))

        selected_leaves = (
            LeaveApproversData.objects.filter(
                id__in=selected_ids,
                approver_id=approver_faculty
            ).select_related(
                "leave_application",
                "leave_application__faculty",
                "leave_application__designation",
                "leave_application__leave_type",
            )
        )

        if action == "approve":
            skipped_proof = 0
            with transaction.atomic():
                for lad in selected_leaves:
                    if lad.status != LeaveApproversData.Status.PENDING:
                        continue

                    # hierarchy safety recheck
                    lower_pending = LeaveApproversData.objects.filter(
                        leave_application=lad.leave_application,
                        approver_level__lt=lad.approver_level,
                    ).exclude(status=LeaveApproversData.Status.APPROVED).exists()

                    if lower_pending:
                        continue

                    # On Duty / Research On Duty cannot be approved until proof is uploaded.
                    if _od_awaiting_proof(lad.leave_application):
                        skipped_proof += 1
                        continue

                    lad.status = LeaveApproversData.Status.APPROVED
                    lad.approved_date = timezone.now()
                    lad.save(update_fields=["status", "approved_date"])

                    app = lad.leave_application

                    all_approved = not LeaveApproversData.objects.filter(
                        leave_application=app
                    ).exclude(status=LeaveApproversData.Status.APPROVED).exists()

                    if all_approved and app.status != "Approved":
                        app.status = "Approved"
                        app.save(update_fields=["status"])

            if skipped_proof:
                messages.warning(
                    request,
                    f"{skipped_proof} On Duty / Research On Duty application(s) could not be approved — "
                    f"the proof PDF has not been uploaded yet. Please view/await the proof, then approve."
                )
            else:
                messages.success(request, "Selected leave applications approved successfully.")
            return redirect(f"{reverse('leave_approvals')}?tab=pending")

        elif action == "reject":
            reason = request.POST.get("rejection_reason", "").strip()
            if not reason:
                messages.error(request, "Rejection reason is required.")
                return redirect(f"{reverse('leave_approvals')}?tab=pending")

            with transaction.atomic():
                for lad in selected_leaves:
                    if lad.status != LeaveApproversData.Status.PENDING:
                        continue

                    # hierarchy safety recheck
                    lower_pending = LeaveApproversData.objects.filter(
                        leave_application=lad.leave_application,
                        approver_level__lt=lad.approver_level,
                    ).exclude(status=LeaveApproversData.Status.APPROVED).exists()

                    if lower_pending:
                        continue

                    app = lad.leave_application
                    days = 0
                    if app and app.from_date and app.to_date:
                        days = calculate_leave_days(
                            app.from_date, app.to_date, app.session_id, app.to_session_id
                        )

                    # Refund leave balance
                    lb = LeaveBalance.objects.select_for_update().filter(
                        faculty=app.faculty,
                        designation=app.designation,
                        leave_type=app.leave_type,
                        academic_year=app.academic_year,
                        start_date__lte=app.from_date,
                        end_date__gte=app.to_date,
                    ).first()

                    if not lb:
                        # Fall back to any balance row for this academic year — a
                        # CCL balance row credited with a narrow earn-date window
                        # (see _credit_ccl_on_approval) would otherwise never match
                        # the exact date-range filter above, silently skipping the
                        # refund entirely.
                        lb = LeaveBalance.objects.select_for_update().filter(
                            faculty=app.faculty,
                            designation=app.designation,
                            leave_type=app.leave_type,
                            academic_year=app.academic_year,
                        ).first()

                    if lb:
                        _before = (lb.available, lb.used)
                        lb.available = (lb.available or 0) + days
                        lb.used = max(0, (lb.used or 0) - days)
                        lb.save(update_fields=["available", "used"])
                        _log_balance_change("leave_approvals:reject_refund", lb, before=_before,
                                             extra=f"leave_id={app.id} days={days}")

                    app.status = "Rejected"
                    app.save(update_fields=["status"])

                    lad.status = LeaveApproversData.Status.REJECTED
                    lad.reason = reason
                    lad.approved_date = timezone.now()
                    lad.save(update_fields=["status", "reason", "approved_date"])

            messages.success(request, "Selected applications rejected and balances refunded.")
            return redirect(f"{reverse('leave_approvals')}?tab=pending")

        else:
            messages.error(request, "Invalid action.")
            return redirect(reverse("leave_approvals"))

    # -------------------------------
    # Pagination
    # -------------------------------
    paginator = Paginator(approvals_qs, per_page)
    page_obj = paginator.get_page(page_number)

    paginated_approvals = list(page_obj.object_list)

    # -------------------------------
    # Build related maps
    # -------------------------------
    creator_ids = [
        lad.creator_id_id
        for lad in paginated_approvals
        if lad.creator_id_id
    ]

    external_users = general_information.objects.filter(id__in=creator_ids)
    user_map = {u.id: u for u in external_users}

    creator_emp_ids = [u.faculty_id for u in external_users if u.faculty_id]

    ext_users = USER.objects.using("rit_approval_system").filter(
        Employee_id__in=creator_emp_ids,
        is_active=True
    )
    ext_user_map = {u.Employee_id: u for u in ext_users}

    role_ids = [u.role_id for u in ext_users if u.role_id]
    external_roles = Role.objects.using("rit_approval_system").filter(id__in=role_ids)
    role_map = {r.id: r for r in external_roles}

    # -------------------------------
    # Group by approver level
    # -------------------------------
    grouped_approvals = {}
    for lad in paginated_approvals:
        # On Duty / Research On Duty cannot be approved until proof is uploaded —
        # flag it so the template can disable selection and show "Awaiting proof".
        lad.awaiting_proof = _od_awaiting_proof(lad.leave_application)
        grouped_approvals.setdefault(lad.approver_level, []).append(lad)

    return render(
        request,
        "faculty_leave_management/leave_approval.html",
        {
            "page_obj": page_obj,
            "grouped_approvals": grouped_approvals,
            "user_map": user_map,
            "ext_user_map": ext_user_map,
            "role_map": role_map,
            "q": q,
            "level_filter": level_filter,
            "available_levels": available_levels,
            "available_academic_years": available_academic_years,
            "academic_year_filter": academic_year_filter,
            "tab": tab,
            "pending_count": pending_count,
            "approved_count": approved_count,
            "rejected_count": rejected_count,
        },
    )




def export_leave_approvals_to_excel(approvals_qs, tab="pending"):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Leave Report"

    headers = [
        "Approver Level",
        "Faculty ID",
        "Faculty Name",
        "Department",
        "Designation",
        "Academic Year",
        "Leave Type",
        "From Date",
        "To Date",
        "Days",
        "Reason",
        "Applied On",
        "Status",
    ]

    if tab == "rejected":
        headers.append("Remarks")

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_num = 2

    for approval in approvals_qs:
        app = approval.leave_application
        faculty = app.faculty if app else None

        faculty_id = getattr(faculty, "faculty_id", "")
        faculty_name = getattr(faculty, "name", "")

        department = ""
        try:
            department = str(faculty.department) if faculty and faculty.department else ""
        except Exception:
            department = ""

        designation = ""
        try:
            designation = str(app.designation) if app and app.designation else ""
        except Exception:
            designation = ""

        leave_type = ""
        try:
            leave_type = str(app.leave_type) if app and app.leave_type else ""
        except Exception:
            leave_type = ""

        days = ""
        try:
            if app and app.from_date and app.to_date:
                days = (app.to_date - app.from_date).days + 1
        except Exception:
            days = ""

        row = [
            approval.approver_level,
            faculty_id,
            faculty_name,
            department,
            designation,
            app.academic_year if app else "",
            leave_type,
            app.from_date.strftime("%d-%m-%Y") if app and app.from_date else "",
            app.to_date.strftime("%d-%m-%Y") if app and app.to_date else "",
            days,
            app.reason if app else "",
            app.requested_date.strftime("%d-%m-%Y %H:%M") if app and app.requested_date else "",
            approval.status,
        ]

        if tab == "rejected":
            row.append(approval.reason or "")

        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        row_num += 1

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    filename = f"leave_{tab}_report.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'inline; filename="{filename}"'

    workbook.save(response)
    return response


def export_leave_approvals_to_pdf(approvals_qs, tab="pending"):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="leave_{tab}_report.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=36 * mm,
        bottomMargin=16 * mm,
    )
    doc.title = f"Leave {tab.capitalize()} Report"

    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=15,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f2f57"),
        spaceAfter=4,
    )

    sub_title_style = ParagraphStyle(
        "SubTitleStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#4b5563"),
        spaceAfter=10,
    )

    dept_band_style = ParagraphStyle(
        "DeptBandStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11,
        alignment=TA_LEFT,
        textColor=colors.white,
        backColor=colors.HexColor("#1F4E78"),
        borderPadding=6,
        leading=14,
        spaceBefore=8,
        spaceAfter=2,
    )

    year_style = ParagraphStyle(
        "YearStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        textColor=colors.HexColor("#163A5F"),
        spaceBefore=12,
        spaceAfter=5,
    )

    cell_style = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.black,
        alignment=TA_LEFT,
    )

    cell_style_center = ParagraphStyle(
        "CellStyleCenter",
        parent=cell_style,
        alignment=TA_CENTER,
    )

    header_left_style = ParagraphStyle(
        "HeaderLeftStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.5,
        textColor=colors.HexColor("#163A5F"),
        alignment=TA_LEFT,
    )

    header_center_style = ParagraphStyle(
        "HeaderCenterStyle",
        parent=header_left_style,
        alignment=TA_CENTER,
    )

    local_time = timezone.localtime()

    elements.append(Paragraph(f"Leave {tab.capitalize()} Report", title_style))
    elements.append(Paragraph(f"Generated on {local_time.strftime('%d-%m-%Y %I:%M %p')}", sub_title_style))
    elements.append(Spacer(1, 4))

    if not approvals_qs.exists():
        elements.append(Paragraph("No records found.", styles["Normal"]))
        doc.build(elements, onFirstPage=leave_report_header, onLaterPages=leave_report_header)
        return response

    grouped = defaultdict(lambda: defaultdict(list))

    for approval in approvals_qs:
        app = getattr(approval, "leave_application", None)
        faculty = getattr(app, "faculty", None) if app else None

        if faculty and getattr(faculty, "department", None):
            dept = faculty.department
            degree = getattr(dept, "degree", None)
            dept_text = str(dept)
            degree_code = getattr(degree, "degree_code", "") if degree else ""
        else:
            dept_text = "No Department"
            degree_code = ""

        academic_year = str(getattr(app, "academic_year", "No Academic Year")) if app else "No Academic Year"
        grouped[(dept_text, degree_code)][academic_year].append(approval)

    for (dept_text, degree_code) in sorted(grouped.keys(), key=lambda x: (x[0].lower(), x[1].lower())):
        dept_years = grouped[(dept_text, degree_code)]

        elements.append(
            Paragraph(
                f"""
                <para alignment="center">
                    <b>{degree_code}</b> &nbsp;&nbsp; | &nbsp;&nbsp; {dept_text}
                </para>
                """ if degree_code else f"""
                <para alignment="center">
                    {dept_text}
                </para>
                """,
                dept_band_style
            )
        )

        for academic_year in sorted(dept_years.keys()):
            year_records = dept_years[academic_year]

            elements.append(
                Paragraph(
                    f"Academic Year : {academic_year}   |   Records : {len(year_records)}",
                    year_style
                )
            )

            data = [[
                Paragraph("Faculty ID", header_left_style),
                Paragraph("Faculty Name", header_left_style),
                Paragraph("Leave Type", header_left_style),
                Paragraph("From", header_center_style),
                Paragraph("To", header_center_style),
                Paragraph("Days", header_center_style),
                Paragraph("Status", header_center_style),
                Paragraph("Designation", header_left_style),
            ]]

            total_days = 0

            for approval in year_records:
                app = getattr(approval, "leave_application", None)
                faculty = getattr(app, "faculty", None) if app else None

                from_date = getattr(app, "from_date", None)
                to_date = getattr(app, "to_date", None)

                from_date_str = from_date.strftime("%d-%m-%Y") if from_date else "-"
                to_date_str = to_date.strftime("%d-%m-%Y") if to_date else "-"

                days_val = 0
                try:
                    if from_date and to_date:
                        days_val = (to_date - from_date).days + 1
                except Exception:
                    days_val = 0

                try:
                    total_days += float(days_val or 0)
                except Exception:
                    pass

                leave_type = ""
                try:
                    leave_type = str(app.leave_type) if app and app.leave_type else ""
                except Exception:
                    leave_type = ""

                designation = ""
                try:
                    designation = str(app.designation) if app and app.designation else ""
                except Exception:
                    designation = ""

                data.append([
                    Paragraph(str(getattr(faculty, "faculty_id", "")), cell_style),
                    Paragraph(str(getattr(faculty, "name", "")), cell_style),
                    Paragraph(leave_type, cell_style),
                    Paragraph(from_date_str, cell_style_center),
                    Paragraph(to_date_str, cell_style_center),
                    Paragraph(str(days_val), cell_style_center),
                    Paragraph(str(getattr(approval, "status", "")), cell_style_center),
                    Paragraph(designation, cell_style),
                ])

            data.append([
                Paragraph("<b>Total</b>", cell_style),
                Paragraph("", cell_style),
                Paragraph("", cell_style),
                Paragraph("", cell_style_center),
                Paragraph("", cell_style_center),
                Paragraph(f"<b>{total_days:g}</b>", cell_style_center),
                Paragraph("", cell_style_center),
                Paragraph("", cell_style),
            ])

            col_widths = [
                24 * mm,  # Faculty ID
                46 * mm,  # Faculty Name
                34 * mm,  # Leave Type
                22 * mm,  # From
                22 * mm,  # To
                16 * mm,  # Days
                22 * mm,  # Status
                46 * mm,  # Designation
            ]

            table = Table(data, colWidths=col_widths, repeatRows=1)

            table_style = TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DCEAF7")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#163A5F")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8.5),

                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#222222")),

                ("ALIGN", (0, 0), (2, -1), "LEFT"),
                ("ALIGN", (3, 0), (6, -1), "CENTER"),
                ("ALIGN", (7, 0), (7, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),

                ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#9DB9D5")),
                ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#7A9CBD")),
                ("LINEBELOW", (0, -1), (-1, -1), 1, colors.HexColor("#7A9CBD")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B8C7D6")),
                ("INNERGRID", (0, 1), (-1, -2), 0.35, colors.HexColor("#D5DEE8")),
            ])

            for row_idx in range(1, len(data) - 1):
                if row_idx % 2 == 1:
                    table_style.add("BACKGROUND", (0, row_idx), (-1, row_idx), colors.white)
                else:
                    table_style.add("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#F7FBFF"))

            table_style.add("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1), colors.HexColor("#EEF4FA"))
            table_style.add("FONTNAME", (0, len(data) - 1), (-1, len(data) - 1), "Helvetica-Bold")
            table_style.add("TEXTCOLOR", (0, len(data) - 1), (-1, len(data) - 1), colors.HexColor("#163A5F"))

            table.setStyle(table_style)

            elements.append(table)
            elements.append(Spacer(1, 8))

    doc.build(elements, onFirstPage=leave_report_header, onLaterPages=leave_report_header)
    return response





@no_cache
@check_permission("pending_leave_applications")
def pending_leave_applications(request):
    ext_user = USER.objects.using("rit_approval_system").filter(Employee_id=request.user.Employee_id, is_active=True).first()
    if not ext_user:
        messages.error(request, "User not found in external system.")
        return redirect("dashboard")

    approver_role_id = ext_user.role_id

    approver_faculty = general_information.objects.filter(faculty_id=request.user.Employee_id).first()
    if not approver_faculty:
        messages.error(request, "Approver faculty record not found.")
        return redirect("dashboard")

    approver_faculty_id = approver_faculty.id

    all_hierarchies = list(
        LeaveApprovers.objects.all()
        .values("creator_role_id", "approver_role_id", "approver_level")
        .order_by("creator_role_id", "approver_level")
    )

    related_creator_roles = {
        h["creator_role_id"]
        for h in all_hierarchies
        if h["approver_role_id"] == approver_role_id
    }

    pending_approvals = []

    for creator_role_id in related_creator_roles:
        hierarchy = [h for h in all_hierarchies if h["creator_role_id"] == creator_role_id]

        current_level = next(
            (h["approver_level"] for h in hierarchy if h["approver_role_id"] == approver_role_id),
            None,
        )
        if not current_level:
            continue

        pending_qs = (
            LeaveApproversData.objects.filter(
                approver_id_id=approver_faculty_id,
                approver_level=current_level,
                status=LeaveApproversData.Status.PENDING,
            )
            .select_related("leave_application", "creator_id", "approver_id")
        )

        for lad in pending_qs:
            lower_pending = (
                LeaveApproversData.objects.filter(
                    leave_application=lad.leave_application,
                    approver_level__lt=current_level,
                )
                .exclude(status=LeaveApproversData.Status.APPROVED)
                .exists()
            )
            if not lower_pending:
                pending_approvals.append(lad)

    grouped_pending = {}
    for lad in pending_approvals:
        grouped_pending.setdefault(lad.approver_level, []).append(lad)

    # ✅ map general_information -> external USER via faculty_id/Employee_id
    creator_emp_ids = [
        lad.creator_id.faculty_id for lad in pending_approvals
        if lad.creator_id and lad.creator_id.faculty_id
    ]

    ext_users = USER.objects.using("rit_approval_system").filter(Employee_id__in=creator_emp_ids, is_active=True)
    user_map = {u.Employee_id: u for u in ext_users}   # key = Employee_id

    role_ids = [u.role_id for u in ext_users if u.role_id]
    roles = Role.objects.using("rit_approval_system").filter(id__in=role_ids)
    role_map = {r.id: r for r in roles}

    return render(
        request,
        "faculty_leave_management/pending_leave_applications.html",
        {
            "grouped_pending": grouped_pending,
            "user_map": user_map,
            "role_map": role_map,
        },
    )


@no_cache
@check_permission("approved_leave_applications")
def approved_leave_applications(request):
    """
    Shows all leave applications approved by the currently logged-in approver.
    Grouped by approval level.
    """

    ext_user = USER.objects.using("rit_approval_system").filter(Employee_id=request.user.Employee_id, is_active=True).first()
    if not ext_user:
        messages.error(request, "User not found in external system.")
        return redirect("dashboard")

    approver_faculty = general_information.objects.filter(faculty_id=request.user.Employee_id).first()
    if not approver_faculty:
        messages.error(request, "Approver faculty record not found.")
        return redirect("dashboard")

    approver_faculty_id = approver_faculty.id

    approved_qs = (
        LeaveApproversData.objects.filter(
            approver_id_id=approver_faculty_id,   # ✅ general_information id
            status=LeaveApproversData.Status.APPROVED,
        )
        .select_related("leave_application", "creator_id", "approver_id")
        .order_by("-approved_date")
    )

    grouped_approved = {}
    for lad in approved_qs:
        grouped_approved.setdefault(lad.approver_level, []).append(lad)

    # ✅ map creator general_information -> external USER using faculty_id/Employee_id
    creator_emp_ids = [
        lad.creator_id.faculty_id for lad in approved_qs
        if lad.creator_id and lad.creator_id.faculty_id
    ]

    ext_users = USER.objects.using("rit_approval_system").filter(Employee_id__in=creator_emp_ids, is_active=True)
    user_map = {u.Employee_id: u for u in ext_users}   # key = Employee_id

    role_ids = [u.role_id for u in ext_users if u.role_id]
    
    roles = Role.objects.using("rit_approval_system").filter(id__in=role_ids)
    role_map = {r.id: r for r in roles}

    return render(
        request,
        "faculty_leave_management/approved_leave_applications.html",
        {
            "grouped_approved": grouped_approved,
            "user_map": user_map,
            "role_map": role_map,
        },
    )


@no_cache
@check_permission("rejected_leave_applications")
def rejected_leave_applications(request):
    """
    Shows all leave applications rejected by the currently logged-in approver.
    Grouped by approval level.
    """

    ext_user = USER.objects.using("rit_approval_system").filter(Employee_id=request.user.Employee_id, is_active=True).first()
    if not ext_user:
        messages.error(request, "User not found in external system.")
        return redirect("dashboard")

    approver_faculty = general_information.objects.filter(faculty_id=request.user.Employee_id).first()
    if not approver_faculty:
        messages.error(request, "Approver faculty record not found.")
        return redirect("dashboard")

    approver_faculty_id = approver_faculty.id

    # ---- Fetch Rejected Applications ----
    rejected_qs = (
        LeaveApproversData.objects.filter(
            approver_id_id=approver_faculty_id,  # ✅ general_information id
            status=LeaveApproversData.Status.REJECTED,
        )
        .select_related("leave_application", "creator_id", "approver_id")
        .order_by("-approved_date")
    )

    grouped_rejected = {}
    for lad in rejected_qs:
        grouped_rejected.setdefault(lad.approver_level, []).append(lad)

    # ✅ map creator general_information -> external USER via faculty_id/Employee_id
    creator_emp_ids = [
        lad.creator_id.faculty_id for lad in rejected_qs
        if lad.creator_id and lad.creator_id.faculty_id
    ]

    ext_users = USER.objects.using("rit_approval_system").filter(Employee_id__in=creator_emp_ids, is_active=True)
    user_map = {u.Employee_id: u for u in ext_users}  # key = Employee_id

    role_ids = [u.role_id for u in ext_users if u.role_id]
    roles = Role.objects.using("rit_approval_system").filter(id__in=role_ids)
    role_map = {r.id: r for r in roles}

    return render(
        request,
        "faculty_leave_management/rejected_leave_applications.html",
        {
            "grouped_rejected": grouped_rejected,
            "user_map": user_map,
            "role_map": role_map,
        },
    )
  



def _resolve_leave_allotment(faculty, academic_year, leave_type):
    """
    Resolve the LeaveAllotment governing a faculty member for a given leave type.

    Allotments are configured either category-wise or role-wise. A category-based
    allotment takes precedence over a role-based one, so faculty grouped under a
    category get the category quota even if their designation has its own row.

    `leave_type` may be a LeaveType instance or its id.
    Returns the LeaveAllotment instance, or None if no active allotment matches.
    """
    allotment = None
    if faculty.category_id:
        allotment = LeaveAllotment.objects.filter(
            academic_year=academic_year,
            category_id=faculty.category_id,
            leave_type=leave_type,
            active=True,
        ).first()
    if not allotment:
        allotment = LeaveAllotment.objects.filter(
            academic_year=academic_year,
            role=faculty.designation,
            leave_type=leave_type,
            active=True,
        ).first()
    return allotment


# ---------------------------------------------------------------------------
# CCL earn-flow helpers: holiday validation, biometric punches, timing mapping
# ---------------------------------------------------------------------------
def _ccl_holiday_on(role_id, on_date, category_id=None):
    """Return the Employee_Holidays row that makes ``on_date`` a holiday for a
    faculty of ``role_id`` (optionally scoped to ``category_id``), else None."""
    qs = Employee_Holidays.objects.filter(role_id=role_id, holiday_date=on_date)
    if category_id:
        qs = qs.filter(Q(category_id=category_id) | Q(category__isnull=True))
    else:
        qs = qs.filter(category__isnull=True)
    return qs.order_by("-category_id").first()


def _ccl_punch_in_out(faculty_id, on_date):
    """Read the faculty's biometric punches for ``on_date`` from attendance_db.

    Returns (in_datetime, out_datetime, punch_count). The first punch of the day
    is treated as IN and the last as OUT. Returns (None, None, 0) when the
    monthly log table is missing or there are no punches.
    """
    table_name = f"DeviceLogs_{on_date.month}_{on_date.year}"
    try:
        with connections["attendance_db"].cursor() as cursor:
            cursor.execute(
                f"SELECT LogDate FROM {table_name} WHERE UserId = %s ORDER BY LogDate ASC",
                [str(faculty_id)],
            )
            log_dates = [row[0] for row in cursor.fetchall()]
    except Exception:
        # Table for that month may not exist yet, or DB unreachable.
        return (None, None, 0)

    day_logs = [d for d in log_dates if d and d.date() == on_date]
    if not day_logs:
        return (None, None, 0)
    return (day_logs[0], day_logs[-1], len(day_logs))


def _ccl_award_from_hours(worked_hours):
    """Map worked hours to (ccl_days, session_name) using CCLTimingMaster.

    Matches the tightest band where ``min_hours < worked_hours <= max_hours``
    (max_hours None = open-ended). Returns (None, None) if nothing qualifies.
    """
    for rule in CCLTimingMaster.objects.filter(is_active=True).order_by("-min_hours"):
        if worked_hours > (rule.min_hours or 0) and (
            rule.max_hours is None or worked_hours <= rule.max_hours
        ):
            return (rule.ccl_days, rule.session_name)
    return (None, None)


def _credit_ccl_on_approval(ccl_app):
    """Credit earned comp-off when a CCL application becomes fully approved.

    Adds ``days`` to the faculty's CCL LeaveBalance (what the leave form spends
    from and what the applicant balance card shows) and to the CCL_Claim ledger.
    Guarded by ``is_claimed`` so a given application is only ever credited once.
    """
    if not ccl_app or ccl_app.is_claimed:
        return

    days = Decimal(str(ccl_app.days or 0))
    faculty = ccl_app.faculty
    academic_year = ccl_app.academic_year

    if days > 0 and faculty:
        ccl_leave_type = LeaveType.objects.filter(code__iexact="CCL").first()
        if ccl_leave_type:
            lb = LeaveBalance.objects.select_for_update().filter(
                faculty=faculty,
                leave_type=ccl_leave_type,
                academic_year=academic_year,
            ).first()
            if not lb:
                allotment = _resolve_leave_allotment(faculty, academic_year, ccl_leave_type)
                lb = LeaveBalance.objects.create(
                    faculty=faculty,
                    designation=faculty.designation,
                    leave_type=ccl_leave_type,
                    academic_year=academic_year,
                    available=0,
                    used=0,
                    start_date=allotment.start_date if allotment else ccl_app.date,
                    end_date=allotment.end_date if allotment else ccl_app.date,
                )
            _before = (lb.available, lb.used)
            lb.available = (lb.available or 0) + days
            lb.save(update_fields=["available"])
            _log_balance_change("_credit_ccl_on_approval", lb, faculty=faculty, before=_before,
                                 extra=f"ccl_app_id={ccl_app.id} credited_days={days}")

        claim, _created = CCL_Claim.objects.get_or_create(
            faculty=faculty,
            designation=faculty.designation,
            academic_year=academic_year,
            defaults={"claimed": Decimal("0.0"), "used": Decimal("0.0"), "total_claimed": Decimal("0.0")},
        )
        claim.claimed = (claim.claimed or 0) + days
        claim.total_claimed = (claim.total_claimed or 0) + days
        claim.save()

    ccl_app.is_claimed = True
    ccl_app.save(update_fields=["is_claimed"])


def _ccl_approver_status_rows(ccl):
    """Approver breakdown for one CCL_Application, in template/JSON-ready shape.
    Works whether or not `approver_rows` was prefetched."""
    rows = []
    for row in ccl.approver_rows.all():
        appr = row.approver_id
        rows.append({
            "level": row.approver_level,
            "emp": getattr(appr, "faculty_id", "—"),
            "name": getattr(appr, "name", "") or getattr(appr, "username", ""),
            "desig": (
                getattr(getattr(appr, "designation", None), "designation_name", None)
                or str(getattr(appr, "designation", "") or "")
            ),
            "status": row.status,
        })
    return rows


def _ccl_is_locked_from_rows(ccl, approver_rows):
    """Lock check reusing already-fetched approver rows (bulk/listing path)."""
    if (ccl.status or "").lower() != "pending":
        return True
    return any(r["status"] != CCL_Approvers_Data.Status.PENDING for r in approver_rows)


def _ccl_is_locked_db(ccl):
    """Lock check for a single freshly-fetched CCL_Application (save/delete path)."""
    if (ccl.status or "").lower() != "pending":
        return True
    return CCL_Approvers_Data.objects.filter(ccl_application=ccl).exclude(
        status=CCL_Approvers_Data.Status.PENDING
    ).exists()


def _ccl_is_admin_awarded(ccl, approver_rows):
    # Admin-awarded: Approved with no approver rows, or all approver rows still
    # Pending (admin set status directly without going through the approval chain).
    return (
        (ccl.status or "") == "Approved"
        and (not approver_rows or all(r["status"] == "Pending" for r in approver_rows))
    )


def _serialize_ccl_row(ccl, is_locked, approver_rows, is_admin_awarded):
    return {
        "id": ccl.id,
        "academic_year": ccl.academic_year or "",
        "date": ccl.date.strftime("%Y-%m-%d") if ccl.date else "",
        "days": float(ccl.days) if ccl.days is not None else 0.0,
        "reason": ccl.reason or "",
        "status": ccl.status or "",
        "session": ccl.session or "",
        "worked_hours": float(ccl.worked_hours) if ccl.worked_hours is not None else 0.0,
        "is_locked": is_locked,
        "is_admin_awarded": is_admin_awarded,
        "approver_rows": approver_rows,
    }


def _validate_and_build_ccl_submission(faculty, creator_role_id, date, reason):
    """CCL earn-flow validation (holiday + biometric punches + timing), shared by
    create and update so the two paths can never drift apart. Returns the
    computed fields dict, or raises ValidationError with a user-facing message.
    """
    holiday = _ccl_holiday_on(
        creator_role_id, date, getattr(faculty, "category_id", None)
    )
    if not holiday:
        raise ValidationError(
            "CCL can be claimed only for a declared holiday. The selected date is not a holiday."
        )

    in_dt, out_dt, punch_count = _ccl_punch_in_out(faculty.faculty_id, date)
    if punch_count == 0:
        raise ValidationError(
            "No punch found for the selected date in the attendance system. "
            "Please contact the admin to enter your punch and then reapply."
        )
    if punch_count < 2:
        raise ValidationError(
            "Only a single punch is available for the selected date. "
            "Please contact the admin for entry of the punch and then reapply."
        )

    worked_hours = Decimal(
        str((out_dt - in_dt).total_seconds() / 3600)
    ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    ccl_days_val, ccl_session_name = _ccl_award_from_hours(worked_hours)
    if ccl_days_val is None:
        raise ValidationError(
            f"The worked duration ({worked_hours} hrs) does not qualify for any CCL "
            f"as per the CCL timing master."
        )

    return {
        "from_time": in_dt.time(),
        "to_time": out_dt.time(),
        "worked_hours": worked_hours,
        "session": ccl_session_name,
        "days": ccl_days_val,
    }


@check_permission("add_ccl_claim")
def ccl_application_form(request):
    """CCL Application self-service page shell.

    Renders the apply/edit form, filters, academic-year dropdown, and balance
    summary card. The applications table itself loads and paginates client-side
    via ccl_application_data(); create/update go through ccl_application_save(),
    delete through ccl_application_delete() — no more full-page POST/redirect.
    """

    # ----------------------------
    # Page access permission
    # ----------------------------
    if Faculty_Leave_Page_Permission.objects.filter(
        user_id=request.user.id, is_hidden=True
    ).exists():
        role_obj = Role.objects.using("rit_approval_system").get(id=request.user.role_id)
        role_label = getattr(role_obj, "role", None) or "current"
        return render(
            request,
            "faculty_leave_management/leave_access_restricted.html",
            {"role_label": role_label},
            status=403,
        )

    ext_user = USER.objects.using("rit_approval_system").filter(
        Employee_id=request.user.Employee_id, is_active=True
    ).first()

    faculty = general_information.objects.filter(
        faculty_id=request.user.Employee_id
    ).select_related("designation").first()

    if not ext_user:
        messages.error(request, "User information not found in external system.")
        return redirect("dashboard")

    if not faculty:
        messages.error(request, "Faculty record not found in local database.")
        return redirect("dashboard")

    academic_years = (
        CCL_Application.objects.filter(faculty=faculty)
        .values_list("academic_year", flat=True)
        .distinct()
        .order_by("academic_year")
    )

    ccl_leave_type = LeaveType.objects.filter(code__iexact="CCL").first()
    ccl_balance_summary = []
    if ccl_leave_type:
        ccl_lb = LeaveBalance.objects.filter(
            faculty=faculty,
            leave_type=ccl_leave_type,
            academic_year=settings.ACADEMIC_YEAR,
        ).first()

        if ccl_lb:
            ccl_used = Decimal(ccl_lb.used or 0)
            ccl_available = Decimal(ccl_lb.available or 0)
            ccl_total = ccl_available + ccl_used
        else:
            ccl_used = Decimal("0.0")
            ccl_available = Decimal("0.0")
            ccl_total = Decimal("0.0")

        ccl_balance_summary.append({
            "leave_type": ccl_leave_type.name,
            "total": ccl_total,
            "used": ccl_used,
            "available": ccl_available,
        })

    return render(
        request,
        "faculty_leave_management/ccl_application_form.html",
        {
            "academic_years": academic_years,
            "ccl_balance_summary": ccl_balance_summary,
            "academic_year": settings.ACADEMIC_YEAR,
        },
    )


@_json_forbidden
@check_permission("add_ccl_claim")
def ccl_application_data(request):
    """JSON listing endpoint for the CCL Application self-service page —
    server-side search / filter / sort / pagination."""
    hidden = _hidden_page_json_or_none(request)
    if hidden:
        return hidden

    faculty = general_information.objects.filter(
        faculty_id=request.user.Employee_id
    ).select_related("designation").first()
    if not faculty:
        return JsonResponse({"error": "Faculty record not found."}, status=404)

    q = (request.GET.get("q") or "").strip()
    fy = (request.GET.get("fy") or "").strip()
    status_f = (request.GET.get("st") or "").strip()
    _raw_df = (request.GET.get("df") or "").strip()
    _raw_dt = (request.GET.get("dt") or "").strip()
    df = parse_date(_raw_df) if _raw_df else None
    dt = parse_date(_raw_dt) if _raw_dt else None
    sort = (request.GET.get("sort") or "requested_date").strip()
    order = (request.GET.get("order") or "desc").strip().lower()

    qs = (
        CCL_Application.objects.filter(faculty=faculty)
        .select_related("designation")
        .prefetch_related(
            Prefetch(
                "approver_rows",
                queryset=CCL_Approvers_Data.objects.select_related(
                    "approver_id", "approver_id__designation"
                ).order_by("approver_level", "id"),
            )
        )
    )

    if fy:
        qs = qs.filter(academic_year=fy)
    if status_f:
        qs = qs.filter(status=status_f)
    if df and dt:
        qs = qs.filter(date__range=[df, dt])
    elif df:
        qs = qs.filter(date__gte=df)
    elif dt:
        qs = qs.filter(date__lte=dt)

    if q:
        # Search across Academic Year / Date / Days / Reason / Status / Session /
        # Worked Hours in one filter by casting every field to text.
        qs = qs.annotate(
            search_blob=Concat(
                Cast("academic_year", CharField()), Value(" "),
                Cast("date", CharField()), Value(" "),
                Cast("days", CharField()), Value(" "),
                Cast("reason", CharField()), Value(" "),
                Cast("status", CharField()), Value(" "),
                Cast("session", CharField()), Value(" "),
                Cast("worked_hours", CharField()),
                output_field=CharField(),
            )
        ).filter(search_blob__icontains=q)

    sort_fields = {
        "academic_year", "date", "days", "status", "session",
        "worked_hours", "requested_date",
    }
    sort_field = sort if sort in sort_fields else "requested_date"
    qs = qs.order_by(sort_field if order == "asc" else f"-{sort_field}", "-id")

    stats = {
        "total": qs.count(),
        "pending": qs.filter(status="Pending").count(),
        "approved": qs.filter(status="Approved").count(),
        "rejected": qs.filter(status="Rejected").count(),
    }

    page_obj, paginator, page_size = _paginate(request, qs, default_page_size=100)

    results = []
    for ccl in page_obj.object_list:
        approver_rows = _ccl_approver_status_rows(ccl)
        results.append(_serialize_ccl_row(
            ccl,
            _ccl_is_locked_from_rows(ccl, approver_rows),
            approver_rows,
            _ccl_is_admin_awarded(ccl, approver_rows),
        ))

    academic_years = list(
        CCL_Application.objects.filter(faculty=faculty)
        .values_list("academic_year", flat=True).distinct().order_by("academic_year")
    )

    return JsonResponse({
        "results": results,
        "stats": stats,
        "filters": {"academic_years": academic_years},
        "pagination": _pagination_meta(page_obj, paginator, page_size),
    })


@_json_forbidden
@check_permission("add_ccl_claim")
def ccl_application_save(request):
    """JSON create/update endpoint for the CCL Application self-service page."""
    if request.method != "POST":
        return JsonResponse({"error": "POST required."}, status=405)

    hidden = _hidden_page_json_or_none(request)
    if hidden:
        return hidden

    ext_user = USER.objects.using("rit_approval_system").filter(
        Employee_id=request.user.Employee_id, is_active=True
    ).first()
    faculty = general_information.objects.filter(
        faculty_id=request.user.Employee_id
    ).select_related("designation").first()

    if not ext_user or not faculty:
        return JsonResponse(
            {"success": False, "message": "User/faculty record not found."}, status=404
        )

    # Use the role of the currently logged-in session user, not ext_user.role_id.
    # ext_user is fetched by Employee_id with .first(), so for an employee that has
    # multiple USER rows (multiple roles) it would pick an arbitrary role — which
    # can build the approver chain configured for the WRONG role (e.g. inserting
    # an extra level that isn't in the hierarchy configured for the role the
    # applicant is actually logged in as). Same fix already applied in
    # leave_application_form / permission_form.
    creator_role_id = request.user.role_id
    data = _json_body(request)

    action = (data.get("action") or "create").strip()
    ccl_id = data.get("ccl_id")
    academic_year = (data.get("academic_year") or "").strip()
    _raw_date = (data.get("date") or "").strip()
    date = parse_date(_raw_date) if _raw_date else None
    reason = (data.get("reason") or "").strip()

    if not academic_year or not date or not reason:
        return JsonResponse(
            {"success": False, "message": "Please fill in all required fields."}, status=400
        )

    try:
        with transaction.atomic():
            fields = _validate_and_build_ccl_submission(faculty, creator_role_id, date, reason)

            if action == "update":
                ccl = get_object_or_404(CCL_Application, id=ccl_id, faculty=faculty)

                if _ccl_is_locked_db(ccl):
                    return JsonResponse(
                        {"success": False, "message": "This CCL application cannot be edited (already processed / not pending)."},
                        status=409,
                    )

                overlap_qs = CCL_Application.objects.filter(
                    faculty=faculty, status__in=["Pending", "Approved"], date=date
                ).exclude(id=ccl.id)
                if overlap_qs.exists():
                    return JsonResponse(
                        {"success": False, "message": "You already have a CCL application for this date."},
                        status=409,
                    )

                ccl.academic_year = academic_year
                ccl.date = date
                ccl.reason = reason
                ccl.designation = faculty.designation
                ccl.status = "Pending"
                ccl.from_time = fields["from_time"]
                ccl.to_time = fields["to_time"]
                ccl.worked_hours = fields["worked_hours"]
                ccl.session = fields["session"]
                ccl.days = fields["days"]
                ccl.is_claimed = False
                ccl.save()

                CCL_Approvers_Data.objects.filter(ccl_application=ccl).delete()
                created_rows = _create_approver_chain_for_ccl(
                    ccl=ccl, creator_role_id=creator_role_id, creator_faculty=faculty
                )

                message = (
                    f"CCL application updated successfully. Approval chain recreated with {created_rows} level(s)."
                    if created_rows else
                    "CCL application updated but no approval chain configured for your role. Please contact administrator."
                )
                target = ccl
            else:
                overlap_qs = CCL_Application.objects.filter(
                    faculty=faculty, status__in=["Pending", "Approved"], date=date
                )
                if overlap_qs.exists():
                    return JsonResponse(
                        {"success": False, "message": "You already have a CCL application for this date."},
                        status=409,
                    )

                new_ccl = CCL_Application.objects.create(
                    user_id=ext_user.id,
                    faculty=faculty,
                    designation=faculty.designation,
                    academic_year=academic_year,
                    date=date,
                    reason=reason,
                    status="Pending",
                    from_time=fields["from_time"],
                    to_time=fields["to_time"],
                    worked_hours=fields["worked_hours"],
                    session=fields["session"],
                    days=fields["days"],
                    is_claimed=False,
                )

                created_rows = _create_approver_chain_for_ccl(
                    ccl=new_ccl, creator_role_id=creator_role_id, creator_faculty=faculty
                )

                message = (
                    f"CCL claim submitted for {fields['worked_hours']} hrs → {fields['days']} day(s) "
                    f"({fields['session']}). Approval chain created with {created_rows} level(s)."
                    if created_rows else
                    "CCL application created but no approval chain configured for your role. Please contact administrator."
                )
                target = new_ccl

            approver_rows = _ccl_approver_status_rows(target)
            row = _serialize_ccl_row(
                target,
                _ccl_is_locked_db(target),
                approver_rows,
                _ccl_is_admin_awarded(target, approver_rows),
            )

    except ValidationError as e:
        return JsonResponse(
            {"success": False, "message": e.messages[0] if e.messages else "Validation error."},
            status=400,
        )
    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"An unexpected error occurred: {str(e)}"}, status=500
        )

    return JsonResponse({"success": True, "message": message, "row": row})


@_json_forbidden
@check_permission("add_ccl_claim")
def ccl_application_delete(request):
    """JSON delete endpoint for the CCL Application self-service page."""
    if request.method != "POST":
        return JsonResponse({"error": "POST required."}, status=405)

    hidden = _hidden_page_json_or_none(request)
    if hidden:
        return hidden

    faculty = general_information.objects.filter(
        faculty_id=request.user.Employee_id
    ).select_related("designation").first()
    if not faculty:
        return JsonResponse({"success": False, "message": "Faculty record not found."}, status=404)

    data = _json_body(request)
    ccl = get_object_or_404(CCL_Application, id=data.get("ccl_id"), faculty=faculty)

    if _ccl_is_locked_db(ccl):
        return JsonResponse(
            {"success": False, "message": "This CCL application cannot be deleted (already processed / not pending)."},
            status=409,
        )

    try:
        with transaction.atomic():
            # Earn flow: a deletable claim is still Pending, so nothing was
            # credited to the balance yet — no refund needed.
            CCL_Approvers_Data.objects.filter(ccl_application=ccl).delete()
            ccl.delete()
    except Exception as e:
        return JsonResponse(
            {"success": False, "message": f"Failed to delete CCL application: {str(e)}"}, status=500
        )

    return JsonResponse({"success": True, "message": "CCL application deleted successfully."})





def _create_approver_chain_for_ccl(ccl, creator_role_id, creator_faculty):
    """
    Reuse LeaveApprovers master table for CCL approval flow.
    Store generated rows in CCL_Approvers_Data only.

    Logic:
    - For HOD-type approvers (same-dept, cross_dept=NO, no dept override):
        uses the creator's own Department_id from rit_approval_system to find
        the correct HOD — not just .first() across all HODs.
    - For cross-dept approvers: uses the configured approver_department.
    - Principal / Director (single-user roles): matched by role_id only.
    """
    approvers_qs = LeaveApprovers.objects.filter(
        creator_role_id=creator_role_id
    ).order_by("approver_level")

    created_rows = 0
    created_levels = set()

    # Get creator's Department_id from rit_approval_system (for HOD matching)
    creator_ext_user = USER.objects.using("rit_approval_system").filter(
        Employee_id=creator_faculty.faculty_id, is_active=True
    ).first()
    creator_dept_id = getattr(creator_ext_user, "Department_id", None) if creator_ext_user else None

    for approver in approvers_qs:
        level = approver.approver_level
        role_id = approver.approver_role_id
        local_dept_id = approver.approver_department_id
        is_cross = (approver.is_cross_department_approver or "NO").upper() == "YES"

        if level in created_levels:
            continue

        approver_filter = {"role_id": role_id, "is_active": True}

        if is_cross and local_dept_id:
            # Cross-dept: use the configured department
            local_dept = Add_Department.objects.filter(id=local_dept_id).first()
            dept_code = getattr(local_dept, "Department_code", None)
            if dept_code:
                ext_dept_id = (
                    Department.objects.using("rit_approval_system")
                    .filter(Department_code=dept_code)
                    .values_list("id", flat=True)
                    .first()
                )
                if ext_dept_id:
                    approver_filter["Department_id"] = ext_dept_id
        elif not is_cross and not local_dept_id and creator_dept_id:
            # Same-dept approver (e.g. HOD): match by creator's department
            approver_filter["Department_id"] = creator_dept_id

        approver_user = USER.objects.using("rit_approval_system").filter(
            **approver_filter
        ).first()

        if not approver_user:
            continue

        approver_faculty = general_information.objects.filter(
            faculty_id=approver_user.Employee_id
        ).select_related("designation").first()

        if not approver_faculty:
            continue

        CCL_Approvers_Data.objects.create(
            ccl_application=ccl,
            approver_id=approver_faculty,
            approver_level=level,
            approver_role_id=role_id,
            creator_role_id=creator_role_id,
            status=CCL_Approvers_Data.Status.PENDING,
            action_date=None,
            remarks=f"Approver: {approver_user.Employee_id}",
        )

        created_rows += 1
        created_levels.add(level)

    return created_rows





from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone

from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils import timezone


def _resolve_ccl_approver(request):
    """Returns (approver_user, approver_faculty); either may be None if not
    resolvable — callers decide how to respond (HTML redirect vs JSON error)."""
    approver_user = USER.objects.using("rit_approval_system").filter(
        Employee_id=request.user.Employee_id, is_active=True
    ).first()
    approver_faculty = general_information.objects.filter(
        faculty_id=request.user.Employee_id
    ).first()
    return approver_user, approver_faculty


def _ccl_role_scope(request, approver_user):
    """Scope to the role the approver is logged in as, while still supporting
    legacy rows created before approver_role_id was populated and cases where
    the local session role and rit_approval_system role do not line up."""
    current_role_ids = {
        rid for rid in (
            getattr(request.user, "role_id", None),
            getattr(approver_user, "role_id", None),
        ) if rid
    }
    if current_role_ids:
        return Q(approver_role_id__in=current_role_ids) | Q(approver_role_id__isnull=True)
    return Q(approver_role_id__isnull=True)


def _ccl_pending_base_qs(approver_faculty, role_scope):
    """CCL_Approvers_Data rows actionable by this approver right now: their own
    Pending row, with no lower-level row for the same application still
    unresolved. A single Exists() subquery replacing the old per-row N+1
    Python loop (same semantics: a row is hidden from "pending" until every
    level below it is Approved)."""
    lower_unresolved = CCL_Approvers_Data.objects.filter(
        ccl_application=OuterRef("ccl_application"),
        approver_level__lt=OuterRef("approver_level"),
    ).exclude(status=CCL_Approvers_Data.Status.APPROVED)

    return (
        CCL_Approvers_Data.objects.filter(
            role_scope,
            approver_id=approver_faculty,
            status=CCL_Approvers_Data.Status.PENDING,
        )
        .annotate(lower_pending=Exists(lower_unresolved))
        .filter(lower_pending=False)
    )


def _ccl_approval_queryset(request, approver_faculty, approver_user, tab):
    """Builds the (unpaginated) CCL_Approvers_Data queryset for one tab, with
    search/level/academic_year filters applied — the single source of truth
    shared by the list, action, and export endpoints so filter/lock logic can
    never drift between them. Returns (queryset, tab_counts, role_scope)."""
    role_scope = _ccl_role_scope(request, approver_user)

    pending_base_qs = _ccl_pending_base_qs(approver_faculty, role_scope)
    approved_base_qs = CCL_Approvers_Data.objects.filter(
        role_scope, approver_id=approver_faculty, status=CCL_Approvers_Data.Status.APPROVED,
    )
    rejected_base_qs = CCL_Approvers_Data.objects.filter(
        role_scope, approver_id=approver_faculty, status=CCL_Approvers_Data.Status.REJECTED,
    )

    # Tab badge counts reflect the base (role-scoped) sets, not the
    # search/level/academic_year filters applied below — matches prior behavior.
    counts = {
        "pending": pending_base_qs.count(),
        "approved": approved_base_qs.count(),
        "rejected": rejected_base_qs.count(),
    }

    if tab == "approved":
        qs = approved_base_qs
    elif tab == "rejected":
        qs = rejected_base_qs
    else:
        qs = pending_base_qs

    qs = qs.select_related(
        "ccl_application", "ccl_application__faculty", "approver_id"
    ).order_by("approver_level", "-ccl_application__requested_date", "-id")

    q = (request.GET.get("q") or "").strip()
    level_filter = (request.GET.get("level") or "").strip()
    academic_year_filter = (request.GET.get("academic_year") or "").strip()

    if q:
        qs = qs.filter(
            Q(ccl_application__date__icontains=q) |
            Q(ccl_application__academic_year__icontains=q) |
            Q(ccl_application__reason__icontains=q) |
            Q(ccl_application__status__icontains=q) |
            Q(ccl_application__faculty__faculty_id__icontains=q) |
            Q(ccl_application__faculty__name__icontains=q)
        )

    if level_filter:
        try:
            qs = qs.filter(approver_level=int(level_filter))
        except ValueError:
            pass

    if academic_year_filter:
        qs = qs.filter(ccl_application__academic_year=academic_year_filter)

    sort = (request.GET.get("sort") or "").strip()
    order = (request.GET.get("order") or "asc").strip().lower()
    sort_map = {
        "date": "ccl_application__date",
        "days": "ccl_application__days",
        "academic_year": "ccl_application__academic_year",
        "status": "ccl_application__status",
    }
    if sort in sort_map:
        field = sort_map[sort]
        qs = qs.order_by(field if order == "asc" else f"-{field}", "approver_level", "-id")

    return qs, counts, role_scope


def _maybe_complete_ccl_application(app):
    """After approving one CCL_Approvers_Data row, flip the parent
    CCL_Application to Approved and credit the balance exactly once, if every
    level for it is now approved. Shared by every approve path so the balance
    can never be credited more than once (or via inconsistent status casing)."""
    if not app:
        return
    all_approved = not CCL_Approvers_Data.objects.filter(
        ccl_application=app
    ).exclude(status=CCL_Approvers_Data.Status.APPROVED).exists()
    if all_approved and app.status != "Approved":
        app.status = "Approved"
        app.save(update_fields=["status"])
        _credit_ccl_on_approval(app)


def _ccl_approval_dropdowns(approver_faculty):
    all_my_rows = CCL_Approvers_Data.objects.filter(approver_id=approver_faculty)
    available_levels = list(
        all_my_rows.values_list("approver_level", flat=True).distinct().order_by("approver_level")
    )
    available_academic_years = list(
        all_my_rows.values_list("ccl_application__academic_year", flat=True)
        .distinct().order_by("ccl_application__academic_year")
    )
    return available_levels, available_academic_years


@no_cache
@check_permission("ccl_approval")
@faculty_leave_management
def ccl_approval(request):
    """CCL Approval page shell.

    Displays hierarchical CCL applications for the current approver. The
    approvals table loads and paginates client-side via ccl_approval_data();
    approve/reject/approve-all go through ccl_approval_action(); exports go
    through ccl_approval_export().
    """
    approver_user, approver_faculty = _resolve_ccl_approver(request)
    if not approver_user:
        messages.error(request, "Approver not found.")
        return redirect("dashboard")
    if not approver_faculty:
        messages.error(request, "Approver faculty record not found.")
        return redirect("dashboard")

    return render(request, "faculty_leave_management/ccl_approval.html", {})


@_json_forbidden
@no_cache
@check_permission("ccl_approval")
@faculty_leave_management
def ccl_approval_data(request):
    """JSON listing endpoint for the CCL Approval page — server-side
    tab/search/level/academic-year filtering, sorting, and pagination."""
    approver_user, approver_faculty = _resolve_ccl_approver(request)
    if not approver_user:
        return JsonResponse({"error": "Approver not found."}, status=404)
    if not approver_faculty:
        return JsonResponse({"error": "Approver faculty record not found."}, status=404)

    tab = (request.GET.get("tab") or "pending").strip().lower()
    if tab not in ("pending", "approved", "rejected"):
        tab = "pending"

    qs, counts, _role_scope = _ccl_approval_queryset(request, approver_faculty, approver_user, tab)
    page_obj, paginator, page_size = _paginate(request, qs)
    available_levels, available_academic_years = _ccl_approval_dropdowns(approver_faculty)

    results = []
    for cad in page_obj.object_list:
        app = cad.ccl_application
        faculty_obj = app.faculty if app else None
        results.append({
            "id": cad.id,
            "level": cad.approver_level,
            "status": cad.status,
            "remarks": cad.remarks or "",
            "ccl_application_id": app.id if app else None,
            "academic_year": app.academic_year if app else "",
            "date": app.date.strftime("%Y-%m-%d") if app and app.date else "",
            "days": float(app.days) if app and app.days is not None else 0.0,
            "reason": app.reason if app else "",
            "requested_date": app.requested_date.strftime("%Y-%m-%d %I:%M %p") if app and app.requested_date else "",
            "faculty_id": getattr(faculty_obj, "faculty_id", None),
            "faculty_name": getattr(faculty_obj, "name", "") or getattr(faculty_obj, "username", ""),
            "designation": str(getattr(faculty_obj, "designation", "") or ""),
        })

    return JsonResponse({
        "results": results,
        "stats": counts,
        "filters": {
            "available_levels": available_levels,
            "available_academic_years": available_academic_years,
        },
        "pagination": _pagination_meta(page_obj, paginator, page_size),
        "tab": tab,
    })


@_json_forbidden
@no_cache
@check_permission("ccl_approval")
@faculty_leave_management
def ccl_approval_action(request):
    """JSON approve/reject/approve-all endpoint for the CCL Approval page.
    Consolidates the old approve/approve_all/reject POST branches of
    ccl_approval and the separate, duplicate bulk_approve_ccl view into one
    place, normalized on the model's title-case Status values."""
    if request.method != "POST":
        return JsonResponse({"error": "POST required."}, status=405)

    approver_user, approver_faculty = _resolve_ccl_approver(request)
    if not approver_user:
        return JsonResponse({"error": "Approver not found."}, status=404)
    if not approver_faculty:
        return JsonResponse({"error": "Approver faculty record not found."}, status=404)

    data = _json_body(request)
    action = (data.get("action") or "").strip()
    selected_ids = data.get("selected_ids") or []
    rejection_reason = (data.get("rejection_reason") or "").strip()

    if action == "approve_all":
        role_scope = _ccl_role_scope(request, approver_user)
        pending_qs = _ccl_pending_base_qs(approver_faculty, role_scope).select_related("ccl_application")
        approved_count = 0
        with transaction.atomic():
            for cad in pending_qs:
                cad.status = CCL_Approvers_Data.Status.APPROVED
                cad.action_date = timezone.now()
                cad.save(update_fields=["status", "action_date"])
                approved_count += 1
                _maybe_complete_ccl_application(cad.ccl_application)
        return JsonResponse({
            "success": True,
            "approved_count": approved_count,
            "message": f"{approved_count} CCL application(s) approved successfully.",
        })

    if not selected_ids:
        return JsonResponse({"success": False, "message": "No CCL applications selected."}, status=400)

    selected_qs = CCL_Approvers_Data.objects.filter(
        id__in=selected_ids, approver_id=approver_faculty
    ).select_related("ccl_application")

    if action == "approve":
        approved_count = 0
        with transaction.atomic():
            for cad in selected_qs:
                if cad.status != CCL_Approvers_Data.Status.PENDING:
                    continue
                cad.status = CCL_Approvers_Data.Status.APPROVED
                cad.action_date = timezone.now()
                cad.save(update_fields=["status", "action_date"])
                approved_count += 1
                _maybe_complete_ccl_application(cad.ccl_application)
        return JsonResponse({
            "success": True,
            "approved_count": approved_count,
            "message": "Selected application(s) approved successfully.",
        })

    if action == "reject":
        if not rejection_reason:
            return JsonResponse({"success": False, "message": "Rejection reason is required."}, status=400)

        rejected_count = 0
        with transaction.atomic():
            for cad in selected_qs:
                if cad.status != CCL_Approvers_Data.Status.PENDING:
                    continue

                app = cad.ccl_application
                app.status = "Rejected"
                app.save(update_fields=["status"])

                cad.status = CCL_Approvers_Data.Status.REJECTED
                cad.remarks = rejection_reason
                cad.action_date = timezone.now()
                cad.save(update_fields=["status", "remarks", "action_date"])
                rejected_count += 1
        return JsonResponse({
            "success": True,
            "rejected_count": rejected_count,
            "message": "Selected application(s) rejected successfully.",
        })

    return JsonResponse({"success": False, "message": "Invalid action."}, status=400)


@no_cache
@check_permission("ccl_approval")
@faculty_leave_management
def ccl_approval_export(request):
    """Excel/PDF export for the CCL Approval page — reflects the full filtered
    (unpaginated) result set for the active tab, not just the current page.
    Kept as a plain GET/download endpoint (not JSON) since it returns a file."""
    approver_user, approver_faculty = _resolve_ccl_approver(request)
    if not approver_user or not approver_faculty:
        messages.error(request, "Approver not found.")
        return redirect("ccl_approval")

    tab = (request.GET.get("tab") or "pending").strip().lower()
    if tab not in ("pending", "approved", "rejected"):
        tab = "pending"

    qs, _counts, _role_scope = _ccl_approval_queryset(request, approver_faculty, approver_user, tab)

    if (request.GET.get("format") or "").strip().lower() == "pdf":
        return export_ccl_approvals_to_pdf(qs, tab)
    return export_ccl_approvals_to_excel(qs, tab)





from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment





def export_ccl_approvals_to_excel(approvals_qs, tab="pending"):

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CCL Report"

    headers = [
        "Approver Level",
        "Faculty ID",
        "Faculty Name",
        "Department",
        "Designation",
        "Academic Year",
        "CCL Date",
        "Days",
        "Reason",
        "Requested On",
        "Status",
        "Is Claimed",
        "Claimed (Year)",
        "Used",
        "Remaining",
        "Total Claimed",
    ]

    if tab == "rejected":
        headers.append("Remarks")

    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_num = 2

    for approval in approvals_qs:

        app = approval.ccl_application
        faculty = app.faculty if app else None

        # ---------------- Faculty Info ----------------
        faculty_id = getattr(faculty, "faculty_id", "")
        faculty_name = getattr(faculty, "name", "")

        department = ""
        try:
            department = str(faculty.department) if faculty and faculty.department else ""
        except:
            department = ""

        designation = ""
        try:
            designation = str(app.designation) if app and app.designation else ""
        except:
            designation = ""

        # ---------------- Claim Info ----------------
        claim = None
        if faculty and app:
            claim = CCL_Claim.objects.filter(
                faculty=faculty,
                academic_year=app.academic_year
            ).first()

        is_claimed = "Yes" if getattr(app, "is_claimed", False) else "No"

        claimed = claim.claimed if claim else 0
        used = claim.used if claim else 0
        remaining = claim.remaining if claim else 0
        total_claimed = claim.total_claimed if claim else 0

        # ---------------- Row ----------------
        row = [
            approval.approver_level,
            faculty_id,
            faculty_name,
            department,
            designation,
            app.academic_year if app else "",
            app.date.strftime("%d-%m-%Y") if app and app.date else "",
            getattr(app, "days", ""),
            app.reason if app else "",
            app.requested_date.strftime("%d-%m-%Y %H:%M") if app and app.requested_date else "",
            approval.status,
            is_claimed,
            claimed,
            used,
            remaining,
            total_claimed,
        ]

        if tab == "rejected":
            row.append(approval.remarks or "")

        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        row_num += 1

    # Auto width
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    filename = f"ccl_{tab}_report.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'inline; filename="{filename}"'

    workbook.save(response)
    return response


from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from django.contrib.staticfiles import finders
from django.conf import settings
import os


# from .models import CCL_Claim


def header(canvas, doc):
    canvas.saveState()
    canvas.setTitle(getattr(doc, "title", "CCL Report"))
    page_w, page_h = landscape(A4)
    left = doc.leftMargin
    right = page_w - doc.rightMargin
    center_x = page_w / 2
    
    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    ACCENT_RED = colors.HexColor("#b91c1c")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    logo_rel = "images/ritlogo.png"
    logo_path = finders.find(logo_rel)

    if not logo_path:
        static_root = getattr(settings, "STATIC_ROOT", "")
        if static_root:
            candidate = os.path.join(static_root, logo_rel)
            if os.path.exists(candidate):
                logo_path = candidate

    if logo_path and os.path.exists(logo_path):
        try:
            canvas.drawImage(
                ImageReader(logo_path),
                left,
                page_h - 24 * mm,
                width=30 * mm,
                height=18 * mm,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass

    canvas.setFillColor(PRIMARY_BLUE)
    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawCentredString(center_x, page_h - 8 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

    canvas.setFillColor(ACCENT_RED)
    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawCentredString(center_x, page_h - 14 * mm, "An Autonomous Institution")

    canvas.setFillColor(MEDIUM_GRAY)
    canvas.setFont("Helvetica", 8.2)
    canvas.drawCentredString(center_x, page_h - 19 * mm, "Approved by AICTE, New Delhi")
    canvas.setFont("Helvetica", 8)
    canvas.drawCentredString(center_x, page_h - 23.5 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
    canvas.drawCentredString(center_x, page_h - 28 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

    canvas.setStrokeColor(BORDER_GRAY)
    canvas.setLineWidth(0.8)
    canvas.line(left, page_h - 31.5 * mm, right, page_h - 31.5 * mm)

    footer_y = 8 * mm
    canvas.line(left, footer_y + 7 * mm, right, footer_y + 7 * mm)

    local_time = timezone.localtime()
    canvas.setFillColor(LIGHT_GRAY)
    canvas.setFont("Helvetica", 8)
    canvas.drawString(left, footer_y, f"Generated: {local_time.strftime('%d %b %Y, %I:%M %p')}")
    canvas.drawCentredString(center_x, footer_y, "CCL Approval Report")
    canvas.drawRightString(right, footer_y, f"Page {doc.page}")

    canvas.restoreState()


def export_ccl_approvals_to_pdf(approvals_qs, tab="pending"):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="ccl_{tab}_report.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=36 * mm,
        bottomMargin=16 * mm,
    )
    doc.title = f"CCL {tab.capitalize()} Report"
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=15,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f2f57"),
        spaceAfter=4,
    )

    sub_title_style = ParagraphStyle(
        "SubTitleStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#4b5563"),
        spaceAfter=10,
    )

    dept_band_style = ParagraphStyle(
        "DeptBandStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11,
        alignment=TA_LEFT,
        textColor=colors.white,
        backColor=colors.HexColor("#1F4E78"),
        borderPadding=6,
        leading=14,
        spaceBefore=8,
        spaceAfter=2,
    )

    year_style = ParagraphStyle(
        "YearStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        textColor=colors.HexColor("#163A5F"),
        spaceBefore=12,
        spaceAfter=5,
    )

    cell_style = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.black,
        alignment=TA_LEFT,
    )

    cell_style_center = ParagraphStyle(
        "CellStyleCenter",
        parent=cell_style,
        alignment=TA_CENTER,
    )

    header_left_style = ParagraphStyle(
        "HeaderLeftStyle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.5,
        textColor=colors.HexColor("#163A5F"),
        alignment=TA_LEFT,
    )

    header_center_style = ParagraphStyle(
        "HeaderCenterStyle",
        parent=header_left_style,
        alignment=TA_CENTER,
    )

    local_time = timezone.localtime()

    elements.append(Paragraph(f"CCL {tab.capitalize()} Report", title_style))
    elements.append(Paragraph(f"Generated on {local_time.strftime('%d-%m-%Y %I:%M %p')}", sub_title_style))
    elements.append(Spacer(1, 4))

    if not approvals_qs.exists():
        elements.append(Paragraph("No records found.", styles["Normal"]))
        doc.build(elements, onFirstPage=header, onLaterPages=header)
        return response

    grouped = defaultdict(lambda: defaultdict(list))

    for approval in approvals_qs:
        app = getattr(approval, "ccl_application", None)
        faculty = getattr(app, "faculty", None) if app else None

        if faculty and getattr(faculty, "department", None):
            dept = faculty.department
            degree = getattr(dept, "degree", None)
            dept_text = str(dept)
            degree_code = getattr(degree, "degree_code", "") if degree else ""
        else:
            dept_text = "No Department"
            degree_code = ""

        academic_year = str(getattr(app, "academic_year", "No Academic Year")) if app else "No Academic Year"
        grouped[(dept_text, degree_code)][academic_year].append(approval)

    for (dept_text, degree_code) in sorted(grouped.keys(), key=lambda x: (x[0].lower(), x[1].lower())):
        dept_years = grouped[(dept_text, degree_code)]
        display_dept = f"{degree_code} - {dept_text}" if degree_code else dept_text

        elements.append(
            Paragraph(
                f"""
                <para alignment="center">
                    <b>{degree_code}</b> &nbsp;&nbsp; | &nbsp;&nbsp; {dept_text}
                </para>
                """ if degree_code else f"""
                <para alignment="center">
                    {dept_text}
                </para>
                """,
                dept_band_style
            )
        )

        for academic_year in sorted(dept_years.keys()):
            year_records = dept_years[academic_year]

            elements.append(
                Paragraph(
                    f"Academic Year : {academic_year}   |   Records : {len(year_records)}",
                    year_style
                )
            )

            data = [[
                Paragraph("Faculty ID", header_left_style),
                Paragraph("Faculty Name", header_left_style),
                Paragraph("Designation", header_left_style),
                Paragraph("Date", header_center_style),
                Paragraph("Days", header_center_style),
                Paragraph("Status", header_center_style),
                Paragraph("Claimed", header_center_style),
                Paragraph("Used", header_center_style),
                Paragraph("Remaining", header_center_style),
            ]]

            total_days = 0
            total_claimed = 0
            total_used = 0
            total_remaining = 0

            for approval in year_records:
                app = getattr(approval, "ccl_application", None)
                faculty = getattr(app, "faculty", None) if app else None

                claim = None
                if faculty and app:
                    claim = CCL_Claim.objects.filter(
                        faculty=faculty,
                        academic_year=app.academic_year
                    ).first()

                days_val = getattr(app, "days", 0) if app else 0
                claimed_val = getattr(claim, "claimed", 0) if claim else 0
                used_val = getattr(claim, "used", 0) if claim else 0
                remaining_val = getattr(claim, "remaining", 0) if claim else 0

                app_date = getattr(app, "date", None)
                app_date_str = app_date.strftime("%d-%m-%Y") if app_date else "-"

                try:
                    total_days += float(days_val or 0)
                except Exception:
                    pass

                try:
                    total_claimed += float(claimed_val or 0)
                except Exception:
                    pass

                try:
                    total_used += float(used_val or 0)
                except Exception:
                    pass

                try:
                    total_remaining += float(remaining_val or 0)
                except Exception:
                    pass

                data.append([
                    Paragraph(str(getattr(faculty, "faculty_id", "")), cell_style),
                    Paragraph(str(getattr(faculty, "name", "")), cell_style),
                    Paragraph(str(getattr(app, "designation", "")), cell_style),
                    Paragraph(app_date_str, cell_style_center),
                    Paragraph(str(days_val), cell_style_center),
                    Paragraph(str(getattr(approval, "status", "")), cell_style_center),
                    Paragraph(str(claimed_val), cell_style_center),
                    Paragraph(str(used_val), cell_style_center),
                    Paragraph(str(remaining_val), cell_style_center),
                ])

            data.append([
                Paragraph("<b>Total</b>", cell_style),
                Paragraph("", cell_style),
                Paragraph("", cell_style),
                Paragraph("", cell_style_center),
                Paragraph(f"<b>{total_days:g}</b>", cell_style_center),
                Paragraph("", cell_style_center),
                Paragraph(f"<b>{total_claimed:g}</b>", cell_style_center),
                Paragraph(f"<b>{total_used:g}</b>", cell_style_center),
                Paragraph(f"<b>{total_remaining:g}</b>", cell_style_center),
            ])

            col_widths = [
                24 * mm,  # Faculty ID
                48 * mm,  # Faculty Name
                38 * mm,  # Designation
                24 * mm,  # Date
                16 * mm,  # Days
                22 * mm,  # Status
                20 * mm,  # Claimed
                18 * mm,  # Used
                22 * mm,  # Remaining
            ]

            table = Table(data, colWidths=col_widths, repeatRows=1)

            table_style = TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DCEAF7")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#163A5F")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8.5),

                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#222222")),

                ("ALIGN", (0, 0), (2, -1), "LEFT"),
                ("ALIGN", (3, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, 0), 8),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),

                ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#9DB9D5")),
                ("LINEABOVE", (0, -1), (-1, -1), 1, colors.HexColor("#7A9CBD")),
                ("LINEBELOW", (0, -1), (-1, -1), 1, colors.HexColor("#7A9CBD")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#B8C7D6")),
                ("INNERGRID", (0, 1), (-1, -2), 0.35, colors.HexColor("#D5DEE8")),
            ])

            for row_idx in range(1, len(data) - 1):
                if row_idx % 2 == 1:
                    table_style.add("BACKGROUND", (0, row_idx), (-1, row_idx), colors.white)
                else:
                    table_style.add("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#F7FBFF"))

            table_style.add("BACKGROUND", (0, len(data) - 1), (-1, len(data) - 1), colors.HexColor("#EEF4FA"))
            table_style.add("FONTNAME", (0, len(data) - 1), (-1, len(data) - 1), "Helvetica-Bold")
            table_style.add("TEXTCOLOR", (0, len(data) - 1), (-1, len(data) - 1), colors.HexColor("#163A5F"))

            table.setStyle(table_style)

            elements.append(table)
            elements.append(Spacer(1, 8))

    doc.build(elements, onFirstPage=header, onLaterPages=header)
    return response












@login_required
@is_super_user('faculty_leave_management')
def regenerate_ccl_approvers(request):
    """
    Utility view to regenerate CCL_Approvers_Data for existing CCL applications
    that don't have approver records.
    """
    from django.contrib import messages
    from django.shortcuts import redirect
    
    regenerated_count = 0
    error_count = 0
    
    # Get all CCL applications that don't have any approver records
    ccl_applications = CCL_Application.objects.filter(
        status__in=['PENDING', 'APPROVED']  # Only process pending or approved applications
    ).select_related('faculty')
    
    for ccl in ccl_applications:
        # Check if this CCL already has approver records
        existing_approvers = CCL_Approvers_Data.objects.filter(ccl_application=ccl).count()
        
        if existing_approvers > 0:
            continue  # Skip if approvers already exist
        
        try:
            # Get creator's role
            creator_faculty = ccl.faculty
            creator_ext_user = USER.objects.using("rit_approval_system").filter(
                Employee_id=creator_faculty.faculty_id,
                is_active=True
            ).first()
            
            if not creator_ext_user:
                error_count += 1
                continue
            
            creator_role_id = creator_ext_user.role_id
            
            # Create approver chain
            created = _create_approver_chain_for_ccl(ccl, creator_role_id, creator_faculty)
            
            if created > 0:
                regenerated_count += 1
        
        except Exception as e:
            error_count += 1
            print(f"Error regenerating approvers for CCL {ccl.id}: {str(e)}")
            continue
    
    if regenerated_count > 0:
        messages.success(
            request,
            f"Successfully regenerated approver records for {regenerated_count} CCL applications."
        )
    
    if error_count > 0:
        messages.warning(
            request,
            f"Failed to regenerate approvers for {error_count} CCL applications."
        )
    
    if regenerated_count == 0 and error_count == 0:
        messages.info(request, "All CCL applications already have approver records.")
    
    return redirect('add_ccl_claim')



@login_required
@is_super_user('faculty_leave_management')
def debug_ccl_approvals(request):
    """
    Debug view to check CCL applications and approver data
    """
    from django.http import HttpResponse
    
    output = []
    output.append("<html><head><style>")
    output.append("body { font-family: monospace; padding: 20px; }")
    output.append("table { border-collapse: collapse; margin: 20px 0; }")
    output.append("th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }")
    output.append("th { background-color: #4CAF50; color: white; }")
    output.append("h2 { color: #333; margin-top: 30px; }")
    output.append(".error { color: red; }")
    output.append(".success { color: green; }")
    output.append("</style></head><body>")
    
    output.append("<h1>CCL Approval Debug Information</h1>")
    
    # 1. Check CCL Applications
    output.append("<h2>1. CCL Applications (PENDING status)</h2>")
    ccl_apps = CCL_Application.objects.filter(status='PENDING').select_related('faculty')
    output.append(f"<p>Total PENDING CCL Applications: <strong>{ccl_apps.count()}</strong></p>")
    
    if ccl_apps.exists():
        output.append("<table>")
        output.append("<tr><th>ID</th><th>Faculty ID</th><th>Faculty Name</th><th>Date</th><th>Academic Year</th><th>Status</th></tr>")
        for ccl in ccl_apps[:10]:  # Show first 10
            output.append(f"<tr>")
            output.append(f"<td>{ccl.id}</td>")
            output.append(f"<td>{ccl.faculty.faculty_id if ccl.faculty else 'N/A'}</td>")
            output.append(f"<td>{ccl.faculty.name if ccl.faculty else 'N/A'}</td>")
            output.append(f"<td>{ccl.date}</td>")
            output.append(f"<td>{ccl.academic_year}</td>")
            output.append(f"<td>{ccl.status}</td>")
            output.append(f"</tr>")
        output.append("</table>")
        if ccl_apps.count() > 10:
            output.append(f"<p><em>Showing first 10 of {ccl_apps.count()} applications</em></p>")
    
    # 2. Check CCL_Approvers_Data
    output.append("<h2>2. CCL Approvers Data</h2>")
    approvers_data = CCL_Approvers_Data.objects.all().select_related('ccl_application', 'approver_id')
    output.append(f"<p>Total CCL_Approvers_Data records: <strong>{approvers_data.count()}</strong></p>")
    
    if approvers_data.exists():
        output.append("<table>")
        output.append("<tr><th>ID</th><th>CCL App ID</th><th>Approver ID</th><th>Approver Name</th><th>Level</th><th>Status</th></tr>")
        for ad in approvers_data[:10]:
            output.append(f"<tr>")
            output.append(f"<td>{ad.id}</td>")
            output.append(f"<td>{ad.ccl_application.id if ad.ccl_application else 'N/A'}</td>")
            output.append(f"<td>{ad.approver_id.faculty_id if ad.approver_id else 'N/A'}</td>")
            output.append(f"<td>{ad.approver_id.name if ad.approver_id else 'N/A'}</td>")
            output.append(f"<td>{ad.approver_level}</td>")
            output.append(f"<td>{ad.status}</td>")
            output.append(f"</tr>")
        output.append("</table>")
    
    # 3. Check LeaveApprovers configuration
    output.append("<h2>3. LeaveApprovers Configuration</h2>")
    leave_approvers = LeaveApprovers.objects.all()
    output.append(f"<p>Total LeaveApprovers records: <strong>{leave_approvers.count()}</strong></p>")
    
    if leave_approvers.exists():
        output.append("<table>")
        output.append("<tr><th>ID</th><th>Creator Role ID</th><th>Approver Role ID</th><th>Level</th><th>Dept ID</th><th>Cross Dept</th></tr>")
        for la in leave_approvers:
            output.append(f"<tr>")
            output.append(f"<td>{la.id}</td>")
            output.append(f"<td>{la.creator_role_id}</td>")
            output.append(f"<td>{la.approver_role_id}</td>")
            output.append(f"<td>{la.approver_level}</td>")
            output.append(f"<td>{la.approver_department_id or 'N/A'}</td>")
            output.append(f"<td>{la.is_cross_department_approver or 'NO'}</td>")
            output.append(f"</tr>")
        output.append("</table>")
    else:
        output.append("<p class='error'>⚠️ No LeaveApprovers configuration found! This is the problem.</p>")
    
    # 4. Check current user (HOD)
    output.append("<h2>4. Current User Information</h2>")
    output.append(f"<p>Logged in as: <strong>{request.user.Employee_id}</strong></p>")
    
    ext_user = USER.objects.using("rit_approval_system").filter(
        Employee_id=request.user.Employee_id,
        is_active=True
    ).first()
    
    if ext_user:
        output.append(f"<p>Role ID: <strong>{ext_user.role_id}</strong></p>")
        output.append(f"<p>Department ID: <strong>{ext_user.Department_id}</strong></p>")
    else:
        output.append("<p class='error'>⚠️ User not found in rit_approval_system database</p>")
    
    # 5. Check which CCL apps should be visible to current user
    if ext_user:
        output.append("<h2>5. CCL Applications That Should Be Visible to Current User</h2>")
        
        # Find hierarchies where current user is an approver
        related_creator_roles = set()
        for la in leave_approvers:
            if la.approver_role_id == ext_user.role_id:
                related_creator_roles.add(la.creator_role_id)
        
        output.append(f"<p>Current user is approver for creator roles: <strong>{list(related_creator_roles)}</strong></p>")
        
        if related_creator_roles:
            # Find pending approver data for current user
            pending_for_user = CCL_Approvers_Data.objects.filter(
                approver_id__faculty_id=request.user.Employee_id,
                status=CCL_Approvers_Data.Status.PENDING
            ).select_related('ccl_application', 'ccl_application__faculty')
            
            output.append(f"<p>Pending approvals for current user: <strong>{pending_for_user.count()}</strong></p>")
            
            if pending_for_user.exists():
                output.append("<table>")
                output.append("<tr><th>CCL App ID</th><th>Faculty</th><th>Level</th><th>Status</th></tr>")
                for ad in pending_for_user:
                    output.append(f"<tr>")
                    output.append(f"<td>{ad.ccl_application.id}</td>")
                    output.append(f"<td>{ad.ccl_application.faculty.name if ad.ccl_application.faculty else 'N/A'}</td>")
                    output.append(f"<td>{ad.approver_level}</td>")
                    output.append(f"<td>{ad.status}</td>")
                    output.append(f"</tr>")
                output.append("</table>")
    
    # 6. Check for CCL apps without approvers
    output.append("<h2>6. CCL Applications Without Approvers</h2>")
    ccl_without_approvers = []
    for ccl in ccl_apps:
        approver_count = CCL_Approvers_Data.objects.filter(ccl_application=ccl).count()
        if approver_count == 0:
            ccl_without_approvers.append(ccl)
    
    output.append(f"<p>CCL Applications without any approvers: <strong class='error'>{len(ccl_without_approvers)}</strong></p>")
    
    if ccl_without_approvers:
        output.append("<table>")
        output.append("<tr><th>CCL ID</th><th>Faculty ID</th><th>Faculty Name</th><th>Date</th></tr>")
        for ccl in ccl_without_approvers[:10]:
            output.append(f"<tr>")
            output.append(f"<td>{ccl.id}</td>")
            output.append(f"<td>{ccl.faculty.faculty_id if ccl.faculty else 'N/A'}</td>")
            output.append(f"<td>{ccl.faculty.name if ccl.faculty else 'N/A'}</td>")
            output.append(f"<td>{ccl.date}</td>")
            output.append(f"</tr>")
        output.append("</table>")
        
        output.append("<p class='error'>⚠️ These applications need approver records created!</p>")
        output.append(f"<p><a href='/faculty_leave_management/faculty/leave/regenerate_ccl_approvers/'>Click here to regenerate approvers</a></p>")
    
    output.append("</body></html>")
    
    return HttpResponse("\n".join(output))



# bulk_approve_ccl was retired: its "approve every actionable pending row"
# behavior (including the hierarchy/role-scope check and the balance credit)
# is now handled by ccl_approval_action(request) with action="approve_all",
# which also fixed a pre-existing status-casing bug (this view used to write
# ccl_app.status = 'APPROVED', while every other approve path wrote "Approved").


@check_permission("employee_leave_dashboard")
def employee_leave_dashboard(request):
    from django.core.paginator import Paginator
    from django.db.models import Count
    from faculty_management.models import (
        DesignationMaster,
        FacultyCategory,
        Faculty_Data_Permission,
        general_information as FacultyInfo,
    )

    request_data = request.GET
    q = (request_data.get("q") or "").strip()
    academic_year_filter = (request_data.get("academic_year") or "").strip()
    department_filter = (request_data.get("department") or "").strip()
    designation_filter = (request_data.get("designation") or "").strip()
    category_filter = (request_data.get("category") or "").strip()
    leave_type_filter = (request_data.get("leave_type") or "").strip()
    leave_kind_filter = (request_data.get("leave_kind") or "").strip()
    status_filter = (request_data.get("status") or "").strip()
    from_date_filter = (request_data.get("from_date") or "").strip()
    to_date_filter = (request_data.get("to_date") or "").strip()
    page_number = request_data.get("page", 1)

    try:
        per_page = int(request_data.get("per_page") or 50)
    except (TypeError, ValueError):
        per_page = 50
    if per_page not in [25, 50, 100]:
        per_page = 50

    role_id = getattr(request.user, "role_id", None)
    data_permission = Faculty_Data_Permission.objects.filter(role_id=role_id).first()
    user_faculty = FacultyInfo.objects.filter(
        faculty_id=getattr(request.user, "Employee_id", None)
    ).select_related("department").first()
    user_dept_id = user_faculty.department_id if user_faculty else None

    permitted_qs = LeaveApplication.objects.select_related(
        "faculty",
        "faculty__department",
        "faculty__designation",
        "faculty__category",
        "designation",
        "leave_type",
        "session",
    )

    permission_note = "No faculty data permission is assigned for this role."
    can_view_all_faculty_data = bool(getattr(request.user, "is_superuser", False))
    can_view_department_faculty_data = False

    if getattr(request.user, "is_superuser", False) or (
        data_permission and data_permission.can_view_all_faculty_data
    ):
        can_view_all_faculty_data = True
        permission_note = "Showing leave records for all departments."
    elif data_permission and data_permission.can_view_department_faculty_data:
        can_view_department_faculty_data = True
        if user_dept_id:
            permitted_qs = permitted_qs.filter(faculty__department_id=user_dept_id)
            department_filter = str(user_dept_id)
            permission_note = "Showing leave records for your department."
        else:
            permitted_qs = permitted_qs.none()
            permission_note = "Your faculty record has no department mapped."
    else:
        permitted_qs = permitted_qs.none()

    available_academic_years = (
        permitted_qs.exclude(academic_year__isnull=True)
        .exclude(academic_year="")
        .values_list("academic_year", flat=True)
        .distinct()
        .order_by("-academic_year")
    )

    leave_qs = permitted_qs

    if academic_year_filter:
        leave_qs = leave_qs.filter(academic_year=academic_year_filter)

    if department_filter:
        if can_view_all_faculty_data or str(department_filter) == str(user_dept_id):
            leave_qs = leave_qs.filter(faculty__department_id=department_filter)

    if designation_filter:
        leave_qs = leave_qs.filter(
            Q(faculty__designation_id=designation_filter) |
            Q(designation_id=designation_filter)
        )

    if category_filter:
        leave_qs = leave_qs.filter(faculty__category_id=category_filter)

    if leave_type_filter:
        leave_qs = leave_qs.filter(leave_type_id=leave_type_filter)

    if leave_kind_filter == "leave":
        leave_qs = leave_qs.filter(leave_type__is_leave=True)
    elif leave_kind_filter == "permission":
        leave_qs = leave_qs.filter(leave_type__is_leave=False)

    if status_filter:
        leave_qs = leave_qs.filter(status__iexact=status_filter)

    parsed_from_date = parse_date(from_date_filter) if from_date_filter else None
    parsed_to_date = parse_date(to_date_filter) if to_date_filter else None
    if parsed_from_date:
        leave_qs = leave_qs.filter(to_date__gte=parsed_from_date)
    if parsed_to_date:
        leave_qs = leave_qs.filter(from_date__lte=parsed_to_date)

    if q:
        leave_qs = leave_qs.filter(
            Q(faculty__faculty_id__icontains=q) |
            Q(faculty__name__icontains=q) |
            Q(faculty__department__Department__icontains=q) |
            Q(faculty__designation__designation_name__icontains=q) |
            Q(faculty__category__category_name__icontains=q) |
            Q(leave_type__name__icontains=q) |
            Q(leave_type__code__icontains=q) |
            Q(academic_year__icontains=q) |
            Q(reason__icontains=q) |
            Q(status__icontains=q)
        )

    leave_qs = leave_qs.order_by("-from_date", "-requested_date", "faculty__name")

    total_applications = leave_qs.count()
    total_days = sum(float(app.days or 0) for app in leave_qs)
    approved_count = leave_qs.filter(status__iexact="Approved").count()
    pending_count = leave_qs.filter(status__iexact="Pending").count()
    rejected_count = leave_qs.filter(status__iexact="Rejected").count()
    pre_approved_count = leave_qs.filter(status__iexact="Pre-approved").count()

    today_date = timezone.localdate()
    today_leave_qs = leave_qs.filter(
        from_date__lte=today_date,
        to_date__gte=today_date,
    ).order_by("faculty__department__Department", "faculty__name", "from_date")
    today_total = today_leave_qs.count()
    today_approved_count = today_leave_qs.filter(status__iexact="Approved").count()
    today_pending_count = today_leave_qs.filter(status__iexact="Pending").count()
    today_rejected_count = today_leave_qs.filter(status__iexact="Rejected").count()
    today_pre_approved_count = today_leave_qs.filter(status__iexact="Pre-approved").count()
    today_permission_count = today_leave_qs.filter(leave_type__is_leave=False).count()
    today_leave_records = list(today_leave_qs[:12])

    leave_type_summary = list(
        leave_qs.values("leave_type__name", "leave_type__code")
        .annotate(total=Count("id"))
        .order_by("-total", "leave_type__name")[:10]
    )
    academic_year_summary = list(
        leave_qs.values("academic_year")
        .annotate(total=Count("id"))
        .order_by("-academic_year")
    )

    paginator = Paginator(leave_qs, per_page)
    page_obj = paginator.get_page(page_number)
    page_records = list(page_obj.object_list)

    hierarchy_record_ids = {record.id for record in page_records}
    hierarchy_record_ids.update(record.id for record in today_leave_records)
    approvals_qs = (
        LeaveApproversData.objects.filter(leave_application_id__in=hierarchy_record_ids)
        .select_related("approver_id", "approver_id__designation")
        .order_by("approver_level", "id")
    )
    approver_status_map = {}
    for row in approvals_qs:
        approver = row.approver_id
        approver_emp = getattr(approver, "faculty_id", "-")
        approver_name = getattr(approver, "name", "") or getattr(approver, "username", "")
        approver_desig = (
            getattr(getattr(approver, "designation", None), "designation_name", None)
            or str(getattr(approver, "designation", "") or "")
        )
        approver_status_map.setdefault(row.leave_application_id, []).append({
            "level": row.approver_level,
            "emp": approver_emp,
            "name": approver_name,
            "desig": approver_desig,
            "status": row.status,
        })

    departments = Add_Department.objects.all().order_by("Department")
    if not can_view_all_faculty_data and user_dept_id:
        departments = departments.filter(id=user_dept_id)

    query_params = request.GET.copy()
    query_params.pop("page", None)

    context = {
        "page_obj": page_obj,
        "leave_records": page_records,
        "approver_status_map": approver_status_map,
        "available_academic_years": available_academic_years,
        "departments": departments,
        "designations": DesignationMaster.objects.all().order_by("designation_name"),
        "categories": FacultyCategory.objects.filter(is_active=True).order_by("category_name"),
        "leave_types": LeaveType.objects.filter(is_active=True).order_by("name"),
        "status_choices": LeaveApplication.STATUS_CHOICES,
        "filters": {
            "q": q,
            "academic_year": academic_year_filter,
            "department": department_filter,
            "designation": designation_filter,
            "category": category_filter,
            "leave_type": leave_type_filter,
            "leave_kind": leave_kind_filter,
            "status": status_filter,
            "from_date": from_date_filter,
            "to_date": to_date_filter,
            "per_page": per_page,
        },
        "total_applications": total_applications,
        "total_days": total_days,
        "approved_count": approved_count,
        "pending_count": pending_count,
        "rejected_count": rejected_count,
        "pre_approved_count": pre_approved_count,
        "today_date": today_date,
        "today_total": today_total,
        "today_approved_count": today_approved_count,
        "today_pending_count": today_pending_count,
        "today_rejected_count": today_rejected_count,
        "today_pre_approved_count": today_pre_approved_count,
        "today_permission_count": today_permission_count,
        "today_leave_records": today_leave_records,
        "leave_type_summary": leave_type_summary,
        "academic_year_summary": academic_year_summary,
        "permission_note": permission_note,
        "can_view_all_faculty_data": can_view_all_faculty_data,
        "can_view_department_faculty_data": can_view_department_faculty_data,
        "query_string": query_params.urlencode(),
    }
    return render(request, "faculty_leave_management/employee_leave_dashboard.html", context)

def _college_data_filter_summary(context):
    parts = []
    filters = context.get("filters", {})
    dept_id = (filters.get("department") or "").strip()
    if dept_id:
        for department in context.get("filter_departments", []):
            if str(department.id) == dept_id:
                parts.append(f"Department: {department.Department}")
                break
    year = (filters.get("year") or "").strip()
    if year:
        year_label = dict(context.get("year_headers", [])).get(year, year)
        parts.append(f"Year: {year_label}")
    gender = (filters.get("gender") or "").strip()
    if gender:
        parts.append(f"Gender: {gender.capitalize()}")
    return " | ".join(parts) if parts else "All Departments"


def export_college_data_to_excel(context):
    """Single-sheet, print-friendly black & white workbook (attendance-report style)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    workbook = Workbook()
    ws = workbook.active
    ws.title = "College Data"

    thin = Side(style="thin", color="000000")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hcell(r, c, val, sz=9):
        cell = ws.cell(r, c, val)
        cell.font = Font(bold=True, size=sz, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr
        return cell

    def dcell(r, c, val, sz=9, center=True, bold=False):
        cell = ws.cell(r, c, val)
        cell.font = Font(size=sz, bold=bold, color="000000")
        cell.alignment = Alignment(
            horizontal="center" if center else "left", vertical="center"
        )
        cell.border = bdr
        return cell

    year_headers = context.get("year_headers", [])
    designations = context.get("teaching_designations", [])
    student_cols = 1 + len(year_headers) * 3 + 4
    teaching_cols = 1 + len(designations) * 3 + 3 + 1
    staff_cols = 5
    total_cols = max(student_cols, teaching_cols, staff_cols)

    # ══════════════════════════════════════════
    # INSTITUTIONAL HEADER ROWS
    # ══════════════════════════════════════════
    def banner(r, text, sz, bold=True, align="center", height=None):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=total_cols)
        cell = ws.cell(r, 1, text)
        cell.font = Font(bold=bold, size=sz, color="000000")
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if height:
            ws.row_dimensions[r].height = height

    banner(1, "RAMCO INSTITUTE OF TECHNOLOGY", 14, height=22)
    banner(2, "An Autonomous Institution", 10, height=16)
    banner(3, "Rajapalayam", 11, height=18)
    banner(4, "College Data Report", 11, height=18)
    ws.row_dimensions[5].height = 6

    generated_time = timezone.localtime(context.get("generated_at")).strftime("%d-%m-%Y %I:%M %p")
    banner(6, f"Generated on: {generated_time}", 9, bold=False, align="left", height=15)
    banner(7, f"Filter: {_college_data_filter_summary(context)}", 9, bold=False, align="left", height=15)
    ws.row_dimensions[8].height = 6

    def section_title(r, text, span):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
        cell = ws.cell(r, 1, text)
        cell.font = Font(bold=True, size=11, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bdr
        ws.row_dimensions[r].height = 18
        return r + 1

    def mfns_group_header(top, start_col, label):
        """Write a 'Label' spanning three columns over Male/Female/Not Set."""
        ws.merge_cells(start_row=top, start_column=start_col, end_row=top, end_column=start_col + 2)
        hcell(top, start_col, label)
        hcell(top + 1, start_col, "Male")
        hcell(top + 1, start_col + 1, "Female")
        hcell(top + 1, start_col + 2, "Not Set")

    r = 9

    # ---------------- Students ----------------
    r = section_title(r, "Current Student Strength", student_cols)
    top, bottom = r, r + 1
    ws.merge_cells(start_row=top, start_column=1, end_row=bottom, end_column=1)
    hcell(top, 1, "Department")
    col = 2
    for _, label in year_headers:
        mfns_group_header(top, col, label)
        col += 3
    for label in ("Total", "Total Male", "Total Female", "Not Set"):
        ws.merge_cells(start_row=top, start_column=col, end_row=bottom, end_column=col)
        hcell(top, col, label)
        col += 1
    r = bottom + 1
    for row in context.get("student_rows", []):
        dcell(r, 1, row["department"], center=False)
        c = 2
        for year in row["years"]:
            for key in ("male", "female", "other"):
                dcell(r, c, year.get(key, 0))
                c += 1
        for key in ("total", "male", "female", "other"):
            dcell(r, c, row["total"].get(key, 0))
            c += 1
        r += 1
    dcell(r, 1, "Sub Total", center=False, bold=True)
    c = 2
    for year in context.get("student_year_totals", []):
        for key in ("male", "female", "other"):
            dcell(r, c, year.get(key, 0), bold=True)
            c += 1
    grand = context.get("student_grand_total", {})
    for key in ("total", "male", "female", "other"):
        dcell(r, c, grand.get(key, 0), bold=True)
        c += 1
    r += 2  # blank spacer row

    # ---------------- Teaching ----------------
    r = section_title(r, "Teaching Faculty Strength", teaching_cols)
    top, bottom = r, r + 1
    ws.merge_cells(start_row=top, start_column=1, end_row=bottom, end_column=1)
    hcell(top, 1, "Department")
    col = 2
    for name in designations:
        mfns_group_header(top, col, name)
        col += 3
    mfns_group_header(top, col, "Department")
    col += 3
    ws.merge_cells(start_row=top, start_column=col, end_row=bottom, end_column=col)
    hcell(top, col, "Total")
    r = bottom + 1
    for row in context.get("teaching_rows", []):
        dcell(r, 1, row["department"], center=False)
        c = 2
        for designation in row["designations"]:
            for key in ("male", "female", "other"):
                dcell(r, c, designation.get(key, 0))
                c += 1
        for key in ("male", "female", "other", "total"):
            dcell(r, c, row["total"].get(key, 0))
            c += 1
        r += 1
    teaching_total = context.get("teaching_total_row", {})
    dcell(r, 1, "Total", center=False, bold=True)
    c = 2
    for designation in teaching_total.get("designations", []):
        for key in ("male", "female", "other"):
            dcell(r, c, designation.get(key, 0), bold=True)
            c += 1
    tt = teaching_total.get("total", {})
    for key in ("male", "female", "other", "total"):
        dcell(r, c, tt.get(key, 0), bold=True)
        c += 1
    r += 2  # blank spacer row

    # ---------------- Staff ----------------
    r = section_title(r, "Staff Strength", staff_cols)
    for idx, label in enumerate(("Staff Category", "Male", "Female", "Not Set", "Total"), 1):
        hcell(r, idx, label)
    r += 1
    for row in context.get("staff_rows", []):
        dcell(r, 1, row["department"], center=False)
        for idx, key in enumerate(("male", "female", "other", "total"), 2):
            dcell(r, idx, row.get(key, 0))
        r += 1
    staff_grand = context.get("staff_grand_total", {})
    dcell(r, 1, "Total", center=False, bold=True)
    for idx, key in enumerate(("male", "female", "other", "total"), 2):
        dcell(r, idx, staff_grand.get(key, 0), bold=True)

    # ══════════════════════════════════════════
    # COLUMN WIDTHS & PAGE SETUP
    # ══════════════════════════════════════════
    ws.column_dimensions["A"].width = 28
    for i in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 9

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4, header=0.1, footer=0.1)
    ws.print_options.horizontalCentered = True

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="college_data_report.xlsx"'
    workbook.save(response)
    return response


def export_college_data_to_pdf(context):
    """Professional PDF matching the Subject Allocation report style."""
    import os
    from reportlab.platypus import (
        BaseDocTemplate, Frame, PageTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.utils import ImageReader
    from django.contrib.staticfiles import finders
    from django.conf import settings

    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1a4b8c")
    ACCENT_RED = colors.HexColor("#b91c1c")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BG_GRAY = colors.HexColor("#f8fafc")
    BORDER_GRAY = colors.HexColor("#e5e7eb")
    TOTAL_BG = colors.HexColor("#e5e7eb")

    page_w, page_h = landscape(A4)

    def header_footer(canvas, doc):
        canvas.saveState()
        left = doc.leftMargin
        right = page_w - doc.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 8 * mm

        logo_path = finders.find("images/ritlogo.png")
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, "images/ritlogo.png")
                if os.path.exists(cand):
                    logo_path = cand
        if logo_path and os.path.exists(logo_path):
            try:
                canvas.drawImage(
                    ImageReader(logo_path), left, top_y - 20 * mm,
                    width=32 * mm, height=20 * mm, preserveAspectRatio=True, mask="auto",
                )
            except Exception:
                pass

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawCentredString(center_x, top_y - 6 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")
        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "An Autonomous Institution")
        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 8.5)
        canvas.drawCentredString(center_x, top_y - 18.5 * mm, "Approved by AICTE, New Delhi")
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(center_x, top_y - 23 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
        canvas.drawCentredString(center_x, top_y - 27.5 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

        footer_y = 12 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.8)
        canvas.line(left, footer_y + 6 * mm, right, footer_y + 6 * mm)
        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 8)
        gen_time = timezone.localtime(context.get("generated_at")).strftime("%d %b %Y, %I:%M %p")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawCentredString(center_x, footer_y, "College Data Report")
        canvas.drawRightString(right, footer_y, f"Page {doc.page}")
        canvas.restoreState()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="college_data_report.pdf"'

    doc = BaseDocTemplate(
        response,
        pagesize=landscape(A4),
        title="College Data Report",
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=34 * mm,
        bottomMargin=20 * mm,
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=header_footer)])

    styles = getSampleStyleSheet()
    section_style = ParagraphStyle(
        "CDSection", parent=styles["Heading2"], fontName="Helvetica-Bold",
        fontSize=12, textColor=PRIMARY_BLUE, alignment=TA_LEFT,
        spaceBefore=12, spaceAfter=6, leading=15,
    )
    sub_style = ParagraphStyle(
        "CDSub", parent=styles["Normal"], fontName="Helvetica",
        fontSize=9, alignment=TA_CENTER, textColor=MEDIUM_GRAY, spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        "CDCell", parent=styles["Normal"], fontName="Helvetica",
        fontSize=7.5, leading=9, alignment=TA_LEFT, textColor=colors.black,
    )
    head_cell_style = ParagraphStyle(
        "CDHeadCell", parent=styles["Normal"], fontName="Helvetica-Bold",
        fontSize=7.5, leading=9, alignment=TA_CENTER, textColor=colors.white,
    )

    def P(text):
        return Paragraph(str(text), cell_style)

    def H(text):
        return Paragraph(str(text), head_cell_style)

    def base_table_style(header_rows, total_row):
        cmds = [
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("BACKGROUND", (0, 0), (-1, header_rows - 1), SECONDARY_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, header_rows - 1), colors.white),
            ("FONTNAME", (0, 0), (-1, header_rows - 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("ROWBACKGROUNDS", (0, header_rows), (-1, total_row - 1), [colors.white, BG_GRAY]),
            ("BACKGROUND", (0, total_row), (-1, total_row), TOTAL_BG),
            ("FONTNAME", (0, total_row), (-1, total_row), "Helvetica-Bold"),
        ]
        return cmds

    elements = [
        Spacer(1, 2 * mm),
        Paragraph(_college_data_filter_summary(context), sub_style),
    ]

    year_headers = context.get("year_headers", [])
    avail_w = doc.width

    # ---------------- Students ----------------
    elements.append(Paragraph("Current Student Strength", section_style))
    student_rows = context.get("student_rows", [])
    if student_rows:
        row1 = [H("Department")]
        for _, label in year_headers:
            row1 += [H(label), "", ""]
        row1 += [H("Total"), H("Total Male"), H("Total Female"), H("Not Set")]
        row2 = [""]
        for _ in year_headers:
            row2 += [H("Male"), H("Female"), H("Not Set")]
        row2 += ["", "", "", ""]
        data = [row1, row2]
        for row in student_rows:
            r = [P(row["department"])]
            for year in row["years"]:
                r += [year.get("male", 0), year.get("female", 0), year.get("other", 0)]
            r += [row["total"].get("total", 0), row["total"].get("male", 0),
                  row["total"].get("female", 0), row["total"].get("other", 0)]
            data.append(r)
        sub = [P("Sub Total")]
        for year in context.get("student_year_totals", []):
            sub += [year.get("male", 0), year.get("female", 0), year.get("other", 0)]
        grand = context.get("student_grand_total", {})
        sub += [grand.get("total", 0), grand.get("male", 0), grand.get("female", 0), grand.get("other", 0)]
        data.append(sub)

        ncols = len(row1)
        dept_w = 42 * mm
        col_w = [dept_w] + [(avail_w - dept_w) / (ncols - 1)] * (ncols - 1)
        table = Table(data, colWidths=col_w, repeatRows=2)
        span_cmds = [("SPAN", (0, 0), (0, 1))]
        col = 1
        for _ in year_headers:
            span_cmds.append(("SPAN", (col, 0), (col + 2, 0)))
            col += 3
        for _ in range(4):
            span_cmds.append(("SPAN", (col, 0), (col, 1)))
            col += 1
        table.setStyle(TableStyle(base_table_style(2, len(data) - 1) + span_cmds))
        elements.append(table)
    else:
        elements.append(Paragraph("No active student records found.", cell_style))

    # ---------------- Teaching ----------------
    elements.append(Paragraph("Teaching Faculty Strength", section_style))
    teaching_rows = context.get("teaching_rows", [])
    designations = context.get("teaching_designations", [])
    if teaching_rows:
        row1 = [H("Department")]
        for name in designations:
            row1 += [H(name), "", ""]
        row1 += [H("Department"), "", "", H("Total")]
        row2 = [""]
        for _ in designations:
            row2 += [H("Male"), H("Female"), H("Not Set")]
        row2 += [H("Male"), H("Female"), H("Not Set"), ""]
        data = [row1, row2]
        for row in teaching_rows:
            r = [P(row["department"])]
            for designation in row["designations"]:
                r += [designation.get("male", 0), designation.get("female", 0), designation.get("other", 0)]
            r += [row["total"].get("male", 0), row["total"].get("female", 0),
                  row["total"].get("other", 0), row["total"].get("total", 0)]
            data.append(r)
        teaching_total = context.get("teaching_total_row", {})
        tr = [P("Total")]
        for designation in teaching_total.get("designations", []):
            tr += [designation.get("male", 0), designation.get("female", 0), designation.get("other", 0)]
        tt = teaching_total.get("total", {})
        tr += [tt.get("male", 0), tt.get("female", 0), tt.get("other", 0), tt.get("total", 0)]
        data.append(tr)

        ncols = len(row1)
        dept_w = 42 * mm
        col_w = [dept_w] + [(avail_w - dept_w) / (ncols - 1)] * (ncols - 1)
        table = Table(data, colWidths=col_w, repeatRows=2)
        span_cmds = [("SPAN", (0, 0), (0, 1)), ("SPAN", (ncols - 1, 0), (ncols - 1, 1))]
        col = 1
        for _ in designations:
            span_cmds.append(("SPAN", (col, 0), (col + 2, 0)))
            col += 3
        span_cmds.append(("SPAN", (col, 0), (col + 2, 0)))  # Department block
        table.setStyle(TableStyle(base_table_style(2, len(data) - 1) + span_cmds))
        elements.append(table)
    else:
        elements.append(Paragraph("No teaching faculty records found.", cell_style))

    # ---------------- Staff ----------------
    elements.append(Paragraph("Staff Strength", section_style))
    staff_rows = context.get("staff_rows", [])
    if staff_rows:
        data = [[H("Staff Category"), H("Male"), H("Female"), H("Not Set"), H("Total")]]
        for row in staff_rows:
            data.append([
                P(row["department"]), row.get("male", 0), row.get("female", 0),
                row.get("other", 0), row.get("total", 0),
            ])
        staff_grand = context.get("staff_grand_total", {})
        data.append([
            P("Total"), staff_grand.get("male", 0), staff_grand.get("female", 0),
            staff_grand.get("other", 0), staff_grand.get("total", 0),
        ])
        col_w = [80 * mm] + [40 * mm] * 4
        table = Table(data, colWidths=col_w, repeatRows=1)
        table.setStyle(TableStyle(base_table_style(1, len(data) - 1)))
        elements.append(table)
    else:
        elements.append(Paragraph("No staff records found.", cell_style))

    doc.build(elements)
    return response


@check_permission("college_data")
def college_data(request):
    from course_management.models import Discontinued_Student, PassOutStudents
    from faculty_management.models import (
        DesignationMaster,
        general_information as FacultyInfo,
    )
    from user_accounts.models import StudentDetails

    selected_department = (request.GET.get("department") or "").strip()
    selected_year = (request.GET.get("year") or "").strip()
    selected_gender = (request.GET.get("gender") or "").strip()
    if not selected_department.isdigit():
        selected_department = ""
    if selected_year not in {"", "1", "2", "3", "4"}:
        selected_year = ""
    if selected_gender not in {"", "male", "female", "other"}:
        selected_gender = ""

    year_headers = [
        ("1", "I Year"),
        ("2", "II Year"),
        ("3", "III Year"),
        ("4", "IV Year"),
    ]

    def empty_gender_counts():
        return {"male": 0, "female": 0, "other": 0, "total": 0}

    def gender_key(value):
        value = (value or "").strip().lower()
        if value in {"m", "male"}:
            return "male"
        if value in {"f", "female"}:
            return "female"
        return "other"

    def year_key(year_value, semester_value=None):
        value = (year_value or "").strip().lower()
        year_map = {
            "1": "1", "i": "1", "first": "1", "first year": "1", "1st": "1", "1st year": "1",
            "2": "2", "ii": "2", "second": "2", "second year": "2", "2nd": "2", "2nd year": "2",
            "3": "3", "iii": "3", "third": "3", "third year": "3", "3rd": "3", "3rd year": "3",
            "4": "4", "iv": "4", "fourth": "4", "fourth year": "4", "final": "4", "final year": "4", "4th": "4", "4th year": "4",
        }
        if value in year_map:
            return year_map[value]

        for key, mapped in year_map.items():
            if key and value.startswith(key):
                return mapped

        semester_text = (semester_value or "").strip()
        semester_digits = "".join(ch for ch in semester_text if ch.isdigit())
        if semester_digits:
            try:
                semester = int(semester_digits)
                if 1 <= semester <= 8:
                    return str(((semester - 1) // 2) + 1)
            except ValueError:
                pass
        return None

    def department_label(department):
        if not department:
            return "Not Set"
        return (
            getattr(department, "Department", None)
            or getattr(department, "department_label", None)
            or getattr(department, "Department_code", None)
            or "Not Set"
        )

    def build_department_order(base_departments, keys_with_data):
        ordered_keys = []
        labels = {}
        for dept in base_departments:
            key = dept.id
            ordered_keys.append(key)
            labels[key] = department_label(dept)
        for key, label in sorted(keys_with_data.items(), key=lambda item: item[1].upper()):
            if key not in labels:
                ordered_keys.append(key)
                labels[key] = label
        return ordered_keys, labels

    academic_departments = list(
        Add_Department.objects.filter(is_active=True, is_academic=True)
        .order_by("Department_code", "Department")
    )
    non_academic_departments = list(
        Add_Department.objects.filter(is_active=True, is_academic=False)
        .order_by("Department_code", "Department")
    )
    filter_departments = list(
        Add_Department.objects.filter(is_active=True)
        .order_by("-is_academic", "Department_code", "Department")
    )

    student_counts = defaultdict(lambda: defaultdict(empty_gender_counts))
    student_data_departments = {}
    passed_out_student_ids = PassOutStudents.objects.exclude(student_id__isnull=True).values_list("student_id", flat=True)
    discontinued_student_ids = Discontinued_Student.objects.exclude(student_id__isnull=True).values_list("student_id", flat=True)
    student_qs = (
        StudentDetails.objects.select_related("department")
        .filter(Q(is_active=True) | Q(is_active__isnull=True))
        .filter(Q(is_break_of_study=False) | Q(is_break_of_study__isnull=True))
        .filter(Q(is_discontinued=False) | Q(is_discontinued__isnull=True))
        .exclude(id__in=passed_out_student_ids)
        .exclude(id__in=discontinued_student_ids)
    )
    if selected_department:
        student_qs = student_qs.filter(department_id=selected_department)
    if selected_gender:
        if selected_gender == "other":
            student_qs = student_qs.exclude(gender__iexact="Male").exclude(gender__iexact="Female")
        else:
            student_qs = student_qs.filter(gender__iexact=selected_gender)

    for student in student_qs:
        dept_key = student.department_id or "not_set"
        student_data_departments[dept_key] = department_label(student.department)
        current_year = year_key(student.year, student.semester)
        if not current_year:
            continue
        if selected_year and current_year != selected_year:
            continue

        counts = student_counts[dept_key][current_year]
        counts["total"] += 1
        counts[gender_key(student.gender)] += 1

    student_dept_order, student_dept_labels = build_department_order(
        academic_departments,
        student_data_departments,
    )

    student_rows = []
    student_year_totals = {year: empty_gender_counts() for year, _ in year_headers}
    student_grand_total = empty_gender_counts()

    for dept_key in student_dept_order:
        year_values = []
        row_total = empty_gender_counts()
        for year, _ in year_headers:
            counts = dict(student_counts[dept_key][year])
            year_values.append(counts)
            for key in row_total:
                row_total[key] += counts[key]
                student_year_totals[year][key] += counts[key]
                student_grand_total[key] += counts[key]

        if row_total["total"] or dept_key in student_data_departments:
            student_rows.append({
                "department": student_dept_labels.get(dept_key, "Not Set"),
                "years": year_values,
                "total": row_total,
            })

    student_year_total_list = [student_year_totals[year] for year, _ in year_headers]

    all_faculty = list(
        FacultyInfo.objects.select_related("department", "designation", "category")
        .order_by("department__Department_code", "department__Department", "name")
    )
    if selected_department:
        all_faculty = [
            faculty for faculty in all_faculty
            if str(faculty.department_id or "") == selected_department
        ]
    if selected_gender:
        all_faculty = [
            faculty for faculty in all_faculty
            if gender_key(faculty.gender) == selected_gender
        ]

    teaching_faculty = []
    staff_faculty = []
    teaching_designation_keywords = (
        "professor",
        "associate professor",
        "assistant professor",
        "lecturer",
    )
    for faculty in all_faculty:
        designation = getattr(faculty, "designation", None)
        designation_name = (getattr(designation, "designation_name", "") or "").strip()
        designation_name_lower = designation_name.lower()
        if "principal" in designation_name_lower:
            continue

        category_name = (getattr(getattr(faculty, "category", None), "category_name", "") or "").strip().lower()
        is_teaching_designation_name = any(
            keyword == designation_name_lower or keyword in designation_name_lower
            for keyword in teaching_designation_keywords
        ) or designation_name_lower in {"ap", "a.p", "a.p."} or designation_name_lower.startswith("ap ")
        is_teaching_faculty = (
            bool(designation and designation.is_teaching)
            or category_name == "teaching"
            or is_teaching_designation_name
        )
        if is_teaching_faculty:
            teaching_faculty.append(faculty)
        else:
            staff_faculty.append(faculty)

    teaching_designation_names = []
    designation_order = list(
        DesignationMaster.objects.filter(
            Q(is_teaching=True)
            | Q(designation_name__icontains="Professor")
            | Q(designation_name__icontains="Lecturer")
            | Q(designation_name__iexact="AP")
            | Q(designation_name__icontains="AP ")
        )
        .exclude(designation_name__icontains="Principal")
        .order_by("designation_name")
        .values_list("designation_name", flat=True)
    )
    designation_names_with_data = {
        (faculty.designation.designation_name if faculty.designation else "Not Set")
        for faculty in teaching_faculty
    }
    for designation_name in designation_order:
        if designation_name in designation_names_with_data:
            teaching_designation_names.append(designation_name)
    for designation_name in sorted(designation_names_with_data):
        if designation_name not in teaching_designation_names:
            teaching_designation_names.append(designation_name)

    teaching_counts = defaultdict(lambda: defaultdict(empty_gender_counts))
    teaching_data_departments = {}
    for faculty in teaching_faculty:
        dept_key = faculty.department_id or "not_set"
        teaching_data_departments[dept_key] = department_label(faculty.department)
        designation_name = faculty.designation.designation_name if faculty.designation else "Not Set"
        counts = teaching_counts[dept_key][designation_name]
        counts["total"] += 1
        counts[gender_key(faculty.gender)] += 1

    teaching_dept_order, teaching_dept_labels = build_department_order(
        academic_departments,
        teaching_data_departments,
    )
    teaching_rows = []
    teaching_designation_totals = {name: empty_gender_counts() for name in teaching_designation_names}
    teaching_grand_total = empty_gender_counts()

    for dept_key in teaching_dept_order:
        designation_values = []
        row_total = empty_gender_counts()
        for designation_name in teaching_designation_names:
            counts = dict(teaching_counts[dept_key][designation_name])
            designation_values.append({"name": designation_name, **counts})
            for key in row_total:
                row_total[key] += counts[key]
                teaching_designation_totals[designation_name][key] += counts[key]
                teaching_grand_total[key] += counts[key]

        if row_total["total"] or dept_key in teaching_data_departments:
            teaching_rows.append({
                "department": teaching_dept_labels.get(dept_key, "Not Set"),
                "designations": designation_values,
                "total": row_total,
            })

    teaching_total_row = {
        "designations": [
            {"name": name, **teaching_designation_totals[name]}
            for name in teaching_designation_names
        ],
        "total": teaching_grand_total,
    }

    def staff_group_label(faculty):
        category_name = (getattr(getattr(faculty, "category", None), "category_name", "") or "").strip()
        designation_name = (getattr(getattr(faculty, "designation", None), "designation_name", "") or "").strip()
        designation_lower = designation_name.lower()
        category_lower = category_name.lower()

        if category_lower and category_lower not in {"non teaching", "non-teaching"}:
            category_map = {
                "civil labour": "Labour",
                "house keeping": "HK",
                "housekeeping": "HK",
            }
            return category_map.get(category_lower, category_name)

        if any(key in designation_lower for key in ["lab technician", "lt", "librarian", "pd", "physical", "placement"]):
            return "LT, PD, Lib, TPO"
        if designation_lower in {"oa"}:
            return "OA"
        if any(key in designation_lower for key in ["junior assistant", "ja", "office", "accountant", "receptionist", "system admin", "gm(a)"]):
            return "Office"
        if any(key in designation_lower for key in ["driver", "transport"]):
            return "Transport"
        if any(key in designation_lower for key in ["power", "plumber", "ro operator"]):
            return "Power House"
        if "security" in designation_lower:
            return "Security"
        if any(key in designation_lower for key in ["mess", "cook", "food server"]):
            return "Mess"
        if any(key in designation_lower for key in ["labour", "civil", "gardener", "carpenter"]):
            return "Labour"
        if any(key in designation_lower for key in ["housekeeping", "sweeper", "toilet", "vessel"]):
            return "HK"

        return designation_name or category_name or department_label(faculty.department)

    staff_counts = defaultdict(empty_gender_counts)
    for faculty in staff_faculty:
        group_label = staff_group_label(faculty)
        counts = staff_counts[group_label]
        counts["total"] += 1
        counts[gender_key(faculty.gender)] += 1

    staff_group_order = [
        "LT, PD, Lib, TPO",
        "Office",
        "OA",
        "Transport",
        "Power House",
        "Security",
        "Mess",
        "Labour",
        "HK",
    ]
    staff_group_order.extend(
        sorted(label for label in staff_counts.keys() if label not in staff_group_order)
    )
    staff_rows = []
    staff_grand_total = empty_gender_counts()
    for group_label in staff_group_order:
        counts = dict(staff_counts[group_label])
        if counts["total"]:
            staff_rows.append({
                "department": group_label,
                **counts,
            })
            for key in staff_grand_total:
                staff_grand_total[key] += counts[key]

    context = {
        "generated_at": timezone.now(),
        "filters": {
            "department": selected_department,
            "year": selected_year,
            "gender": selected_gender,
        },
        "filter_departments": filter_departments,
        "year_headers": year_headers,
        "student_rows": student_rows,
        "student_year_totals": student_year_total_list,
        "student_grand_total": student_grand_total,
        "teaching_designations": teaching_designation_names,
        "teaching_rows": teaching_rows,
        "teaching_total_row": teaching_total_row,
        "staff_rows": staff_rows,
        "staff_grand_total": staff_grand_total,
        "student_total": student_grand_total["total"],
        "teaching_total": teaching_grand_total["total"],
        "staff_total": staff_grand_total["total"],
        "other_gender_total": (
            student_grand_total["other"]
            + teaching_grand_total["other"]
            + staff_grand_total["other"]
        ),
    }

    export_format = (request.GET.get("export") or "").strip().lower()
    if export_format == "excel":
        return export_college_data_to_excel(context)
    if export_format == "pdf":
        return export_college_data_to_pdf(context)

    return render(request, "faculty_leave_management/college_data.html", context)




def releaving_order(request):
    from faculty_management.models import DesignationMaster, general_information as FacultyInfo

    selected_department = (request.GET.get("department") or "").strip()
    selected_designation = (request.GET.get("designation") or "").strip()
    search_employee_id = (request.GET.get("employee_id") or "").strip()
    if not selected_department.isdigit():
        selected_department = ""
    if not selected_designation.isdigit():
        selected_designation = ""

    departments = Add_Department.objects.filter(is_active=True).order_by("Department_code", "Department")
    designations = DesignationMaster.objects.all().order_by("designation_name")

    faculty_list = FacultyInfo.objects.select_related("department", "designation").order_by(
        "department__Department", "name"
    )
    if selected_department:
        faculty_list = faculty_list.filter(department_id=selected_department)
    if selected_designation:
        faculty_list = faculty_list.filter(designation_id=selected_designation)
    if search_employee_id:
        faculty_list = faculty_list.filter(faculty_id__icontains=search_employee_id)

    context = {
        "departments": departments,
        "designations": designations,
        "faculty_list": faculty_list,
        "filters": {
            "department": selected_department,
            "designation": selected_designation,
            "employee_id": search_employee_id,
        },
    }
    return render(request, "faculty_leave_management/releaving_order.html", context)


@login_required
def relieve_faculty(request, faculty_id):
    if request.method != "POST":
        return redirect("releaving_order")

    faculty = get_object_or_404(general_information, pk=faculty_id)
    relieve_date = parse_date((request.POST.get("relieve_date") or "").strip())

    if not relieve_date:
        messages.error(request, "Please provide a valid relieving date.")
    else:
        from faculty_management.models import Faculty_Academic_Experience

        faculty.dor = relieve_date
        faculty.save(update_fields=["dor"])
        USER.objects.using("rit_approval_system").filter(
            Employee_id=str(faculty.faculty_id)
        ).update(is_active=False)

        exp_faculty_id = faculty.faculty_id or faculty.id
        latest_experience = (
            Faculty_Academic_Experience.objects.using("rit_academic_system")
            .filter(faculty_id=exp_faculty_id, institute_name__icontains="ramco")
            .order_by("-from_date")
            .first()
        )
        if latest_experience:
            latest_experience.to_date = relieve_date
            latest_experience.save(using="rit_academic_system", update_fields=["to_date"])

        messages.success(request, f"{faculty.name} has been marked as relieved on {relieve_date}.")

    redirect_url = reverse("releaving_order")
    query_string = request.GET.urlencode()
    if query_string:
        redirect_url = f"{redirect_url}?{query_string}"
    return redirect(redirect_url)


@login_required
def relieving_order_pdf(request, faculty_id):
    faculty = get_object_or_404(
        general_information.objects.select_related("department", "designation"), pk=faculty_id
    )
    if not faculty.dor:
        messages.error(request, "This faculty has not been relieved yet.")
        return redirect("releaving_order")

    def as_amount(value):
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    def as_remark(value):
        text = (value or "").strip()
        if not text:
            return "Good"
        return text[:1].upper() + text[1:]

    emoluments = {
        "basic_pay": as_amount(request.GET.get("basic_pay")),
        "agp": as_amount(request.GET.get("agp")),
        "da": as_amount(request.GET.get("da")),
        "hra": as_amount(request.GET.get("hra")),
        "special_allowance": as_amount(request.GET.get("special_allowance")),
    }
    remarks = {
        "on_work": as_remark(request.GET.get("on_work")),
        "on_conduct": as_remark(request.GET.get("on_conduct")),
    }

    return _build_relieving_order_pdf(faculty, emoluments, remarks)


def _build_relieving_order_pdf(faculty, emoluments, remarks):
    from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
    from faculty_management.models import Faculty_Academic_Experience

    PRIMARY_BLUE = colors.HexColor("#1a4b8c")
    BORDER_GRAY = colors.HexColor("#9ca3af")

    response = HttpResponse(content_type="application/pdf")
    filename = f"relieving_order_{faculty.faculty_id or faculty.id}.pdf"
    response["Content-Disposition"] = f'inline; filename="{filename}"'

    doc = BaseDocTemplate(
        response,
        pagesize=A4,
        title="Relieving Order-cum-Service Certificate",
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=16 * mm,
        bottomMargin=20 * mm,
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="normal")
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame])])

    styles = getSampleStyleSheet()
    ref_style = ParagraphStyle("Ref", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9.5)
    ref_right_style = ParagraphStyle("RefRight", parent=ref_style, alignment=TA_RIGHT)
    title_style = ParagraphStyle(
        "Title", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=12,
        alignment=TA_CENTER, spaceBefore=8, spaceAfter=12,
    )
    body_style = ParagraphStyle(
        "Body", parent=styles["Normal"], fontName="Helvetica", fontSize=10.5,
        alignment=TA_JUSTIFY, leading=15, spaceAfter=10,
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=10,
        spaceBefore=8, spaceAfter=5,
    )
    remark_style = ParagraphStyle("Remark", parent=styles["Normal"], fontName="Helvetica", fontSize=10, leading=15)
    head_cell_style = ParagraphStyle(
        "HeadCell", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=9,
        alignment=TA_CENTER, textColor=colors.white,
    )
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontName="Helvetica", fontSize=9, alignment=TA_CENTER)
    sig_style_left = ParagraphStyle(
        "SigLeft", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=10, alignment=TA_LEFT,
    )
    sig_style_right = ParagraphStyle(
        "SigRight", parent=styles["Normal"], fontName="Helvetica-Bold", fontSize=10, alignment=TA_RIGHT,
    )

    def H(text):
        return Paragraph(str(text), head_cell_style)

    def C(text):
        return Paragraph(str(text), cell_style)

    def fmt(value):
        return value.strftime("%d.%m.%Y") if value else "-"

    department = faculty.department
    department_name = department.Department if department else "-"
    department_label = (department.department_label if department else "") or "GEN"
    current_designation = faculty.designation.designation_name if faculty.designation else ""

    today = timezone.localdate()
    ref_year = faculty.dor.year if faculty.dor else today.year
    peer_ids = list(
        general_information.objects.filter(
            department_id=faculty.department_id,
            dor__year=ref_year,
        ).order_by("dor", "pk").values_list("pk", flat=True)
    )
    sequence = (peer_ids.index(faculty.pk) + 1) if faculty.pk in peer_ids else len(peer_ids) + 1
    ref_no = f"RIT/ES/Faculty/{department_label}/RO/{ref_year}/{sequence:02d}"
    faculty_id = faculty.faculty_id or faculty.id
    experiences = list(
        Faculty_Academic_Experience.objects.using("rit_academic_system")
        .filter(faculty_id=faculty_id, institute_name__icontains="ramco")
        .order_by("from_date")
    )
    designation_sequence = [exp.designation for exp in experiences if exp.designation]
    if not designation_sequence and current_designation:
        designation_sequence = [current_designation]

    if len(designation_sequence) > 1:
        worked_as = designation_sequence[0] + " and later as " + " and ".join(designation_sequence[1:])
    elif designation_sequence:
        worked_as = designation_sequence[0]
    else:
        worked_as = "Faculty"

    overall_from = experiences[0].from_date if experiences else faculty.doj
    overall_to = faculty.dor

    gender = (faculty.gender or "").strip().lower()
    if gender == "male":
        pronoun_subject, pronoun_possessive, pronoun_verb = "he", "his", "is"
    elif gender == "female":
        pronoun_subject, pronoun_possessive, pronoun_verb = "she", "her", "is"
    else:
        pronoun_subject, pronoun_possessive, pronoun_verb = "they", "their", "are"

    body_text = (
        f"This is to certify that {faculty.name} (Employee Code: {faculty.faculty_id}) has worked as "
        f"{worked_as} of the Department of {department_name} of our college from "
        f"{fmt(overall_from)} to {fmt(overall_to)}. Based upon {pronoun_possessive} request, "
        f"{pronoun_subject} {pronoun_verb} relieved from {pronoun_possessive} duties and service "
        f"with effect from {fmt(overall_to)}."
    )

    elements = [
        Table(
            [[Paragraph(f"Ref: {ref_no}", ref_style), Paragraph(f"Date: {fmt(today)}", ref_right_style)]],
            colWidths=[doc.width * 0.6, doc.width * 0.4],
        ),
        Spacer(1, 4 * mm),
        Paragraph("<u>Relieving Order-cum-Service Certificate</u>", title_style),
        Paragraph(body_text, body_style),
        Paragraph("Service Particulars:", section_style),
    ]

    if experiences:
        data = [[H("Designation"), H("From"), H("To")]]
        for exp in reversed(experiences):
            data.append([C(exp.designation or "-"), C(fmt(exp.from_date)), C(fmt(exp.to_date))])
        table = Table(data, colWidths=[doc.width * 0.5, doc.width * 0.25, doc.width * 0.25])
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.6, BORDER_GRAY),
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_BLUE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("No academic experience records found for this faculty.", remark_style))

    emo_items = [
        ("Basic Pay", emoluments.get("basic_pay", 0.0)),
        ("AGP", emoluments.get("agp", 0.0)),
        ("DA", emoluments.get("da", 0.0)),
        ("HRA", emoluments.get("hra", 0.0)),
        ("Special Allowance", emoluments.get("special_allowance", 0.0)),
    ]
    entered_items = [(label, value) for label, value in emo_items if value]

    if entered_items:
        total = sum(value for _, value in entered_items)
        col_count = len(entered_items) + 1
        emo_table = Table(
            [
                [H(label) for label, _ in entered_items] + [H("Total")],
                [C(f"Rs {value:,.0f}") for _, value in entered_items] + [C(f"Rs {total:,.0f}")],
            ],
            colWidths=[doc.width / col_count] * col_count,
        )
        emo_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.6, BORDER_GRAY),
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY_BLUE),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(Spacer(1, 6 * mm))
        elements.append(Paragraph("Emoluments last drawn:", section_style))
        elements.append(emo_table)

    elements.append(Spacer(1, 8 * mm))
    elements.append(Paragraph("<b>General Remarks:</b>", remark_style))
    elements.append(Paragraph(f"On Work &nbsp;&nbsp;&nbsp;: {remarks.get('on_work', 'Good')}", remark_style))
    elements.append(Paragraph(f"On Conduct : {remarks.get('on_conduct', 'Good')}", remark_style))

    elements.append(Spacer(1, 36 * mm))
    sig_table = Table(
        [[Paragraph("PRINCIPAL", sig_style_left), Paragraph("DIRECTOR", sig_style_right)]],
        colWidths=[doc.width / 2] * 2,
    )
    elements.append(sig_table)

    doc.build(elements)
    return response
@check_permission("punch_dashboard")
def punch_dashboard(request):
    selected_date = parse_date((request.GET.get("date") or "").strip()) or timezone.localdate()
    selected_status = (request.GET.get("status") or "").strip()
    search = (request.GET.get("search") or "").strip()

    if selected_status not in {"", "present", "single", "absent"}:
        selected_status = ""

    login_employee_id = (
        getattr(request.user, "Employee_id", None)
        or getattr(request.user, "employee_id", None)
    )
    login_employee_id = str(login_employee_id or "").strip()
    logged_faculty = None
    if login_employee_id.isdigit():
        logged_faculty = (
            general_information.objects
            .select_related("department")
            .filter(faculty_id=int(login_employee_id))
            .first()
        )

    logged_department = getattr(logged_faculty, "department", None)
    selected_department = str(getattr(logged_department, "id", "") or "")

    employees_qs = (
        general_information.objects
        .select_related("department", "designation")
        .filter(faculty_id__isnull=False)
        .filter(Q(dor__isnull=True) | Q(dor__gt=selected_date))
        .order_by("department__Department", "name", "faculty_id")
    )

    if logged_department:
        employees_qs = employees_qs.filter(department_id=logged_department.id)
    else:
        employees_qs = employees_qs.none()

    if search:
        search_filter = (
            Q(name__icontains=search)
            | Q(designation__designation_name__icontains=search)
        )
        if search.isdigit():
            search_filter |= Q(faculty_id=int(search))
        employees_qs = employees_qs.filter(search_filter)

    employees = list(employees_qs)
    faculty_ids = []
    for emp in employees:
        try:
            faculty_ids.append(int(emp.faculty_id))
        except (TypeError, ValueError):
            continue

    attendance_device_ids = [
        str(device_id).strip()
        for device_id in DeviceInfo.objects.filter(
            is_active=True,
            is_attendance=True,
        ).values_list("deviceid", flat=True)
        if str(device_id or "").strip()
    ]

    punch_errors = []
    if not login_employee_id:
        punch_errors.append("Could not find the logged-in employee ID.")
    elif not logged_faculty:
        punch_errors.append("Could not find your faculty general information record.")
    elif not logged_department:
        punch_errors.append("Your faculty record does not have a department assigned.")
    punches_by_uid = defaultdict(list)

    if not attendance_device_ids:
        punch_errors.append("No active attendance devices are configured.")
    elif faculty_ids:
        table_name = f"DeviceLogs_{selected_date.month}_{selected_date.year}"
        try:
            with connections["attendance_db"].cursor() as cursor:
                cursor.execute(
                    "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = %s",
                    [table_name],
                )
                table_exists = cursor.fetchone() is not None

            if not table_exists:
                punch_errors.append(f"Attendance table '{table_name}' was not found.")
            else:
                user_placeholders = ",".join(["%s"] * len(faculty_ids))
                device_placeholders = ",".join(["%s"] * len(attendance_device_ids))
                day_start = datetime.combine(selected_date, datetime.min.time())
                day_end = datetime.combine(selected_date + timedelta(days=1), datetime.min.time())
                sql = f"""
                    SELECT UserId, LogDate, DeviceId
                    FROM {table_name}
                    WHERE CAST(UserId AS VARCHAR(50)) IN ({user_placeholders})
                      AND LogDate >= %s
                      AND LogDate < %s
                      AND CAST(DeviceId AS VARCHAR(50)) IN ({device_placeholders})
                    ORDER BY UserId, LogDate ASC
                """
                params = [str(fid) for fid in faculty_ids]
                params.extend([day_start, day_end])
                params.extend(attendance_device_ids)

                with connections["attendance_db"].cursor() as cursor:
                    cursor.execute(sql, params)
                    columns = [column[0] for column in cursor.description]
                    for row in cursor.fetchall():
                        item = dict(zip(columns, row))
                        try:
                            uid = int(item.get("UserId"))
                        except (TypeError, ValueError):
                            continue
                        if item.get("LogDate"):
                            punches_by_uid[uid].append(item)
        except Exception as exc:
            punch_errors.append(f"Attendance database unavailable: {exc}")

    rows = []
    dept_summary = {}
    overall = {
        "total": 0,
        "punched": 0,
        "present": 0,
        "single": 0,
        "absent": 0,
    }

    for emp in employees:
        try:
            faculty_id = int(emp.faculty_id)
        except (TypeError, ValueError):
            faculty_id = None

        punches = sorted(
            punches_by_uid.get(faculty_id, []),
            key=lambda item: item.get("LogDate") or datetime.min,
        ) if faculty_id else []
        punch_count = len(punches)
        first_dt = punches[0].get("LogDate") if punch_count else None
        last_dt = punches[-1].get("LogDate") if punch_count > 1 else None

        if punch_count >= 2:
            status, badge, status_group = "Punched", "badge-present", "present"
        elif punch_count == 1:
            status, badge, status_group = "Single Punch", "badge-single", "single"
        else:
            status, badge, status_group = "Absent", "badge-absent", "absent"

        if selected_status and selected_status != status_group:
            continue

        duration = "-"
        if first_dt and last_dt and last_dt > first_dt:
            total_minutes = int((last_dt - first_dt).total_seconds() // 60)
            hours, minutes = divmod(total_minutes, 60)
            duration = f"{hours}h {minutes:02d}m"

        department_name = getattr(getattr(emp, "department", None), "Department", None) or "No Department"
        dept = dept_summary.setdefault(
            department_name,
            {
                "department": department_name,
                "total": 0,
                "punched": 0,
                "present": 0,
                "single": 0,
                "absent": 0,
            },
        )

        dept["total"] += 1
        overall["total"] += 1
        if punch_count > 0:
            dept["punched"] += 1
            overall["punched"] += 1
        dept[status_group] += 1
        overall[status_group] += 1

        rows.append({
            "faculty": emp,
            "department": department_name,
            "designation": getattr(getattr(emp, "designation", None), "designation_name", None) or "-",
            "punch_count": punch_count,
            "first_in": first_dt.strftime("%I:%M %p") if first_dt else "-",
            "last_out": last_dt.strftime("%I:%M %p") if last_dt else "-",
            "duration": duration,
            "status": status,
            "badge": badge,
        })

    context = {
        "selected_date": selected_date,
        "selected_department": selected_department,
        "selected_status": selected_status,
        "search": search,
        "dept_rows": sorted(dept_summary.values(), key=lambda item: item["department"]),
        "rows": rows,
        "overall": overall,
        "punch_errors": punch_errors,
        "attendance_device_count": len(attendance_device_ids),
    }
    return render(request, "faculty_leave_management/punch_dashboard.html", context)

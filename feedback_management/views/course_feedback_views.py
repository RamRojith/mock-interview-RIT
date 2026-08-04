from django.shortcuts import render
from django.http import HttpResponse
from django.contrib import messages
from django.shortcuts import redirect
from faculty_management.decorators import faculty_management
from user_accounts.models import Role, Department
import re
from user_accounts.decorators import faculty_login_required, no_cache, is_super_user, check_permission
from feedback_management.models import FeedbackPermission
from feedback_management.decorators import feedback_management
from course_management.models import CourseEnrollment, Course
from course_management.models import AssignSubjectFaculty
from django.utils import timezone
from user_accounts.models import StudentDetails
from datetime import date
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib import messages
from django.utils import timezone
from faculty_management.models import *
from datetime import date

def get_academic_year():
    """
    Dynamically returns academic year string.
    Example:
      If current month >= June → '2025-2026'
      Else (Jan–May) → '2024-2025'
    """
    today = date.today()
    current_year = today.year
    if today.month >= 6:  # June or later
        return f"{current_year}-{current_year + 1}"
    else:  # Before June → part of previous cycle
        return f"{current_year - 1}-{current_year}"

from feedback_management.models import *



from datetime import datetime
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.dateparse import parse_date




from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.dateparse import parse_date
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.dateparse import parse_date

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.dateparse import parse_date

@check_permission("course_survey_entry")
def course_feedback_entry(request):
    question = None

    user = request.user.Employee_id
    faculty = get_object_or_404(general_information, faculty_id=user)
    department = faculty.department

    edit_id = request.GET.get("edit")
    if edit_id:
        question = get_object_or_404(FeedbackQuestion, pk=edit_id)

    if request.method == "POST":
        action_type = request.POST.get("action_type")

        # =====================================================
        # 1) COMMON SURVEY WINDOW FOR ALL COMMON QUESTIONS
        # =====================================================
        if action_type == "set_window":
            start_date = parse_date(request.POST.get("common_start_date") or "")
            end_date = parse_date(request.POST.get("common_end_date") or "")

            if start_date and end_date and end_date < start_date:
                messages.error(request, "End date cannot be earlier than start date.")
                return redirect("course_feedback_entry")

            try:
                # update all common questions
                FeedbackQuestion.objects.filter(department__isnull=True).update(
                    start_date=start_date,
                    end_date=end_date
                )
                messages.success(request, "Common survey window updated for all departments successfully.")
            except Exception as e:
                messages.error(request, f"Error updating survey window: {str(e)}")

            return redirect("course_feedback_entry")

        # =====================================================
        # 2) ADD / EDIT QUESTION
        # department=None => visible to all departments
        # =====================================================
        question_text = (request.POST.get("question_text") or "").strip()
        category = (request.POST.get("category") or "").strip()

        if not question_text:
            messages.error(request, "Question text is required.")
            return redirect("course_feedback_entry")

        # get common date from any existing common question
        existing_question_with_window = (
            FeedbackQuestion.objects
            .filter(department__isnull=True)
            .exclude(start_date__isnull=True, end_date__isnull=True)
            .order_by("-id")
            .first()
        )

        inherited_start_date = existing_question_with_window.start_date if existing_question_with_window else None
        inherited_end_date = existing_question_with_window.end_date if existing_question_with_window else None

        if question:
            question.question_text = question_text
            question.category = category
            question.department = None   # common for all departments

            if inherited_start_date or inherited_end_date:
                question.start_date = inherited_start_date
                question.end_date = inherited_end_date

            action = "updated"
        else:
            question = FeedbackQuestion.objects.create(
                question_text=question_text,
                category=category,
                department=None,   # common for all departments
                start_date=inherited_start_date,
                end_date=inherited_end_date
            )
            action = "added"

        try:
            question.save()
            messages.success(request, f"Feedback question {action} successfully for all departments!")
            return redirect("course_feedback_entry")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
            if question and question.pk:
                return redirect(f"{request.path}?edit={question.pk}")
            return redirect("course_feedback_entry")

    # show all common questions only
    feedback_questions = FeedbackQuestion.objects.filter(department__isnull=True).order_by("-id")

    # common survey window display
    common_window_obj = (
        FeedbackQuestion.objects
        .filter(department__isnull=True)
        .exclude(start_date__isnull=True, end_date__isnull=True)
        .order_by("-id")
        .first()
    )

    common_start_date = common_window_obj.start_date if common_window_obj else None
    common_end_date = common_window_obj.end_date if common_window_obj else None

    return render(
        request,
        "feedback_management/faculty/entry/course_feedback_entry.html",
        {
            "feedback_questions": feedback_questions,
            "question": question,
            "faculty": faculty,
            "common_start_date": common_start_date,
            "common_end_date": common_end_date,
        }
    )






@check_permission("course_survey_entry")
def delete_feedback_question(request, pk):
    question = get_object_or_404(FeedbackQuestion, pk=pk)

    try:
        question.delete()
        messages.success(request, "Feedback question deleted successfully!")
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")

    return redirect('course_feedback_entry') 













from collections import OrderedDict
from datetime import date

from django.contrib import messages
from django.db import transaction
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone

from course_management.models import CourseEnrollment
from feedback_management.models import FeedbackQuestion, FeedbackSubmission, FeedbackAnswer


from collections import OrderedDict
from datetime import date

from django.contrib import messages
from django.db import transaction
from django.shortcuts import render, get_object_or_404, redirect
from django.utils.dateparse import parse_date

from user_accounts.models import StudentDetails
from course_management.models import CourseEnrollment

from feedback_management.models import (
    FeedbackQuestion,
    FeedbackSubmission,
    FeedbackAnswer,
)


# -------------------------------------------------
# Helper: Get Logged Student (matches your model)
# -------------------------------------------------
def get_student_details(request):
    emp = getattr(request.user, "Employee_id", None)
    uname = getattr(request.user, "username", None)

    if emp:
        s = StudentDetails.objects.filter(reg_no=emp).first() or StudentDetails.objects.filter(umis_id=emp).first()
        if s:
            return s

    if uname:
        s = StudentDetails.objects.filter(reg_no=uname).first() or StudentDetails.objects.filter(umis_id=uname).first()
        if s:
            return s

    return None


def group_questions_by_category(questions):
    grouped = OrderedDict()
    for q in questions:
        cat = (q.category or "General").strip()
        grouped.setdefault(cat, []).append(q)
    return grouped


# -------------------------------------------------
# Page 1: Student Course List (current context)
# -------------------------------------------------
from datetime import date
from django.db.models import Q, Min, Max
from django.contrib import messages
from django.shortcuts import render, redirect

from datetime import date
from django.db.models import Q, Min, Max
from django.contrib import messages
from django.shortcuts import render, redirect

from datetime import date
from django.contrib import messages
from django.db.models import Q, Min, Max
from django.shortcuts import redirect, render

@check_permission("course_feedback")
def course_feedback(request):
    student = get_student_details(request)
    if not student:
        messages.error(request, "Student record not found.")
        return redirect("home")

    dept = student.department
    batch = (student.batch or "").strip()
    section = (student.section or "").strip()
    sem = str(student.semester or "").strip()
    year = str(student.year or "").strip()

    if not dept:
        messages.error(request, "Student department not set.")
        return redirect("home")

    # ---------------------------------------------------------
    # Only CURRENT SEM/YEAR enrolled courses
    # ---------------------------------------------------------
    qs = (
        CourseEnrollment.objects
        .select_related("course", "faculty", "department", "regulation")
        .filter(
            student=student,
            enroll=True,
            department=dept,
            course__is_active=True,
        )
    )

    if batch:
        qs = qs.filter(batch=batch)

    if section:
        qs = qs.filter(section=section)

    if sem:
        qs = qs.filter(course__semester=str(sem))

    if year:
        qs = qs.filter(course__year=str(year))

    enrollments_list = list(qs.order_by("course__course_code", "id"))

    # ---------------------------------------------------------
    # Submitted enrollments for this student
    # ---------------------------------------------------------
    submitted_ids = set(
        FeedbackSubmission.objects
        .filter(student=student, department=dept)
        .values_list("enrollment_id", flat=True)
    )

    # ---------------------------------------------------------
    # Feedback window logic
    # COMMON QUESTIONS ONLY => department is NULL
    # ---------------------------------------------------------
    today = date.today()

    common_questions = FeedbackQuestion.objects.filter(department__isnull=True)

    open_exists = common_questions.filter(
        Q(start_date__isnull=True) | Q(start_date__lte=today),
        Q(end_date__isnull=True) | Q(end_date__gte=today),
    ).exists()

    window = common_questions.aggregate(
        w_start=Min("start_date"),
        w_end=Max("end_date"),
    )

    # ---------------------------------------------------------
    # Resolve faculty name from AssignSubjectFaculty
    # ---------------------------------------------------------
    enrollment_course_ids = [e.course_id for e in enrollments_list if e.course_id]

    assign_qs = AssignSubjectFaculty.objects.select_related("faculty", "course").filter(
        department=dept,
        course_id__in=enrollment_course_ids,
        is_active=True,
    )

    # Prefer exact batch/section match when available
    assignment_map = {}
    for a in assign_qs.order_by("course_id", "id"):
        key_exact = (a.course_id, (a.batch or "").strip(), (a.section or "").strip())
        key_common_batch = (a.course_id, "", (a.section or "").strip())
        key_common_section = (a.course_id, (a.batch or "").strip(), "")
        key_common_both = (a.course_id, "", "")

        # keep first-found mapping only
        if key_exact not in assignment_map:
            assignment_map[key_exact] = a
        if key_common_batch not in assignment_map:
            assignment_map[key_common_batch] = a
        if key_common_section not in assignment_map:
            assignment_map[key_common_section] = a
        if key_common_both not in assignment_map:
            assignment_map[key_common_both] = a

    # ---------------------------------------------------------
    # Build cards
    # ---------------------------------------------------------
    enrollments = []
    for e in enrollments_list:
        submitted = (e.id in submitted_ids)

        lookup_keys = [
            (e.course_id, (e.batch or "").strip(), (e.section or "").strip()),
            (e.course_id, "", (e.section or "").strip()),
            (e.course_id, (e.batch or "").strip(), ""),
            (e.course_id, "", ""),
        ]

        assigned_obj = None
        for key in lookup_keys:
            if key in assignment_map:
                assigned_obj = assignment_map[key]
                break

        if assigned_obj and assigned_obj.faculty:
            display_faculty_name = assigned_obj.faculty.name
        elif getattr(e, "faculty", None):
            display_faculty_name = getattr(e.faculty, "name", None) or str(e.faculty)
        else:
            display_faculty_name = "-"

        enrollments.append({
            "obj": e,
            "submitted": submitted,
            "is_open": True if submitted else open_exists,
            "win_start": window.get("w_start"),
            "win_end": window.get("w_end"),
            "display_faculty_name": display_faculty_name,
        })

    return render(
        request,
        "feedback_management/student/feedback/feedback_course_list.html",
        {
            "student": student,
            "enrollments": enrollments,
            "current_sem": sem,
            "current_year": year,
        }
    )










# -------------------------------------------------
# Page 2: Feedback Form (Submit + View Marks)
# -------------------------------------------------
from datetime import date
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404, render, redirect

from course_management.models import CourseEnrollment
from feedback_management.models import FeedbackQuestion, FeedbackSubmission, FeedbackAnswer
from datetime import date
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render

from feedback_management.models import FeedbackQuestion, FeedbackSubmission, FeedbackAnswer, gradeupload



from datetime import date
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404, render, redirect

from datetime import date
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render

from datetime import date

from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render

from course_management.models import CourseEnrollment
from feedback_management.models import FeedbackAnswer, FeedbackQuestion, FeedbackSubmission, gradeupload
from faculty_management.models import general_information
from user_accounts.models import StudentDetails

# keep your existing helper imports
# from feedback_management.views.course_feedback_views import get_student_details
# from .... import group_questions_by_category


from datetime import date

from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render

from course_management.models import CourseEnrollment
from feedback_management.models import FeedbackQuestion, FeedbackSubmission, FeedbackAnswer, gradeupload


from datetime import date
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render


def student_feedback_form(request, enrollment_id):
    student = get_student_details(request)
    if not student:
        messages.error(request, "Student record not found.")
        return redirect("home")

    enrollment = get_object_or_404(
        CourseEnrollment.objects.select_related("course", "faculty", "department"),
        pk=enrollment_id,
        student=student,
        enroll=True
    )

    department = enrollment.department or student.department
    if not department:
        messages.error(request, "Department not found.")
        return redirect("course_feedback")

    today = date.today()

    # ---------------------------------------------------------
    # Resolve actual assigned faculty
    # ---------------------------------------------------------
    assigned_faculty_obj = (
        AssignSubjectFaculty.objects
        .select_related("faculty", "course")
        .filter(
            department=department,
            course_id=enrollment.course_id,
            is_active=True,
        )
        .filter(
            Q(batch=enrollment.batch) | Q(batch__isnull=True) | Q(batch__exact="")
        )
        .filter(
            Q(section=enrollment.section) | Q(section__isnull=True) | Q(section__exact="")
        )
        .filter(
            Q(course__year=str(student.year)) | Q(course__year__isnull=True) | Q(course__year__exact="")
        )
        .filter(
            Q(course__semester=str(student.semester)) | Q(course__semester__isnull=True) | Q(course__semester__exact="")
        )
        .order_by("-id")
        .first()
    )

    actual_faculty = None
    if assigned_faculty_obj and assigned_faculty_obj.faculty:
        actual_faculty = assigned_faculty_obj.faculty
    else:
        actual_faculty = enrollment.faculty

    # ---------------------------------------------------------
    # COMMON QUESTIONS ONLY => department is NULL
    # ---------------------------------------------------------
    qs = (
        FeedbackQuestion.objects
        .filter(department__isnull=True)
        .filter(
            Q(start_date__isnull=True) | Q(start_date__lte=today),
            Q(end_date__isnull=True) | Q(end_date__gte=today),
        )
        .order_by("category", "id")
    )
    active_questions = list(qs)

    if not active_questions:
        messages.warning(request, "Feedback is not open for this course (date window closed).")
        return redirect("course_feedback")

    grade_options = list(gradeupload.objects.all().order_by("-marks", "grade"))
    if not grade_options:
        messages.warning(request, "Grade options are not configured.")
        return redirect("course_feedback")

    window_start = min([q.start_date for q in active_questions if q.start_date], default=None)
    window_end = max([q.end_date for q in active_questions if q.end_date], default=None)

    grouped = group_questions_by_category(active_questions)

    existing = (
        FeedbackSubmission.objects
        .filter(student=student, enrollment=enrollment)
        .prefetch_related("answers")
        .first()
    )

    # ---------------------------------------------------------
    # Already submitted
    # ---------------------------------------------------------
    if existing:
        answers_map = {a.question_id: a.score for a in existing.answers.all()}

        score_grade_map = {}
        for g in grade_options:
            score_grade_map[g.marks] = g.grade

        return render(
            request,
            "feedback_management/student/feedback/feedback_form.html",
            {
                "student": student,
                "enrollment": enrollment,
                "display_faculty": actual_faculty,
                "assigned_faculty_obj": assigned_faculty_obj,
                "grouped": grouped,
                "already_submitted": True,
                "answers_map": answers_map,
                "submission": existing,
                "window_start": window_start,
                "window_end": window_end,
                "grade_options": grade_options,
                "score_grade_map": score_grade_map,
                "overall_effectiveness_value": existing.overall_effectiveness_percentage if existing.overall_effectiveness_percentage is not None else "",
                "student_satisfaction_value": existing.student_satisfaction,
                "recommendation_value": existing.recommendation_to_continue_improve or "",
                "open_comments_value": existing.open_comments_for_improvement or "",
            }
        )

    # ---------------------------------------------------------
    # Submit
    # ---------------------------------------------------------
    if request.method == "POST":
        overall_effectiveness_raw = (request.POST.get("overall_effectiveness_percentage") or "").strip()
        student_satisfaction_raw = (request.POST.get("student_satisfaction") or "").strip().lower()
        recommendation = (request.POST.get("recommendation_to_continue_improve") or "").strip()
        open_comments = (request.POST.get("open_comments_for_improvement") or "").strip()

        total = 0
        answers = []

        # Validate all question grades
        for q in active_questions:
            grade_id_raw = (request.POST.get(f"q_{q.id}") or "").strip()

            if not grade_id_raw:
                messages.error(request, f"Please select a grade for: {q.question_text}")
                return redirect("student_feedback_form", enrollment_id=enrollment.id)

            try:
                grade_obj = gradeupload.objects.get(pk=int(grade_id_raw))
            except (ValueError, gradeupload.DoesNotExist):
                messages.error(request, f"Invalid grade selected for: {q.question_text}")
                return redirect("student_feedback_form", enrollment_id=enrollment.id)

            val = int(grade_obj.marks or 0)
            total += val
            answers.append(
                FeedbackAnswer(
                    question=q,
                    score=val
                )
            )

        # ---------------------------------------------------------
        # Validate extra 4 fields
        # ---------------------------------------------------------
        if not overall_effectiveness_raw:
            messages.error(request, "Please enter overall effectiveness of the course percentage.")
            return redirect("student_feedback_form", enrollment_id=enrollment.id)

        try:
            overall_effectiveness_percentage = int(overall_effectiveness_raw)
        except ValueError:
            messages.error(request, "Overall effectiveness percentage must be a valid number.")
            return redirect("student_feedback_form", enrollment_id=enrollment.id)

        if overall_effectiveness_percentage < 0 or overall_effectiveness_percentage > 100:
            messages.error(request, "Overall effectiveness percentage must be between 0 and 100 only.")
            return redirect("student_feedback_form", enrollment_id=enrollment.id)

        if student_satisfaction_raw not in ["yes", "no"]:
            messages.error(request, "Please choose student satisfaction as Yes or No.")
            return redirect("student_feedback_form", enrollment_id=enrollment.id)

        student_satisfaction = True if student_satisfaction_raw == "yes" else False

        if not recommendation:
            messages.error(request, "Please enter recommendation to continue / improve course.")
            return redirect("student_feedback_form", enrollment_id=enrollment.id)

        if not open_comments:
            messages.error(request, "Please enter open comments for improvement.")
            return redirect("student_feedback_form", enrollment_id=enrollment.id)

        # ---------------------------------------------------------
        # Save submission + answers
        # ---------------------------------------------------------
        try:
            with transaction.atomic():
                sub = FeedbackSubmission.objects.create(
                    student=student,
                    enrollment=enrollment,
                    department=department,
                    course=enrollment.course,
                    faculty=actual_faculty,
                    total_score=total,
                    window_start=window_start,
                    window_end=window_end,
                    overall_effectiveness_percentage=overall_effectiveness_percentage,
                    student_satisfaction=student_satisfaction,
                    recommendation_to_continue_improve=recommendation,
                    open_comments_for_improvement=open_comments,
                )

                for ans in answers:
                    ans.submission = sub

                FeedbackAnswer.objects.bulk_create(answers)

            messages.success(request, "Feedback submitted successfully!")
            return redirect("course_feedback")

        except Exception as e:
            messages.error(request, f"Error saving feedback: {str(e)}")
            return redirect("student_feedback_form", enrollment_id=enrollment.id)

    # ---------------------------------------------------------
    # Initial page load
    # ---------------------------------------------------------
    return render(
        request,
        "feedback_management/student/feedback/feedback_form.html",
        {
            "student": student,
            "enrollment": enrollment,
            "display_faculty": actual_faculty,
            "assigned_faculty_obj": assigned_faculty_obj,
            "grouped": grouped,
            "already_submitted": False,
            "answers_map": {},
            "submission": None,
            "window_start": window_start,
            "window_end": window_end,
            "grade_options": grade_options,
            "score_grade_map": {},
            "overall_effectiveness_value": "",
            "student_satisfaction_value": None,
            "recommendation_value": "",
            "open_comments_value": "",
        }
    )











def group_questions_by_category(questions):
    """
    Returns ordered dict like:
    {
      "Teaching Effectiveness(40)": [q1,q2,q3,q4],
      "Maturity Level(30)": [q5,q6,q7],
      ...
    }
    """
    grouped = OrderedDict()
    for q in questions:
        cat = (q.category or "General").strip()
        grouped.setdefault(cat, []).append(q)
    return grouped





from collections import defaultdict
from datetime import datetime

from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponseBadRequest
from django.shortcuts import get_object_or_404, redirect, render

from faculty_management.models import general_information
from course_management.models import Course, CourseEnrollment, AssignSubjectFaculty
from feedback_management.models import (
    FeedbackQuestion,
    FeedbackSubmission,
    FeedbackAnswer,
    CourseFeedbackRemark,
)
from user_accounts.models import StudentDetails


@check_permission("course_feedbacks")
def course_feedbacks(request):
    from collections import defaultdict
    from django.db.models import Q
    from django.shortcuts import get_object_or_404, redirect, render
    from django.contrib import messages

    user = request.user.Employee_id
    faculty = get_object_or_404(
        general_information.objects.select_related("department"),
        faculty_id=user
    )
    department = faculty.department

    if not department:
        messages.error(request, "Department not assigned for this faculty.")
        return redirect("home")

    # IMPORTANT:
    # Remove discontinued students globally.
    # Your model field is `is_discontinued`, not `is_continue`.
    dept_students_qs = StudentDetails.objects.filter(
        department=department,
        is_discontinued=False
    )

    sel_batch = (request.GET.get("batch") or "").strip()
    sel_year = (request.GET.get("year") or "").strip()
    sel_sem = (request.GET.get("semester") or "").strip()
    sel_section = (request.GET.get("section") or "").strip()
    sel_course_id = (request.GET.get("course_id") or "").strip()
    view_mode = (request.GET.get("view") or "").strip()

    filters_applied = bool(sel_year and sel_sem)
    bulk_excel_mode = (view_mode == "bulk_excel" and sel_year and sel_sem)

    # ---------------------------------------------------------
    # Courses dropdown
    # ---------------------------------------------------------
    courses_qs = Course.objects.filter(
        department=department,
        is_active=True
    )

    if sel_year:
        courses_qs = courses_qs.filter(year=str(sel_year))
    if sel_sem:
        courses_qs = courses_qs.filter(semester=str(sel_sem))

    courses = list(
        courses_qs.values("id", "title", "course_code").order_by("course_code")
    )

    # ---------------------------------------------------------
    # Questions
    # ---------------------------------------------------------
    ordered_questions = list(
        FeedbackQuestion.objects
        .filter(Q(department__isnull=True) | Q(department=department))
        .order_by("category", "id")
    )
    question_ids = [q.id for q in ordered_questions]

    for q in ordered_questions:
        q.scope_label = "Common" if q.department_id is None else (
            q.department.name if q.department else "Department"
        )

    # ---------------------------------------------------------
    # Category spans
    # ---------------------------------------------------------
    category_spans = []
    question_number_map = {}
    question_legend = []

    if ordered_questions:
        cur = ordered_questions[0].category or "General"
        span = 0

        for idx, q in enumerate(ordered_questions, start=1):
            qcat = q.category or "General"
            question_number_map[q.id] = idx
            question_legend.append({
                "number": idx,
                "category": qcat,
                "question_text": q.question_text,
                "question_id": q.id,
            })

            if qcat != cur:
                category_spans.append({"category": cur, "span": span})
                cur = qcat
                span = 1
            else:
                span += 1

        category_spans.append({"category": cur, "span": span})

    # ---------------------------------------------------------
    # NORMAL SINGLE COURSE VIEW
    # ---------------------------------------------------------
    rows = []
    submitted_student_ids = set()
    students = []

    if not bulk_excel_mode:
        if sel_course_id:
            enroll_qs = CourseEnrollment.objects.select_related("student").filter(
                department=department,
                course_id=sel_course_id,
                enroll=True,
                student__is_discontinued=False,
            )

            if sel_batch:
                enroll_qs = enroll_qs.filter(batch=sel_batch)
            if sel_year:
                enroll_qs = enroll_qs.filter(student__year=str(sel_year))
            if sel_sem:
                enroll_qs = enroll_qs.filter(student__semester=str(sel_sem))
            if sel_section:
                enroll_qs = enroll_qs.filter(section=sel_section)

            students = [
                e.student for e in enroll_qs.order_by("student__reg_no")
                if e.student and not e.student.is_discontinued
            ]
        else:
            base_students = dept_students_qs

            if sel_batch:
                base_students = base_students.filter(batch=sel_batch)
            if sel_year:
                base_students = base_students.filter(year=str(sel_year))
            if sel_sem:
                base_students = base_students.filter(semester=str(sel_sem))
            if sel_section:
                base_students = base_students.filter(section=sel_section)

            students = list(base_students.order_by("reg_no"))

        student_ids = [s.id for s in students]

        submissions_qs = FeedbackSubmission.objects.none()
        submission_ids = []
        submission_map = {}

        if sel_course_id and student_ids:
            submissions_qs = (
                FeedbackSubmission.objects
                .filter(
                    student_id__in=student_ids,
                    enrollment__course_id=sel_course_id,
                    enrollment__department=department,
                    student__is_discontinued=False,
                )
                .select_related("student", "course", "faculty", "enrollment")
            )

            submission_ids = list(submissions_qs.values_list("id", flat=True))
            submission_map = {sub.student_id: sub for sub in submissions_qs}

        student_q_marks = defaultdict(dict)
        student_total = defaultdict(int)

        if submission_ids and question_ids:
            answers = (
                FeedbackAnswer.objects
                .filter(
                    submission_id__in=submission_ids,
                    question_id__in=question_ids,
                    submission__student__is_discontinued=False,
                )
                .select_related("submission", "question")
                .order_by("submission__student_id", "question_id")
            )

            for ans in answers:
                sid = ans.submission.student_id
                qid = ans.question_id
                score_value = int(ans.score or 0)

                submitted_student_ids.add(sid)
                student_q_marks[sid][qid] = score_value
                student_total[sid] += score_value

        for idx, st in enumerate(students, start=1):
            submission = submission_map.get(st.id)

            rows.append({
                "sno": idx,
                "student_obj": st,
                "reg_no": st.reg_no or "-",
                "student_name": getattr(st, "student_name", "") or getattr(st, "name", "") or "-",
                "marks": student_q_marks.get(st.id, {}),
                "total": student_total.get(st.id, 0),
                "is_submitted": st.id in submitted_student_ids,
                "overall_effectiveness_percentage": (
                    submission.overall_effectiveness_percentage
                    if submission and submission.overall_effectiveness_percentage is not None
                    else "-"
                ),
                "student_satisfaction": (
                    "Yes" if submission and submission.student_satisfaction is True
                    else "No" if submission and submission.student_satisfaction is False
                    else "-"
                ),
                "recommendation_to_continue_improve": (
                    submission.recommendation_to_continue_improve
                    if submission and submission.recommendation_to_continue_improve
                    else "-"
                ),
                "open_comments_for_improvement": (
                    submission.open_comments_for_improvement
                    if submission and submission.open_comments_for_improvement
                    else "-"
                ),
            })

    # ---------------------------------------------------------
    # BULK EXCEL PREVIEW MODE
    # ---------------------------------------------------------
    bulk_course_sections = []
    bulk_summary_rows = []
    legend_grouped = defaultdict(list)

    if bulk_excel_mode:
        bulk_courses_qs = Course.objects.filter(
            department=department,
            year=str(sel_year),
            semester=str(sel_sem),
            is_active=True
        )

        if sel_batch or sel_section:
            enroll_course_qs = CourseEnrollment.objects.filter(
                department=department,
                enroll=True,
                student__is_discontinued=False,
            )

            if sel_batch:
                enroll_course_qs = enroll_course_qs.filter(batch=sel_batch)
            if sel_section:
                enroll_course_qs = enroll_course_qs.filter(section=sel_section)

            enroll_course_qs = enroll_course_qs.filter(
                student__year=str(sel_year),
                student__semester=str(sel_sem)
            )

            filtered_course_ids = enroll_course_qs.values_list(
                "course_id", flat=True
            ).distinct()

            bulk_courses_qs = bulk_courses_qs.filter(id__in=filtered_course_ids)

        bulk_courses = list(bulk_courses_qs.order_by("course_code"))

        for course_index, course in enumerate(bulk_courses, start=1):
            assign_qs = AssignSubjectFaculty.objects.select_related(
                "faculty", "course"
            ).filter(
                department=department,
                course_id=course.id,
                is_active=True,
            )

            if sel_batch:
                assign_qs = assign_qs.filter(batch=sel_batch)
            if sel_section:
                assign_qs = assign_qs.filter(section=sel_section)

            assign_qs = assign_qs.filter(
                course__year=str(sel_year),
                course__semester=str(sel_sem)
            )

            assign_obj = assign_qs.first()
            mapped_faculty = assign_obj.faculty if assign_obj and assign_obj.faculty else None
            filtered_faculty_name = mapped_faculty.name if mapped_faculty else faculty.name

            enroll_qs = CourseEnrollment.objects.select_related("student").filter(
                department=department,
                course_id=course.id,
                enroll=True,
                student__is_discontinued=False,
            )

            if sel_batch:
                enroll_qs = enroll_qs.filter(batch=sel_batch)
            if sel_section:
                enroll_qs = enroll_qs.filter(section=sel_section)

            enroll_qs = enroll_qs.filter(
                student__year=str(sel_year),
                student__semester=str(sel_sem)
            )

            course_students = [
                e.student for e in enroll_qs.order_by("student__reg_no")
                if e.student and not e.student.is_discontinued
            ]

            course_student_ids = [s.id for s in course_students]
            student_count = len(course_students)

            submissions_qs = FeedbackSubmission.objects.none()
            submission_ids = []
            submission_map = {}

            if course_student_ids:
                submissions_qs = (
                    FeedbackSubmission.objects
                    .filter(
                        student_id__in=course_student_ids,
                        enrollment__course_id=course.id,
                        enrollment__department=department,
                        student__is_discontinued=False,
                    )
                    .select_related("student", "course", "faculty", "enrollment")
                )

                submission_ids = list(submissions_qs.values_list("id", flat=True))
                submission_map = {sub.student_id: sub for sub in submissions_qs}

            course_student_q_marks = defaultdict(dict)
            course_student_total = defaultdict(int)
            course_question_totals = defaultdict(int)
            course_submitted_student_ids = set()

            if submission_ids:
                answers = (
                    FeedbackAnswer.objects
                    .filter(
                        submission_id__in=submission_ids,
                        question_id__in=question_ids,
                        submission__student__is_discontinued=False,
                    )
                    .select_related("submission")
                    .order_by("submission__student_id", "question_id")
                )

                for ans in answers:
                    sid = ans.submission.student_id
                    qid = ans.question_id
                    sc = int(ans.score or 0)

                    course_submitted_student_ids.add(sid)
                    course_student_q_marks[sid][qid] = sc
                    course_student_total[sid] += sc
                    course_question_totals[qid] += sc

            course_rows = []

            for i, st in enumerate(course_students, start=1):
                submission = submission_map.get(st.id)

                course_rows.append({
                    "sno": i,
                    "reg_no": st.reg_no or "-",
                    "student_name": getattr(st, "student_name", "") or getattr(st, "name", "") or "-",
                    "marks": course_student_q_marks.get(st.id, {}),
                    "total": course_student_total.get(st.id, 0),
                    "overall_effectiveness_percentage": (
                        submission.overall_effectiveness_percentage
                        if submission and submission.overall_effectiveness_percentage is not None
                        else "-"
                    ),
                    "student_satisfaction": (
                        "Yes" if submission and submission.student_satisfaction is True
                        else "No" if submission and submission.student_satisfaction is False
                        else "-"
                    ),
                    "recommendation_to_continue_improve": (
                        submission.recommendation_to_continue_improve
                        if submission and submission.recommendation_to_continue_improve
                        else "-"
                    ),
                    "open_comments_for_improvement": (
                        submission.open_comments_for_improvement
                        if submission and submission.open_comments_for_improvement
                        else "-"
                    ),
                })

            all_feedback_notes = CourseFeedbackRemark.objects.select_related("faculty").filter(
                department=department,
                course=course
            ).order_by("faculty_id", "-updated_at", "-id")

            unique_notes = []
            seen_faculty_ids = set()

            for note in all_feedback_notes:
                if note.faculty_id not in seen_faculty_ids:
                    unique_notes.append(note)
                    seen_faculty_ids.add(note.faculty_id)

            remarks_rows = []

            if unique_notes:
                for note_index, note in enumerate(unique_notes, start=1):
                    remarks_rows.append({
                        "sno": note_index,
                        "faculty_name": note.faculty.name if note.faculty else "Unknown User",
                        "remarks": note.remarks or "-",
                        "action_taken": note.action_taken or "-",
                    })
            else:
                remarks_rows.append({
                    "sno": 1,
                    "faculty_name": "-",
                    "remarks": "-",
                    "action_taken": "-",
                })

            course_grand_total = sum(course_question_totals.values())

            course_question_averages = {}

            if student_count > 0:
                for q in ordered_questions:
                    course_question_averages[q.id] = round(
                        course_question_totals.get(q.id, 0) / student_count,
                        2
                    )

                course_grand_average = round(course_grand_total / student_count, 2)
            else:
                for q in ordered_questions:
                    course_question_averages[q.id] = 0

                course_grand_average = 0

            bulk_summary_rows.append({
                "course_code": course.course_code or "-",
                "course_title": course.title or "-",
                "student_count": student_count,
                "question_averages": course_question_averages,
                "grand_average": course_grand_average,
            })

            bulk_course_sections.append({
                "course_index": course_index,
                "course_code": course.course_code or "-",
                "course_title": course.title or "-",
                "faculty_name": filtered_faculty_name,
                "student_count": student_count,
                "submitted_students": len(course_submitted_student_ids),
                "not_submitted_students": student_count - len(course_submitted_student_ids),
                "rows": course_rows,
                "remarks_rows": remarks_rows,
            })

        for item in question_legend:
            legend_grouped[item["category"]].append(item)

    return render(
        request,
        "feedback_management/faculty/feedback/course_feedbacks.html",
        {
            "faculty": faculty,
            "department": department,

            "batches": dept_students_qs.values_list("batch", flat=True).distinct().order_by("batch"),
            "years": dept_students_qs.values_list("year", flat=True).distinct().order_by("year"),
            "semesters": dept_students_qs.values_list("semester", flat=True).distinct().order_by("semester"),
            "sections": dept_students_qs.values_list("section", flat=True).distinct().order_by("section"),

            "sel_batch": sel_batch,
            "sel_year": sel_year,
            "sel_sem": sel_sem,
            "sel_section": sel_section,
            "sel_course_id": sel_course_id,

            "filters_applied": filters_applied,
            "courses": courses,

            "total_students": len(students),
            "submitted_students": len(submitted_student_ids),
            "not_submitted_students": len(students) - len(submitted_student_ids),

            "category_spans": category_spans,
            "ordered_questions": ordered_questions,
            "question_number_map": question_number_map,
            "question_legend": question_legend,

            "rows": rows,

            "bulk_excel_mode": bulk_excel_mode,
            "bulk_course_sections": bulk_course_sections,
            "bulk_summary_rows": bulk_summary_rows,
            "legend_grouped": dict(legend_grouped),
        }
    )






import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse


from collections import defaultdict
from datetime import datetime
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import get_object_or_404
from django.db.models import Q

from faculty_management.models import general_information
from course_management.models import Course, CourseEnrollment, AssignSubjectFaculty
from feedback_management.models import (
    FeedbackQuestion,
    FeedbackSubmission,
    FeedbackAnswer,
    CourseFeedbackRemark,
)


@check_permission("course_feedbacks")
def course_feedbacks_bulk_excel_download(request):
    user = request.user.Employee_id
    faculty = get_object_or_404(
        general_information.objects.select_related("department"),
        faculty_id=user
    )
    department = faculty.department

    sel_batch = (request.GET.get("batch") or "").strip()
    sel_year = (request.GET.get("year") or "").strip()
    sel_sem = (request.GET.get("semester") or "").strip()
    sel_section = (request.GET.get("section") or "").strip()

    if not department:
        return HttpResponseBadRequest("Invalid department.")

    if not (sel_year and sel_sem):
        return HttpResponseBadRequest("Select Year and Semester.")

    # ---------------------------------------------------------
    # Filtered courses
    # ---------------------------------------------------------
    courses_qs = Course.objects.filter(
        department=department,
        year=str(sel_year),
        semester=str(sel_sem),
        is_active=True
    )

    if sel_batch or sel_section:
        enroll_course_qs = CourseEnrollment.objects.filter(
            department=department,
            enroll=True,
        )

        if sel_batch:
            enroll_course_qs = enroll_course_qs.filter(batch=sel_batch)
        if sel_section:
            enroll_course_qs = enroll_course_qs.filter(section=sel_section)

        enroll_course_qs = enroll_course_qs.filter(
            student__year=str(sel_year),
            student__semester=str(sel_sem)
        )

        filtered_course_ids = enroll_course_qs.values_list("course_id", flat=True).distinct()
        courses_qs = courses_qs.filter(id__in=filtered_course_ids)

    courses = list(courses_qs.order_by("course_code"))

    if not courses:
        return HttpResponseBadRequest("No courses found for the selected filters.")

    # ---------------------------------------------------------
    # Questions
    # ---------------------------------------------------------
    ordered_questions = list(
        FeedbackQuestion.objects
        .filter(Q(department__isnull=True) | Q(department=department))
        .order_by("category", "id")
    )
    if not ordered_questions:
        return HttpResponseBadRequest("No feedback questions configured.")

    question_ids = [q.id for q in ordered_questions]

    # ---------------------------------------------------------
    # Category spans + question numbering
    # ---------------------------------------------------------
    category_spans = []
    question_number_map = {}
    question_legend = []

    cur = ordered_questions[0].category or "General"
    span = 0

    for idx, q in enumerate(ordered_questions, start=1):
        qcat = q.category or "General"
        question_number_map[q.id] = idx
        question_legend.append({
            "number": idx,
            "category": qcat,
            "question_text": q.question_text,
            "question_id": q.id,
        })

        if qcat != cur:
            category_spans.append({"category": cur, "span": span})
            cur = qcat
            span = 1
        else:
            span += 1

    category_spans.append({"category": cur, "span": span})

    academic_year_text = (
        f"Academic Year: {datetime.now().year}-{datetime.now().year + 1} "
        f"({'Odd' if str(sel_sem) in ['1', '3', '5', '7'] else 'Even'} Semester)"
    )

    # ---------------------------------------------------------
    # Workbook setup
    # ---------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_fill = PatternFill("solid", fgColor="DCE6F1")
    head_fill = PatternFill("solid", fgColor="EAF0FF")
    sub_fill = PatternFill("solid", fgColor="F8FAFC")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    bold = Font(bold=True)
    bold_big = Font(bold=True, size=14)
    bold_mid = Font(bold=True, size=12)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def style_cell(cell, font=None, fill=None, alignment=None, border_on=True):
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border_on:
            cell.border = border

    def safe_sheet_name(name):
        name = re.sub(r'[:\\/*?\[\]]', '_', name or "Sheet")
        return name[:31] if len(name) > 31 else name

    # ---------------------------------------------------------
    # Summary sheet
    # ---------------------------------------------------------
    row_no = 1
    ws.cell(row=row_no, column=1, value="Course Feedback Bulk Report")
    style_cell(ws.cell(row=row_no, column=1), font=bold_big)
    row_no += 1

    ws.cell(row=row_no, column=1, value=f"Department: {department.Department}")
    row_no += 1
    ws.cell(row=row_no, column=1, value=f"Year / Semester: {sel_year} / {sel_sem}")
    row_no += 1
    ws.cell(row=row_no, column=1, value=f"Section: {sel_section or '-'} | Batch: {sel_batch or '-'}")
    row_no += 1
    ws.cell(row=row_no, column=1, value=academic_year_text)
    row_no += 2

    summary_header_row_1 = row_no
    summary_header_row_2 = row_no + 1

    # Header row 1
    ws.cell(summary_header_row_1, 1, "Course Code")
    style_cell(ws.cell(summary_header_row_1, 1), font=bold, fill=head_fill, alignment=center)
    ws.merge_cells(start_row=summary_header_row_1, start_column=1, end_row=summary_header_row_2, end_column=1)

    ws.cell(summary_header_row_1, 2, "Course Title")
    style_cell(ws.cell(summary_header_row_1, 2), font=bold, fill=head_fill, alignment=center)
    ws.merge_cells(start_row=summary_header_row_1, start_column=2, end_row=summary_header_row_2, end_column=2)

    col_ptr = 3
    for c in category_spans:
        start_col = col_ptr
        end_col = col_ptr + c["span"] - 1
        ws.cell(summary_header_row_1, start_col, c["category"])
        style_cell(ws.cell(summary_header_row_1, start_col), font=bold, fill=head_fill, alignment=center)
        if start_col != end_col:
            ws.merge_cells(
                start_row=summary_header_row_1,
                start_column=start_col,
                end_row=summary_header_row_1,
                end_column=end_col
            )
        for q_index in range(start_col, end_col + 1):
            style_cell(ws.cell(summary_header_row_1, q_index), font=bold, fill=head_fill, alignment=center)
        col_ptr = end_col + 1

    avg_total_col = col_ptr
    students_col = col_ptr + 1

    ws.cell(summary_header_row_1, avg_total_col, "Avg Total")
    style_cell(ws.cell(summary_header_row_1, avg_total_col), font=bold, fill=head_fill, alignment=center)
    ws.merge_cells(start_row=summary_header_row_1, start_column=avg_total_col, end_row=summary_header_row_2, end_column=avg_total_col)

    ws.cell(summary_header_row_1, students_col, "Students")
    style_cell(ws.cell(summary_header_row_1, students_col), font=bold, fill=head_fill, alignment=center)
    ws.merge_cells(start_row=summary_header_row_1, start_column=students_col, end_row=summary_header_row_2, end_column=students_col)

    # Header row 2
    col_ptr = 3
    for q in ordered_questions:
        ws.cell(summary_header_row_2, col_ptr, f"Q{question_number_map[q.id]}")
        style_cell(ws.cell(summary_header_row_2, col_ptr), font=bold, fill=sub_fill, alignment=center)
        col_ptr += 1

    # ---------------------------------------------------------
    # Store summary data
    # ---------------------------------------------------------
    course_summary_rows = []

    # ---------------------------------------------------------
    # Course sheets
    # ---------------------------------------------------------
    for course_index, course in enumerate(courses, start=1):

        # -----------------------------------------------------
        # Faculty name mapped to this course
        # -----------------------------------------------------
        assign_qs = AssignSubjectFaculty.objects.select_related("faculty", "course").filter(
            department=department,
            course_id=course.id,
            is_active=True,
        )

        if sel_batch:
            assign_qs = assign_qs.filter(batch=sel_batch)
        if sel_section:
            assign_qs = assign_qs.filter(section=sel_section)

        assign_qs = assign_qs.filter(
            course__year=str(sel_year),
            course__semester=str(sel_sem)
        )

        assign_obj = assign_qs.first()
        mapped_faculty = assign_obj.faculty if assign_obj and assign_obj.faculty else None
        filtered_faculty_name = mapped_faculty.name if mapped_faculty else faculty.name

        # -----------------------------------------------------
        # Students for this course
        # -----------------------------------------------------
        enroll_qs = CourseEnrollment.objects.select_related("student").filter(
            department=department,
            course_id=course.id,
            enroll=True,
        )

        if sel_batch:
            enroll_qs = enroll_qs.filter(batch=sel_batch)
        if sel_section:
            enroll_qs = enroll_qs.filter(section=sel_section)

        enroll_qs = enroll_qs.filter(
            student__year=str(sel_year),
            student__semester=str(sel_sem)
        )

        students = [e.student for e in enroll_qs.order_by("student__reg_no") if e.student]
        student_ids = [s.id for s in students]
        student_count = len(students)

        # -----------------------------------------------------
        # Submissions
        # -----------------------------------------------------
        submission_ids = []
        submissions_qs = FeedbackSubmission.objects.none()
        submission_map = {}

        if student_ids:
            submissions_qs = (
                FeedbackSubmission.objects
                .filter(
                    student_id__in=student_ids,
                    enrollment__course_id=course.id,
                    enrollment__department=department,
                )
                .select_related("student", "course", "faculty", "enrollment")
            )
            submission_ids = list(submissions_qs.values_list("id", flat=True))
            submission_map = {sub.student_id: sub for sub in submissions_qs}

        # -----------------------------------------------------
        # Marks
        # -----------------------------------------------------
        student_q_marks = defaultdict(dict)
        student_total = defaultdict(int)
        course_question_totals = defaultdict(int)

        if submission_ids:
            answers = (
                FeedbackAnswer.objects
                .filter(submission_id__in=submission_ids, question_id__in=question_ids)
                .select_related("submission")
            )
            for ans in answers:
                sid = ans.submission.student_id
                qid = ans.question_id
                sc = int(ans.score or 0)

                student_q_marks[sid][qid] = sc
                student_total[sid] += sc
                course_question_totals[qid] += sc

        course_grand_total = sum(course_question_totals.values())

        course_question_averages = {}
        if student_count > 0:
            for q in ordered_questions:
                course_question_averages[q.id] = round(
                    course_question_totals.get(q.id, 0) / student_count, 2
                )
            course_grand_average = round(course_grand_total / student_count, 2)
        else:
            for q in ordered_questions:
                course_question_averages[q.id] = 0
            course_grand_average = 0

        course_summary_rows.append({
            "course_code": course.course_code or "-",
            "course_title": course.title or "-",
            "student_count": student_count,
            "question_averages": course_question_averages,
            "grand_average": course_grand_average,
        })

        # -----------------------------------------------------
        # Create course sheet
        # -----------------------------------------------------
        sheet_name = safe_sheet_name(f"{course.course_code}_{course_index}")
        cws = wb.create_sheet(title=sheet_name)

        r = 1
        cws.cell(r, 1, f"Course {course_index}: {course.course_code} - {course.title}")
        style_cell(cws.cell(r, 1), font=bold_big)
        r += 1

        cws.cell(r, 1, f"Department: {department.Department}")
        r += 1
        cws.cell(r, 1, f"Name of the Faculty: {filtered_faculty_name}")
        r += 1
        cws.cell(r, 1, f"Year / Semester: {sel_year} / {sel_sem}")
        r += 1
        cws.cell(r, 1, f"Section: {sel_section or '-'} | Batch: {sel_batch or '-'}")
        r += 1
        cws.cell(r, 1, academic_year_text)
        r += 2

        header_row_1 = r
        header_row_2 = r + 1

        # fixed columns
        cws.cell(header_row_1, 1, "Sl.No")
        style_cell(cws.cell(header_row_1, 1), font=bold, fill=head_fill, alignment=center)
        cws.merge_cells(start_row=header_row_1, start_column=1, end_row=header_row_2, end_column=1)

        cws.cell(header_row_1, 2, "Reg No")
        style_cell(cws.cell(header_row_1, 2), font=bold, fill=head_fill, alignment=center)
        cws.merge_cells(start_row=header_row_1, start_column=2, end_row=header_row_2, end_column=2)

        cws.cell(header_row_1, 3, "Student Name")
        style_cell(cws.cell(header_row_1, 3), font=bold, fill=head_fill, alignment=center)
        cws.merge_cells(start_row=header_row_1, start_column=3, end_row=header_row_2, end_column=3)

        col_ptr = 4
        for c in category_spans:
            start_col = col_ptr
            end_col = col_ptr + c["span"] - 1
            cws.cell(header_row_1, start_col, c["category"])
            style_cell(cws.cell(header_row_1, start_col), font=bold, fill=head_fill, alignment=center)
            if start_col != end_col:
                cws.merge_cells(
                    start_row=header_row_1,
                    start_column=start_col,
                    end_row=header_row_1,
                    end_column=end_col
                )
            for mcol in range(start_col, end_col + 1):
                style_cell(cws.cell(header_row_1, mcol), font=bold, fill=head_fill, alignment=center)
            col_ptr = end_col + 1

        total_col = col_ptr
        eff_col = col_ptr + 1
        sat_col = col_ptr + 2
        rec_col = col_ptr + 3
        open_col = col_ptr + 4

        extra_headers = [
            (total_col, "Total"),
            (eff_col, "Overall Effectiveness (%)"),
            (sat_col, "Student Satisfaction"),
            (rec_col, "Recommendation to Continue / Improve"),
            (open_col, "Open Comments for Improvement"),
        ]

        for col_no, head in extra_headers:
            cws.cell(header_row_1, col_no, head)
            style_cell(cws.cell(header_row_1, col_no), font=bold, fill=head_fill, alignment=center)
            cws.merge_cells(start_row=header_row_1, start_column=col_no, end_row=header_row_2, end_column=col_no)

        # question row
        col_ptr = 4
        for q in ordered_questions:
            cws.cell(header_row_2, col_ptr, f"Q{question_number_map[q.id]}")
            style_cell(cws.cell(header_row_2, col_ptr), font=bold, fill=sub_fill, alignment=center)
            col_ptr += 1

        data_start_row = header_row_2 + 1

        if students:
            for i, st in enumerate(students, start=1):
                submission = submission_map.get(st.id)

                cws.cell(data_start_row, 1, i)
                cws.cell(data_start_row, 2, st.reg_no or "-")
                cws.cell(data_start_row, 3, getattr(st, "student_name", "") or getattr(st, "name", "") or "-")

                style_cell(cws.cell(data_start_row, 1), alignment=center)
                style_cell(cws.cell(data_start_row, 2), alignment=center)
                style_cell(cws.cell(data_start_row, 3), alignment=left)

                col_ptr = 4
                for q in ordered_questions:
                    cws.cell(data_start_row, col_ptr, student_q_marks.get(st.id, {}).get(q.id, 0))
                    style_cell(cws.cell(data_start_row, col_ptr), alignment=center)
                    col_ptr += 1

                overall_effectiveness = (
                    submission.overall_effectiveness_percentage
                    if submission and submission.overall_effectiveness_percentage is not None
                    else "-"
                )
                student_satisfaction = (
                    "Yes" if submission and submission.student_satisfaction is True
                    else "No" if submission and submission.student_satisfaction is False
                    else "-"
                )
                recommendation = (
                    submission.recommendation_to_continue_improve
                    if submission and submission.recommendation_to_continue_improve
                    else "-"
                )
                open_comments = (
                    submission.open_comments_for_improvement
                    if submission and submission.open_comments_for_improvement
                    else "-"
                )

                cws.cell(data_start_row, total_col, student_total.get(st.id, 0))
                cws.cell(data_start_row, eff_col, overall_effectiveness)
                cws.cell(data_start_row, sat_col, student_satisfaction)
                cws.cell(data_start_row, rec_col, recommendation)
                cws.cell(data_start_row, open_col, open_comments)

                style_cell(cws.cell(data_start_row, total_col), font=bold, fill=sub_fill, alignment=center)
                style_cell(cws.cell(data_start_row, eff_col), alignment=center)
                style_cell(cws.cell(data_start_row, sat_col), alignment=center)
                style_cell(cws.cell(data_start_row, rec_col), alignment=left)
                style_cell(cws.cell(data_start_row, open_col), alignment=left)

                data_start_row += 1
        else:
            cws.cell(data_start_row, 1, "No students found.")
            style_cell(cws.cell(data_start_row, 1), font=bold, alignment=center)
            cws.merge_cells(
                start_row=data_start_row,
                start_column=1,
                end_row=data_start_row,
                end_column=open_col
            )
            for ccol in range(1, open_col + 1):
                style_cell(cws.cell(data_start_row, ccol), alignment=center)
            data_start_row += 1

        # -----------------------------------------------------
        # Remarks / Action Taken
        # -----------------------------------------------------
        data_start_row += 2
        cws.cell(data_start_row, 1, "Remarks / Action Taken")
        style_cell(cws.cell(data_start_row, 1), font=bold_mid)
        data_start_row += 1

        remark_headers = ["S.No", "Faculty", "Remarks", "Action Plan"]
        for idx, head in enumerate(remark_headers, start=1):
            cws.cell(data_start_row, idx, head)
            style_cell(cws.cell(data_start_row, idx), font=bold, fill=sub_fill, alignment=center)

        all_feedback_notes = CourseFeedbackRemark.objects.select_related("faculty").filter(
            department=department,
            course=course
        ).order_by("faculty_id", "-updated_at", "-id")

        unique_notes = []
        seen_faculty_ids = set()
        for note in all_feedback_notes:
            if note.faculty_id not in seen_faculty_ids:
                unique_notes.append(note)
                seen_faculty_ids.add(note.faculty_id)

        data_start_row += 1
        if unique_notes:
            for note_index, note in enumerate(unique_notes, start=1):
                note_faculty_name = note.faculty.name if note.faculty else "Unknown User"

                values = [
                    note_index,
                    note_faculty_name,
                    note.remarks if note.remarks else "-",
                    note.action_taken if note.action_taken else "-",
                ]
                for idx, val in enumerate(values, start=1):
                    cws.cell(data_start_row, idx, val)
                    style_cell(
                        cws.cell(data_start_row, idx),
                        alignment=center if idx == 1 else left,
                        fill=white_fill
                    )
                data_start_row += 1
        else:
            cws.cell(data_start_row, 1, 1)
            cws.cell(data_start_row, 2, "-")
            cws.cell(data_start_row, 3, "-")
            cws.cell(data_start_row, 4, "-")
            for idx in range(1, 5):
                style_cell(
                    cws.cell(data_start_row, idx),
                    alignment=center if idx == 1 else left,
                    fill=white_fill
                )

        # -----------------------------------------------------
        # Column widths
        # -----------------------------------------------------
        width_map = {
            1: 8,
            2: 18,
            3: 28,
        }

        for idx in range(4, total_col):
            width_map[idx] = 9

        width_map[total_col] = 10
        width_map[eff_col] = 18
        width_map[sat_col] = 18
        width_map[rec_col] = 35
        width_map[open_col] = 35

        for col_idx, width in width_map.items():
            cws.column_dimensions[get_column_letter(col_idx)].width = width

        cws.column_dimensions["A"].width = 8
        cws.freeze_panes = f"A{header_row_2 + 1}"

    # ---------------------------------------------------------
    # Fill summary rows
    # ---------------------------------------------------------
    current_summary_row = summary_header_row_2 + 1
    for item in course_summary_rows:
        ws.cell(current_summary_row, 1, item["course_code"])
        ws.cell(current_summary_row, 2, item["course_title"])
        style_cell(ws.cell(current_summary_row, 1), font=bold, alignment=center)
        style_cell(ws.cell(current_summary_row, 2), alignment=left)

        col_ptr = 3
        for q in ordered_questions:
            ws.cell(current_summary_row, col_ptr, item["question_averages"].get(q.id, 0))
            style_cell(ws.cell(current_summary_row, col_ptr), alignment=center)
            col_ptr += 1

        ws.cell(current_summary_row, avg_total_col, item["grand_average"])
        ws.cell(current_summary_row, students_col, item["student_count"])
        style_cell(ws.cell(current_summary_row, avg_total_col), font=bold, fill=sub_fill, alignment=center)
        style_cell(ws.cell(current_summary_row, students_col), font=bold, fill=sub_fill, alignment=center)

        current_summary_row += 1

    # column widths summary
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 32

    for idx in range(3, avg_total_col):
        ws.column_dimensions[get_column_letter(idx)].width = 9

    ws.column_dimensions[get_column_letter(avg_total_col)].width = 12
    ws.column_dimensions[get_column_letter(students_col)].width = 12
    ws.freeze_panes = f"A{summary_header_row_2 + 1}"

    # ---------------------------------------------------------
    # Legend sheet
    # ---------------------------------------------------------
    lws = wb.create_sheet(title="Question Legend")
    lr = 1

    lws.cell(lr, 1, "Question Reference / Legend")
    style_cell(lws.cell(lr, 1), font=bold_big)
    lr += 1

    lws.cell(lr, 1, f"Department: {department.Department}")
    lr += 1
    lws.cell(lr, 1, f"Year / Semester: {sel_year} / {sel_sem}")
    lr += 1
    lws.cell(lr, 1, f"Section: {sel_section or '-'} | Batch: {sel_batch or '-'}")
    lr += 2

    lws.cell(lr, 1, "No.")
    lws.cell(lr, 2, "Category")
    lws.cell(lr, 3, "Question")

    for c in range(1, 4):
        style_cell(lws.cell(lr, c), font=bold, fill=head_fill, alignment=center)

    for idx, q in enumerate(ordered_questions, start=1):
        lr += 1
        lws.cell(lr, 1, idx)
        lws.cell(lr, 2, q.category or "General")
        lws.cell(lr, 3, q.question_text)

        style_cell(lws.cell(lr, 1), font=bold, alignment=center)
        style_cell(lws.cell(lr, 2), alignment=center)
        style_cell(lws.cell(lr, 3), alignment=left)

    lws.column_dimensions["A"].width = 10
    lws.column_dimensions["B"].width = 24
    lws.column_dimensions["C"].width = 90
    lws.freeze_panes = "A7"

    # ---------------------------------------------------------
    # Download response
    # ---------------------------------------------------------
    filename = f"Course_Feedbacks_Bulk_{department.Department}_{sel_year}_Sem{sel_sem}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response














import os
from datetime import datetime
from collections import defaultdict

from django.conf import settings
from django.http import HttpResponse, HttpResponseBadRequest
from django.contrib.staticfiles import finders
from django.views.decorators.http import require_GET

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

from faculty_management.models import general_information
from course_management.models import Course, CourseEnrollment
from feedback_management.models import FeedbackQuestion, FeedbackSubmission, FeedbackAnswer


def _draw_rit_header_footer(c: canvas.Canvas, page_w, page_h, title="STUDENTS FEEDBACK", subtitle=""):
    left_margin = 18 * mm
    right_margin = page_w - 12 * mm
    top_margin = page_h - 12 * mm
    bottom_margin = 14 * mm

    logo_height = 22 * mm
    logo_rel = "images/ritlogo.png"
    logo_path = finders.find(logo_rel)

    if not logo_path:
        for d in getattr(settings, "STATICFILES_DIRS", []):
            cand = os.path.join(d, logo_rel)
            if os.path.exists(cand):
                logo_path = cand
                break

    logo_bottom_y = None

    if logo_path and os.path.exists(logo_path):
        try:
            img = ImageReader(logo_path)
            iw, ih = img.getSize()
            tw = logo_height * (iw / float(ih))
            logo_y = top_margin - logo_height + 11 * mm
            c.drawImage(img, left_margin, logo_y, width=tw, height=logo_height, mask='auto')
            logo_bottom_y = logo_y
        except Exception:
            pass

    tuv_logo_rel = "images/tuvlogo.png"
    tuv_path = finders.find(tuv_logo_rel)

    if tuv_path and os.path.exists(tuv_path):
        try:
            img2 = ImageReader(tuv_path)
            iw, ih = img2.getSize()
            th = 16 * mm
            tw = th * (iw / float(ih))
            tuv_y = top_margin - th + 6 * mm
            c.drawImage(img2, right_margin - tw, tuv_y, width=tw, height=th, mask='auto')
        except Exception:
            pass

    c.setFillColor(colors.HexColor("#2C3E50"))

    c.setFont("Helvetica-Bold", 15)
    c.drawCentredString(page_w / 2.0, top_margin, "RAMCO INSTITUTE OF TECHNOLOGY")

    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(page_w / 2.0, top_margin - 7 * mm, title)

    if subtitle:
        c.setFont("Helvetica", 10)
        c.drawCentredString(page_w / 2.0, top_margin - 13 * mm, subtitle)

    c.setStrokeColor(colors.black)
    c.setLineWidth(0.5)

    if logo_bottom_y:
        line_y = logo_bottom_y - 6 * mm
    else:
        line_y = top_margin - 24 * mm

    c.line(left_margin, line_y, right_margin, line_y)

    footer_y = bottom_margin

    c.setStrokeColor(colors.HexColor("#C0392B"))
    c.setLineWidth(0.8)
    c.line(left_margin, footer_y + 7 * mm, right_margin, footer_y + 7 * mm)

    c.setFont("Helvetica", 8.8)
    c.setFillColor(colors.HexColor("#2C3E50"))

    c.drawCentredString(
        page_w / 2.0,
        footer_y + 2.5 * mm,
        "North Venganallur, Ayyanarkovil Road, Rajapalayam - 626 117, Virudhunagar District, Tamil Nadu."
    )

    c.drawCentredString(
        page_w / 2.0,
        footer_y - 2.0 * mm,
        "Tel: 04563 233400 | E-mail: rit@ritrjpm.ac.in | Web: www.ritrjpm.ac.in"
    )







from collections import defaultdict
from datetime import datetime

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest
from django.views.decorators.http import require_GET

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)


from collections import defaultdict
from datetime import datetime

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest
from django.views.decorators.http import require_GET

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)



@require_GET
@check_permission("course_feedbacks")
def course_feedbacks_pdf(request):
    from collections import defaultdict

    user = request.user.Employee_id
    login_faculty = general_information.objects.select_related("department").get(
        faculty_id=user
    )
    department = login_faculty.department

    sel_batch = (request.GET.get("batch") or "").strip()
    sel_year = (request.GET.get("year") or "").strip()
    sel_sem = (request.GET.get("semester") or "").strip()
    sel_section = (request.GET.get("section") or "").strip()
    sel_course_id = (request.GET.get("course_id") or "").strip()

    if not (sel_year and sel_sem and sel_course_id):
        return HttpResponseBadRequest("Select Year, Semester and Course to generate PDF.")

    try:
        course = Course.objects.get(id=sel_course_id, department=department)
    except Course.DoesNotExist:
        return HttpResponseBadRequest("Invalid course.")

    # ---------------------------------------------------------
    # Get faculty name for selected filtered subject
    # ---------------------------------------------------------
    assign_qs = (
        AssignSubjectFaculty.objects
        .select_related("faculty", "course")
        .filter(
            department=department,
            course_id=sel_course_id,
            is_active=True,
        )
    )

    if sel_batch:
        assign_qs = assign_qs.filter(batch=sel_batch)
    if sel_section:
        assign_qs = assign_qs.filter(section=sel_section)
    if sel_year:
        assign_qs = assign_qs.filter(course__year=str(sel_year))
    if sel_sem:
        assign_qs = assign_qs.filter(course__semester=str(sel_sem))

    assign_obj = assign_qs.first()
    subject_faculty_name = (
        assign_obj.faculty.name
        if assign_obj and assign_obj.faculty
        else login_faculty.name
    )

    # ---------------------------------------------------------
    # Students - discontinued students removed
    # ---------------------------------------------------------
    enroll_qs = (
        CourseEnrollment.objects
        .select_related("student")
        .filter(
            department=department,
            course_id=sel_course_id,
            enroll=True,
            student__is_discontinued=False,
        )
    )

    if sel_batch:
        enroll_qs = enroll_qs.filter(batch=sel_batch)
    if sel_year:
        enroll_qs = enroll_qs.filter(student__year=str(sel_year))
    if sel_sem:
        enroll_qs = enroll_qs.filter(student__semester=str(sel_sem))
    if sel_section:
        enroll_qs = enroll_qs.filter(section=sel_section)

    students = [
        e.student for e in enroll_qs.order_by("student__reg_no")
        if e.student and not e.student.is_discontinued
    ]
    student_ids = [s.id for s in students]

    # ---------------------------------------------------------
    # Questions
    # ---------------------------------------------------------
    ordered_questions = list(
        FeedbackQuestion.objects
        .filter(Q(department__isnull=True) | Q(department=department))
        .order_by("category", "id")
    )

    if not ordered_questions:
        return HttpResponseBadRequest("No feedback questions configured.")

    question_ids = [q.id for q in ordered_questions]

    # ---------------------------------------------------------
    # Category spans + numbering
    # ---------------------------------------------------------
    category_spans = []
    question_number_map = {}
    question_legend = []

    cur = ordered_questions[0].category or "General"
    span = 0

    for idx, q in enumerate(ordered_questions, start=1):
        qcat = q.category or "General"

        question_number_map[q.id] = idx
        question_legend.append({
            "number": idx,
            "category": qcat,
            "question_text": q.question_text,
            "question_id": q.id,
        })

        if qcat != cur:
            category_spans.append({"category": cur, "span": span})
            cur = qcat
            span = 1
        else:
            span += 1

    category_spans.append({"category": cur, "span": span})

    # ---------------------------------------------------------
    # Submissions - discontinued students removed
    # ---------------------------------------------------------
    submissions_qs = FeedbackSubmission.objects.none()
    submission_ids = []
    submission_map = {}

    if student_ids:
        submissions_qs = (
            FeedbackSubmission.objects
            .filter(
                student_id__in=student_ids,
                enrollment__course_id=sel_course_id,
                enrollment__department=department,
                student__is_discontinued=False,
            )
            .select_related("student", "course", "faculty", "enrollment")
        )

        submission_ids = list(submissions_qs.values_list("id", flat=True))
        submission_map = {sub.student_id: sub for sub in submissions_qs}

    # ---------------------------------------------------------
    # Marks - discontinued students removed
    # ---------------------------------------------------------
    student_q_marks = defaultdict(dict)
    student_total = defaultdict(int)

    if submission_ids:
        answers = (
            FeedbackAnswer.objects
            .filter(
                submission_id__in=submission_ids,
                question_id__in=question_ids,
                submission__student__is_discontinued=False,
            )
            .select_related("submission")
        )

        for ans in answers:
            sid = ans.submission.student_id
            qid = ans.question_id
            sc = int(ans.score or 0)

            student_q_marks[sid][qid] = sc
            student_total[sid] += sc

    # ---------------------------------------------------------
    # PDF setup
    # ---------------------------------------------------------
    PDF_PAGE_SIZE = landscape(A3)

    filename = f"Feedback_{course.course_code}_{sel_year}_Sem{sel_sem}.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=PDF_PAGE_SIZE,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=50 * mm,
        bottomMargin=20 * mm,
        title="Students Feedback"
    )

    # ---------------------------------------------------------
    # Styles
    # ---------------------------------------------------------
    styles = getSampleStyleSheet()

    p_center = ParagraphStyle(
        "p_center",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=5.2,
        leading=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a"),
    )

    p_center_bold = ParagraphStyle(
        "p_center_bold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=5.4,
        leading=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a"),
    )

    p_small_left = ParagraphStyle(
        "p_small_left",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=5.8,
        leading=7,
        alignment=0,
        textColor=colors.HexColor("#0f172a"),
        wordWrap="CJK",
    )

    info_style = ParagraphStyle(
        "info",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#0f172a"),
    )

    legend_heading_style = ParagraphStyle(
        "legend_heading_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=6,
    )

    legend_category_style = ParagraphStyle(
        "legend_category_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=4,
        spaceBefore=6,
    )

    legend_text_style = ParagraphStyle(
        "legend_text_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=12,
        textColor=colors.HexColor("#111827"),
    )

    # ---------------------------------------------------------
    # Main table rows
    # ---------------------------------------------------------
    row1 = ["Sl.No"]

    for c in category_spans:
        row1.append(Paragraph(str(c["category"]), p_center))
        for _ in range(c["span"] - 1):
            row1.append("")

    row1.extend([
        "Total",
        Paragraph("Overall Effectiveness (%)", p_center),
        Paragraph("Student Satisfaction", p_center),
        Paragraph("Recommendation to Continue / Improve", p_center),
        Paragraph("Open Comments for Improvement", p_center),
    ])

    row2 = [""]

    for q in ordered_questions:
        row2.append(Paragraph(str(question_number_map.get(q.id, "")), p_center_bold))

    row2.extend(["", "", "", "", ""])

    data = [row1, row2]

    for i, st in enumerate(students, start=1):
        submission = submission_map.get(st.id)

        overall_effectiveness = (
            str(submission.overall_effectiveness_percentage)
            if submission and submission.overall_effectiveness_percentage is not None
            else "-"
        )

        student_satisfaction = (
            "Yes" if submission and submission.student_satisfaction is True
            else "No" if submission and submission.student_satisfaction is False
            else "-"
        )

        recommendation = (
            submission.recommendation_to_continue_improve
            if submission and submission.recommendation_to_continue_improve
            else "-"
        )

        open_comments = (
            submission.open_comments_for_improvement
            if submission and submission.open_comments_for_improvement
            else "-"
        )

        r = [str(i)]

        for q in ordered_questions:
            r.append(str(student_q_marks.get(st.id, {}).get(q.id, 0)))

        r.extend([
            str(student_total.get(st.id, 0)),
            overall_effectiveness,
            student_satisfaction,
            Paragraph(recommendation, p_small_left),
            Paragraph(open_comments, p_small_left),
        ])

        data.append(r)

    if not students:
        no_data_row = ["No students found."]
        no_data_row += [""] * len(ordered_questions)
        no_data_row += ["", "", "", "", ""]
        data.append(no_data_row)

    # ---------------------------------------------------------
    # Table widths
    # ---------------------------------------------------------
    page_w, page_h = doc.pagesize
    usable_w = page_w - doc.leftMargin - doc.rightMargin

    col_w_sno = 10 * mm
    col_w_total = 12 * mm
    col_w_eff = 16 * mm
    col_w_sat = 18 * mm
    col_w_rec = 32 * mm
    col_w_open = 32 * mm

    q_count = len(ordered_questions)

    fixed_width = (
        col_w_sno +
        col_w_total +
        col_w_eff +
        col_w_sat +
        col_w_rec +
        col_w_open
    )

    q_w = (usable_w - fixed_width) / float(max(q_count, 1))
    q_w = max(q_w, 6 * mm)

    col_widths = (
        [col_w_sno] +
        [q_w] * q_count +
        [col_w_total, col_w_eff, col_w_sat, col_w_rec, col_w_open]
    )

    tbl = Table(data, colWidths=col_widths, repeatRows=2)

    table_style = TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d1d5db")),
        ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f8fafc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 1), 5.5),

        ("FONTNAME", (0, 2), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 2), (-3, -1), 5.5),
        ("ALIGN", (0, 2), (-3, -1), "CENTER"),

        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),

        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f8fafc")),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#111827")),

        ("ALIGN", (-2, 2), (-1, -1), "LEFT"),
        ("VALIGN", (-2, 2), (-1, -1), "TOP"),
        ("LEFTPADDING", (-2, 2), (-1, -1), 3),
        ("RIGHTPADDING", (-2, 2), (-1, -1), 3),

        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])

    table_style.add("SPAN", (0, 0), (0, 1))
    table_style.add("SPAN", (q_count + 1, 0), (q_count + 1, 1))
    table_style.add("SPAN", (q_count + 2, 0), (q_count + 2, 1))
    table_style.add("SPAN", (q_count + 3, 0), (q_count + 3, 1))
    table_style.add("SPAN", (q_count + 4, 0), (q_count + 4, 1))
    table_style.add("SPAN", (q_count + 5, 0), (q_count + 5, 1))

    start_col = 1

    for c in category_spans:
        end_col = start_col + c["span"] - 1
        table_style.add("SPAN", (start_col, 0), (end_col, 0))
        start_col = end_col + 1

    table_style.add(
        "BACKGROUND",
        (q_count + 1, 2),
        (q_count + 1, -1),
        colors.HexColor("#f8fafc")
    )
    table_style.add(
        "FONTNAME",
        (q_count + 1, 2),
        (q_count + 1, -1),
        "Helvetica-Bold"
    )

    if not students:
        table_style.add("SPAN", (0, 2), (-1, 2))
        table_style.add("ALIGN", (0, 2), (-1, 2), "CENTER")
        table_style.add("FONTNAME", (0, 2), (-1, 2), "Helvetica-Oblique")

    tbl.setStyle(table_style)

    semester_type = "Odd" if str(sel_sem) in ["1", "3", "5", "7"] else "Even"
    academic_year_text = f"Academic Year: {get_academic_year()} ({semester_type} Semester)"
    subtitle = academic_year_text

    info_lines = [
        Paragraph(f"Name of the Subject: {course.course_code} - {course.title}", info_style),
        Paragraph(f"Name of the Faculty: {subject_faculty_name}", info_style),
        Paragraph(
            f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
            f"Section: {sel_section or '-'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or '-'}",
            info_style
        ),
        Spacer(1, 8),
    ]

    # ---------------------------------------------------------
    # Story
    # ---------------------------------------------------------
    story = []
    story.extend(info_lines)
    story.append(tbl)

    # ---------------------------------------------------------
    # Final legend page
    # ---------------------------------------------------------
    story.append(PageBreak())
    story.append(Paragraph("Question Reference / Legend", legend_heading_style))
    story.append(Paragraph(f"Name of the Subject: {course.course_code} - {course.title}", info_style))
    story.append(Paragraph(f"Name of the Faculty: {subject_faculty_name}", info_style))
    story.append(Paragraph(
        f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
        f"Section: {sel_section or '-'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or '-'}",
        info_style
    ))
    story.append(Spacer(1, 8))

    legend_grouped = defaultdict(list)

    for item in question_legend:
        legend_grouped[item["category"]].append(item)

    legend_usable_w = page_w - doc.leftMargin - doc.rightMargin

    for cat in legend_grouped:
        story.append(Paragraph(str(cat), legend_category_style))

        legend_data = [["No.", "Question"]]

        for item in legend_grouped[cat]:
            legend_data.append([
                str(item["number"]),
                Paragraph(item["question_text"], legend_text_style)
            ])

        legend_tbl = Table(
            legend_data,
            colWidths=[18 * mm, legend_usable_w - (18 * mm)],
            repeatRows=1
        )

        legend_tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8.5),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 1), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8.3),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ]))

        story.append(legend_tbl)
        story.append(Spacer(1, 8))

    def _first_page(canv, doc_):
        pw, ph = doc_.pagesize
        _draw_rit_header_footer(
            canv,
            pw,
            ph,
            title="STUDENTS FEEDBACK",
            subtitle=subtitle
        )

    def _later_pages(canv, doc_):
        pw, ph = doc_.pagesize
        _draw_rit_header_footer(
            canv,
            pw,
            ph,
            title="STUDENTS FEEDBACK",
            subtitle=subtitle
        )

    doc.build(story, onFirstPage=_first_page, onLaterPages=_later_pages)

    return response








from collections import defaultdict
from datetime import datetime

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
)


from collections import defaultdict
from datetime import datetime

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)








@check_permission("course_feedbacks")
def course_feedbacks_bulk_pdf(request):
    user = request.user.Employee_id
    faculty = general_information.objects.select_related("department").get(faculty_id=user)
    faculty_department = faculty.department

    permission_scope = _get_feedback_permission_scope(request)
    if not permission_scope["has_access"]:
        return HttpResponseForbidden("You do not have permission to view feedback data.")

    sel_department_id = (request.GET.get("department_id") or "").strip()
    sel_batch = (request.GET.get("batch") or "").strip()
    sel_year = (request.GET.get("year") or "").strip()
    sel_sem = (request.GET.get("semester") or "").strip()
    sel_section = (request.GET.get("section") or "").strip()

    # ---------------------------------------------------------
    # Department resolve
    # ---------------------------------------------------------
    if permission_scope.get("can_view_all"):
        if sel_department_id:
            department = Add_Department.objects.filter(
                id=sel_department_id,
                is_active=True
            ).first()
        else:
            department = faculty_department
    else:
        department = faculty_department

    if not department:
        return HttpResponseBadRequest("Invalid department.")

    if not (sel_year and sel_sem):
        return HttpResponseBadRequest("Select Year and Semester to generate Bulk PDF.")

    # ---------------------------------------------------------
    # Filtered courses
    # ---------------------------------------------------------
    courses_qs = Course.objects.filter(
        department=department,
        year=str(sel_year),
        semester=str(sel_sem),
        is_active=True
    )

    if sel_batch or sel_section:
        enroll_course_qs = CourseEnrollment.objects.filter(
            department=department,
            enroll=True,
            student__is_discontinued=False,
        )

        if sel_batch:
            enroll_course_qs = enroll_course_qs.filter(batch=sel_batch)
        if sel_section:
            enroll_course_qs = enroll_course_qs.filter(section=sel_section)

        enroll_course_qs = enroll_course_qs.filter(
            student__year=str(sel_year),
            student__semester=str(sel_sem)
        )

        filtered_course_ids = enroll_course_qs.values_list("course_id", flat=True).distinct()
        courses_qs = courses_qs.filter(id__in=filtered_course_ids)

    courses = list(courses_qs.order_by("course_code"))

    if not courses:
        return HttpResponseBadRequest("No courses found for the selected filters.")

    # ---------------------------------------------------------
    # Questions
    # ---------------------------------------------------------
    ordered_questions = list(
        FeedbackQuestion.objects
        .filter(Q(department__isnull=True) | Q(department=department))
        .order_by("category", "id")
    )

    if not ordered_questions:
        return HttpResponseBadRequest("No feedback questions configured.")

    question_ids = [q.id for q in ordered_questions]

    # ---------------------------------------------------------
    # Category spans + question numbering + legend
    # ---------------------------------------------------------
    category_spans = []
    question_number_map = {}
    question_legend = []

    cur = ordered_questions[0].category or "General"
    span = 0

    for idx, q in enumerate(ordered_questions, start=1):
        qcat = q.category or "General"

        question_number_map[q.id] = idx
        question_legend.append({
            "number": idx,
            "category": qcat,
            "question_text": q.question_text,
            "question_id": q.id,
        })

        if qcat != cur:
            category_spans.append({"category": cur, "span": span})
            cur = qcat
            span = 1
        else:
            span += 1

    category_spans.append({"category": cur, "span": span})

    # ---------------------------------------------------------
    # PDF setup - LANDSCAPE A3
    # ---------------------------------------------------------
    PDF_PAGE_SIZE = landscape(A3)

    filename = f"Course_Feedbacks_Bulk_{department.Department}_{sel_year}_Sem{sel_sem}.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=PDF_PAGE_SIZE,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=50 * mm,
        bottomMargin=25 * mm,
        title="Course Feedback Bulk Report"
    )

    styles = getSampleStyleSheet()

    info_style = ParagraphStyle(
        "info",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#0f172a"),
        alignment=0
    )

    heading_style = ParagraphStyle(
        "heading_style",
        parent=styles["Heading4"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=6
    )

    sub_heading_style = ParagraphStyle(
        "sub_heading_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#111827"),
        spaceAfter=6
    )

    p_center = ParagraphStyle(
        "p_center",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=5.2,
        leading=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a")
    )

    p_center_bold = ParagraphStyle(
        "p_center_bold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=5.4,
        leading=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a")
    )

    p_summary_course = ParagraphStyle(
        "p_summary_course",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=6.2,
        leading=8,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#0f172a"),
        wordWrap="CJK"
    )

    p_small_left = ParagraphStyle(
        "p_small_left",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=6.2,
        leading=8.5,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#0f172a"),
        wordWrap="CJK"
    )

    remarks_style = ParagraphStyle(
        "remarks_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=12,
        alignment=0,
        textColor=colors.HexColor("#0f172a")
    )

    legend_category_style = ParagraphStyle(
        "legend_category_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=4,
        spaceBefore=6
    )

    legend_text_style = ParagraphStyle(
        "legend_text_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=12,
        textColor=colors.HexColor("#111827")
    )

    summary_title_style = ParagraphStyle(
        "summary_title_style",
        parent=styles["Heading4"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=6
    )

    semester_type = "Odd" if str(sel_sem) in ["1", "3", "5", "7"] else "Even"
    academic_year_text = f"Academic Year: {get_academic_year()} ({semester_type} Semester)"
    subtitle = academic_year_text

    story = []

    # ---------------------------------------------------------
    # Collect data for Summary First Page
    # ---------------------------------------------------------
    course_summary_rows = []

    for course in courses:
        assign_qs = AssignSubjectFaculty.objects.select_related("faculty").filter(
            department=department,
            course_id=course.id,
            is_active=True,
            course__year=str(sel_year),
            course__semester=str(sel_sem)
        )

        if sel_batch:
            assign_qs = assign_qs.filter(batch=sel_batch)
        if sel_section:
            assign_qs = assign_qs.filter(section=sel_section)

        assign_obj = assign_qs.first()
        filtered_faculty_name = (
            assign_obj.faculty.name
            if assign_obj and assign_obj.faculty
            else faculty.name
        )

        enroll_qs = CourseEnrollment.objects.select_related("student").filter(
            department=department,
            course_id=course.id,
            enroll=True,
            student__year=str(sel_year),
            student__semester=str(sel_sem),
            student__is_discontinued=False,
        )

        if sel_batch:
            enroll_qs = enroll_qs.filter(batch=sel_batch)
        if sel_section:
            enroll_qs = enroll_qs.filter(section=sel_section)

        students = [
            e.student for e in enroll_qs.order_by("student__reg_no")
            if e.student and not e.student.is_discontinued
        ]

        student_ids = [s.id for s in students]
        student_count = len(students)

        submissions_qs = FeedbackSubmission.objects.filter(
            student_id__in=student_ids,
            enrollment__course_id=course.id,
            department=department,
            enrollment__department=department,
            student__is_discontinued=False,
        ).select_related("student")

        submission_ids = list(submissions_qs.values_list("id", flat=True))
        actual_responses = len(submission_ids)

        course_question_totals = defaultdict(int)

        if submission_ids:
            answers = FeedbackAnswer.objects.filter(
                submission_id__in=submission_ids,
                question_id__in=question_ids,
                submission__student__is_discontinued=False,
            )

            for ans in answers:
                course_question_totals[ans.question_id] += int(ans.score or 0)

        course_grand_total = sum(course_question_totals.values())

        course_question_averages = {}

        if actual_responses > 0:
            for q in ordered_questions:
                course_question_averages[q.id] = round(
                    course_question_totals.get(q.id, 0) / actual_responses,
                    2
                )

            course_grand_average = round(course_grand_total / actual_responses, 2)
        else:
            for q in ordered_questions:
                course_question_averages[q.id] = 0.0

            course_grand_average = 0.0

        course_summary_rows.append({
            "course_code": course.course_code or "-",
            "course_title": course.title or "-",
            "student_count": student_count,
            "response_count": actual_responses,
            "question_averages": course_question_averages,
            "grand_average": course_grand_average,
            "faculty_name": filtered_faculty_name,
            "course_faculty_display": f"{course.course_code or '-'}<br/>{filtered_faculty_name or '-'}",
        })

    # ---------------------------------------------------------
    # 1. Summary Table - FIRST PAGE
    # ---------------------------------------------------------
    story.append(Paragraph("Course-wise Question Average Summary", summary_title_style))
    story.append(Paragraph(f"Department: {department.Department}", info_style))
    story.append(Paragraph(
        f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
        f"Section: {sel_section or '-'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or '-'}",
        info_style
    ))
    story.append(Spacer(1, 12))

    summary_row1 = ["Course Code / Faculty Name"]

    for c in category_spans:
        summary_row1.append(Paragraph(str(c["category"]), p_center))
        for _ in range(c["span"] - 1):
            summary_row1.append("")

    summary_row1.extend(["Avg Total", "Responses", "Enrolled"])

    summary_row2 = [""]

    for q in ordered_questions:
        summary_row2.append(Paragraph(str(question_number_map.get(q.id, "")), p_center_bold))

    summary_row2.extend(["", "", ""])

    summary_data = [summary_row1, summary_row2]

    for item in course_summary_rows:
        row = [Paragraph(item["course_faculty_display"], p_summary_course)]

        for q in ordered_questions:
            row.append(str(item["question_averages"].get(q.id, 0)))

        row.extend([
            str(item["grand_average"]),
            str(item["response_count"]),
            str(item["student_count"]),
        ])

        summary_data.append(row)

    usable_w = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

    summary_course_col_w = 48 * mm
    summary_avg_total_col_w = 18 * mm
    summary_responses_col_w = 18 * mm
    summary_enrolled_col_w = 18 * mm

    q_count = len(ordered_questions)

    summary_fixed_w = (
        summary_course_col_w +
        summary_avg_total_col_w +
        summary_responses_col_w +
        summary_enrolled_col_w
    )

    summary_q_w = max((usable_w - summary_fixed_w) / max(q_count, 1), 7 * mm)

    summary_col_widths = (
        [summary_course_col_w] +
        [summary_q_w] * q_count +
        [summary_avg_total_col_w, summary_responses_col_w, summary_enrolled_col_w]
    )

    summary_tbl = Table(summary_data, colWidths=summary_col_widths, repeatRows=2)

    summary_style = TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d1d5db")),
        ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f8fafc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 1), 6),
        ("FONTNAME", (0, 2), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 2), (-4, -1), "Helvetica"),
        ("FONTNAME", (-3, 2), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 2), (-1, -1), 6),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("BACKGROUND", (-3, 2), (-1, -1), colors.HexColor("#f8fafc")),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ALIGN", (0, 2), (0, -1), "LEFT"),
        ("VALIGN", (0, 2), (0, -1), "MIDDLE"),
    ])

    summary_style.add("SPAN", (0, 0), (0, 1))
    summary_style.add("SPAN", (-3, 0), (-3, 1))
    summary_style.add("SPAN", (-2, 0), (-2, 1))
    summary_style.add("SPAN", (-1, 0), (-1, 1))

    start_col = 1

    for c in category_spans:
        end_col = start_col + c["span"] - 1
        summary_style.add("SPAN", (start_col, 0), (end_col, 0))
        start_col = end_col + 1

    summary_tbl.setStyle(summary_style)

    story.append(summary_tbl)
    story.append(Spacer(1, 15))
    story.append(PageBreak())

    # ---------------------------------------------------------
    # Individual Course Pages
    # ---------------------------------------------------------
    for course_index, course in enumerate(courses, start=1):
        item = course_summary_rows[course_index - 1]
        filtered_faculty_name = item["faculty_name"]

        enroll_qs = CourseEnrollment.objects.select_related("student").filter(
            department=department,
            course_id=course.id,
            enroll=True,
            student__year=str(sel_year),
            student__semester=str(sel_sem),
            student__is_discontinued=False,
        )

        if sel_batch:
            enroll_qs = enroll_qs.filter(batch=sel_batch)
        if sel_section:
            enroll_qs = enroll_qs.filter(section=sel_section)

        students = [
            e.student for e in enroll_qs.order_by("student__reg_no")
            if e.student and not e.student.is_discontinued
        ]

        student_ids = [s.id for s in students]
        student_count = len(students)

        submissions_qs = FeedbackSubmission.objects.filter(
            student_id__in=student_ids,
            enrollment__course_id=course.id,
            department=department,
            enrollment__department=department,
            student__is_discontinued=False,
        ).select_related("student")

        submission_ids = list(submissions_qs.values_list("id", flat=True))
        submission_map = {sub.student_id: sub for sub in submissions_qs}

        student_q_marks = defaultdict(dict)
        student_total = defaultdict(int)

        if submission_ids:
            answers = FeedbackAnswer.objects.filter(
                submission_id__in=submission_ids,
                question_id__in=question_ids,
                submission__student__is_discontinued=False,
            ).select_related("submission")

            for ans in answers:
                sid = ans.submission.student_id
                qid = ans.question_id
                sc = int(ans.score or 0)

                student_q_marks[sid][qid] = sc
                student_total[sid] += sc

        story.append(Paragraph(f"Course {course_index}: {course.course_code} - {course.title}", heading_style))
        story.append(Paragraph(f"Department: {department.Department}", info_style))
        story.append(Paragraph(f"Name of the Faculty: {filtered_faculty_name}", info_style))
        story.append(Paragraph(
            f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
            f"Section: {sel_section or '-'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or '-'}<br/>"
            f"Students Enrolled: {student_count} | Responses Received: {item['response_count']} "
            f"({round((item['response_count'] / student_count * 100), 1) if student_count > 0 else 0}%)",
            info_style
        ))
        story.append(Spacer(1, 8))

        row1 = ["Sl.No"]

        for c in category_spans:
            row1.append(Paragraph(str(c["category"]), p_center))
            for _ in range(c["span"] - 1):
                row1.append("")

        row1.extend([
            Paragraph("Total", p_center),
            Paragraph("Overall Effectiveness (%)", p_center),
            Paragraph("Student Satisfaction", p_center),
            Paragraph("Recommendation to Continue / Improve", p_center),
            Paragraph("Open Comments for Improvement", p_center),
        ])

        row2 = [""]

        for q in ordered_questions:
            row2.append(Paragraph(str(question_number_map.get(q.id, "")), p_center_bold))

        row2.extend(["", "", "", "", ""])

        data = [row1, row2]

        for i, st in enumerate(students, start=1):
            submission = submission_map.get(st.id)

            overall_effectiveness = (
                str(submission.overall_effectiveness_percentage)
                if submission and submission.overall_effectiveness_percentage is not None
                else "-"
            )

            student_satisfaction = (
                "Yes" if submission and submission.student_satisfaction is True
                else "No" if submission and submission.student_satisfaction is False
                else "-"
            )

            recommendation = (
                submission.recommendation_to_continue_improve
                if submission and submission.recommendation_to_continue_improve
                else "-"
            )

            open_comments = (
                submission.open_comments_for_improvement
                if submission and submission.open_comments_for_improvement
                else "-"
            )

            r = [str(i)]

            for q in ordered_questions:
                r.append(str(student_q_marks.get(st.id, {}).get(q.id, 0)))

            r.extend([
                str(student_total.get(st.id, 0)),
                overall_effectiveness,
                student_satisfaction,
                Paragraph(recommendation, p_small_left),
                Paragraph(open_comments, p_small_left),
            ])

            data.append(r)

        if not students:
            no_data_row = ["No students found."] + [""] * (len(ordered_questions) + 5)
            data.append(no_data_row)

        usable_w = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

        col_w_sno = 11 * mm
        col_w_total = 15 * mm
        col_w_eff = 23 * mm
        col_w_sat = 23 * mm
        col_w_rec = 48 * mm
        col_w_open = 55 * mm

        q_count = len(ordered_questions)

        fixed_width = (
            col_w_sno +
            col_w_total +
            col_w_eff +
            col_w_sat +
            col_w_rec +
            col_w_open
        )

        q_w = max((usable_w - fixed_width) / max(q_count, 1), 7.5 * mm)

        col_widths = (
            [col_w_sno] +
            [q_w] * q_count +
            [col_w_total, col_w_eff, col_w_sat, col_w_rec, col_w_open]
        )

        tbl = Table(data, colWidths=col_widths, repeatRows=2)

        table_style = TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f8fafc")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, 1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 1), 5.8),
            ("FONTNAME", (0, 2), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 2), (-5, -1), 5.8),
            ("ALIGN", (0, 2), (-5, -1), "CENTER"),
            ("ALIGN", (-4, 2), (-1, -1), "LEFT"),
            ("VALIGN", (-4, 2), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f8fafc")),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])

        table_style.add("SPAN", (0, 0), (0, 1))
        table_style.add("SPAN", (q_count + 1, 0), (q_count + 1, 1))
        table_style.add("SPAN", (q_count + 2, 0), (q_count + 2, 1))
        table_style.add("SPAN", (q_count + 3, 0), (q_count + 3, 1))
        table_style.add("SPAN", (q_count + 4, 0), (q_count + 4, 1))
        table_style.add("SPAN", (q_count + 5, 0), (q_count + 5, 1))

        start_col = 1

        for c in category_spans:
            end_col = start_col + c["span"] - 1
            table_style.add("SPAN", (start_col, 0), (end_col, 0))
            start_col = end_col + 1

        if not students:
            table_style.add("SPAN", (0, 2), (-1, 2))
            table_style.add("ALIGN", (0, 2), (-1, 2), "CENTER")
            table_style.add("FONTNAME", (0, 2), (-1, 2), "Helvetica-Oblique")

        tbl.setStyle(table_style)
        story.append(tbl)

        # ---------------------------------------------------------
        # Remarks Page
        # ---------------------------------------------------------
        story.append(PageBreak())
        story.append(Paragraph(f"Course {course_index}: {course.course_code} - {course.title}", heading_style))
        story.append(Paragraph("Remarks / Action Taken", sub_heading_style))
        story.append(Paragraph(f"Department: {department.Department}", info_style))
        story.append(Paragraph(f"Name of the Faculty: {filtered_faculty_name}", info_style))
        story.append(Paragraph(
            f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
            f"Section: {sel_section or '-'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or '-'}",
            info_style
        ))
        story.append(Spacer(1, 8))

        all_feedback_notes = CourseFeedbackRemark.objects.select_related("faculty").filter(
            department=department,
            course=course
        ).order_by("-updated_at", "-id")

        remarks_box_width = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

        if all_feedback_notes.exists():
            for note_index, note in enumerate(all_feedback_notes, start=1):
                note_faculty_name = note.faculty.name if note.faculty else "Unknown User"

                note_data = [
                    [Paragraph(f"<b>{note_index}. Faculty:</b> {note_faculty_name}", remarks_style)],
                    [Paragraph(f"<b>Remarks:</b> {note.remarks if note.remarks else '-'}", remarks_style)],
                    [Paragraph(f"<b>Action Plan:</b> {note.action_taken if note.action_taken else '-'}", remarks_style)],
                ]

                note_tbl = Table(note_data, colWidths=[remarks_box_width])

                note_tbl.setStyle(TableStyle([
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]))

                story.append(note_tbl)
                story.append(Spacer(1, 6))
        else:
            empty_tbl = Table(
                [
                    [Paragraph("<b>Remarks:</b> -", remarks_style)],
                    [Paragraph("<b>Action Plan:</b> -", remarks_style)]
                ],
                colWidths=[remarks_box_width]
            )

            empty_tbl.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            story.append(empty_tbl)

        if course_index != len(courses):
            story.append(PageBreak())

    # ---------------------------------------------------------
    # Final Legend Page
    # ---------------------------------------------------------
    story.append(PageBreak())
    story.append(Paragraph("Question Reference / Legend", heading_style))
    story.append(Paragraph(f"Department: {department.Department}", info_style))
    story.append(Paragraph(
        f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
        f"Section: {sel_section or '-'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or '-'}",
        info_style
    ))
    story.append(Spacer(1, 8))

    legend_grouped = defaultdict(list)

    for item in question_legend:
        legend_grouped[item["category"]].append(item)

    legend_usable_w = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

    for cat in legend_grouped:
        story.append(Paragraph(str(cat), legend_category_style))

        legend_data = [["No.", "Question"]]

        for item in legend_grouped[cat]:
            legend_data.append([
                str(item["number"]),
                Paragraph(item["question_text"], legend_text_style)
            ])

        legend_tbl = Table(
            legend_data,
            colWidths=[18 * mm, legend_usable_w - 18 * mm],
            repeatRows=1
        )

        legend_tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8.5),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 1), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8.3),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))

        story.append(legend_tbl)
        story.append(Spacer(1, 8))

    def _on_page(canv, doc_):
        page_w, page_h = doc_.pagesize
        _draw_rit_header_footer(
            canv,
            page_w,
            page_h,
            title="COURSE FEEDBACKS",
            subtitle=subtitle
        )

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    return response













import os
from datetime import datetime
from collections import defaultdict

from django.conf import settings
from django.http import HttpResponse, HttpResponseBadRequest
from django.shortcuts import render
from django.contrib.staticfiles import finders

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

from faculty_management.models import general_information
from user_accounts.models import StudentDetails
from course_management.models import Course, CourseEnrollment, AssignSubjectFaculty
from feedback_management.models import FeedbackQuestion, FeedbackSubmission, FeedbackAnswer



# -----------------------------------------------------------------------------------
# ✅ HEADER / FOOTER
# -----------------------------------------------------------------------------------
def _draw_rit_header_footer(c: canvas.Canvas, page_w, page_h, title="STUDENTS FEEDBACK", subtitle=""):
    left_margin = 18 * mm
    right_margin = page_w - 18 * mm
    top_margin = page_h - 12 * mm
    bottom_margin = 14 * mm

    # --- Left Logo (RIT) ---
    logo_rel = "images/ritlogo.png"
    logo_path = finders.find(logo_rel)
    if not logo_path:
        for d in getattr(settings, "STATICFILES_DIRS", []):
            cand = os.path.join(d, logo_rel)
            if os.path.exists(cand):
                logo_path = cand
                break

    if logo_path and os.path.exists(logo_path):
        try:
            img = ImageReader(logo_path)
            iw, ih = img.getSize()
            th = 18 * mm
            tw = th * (iw / float(ih))
            c.drawImage(img, left_margin, top_margin - th + 2 * mm, width=tw, height=th, mask='auto')
        except Exception:
            pass

    # --- Right Logo (TUV) ---
    tuv_logo_rel = "images/tuvlogo.png"
    tuv_path = finders.find(tuv_logo_rel)
    if tuv_path and os.path.exists(tuv_path):
        try:
            img2 = ImageReader(tuv_path)
            iw, ih = img2.getSize()
            th = 13 * mm
            tw = th * (iw / float(ih))
            c.drawImage(img2, right_margin - tw, top_margin - th + 4 * mm, width=tw, height=th, mask='auto')
        except Exception:
            pass

    # --- Institute header ---
    c.setFillColor(colors.HexColor("#2C3E50"))
    c.setFont("Helvetica-Bold", 15)
    c.drawCentredString(page_w / 2.0, top_margin, "RAMCO INSTITUTE OF TECHNOLOGY")

    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(page_w / 2.0, top_margin - 7 * mm, title)

    if subtitle:
        c.setFont("Helvetica", 10)
        c.drawCentredString(page_w / 2.0, top_margin - 13 * mm, subtitle)

    # Line
    c.setStrokeColor(colors.black)
    c.setLineWidth(0.5)
    c.line(left_margin, top_margin - 16 * mm, right_margin, top_margin - 16 * mm)

    # Footer
    footer_y = bottom_margin
    c.setStrokeColor(colors.HexColor("#C0392B"))
    c.setLineWidth(0.8)
    c.line(left_margin, footer_y + 7 * mm, right_margin, footer_y + 7 * mm)

    c.setFont("Helvetica", 8.8)
    c.setFillColor(colors.HexColor("#2C3E50"))
    c.drawCentredString(
        page_w / 2.0,
        footer_y + 2.5 * mm,
        "North Venganallur, Ayyanarkovil Road, Rajapalayam - 626 117, Virudhunagar District, Tamil Nadu."
    )
    c.drawCentredString(
        page_w / 2.0,
        footer_y - 2.0 * mm,
        "Tel: 04563 233400 | E-mail: rit@ritrjpm.ac.in | Web: www.ritrjpm.ac.in"
    )


# -----------------------------------------------------------------------------------
# ✅ SUBJECT FEEDBACK VIEW
# -----------------------------------------------------------------------------------
from collections import defaultdict
import os

from django.contrib import messages
from django.shortcuts import render, redirect
from django.core.files.storage import default_storage

import os
from collections import defaultdict

from django.contrib import messages
from django.db.models import Q
from django.shortcuts import render
from user_accounts.decorators import check_permission

from faculty_management.models import general_information
from user_accounts.models import StudentDetails
from course_management.models import Course, CourseEnrollment, AssignSubjectFaculty
from feedback_management.models import (
    FeedbackQuestion,
    FeedbackSubmission,
    FeedbackAnswer,
    CourseFeedbackRemark,
)


import os
from collections import defaultdict

from django.contrib import messages
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render


import os
from collections import defaultdict

from django.contrib import messages
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render


from collections import defaultdict
import os

from django.contrib import messages
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render

@check_permission("subject_feedback")
def subject_feedback(request):
    # ---------------------------------------------------------
    # FACULTY
    # ---------------------------------------------------------
    user_employee_id = request.user.Employee_id
    faculty = get_object_or_404(
        general_information.objects.select_related("department"),
        faculty_id=user_employee_id
    )
    department = faculty.department

    # ---------------------------------------------------------
    # ASSIGNED SUBJECTS (same source as course_end_survey_entry)
    # ---------------------------------------------------------
    assigned_subjects = (
        AssignSubjectFaculty.objects
        .select_related("course", "department", "regulation", "faculty", "skilled_faculty")
        .filter(
            Q(faculty=faculty) | Q(skilled_faculty=faculty),
            is_active=True,
            course__isnull=False
        )
        .order_by(
            "academic_year",
            "course__year",
            "course__semester",
            "course__course_code",
            "section",
            "batch",
            "id"
        )
    )

    # ---------------------------------------------------------
    # FILE UPLOAD POST
    # ---------------------------------------------------------
    if request.method == "POST" and request.FILES.get("report_file") is not None:
        remark_id = (request.POST.get("remark_id") or "").strip()
        current_batch = (request.POST.get("sel_batch") or "").strip()
        current_year = (request.POST.get("sel_year") or "").strip()
        current_sem = (request.POST.get("sel_sem") or "").strip()
        current_section = (request.POST.get("sel_section") or "").strip()
        current_assign_id = (request.POST.get("sel_assign_id") or "").strip()

        uploaded_file = request.FILES.get("report_file")

        try:
            remark_obj = CourseFeedbackRemark.objects.get(
                id=remark_id,
                department=department
            )
        except CourseFeedbackRemark.DoesNotExist:
            messages.error(request, "Invalid remark selected.")
            return redirect(
                f"{request.path}?batch={current_batch}&year={current_year}&semester={current_sem}&section={current_section}&assign_id={current_assign_id}"
            )

        if not uploaded_file:
            messages.error(request, "Please choose a file to upload.")
            return redirect(
                f"{request.path}?batch={current_batch}&year={current_year}&semester={current_sem}&section={current_section}&assign_id={current_assign_id}"
            )

        allowed_ext = [".pdf", ".doc", ".docx"]
        ext = os.path.splitext(uploaded_file.name)[1].lower()

        if ext not in allowed_ext:
            messages.error(request, "Only PDF, DOC, and DOCX files are allowed.")
            return redirect(
                f"{request.path}?batch={current_batch}&year={current_year}&semester={current_sem}&section={current_section}&assign_id={current_assign_id}"
            )

        if uploaded_file.size > 10 * 1024 * 1024:
            messages.error(request, "File size must be less than 10 MB.")
            return redirect(
                f"{request.path}?batch={current_batch}&year={current_year}&semester={current_sem}&section={current_section}&assign_id={current_assign_id}"
            )

        if remark_obj.report_file:
            remark_obj.report_file.delete(save=False)

        remark_obj.report_file = uploaded_file
        remark_obj.save()

        messages.success(request, "Report uploaded successfully.")
        return redirect(
            f"{request.path}?batch={current_batch}&year={current_year}&semester={current_sem}&section={current_section}&assign_id={current_assign_id}"
        )

    # ---------------------------------------------------------
    # FILTERS
    # ---------------------------------------------------------
    sel_batch = (request.GET.get("batch") or "").strip()
    sel_year = (request.GET.get("year") or "").strip()
    sel_sem = (request.GET.get("semester") or "").strip()
    sel_section = (request.GET.get("section") or "").strip()
    sel_assign_id = (request.GET.get("assign_id") or "").strip()

    filters_applied = bool(sel_year and sel_sem)

    # ---------------------------------------------------------
    # DROPDOWNS FROM ASSIGNMENTS
    # ---------------------------------------------------------
    batches = (
        assigned_subjects.exclude(batch__isnull=True)
        .exclude(batch__exact="")
        .values_list("batch", flat=True)
        .distinct()
        .order_by("batch")
    )

    years = (
        assigned_subjects.exclude(course__year__isnull=True)
        .exclude(course__year__exact="")
        .values_list("course__year", flat=True)
        .distinct()
        .order_by("course__year")
    )

    semesters = (
        assigned_subjects.exclude(course__semester__isnull=True)
        .exclude(course__semester__exact="")
        .values_list("course__semester", flat=True)
        .distinct()
        .order_by("course__semester")
    )

    sections = (
        assigned_subjects.exclude(section__isnull=True)
        .exclude(section__exact="")
        .values_list("section", flat=True)
        .distinct()
        .order_by("section")
    )

    # ---------------------------------------------------------
    # FILTERED ASSIGNMENTS
    # ---------------------------------------------------------
    filtered_assignments = assigned_subjects

    if sel_batch:
        filtered_assignments = filtered_assignments.filter(batch=sel_batch)

    if sel_section:
        filtered_assignments = filtered_assignments.filter(section=sel_section)

    if sel_year:
        filtered_assignments = filtered_assignments.filter(course__year=str(sel_year))

    if sel_sem:
        filtered_assignments = filtered_assignments.filter(course__semester=str(sel_sem))

    course_assignments = list(filtered_assignments)

    # auto select first assignment after year + semester
    if filters_applied and not sel_assign_id and course_assignments:
        sel_assign_id = str(course_assignments[0].id)

    valid_assign_ids = {str(a.id) for a in course_assignments}
    if sel_assign_id and sel_assign_id not in valid_assign_ids:
        sel_assign_id = ""

    # ---------------------------------------------------------
    # SELECTED ASSIGNMENT
    # ---------------------------------------------------------
    selected_assignment = None
    selected_course = None
    selected_course_id = ""
    selected_course_faculty_name = faculty.name

    if sel_assign_id:
        selected_assignment = filtered_assignments.filter(id=sel_assign_id).first()

        if selected_assignment:
            selected_course = selected_assignment.course
            selected_course_id = str(selected_assignment.course_id)
            if selected_assignment.faculty:
                selected_course_faculty_name = selected_assignment.faculty.name

    # ---------------------------------------------------------
    # STUDENTS FOR SELECTED ASSIGNMENT
    # ---------------------------------------------------------
    if selected_assignment:
        enroll_qs = (
            CourseEnrollment.objects
            .select_related("student")
            .filter(
                department=selected_assignment.department,
                course_id=selected_assignment.course_id,
                enroll=True,
            )
        )

        if selected_assignment.batch:
            enroll_qs = enroll_qs.filter(batch=selected_assignment.batch)

        if selected_assignment.section:
            enroll_qs = enroll_qs.filter(section=selected_assignment.section)

        if selected_assignment.course and selected_assignment.course.year:
            enroll_qs = enroll_qs.filter(student__year=str(selected_assignment.course.year))

        if selected_assignment.course and selected_assignment.course.semester:
            enroll_qs = enroll_qs.filter(student__semester=str(selected_assignment.course.semester))

        students = [e.student for e in enroll_qs.order_by("student__reg_no") if e.student]
    else:
        students = []

    student_ids = [s.id for s in students]

    # ---------------------------------------------------------
    # QUESTIONS
    # ---------------------------------------------------------
    ordered_questions = list(
        FeedbackQuestion.objects
        .filter(Q(department__isnull=True) | Q(department=department))
        .order_by("category", "id")
    )
    question_ids = [q.id for q in ordered_questions]

    for q in ordered_questions:
        q.scope_label = "Common" if q.department_id is None else "Department"

    # ---------------------------------------------------------
    # CATEGORY SPANS
    # ---------------------------------------------------------
    category_spans = []
    if ordered_questions:
        cur = ordered_questions[0].category or "General"
        span = 0
        for q in ordered_questions:
            qcat = q.category or "General"
            if qcat != cur:
                category_spans.append({"category": cur, "span": span})
                cur = qcat
                span = 1
            else:
                span += 1
        category_spans.append({"category": cur, "span": span})

    # ---------------------------------------------------------
    # SUBMISSIONS
    # ---------------------------------------------------------
    submission_ids = []
    submissions_qs = FeedbackSubmission.objects.none()
    submission_map = {}

    if selected_assignment and student_ids:
        submissions_qs = (
            FeedbackSubmission.objects
            .filter(
                student_id__in=student_ids,
                enrollment__course_id=selected_assignment.course_id,
                faculty=selected_assignment.faculty,
                department=selected_assignment.department,
                enrollment__department=selected_assignment.department,
            )
            .select_related("student", "course", "faculty", "enrollment")
        )

        if selected_assignment.batch:
            submissions_qs = submissions_qs.filter(enrollment__batch=selected_assignment.batch)

        if selected_assignment.section:
            submissions_qs = submissions_qs.filter(enrollment__section=selected_assignment.section)

        submission_ids = list(submissions_qs.values_list("id", flat=True))
        submission_map = {sub.student_id: sub for sub in submissions_qs}

    # ---------------------------------------------------------
    # ANSWERS
    # ---------------------------------------------------------
    student_q_marks = defaultdict(dict)
    student_total = defaultdict(int)
    submitted_student_ids = set()

    if submission_ids:
        answers = (
            FeedbackAnswer.objects
            .filter(submission_id__in=submission_ids, question_id__in=question_ids)
            .select_related("submission")
        )

        for ans in answers:
            sid = ans.submission.student_id
            qid = ans.question_id
            sc = int(ans.score or 0)

            submitted_student_ids.add(sid)
            student_q_marks[sid][qid] = sc
            student_total[sid] += sc

    # ---------------------------------------------------------
    # TABLE ROWS
    # ---------------------------------------------------------
    rows = []
    for idx, st in enumerate(students, start=1):
        submission = submission_map.get(st.id)

        rows.append({
            "sno": idx,
            "reg_no": st.reg_no or "-",
            "student_name": getattr(st, "student_name", "") or getattr(st, "name", "") or "-",
            "marks": student_q_marks.get(st.id, {}),
            "total": student_total.get(st.id, 0),
            "is_submitted": st.id in submitted_student_ids,
            "overall_effectiveness_percentage": (
                submission.overall_effectiveness_percentage
                if submission and submission.overall_effectiveness_percentage is not None
                else "-"
            ),
            "student_satisfaction": (
                "Yes" if submission and submission.student_satisfaction is True
                else "No" if submission and submission.student_satisfaction is False
                else "-"
            ),
            "recommendation_to_continue_improve": (
                submission.recommendation_to_continue_improve
                if submission and submission.recommendation_to_continue_improve
                else "-"
            ),
            "open_comments_for_improvement": (
                submission.open_comments_for_improvement
                if submission and submission.open_comments_for_improvement
                else "-"
            ),
        })

    # ---------------------------------------------------------
    # REMARKS / ACTIONS
    # ---------------------------------------------------------
    all_feedback_notes = []
    if selected_assignment:
        all_feedback_notes = (
            CourseFeedbackRemark.objects
            .select_related("faculty")
            .filter(
                department=selected_assignment.department,
                course_id=selected_assignment.course_id
            )
            .order_by("-updated_at", "-id")
        )

    return render(
        request,
        "feedback_management/faculty/feedback/subject_feedback.html",
        {
            "faculty": faculty,
            "department": department,

            "batches": batches,
            "years": years,
            "semesters": semesters,
            "sections": sections,

            "sel_batch": sel_batch,
            "sel_year": sel_year,
            "sel_sem": sel_sem,
            "sel_section": sel_section,
            "sel_assign_id": sel_assign_id,

            "filters_applied": filters_applied,
            "course_assignments": course_assignments,

            "selected_assignment": selected_assignment,
            "selected_course": selected_course,
            "selected_course_id": selected_course_id,
            "selected_course_faculty_name": selected_course_faculty_name,

            "total_students": len(students),
            "submitted_students": len(submitted_student_ids),
            "not_submitted_students": len(students) - len(submitted_student_ids),

            "category_spans": category_spans,
            "ordered_questions": ordered_questions,
            "rows": rows,

            "all_feedback_notes": all_feedback_notes,
        }
    )





# -----------------------------------------------------------------------------------
# ✅ SUBJECT FEEDBACK PDF
# -----------------------------------------------------------------------------------
from collections import defaultdict
from datetime import datetime
import os

from django.conf import settings
from django.contrib.staticfiles import finders
from django.http import HttpResponse, HttpResponseBadRequest
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle




from collections import defaultdict
from datetime import datetime

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from collections import defaultdict
from datetime import datetime

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


@check_permission("subject_feedback")
def subject_feedback_pdf(request):
    """
    PDF for subject_feedback page.
    Uses SAME assignment-based logic as subject_feedback page.
    Fixes section/batch data issue by using selected assignment context.
    """

    user = request.user.Employee_id
    login_faculty = get_object_or_404(
        general_information.objects.select_related("department"),
        faculty_id=user
    )
    department = login_faculty.department

    sel_batch = (request.GET.get("batch") or "").strip()
    sel_year = (request.GET.get("year") or "").strip()
    sel_sem = (request.GET.get("semester") or "").strip()
    sel_section = (request.GET.get("section") or "").strip()
    sel_course_id = (request.GET.get("course_id") or "").strip()
    sel_assign_id = (request.GET.get("assign_id") or "").strip()

    filters_applied = bool(sel_year and sel_sem)

    assign_qs = (
        AssignSubjectFaculty.objects
        .select_related("course", "department", "regulation", "faculty")
        .filter(
            faculty=login_faculty,
            is_active=True,
            course__isnull=False,
            course__is_active=True,
        )
        .order_by(
            "academic_year",
            "course__year",
            "course__semester",
            "course__course_code",
            "section",
            "batch",
            "id"
        )
    )

    if sel_batch:
        assign_qs = assign_qs.filter(batch=sel_batch)

    if sel_section:
        assign_qs = assign_qs.filter(section=sel_section)

    if sel_year:
        assign_qs = assign_qs.filter(course__year=str(sel_year))

    if sel_sem:
        assign_qs = assign_qs.filter(course__semester=str(sel_sem))

    course_assignments = list(assign_qs)

    if filters_applied and not sel_assign_id and course_assignments:
        if sel_course_id:
            matched_assign = None
            for a in course_assignments:
                if str(a.course_id) == str(sel_course_id):
                    matched_assign = a
                    break

            if matched_assign:
                sel_assign_id = str(matched_assign.id)
            else:
                sel_assign_id = str(course_assignments[0].id)
        else:
            sel_assign_id = str(course_assignments[0].id)

    valid_assign_ids = {str(a.id) for a in course_assignments}

    if sel_assign_id and sel_assign_id not in valid_assign_ids:
        sel_assign_id = ""

    selected_assignment = None

    if sel_assign_id:
        selected_assignment = assign_qs.filter(id=sel_assign_id).first()

    if not selected_assignment and sel_course_id:
        selected_assignment = assign_qs.filter(course_id=sel_course_id).first()

    if not selected_assignment:
        return HttpResponseBadRequest("Select Year, Semester and Course to generate PDF.")

    course = selected_assignment.course

    if not course or not course.is_active:
        return HttpResponseBadRequest("Invalid course.")

    selected_department = selected_assignment.department
    selected_faculty = selected_assignment.faculty or login_faculty

    subject_faculty_name = selected_faculty.name if selected_faculty else login_faculty.name

    selected_year = str(course.year or sel_year or "")
    selected_sem = str(course.semester or sel_sem or "")
    selected_section = str(selected_assignment.section or sel_section or "")
    selected_batch = str(selected_assignment.batch or sel_batch or "")

    enroll_qs = (
        CourseEnrollment.objects
        .select_related("student")
        .filter(
            department=selected_department,
            course_id=selected_assignment.course_id,
            enroll=True,
        )
    )

    if selected_batch:
        enroll_qs = enroll_qs.filter(batch=selected_batch)

    if selected_section:
        enroll_qs = enroll_qs.filter(section=selected_section)

    if selected_year:
        enroll_qs = enroll_qs.filter(student__year=str(selected_year))

    if selected_sem:
        enroll_qs = enroll_qs.filter(student__semester=str(selected_sem))

    students = [e.student for e in enroll_qs.order_by("student__reg_no") if e.student]
    student_ids = [s.id for s in students]

    ordered_questions = list(
        FeedbackQuestion.objects
        .filter(Q(department__isnull=True) | Q(department=selected_department))
        .order_by("category", "id")
    )

    if not ordered_questions:
        return HttpResponseBadRequest("No feedback questions configured.")

    question_ids = [q.id for q in ordered_questions]

    category_spans = []
    question_number_map = {}
    question_legend = []

    cur = ordered_questions[0].category or "General"
    span = 0

    for idx, q in enumerate(ordered_questions, start=1):
        qcat = q.category or "General"

        question_number_map[q.id] = idx
        question_legend.append({
            "number": idx,
            "category": qcat,
            "question_text": q.question_text,
            "question_id": q.id,
        })

        if qcat != cur:
            category_spans.append({"category": cur, "span": span})
            cur = qcat
            span = 1
        else:
            span += 1

    category_spans.append({"category": cur, "span": span})

    submission_ids = []
    submissions_qs = FeedbackSubmission.objects.none()
    submission_map = {}

    if student_ids:
        submissions_qs = (
            FeedbackSubmission.objects
            .filter(
                student_id__in=student_ids,
                enrollment__course_id=selected_assignment.course_id,
                faculty=selected_faculty,
                department=selected_department,
                enrollment__department=selected_department,
            )
            .select_related("student", "course", "faculty", "enrollment")
        )

        if selected_batch:
            submissions_qs = submissions_qs.filter(enrollment__batch=selected_batch)

        if selected_section:
            submissions_qs = submissions_qs.filter(enrollment__section=selected_section)

        submission_ids = list(submissions_qs.values_list("id", flat=True))
        submission_map = {sub.student_id: sub for sub in submissions_qs}

    student_q_marks = defaultdict(dict)
    student_total = defaultdict(int)

    if submission_ids:
        answers = (
            FeedbackAnswer.objects
            .filter(
                submission_id__in=submission_ids,
                question_id__in=question_ids
            )
            .select_related("submission")
        )

        for ans in answers:
            sid = ans.submission.student_id
            qid = ans.question_id
            sc = int(ans.score or 0)

            student_q_marks[sid][qid] = sc
            student_total[sid] += sc

    all_feedback_notes = (
        CourseFeedbackRemark.objects
        .select_related("faculty")
        .filter(
            department=selected_department,
            course=course
        )
        .order_by("-updated_at", "-id")
    )

    filename = f"SubjectFeedback_{course.course_code}_{selected_year}_Sem{selected_sem}.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'

    PDF_PAGE_SIZE = landscape(A3)

    semester_type = "Odd" if str(selected_sem) in ["1", "3", "5", "7"] else "Even"
    subtitle = f"Academic Year: {get_academic_year()} ({semester_type} Semester)"

    doc = SimpleDocTemplate(
        response,
        pagesize=PDF_PAGE_SIZE,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=50 * mm,
        bottomMargin=20 * mm,
        title="Students Feedback"
    )

    styles = getSampleStyleSheet()

    info_style = ParagraphStyle(
        "info",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#0f172a"),
        alignment=0,
    )

    heading_style = ParagraphStyle(
        "heading_style",
        parent=styles["Heading4"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=6,
    )

    sub_heading_style = ParagraphStyle(
        "sub_heading_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#111827"),
        spaceAfter=6,
    )

    p_center = ParagraphStyle(
        "p_center",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=5.2,
        leading=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a"),
    )

    p_center_bold = ParagraphStyle(
        "p_center_bold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=5.4,
        leading=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a"),
    )

    p_small_left = ParagraphStyle(
        "p_small_left",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=5.8,
        leading=7,
        alignment=0,
        textColor=colors.HexColor("#0f172a"),
        wordWrap="CJK",
    )

    remarks_style = ParagraphStyle(
        "remarks_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=12,
        alignment=0,
        textColor=colors.HexColor("#0f172a"),
    )

    legend_category_style = ParagraphStyle(
        "legend_category_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=4,
        spaceBefore=6,
    )

    legend_text_style = ParagraphStyle(
        "legend_text_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=12,
        textColor=colors.HexColor("#111827"),
    )

    story = []

    story.append(Paragraph(f"Name of the Subject: {course.course_code} - {course.title}", info_style))
    story.append(Paragraph(f"Name of the Faculty: {subject_faculty_name}", info_style))
    story.append(
        Paragraph(
            f"Year/Semester: {selected_year or '-'} / {selected_sem or '-'} "
            f"&nbsp;&nbsp;&nbsp; Section: {selected_section or '-'} "
            f"&nbsp;&nbsp;&nbsp; Batch: {selected_batch or '-'}",
            info_style
        )
    )
    story.append(Spacer(1, 8))

    row1 = ["Sl.No"]

    for c in category_spans:
        row1.append(Paragraph(str(c["category"]), p_center))
        for _ in range(c["span"] - 1):
            row1.append("")

    row1.extend([
        "Total",
        Paragraph("Overall Effectiveness (%)", p_center),
        Paragraph("Student Satisfaction", p_center),
        Paragraph("Recommendation to Continue / Improve", p_center),
        Paragraph("Open Comments for Improvement", p_center),
    ])

    row2 = [""]

    for q in ordered_questions:
        row2.append(Paragraph(str(question_number_map.get(q.id, "")), p_center_bold))

    row2.extend(["", "", "", "", ""])

    data = [row1, row2]

    for i, st in enumerate(students, start=1):
        submission = submission_map.get(st.id)

        overall_effectiveness = (
            str(submission.overall_effectiveness_percentage)
            if submission and submission.overall_effectiveness_percentage is not None
            else "-"
        )

        student_satisfaction = (
            "Yes" if submission and submission.student_satisfaction is True
            else "No" if submission and submission.student_satisfaction is False
            else "-"
        )

        recommendation = (
            submission.recommendation_to_continue_improve
            if submission and submission.recommendation_to_continue_improve
            else "-"
        )

        open_comments = (
            submission.open_comments_for_improvement
            if submission and submission.open_comments_for_improvement
            else "-"
        )

        row = [str(i)]

        for q in ordered_questions:
            row.append(str(student_q_marks.get(st.id, {}).get(q.id, 0)))

        row.extend([
            str(student_total.get(st.id, 0)),
            overall_effectiveness,
            student_satisfaction,
            Paragraph(recommendation, p_small_left),
            Paragraph(open_comments, p_small_left),
        ])

        data.append(row)

    if not students:
        no_data_row = ["No students found."]
        no_data_row += [""] * len(ordered_questions)
        no_data_row += ["", "", "", "", ""]
        data.append(no_data_row)

    page_w, page_h = doc.pagesize
    usable_w = page_w - doc.leftMargin - doc.rightMargin

    col_w_sno = 10 * mm
    col_w_total = 12 * mm
    col_w_eff = 16 * mm
    col_w_sat = 18 * mm
    col_w_rec = 32 * mm
    col_w_open = 32 * mm

    q_count = len(ordered_questions)
    fixed_width = col_w_sno + col_w_total + col_w_eff + col_w_sat + col_w_rec + col_w_open

    q_w = (usable_w - fixed_width) / float(max(q_count, 1))
    q_w = max(q_w, 6 * mm)

    col_widths = (
        [col_w_sno]
        + [q_w] * q_count
        + [col_w_total, col_w_eff, col_w_sat, col_w_rec, col_w_open]
    )

    tbl = Table(data, colWidths=col_widths, repeatRows=2)

    ts = TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d1d5db")),
        ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f8fafc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 1), 5.5),
        ("FONTNAME", (0, 2), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 2), (-3, -1), 5.5),
        ("ALIGN", (0, 2), (-3, -1), "CENTER"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f8fafc")),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#111827")),

        ("ALIGN", (-2, 2), (-1, -1), "LEFT"),
        ("VALIGN", (-2, 2), (-1, -1), "TOP"),
        ("LEFTPADDING", (-2, 2), (-1, -1), 3),
        ("RIGHTPADDING", (-2, 2), (-1, -1), 3),

        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])

    ts.add("SPAN", (0, 0), (0, 1))
    ts.add("SPAN", (q_count + 1, 0), (q_count + 1, 1))
    ts.add("SPAN", (q_count + 2, 0), (q_count + 2, 1))
    ts.add("SPAN", (q_count + 3, 0), (q_count + 3, 1))
    ts.add("SPAN", (q_count + 4, 0), (q_count + 4, 1))
    ts.add("SPAN", (q_count + 5, 0), (q_count + 5, 1))

    start_col = 1

    for c in category_spans:
        end_col = start_col + c["span"] - 1
        ts.add("SPAN", (start_col, 0), (end_col, 0))
        start_col = end_col + 1

    ts.add("BACKGROUND", (q_count + 1, 2), (q_count + 1, -1), colors.HexColor("#f8fafc"))
    ts.add("FONTNAME", (q_count + 1, 2), (q_count + 1, -1), "Helvetica-Bold")

    if not students:
        ts.add("SPAN", (0, 2), (-1, 2))
        ts.add("ALIGN", (0, 2), (-1, 2), "CENTER")
        ts.add("FONTNAME", (0, 2), (-1, 2), "Helvetica-Oblique")

    tbl.setStyle(ts)
    story.append(tbl)

    story.append(PageBreak())
    story.append(Paragraph(f"{course.course_code} - {course.title}", heading_style))
    story.append(Paragraph("Remarks / Action Taken", sub_heading_style))
    story.append(Paragraph(f"Course Faculty: {subject_faculty_name}", info_style))
    story.append(
        Paragraph(
            f"Year/Semester: {selected_year or '-'} / {selected_sem or '-'} "
            f"&nbsp;&nbsp;&nbsp; Section: {selected_section or '-'} "
            f"&nbsp;&nbsp;&nbsp; Batch: {selected_batch or '-'}",
            info_style
        )
    )
    story.append(Spacer(1, 8))

    remarks_box_width = page_w - doc.leftMargin - doc.rightMargin

    if all_feedback_notes.exists():
        for note_index, note in enumerate(all_feedback_notes, start=1):
            entered_faculty_name = note.faculty.name if note.faculty else "Unknown User"

            note_data = [
                [Paragraph(f"<b>{note_index}. Entered By (Faculty):</b> {entered_faculty_name}", remarks_style)],
                [Paragraph(f"<b>Remarks:</b> {note.remarks if note.remarks else '-'}", remarks_style)],
                [Paragraph(f"<b>Action Plan:</b> {note.action_taken if note.action_taken else '-'}", remarks_style)],
            ]

            note_tbl = Table(note_data, colWidths=[remarks_box_width])
            note_tbl.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            story.append(note_tbl)
            story.append(Spacer(1, 6))
    else:
        empty_tbl = Table(
            [
                [Paragraph("<b>Remarks:</b> -", remarks_style)],
                [Paragraph("<b>Action Plan:</b> -", remarks_style)],
            ],
            colWidths=[remarks_box_width]
        )

        empty_tbl.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))

        story.append(empty_tbl)

    story.append(PageBreak())
    story.append(Paragraph("Question Reference / Legend", heading_style))
    story.append(Paragraph(f"Name of the Subject: {course.course_code} - {course.title}", info_style))
    story.append(Paragraph(f"Name of the Faculty: {subject_faculty_name}", info_style))
    story.append(
        Paragraph(
            f"Year/Semester: {selected_year or '-'} / {selected_sem or '-'} "
            f"&nbsp;&nbsp;&nbsp; Section: {selected_section or '-'} "
            f"&nbsp;&nbsp;&nbsp; Batch: {selected_batch or '-'}",
            info_style
        )
    )
    story.append(Spacer(1, 8))

    legend_grouped = defaultdict(list)

    for item in question_legend:
        legend_grouped[item["category"]].append(item)

    legend_usable_w = page_w - doc.leftMargin - doc.rightMargin

    for cat in legend_grouped:
        story.append(Paragraph(str(cat), legend_category_style))

        legend_data = [["No.", "Question"]]

        for item in legend_grouped[cat]:
            legend_data.append([
                str(item["number"]),
                Paragraph(item["question_text"], legend_text_style)
            ])

        legend_tbl = Table(
            legend_data,
            colWidths=[18 * mm, legend_usable_w - (18 * mm)],
            repeatRows=1
        )

        legend_tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8.5),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 1), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8.3),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ]))

        story.append(legend_tbl)
        story.append(Spacer(1, 8))

    def _on_page(canv, doc_):
        pw, ph = doc_.pagesize
        _draw_rit_header_footer(
            canv,
            pw,
            ph,
            title="STUDENTS FEEDBACK",
            subtitle=subtitle
        )

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    return response





def get_current_semester_values():
    """
    Returns list of semester values for current semester cycle.

    Assumption:
    - June to November  => Odd semester cycle
    - December to May   => Even semester cycle
    """
    today = date.today()

    if 6 <= today.month <= 11:
        return ["1", "3", "5", "7"]
    return ["2", "4", "6", "8"]


def get_current_semester_label():
    """
    For display purpose only.
    """
    today = date.today()
    return "Odd" if 6 <= today.month <= 11 else "Even"

def get_current_academic_year():
    """
    Returns current academic year string like '2025-2026'.

    Change ACADEMIC_YEAR_START_MONTH if your institution starts
    the academic year in a different month.
    """
    ACADEMIC_YEAR_START_MONTH = 6  # June
    today = date.today()

    if today.month >= ACADEMIC_YEAR_START_MONTH:
        start_year = today.year
        end_year = today.year + 1
    else:
        start_year = today.year - 1
        end_year = today.year

    return f"{start_year}-{end_year}"






@check_permission("subject_feedback")
def subject_feedback_bulk_pdf(request):
    """
    Bulk PDF for subject_feedback page.
    Uses same filter logic as subject_feedback page:
    batch / year / semester / section / assign_id.
    """

    user_employee_id = request.user.Employee_id
    faculty = get_object_or_404(
        general_information.objects.select_related("department"),
        faculty_id=user_employee_id
    )
    department = faculty.department

    current_academic_year = get_academic_year()
    current_semester_label = get_current_semester_label()

    sel_batch = (request.GET.get("batch") or "").strip()
    sel_year = (request.GET.get("year") or "").strip()
    sel_sem = (request.GET.get("semester") or "").strip()
    sel_section = (request.GET.get("section") or "").strip()
    sel_assign_id = (request.GET.get("assign_id") or "").strip()

    assigned_subjects = (
        AssignSubjectFaculty.objects
        .select_related("course", "department", "regulation", "faculty")
        .filter(
            faculty=faculty,
            is_active=True,
            course__isnull=False,
            course__is_active=True,
        )
        .order_by(
            "academic_year",
            "course__year",
            "course__semester",
            "course__course_code",
            "section",
            "batch",
            "id"
        )
    )

    filtered_assignments = assigned_subjects

    if sel_batch:
        filtered_assignments = filtered_assignments.filter(batch=sel_batch)

    if sel_section:
        filtered_assignments = filtered_assignments.filter(section=sel_section)

    if sel_year:
        filtered_assignments = filtered_assignments.filter(course__year=str(sel_year))

    if sel_sem:
        filtered_assignments = filtered_assignments.filter(course__semester=str(sel_sem))

    if sel_assign_id:
        filtered_assignments = filtered_assignments.filter(id=sel_assign_id)

    filtered_assignments = list(filtered_assignments)

    if not filtered_assignments:
        return HttpResponseBadRequest(
            "No faculty-handled active subjects found for the selected filters."
        )

    ordered_questions = list(
        FeedbackQuestion.objects
        .filter(Q(department__isnull=True) | Q(department=department))
        .order_by("category", "id")
    )

    if not ordered_questions:
        return HttpResponseBadRequest("No feedback questions configured.")

    question_ids = [q.id for q in ordered_questions]

    category_spans = []
    question_number_map = {}
    question_legend = []

    cur = ordered_questions[0].category or "General"
    span = 0

    for idx, q in enumerate(ordered_questions, start=1):
        qcat = q.category or "General"
        question_number_map[q.id] = idx
        question_legend.append({
            "number": idx,
            "category": qcat,
            "question_text": q.question_text,
            "question_id": q.id,
        })

        if qcat != cur:
            category_spans.append({"category": cur, "span": span})
            cur = qcat
            span = 1
        else:
            span += 1

    category_spans.append({"category": cur, "span": span})

    PDF_PAGE_SIZE = landscape(A3)

    filename = f"Subject_Feedback_Bulk_{faculty.name}_{current_academic_year}.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=PDF_PAGE_SIZE,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=50 * mm,
        bottomMargin=25 * mm,
        title="Subject Feedback Bulk Report"
    )

    styles = getSampleStyleSheet()

    info_style = ParagraphStyle(
        "info",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#0f172a"),
        alignment=0
    )

    heading_style = ParagraphStyle(
        "heading_style",
        parent=styles["Heading4"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=6
    )

    sub_heading_style = ParagraphStyle(
        "sub_heading_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#111827"),
        spaceAfter=6
    )

    p_center = ParagraphStyle(
        "p_center",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=5.2,
        leading=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a")
    )

    p_center_bold = ParagraphStyle(
        "p_center_bold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=5.4,
        leading=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a")
    )

    p_summary_course = ParagraphStyle(
        "p_summary_course",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=6.0,
        leading=8,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#0f172a"),
        wordWrap="CJK"
    )

    p_small_left = ParagraphStyle(
        "p_small_left",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=6.2,
        leading=8.5,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#0f172a"),
        wordWrap="CJK"
    )

    remarks_style = ParagraphStyle(
        "remarks_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=12,
        alignment=0,
        textColor=colors.HexColor("#0f172a")
    )

    legend_category_style = ParagraphStyle(
        "legend_category_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=4,
        spaceBefore=6
    )

    legend_text_style = ParagraphStyle(
        "legend_text_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=12,
        textColor=colors.HexColor("#111827")
    )

    summary_title_style = ParagraphStyle(
        "summary_title_style",
        parent=styles["Heading4"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=6
    )

    subtitle = f"Academic Year: {current_academic_year} ({current_semester_label} Semester)"
    story = []

    handled_entries = []
    seen_keys = set()

    for assign in filtered_assignments:
        if not assign.course:
            continue

        unique_key = (
            str(assign.id),
            str(assign.course_id or ""),
            str(assign.course.year or ""),
            str(assign.course.semester or ""),
            str(assign.section or ""),
            str(assign.batch or ""),
            str(assign.faculty_id or ""),
        )

        if unique_key in seen_keys:
            continue

        seen_keys.add(unique_key)

        handled_entries.append({
            "assign": assign,
            "assign_id": assign.id,
            "course": assign.course,
            "faculty": assign.faculty,
            "faculty_name": assign.faculty.name if assign.faculty else faculty.name,
            "department": assign.department,
            "section": assign.section or "",
            "batch": assign.batch or "",
            "year": str(assign.course.year or ""),
            "semester": str(assign.course.semester or ""),
        })

    handled_entries = sorted(
        handled_entries,
        key=lambda x: (
            x["year"],
            x["semester"],
            x["course"].course_code or "",
            x["section"],
            x["batch"],
            x["assign_id"]
        )
    )

    if not handled_entries:
        return HttpResponseBadRequest(
            "No handled subject entries found for the selected filters."
        )

    course_summary_rows = []

    for entry in handled_entries:
        assign = entry["assign"]
        course = entry["course"]
        entry_department = entry["department"]
        entry_faculty = entry["faculty"]
        filtered_faculty_name = entry["faculty_name"]
        entry_year = entry["year"]
        entry_sem = entry["semester"]
        entry_section = entry["section"]
        entry_batch = entry["batch"]

        enroll_qs = (
            CourseEnrollment.objects
            .select_related("student")
            .filter(
                department=entry_department,
                course_id=course.id,
                enroll=True,
            )
        )

        if entry_batch:
            enroll_qs = enroll_qs.filter(batch=entry_batch)

        if entry_section:
            enroll_qs = enroll_qs.filter(section=entry_section)

        if entry_year:
            enroll_qs = enroll_qs.filter(student__year=str(entry_year))

        if entry_sem:
            enroll_qs = enroll_qs.filter(student__semester=str(entry_sem))

        students = [e.student for e in enroll_qs.order_by("student__reg_no") if e.student]
        student_ids = [s.id for s in students]
        student_count = len(students)

        submissions_qs = FeedbackSubmission.objects.none()

        if student_ids:
            submissions_qs = (
                FeedbackSubmission.objects
                .filter(
                    student_id__in=student_ids,
                    enrollment__course_id=course.id,
                    faculty=entry_faculty,
                    department=entry_department,
                    enrollment__department=entry_department,
                )
                .select_related("student", "course", "faculty", "enrollment")
            )

            if entry_batch:
                submissions_qs = submissions_qs.filter(enrollment__batch=entry_batch)

            if entry_section:
                submissions_qs = submissions_qs.filter(enrollment__section=entry_section)

        submission_ids = list(submissions_qs.values_list("id", flat=True))
        actual_responses = len(submission_ids)

        course_question_totals = defaultdict(int)

        if submission_ids:
            answers = FeedbackAnswer.objects.filter(
                submission_id__in=submission_ids,
                question_id__in=question_ids
            )

            for ans in answers:
                course_question_totals[ans.question_id] += int(ans.score or 0)

        course_grand_total = sum(course_question_totals.values())

        course_question_averages = {}

        if actual_responses > 0:
            for q in ordered_questions:
                course_question_averages[q.id] = round(
                    course_question_totals.get(q.id, 0) / actual_responses,
                    2
                )
            course_grand_average = round(course_grand_total / actual_responses, 2)
        else:
            for q in ordered_questions:
                course_question_averages[q.id] = 0.0
            course_grand_average = 0.0

        course_summary_rows.append({
            "entry": entry,
            "course_code": course.course_code or "-",
            "course_title": course.title or "-",
            "student_count": student_count,
            "response_count": actual_responses,
            "question_averages": course_question_averages,
            "grand_average": course_grand_average,
            "faculty_name": filtered_faculty_name,
            "course_faculty_display": (
                f"{course.course_code or '-'}<br/>"
                f"{filtered_faculty_name or '-'}<br/>"
                f"Year: {entry_year or '-'} | Sem: {entry_sem or '-'}<br/>"
                f"Sec: {entry_section or '-'} | Batch: {entry_batch or '-'}"
            ),
        })

    story.append(Paragraph("Course-wise Question Average Summary", summary_title_style))
    story.append(Paragraph(
        f"Department: {department.department_name if hasattr(department, 'department_name') else department}",
        info_style
    ))
    story.append(Paragraph(f"Faculty: {faculty.name}", info_style))
    story.append(Paragraph(f"Academic Year: {current_academic_year}", info_style))
    story.append(Paragraph(f"Selected Year: {sel_year or 'All'}", info_style))
    story.append(Paragraph(f"Selected Semester: {sel_sem or 'All'}", info_style))
    story.append(Paragraph(
        f"Section: {sel_section or 'All'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or 'All'}",
        info_style
    ))
    story.append(Spacer(1, 12))

    summary_row1 = ["Course / Faculty / Year / Sem / Sec / Batch"]

    for c in category_spans:
        summary_row1.append(Paragraph(str(c["category"]), p_center))
        for _ in range(c["span"] - 1):
            summary_row1.append("")

    summary_row1.extend(["Avg Total", "Responses", "Enrolled"])

    summary_row2 = [""]

    for q in ordered_questions:
        summary_row2.append(Paragraph(str(question_number_map.get(q.id, "")), p_center_bold))

    summary_row2.extend(["", "", ""])

    summary_data = [summary_row1, summary_row2]

    for item in course_summary_rows:
        row = [Paragraph(item["course_faculty_display"], p_summary_course)]

        for q in ordered_questions:
            row.append(str(item["question_averages"].get(q.id, 0)))

        row.extend([
            str(item["grand_average"]),
            str(item["response_count"]),
            str(item["student_count"]),
        ])

        summary_data.append(row)

    usable_w = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

    summary_course_col_w = 62 * mm
    summary_avg_total_col_w = 18 * mm
    summary_responses_col_w = 18 * mm
    summary_enrolled_col_w = 18 * mm

    q_count = len(ordered_questions)
    summary_fixed_w = (
        summary_course_col_w
        + summary_avg_total_col_w
        + summary_responses_col_w
        + summary_enrolled_col_w
    )
    summary_q_w = max((usable_w - summary_fixed_w) / max(q_count, 1), 7 * mm)

    summary_col_widths = [summary_course_col_w] + [summary_q_w] * q_count + [
        summary_avg_total_col_w,
        summary_responses_col_w,
        summary_enrolled_col_w
    ]

    summary_tbl = Table(summary_data, colWidths=summary_col_widths, repeatRows=2)

    summary_style = TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d1d5db")),
        ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f8fafc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 1), 6),
        ("FONTNAME", (0, 2), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 2), (-4, -1), "Helvetica"),
        ("FONTNAME", (-3, 2), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 2), (-1, -1), 6),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("BACKGROUND", (-3, 2), (-1, -1), colors.HexColor("#f8fafc")),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ALIGN", (0, 2), (0, -1), "LEFT"),
        ("VALIGN", (0, 2), (0, -1), "MIDDLE"),
    ])

    summary_style.add("SPAN", (0, 0), (0, 1))
    summary_style.add("SPAN", (-3, 0), (-3, 1))
    summary_style.add("SPAN", (-2, 0), (-2, 1))
    summary_style.add("SPAN", (-1, 0), (-1, 1))

    start_col = 1

    for c in category_spans:
        end_col = start_col + c["span"] - 1
        summary_style.add("SPAN", (start_col, 0), (end_col, 0))
        start_col = end_col + 1

    summary_tbl.setStyle(summary_style)
    story.append(summary_tbl)
    story.append(Spacer(1, 15))
    story.append(PageBreak())

    for course_index, item in enumerate(course_summary_rows, start=1):
        entry = item["entry"]
        assign = entry["assign"]
        course = entry["course"]
        entry_department = entry["department"]
        entry_faculty = entry["faculty"]
        filtered_faculty_name = item["faculty_name"]
        entry_year = entry["year"]
        entry_sem = entry["semester"]
        entry_section = entry["section"]
        entry_batch = entry["batch"]

        enroll_qs = (
            CourseEnrollment.objects
            .select_related("student")
            .filter(
                department=entry_department,
                course_id=course.id,
                enroll=True,
            )
        )

        if entry_batch:
            enroll_qs = enroll_qs.filter(batch=entry_batch)

        if entry_section:
            enroll_qs = enroll_qs.filter(section=entry_section)

        if entry_year:
            enroll_qs = enroll_qs.filter(student__year=str(entry_year))

        if entry_sem:
            enroll_qs = enroll_qs.filter(student__semester=str(entry_sem))

        students = [e.student for e in enroll_qs.order_by("student__reg_no") if e.student]
        student_ids = [s.id for s in students]
        student_count = len(students)

        submissions_qs = FeedbackSubmission.objects.none()

        if student_ids:
            submissions_qs = (
                FeedbackSubmission.objects
                .filter(
                    student_id__in=student_ids,
                    enrollment__course_id=course.id,
                    faculty=entry_faculty,
                    department=entry_department,
                    enrollment__department=entry_department,
                )
                .select_related("student", "course", "faculty", "enrollment")
            )

            if entry_batch:
                submissions_qs = submissions_qs.filter(enrollment__batch=entry_batch)

            if entry_section:
                submissions_qs = submissions_qs.filter(enrollment__section=entry_section)

        submission_ids = list(submissions_qs.values_list("id", flat=True))
        submission_map = {sub.student_id: sub for sub in submissions_qs}

        student_q_marks = defaultdict(dict)
        student_total = defaultdict(int)

        if submission_ids:
            answers = (
                FeedbackAnswer.objects
                .filter(
                    submission_id__in=submission_ids,
                    question_id__in=question_ids
                )
                .select_related("submission")
            )

            for ans in answers:
                sid = ans.submission.student_id
                qid = ans.question_id
                sc = int(ans.score or 0)

                student_q_marks[sid][qid] = sc
                student_total[sid] += sc

        story.append(Paragraph(
            f"Course {course_index}: {course.course_code} - {course.title}",
            heading_style
        ))
        story.append(Paragraph(
            f"Department: {entry_department.department_name if hasattr(entry_department, 'department_name') else entry_department}",
            info_style
        ))
        story.append(Paragraph(f"Name of the Faculty: {filtered_faculty_name}", info_style))
        story.append(Paragraph(f"Academic Year: {assign.academic_year or current_academic_year}", info_style))
        story.append(Paragraph(
            f"Year/Semester: {entry_year or '-'} / {entry_sem or '-'} &nbsp;&nbsp;&nbsp; "
            f"Section: {entry_section or '-'} &nbsp;&nbsp;&nbsp; Batch: {entry_batch or '-'}<br/>"
            f"Students Enrolled: {student_count} | Responses Received: {item['response_count']} "
            f"({round((item['response_count'] / student_count * 100), 1) if student_count > 0 else 0}%)",
            info_style
        ))
        story.append(Spacer(1, 8))

        row1 = ["Sl.No"]

        for c in category_spans:
            row1.append(Paragraph(str(c["category"]), p_center))
            for _ in range(c["span"] - 1):
                row1.append("")

        row1.extend([
            Paragraph("Total", p_center),
            Paragraph("Overall Effectiveness (%)", p_center),
            Paragraph("Student Satisfaction", p_center),
            Paragraph("Recommendation to Continue / Improve", p_center),
            Paragraph("Open Comments for Improvement", p_center),
        ])

        row2 = [""]

        for q in ordered_questions:
            row2.append(Paragraph(str(question_number_map.get(q.id, "")), p_center_bold))

        row2.extend(["", "", "", "", ""])

        data = [row1, row2]

        for i, st in enumerate(students, start=1):
            submission = submission_map.get(st.id)

            overall_effectiveness = (
                str(submission.overall_effectiveness_percentage)
                if submission and submission.overall_effectiveness_percentage is not None
                else "-"
            )

            student_satisfaction = (
                "Yes" if submission and submission.student_satisfaction is True
                else "No" if submission and submission.student_satisfaction is False
                else "-"
            )

            recommendation = (
                submission.recommendation_to_continue_improve
                if submission and submission.recommendation_to_continue_improve
                else "-"
            )

            open_comments = (
                submission.open_comments_for_improvement
                if submission and submission.open_comments_for_improvement
                else "-"
            )

            r = [str(i)]

            for q in ordered_questions:
                r.append(str(student_q_marks.get(st.id, {}).get(q.id, 0)))

            r.extend([
                str(student_total.get(st.id, 0)),
                overall_effectiveness,
                student_satisfaction,
                Paragraph(recommendation, p_small_left),
                Paragraph(open_comments, p_small_left),
            ])

            data.append(r)

        if not students:
            no_data_row = ["No students found."] + [""] * (len(ordered_questions) + 5)
            data.append(no_data_row)

        usable_w = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

        col_w_sno = 11 * mm
        col_w_total = 15 * mm
        col_w_eff = 23 * mm
        col_w_sat = 23 * mm
        col_w_rec = 48 * mm
        col_w_open = 55 * mm

        q_count = len(ordered_questions)

        fixed_width = (
            col_w_sno
            + col_w_total
            + col_w_eff
            + col_w_sat
            + col_w_rec
            + col_w_open
        )

        q_w = max((usable_w - fixed_width) / max(q_count, 1), 7.5 * mm)

        col_widths = [col_w_sno] + [q_w] * q_count + [
            col_w_total,
            col_w_eff,
            col_w_sat,
            col_w_rec,
            col_w_open
        ]

        tbl = Table(data, colWidths=col_widths, repeatRows=2)

        table_style = TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f8fafc")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, 1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 1), 5.8),
            ("FONTNAME", (0, 2), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 2), (-5, -1), 5.8),
            ("ALIGN", (0, 2), (-5, -1), "CENTER"),
            ("ALIGN", (-4, 2), (-1, -1), "LEFT"),
            ("VALIGN", (-4, 2), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f8fafc")),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])

        table_style.add("SPAN", (0, 0), (0, 1))
        table_style.add("SPAN", (q_count + 1, 0), (q_count + 1, 1))
        table_style.add("SPAN", (q_count + 2, 0), (q_count + 2, 1))
        table_style.add("SPAN", (q_count + 3, 0), (q_count + 3, 1))
        table_style.add("SPAN", (q_count + 4, 0), (q_count + 4, 1))
        table_style.add("SPAN", (q_count + 5, 0), (q_count + 5, 1))

        start_col = 1

        for c in category_spans:
            end_col = start_col + c["span"] - 1
            table_style.add("SPAN", (start_col, 0), (end_col, 0))
            start_col = end_col + 1

        tbl.setStyle(table_style)
        story.append(tbl)

        story.append(PageBreak())
        story.append(Paragraph(
            f"Course {course_index}: {course.course_code} - {course.title}",
            heading_style
        ))
        story.append(Paragraph("Remarks / Action Taken", sub_heading_style))
        story.append(Paragraph(
            f"Department: {entry_department.department_name if hasattr(entry_department, 'department_name') else entry_department}",
            info_style
        ))
        story.append(Paragraph(f"Name of the Faculty: {filtered_faculty_name}", info_style))
        story.append(Paragraph(f"Academic Year: {assign.academic_year or current_academic_year}", info_style))
        story.append(Paragraph(
            f"Year/Semester: {entry_year or '-'} / {entry_sem or '-'} &nbsp;&nbsp;&nbsp; "
            f"Section: {entry_section or '-'} &nbsp;&nbsp;&nbsp; Batch: {entry_batch or '-'}",
            info_style
        ))
        story.append(Spacer(1, 8))

        all_feedback_notes = (
            CourseFeedbackRemark.objects
            .select_related("faculty")
            .filter(
                department=entry_department,
                course_id=course.id
            )
            .order_by("-updated_at", "-id")
        )

        remarks_box_width = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

        if all_feedback_notes.exists():
            for note_index, note in enumerate(all_feedback_notes, start=1):
                note_faculty_name = note.faculty.name if note.faculty else "Unknown User"

                note_data = [
                    [Paragraph(f"<b>{note_index}. Faculty:</b> {note_faculty_name}", remarks_style)],
                    [Paragraph(f"<b>Remarks:</b> {note.remarks if note.remarks else '-'}", remarks_style)],
                    [Paragraph(f"<b>Action Plan:</b> {note.action_taken if note.action_taken else '-'}", remarks_style)],
                ]

                note_tbl = Table(note_data, colWidths=[remarks_box_width])
                note_tbl.setStyle(TableStyle([
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]))

                story.append(note_tbl)
                story.append(Spacer(1, 6))
        else:
            empty_tbl = Table(
                [
                    [Paragraph("<b>Remarks:</b> -", remarks_style)],
                    [Paragraph("<b>Action Plan:</b> -", remarks_style)]
                ],
                colWidths=[remarks_box_width]
            )

            empty_tbl.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            story.append(empty_tbl)

        if course_index != len(course_summary_rows):
            story.append(PageBreak())

    story.append(PageBreak())
    story.append(Paragraph("Question Reference / Legend", heading_style))
    story.append(Paragraph(
        f"Department: {department.department_name if hasattr(department, 'department_name') else department}",
        info_style
    ))
    story.append(Paragraph(f"Faculty: {faculty.name}", info_style))
    story.append(Paragraph(f"Academic Year: {current_academic_year}", info_style))
    story.append(Paragraph(f"Selected Year: {sel_year or 'All'}", info_style))
    story.append(Paragraph(f"Selected Semester: {sel_sem or 'All'}", info_style))
    story.append(Paragraph(
        f"Section: {sel_section or 'All'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or 'All'}",
        info_style
    ))
    story.append(Spacer(1, 8))

    legend_grouped = defaultdict(list)

    for item in question_legend:
        legend_grouped[item["category"]].append(item)

    legend_usable_w = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

    for cat in legend_grouped:
        story.append(Paragraph(str(cat), legend_category_style))

        legend_data = [["No.", "Question"]]

        for item in legend_grouped[cat]:
            legend_data.append([
                str(item["number"]),
                Paragraph(item["question_text"], legend_text_style)
            ])

        legend_tbl = Table(
            legend_data,
            colWidths=[18 * mm, legend_usable_w - 18 * mm],
            repeatRows=1
        )

        legend_tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8.5),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 1), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8.3),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))

        story.append(legend_tbl)
        story.append(Spacer(1, 8))

    def _on_page(canv, doc_):
        page_w, page_h = doc_.pagesize
        _draw_rit_header_footer(
            canv,
            page_w,
            page_h,
            title="SUBJECT FEEDBACKS",
            subtitle=subtitle
        )

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    return response







from collections import defaultdict
from datetime import datetime

from django.contrib import messages
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.shortcuts import render, redirect
from django.urls import reverse

from faculty_management.models import general_information
from user_accounts.models import StudentDetails, Add_Department
from course_management.models import Course, CourseEnrollment
from feedback_management.models import (
    FeedbackQuestion,
    FeedbackSubmission,
    FeedbackAnswer,
    CourseFeedbackRemark,
    feedback_data_Permission,
)


from collections import defaultdict
from django.contrib import messages
from django.http import HttpResponseForbidden
from django.shortcuts import render, redirect
from django.urls import reverse

from faculty_management.models import general_information
from user_accounts.models import StudentDetails, Add_Department
from course_management.models import Course, CourseEnrollment
from feedback_management.models import (
    FeedbackQuestion,
    FeedbackSubmission,
    FeedbackAnswer,
    CourseFeedbackRemark,
    feedback_data_Permission,
)

from collections import defaultdict
import os

from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponseForbidden
from django.shortcuts import render, redirect
from django.urls import reverse


def _get_feedback_permission_scope(request):
    scope = {
        "has_access": False,
        "can_view_all": False,
        "can_view_department": False,
    }

    role_id = getattr(request.user, "role_id", None)

    if not role_id:
        return scope

    perm = feedback_data_Permission.objects.filter(role_id=role_id).first()
    if not perm:
        return scope

    scope["can_view_all"] = bool(perm.can_view_all_feedback_data)
    scope["can_view_department"] = bool(perm.can_view_department_feedback_data)
    scope["has_access"] = scope["can_view_all"] or scope["can_view_department"]

    return scope


import os
from collections import defaultdict

from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from faculty_management.models import general_information
from user_accounts.models import StudentDetails, Add_Department
from course_management.models import Course, CourseEnrollment
from feedback_management.models import (
    FeedbackQuestion,
    FeedbackSubmission,
    FeedbackAnswer,
    CourseFeedbackRemark,
)
from user_accounts.decorators import check_permission


from collections import defaultdict
from django.db.models import Q
from django.http import HttpResponseForbidden
from django.shortcuts import render

from collections import defaultdict
from django.db.models import Q
from django.http import HttpResponseForbidden
from django.shortcuts import render


@check_permission("view_feedbacks")
def view_feedbacks(request):
    from collections import defaultdict

    from django.db.models import Q
    from django.http import HttpResponseForbidden
    from django.shortcuts import render

    user = request.user.Employee_id

    faculty = (
        general_information.objects
        .select_related("department")
        .get(faculty_id=user)
    )

    faculty_department = faculty.department

    permission_scope = _get_feedback_permission_scope(request)

    if not permission_scope["has_access"]:
        return HttpResponseForbidden(
            "You do not have permission to view feedback data."
        )

    # ---------------------------------------------------------
    # Read request params
    # ---------------------------------------------------------
    if request.method == "POST":
        sel_department_id = (
            request.POST.get("department_id") or ""
        ).strip()

        sel_batch = (
            request.POST.get("batch") or ""
        ).strip()

        sel_year = (
            request.POST.get("year") or ""
        ).strip()

        sel_sem = (
            request.POST.get("semester") or ""
        ).strip()

        sel_section = (
            request.POST.get("section") or ""
        ).strip()

        sel_course_id = (
            request.POST.get("course_id") or ""
        ).strip()

    else:
        sel_department_id = (
            request.GET.get("department_id") or ""
        ).strip()

        sel_batch = (
            request.GET.get("batch") or ""
        ).strip()

        sel_year = (
            request.GET.get("year") or ""
        ).strip()

        sel_sem = (
            request.GET.get("semester") or ""
        ).strip()

        sel_section = (
            request.GET.get("section") or ""
        ).strip()

        sel_course_id = (
            request.GET.get("course_id") or ""
        ).strip()

    # ---------------------------------------------------------
    # Department permission scope
    # ---------------------------------------------------------
    if permission_scope.get("can_view_all"):

        departments = (
            Add_Department.objects
            .filter(is_active=True)
            .order_by("Department")
        )

        if sel_department_id:
            department = (
                Add_Department.objects
                .filter(
                    id=sel_department_id,
                    is_active=True
                )
                .first()
            )
        else:
            department = faculty_department or departments.first()

    else:
        department = faculty_department

        departments = (
            Add_Department.objects.filter(id=faculty_department.id)
            if faculty_department
            else Add_Department.objects.none()
        )

    sel_department_id = str(department.id) if department else ""

    # ---------------------------------------------------------
    # Base students
    #
    # General filtering must exclude discontinued students.
    #
    # A discontinued student is added later only when:
    # 1. The student is enrolled in the selected course.
    # 2. The student has submitted feedback for that course.
    # ---------------------------------------------------------
    if department:
        base_students_qs = StudentDetails.objects.filter(
            department=department,
            is_discontinued=False
        )
    else:
        base_students_qs = StudentDetails.objects.none()

    # ---------------------------------------------------------
    # Filter dropdowns
    # ---------------------------------------------------------
    students_for_filters = base_students_qs

    if sel_batch:
        students_for_filters = students_for_filters.filter(
            batch=sel_batch
        )

    batches = (
        base_students_qs
        .exclude(batch__isnull=True)
        .exclude(batch__exact="")
        .values_list("batch", flat=True)
        .distinct()
        .order_by("batch")
    )

    years = (
        students_for_filters
        .exclude(year__isnull=True)
        .exclude(year__exact="")
        .values_list("year", flat=True)
        .distinct()
        .order_by("year")
    )

    sem_filter_qs = students_for_filters

    if sel_year:
        sem_filter_qs = sem_filter_qs.filter(
            year=str(sel_year)
        )

    semesters = (
        sem_filter_qs
        .exclude(semester__isnull=True)
        .exclude(semester__exact="")
        .values_list("semester", flat=True)
        .distinct()
        .order_by("semester")
    )

    section_filter_qs = sem_filter_qs

    if sel_sem:
        section_filter_qs = section_filter_qs.filter(
            semester=str(sel_sem)
        )

    sections = (
        section_filter_qs
        .exclude(section__isnull=True)
        .exclude(section__exact="")
        .values_list("section", flat=True)
        .distinct()
        .order_by("section")
    )

    filters_applied = bool(
        department and sel_year and sel_sem
    )

    # ---------------------------------------------------------
    # Courses dropdown
    # ---------------------------------------------------------
    if department:
        courses_qs = Course.objects.filter(
            department=department,
            is_active=True
        )
    else:
        courses_qs = Course.objects.none()

    if sel_year:
        courses_qs = courses_qs.filter(
            year=str(sel_year)
        )

    if sel_sem:
        courses_qs = courses_qs.filter(
            semester=str(sel_sem)
        )

    if department and (sel_batch or sel_section):

        # Active student enrollments
        active_enrollment_qs = CourseEnrollment.objects.filter(
            department=department,
            enroll=True,
            student__is_discontinued=False,
        )

        if sel_batch:
            active_enrollment_qs = active_enrollment_qs.filter(
                batch=sel_batch
            )

        if sel_section:
            active_enrollment_qs = active_enrollment_qs.filter(
                section=sel_section
            )

        if sel_year:
            active_enrollment_qs = active_enrollment_qs.filter(
                student__year=str(sel_year)
            )

        if sel_sem:
            active_enrollment_qs = active_enrollment_qs.filter(
                student__semester=str(sel_sem)
            )

        active_course_ids = (
            active_enrollment_qs
            .values_list("course_id", flat=True)
            .distinct()
        )

        # Discontinued students are considered only when they have
        # submitted feedback for an enrolled course.
        discontinued_submission_qs = FeedbackSubmission.objects.filter(
            department=department,
            enrollment__department=department,
            enrollment__enroll=True,
            student__is_discontinued=True,
        )

        if sel_batch:
            discontinued_submission_qs = (
                discontinued_submission_qs.filter(
                    enrollment__batch=sel_batch
                )
            )

        if sel_section:
            discontinued_submission_qs = (
                discontinued_submission_qs.filter(
                    enrollment__section=sel_section
                )
            )

        if sel_year:
            discontinued_submission_qs = (
                discontinued_submission_qs.filter(
                    student__year=str(sel_year)
                )
            )

        if sel_sem:
            discontinued_submission_qs = (
                discontinued_submission_qs.filter(
                    student__semester=str(sel_sem)
                )
            )

        discontinued_feedback_course_ids = (
            discontinued_submission_qs
            .values_list("enrollment__course_id", flat=True)
            .distinct()
        )

        courses_qs = courses_qs.filter(
            Q(id__in=active_course_ids)
            | Q(id__in=discontinued_feedback_course_ids)
        )

    courses = (
        courses_qs
        .values("id", "title", "course_code")
        .order_by("course_code")
    )

    selected_course = None

    if department and sel_course_id:
        selected_course = (
            Course.objects
            .filter(
                id=sel_course_id,
                department=department
            )
            .first()
        )

    # ---------------------------------------------------------
    # Questions
    # ---------------------------------------------------------
    ordered_questions = []
    question_ids = []
    category_spans = []

    if department:
        ordered_questions = list(
            FeedbackQuestion.objects
            .filter(
                Q(department__isnull=True)
                | Q(department=department)
            )
            .order_by("category", "id")
        )

        question_ids = [
            question.id
            for question in ordered_questions
        ]

        if ordered_questions:
            current_category = (
                ordered_questions[0].category or "General"
            )

            span = 0

            for question in ordered_questions:
                question_category = (
                    question.category or "General"
                )

                if question_category != current_category:
                    category_spans.append({
                        "category": current_category,
                        "span": span,
                    })

                    current_category = question_category
                    span = 1

                else:
                    span += 1

            category_spans.append({
                "category": current_category,
                "span": span,
            })

    # ---------------------------------------------------------
    # Dashboard Summary Data
    # ---------------------------------------------------------
    dashboard_overall = {
        "total_courses": 0,
        "total_responses": 0,
        "total_enrolled": 0,
        "overall_avg_total": 0.0,
        "overall_response_percentage": 0.0,
    }

    dashboard_category_averages = []
    dashboard_course_boxes = []

    # ---------------------------------------------------------
    # Course-wise Question Average Summary
    # ---------------------------------------------------------
    summary_rows = []

    if department and sel_year and sel_sem and question_ids:

        summary_courses_qs = Course.objects.filter(
            department=department,
            year=str(sel_year),
            semester=str(sel_sem),
            is_active=True
        )

        if sel_batch or sel_section:

            # Active student enrollments
            active_summary_enrollment_qs = (
                CourseEnrollment.objects.filter(
                    department=department,
                    enroll=True,
                    student__is_discontinued=False,
                    student__year=str(sel_year),
                    student__semester=str(sel_sem)
                )
            )

            if sel_batch:
                active_summary_enrollment_qs = (
                    active_summary_enrollment_qs.filter(
                        batch=sel_batch
                    )
                )

            if sel_section:
                active_summary_enrollment_qs = (
                    active_summary_enrollment_qs.filter(
                        section=sel_section
                    )
                )

            active_summary_course_ids = (
                active_summary_enrollment_qs
                .values_list("course_id", flat=True)
                .distinct()
            )

            # Discontinued students having submitted feedback
            discontinued_summary_submission_qs = (
                FeedbackSubmission.objects.filter(
                    department=department,
                    enrollment__department=department,
                    enrollment__enroll=True,
                    student__is_discontinued=True,
                    student__year=str(sel_year),
                    student__semester=str(sel_sem),
                )
            )

            if sel_batch:
                discontinued_summary_submission_qs = (
                    discontinued_summary_submission_qs.filter(
                        enrollment__batch=sel_batch
                    )
                )

            if sel_section:
                discontinued_summary_submission_qs = (
                    discontinued_summary_submission_qs.filter(
                        enrollment__section=sel_section
                    )
                )

            discontinued_summary_course_ids = (
                discontinued_summary_submission_qs
                .values_list(
                    "enrollment__course_id",
                    flat=True
                )
                .distinct()
            )

            summary_courses_qs = summary_courses_qs.filter(
                Q(id__in=active_summary_course_ids)
                | Q(id__in=discontinued_summary_course_ids)
            )

        summary_courses = list(
            summary_courses_qs.order_by("course_code")
        )

        dashboard_total_of_course_averages = 0.0

        category_total_scores = defaultdict(int)
        category_total_question_slots = defaultdict(int)

        for course in summary_courses:

            # -------------------------------------------------
            # Submitted discontinued student IDs for this course
            # -------------------------------------------------
            discontinued_submitted_student_ids_qs = (
                FeedbackSubmission.objects.filter(
                    department=department,
                    enrollment__department=department,
                    enrollment__course_id=course.id,
                    enrollment__enroll=True,
                    student__is_discontinued=True,
                    student__year=str(sel_year),
                    student__semester=str(sel_sem),
                )
            )

            if sel_batch:
                discontinued_submitted_student_ids_qs = (
                    discontinued_submitted_student_ids_qs.filter(
                        enrollment__batch=sel_batch
                    )
                )

            if sel_section:
                discontinued_submitted_student_ids_qs = (
                    discontinued_submitted_student_ids_qs.filter(
                        enrollment__section=sel_section
                    )
                )

            discontinued_submitted_student_ids = (
                discontinued_submitted_student_ids_qs
                .values_list("student_id", flat=True)
                .distinct()
            )

            # -------------------------------------------------
            # Enrolled students
            #
            # Include:
            # - All active enrolled students
            # - Discontinued enrolled students who submitted
            # -------------------------------------------------
            enroll_qs = CourseEnrollment.objects.filter(
                department=department,
                course_id=course.id,
                enroll=True,
                student__year=str(sel_year),
                student__semester=str(sel_sem)
            ).filter(
                Q(student__is_discontinued=False)
                | Q(
                    student_id__in=
                    discontinued_submitted_student_ids
                )
            )

            if sel_batch:
                enroll_qs = enroll_qs.filter(
                    batch=sel_batch
                )

            if sel_section:
                enroll_qs = enroll_qs.filter(
                    section=sel_section
                )

            student_count = (
                enroll_qs
                .values("student_id")
                .distinct()
                .count()
            )

            included_student_ids = (
                enroll_qs
                .values_list("student_id", flat=True)
                .distinct()
            )

            # -------------------------------------------------
            # Submissions
            #
            # Do not globally exclude discontinued students here.
            # Only students included in the enrollment queryset are
            # allowed, which already applies the required rule.
            # -------------------------------------------------
            submissions_qs = FeedbackSubmission.objects.filter(
                enrollment__course_id=course.id,
                enrollment__enroll=True,
                department=department,
                enrollment__department=department,
                student_id__in=included_student_ids,
            )

            if sel_batch:
                submissions_qs = submissions_qs.filter(
                    enrollment__batch=sel_batch
                )

            if sel_section:
                submissions_qs = submissions_qs.filter(
                    enrollment__section=sel_section
                )

            submissions_qs = submissions_qs.filter(
                student__year=str(sel_year),
                student__semester=str(sel_sem)
            )

            actual_responses = (
                submissions_qs
                .values("student_id")
                .distinct()
                .count()
            )

            course_question_totals = defaultdict(int)
            course_category_scores = defaultdict(int)
            course_category_slots = defaultdict(int)

            if actual_responses > 0:
                answers = (
                    FeedbackAnswer.objects
                    .filter(
                        submission__in=submissions_qs,
                        question_id__in=question_ids,
                    )
                    .select_related("question")
                )

                for answer in answers:
                    score_value = int(
                        answer.score or 0
                    )

                    course_question_totals[
                        answer.question_id
                    ] += score_value

                    question_object = answer.question

                    category_name = (
                        question_object.category or "General"
                        if question_object
                        else "General"
                    )

                    category_total_scores[
                        category_name
                    ] += score_value

                    course_category_scores[
                        category_name
                    ] += score_value

            course_grand_total = sum(
                course_question_totals.values()
            )

            course_question_averages = {}
            course_category_averages = []

            if actual_responses > 0:

                for question in ordered_questions:
                    average_value = round(
                        course_question_totals.get(
                            question.id,
                            0
                        ) / actual_responses,
                        2
                    )

                    course_question_averages[
                        question.id
                    ] = average_value

                    category_name = (
                        question.category or "General"
                    )

                    category_total_question_slots[
                        category_name
                    ] += actual_responses

                    course_category_slots[
                        category_name
                    ] += actual_responses

                grand_average = round(
                    course_grand_total / actual_responses,
                    2
                )

            else:
                for question in ordered_questions:
                    course_question_averages[
                        question.id
                    ] = 0.0

                grand_average = 0.0

            for category_data in category_spans:

                category_name = category_data["category"]

                category_score = (
                    course_category_scores.get(
                        category_name,
                        0
                    )
                )

                category_slots = (
                    course_category_slots.get(
                        category_name,
                        0
                    )
                )

                category_average = (
                    round(
                        category_score / category_slots,
                        2
                    )
                    if category_slots > 0
                    else 0.0
                )

                course_category_averages.append({
                    "category": category_name,
                    "average": category_average,
                })

            summary_rows.append({
                "course_code": course.course_code or "-",
                "course_title": course.title or "-",
                "student_count": student_count,
                "response_count": actual_responses,
                "question_averages": (
                    course_question_averages
                ),
                "grand_average": grand_average,
            })

            dashboard_course_boxes.append({
                "course_code": course.course_code or "-",
                "course_title": course.title or "-",
                "grand_average": grand_average,
                "response_count": actual_responses,
                "student_count": student_count,
                "category_averages": (
                    course_category_averages
                ),
            })

            dashboard_overall["total_courses"] += 1

            dashboard_overall[
                "total_responses"
            ] += actual_responses

            dashboard_overall[
                "total_enrolled"
            ] += student_count

            dashboard_total_of_course_averages += (
                grand_average
            )

        if dashboard_overall["total_courses"] > 0:
            dashboard_overall["overall_avg_total"] = round(
                dashboard_total_of_course_averages
                / dashboard_overall["total_courses"],
                2
            )

        if dashboard_overall["total_enrolled"] > 0:
            dashboard_overall[
                "overall_response_percentage"
            ] = round(
                (
                    dashboard_overall["total_responses"]
                    / dashboard_overall["total_enrolled"]
                ) * 100,
                2
            )

        for category_data in category_spans:

            category_name = category_data["category"]

            total_score = category_total_scores.get(
                category_name,
                0
            )

            total_slots = (
                category_total_question_slots.get(
                    category_name,
                    0
                )
            )

            average_score = (
                round(
                    total_score / total_slots,
                    2
                )
                if total_slots > 0
                else 0.0
            )

            dashboard_category_averages.append({
                "category": category_name,
                "average": average_score,
            })

    # ---------------------------------------------------------
    # Students list
    #
    # Selected course:
    # - Active enrolled students are included.
    # - Discontinued enrolled students are included only when
    #   they submitted feedback.
    #
    # No selected course:
    # - Discontinued students are removed.
    # ---------------------------------------------------------
    if department and sel_course_id:

        # Find discontinued students who actually submitted
        # feedback for the currently selected course.
        discontinued_submitted_qs = (
            FeedbackSubmission.objects
            .filter(
                department=department,
                enrollment__department=department,
                enrollment__course_id=sel_course_id,
                enrollment__enroll=True,
                student__is_discontinued=True,
            )
        )

        if sel_batch:
            discontinued_submitted_qs = (
                discontinued_submitted_qs.filter(
                    enrollment__batch=sel_batch
                )
            )

        if sel_year:
            discontinued_submitted_qs = (
                discontinued_submitted_qs.filter(
                    student__year=str(sel_year)
                )
            )

        if sel_sem:
            discontinued_submitted_qs = (
                discontinued_submitted_qs.filter(
                    student__semester=str(sel_sem)
                )
            )

        if sel_section:
            discontinued_submitted_qs = (
                discontinued_submitted_qs.filter(
                    enrollment__section=sel_section
                )
            )

        discontinued_submitted_student_ids = (
            discontinued_submitted_qs
            .values_list("student_id", flat=True)
            .distinct()
        )

        enroll_qs = (
            CourseEnrollment.objects
            .select_related("student")
            .filter(
                department=department,
                course_id=sel_course_id,
                enroll=True,
            )
            .filter(
                Q(student__is_discontinued=False)
                | Q(
                    student_id__in=
                    discontinued_submitted_student_ids
                )
            )
        )

        if sel_batch:
            enroll_qs = enroll_qs.filter(
                batch=sel_batch
            )

        if sel_year:
            enroll_qs = enroll_qs.filter(
                student__year=str(sel_year)
            )

        if sel_sem:
            enroll_qs = enroll_qs.filter(
                student__semester=str(sel_sem)
            )

        if sel_section:
            enroll_qs = enroll_qs.filter(
                section=sel_section
            )

        # Use a dictionary to avoid duplicate students when duplicate
        # enrollment records exist.
        student_map = {}

        for enrollment in enroll_qs.order_by(
            "student__reg_no"
        ):
            if enrollment.student:
                student_map[
                    enrollment.student_id
                ] = enrollment.student

        students = list(student_map.values())

    else:
        students_qs = base_students_qs

        if sel_batch:
            students_qs = students_qs.filter(
                batch=sel_batch
            )

        if sel_year:
            students_qs = students_qs.filter(
                year=str(sel_year)
            )

        if sel_sem:
            students_qs = students_qs.filter(
                semester=str(sel_sem)
            )

        if sel_section:
            students_qs = students_qs.filter(
                section=sel_section
            )

        students = list(
            students_qs.order_by("reg_no")
        )

    student_ids = [
        student.id
        for student in students
    ]

    # ---------------------------------------------------------
    # Submissions
    #
    # Do not add student__is_discontinued=False here because the
    # students list already contains only:
    # - Active students
    # - Qualified discontinued students having submitted feedback
    # ---------------------------------------------------------
    submissions_qs = FeedbackSubmission.objects.none()
    submission_ids = []
    submission_map = {}

    if department and sel_course_id and student_ids:

        submissions_qs = (
            FeedbackSubmission.objects
            .filter(
                student_id__in=student_ids,
                enrollment__course_id=sel_course_id,
                enrollment__enroll=True,
                department=department,
                enrollment__department=department,
            )
            .select_related(
                "student",
                "course",
                "faculty",
                "enrollment"
            )
            .order_by("student_id", "-id")
        )

        submission_ids = list(
            submissions_qs.values_list(
                "id",
                flat=True
            )
        )

        # Preserve the latest submission for each student.
        submission_map = {}

        for submission in submissions_qs:
            if submission.student_id not in submission_map:
                submission_map[
                    submission.student_id
                ] = submission

    student_q_marks = defaultdict(dict)
    student_total = defaultdict(int)
    submitted_student_ids = set()

    if submission_ids and question_ids:

        answers = (
            FeedbackAnswer.objects
            .filter(
                submission_id__in=submission_ids,
                question_id__in=question_ids,
            )
            .select_related(
                "submission",
                "question"
            )
            .order_by(
                "submission__student_id",
                "question_id"
            )
        )

        for answer in answers:
            student_id = (
                answer.submission.student_id
            )

            question_id = answer.question_id

            score_value = int(
                answer.score or 0
            )

            submitted_student_ids.add(
                student_id
            )

            student_q_marks[
                student_id
            ][question_id] = score_value

            student_total[
                student_id
            ] += score_value

    # A submission may exist even when no answer row exists.
    # Include those students in the submitted count also.
    submitted_student_ids.update(
        submission_map.keys()
    )

    rows = []

    for index, student in enumerate(
        students,
        start=1
    ):
        submission = submission_map.get(
            student.id
        )

        rows.append({
            "sno": index,

            "student_id": student.id,

            "student_name": (
                getattr(
                    student,
                    "student_name",
                    ""
                )
                or getattr(
                    student,
                    "name",
                    ""
                )
                or "-"
            ),

            "reg_no": student.reg_no or "-",

            "marks": student_q_marks.get(
                student.id,
                {}
            ),

            "total": student_total.get(
                student.id,
                0
            ),

            "is_submitted": (
                student.id in submitted_student_ids
            ),

            # Optional flag for showing discontinued status
            # inside the template when required.
            "is_discontinued": bool(
                student.is_discontinued
            ),

            "overall_effectiveness_percentage": (
                submission.overall_effectiveness_percentage
                if (
                    submission
                    and submission
                    .overall_effectiveness_percentage
                    is not None
                )
                else "-"
            ),

            "student_satisfaction": (
                "Yes"
                if (
                    submission
                    and submission.student_satisfaction
                    is True
                )
                else "No"
                if (
                    submission
                    and submission.student_satisfaction
                    is False
                )
                else "-"
            ),

            "recommendation_to_continue_improve": (
                submission
                .recommendation_to_continue_improve
                if (
                    submission
                    and submission
                    .recommendation_to_continue_improve
                )
                else "-"
            ),

            "open_comments_for_improvement": (
                submission.open_comments_for_improvement
                if (
                    submission
                    and submission
                    .open_comments_for_improvement
                )
                else "-"
            ),
        })

    # ---------------------------------------------------------
    # OTHER USERS latest remarks/action only
    # ---------------------------------------------------------
    other_feedback_notes = []

    if department and sel_course_id:

        raw_other_notes = (
            CourseFeedbackRemark.objects
            .select_related("faculty")
            .filter(
                department=department,
                course_id=sel_course_id,
            )
            .exclude(faculty=faculty)
            .order_by(
                "faculty_id",
                "-updated_at",
                "-id"
            )
        )

        seen_faculty_ids = set()

        for note in raw_other_notes:
            if note.faculty_id not in seen_faculty_ids:
                other_feedback_notes.append(note)
                seen_faculty_ids.add(note.faculty_id)

    # ---------------------------------------------------------
    # Current user's own note
    # ---------------------------------------------------------
    feedback_note = None

    if department and sel_course_id:
        feedback_note = (
            CourseFeedbackRemark.objects
            .filter(
                faculty=faculty,
                department=department,
                course_id=sel_course_id,
            )
            .first()
        )

    # ---------------------------------------------------------
    # Render
    # ---------------------------------------------------------
    return render(
        request,
        "feedback_management/faculty/feedback/view_feedbacks.html",
        {
            "faculty": faculty,
            "department": department,
            "faculty_department": faculty_department,
            "departments": departments,

            "can_view_all_feedback_data": (
                permission_scope.get(
                    "can_view_all",
                    False
                )
            ),

            "can_view_department_feedback_data": (
                permission_scope.get(
                    "can_view_department",
                    False
                )
            ),

            "batches": batches,
            "years": years,
            "semesters": semesters,
            "sections": sections,

            "sel_department_id": sel_department_id,
            "sel_batch": sel_batch,
            "sel_year": sel_year,
            "sel_sem": sel_sem,
            "sel_section": sel_section,
            "sel_course_id": sel_course_id,

            "filters_applied": filters_applied,
            "courses": courses,
            "selected_course": selected_course,

            "total_students": len(students),

            "submitted_students": len(
                submitted_student_ids
            ),

            "not_submitted_students": max(
                len(students)
                - len(submitted_student_ids),
                0
            ),

            "category_spans": category_spans,
            "ordered_questions": ordered_questions,
            "rows": rows,

            "feedback_note": feedback_note,
            "other_feedback_notes": other_feedback_notes,

            "summary_rows": summary_rows,
            "show_summary": bool(summary_rows),

            "dashboard_overall": dashboard_overall,

            "dashboard_category_averages": (
                dashboard_category_averages
            ),

            "dashboard_course_boxes": (
                dashboard_course_boxes
            ),
        }
    )



from collections import defaultdict
from io import BytesIO

from django.db.models import Q
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import render
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from collections import defaultdict
from datetime import datetime

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT







@check_permission("view_feedbacks")
def view_feedbacks_summary_pdf(request):
    """
    PDF for Course-wise Question Average Summary
    - Opens in browser first (not direct download)
    - A4 landscape
    - Shows course-wise category cumulative averages only
    - Full category names shown dynamically with word-based wrapping
    """

    # ---------------------------------------------------------
    # Logged-in faculty
    # ---------------------------------------------------------
    user = request.user.Employee_id
    faculty = general_information.objects.select_related("department").get(faculty_id=user)
    faculty_department = faculty.department

    permission_scope = _get_feedback_permission_scope(request)
    if not permission_scope["has_access"]:
        return HttpResponseForbidden("You do not have permission to view feedback data.")

    # ---------------------------------------------------------
    # Read request params
    # ---------------------------------------------------------
    sel_department_id = (request.GET.get("department_id") or "").strip()
    sel_batch = (request.GET.get("batch") or "").strip()
    sel_year = (request.GET.get("year") or "").strip()
    sel_sem = (request.GET.get("semester") or "").strip()
    sel_section = (request.GET.get("section") or "").strip()

    if not (sel_year and sel_sem):
        return HttpResponseBadRequest("Select Year and Semester to generate summary PDF.")

    # ---------------------------------------------------------
    # Department permission scope
    # ---------------------------------------------------------
    if permission_scope.get("can_view_all"):
        departments = Add_Department.objects.filter(is_active=True).order_by("Department")
        if sel_department_id:
            department = Add_Department.objects.filter(id=sel_department_id, is_active=True).first()
        else:
            department = faculty_department or departments.first()
    else:
        department = faculty_department

    if not department:
        return HttpResponseBadRequest("Department not found.")

    # ---------------------------------------------------------
    # Dynamic wrapper for category headings
    # ---------------------------------------------------------
    def wrap_heading_text(text, max_lines=3):
        """
        Dynamically wrap heading text by words.
        No manual hardcoding of category names.
        Produces balanced 2 or 3 lines using word groups.
        """
        text = (text or "").strip()
        if not text:
            return "General"

        words = text.split()
        if len(words) <= 2:
            return "<br/>".join(words)

        if len(words) <= 4:
            mid = (len(words) + 1) // 2
            line1 = " ".join(words[:mid])
            line2 = " ".join(words[mid:])
            return f"{line1}<br/>{line2}"

        if max_lines == 2:
            mid = (len(words) + 1) // 2
            line1 = " ".join(words[:mid])
            line2 = " ".join(words[mid:])
            return f"{line1}<br/>{line2}"

        # 3-line balanced split
        total = len(words)
        first = (total + 2) // 3
        second = (total - first + 1) // 2
        third = total - first - second

        line1 = " ".join(words[:first])
        line2 = " ".join(words[first:first + second])
        line3 = " ".join(words[first + second:first + second + third])

        lines = [line1, line2, line3]
        lines = [ln for ln in lines if ln.strip()]
        return "<br/>".join(lines)

    # ---------------------------------------------------------
    # Questions / categories
    # ---------------------------------------------------------
    ordered_questions = list(
        FeedbackQuestion.objects
        .filter(Q(department__isnull=True) | Q(department=department))
        .order_by("category", "id")
    )

    if not ordered_questions:
        return HttpResponseBadRequest("No feedback questions configured.")

    question_ids = [q.id for q in ordered_questions]

    category_question_map = defaultdict(list)
    for q in ordered_questions:
        cat_name = q.category or "General"
        category_question_map[cat_name].append(q.id)

    category_names = list(category_question_map.keys())

    # ---------------------------------------------------------
    # Course list based on year/semester + optional batch/section
    # ---------------------------------------------------------
    summary_courses_qs = Course.objects.filter(
        department=department,
        year=str(sel_year),
        semester=str(sel_sem),
        is_active=True
    )

    if sel_batch or sel_section:
        enroll_course_qs = CourseEnrollment.objects.filter(
            department=department,
            enroll=True,
            student__year=str(sel_year),
            student__semester=str(sel_sem)
        )
        if sel_batch:
            enroll_course_qs = enroll_course_qs.filter(batch=sel_batch)
        if sel_section:
            enroll_course_qs = enroll_course_qs.filter(section=sel_section)

        filtered_ids = enroll_course_qs.values_list("course_id", flat=True).distinct()
        summary_courses_qs = summary_courses_qs.filter(id__in=filtered_ids)

    summary_courses = list(summary_courses_qs.order_by("course_code"))

    # ---------------------------------------------------------
    # Build summary rows
    # ---------------------------------------------------------
    summary_rows = []

    for course in summary_courses:
        enroll_qs = CourseEnrollment.objects.filter(
            department=department,
            course_id=course.id,
            enroll=True,
            student__year=str(sel_year),
            student__semester=str(sel_sem)
        )

        if sel_batch:
            enroll_qs = enroll_qs.filter(batch=sel_batch)
        if sel_section:
            enroll_qs = enroll_qs.filter(section=sel_section)

        student_count = enroll_qs.count()

        submissions_qs = FeedbackSubmission.objects.filter(
            enrollment__course_id=course.id,
            department=department,
            enrollment__department=department
        )

        if sel_batch:
            submissions_qs = submissions_qs.filter(enrollment__batch=sel_batch)
        if sel_section:
            submissions_qs = submissions_qs.filter(enrollment__section=sel_section)

        actual_responses = submissions_qs.count()

        answers = FeedbackAnswer.objects.filter(
            submission__in=submissions_qs,
            question_id__in=question_ids
        )

        question_totals = defaultdict(int)
        for ans in answers:
            question_totals[ans.question_id] += int(ans.score or 0)

        category_averages = {}
        grand_average = 0.0

        if actual_responses > 0:
            total_all_scores = 0

            for category_name, q_ids in category_question_map.items():
                category_total = sum(question_totals.get(qid, 0) for qid in q_ids)
                total_all_scores += category_total

                divisor = actual_responses * len(q_ids)
                category_avg = round(category_total / divisor, 2) if divisor > 0 else 0.0
                category_averages[category_name] = category_avg

            grand_average = round(total_all_scores / actual_responses, 2)
        else:
            for category_name in category_names:
                category_averages[category_name] = 0.0
            grand_average = 0.0

        summary_rows.append({
            "course_code": course.course_code or "-",
            "course_title": course.title or "-",
            "category_averages": category_averages,
            "grand_average": grand_average,
            "response_count": actual_responses,
            "student_count": student_count,
        })

    # ---------------------------------------------------------
    # PDF response
    # ---------------------------------------------------------
    filename = f"CourseWiseSummary_{department.Department}_{sel_year}_Sem{sel_sem}.pdf"
    response = HttpResponse(content_type="application/pdf")

    # Open in browser first
    response["Content-Disposition"] = f'inline; filename="{filename}"'

    PDF_PAGE_SIZE = landscape(A4)

    semester_type = "Odd" if str(sel_sem) in ["1", "3", "5", "7"] else "Even"
    subtitle = f"Academic Year: {get_academic_year()} ({semester_type} Semester)"

    doc = SimpleDocTemplate(
        response,
        pagesize=PDF_PAGE_SIZE,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=42 * mm,
        bottomMargin=16 * mm,
        title="Course-wise Question Average Summary"
    )

    styles = getSampleStyleSheet()

    info_style = ParagraphStyle(
        "info_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#0f172a"),
        alignment=TA_LEFT,
    )

    heading_style = ParagraphStyle(
        "heading_style",
        parent=styles["Heading4"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=6,
    )

    header_style = ParagraphStyle(
        "header_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=6.9,
        leading=8.0,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#1e3a8a"),
        wordWrap="LTR",
    )

    body_center_style = ParagraphStyle(
        "body_center_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
    )

    body_left_style = ParagraphStyle(
        "body_left_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.5,
        leading=9,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#111827"),
        wordWrap="CJK",
    )

    story = []

    story.append(Paragraph("Course-wise Question Average Summary", heading_style))
    story.append(Paragraph(f"Department: {department.Department}", info_style))
    story.append(
        Paragraph(
            f"Year / Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
            f"Section: {sel_section or '-'} &nbsp;&nbsp;&nbsp; "
            f"Batch: {sel_batch or '-'}",
            info_style
        )
    )
    story.append(Spacer(1, 8))

    # ---------------------------------------------------------
    # Table header with dynamic full-name wrapping
    # ---------------------------------------------------------
    header = [Paragraph("Course Code - Title", header_style)]
    for cat in category_names:
        header.append(Paragraph(wrap_heading_text(cat, max_lines=3), header_style))
    header.extend([
        Paragraph("Avg Total", header_style),
        Paragraph("Responses", header_style),
        Paragraph("Enrolled", header_style),
    ])

    data = [header]

    for row in summary_rows:
        row_data = [
            Paragraph(f"<b>{row['course_code']}</b><br/>{row['course_title']}", body_left_style)
        ]
        for cat in category_names:
            row_data.append(Paragraph(f"{row['category_averages'].get(cat, 0.0):.2f}", body_center_style))

        row_data.extend([
            Paragraph(f"{row['grand_average']:.2f}", body_center_style),
            Paragraph(str(row["response_count"]), body_center_style),
            Paragraph(str(row["student_count"]), body_center_style),
        ])
        data.append(row_data)

    if not summary_rows:
        empty_row = [Paragraph("No data found.", body_left_style)] + [Paragraph("-", body_center_style)] * (len(header) - 1)
        data.append(empty_row)

    # ---------------------------------------------------------
    # Width calculation for A4 landscape
    # ---------------------------------------------------------
    page_w, _ = doc.pagesize
    usable_w = page_w - doc.leftMargin - doc.rightMargin

    course_col_w = 60 * mm
    avg_total_col_w = 16 * mm
    responses_col_w = 16 * mm
    enrolled_col_w = 16 * mm

    fixed_width = course_col_w + avg_total_col_w + responses_col_w + enrolled_col_w
    remaining_w = usable_w - fixed_width

    category_col_count = max(len(category_names), 1)
    category_col_w = remaining_w / category_col_count
    category_col_w = max(category_col_w, 24 * mm)

    col_widths = [course_col_w] + [category_col_w] * len(category_names) + [
        avg_total_col_w, responses_col_w, enrolled_col_w
    ]

    # Taller header row so wrapped headings fit neatly
    row_heights = [18 * mm] + [None] * (len(data) - 1)

    tbl = Table(
        data,
        colWidths=col_widths,
        rowHeights=row_heights,
        repeatRows=1
    )

    ts = TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d1d5db")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "LEFT"),

        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),

        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("BACKGROUND", (-3, 1), (-3, -1), colors.HexColor("#f8fafc")),
        ("FONTNAME", (-3, 1), (-3, -1), "Helvetica-Bold"),
    ])

    tbl.setStyle(ts)
    story.append(tbl)
    story.append(Spacer(1, 8))
    story.append(
        Paragraph(
            "* Category values are cumulative averages calculated from the questions under each category.",
            info_style
        )
    )

    # ---------------------------------------------------------
    # Header / Footer
    # ---------------------------------------------------------
    def _on_page(canv, doc_):
        pw, ph = doc_.pagesize
        _draw_rit_header_footer(
            canv,
            pw,
            ph,
            title="STUDENTS FEEDBACK",
            subtitle=subtitle
        )

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    return response








from collections import defaultdict
from datetime import datetime

from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)


from collections import defaultdict
from datetime import datetime

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)


from collections import defaultdict
from datetime import datetime

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
)

# make sure these models are already imported in your file:
# Add_Department, Course, CourseEnrollment, FeedbackQuestion,
# FeedbackSubmission, FeedbackAnswer, AssignSubjectFaculty,
# CourseFeedbackRemark, general_information


from collections import defaultdict
from datetime import datetime

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from collections import defaultdict
from datetime import datetime

from django.db.models import Q
from django.http import (
    HttpResponse,
    HttpResponseBadRequest,
    HttpResponseForbidden,
)

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)






from reportlab.lib.pagesizes import landscape, A3
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT   

from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from datetime import datetime
from collections import defaultdict
from django.db.models import Q


@check_permission("view_feedbacks")
def view_feedbacks_bulk_pdf(request):
    user = request.user.Employee_id
    faculty = general_information.objects.select_related("department").get(faculty_id=user)
    faculty_department = faculty.department

    permission_scope = _get_feedback_permission_scope(request)
    if not permission_scope["has_access"]:
        return HttpResponseForbidden("You do not have permission to view feedback data.")

    sel_department_id = (request.GET.get("department_id") or "").strip()
    sel_batch = (request.GET.get("batch") or "").strip()
    sel_year = (request.GET.get("year") or "").strip()
    sel_sem = (request.GET.get("semester") or "").strip()
    sel_section = (request.GET.get("section") or "").strip()

    # ---------------------------------------------------------
    # Department resolve
    # ---------------------------------------------------------
    if permission_scope["can_view_all"]:
        if sel_department_id:
            department = Add_Department.objects.filter(
                id=sel_department_id,
                is_active=True
            ).first()
        else:
            department = faculty_department
    else:
        department = faculty_department

    if not department:
        return HttpResponseBadRequest("Invalid department.")

    if not (sel_year and sel_sem):
        return HttpResponseBadRequest("Select Year and Semester to generate Bulk PDF.")

    # ---------------------------------------------------------
    # Filtered courses
    # ---------------------------------------------------------
    courses_qs = Course.objects.filter(
        department=department,
        year=str(sel_year),
        semester=str(sel_sem),
        is_active=True
    )

    if sel_batch or sel_section:
        enroll_course_qs = CourseEnrollment.objects.filter(
            department=department,
            enroll=True,
            student__is_discontinued=False,
        )

        if sel_batch:
            enroll_course_qs = enroll_course_qs.filter(batch=sel_batch)
        if sel_section:
            enroll_course_qs = enroll_course_qs.filter(section=sel_section)

        enroll_course_qs = enroll_course_qs.filter(
            student__year=str(sel_year),
            student__semester=str(sel_sem)
        )

        filtered_course_ids = enroll_course_qs.values_list("course_id", flat=True).distinct()
        courses_qs = courses_qs.filter(id__in=filtered_course_ids)

    courses = list(courses_qs.order_by("course_code"))

    if not courses:
        return HttpResponseBadRequest("No courses found for the selected filters.")

    # ---------------------------------------------------------
    # Questions
    # ---------------------------------------------------------
    ordered_questions = list(
        FeedbackQuestion.objects
        .filter(Q(department__isnull=True) | Q(department=department))
        .order_by("category", "id")
    )

    if not ordered_questions:
        return HttpResponseBadRequest("No feedback questions configured.")

    question_ids = [q.id for q in ordered_questions]

    # ---------------------------------------------------------
    # Category spans + question numbering + legend
    # ---------------------------------------------------------
    category_spans = []
    question_number_map = {}
    question_legend = []

    cur = ordered_questions[0].category or "General"
    span = 0

    for idx, q in enumerate(ordered_questions, start=1):
        qcat = q.category or "General"

        question_number_map[q.id] = idx
        question_legend.append({
            "number": idx,
            "category": qcat,
            "question_text": q.question_text,
            "question_id": q.id,
        })

        if qcat != cur:
            category_spans.append({"category": cur, "span": span})
            cur = qcat
            span = 1
        else:
            span += 1

    category_spans.append({"category": cur, "span": span})

    # ---------------------------------------------------------
    # PDF setup - LANDSCAPE A3
    # ---------------------------------------------------------
    PDF_PAGE_SIZE = landscape(A3)

    filename = f"Bulk_Feedbacks_{department.Department}_{sel_year}_Sem{sel_sem}.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=PDF_PAGE_SIZE,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=50 * mm,
        bottomMargin=25 * mm,
        title="Students Feedback Bulk Report"
    )

    styles = getSampleStyleSheet()

    info_style = ParagraphStyle(
        "info",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#0f172a"),
        alignment=0
    )

    heading_style = ParagraphStyle(
        "heading_style",
        parent=styles["Heading4"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=6
    )

    sub_heading_style = ParagraphStyle(
        "sub_heading_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#111827"),
        spaceAfter=6
    )

    p_center = ParagraphStyle(
        "p_center",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=5.2,
        leading=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a")
    )

    p_center_bold = ParagraphStyle(
        "p_center_bold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=5.4,
        leading=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a")
    )

    p_summary_course = ParagraphStyle(
        "p_summary_course",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=6.2,
        leading=8,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#0f172a"),
        wordWrap="CJK"
    )

    p_small_left = ParagraphStyle(
        "p_small_left",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=6.2,
        leading=8.5,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#0f172a"),
        wordWrap="CJK"
    )

    remarks_style = ParagraphStyle(
        "remarks_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=12,
        alignment=0,
        textColor=colors.HexColor("#0f172a")
    )

    legend_category_style = ParagraphStyle(
        "legend_category_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=4,
        spaceBefore=6
    )

    legend_text_style = ParagraphStyle(
        "legend_text_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=12,
        textColor=colors.HexColor("#111827")
    )

    summary_title_style = ParagraphStyle(
        "summary_title_style",
        parent=styles["Heading4"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=6
    )

    summary_section_style = ParagraphStyle(
        "summary_section_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9.5,
        leading=12,
        textColor=colors.HexColor("#111827"),
        spaceAfter=4,
        spaceBefore=8
    )

    semester_type = "Odd" if str(sel_sem) in ["1", "3", "5", "7"] else "Even"
    academic_year_text = f"Academic Year: {get_academic_year()} ({semester_type} Semester)"
    subtitle = academic_year_text

    story = []

    # ---------------------------------------------------------
    # BUILD SEPARATE ENTRIES COURSE + SECTION + BATCH + FACULTY
    # ---------------------------------------------------------
    handled_entries = []
    seen_keys = set()

    for course in courses:
        assign_qs = AssignSubjectFaculty.objects.select_related("faculty", "course").filter(
            department=department,
            course_id=course.id,
            is_active=True,
            course__year=str(sel_year),
            course__semester=str(sel_sem)
        )

        if sel_batch:
            assign_qs = assign_qs.filter(batch=sel_batch)

        if sel_section:
            assign_qs = assign_qs.filter(section=sel_section)

        assign_list = list(assign_qs.order_by("section", "batch", "id"))

        if not assign_list:
            unique_key = (
                str(course.id),
                "",
                str(sel_batch or ""),
                "",
            )

            if unique_key not in seen_keys:
                seen_keys.add(unique_key)
                handled_entries.append({
                    "course": course,
                    "faculty_name": faculty.name,
                    "section": sel_section or "",
                    "batch": sel_batch or "",
                })

            continue

        for assign_obj in assign_list:
            entry_section = assign_obj.section or ""
            entry_batch = assign_obj.batch or ""
            entry_faculty_name = assign_obj.faculty.name if assign_obj.faculty else faculty.name

            unique_key = (
                str(course.id),
                entry_section,
                entry_batch,
                entry_faculty_name,
            )

            if unique_key in seen_keys:
                continue

            seen_keys.add(unique_key)

            handled_entries.append({
                "course": course,
                "faculty_name": entry_faculty_name,
                "section": entry_section,
                "batch": entry_batch,
            })

    handled_entries = sorted(
        handled_entries,
        key=lambda x: (
            x["section"] or "",
            x["course"].course_code or "",
            x["batch"] or "",
            x["faculty_name"] or "",
        )
    )

    if not handled_entries:
        return HttpResponseBadRequest("No faculty/course entries found for the selected filters.")

    # ---------------------------------------------------------
    # 1. Collect data for Summary Table First Page
    # ---------------------------------------------------------
    course_summary_rows = []

    for entry in handled_entries:
        course = entry["course"]
        filtered_faculty_name = entry["faculty_name"]
        entry_section = entry["section"]
        entry_batch = entry["batch"]

        enroll_qs = CourseEnrollment.objects.select_related("student").filter(
            department=department,
            course_id=course.id,
            enroll=True,
            student__year=str(sel_year),
            student__semester=str(sel_sem),
            student__is_discontinued=False,
        )

        if entry_batch:
            enroll_qs = enroll_qs.filter(batch=entry_batch)
        elif sel_batch:
            enroll_qs = enroll_qs.filter(batch=sel_batch)

        if entry_section:
            enroll_qs = enroll_qs.filter(section=entry_section)
        elif sel_section:
            enroll_qs = enroll_qs.filter(section=sel_section)

        students = [
            e.student for e in enroll_qs.order_by("student__reg_no")
            if e.student and not e.student.is_discontinued
        ]

        student_ids = [s.id for s in students]
        student_count = len(students)

        submissions_qs = FeedbackSubmission.objects.filter(
            student_id__in=student_ids,
            enrollment__course_id=course.id,
            department=department,
            enrollment__department=department,
            student__is_discontinued=False,
        ).select_related("student", "faculty", "enrollment")

        if entry_batch:
            submissions_qs = submissions_qs.filter(enrollment__batch=entry_batch)
        elif sel_batch:
            submissions_qs = submissions_qs.filter(enrollment__batch=sel_batch)

        if entry_section:
            submissions_qs = submissions_qs.filter(enrollment__section=entry_section)
        elif sel_section:
            submissions_qs = submissions_qs.filter(enrollment__section=sel_section)

        if filtered_faculty_name:
            submissions_qs = submissions_qs.filter(faculty__name=filtered_faculty_name)

        submission_ids = list(submissions_qs.values_list("id", flat=True))
        actual_responses = len(submission_ids)

        course_question_totals = defaultdict(int)

        if submission_ids:
            answers = FeedbackAnswer.objects.filter(
                submission_id__in=submission_ids,
                question_id__in=question_ids,
                submission__student__is_discontinued=False,
            )

            for ans in answers:
                course_question_totals[ans.question_id] += int(ans.score or 0)

        course_grand_total = sum(course_question_totals.values())

        course_question_averages = {}

        if actual_responses > 0:
            for q in ordered_questions:
                course_question_averages[q.id] = round(
                    course_question_totals.get(q.id, 0) / actual_responses,
                    2
                )

            course_grand_average = round(course_grand_total / actual_responses, 2)
        else:
            for q in ordered_questions:
                course_question_averages[q.id] = 0.0

            course_grand_average = 0.0

        course_summary_rows.append({
            "course": course,
            "course_code": course.course_code or "-",
            "course_title": course.title or "-",
            "student_count": student_count,
            "response_count": actual_responses,
            "question_averages": course_question_averages,
            "grand_average": course_grand_average,
            "faculty_name": filtered_faculty_name,
            "section": entry_section,
            "batch": entry_batch,
            "course_faculty_display": (
                f"{course.course_code or '-'}<br/>"
                f"{filtered_faculty_name or '-'}<br/>"
                f"Batch: {entry_batch or '-'}"
            ),
        })

    # ---------------------------------------------------------
    # 2. Summary Table - FIRST PAGE SECTION-WISE SEPARATE
    # ---------------------------------------------------------
    story.append(Paragraph("Course-wise Question Average Summary", summary_title_style))
    story.append(Paragraph(f"Department: {department.Department}", info_style))
    story.append(Paragraph(
        f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
        f"Section: {sel_section or 'All'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or 'All'}",
        info_style
    ))
    story.append(Spacer(1, 12))

    summary_grouped = defaultdict(list)

    for item in course_summary_rows:
        summary_grouped[item["section"] or "-"].append(item)

    ordered_sections = sorted(summary_grouped.keys())

    usable_w = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

    summary_course_col_w = 55 * mm
    summary_avg_total_col_w = 18 * mm
    summary_responses_col_w = 18 * mm
    summary_enrolled_col_w = 18 * mm

    q_count = len(ordered_questions)

    summary_fixed_w = (
        summary_course_col_w +
        summary_avg_total_col_w +
        summary_responses_col_w +
        summary_enrolled_col_w
    )

    summary_q_w = max((usable_w - summary_fixed_w) / max(q_count, 1), 7 * mm)

    summary_col_widths = (
        [summary_course_col_w] +
        [summary_q_w] * q_count +
        [
            summary_avg_total_col_w,
            summary_responses_col_w,
            summary_enrolled_col_w,
        ]
    )

    for sec_index, section_key in enumerate(ordered_sections, start=1):
        story.append(Paragraph(f"Section: {section_key}", summary_section_style))

        summary_row1 = ["Course / Faculty / Batch"]

        for c in category_spans:
            summary_row1.append(Paragraph(str(c["category"]), p_center))
            for _ in range(c["span"] - 1):
                summary_row1.append("")

        summary_row1.extend(["Avg Total", "Responses", "Enrolled"])

        summary_row2 = [""]

        for q in ordered_questions:
            summary_row2.append(Paragraph(str(question_number_map.get(q.id, "")), p_center_bold))

        summary_row2.extend(["", "", ""])

        summary_data = [summary_row1, summary_row2]

        for item in summary_grouped[section_key]:
            row = [Paragraph(item["course_faculty_display"], p_summary_course)]

            for q in ordered_questions:
                row.append(str(item["question_averages"].get(q.id, 0)))

            row.extend([
                str(item["grand_average"]),
                str(item["response_count"]),
                str(item["student_count"]),
            ])

            summary_data.append(row)

        summary_tbl = Table(summary_data, colWidths=summary_col_widths, repeatRows=2)

        summary_style = TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d1d5db")),
            ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f8fafc")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 1), 6),
            ("FONTNAME", (0, 2), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 2), (-4, -1), "Helvetica"),
            ("FONTNAME", (-3, 2), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 2), (-1, -1), 6),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("BACKGROUND", (-3, 2), (-1, -1), colors.HexColor("#f8fafc")),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("ALIGN", (0, 2), (0, -1), "LEFT"),
            ("VALIGN", (0, 2), (0, -1), "MIDDLE"),
        ])

        summary_style.add("SPAN", (0, 0), (0, 1))
        summary_style.add("SPAN", (-3, 0), (-3, 1))
        summary_style.add("SPAN", (-2, 0), (-2, 1))
        summary_style.add("SPAN", (-1, 0), (-1, 1))

        start_col = 1

        for c in category_spans:
            end_col = start_col + c["span"] - 1
            summary_style.add("SPAN", (start_col, 0), (end_col, 0))
            start_col = end_col + 1

        summary_tbl.setStyle(summary_style)

        story.append(summary_tbl)
        story.append(Spacer(1, 10))

    story.append(PageBreak())

    # ---------------------------------------------------------
    # 3. Individual Course Feedback Tables
    # ---------------------------------------------------------
    for course_index, item in enumerate(course_summary_rows, start=1):
        course = item["course"]
        filtered_faculty_name = item["faculty_name"]
        entry_section = item["section"]
        entry_batch = item["batch"]

        enroll_qs = CourseEnrollment.objects.select_related("student").filter(
            department=department,
            course_id=course.id,
            enroll=True,
            student__year=str(sel_year),
            student__semester=str(sel_sem),
            student__is_discontinued=False,
        )

        if entry_batch:
            enroll_qs = enroll_qs.filter(batch=entry_batch)
        elif sel_batch:
            enroll_qs = enroll_qs.filter(batch=sel_batch)

        if entry_section:
            enroll_qs = enroll_qs.filter(section=entry_section)
        elif sel_section:
            enroll_qs = enroll_qs.filter(section=sel_section)

        students = [
            e.student for e in enroll_qs.order_by("student__reg_no")
            if e.student and not e.student.is_discontinued
        ]

        student_ids = [s.id for s in students]
        student_count = len(students)

        submissions_qs = FeedbackSubmission.objects.filter(
            student_id__in=student_ids,
            enrollment__course_id=course.id,
            department=department,
            enrollment__department=department,
            student__is_discontinued=False,
        ).select_related("student", "faculty", "enrollment")

        if entry_batch:
            submissions_qs = submissions_qs.filter(enrollment__batch=entry_batch)
        elif sel_batch:
            submissions_qs = submissions_qs.filter(enrollment__batch=sel_batch)

        if entry_section:
            submissions_qs = submissions_qs.filter(enrollment__section=entry_section)
        elif sel_section:
            submissions_qs = submissions_qs.filter(enrollment__section=sel_section)

        if filtered_faculty_name:
            submissions_qs = submissions_qs.filter(faculty__name=filtered_faculty_name)

        submission_ids = list(submissions_qs.values_list("id", flat=True))
        submission_map = {sub.student_id: sub for sub in submissions_qs}

        student_q_marks = defaultdict(dict)
        student_total = defaultdict(int)

        if submission_ids:
            answers = FeedbackAnswer.objects.filter(
                submission_id__in=submission_ids,
                question_id__in=question_ids,
                submission__student__is_discontinued=False,
            ).select_related("submission")

            for ans in answers:
                sid = ans.submission.student_id
                qid = ans.question_id
                sc = int(ans.score or 0)

                student_q_marks[sid][qid] = sc
                student_total[sid] += sc

        story.append(Paragraph(
            f"Course {course_index}: {course.course_code} - {course.title}",
            heading_style
        ))
        story.append(Paragraph(f"Department: {department.Department}", info_style))
        story.append(Paragraph(f"Name of the Faculty: {filtered_faculty_name}", info_style))
        story.append(Paragraph(
            f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
            f"Section: {entry_section or 'All'} &nbsp;&nbsp;&nbsp; Batch: {entry_batch or 'All'}<br/>"
            f"Students Enrolled: {student_count} | Responses Received: {item['response_count']} "
            f"({round((item['response_count'] / student_count * 100), 1) if student_count > 0 else 0}%)",
            info_style
        ))
        story.append(Spacer(1, 8))

        row1 = ["Sl.No"]

        for c in category_spans:
            row1.append(Paragraph(str(c["category"]), p_center))
            for _ in range(c["span"] - 1):
                row1.append("")

        row1.extend([
            Paragraph("Total", p_center),
            Paragraph("Overall Effectiveness (%)", p_center),
            Paragraph("Student Satisfaction", p_center),
            Paragraph("Recommendation to Continue / Improve", p_center),
            Paragraph("Open Comments for Improvement", p_center),
        ])

        row2 = [""]

        for q in ordered_questions:
            row2.append(Paragraph(str(question_number_map.get(q.id, "")), p_center_bold))

        row2.extend(["", "", "", "", ""])

        data = [row1, row2]

        for i, st in enumerate(students, start=1):
            submission = submission_map.get(st.id)

            overall_effectiveness = (
                str(submission.overall_effectiveness_percentage)
                if submission and submission.overall_effectiveness_percentage is not None
                else "-"
            )

            student_satisfaction = (
                "Yes" if submission and submission.student_satisfaction is True
                else "No" if submission and submission.student_satisfaction is False
                else "-"
            )

            recommendation = (
                submission.recommendation_to_continue_improve
                if submission and submission.recommendation_to_continue_improve
                else "-"
            )

            open_comments = (
                submission.open_comments_for_improvement
                if submission and submission.open_comments_for_improvement
                else "-"
            )

            r = [str(i)]

            for q in ordered_questions:
                r.append(str(student_q_marks.get(st.id, {}).get(q.id, 0)))

            r.extend([
                str(student_total.get(st.id, 0)),
                overall_effectiveness,
                student_satisfaction,
                Paragraph(recommendation, p_small_left),
                Paragraph(open_comments, p_small_left),
            ])

            data.append(r)

        if not students:
            no_data_row = ["No students found."] + [""] * (len(ordered_questions) + 5)
            data.append(no_data_row)

        usable_w = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

        col_w_sno = 11 * mm
        col_w_total = 15 * mm
        col_w_eff = 23 * mm
        col_w_sat = 23 * mm
        col_w_rec = 48 * mm
        col_w_open = 55 * mm

        q_count = len(ordered_questions)

        fixed_width = (
            col_w_sno +
            col_w_total +
            col_w_eff +
            col_w_sat +
            col_w_rec +
            col_w_open
        )

        q_w = max((usable_w - fixed_width) / max(q_count, 1), 7.5 * mm)

        col_widths = (
            [col_w_sno] +
            [q_w] * q_count +
            [col_w_total, col_w_eff, col_w_sat, col_w_rec, col_w_open]
        )

        tbl = Table(data, colWidths=col_widths, repeatRows=2)

        table_style = TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f8fafc")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, 1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 1), 5.8),
            ("FONTNAME", (0, 2), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 2), (-5, -1), 5.8),
            ("ALIGN", (0, 2), (-5, -1), "CENTER"),
            ("ALIGN", (-4, 2), (-1, -1), "LEFT"),
            ("VALIGN", (-4, 2), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f8fafc")),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])

        table_style.add("SPAN", (0, 0), (0, 1))
        table_style.add("SPAN", (q_count + 1, 0), (q_count + 1, 1))
        table_style.add("SPAN", (q_count + 2, 0), (q_count + 2, 1))
        table_style.add("SPAN", (q_count + 3, 0), (q_count + 3, 1))
        table_style.add("SPAN", (q_count + 4, 0), (q_count + 4, 1))
        table_style.add("SPAN", (q_count + 5, 0), (q_count + 5, 1))

        start_col = 1

        for c in category_spans:
            end_col = start_col + c["span"] - 1
            table_style.add("SPAN", (start_col, 0), (end_col, 0))
            start_col = end_col + 1

        if not students:
            table_style.add("SPAN", (0, 2), (-1, 2))
            table_style.add("ALIGN", (0, 2), (-1, 2), "CENTER")
            table_style.add("FONTNAME", (0, 2), (-1, 2), "Helvetica-Oblique")

        tbl.setStyle(table_style)
        story.append(tbl)

        # ---------------------------------------------------------
        # Remarks Page
        # ---------------------------------------------------------
        story.append(PageBreak())
        story.append(Paragraph(
            f"Course {course_index}: {course.course_code} - {course.title}",
            heading_style
        ))
        story.append(Paragraph("Remarks / Action Taken", sub_heading_style))
        story.append(Paragraph(f"Department: {department.Department}", info_style))
        story.append(Paragraph(f"Name of the Faculty: {filtered_faculty_name}", info_style))
        story.append(Paragraph(
            f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
            f"Section: {entry_section or 'All'} &nbsp;&nbsp;&nbsp; Batch: {entry_batch or 'All'}",
            info_style
        ))
        story.append(Spacer(1, 8))

        all_feedback_notes = CourseFeedbackRemark.objects.select_related("faculty").filter(
            department=department,
            course=course
        ).order_by("-updated_at", "-id")

        remarks_box_width = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

        if all_feedback_notes.exists():
            for note_index, note in enumerate(all_feedback_notes, start=1):
                note_faculty_name = note.faculty.name if note.faculty else "Unknown User"

                note_data = [
                    [Paragraph(f"<b>{note_index}. Faculty:</b> {note_faculty_name}", remarks_style)],
                    [Paragraph(f"<b>Remarks:</b> {note.remarks if note.remarks else '-'}", remarks_style)],
                    [Paragraph(f"<b>Action Plan:</b> {note.action_taken if note.action_taken else '-'}", remarks_style)],
                ]

                note_tbl = Table(note_data, colWidths=[remarks_box_width])

                note_tbl.setStyle(TableStyle([
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]))

                story.append(note_tbl)
                story.append(Spacer(1, 6))
        else:
            empty_tbl = Table(
                [
                    [Paragraph("<b>Remarks:</b> -", remarks_style)],
                    [Paragraph("<b>Action Plan:</b> -", remarks_style)]
                ],
                colWidths=[remarks_box_width]
            )

            empty_tbl.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            story.append(empty_tbl)

        if course_index != len(course_summary_rows):
            story.append(PageBreak())

    # ---------------------------------------------------------
    # 4. Final Legend Page
    # ---------------------------------------------------------
    story.append(PageBreak())
    story.append(Paragraph("Question Reference / Legend", heading_style))
    story.append(Paragraph(f"Department: {department.Department}", info_style))
    story.append(Paragraph(
        f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
        f"Section: {sel_section or 'All'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or 'All'}",
        info_style
    ))
    story.append(Spacer(1, 8))

    legend_grouped = defaultdict(list)

    for item in question_legend:
        legend_grouped[item["category"]].append(item)

    legend_usable_w = doc.pagesize[0] - doc.leftMargin - doc.rightMargin

    for cat in legend_grouped:
        story.append(Paragraph(str(cat), legend_category_style))

        legend_data = [["No.", "Question"]]

        for item in legend_grouped[cat]:
            legend_data.append([
                str(item["number"]),
                Paragraph(item["question_text"], legend_text_style)
            ])

        legend_tbl = Table(
            legend_data,
            colWidths=[18 * mm, legend_usable_w - 18 * mm],
            repeatRows=1
        )

        legend_tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8.5),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 1), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8.3),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))

        story.append(legend_tbl)
        story.append(Spacer(1, 8))

    def _on_page(canv, doc_):
        page_w, page_h = doc_.pagesize
        _draw_rit_header_footer(
            canv,
            page_w,
            page_h,
            title="STUDENTS FEEDBACK",
            subtitle=subtitle
        )

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    return response



from collections import defaultdict
from datetime import datetime
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.shortcuts import get_object_or_404

from faculty_management.models import general_information
from user_accounts.models import Add_Department
from course_management.models import Course, CourseEnrollment, AssignSubjectFaculty
from feedback_management.models import (
    FeedbackQuestion,
    FeedbackSubmission,
    FeedbackAnswer,
    CourseFeedbackRemark,
)




from collections import defaultdict
from datetime import datetime
import re

from django.db.models import Q
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.shortcuts import get_object_or_404

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from faculty_management.models import general_information
from user_accounts.decorators import check_permission
from user_accounts.models import Add_Department
from course_management.models import Course, CourseEnrollment, AssignSubjectFaculty
from feedback_management.models import (
    FeedbackQuestion,
    FeedbackSubmission,
    FeedbackAnswer,
    CourseFeedbackRemark,
)







@check_permission("view_feedbacks")
def view_feedbacks_bulk_excel_download(request):
    user = request.user.Employee_id
    faculty = general_information.objects.select_related("department").get(faculty_id=user)
    faculty_department = faculty.department

    permission_scope = _get_feedback_permission_scope(request)
    if not permission_scope["has_access"]:
        return HttpResponseForbidden("You do not have permission to view feedback data.")

    sel_department_id = (request.GET.get("department_id") or "").strip()
    sel_batch = (request.GET.get("batch") or "").strip()
    sel_year = (request.GET.get("year") or "").strip()
    sel_sem = (request.GET.get("semester") or "").strip()
    sel_section = (request.GET.get("section") or "").strip()

    # ---------------------------------------------------------
    # Department resolve
    # ---------------------------------------------------------
    if permission_scope["can_view_all"]:
        if sel_department_id:
            department = Add_Department.objects.filter(
                id=sel_department_id,
                is_active=True
            ).first()
        else:
            department = faculty_department
    else:
        department = faculty_department

    if not department:
        return HttpResponseBadRequest("Invalid department.")

    if not (sel_year and sel_sem):
        return HttpResponseBadRequest("Select Year and Semester.")

    # ---------------------------------------------------------
    # Filtered courses
    # ---------------------------------------------------------
    courses_qs = Course.objects.filter(
        department=department,
        year=str(sel_year),
        semester=str(sel_sem),
        is_active=True
    )

    if sel_batch or sel_section:
        enroll_course_qs = CourseEnrollment.objects.filter(
            department=department,
            enroll=True,
        )

        if sel_batch:
            enroll_course_qs = enroll_course_qs.filter(batch=sel_batch)
        if sel_section:
            enroll_course_qs = enroll_course_qs.filter(section=sel_section)

        enroll_course_qs = enroll_course_qs.filter(
            student__year=str(sel_year),
            student__semester=str(sel_sem)
        )

        filtered_course_ids = enroll_course_qs.values_list("course_id", flat=True).distinct()
        courses_qs = courses_qs.filter(id__in=filtered_course_ids)

    courses = list(courses_qs.order_by("course_code"))

    if not courses:
        return HttpResponseBadRequest("No courses found for the selected filters.")

    # ---------------------------------------------------------
    # Questions
    # ---------------------------------------------------------
    ordered_questions = list(
        FeedbackQuestion.objects
        .filter(Q(department__isnull=True) | Q(department=department))
        .order_by("category", "id")
    )
    if not ordered_questions:
        return HttpResponseBadRequest("No feedback questions configured.")

    question_ids = [q.id for q in ordered_questions]

    # ---------------------------------------------------------
    # Category spans + question numbering
    # ---------------------------------------------------------
    category_spans = []
    question_number_map = {}
    question_legend = []

    cur = ordered_questions[0].category or "General"
    span = 0

    for idx, q in enumerate(ordered_questions, start=1):
        qcat = q.category or "General"

        question_number_map[q.id] = idx
        question_legend.append({
            "number": idx,
            "category": qcat,
            "question_text": q.question_text,
            "question_id": q.id,
        })

        if qcat != cur:
            category_spans.append({"category": cur, "span": span})
            cur = qcat
            span = 1
        else:
            span += 1

    category_spans.append({"category": cur, "span": span})

    academic_year_text = (
        f"Academic Year: {datetime.now().year}-{datetime.now().year + 1} "
        f"({'Odd' if str(sel_sem) in ['1', '3', '5', '7'] else 'Even'} Semester)"
    )

    # ---------------------------------------------------------
    # Workbook setup
    # ---------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    thin = Side(border_style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_fill = PatternFill("solid", fgColor="DCE6F1")
    head_fill = PatternFill("solid", fgColor="EEF2FF")
    sub_fill = PatternFill("solid", fgColor="F8FAFC")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    bold = Font(bold=True)
    bold_big = Font(bold=True, size=14)
    bold_mid = Font(bold=True, size=12)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def style_cell(cell, font=None, fill=None, alignment=None, border_on=True):
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border_on:
            cell.border = border

    def safe_sheet_name(name):
        name = re.sub(r'[:\\/*?\[\]]', '_', name or "Sheet")
        return name[:31] if len(name) > 31 else name

    # ---------------------------------------------------------
    # Summary sheet heading
    # ---------------------------------------------------------
    row_no = 1
    ws.cell(row=row_no, column=1, value="Students Feedback Bulk Report")
    style_cell(ws.cell(row=row_no, column=1), font=bold_big, fill=title_fill)
    row_no += 1

    ws.cell(row=row_no, column=1, value=f"Department: {department.Department}")
    row_no += 1
    ws.cell(row=row_no, column=1, value=f"Faculty: {faculty.name}")
    row_no += 1
    ws.cell(row=row_no, column=1, value=f"Year / Semester: {sel_year} / {sel_sem}")
    row_no += 1
    ws.cell(row=row_no, column=1, value=f"Section: {sel_section or '-'} | Batch: {sel_batch or '-'}")
    row_no += 1
    ws.cell(row=row_no, column=1, value=academic_year_text)
    row_no += 2

    summary_header_row_1 = row_no
    summary_header_row_2 = row_no + 1

    ws.cell(summary_header_row_1, 1, "Course Code")
    style_cell(ws.cell(summary_header_row_1, 1), font=bold, fill=head_fill, alignment=center)
    ws.merge_cells(start_row=summary_header_row_1, start_column=1, end_row=summary_header_row_2, end_column=1)

    ws.cell(summary_header_row_1, 2, "Course Title")
    style_cell(ws.cell(summary_header_row_1, 2), font=bold, fill=head_fill, alignment=center)
    ws.merge_cells(start_row=summary_header_row_1, start_column=2, end_row=summary_header_row_2, end_column=2)

    col_ptr = 3
    for c in category_spans:
        start_col = col_ptr
        end_col = col_ptr + c["span"] - 1

        ws.cell(summary_header_row_1, start_col, c["category"])
        style_cell(ws.cell(summary_header_row_1, start_col), font=bold, fill=head_fill, alignment=center)

        if start_col != end_col:
            ws.merge_cells(
                start_row=summary_header_row_1,
                start_column=start_col,
                end_row=summary_header_row_1,
                end_column=end_col
            )

        for ccol in range(start_col, end_col + 1):
            style_cell(ws.cell(summary_header_row_1, ccol), font=bold, fill=head_fill, alignment=center)

        col_ptr = end_col + 1

    avg_total_col = col_ptr
    students_col = col_ptr + 1

    ws.cell(summary_header_row_1, avg_total_col, "Avg Total")
    style_cell(ws.cell(summary_header_row_1, avg_total_col), font=bold, fill=head_fill, alignment=center)
    ws.merge_cells(
        start_row=summary_header_row_1,
        start_column=avg_total_col,
        end_row=summary_header_row_2,
        end_column=avg_total_col
    )

    ws.cell(summary_header_row_1, students_col, "Students")
    style_cell(ws.cell(summary_header_row_1, students_col), font=bold, fill=head_fill, alignment=center)
    ws.merge_cells(
        start_row=summary_header_row_1,
        start_column=students_col,
        end_row=summary_header_row_2,
        end_column=students_col
    )

    col_ptr = 3
    for q in ordered_questions:
        ws.cell(summary_header_row_2, col_ptr, f"Q{question_number_map[q.id]}")
        style_cell(ws.cell(summary_header_row_2, col_ptr), font=bold, fill=sub_fill, alignment=center)
        col_ptr += 1

    # ---------------------------------------------------------
    # Store summary data
    # ---------------------------------------------------------
    course_summary_rows = []

    # ---------------------------------------------------------
    # Build each course sheet
    # ---------------------------------------------------------
    for course_index, course in enumerate(courses, start=1):

        # -----------------------------------------------------
        # Faculty name for this filtered course
        # -----------------------------------------------------
        assign_qs = AssignSubjectFaculty.objects.select_related("faculty", "course").filter(
            department=department,
            course_id=course.id,
            is_active=True,
        )

        if sel_batch:
            assign_qs = assign_qs.filter(batch=sel_batch)
        if sel_section:
            assign_qs = assign_qs.filter(section=sel_section)

        assign_qs = assign_qs.filter(
            course__year=str(sel_year),
            course__semester=str(sel_sem)
        )

        assign_obj = assign_qs.first()
        mapped_faculty = assign_obj.faculty if assign_obj and assign_obj.faculty else None
        filtered_faculty_name = mapped_faculty.name if mapped_faculty else faculty.name

        # -----------------------------------------------------
        # Students
        # -----------------------------------------------------
        enroll_qs = CourseEnrollment.objects.select_related("student").filter(
            department=department,
            course_id=course.id,
            enroll=True,
        )

        if sel_batch:
            enroll_qs = enroll_qs.filter(batch=sel_batch)
        if sel_section:
            enroll_qs = enroll_qs.filter(section=sel_section)

        enroll_qs = enroll_qs.filter(
            student__year=str(sel_year),
            student__semester=str(sel_sem)
        )

        students = [e.student for e in enroll_qs.order_by("student__reg_no") if e.student]
        student_ids = [s.id for s in students]
        student_count = len(students)

        # -----------------------------------------------------
        # Submissions
        # -----------------------------------------------------
        submission_ids = []
        submissions_qs = FeedbackSubmission.objects.none()
        submission_map = {}

        if student_ids:
            submissions_qs = (
                FeedbackSubmission.objects
                .filter(
                    student_id__in=student_ids,
                    enrollment__course_id=course.id,
                    department=department,
                    enrollment__department=department,
                )
                .select_related("student", "course", "faculty", "enrollment")
            )
            submission_ids = list(submissions_qs.values_list("id", flat=True))
            submission_map = {sub.student_id: sub for sub in submissions_qs}

        # -----------------------------------------------------
        # Marks
        # -----------------------------------------------------
        student_q_marks = defaultdict(dict)
        student_total = defaultdict(int)
        course_question_totals = defaultdict(int)

        if submission_ids:
            answers = (
                FeedbackAnswer.objects
                .filter(submission_id__in=submission_ids, question_id__in=question_ids)
                .select_related("submission")
            )
            for ans in answers:
                sid = ans.submission.student_id
                qid = ans.question_id
                sc = int(ans.score or 0)

                student_q_marks[sid][qid] = sc
                student_total[sid] += sc
                course_question_totals[qid] += sc

        course_grand_total = sum(course_question_totals.values())

        # -----------------------------------------------------
        # Average summary
        # -----------------------------------------------------
        course_question_averages = {}
        if student_count > 0:
            for q in ordered_questions:
                course_question_averages[q.id] = round(
                    course_question_totals.get(q.id, 0) / student_count, 2
                )
            course_grand_average = round(course_grand_total / student_count, 2)
        else:
            for q in ordered_questions:
                course_question_averages[q.id] = 0
            course_grand_average = 0

        course_summary_rows.append({
            "course_code": course.course_code or "-",
            "course_title": course.title or "-",
            "student_count": student_count,
            "question_averages": course_question_averages,
            "grand_average": course_grand_average,
        })

        # -----------------------------------------------------
        # Course sheet
        # -----------------------------------------------------
        sheet_name = safe_sheet_name(f"{course.course_code}_{course_index}")
        cws = wb.create_sheet(title=sheet_name)

        r = 1
        cws.cell(r, 1, f"Course {course_index}: {course.course_code} - {course.title}")
        style_cell(cws.cell(r, 1), font=bold_big, fill=title_fill)
        r += 1

        cws.cell(r, 1, f"Department: {department.Department}")
        r += 1
        cws.cell(r, 1, f"Name of the Faculty: {filtered_faculty_name}")
        r += 1
        cws.cell(r, 1, f"Year / Semester: {sel_year} / {sel_sem}")
        r += 1
        cws.cell(r, 1, f"Section: {sel_section or '-'} | Batch: {sel_batch or '-'}")
        r += 1
        cws.cell(r, 1, academic_year_text)
        r += 2

        header_row_1 = r
        header_row_2 = r + 1

        cws.cell(header_row_1, 1, "Sl.No")
        style_cell(cws.cell(header_row_1, 1), font=bold, fill=head_fill, alignment=center)
        cws.merge_cells(start_row=header_row_1, start_column=1, end_row=header_row_2, end_column=1)

        col_ptr = 2
        for c in category_spans:
            start_col = col_ptr
            end_col = col_ptr + c["span"] - 1

            cws.cell(header_row_1, start_col, c["category"])
            style_cell(cws.cell(header_row_1, start_col), font=bold, fill=head_fill, alignment=center)

            if start_col != end_col:
                cws.merge_cells(
                    start_row=header_row_1,
                    start_column=start_col,
                    end_row=header_row_1,
                    end_column=end_col
                )

            for mcol in range(start_col, end_col + 1):
                style_cell(cws.cell(header_row_1, mcol), font=bold, fill=head_fill, alignment=center)

            col_ptr = end_col + 1

        total_col = col_ptr
        eff_col = col_ptr + 1
        sat_col = col_ptr + 2
        rec_col = col_ptr + 3
        open_col = col_ptr + 4

        extra_headers = [
            (total_col, "Total"),
            (eff_col, "Overall Effectiveness (%)"),
            (sat_col, "Student Satisfaction"),
            (rec_col, "Recommendation to Continue / Improve"),
            (open_col, "Open Comments for Improvement"),
        ]

        for col_no, head in extra_headers:
            cws.cell(header_row_1, col_no, head)
            style_cell(cws.cell(header_row_1, col_no), font=bold, fill=head_fill, alignment=center)
            cws.merge_cells(
                start_row=header_row_1,
                start_column=col_no,
                end_row=header_row_2,
                end_column=col_no
            )

        col_ptr = 2
        for q in ordered_questions:
            cws.cell(header_row_2, col_ptr, f"Q{question_number_map[q.id]}")
            style_cell(cws.cell(header_row_2, col_ptr), font=bold, fill=sub_fill, alignment=center)
            col_ptr += 1

        data_row = header_row_2 + 1

        if students:
            for i, st in enumerate(students, start=1):
                submission = submission_map.get(st.id)

                overall_effectiveness = (
                    submission.overall_effectiveness_percentage
                    if submission and submission.overall_effectiveness_percentage is not None
                    else "-"
                )
                student_satisfaction = (
                    "Yes" if submission and submission.student_satisfaction is True
                    else "No" if submission and submission.student_satisfaction is False
                    else "-"
                )
                recommendation = (
                    submission.recommendation_to_continue_improve
                    if submission and submission.recommendation_to_continue_improve
                    else "-"
                )
                open_comments = (
                    submission.open_comments_for_improvement
                    if submission and submission.open_comments_for_improvement
                    else "-"
                )

                cws.cell(data_row, 1, i)
                style_cell(cws.cell(data_row, 1), alignment=center)

                col_ptr = 2
                for q in ordered_questions:
                    cws.cell(data_row, col_ptr, student_q_marks.get(st.id, {}).get(q.id, 0))
                    style_cell(cws.cell(data_row, col_ptr), alignment=center)
                    col_ptr += 1

                cws.cell(data_row, total_col, student_total.get(st.id, 0))
                cws.cell(data_row, eff_col, overall_effectiveness)
                cws.cell(data_row, sat_col, student_satisfaction)
                cws.cell(data_row, rec_col, recommendation)
                cws.cell(data_row, open_col, open_comments)

                style_cell(cws.cell(data_row, total_col), font=bold, fill=sub_fill, alignment=center)
                style_cell(cws.cell(data_row, eff_col), alignment=center)
                style_cell(cws.cell(data_row, sat_col), alignment=center)
                style_cell(cws.cell(data_row, rec_col), alignment=left)
                style_cell(cws.cell(data_row, open_col), alignment=left)

                data_row += 1
        else:
            cws.cell(data_row, 1, "No students found.")
            cws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=open_col)
            for ccol in range(1, open_col + 1):
                style_cell(cws.cell(data_row, ccol), font=bold, alignment=center)
            data_row += 1

        # -----------------------------------------------------
        # Remarks / Action Taken
        # -----------------------------------------------------
        data_row += 2
        cws.cell(data_row, 1, "Remarks / Action Taken")
        style_cell(cws.cell(data_row, 1), font=bold_mid)
        data_row += 1

        remark_headers = ["S.No", "Faculty", "Remarks", "Action Plan"]
        for idx, head in enumerate(remark_headers, start=1):
            cws.cell(data_row, idx, head)
            style_cell(cws.cell(data_row, idx), font=bold, fill=sub_fill, alignment=center)

        all_feedback_notes = CourseFeedbackRemark.objects.select_related("faculty").filter(
            department=department,
            course=course
        ).order_by("-updated_at", "-id")

        data_row += 1
        if all_feedback_notes.exists():
            for note_index, note in enumerate(all_feedback_notes, start=1):
                note_faculty_name = note.faculty.name if note.faculty else "Unknown User"

                values = [
                    note_index,
                    note_faculty_name,
                    note.remarks if note.remarks else "-",
                    note.action_taken if note.action_taken else "-",
                ]
                for idx, val in enumerate(values, start=1):
                    cws.cell(data_row, idx, val)
                    style_cell(
                        cws.cell(data_row, idx),
                        alignment=center if idx == 1 else left,
                        fill=white_fill
                    )
                data_row += 1
        else:
            values = [1, "-", "-", "-"]
            for idx, val in enumerate(values, start=1):
                cws.cell(data_row, idx, val)
                style_cell(
                    cws.cell(data_row, idx),
                    alignment=center if idx == 1 else left,
                    fill=white_fill
                )

        # -----------------------------------------------------
        # Column widths
        # -----------------------------------------------------
        width_map = {
            1: 8,
        }

        for idx in range(2, total_col):
            width_map[idx] = 9

        width_map[total_col] = 10
        width_map[eff_col] = 18
        width_map[sat_col] = 18
        width_map[rec_col] = 35
        width_map[open_col] = 35

        for col_idx, width in width_map.items():
            cws.column_dimensions[get_column_letter(col_idx)].width = width

        cws.freeze_panes = f"A{header_row_2 + 1}"

    # ---------------------------------------------------------
    # Fill summary rows
    # ---------------------------------------------------------
    current_summary_row = summary_header_row_2 + 1
    for item in course_summary_rows:
        ws.cell(current_summary_row, 1, item["course_code"])
        ws.cell(current_summary_row, 2, item["course_title"])

        style_cell(ws.cell(current_summary_row, 1), font=bold, alignment=center)
        style_cell(ws.cell(current_summary_row, 2), alignment=left)

        col_ptr = 3
        for q in ordered_questions:
            ws.cell(current_summary_row, col_ptr, item["question_averages"].get(q.id, 0))
            style_cell(ws.cell(current_summary_row, col_ptr), alignment=center)
            col_ptr += 1

        ws.cell(current_summary_row, avg_total_col, item["grand_average"])
        ws.cell(current_summary_row, students_col, item["student_count"])
        style_cell(ws.cell(current_summary_row, avg_total_col), font=bold, fill=sub_fill, alignment=center)
        style_cell(ws.cell(current_summary_row, students_col), font=bold, fill=sub_fill, alignment=center)

        current_summary_row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 32

    for idx in range(3, avg_total_col):
        ws.column_dimensions[get_column_letter(idx)].width = 9

    ws.column_dimensions[get_column_letter(avg_total_col)].width = 12
    ws.column_dimensions[get_column_letter(students_col)].width = 12
    ws.freeze_panes = f"A{summary_header_row_2 + 1}"

    # ---------------------------------------------------------
    # Legend sheet
    # ---------------------------------------------------------
    lws = wb.create_sheet(title="Question Legend")
    lr = 1

    lws.cell(lr, 1, "Question Reference / Legend")
    style_cell(lws.cell(lr, 1), font=bold_big, fill=title_fill)
    lr += 1

    lws.cell(lr, 1, f"Department: {department.Department}")
    lr += 1
    lws.cell(lr, 1, f"Faculty: {faculty.name}")
    lr += 1
    lws.cell(lr, 1, f"Year / Semester: {sel_year} / {sel_sem}")
    lr += 1
    lws.cell(lr, 1, f"Section: {sel_section or '-'} | Batch: {sel_batch or '-'}")
    lr += 2

    lws.cell(lr, 1, "No.")
    lws.cell(lr, 2, "Category")
    lws.cell(lr, 3, "Question")

    for c in range(1, 4):
        style_cell(lws.cell(lr, c), font=bold, fill=head_fill, alignment=center)

    for idx, q in enumerate(ordered_questions, start=1):
        lr += 1
        lws.cell(lr, 1, idx)
        lws.cell(lr, 2, q.category or "General")
        lws.cell(lr, 3, q.question_text)

        style_cell(lws.cell(lr, 1), font=bold, alignment=center)
        style_cell(lws.cell(lr, 2), alignment=center)
        style_cell(lws.cell(lr, 3), alignment=left)

    lws.column_dimensions["A"].width = 10
    lws.column_dimensions["B"].width = 24
    lws.column_dimensions["C"].width = 90
    lws.freeze_panes = "A7"

    # ---------------------------------------------------------
    # Download response
    # ---------------------------------------------------------
    filename = f"Bulk_Feedbacks_{department.Department}_{sel_year}_Sem{sel_sem}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response














from collections import defaultdict
from datetime import datetime

from django.db.models import Q
from django.http import (
    HttpResponse,
    HttpResponseBadRequest,
    HttpResponseForbidden,
)

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


@check_permission("view_feedbacks")
def view_feedbacks_pdf(request):
    user = request.user.Employee_id
    faculty = general_information.objects.select_related("department").get(faculty_id=user)
    faculty_department = faculty.department

    permission_scope = _get_feedback_permission_scope(request)
    if not permission_scope["has_access"]:
        return HttpResponseForbidden("You do not have permission to view feedback data.")

    sel_department_id = (request.GET.get("department_id") or "").strip()
    sel_batch = (request.GET.get("batch") or "").strip()
    sel_year = (request.GET.get("year") or "").strip()
    sel_sem = (request.GET.get("semester") or "").strip()
    sel_section = (request.GET.get("section") or "").strip()
    sel_course_id = (request.GET.get("course_id") or "").strip()

    # ---------------------------------------------------------
    # Department resolve
    # ---------------------------------------------------------
    if permission_scope["can_view_all"]:
        if sel_department_id:
            department = Add_Department.objects.filter(
                id=sel_department_id,
                is_active=True
            ).first() or faculty_department
        else:
            department = faculty_department
    else:
        department = faculty_department

    if not department:
        return HttpResponseBadRequest("Invalid department.")

    if not (sel_year and sel_sem and sel_course_id):
        return HttpResponseBadRequest("Select Year, Semester and Course to generate PDF.")

    try:
        course = Course.objects.get(id=sel_course_id, department=department)
    except Course.DoesNotExist:
        return HttpResponseBadRequest("Invalid course.")

    # ---------------------------------------------------------
    # Get filtered faculty name for selected subject
    # ---------------------------------------------------------
    assign_qs = AssignSubjectFaculty.objects.select_related("faculty", "course").filter(
        department=department,
        course_id=sel_course_id,
        is_active=True,
    )

    if sel_batch:
        assign_qs = assign_qs.filter(batch=sel_batch)
    if sel_section:
        assign_qs = assign_qs.filter(section=sel_section)
    if sel_year:
        assign_qs = assign_qs.filter(course__year=str(sel_year))
    if sel_sem:
        assign_qs = assign_qs.filter(course__semester=str(sel_sem))

    assign_obj = assign_qs.first()
    mapped_faculty = assign_obj.faculty if assign_obj and assign_obj.faculty else None
    filtered_faculty_name = mapped_faculty.name if mapped_faculty else faculty.name

    # ---------------------------------------------------------
    # Enrolled students
    # ---------------------------------------------------------
    enroll_qs = CourseEnrollment.objects.select_related("student").filter(
        department=department,
        course_id=sel_course_id,
        enroll=True,
    )

    if sel_batch:
        enroll_qs = enroll_qs.filter(batch=sel_batch)
    if sel_year:
        enroll_qs = enroll_qs.filter(student__year=str(sel_year))
    if sel_sem:
        enroll_qs = enroll_qs.filter(student__semester=str(sel_sem))
    if sel_section:
        enroll_qs = enroll_qs.filter(section=sel_section)

    students = [e.student for e in enroll_qs.order_by("student__reg_no") if e.student]
    student_ids = [s.id for s in students]

    # ---------------------------------------------------------
    # Questions
    # common questions + selected department questions
    # ---------------------------------------------------------
    ordered_questions = list(
        FeedbackQuestion.objects
        .filter(Q(department__isnull=True) | Q(department=department))
        .order_by("category", "id")
    )
    if not ordered_questions:
        return HttpResponseBadRequest("No feedback questions configured.")

    question_ids = [q.id for q in ordered_questions]

    # ---------------------------------------------------------
    # Category spans + question numbering
    # ---------------------------------------------------------
    category_spans = []
    question_number_map = {}
    question_legend = []

    cur = ordered_questions[0].category or "General"
    span = 0
    for idx, q in enumerate(ordered_questions, start=1):
        qcat = q.category or "General"

        question_number_map[q.id] = idx
        question_legend.append({
            "number": idx,
            "category": qcat,
            "question_text": q.question_text,
            "question_id": q.id,
        })

        if qcat != cur:
            category_spans.append({"category": cur, "span": span})
            cur = qcat
            span = 1
        else:
            span += 1
    category_spans.append({"category": cur, "span": span})

    # ---------------------------------------------------------
    # Submissions
    # ---------------------------------------------------------
    submission_ids = []
    submissions_qs = FeedbackSubmission.objects.none()
    submission_map = {}

    if student_ids:
        submissions_qs = (
            FeedbackSubmission.objects
            .filter(
                student_id__in=student_ids,
                enrollment__course_id=sel_course_id,
                department=department,
                enrollment__department=department,
            )
            .select_related("student", "course", "faculty", "enrollment")
        )
        submission_ids = list(submissions_qs.values_list("id", flat=True))
        submission_map = {sub.student_id: sub for sub in submissions_qs}

    student_q_marks = defaultdict(dict)
    student_total = defaultdict(int)

    if submission_ids:
        answers = (
            FeedbackAnswer.objects
            .filter(submission_id__in=submission_ids, question_id__in=question_ids)
            .select_related("submission")
        )
        for ans in answers:
            sid = ans.submission.student_id
            qid = ans.question_id
            sc = int(ans.score or 0)
            student_q_marks[sid][qid] = sc
            student_total[sid] += sc

    # ---------------------------------------------------------
    # PDF setup
    # ---------------------------------------------------------
    PDF_PAGE_SIZE = landscape(A3)

    filename = f"View_Feedbacks_{course.course_code}_{sel_year}_Sem{sel_sem}.pdf"
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{filename}"'

    doc = SimpleDocTemplate(
        response,
        pagesize=PDF_PAGE_SIZE,
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=50 * mm,
        bottomMargin=20 * mm,
        title="Students Feedback"
    )

    styles = getSampleStyleSheet()

    p_center = ParagraphStyle(
        "p_center",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=5.2,
        leading=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a"),
    )

    p_center_bold = ParagraphStyle(
        "p_center_bold",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=5.4,
        leading=6,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a"),
    )

    p_small_left = ParagraphStyle(
        "p_small_left",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=5.8,
        leading=7,
        alignment=0,
        textColor=colors.HexColor("#0f172a"),
        wordWrap="CJK",
    )

    info_style = ParagraphStyle(
        "info",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#0f172a"),
        alignment=0,
    )

    heading_style = ParagraphStyle(
        "heading_style",
        parent=styles["Heading4"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=6,
    )

    sub_heading_style = ParagraphStyle(
        "sub_heading_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#111827"),
        spaceAfter=6,
    )

    remarks_style = ParagraphStyle(
        "remarks_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=12,
        alignment=0,
        textColor=colors.HexColor("#0f172a"),
    )

    legend_category_style = ParagraphStyle(
        "legend_category_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=4,
        spaceBefore=6,
    )

    legend_text_style = ParagraphStyle(
        "legend_text_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.8,
        leading=12,
        textColor=colors.HexColor("#111827"),
    )

    # ---------------------------------------------------------
    # Main feedback table data
    # ---------------------------------------------------------
    row1 = ["Sl.No"]
    for c in category_spans:
        row1.append(Paragraph(str(c["category"]), p_center))
        for _ in range(c["span"] - 1):
            row1.append("")
    row1.extend([
        "Total",
        Paragraph("Overall Effectiveness (%)", p_center),
        Paragraph("Student Satisfaction", p_center),
        Paragraph("Recommendation to Continue / Improve", p_center),
        Paragraph("Open Comments for Improvement", p_center),
    ])

    row2 = [""]
    for q in ordered_questions:
        row2.append(Paragraph(str(question_number_map.get(q.id, "")), p_center_bold))
    row2.extend(["", "", "", "", ""])

    data = [row1, row2]

    for i, st in enumerate(students, start=1):
        submission = submission_map.get(st.id)

        overall_effectiveness = (
            str(submission.overall_effectiveness_percentage)
            if submission and submission.overall_effectiveness_percentage is not None
            else "-"
        )
        student_satisfaction = (
            "Yes" if submission and submission.student_satisfaction is True
            else "No" if submission and submission.student_satisfaction is False
            else "-"
        )
        recommendation = (
            submission.recommendation_to_continue_improve
            if submission and submission.recommendation_to_continue_improve
            else "-"
        )
        open_comments = (
            submission.open_comments_for_improvement
            if submission and submission.open_comments_for_improvement
            else "-"
        )

        r = [str(i)]
        for q in ordered_questions:
            r.append(str(student_q_marks.get(st.id, {}).get(q.id, 0)))
        r.extend([
            str(student_total.get(st.id, 0)),
            overall_effectiveness,
            student_satisfaction,
            Paragraph(recommendation, p_small_left),
            Paragraph(open_comments, p_small_left),
        ])
        data.append(r)

    if not students:
        no_data_row = ["No students found."]
        no_data_row += [""] * len(ordered_questions)
        no_data_row += ["", "", "", "", ""]
        data.append(no_data_row)

    # ---------------------------------------------------------
    # Width calculation fix
    # ---------------------------------------------------------
    page_w, page_h = doc.pagesize
    usable_w = page_w - doc.leftMargin - doc.rightMargin

    col_w_sno = 10 * mm
    col_w_total = 12 * mm
    col_w_eff = 16 * mm
    col_w_sat = 18 * mm
    col_w_rec = 32 * mm
    col_w_open = 32 * mm

    q_count = len(ordered_questions)
    fixed_width = col_w_sno + col_w_total + col_w_eff + col_w_sat + col_w_rec + col_w_open
    q_w = (usable_w - fixed_width) / float(max(q_count, 1))
    q_w = max(q_w, 6 * mm)

    col_widths = (
        [col_w_sno] +
        [q_w] * q_count +
        [col_w_total, col_w_eff, col_w_sat, col_w_rec, col_w_open]
    )

    tbl = Table(data, colWidths=col_widths, repeatRows=2)

    table_style = TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d1d5db")),
        ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#f8fafc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 1), 5.5),
        ("FONTNAME", (0, 2), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 2), (-3, -1), 5.5),
        ("ALIGN", (0, 2), (-3, -1), "CENTER"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f8fafc")),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#111827")),

        ("ALIGN", (-2, 2), (-1, -1), "LEFT"),
        ("VALIGN", (-2, 2), (-1, -1), "TOP"),
        ("LEFTPADDING", (-2, 2), (-1, -1), 3),
        ("RIGHTPADDING", (-2, 2), (-1, -1), 3),

        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])

    table_style.add("SPAN", (0, 0), (0, 1))
    table_style.add("SPAN", (q_count + 1, 0), (q_count + 1, 1))
    table_style.add("SPAN", (q_count + 2, 0), (q_count + 2, 1))
    table_style.add("SPAN", (q_count + 3, 0), (q_count + 3, 1))
    table_style.add("SPAN", (q_count + 4, 0), (q_count + 4, 1))
    table_style.add("SPAN", (q_count + 5, 0), (q_count + 5, 1))

    start_col = 1
    for c in category_spans:
        end_col = start_col + c["span"] - 1
        table_style.add("SPAN", (start_col, 0), (end_col, 0))
        start_col = end_col + 1

    table_style.add("BACKGROUND", (q_count + 1, 2), (q_count + 1, -1), colors.HexColor("#f8fafc"))
    table_style.add("FONTNAME", (q_count + 1, 2), (q_count + 1, -1), "Helvetica-Bold")

    if not students:
        table_style.add("SPAN", (0, 2), (-1, 2))
        table_style.add("ALIGN", (0, 2), (-1, 2), "CENTER")
        table_style.add("FONTNAME", (0, 2), (-1, 2), "Helvetica-Oblique")

    tbl.setStyle(table_style)

    # ---------------------------------------------------------
    # Get ALL remarks / action rows for this course
    # ---------------------------------------------------------
    all_feedback_notes = CourseFeedbackRemark.objects.select_related("faculty").filter(
        department=department,
        course=course
    ).order_by("-updated_at", "-id")

    semester_type = "Odd" if str(sel_sem) in ["1", "3", "5", "7"] else "Even"
    academic_year_text = f"Academic Year: {get_academic_year()} ({semester_type} Semester)"
    subtitle = academic_year_text

    story = [
        Paragraph(f"Department: {department.Department}", info_style),
        Paragraph(f"Name of the Subject: {course.course_code} - {course.title}", info_style),
        Paragraph(f"Name of the Faculty: {filtered_faculty_name}", info_style),
        Paragraph(
            f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
            f"Section: {sel_section or '-'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or '-'}",
            info_style
        ),
        Spacer(1, 6),
        tbl,
        PageBreak(),
        Paragraph(f"{course.course_code} - {course.title}", heading_style),
        Paragraph("Remarks / Action Taken", sub_heading_style),
        Paragraph(f"Department: {department.Department}", info_style),
        Paragraph(f"Name of the Faculty: {filtered_faculty_name}", info_style),
        Paragraph(
            f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
            f"Section: {sel_section or '-'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or '-'}",
            info_style
        ),
        Spacer(1, 8),
    ]

    remarks_box_width = page_w - doc.leftMargin - doc.rightMargin

    if all_feedback_notes.exists():
        for note_index, note in enumerate(all_feedback_notes, start=1):
            note_faculty_name = note.faculty.name if note.faculty else "Unknown User"

            note_data = [
                [Paragraph(f"<b>{note_index}. Faculty:</b> {note_faculty_name}", remarks_style)],
                [Paragraph(f"<b>Remarks:</b> {note.remarks if note.remarks else '-'}", remarks_style)],
                [Paragraph(f"<b>Action Plan:</b> {note.action_taken if note.action_taken else '-'}", remarks_style)],
            ]

            note_tbl = Table(note_data, colWidths=[remarks_box_width])
            note_tbl.setStyle(TableStyle([
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))

            story.append(note_tbl)
            story.append(Spacer(1, 6))
    else:
        empty_tbl = Table(
            [
                [Paragraph("<b>Remarks:</b> -", remarks_style)],
                [Paragraph("<b>Action Plan:</b> -", remarks_style)],
            ],
            colWidths=[remarks_box_width]
        )
        empty_tbl.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(empty_tbl)

    # ---------------------------------------------------------
    # Final legend page
    # ---------------------------------------------------------
    story.append(PageBreak())
    story.append(Paragraph("Question Reference / Legend", heading_style))
    story.append(Paragraph(f"Department: {department.Department}", info_style))
    story.append(Paragraph(f"Name of the Subject: {course.course_code} - {course.title}", info_style))
    story.append(Paragraph(f"Name of the Faculty: {filtered_faculty_name}", info_style))
    story.append(Paragraph(
        f"Year/Semester: {sel_year} / {sel_sem} &nbsp;&nbsp;&nbsp; "
        f"Section: {sel_section or '-'} &nbsp;&nbsp;&nbsp; Batch: {sel_batch or '-'}",
        info_style
    ))
    story.append(Spacer(1, 8))

    legend_grouped = defaultdict(list)
    for item in question_legend:
        legend_grouped[item["category"]].append(item)

    legend_usable_w = page_w - doc.leftMargin - doc.rightMargin

    for cat in legend_grouped:
        story.append(Paragraph(str(cat), legend_category_style))

        legend_data = [["No.", "Question"]]
        for item in legend_grouped[cat]:
            legend_data.append([
                str(item["number"]),
                Paragraph(item["question_text"], legend_text_style)
            ])

        legend_tbl = Table(
            legend_data,
            colWidths=[18 * mm, legend_usable_w - (18 * mm)],
            repeatRows=1
        )
        legend_tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#d1d5db")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8.5),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (1, 1), (1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8.3),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ]))
        story.append(legend_tbl)
        story.append(Spacer(1, 8))

    def _first_page(canv, doc_):
        page_w_, page_h_ = doc_.pagesize
        _draw_rit_header_footer(canv, page_w_, page_h_, title="STUDENTS FEEDBACK", subtitle=subtitle)

    def _later_pages(canv, doc_):
        page_w_, page_h_ = doc_.pagesize
        _draw_rit_header_footer(canv, page_w_, page_h_, title="STUDENTS FEEDBACK", subtitle=subtitle)

    doc.build(story, onFirstPage=_first_page, onLaterPages=_later_pages)
    return response












from django.shortcuts import render,redirect
from django.contrib import messages
import re
from requests import request
from user_accounts.models import Role
from django.shortcuts import render, get_object_or_404
from course_management.models import *
from user_accounts.decorators import check_permission, is_super_user, no_cache, faculty_login_required
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
@faculty_login_required
@no_cache
@is_super_user('course_management')
def cm_assign_permission(request):
    if request.method == 'POST':  
        permissions = request.POST
        for role_name, role_permissions in permissions.items():
            if role_name.startswith('permissions'):
                try:
                    # Extract data from role_name using regex
                    extract_data = list(re.findall(r'\[([^\]]+)\]', role_name))
                    if len(extract_data) < 2:  # Ensure there are at least role and function
                        messages.warning(request,f"Invalid format in role_name: {role_name}. Skipping.")
                        continue
                    
                    extract_data.append(role_permissions)

                    # Retrieve the role (Handle if the role does not exist)
                    try:
                        role = Role.objects.using("rit_approval_system").get(role=extract_data[0])

                    except Role.DoesNotExist:
                        messages.error(request,f"Role {extract_data[0]} does not exist.")
                        messages.error(request, f"Role '{extract_data[0]}' does not exist. Skipping this entry.")
                        continue
                    
                    # Parse permissions - handle the case where role_permissions is a list (unlikely with POST data)
                    if isinstance(role_permissions, list):  # Handle list case
                        role_permissions = role_permissions[0]
                    
                    # Convert permission to boolean (True/False)
                    permission = extract_data[2] == 'true'
       
                    
                    # Find or create ApprovalPermissionFunction object
                    permission_obj = CourseandexaminationFunction.objects.filter(
                        role=role, function=extract_data[1]
                    ).first()
                    
                    if permission_obj:
                        permission_obj.permission = permission
                        permission_obj.save()
                    else:
                        # Create a new ApprovalPermissionFunction object
                        CourseandexaminationFunction.objects.create(
                            role=role,
                            function=extract_data[1],
                            permission=permission
                        )
                except Exception as e:
                    # Catch unexpected errors and log them
                    messages.error(request,f"Error processing role '{role_name}': {str(e)}")
                    messages.error(request, f"An error occurred while processing '{role_name}': {str(e)}")

    # Redirect to admin dashboard after processing
    messages.success(request,"The permission changes have been successfully applied.")
    return redirect('course_management')




from student_management.models import Student_cgpa
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import datetime, timedelta
from user_accounts.models import *
from student_management.models import *
from course_management.models import  Course, PassOutStudents, SectionMaster
from user_accounts.models import USER, PersonalDetails, AcademicDetails
from django.utils import timezone
from collections import defaultdict
from nba.models import SanctionedIntake

from course_management.models import Semester_Cooldown_Period

from collections import defaultdict
from datetime import timedelta
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

# If available, this gives true month-add/subtract (recommended)
try:
    from dateutil.relativedelta import relativedelta
except Exception:
    relativedelta = None

def get_year_from_semester(semester: int) -> int:
    return (semester + 1) // 2


def semester_upgrade(request):
    # ── Dropdown / static data ───────────────────────────────────────────────
    batches = StudentDetails.objects.values_list("batch", flat=True).distinct().order_by("batch")
    regulations = Regulations.objects.values_list("year", flat=True).distinct().order_by("year")
    semesters = list(range(1, 9))
    sections = SectionMaster.objects.all().order_by("section")

    selected_batch = request.GET.get("batch")

    # ── Cooldown configuration lookup (latest per degree) ────────────────────
    cooldown_qs = (
        Semester_Cooldown_Period.objects
        .select_related("degree")
        .order_by("degree_id", "-updated_at", "-created_at")
    )
    cooldown_months_by_degree = {}
    for row in cooldown_qs:
        if row.degree_id not in cooldown_months_by_degree:
            cooldown_months_by_degree[row.degree_id] = int(row.no_of_months or 0)

    DEFAULT_COOLDOWN_MONTHS = 4

    def _get_student_degree_id(student):
        if hasattr(student, 'degree_id') and student.degree_id:
            return student.degree_id
        if hasattr(student, 'department') and student.department and student.department.degree_id:
            return student.department.degree_id
        return None

    def _cooldown_months_for_student(student):
        deg_id = _get_student_degree_id(student)
        return cooldown_months_by_degree.get(deg_id, DEFAULT_COOLDOWN_MONTHS)

    def _add_months(dt, months):
        if not dt:
            return None
        try:
            return dt + relativedelta(months=months)
        except:
            return dt + timedelta(days=months * 30)  # rough fallback

    def _sub_months(dt, months):
        if not dt:
            return None
        try:
            return dt - relativedelta(months=months)
        except:
            return dt - timedelta(days=months * 30)

    # ── Section CRUD actions ─────────────────────────────────────────────────
    if request.method == "POST" and request.POST.get("section_action"):
        action = request.POST.get("section_action")
        sec_name = request.POST.get("section_name", "").strip()
        sec_id = request.POST.get("section_id")

        if action == "add" and sec_name:
            SectionMaster.objects.get_or_create(section=sec_name)
            messages.success(request, f"Section '{sec_name}' added.")
        elif action == "edit" and sec_id and sec_name:
            sec = get_object_or_404(SectionMaster, id=sec_id)
            sec.section = sec_name
            sec.save()
            messages.success(request, f"Section updated to '{sec_name}'.")
        elif action == "delete" and sec_id:
            sec = get_object_or_404(SectionMaster, id=sec_id)
            label = sec.section
            sec.delete()
            messages.warning(request, f"Section '{label}' deleted.")

        return redirect("semester_upgrade")

    # ── Main student data when batch is selected ─────────────────────────────
    department_students = defaultdict(list)
    current_semester = current_year = current_regulation = "Not assigned"
    last_upgrade_date = next_upgrade_date = None
    today = timezone.now().date()

    if selected_batch:
        students_qs = StudentDetails.objects.filter(batch=selected_batch,
        is_active=True)\
            .select_related("department", "department__degree")

        # Department display name cache
        dept_lookup = {
            d.id: (d.Department, d.Department_code)
            for d in Add_Department.objects.all()
        }

        for student in students_qs:
            dep_name, dep_code = dept_lookup.get(
                getattr(student, "department_id", None),
                ("Unknown", f"ID-{getattr(student, 'department_id', '??')}")
            )
            dept_display = f"{dep_name} ({dep_code})"

            cooldown_months = _cooldown_months_for_student(student)
            last_upd = getattr(student, "last_updated", None)
            ready_on = _add_months(last_upd, cooldown_months) if last_upd else None

            is_locked = bool(last_upd and ready_on and ready_on > timezone.now())
            remaining_days = None
            if is_locked and ready_on:
                remaining_days = max(0, (ready_on.date() - today).days)

            department_students[dept_display].append({
                "id": student.id,
                "reg_no": student.reg_no,
                "name": student.name,
                "section": student.section or "N/A",
                "batch": student.batch,
                "year": student.year or "Not assigned",
                "semester": student.semester or "Not assigned",
                "regulation": student.regulation or "Not assigned",

                "last_updated": last_upd,
                "cooldown_months": cooldown_months,
                "cooldown_locked": is_locked,
                "ready_on": ready_on,
                "remaining_days": remaining_days,
            })

        if students_qs.exists():
            first = students_qs.first()
            current_semester = first.semester or "Not assigned"
            current_year     = first.year     or "Not assigned"
            current_regulation = first.regulation or "Not assigned"

            # Informational: most recent update across the batch
            most_recent = students_qs.order_by("-last_updated").first()
            if most_recent and most_recent.last_updated:
                cm = _cooldown_months_for_student(most_recent)
                last_upgrade_date = most_recent.last_updated.strftime("%d-%m-%Y %H:%M")
                next_dt = _add_months(most_recent.last_updated, cm)
                next_upgrade_date = next_dt.strftime("%d-%m-%Y %H:%M") if next_dt else None

    # ── POST: Upgrade or Passout ─────────────────────────────────────────────
    if request.method == "POST" and not request.POST.get("section_action"):
        action = request.POST.get("action")
        batch = request.POST.get("batch")
        now = timezone.now()

        selected_ids = request.POST.getlist("selected_ids")
        if selected_ids:
            students = StudentDetails.objects.filter(id__in=selected_ids,
        is_active=True)\
                .select_related("department", "department__degree")
        else:
            students = StudentDetails.objects.filter(batch=batch,
        is_active=True)\
                .select_related("department", "department__degree")

        if action == "passout":
            moved = 0
            for s in students:
                dept = s.department
                if not dept:
                    continue

                sanctioned = SanctionedIntake.objects.filter(
                    department=dept,
                    year__lte=now.year,
                    degree=dept.degree
                ).order_by("-year").first()

                sanctioned_value = sanctioned.sanctioned_intake if sanctioned else 60
                pad_length = 3 if sanctioned_value >= 100 else 2

                reg_no = (s.reg_no or "").strip()
                last_digits = reg_no[-pad_length:] if reg_no and len(reg_no) >= pad_length else str(moved + 1).zfill(pad_length)

                dept_label = getattr(dept, "department_label", None) or getattr(dept, "Department_code", "XX")
                certificate_no = f"{now.year}/{dept_label}/{last_digits}"

                PassOutStudents.objects.get_or_create(
                    student=s,
                    department=dept,
                    defaults={
                        "year_of_passing": now.year,
                        "certificate_number": certificate_no,
                    }
                )
                moved += 1

            scope = "selected students" if selected_ids else f"batch {batch}"
            messages.success(request, f"{moved} students moved to PassOut from {scope}.")
            return redirect("semester_upgrade")

        elif action == "upgrade":
            regulation = request.POST.get("regulation")
            semester_str = request.POST.get("semester")
            semester = int(semester_str) if semester_str and semester_str.isdigit() else None

            if not semester or not regulation:
                messages.error(request, "Semester and Regulation are required.")
                return redirect("semester_upgrade")

            updated = []
            skipped = []

            for s in students:
                cooldown_months = _cooldown_months_for_student(s)
                cutoff = _sub_months(now, cooldown_months)

                if s.last_updated and cutoff and s.last_updated > cutoff:
                    ready_on_dt = _add_months(s.last_updated, cooldown_months)
                    skipped.append({
                        "reg_no": s.reg_no,
                        "name": s.name,
                        "months": cooldown_months,
                        "ready_on": ready_on_dt.strftime("%d-%m-%Y %H:%M") if ready_on_dt else "--",
                    })
                    continue

                s.semester = semester
                s.year = get_year_from_semester(semester)
                s.regulation = regulation
                # s.section = request.POST.get("section") or s.section   # optional — uncomment if needed

                s.last_updated = now
                s.save(update_fields=["semester", "year", "regulation", "last_updated"])

                updated.append({"reg_no": s.reg_no, "name": s.name})

            scope = "selected students" if selected_ids else f"batch {batch}"
            if updated:
                messages.success(
                    request,
                    f"Upgraded to Semester {semester} (Year {get_year_from_semester(semester)}) "
                    f"for {len(updated)} {scope}."
                )
            if skipped:
                preview = ", ".join(f"{x['reg_no']} ({x['ready_on']})" for x in skipped[:5])
                more = f" + {len(skipped)-5} more" if len(skipped) > 5 else ""
                messages.warning(request, f"{len(skipped)} students skipped (cooldown): {preview}{more}")

            return redirect("semester_upgrade")

    # ── Context for template ─────────────────────────────────────────────────
    context = {
        "batches": batches,
        "regulations": regulations,
        "semesters": semesters,
        "sections": sections,
        "selected_batch": selected_batch,
        "current_semester": current_semester,
        "current_year": current_year,
        "current_regulation": current_regulation,
        "last_upgrade_date": last_upgrade_date,
        "next_upgrade_date": next_upgrade_date,
        "department_students": dict(department_students),
        "default_cooldown_months": DEFAULT_COOLDOWN_MONTHS,
    }

    return render(request, "course_management/semester_upgrade.html", context)
# def admin_course_management(request):
#     departments = AddDepartment.objects.all()
#     department_stats = {}

#     for dept in departments:
#         courses = AddCourse.objects.filter(department=dept)
#         course_ids = courses.values_list('id', flat=True)
#         semester_courses = SemesterCourse.objects.filter(course_id__in=course_ids)

#         # Initialize stats dictionary
#         stats = {
#             'total_credits': 0,
#             'integrated_credits': 0,  # Both lecture and lab
#             'theory_credits': 0,      # Only lecture
#             'laboratory_credits': 0,   # Only lab
#             'hsmc_credits': 0,        # HSMC electives
#             'bsc_credits': 0,         # BSC electives
#             'eec_credits': 0,         # EEC electives
#             'pcc_credits': 0,         # PCC electives
#             'mc_credits': 0,          # MC electives
#             'oec_credits': 0,         # OEC electives
#             'pec_credits': 0,         # PEC electives
#         }

#         for sc in semester_courses:
#             try:
#                 # Convert credits to float, default to 0 if empty/invalid
#                 credits = float(sc.credits) if sc.credits and sc.credits.strip() else 0
                
#                 # Check for lecture and lab hours
#                 # Convert to float and check if greater than 0
#                 lecture_hours = float(sc.leture_hpwk) if sc.leture_hpwk and sc.leture_hpwk.strip() else 0
#                 lab_hours = float(sc.laboratory_hpwk) if sc.laboratory_hpwk and sc.laboratory_hpwk.strip() else 0

#                 # Add to total credits
#                 stats['total_credits'] += credits

#                 # Categorize based on lecture/lab hours
#                 if lecture_hours > 0 and lab_hours > 0:
#                     stats['integrated_credits'] += credits
#                 elif lecture_hours > 0 and lab_hours == 0:
#                     stats['theory_credits'] += credits
#                 elif lecture_hours == 0 and lab_hours > 0:
#                     stats['laboratory_credits'] += credits

#                 # Categorize elective credits if course has elective type
#                 if sc.course.elective:
#                     elective_type = sc.course.elective.upper().strip()
#                     if elective_type == 'HSMC':
#                         stats['hsmc_credits'] += credits
#                     elif elective_type == 'BSC':
#                         stats['bsc_credits'] += credits
#                     elif elective_type == 'EEC':
#                         stats['eec_credits'] += credits
#                     elif elective_type == 'PCC':
#                         stats['pcc_credits'] += credits
#                     elif elective_type == 'MC':
#                         stats['mc_credits'] += credits
#                     elif elective_type == 'OEC':
#                         stats['oec_credits'] += credits
#                     elif elective_type == 'PEC':
#                         stats['pec_credits'] += credits

#             except (ValueError, TypeError):
#                 # Skip this course if there are any conversion errors
#                 continue

#         # Round all values to 1 decimal place and store in department_stats
#         department_stats[dept.id] = {
#             'name': dept.department_name,
#             'total_credits': round(stats['total_credits'], 1),
#             'integrated_credits': round(stats['integrated_credits'], 1),
#             'theory_credits': round(stats['theory_credits'], 1),
#             'laboratory_credits': round(stats['laboratory_credits'], 1),
#             'hsmc_credits': round(stats['hsmc_credits'], 1),
#             'bsc_credits': round(stats['bsc_credits'], 1),
#             'eec_credits': round(stats['eec_credits'], 1),
#             'pcc_credits': round(stats['pcc_credits'], 1),
#             'mc_credits': round(stats['mc_credits'], 1),
#             'oec_credits': round(stats['oec_credits'], 1),
#             'pec_credits': round(stats['pec_credits'], 1),
#         }

#     return render(request, "course_management/admin_course_management.html", {
#         'department_stats': department_stats
#     })

# from course_management.models import AddRegulation



# def add_regulation(request):
#     message = ""
    
#     if request.method == 'POST':
#         action = request.POST.get('action')
#         year = request.POST.get('year')
#         new_year = request.POST.get('new_year')  # for editing

#         if action == 'add' and year:
#             obj, created = AddRegulation.objects.get_or_create(year=year)
#             message = "Added successfully" if created else "Regulation already exists."

#         elif action == 'edit' and year and new_year:
#             try:
#                 regulation = AddRegulation.objects.get(year=year)
#                 regulation.year = new_year
#                 regulation.save()
#                 message = "Edited successfully"
#             except AddRegulation.DoesNotExist:
#                 message = "Regulation not found"

#         elif action == 'delete' and year:
#             try:
#                 regulation = AddRegulation.objects.get(year=year)
#                 regulation.delete()
#                 message = "Deleted successfully"
#             except AddRegulation.DoesNotExist:
#                 message = "Regulation not found"

#     regulations = AddRegulation.objects.all().order_by('year')
#     return render(request, "course_management/add_regulation.html", {"regulations": regulations, "message": message})


# from faculty_management.models import AddD/epartment

from django.shortcuts import render, redirect
from django.contrib import messages
from course_management.models import Regulations


import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook


def add_regulation(request):
    # ---------- EXPORT CSV ----------
    if request.method == "GET" and request.GET.get("export") == "csv":
        regulations = Regulations.objects.all().order_by("year")

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="regulations.csv"'

        writer = csv.writer(response)
        writer.writerow(["regulation_number", "year"])

        for reg in regulations:
            writer.writerow([reg.regulation_number, reg.year])

        return response

    # ---------- EXPORT EXCEL ----------
    if request.method == "GET" and request.GET.get("export") == "excel":
        regulations = Regulations.objects.all().order_by("year")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Regulations"

        # Header row
        worksheet.append(["regulation_number", "year"])

        # Data rows
        for reg in regulations:
            worksheet.append([reg.regulation_number, reg.year])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="regulations.xlsx"'

        workbook.save(response)
        return response

    if request.method == "POST":
        action = request.POST.get("action")

        # ---------- ADD ----------
        if action == "add":
            regulation_number = request.POST.get("regulation_number")
            year = request.POST.get("year")

            if not regulation_number or not year:
                messages.error(request, "All fields are required.")
            elif Regulations.objects.filter(year=year).exists():
                messages.warning(request, "⚠ Regulation for this year already exists.")
            else:
                Regulations.objects.create(
                    regulation_number=int(regulation_number),
                    year=int(year)
                )
                messages.success(request, "Regulation added successfully.")

        # ---------- EDIT ----------
        elif action == "edit":
            reg_id = request.POST.get("reg_id")
            new_regulation_number = request.POST.get("new_regulation_number")
            new_year = request.POST.get("new_year")

            if not reg_id or not new_regulation_number or not new_year:
                messages.error(request, "Invalid edit request.")
                return redirect("add_regulation")

            regulation = get_object_or_404(Regulations, id=reg_id)

            # Prevent duplicate year (except same record)
            if Regulations.objects.exclude(id=reg_id).filter(year=new_year).exists():
                messages.warning(request, "⚠ Regulation for this year already exists.")
                return redirect("add_regulation")

            regulation.regulation_number = int(new_regulation_number)
            regulation.year = int(new_year)
            regulation.save()

            messages.success(request, "Regulation updated successfully.")

        # ---------- DELETE ----------
        elif action == "delete":
            reg_id = request.POST.get("reg_id")
            Regulations.objects.filter(id=reg_id).delete()
            messages.success(request, "🗑 Regulation deleted successfully.")

        return redirect("add_regulation")

    regulations = Regulations.objects.all().order_by("year")
    return render(
        request,
        "course_management/admin/add_regulation.html",
        {"regulations": regulations}
    )





from django.shortcuts import render, redirect
from django.contrib import messages
from user_accounts.models import Department
from django.db.models import Q
from course_management.models import Course_category

@check_permission("add_new_course")
def add_new_course(request):
    degrees = Degree.objects.filter(is_active=True)

    # ✅ Department comes from user (NOT UI)
    user_department = general_information.objects.filter(faculty_id=request.user.Employee_id).first().department
    if not user_department:
        messages.error(request, "Your account is not mapped to any active department. Contact admin.")
        return redirect("home")  # or redirect('add_new_course')

    if request.method == 'POST':
        regulation_id = request.POST.get("regulation")
        year_number = request.POST.get("year")
        elective_id = request.POST.get("elective")
        semester_number = request.POST.get("semester")
        course_codes = request.POST.getlist("course_code[]")
        course_names = request.POST.getlist("course_name[]")

        # Validate required
        if not regulation_id or not elective_id or not year_number or not semester_number:
            messages.error(request, "Please fill Degree/Regulation/Category/Year/Semester.")
            return redirect('add_new_course')

        try:
            course_category = Course_category.objects.get(id=elective_id)
        except Course_category.DoesNotExist:
            messages.error(request, "Invalid category selected.")
            return redirect('add_new_course')

        try:
            regulation = Regulations.objects.get(id=regulation_id)
        except Regulations.DoesNotExist:
            messages.error(request, "Invalid regulation selected.")
            return redirect('add_new_course')

        created_count, updated_count = 0, 0

        # ✅ Only ONE department (logged-in user's)
        for code, name in zip(course_codes, course_names):
            code = (code or '').strip()
            name = (name or '').strip()
            if not code or not name:
                continue

            course_obj, created = Course.objects.update_or_create(
                department=user_department,
                regulation=regulation,
                year=year_number,
                semester=semester_number,
                course_code=code,
                defaults={'title': name, 'elective': course_category}
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

        messages.success(
            request,
            f"✅ {created_count} course(s) added, {updated_count} course(s) updated for {user_department.Department}."
        )
        return redirect('add_new_course')

    # -------- GET (preview existing courses) --------
    regulations = Regulations.objects.all()
    course_category = Course_category.objects.all()

    q_degree = request.GET.get('degree')
    q_regulation = request.GET.get('regulation')
    q_category = request.GET.get('elective')
    q_year = request.GET.get('year')
    q_semester = request.GET.get('semester')

    existing_courses = None
    if all([q_regulation, q_category, q_year, q_semester]):
        qs = Course.objects.select_related('department', 'regulation', 'elective').filter(
            department=user_department
        )
        try:
            qs = qs.filter(regulation_id=int(q_regulation), elective_id=int(q_category))
        except (TypeError, ValueError):
            qs = Course.objects.none()

        qs = qs.filter(year=str(q_year), semester=str(q_semester))
        existing_courses = qs.order_by('course_code')

    context = {
        'degrees': degrees,
        'regulations': regulations,
        'course_category': course_category,
        'existing_courses': existing_courses,

        # show user dept (optional)
        'user_department': user_department,

        # Echo selections
        'sel_degree': q_degree or '',
        'sel_regulation': q_regulation or '',
        'sel_category': q_category or '',
        'sel_year': q_year or '',
        'sel_semester': q_semester or '',
    }
    return render(request, "course_management/add_course.html", context)
 



def get_departments(request):
    degree_id = request.GET.get('degree_id')
    degree = get_object_or_404(Degree, id=degree_id)
    departments = Add_Department.objects.filter(
        degree_id=degree_id,
        is_active=True
    )

    data = {
        "duration": degree.duration,   # 👈 KEY PART
        "departments": [
            {"id": d.id, "name": d.Department}
            for d in departments
        ]
    }

    return JsonResponse(data, safe=False)

 

 

def get_course_categories(request):
    regulation_id = request.GET.get('regulation_id')
    categories = Course_category.objects.filter(regulation_id=regulation_id)
    data = [{'id': cat.id, 'name': cat.Course_category_name} for cat in categories]
    return JsonResponse(data, safe=False)






# ---------------- AJAX: Get Programmes for Selected Degree ----------------
def ajax_get_programmes(request):
    degree_id = request.GET.get('degree_id')
    if not degree_id:
        return JsonResponse([], safe=False)

    programmes = DegreeDepartment.objects.filter(degree_id=degree_id)
    data = [{'id': p.id, 'name': p.degree_department} for p in programmes]
    return JsonResponse(data, safe=False)



def ajax_get_departments(request):
    degree_id = request.GET.get('degree_id')
    if not degree_id:
        return JsonResponse([], safe=False)

    departments = Add_Department.objects.filter(degree_id=degree_id)
    data = [{'id': d.id, 'name': d.Department} for d in departments]
    return JsonResponse(data, safe=False)



def ajax_get_departments(request):
    degree_id = request.GET.get("degree_id")
    if not degree_id:
        return JsonResponse([], safe=False)
    
    depts = DegreeDepartment.objects.filter(degree_code=degree_id)
    data = [{"id": d.department_id, "name": d.degree_department} for d in depts]
    return JsonResponse(data, safe=False)
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib import messages

def admin_update_course(request, course_id=None, action=None):
    # ===================== INITIAL DATA =====================
    courses = Course.objects.select_related(
        'department', 'regulation', 'elective'
    ).all()

    departments = Add_Department.objects.filter(is_active=True)
    degree = Degree.objects.filter(is_active=True)
    regulations = Regulations.objects.all()

    course = None
    if course_id and action == 'edit':
        course = get_object_or_404(Course, id=course_id)

    # ===================== POST (UPDATE / DELETE) =====================
    if request.method == "POST":
        action = request.POST.get("action")
        course_id = request.POST.get("course_id") or course_id

        if action not in ["update", "delete"]:
            messages.error(request, "Invalid action specified.")
            return redirect('admin_update_course')

        if not course_id:
            messages.error(request, "No course ID provided.")
            return redirect('admin_update_course')

        course = get_object_or_404(Course, id=course_id)

        if action == "update":
            course_code = request.POST.get("course_code")
            title = request.POST.get("title")
            is_active = request.POST.get("is_active") == "on"

            if not all([course_code, title]):
                messages.error(request, "Please fill in all required fields.")
                return render(request, 'course_management/admin_update_course.html', {
                    'courses': courses,
                    'degrees': degree,
                    'departments': departments,
                    'course': course,
                    'regulations': regulations
                })

            course.course_code = course_code
            course.title = title
            course.is_active = is_active
            course.save()

            messages.success(request, f"Course '{course.title}' updated successfully.")
            return redirect('admin_update_course')

        elif action == "delete":
            course_title = course.title
            course.delete()
            messages.success(request, f"Course '{course_title}' deleted successfully.")
            return redirect('admin_update_course')

    # ===================== AJAX FILTER =====================
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        courses = Course.objects.select_related(
            'department', 'regulation', 'elective'
        ).all()

        degree_id = request.GET.get('degree')
        department_id = request.GET.get('department')
        regulation_id = request.GET.get('regulation')
        year = request.GET.get('year')
        semester = request.GET.get('semester')

        if degree_id:
            courses = courses.filter(department__degree_id=degree_id)

        if department_id:
            courses = courses.filter(department_id=department_id)

        if regulation_id:
            courses = courses.filter(regulation_id=regulation_id)

        if year:
            courses = courses.filter(year=year)

        if semester:
            courses = courses.filter(semester=semester)

        data = []
        for c in courses:
            data.append({
                'id': c.id,
                'course_code': c.course_code,
                'title': c.title,
                'department': c.department.Department,
                'regulation': c.regulation.year,
                'year': c.year,
                'semester': c.semester,
                'category': c.elective.Course_category_name,
                'is_active': c.is_active,
            })

        return JsonResponse(data, safe=False)



    

    # ===================== Normal Page Render =====================
    regulations = Regulations.objects.all()
    
    context = {
        'courses': courses,
        'degrees': degree,
        'departments': departments,
        'course': course,
        'regulations': regulations,
    }
    return render(request, 'course_management/admin_update_course.html', context)






from django.http import JsonResponse
from django.shortcuts import render
from django.http import JsonResponse
from django.shortcuts import render

@check_permission("add_hours")
def add_hours(request):
    department = Add_Department.objects.filter(is_active=True)

    if request.method == "POST":
        department_id = request.POST.get("department")
        year = request.POST.get("year")
        regulation_year = request.POST.get("regulation_year")
        semester_number = request.POST.get("semester")

        course_code = request.POST.getlist("course_code[]") or []
        course_title = request.POST.getlist("course_title[]") or []

        hour_config_ids = request.POST.getlist("hour_config_id[]") or []

        lecture_npwk = request.POST.getlist("leture_npwk[]") or []
        tutorial_npwk = request.POST.getlist("tutorial_npwk[]") or []
        laboratory_npwk = request.POST.getlist("laboratory_npwk[]") or []

        total_hours = request.POST.getlist("total_hours[]") or []
        credits = request.POST.getlist("credits[]") or []
        regulation = Regulations.objects.filter(year=regulation_year).first()
        for i in range(len(course_code)):
            try:
                course = Course.objects.get(
                    course_code=course_code[i],
                    title=course_title[i],
                    department_id=department_id,
                    year=year,
                    semester=semester_number,
                    regulation=regulation,
                    is_active = True

                )
            except Course.DoesNotExist:
                continue

            cfg = None
            cfg_id = (hour_config_ids[i] if i < len(hour_config_ids) else None)
            if cfg_id:
                cfg = CourseHourConfig.objects.filter(id=cfg_id).first()

            defaults = {
                "hour_config": cfg,
                "leture_npwk": lecture_npwk[i] if i < len(lecture_npwk) else None,
                "tutorial_npwk": tutorial_npwk[i] if i < len(tutorial_npwk) else None,
                "laboratory_npwk": laboratory_npwk[i] if i < len(laboratory_npwk) else None,
                "total_hours": total_hours[i] if i < len(total_hours) else None,
                "credits": credits[i] if i < len(credits) else None,
            }

            semester_course, created = CourseHours.objects.get_or_create(course=course, defaults=defaults)

            if not created:
                for k, v in defaults.items():
                    setattr(semester_course, k, v)
                semester_course.save()

        return JsonResponse({'status': 'success', 'message': 'Hours updated successfully'})

    regulations = Regulations.objects.all()
    return render(request, "course_management/add_hours.html", {
        "department": department,
        "regulations": regulations
    })





def get_courses(request):
    department_id = request.GET.get("department_id")
    year = request.GET.get("year")
    semester = request.GET.get("semester")
    regulation = request.GET.get("regulation_year")

    if not (department_id and year and semester):
        return JsonResponse({"courses": []}, safe=False)
    regulation = Regulations.objects.filter(year=regulation).first()

    courses = Course.objects.filter(
        department_id=department_id,
        year=year,
        semester=semester,
        regulation=regulation,
        is_active = True
    ).values("id", "course_code", "title")

    courses_list = list(courses)

    semester_courses = CourseHours.objects.filter(course_id__in=[c["id"] for c in courses_list])
    semester_course_map = {sc.course_id: sc for sc in semester_courses}

    full_courses_list = []
    for course in courses_list:
        sc = semester_course_map.get(course["id"])
        full_courses_list.append({
            "id": course["id"],
            "course_code": course["course_code"],
            "title": course["title"],

            # ✅ send saved config id
            "hour_config_id": getattr(sc, "hour_config_id", ""),

            "lecture_npwk": getattr(sc, "leture_npwk", ""),
            "tutorial_npwk": getattr(sc, "tutorial_npwk", ""),
            "laboratory_npwk": getattr(sc, "laboratory_npwk", ""),
            "total_hours": getattr(sc, "total_hours", ""),
            "credits": getattr(sc, "credits", ""),
        })

    return JsonResponse({"courses": full_courses_list}, safe=False)


# ============================================================
# HOUR ALLOCATION  (Course & Examination → sidebar)
# ------------------------------------------------------------
# Department-scoped, inline-editable listing of each course's hour
# allocation (course_management_coursehours) joined with its linked
# hour configuration (examination_management_coursehourconfig).
# Only courses mapped to the logged-in user's department are shown.
# ============================================================
from course_management.decorators import course_management


def _user_add_department(request):
    """
    Resolve the logged-in user's Add_Department.

    Course.department is an FK to Add_Department, but request.user.Department
    is the control-room Department model — so we bridge across by code, then
    by name. Returns an Add_Department instance or None.
    """
    udept = getattr(request.user, "Department", None)
    if udept is None:
        return None
    code = (getattr(udept, "Department_code", "") or "").strip()
    name = (getattr(udept, "Department", "") or "").strip()
    add_dept = None
    if code:
        add_dept = Add_Department.objects.filter(Department_code=code).first()
    if add_dept is None and name:
        add_dept = Add_Department.objects.filter(Department=name).first()
    return add_dept


@check_permission("hour_allocation")
@course_management
def hour_allocation(request):
    """
    Course Management -> Hour Allocation.
    Shows only the courses mapped to the logged-in user's department, with
    their hour allocation (course_management_coursehours) and linked hour
    config (examination_management_coursehourconfig). Both are editable inline.
    """
    from collections import defaultdict

    # ---- Department scope ----
    # Course.department uses Add_Department; map the logged-in user's
    # control-room Department to it. HODs are locked to their own department;
    # superusers may pick one via ?department=.
    department_id = (request.GET.get("department") or "").strip()
    is_super = getattr(request.user, "is_superuser", False)

    qs = (
        CourseHours.objects
        .select_related(
            "course",
            "course__department",
            "course__regulation",
            "hour_config",
            "hour_config__degree",
            "hour_config__regulation",
        )
        .filter(course__isnull=False)
        .order_by("course__year", "course__semester", "course__course_code")
    )

    if not is_super:
        # Regular user (e.g. HOD) -> only their own department's courses.
        add_dept = _user_add_department(request)
        dept_locked = True
        if add_dept is not None:
            qs = qs.filter(course__department=add_dept)
            dept_name = add_dept.Department or ""
        else:
            qs = qs.none()
            raw_dept = getattr(request.user, "Department", None)
            dept_name = getattr(raw_dept, "Department", "") or ""
    else:
        # Superuser -> choose a department to view.
        dept_locked = False
        if department_id:
            qs = qs.filter(course__department_id=department_id)
            dept_obj = Add_Department.objects.filter(id=department_id).first()
            dept_name = getattr(dept_obj, "Department", "") or ""
        else:
            qs = qs.none()
            dept_name = ""

    course_hours = list(qs)
    course_ids = [ch.course_id for ch in course_hours if ch.course_id]

    # ---- Faculty assigned per course (so the search can match faculty name) ----
    fac_by_course = defaultdict(list)
    seen_fac = defaultdict(set)
    if course_ids:
        for a in (
            AssignSubjectFaculty.objects
            .filter(course_id__in=course_ids, is_active=True)
            .select_related("faculty")
        ):
            fname = getattr(a.faculty, "name", None)
            if fname and fname not in seen_fac[a.course_id]:
                seen_fac[a.course_id].add(fname)
                fac_by_course[a.course_id].append(fname)

    rows = []
    with_config = 0
    for ch in course_hours:
        course = ch.course
        cfg = ch.hour_config
        if cfg is not None:
            with_config += 1
        rows.append({
            "ch_id": ch.id,
            "cfg_id": cfg.id if cfg else "",
            "course_code": getattr(course, "course_code", "") or "",
            "title": getattr(course, "title", "") or "",
            "regulation": getattr(getattr(course, "regulation", None), "year", "") or "",
            "year": getattr(course, "year", "") or "",
            "semester": getattr(course, "semester", "") or "",
            "faculty": ", ".join(fac_by_course.get(ch.course_id, [])),

            # ---- course_management_coursehours ----
            "lecture_npwk": ch.leture_npwk or "",
            "tutorial_npwk": ch.tutorial_npwk or "",
            "laboratory_npwk": ch.laboratory_npwk or "",
            "total_hours": ch.total_hours or "",
            "credits": ch.credits or "",

            # ---- examination_management_coursehourconfig ----
            "has_config": cfg is not None,
            "cfg_degree": getattr(getattr(cfg, "degree", None), "degree", "") if cfg else "",
            "cfg_regulation": getattr(getattr(cfg, "regulation", None), "year", "") if cfg else "",
            "cfg_lecture": getattr(cfg, "lecture_hours", "") if cfg else "",
            "cfg_tutorial": getattr(cfg, "tutorial_hours", "") if cfg else "",
            "cfg_laboratory": getattr(cfg, "laboratory_hours", "") if cfg else "",
            "cfg_theory_pct": getattr(cfg, "theory_percentage", "") if cfg else "",
            "cfg_practical_pct": getattr(cfg, "practical_percentage", "") if cfg else "",
            "cfg_activity_pct": getattr(cfg, "activity_percentage", "") if cfg else "",
        })

    context = {
        "rows": rows,
        "total_count": len(rows),
        "with_config": with_config,
        "without_config": len(rows) - with_config,
        "dept_name": dept_name,
        "dept_locked": dept_locked,
        "departments": Add_Department.objects.filter(is_active=True).order_by("Department"),
        "selected_department": department_id,
    }
    return render(request, "course_management/hour_allocation.html", context)


@check_permission("hour_allocation")
@course_management
def hour_allocation_save(request):
    """
    AJAX save for a single Hour Allocation row.
    Updates the CourseHours record and (when linked) its CourseHourConfig.
    The row must belong to the logged-in user's department.
    """
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid request method."}, status=405)

    ch_id = (request.POST.get("ch_id") or "").strip()
    if not ch_id:
        return JsonResponse({"status": "error", "message": "Missing record id."}, status=400)

    ch = (
        CourseHours.objects
        .select_related("course", "course__department", "hour_config")
        .filter(id=ch_id)
        .first()
    )
    if not ch:
        return JsonResponse({"status": "error", "message": "Record not found."}, status=404)

    # ---- Security: the row must be in the user's department ----
    if not getattr(request.user, "is_superuser", False):
        add_dept = _user_add_department(request)
        if add_dept is None or getattr(ch.course, "department_id", None) != add_dept.id:
            return JsonResponse({"status": "error", "message": "Not allowed for this department."}, status=403)

    def _val(name):
        v = request.POST.get(name)
        return v.strip() if isinstance(v, str) else v

    # ---- Course Hours (course_management_coursehours) ----
    ch.leture_npwk = _val("lecture_npwk")
    ch.tutorial_npwk = _val("tutorial_npwk")
    ch.laboratory_npwk = _val("laboratory_npwk")
    ch.total_hours = _val("total_hours")
    ch.credits = _val("credits")
    ch.save()

    # ---- Hour Config (examination_management_coursehourconfig) ----
    cfg = ch.hour_config
    cfg_payload = None
    if cfg is not None:
        def _int(name):
            raw = (request.POST.get(name) or "").strip()
            try:
                return int(float(raw)) if raw != "" else 0
            except (TypeError, ValueError):
                return 0

        cfg.lecture_hours = _int("cfg_lecture")
        cfg.tutorial_hours = _int("cfg_tutorial")
        cfg.laboratory_hours = _int("cfg_laboratory")
        cfg.theory_percentage = _int("cfg_theory_pct")
        cfg.practical_percentage = _int("cfg_practical_pct")
        cfg.activity_percentage = _int("cfg_activity_pct")
        cfg.save()
        cfg_payload = {
            "cfg_id": cfg.id,
            "cfg_lecture": cfg.lecture_hours,
            "cfg_tutorial": cfg.tutorial_hours,
            "cfg_laboratory": cfg.laboratory_hours,
            "cfg_theory_pct": cfg.theory_percentage,
            "cfg_practical_pct": cfg.practical_percentage,
            "cfg_activity_pct": cfg.activity_percentage,
        }

    return JsonResponse({
        "status": "success",
        "message": "Saved successfully.",
        "ch_id": ch.id,
        "config": cfg_payload,
    })


# ============================================================
# WORKLOAD DASHBOARD  (Course & Examination → sidebar)
# ------------------------------------------------------------
# Permission/function registered for now (so it can be granted from
# the Course & Examination permission matrix and appear in the sidebar).
# The dashboard content will be built next — this is a placeholder.
# ============================================================
@check_permission("workload_dashboard")
@course_management
def workload_dashboard(request):
    """Course Management → Workload Dashboard (placeholder)."""
    return render(request, "course_management/workload_dashboard.html", {})



from django.http import JsonResponse
from examination_management.models import CourseHourConfig



def get_ltp_config(request):
    regulation_year = request.GET.get("regulation_year")
    if not regulation_year:
        return JsonResponse({"ltp": []})

    regulation = Regulations.objects.filter(year=regulation_year).first()
    if not regulation:
        return JsonResponse({"ltp": []})

    configs = CourseHourConfig.objects.filter(regulation=regulation)

    ltp_list = []
    for c in configs:
        ltp_list.append({
            "id": c.id,  # ✅ send ID
            "lecture": c.lecture_hours,
            "tutorial": c.tutorial_hours,
            "laboratory": c.laboratory_hours,
        })
    return JsonResponse({"ltp": ltp_list})



from django.http import JsonResponse
from django.shortcuts import render
from django.contrib import messages
from faculty_management.models import general_information


@check_permission("add_period_allocation")
def add_period_allocation(request):
    # ------------------ LOGGED IN USER DEPARTMENT ONLY ------------------
    emp_id = getattr(request.user, "Employee_id", None)

    # get faculty general info (adjust model field names if needed)
    faculty_info = get_object_or_404(general_information, faculty_id=emp_id)

    # If faculty_info.department is FK to Add_Department, this works directly.
    # If it stores department_id as integer, handle both safely:
    user_dept_id = getattr(faculty_info.department, "id", None) or getattr(faculty_info, "department_id", None)

    # ✅ show only that department in dropdown
    departments = Add_Department.objects.filter(is_active=True, id=user_dept_id)

    regulations = Regulations.objects.all()
    degrees = Degree.objects.filter(is_active=True)

    if request.method == "POST":
        # ✅ enforce department from logged-in user (ignore POST tampering)
        department_id = str(user_dept_id)  # keep as string like POST

        section = request.POST.get("section")
        year = request.POST.get("year")
        semester = request.POST.get("semester")
        regulation = request.POST.get("regulation")

        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
        periods = [
            "first_period","second_period","third_period","fourth_period",
            "fifth_period","sixth_period","seventh_period","eighth_period",
            "nineth_period","tenth_period"
        ]

        errors, saved_days = [], []

        for day_index, day in enumerate(days):
            period_data, period_filled = {}, False

            for period in periods:
                selected_values = request.POST.getlist(f"{period}[]")
                if len(selected_values) > day_index:
                    selected_period = selected_values[day_index]
                    if selected_period:
                        period_data[period] = selected_period
                        period_filled = True

            if period_filled:
                try:
                    existing = PeriodAllocation.objects.filter(
                        department_id=department_id,
                        section=section,
                        year=year,
                        semester=semester,
                        day=day
                    ).first()

                    if existing:
                        for period, value in period_data.items():
                            setattr(existing, period, value)
                        existing.save()
                        saved_days.append(f"{day} (Updated)")
                    else:
                        PeriodAllocation.objects.create(
                            department_id=department_id,
                            section=section,
                            year=year,
                            semester=semester,
                            day=day,
                            **period_data
                        )
                        saved_days.append(f"{day} (New)")

                except Exception as e:
                    errors.append(f"Error saving {day}: {str(e)}")

        if errors:
            messages.error(request, "Some data could not be saved: " + ", ".join(errors))
        if saved_days:
            messages.success(request, "Data saved/updated for: " + ", ".join(saved_days))

    return render(request, "course_management/period_allocation.html", {
        "departments": departments,          # ✅ now only one department
        "regulations": regulations,
        "degrees": degrees,
        "user_department": faculty_info.department,  # optional (if you want to show name somewhere)
    })



# -------------------- AJAX ENDPOINTS --------------------

def get_sections(request):
    sections = SectionMaster.objects.all()
    data = [{"id": sec.id, "name": sec.section} for sec in sections]
    return JsonResponse(data, safe=False)


def get_courses_only(request):
    department_id = request.GET.get("department_id")
    section = request.GET.get("section")
    year = request.GET.get("year")
    semester = request.GET.get("semester")
    regulation_year = request.GET.get("regulation_year")

    if not all([department_id, section, year, semester, regulation_year]):
        return JsonResponse({"courses": [], "existing_allocations": {}}, safe=False)

    try:
        regulation_obj = Regulations.objects.get(year=regulation_year)
    except Regulations.DoesNotExist:
        return JsonResponse({"courses": [], "existing_allocations": {}}, safe=False)

    courses_query = Course.objects.filter(
        department_id=int(department_id),
        year=str(year),
        semester=str(semester),
        regulation=regulation_obj,
        is_active=True
    ).order_by("course_code")

    courses_list = list(courses_query.values("id","course_code","title"))

    # Existing allocations
    allocation_query = PeriodAllocation.objects.filter(
        department_id=int(department_id),
        section=section,
        year=str(year),
        semester=str(semester)
    )
    period_allocations = allocation_query.values(
        "day","first_period","second_period","third_period","fourth_period",
        "fifth_period","sixth_period","seventh_period","eighth_period",
        "nineth_period","tenth_period"
    )
    existing_allocations = {
        alloc["day"]: {k: alloc[k] for k in alloc if k != "day"}
        for alloc in period_allocations
    }

    return JsonResponse({
        "courses": courses_list,
        "existing_allocations": existing_allocations
    }, safe=False)


from course_management.models import Course_category
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def course_category(request):

    if request.method == "POST":
        action = request.POST.get("action")
        category_id = request.POST.get("id")
        name = request.POST.get("Course_category_name")
        code = request.POST.get("category_code")
        description = request.POST.get("category_description")
        regulation_id = request.POST.get("regulation")

        if action in ["add", "edit"] and (not name or not code or not regulation_id):
            return JsonResponse({
                "status": "error",
                "message": "All required fields must be filled!"
            })

        # ADD
        if action == "add":
            reg = get_object_or_404(Regulations, id=regulation_id)

            if Course_category.objects.filter(category_code=code).exists():
                return JsonResponse({
                    "status": "error",
                    "message": "Category code already exists!"
                })

            Course_category.objects.create(
                Course_category_name=name,
                category_code=code,
                category_description=description,
                regulation=reg
            )

            return JsonResponse({
                "status": "success",
                "message": "Category added successfully!"
            })

        # EDIT
        elif action == "edit":
            obj = get_object_or_404(Course_category, id=category_id)
            reg = get_object_or_404(Regulations, id=regulation_id)

            if Course_category.objects.filter(category_code=code).exclude(id=obj.id).exists():
                return JsonResponse({
                    "status": "error",
                    "message": "Category code already exists!"
                })

            obj.Course_category_name = name
            obj.category_code = code
            obj.category_description = description
            obj.regulation = reg
            obj.save()

            return JsonResponse({
                "status": "success",
                "message": "Category updated successfully!"
            })

        # DELETE
        elif action == "delete":
            obj = get_object_or_404(Course_category, id=category_id)
            obj.delete()
            return JsonResponse({
                "status": "success",
                "message": "Category deleted!"
            })

    # GET
    categories = [
        {
            "id": cat.id,
            "Course_category_name": cat.Course_category_name,
            "category_code": cat.category_code,
            "category_description": cat.category_description,
            "regulation": cat.regulation.year if cat.regulation else "",
            "regulation_id": cat.regulation.id if cat.regulation else ""
        }
        for cat in Course_category.objects.select_related("regulation")
    ]

    return JsonResponse({"status": "success", "data": categories})


def export_course_category_excel(request):
    categories = Course_category.objects.select_related("regulation").all()

    regulation_id = request.GET.get("regulation")

    if regulation_id:
        categories = categories.filter(regulation_id=regulation_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Course Categories"

    # Header
    ws.append([
        "course_category_name",
        "category_code",
        "category_description",
        "regulation",
    ])

    # Data
    for cat in categories:
        ws.append([
            cat.Course_category_name,
            cat.category_code,
            cat.category_description or "",
            cat.regulation.year if cat.regulation else "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="course_categories.xlsx"'

    wb.save(response)
    return response


# =============================================================================
# COURSE ENROLLMENT CRUD VIEWS
# =============================================================================
from django.shortcuts import render
from django.db.models import Q
from examination_management.models import ConsolidatedAssessmentResult
from django.db.models import Q
from examination_management.models import ConsolidatedAssessmentResult
      # adjust app names if different

   # adjust import paths to your project

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import render

# views.py  (corresponding to the HTML with Degree -> Dept -> Batch -> Section(optional) -> Semester -> Courses)

from django.db.models import Q, Min
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.http import require_GET

 # change import path if needed


# ---------------------------
# PAGE: load only degrees
# ---------------------------
def admin_consolidate(request):
    """
    Filter page only.
    Load ONLY degrees initially. Everything else loads via AJAX.
    """
    qs = ConsolidatedAssessmentResult.objects.select_related(
        "degree", "department", "course", "student", "student__department__degree"
    )

    deg_ids_1 = list(qs.exclude(degree_id__isnull=True).values_list("degree_id", flat=True))
    deg_ids_2 = list(
        qs.exclude(student__department__degree_id__isnull=True)
          .values_list("student__department__degree_id", flat=True)
    )
    degree_ids = sorted(set(deg_ids_1 + deg_ids_2))

    degrees = list(
        Degree.objects.filter(id__in=degree_ids)
        .values("id", "degree_code", "degree")
        .order_by("degree_code", "degree")
    )

    return render(request, "examination_management/admin/consolidate_admin.html", {
        "degrees": degrees,
        "departments": [],
        "batches": [],
        "sections": [],
        "courses": [],
    })


# ---------------------------
# AJAX: Degree -> Departments
# ---------------------------
@require_GET
def api_consolidate_departments(request):
    degree_id = (request.GET.get("degree") or "").strip()
    qs = ConsolidatedAssessmentResult.objects.select_related("department", "student__department")

    if degree_id:
        qs = qs.filter(Q(degree_id=degree_id) | Q(student__department__degree_id=degree_id))

    dep_ids_1 = list(qs.exclude(department_id__isnull=True).values_list("department_id", flat=True))
    dep_ids_2 = list(qs.exclude(student__department_id__isnull=True).values_list("student__department_id", flat=True))
    department_ids = sorted(set(dep_ids_1 + dep_ids_2))

    departments = list(
        Add_Department.objects.filter(id__in=department_ids)
        .values("id", "Department_code", "Department")
        .order_by("Department_code", "Department")
    )
    return JsonResponse({"departments": departments})


# ---------------------------
# AJAX: Degree+Department -> Batches
# ---------------------------
@require_GET
def api_consolidate_batches(request):
    degree_id = (request.GET.get("degree") or "").strip()
    department_id = (request.GET.get("department") or "").strip()

    qs = ConsolidatedAssessmentResult.objects.all()

    if degree_id:
        qs = qs.filter(Q(degree_id=degree_id) | Q(student__department__degree_id=degree_id))
    if department_id:
        qs = qs.filter(Q(department_id=department_id) | Q(student__department_id=department_id))

    batches = list(
        qs.exclude(batch__isnull=True).exclude(batch="")
          .values_list("batch", flat=True)
          .distinct()
          .order_by("batch")
    )
    return JsonResponse({"batches": batches})


# ---------------------------
# AJAX: Degree+Department+Batch -> Sections (OPTIONAL)
# ---------------------------
@require_GET
def api_consolidate_sections(request):
    degree_id = (request.GET.get("degree") or "").strip()
    department_id = (request.GET.get("department") or "").strip()
    batch = (request.GET.get("batch") or "").strip()

    qs = ConsolidatedAssessmentResult.objects.all()

    if degree_id:
        qs = qs.filter(Q(degree_id=degree_id) | Q(student__department__degree_id=degree_id))
    if department_id:
        qs = qs.filter(Q(department_id=department_id) | Q(student__department_id=department_id))
    if batch:
        qs = qs.filter(batch=batch)

    sections = list(
        qs.exclude(section__isnull=True).exclude(section="")
          .values_list("section", flat=True)
          .distinct()
          .order_by("section")
    )

    return JsonResponse({"sections": sections})


# ---------------------------
# AJAX: Degree+Dept+Batch+(Section optional) -> Semesters (Dynamic)
# uses ConsolidatedAssessmentResult.current_semester
# ---------------------------
@require_GET
def api_consolidate_semesters(request):
    degree_id = (request.GET.get("degree") or "").strip()
    department_id = (request.GET.get("department") or "").strip()
    batch = (request.GET.get("batch") or "").strip()
    section = (request.GET.get("section") or "").strip()  # OPTIONAL

    qs = ConsolidatedAssessmentResult.objects.exclude(current_semester__isnull=True)

    if degree_id:
        qs = qs.filter(Q(degree_id=degree_id) | Q(student__department__degree_id=degree_id))
    if department_id:
        qs = qs.filter(Q(department_id=department_id) | Q(student__department_id=department_id))
    if batch:
        qs = qs.filter(batch=batch)

    # ✅ section optional: filter only if section provided
    if section:
        qs = qs.filter(section=section)

    semesters = list(
        qs.values_list("current_semester", flat=True)
          .distinct()
          .order_by("current_semester")
    )

    # return as plain ints
    return JsonResponse({"semesters": semesters})



def api_consolidate_courses(request):
    degree_id = (request.GET.get("degree") or "").strip()
    department_id = (request.GET.get("department") or "").strip()
    batch = (request.GET.get("batch") or "").strip()
    section = (request.GET.get("section") or "").strip()       # OPTIONAL
    semester = (request.GET.get("semester") or "").strip()     # REQUIRED in UI

    qs = ConsolidatedAssessmentResult.objects.select_related("course", "course__course")

    if degree_id:
        qs = qs.filter(Q(degree_id=degree_id) | Q(student__department__degree_id=degree_id))
    if department_id:
        qs = qs.filter(Q(department_id=department_id) | Q(student__department_id=department_id))
    if batch:
        qs = qs.filter(batch=batch)
    if section:
        qs = qs.filter(section=section)
    if semester:
        qs = qs.filter(current_semester=semester)

    # ✅ IMPORTANT:
    # Return MASTER Course id (course__course_id) so PDF can fetch ALL students
    courses = list(
        qs.exclude(course__isnull=True)
          .exclude(course__course__isnull=True)
          .values(
              "course__course_id",                 # ✅ master course id
              "course__course__course_code",
              "course__course__title"
          )
          .distinct()
          .order_by("course__course__course_code")
    )

    return JsonResponse({"courses": courses})








import io
from decimal import Decimal, ROUND_HALF_UP

from django.http import HttpResponse

from django.db.models import Q
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

import io
import os
from decimal import Decimal, ROUND_HALF_UP

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.staticfiles import finders

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from course_management.models import CourseHours
from examination_management.models import CourseHourConfig


def download_consolidated_pdf(request):
    # ---------------- GET PARAMS ----------------
    degree_id = request.GET.get("degree")
    department_id = request.GET.get("department")
    batch = request.GET.get("batch")
    section = request.GET.get("section", "")  # optional
    semester = request.GET.get("semester")
    course_ids = request.GET.getlist("courses")  # from UI (now single select but getlist works)

    if not (degree_id and department_id and batch and semester and course_ids):
        return HttpResponse("Missing required filters.", status=400)

   

    # ---------------- FILTER DB ----------------
    qs = (
        ConsolidatedAssessmentResult.objects
        .select_related("student", "department", "degree", "course", "course__course", "hour_config")
        .filter(
            degree_id=degree_id,
            department_id=department_id,
            batch=batch,
            current_semester=semester,

            # ✅ FIX: filter by CourseEnrollment -> Course master id
            course__course_id__in=course_ids,
        )
        .order_by("course__course_id", "student__reg_no")
    )

    if section:
        qs = qs.filter(section=section)

    

    results = list(qs)
    if not results:
        return HttpResponse("No consolidated records found for selected filter.", status=404)

    # ---------------- HELPERS ----------------
    styles = getSampleStyleSheet()
    style_small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    def _split_csv(s):
        if not s:
            return []
        return [x.strip() for x in str(s).split(",") if str(x).strip() != ""]

    def _fmt_int_or_blank(x):
        if x is None or x == "":
            return ""
        try:
            return str(int(round(float(x))))
        except Exception:
            return str(x)

    def _fmt_num2(x):
        if x is None or x == "":
            return ""
        try:
            return str(Decimal(str(x)).quantize(Decimal("0.00"), rounding=ROUND_HALF_UP))
        except Exception:
            return str(x)

    # ✅ LTP PRINT + CONV TOTAL (ONLY ADDED; REST UNCHANGED)
    def _print_course_ltp_and_get_conv_total(course_master_obj):
        ch = CourseHours.objects.select_related("hour_config").filter(course=course_master_obj).first()

        if not ch:
           
            return 100  # safe fallback

        cfg = ch.hour_config
        if not cfg:
            
            return 100  # safe fallback

        lec = int(cfg.lecture_hours or 0)
        lab = int(cfg.laboratory_hours or 0)

        if lec > 0 and lab == 0:
            tot_to_convert = 40
        elif lec == 0 and lab > 0:
            tot_to_convert = 60
        elif lec > 0 and lab > 0:
            tot_to_convert = 50
        else:
            tot_to_convert = 100  # safe fallback

       

        return tot_to_convert

    # ---------------- GROUP BY COURSE ----------------
    # ✅ FIX: group by master course id so both students fall under same course table
    by_course = {}
    for r in results:
        master_course_id = r.course.course_id if (r.course and getattr(r.course, "course_id", None)) else None
        by_course.setdefault(master_course_id, []).append(r)

    # ---------------- PDF SETUP (EXACT LIKE OLD) ----------------
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=42 * mm,
        bottomMargin=15 * mm
    )

    # ---------------- LOGO / HEADER FOOTER (EXACT LIKE OLD) ----------------
    def _find_logo():
        p = finders.find("images/ritlogo.png")
        return p if p and os.path.exists(p) else None

    logo_path = _find_logo()

    first = results[0]
    degree_code = getattr(first.degree, "degree_code", "") if first.degree else ""
    degree_name = getattr(first.degree, "degree", "") if first.degree else ""

    dept_name = ""
    if first.department:
        dept_name = getattr(first.department, "Department", "") or str(first.department)

    def _on_page(canvas, doc_obj):
        canvas.saveState()
        w, h = landscape(A4)
        left = 15 * mm

        if logo_path:
            try:
                img = ImageReader(logo_path)
                canvas.drawImage(img, left, h - 26 * mm, height=18 * mm, preserveAspectRatio=True)
            except:
                pass

        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(w / 2, h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")
        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w / 2, h - 15 * mm, "Rajapalayam - 626117")
        canvas.drawCentredString(w / 2, h - 19 * mm, "Affiliated to Anna University, Chennai")
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(w / 2, h - 26 * mm, "Overall Consolidated Statement")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w / 2, h - 31 * mm, f"{degree_code} — {degree_name}")
        canvas.drawCentredString(w / 2, h - 35 * mm, dept_name)

        course_line = getattr(doc_obj, "_course_line", "")
        if course_line:
            canvas.drawCentredString(w / 2, h - 39 * mm, course_line)

        canvas.line(left, h - 42 * mm, w - 15 * mm, h - 42 * mm)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(w - 15 * mm, 12 * mm, f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    # ---------------- BUILD STORY ----------------
    story = []

    meta_lines = [
        f"<b>Batch-Section:</b> {batch} - {(section or 'ALL')}",
        f"<b>Semester:</b> {semester}",
    ]
    story.append(Paragraph("<br/>".join(meta_lines), style_small))
    story.append(Spacer(1, 10))

    # ---------------- TABLE(S) PER COURSE ----------------
    for idx, (master_course_id, rows) in enumerate(by_course.items()):
        course_obj = rows[0].course.course if (rows[0].course and hasattr(rows[0].course, "course")) else None
        course_code = getattr(course_obj, "course_code", "") if course_obj else ""
        course_title = getattr(course_obj, "title", "") if course_obj else ""
        course_line = f"{course_code} — {course_title}".strip(" —")
        doc._course_line = course_line

        tot_to_convert = 100
        if course_obj:
            tot_to_convert = _print_course_ltp_and_get_conv_total(course_obj)

        r0 = rows[0]
        theory_names = _split_csv(r0.theory_assessment_name)
        activity_names = _split_csv(r0.activity_assessment_name)
        practical_names = _split_csv(r0.practical_assessment_name)

        theory_max_list = _split_csv(r0.theory_display_max_mark)
        activity_max_list = _split_csv(r0.activity_display_max_mark)
        practical_max_list = _split_csv(r0.practical_display_max_mark)

        header = ["Reg No", "Name"]

        for i, nm in enumerate(theory_names):
            mx = theory_max_list[i] if i < len(theory_max_list) else ""
            header.append(f"{nm} ({mx})" if mx else nm)

        for i, nm in enumerate(activity_names):
            mx = activity_max_list[i] if i < len(activity_max_list) else ""
            header.append(f"{nm} ({mx})" if mx else nm)

        for i, nm in enumerate(practical_names):
            mx = practical_max_list[i] if i < len(practical_max_list) else ""
            header.append(f"{nm} ({mx})" if mx else nm)

        grand_total = (r0.theory_max_mark or 0) + (r0.activity_max_mark or 0) + (r0.practical_max_mark or 0)
        header.append(f"Total ({_fmt_int_or_blank(grand_total)})")
        header.append(f"Conv ({tot_to_convert})")

        data = [header]

        for r in rows:
            row = [
                r.student.reg_no if r.student else "",
                r.student.name if r.student else "",
            ]

            t_act = _split_csv(r.theory_display_actual_mark)
            a_act = _split_csv(r.activity_display_actual_mark)
            p_act = _split_csv(r.practical_display_actual_mark)

            for i in range(len(theory_names)):
                row.append(t_act[i] if i < len(t_act) else "")

            for i in range(len(activity_names)):
                row.append(a_act[i] if i < len(a_act) else "")

            for i in range(len(practical_names)):
                row.append(p_act[i] if i < len(p_act) else "")

            th = r.theory_actual_mark or 0
            ac = r.activity_actual_mark or 0
            pr = r.practical_actual_mark or 0
            total_val = th + ac + pr

            

            row.append(_fmt_int_or_blank(total_val))

            conv_val = ""
            try:
                gt = float(grand_total) if grand_total else 0.0
                tv = float(total_val) if total_val is not None else 0.0
                if gt > 0:
                    conv_val = str(int(round((tv / gt) * float(tot_to_convert))))
            except Exception:
                conv_val = ""

            row.append(conv_val)
            data.append(row)

        total_width = 297 * mm - 30 * mm
        w_reg = 30 * mm
        w_name = 55 * mm
        w_total = 22 * mm
        w_conv = 22 * mm

        col_widths = [w_reg, w_name]
        remaining = total_width - (w_reg + w_name + w_total + w_conv)

        assessment_cols_count = len(header) - 4
        if assessment_cols_count > 0:
            each = remaining / assessment_cols_count
            col_widths += [each] * assessment_cols_count
            col_widths += [w_total, w_conv]
        else:
            col_widths += [w_total, w_conv]

        table = Table(data, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
            ("ALIGN", (2, 1), (-1, -1), "CENTER"),
            ("ALIGN", (0, 1), (1, -1), "LEFT"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))

        story.append(table)
        if idx < len(by_course) - 1:
            story.append(Spacer(1, 14))

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buffer.getvalue()
    buffer.close()

    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = 'inline; filename="Consolidated_Assessment_Result.pdf"'
    return resp



def internal_admin(request):
    qs = ConsolidatedAssessmentResult.objects.select_related(
        "degree", "department", "course", "student", "student__department__degree"
    )

    deg_ids_1 = list(qs.exclude(degree_id__isnull=True).values_list("degree_id", flat=True))
    deg_ids_2 = list(
        qs.exclude(student__department__degree_id__isnull=True)
          .values_list("student__department__degree_id", flat=True)
    )
    degree_ids = sorted(set(deg_ids_1 + deg_ids_2))

    degrees = list(
        Degree.objects.filter(id__in=degree_ids)
        .values("id", "degree_code", "degree")
        .order_by("degree_code", "degree")
    )

    return render(request, "examination_management/admin/internal_admin.html", {
        "degrees": degrees,
    })



@require_GET
def api_internal_departments(request):
    degree_id = (request.GET.get("degree_id") or "").strip()

    qs = ConsolidatedAssessmentResult.objects.select_related("department", "student__department")

    if degree_id:
        qs = qs.filter(Q(degree_id=degree_id) | Q(student__department__degree_id=degree_id))

    dep_ids_1 = list(qs.exclude(department_id__isnull=True).values_list("department_id", flat=True))
    dep_ids_2 = list(qs.exclude(student__department_id__isnull=True).values_list("student__department_id", flat=True))
    department_ids = sorted(set(dep_ids_1 + dep_ids_2))

    departments = list(
        Add_Department.objects.filter(id__in=department_ids)
        .values("id", "Department_code", "Department")
        .order_by("Department_code", "Department")
    )

    # return in SAME "items" format expected by your HTML JS
    items = [{"id": d["id"], "text": f'{d["Department_code"]} - {d["Department"]}'} for d in departments]
    return JsonResponse({"items": items})




@require_GET
def api_internal_batches(request):
    degree_id = (request.GET.get("degree_id") or "").strip()
    department_id = (request.GET.get("department_id") or "").strip()

    qs = ConsolidatedAssessmentResult.objects.all()

    if degree_id:
        qs = qs.filter(Q(degree_id=degree_id) | Q(student__department__degree_id=degree_id))
    if department_id:
        qs = qs.filter(Q(department_id=department_id) | Q(student__department_id=department_id))

    batches = list(
        qs.exclude(batch__isnull=True).exclude(batch="")
          .values_list("batch", flat=True)
          .distinct()
          .order_by("batch")
    )

    items = [{"id": b, "text": b} for b in batches]
    return JsonResponse({"items": items})



@require_GET
def api_internal_sections(request):
    degree_id = (request.GET.get("degree_id") or "").strip()
    department_id = (request.GET.get("department_id") or "").strip()
    batch = (request.GET.get("batch") or "").strip()

    qs = ConsolidatedAssessmentResult.objects.all()

    if degree_id:
        qs = qs.filter(Q(degree_id=degree_id) | Q(student__department__degree_id=degree_id))
    if department_id:
        qs = qs.filter(Q(department_id=department_id) | Q(student__department_id=department_id))
    if batch:
        qs = qs.filter(batch=batch)

    sections = list(
        qs.exclude(section__isnull=True).exclude(section="")
          .values_list("section", flat=True)
          .distinct()
          .order_by("section")
    )

    items = [{"id": s, "text": s} for s in sections]
    return JsonResponse({"items": items})



@require_GET
def api_internal_semesters(request):
    degree_id = (request.GET.get("degree_id") or "").strip()
    department_id = (request.GET.get("department_id") or "").strip()
    batch = (request.GET.get("batch") or "").strip()
    section = (request.GET.get("section") or "").strip()

    qs = ConsolidatedAssessmentResult.objects.exclude(current_semester__isnull=True)

    if degree_id:
        qs = qs.filter(Q(degree_id=degree_id) | Q(student__department__degree_id=degree_id))
    if department_id:
        qs = qs.filter(Q(department_id=department_id) | Q(student__department_id=department_id))
    if batch:
        qs = qs.filter(batch=batch)
    if section:
        qs = qs.filter(section=section)

    semesters = list(
        qs.values_list("current_semester", flat=True)
          .distinct()
          .order_by("current_semester")
    )

    items = [{"id": sem, "text": f"Semester {sem}"} for sem in semesters]
    return JsonResponse({"items": items})


@require_GET
def api_internal_courses(request):
    degree_id = (request.GET.get("degree_id") or "").strip()
    department_id = (request.GET.get("department_id") or "").strip()
    batch = (request.GET.get("batch") or "").strip()
    section = (request.GET.get("section") or "").strip()
    semester = (request.GET.get("semester") or "").strip()

    qs = ConsolidatedAssessmentResult.objects.select_related("course", "course__course")

    if degree_id:
        qs = qs.filter(Q(degree_id=degree_id) | Q(student__department__degree_id=degree_id))
    if department_id:
        qs = qs.filter(Q(department_id=department_id) | Q(student__department_id=department_id))
    if batch:
        qs = qs.filter(batch=batch)
    if section:
        qs = qs.filter(section=section)
    if semester:
        qs = qs.filter(current_semester=semester)

    courses = list(
        qs.exclude(course__isnull=True)
          .exclude(course__course__isnull=True)
          .values(
              "course__course_id",            # ✅ master course id
              "course__course__course_code",
              "course__course__title"
          )
          .distinct()
          .order_by("course__course__course_code")
    )

    items = [
        {
            "id": c["course__course_id"],  # ✅ master id returned
            "text": f'{c["course__course__course_code"]} - {c["course__course__title"]}'
        }
        for c in courses
    ]
    return JsonResponse({"items": items})




# views.py  (Internal Admin PDF - EXACT OLD DESIGN)
import io, os
from decimal import Decimal, ROUND_HALF_UP

from django.http import HttpResponse
from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.contrib.staticfiles import finders

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader



def internal_admin_download_pdf(request):
 

    # ---------------- GET PARAMS (from your HTML) ----------------
    degree_id = request.GET.get("degree") or request.GET.get("degree_id")
    department_id = request.GET.get("department") or request.GET.get("department_id")
    batch = request.GET.get("batch")
    section = request.GET.get("section", "")  # optional
    semester = request.GET.get("semester")

    # your HTML uses courses[]
    course_ids = request.GET.getlist("courses[]") or request.GET.getlist("courses")
    

    if not (degree_id and department_id and batch and semester and course_ids):
        return HttpResponse("Missing required filters.", status=400)
    
    selected_assessments = request.GET.getlist("assessment")

    if not selected_assessments:
        return HttpResponse("Please select at least one assessment.", status=400)
    
    def get_header_label():
        mapping = {
            "theory": "T",
            "practical": "P",
            "activity": "A",
        }
        return "".join(mapping[a] for a in selected_assessments if a in mapping)


    def sum_marks(obj, mark_type="max"):
        total = Decimal("0.00")

        if "theory" in selected_assessments:
            val = getattr(obj, f"theory_{mark_type}_mark", None)
            if val is not None:
                total += val

        if "practical" in selected_assessments:
            val = getattr(obj, f"practical_{mark_type}_mark", None)
            if val is not None:
                total += val

        if "activity" in selected_assessments:
            val = getattr(obj, f"activity_{mark_type}_mark", None)
            if val is not None:
                total += val

        return total.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


    # ---------------- FILTER DB (match your old logic) ----------------
    qs = (
        ConsolidatedAssessmentResult.objects
        .select_related("student", "department", "degree", "course", "course__course", "hour_config")
        .filter(
            degree_id=degree_id,
            department_id=department_id,
            batch=batch,
            current_semester=semester,
            course__course_id__in=course_ids,
        )
        .order_by("course__course_id", "student__reg_no")
    )

    if section:
        qs = qs.filter(section=section)

    results = list(qs)
    if not results:
        return HttpResponse("No consolidated records found for selected filter.", status=404)
    from collections import OrderedDict

    students_map = OrderedDict()
    course_codes = []

    for r in results:
        reg = getattr(r.student, "reg_no", "")
        name = getattr(r.student, "name", "")
        course_code = r.course.course.course_code

        if course_code not in course_codes:
            course_codes.append(course_code)

        if reg not in students_map:
            students_map[reg] = {
                "name": name,
                "courses": {}
            }

        # value can be empty for now
        students_map[reg]["courses"][course_code] = {
    "max": sum_marks(r, "max"),
    "actual": sum_marks(r, "actual"),
}



    # ---------------- HELPERS (same as old) ----------------
    styles = getSampleStyleSheet()
    style_small = ParagraphStyle("small", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

   

    # Grouping by course
    
    # ---------------- PDF SETUP ----------------
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=42 * mm,
        bottomMargin=15 * mm
    )

    # Logo and header/footer setup
    def _find_logo():
        p = finders.find("images/ritlogo.png")
        return p if p and os.path.exists(p) else None

    logo_path = _find_logo()

    first = results[0]
    degree_code = getattr(first.degree, "degree_code", "") if first.degree else ""
    degree_name = getattr(first.degree, "degree", "") if first.degree else ""

    dept_name = ""
    if first.department:
        dept_name = getattr(first.department, "Department", "") or str(first.department)

    def _on_page(canvas, doc_obj):
        canvas.saveState()
        w, h = landscape(A4)
        left = 15 * mm

        # logo
        if logo_path:
            try:
                img = ImageReader(logo_path)
                canvas.drawImage(img, left, h - 26 * mm, height=18 * mm, preserveAspectRatio=True)
            except Exception:
                pass

        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(w / 2, h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w / 2, h - 15 * mm, "Rajapalayam - 626117")
        canvas.drawCentredString(w / 2, h - 19 * mm, "Affiliated to Anna University, Chennai")

        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(w / 2, h - 26 * mm, "Internalmark Statement")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w / 2, h - 31 * mm, f"{degree_code} — {degree_name}")
        canvas.drawCentredString(w / 2, h - 35 * mm, dept_name)

        # per-course header line
        course_line = getattr(doc_obj, "_course_line", "")
        if course_line:
            canvas.drawCentredString(w / 2, h - 39 * mm, course_line)

        # separator
        canvas.line(left, h - 42 * mm, w - 15 * mm, h - 42 * mm)

        # page no
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(w - 15 * mm, 12 * mm, f"Page {canvas.getPageNumber()}")

        canvas.restoreState()
    

    # ---------------- BUILD STORY ----------------
    story = []

    meta_lines = [
        f"<b>Batch-Section:</b> {batch} - {(section or 'ALL')}",
        f"<b>Semester:</b> {semester}",
    ]
    story.append(Paragraph("<br/>".join(meta_lines), style_small))
    story.append(Spacer(1, 10))
    
    # ---------------- TABLE(S) PER COURSE ----------------
    # ---------------- SINGLE TABLE ----------------

    header_label = get_header_label()

    data = [["Reg No", "Name"] + [
        f"{code} ({int(students_map[next(iter(students_map))]['courses'][code]['max'])})\n{header_label}"
        for code in course_codes
    ]]




    for reg, info in students_map.items():
        row = [reg, info["name"]]
        for code in course_codes:
            cell = info["courses"].get(code)

            if cell:
                row.append(str(int(cell["actual"])))

            else:
                row.append("")
            # empty cell
        data.append(row)

    # column widths
    total_width = 297 * mm - 30 * mm
    w_reg = 30 * mm
    w_name = 55 * mm
    remaining = total_width - (w_reg + w_name)
    w_course = remaining / len(course_codes)

    col_widths = [w_reg, w_name] + [w_course] * len(course_codes)

    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ALIGN", (2, 1), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    story.append(table)

        
    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buffer.getvalue()
    buffer.close()

    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = 'inline; filename="Consolidated_Assessment_Result.pdf"'
    return resp

from django.shortcuts import render, redirect
from django.contrib import messages
from course_management.models import Hall

# @check_permission("Hall_entry")
def hall_entry(request):
    halls = Hall.objects.all().order_by('hall_name')

    if request.method == "POST" and request.POST.get("action") == "add":
        hall_name = request.POST.get("hall_name")
        benches = request.POST.get("benches")

        if hall_name and benches:
            if Hall.objects.filter(hall_name=hall_name).exists():
                messages.error(request, "Hall already exists")
            else:
                Hall.objects.create(
                    hall_name=hall_name,
                    benches=benches
                )
                messages.success(request, "Hall added successfully")
                return redirect('Hall_entry_old')
        else:
            messages.error(request, "All fields are required")

    return render(request, 'course_management/admin/admin_hall.html', {
        'halls': halls
    })



def edit_hall(request, id):
    halls = Hall.objects.all()
    edit_hall_obj = get_object_or_404(Hall, id=id)

    if request.method == "POST":
        hall_name = request.POST.get("hall_name")
        benches = request.POST.get("benches")

        if hall_name and benches:
            edit_hall_obj.hall_name = hall_name
            edit_hall_obj.benches = benches
            edit_hall_obj.save()
            messages.success(request, "Hall updated successfully")
            return redirect("Hall_entry")
        else:
            messages.error(request, "All fields are required")

    return render(request, "course_management/admin/admin_hall.html", {
        "halls": halls,
        "edit_hall": edit_hall_obj
    })


def delete_hall(request, id):
    hall_obj = get_object_or_404(Hall, id=id)
    hall_obj.delete()
    messages.success(request, "Hall deleted successfully")
    return redirect('Hall_entry')



from django.http import JsonResponse

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib import messages

def admin_update_course(request, course_id=None, action=None):
    # ===================== INITIAL DATA =====================
    courses = Course.objects.select_related(
        'department', 'regulation', 'elective'
    ).all()

    departments = Add_Department.objects.filter(is_active=True)
    degree = Degree.objects.filter(is_active=True)
    regulations = Regulations.objects.all()

    course = None
    if course_id and action == 'edit':
        course = get_object_or_404(Course, id=course_id)

    # ===================== POST (UPDATE / DELETE) =====================
    if request.method == "POST":
        action = request.POST.get("action")
        course_id = request.POST.get("course_id") or course_id

        if action not in ["update", "delete"]:
            messages.error(request, "Invalid action specified.")
            return redirect('admin_update_course')

        if not course_id:
            messages.error(request, "No course ID provided.")
            return redirect('admin_update_course')

        course = get_object_or_404(Course, id=course_id)

        if action == "update":
            course_code = request.POST.get("course_code")
            title = request.POST.get("title")
            is_active = request.POST.get("is_active") == "on"

            if not all([course_code, title]):
                messages.error(request, "Please fill in all required fields.")
                return render(request, 'course_management/admin_update_course.html', {
                    'courses': courses,
                    'degrees': degree,
                    'departments': departments,
                    'course': course,
                    'regulations': regulations
                })

            course.course_code = course_code
            course.title = title
            course.is_active = is_active
            course.save()

            messages.success(request, f"Course '{course.title}' updated successfully.")
            return redirect('admin_update_course')

        elif action == "delete":
            course_title = course.title
            course.delete()
            messages.success(request, f"Course '{course_title}' deleted successfully.")
            return redirect('admin_update_course')

    # ===================== AJAX FILTER =====================
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        courses = Course.objects.select_related(
            'department', 'regulation', 'elective'
        ).all()

        degree_id = request.GET.get('degree')
        department_id = request.GET.get('department')
        regulation_id = request.GET.get('regulation')
        year = request.GET.get('year')
        semester = request.GET.get('semester')

        if degree_id:
            courses = courses.filter(department__degree_id=degree_id)

        if department_id:
            courses = courses.filter(department_id=department_id)

        if regulation_id:
            courses = courses.filter(regulation_id=regulation_id)

        if year:
            courses = courses.filter(year=year)

        if semester:
            courses = courses.filter(semester=semester)

        data = []
        for c in courses:
            data.append({
                'id': c.id,
                'course_code': c.course_code,
                'title': c.title,
                'department': c.department.Department,
                'regulation': c.regulation.year,
                'year': c.year,
                'semester': c.semester,
                'category': c.elective.Course_category_name,
                'is_active': c.is_active,
            })

        return JsonResponse(data, safe=False)

    # ===================== NORMAL PAGE LOAD =====================
    context = {
        'courses': courses,
        'degrees': degree,
        'departments': departments,
        'course': course,
        'regulations': regulations,
    }

    return render(request, 'course_management/admin_update_course.html', context)



def export_filtered_courses_excel(request):
    courses = Course.objects.select_related(
        'department', 'regulation', 'elective'
    ).all()

    degree_id = request.GET.get('degree')
    department_id = request.GET.get('department')
    regulation_id = request.GET.get('regulation')
    year = request.GET.get('year')
    semester = request.GET.get('semester')

    if degree_id:
        courses = courses.filter(department__degree_id=degree_id)

    if department_id:
        courses = courses.filter(department_id=department_id)

    if regulation_id:
        courses = courses.filter(regulation_id=regulation_id)

    if year:
        courses = courses.filter(year=year)

    if semester:
        courses = courses.filter(semester=semester)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Courses"

    # Header row
    worksheet.append([
        "department_code",
        "course_code",
        "title",
        "year",
        "semester",
        "regulation",
        "course_category",
    ])

    # Data rows
    for c in courses:
        worksheet.append([
            getattr(c.department, 'Department_code', ''),   # change field name if needed
            c.course_code,
            c.title,
            c.year,
            c.semester,
            c.regulation.year,
            c.elective.Course_category_name if c.elective else '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=filtered_courses.xlsx'
    workbook.save(response)

    return response



def departments_by_degree(request):
    degree_ids = request.GET.get("degree_id")

    if not degree_ids:
        return JsonResponse({"items": []})

    degree_list = [x.strip() for x in degree_ids.split(",")]

    departments = (
        Add_Department.objects.filter(is_active=True, degree_id__in=degree_list)
        .order_by("Department")
        .distinct()
    )

    items = []

    # ✅ Add "All Departments" option
    if departments.exists():
        items.append({
            "id": "ALL",
            "text": "All Departments"
        })

    # ✅ Add actual departments
    items += [{"id": d.id, "text": d.Department} for d in departments]

    return JsonResponse({"items": items})



from examination_management.models import InternalAssessment
from collections import defaultdict
from django.http import JsonResponse


def internal_assessments_by_degree(request):
    degree_ids = request.GET.get("degree_id")

    if not degree_ids:
        return JsonResponse({"items": []})

    degree_list = [x.strip() for x in degree_ids.split(",")]

    qs = (
        InternalAssessment.objects
        .filter(degree_id__in=degree_list)
        .select_related("degree")
        .exclude(iat__isnull=True)
        .exclude(iat__exact="")
        .order_by("iat")
    )

    # ----------------------------------------------------
    # Group IATs
    # ----------------------------------------------------
    iat_map = defaultdict(list)

    for obj in qs:
        key = obj.iat.strip().upper()
        iat_map[key].append(obj)

    items = []

    # ----------------------------------------------------
    # Build dropdown items
    # ----------------------------------------------------
    for iat_name, records in sorted(iat_map.items()):

        # degrees available for this IAT
        degree_names = list(set(r.degree.degree for r in records))

        # If IAT exists in ALL selected degrees → show once
        if len(degree_names) == len(degree_list):
            items.append({
                "id": records[0].id,  # any id is fine
                "text": iat_name
            })

        else:
            # IAT exists only for specific degree(s)
            for r in records:
                items.append({
                    "id": r.id,
                    "text": f"{iat_name} - {r.degree.degree} Only"

                })

    return JsonResponse({"items": items})


 # ✅ change import based on your project

from examination_management.models import InternalExamSchedule

from django.shortcuts import render
from django.utils.dateparse import parse_date
from datetime import timedelta
from django.shortcuts import render
from django.contrib import messages
from django.utils.dateparse import parse_date
from django.db.models import Min
from datetime import timedelta
from django.shortcuts import render
from django.contrib import messages
from django.utils.dateparse import parse_date
from django.db.models import Min

from datetime import timedelta
from django.shortcuts import render
from django.contrib import messages
from django.utils.dateparse import parse_date
from django.db.models import Min
from itertools import groupby
def next_working_day(d):
    d = d + timedelta(days=1)
    while d.weekday() == 6:  # Sunday
        d = d + timedelta(days=1)
    return d


def internal_exam_schedule(request):
    degrees_qs = Degree.objects.filter(is_active=True).order_by("duration", "degree")

    grouped_degrees = []
    final_degree_options = []

    for duration, group in groupby(degrees_qs, key=lambda x: x.duration):
        group_list = list(group)

        # ✅ Add individual degrees
        for d in group_list:
            final_degree_options.append({
                "ids": str(d.id),
                "name": d.degree,
            })

        # ✅ Add grouped option ONLY if more than one degree exists
        if len(group_list) > 1:
            final_degree_options.append({
                "ids": ",".join(str(d.id) for d in group_list),
                "name": " / ".join(d.degree for d in group_list),
            })

        regulations = Regulations.objects.all().order_by("year")
    internal_assessment_id = request.POST.get("internal_assessment")

    if request.method == "POST":
        degree_id = request.POST.get("degree")
        department_id = request.POST.get("department")
        regulation_id = request.POST.get("regulation")
        schedule_date = request.POST.get("schedule_date")

        exam_date = parse_date(schedule_date) if schedule_date else None

        degree_ids = degree_id.split(",")
        degree_ids = (degree_id or "").split(",")

        def get_degree_for_department(dept_obj):
            # If Add_Department has FK like dept_obj.degree -> use it
            if hasattr(dept_obj, "degree") and dept_obj.degree:
                return dept_obj.degree

            # If Add_Department has degree_id column
            if hasattr(dept_obj, "degree_id") and dept_obj.degree_id:
                return Degree.objects.filter(id=dept_obj.degree_id).first()

            return None

        # ✅ GLOBAL KEY NORMALIZER (ONLY USED FOR "ALL" SAME-COURSE MATCHING)
        def _norm(v):
            return " ".join((v or "").strip().upper().split())

        department_fk = None
        selected_departments = []

        if department_id == "ALL":
            selected_departments = Add_Department.objects.filter(
                degree_id__in=degree_ids,
                is_active=True
            )
        else:
            department_fk = Add_Department.objects.filter(id=department_id).first()
            if department_fk:
                selected_departments = [department_fk]

        regulation_fk = Regulations.objects.filter(id=regulation_id).first() if regulation_id else None
        internal_assessment_fk = (
            InternalAssessment.objects.filter(id=internal_assessment_id).first()
            if internal_assessment_id else None
        )

        if not exam_date:
            messages.error(request, "Please select a start date to generate schedule.")
            return render(
                request,
                "examination_management/admin/internal_exam_schedule.html",
                {
                    "degrees": final_degree_options,
                    "regulations": regulations
                }
            )

        # ✅ 1) Find semesters and batches from StudentDetails
        student_qs = StudentDetails.objects.all()
        if department_id == "ALL":
            student_qs = student_qs.filter(department_id__in=[d.id for d in selected_departments])
        elif department_id:
            student_qs = student_qs.filter(department_id=department_id)

        if regulation_id:
            try:
                student_qs = student_qs.filter(regulation_id=regulation_id)
            except Exception:
                pass

        semesters = (
            student_qs.exclude(semester__isnull=True)
            .values_list("semester", flat=True)
            .distinct()
        )

        semester_list = list(semesters)

        # 🔥 Collect batches per department + semester
        batch_map = {}  # (dept_id, semester) -> set(batch_names)

        student_batch_qs = StudentDetails.objects.all()

        if department_id == "ALL":
            student_batch_qs = student_batch_qs.filter(
                department_id__in=[d.id for d in selected_departments]
            )
        elif department_id:
            student_batch_qs = student_batch_qs.filter(
                department_id=department_id
            )

        if regulation_id:
            student_batch_qs = student_batch_qs.filter(
                regulation=regulation_fk.year  # adjust if regulation stored differently
            )

        student_batch_qs = student_batch_qs.exclude(batch__isnull=True).exclude(batch="")

        for row in student_batch_qs.values("department_id", "semester", "batch").distinct():
            key = (int(row["department_id"]), (row["semester"] or "").strip())
            batch_map.setdefault(key, set()).add(row["batch"])

        # ✅ 2) Fetch courses for THIS department
        semester_str_list = [str(s) for s in semester_list]

        course_qs = Course.objects.filter(is_active=True)

        if department_id == "ALL":
            course_qs = course_qs.filter(department_id__in=[d.id for d in selected_departments])
        elif department_id:
            course_qs = course_qs.filter(department_id=department_id)

        if regulation_id:
            course_qs = course_qs.filter(regulation_id=regulation_id)

        if semester_str_list:
            course_qs = course_qs.filter(semester__in=semester_str_list)
        else:
            course_qs = course_qs.none()

        # ✅ REMOVE duplicate courses per department
        def normalize_text(v):
            return " ".join((v or "").lower().split())

        unique_courses = {}
        filtered_courses = []

        for c in course_qs:
            key = (
                c.department_id,
                normalize_text(c.semester),
                normalize_text(c.course_code),
                normalize_text(c.title),
                c.regulation_id,
            )

            if key not in unique_courses:
                unique_courses[key] = c
                filtered_courses.append(c)

        course_qs = filtered_courses

        default_session = "FN"

        # ---------------------------------------------------------------------
        # ✅ SAME COURSE → SAME DATE (across departments) (DB lookup)
        # ---------------------------------------------------------------------
        db_course_date_map = {}                  # single dept usage
        db_course_date_map_by_dept = {}          # dept_id -> {course_key: date}

        if department_id == "ALL":
            for dept in selected_departments:
                existing_course_dates = (
                    InternalExamSchedule.objects
                    .filter(
                        regulation=regulation_fk,
                        internal_assessment=internal_assessment_fk,
                        session=default_session,
                        department=dept,
                    )
                    .exclude(exam_date__isnull=True)
                    .values("course__course_code", "course__title", "semester")
                    .annotate(first_date=Min("exam_date"))
                )

                db_course_date_map_by_dept[int(dept.id)] = {
                    f"{(r['semester'] or '').strip()}||{(r['course__course_code'] or '').strip()}||{(r['course__title'] or '').strip()}".lower(): r["first_date"]
                    for r in existing_course_dates
                }
        else:
            existing_course_dates = (
                InternalExamSchedule.objects
                .filter(
                    regulation=regulation_fk,
                    internal_assessment=internal_assessment_fk,
                    session=default_session,
                )
                .exclude(exam_date__isnull=True)
                .values("course__course_code", "course__title", "semester")
                .annotate(first_date=Min("exam_date"))
            )

            db_course_date_map = {
                f"{(r['semester'] or '').strip()}||{(r['course__course_code'] or '').strip()}||{(r['course__title'] or '').strip()}".lower(): r["first_date"]
                for r in existing_course_dates
            }

        # ---------------------------------------------------------------------
        # ✅ Reserve dates ONLY for reused courses that are present in THIS course_qs.
        # ---------------------------------------------------------------------
        reserved_dates_by_sem = {}                   # single dept usage
        reserved_dates_by_dept_sem = {}              # (dept_id, sem) -> set(dates)

        if department_id == "ALL":
            for c in course_qs:
                sem_val = (c.semester or "").strip()
                course_key = f"{sem_val}||{(c.course_code or '').strip()}||{(c.title or '').strip()}".lower()

                dept_id_val = int(c.department_id) if c.department_id else None
                if dept_id_val is None:
                    continue

                dt = (db_course_date_map_by_dept.get(dept_id_val) or {}).get(course_key)
                if dt:
                    reserved_dates_by_dept_sem.setdefault((dept_id_val, sem_val), set()).add(dt)
        else:
            # Build reserved dates by scanning dept's own courses
            for c in course_qs:
                sem_val = (c.semester or "").strip()
                k = f"{sem_val}||{(c.course_code or '').strip()}||{(c.title or '').strip()}".lower()
                dt = db_course_date_map.get(k)
                if dt:
                    reserved_dates_by_sem.setdefault(sem_val, set()).add(dt)

        # ---------------------------------------------------------------------
        # ✅ Generate schedules
        # ---------------------------------------------------------------------
        to_create = []
        created_count = 0
        skipped_count = 0
        dept_course_counter = {}

        # ✅ for department=ALL
        sem_cursor_by_dept_sem = {}             # (dept_id, sem) -> cursor date
        local_course_date_map_by_dept_sem = {}  # (dept_id, sem) -> {course_key: date}
        blocked_dates_by_dept_sem = {}          # (dept_id, sem) -> set(dates)

        # ✅ IMPORTANT: global map uses course_code-based key (so same code => same date)
        global_common_course_date_map = {}      # global_key -> date

        # ✅ for single department (your existing)
        sem_cursor = {}                  # cursor for NEW course timeline per sem
        local_course_date_map = {}       # course_key -> date in this run
        blocked_dates_by_sem = {}        # sem -> set of blocked dates (reserved + new)

        course_qs_ordered = sorted(
            course_qs,
            key=lambda x: (x.semester or "", x.course_code or "")
        )

        current_semester = None

        for course in course_qs_ordered:
            sem_val = (course.semester or "").strip()

            # ✅ lecture hour rule
            ch = CourseHours.objects.filter(course=course).select_related("hour_config").first()
            if ch and ch.hour_config:
                if (ch.hour_config.lecture_hours or 0) == 0:
                    continue

            course_key = f"{sem_val}||{(course.course_code or '').strip()}||{(course.title or '').strip()}".lower()

            # ✅ NEW: global_key (ONLY used when department_id == "ALL")
            global_key = f"{_norm(sem_val)}||{_norm(course.course_code)}"

            # -----------------------------------------------------------------
            # ✅ CASE A: Single department (UNCHANGED behavior)
            # -----------------------------------------------------------------
            if department_id != "ALL":

                if sem_val != current_semester:
                    current_semester = sem_val
                    sem_cursor[sem_val] = exam_date
                    local_course_date_map = {}

                    # ✅ block reserved dates (only reused courses in THIS dept)
                    blocked_dates_by_sem[sem_val] = set(reserved_dates_by_sem.get(sem_val, set()))

                # 1️⃣ same course in THIS run
                if course_key in local_course_date_map:
                    course_exam_date = local_course_date_map[course_key]

                    

                # 2️⃣ same course exists in DB → reuse (fixed date)
                elif course_key in db_course_date_map:
                    course_exam_date = db_course_date_map[course_key]
                    local_course_date_map[course_key] = course_exam_date
                    blocked_dates_by_sem[sem_val].add(course_exam_date)  # keep blocked

                    

                # 3️⃣ NEW course → start from selected date, skip reserved dates
                else:
                    blocked = blocked_dates_by_sem.get(sem_val, set())
                    probe = sem_cursor.get(sem_val, exam_date)

                    while probe in blocked:
                        probe = next_working_day(probe)

                    course_exam_date = probe
                    local_course_date_map[course_key] = course_exam_date

                    # block date so next NEW course won't reuse it
                    blocked_dates_by_sem[sem_val].add(course_exam_date)

                    # move cursor forward for next NEW course
                    sem_cursor[sem_val] = next_working_day(course_exam_date)
                    dept_degree_fk = get_degree_for_department(dept)

                   

                # ✅ duplicate check
                exists = InternalExamSchedule.objects.filter(
                    degree=dept_degree_fk,
                    department=department_fk,
                    regulation=regulation_fk,
                    course=course,
                    semester=sem_val,
                    exam_date=course_exam_date,
                    session=default_session,
                    internal_assessment=internal_assessment_fk,
                ).exists()

                if exists:
                    skipped_count += 1
                    continue

                for dept in selected_departments:
                    dept_degree_fk = get_degree_for_department(dept)
                    if dept.id not in dept_course_counter:
                        dept_course_counter[dept.id] = 0

                    if course.department_id and int(course.department_id) != int(dept.id):
                        continue

                    exists = InternalExamSchedule.objects.filter(
                        degree=dept_degree_fk,
                        department=dept,
                        regulation=regulation_fk,
                        course=course,
                        semester=sem_val,
                        exam_date=course_exam_date,
                        session=default_session,
                        internal_assessment=internal_assessment_fk,
                    ).exists()

                    if exists:
                        skipped_count += 1
                        continue
                    dept_course_counter[dept.id] += 1
                    ds_key = (int(dept.id), sem_val)
                    batches = batch_map.get(ds_key, [])

                    ds_key = (int(dept.id), sem_val)
                    batches = batch_map.get(ds_key, [])

                   

                    if not batches:
                        
                        continue

                    for batch_name in batches:
                        

                        exists = InternalExamSchedule.objects.filter(
                            degree=dept_degree_fk,
                            department=dept,
                            regulation=regulation_fk,
                            course=course,
                            semester=sem_val,
                            exam_date=course_exam_date,
                            session=default_session,
                            internal_assessment=internal_assessment_fk,
                            batch=batch_name,
                        ).exists()

                        if exists:
                            
                            skipped_count += 1
                            continue

                        to_create.append(
                            InternalExamSchedule(
                                degree=dept_degree_fk,
                                department=dept,
                                regulation=regulation_fk,
                                course=course,
                                semester=sem_val,
                                exam_date=course_exam_date,
                                batch=batch_name,   # ⭐ STORE BATCH HERE
                                session=default_session,
                                internal_assessment=internal_assessment_fk,
                            )
                        )

                        dept_course_counter[dept.id] += 1

            # -----------------------------------------------------------------
            # ✅ CASE B: Department ALL (NEW behavior)
            #   ✅ SAME COURSE CODE across depts => same date (global_key)
            #   ✅ NEW RULE: NO TWO DIFFERENT COURSES ON SAME DATE (per dept+sem)
            # -----------------------------------------------------------------
            else:
                for dept in selected_departments:
                    dept_degree_fk = get_degree_for_department(dept)
                    if dept.id not in dept_course_counter:
                        dept_course_counter[dept.id] = 0

                    if course.department_id and int(course.department_id) != int(dept.id):
                        continue

                    dept_id_val = int(dept.id)
                    ds_key = (dept_id_val, sem_val)

                    if ds_key not in sem_cursor_by_dept_sem:
                        sem_cursor_by_dept_sem[ds_key] = exam_date
                        local_course_date_map_by_dept_sem[ds_key] = {}
                        blocked_dates_by_dept_sem[ds_key] = set(
                            reserved_dates_by_dept_sem.get(ds_key, set())
                        )

                   

                    # ---------- DATE DECISION ----------
                    got_from_global = False  # ⭐ used to update global date if we must shift due to collision

                    if global_key in global_common_course_date_map:
                        course_exam_date = global_common_course_date_map[global_key]
                        got_from_global = True


                    elif course_key in local_course_date_map_by_dept_sem[ds_key]:
                        course_exam_date = local_course_date_map_by_dept_sem[ds_key][course_key]

                       

                    elif course_key in (db_course_date_map_by_dept.get(dept_id_val) or {}):
                        course_exam_date = db_course_date_map_by_dept[dept_id_val][course_key]
                        global_common_course_date_map[global_key] = course_exam_date
                        got_from_global = True  # treated as global fixed for same course code


                    else:
                        blocked = blocked_dates_by_dept_sem.get(ds_key, set())
                        probe = sem_cursor_by_dept_sem.get(ds_key, exam_date)

                        while probe in blocked:
                            probe = next_working_day(probe)

                        course_exam_date = probe
                        global_common_course_date_map[global_key] = course_exam_date
                        sem_cursor_by_dept_sem[ds_key] = next_working_day(course_exam_date)
                        got_from_global = True  # global saved now

                       

                    # -----------------------------------------------------------------
                    # ⭐ IMPORTANT FIX: prevent TWO DIFFERENT COURSES in same dept+sem
                    # If chosen date is already blocked for this dept+sem, shift forward.
                    # If we shift, also update global map so other depts keep the SAME shifted date.
                    # -----------------------------------------------------------------
                    blocked_here = blocked_dates_by_dept_sem.get(ds_key, set())
                    if course_key not in local_course_date_map_by_dept_sem[ds_key]:
                        # only protect "new assignment for this dept+sem"
                        if course_exam_date in blocked_here:
                            old_dt = course_exam_date
                            probe = course_exam_date
                            while probe in blocked_here:
                                probe = next_working_day(probe)
                            course_exam_date = probe

                           
                            if got_from_global:
                                global_common_course_date_map[global_key] = course_exam_date

                    local_course_date_map_by_dept_sem[ds_key][course_key] = course_exam_date
                    blocked_dates_by_dept_sem[ds_key].add(course_exam_date)

                    # ---------- ⭐ BATCH CREATION MOVED HERE ----------
                    batches = batch_map.get(ds_key, [])

                    if not batches:
                       
                        continue

                    for batch_name in batches:

                        exists = InternalExamSchedule.objects.filter(
                            degree=dept_degree_fk,
                            department=dept,
                            regulation=regulation_fk,
                            course=course,
                            semester=sem_val,
                            exam_date=course_exam_date,
                            session=default_session,
                            internal_assessment=internal_assessment_fk,
                            batch=batch_name,
                        ).exists()

                        if exists:
                            
                            skipped_count += 1
                            continue

                       
                        to_create.append(
                            InternalExamSchedule(
                                degree=dept_degree_fk,
                                department=dept,
                                regulation=regulation_fk,
                                course=course,
                                semester=sem_val,
                                exam_date=course_exam_date,
                                batch=batch_name,
                                session=default_session,
                                internal_assessment=internal_assessment_fk,
                            )
                        )

                        dept_course_counter[dept.id] += 1

        if to_create:
            InternalExamSchedule.objects.bulk_create(to_create)
            created_count = len(to_create)

        messages.success(
            request,
            f"Schedule generated. Created: {created_count}, Skipped: {skipped_count}"
        )

        return render(
            request,
            "examination_management/admin/internal_exam_schedule.html",
            {
                "degrees": grouped_degrees,
                "regulations": regulations,
                "posted": True,
                "selected_degree": degree_id,
                "selected_department": department_id,
                "selected_regulation": regulation_id,
                "selected_date": schedule_date,
                "common_semesters": semester_list,
                "matched_courses": course_qs,
                "created_count": created_count,
                "skipped_count": skipped_count,
            }
        )

    return render(
        request,
        "examination_management/admin/internal_exam_schedule.html",
        {"degrees": final_degree_options, "regulations": regulations}
    )






def request_RELAX(v):
    v = (v or "").strip()
    return v if v else ""





from examination_management.models import Add_Department, InternalExamSchedule

def view_internal_exam_schedule(request):
    degree_id = request.GET.get("degree") or ""
    department_id = request.GET.get("department") or ""
    regulation_id = request.GET.get("regulation") or ""
    semester = request.GET.get("semester") or ""
    internal_assessment_id = request.GET.get("internal_assessment") or ""

    qs = InternalExamSchedule.objects.select_related(
        "degree", "department", "regulation", "course"
    ).order_by("department_id", "semester", "exam_date" )

    if degree_id:
        qs = qs.filter(degree_id=degree_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if semester:
        qs = qs.filter(semester=semester)
    if internal_assessment_id:
        qs = qs.filter(internal_assessment_id=internal_assessment_id)

    return render(
        request,
        "examination_management/admin/view_internal_exam_schedule.html",
        {
            "schedules": qs,

            # ✅ ONLY departments that have schedules
            "departments": Add_Department.objects.filter(
                id__in=InternalExamSchedule.objects
                    .values_list("department_id", flat=True)
                    .distinct()
            ),

            # ✅ ONLY semesters that exist in schedules
            "semesters": sorted(
                InternalExamSchedule.objects
                .values_list("semester", flat=True)
                .distinct()
            ),

            "selected_degree": degree_id,
            "selected_department": department_id,
            "selected_regulation": regulation_id,
            "selected_internal_assessment": internal_assessment_id,
            "selected_semester": semester,
        }
    )




# ---------------------------
# ✅ AJAX FILTER ENDPOINTS
# All come ONLY from InternalExamSchedule table
# ---------------------------

def ies_degrees(request):
    qs = InternalExamSchedule.objects.select_related("degree").exclude(degree__isnull=True)
    items = qs.values("degree_id", "degree__degree").distinct().order_by("degree__degree")
    return JsonResponse({"items": [{"id": x["degree_id"], "text": x["degree__degree"]} for x in items]})


def ies_departments(request):
    degree_id = request.GET.get("degree")
    qs = InternalExamSchedule.objects.select_related("department").exclude(department__isnull=True)

    if degree_id:
        qs = qs.filter(degree_id=degree_id)

    items = qs.values("department_id", "department__Department").distinct().order_by("department__Department")
    return JsonResponse({"items": [{"id": x["department_id"], "text": x["department__Department"]} for x in items]})


def ies_regulations(request):
    degree_id = request.GET.get("degree")
    department_id = request.GET.get("department")

    qs = InternalExamSchedule.objects.select_related("regulation").exclude(regulation__isnull=True)
    if degree_id:
        qs = qs.filter(degree_id=degree_id)
    if department_id:
        qs = qs.filter(department_id=department_id)

    items = qs.values("regulation_id", "regulation__year").distinct().order_by("regulation__year")
    return JsonResponse({"items": [{"id": x["regulation_id"], "text": str(x["regulation__year"])} for x in items]})


def ies_semesters(request):
    degree_id = request.GET.get("degree")
    department_id = request.GET.get("department")
    regulation_id = request.GET.get("regulation")

    qs = InternalExamSchedule.objects.exclude(semester__isnull=True).exclude(semester__exact="")
    if degree_id:
        qs = qs.filter(degree_id=degree_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)

    sems = qs.values_list("semester", flat=True).distinct()
    # sort numeric semesters nicely
    sems_sorted = sorted(sems, key=lambda x: int(x) if str(x).isdigit() else 999)

    return JsonResponse({"items": [{"id": s, "text": f"Sem {s}"} for s in sems_sorted]})


def ies_courses(request):
    degree_id = request.GET.get("degree")
    department_id = request.GET.get("department")
    regulation_id = request.GET.get("regulation")
    semester = request.GET.get("semester")

    qs = InternalExamSchedule.objects.select_related("course").exclude(course__isnull=True)
    if degree_id:
        qs = qs.filter(degree_id=degree_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if semester:
        qs = qs.filter(semester=semester)

    items = qs.values(
        "course_id", "course__course_code", "course__title"
    ).distinct().order_by("course__course_code")

    return JsonResponse({
        "items": [{"id": x["course_id"], "text": f'{x["course__course_code"]} - {x["course__title"]}'} for x in items]
    })


def ies_dates(request):
    degree_id = request.GET.get("degree")
    department_id = request.GET.get("department")
    regulation_id = request.GET.get("regulation")
    semester = request.GET.get("semester")
    course_id = request.GET.get("course")

    qs = InternalExamSchedule.objects.exclude(exam_date__isnull=True)
    if degree_id:
        qs = qs.filter(degree_id=degree_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if semester:
        qs = qs.filter(semester=semester)
    if course_id:
        qs = qs.filter(course_id=course_id)

    dates = qs.values_list("exam_date", flat=True).distinct().order_by("exam_date")
    return JsonResponse({"items": [{"id": str(d), "text": d.strftime("%d-%m-%Y")} for d in dates]})






from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from course_management.models import Degree, Semester_Cooldown_Period

def ies_update(request):
    row_id = request.POST.get("id")
    exam_date = request.POST.get("exam_date")
    session = request.POST.get("session")
    next_url = request.POST.get("next") or "view_internal_exam_schedule"

    obj = get_object_or_404(InternalExamSchedule, id=row_id)

    # ✅ update only allowed fields
    obj.exam_date = exam_date
    obj.session = session
    obj.save(update_fields=["exam_date", "session", "updated_at"])

    messages.success(request, "Schedule updated successfully.")
    return redirect(next_url)



def ies_delete(request, pk):
    obj = get_object_or_404(InternalExamSchedule, pk=pk)
    obj.delete()
    messages.success(request, "Schedule deleted successfully.")
    return redirect(request.META.get("HTTP_REFERER", "view_internal_exam_schedule"))


def ies_delete_all(request):
    count = InternalExamSchedule.objects.count()
    InternalExamSchedule.objects.all().delete()
    messages.success(request, f"Deleted ALL schedules ({count} rows).")
    return redirect("view_internal_exam_schedule")
def request_RELAX(v):
    v = (v or "").strip()
    return v if v else ""


from examination_management.models import InternalExamSchedule, InternalTimeTable

from django.shortcuts import redirect
from django.contrib import messages
from django.utils.timezone import now
from django.db import IntegrityError
from django.utils.timezone import now
from django.contrib import messages
from django.db import IntegrityError

def ies_publish(request):
    if request.method != "POST":
        return redirect("view_internal_exam_schedule")

    # ---------------- FILTERS ----------------
    degree = request.POST.get("degree")
    department = request.POST.get("department")
    regulation = request.POST.get("regulation")
    internal_assessment = request.POST.get("internal_assessment")
    semester = request.POST.get("semester")
    course = request.POST.get("course")
    exam_date = request.POST.get("exam_date")
    session = request.POST.get("session")

    qs = InternalExamSchedule.objects.all()

    if degree:
        qs = qs.filter(degree_id=degree)
    if department:
        qs = qs.filter(department_id=department)
    if regulation:
        qs = qs.filter(regulation_id=regulation)
    if internal_assessment:
        qs = qs.filter(internal_assessment_id=internal_assessment)
    if semester:
        qs = qs.filter(semester=semester)
    if course:
        qs = qs.filter(course_id=course)
    if exam_date:
        qs = qs.filter(exam_date=exam_date)
    if session:
        qs = qs.filter(session=session)

    if not qs.exists():
        messages.warning(request, "No schedules found to publish.")
        return redirect(request.META.get("HTTP_REFERER", "view_internal_exam_schedule"))

    publish_date = now().date()

    created_count = 0
    updated_count = 0

    for s in qs:
        obj, created = InternalTimeTable.objects.update_or_create(
    # 🔑 STABLE IDENTIFIER FIELDS
            degree=s.degree,
            department=s.department,
            regulation=s.regulation,
            course=s.course,
            semester=s.semester,
            batch=s.batch,
            internal_assessment=s.internal_assessment,

            # ✏️ FIELDS THAT CAN CHANGE
            defaults={
                "exam_date": s.exam_date,
                "session": s.session,
                "published_date": publish_date,
            }
        )


        if created:
            created_count += 1
        else:
            updated_count += 1

    messages.success(
        request,
        f"Publish completed: {created_count} created, {updated_count} updated."
    )

    return redirect(request.META.get("HTTP_REFERER", "view_internal_exam_schedule"))






def ies_bulk_session_update(request):
    if request.method != "POST":
        return redirect("view_internal_exam_schedule")

    department_id = request.POST.get("department")
    semester = request.POST.get("semester")
    session = request.POST.get("session")

    if not (department_id and semester and session):
        messages.error(request, "All fields are required.")
        return redirect(request.META.get("HTTP_REFERER"))

    qs = InternalExamSchedule.objects.filter(
        department_id=department_id,
        semester=semester
    )

    if not qs.exists():
        messages.warning(
            request,
            "No exam schedules found for selected department and semester."
        )
        return redirect(request.META.get("HTTP_REFERER"))

    updated = qs.update(session=session)

    messages.success(
        request,
        f"Session updated to '{session}' for {updated} exam schedules."
    )

    return redirect(request.META.get("HTTP_REFERER"))




def ies_semesters_by_department(request):
    department_id = request.GET.get("department_id")

    semesters = []

    if department_id:
        semesters = (
            InternalExamSchedule.objects
            .filter(department_id=department_id)
            .values_list("semester", flat=True)
            .distinct()
            .order_by("semester")
        )

    return JsonResponse({
        "items": [{"id": s, "text": f"Semester {s}"} for s in semesters]
    })










from django.http import JsonResponse


def load_adi_degrees(request):
    # degrees that actually exist in InternalTimeTable
    degree_ids = list(
        InternalTimeTable.objects
        .exclude(degree_id__isnull=True)
        .values_list("degree_id", flat=True)
        .distinct()
    )

    degrees_qs = (
        Degree.objects
        .filter(id__in=degree_ids, is_active=True)
        .order_by("duration", "degree")
    )

    data = []

    # group by duration
    for duration, grp in groupby(degrees_qs, key=lambda x: x.duration):
        grp = list(grp)

        # ✅ individual degrees
        for d in grp:
            data.append({"id": str(d.id), "text": d.degree})

        # ✅ grouped option only if more than one degree in same duration
        if len(grp) > 1:
            data.append({
                "id": ",".join(str(d.id) for d in grp),     # ⭐ comma ids
                "text": " / ".join(d.degree for d in grp)   # ex: BE / BTECH
            })

    return JsonResponse(data, safe=False)






def _split_ids(v):
    return [x.strip() for x in (v or "").split(",") if x.strip()]

def load_adi_departments(request):
    degree_ids = _split_ids(request.GET.get("degree_id"))

    qs = InternalTimeTable.objects.all()
    if degree_ids:
        qs = qs.filter(degree_id__in=degree_ids)

    qs = (qs.filter(department__isnull=False)
            .values("department_id", "department__Department")
            .distinct()
            .order_by("department__Department"))

    data = [{"id": q["department_id"], "text": q["department__Department"]} for q in qs]
    return JsonResponse(data, safe=False)



def load_adi_batches(request):
    degree_ids = _split_ids(request.GET.get("degree_id"))
    department_id = request.GET.get("department_id")

    qs = InternalTimeTable.objects.all()
    if degree_ids:
        qs = qs.filter(degree_id__in=degree_ids)
    if department_id:
        qs = qs.filter(department_id=department_id)

    qs = (qs.exclude(batch__isnull=True)
            .exclude(batch__exact="")
            .values_list("batch", flat=True)
            .distinct()
            .order_by("batch"))

    return JsonResponse(list(qs), safe=False)



def load_adi_semesters(request):
    degree_ids = _split_ids(request.GET.get("degree_id"))
    batch = request.GET.get("batch")
    department_id = request.GET.get("department_id")

    qs = InternalTimeTable.objects.all()
    if degree_ids:
        qs = qs.filter(degree_id__in=degree_ids)
    if batch:
        qs = qs.filter(batch=batch)
    if department_id:
        qs = qs.filter(department_id=department_id)

    qs = (qs.exclude(semester__isnull=True)
            .exclude(semester__exact="")
            .values_list("semester", flat=True)
            .distinct()
            .order_by("semester"))

    return JsonResponse(list(qs), safe=False)



def load_adi_iats(request):
    degree_ids = _split_ids(request.GET.get("degree_id"))
    batch = request.GET.get("batch")
    semester = request.GET.get("semester")
    department_id = request.GET.get("department_id")

    qs = InternalTimeTable.objects.filter(internal_assessment__isnull=False)
    if degree_ids:
        qs = qs.filter(degree_id__in=degree_ids)
    if batch:
        qs = qs.filter(batch=batch)
    if semester:
        qs = qs.filter(semester=semester)
    if department_id:
        qs = qs.filter(department_id=department_id)

    qs = (qs.values("internal_assessment_id", "internal_assessment__iat")
            .distinct()
            .order_by("internal_assessment__iat"))

    data = [{"id": q["internal_assessment_id"], "text": q["internal_assessment__iat"]} for q in qs]
    return JsonResponse(data, safe=False)



def published_internal_exam_schedule(request):
    return render(
        request,
        "examination_management/admin/published_internal_exam_schedule.html"
    )


from django.http import JsonResponse
from django.utils.dateformat import format as date_format
def load_adi_timetable(request):
    degree_ids = _split_ids(request.GET.get("degree_id"))
    department_id = request.GET.get("department_id")
    batch = request.GET.get("batch")
    semester = request.GET.get("semester")
    iat_id = request.GET.get("iat_id")

    qs = InternalTimeTable.objects.select_related(
        "degree", "department", "course", "internal_assessment"
    )

    if degree_ids:
        qs = qs.filter(degree_id__in=degree_ids)
    if department_id:
        qs = qs.filter(department_id=department_id)
    if batch:
        qs = qs.filter(batch=batch)
    if semester:
        qs = qs.filter(semester=semester)
    if iat_id:
        qs = qs.filter(internal_assessment_id=iat_id)

    qs = qs.order_by("exam_date", "session", "course__course_code")

    data = []
    for row in qs:
        data.append({
            "exam_date": row.exam_date.strftime("%d-%m-%Y") if row.exam_date else "",
            "session": row.session or "",
            "degree": getattr(row.degree, "degree", "") if row.degree else "",
            "department": getattr(row.department, "Department", "") if row.department else "",
            "batch": row.batch or "",
            "semester": row.semester or "",
            "iat": getattr(row.internal_assessment, "iat", "") if row.internal_assessment else "",
            "course_code": getattr(row.course, "course_code", "") if row.course else "",
            "course_title": getattr(row.course, "title", "") if row.course else "",
            "department_code": getattr(row.department, "department_label", "") if row.department else "",
        })

    return JsonResponse(data, safe=False)


def semester_cooldown_period(request):
    degrees = Degree.objects.filter(is_active=True)
    cooldowns = Semester_Cooldown_Period.objects.select_related('degree').all()

    if request.method == "POST":
        action = request.POST.get("action")
        degree_id = request.POST.get("degree")
        months = request.POST.get("no_of_months")
        cooldown_id = request.POST.get("cooldown_id")

        # DELETE
        if action == "delete":
            if cooldown_id:
                Semester_Cooldown_Period.objects.filter(id=cooldown_id).delete()
                messages.success(request, "Cooldown deleted successfully.")
            else:
                messages.error(request, "Invalid cooldown ID.")
            return redirect("semester_cooldown_period")

        # VALIDATION
        if not degree_id or not months:
            messages.error(request, "All fields are required.")
            return redirect("semester_cooldown_period")

        try:
            months = int(months)
        except ValueError:
            messages.error(request, "Months must be a number.")
            return redirect("semester_cooldown_period")

        degree = get_object_or_404(Degree, id=degree_id)

        # ADD
        if action == "add":
            if Semester_Cooldown_Period.objects.filter(degree=degree).exists():
                messages.error(request, "Cooldown already exists for this degree.")
            else:
                Semester_Cooldown_Period.objects.create(
                    degree=degree,
                    no_of_months=months
                )
                messages.success(request, "Cooldown added successfully.")

        # UPDATE
        elif action == "update":
            if not cooldown_id:
                messages.error(request, "Invalid cooldown ID.")
                return redirect("semester_cooldown_period")

            cooldown = get_object_or_404(Semester_Cooldown_Period, id=cooldown_id)

            # Prevent duplicate on update
            if Semester_Cooldown_Period.objects.filter(degree=degree).exclude(id=cooldown.id).exists():
                messages.error(request, "Cooldown already exists for this degree.")
            else:
                cooldown.degree = degree
                cooldown.no_of_months = months
                cooldown.save()
                messages.success(request, "Cooldown updated successfully.")

        return redirect("semester_cooldown_period")

    return render(request, 'course_management/admin/semester_cooldown_period.html', {
        "degrees": degrees,
        "cooldowns": cooldowns
    })



from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.conf import settings
from django.contrib.staticfiles import finders

# KEEP YOUR EXISTING check_permission IMPORT HERE
# Example:
# from user_accounts.decorators import check_permission

import math
import os
from io import BytesIO
from datetime import datetime
from collections import defaultdict

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.utils import ImageReader

from course_management.models import Hall, Degree, Add_Department, SectionMaster, Hall_Allotment
from user_accounts.models import StudentDetails
from examination_management.models import InternalAssessment

EMPTY = "__EMPTY__"


def _choose_cols(total_seats: int) -> int:
    if total_seats <= 15:
        return 3
    if total_seats <= 30:
        return 5
    if total_seats <= 40:
        return 6
    if total_seats <= 60:
        return 8
    if total_seats <= 80:
        return 10
    return 12


def _build_adjacency(total_seats: int, cols: int):
    """
    Rule:
    - no same department on LEFT
    - no same department on RIGHT
    - no same department on FRONT
    - no same department on BACK
    """
    if total_seats <= 0 or cols <= 0:
        return {}

    rows = int(math.ceil(total_seats / float(cols)))
    adj = {i: set() for i in range(1, total_seats + 1)}

    visual_columns = []
    next_seat = 1

    for col_idx in range(cols):
        col_vals = []
        for _ in range(rows):
            if next_seat <= total_seats:
                col_vals.append(next_seat)
                next_seat += 1
            else:
                break

        if col_idx % 2 == 1:
            col_vals.reverse()

        visual_columns.append(col_vals)

    grid = []
    for row_idx in range(rows):
        row_cells = []
        for col_idx in range(cols):
            if row_idx < len(visual_columns[col_idx]):
                row_cells.append(visual_columns[col_idx][row_idx])
            else:
                row_cells.append(None)
        grid.append(row_cells)

    for r in range(rows):
        for c in range(cols):
            seat = grid[r][c]
            if not seat:
                continue

            if c - 1 >= 0 and grid[r][c - 1]:
                adj[seat].add(grid[r][c - 1])

            if c + 1 < cols and grid[r][c + 1]:
                adj[seat].add(grid[r][c + 1])

            if r - 1 >= 0 and grid[r - 1][c]:
                adj[seat].add(grid[r - 1][c])

            if r + 1 < rows and grid[r + 1][c]:
                adj[seat].add(grid[r + 1][c])

    return adj


def arrange_full_hall(hall_seats: int, dept_counts: dict):
    total_students = sum(dept_counts.values())
    if total_students > hall_seats:
        raise ValueError("Students exceed hall seats")

    cols = _choose_cols(hall_seats)
    adj = _build_adjacency(hall_seats, cols)

    remaining = dict(dept_counts)
    remaining[EMPTY] = hall_seats - total_students

    seats = list(range(1, hall_seats + 1))
    seats.sort(key=lambda s: len(adj[s]), reverse=True)

    assignment = {}

    def candidates(seat):
        used = {assignment[n] for n in adj[seat] if n in assignment}
        opts = []

        for d, cnt in remaining.items():
            if cnt <= 0:
                continue
            if d != EMPTY and d in used:
                continue
            opts.append(d)

        def score(d):
            return -999 if d == EMPTY else remaining[d]

        opts.sort(key=score, reverse=True)
        return opts

    def backtrack(i=0):
        if i == len(seats):
            return True

        seat = seats[i]
        for dept in candidates(seat):
            assignment[seat] = dept
            remaining[dept] -= 1

            if backtrack(i + 1):
                return True

            remaining[dept] += 1
            del assignment[seat]

        return False

    if not backtrack(0):
        raise ValueError("No valid arrangement possible with front/back/left/right department rule")

    return assignment, cols


def _build_visual_grid(total_seats, cols, seat_map):
    if total_seats <= 0 or cols <= 0:
        return [], 0

    rows = int(math.ceil(total_seats / float(cols)))

    visual_columns = []
    next_seat = 1

    for col_idx in range(cols):
        col_vals = []
        for _ in range(rows):
            if next_seat <= total_seats:
                col_vals.append(next_seat)
                next_seat += 1
            else:
                break

        if col_idx % 2 == 1:
            col_vals.reverse()

        visual_columns.append(col_vals)

    grid = []
    for row_idx in range(rows):
        row_cells = []
        for col_idx in range(cols):
            if row_idx < len(visual_columns[col_idx]):
                seat_no = visual_columns[col_idx][row_idx]
                row_cells.append({
                    "seat_no": seat_no,
                    "allot": seat_map.get(seat_no)
                })
            else:
                qs = StudentDetails.objects.filter(
                    department_id=selected_department_id,
                    batch=selected_batch,
        is_active=True
                )
                if selected_section:
                    qs = qs.filter(section=selected_section)

    return grid, rows


def _find_logo():
    logo_rel = "images/ritlogo.png"
    path = finders.find(logo_rel)
    if path and os.path.exists(path):
        return path

    static_root = getattr(settings, "STATIC_ROOT", "")
    if static_root:
        cand = os.path.join(static_root, logo_rel)
        if os.path.exists(cand):
            return cand

    for d in getattr(settings, "STATICFILES_DIRS", []):
        cand = os.path.join(d, logo_rel)
        if os.path.exists(cand):
            return cand

    return None


def _build_filtered_students_queryset(selected_department_id, selected_batch, selected_section):
    qs = StudentDetails.objects.filter(
        department_id=selected_department_id,
        batch=selected_batch
    )

    if selected_section:
        qs = qs.filter(section=selected_section)

    qs = qs.exclude(
        id__in=Hall_Allotment.objects.values_list("student_id", flat=True)
    )

    return qs.order_by("reg_no")


def _rebuild_hall_seating_with_rule(hall, total_seats, students_for_hall):
    if len(students_for_hall) > total_seats:
        raise ValueError(f"Not enough seats. Need {len(students_for_hall)}, seats={total_seats}")

    dept_students = {}
    for s in students_for_hall:
        dept_name = s.department.Department if s.department else "No Department"
        dept_students.setdefault(dept_name, []).append(s)

    dept_counts = {d: len(v) for d, v in dept_students.items()}
    seat_plan, cols_used = arrange_full_hall(total_seats, dept_counts)

    dept_index = {d: 0 for d in dept_students}
    to_create = []

    for seat_no in range(1, total_seats + 1):
        dept = seat_plan.get(seat_no)
        if not dept or dept == EMPTY:
            continue

        student_obj = dept_students[dept][dept_index[dept]]
        dept_index[dept] += 1

        to_create.append(
            Hall_Allotment(hall_id=hall.id, student=student_obj, seat_no=seat_no)
        )

    Hall_Allotment.objects.filter(hall_id=hall.id).delete()
    Hall_Allotment.objects.bulk_create(to_create)

    return len(to_create), cols_used


def _compress_reg_numbers(reg_numbers):
    cleaned = []
    for r in reg_numbers:
        r = str(r).strip()
        if r:
            cleaned.append(r)

    if not cleaned:
        return "-"

    parsed = []
    for reg in cleaned:
        digits = "".join(ch for ch in reg if ch.isdigit())
        if not digits:
            return ", ".join(cleaned)
        parsed.append((reg, digits))

    try:
        full_len = max(len(d) for _, d in parsed)
        nums = []
        reg_map = {}
        for reg, digits in parsed:
            n = int(digits)
            nums.append(n)
            reg_map[n] = reg

        nums = sorted(set(nums))

        ranges = []
        start = nums[0]
        prev = nums[0]

        for n in nums[1:]:
            if n == prev + 1:
                prev = n
            else:
                ranges.append((start, prev))
                start = prev = n
        ranges.append((start, prev))

        out = []
        for s, e in ranges:
            if s == e:
                out.append(reg_map[s])
            elif e == s + 1:
                out.append(f"{reg_map[s]}, {reg_map[e]}")
            else:
                s_txt = str(s).zfill(full_len)
                e_txt = str(e).zfill(full_len)
                out.append(f"{s_txt} - {e_txt}")

        return ", ".join(out)
    except Exception:
        return ", ".join(cleaned)


def _pdf_brand_header_footer(canvas, doc, title, page_size=A4):
    """
    Clean institute header without blue top line.
    """
    page_w, page_h = page_size
    left = doc.leftMargin
    right = doc.rightMargin

    canvas.saveState()

    try:
        logo_path = _find_logo()
        if logo_path and os.path.exists(logo_path):
            img = ImageReader(logo_path)
            iw, ih = img.getSize()
            target_h = 18 * mm
            target_w = target_h * (iw / float(ih))

            canvas.drawImage(
                img,
                left,
                page_h - 30 * mm,
                width=target_w,
                height=target_h,
                preserveAspectRatio=True,
                mask="auto"
            )
    except Exception:
        pass

    center_x = page_w / 2.0

    canvas.setFillColor(colors.HexColor("#1e3a8a"))
    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawCentredString(center_x, page_h - 15 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

    canvas.setFillColor(colors.red)
    canvas.setFont("Helvetica-Bold", 9)
    canvas.drawCentredString(center_x, page_h - 20 * mm, "An Autonomous Institution")

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawCentredString(center_x, page_h - 25 * mm, "EXAMINATION CONTROL OFFICE")

    canvas.setFont("Helvetica", 8)
    canvas.drawCentredString(center_x, page_h - 30 * mm, "Approved by AICTE, New Delhi")
    canvas.drawCentredString(center_x, page_h - 34 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
    canvas.drawCentredString(center_x, page_h - 38 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

    canvas.setFont("Helvetica-Bold", 11)
    canvas.drawCentredString(center_x, page_h - 46 * mm, title)

    canvas.setStrokeColor(colors.grey)
    canvas.setLineWidth(0.6)
    canvas.line(left, page_h - 50 * mm, page_w - right, page_h - 50 * mm)

    canvas.line(left, 12 * mm, page_w - right, 12 * mm)

    canvas.setFont("Helvetica", 8)
    canvas.drawString(left, 7 * mm, f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}")
    canvas.drawRightString(page_w - right, 7 * mm, f"Page {canvas.getPageNumber()}")

    canvas.restoreState()


@check_permission("hall_allotment")
def hall_allotment(request):
    halls = Hall.objects.all().order_by("hall_name")
    degrees = Degree.objects.filter(is_active=True).order_by("degree_code")
    department = Add_Department.objects.filter(is_active=True).order_by("Department_code")

    batch = StudentDetails.objects.order_by("batch").values_list("batch", flat=True).distinct()
    semesters = StudentDetails.objects.order_by("semester").values_list("semester", flat=True).distinct()
    years = StudentDetails.objects.order_by("year").values_list("year", flat=True).distinct()

    iat = InternalAssessment.objects.all().order_by("degree_id", "iat")

    context = {
        "halls": halls,
        "degrees": degrees,
        "departments": department,
        "batches": batch,
        "semesters": semesters,
        "years": years,
        "iat": iat,
    }
    return render(request, "course_management/admin/hall_allotment.html", context)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q
from django.db.utils import ProgrammingError, OperationalError

from course_management.models import Hall_Allotment, Hall
from examination_management.models import InternalExamSchedule

# keep your existing imports also
# from user_accounts.models import StudentDetails
# from masters.models import Degree, Add_Department, SectionMaster


def hall_allotment_detail(request):
    hall_id = (request.GET.get("hall_id") or request.POST.get("hall_id") or "").strip()

    halls = Hall.objects.all().order_by("hall_name")
    degrees = Degree.objects.filter(is_active=True).order_by("degree_code")
    sections = SectionMaster.objects.filter(section__isnull=False).exclude(section="").order_by("section")

    batches = (
        StudentDetails.objects
        .order_by("batch")
        .values_list("batch", flat=True)
        .distinct()
    )

    selected_degree_id = request.GET.get("degree_id") or request.POST.get("degree_id") or ""
    selected_department_id = request.GET.get("department_id") or request.POST.get("department_id") or ""
    selected_batch = request.GET.get("batch") or request.POST.get("batch") or ""
    selected_section = (request.GET.get("section") or request.POST.get("section") or "").strip()

    selected_exam_date = (request.GET.get("exam_date") or request.POST.get("exam_date") or "").strip()
    selected_session = (request.GET.get("session") or request.POST.get("session") or "").strip()

    if selected_degree_id:
        departments = Add_Department.objects.filter(
            is_active=True,
            degree_id=selected_degree_id
        ).order_by("Department_code")
    else:
        departments = Add_Department.objects.none()

    exam_dates = []
    sessions = []

    try:
        exam_schedule_qs = InternalExamSchedule.objects.all()

        if selected_degree_id:
            exam_schedule_qs = exam_schedule_qs.filter(degree_id=selected_degree_id)

        if selected_department_id:
            exam_schedule_qs = exam_schedule_qs.filter(department_id=selected_department_id)

        if selected_batch:
            exam_schedule_qs = exam_schedule_qs.filter(batch=selected_batch)

        exam_dates = list(
            exam_schedule_qs
            .exclude(exam_date__isnull=True)
            .order_by("exam_date")
            .values_list("exam_date", flat=True)
            .distinct()
        )

        session_qs = exam_schedule_qs
        if selected_exam_date:
            session_qs = session_qs.filter(exam_date=selected_exam_date)

        sessions = list(
            session_qs
            .exclude(session__isnull=True)
            .order_by("session")
            .values_list("session", flat=True)
            .distinct()
        )

    except (ProgrammingError, OperationalError):
        exam_dates = []
        sessions = []

    hall = None
    total_seats = 0
    students = StudentDetails.objects.none()
    allotted_students = Hall_Allotment.objects.none()
    allocated_count = 0
    remaining_seats = 0

    if hall_id:
        hall = get_object_or_404(Hall, id=hall_id)

        total_seats = getattr(hall, "benches", 0) or 0
        try:
            total_seats = int(total_seats)
        except Exception:
            total_seats = 0

        allotted_students = (
            Hall_Allotment.objects
            .select_related("student", "student__department", "student__department__degree")
            .filter(hall_id=hall.id)
            .order_by("seat_no")
        )

        if selected_exam_date:
            allotted_students = allotted_students.filter(exam_date=selected_exam_date)

        if selected_session:
            allotted_students = allotted_students.filter(session=selected_session)

        allocated_count = allotted_students.count()
        remaining_seats = max(total_seats - allocated_count, 0)

    action_get = (request.GET.get("action") or "").strip()

    if request.method == "GET" and action_get == "load" and hall:
        if not (selected_degree_id and selected_department_id and selected_batch):
            messages.error(request, "Select Degree, Department and Batch.")
        else:
            students = _build_filtered_students_queryset(
                selected_department_id,
                selected_batch,
                selected_section
            )

            if not students.exists():
                messages.info(request, "No students found for selected filters.")

    if request.method == "POST":
        action = request.POST.get("action")

        if not hall:
            messages.error(request, "Please select Hall first.")
            return redirect(request.path)

        def _reload_url():
            return (
                f"{request.path}?hall_id={hall.id}"
                f"&degree_id={selected_degree_id}"
                f"&department_id={selected_department_id}"
                f"&batch={selected_batch}"
                f"&section={selected_section}"
                f"&exam_date={selected_exam_date}"
                f"&session={selected_session}"
                f"&action=load"
            )

        if action == "clear_hall":
            delete_qs = Hall_Allotment.objects.filter(hall_id=hall.id)

            if selected_exam_date:
                delete_qs = delete_qs.filter(exam_date=selected_exam_date)
            if selected_session:
                delete_qs = delete_qs.filter(session=selected_session)

            deleted, _ = delete_qs.delete()
            messages.success(request, f"Cleared hall. Removed {deleted} allotment(s).")
            return redirect(_reload_url())

        if action == "remove":
            student_id = request.POST.get("remove_student_id")
            if not student_id:
                messages.error(request, "Student not selected for removal.")
                return redirect(_reload_url())

            remove_qs = Hall_Allotment.objects.filter(
                hall_id=hall.id,
                student_id=student_id
            )

            if selected_exam_date:
                remove_qs = remove_qs.filter(exam_date=selected_exam_date)
            if selected_session:
                remove_qs = remove_qs.filter(session=selected_session)

            deleted, _ = remove_qs.delete()

            if deleted:
                messages.success(request, "Student removed.")
            else:
                messages.warning(request, "This student is not allotted in this hall.")

            return redirect(_reload_url())

        if action == "clear":
            return redirect(_reload_url())

        if action == "load":
            if not (selected_degree_id and selected_department_id and selected_batch):
                messages.error(request, "Select Degree, Department and Batch.")
            else:
                students = _build_filtered_students_queryset(
                    selected_department_id,
                    selected_batch,
                    selected_section
                )
                if not students.exists():
                    messages.info(request, "No students found for selected filters.")

        elif action == "allot":
            student_ids = request.POST.getlist("student_ids")

            if not student_ids:
                messages.error(request, "Select at least one student.")
                return redirect(_reload_url())

            existing_allotments_qs = Hall_Allotment.objects.filter(hall_id=hall.id)

            if selected_exam_date:
                existing_allotments_qs = existing_allotments_qs.filter(exam_date=selected_exam_date)
            if selected_session:
                existing_allotments_qs = existing_allotments_qs.filter(session=selected_session)

            existing_students = list(
                StudentDetails.objects
                .filter(
                    id__in=existing_allotments_qs.values_list("student_id", flat=True)
                )
                .select_related("department", "department__degree")
            )

            existing_ids = {s.id for s in existing_students}
            new_ids = [int(sid) for sid in student_ids if int(sid) not in existing_ids]

            new_students = list(
                StudentDetails.objects
                .filter(id__in=new_ids)
                .exclude(id__in=Hall_Allotment.objects.values_list("student_id", flat=True))
                .select_related("department", "department__degree")
                .order_by("reg_no")
            )

            all_students = existing_students + new_students

            try:
                created_count, cols_used = _rebuild_hall_seating_with_rule(hall, total_seats, all_students)
            except ValueError as e:
                messages.error(request, str(e))
                return redirect(_reload_url())

            if selected_exam_date or selected_session:
                update_qs = Hall_Allotment.objects.filter(hall=hall)

                if selected_exam_date:
                    update_qs = update_qs.filter(exam_date__isnull=True)

                if selected_session:
                    update_qs = update_qs.filter(Q(session__isnull=True) | Q(session=""))

                update_data = {}
                if selected_exam_date:
                    update_data["exam_date"] = selected_exam_date
                if selected_session:
                    update_data["session"] = selected_session

                if update_data:
                    update_qs.update(**update_data)

            messages.success(
                request,
                f"Allotted {created_count} student(s) with department seating rule (front/back/left/right) (cols={cols_used})."
            )
            return redirect(_reload_url())

        elif action == "reseat_allot":
            student_ids = request.POST.getlist("student_ids")
            if not student_ids:
                messages.error(request, "Select at least one student.")
                return redirect(_reload_url())

            existing_allotments_qs = Hall_Allotment.objects.filter(hall_id=hall.id)

            if selected_exam_date:
                existing_allotments_qs = existing_allotments_qs.filter(exam_date=selected_exam_date)
            if selected_session:
                existing_allotments_qs = existing_allotments_qs.filter(session=selected_session)

            existing_students = list(
                StudentDetails.objects
                .filter(
                    id__in=existing_allotments_qs.values_list("student_id", flat=True)
                )
                .select_related("department", "department__degree")
            )

            existing_ids = {s.id for s in existing_students}
            new_ids = [int(sid) for sid in student_ids if int(sid) not in existing_ids]

            new_students = list(
                StudentDetails.objects
                .filter(id__in=new_ids)
                .exclude(id__in=Hall_Allotment.objects.values_list("student_id", flat=True))
                .select_related("department", "department__degree")
                .order_by("reg_no")
            )

            all_students = existing_students + new_students

            reseat_clear_qs = Hall_Allotment.objects.filter(hall=hall)
            if selected_exam_date:
                reseat_clear_qs = reseat_clear_qs.filter(exam_date=selected_exam_date)
            if selected_session:
                reseat_clear_qs = reseat_clear_qs.filter(session=selected_session)

            reseat_clear_qs.delete()

            try:
                created_count, cols_used = _rebuild_hall_seating_with_rule(hall, total_seats, all_students)
            except ValueError:
                messages.error(request, "Re-seat not possible. Reduce one dept count or use bigger hall.")
                return redirect(_reload_url())

            if selected_exam_date or selected_session:
                update_qs = Hall_Allotment.objects.filter(hall=hall)

                update_data = {}
                if selected_exam_date:
                    update_data["exam_date"] = selected_exam_date
                if selected_session:
                    update_data["session"] = selected_session

                if update_data:
                    update_qs.update(**update_data)

            messages.success(request, f"Re-seated & allotted {created_count} students (cols={cols_used}).")
            return redirect(_reload_url())

    context = {
        "hall": hall,
        "halls": halls,
        "degrees": degrees,
        "departments": departments,
        "batches": batches,
        "sections": sections,
        "students": students,
        "allotted_students": allotted_students,
        "selected_degree_id": selected_degree_id,
        "selected_department_id": selected_department_id,
        "selected_batch": selected_batch,
        "selected_section": selected_section,
        "selected_exam_date": selected_exam_date,
        "selected_session": selected_session,
        "exam_dates": exam_dates,
        "sessions": sessions,
        "total_seats": total_seats,
        "allocated_count": allocated_count,
        "remaining_seats": remaining_seats,
    }
    return render(request, "course_management/admin/hall_allotment_detail.html", context)


def hall_students(request, hall_id):
    hall = get_object_or_404(Hall, id=hall_id)

    allotments = list(
        Hall_Allotment.objects
        .select_related("student", "student__department", "student__department__degree")
        .filter(hall_id=hall.id)
        .order_by("seat_no")
    )

    total_seats = getattr(hall, "benches", 0) or 0
    try:
        total_seats = int(total_seats)
    except Exception:
        total_seats = 0

    seat_map = {}
    for allot in allotments:
        seat_no = getattr(allot, "seat_no", None)
        if seat_no:
            seat_map[int(seat_no)] = allot

    cols = _choose_cols(total_seats)
    grid, _rows = _build_visual_grid(total_seats, cols, seat_map)

    context = {
        "hall": hall,
        "allotments": allotments,
        "total_seats": total_seats,
        "grid": grid,
    }
    return render(request, "course_management/admin/hall_students.html", context)



from io import BytesIO
from collections import defaultdict
from datetime import datetime

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.staticfiles import finders

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepInFrame
)

from course_management.models import Hall, Hall_Allotment


def _pdf_brand_header_footer(canvas, doc, report_title, page_size):
    page_w, page_h = page_size
    canvas.saveState()

    # Outer border
    canvas.setStrokeColor(colors.black)
    canvas.setLineWidth(0.8)
    canvas.rect(8 * mm, 8 * mm, page_w - 16 * mm, page_h - 16 * mm)

    # Logo path
    logo_path = (
        finders.find("images/ritlogo.png")
        or finders.find("images/rit_logo.png")
        or finders.find("img/ritlogo.png")
        or finders.find("img/rit_logo.png")
        or finders.find("logo.png")
    )

    # ---------------- Header alignment ----------------
    # Keep logo clearly on the left of institute title
    logo_x = 88 * mm
    logo_y = page_h - 36 * mm
    logo_w = 16 * mm
    logo_h = 20 * mm

    if logo_path:
        try:
            canvas.drawImage(
                logo_path,
                logo_x,
                logo_y,
                width=logo_w,
                height=logo_h,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass

    text_center_x = page_w / 2 + 22 * mm

    canvas.setFillColor(colors.HexColor("#223f8d"))
    canvas.setFont("Helvetica-Bold", 17)
    canvas.drawCentredString(text_center_x, page_h - 18 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawCentredString(text_center_x, page_h - 24 * mm, "(An Autonomous Institution)")

    canvas.setFillColor(colors.red)
    canvas.setFont("Helvetica-Bold", 8.8)
    canvas.drawCentredString(
        text_center_x,
        page_h - 30 * mm,
        "Approved by AICTE, New Delhi & Affiliated to Anna University"
    )

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 8.0)
    canvas.drawCentredString(
        text_center_x,
        page_h - 35 * mm,
        "NAAC Accredited with 'A+' Grade & An ISO 9001:2015 Certified Institution"
    )

    canvas.setFillColor(colors.red)
    canvas.setFont("Helvetica-Bold", 8.0)
    canvas.drawCentredString(
        text_center_x,
        page_h - 40 * mm,
        "NBA Accredited UG Programs: CSE, EEE, ECE, MECH and CIVIL"
    )

    # Report title
    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 12.5)
    canvas.drawCentredString(page_w / 2, page_h - 48 * mm, report_title)

    # Divider line
    canvas.setLineWidth(0.5)
    canvas.line(12 * mm, page_h - 54 * mm, page_w - 12 * mm, page_h - 54 * mm)

    # Footer
    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica", 7.5)
    canvas.drawString(12 * mm, 12 * mm, f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}")
    canvas.drawRightString(page_w - 12 * mm, 12 * mm, f"Page {canvas.getPageNumber()}")

    canvas.restoreState()


def hall_students_pdf(request, hall_id):
    hall = get_object_or_404(Hall, id=hall_id)

    allotments = (
        Hall_Allotment.objects
        .select_related("student", "student__department", "student__department__degree")
        .filter(hall_id=hall_id)
        .order_by("seat_no", "id")
    )

    total_seats = getattr(hall, "benches", 0) or 0
    try:
        total_seats = int(total_seats)
    except Exception:
        total_seats = 0

    cols = _choose_cols(total_seats)
    seat_map = {a.seat_no: a for a in allotments if a.seat_no}
    grid, _rows = _build_visual_grid(total_seats, cols, seat_map)

    buf = BytesIO()
    page_size = landscape(A4)
    page_w, page_h = page_size

    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=7 * mm,
        rightMargin=7 * mm,
        topMargin=59 * mm,
        bottomMargin=14 * mm,   # increased to avoid footer overlap
        title="Allotted Students Report",
    )

    styles = getSampleStyleSheet()

    if "MetaBold" not in styles:
        styles.add(ParagraphStyle(
            name="MetaBold",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8.5,
            leading=9,
            alignment=TA_LEFT,
            textColor=colors.black,
        ))

    if "Meta" not in styles:
        styles.add(ParagraphStyle(
            name="Meta",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=9,
            alignment=TA_LEFT,
            textColor=colors.black,
        ))

    if "SmallBold" not in styles:
        styles.add(ParagraphStyle(
            name="SmallBold",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8.8,
            leading=9.4,
            alignment=TA_LEFT,
            textColor=colors.black,
        ))

    if "SummaryHead" not in styles:
        styles.add(ParagraphStyle(
            name="SummaryHead",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=7.4,
            leading=8.2,
            alignment=TA_CENTER,
            textColor=colors.black,
        ))

    if "SummaryCellLeft" not in styles:
        styles.add(ParagraphStyle(
            name="SummaryCellLeft",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7.2,
            leading=8.1,
            alignment=TA_LEFT,
            textColor=colors.black,
        ))

    if "SummaryCellCenter" not in styles:
        styles.add(ParagraphStyle(
            name="SummaryCellCenter",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7.2,
            leading=8.1,
            alignment=TA_CENTER,
            textColor=colors.black,
        ))

    if "CenterSeat" not in styles:
        styles.add(ParagraphStyle(
            name="CenterSeat",
            parent=styles["Normal"],
            alignment=TA_LEFT,
            fontName="Helvetica",
            fontSize=6.8,
            leading=8.0,
            textColor=colors.black,
            wordWrap="CJK",
        ))

    story = []

    # ---------------- Meta ----------------
    meta = Table(
        [[
            Paragraph("Hall:", styles["MetaBold"]),
            Paragraph(str(getattr(hall, "hall_name", "-")), styles["Meta"]),
            Paragraph("Total<br/>Seats:", styles["MetaBold"]),
            Paragraph(str(total_seats), styles["Meta"]),
            Paragraph("Allotted:", styles["MetaBold"]),
            Paragraph(str(allotments.count()), styles["Meta"]),
        ]],
        colWidths=[10 * mm, 45 * mm, 16 * mm, 14 * mm, 14 * mm, 14 * mm],
        hAlign="LEFT",
    )
    meta.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(meta)

    story.append(Spacer(1, 2.5))
    story.append(Paragraph("<b>Department-wise Count Summary</b>", styles["SmallBold"]))
    story.append(Spacer(1, 2))

    # ---------------- Department Summary ----------------
    dept_counts = defaultdict(int)
    for a in allotments:
        st = a.student
        dept_obj = getattr(st, "department", None)
        dept_name = "No Department"
        if dept_obj:
            dept_name = getattr(dept_obj, "Department", None) or "No Department"
        dept_counts[dept_name] += 1

    dept_items = sorted(dept_counts.items(), key=lambda x: x[0].lower())

    summary_data = [[
        Paragraph("S.No", styles["SummaryHead"]),
        Paragraph("Department", styles["SummaryHead"]),
        Paragraph("Count", styles["SummaryHead"]),
    ]]

    for i, (dept_name, cnt) in enumerate(dept_items, start=1):
        summary_data.append([
            Paragraph(str(i), styles["SummaryCellCenter"]),
            Paragraph(dept_name, styles["SummaryCellLeft"]),
            Paragraph(str(cnt), styles["SummaryCellCenter"]),
        ])

    summary_tbl = Table(
        summary_data,
        colWidths=[10 * mm, 116 * mm, 14 * mm],
        repeatRows=1,
        hAlign="LEFT",
    )
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9eef5")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.45, colors.black),
        ("BOX", (0, 0), (-1, -1), 0.65, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (-1, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2.5),
        ("TOPPADDING", (0, 0), (-1, -1), 2.4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.4),
    ]))
    story.append(summary_tbl)

    story.append(Spacer(1, 4))
    story.append(Paragraph("<b>Exam Hall Seating</b>", styles["SmallBold"]))
    story.append(Spacer(1, 2))

    # ---------------- Seating Grid ----------------
    grid_data = []

    for row in grid:
        pdf_row = []
        for cell in row:
            if not cell:
                pdf_row.append(Paragraph("", styles["CenterSeat"]))
                continue

            seat_no = cell["seat_no"]
            allot = cell.get("allot")

            if allot:
                st = allot.student
                reg = getattr(st, "reg_no", "") or getattr(st, "register_no", "") or "-"
                name = getattr(st, "name", None) or getattr(st, "student_name", None) or "-"
                dept = "-"
                if getattr(st, "department", None):
                    dept = getattr(st.department, "Department", None) or "-"

                if len(name) > 16:
                    name = name[:16] + "..."

                txt = (
                    f"<b><font color='#1f4aa8'>Seat {seat_no}</font></b><br/>"
                    f"<b>{reg}</b><br/>"
                    f"{name.upper()}<br/>"
                    f"<font size='6'>{dept}</font>"
                )
                pdf_row.append(Paragraph(txt, styles["CenterSeat"]))
            else:
                pdf_row.append(
                    Paragraph(
                        f"<b><font color='#1f4aa8'>Seat {seat_no}</font></b><br/><font color='#888888'>Empty</font>",
                        styles["CenterSeat"]
                    )
                )

        grid_data.append(pdf_row)

    usable_w = page_w - (doc.leftMargin + doc.rightMargin)
    usable_h = page_h - (doc.topMargin + doc.bottomMargin)

    summary_rows = max(len(summary_data), 1)
    estimated_meta_h = 9 * mm
    estimated_summary_title_h = 5 * mm
    estimated_summary_table_h = 8 * mm + (summary_rows * 5.2 * mm)
    estimated_grid_title_h = 5 * mm

    # Leave more safe space for footer
    remaining_grid_h = usable_h - (
        estimated_meta_h +
        estimated_summary_title_h +
        estimated_summary_table_h +
        estimated_grid_title_h +
        10 * mm
    )

    grid_row_count = max(len(grid_data), 1)
    col_w = usable_w / max(cols, 1)
    row_h = max(21 * mm, remaining_grid_h / grid_row_count)

    grid_table = Table(
        grid_data,
        colWidths=[col_w] * max(cols, 1),
        rowHeights=[row_h] * grid_row_count,
        hAlign="CENTER"
    )
    grid_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.8, colors.HexColor("#9fb8e6")),
        ("BOX", (0, 0), (-1, -1), 0.9, colors.HexColor("#9fb8e6")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(grid_table)

    available_width = page_w - doc.leftMargin - doc.rightMargin
    available_height = page_h - doc.topMargin - doc.bottomMargin

    content = KeepInFrame(
        available_width,
        available_height,
        story,
        mode="shrink",
        hAlign="LEFT",
        vAlign="TOP",
    )

    doc.build(
        [content],
        onFirstPage=lambda canvas, d: _pdf_brand_header_footer(canvas, d, "ALLOTTED STUDENTS REPORT", page_size),
        onLaterPages=lambda canvas, d: _pdf_brand_header_footer(canvas, d, "ALLOTTED STUDENTS REPORT", page_size),
    )

    pdf = buf.getvalue()
    buf.close()

    fname = f"{getattr(hall, 'hall_name', 'hall')}_allotted_students.pdf".replace(" ", "_")
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{fname}"'
    return resp

from io import BytesIO
from collections import defaultdict
import os

from django.http import HttpResponse
from django.contrib.staticfiles import finders
from django.utils import timezone

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from course_management.models import Hall_Allotment


def _compress_reg_numbers(reg_numbers):
    reg_numbers = [str(r).strip() for r in reg_numbers if str(r).strip() and str(r).strip() != "-"]
    if not reg_numbers:
        return "-"

    def split_prefix_num(x):
        i = len(x)
        while i > 0 and x[i - 1].isdigit():
            i -= 1
        return x[:i], x[i:]

    parts = [split_prefix_num(r) for r in reg_numbers]
    result = []
    i = 0

    while i < len(parts):
        prefix, num = parts[i]
        if not num:
            result.append(reg_numbers[i])
            i += 1
            continue

        start = int(num)
        width = len(num)
        j = i

        while j + 1 < len(parts):
            nprefix, nnum = parts[j + 1]
            if nprefix != prefix or not nnum:
                break
            if int(nnum) == int(parts[j][1]) + 1:
                j += 1
            else:
                break

        if j > i:
            end_num = int(parts[j][1])
            result.append(f"{prefix}{str(start).zfill(width)} - {prefix}{str(end_num).zfill(width)}")
        else:
            result.append(f"{prefix}{str(start).zfill(width)}")

        i = j + 1

    return ", ".join(result)


def _pdf_brand_header_footer(canvas, doc, title, page_size=A4):
    canvas.saveState()

    page_w, page_h = page_size

    border_margin = 9 * mm
    canvas.setStrokeColor(colors.black)
    canvas.setLineWidth(0.8)
    canvas.rect(
        border_margin,
        border_margin,
        page_w - 2 * border_margin,
        page_h - 2 * border_margin
    )

    inner_left = 14 * mm
    inner_right = page_w - 14 * mm

    text_center_x = page_w / 2
    header_top_y = page_h - 24 * mm

    logo_path = finders.find("images/ritlogo.png")
    logo_w = 20 * mm
    logo_h = 20 * mm
    logo_x = (page_w / 2) - 65 * mm
    logo_y = page_h - 42 * mm

    if logo_path and os.path.exists(logo_path):
        try:
            canvas.drawImage(
                logo_path,
                logo_x,
                logo_y,
                width=logo_w,
                height=logo_h,
                preserveAspectRatio=True,
                mask="auto"
            )
        except Exception:
            pass

    y = header_top_y

    canvas.setFillColor(colors.HexColor("#1f4aa8"))
    canvas.setFont("Helvetica-Bold", 18)
    txt = "RAMCO INSTITUTE OF TECHNOLOGY"
    w = stringWidth(txt, "Helvetica-Bold", 18)
    canvas.drawString(text_center_x - w / 2, y, txt)

    y -= 7.2 * mm
    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica", 11.5)
    txt = "(An Autonomous Institution)"
    w = stringWidth(txt, "Helvetica", 11.5)
    canvas.drawString(text_center_x - w / 2, y, txt)

    y -= 6.2 * mm
    canvas.setFillColor(colors.red)
    canvas.setFont("Helvetica", 9.8)
    txt = "Approved by AICTE, New Delhi & Affiliated to Anna University, Chennai"
    w = stringWidth(txt, "Helvetica", 9.8)
    canvas.drawString(text_center_x - w / 2, y, txt)

    y -= 5.2 * mm
    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica", 9.6)
    txt = "NAAC Accredited with 'A+' Grade & An ISO 9001:2015 Certified Institution"
    w = stringWidth(txt, "Helvetica", 9.6)
    canvas.drawString(text_center_x - w / 2, y, txt)

    y -= 5.2 * mm
    canvas.setFillColor(colors.red)
    canvas.setFont("Helvetica-Bold", 9.6)
    txt = "NBA Accredited UG Programs: CSE, EEE, ECE, MECH and CIVIL"
    w = stringWidth(txt, "Helvetica-Bold", 9.6)
    canvas.drawString(text_center_x - w / 2, y, txt)

    y -= 11 * mm
    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 12.5)
    canvas.drawString(inner_left, y, title)

    line_y = y - 4 * mm
    canvas.setStrokeColor(colors.black)
    canvas.setLineWidth(0.8)
    canvas.line(inner_left, line_y, inner_right, line_y)

    footer_y = 12 * mm
    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica", 8)
    canvas.drawString(inner_left, footer_y, f"Generated: {timezone.now().strftime('%d-%m-%Y %H:%M')}")
    canvas.drawRightString(inner_right, footer_y, f"Page {doc.page}")

    canvas.restoreState()


def course_hall_arrangement_statement_pdf(request):
    all_allotments = (
        Hall_Allotment.objects
        .select_related("hall", "student", "student__department", "student__department__degree")
        .order_by("hall__hall_name", "seat_no", "student__reg_no")
    )

    if not all_allotments.exists():
        return HttpResponse("No hall allotment data found.", status=400)

    hall_map = defaultdict(lambda: {
        "hall_name": "-",
        "groups": defaultdict(list),
        "total": 0,
    })

    for a in all_allotments:
        hall_obj = getattr(a, "hall", None)
        student = getattr(a, "student", None)

        hall_name = getattr(hall_obj, "hall_name", None) or f"Hall-{getattr(a, 'hall_id', '-')}"
        reg_no = getattr(student, "reg_no", None) or getattr(student, "register_no", None) or "-"
        semester = getattr(student, "semester", None) or "-"
        dept_name = "-"

        dept_obj = getattr(student, "department", None)
        if dept_obj:
            dept_name = getattr(dept_obj, "Department", None) or getattr(dept_obj, "department_label", None) or "-"

        branch = dept_name
        x = branch.lower()

        if x.startswith("computer science and engineering"):
            branch = "CSE"
        elif x.startswith("electronics and communication engineering"):
            branch = "ECE"
        elif x.startswith("electrical and electronics engineering"):
            branch = "EEE"
        elif x.startswith("mechanical engineering"):
            branch = "MECH"
        elif x.startswith("civil engineering"):
            branch = "CIVIL"
        elif x.startswith("artificial intelligence and data science"):
            branch = "AI&DS"
        elif x.startswith("information technology"):
            branch = "IT"
        elif x.startswith("computer science and business systems"):
            branch = "CSBS"
        elif x.startswith("artificial intelligence"):
            branch = "AI"

        sem_branch = f"{semester} Sem. {branch}"

        hall_map[hall_name]["hall_name"] = hall_name
        hall_map[hall_name]["groups"][sem_branch].append(reg_no)
        hall_map[hall_name]["total"] += 1

    hall_items = sorted(hall_map.items(), key=lambda x: str(x[0]).lower())

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=82 * mm,
        bottomMargin=18 * mm,
        title="Hall Arrangement Statement",
    )

    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name="Cell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.3,
        leading=10,
        alignment=TA_LEFT,
        textColor=colors.black,
    ))

    styles.add(ParagraphStyle(
        name="CellCenter",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.3,
        leading=10,
        alignment=TA_CENTER,
        textColor=colors.black,
    ))

    styles.add(ParagraphStyle(
        name="CellBoldCenter",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.8,
        leading=10,
        alignment=TA_CENTER,
        textColor=colors.black,
    ))

    styles.add(ParagraphStyle(
        name="CellBigCenter",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=16,
        alignment=TA_CENTER,
        textColor=colors.black,
    ))

    elements = []

    data = [[
        Paragraph("<b>Sl.<br/>No.</b>", styles["CellBoldCenter"]),
        Paragraph("<b>Hall No.<br/>(Classroom)</b>", styles["CellBoldCenter"]),
        Paragraph("<b>Semester and<br/>Branch</b>", styles["CellBoldCenter"]),
        Paragraph("<b>Register Number of Candidates</b>", styles["CellBoldCenter"]),
        Paragraph("<b>Total No. of<br/>Candidates</b>", styles["CellBoldCenter"]),
    ]]

    sl_no = 1
    for hall_name, hall_info in hall_items:
        groups = hall_info["groups"]
        total = hall_info["total"]

        sem_items = sorted(groups.items(), key=lambda x: str(x[0]).lower())
        first_row = True

        for sem_branch, regs in sem_items:
            reg_text = _compress_reg_numbers(sorted(regs))
            row = []

            if first_row:
                row.append(Paragraph(str(sl_no), styles["CellCenter"]))
                row.append(Paragraph(hall_name, styles["CellCenter"]))
            else:
                row.append("")
                row.append("")

            row.append(Paragraph(sem_branch, styles["Cell"]))
            row.append(Paragraph(reg_text, styles["Cell"]))

            if first_row:
                row.append(Paragraph(str(total), styles["CellBigCenter"]))
            else:
                row.append("")

            data.append(row)
            first_row = False

        sl_no += 1

    col_widths = [14 * mm, 26 * mm, 36 * mm, 88 * mm, 24 * mm]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e9eef5")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (1, -1), "CENTER"),
        ("ALIGN", (4, 0), (4, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]

    current_row = 1
    for _hall_name, hall_info in hall_items:
        row_span = len(hall_info["groups"])
        if row_span > 1:
            style_cmds.extend([
                ("SPAN", (0, current_row), (0, current_row + row_span - 1)),
                ("SPAN", (1, current_row), (1, current_row + row_span - 1)),
                ("SPAN", (4, current_row), (4, current_row + row_span - 1)),
            ])
        current_row += row_span

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)
    elements.append(Spacer(1, 8))

    footer = Table(
        [[
            Paragraph("Form No. EX 02", styles["Cell"]),
            Paragraph("Rev.No.01", styles["CellCenter"]),
            Paragraph("Effective Date: 16.08.2017", styles["CellCenter"]),
        ]],
        colWidths=[55 * mm, 45 * mm, 70 * mm],
        hAlign="LEFT",
    )
    footer.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (-1, 0), 0.6, colors.black),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(footer)

    doc.build(
        elements,
        onFirstPage=lambda canvas, d: _pdf_brand_header_footer(canvas, d, "Hall Arrangement Statement", A4),
        onLaterPages=lambda canvas, d: _pdf_brand_header_footer(canvas, d, "Hall Arrangement Statement", A4),
    )

    buffer.seek(0)
    return HttpResponse(
        buffer.getvalue(),
        content_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="hall_arrangement_statement.pdf"'}
    )


def hall_dept_students_pdf(request, hall_id, dept_id):
    hall = get_object_or_404(Hall, id=hall_id)
    dept = get_object_or_404(Add_Department, id=dept_id)

    allotments = (
        Hall_Allotment.objects
        .select_related("student", "student__department")
        .filter(hall_id=hall_id, student__department_id=dept_id)
        .order_by("student__reg_no", "seat_no")
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title="Department Signature Sheet",
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleX", parent=styles["Title"], fontSize=14, alignment=TA_CENTER, spaceAfter=8))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=10, leading=12))

    story = []

    logo_path = _find_logo()
    if logo_path:
        try:
            logo = Image(logo_path)
            logo.drawHeight = 14 * mm
            logo.drawWidth = 14 * mm
            head = Table(
                [[logo, Paragraph(
                    "<b>RAMCO INSTITUTE OF TECHNOLOGY</b><br/>"
                    "<font color='red'><b>An Autonomous Institution</b></font><br/>"
                    "<b>EXAMINATION CONTROL OFFICE</b><br/>"
                    "Rajapalayam - 626117",
                    styles["Small"]
                )]],
                colWidths=[16 * mm, None],
                hAlign="LEFT",
            )
            head.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]))
            story.append(head)
        except Exception:
            pass

    story.append(Spacer(1, 6))
    story.append(Paragraph("Allotted Students – Signature Sheet", styles["TitleX"]))
    story.append(Paragraph(f"<b>Hall:</b> {hall.hall_name}", styles["Small"]))
    story.append(Paragraph(f"<b>Department:</b> {dept.Department}", styles["Small"]))
    story.append(Paragraph(f"<b>Total:</b> {allotments.count()}", styles["Small"]))
    story.append(Spacer(1, 8))

    data = [["S.No", "Reg No", "Name", "Signature"]]
    for i, a in enumerate(allotments, start=1):
        st = a.student
        reg = getattr(st, "reg_no", "") or "-"
        name = getattr(st, "name", None) or getattr(st, "student_name", None) or "-"
        data.append([str(i), reg, name, ""])

    tbl = Table(data, colWidths=[15*mm, 45*mm, 90*mm, 40*mm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("TOPPADDING", (0, 1), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 12),
    ]))
    story.append(tbl)

    doc.build(story)

    pdf = buf.getvalue()
    buf.close()

    filename = f"{hall.hall_name}_{dept.Department}_Signature.pdf".replace(" ", "_")
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp


def hall_all_departments_signature_pdf(request, hall_id):
    hall = get_object_or_404(Hall, id=hall_id)

    allotments = (
        Hall_Allotment.objects
        .select_related("student", "student__department")
        .filter(hall_id=hall_id)
        .order_by("student__department__Department", "student__reg_no", "seat_no")
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title="All Departments Signature Sheet",
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleX", parent=styles["Title"], fontSize=14, alignment=TA_CENTER, spaceAfter=6))
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=10, leading=12))

    story = []
    dept_map = {}

    for a in allotments:
        dept = getattr(a.student, "department", None)
        dept_id = getattr(dept, "id", None) or 0
        dept_name = getattr(dept, "Department", None) or "No Department"
        dept_map.setdefault((dept_id, dept_name), []).append(a)

    items = list(dept_map.items())
    items.sort(key=lambda x: x[0][1].lower())

    logo_path = _find_logo()

    for idx, ((dept_id, dept_name), dept_allots) in enumerate(items, start=1):
        if logo_path:
            try:
                logo = Image(logo_path)
                logo.drawHeight = 14 * mm
                logo.drawWidth = 14 * mm
                head = Table(
                    [[logo, Paragraph(
                        "<b>RAMCO INSTITUTE OF TECHNOLOGY</b><br/>"
                        "<font color='red'><b>An Autonomous Institution</b></font><br/>"
                        "<b>EXAMINATION CONTROL OFFICE</b><br/>"
                        "Rajapalayam - 626117",
                        styles["Small"]
                    )]],
                    colWidths=[16 * mm, None],
                    hAlign="LEFT",
                )
                head.setStyle(TableStyle([
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]))
                story.append(head)
            except Exception:
                pass

        story.append(Spacer(1, 6))
        story.append(Paragraph("Allotted Students – Signature Sheet", styles["TitleX"]))
        story.append(Paragraph(f"<b>Hall:</b> {hall.hall_name}", styles["Small"]))
        story.append(Paragraph(f"<b>Department:</b> {dept_name}", styles["Small"]))
        story.append(Paragraph(f"<b>Total:</b> {len(dept_allots)}", styles["Small"]))
        story.append(Spacer(1, 8))

        data = [["S.No", "Reg No", "Name", "Signature"]]
        for i, a in enumerate(dept_allots, start=1):
            st = a.student
            reg = getattr(st, "reg_no", "") or "-"
            name = getattr(st, "name", None) or getattr(st, "student_name", None) or "-"
            data.append([str(i), reg, name, ""])

        tbl = Table(data, colWidths=[15*mm, 45*mm, 90*mm, 40*mm], repeatRows=1)
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("TOPPADDING", (0, 1), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 12),
        ]))
        story.append(tbl)

        if idx != len(items):
            story.append(PageBreak())

    doc.build(story)

    pdf = buf.getvalue()
    buf.close()

    filename = f"{hall.hall_name}_All_Departments_Signature.pdf".replace(" ", "_")
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp

from collections import defaultdict
from io import BytesIO
import os
from datetime import datetime

from django.conf import settings
from django.http import HttpResponse
from django.contrib.staticfiles import finders

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame,
    Table, TableStyle, Paragraph, Spacer
)

from course_management.models import Hall_Allotment


def _clean_hall_name(name):
    name = str(name or "").strip()
    if " (" in name and name.endswith(")"):
        name = name.split(" (")[0].strip()
    return name or "-"


def _get_branch_short_name(dept_name):
    dept_name = str(dept_name or "-").strip()
    x = dept_name.lower()

    if x.startswith("computer science and engineering"):
        return "CSE"
    elif x.startswith("electronics and communication engineering"):
        return "ECE"
    elif x.startswith("electrical and electronics engineering"):
        return "EEE"
    elif x.startswith("mechanical engineering"):
        return "MECH"
    elif x.startswith("civil engineering"):
        return "CIVIL"
    elif x.startswith("artificial intelligence and data science"):
        return "AI&DS"
    elif x.startswith("information technology"):
        return "IT"
    elif x.startswith("computer science and business systems"):
        return "CSBS"
    elif x.startswith("artificial intelligence"):
        return "AI"

    return dept_name


def absentees_statement_pdf(request):
    allotments = (
        Hall_Allotment.objects
        .select_related("hall", "student", "student__department", "student__department__degree")
        .order_by("hall__hall_name", "seat_no", "student__reg_no")
    )

    if not allotments.exists():
        return HttpResponse("No hall allotment data found.", status=400)

    # -----------------------------------------
    # Helpers
    # -----------------------------------------
    def safe_str(value):
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    hall_map = defaultdict(lambda: {
        "hall_name": "-",
        "groups": defaultdict(list),
    })

    for a in allotments:
        hall_obj = getattr(a, "hall", None)
        student = getattr(a, "student", None)

        raw_hall_name = getattr(hall_obj, "hall_name", None) or f"Hall-{getattr(a, 'hall_id', '-')}"
        hall_name = _clean_hall_name(raw_hall_name)

        reg_no = (
            getattr(student, "reg_no", None)
            or getattr(student, "register_no", None)
            or "-"
        )

        semester = getattr(student, "semester", None) or "-"
        dept_name = "-"

        dept_obj = getattr(student, "department", None)
        if dept_obj:
            dept_name = (
                getattr(dept_obj, "Department", None)
                or getattr(dept_obj, "department_label", None)
                or "-"
            )

        branch = _get_branch_short_name(dept_name)
        sem_branch = f"{semester} Sem. {branch}"

        hall_map[hall_name]["hall_name"] = hall_name
        hall_map[hall_name]["groups"][sem_branch].append(reg_no)

    hall_items = sorted(hall_map.items(), key=lambda x: str(x[0]).lower())

    # -----------------------------------------
    # Styles / Colors
    # -----------------------------------------
    styles = getSampleStyleSheet()

    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1a4b8c")
    ACCENT_RED = colors.HexColor("#b91c1c")
    DARK_GRAY = colors.HexColor("#111827")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BG_GRAY = colors.HexColor("#f8fafc")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle(
        "title_style",
        parent=styles["Heading1"],
        fontSize=13,
        textColor=PRIMARY_BLUE,
        alignment=TA_LEFT,
        spaceAfter=4,
        fontName="Helvetica-Bold",
        leading=16
    )

    sub_title_style = ParagraphStyle(
        "sub_title_style",
        parent=styles["Normal"],
        fontSize=9.5,
        textColor=MEDIUM_GRAY,
        alignment=TA_LEFT,
        spaceAfter=6,
        fontName="Helvetica"
    )

    table_header_style = ParagraphStyle(
        "table_header_style",
        parent=styles["Normal"],
        fontSize=8.7,
        textColor=colors.white,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        leading=10
    )

    table_cell_style = ParagraphStyle(
        "table_cell_style",
        parent=styles["Normal"],
        fontSize=8.5,
        textColor=DARK_GRAY,
        alignment=TA_LEFT,
        leading=10,
        wordWrap="CJK"
    )

    table_cell_center_style = ParagraphStyle(
        "table_cell_center_style",
        parent=table_cell_style,
        alignment=TA_CENTER
    )

    # -----------------------------------------
    # Document Setup
    # -----------------------------------------
    buffer = BytesIO()
    doc = BaseDocTemplate(
        buffer,
        pagesize=A4,
        title="Absentees Statement",
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        showBoundary=0
    )

    # -----------------------------------------
    # Header / Footer
    # -----------------------------------------
    HEADER_HEIGHT = 36 * mm

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = A4
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        top_y = page_h - 8 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)

        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        # grouped alignment: logo + text visually together like your sample
        logo_w = 24 * mm
        logo_h = 16 * mm
        gap = 5 * mm
        text_block_w = 120 * mm
        total_group_w = logo_w + gap + text_block_w
        group_left = (page_w - total_group_w) / 2

        logo_x = group_left
        logo_y = top_y - 18 * mm

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(
                ImageReader(logo_path),
                logo_x,
                logo_y,
                width=logo_w,
                height=logo_h,
                preserveAspectRatio=True,
                mask="auto"
            )

        text_center_x = group_left + logo_w + gap + (text_block_w / 2)

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawCentredString(text_center_x, top_y - 5.5 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawCentredString(text_center_x, top_y - 12.5 * mm, "An Autonomous Institution")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 8.2)
        canvas.drawCentredString(text_center_x, top_y - 18 * mm, "Approved by AICTE, New Delhi")
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(text_center_x, top_y - 22.5 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
        canvas.drawCentredString(text_center_x, top_y - 27 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

        footer_y = 18 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.8)
        canvas.line(left, footer_y + 7 * mm, right, footer_y + 7 * mm)

        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 8)
        gen_time = datetime.now().strftime("%d %b %Y, %I:%M %p")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")
        canvas.restoreState()

    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin + 6 * mm,
        doc.width,
        doc.height - HEADER_HEIGHT + 8 * mm,
        leftPadding=0,
        rightPadding=0,
        topPadding=0,
        bottomPadding=0,
        id="normal"
    )

    doc.addPageTemplates([
        PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)
    ])

    # -----------------------------------------
    # Table Helper
    # -----------------------------------------
    def create_table(data, col_widths, header_bg=PRIMARY_BLUE, zebra=True):
        t = Table(data, repeatRows=1, colWidths=col_widths)
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8.7),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),

            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ])
        if zebra and len(data) > 1:
            style.add("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY])
        t.setStyle(style)
        return t

    # -----------------------------------------
    # Elements
    # -----------------------------------------
    elements = []

    elements.append(Spacer(1, 17 * mm))
    elements.append(Paragraph("ABSENTEES STATEMENT", title_style))
    elements.append(Paragraph("Exam Type: SEMESTER", sub_title_style))
    elements.append(Spacer(1, 2 * mm))

    data = [[
        Paragraph("Hall No", table_header_style),
        Paragraph("Semester &amp; Branch", table_header_style),
        Paragraph("Total No. of<br/>Students", table_header_style),
        Paragraph("No. of<br/>Students<br/>Present", table_header_style),
        Paragraph("No. of<br/>Students<br/>Absent", table_header_style),
        Paragraph("Register Numbers of<br/>Absentees", table_header_style),
    ]]

    hall_spans = []
    row_idx = 1

    for hall_name, hall_info in hall_items:
        groups = hall_info["groups"]
        sem_items = sorted(groups.items(), key=lambda x: str(x[0]).lower())

        start_row = row_idx

        for i, (sem_branch, regs) in enumerate(sem_items):
            total = len(regs)

            data.append([
                Paragraph(hall_name, table_cell_center_style) if i == 0 else "",
                Paragraph(sem_branch, table_cell_style),
                Paragraph(str(total), table_cell_center_style),
                Paragraph("", table_cell_center_style),
                Paragraph("", table_cell_center_style),
                Paragraph("", table_cell_style),
            ])
            row_idx += 1

        hall_spans.append((start_row, row_idx - 1))

    col_widths = [
        22 * mm,
        50 * mm,
        22 * mm,
        22 * mm,
        22 * mm,
        doc.width - (22 + 50 + 22 + 22 + 22) * mm,
    ]

    table = create_table(data, col_widths, header_bg=SECONDARY_BLUE, zebra=True)

    span_cmds = []
    for start_row, end_row in hall_spans:
        if end_row > start_row:
            span_cmds.append(("SPAN", (0, start_row), (0, end_row)))

    span_cmds += [
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("ALIGN", (2, 1), (4, -1), "CENTER"),
        ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
    ]

    table.setStyle(TableStyle(span_cmds))
    elements.append(table)

    # -----------------------------------------
    # Build PDF
    # -----------------------------------------
    try:
        doc.build(elements)
    except Exception as e:
        print("PDF Generation Error:", e)
        return HttpResponse("PDF generation failed.", status=500)

    buffer.seek(0)
    return HttpResponse(
        buffer.getvalue(),
        content_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="absentees_statement.pdf"'}
    )
    
    
from io import BytesIO
from collections import defaultdict

from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from course_management.models import Hall_Allotment


def absentees_statement_summary_pdf(request):
    allotments = (
        Hall_Allotment.objects
        .select_related("student", "student__department", "student__department__degree")
        .all()
        .order_by("student__department__Department", "student__section", "seat_no")
    )

    if not allotments.exists():
        return HttpResponse("No allotment data found.", status=400)

    def to_roman_year(value):
        val = str(value).strip().upper()
        mapping = {
            "1": "I",
            "2": "II",
            "3": "III",
            "4": "IV",
            "I": "I",
            "II": "II",
            "III": "III",
            "IV": "IV",
            "FIRST": "I",
            "SECOND": "II",
            "THIRD": "III",
            "FOURTH": "IV",
        }
        return mapping.get(val, "-")

    # department full name -> short readable name
    def get_department_name(student):
        if not student or not getattr(student, "department", None):
            return "-"

        dept_name = (getattr(student.department, "Department", "") or "").strip()

        dept_map = {
            "Artificial Intelligence and Data Science": "AD",
            "Artificial Intelligence & Data Science": "AD",
            "Computer Science and Engineering": "CSE",
            "Computer Science & Engineering": "CSE",
            "Information Technology": "IT",
            "Electronics and Communication Engineering": "ECE",
            "Electronics & Communication Engineering": "ECE",
            "Electrical and Electronics Engineering": "EEE",
            "Electrical & Electronics Engineering": "EEE",
            "Mechanical Engineering": "MECH",
            "Civil Engineering": "CIVIL",
            "Computer Science and Business Systems": "CSBS",
            "Computer Science & Business Systems": "CSBS",
            "Artificial Intelligence": "AI",
            "Biomedical Engineering": "BME",
        }

        return dept_map.get(dept_name, dept_name if dept_name else "-")

    group_map = defaultdict(list)

    for a in allotments:
        dept_name = "-"
        year_text = "-"
        section = "-"

        if a.student:
            dept_name = get_department_name(a.student)

            year_value = (
                getattr(a.student, "year", "") or
                getattr(a.student, "current_year", "") or
                getattr(a.student, "study_year", "") or
                ""
            )
            year_text = to_roman_year(year_value)

            section = (
                getattr(a.student, "section", "") or
                getattr(a.student, "student_section", "") or
                "-"
            )

        # same logic, only showing proper department text instead of id
        branch_section = f"{year_text} {dept_name} {section}".strip()
        group_map[branch_section].append(a)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=50 * mm,
        bottomMargin=14 * mm,
        title="Absentees Statement Summary",
    )

    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        "title_style",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        alignment=TA_CENTER,
        spaceAfter=4,
    )

    header_style = ParagraphStyle(
        "header_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9,
        alignment=TA_CENTER,
        textColor=colors.black,
    )

    left_cell_style = ParagraphStyle(
        "left_cell_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
    )

    center_cell_style = ParagraphStyle(
        "center_cell_style",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        alignment=TA_CENTER,
    )

    total_style = ParagraphStyle(
        "total_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        alignment=TA_LEFT,
    )

    total_center_style = ParagraphStyle(
        "total_center_style",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        alignment=TA_CENTER,
    )

    elements.append(Paragraph("ABSENTEES STATEMENT SUMMARY", title_style))
    elements.append(Spacer(1, 6))

    data = [[
        Paragraph("Sl. No", header_style),
        Paragraph("Branch and Section", header_style),
        Paragraph("Total No. of Students", header_style),
        Paragraph("No. of Students Present", header_style),
        Paragraph("No. of Students Absent", header_style),
        Paragraph("Register No. of Absentees", header_style),
    ]]

    grand_total = 0

    for i, (branch_section, students) in enumerate(group_map.items(), start=1):
        total_students = len(students)
        grand_total += total_students

        data.append([
            Paragraph(str(i), center_cell_style),
            Paragraph(branch_section, left_cell_style),
            Paragraph(str(total_students), center_cell_style),
            Paragraph("", center_cell_style),
            Paragraph("", center_cell_style),
            Paragraph("", left_cell_style),
        ])

    data.append([
        Paragraph("", total_center_style),
        Paragraph("Total No. of Students", total_style),
        Paragraph(str(grand_total), total_center_style),
        Paragraph("", total_center_style),
        Paragraph("", total_center_style),
        Paragraph("", total_style),
    ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[12 * mm, 48 * mm, 25 * mm, 30 * mm, 28 * mm, 42 * mm]
    )

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9e2f3")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 1), (4, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.6, colors.black),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
    ]))

    elements.append(table)

    doc.build(
        elements,
        onFirstPage=lambda canvas, d: _pdf_brand_header_footer(
            canvas, d, "ABSENTEES STATEMENT SUMMARY", A4
        ),
        onLaterPages=lambda canvas, d: _pdf_brand_header_footer(
            canvas, d, "ABSENTEES STATEMENT SUMMARY", A4
        ),
    )

    buffer.seek(0)
    return HttpResponse(
        buffer.getvalue(),
        content_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="absentees_statement_summary.pdf"'}
    )
    
    
from io import BytesIO
from collections import defaultdict

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.staticfiles import finders

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
    KeepTogether,
)

from course_management.models import Hall, Hall_Allotment
from examination_management.models import InternalExamSchedule


def _safe_text(value):
    if value is None:
        return "-"
    value = str(value).strip()
    return value if value else "-"


def _display_count(value):
    try:
        value = int(value)
        return "" if value == 0 else str(value)
    except Exception:
        return ""


def _get_student_reg_no(student):
    if not student:
        return "-"
    return (
        getattr(student, "reg_no", None)
        or getattr(student, "register_no", None)
        or getattr(student, "roll_no", None)
        or getattr(student, "student_rollno", None)
        or "-"
    )


def _get_student_name(student):
    if not student:
        return "-"
    return (
        getattr(student, "name", None)
        or getattr(student, "student_name", None)
        or "-"
    )


def _get_absent_count(records):
    count = 0
    for rec in records:
        try:
            if hasattr(rec, "is_absent") and rec.is_absent:
                count += 1
            elif hasattr(rec, "attendance_status") and str(rec.attendance_status).strip().upper() == "ABSENT":
                count += 1
            elif hasattr(rec, "status") and str(rec.status).strip().upper() == "ABSENT":
                count += 1
        except Exception:
            pass
    return count


def _session_to_time_text(session_value):
    session_value = (session_value or "").strip().upper()
    if session_value == "FN":
        return "Time : 9.45 a.m. to 12.45 p.m."
    if session_value == "AN":
        return "Time : 1.45 p.m. to 4.45 p.m."
    return "Time : -"


def _session_to_display(session_value):
    session_value = (session_value or "").strip().upper()
    if session_value == "FN":
        return "Forenoon"
    if session_value == "AN":
        return "Afternoon"
    return "-"


def _draw_same_style_header(canvas, doc):
    page_w, page_h = landscape(A4)
    canvas.saveState()

    # Outer border
    canvas.setLineWidth(0.8)
    canvas.rect(8 * mm, 8 * mm, page_w - 16 * mm, page_h - 16 * mm)

    # Logo
    logo_path = (
        finders.find("images/rit_logo.png")
        or finders.find("img/rit_logo.png")
        or finders.find("images/logo.png")
        or finders.find("img/logo.png")
        or finders.find("logo.png")
    )
    if logo_path:
        try:
            canvas.drawImage(
                logo_path,
                14 * mm,
                page_h - 31 * mm,
                width=16 * mm,
                height=21 * mm,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass

    # Header text
    canvas.setFillColor(colors.HexColor("#1f4aa8"))
    canvas.setFont("Helvetica-Bold", 16)
    canvas.drawCentredString(page_w / 2, page_h - 14 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

    canvas.setFillColor(colors.red)
    canvas.setFont("Helvetica-Bold", 9.5)
    canvas.drawCentredString(page_w / 2, page_h - 19.5 * mm, "An Autonomous Institution")

    canvas.setFillColor(colors.black)
    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawCentredString(page_w / 2, page_h - 25 * mm, "EXAMINATION CONTROL OFFICE")

    canvas.setFont("Helvetica", 8)
    canvas.drawCentredString(page_w / 2, page_h - 29.2 * mm, "Approved by AICTE, New Delhi")
    canvas.drawCentredString(page_w / 2, page_h - 33.2 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
    canvas.drawCentredString(page_w / 2, page_h - 37.2 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

    canvas.setFont("Helvetica-Bold", 11.5)
    canvas.drawCentredString(page_w / 2, page_h - 44.5 * mm, "ATTENDANCE SHEET FOR INTERNAL ASSESSMENT TEST")

    canvas.setLineWidth(0.6)
    canvas.line(13 * mm, page_h - 48 * mm, page_w - 13 * mm, page_h - 48 * mm)

    # Footer
    canvas.setFont("Helvetica", 7.2)
    canvas.drawString(13 * mm, 11 * mm, "Form No. EX 07")
    canvas.drawCentredString(page_w / 2, 11 * mm, "Rev.No.00")
    canvas.drawRightString(page_w - 13 * mm, 11 * mm, "Effective Date")

    canvas.restoreState()


def _get_layout_config(student_count):
    """
    Auto-fit layout so department page stays in single page.
    """
    if student_count <= 12:
        return {
            "title_fs": 9.4,
            "title_leading": 10.2,
            "meta_fs": 7.8,
            "meta_leading": 8.4,
            "head_fs": 7.6,
            "head_leading": 8.2,
            "cell_fs": 7.4,
            "cell_leading": 8.0,
            "box_fs": 7.4,
            "box_leading": 8.0,
            "sign_fs": 8.0,
            "sign_leading": 9.0,
            "header_h": 7.2 * mm,
            "body_h": 6.3 * mm,
            "summary_h1": 7.0 * mm,
            "summary_h2": 7.0 * mm,
            "summary_h3": 8.5 * mm,
            "sp_title": 0.8 * mm,
            "sp_meta": 1.2 * mm,
            "sp_main": 3.0 * mm,
            "sp_sign": 12.0 * mm,
        }
    elif student_count <= 20:
        return {
            "title_fs": 9.0,
            "title_leading": 9.8,
            "meta_fs": 7.5,
            "meta_leading": 8.0,
            "head_fs": 7.2,
            "head_leading": 7.8,
            "cell_fs": 7.0,
            "cell_leading": 7.6,
            "box_fs": 7.1,
            "box_leading": 7.7,
            "sign_fs": 7.8,
            "sign_leading": 8.7,
            "header_h": 6.8 * mm,
            "body_h": 5.7 * mm,
            "summary_h1": 6.5 * mm,
            "summary_h2": 6.5 * mm,
            "summary_h3": 7.8 * mm,
            "sp_title": 0.6 * mm,
            "sp_meta": 1.0 * mm,
            "sp_main": 2.4 * mm,
            "sp_sign": 10.0 * mm,
        }
    elif student_count <= 26:
        return {
            "title_fs": 8.6,
            "title_leading": 9.2,
            "meta_fs": 7.0,
            "meta_leading": 7.5,
            "head_fs": 6.8,
            "head_leading": 7.2,
            "cell_fs": 6.5,
            "cell_leading": 6.9,
            "box_fs": 6.8,
            "box_leading": 7.2,
            "sign_fs": 7.5,
            "sign_leading": 8.3,
            "header_h": 6.2 * mm,
            "body_h": 4.9 * mm,
            "summary_h1": 6.0 * mm,
            "summary_h2": 6.0 * mm,
            "summary_h3": 7.0 * mm,
            "sp_title": 0.4 * mm,
            "sp_meta": 0.8 * mm,
            "sp_main": 1.8 * mm,
            "sp_sign": 7.0 * mm,
        }
    else:
        return {
            "title_fs": 8.2,
            "title_leading": 8.7,
            "meta_fs": 6.7,
            "meta_leading": 7.0,
            "head_fs": 6.4,
            "head_leading": 6.8,
            "cell_fs": 6.0,
            "cell_leading": 6.3,
            "box_fs": 6.4,
            "box_leading": 6.8,
            "sign_fs": 7.2,
            "sign_leading": 8.0,
            "header_h": 5.8 * mm,
            "body_h": 4.4 * mm,
            "summary_h1": 5.5 * mm,
            "summary_h2": 5.5 * mm,
            "summary_h3": 6.4 * mm,
            "sp_title": 0.3 * mm,
            "sp_meta": 0.6 * mm,
            "sp_main": 1.5 * mm,
            "sp_sign": 5.0 * mm,
        }


def internal_exam_hall_department_pdf(request):
    hall_id = (request.GET.get("hall_id") or "").strip()
    schedule_id = (request.GET.get("schedule_id") or "").strip()

    if not hall_id:
        return HttpResponse("hall_id is required", status=400)

    hall = get_object_or_404(Hall, id=hall_id)

    schedule = None
    if schedule_id:
        schedule = get_object_or_404(
            InternalExamSchedule.objects.select_related(
                "degree", "department", "regulation", "course", "internal_assessment"
            ),
            id=schedule_id
        )

    allotments = (
        Hall_Allotment.objects
        .select_related("hall", "student", "student__department", "student__department__degree")
        .filter(hall_id=hall.id)
        .order_by("student__department__Department", "seat_no", "student__reg_no")
    )

    if not allotments.exists():
        return HttpResponse("No allotment data found for this hall.", status=400)

    dept_map = defaultdict(list)
    for a in allotments:
        dept_name = "-"
        try:
            if a.student and a.student.department:
                dept_name = a.student.department.Department or "-"
        except Exception:
            dept_name = "-"
        dept_map[dept_name].append(a)

    related_schedules = []
    if schedule:
        related_qs = InternalExamSchedule.objects.filter(
            degree=schedule.degree,
            department=schedule.department,
            semester=schedule.semester,
            internal_assessment=schedule.internal_assessment,
        ).order_by("exam_date", "id")
        related_schedules = list(related_qs[:7])

    date_headers = []
    if related_schedules:
        for s in related_schedules:
            if s.exam_date:
                date_headers.append(s.exam_date.strftime("%d.%m.%Y"))
            else:
                date_headers.append("-")
    elif schedule and schedule.exam_date:
        date_headers = [schedule.exam_date.strftime("%d.%m.%Y")]
    else:
        date_headers = ["-", "-", "-", "-", "-", "-", "-"]

    while len(date_headers) < 7:
        date_headers.append("-")
    date_headers = date_headers[:7]

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=54 * mm,
        bottomMargin=10 * mm,
        title=f"Department Wise Hall Report - {_safe_text(getattr(hall, 'hall_name', '-'))}"
    )

    styles = getSampleStyleSheet()
    elements = []
    dept_items = list(dept_map.items())

    for index, (dept_name, records) in enumerate(dept_items, start=1):
        section = []

        total_students = len(records)
        absent_count = _get_absent_count(records)
        present_count = total_students - absent_count

        cfg = _get_layout_config(total_students)

        title_style = ParagraphStyle(
            "title_style",
            parent=styles["Normal"],
            alignment=TA_CENTER,
            fontSize=cfg["title_fs"],
            leading=cfg["title_leading"],
            fontName="Helvetica-Bold",
            spaceAfter=0,
        )

        meta_left = ParagraphStyle(
            "meta_left",
            parent=styles["Normal"],
            alignment=TA_LEFT,
            fontSize=cfg["meta_fs"],
            leading=cfg["meta_leading"],
            fontName="Helvetica-Bold",
        )

        meta_right = ParagraphStyle(
            "meta_right",
            parent=styles["Normal"],
            alignment=TA_RIGHT,
            fontSize=cfg["meta_fs"],
            leading=cfg["meta_leading"],
            fontName="Helvetica-Bold",
        )

        cell_left = ParagraphStyle(
            "cell_left",
            parent=styles["Normal"],
            alignment=TA_LEFT,
            fontSize=cfg["cell_fs"],
            leading=cfg["cell_leading"],
            fontName="Helvetica",
        )

        cell_center = ParagraphStyle(
            "cell_center",
            parent=styles["Normal"],
            alignment=TA_CENTER,
            fontSize=cfg["cell_fs"],
            leading=cfg["cell_leading"],
            fontName="Helvetica",
        )

        head_center = ParagraphStyle(
            "head_center",
            parent=styles["Normal"],
            alignment=TA_CENTER,
            fontSize=cfg["head_fs"],
            leading=cfg["head_leading"],
            fontName="Helvetica-Bold",
        )

        box_left = ParagraphStyle(
            "box_left",
            parent=styles["Normal"],
            alignment=TA_LEFT,
            fontSize=cfg["box_fs"],
            leading=cfg["box_leading"],
            fontName="Helvetica",
        )

        sign_style = ParagraphStyle(
            "sign_style",
            parent=styles["Normal"],
            alignment=TA_RIGHT,
            fontSize=cfg["sign_fs"],
            leading=cfg["sign_leading"],
            fontName="Helvetica-Bold",
        )

        first = records[0] if records else None

        degree_name = "-"
        branch_name = dept_name
        semester_name = "-"
        session_name = "-"
        internal_name = "INTERNAL ASSESSMENT TEST"

        if schedule:
            degree_name = _safe_text(getattr(schedule.degree, "degree_code", None) or schedule.degree)
            branch_name = _safe_text(
                getattr(schedule.department, "Department", None) or schedule.department or dept_name
            )
            semester_name = _safe_text(schedule.semester)
            session_name = _session_to_display(schedule.session)
            internal_name = (
                _safe_text(schedule.internal_assessment)
                if schedule.internal_assessment else "INTERNAL ASSESSMENT TEST"
            )
        else:
            try:
                if first and first.student and first.student.department:
                    branch_name = _safe_text(first.student.department.Department or dept_name)
                    if getattr(first.student.department, "degree", None):
                        degree_name = _safe_text(
                            getattr(first.student.department.degree, "degree_code", None)
                            or first.student.department.degree
                        )
                if hasattr(first, "semester") and getattr(first, "semester", None):
                    semester_name = _safe_text(first.semester)
            except Exception:
                pass

        time_text = _session_to_time_text(schedule.session if schedule else "")

        section.append(
            Paragraph(
                f"ATTENDANCE SHEET FOR {internal_name}".upper(),
                title_style
            )
        )
        section.append(Spacer(1, cfg["sp_title"]))

        meta_table = Table(
            [
                [
                    Paragraph(
                        f"Semester, Degree & Branch: {_safe_text(semester_name)} Semester "
                        f"{_safe_text(degree_name)} {_safe_text(branch_name)}",
                        meta_left
                    ),
                    Paragraph(f"HALL: {_safe_text(getattr(hall, 'hall_name', '-'))}", meta_right),
                ],
                [
                    Paragraph(f"Session : {_safe_text(session_name)}", meta_left),
                    Paragraph(time_text, meta_right),
                ],
            ],
            colWidths=[172 * mm, 92 * mm]
        )
        meta_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0.6),
        ]))
        section.append(meta_table)
        section.append(Spacer(1, cfg["sp_meta"]))

        header_row = [
            Paragraph("<b>Sl. No.</b>", head_center),
            Paragraph("<b>Register No.</b>", head_center),
            Paragraph("<b>Name of the Student</b>", head_center),
        ] + [Paragraph(f"<b>{d}</b>", head_center) for d in date_headers]

        table_data = [header_row]

        for i, rec in enumerate(records, start=1):
            student = getattr(rec, "student", None)
            reg_no = _safe_text(_get_student_reg_no(student))
            student_name = _safe_text(_get_student_name(student)).upper()

            row = [
                Paragraph(str(i), cell_center),
                Paragraph(reg_no, cell_center),
                Paragraph(student_name, cell_left),
            ] + [Paragraph("", cell_center) for _ in date_headers]

            table_data.append(row)

        main_table = Table(
            table_data,
            colWidths=[
                12 * mm,   # Sl. No.
                30 * mm,   # Register No.
                74 * mm,   # Name
                20 * mm,
                20 * mm,
                20 * mm,
                20 * mm,
                20 * mm,
                20 * mm,
                20 * mm,
            ],
            rowHeights=[cfg["header_h"]] + [cfg["body_h"]] * (len(table_data) - 1),
            repeatRows=1
        )
        main_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.55, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (1, -1), "CENTER"),
            ("ALIGN", (3, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 1.3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1.3),
            ("TOPPADDING", (0, 0), (-1, -1), 0.8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0.8),
        ]))
        section.append(main_table)
        section.append(Spacer(1, cfg["sp_main"]))

        summary_table = Table(
            [
                [
                    Paragraph("Total number of Present", box_left),
                    Paragraph(_display_count(present_count), cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                ],
                [
                    Paragraph("Total number of Absent", box_left),
                    Paragraph(_display_count(absent_count), cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                ],
                [
                    Paragraph("Hall supervisor's Signature", box_left),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                    Paragraph("", cell_center),
                ],
            ],
            colWidths=[
                43 * mm,
                10 * mm,
                10 * mm,
                10 * mm,
                10 * mm,
                10 * mm,
                10 * mm,
                10 * mm,
            ],
            rowHeights=[cfg["summary_h1"], cfg["summary_h2"], cfg["summary_h3"]]
        )
        summary_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.55, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 1.5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1.5),
            ("TOPPADDING", (0, 0), (-1, -1), 0.8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0.8),
        ]))

        right_summary_wrap = Table(
            [["", summary_table]],
            colWidths=[151 * mm, 102 * mm]
        )
        right_summary_wrap.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        section.append(right_summary_wrap)

        section.append(Spacer(1, cfg["sp_sign"]))

        controller_wrap = Table(
            [["", Paragraph("Controller of Examinations", sign_style)]],
            colWidths=[163 * mm, 90 * mm]
        )
        controller_wrap.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        section.append(controller_wrap)

        elements.append(KeepTogether(section))

        if index != len(dept_items):
            elements.append(PageBreak())

    doc.build(
        elements,
        onFirstPage=_draw_same_style_header,
        onLaterPages=_draw_same_style_header
    )

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="department_wise_hall_{hall.id}.pdf"'
    response.write(pdf)
    return response






from django.http import HttpResponse
from django.db.models import Q
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from collections import defaultdict

from collections import defaultdict
from django.http import HttpResponse
from django.db.models import Min
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm


from collections import defaultdict
from django.http import HttpResponse
from django.db.models import Min
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm, mm
from reportlab.lib.utils import ImageReader
from django.contrib.staticfiles import finders
import os
from collections import defaultdict
from django.http import HttpResponse
from django.contrib.staticfiles import finders
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm, cm
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
import os

def _split_ids(v):
    return [x.strip() for x in (v or "").split(",") if x.strip()]

def admin_internal_timetable_pdf(request):

    # ---------------- GET FILTERS ----------------
    degree_id = request.GET.get("degree_id")  # can be "2" OR "2,1"
    department_id = (request.GET.get("department_id") or "").strip()
    batch = request.GET.get("batch")
    semester = request.GET.get("semester")
    iat_id = request.GET.get("iat_id")

    if not (degree_id and batch and semester and iat_id):
        return HttpResponse("Missing required filters", status=400)

    degree_ids = _split_ids(degree_id)

    # ---------------- QUERY DATA ----------------
    qs = InternalExamSchedule.objects.select_related(
        "degree", "department", "course", "internal_assessment"
    ).filter(
        degree_id__in=degree_ids,              # ✅ FIX
        batch=batch,
        semester=semester,
        internal_assessment_id=iat_id,
    )

    if department_id:
        qs = qs.filter(department_id=department_id)

    qs = qs.order_by("exam_date", "session", "department__department_label", "course__course_code")

    if not qs.exists():
        return HttpResponse("No data found", status=404)

    # ---------------- PREPARE MATRIX ----------------
    dept_list = []
    dept_seen = set()

    for row in qs:
        dept_obj = getattr(row, "department", None)
        dept_label = (
            getattr(dept_obj, "department_label", None) or
            getattr(dept_obj, "Department", None) or
            "-"
        )
        dept_label = str(dept_label).strip() or "-"
        if dept_label not in dept_seen:
            dept_seen.add(dept_label)
            dept_list.append(dept_label)

    dept_list = sorted(dept_list)

    cell_map = defaultdict(lambda: defaultdict(list))
    date_list = []

    for obj in qs:
        dt = obj.exam_date
        date_key = f"{dt.strftime('%d.%m.%y')}\n{(obj.session or '').upper()}"

        if date_key not in cell_map:
            date_list.append(date_key)

        dept_obj = obj.department
        dept_label = (
            getattr(dept_obj, "department_label", None) or
            getattr(dept_obj, "Department", None) or
            "-"
        )
        dept_label = str(dept_label).strip() or "-"

        course = obj.course
        code = getattr(course, "course_code", "-")
        title = getattr(course, "title", "")
        display = code if not title else f"{code} - {title}"

        cell_map[date_key][dept_label].append(display)

    def _parse_date(d):
        try:
            dp, sp = d.split("\n")
            dd, mm_, yy = map(int, dp.split("."))
            return (yy, mm_, dd, 1 if sp == "FN" else 2)
        except Exception:
            return (9999, 12, 31, 99)

    date_list = sorted(set(date_list), key=_parse_date)

    # ---------------- PDF SETUP ----------------
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "inline; filename=internal_exam_schedule.pdf"

    page_size = landscape(A4)

    doc = SimpleDocTemplate(
        response,
        pagesize=page_size,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=45 * mm,
        bottomMargin=15 * mm
    )

    styles = getSampleStyleSheet()
    elements = []

    # ---------------- HEADER DATA ----------------
    first = qs.first()

    # ✅ show correct degree title even for grouped degree ids
    degree_names = list(
        qs.values_list("degree__degree", flat=True).distinct()
    )
    degree_codes = list(
        qs.values_list("degree__degree_code", flat=True).distinct()
    )

    degree_name = " / ".join([d for d in degree_names if d]) if degree_names else ""
    degree_code = " / ".join([c for c in degree_codes if c]) if degree_codes else ""

    dept_name = (
        (getattr(first.department, "department_label", None) or getattr(first.department, "Department", None))
        if department_id and first.department else "All Departments"
    )
    dept_name = str(dept_name).strip() if dept_name else "All Departments"

    iat_name = first.internal_assessment.iat if first.internal_assessment else "-"

    def _find_logo():
        p = finders.find("images/ritlogo.png")
        return p if p and os.path.exists(p) else None

    logo_path = _find_logo()

    def _on_page(canvas, doc_obj):
        canvas.saveState()
        w, h = page_size
        left = 15 * mm

        if logo_path:
            try:
                canvas.drawImage(
                    ImageReader(logo_path),
                    left, h - 28 * mm,
                    height=18 * mm,
                    preserveAspectRatio=True
                )
            except Exception:
                pass

        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(w / 2, h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w / 2, h - 15 * mm, "Rajapalayam - 626117")
        canvas.drawCentredString(w / 2, h - 19 * mm, "Affiliated to Anna University, Chennai")

        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(w / 2, h - 26 * mm, "Internal Examination Schedule")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(w / 2, h - 31 * mm, f"{degree_code} — {degree_name}")
        canvas.drawCentredString(w / 2, h - 35 * mm, dept_name)
        canvas.drawCentredString(w / 2, h - 39 * mm, f"Batch: {batch} | Semester: {semester} | IAT: {iat_name}")

        canvas.line(left, h - 42 * mm, w - left, h - 42 * mm)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(w - left, 12 * mm, f"Page {canvas.getPageNumber()}")

        canvas.restoreState()

    # ---------------- TABLE ----------------
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontSize=9, leading=11)
    head_style = ParagraphStyle("head", parent=styles["Normal"], fontSize=10, leading=12)

    table_data = [[Paragraph("<b>Date</b>", head_style)] +
                  [Paragraph(f"<b>{d}</b>", head_style) for d in dept_list]]

    for d in date_list:
        row = [Paragraph(f"<b>{d}</b>", cell_style)]
        for dept in dept_list:
            items = cell_map[d].get(dept, [])
            row.append(Paragraph("<br/>".join(items), cell_style) if items else Paragraph("-", cell_style))
        table_data.append(row)

    usable_w = page_size[0] - doc.leftMargin - doc.rightMargin
    date_col_w = 3.2 * cm
    other_w = (usable_w - date_col_w) / max(1, len(dept_list))
    col_widths = [date_col_w] + [other_w] * len(dept_list)

    table = Table(table_data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.7, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)

    doc.build(elements, onFirstPage=_on_page, onLaterPages=_on_page)
    return response






from django.http import JsonResponse

def semester_departments_by_degree(request):
    degree_id = (request.GET.get("degree_id") or "").strip()
    if not degree_id:
        return JsonResponse({"items": []})

    # ✅ supports "1" or "1,2,3"
    try:
        degree_ids = [int(x) for x in degree_id.split(",") if x.strip().isdigit()]
    except ValueError:
        degree_ids = []

    if not degree_ids:
        return JsonResponse({"items": []})

    departments = (
        Add_Department.objects
        .filter(is_active=True, degree_id__in=degree_ids)
        .order_by("Department")
    )

    items = [{"id": d.id, "text": d.Department} for d in departments]
    return JsonResponse({"items": items})



def fetch_semesters(request):
    degree_id = (request.GET.get('degree_id') or "").strip()
    department_id = request.GET.get('department_id')
    regulation_id = request.GET.get('regulation_id')

    student_qs = StudentDetails.objects.all()

    # ✅ supports "1" or "1,2,3"
    if degree_id:
        degree_ids = [int(x) for x in degree_id.split(",") if x.strip().isdigit()]
        if degree_ids:
            student_qs = student_qs.filter(department__degree_id__in=degree_ids)

    if department_id:
        student_qs = student_qs.filter(department_id=department_id)

    if regulation_id:
        reg_obj = Regulations.objects.filter(id=regulation_id).first()
        if reg_obj:
            student_qs = student_qs.filter(regulation=reg_obj.year)
        else:
            student_qs = student_qs.none()

    semesters = (
        student_qs.exclude(semester__isnull=True)
                  .exclude(semester__exact="")
                  .values_list("semester", flat=True)
                  .distinct()
    )

    sem_list = []
    for s in semesters:
        s = str(s).strip()
        if s.isdigit():
            sem_list.append(int(s))

    unique_sorted_semesters = [str(x) for x in sorted(set(sem_list))]
    return JsonResponse({'semesters': unique_sorted_semesters})





from itertools import groupby
from collections import defaultdict
from datetime import datetime, date, timedelta

from django.shortcuts import render
from django.http import JsonResponse
from django.contrib import messages

from course_management.models import Course
from user_accounts.models import StudentDetails, Add_Department
  # ✅ adjust import path if different
from examination_management.models import SemesterExamSchedule, Result


# ==========================================================
# Helpers
# ==========================================================
def parse_date_flexible(s: str):
    """Accepts 'YYYY-MM-DD' or 'DD-MM-YYYY' and returns date object."""
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def next_working_day(d: date):
    d = d + timedelta(days=1)
    while d.weekday() == 6:  # Sunday
        d = d + timedelta(days=1)
    return d


def build_degree_options():
    """
    Creates:
      - Individual degree options: value="id" label="BE"
      - Grouped degree options (same duration): value="1,2" label="BE / BTech"
    """
    degrees_qs = Degree.objects.filter(is_active=True).order_by("duration", "degree")

    options = []
    for duration, grp in groupby(degrees_qs, key=lambda x: x.effective_duration):
        grp_list = list(grp)

        # Individual
        for d in grp_list:
            options.append({"ids": str(d.id), "name": d.degree})

        # Grouped
        if len(grp_list) > 1:
            options.append({
                "ids": ",".join(str(d.id) for d in grp_list),
                "name": " / ".join(d.degree for d in grp_list),
            })

    return options


# ==========================================================
# Main View
# ==========================================================
def semester_exam_schedule(request):
    degree_options = build_degree_options()
    regulations = Regulations.objects.all().order_by("year")

    if request.method == "POST":
        degree_ids_str = (request.POST.get("degree") or "").strip()  # could be "1" or "1,2"
        department_id = request.POST.get("department")
        regulation_id = request.POST.get("regulation")
        schedule_date = request.POST.get("schedule_date")
        selected_semesters = request.POST.getlist("semesters")

        # ----------------------------
        # Basic validations
        # ----------------------------
        if not department_id:
            messages.error(request, "Please select department.")
            return render(request, "examination_management/admin/semester_exam_schedule.html",
                          {"degree_options": degree_options, "regulations": regulations})

        if not regulation_id:
            messages.error(request, "Please select regulation.")
            return render(request, "examination_management/admin/semester_exam_schedule.html",
                          {"degree_options": degree_options, "regulations": regulations})

        if not schedule_date:
            messages.error(request, "Please select schedule date.")
            return render(request, "examination_management/admin/semester_exam_schedule.html",
                          {"degree_options": degree_options, "regulations": regulations})

        if not selected_semesters:
            messages.error(request, "Please select at least one semester.")
            return render(request, "examination_management/admin/semester_exam_schedule.html",
                          {"degree_options": degree_options, "regulations": regulations})

        # ----------------------------
        # Department (also provides single Degree FK)
        # ----------------------------
        dept_obj = Add_Department.objects.filter(id=department_id).select_related("degree").first()
        if not dept_obj:
            messages.error(request, "Invalid department selected.")
            return render(request, "examination_management/admin/semester_exam_schedule.html",
                          {"degree_options": degree_options, "regulations": regulations})

        # ✅ real degree comes from department
        degree_fk = dept_obj.degree
        if not degree_fk:
            messages.error(request, "Selected department has no degree mapped.")
            return render(request, "examination_management/admin/semester_exam_schedule.html",
                          {"degree_options": degree_options, "regulations": regulations})

        # optional safety: ensure selected degree group contains this dept's degree
        selected_ids = [x.strip() for x in degree_ids_str.split(",") if x.strip().isdigit()]
        if selected_ids and str(degree_fk.id) not in selected_ids:
            messages.error(request, "Selected department does not match the selected degree group.")
            return render(request, "examination_management/admin/semester_exam_schedule.html",
                          {"degree_options": degree_options, "regulations": regulations})

        regulation_fk = Regulations.objects.filter(id=regulation_id).first()
        if not regulation_fk:
            messages.error(request, "Invalid regulation selected.")
            return render(request, "examination_management/admin/semester_exam_schedule.html",
                          {"degree_options": degree_options, "regulations": regulations})

        start_date = parse_date_flexible(schedule_date)
        if not start_date:
            messages.error(request, "Invalid date format.")
            return render(request, "examination_management/admin/semester_exam_schedule.html",
                          {"degree_options": degree_options, "regulations": regulations})

        default_session = "FN"

        # ----------------------------
        # compute odd/even selected semesters (for arrears logic)
        # ----------------------------
        odd_semesters = []
        even_semesters = []
        for sem in selected_semesters:
            try:
                sem_int = int(str(sem).strip())
                if sem_int % 2 == 0:
                    even_semesters.append(sem_int)
                else:
                    odd_semesters.append(sem_int)
            except Exception:
                # ignore invalid values
                pass

        # --------------------------------
        # Collect student data
        # --------------------------------
        student_qs = StudentDetails.objects.all()

        # ✅ filter by selected degree(s) - since degree dropdown can be "1,2"
        if degree_ids_str:
            deg_ids = [int(x) for x in degree_ids_str.split(",") if x.strip().isdigit()]
            if deg_ids:
                student_qs = student_qs.filter(department__degree_id__in=deg_ids)

        # department filter
        student_qs = student_qs.filter(department_id=department_id)

        # regulation stored as year (CharField)
        student_qs = student_qs.filter(regulation=regulation_fk.year)

        # only selected semesters
        student_qs = student_qs.filter(semester__in=selected_semesters)

        semesters = (
            student_qs.exclude(semester__isnull=True)
                      .exclude(batch__isnull=True)
                      .values_list("semester", "batch", "department")
                      .distinct()
        )

        grouped_data = {}
        for sem, batch, dept in semesters:
            if str(dept) != str(department_id):
                continue
            sem = str(sem).strip()
            batch = str(batch).strip()
            grouped_data.setdefault(sem, []).append((batch, dept))

        if not grouped_data:
            messages.error(request, "No students found for selected filters.")
            return render(request, "examination_management/admin/semester_exam_schedule.html",
                          {"degree_options": degree_options, "regulations": regulations})

        # --------------------------------------------
        # 1) GLOBAL FIXED MAP for REGULAR (same course_code + regulation => same date)
        # --------------------------------------------
        fixed_course_dates = {}
        existing_fixed = (
            SemesterExamSchedule.objects
            .filter(regulation=regulation_fk)
            .select_related("course")
            .values_list("course__course_code", "exam_date")
        )
        for cc, ed in existing_fixed:
            if cc and ed and cc not in fixed_course_dates:
                fixed_course_dates[cc] = ed

        # --------------------------------------------
        # ARREAR FIXED MAP (same course_code arrear => same date) - separate from regular
        # --------------------------------------------
        fixed_arrear_dates = {}
        existing_arrear_fixed = (
            SemesterExamSchedule.objects
            .filter(regulation=regulation_fk, is_failed=True)
            .select_related("course")
            .values_list("course__course_code", "exam_date")
        )
        for cc, ed in existing_arrear_fixed:
            if cc and ed and cc not in fixed_arrear_dates:
                fixed_arrear_dates[cc] = ed

        # --------------------------------------------
        # 2) Dept courses for selected semesters
        # --------------------------------------------
        dept_courses = list({
            c.course_code: c
            for sem in grouped_data.keys()
            for c in Course.objects.filter(
                department_id=department_id,
                regulation=regulation_fk,
                semester=str(sem),
                is_active=True
            )
        }.values())

        if not dept_courses:
            messages.error(request, "No active courses found for the selected department/semester.")
            return render(request, "examination_management/admin/semester_exam_schedule.html",
                          {"degree_options": degree_options, "regulations": regulations})

        # ==========================================================
        # REGULAR EXAMS (Engine)
        # ==========================================================

        # Dept booked dates (THIS department only)
        dept_booked_dates = set(
            SemesterExamSchedule.objects
            .filter(department_id=department_id, regulation=regulation_fk)
            .values_list("exam_date", flat=True)
        )

        # date -> set(course_codes) in THIS dept (regular + arrear existing)
        dept_date_to_codes = defaultdict(set)
        existing_date_course = (
            SemesterExamSchedule.objects
            .filter(department_id=department_id, regulation=regulation_fk)
            .select_related("course")
            .values_list("exam_date", "course__course_code")
        )
        for ed, cc in existing_date_course:
            if ed and cc:
                dept_date_to_codes[ed].add(cc)

        # Per-semester booked dates for GAP RULE
        sem_booked_dates = {str(sem): set() for sem in grouped_data.keys()}
        existing_sem_dates = (
            SemesterExamSchedule.objects
            .filter(department_id=department_id, regulation=regulation_fk)
            .values_list("semester", "exam_date")
        )
        for sem, ed in existing_sem_dates:
            if sem is not None and ed is not None:
                sem_booked_dates.setdefault(str(sem).strip(), set()).add(ed)

        # Add fixed course dates into THIS department buckets (if those courses exist in this dept/sem)
        for sem in grouped_data.keys():
            sem = str(sem).strip()
            sem_codes = Course.objects.filter(
                department_id=department_id,
                regulation=regulation_fk,
                semester=sem,
                is_active=True
            ).values_list("course_code", flat=True)

            for cc in sem_codes:
                if cc in fixed_course_dates:
                    ed = fixed_course_dates[cc]
                    dept_booked_dates.add(ed)
                    dept_date_to_codes[ed].add(cc)
                    sem_booked_dates.setdefault(sem, set()).add(ed)

        # Build NON-FIXED pending courses semester-wise
        pending_by_sem = {}
        for sem in grouped_data.keys():
            sem = str(sem).strip()

            sem_courses = list({
                c.course_code: c
                for c in Course.objects.filter(
                    department_id=department_id,
                    regulation=regulation_fk,
                    semester=sem,
                    is_active=True
                )
            }.values())

            sem_pending = [c for c in sem_courses if c.course_code not in fixed_course_dates]
            sem_pending.sort(key=lambda x: str(x.course_code or ""))
            if sem_pending:
                pending_by_sem[sem] = sem_pending

        def is_valid_for_semester(sem: str, cand: date) -> bool:
            # skip Sunday
            if cand.weekday() == 6:
                return False

            # SAME DATE rule (dept only)
            if cand in dept_booked_dates:
                return False

            # GAP rule (same semester only)
            prev_cal = cand - timedelta(days=1)
            next_cal = cand + timedelta(days=1)
            if prev_cal in sem_booked_dates.get(sem, set()):
                return False
            if next_cal in sem_booked_dates.get(sem, set()):
                return False

            return True

        allocated_dates = {}  # course_code -> exam_date

        # Greedy day-by-day allocation
        cand = start_date
        sem_cycle = sorted(pending_by_sem.keys(), key=lambda x: int(x) if str(x).isdigit() else 999)

        while any(pending_by_sem.get(s) for s in sem_cycle):
            if cand.weekday() == 6:
                cand = next_working_day(cand)
                continue

            for sem in sem_cycle:
                if not pending_by_sem.get(sem):
                    continue

                course = pending_by_sem[sem][0]
                if is_valid_for_semester(sem, cand):
                    allocated_dates[course.course_code] = cand

                    dept_booked_dates.add(cand)
                    dept_date_to_codes[cand].add(course.course_code)
                    sem_booked_dates.setdefault(sem, set()).add(cand)

                    fixed_course_dates[course.course_code] = cand
                    pending_by_sem[sem].pop(0)
                    break

            cand = next_working_day(cand)

        # --------------------------------------------
        # SAVE REGULAR schedules (selected semesters only)
        # --------------------------------------------
        created_count = 0

        for sem, batch_dept_list in grouped_data.items():
            sem = str(sem).strip()

            sem_courses = list({
                c.course_code: c
                for c in Course.objects.filter(
                    department_id=department_id,
                    regulation=regulation_fk,
                    semester=sem,
                    is_active=True
                )
            }.values())

            for course in sem_courses:
                course_code = course.course_code
                exam_day = fixed_course_dates.get(course_code) or allocated_dates.get(course_code)
                if not exam_day:
                    continue

                for batch, dept_id in batch_dept_list:
                    if str(dept_id) != str(department_id):
                        continue

                    exists = SemesterExamSchedule.objects.filter(
                        degree=degree_fk,
                        department=dept_obj,
                        regulation=regulation_fk,
                        semester=sem,
                        batch=str(batch),
                        course=course
                    ).exists()
                    if exists:
                        continue

                    SemesterExamSchedule.objects.create(
                        degree=degree_fk,
                        department=dept_obj,
                        regulation=regulation_fk,
                        semester=sem,
                        batch=str(batch),
                        course=course,
                        exam_date=exam_day,
                        session=default_session,
                    )
                    created_count += 1

        # ==========================================================
        # ARREARS (same failed course_code => SAME DATE across batches)
        # - Date must be FREE in this dept (not clashing with regular schedule)
        # ==========================================================

        # decide which failed semesters to pull based on selected odd/even set
        target_failed_semesters = set()
        if even_semesters:
            target_failed_semesters.update([str(x) for x in range(1, 13, 2)])  # odd
        if odd_semesters:
            target_failed_semesters.update([str(x) for x in range(2, 13, 2)])  # even

        student_ids = list(student_qs.values_list("id", flat=True))

        failed_qs = (
            Result.objects
            .filter(
                degree=degree_fk,
                department=dept_obj,
                regulation=regulation_fk,
                student_id__in=student_ids,
            )
            .exclude(grade__isnull=True)
            .filter(grade__iexact="U")
        )

        if target_failed_semesters:
            failed_qs = failed_qs.filter(semester__in=list(target_failed_semesters))

        failed_rows = list(
            failed_qs
            .select_related("course")
            .exclude(course__isnull=True)
            .values_list("semester", "batch", "course__course_code", "course_id")
            .distinct()
        )

        def next_free_day_arrear(d: date) -> date:
            x = d
            while x.weekday() == 6 or x in dept_booked_dates:
                x = next_working_day(x)
            return x

        arrear_cur = start_date

        # group by course_code so date decided once
        rows_by_code = defaultdict(list)
        for failed_sem, failed_batch, course_code, course_id in failed_rows:
            if course_code:
                rows_by_code[course_code].append((failed_sem, failed_batch, course_id))

        for course_code in sorted(rows_by_code.keys()):
            # decide ONE date for this arrear course_code
            if course_code in fixed_arrear_dates:
                arrear_exam_day = fixed_arrear_dates[course_code]
                if arrear_exam_day in dept_booked_dates:
                    arrear_exam_day = next_free_day_arrear(arrear_exam_day)
                    fixed_arrear_dates[course_code] = arrear_exam_day
            else:
                arrear_exam_day = next_free_day_arrear(arrear_cur)
                fixed_arrear_dates[course_code] = arrear_exam_day

            # reserve once
            dept_booked_dates.add(arrear_exam_day)
            dept_date_to_codes[arrear_exam_day].add(course_code)
            arrear_cur = next_working_day(arrear_exam_day)

            # create rows for each (semester,batch) for this course_code
            for failed_sem, failed_batch, course_id in rows_by_code[course_code]:
                course_obj = Course.objects.filter(id=course_id).first()
                if not course_obj:
                    continue

                exists = SemesterExamSchedule.objects.filter(
                    degree=degree_fk,
                    department=dept_obj,
                    regulation=regulation_fk,
                    semester=str(failed_sem).strip(),
                    batch=str(failed_batch).strip() if failed_batch else "",
                    course=course_obj
                ).exists()
                if exists:
                    continue

                SemesterExamSchedule.objects.create(
                    degree=degree_fk,
                    department=dept_obj,
                    regulation=regulation_fk,
                    semester=str(failed_sem).strip(),
                    batch=str(failed_batch).strip() if failed_batch else "",
                    course=course_obj,
                    exam_date=arrear_exam_day,
                    session=default_session,
                    is_failed=True
                )
                created_count += 1

        messages.success(request, f"{created_count} schedules created successfully (including arrears).")

        return render(
            request,
            "examination_management/admin/semester_exam_schedule.html",
            {"degree_options": degree_options, "regulations": regulations}
        )

    # GET
    return render(
        request,
        "examination_management/admin/semester_exam_schedule.html",
        {"degree_options": degree_options, "regulations": regulations}
    )




def view_semester_exam_schedule(request):
    degree_id = request.GET.get("degree") or ""
    department_id = request.GET.get("department") or ""
    regulation_id = request.GET.get("regulation") or ""
    semester = request.GET.get("semester") or ""

    qs = (
        SemesterExamSchedule.objects.select_related("degree", "department", "regulation", "course")
        .order_by("department_id", "semester", "exam_date", "course__course_code", "batch")
    )

    if degree_id:
        qs = qs.filter(degree_id=degree_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if semester:
        qs = qs.filter(semester=semester)

    # ✅ SHOW EACH COURSE ONLY ONCE (even if multiple batches exist)
    # keeps the earliest row (because of ordering) per (dept, sem, exam_date, course_code)
    unique_ids = []
    seen = set()
    for row in qs:
        key = (
            row.department_id,
            str(row.semester or "").strip(),
            row.exam_date,
            (row.course.course_code if row.course else None),
        )
        if key in seen:
            continue
        seen.add(key)
        unique_ids.append(row.id)

    qs = (
        SemesterExamSchedule.objects.select_related("degree", "department", "regulation", "course")
        .filter(id__in=unique_ids)
        .order_by("department_id", "semester", "exam_date", "course__course_code")
    )

    # ✅ filter dropdown sources must come from SemesterExamSchedule
    departments = Add_Department.objects.filter(
        id__in=SemesterExamSchedule.objects.values_list("department_id", flat=True).distinct()
    ).order_by("Department")

    semesters = sorted(
        SemesterExamSchedule.objects.values_list("semester", flat=True).distinct()
    )

    degrees = Degree.objects.filter(
        id__in=SemesterExamSchedule.objects.values_list("degree_id", flat=True).distinct()
    ).order_by("degree")

    regulations = Regulations.objects.filter(
        id__in=SemesterExamSchedule.objects.values_list("regulation_id", flat=True).distinct()
    ).order_by("year")

    return render(
        request,
        "examination_management/admin/view_semester_exam_schedule.html",
        {
            "schedules": qs,
            "degrees": degrees,
            "departments": departments,
            "regulations": regulations,
            "semesters": semesters,

            "selected_degree": degree_id,
            "selected_department": department_id,
            "selected_regulation": regulation_id,
            "selected_semester": semester,
        }
    )






def ses_update(request):
    row_id = request.POST.get("id")
    exam_date = request.POST.get("exam_date")
    session = request.POST.get("session")
    next_url = request.POST.get("next") or "view_internal_exam_schedule"

    obj = get_object_or_404(SemesterExamSchedule, id=row_id)

    # ✅ update only allowed fields
    obj.exam_date = exam_date
    obj.session = session
    obj.save(update_fields=["exam_date", "session", "updated_at"])

    messages.success(request, "Schedule updated successfully.")
    return redirect(next_url)



from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages

def ses_delete(request, pk):
    obj = get_object_or_404(SemesterExamSchedule, pk=pk)  # ✅ FIXED
    obj.delete()
    messages.success(request, "Schedule deleted successfully.")
    return redirect(request.META.get("HTTP_REFERER", "view_semester_exam_schedule"))


def ses_delete_all(request):
    if request.method != "POST":
        return redirect("view_semester_exam_schedule")

    # ✅ delete from SemesterExamSchedule (NOT InternalExamSchedule)
    count = SemesterExamSchedule.objects.count()
    SemesterExamSchedule.objects.all().delete()

    messages.success(request, f"Deleted ALL semester exam schedules ({count} rows).")

    # ✅ stay on same page
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or "view_semester_exam_schedule"
    return redirect(next_url)
def request_RELAX(v):
    v = (v or "").strip()
    return v if v else ""





from django.http import JsonResponse

def ses_semesters_by_department(request):
    department_id = (request.GET.get("department_id") or "").strip()

    qs = SemesterExamSchedule.objects.all()
    if department_id:
        qs = qs.filter(department_id=department_id)

    semesters = (
        qs.exclude(semester__isnull=True)
          .exclude(semester__exact="")
          .values_list("semester", flat=True)
          .distinct()
    )

    # sort numeric-safe
    def _k(x):
        try:
            return int(str(x).strip())
        except Exception:
            return 9999

    items = [{"id": s, "text": str(s)} for s in sorted(semesters, key=_k)]
    return JsonResponse({"items": items})



def ses_bulk_session_update(request):
    if request.method != "POST":
        return redirect("view_semester_exam_schedule")

    department_id = request.POST.get("department")
    semester = request.POST.get("semester")
    session = request.POST.get("session")

    if not (department_id and semester and session):
        messages.error(request, "All fields are required.")
        return redirect(request.META.get("HTTP_REFERER"))

    qs = SemesterExamSchedule.objects.filter(
        department_id=department_id,
        semester=semester
    )

    if not qs.exists():
        messages.warning(request, "No semester exam schedules found for selected department and semester.")
        return redirect(request.META.get("HTTP_REFERER"))

    updated = qs.update(session=session)

    messages.success(request, f"Session updated to '{session}' for {updated} semester exam schedules.")
    return redirect(request.META.get("HTTP_REFERER"))





def ses_degrees(request):
    """
    Degrees that exist in SemesterExamSchedule.
    """
    base = SemesterExamSchedule.objects.all()
    degree_ids = base.values_list("degree_id", flat=True).distinct()

    items = list(
        Degree.objects.filter(id__in=degree_ids)
        .order_by("degree")
        .values("id", "degree")
    )
    return JsonResponse(
        {"items": [{"id": x["id"], "text": x["degree"]} for x in items]},
        safe=False
    )


def ses_departments(request):
    """
    Departments that exist in SemesterExamSchedule,
    optionally filtered by degree.
    """
    degree_id = (request.GET.get("degree") or "").strip()

    base = SemesterExamSchedule.objects.all()
    if degree_id:
        base = base.filter(degree_id=degree_id)

    dept_ids = base.values_list("department_id", flat=True).distinct()

    items = list(
        Add_Department.objects.filter(id__in=dept_ids)
        .order_by("Department")
        .values("id", "Department")
    )
    return JsonResponse(
        {"items": [{"id": x["id"], "text": x["Department"]} for x in items]},
        safe=False
    )



def ses_regulations(request):
    """
    Regulations that exist in SemesterExamSchedule,
    filtered by degree + department if provided.
    """
    degree_id = (request.GET.get("degree") or "").strip()
    department_id = (request.GET.get("department") or "").strip()

    base = SemesterExamSchedule.objects.all()
    if degree_id:
        base = base.filter(degree_id=degree_id)
    if department_id:
        base = base.filter(department_id=department_id)

    reg_ids = base.values_list("regulation_id", flat=True).distinct()

    # If your Regulations model uses "year" for sorting, keep it.
    regs = Regulations.objects.filter(id__in=reg_ids).order_by("year")

    return JsonResponse(
        {"items": [{"id": r.id, "text": str(r)} for r in regs]},
        safe=False
    )



def ses_semesters(request):
    """
    Semesters that exist in SemesterExamSchedule,
    filtered by degree + department + regulation if provided.
    Returns list of strings for semester.
    """
    degree_id = (request.GET.get("degree") or "").strip()
    department_id = (request.GET.get("department") or "").strip()
    regulation_id = (request.GET.get("regulation") or "").strip()

    base = SemesterExamSchedule.objects.all()
    if degree_id:
        base = base.filter(degree_id=degree_id)
    if department_id:
        base = base.filter(department_id=department_id)
    if regulation_id:
        base = base.filter(regulation_id=regulation_id)

    sems = (
        base.exclude(semester__isnull=True)
            .exclude(semester__exact="")
            .values_list("semester", flat=True)
            .distinct()
    )

    # Sort numerically if possible, else as string
    def _sort_key(x):
        try:
            return (0, int(str(x).strip()))
        except Exception:
            return (1, str(x).strip())

    sem_list = sorted([str(s).strip() for s in sems], key=_sort_key)

    return JsonResponse(
        {"items": [{"id": s, "text": s} for s in sem_list]},
        safe=False
    )



from django.db.models.functions import TruncMonth
from django.utils.timezone import localdate


def load_sem_adi_degrees(request):
    """
    Always load Degree list from Degree master.
    Returns: [{id: <degree_id>, text: <degree_name>}]
    """
    qs = SemesterExamScheduletimetable.objects.all()


    qs = (
        qs.filter(degree__isnull=False)
          .values("degree_id", "degree__degree")
          .distinct()
          .order_by("degree__degree")
    )

    data = [{"id": q["degree_id"], "text": q["degree__degree"]} for q in qs]



    return JsonResponse(data, safe=False)


def load_sem_adi_departments(request):
    degree_id = request.GET.get("degree_id")

    qs = SemesterExamScheduletimetable.objects.filter(degree_id=degree_id)

    qs = (
        qs.filter(department__isnull=False)
        .values("department_id", "department__Department")
        .distinct()
        .order_by("department__Department")
    )

    data = [{"id": q["department_id"], "text": q["department__Department"]} for q in qs]
    return JsonResponse(data, safe=False)


def load_sem_adi_batches(request):
    degree_id = request.GET.get("degree_id")
    department_id = request.GET.get("department_id")

    qs = SemesterExamScheduletimetable.objects.filter(degree_id=degree_id)

    if department_id:
        qs = qs.filter(department_id=department_id)

    qs = (
        qs.exclude(batch__isnull=True)
          .exclude(batch__exact="")
          .values_list("batch", flat=True)
          .distinct()
          .order_by("batch")
    )

    return JsonResponse(list(qs), safe=False)



def load_sem_adi_semesters(request):
    degree_id = request.GET.get("degree_id")
    batch = request.GET.get("batch")
    department_id = request.GET.get("department_id")

    qs = SemesterExamScheduletimetable.objects.filter(
        degree_id=degree_id,
        batch=batch
    )

    if department_id:
        qs = qs.filter(department_id=department_id)

    qs = (qs.exclude(semester__isnull=True)
            .exclude(semester__exact="")
            .values_list("semester", flat=True)
            .distinct()
            .order_by("semester"))

    return JsonResponse(list(qs), safe=False)


def load_sem_adi_regulations(request):
    """
    Regulations available for (degree, optional department) based on timetable rows.
    return: [{id, text}]
    """
    degree_id = request.GET.get("degree_id")
    department_id = request.GET.get("department_id") or ""

    if not degree_id:
        return JsonResponse([], safe=False)

    filt = Q(degree_id=degree_id)
    if department_id:
        filt &= Q(department_id=department_id)

    regs = Regulations.objects.filter(
        semester_timetable_exam_schedules__in=SemesterExamScheduletimetable.objects.filter(filt)
    ).distinct().order_by("year")

    # If your Regulations display is like "R1 (2021)" you can format here
    data = [{"id": r.id, "text": str(r)} for r in regs]
    return JsonResponse(data, safe=False)


from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models.functions import TruncMonth

@require_http_methods(["GET"])
def load_sem_adi_monthyears(request):
    """
    Month-Year options from created_at for (degree, optional department, regulation).
    return: [{value:"2026-02", text:"Feb-2026"}, ...]
    """
    degree_id = request.GET.get("degree_id")
    department_id = request.GET.get("department_id") or ""
    regulation_id = request.GET.get("regulation_id")

    if not (degree_id and regulation_id):
        return JsonResponse([], safe=False)

    qs = SemesterExamScheduletimetable.objects.filter(
        degree_id=degree_id,
        regulation_id=regulation_id
    )

    if department_id:
        qs = qs.filter(department_id=department_id)

    months = (
        qs.annotate(m=TruncMonth("created_at"))
          .values_list("m", flat=True)
          .distinct()
          .order_by("-m")
    )

    data = []
    for m in months:
        if not m:
            continue
        data.append({
            "value": m.strftime("%Y-%m"),   # ✅ used in filter
            "text":  m.strftime("%b-%Y")    # ✅ shown to user
        })

    return JsonResponse(data, safe=False)



def load_sem_adi_timetable(request):
    degree_id = request.GET.get("degree_id")
    department_id = request.GET.get("department_id") or ""
    regulation_id = request.GET.get("regulation_id")
    monthyear = request.GET.get("monthyear")  # "YYYY-MM" e.g., "2026-02"

    if not (degree_id and regulation_id and monthyear):
        return JsonResponse([], safe=False)

    # Parse "YYYY-MM"
    try:
        year_str, month_str = monthyear.split("-")
        year, month = int(year_str), int(month_str)
        month_start = datetime(year, month, 1, 0, 0, 0)
    except Exception:
        return JsonResponse([], safe=False)

    # next month start
    if month == 12:
        month_end = datetime(year + 1, 1, 1, 0, 0, 0)
    else:
        month_end = datetime(year, month + 1, 1, 0, 0, 0)

    qs = SemesterExamScheduletimetable.objects.select_related("department", "course").filter(
        degree_id=degree_id,
        regulation_id=regulation_id,
        created_at__gte=month_start,
        created_at__lt=month_end,
    )

    if department_id:
        qs = qs.filter(department_id=department_id)

    qs = qs.order_by("semester", "exam_date", "department__Department", "course__course_code", "id")

    rows = []

    # ✅ show same course only one time (for same sem + date + session + dept)
    seen = set()

    for r in qs:
        dept_name = ""
        dept_code = ""
        if r.department:
            dept_name = getattr(r.department, "Department", "") or str(r.department)
            dept_code = getattr(r.department, "Department_code", "") or ""

        dept_key = (dept_code or dept_name or "—").strip()

        course_title = ""
        course_code = ""
        if r.course:
            course_title = getattr(r.course, "course_title", "") or getattr(r.course, "course_name", "") or str(r.course)
            course_code = getattr(r.course, "course_code", "") or ""

        # ✅ dedupe key
        dedupe_key = (
            (r.semester or "").strip(),
            r.exam_date,
            (r.session or "").strip(),
            dept_key,
            (course_code or "").strip(),
        )
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)

        rows.append({
            "exam_date": r.exam_date.strftime("%d-%m-%Y") if r.exam_date else "",
            "session": r.session or "",
            "semester": (r.semester or "").strip(),
            "department_key": dept_key,
            "department": dept_name,
            "department_code": dept_code,
            "course_title": course_title,
            "course_code": course_code,
        })

    return JsonResponse(rows, safe=False)


def published_semester_exam_schedule(request):
   return render(request, "examination_management/admin/published_semester_exam_schedule.html")






from django.contrib import messages
from django.db import transaction
from django.shortcuts import redirect
from django.utils import timezone

from examination_management.models import SemesterExamSchedule, SemesterExamScheduletimetable


def ses_publish(request):

    if request.method != "POST":
        return redirect("view_semester_exam_schedule")

    degree = request.POST.get("degree") or ""
    department = request.POST.get("department") or ""
    regulation = request.POST.get("regulation") or ""
    semester = request.POST.get("semester") or ""

    examsession = (request.POST.get("examsession") or "").strip()
    if not examsession:
        messages.error(request, "Please select Exam Session before publishing.")
        return redirect(request.POST.get("next") or "view_semester_exam_schedule")

    qs = SemesterExamSchedule.objects.select_related(
        "degree", "department", "regulation", "course"
    ).all()

    if degree:
        qs = qs.filter(degree_id=degree)
    if department:
        qs = qs.filter(department_id=department)
    if regulation:
        qs = qs.filter(regulation_id=regulation)
    if semester:
        qs = qs.filter(semester=str(semester))

    rows = list(qs)

    if not rows:
        messages.warning(request, "No schedules found to publish.")
        return redirect(request.POST.get("next") or "view_semester_exam_schedule")

    published_date = timezone.localdate()

    created_count = 0
    updated_count = 0

    for s in rows:

        obj, created = SemesterExamScheduletimetable.objects.update_or_create(
            degree=s.degree,
            department=s.department,
            regulation=s.regulation,
            course=s.course,
            semester=s.semester,
            batch=s.batch,
            defaults={
                "exam_date": s.exam_date,
                "session": s.session,
                "published_date": published_date,
                "examsession": examsession,

                # ✅ Copy failed flag
                "is_failed": bool(s.is_failed),
            },
        )

        if created:
            created_count += 1
        else:
            updated_count += 1

    messages.success(
        request,
        f"✅ Published successfully. Added: {created_count}, Updated: {updated_count}"
    )

    return redirect(request.POST.get("next") or "view_semester_exam_schedule")






from django.http import FileResponse
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from io import BytesIO
from datetime import datetime
from collections import defaultdict
from django.http import FileResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from io import BytesIO
from datetime import datetime
from collections import defaultdict
import os

from django.http import HttpResponse
from django.contrib.staticfiles import finders

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, PageBreak
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.utils import ImageReader

from django.http import HttpResponse
from django.contrib.staticfiles import finders

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, PageBreak, Paragraph
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.utils import ImageReader

import calendar


def admin_semester_timetable_pdf(request):

    degree_id = request.GET.get("degree_id")
    department_id = request.GET.get("department_id") or ""
    regulation_id = request.GET.get("regulation_id")
    monthyear = request.GET.get("monthyear")

    if not monthyear:
        return HttpResponse("Invalid Filters")

    year, month = map(int, monthyear.split("-"))

    month_start = datetime(year, month, 1)

    if month == 12:
        month_end = datetime(year + 1, 1, 1)
    else:
        month_end = datetime(year, month + 1, 1)





    # -------- Month Parsing --------
   
    # -------- Month Label --------
    

    # -------- Degree Fetch --------
    # -------- Degree Fetch --------
    degree_name = ""

    
    degree_obj = Degree.objects.filter(id=degree_id).first()
    if degree_obj:
        degree_name = getattr(degree_obj, "degree", "") or str(degree_obj)

    qs = SemesterExamScheduletimetable.objects.select_related(
        "department", "course"
    ).filter(
        degree_id=degree_id,
        regulation_id=regulation_id,
        exam_date__gte=month_start,
        exam_date__lt=month_end
    )
    examsession = qs.values_list("examsession", flat=True).first()






    if department_id:
        qs = qs.filter(department_id=department_id)

    qs = qs.order_by("department__Department", "semester", "exam_date", "session")

    # -------- Group Department --------
    dept_map = defaultdict(list)
    for r in qs:
        dept_name = getattr(r.department, "Department", "—")
        dept_map[dept_name].append(r)

    # -------- PDF Setup --------
    buf = BytesIO()

    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=38 * mm,
        bottomMargin=15 * mm,
    )

    story = []

    styles = getSampleStyleSheet()
    course_style = styles["Normal"]

    # -------- Table Build --------
    for idx, dept in enumerate(sorted(dept_map.keys())):

        if idx > 0:
            story.append(PageBreak())

        story.append(Spacer(1, 4))

        rows = dept_map[dept]

        # -------- Degree Title --------
        if degree_name:
            story.append(Table([[f"Degree : {degree_name}"]], colWidths=[170 * mm],
                style=[
                    ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
                    ("FONTSIZE", (0,0), (-1,-1), 11),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 4)
                ]
            ))

        # -------- Department Title --------
        story.append(Table([[f"Department : {dept}"]], colWidths=[170 * mm],
            style=[
                ("FONTNAME", (0,0), (-1,-1), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 11),
                ("BOTTOMPADDING", (0,0), (-1,-1), 6)
            ]
        ))

        # Header
        data = [["Semester", "Course", "Exam Date", "Session"]]

        rows.sort(key=lambda x: (
            int(x.semester or 0),
            x.exam_date,
            x.session
        ))

        for r in rows:
            course = ""
            if r.course:
                code = getattr(r.course, "course_code", "") or ""
                title = (
                    getattr(r.course, "title", "") or
                    getattr(r.course, "course_title", "") or ""
                )

                if code and title:
                    course = f"{code} - {title}"
                else:
                    course = code or title or str(r.course)

            data.append([
                r.semester,
                Paragraph(course, course_style),
                r.exam_date.strftime("%d-%m-%Y") if r.exam_date else "",
                r.session
            ])

        table = Table(data, colWidths=[30*mm, 70*mm, 40*mm, 30*mm], repeatRows=1)

        table.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.5, colors.black),
            ("BACKGROUND", (0,0), (-1,0), colors.whitesmoke),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
        ]))

        story.append(table)

    # -------- Header/Footer --------
    def draw_header_footer(canvas, doc):
        canvas.saveState()

        width, height = A4
        left = 15 * mm
        right = 15 * mm

        logo_path = finders.find("images/ritlogo.png")
        if logo_path and os.path.exists(logo_path):
            img = ImageReader(logo_path)
            canvas.drawImage(img, left, height - 28 * mm, width=25 * mm, height=18 * mm)

        canvas.setFont("Helvetica-Bold", 15)
        canvas.drawCentredString(width/2, height - 12*mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFont("Helvetica", 9)
        canvas.drawCentredString(width/2, height - 17*mm, "Rajapalayam - 626117")
        canvas.drawCentredString(width/2, height - 21*mm, "Affiliated to Anna University, Chennai")

        # ✅ Title with Month Year
        title_text = "Semester Examination Schedule"
        if examsession:
            title_text += f" – {examsession}"


        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(width/2, height - 27*mm, title_text)

        canvas.line(left, height - 32*mm, width-right, height - 32*mm)

        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(width-right, 10*mm, f"Page {canvas.getPageNumber()}")

        canvas.restoreState()

    doc.build(story, onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)

    pdf = buf.getvalue()
    buf.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="Semester_Timetable.pdf"'
    return response



def hall_degree(request):

    degrees = Degree.objects.filter(is_active=True).order_by("degree")

    return render(request, "filter_page.html", {
        "degrees": degrees
    })


from django.http import JsonResponse

def hall_departments(request):
    degree_id = request.GET.get("degree_id")

    departments = Add_Department.objects.filter(
        degree_id=degree_id,
        is_active=True
    ).values("id", "Department")

    return JsonResponse(list(departments), safe=False)

def hall_batches(request):
    department_id = request.GET.get("department_id")

    batches = StudentDetails.objects.filter(
        department_id=department_id,
        is_active=True
    ).values_list("batch", flat=True).distinct()

    return JsonResponse(list(batches), safe=False)



from urllib.parse import urlencode

from django.db.models import Count, Max, Q
from django.contrib import messages
from django.shortcuts import redirect, render
from django.urls import reverse
from django.views.decorators.http import require_http_methods

from examination_management.models import HallticketStudent, HallticketStudentCourse
@require_http_methods(["GET", "POST"])
@require_http_methods(["GET", "POST"])
def generate_hallticket(request):
    degrees = Degree.objects.filter(is_active=True)
    students = None

    # -----------------------------
    # Helper to load students (DISPLAY ONLY)
    # -----------------------------
    def fetch_students(degree_id, department_id, batch):
        if not (degree_id and department_id and batch):
            return None
        return (
            StudentDetails.objects.filter(department_id=department_id, batch=batch)
            .select_related("department", "department__degree")
            .order_by("reg_no")
        )

    # -----------------------------
    # GET: just show (no DB write)
    # -----------------------------
    if request.method == "GET":
        degree_id = (request.GET.get("degree") or "").strip()
        department_id = (request.GET.get("department") or "").strip()
        batch = (request.GET.get("batch") or "").strip()

        students = fetch_students(degree_id, department_id, batch)

        return render(request, "examination_management/admin/generate_hallticket.html", {
            "degrees": degrees,
            "students": students,
            "selected_degree": degree_id,
            "selected_department": department_id,
            "selected_batch": batch,
        })

    # -----------------------------
    # POST: SAVE ONLY when all selected
    # -----------------------------
    degree_id = (request.POST.get("degree") or "").strip()
    department_id = (request.POST.get("department") or "").strip()
    batch = (request.POST.get("batch") or "").strip()

    if not (degree_id and department_id and batch):
        messages.error(request, "Please select Degree, Department, and Batch, then click Generate.")
        return render(request, "examination_management/admin/generate_hallticket.html", {
            "degrees": degrees,
            "students": None,
            "selected_degree": degree_id,
            "selected_department": department_id,
            "selected_batch": batch,
        })

    students = fetch_students(degree_id, department_id, batch)

    # ✅ Store into HallticketStudent (ONLY ON POST)
    created = 0
    updated = 0

    ht_map = {}          # student_id -> HallticketStudent obj
    student_meta = {}    # student_id -> {"semester":..., "regulation_id":...}

    for s in students:
        degree_fk = s.department.degree if getattr(s, "department", None) else None
        dept_fk = getattr(s, "department", None)

        if not degree_fk and degree_id:
            degree_fk = Degree.objects.filter(id=degree_id).first()
        if not dept_fk and department_id:
            dept_fk = Add_Department.objects.filter(id=department_id).first()

        if not degree_fk or not dept_fk:
            continue

        sem_val = getattr(s, "semester", None)
        year_val = getattr(s, "year", None)

        obj, was_created = HallticketStudent.objects.update_or_create(
            degree=degree_fk,
            department=dept_fk,
            student=s,
            batch=batch,
            defaults={
                "semester": str(sem_val).strip() if sem_val is not None else None,
                "year": int(year_val) if year_val not in (None, "", "0") else None,
            }
        )
        created += 1 if was_created else 0
        updated += 0 if was_created else 1

        ht_map[s.id] = obj

        reg_id = getattr(s, "regulation_id", None) or ""
        student_meta[s.id] = {
            "semester": (str(sem_val).strip() if sem_val is not None else ""),
            "regulation_id": str(reg_id).strip() if reg_id else "",
        }

    # ==========================================================
    # SAVE COURSES (ONLY ON POST)
    # ==========================================================
    student_ids = list(ht_map.keys())

    groups = defaultdict(list)  # (reg_id, sem) -> [student_id,...]
    for sid, meta in student_meta.items():
        reg_id = meta.get("regulation_id", "")
        sem = meta.get("semester", "")
        if sem:
            groups[(reg_id, sem)].append(sid)

    group_cycle = {}
    group_tt_rows = {}

    for (reg_id, sem), sids in groups.items():
        base_tt = SemesterExamScheduletimetable.objects.select_related("course").filter(
            degree_id=degree_id,
            department_id=department_id,
            batch=batch,
            semester=str(sem).strip(),
            is_failed=False,
        )
        if reg_id:
            base_tt = base_tt.filter(regulation_id=reg_id)

        es = (
            base_tt.values("examsession")
            .exclude(examsession__isnull=True)
            .exclude(examsession__exact="")
            .annotate(cnt=Count("id"))
            .order_by("-cnt")
            .first()
        )
        active_examsession = (es["examsession"] if es else "") or ""

        pub_qs = base_tt.filter(examsession=active_examsession) if active_examsession else base_tt
        active_published_date = pub_qs.aggregate(mx=Max("published_date"))["mx"]

        group_cycle[(reg_id, sem)] = {
            "examsession": active_examsession,
            "published_date": active_published_date,
        }

        save_tt = base_tt
        if active_examsession:
            save_tt = save_tt.filter(examsession=active_examsession)
        if active_published_date:
            save_tt = save_tt.filter(published_date=active_published_date)

        group_tt_rows[(reg_id, sem)] = list(
            save_tt.order_by("exam_date", "session", "course__course_code")
        )

    regular_saved = 0
    for (reg_id, sem), sids in groups.items():
        rows = group_tt_rows.get((reg_id, sem), [])
        for sid in sids:
            ht = ht_map.get(sid)
            if not ht:
                continue
            for row in rows:
                HallticketStudentCourse.objects.get_or_create(
                    hallticket_student=ht,
                    exam_timetable=row,
                    defaults={"semester": str(row.semester or "").strip()}
                )
                regular_saved += 1

    failed_saved = 0
    failed_qs = (
        Result.objects
        .select_related("course", "regulation")
        .filter(
            student_id__in=student_ids,
            degree_id=degree_id,
            department_id=department_id,
            batch=batch,
        )
        .filter(Q(grade__iexact="U") | Q(grade__iexact="RA"))
        .order_by("student_id", "semester", "course__course_code")
    )

    failed_tt_cache = {}

    for r in failed_qs:
        if not r.course_id:
            continue

        sid = r.student_id
        ht = ht_map.get(sid)
        if not ht:
            continue

        meta = student_meta.get(sid, {})
        cur_sem = meta.get("semester", "")
        cur_reg_id = meta.get("regulation_id", "")

        cyc = group_cycle.get((cur_reg_id, cur_sem), {})
        active_examsession = (cyc.get("examsession") or "").strip()
        active_published_date = cyc.get("published_date")

        used_reg_id = (str(getattr(r, "regulation_id", "") or "").strip()) or cur_reg_id

        cache_key = (
            r.course_id,
            str(r.semester or "").strip(),
            (str(r.batch or "").strip()),
            used_reg_id,
            active_examsession,
            active_published_date,
        )

        if cache_key in failed_tt_cache:
            failed_tt = failed_tt_cache[cache_key]
        else:
            q = SemesterExamScheduletimetable.objects.filter(
                course_id=r.course_id,
                degree_id=degree_id,
                department_id=department_id,
                batch=str(r.batch or "").strip(),
                semester=str(r.semester or "").strip(),
                is_failed=True,
            )
            if used_reg_id:
                q = q.filter(regulation_id=used_reg_id)
            if active_examsession:
                q = q.filter(examsession=active_examsession)
            if active_published_date:
                q = q.filter(published_date=active_published_date)

            failed_tt = q.order_by("-published_date", "exam_date", "session").first()
            failed_tt_cache[cache_key] = failed_tt

        if not failed_tt:
            continue

        HallticketStudentCourse.objects.get_or_create(
            hallticket_student=ht,
            exam_timetable=failed_tt,
            defaults={"semester": str(r.semester or "").strip()}
        )
        failed_saved += 1

    messages.success(
        request,
        f"Hallticket students saved: {created} created, {updated} updated. "
        f"Courses saved: Regular={regular_saved}, Failed={failed_saved}."
    )

    # ✅ IMPORTANT: Redirect to GET so refresh won't re-save
    params = urlencode({"degree": degree_id, "department": department_id, "batch": batch})
    return redirect(f"{reverse('generate_hallticket')}?{params}")





from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.db.models import Q, Max, Count
@require_GET
@require_GET

@require_GET
def hallticket_courses(request):

    degree_id = request.GET.get("degree_id")
    department_id = request.GET.get("department_id")
    batch = (request.GET.get("batch") or "").strip()
    semester = (request.GET.get("semester") or "").strip()
    regulation_id = (request.GET.get("regulation_id") or "").strip()
    examsession = (request.GET.get("examsession") or "").strip()
    student_id = (request.GET.get("student_id") or "").strip()



    if not (degree_id and department_id and batch and semester):
        return JsonResponse({"ok": False, "items": [], "error": "Missing filters"}, status=400)

    # ✅ get correct HallticketStudent (this is the KEY FIX)
    ht_student = None
    if student_id:
        ht_student = HallticketStudent.objects.filter(
            student_id=student_id,
            degree_id=degree_id,
            department_id=department_id,
            batch=batch
        ).first()

    # ---------------- 1) TIMETABLE COURSES ----------------
    tt_qs = (
        SemesterExamScheduletimetable.objects
        .select_related("course")
        .filter(
            degree_id=degree_id,
            department_id=department_id,
            batch=batch,
            semester=semester,
            is_failed=False,
        )
    )

    if regulation_id:
        tt_qs = tt_qs.filter(regulation_id=regulation_id)

    if examsession:
        tt_qs = tt_qs.filter(examsession=examsession)

    tt_qs = tt_qs.order_by("exam_date", "session", "course__course_code")

    items = []
    for row in tt_qs:
        c = row.course

        

        items.append({
            "source": "TIMETABLE",
            "semester": str(row.semester or ""),
            "course_id": c.id if c else None,
            "course_code": (c.course_code or "") if c else "",
            "course_name": (c.title or "") if c else "",
            "exam_date": row.exam_date.strftime("%d-%m-%Y") if row.exam_date else "",
            "session": row.session or "",
            "examsession": row.examsession or "",
        })

        # ✅ STORE corresponding student-course mapping
       

    # ---------------- Detect ACTIVE cycle ----------------
    active_examsession = examsession
    if not active_examsession:
        es = (
            tt_qs.values("examsession")
            .exclude(examsession__isnull=True)
            .exclude(examsession__exact="")
            .annotate(cnt=Count("id"))
            .order_by("-cnt")
            .first()
        )
        active_examsession = (es["examsession"] if es else "") or ""


    # ---------------- Detect ACTIVE published date ----------------
    base_pub_qs = SemesterExamScheduletimetable.objects.filter(
        degree_id=degree_id,
        department_id=department_id,
        batch=batch,
        semester=semester,
        is_failed=False,
    )

    if regulation_id:
        base_pub_qs = base_pub_qs.filter(regulation_id=regulation_id)

    if active_examsession:
        base_pub_qs = base_pub_qs.filter(examsession=active_examsession)

    active_published_date = base_pub_qs.aggregate(mx=Max("published_date"))["mx"]


    # ---------------- 2) FAILED COURSES ----------------
    if student_id:

        failed_qs = (
            Result.objects
            .select_related("course", "regulation")
            .filter(
                student_id=student_id,
                degree_id=degree_id,
                department_id=department_id,
                batch=batch,
            )
            .filter(Q(grade__iexact="U") | Q(grade__iexact="RA"))
        )

        if regulation_id:
            failed_qs = failed_qs.filter(regulation_id=regulation_id)

        failed_qs = failed_qs.order_by("semester", "course__course_code")

        for r in failed_qs:
            c = r.course
            if not c:
                continue

            used_regulation_id = regulation_id or (r.regulation_id if r.regulation else None)


            failed_tt = SemesterExamScheduletimetable.objects.filter(
                course_id=c.id,
                department_id=department_id,
                semester=str(r.semester).strip(),
                is_failed=True,
            )

            if degree_id:
                failed_tt = failed_tt.filter(degree_id=degree_id)

            # ✅ use batch from Result (important for arrears)
            result_batch = (r.batch or "").strip()
            if result_batch:
                failed_tt = failed_tt.filter(batch=result_batch)

            if used_regulation_id:
                failed_tt = failed_tt.filter(regulation_id=used_regulation_id)

            if active_examsession:
                failed_tt = failed_tt.filter(examsession=active_examsession)

            if active_published_date:
                failed_tt = failed_tt.filter(published_date=active_published_date)

            failed_tt = failed_tt.order_by("-published_date", "exam_date", "session").first()

            items.append({
                "source": "FAILED",
                "semester": str(r.semester or ""),
                "course_id": c.id,
                "course_code": c.course_code or "",
                "course_name": c.title or "",
                "exam_date": failed_tt.exam_date.strftime("%d-%m-%Y") if failed_tt and failed_tt.exam_date else "",
                "session": failed_tt.session if failed_tt else "",
                "examsession": failed_tt.examsession if failed_tt else (active_examsession or ""),
            })

            # ✅ STORE corresponding student-course mapping (FAILED)
            if failed_tt:
                # ✅ ensure hallticket student exists for the SAME batch as result
                ht_student_failed = HallticketStudent.objects.filter(
                    student_id=student_id,
                    degree_id=degree_id,
                    department_id=department_id,
                    batch=result_batch or batch
                ).first()

             

    # ---------------- De-duplicate ----------------
    seen = set()
    unique_items = []
    for it in items:
        key = (
            it.get("course_id"),
            it.get("semester"),
            it.get("exam_date"),
            it.get("session"),
            it.get("examsession"),
            it.get("source"),
        )
        if key in seen:
            continue
        seen.add(key)
        unique_items.append(it)

    return JsonResponse({"ok": True, "items": unique_items})




@require_GET
def view_generated_halltickets(request):
    # dropdown: only degrees that exist in generated halltickets
    degrees = Degree.objects.filter(
        id__in=HallticketStudent.objects.values_list("degree_id", flat=True).distinct()
    )

    degree_id = (request.GET.get("degree") or "").strip()
    department_id = (request.GET.get("department") or "").strip()
    batch = (request.GET.get("batch") or "").strip()
    semester = (request.GET.get("semester") or "").strip()

    students = None
    if degree_id and department_id and batch and semester:
        students = (
            HallticketStudent.objects
            .select_related("student", "department", "degree")
            .filter(
                degree_id=degree_id,
                department_id=department_id,
                batch=batch,
                semester=str(semester).strip(),
            )
            .order_by("student__reg_no")
        )

    return render(
        request,
        "examination_management/admin/view_generated_halltickets.html",
        {
            "degrees": degrees,
            "students": students,
            "selected_degree": degree_id,
            "selected_department": department_id,
            "selected_batch": batch,
            "selected_semester": semester,
        }
    )



@require_GET
def hallticket_saved_courses(request):
    hallticket_student_id = (request.GET.get("hallticket_student_id") or "").strip()

    if not hallticket_student_id:
        return JsonResponse({"ok": False, "items": [], "error": "Missing hallticket_student_id"}, status=400)

    ht = (
        HallticketStudent.objects
        .select_related("student", "degree", "department")
        .filter(id=hallticket_student_id)
        .first()
    )
    if not ht:
        return JsonResponse({"ok": False, "items": [], "error": "Invalid hallticket_student_id"}, status=404)

    qs = (
        HallticketStudentCourse.objects
        .select_related("exam_timetable", "exam_timetable__course")
        .filter(hallticket_student_id=hallticket_student_id)
        .order_by("semester", "exam_timetable__exam_date", "exam_timetable__session", "exam_timetable__course__course_code")
    )

    items = []
    for obj in qs:
        tt = obj.exam_timetable
        c = getattr(tt, "course", None)

        items.append({
            "source": "SAVED",
            "semester": str(obj.semester or ""),
            "course_id": c.id if c else None,
            "course_code": (c.course_code or "") if c else "",
            "course_name": (c.title or "") if c else "",
            "exam_date": tt.exam_date.strftime("%d-%m-%Y") if tt and tt.exam_date else "",
            "session": tt.session or "",
            "examsession": tt.examsession or "",
        })

    return JsonResponse({"ok": True, "items": items})










from django.http import JsonResponse

def hall_generated_departments(request):

    degree_id = request.GET.get("degree_id")

    departments = Add_Department.objects.filter(
        id__in=HallticketStudent.objects.filter(
            degree_id=degree_id
        ).values_list("department_id", flat=True).distinct()
    )

    data = list(departments.values("id", "Department"))

    return JsonResponse(data, safe=False)


def hall_generated_batches(request):

    degree_id = request.GET.get("degree_id")
    department_id = request.GET.get("department_id")

    batches = (
        HallticketStudent.objects
        .filter(degree_id=degree_id, department_id=department_id)
        .values_list("batch", flat=True)
        .distinct()
        .order_by("batch")
    )

    return JsonResponse(list(batches), safe=False)
def hall_generated_semesters(request):

    degree_id = request.GET.get("degree_id")
    department_id = request.GET.get("department_id")
    batch = request.GET.get("batch")

    semesters = (
        HallticketStudent.objects
        .filter(
            degree_id=degree_id,
            department_id=department_id,
            batch=batch
        )
        .values_list("semester", flat=True)
        .distinct()
        .order_by("semester")
    )

    return JsonResponse(list(semesters), safe=False)






import io
from datetime import datetime

from django.http import FileResponse, HttpResponseBadRequest
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


import io
import os
from datetime import datetime

from django.conf import settings
from django.http import HttpResponseBadRequest, FileResponse
from django.views.decorators.http import require_POST
from django.contrib.staticfiles import finders

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth


def _safe(s):
    return "" if s is None else str(s)


def _wrap_text(text, font_name, font_size, max_width):
    """
    Simple word wrap for canvas.drawString.
    Returns list of lines.
    """
    text = _safe(text).strip()
    if not text:
        return [""]

    words = text.split()
    lines = []
    cur = ""
    for w in words:
        t = (cur + " " + w).strip()
        if stringWidth(t, font_name, font_size) <= max_width:
            cur = t
        else:
            if cur:
                lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return lines


def _draw_cell(c, x, y, w, h, text="", font="Helvetica", size=9, pad=2, align="L"):
    """
    Draw single cell border and text.
    (x,y) is bottom-left.
    """
    c.rect(x, y, w, h, stroke=1, fill=0)
    c.setFont(font, size)

    tx = x + pad
    if align == "C":
        tw = stringWidth(_safe(text), font, size)
        tx = x + (w - tw) / 2
    elif align == "R":
        tw = stringWidth(_safe(text), font, size)
        tx = x + w - tw - pad

    ty = y + (h - size) / 2 - 1
    c.drawString(tx, ty, _safe(text)[:200])


def resolve_logo_path():
    """Find the institution logo (finders + STATIC_ROOT + STATICFILES_DIRS)."""
    logo_paths = [
        "images/ritlogo.png",
        "static/images/ritlogo.png",
        "media/images/ritlogo.png",
        "img/ritlogo.png",
        "img/anna_logo.png",
    ]

    for logo_rel in logo_paths:
        path = finders.find(logo_rel)
        if path and os.path.exists(path):
            return path

        static_root = getattr(settings, "STATIC_ROOT", "")
        if static_root:
            cand = os.path.join(static_root, logo_rel)
            if os.path.exists(cand):
                return cand

        for static_dir in getattr(settings, "STATICFILES_DIRS", []):
            cand = os.path.join(static_dir, logo_rel)
            if os.path.exists(cand):
                return cand

    return None


def draw_hallticket_page(c, ht, logo_path=None):
    """
    Draw one hallticket page (A4) like the sample image.
    """
    W, H = A4
    margin = 10 * mm
    x0, y0 = margin, margin
    w0, h0 = W - 2 * margin, H - 2 * margin

    # outer border
    c.setLineWidth(1)
    c.rect(x0, y0, w0, h0, stroke=1, fill=0)

    # -----------------------------------
    # HEADER BLOCK
    # -----------------------------------
    header_h = 38 * mm
    header_y = H - margin - header_h

    # draw header border WITHOUT bottom line
    c.line(x0, header_y + header_h, x0 + w0, header_y + header_h)  # top
    c.line(x0, header_y, x0, header_y + header_h)                  # left
    c.line(x0 + w0, header_y, x0 + w0, header_y + header_h)        # right

    # logo box (left)
    logo_box_w = 26 * mm

    # draw logo box WITHOUT right line
    c.line(x0, header_y, x0, header_y + header_h)                          # left
    c.line(x0, header_y + header_h, x0 + logo_box_w, header_y + header_h)  # top
    c.line(x0, header_y, x0 + logo_box_w, header_y)                        # bottom

    # LOGO DRAWING
    logo_path = resolve_logo_path()
    if logo_path:
        try:
            img = ImageReader(logo_path)
            iw, ih = img.getSize()

            max_h = header_h - 6 * mm
            max_w = logo_box_w - 6 * mm
            scale = min(max_h / ih, max_w / iw)

            sw = iw * scale
            sh = ih * scale

            lx = x0 + (logo_box_w - sw) / 2
            ly = header_y + (header_h - sh) / 2

            c.drawImage(
                logo_path,
                lx,
                ly,
                width=sw,
                height=sh,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception:
            pass

    # photo box (right)
    photo_box_w = 32 * mm
    c.rect(x0 + w0 - photo_box_w, header_y, photo_box_w, header_h, stroke=1, fill=0)

    c.setFont("Helvetica", 7)
    c.drawString(x0 + w0 - photo_box_w + 4 * mm, header_y + header_h - 5 * mm, "Photo of the candidate")
   

    # candidate photo
        # candidate photo (from StudentDetails.profile_img)
    student = ht.student

    # ✅ your model field is profile_img
    photo_field = getattr(student, "profile_img", None)

    if photo_field and getattr(photo_field, "name", ""):
        try:
            # Works when MEDIA is local filesystem (ImageField)
            photo_path = photo_field.path

            if os.path.exists(photo_path):
                imgp = ImageReader(photo_path)

                img_x = x0 + w0 - photo_box_w + 3 * mm
                img_y = header_y + 3 * mm
                img_w = photo_box_w - 6 * mm
                img_h = header_h - 10 * mm


                c.drawImage(
                    imgp,
                    img_x,
                    img_y,
                    width=img_w,
                    height=img_h,
                    preserveAspectRatio=True,
                    anchor='c',      # ✅ centers inside box
                    mask="auto",
                )
        except Exception:
            pass


    # center header text
    center_x = x0 + logo_box_w
    center_w = w0 - logo_box_w - photo_box_w

    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(center_x + center_w / 2, header_y + header_h - 12 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")
    c.setFont("Helvetica", 8)
    c.drawCentredString(center_x + center_w / 2, header_y + header_h - 18 * mm, "Rajapalayam - 626117")
    c.setFont("Helvetica", 8)
    c.drawCentredString(center_x + center_w / 2, header_y + header_h - 24 * mm, "UNIVERSITY EXAMINATIONS")
    c.setFont("Helvetica", 8)
    c.drawCentredString(center_x + center_w / 2, header_y + header_h - 30 * mm, "HALL TICKET")

    # -----------------------------------
    # META TABLE
    # -----------------------------------
    meta_top = header_y
    meta_h = 28 * mm
    meta_y = meta_top - meta_h

    c.rect(x0, meta_y, w0, meta_h, stroke=1, fill=0)

    row_h = meta_h / 4.0

    reg_no = _safe(getattr(student, "reg_no", ""))
    name = _safe(getattr(student, "name", getattr(student, "student_name", "")))
    dob = getattr(student, "date_of_birth", None)
    dob_txt = dob.strftime("%d-%b-%Y") if dob else ""

    degree_txt = _safe(getattr(ht.degree, "degree", ""))
    dept_txt = _safe(getattr(ht.department, "Department", getattr(ht.department, "name", "")))

    # 1st row
    y = meta_y + meta_h - row_h
    _draw_cell(c, x0, y, 42 * mm, row_h, "Register Number", font="Helvetica-Bold", size=8)
    _draw_cell(c, x0 + 42 * mm, y, 55 * mm, row_h, reg_no, size=8)
    _draw_cell(c, x0 + 97 * mm, y, 50 * mm, row_h, "Current Semester", font="Helvetica-Bold", size=8)
    _draw_cell(c, x0 + 147 * mm, y, w0 - 147 * mm, row_h, _safe(ht.semester), size=8, align="C")

    # 2nd row
    y = meta_y + meta_h - 2 * row_h
    _draw_cell(c, x0, y, 42 * mm, row_h, "Name", font="Helvetica-Bold", size=8)
    _draw_cell(c, x0 + 42 * mm, y, 105 * mm, row_h, name, size=8)
    _draw_cell(c, x0 + 147 * mm, y, 18 * mm, row_h, "D.O.B", font="Helvetica-Bold", size=8, align="C")
    _draw_cell(c, x0 + 165 * mm, y, w0 - 165 * mm, row_h, dob_txt, size=8, align="C")

    # 3rd row
    y = meta_y + meta_h - 3 * row_h
    _draw_cell(c, x0, y, 42 * mm, row_h, "Degree & Branch", font="Helvetica-Bold", size=8)
    _draw_cell(c, x0 + 42 * mm, y, w0 - 42 * mm, row_h, f"{degree_txt} - {dept_txt}", size=8)

    # 4th row
    y = meta_y
    _draw_cell(c, x0, y, 42 * mm, row_h, "Examination\nCentre", font="Helvetica-Bold", size=8)
    centre = _safe(getattr(student, "college_name", "")) or _safe(getattr(student, "institution", ""))
    _draw_cell(c, x0 + 42 * mm, y, w0 - 42 * mm, row_h, centre, size=8)

    # -----------------------------------
    # SUBJECTS BLOCK (2 columns)
    # -----------------------------------
    notes_h = 28 * mm
    notes_y = y0 + 25 * mm
    notes_top = notes_y + notes_h

    subjects_top = meta_y
    subjects_bottom = notes_top
    subjects_h = subjects_top - subjects_bottom

    c.rect(x0, subjects_bottom, w0, subjects_h, stroke=1, fill=0)

    # vertical center line
    mid_x = x0 + w0 / 2
    c.line(mid_x, subjects_bottom, mid_x, subjects_bottom + subjects_h)

    # headers
    hdr_h = 10 * mm
    c.setFont("Helvetica-Bold", 8)

    c.rect(x0, subjects_bottom + subjects_h - hdr_h, w0 / 2, hdr_h, stroke=1, fill=0)
    c.drawString(x0 + 3 * mm, subjects_bottom + subjects_h - 7 * mm, "Sem")
    c.drawString(x0 + 15 * mm, subjects_bottom + subjects_h - 7 * mm, "coursecode")
    c.drawString(x0 + 35 * mm, subjects_bottom + subjects_h - 7 * mm, "courseTitle")

    c.rect(mid_x, subjects_bottom + subjects_h - hdr_h, w0 / 2, hdr_h, stroke=1, fill=0)
    c.drawString(mid_x + 3 * mm, subjects_bottom + subjects_h - 7 * mm, "Sem")
    c.drawString(mid_x + 15 * mm, subjects_bottom + subjects_h - 7 * mm, "coursecode")
    c.drawString(mid_x + 35 * mm, subjects_bottom + subjects_h - 7 * mm, "courseTitle")

    # fetch courses
    course_qs = (
        HallticketStudentCourse.objects
        .select_related("exam_timetable", "exam_timetable__course")
        .filter(hallticket_student_id=ht.id)
        .order_by("semester", "exam_timetable__exam_date", "exam_timetable__session")
    )

    items = []
    for sc in course_qs:
        tt = sc.exam_timetable
        course = getattr(tt, "course", None)

        code = _safe(getattr(course, "course_code", "") or getattr(course, "code", ""))
        title = _safe(getattr(course, "title", "") or getattr(course, "name", ""))
        sem = _safe(sc.semester or ht.semester)

        # show exam_date & session like: Tamil [01-02-2006 & FN]
        ex_date = getattr(tt, "exam_date", None)
        ex_date_txt = ex_date.strftime("%d-%m-%Y") if ex_date else ""
        sess = _safe(getattr(tt, "session", "")).upper()
        if ex_date_txt or sess:
            title = f"{title} [{ex_date_txt} & {sess}]"

        items.append((sem, code, title))

    # row drawing area
    body_top = subjects_bottom + subjects_h - hdr_h
    row_h = 7 * mm
    max_rows = int((subjects_h - hdr_h) // row_h)

    # fill left then overflow to right
    left_items = items[:max_rows]
    right_items = items[max_rows:max_rows * 2]

    def draw_subject_rows(col_x, col_w, rows):
        sem_w = 12 * mm
        code_w = 20 * mm
        title_w = col_w - sem_w - code_w - 2 * mm

        y = body_top - row_h
        c.setFont("Helvetica", 8)

        for sem, code, title in rows:
            c.drawString(col_x + 3 * mm, y + 2.2 * mm, sem[:4])
            c.drawString(col_x + 3 * mm + sem_w, y + 2.2 * mm, code[:15])

            # ✅ FIX: draw MULTI-LINE title (wrap to next line)
            lines = _wrap_text(title, "Helvetica", 8, title_w)
            # We have row_h=7mm, font=8 => safely 2 lines fits (approx).
            lines = lines[:2]

            if lines:
                first_line_y = y + 3.0 * mm
                line_gap = 3 * mm   # 🔥 increase/decrease this for spacing control

                c.drawString(
                    col_x + 3 * mm + sem_w + code_w,
                    first_line_y,
                    lines[0][:90]
                )

            if len(lines) > 1:
                c.drawString(
                    col_x + 3 * mm + sem_w + code_w,
                    first_line_y - line_gap,
                    lines[1][:90]
                )


            y -= row_h

    draw_subject_rows(x0, w0 / 2, left_items)
    draw_subject_rows(mid_x, w0 / 2, right_items)

    c.setFont("Helvetica-Bold", 8)
    c.drawString(x0 + 3 * mm, subjects_bottom + 6 * mm, f"No of Subjects Registered: {len(items)}")

    # -----------------------------------
    # NOTES BLOCK
    # -----------------------------------
    c.rect(x0, notes_y, w0, notes_h, stroke=1, fill=0)

    c.setFont("Helvetica-Bold", 8)
    c.drawString(x0 + 3 * mm, notes_y + notes_h - 6 * mm, "NOTE:")

    c.setFont("Helvetica", 7)
    note_lines = [
        "1. This hall ticket is valid only if the candidate's admission is approved by the University.",
        "2. Correction in Name / Date of Birth / Photograph (if any) to be updated in the web portal on correction window.",
        "3. Instructions printed overleaf are to be followed strictly.",
    ]
    yy = notes_y + notes_h - 12 * mm
    for line in note_lines:
        wrapped = _wrap_text(line, "Helvetica", 7, w0 - 10 * mm)
        for wl in wrapped:
            c.drawString(x0 + 5 * mm, yy, wl)
            yy -= 4 * mm

    c.setFont("Helvetica", 7)
    c.drawString(x0 + 3 * mm, notes_y + 3 * mm, f"Generated on: {datetime.now().strftime('%d-%b-%Y')}")

    # -----------------------------------
    # SIGNATURE BLOCK
    # -----------------------------------
    sig_h = 25 * mm
    sig_y = y0
    c.rect(x0, sig_y, w0, sig_h, stroke=1, fill=0)

    col = w0 / 3
    c.line(x0 + col, sig_y, x0 + col, sig_y + sig_h)
    c.line(x0 + 2 * col, sig_y, x0 + 2 * col, sig_y + sig_h)

    c.setFont("Helvetica", 7)
    c.drawCentredString(x0 + col / 2, sig_y + 3 * mm, "Signature of the Candidate")
    c.drawCentredString(x0 + col + col / 2, sig_y + 3 * mm, "Signature of the Principal with seal")
    c.drawCentredString(x0 + 2 * col + col / 2, sig_y + 3 * mm, "Controller of Examinations")


@require_POST
def hallticket_bulk_pdf(request):
    """
    Download combined Hallticket PDF for selected HallticketStudent IDs.
    POST: ids="1,2,3"
    """
    ids_raw = (request.POST.get("ids") or "").strip()
    if not ids_raw:
        return HttpResponseBadRequest("No students selected.")

    ids = []
    for x in ids_raw.split(","):
        x = x.strip()
        if x.isdigit():
            ids.append(int(x))
    ids = list(dict.fromkeys(ids))
    if not ids:
        return HttpResponseBadRequest("Invalid selection.")

    ht_qs = (
        HallticketStudent.objects
        .select_related("degree", "department", "student")
        .filter(id__in=ids)
    )
    ht_map = {h.id: h for h in ht_qs}
    halltickets = [ht_map[i] for i in ids if i in ht_map]
    if not halltickets:
        return HttpResponseBadRequest("No hallticket records found for selected IDs.")

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)

    logo_path = os.path.join(settings.BASE_DIR, "static", "img", "anna_logo.png")

    for ht in halltickets:
        draw_hallticket_page(c, ht, logo_path=logo_path)
        c.showPage()

    c.save()
    buf.seek(0)

    filename = f"halltickets_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    resp = FileResponse(buf, content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    return resp













from django.http import JsonResponse

def ajax_load_prac_departments(request):
    degree_id = request.GET.get("degree_id")

    if not degree_id:
        return JsonResponse({"departments": []})

    qs = (
        Add_Department.objects
        .filter(degree_id=degree_id, is_active=True)
        .order_by("Department")   # only Department
        .values("id", "Department")
    )

    departments = [
        {"id": d["id"], "name": d["Department"] or "Unnamed"}
        for d in qs
    ]

    return JsonResponse({"departments": departments})

def ajax_load_prac_batches(request):
    department_id = request.GET.get("department_id")
    if not department_id:
        return JsonResponse({"batches": []})

    batches_qs = (
        StudentDetails.objects
        .filter(department_id=department_id)
        .exclude(batch__isnull=True)
        .exclude(batch__exact="")
        .values_list("batch", flat=True)
        .distinct()
        .order_by("batch")
    )
    return JsonResponse({"batches": list(batches_qs)})


def ajax_load_prac_courses(request):
    department_id = request.GET.get("department_id")
    batch = request.GET.get("batch")

    if not department_id or not batch:
        return JsonResponse({"courses": []})

    sem = (
        StudentDetails.objects
        .filter(department_id=department_id, batch=batch)
        .aggregate(m=Max("semester"))
        .get("m")
    )

    ce_qs = (
        CourseEnrollment.objects
        .filter(department_id=department_id, batch=batch, enroll=True)
        .exclude(course__isnull=True)
        .select_related("course")
    )

    if sem is not None:
        ce_qs = ce_qs.filter(course__semester=sem)

    courses_qs = (
        ce_qs
        .values("course_id", "course__course_code", "course__title")
        .distinct()
        .order_by("course__course_code")
    )

    course_ids = list(courses_qs.values_list("course_id", flat=True))
    config_map = dict(
        CourseHours.objects
        .filter(course_id__in=course_ids)
        .exclude(hour_config_id__isnull=True)
        .values_list("course_id", "hour_config_id")
    )


    config_ids = sorted({cid for cid in config_map.values() if cid})
    config_hours_map = {
        row["id"]: {
            "lecture_hours": row["lecture_hours"],
            "tutorial_hours": row["tutorial_hours"],
            "laboratory_hours": row["laboratory_hours"],
        }
        for row in CourseHourConfig.objects
            .filter(id__in=config_ids)
            .values("id", "lecture_hours", "tutorial_hours", "laboratory_hours")
    }


    courses = []
    for c in courses_qs:
        course_id = c["course_id"]
        hour_config_id = config_map.get(course_id)

        ltp = config_hours_map.get(hour_config_id, {
            "lecture_hours": 0,
            "tutorial_hours": 0,
            "laboratory_hours": 0,
        })

        p_hours = int(ltp.get("laboratory_hours") or 0)

        

        # ✅ FILTER: show ONLY courses where P != 0
        if p_hours == 0:
            continue

        code = (c.get("course__course_code") or "").strip()
        title = (c.get("course__title") or "").strip()
        display_text = f"{code} - {title}" if code and title else code or title or "Unnamed Course"

        courses.append({
            "id": course_id,
            "text": display_text,
            "hour_config_id": hour_config_id,
            "lecture_hours": ltp["lecture_hours"],
            "tutorial_hours": ltp["tutorial_hours"],
            "laboratory_hours": p_hours,
        })


    return JsonResponse({
        "semester": sem,
        "courses": courses
    })



def ajax_load_prac_students(request):
    department_id = request.GET.get("department_id")
    batch = request.GET.get("batch")

    if not department_id or not batch:
        return JsonResponse({"students": []})

    sem = (
        StudentDetails.objects
        .filter(department_id=department_id, batch=batch)
        .aggregate(m=Max("semester"))
        .get("m")
    )

    students_qs = (
        StudentDetails.objects
        .filter(department_id=department_id, batch=batch)
        .exclude(reg_no__isnull=True)
        .exclude(reg_no__exact="")
        .values("id", "reg_no")
        .order_by("reg_no")
    )

    return JsonResponse({
        "semester": sem,
        "students": list(students_qs)   # ✅ now list of {id, reg_no}
    })


def ajax_load_halls(request):
    qs = Hall.objects.filter(is_active=True).order_by("hall_name")
    halls = [
        {
            "id": h.id,
            "hall_name": h.hall_name,
            "benches": h.benches,
        }
        for h in qs
    ]
    return JsonResponse({"halls": halls})

def practicalexamschedule(request):
    degrees = Degree.objects.filter(is_active=True).order_by("degree")
    return render(request, "examination_management/admin/practicalexamschedule.html",{"degrees": degrees})

import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from examination_management.models import PracticalExamStudent, PracticalExamStudentSchedule
@require_POST
def ajax_save_prac_schedule(request):
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "message": "Invalid JSON"}, status=400)

    degree_id = payload.get("degree_id")
    department_id = payload.get("department_id")
    batch = payload.get("batch")
    course_id = payload.get("course_id")
    semester = payload.get("semester")
    hall_id = payload.get("hall_id")
    assignments = payload.get("assignments", [])  # list of per-student schedule rows

    if not (degree_id and department_id and batch and course_id):
        return JsonResponse({"ok": False, "message": "Missing required fields"}, status=400)

    if not assignments:
        return JsonResponse({"ok": False, "message": "No assignments to save"}, status=400)

    # Resolve FK objects
    degree = Degree.objects.get(id=degree_id)
    department = Add_Department.objects.get(id=department_id)
    course = Course.objects.get(id=course_id)

    hall_obj = None
    if hall_id:
        hall_obj = Hall.objects.filter(id=hall_id).first()

    # Optional: prevent duplicates by deleting old saved plan for same context
    # (choose your rule: course+batch+department)
    with transaction.atomic():
        # delete previous saved schedule for the same course+batch+dept
        old_rows = PracticalExamStudent.objects.filter(
            department_id=department_id,
            batch=batch,
            course_id=course_id,
        )
        PracticalExamStudentSchedule.objects.filter(prac_student__in=old_rows).delete()
        old_rows.delete()

        # Build PracticalExamStudent rows
        # assignments item must contain: student_id, batch_no, exam_date, session, exam_time
        student_ids = [a.get("student_id") for a in assignments if a.get("student_id")]
        students = {s.id: s for s in StudentDetails.objects.filter(id__in=student_ids,
        is_active=True)}

        prac_rows = []
        for sid in student_ids:
            s = students.get(sid)
            if not s:
                continue
            prac_rows.append(
                PracticalExamStudent(
                    student=s,
                    degree=degree,
                    department=department,
                    batch=batch,
                    course=course,
                    semester=semester or None,
                )
            )

        # bulk create model 1
        created_prac = PracticalExamStudent.objects.bulk_create(prac_rows)

        # map student_id -> created PracticalExamStudent
        map_prac = {row.student_id: row for row in created_prac}

        # Build schedule rows
        sched_rows = []
        for a in assignments:
            sid = a.get("student_id")
            prac_obj = map_prac.get(sid)
            if not prac_obj:
                continue

            sched_rows.append(
                PracticalExamStudentSchedule(
                    prac_student=prac_obj,
                    batch_no=int(a.get("batch_no")),
                    exam_date=a.get("exam_date"),
                    session=a.get("session"),
                    exam_time=a.get("exam_time") or "",
                    hall=hall_obj,
                )
            )

        PracticalExamStudentSchedule.objects.bulk_create(sched_rows)

    return JsonResponse({"ok": True, "message": f"Saved {len(sched_rows)} students schedule successfully."})




# views.py  ✅ NEW API (GET) to load already saved schedule for selected dept+batch+course
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.db.models import Prefetch

@require_GET
def ajax_get_prac_saved_schedule(request):
    department_id = request.GET.get("department_id")
    batch = request.GET.get("batch")
    course_id = request.GET.get("course_id")

    if not (department_id and batch and course_id):
        return JsonResponse({"ok": False, "message": "Missing params"}, status=400)

    qs = (
        PracticalExamStudent.objects
        .filter(department_id=department_id, batch=batch, course_id=course_id)
        .select_related("student")
        .prefetch_related(
            Prefetch("schedules", queryset=PracticalExamStudentSchedule.objects.select_related("hall"))
        )
        .order_by("student__reg_no")
    )

    if not qs.exists():
        return JsonResponse({"ok": True, "exists": False, "message": "No saved schedule found."})

    # hall: take from first schedule row (same hall for all in your save)
    first = qs.first()
    first_sched = first.schedules.first()
    hall_id = first_sched.hall_id if first_sched else None

    # build student assignments
    assignments = []
    for row in qs:
        sched = row.schedules.first()
        if not sched:
            continue
        assignments.append({
            "student_id": row.student_id,
            "reg_no": row.student.reg_no,
            "batch_no": sched.batch_no,
            "exam_date": sched.exam_date.strftime("%Y-%m-%d"),
            "session": sched.session,
            "exam_time": sched.exam_time or "",
            "hall_id": sched.hall_id
        })

    # build batch boxes from assignments (group by batch_no)
    batch_map = {}
    for a in assignments:
        bno = int(a["batch_no"])
        if bno not in batch_map:
            batch_map[bno] = {"count": 0, "exam_date": a["exam_date"], "session": a["session"], "exam_time": a["exam_time"]}
        batch_map[bno]["count"] += 1

    batch_boxes = [
        {"batch_no": bno, **vals}
        for bno, vals in sorted(batch_map.items(), key=lambda x: x[0])
    ]

    return JsonResponse({
        "ok": True,
        "exists": True,
        "hall_id": hall_id,
        "batch_boxes": batch_boxes,
        "assignments": assignments,  # per student
        "message": f"Loaded saved schedule ({len(assignments)} students)."
    })



from examination_management.models import PassValue

def passvalue(request):
    degrees = Degree.objects.filter(is_active=True)
    regulations = Regulations.objects.all()

    selected_degree = request.GET.get("degree") or request.POST.get("degree")
    selected_regulation = request.GET.get("regulation") or request.POST.get("regulation")
    

    pass_obj = None

    if selected_degree and selected_regulation:
        pass_obj = PassValue.objects.filter(
            degree_id=selected_degree,
            regulation_id=selected_regulation
        ).first()

    if request.method == "POST":
        degree_id = request.POST.get("degree")
        regulation_id = request.POST.get("regulation")
        iat_pass_value = request.POST.get("iat_pass_value")
        university_iat_pass_value = request.POST.get("university_iat_pass_value")

        if degree_id and regulation_id:
            pass_obj, created = PassValue.objects.update_or_create(
                degree_id=degree_id,
                regulation_id=regulation_id,
                defaults={
                    "iat_pass_value": iat_pass_value or None,
                    "university_iat_pass_value": university_iat_pass_value or None,
                }
            )
            messages.success(request, "Pass values saved successfully.")
            return redirect(f"{request.path}?degree={degree_id}&regulation={regulation_id}")

    context = {
        "degrees": degrees,
        "regulations": regulations,
        "selected_degree": selected_degree,
        "selected_regulation": selected_regulation,
        "pass_obj": pass_obj,
    }
    return render(request, "examination_management/admin/passvalue.html", context)



from examination_management.models import StudentInternalMark

from django.http import JsonResponse
from django.db.models import Sum, F

from collections import defaultdict
from django.http import JsonResponse
from django.db.models import Sum, F

from collections import defaultdict
from django.http import JsonResponse
from django.db.models import Sum, F

from collections import defaultdict
from django.http import JsonResponse
from django.db.models import Sum, F


from django.http import HttpResponse
from django.conf import settings
from django.contrib.staticfiles import finders

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.utils import ImageReader

from io import BytesIO
import json
import os


def admin_iat_result_analysis(request):

    degree_ids = StudentInternalMark.objects.values_list(
        "degree_id", flat=True
    ).distinct()

    degrees = Degree.objects.filter(id__in=degree_ids, is_active=True)

    context = {
        "degrees": degrees
    }

    return render(
        request,
        "examination_management/admin/admin_iat_result_analysis.html",
        context
    )



def load_departments(request):

    degree_id = request.GET.get("degree_id")

    dept_ids = StudentInternalMark.objects.filter(
        degree_id=degree_id
    ).values_list(
        "department_id", flat=True
    ).distinct()

    departments = Add_Department.objects.filter(
        id__in=dept_ids,
        is_active=True
    )

    data = []

    for d in departments:
        data.append({
            "id": d.id,
            "name": d.Department
        })

    return JsonResponse({"departments": data})



from examination_management.models import StudentInternalMark
from django.db.models import Sum, F

def is_all_value(value):
    return value in [None, "", "All", "ALL", "all", "null", "None"]


def admin_iat_result_analysis(request):
    degree_ids = StudentInternalMark.objects.exclude(
        degree_id__isnull=True
    ).values_list("degree_id", flat=True).distinct()

    degrees = Degree.objects.filter(
        id__in=degree_ids,
        is_active=True
    ).order_by("degree")

    context = {
        "degrees": degrees
    }

    return render(
        request,
        "examination_management/admin/admin_iat_result_analysis.html",
        context
    )


def load_admin_iat_result_departments(request):
    degree_id = request.GET.get("degree_id")

    qs = StudentInternalMark.objects.all()

    if not is_all_value(degree_id):
        qs = qs.filter(degree_id=degree_id)

    department_ids = qs.exclude(
        department_id__isnull=True
    ).values_list("department_id", flat=True).distinct()

    departments = Add_Department.objects.filter(
        id__in=department_ids
    ).order_by("Department")

    data = {
        "departments": list(
            departments.values("id", "Department")
        )
    }

    return JsonResponse(data)

def load_filter_values(request):
    degree_id = request.GET.get("degree_id")
    department_id = request.GET.get("department_id")
    batch = request.GET.get("batch")

    qs = StudentInternalMark.objects.all()

    if not is_all_value(degree_id):
        qs = qs.filter(degree_id=degree_id)

    if not is_all_value(department_id):
        qs = qs.filter(department_id=department_id)

    if not is_all_value(batch):
        qs = qs.filter(batch=batch)

    batch_list = list(
        qs.exclude(batch__isnull=True)
          .exclude(batch__exact="")
          .values_list("batch", flat=True)
          .distinct()
          .order_by("batch")
    )

    section_list = list(
        qs.exclude(section__isnull=True)
          .exclude(section__exact="")
          .values_list("section", flat=True)
          .distinct()
          .order_by("section")
    )

    semester_list = list(
        qs.exclude(semester__isnull=True)
          .exclude(semester__exact="")
          .values_list("semester", flat=True)
          .distinct()
          .order_by("semester")
    )

    exam_name_list = list(
        qs.exclude(exam_name__isnull=True)
          .exclude(exam_name__exact="")
          .values_list("exam_name", flat=True)
          .distinct()
          .order_by("exam_name")
    )

    return JsonResponse({
        "batch": batch_list,
        "section": section_list,
        "semester": semester_list,
        "exam_name": exam_name_list
    })


def is_all_value(value):
    return value in [None, "", "All", "ALL", "all", "null", "None"]





def is_all_value(value):
    return value in [None, "", "All", "all"]


def get_pass_percentage_table(request):
    degree = request.GET.get("degree")
    department = request.GET.get("department")
    batch = (request.GET.get("batch") or "").strip()
    section = (request.GET.get("section") or "").strip()
    semester = (request.GET.get("semester") or "").strip()
    exam_name = (request.GET.get("exam_name") or "").strip()

    if not semester or not exam_name:
        return JsonResponse({
            "status": False,
            "message": "Semester and Exam Name are required."
        }, status=400)

    qs = StudentInternalMark.objects.select_related(
        "student", "degree", "department", "course"
    ).all()

    if not is_all_value(degree):
        qs = qs.filter(degree_id=degree)

    if not is_all_value(department):
        qs = qs.filter(department_id=department)

    if not is_all_value(exam_name):
        qs = qs.filter(exam_name=exam_name)

    if not is_all_value(batch):
        qs = qs.filter(batch=batch)

    if not is_all_value(section):
        qs = qs.filter(section=section)

    if semester == "CURRENT":
        qs = qs.filter(
            student__semester=F("semester")
        ).exclude(
            student__semester__isnull=True
        ).exclude(
            student__semester__exact=""
        )
    else:
        qs = qs.filter(semester=semester)

    degree_name = "All"
    if not is_all_value(degree):
        degree_obj = Degree.objects.filter(id=degree).first()
        degree_name = degree_obj.degree if degree_obj else "All"

    department_name = "All"
    if not is_all_value(department):
        dept_obj = Add_Department.objects.filter(id=department).first()
        if dept_obj:
            department_name = getattr(dept_obj, "Department", "") or "All"

    pass_mark = 50
    rows = []

    def _to_int(value):
        try:
            if value is None or value == "":
                return 0
            return int(value)
        except Exception:
            return 0

    def calculate_pass_fail(group_qs):
        # --------------------------------------------------
        # COURSE TYPE COUNT
        # Theory:
        #   lecture_hours != 0 and practical_hours == 0
        # Integrated:
        #   lecture_hours != 0 and practical_hours != 0
        # --------------------------------------------------
        unique_courses = group_qs.values(
            "course_id",
            "course_code",
            "course__title",
        ).distinct()

        theory_course_count = 0
        integrated_course_count = 0

       

        for item in unique_courses:
            course_id = item.get("course_id")
            course_code = item.get("course_code")
            course_title = item.get("course__title")

            

            course_hours_obj = CourseHours.objects.filter(course_id=course_id).first()

            if not course_hours_obj:
                
                continue

            hour_config_id = course_hours_obj.hour_config_id
            

            if not hour_config_id:
                
                continue

            config = CourseHourConfig.objects.filter(id=hour_config_id).first()

            if not config:
                
                continue

            lecture_hours = config.lecture_hours or 0
            practical_hours = config.laboratory_hours or 0



            if lecture_hours != 0 and practical_hours != 0:
                integrated_course_count += 1
            elif lecture_hours != 0 and practical_hours == 0:
                theory_course_count += 1



        # --------------------------------------------------
        # STEP 1: aggregate full option totals
        # --------------------------------------------------
        option_level_rows = list(
            group_qs.values(
                "student_id",
                "student__reg_no",
                "student__name",
                "course_id",
                "course_code",
                "exam_name",
                "part_name",
                "question_number",
                "option_letter",
            )
            .annotate(
                option_obtained=Sum("marks_obtained"),
                option_max=Sum("max_marks"),
            )
            .order_by(
                "student__reg_no",
                "course_code",
                "part_name",
                "question_number",
                "option_letter",
            )
        )

        # --------------------------------------------------
        # STEP 2: choose only one full option per question
        # --------------------------------------------------
        question_choice_map = {}

        for row in option_level_rows:
            question_key = (
                row.get("student_id"),
                row.get("student__reg_no"),
                row.get("student__name"),
                row.get("course_id"),
                row.get("course_code"),
                (row.get("exam_name") or "").strip(),
                row.get("part_name"),
                row.get("question_number"),
            )

            option_obtained = _to_int(row.get("option_obtained"))
            option_max = _to_int(row.get("option_max"))
            option_letter = (row.get("option_letter") or "").strip()

            if question_key not in question_choice_map:
                question_choice_map[question_key] = {
                    "chosen_option": option_letter,
                    "chosen_obtained": option_obtained,
                    "chosen_max": option_max,
                }
            else:
                existing = question_choice_map[question_key]

                if option_max > existing["chosen_max"]:
                    question_choice_map[question_key] = {
                        "chosen_option": option_letter,
                        "chosen_obtained": option_obtained,
                        "chosen_max": option_max,
                    }
                elif option_max == existing["chosen_max"] and option_obtained > existing["chosen_obtained"]:
                    question_choice_map[question_key] = {
                        "chosen_option": option_letter,
                        "chosen_obtained": option_obtained,
                        "chosen_max": option_max,
                    }

        # --------------------------------------------------
        # STEP 3: sum chosen question totals course-wise
        # --------------------------------------------------
        course_totals_map = defaultdict(lambda: {
            "student_name": "",
            "course_id": None,
            "total_max_marks": 0,
            "total_marks_obtained": 0
        })

        for key, chosen in question_choice_map.items():
            student_id, reg_no, student_name, course_id, course_code, exam_name_value, part_name, question_number = key

            course_key = (reg_no or "-", course_code or "-")

            course_totals_map[course_key]["student_name"] = student_name or ""
            course_totals_map[course_key]["course_id"] = course_id
            course_totals_map[course_key]["total_max_marks"] += _to_int(chosen["chosen_max"])
            course_totals_map[course_key]["total_marks_obtained"] += _to_int(chosen["chosen_obtained"])

        student_result = {}
        student_course_details = []

        for (reg_no, course_code), totals in sorted(course_totals_map.items()):
            student_name = totals["student_name"]
            course_id = totals["course_id"]
            max_mark = totals["total_max_marks"]
            obtained = totals["total_marks_obtained"]

            converted_mark = (obtained / max_mark) * 100 if max_mark > 0 else 0
            status = "PASS" if converted_mark >= pass_mark else "FAIL"

            student_course_details.append({
                "reg_no": reg_no,
                "student_name": student_name,
                "course_code": course_code,
                "max_mark": max_mark,
                "mark_obtained": obtained,
                "status": status
            })

            if reg_no not in student_result:
                student_result[reg_no] = {
                    "student_name": student_name,
                    "failed_any_course": False,
                    "failed_subject_count": 0,
                    "courses": []
                }

            student_result[reg_no]["courses"].append({
                "course_code": course_code,
                "max_mark": max_mark,
                "mark_obtained": obtained,
                "status": status
            })

            if status == "FAIL":
                student_result[reg_no]["failed_any_course"] = True
                student_result[reg_no]["failed_subject_count"] += 1

        total_students = len(student_result)
        failed_students = sum(
            1 for info in student_result.values() if info["failed_any_course"]
        )
        passed_students = sum(
            1 for info in student_result.values() if not info["failed_any_course"]
        )
        percentage = round((passed_students / total_students) * 100, 2) if total_students else 0

        # --------------------------------------------------
        # Failed subject summary
        # Example:
        # 1 subject failed -> 10 students
        # 2 subjects failed -> 5 students
        # --------------------------------------------------
        failed_subject_summary = defaultdict(int)

        for reg_no, info in student_result.items():
            fail_count = info.get("failed_subject_count", 0)
            if fail_count > 0:
                failed_subject_summary[fail_count] += 1

        failed_subject_summary = dict(sorted(failed_subject_summary.items()))

        return (
            total_students,
            failed_students,
            passed_students,
            percentage,
            student_course_details,
            student_result,
            theory_course_count,
            integrated_course_count,
            failed_subject_summary,
        )

    if semester == "CURRENT":
        grouped_keys = list(
            qs.values_list(
                "degree_id",
                "department_id",
                "batch",
                "section",
                "semester"
            ).distinct().order_by(
                "degree_id",
                "department_id",
                "batch",
                "section",
                "semester"
            )
        )

        for degree_id, department_id, batch_value, section_value, semester_value in grouped_keys:
            group_qs = qs.filter(
                degree_id=degree_id,
                department_id=department_id,
                batch=batch_value,
                section=section_value,
                semester=semester_value
            )

            degree_obj = Degree.objects.filter(id=degree_id).first()
            dept_obj = Add_Department.objects.filter(id=department_id).first()

            row_degree = degree_obj.degree if degree_obj else "-"
            row_department = getattr(dept_obj, "Department", "") if dept_obj else "-"

            (
                total_students,
                failed_students,
                passed_students,
                percentage,
                student_course_details,
                student_result,
                theory_course_count,
                integrated_course_count,
                failed_subject_summary,
            ) = calculate_pass_fail(group_qs)

            rows.append({
                "degree": row_degree,
                "department": row_department or "-",
                "batch": batch_value or "-",
                "section": section_value or "-",
                "semester": semester_value or "-",
                "total_students": total_students,
                "failed_students": failed_students,
                "passed_students": passed_students,
                "pass_percentage": f"{percentage}%",
                "degree_id": degree_id or "",
                "department_id": department_id or "",
                "batch_value": batch_value or "",
                "section_value": section_value or "",
                "semester_value": semester_value or "",
                "exam_name": exam_name,
                "theory_course_count": theory_course_count,
                "integrated_course_count": integrated_course_count,
                "failed_subject_summary": failed_subject_summary,
                "student_course_details": student_course_details,
                "student_wise_result": [
                    {
                        "reg_no": reg_no,
                        "student_name": info["student_name"],
                        "overall_status": "FAILED" if info["failed_any_course"] else "PASSED",
                        "failed_subject_count": info["failed_subject_count"],
                        "courses": info["courses"]
                    }
                    for reg_no, info in student_result.items()
                ]
            })
    else:
        grouped_keys = list(
            qs.values_list(
                "degree_id",
                "department_id",
                "batch",
                "section"
            ).distinct().order_by(
                "degree_id",
                "department_id",
                "batch",
                "section"
            )
        )

        for degree_id, department_id, batch_value, section_value in grouped_keys:
            group_qs = qs.filter(
                degree_id=degree_id,
                department_id=department_id,
                batch=batch_value,
                section=section_value
            )

            degree_obj = Degree.objects.filter(id=degree_id).first()
            dept_obj = Add_Department.objects.filter(id=department_id).first()

            row_degree = degree_obj.degree if degree_obj else "-"
            row_department = getattr(dept_obj, "Department", "") if dept_obj else "-"

            (
                total_students,
                failed_students,
                passed_students,
                percentage,
                student_course_details,
                student_result,
                theory_course_count,
                integrated_course_count,
                failed_subject_summary,
            ) = calculate_pass_fail(group_qs)

            rows.append({
                "degree": row_degree,
                "department": row_department or "-",
                "batch": batch_value or "-",
                "section": section_value or "-",
                "semester": semester,
                "total_students": total_students,
                "failed_students": failed_students,
                "passed_students": passed_students,
                "pass_percentage": f"{percentage}%",
                "degree_id": degree_id or "",
                "department_id": department_id or "",
                "batch_value": batch_value or "",
                "section_value": section_value or "",
                "semester_value": semester,
                "exam_name": exam_name,
                "theory_course_count": theory_course_count,
                "integrated_course_count": integrated_course_count,
                "failed_subject_summary": failed_subject_summary,
                "student_course_details": student_course_details,
                "student_wise_result": [
                    {
                        "reg_no": reg_no,
                        "student_name": info["student_name"],
                        "overall_status": "FAILED" if info["failed_any_course"] else "PASSED",
                        "failed_subject_count": info["failed_subject_count"],
                        "courses": info["courses"]
                    }
                    for reg_no, info in student_result.items()
                ]
            })

    if not rows:
        rows.append({
            "degree": degree_name,
            "department": department_name,
            "batch": "-",
            "section": "-",
            "semester": "Current Semester" if semester == "CURRENT" else semester,
            "total_students": 0,
            "failed_students": 0,
            "passed_students": 0,
            "pass_percentage": "0%",
            "degree_id": degree or "",
            "department_id": department or "",
            "batch_value": batch or "",
            "section_value": section or "",
            "semester_value": semester or "",
            "exam_name": exam_name,
            "theory_course_count": 0,
            "integrated_course_count": 0,
            "failed_subject_summary": {},
            "student_course_details": [],
            "student_wise_result": []
        })

    # --------------------------------------------------
    # OVERALL SUMMARY FOR ALL RETURNED ROWS
    # --------------------------------------------------
    overall_theory_course_count = sum(_to_int(row.get("theory_course_count")) for row in rows)
    overall_integrated_course_count = sum(_to_int(row.get("integrated_course_count")) for row in rows)

    overall_failed_subject_summary = defaultdict(int)

    for row in rows:
        row_failed_summary = row.get("failed_subject_summary", {}) or {}
        for fail_count, student_count in row_failed_summary.items():
            overall_failed_subject_summary[_to_int(fail_count)] += _to_int(student_count)

    overall_failed_subject_summary = dict(sorted(overall_failed_subject_summary.items()))

    return JsonResponse({
        "status": True,
        "filters": {
            "degree": degree_name,
            "department": department_name,
            "batch": batch if not is_all_value(batch) else "All",
            "section": section if not is_all_value(section) else "All",
            "semester": "Current Semester" if semester == "CURRENT" else semester,
            "exam_name": exam_name
        },
        "summary": {
            "theory_course_count": overall_theory_course_count,
            "integrated_course_count": overall_integrated_course_count,
            "failed_subject_summary": overall_failed_subject_summary
        },
        "rows": rows
    })


def get_course_wise_details(request):
    degree = request.GET.get("degree")
    department = request.GET.get("department")
    batch = (request.GET.get("batch") or "").strip()
    section = (request.GET.get("section") or "").strip()
    semester = (request.GET.get("semester") or "").strip()
    exam_name = (request.GET.get("exam_name") or "").strip()

    if not semester or not exam_name:
        return JsonResponse({
            "status": False,
            "message": "Semester and Exam Name are required."
        }, status=400)

    qs = StudentInternalMark.objects.select_related(
        "course", "degree", "department", "student"
    ).all()

    if not is_all_value(degree):
        qs = qs.filter(degree_id=degree)

    if not is_all_value(department):
        qs = qs.filter(department_id=department)

    if not is_all_value(batch):
        qs = qs.filter(batch=batch)

    if not is_all_value(section):
        qs = qs.filter(section=section)

    if not is_all_value(exam_name):
        qs = qs.filter(exam_name=exam_name)

    if semester == "CURRENT":
        qs = qs.filter(
            student__semester=F("semester")
        ).exclude(
            student__semester__isnull=True
        ).exclude(
            student__semester__exact=""
        )
    else:
        qs = qs.filter(semester=semester)

    degree_name = "All"
    if not is_all_value(degree):
        degree_obj = Degree.objects.filter(id=degree).first()
        degree_name = degree_obj.degree if degree_obj else "All"

    department_name = "All"
    if not is_all_value(department):
        dept_obj = Add_Department.objects.filter(id=department).first()
        if dept_obj:
            department_name = getattr(dept_obj, "Department", "") or "All"

    def _to_int(value):
        try:
            if value is None or value == "":
                return 0
            return int(value)
        except Exception:
            return 0

    option_level_rows = list(
        qs.values(
            "course_id",
            "course_code",
            "course__title",
            "department_id",
            "batch",
            "section",
            "student_id",
            "student__reg_no",
            "degree_id",
            "student__regulation",
            "part_name",
            "question_number",
            "option_letter",
        ).annotate(
            option_obtained=Sum("marks_obtained"),
            option_max=Sum("max_marks"),
        ).order_by(
            "course_code",
            "student__reg_no",
            "part_name",
            "question_number",
            "option_letter",
        )
    )

    question_choice_map = {}

    for row in option_level_rows:
        question_key = (
            row.get("course_id"),
            row.get("course_code"),
            row.get("course__title"),
            row.get("department_id"),
            row.get("batch"),
            row.get("section"),
            row.get("student_id"),
            row.get("student__reg_no"),
            row.get("degree_id"),
            row.get("student__regulation"),
            row.get("part_name"),
            row.get("question_number"),
        )

        option_obtained = _to_int(row.get("option_obtained"))
        option_max = _to_int(row.get("option_max"))
        option_letter = (row.get("option_letter") or "").strip()

        if question_key not in question_choice_map:
            question_choice_map[question_key] = {
                "chosen_option": option_letter,
                "chosen_obtained": option_obtained,
                "chosen_max": option_max,
            }
        else:
            existing = question_choice_map[question_key]

            if option_max > existing["chosen_max"]:
                question_choice_map[question_key] = {
                    "chosen_option": option_letter,
                    "chosen_obtained": option_obtained,
                    "chosen_max": option_max,
                }
            elif option_max == existing["chosen_max"] and option_obtained > existing["chosen_obtained"]:
                question_choice_map[question_key] = {
                    "chosen_option": option_letter,
                    "chosen_obtained": option_obtained,
                    "chosen_max": option_max,
                }

    student_course_totals = {}

    for key, chosen in question_choice_map.items():
        (
            course_id,
            course_code,
            course_title,
            department_id,
            batch_value,
            section_value,
            student_id,
            reg_no,
            degree_id,
            regulation,
            part_name,
            question_number
        ) = key

        student_course_key = (
            course_id,
            course_code or "-",
            course_title or "-",
            department_id,
            batch_value or "",
            section_value or "",
            reg_no or "-",
            degree_id,
            regulation,
        )

        if student_course_key not in student_course_totals:
            student_course_totals[student_course_key] = {
                "total_max_marks": 0,
                "total_marks_obtained": 0,
            }

        student_course_totals[student_course_key]["total_max_marks"] += _to_int(chosen["chosen_max"])
        student_course_totals[student_course_key]["total_marks_obtained"] += _to_int(chosen["chosen_obtained"])

    pass_value_map = {
        (pv.degree_id, pv.regulation_id): _to_int(pv.iat_pass_value)
        for pv in PassValue.objects.all()
    }

    # -----------------------------
    # FACULTY MAP
    # -----------------------------
    course_ids = {key[0] for key in student_course_totals.keys() if key[0]}
    department_ids = {key[3] for key in student_course_totals.keys() if key[3]}
    batch_values = {key[4] for key in student_course_totals.keys() if key[4]}
    section_values = {key[5] for key in student_course_totals.keys() if key[5]}

    faculty_qs = AssignSubjectFaculty.objects.select_related("faculty", "course").filter(
        is_active=True
    )

    if department_ids:
        faculty_qs = faculty_qs.filter(department_id__in=department_ids)
    if course_ids:
        faculty_qs = faculty_qs.filter(course_id__in=course_ids)
    if batch_values:
        faculty_qs = faculty_qs.filter(batch__in=batch_values)
    if section_values:
        faculty_qs = faculty_qs.filter(section__in=section_values)

    faculty_map = {}
    for obj in faculty_qs:
        faculty_key = (
            obj.department_id,
            (obj.batch or "").strip(),
            (obj.section or "").strip(),
            obj.course_id,
        )
        faculty_map[faculty_key] = obj.faculty.name if obj.faculty else "-"

    course_summary = {}

    for (
        course_id,
        course_code,
        course_title,
        department_id,
        batch_value,
        section_value,
        reg_no,
        degree_id,
        regulation
    ), totals in student_course_totals.items():
        max_mark = totals["total_max_marks"]
        obtained = totals["total_marks_obtained"]

        pass_mark = pass_value_map.get((degree_id, regulation), 50)
        converted_mark = (obtained / max_mark) * 100 if max_mark > 0 else 0

        faculty_name = faculty_map.get(
            (
                department_id,
                (batch_value or "").strip(),
                (section_value or "").strip(),
                course_id,
            ),
            "-"
        )

        key = course_code

        if key not in course_summary:
            course_summary[key] = {
                "course_code": course_code,
                "course_title": course_title,
                "faculty_name": faculty_name,
                "total_students": 0,
                "passed_students": 0,
                "failed_students": 0
            }

        course_summary[key]["total_students"] += 1

        if converted_mark >= pass_mark:
            course_summary[key]["passed_students"] += 1
        else:
            course_summary[key]["failed_students"] += 1

    rows = []
    for _, data in course_summary.items():
        total_students = data["total_students"]
        passed_students = data["passed_students"]
        failed_students = data["failed_students"]
        percentage = round((passed_students / total_students) * 100, 2) if total_students else 0

        rows.append({
            "course_code": data["course_code"],
            "course_title": data["course_title"],
            "faculty_name": data["faculty_name"],
            "total_students": total_students,
            "passed_students": passed_students,
            "failed_students": failed_students,
            "pass_percentage": f"{percentage}%"
        })

    rows = sorted(rows, key=lambda x: x["course_code"])

    return JsonResponse({
        "status": True,
        "filters": {
            "degree": degree_name,
            "department": department_name,
            "batch": batch if not is_all_value(batch) else "All",
            "section": section if not is_all_value(section) else "All",
            "semester": "Current Semester" if semester == "CURRENT" else semester,
            "exam_name": exam_name
        },
        "rows": rows
    })



def download_student_wise_pdf(request):
    response_json = get_pass_percentage_table(request)
    data = json.loads(response_json.content.decode("utf-8"))

    filters = data.get("filters", {})
    rows = data.get("rows", [])

    selected_row = rows[0] if rows else None
    student_wise_result = selected_row.get("student_wise_result", []) if selected_row else []

    all_course_codes = []
    for student in student_wise_result:
        for course in student.get("courses", []):
            code = course.get("course_code", "-")
            if code not in all_course_codes:
                all_course_codes.append(code)

    all_course_codes.sort()

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=34 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="CenterTitlePDF",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        fontSize=15,
        spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        name="CenterSubTitlePDF",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontName="Helvetica",
        fontSize=10,
        leading=12,
    ))
    styles.add(ParagraphStyle(
        name="FilterTextPDF",
        parent=styles["Normal"],
        alignment=TA_LEFT,
        fontName="Helvetica",
        fontSize=9,
        leading=12,
    ))
    styles.add(ParagraphStyle(
        name="TableCellPDF",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
    ))
    styles.add(ParagraphStyle(
        name="TableCellCenterPDF",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        alignment=TA_CENTER,
        leading=10,
    ))

    elements = []

    elements.append(Spacer(1, 4))
    elements.append(Paragraph("Student Wise Pass Details", styles["CenterSubTitlePDF"]))
    elements.append(Spacer(1, 6))

    filter_lines = [
        f"<b>Degree :</b> {filters.get('degree', '-')}",
        f"<b>Department :</b> {filters.get('department', '-')}",
        f"<b>Batch :</b> {filters.get('batch', '-')}",
        f"<b>Section :</b> {filters.get('section', '-')}",
        f"<b>Semester :</b> {filters.get('semester', '-')}",
        f"<b>Exam :</b> {filters.get('exam_name', '-')}",
    ]
    elements.append(Paragraph(" | ".join(filter_lines), styles["FilterTextPDF"]))
    elements.append(Spacer(1, 8))

    table_data = [[
        Paragraph("<b>S.No</b>", styles["TableCellCenterPDF"]),
        Paragraph("<b>Reg No</b>", styles["TableCellCenterPDF"]),
        Paragraph("<b>Name</b>", styles["TableCellCenterPDF"]),
    ]]

    for code in all_course_codes:
        table_data[0].append(Paragraph(f"<b>{code}</b>", styles["TableCellCenterPDF"]))

    table_data[0].append(Paragraph("<b>Overall Status</b>", styles["TableCellCenterPDF"]))

    # count trackers
    gt_90_counts = {code: 0 for code in all_course_codes}
    gt_60_counts = {code: 0 for code in all_course_codes}
    lt_50_counts = {code: 0 for code in all_course_codes}

    if student_wise_result:
        for idx, student in enumerate(student_wise_result, start=1):
            course_marks_map = {}
            for course in student.get("courses", []):
                code = course.get("course_code", "-")
                mark = course.get("mark_obtained", "-")
                course_marks_map[code] = mark

                try:
                    numeric_mark = float(mark)
                    if numeric_mark > 90:
                        gt_90_counts[code] += 1
                    if numeric_mark > 60:
                        gt_60_counts[code] += 1
                    if numeric_mark < 50:
                        lt_50_counts[code] += 1
                except Exception:
                    pass

            row_data = [
                Paragraph(str(idx), styles["TableCellCenterPDF"]),
                Paragraph(str(student.get("reg_no", "-")), styles["TableCellCenterPDF"]),
                Paragraph(str(student.get("student_name", "-")), styles["TableCellPDF"]),
            ]

            for code in all_course_codes:
                mark = course_marks_map.get(code, "-")
                row_data.append(Paragraph(str(mark), styles["TableCellCenterPDF"]))

            row_data.append(Paragraph(str(student.get("overall_status", "-")), styles["TableCellCenterPDF"]))
            table_data.append(row_data)
    else:
        no_data_row = [
            Paragraph("-", styles["TableCellCenterPDF"]),
            Paragraph("-", styles["TableCellCenterPDF"]),
            Paragraph("No student wise data found.", styles["TableCellCenterPDF"]),
        ]
        for _ in all_course_codes:
            no_data_row.append(Paragraph("-", styles["TableCellCenterPDF"]))
        no_data_row.append(Paragraph("-", styles["TableCellCenterPDF"]))
        table_data.append(no_data_row)

    # summary rows
    summary_90_row = [
        Paragraph("<b></b>", styles["TableCellCenterPDF"]),
        Paragraph("<b></b>", styles["TableCellCenterPDF"]),
        Paragraph("<b>> 90</b>", styles["TableCellPDF"]),
    ]
    for code in all_course_codes:
        summary_90_row.append(Paragraph(str(gt_90_counts[code]), styles["TableCellCenterPDF"]))
    summary_90_row.append(Paragraph("-", styles["TableCellCenterPDF"]))
    table_data.append(summary_90_row)

    summary_60_row = [
        Paragraph("<b></b>", styles["TableCellCenterPDF"]),
        Paragraph("<b></b>", styles["TableCellCenterPDF"]),
        Paragraph("<b>> 60</b>", styles["TableCellPDF"]),
    ]
    for code in all_course_codes:
        summary_60_row.append(Paragraph(str(gt_60_counts[code]), styles["TableCellCenterPDF"]))
    summary_60_row.append(Paragraph("-", styles["TableCellCenterPDF"]))
    table_data.append(summary_60_row)

    summary_50_row = [
        Paragraph("<b></b>", styles["TableCellCenterPDF"]),
        Paragraph("<b></b>", styles["TableCellCenterPDF"]),
        Paragraph("<b>< 50</b>", styles["TableCellPDF"]),
    ]
    for code in all_course_codes:
        summary_50_row.append(Paragraph(str(lt_50_counts[code]), styles["TableCellCenterPDF"]))
    summary_50_row.append(Paragraph("-", styles["TableCellCenterPDF"]))
    table_data.append(summary_50_row)

    fixed_cols = [15 * mm, 28 * mm, 45 * mm]
    remaining_width = 273 * mm - sum(fixed_cols) - 28 * mm
    course_col_count = max(len(all_course_codes), 1)
    each_course_width = remaining_width / course_col_count

    col_widths = fixed_cols + ([each_course_width] * len(all_course_codes)) + [28 * mm]

    table = Table(
        table_data,
        colWidths=col_widths,
        repeatRows=1,
        hAlign="LEFT",
    )

    total_rows = len(table_data)
    summary_start = total_rows - 3

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9e2f3")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ROWBACKGROUNDS", (0, 1), (-1, summary_start - 1), [colors.white, colors.HexColor("#f7f7f7")]),
        ("BACKGROUND", (0, summary_start), (-1, -1), colors.HexColor("#fff2cc")),
        ("FONTNAME", (0, summary_start), (-1, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)

    def _on_page(canvas, _doc):
        canvas.saveState()

        page_w, page_h = landscape(A4)
        left = 12 * mm
        right = 12 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)

        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                candidate = os.path.join(static_root, logo_rel)
                if os.path.exists(candidate):
                    logo_path = candidate

        if not logo_path:
            for static_dir in getattr(settings, "STATICFILES_DIRS", []):
                candidate = os.path.join(static_dir, logo_rel)
                if os.path.exists(candidate):
                    logo_path = candidate
                    break

        if logo_path and os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                target_h = 18 * mm
                target_w = target_h * (iw / float(ih))
                canvas.drawImage(
                    img,
                    left,
                    page_h - (target_h + 8 * mm),
                    width=target_w,
                    height=target_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass

        canvas.setFillColor(colors.black)
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(page_w / 2.0, page_h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(page_w / 2.0, page_h - 15 * mm, "Rajapalayam - 626117")
        canvas.drawCentredString(page_w / 2.0, page_h - 19 * mm, "Affiliated to Anna University, Chennai")

        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(page_w / 2.0, page_h - 26 * mm, "Student Wise Pass Details")

        rule_y = page_h - 30 * mm
        canvas.setStrokeColor(colors.black)
        canvas.setLineWidth(0.5)
        canvas.line(left, rule_y, page_w - right, rule_y)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(page_w - right, 8 * mm, f"Page {canvas.getPageNumber()}")

        canvas.restoreState()

    doc.build(elements, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buf.getvalue()
    buf.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="student_wise_pass_details.pdf"'
    return response




from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import Sum, F
from collections import defaultdict


def download_student_wise_excel(request):
    degree = request.GET.get("degree")
    department = request.GET.get("department")
    batch = (request.GET.get("batch") or "").strip()
    section = (request.GET.get("section") or "").strip()
    semester = (request.GET.get("semester") or "").strip()
    exam_name = (request.GET.get("exam_name") or "").strip()

    qs = StudentInternalMark.objects.select_related(
        "student", "degree", "department", "course"
    ).all()

    if not is_all_value(degree):
        qs = qs.filter(degree_id=degree)

    if not is_all_value(department):
        qs = qs.filter(department_id=department)

    if not is_all_value(batch):
        qs = qs.filter(batch=batch)

    if not is_all_value(section):
        qs = qs.filter(section=section)

    if not is_all_value(exam_name):
        qs = qs.filter(exam_name=exam_name)

    if semester == "CURRENT":
        qs = qs.filter(
            student__semester=F("semester")
        ).exclude(student__semester__isnull=True).exclude(student__semester__exact="")
    else:
        qs = qs.filter(semester=semester)

    def _to_int(value):
        try:
            if value is None or value == "":
                return 0
            return int(value)
        except Exception:
            return 0

    option_rows = list(
        qs.values(
            "student_id",
            "student__reg_no",
            "student__year",
            "student__semester",
            "student__regulation",
            "department__Department_code",
            "faculty_assignment__faculty__faculty_id",
            "batch",
            "section",
            "semester",
            "course_code",
            "academic_year",
            "part_name",
            "question_number",
            "option_letter",
        )
        .annotate(
            option_obtained=Sum("marks_obtained"),
            option_max=Sum("max_marks"),
        )
        .order_by(
            "student__reg_no",
            "course_code",
            "part_name",
            "question_number",
            "option_letter",
        )
    )

    question_choice_map = {}

    for row in option_rows:
        question_key = (
            row.get("student_id"),
            row.get("student__reg_no"),
            row.get("course_code"),
            row.get("part_name"),
            row.get("question_number"),
        )

        option_obtained = _to_int(row.get("option_obtained"))
        option_max = _to_int(row.get("option_max"))

        if question_key not in question_choice_map:
            question_choice_map[question_key] = {
                "row": row,
                "marks": option_obtained,
                "max": option_max,
            }
        else:
            old = question_choice_map[question_key]
            if option_max > old["max"] or (
                option_max == old["max"] and option_obtained > old["marks"]
            ):
                question_choice_map[question_key] = {
                    "row": row,
                    "marks": option_obtained,
                    "max": option_max,
                }

    final_rows = defaultdict(lambda: {
        "marks": 0,
        "row": None,
    })

    for item in question_choice_map.values():
        row = item["row"]

        key = (
            row.get("faculty_assignment__faculty__faculty_id"),
            row.get("department__Department_code"),
            row.get("student__reg_no"),
            row.get("student__year"),
            row.get("semester"),
            row.get("section"),
            row.get("student__regulation"),
            row.get("batch"),
            row.get("course_code"),
            row.get("academic_year"),
        )

        final_rows[key]["marks"] += _to_int(item["marks"])
        final_rows[key]["row"] = row

    wb = Workbook()
    ws = wb.active
    ws.title = "Student Wise Marks"

    headers = [
        "faculty_id",
        "department_code",
        "reg_no",
        "year",
        "semester",
        "section",
        "regulation",
        "batch",
        "course_code",
        "marks",
        "academic_year",
    ]

    ws.append(headers)

    for key, data in final_rows.items():
        (
            faculty_id,
            department_code,
            reg_no,
            year,
            sem,
            sec,
            regulation,
            batch_value,
            course_code,
            academic_year,
        ) = key

        ws.append([
            faculty_id or "",
            department_code or "",
            reg_no or "",
            year or "",
            sem or "",
            sec or "",
            regulation or "",
            batch_value or "",
            course_code or "",
            data["marks"],
            academic_year or "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="student_wise_marks.xlsx"'

    wb.save(response)
    return response




from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from django.http import HttpResponse
from django.conf import settings
from django.contrib.staticfiles import finders

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.utils import ImageReader

from io import BytesIO
import json
import os


def download_course_wise_pdf(request):
    response_json = get_course_wise_details(request)
    data = json.loads(response_json.content.decode("utf-8"))

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=34 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="CenterTitle",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        fontSize=15,
        spaceAfter=4,
    ))
    styles.add(ParagraphStyle(
        name="CenterSubTitle",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontName="Helvetica",
        fontSize=10,
        leading=12,
    ))
    styles.add(ParagraphStyle(
        name="FilterText",
        parent=styles["Normal"],
        alignment=TA_LEFT,
        fontName="Helvetica",
        fontSize=9,
        leading=12,
    ))
    styles.add(ParagraphStyle(
        name="TableCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
    ))
    styles.add(ParagraphStyle(
        name="TableCellCenter",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        alignment=TA_CENTER,
        leading=10,
    ))

    elements = []

    elements.append(Spacer(1, 4))
    elements.append(Paragraph("Course-wise Pass Details", styles["CenterSubTitle"]))
    elements.append(Spacer(1, 6))

    filters = data.get("filters", {})
    filter_lines = [
        f"<b>Degree :</b> {filters.get('degree', '-')}",
        f"<b>Department :</b> {filters.get('department', '-')}",
        f"<b>Batch :</b> {filters.get('batch', '-')}",
        f"<b>Section :</b> {filters.get('section', '-')}",
        f"<b>Semester :</b> {filters.get('semester', '-')}",
        f"<b>Exam :</b> {filters.get('exam_name', '-')}",
    ]
    elements.append(Paragraph(" | ".join(filter_lines), styles["FilterText"]))
    elements.append(Spacer(1, 8))

    table_data = [[
        Paragraph("<b>S.No</b>", styles["TableCellCenter"]),
        Paragraph("<b>Course & Course Name</b>", styles["TableCellCenter"]),
        Paragraph("<b>Faculty</b>", styles["TableCellCenter"]),
        Paragraph("<b>No of Students</b>", styles["TableCellCenter"]),
        Paragraph("<b>No of Passed</b>", styles["TableCellCenter"]),
        Paragraph("<b>No of Failed</b>", styles["TableCellCenter"]),
        Paragraph("<b>Pass Percentage</b>", styles["TableCellCenter"]),
    ]]

    rows = data.get("rows", [])
    for idx, row in enumerate(rows, start=1):
        course_text = f"{row.get('course_code', '-')} - {row.get('course_title', '-')}"
        table_data.append([
            Paragraph(str(idx), styles["TableCellCenter"]),
            Paragraph(course_text, styles["TableCell"]),
            Paragraph(row.get("faculty_name", "-"), styles["TableCell"]),
            Paragraph(str(row.get("total_students", 0)), styles["TableCellCenter"]),
            Paragraph(str(row.get("passed_students", 0)), styles["TableCellCenter"]),
            Paragraph(str(row.get("failed_students", 0)), styles["TableCellCenter"]),
            Paragraph(row.get("pass_percentage", "0%"), styles["TableCellCenter"]),
        ])

    if not rows:
        table_data.append([
            Paragraph("-", styles["TableCellCenter"]),
            Paragraph("No course data found.", styles["TableCellCenter"]),
            Paragraph("-", styles["TableCellCenter"]),
            Paragraph("-", styles["TableCellCenter"]),
            Paragraph("-", styles["TableCellCenter"]),
            Paragraph("-", styles["TableCellCenter"]),
            Paragraph("-", styles["TableCellCenter"]),
        ])

    table = Table(
        table_data,
        colWidths=[
            15 * mm,
            85 * mm,
            55 * mm,
            28 * mm,
            28 * mm,
            28 * mm,
            30 * mm,
        ],
        repeatRows=1,
        hAlign="LEFT",
    )

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9e2f3")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (3, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f7f7")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)

    def _on_page(canvas, _doc):
        canvas.saveState()

        page_w, page_h = landscape(A4)
        left = 12 * mm
        right = 12 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)

        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                candidate = os.path.join(static_root, logo_rel)
                if os.path.exists(candidate):
                    logo_path = candidate

        if not logo_path:
            for static_dir in getattr(settings, "STATICFILES_DIRS", []):
                candidate = os.path.join(static_dir, logo_rel)
                if os.path.exists(candidate):
                    logo_path = candidate
                    break

        if logo_path and os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                target_h = 18 * mm
                target_w = target_h * (iw / float(ih))
                canvas.drawImage(
                    img,
                    left,
                    page_h - (target_h + 8 * mm),
                    width=target_w,
                    height=target_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass

        canvas.setFillColor(colors.black)
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(page_w / 2.0, page_h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(page_w / 2.0, page_h - 15 * mm, "Rajapalayam - 626117")
        canvas.drawCentredString(page_w / 2.0, page_h - 19 * mm, "Affiliated to Anna University, Chennai")

        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(page_w / 2.0, page_h - 26 * mm, "Course-wise Pass Details")

        rule_y = page_h - 30 * mm
        canvas.setStrokeColor(colors.black)
        canvas.setLineWidth(0.5)
        canvas.line(left, rule_y, page_w - right, rule_y)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(page_w - right, 8 * mm, f"Page {canvas.getPageNumber()}")

        canvas.restoreState()

    doc.build(elements, onFirstPage=_on_page, onLaterPages=_on_page)

    pdf = buf.getvalue()
    buf.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="course_wise_pass_details.pdf"'
    return response








from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.db.models import Sum, F
from io import BytesIO
import os

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader

from django.contrib.staticfiles import finders
from django.conf import settings


def download_pass_percentage_pdf(request):
    degree = request.GET.get("degree")
    department = request.GET.get("department")
    batch = (request.GET.get("batch") or "").strip()
    section = (request.GET.get("section") or "").strip()
    semester = (request.GET.get("semester") or "").strip()
    exam_name = (request.GET.get("exam_name") or "").strip()

    if not semester or not exam_name:
        return HttpResponse("Semester and Exam Name are required.", status=400)

    base_qs = StudentInternalMark.objects.select_related("student").all()

    if not is_all_value(exam_name):
        base_qs = base_qs.filter(exam_name=exam_name)

    if not is_all_value(batch):
        base_qs = base_qs.filter(batch=batch)

    if not is_all_value(section):
        base_qs = base_qs.filter(section=section)

    if semester == "CURRENT":
        base_qs = base_qs.filter(
            student__semester=F("semester")
        ).exclude(
            student__semester__isnull=True
        ).exclude(
            student__semester__exact=""
        )
    else:
        base_qs = base_qs.filter(semester=semester)

    # decide grouping
    if is_all_value(degree) and is_all_value(department):
        group_pairs = list(
            base_qs.exclude(degree_id__isnull=True)
                   .exclude(department_id__isnull=True)
                   .values_list("degree_id", "department_id")
                   .distinct()
                   .order_by("degree_id", "department_id")
        )
    elif not is_all_value(degree) and is_all_value(department):
        group_pairs = list(
            base_qs.filter(degree_id=degree)
                   .exclude(department_id__isnull=True)
                   .values_list("degree_id", "department_id")
                   .distinct()
                   .order_by("degree_id", "department_id")
        )
    elif is_all_value(degree) and not is_all_value(department):
        group_pairs = list(
            base_qs.filter(department_id=department)
                   .exclude(degree_id__isnull=True)
                   .values_list("degree_id", "department_id")
                   .distinct()
                   .order_by("degree_id", "department_id")
        )
    else:
        group_pairs = [(degree, department)]

    pass_mark = 50
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=45 * mm,
        bottomMargin=20 * mm
    )

    styles = getSampleStyleSheet()
    story = []

    def calculate_pass_fail(group_qs):
        course_totals = list(
            group_qs.values("student__reg_no", "course_code")
                    .annotate(total_marks=Sum("marks_obtained"))
                    .order_by("student__reg_no", "course_code")
        )

        

            

        student_result = {}

        for item in course_totals:
            reg_no = item["student__reg_no"]
            total_marks = item["total_marks"] or 0

            if reg_no not in student_result:
                student_result[reg_no] = {
                    "failed_any_course": False,
                    "courses": []
                }

            student_result[reg_no]["courses"].append({
                "course_code": item["course_code"],
                "total_marks": total_marks
            })

            if total_marks < pass_mark:
                student_result[reg_no]["failed_any_course"] = True

        
        for reg_no, info in student_result.items():
            status = "FAILED" if info["failed_any_course"] else "PASSED"
            print(f"Student: {reg_no} -> {status}")
            for c in info["courses"]:
                print(f"   Course: {c['course_code']} | Marks: {c['total_marks']}")

        total_students = len(student_result)
        failed_students = sum(1 for info in student_result.values() if info["failed_any_course"])
        passed_students = sum(1 for info in student_result.values() if not info["failed_any_course"])
        percentage = round((passed_students / total_students) * 100, 2) if total_students else 0



        return total_students, failed_students, passed_students, percentage

    total_groups = len(group_pairs)

    for idx, (degree_id, department_id) in enumerate(group_pairs, start=1):
        qs = base_qs.all()

        if not is_all_value(degree_id):
            qs = qs.filter(degree_id=degree_id)

        if not is_all_value(department_id):
            qs = qs.filter(department_id=department_id)

        degree_name = "All"
        if not is_all_value(degree_id):
            degree_obj = Degree.objects.filter(id=degree_id).first()
            degree_name = degree_obj.degree if degree_obj else "All"

        department_name = "All"
        if not is_all_value(department_id):
            dept_obj = Add_Department.objects.filter(id=department_id).first()
            if dept_obj:
                department_name = getattr(dept_obj, "Department", None) or "All"

        table_rows = []

        if semester == "CURRENT":
            grouped_keys = list(
                qs.values_list("batch", "section", "semester")
                  .distinct()
                  .order_by("batch", "section", "semester")
            )

            for batch_value, section_value, semester_value in grouped_keys:
                group_qs = qs.filter(
                    batch=batch_value,
                    section=section_value,
                    semester=semester_value
                )

                
                total_students, failed_students, passed_students, percentage = calculate_pass_fail(group_qs)

                table_rows.append([
                    batch_value or "-",
                    section_value or "-",
                    semester_value or "-",
                    total_students,
                    failed_students,
                    passed_students,
                    f"{percentage}%"
                ])
        else:
            grouped_keys = list(
                qs.values_list("batch", "section")
                  .distinct()
                  .order_by("batch", "section")
            )

            for batch_value, section_value in grouped_keys:
                group_qs = qs.filter(
                    batch=batch_value,
                    section=section_value
                )

                
                total_students, failed_students, passed_students, percentage = calculate_pass_fail(group_qs)

                table_rows.append([
                    batch_value or "-",
                    section_value or "-",
                    semester,
                    total_students,
                    failed_students,
                    passed_students,
                    f"{percentage}%"
                ])

        if not table_rows:
            table_rows.append([
                "-",
                "-",
                "Current Semester" if semester == "CURRENT" else semester,
                0,
                0,
                0,
                "0%"
            ])

        story.append(Paragraph(f"<b>Degree :</b> {degree_name}", styles["Normal"]))
        story.append(Paragraph(f"<b>Department :</b> {department_name}", styles["Normal"]))
        story.append(Paragraph(f"<b>Exam :</b> {exam_name}", styles["Normal"]))
        story.append(Paragraph(f"<b>Batch :</b> {batch if not is_all_value(batch) else 'All'}", styles["Normal"]))
        story.append(Paragraph(f"<b>Section :</b> {section if not is_all_value(section) else 'All'}", styles["Normal"]))
        story.append(Paragraph(f"<b>Semester :</b> {'Current Semester' if semester == 'CURRENT' else semester}", styles["Normal"]))
        story.append(Spacer(1, 15))

        data = [
            ["Batch", "Section", "Semester", "Total Students", "No of Failed", "No of Passed", "Pass Percentage"]
        ]
        data.extend(table_rows)

        table = Table(
            data,
            repeatRows=1,
            colWidths=[22 * mm, 22 * mm, 24 * mm, 30 * mm, 28 * mm, 28 * mm, 34 * mm]
        )

        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        story.append(table)

        if idx < total_groups:
            story.append(PageBreak())

    def header(canvas, doc):
        canvas.saveState()

        page_w, page_h = A4
        left = 18 * mm
        right = 18 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)

        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                candidate = os.path.join(static_root, logo_rel)
                if os.path.exists(candidate):
                    logo_path = candidate

            if not logo_path:
                for d in getattr(settings, "STATICFILES_DIRS", []):
                    candidate = os.path.join(d, logo_rel)
                    if os.path.exists(candidate):
                        logo_path = candidate
                        break

        if logo_path and os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                target_h = 18 * mm
                target_w = target_h * (iw / float(ih))

                canvas.drawImage(
                    img,
                    left,
                    page_h - (target_h + 8 * mm),
                    width=target_w,
                    height=target_h,
                    preserveAspectRatio=True,
                    mask="auto"
                )
            except Exception:
                pass

        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(page_w / 2, page_h - 10 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFont("Helvetica", 10)
        canvas.drawCentredString(page_w / 2, page_h - 15 * mm, "Rajapalayam - 626117")
        canvas.drawCentredString(page_w / 2, page_h - 19 * mm, "Affiliated to Anna University, Chennai")

        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(page_w / 2, page_h - 26 * mm, "IAT Result Analysis - Pass Percentage Report")

        canvas.line(left, page_h - 30 * mm, page_w - right, page_h - 30 * mm)

        canvas.restoreState()

    doc.build(story, onFirstPage=header, onLaterPages=header)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = "inline; filename=pass_percentage_report.pdf"
    return response

from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from openpyxl import Workbook

def get_academic_year():
    """
    Dynamically returns academic year string.
    Example:
      If current month >= June → '2025-2026'
      Else (Jan–May) → '2024-2025'
    """
    today = date.today()
    current_year = today.year
    if today.month >= 6:  # June or later
        return f"{current_year}-{current_year + 1}"
    else:  # Before June → part of previous cycle
        return f"{current_year - 1}-{current_year}"


# ==================================================================
# Honours Course Master
# Regulation -> Degree -> Department -> Year -> Semester -> Academic Year,
# then tick which of the matching courses count as Honours for that
# academic year (kept in its own HonoursCourse table, not a flag on Course,
# since the same course can be Honours in one academic year and not another).
# ==================================================================

def _hcm_safe_str(v):
    return (v or "").strip() if isinstance(v, str) else ("" if v is None else str(v).strip())


def _hcm_resolve_scope(request):
    """
    Global users (Create Global Users admin screen — same mechanism as
    Fee View / Course Enrollment Dashboard) can pick any department.
    Everyone else is locked to their own department, mirroring add_new_course.
    Returns (is_global, locked_department_or_None).
    """
    from course_management.views.faculty_control_cm import _is_global_course_user

    if _is_global_course_user(request.user):
        return True, None

    faculty = (
        general_information.objects
        .filter(faculty_id=request.user.Employee_id)
        .select_related("department")
        .first()
    )
    return False, (faculty.department if faculty else None)


@check_permission("honours_course_master")
def honours_course_master(request):
    is_global, locked_department = _hcm_resolve_scope(request)

    if not is_global and not locked_department:
        messages.error(request, "Your account is not mapped to any active department. Contact admin.")
        return redirect("home")

    if request.method == "POST":
        regulation_id = _hcm_safe_str(request.POST.get("regulation"))
        degree_id = _hcm_safe_str(request.POST.get("degree"))
        department_id = _hcm_safe_str(request.POST.get("department"))
        year = _hcm_safe_str(request.POST.get("year"))
        semester = _hcm_safe_str(request.POST.get("semester"))
        academic_year = _hcm_safe_str(request.POST.get("academic_year")) or get_academic_year()
        selected_ids = set(request.POST.getlist("honours_course[]"))

        if not is_global:
            department_id = str(locked_department.id)

        redirect_url = (
            f"{reverse('honours_course_master')}?regulation={regulation_id}&degree={degree_id}"
            f"&department={department_id}&year={year}&semester={semester}&academic_year={academic_year}"
        )

        if not (regulation_id and department_id and year and semester):
            messages.error(request, "Please select Regulation, Department, Year and Semester.")
            return redirect(redirect_url)

        courses_qs = Course.objects.filter(
            is_active=True,
            regulation_id=regulation_id,
            department_id=department_id,
            year=year,
            semester=semester,
        )
        if degree_id:
            courses_qs = courses_qs.filter(department__degree_id=degree_id)

        all_ids = {str(cid) for cid in courses_qs.values_list("id", flat=True)}
        to_mark = selected_ids & all_ids

        existing = HonoursCourse.objects.filter(course_id__in=all_ids, academic_year=academic_year)
        existing.exclude(course_id__in=to_mark).delete()

        already_marked = {str(cid) for cid in existing.values_list("course_id", flat=True)}
        for course_id in to_mark - already_marked:
            HonoursCourse.objects.create(
                course_id=course_id,
                department_id=department_id,
                regulation_id=regulation_id,
                year=year,
                semester=semester,
                academic_year=academic_year,
            )

        messages.success(request, f"Honours courses updated for {academic_year}.")
        return redirect(redirect_url)

    regulation_id = _hcm_safe_str(request.GET.get("regulation"))
    degree_id = _hcm_safe_str(request.GET.get("degree"))
    department_id = _hcm_safe_str(request.GET.get("department"))
    year = _hcm_safe_str(request.GET.get("year"))
    semester = _hcm_safe_str(request.GET.get("semester"))
    academic_year = _hcm_safe_str(request.GET.get("academic_year")) or get_academic_year()

    if not is_global:
        department_id = str(locked_department.id)
        degree_id = str(locked_department.degree_id) if locked_department.degree_id else degree_id

    regulations = Regulations.objects.all().order_by("year")

    courses = Course.objects.none()
    honours_ids = set()
    if regulation_id and department_id and year and semester:
        courses = (
            Course.objects.filter(
                is_active=True,
                regulation_id=regulation_id,
                department_id=department_id,
                year=year,
                semester=semester,
            )
            .select_related("elective")
            .order_by("course_code")
        )
        honours_ids = {
            str(cid) for cid in
            HonoursCourse.objects.filter(course__in=courses, academic_year=academic_year)
            .values_list("course_id", flat=True)
        }

    context = {
        "is_global": is_global,
        "locked_department": locked_department,
        "regulations": regulations,
        "courses": courses,
        "honours_ids": honours_ids,
        "sel_regulation": regulation_id,
        "sel_degree": degree_id,
        "sel_department": department_id,
        "sel_year": year,
        "sel_semester": semester,
        "sel_academic_year": academic_year,
        "default_academic_year": get_academic_year(),
    }
    return render(request, "course_management/admin/honours_course_master.html", context)


@check_permission("honours_course_master")
def api_honours_degrees(request):
    is_global, locked_department = _hcm_resolve_scope(request)
    if not is_global:
        if locked_department and locked_department.degree_id:
            degrees = Degree.objects.filter(id=locked_department.degree_id)
            return JsonResponse({"results": list(degrees.values("id", "degree", "degree_code"))})
        return JsonResponse({"results": []})

    regulation_id = _hcm_safe_str(request.GET.get("regulation_id"))
    qs = Course.objects.filter(is_active=True)
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    degree_ids = qs.exclude(department__degree_id__isnull=True).values_list("department__degree_id", flat=True).distinct()
    degrees = Degree.objects.filter(id__in=degree_ids, is_active=True).order_by("degree")
    return JsonResponse({"results": list(degrees.values("id", "degree", "degree_code"))})


@check_permission("honours_course_master")
def api_honours_departments(request):
    is_global, locked_department = _hcm_resolve_scope(request)
    if not is_global:
        departments = Add_Department.objects.filter(id=locked_department.id) if locked_department else Add_Department.objects.none()
        return JsonResponse({"results": list(departments.values("id", "Department", "Department_code"))})

    regulation_id = _hcm_safe_str(request.GET.get("regulation_id"))
    degree_id = _hcm_safe_str(request.GET.get("degree_id"))
    qs = Course.objects.filter(is_active=True)
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if degree_id:
        qs = qs.filter(department__degree_id=degree_id)
    dept_ids = qs.exclude(department_id__isnull=True).values_list("department_id", flat=True).distinct()
    departments = Add_Department.objects.filter(id__in=dept_ids, is_active=True).order_by("Department")
    return JsonResponse({"results": list(departments.values("id", "Department", "Department_code"))})


@check_permission("honours_course_master")
def api_honours_years(request):
    is_global, locked_department = _hcm_resolve_scope(request)
    regulation_id = _hcm_safe_str(request.GET.get("regulation_id"))
    degree_id = _hcm_safe_str(request.GET.get("degree_id"))
    department_id = _hcm_safe_str(request.GET.get("department_id"))
    if not is_global:
        department_id = str(locked_department.id) if locked_department else ""

    qs = Course.objects.filter(is_active=True)
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if degree_id:
        qs = qs.filter(department__degree_id=degree_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    years = qs.exclude(year__isnull=True).exclude(year="").values_list("year", flat=True).distinct().order_by("year")
    return JsonResponse({"results": list(years)})


@check_permission("honours_course_master")
def api_honours_semesters(request):
    is_global, locked_department = _hcm_resolve_scope(request)
    regulation_id = _hcm_safe_str(request.GET.get("regulation_id"))
    degree_id = _hcm_safe_str(request.GET.get("degree_id"))
    department_id = _hcm_safe_str(request.GET.get("department_id"))
    year = _hcm_safe_str(request.GET.get("year"))
    if not is_global:
        department_id = str(locked_department.id) if locked_department else ""

    qs = Course.objects.filter(is_active=True)
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if degree_id:
        qs = qs.filter(department__degree_id=degree_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    if year:
        qs = qs.filter(year=year)
    semesters = qs.exclude(semester__isnull=True).exclude(semester="").values_list("semester", flat=True).distinct().order_by("semester")
    return JsonResponse({"results": list(semesters)})


def get_filtered_enrollments(request):
    department_id = request.GET.get("department_id", "").strip()
    section = request.GET.get("section", "").strip()
    batch = request.GET.get("batch", "").strip()
    course_id = request.GET.get("course_id", "").strip()
    regulation_id = request.GET.get("regulation_id", "").strip()
    academic_year = request.GET.get("academic_year", "").strip()
    search = request.GET.get("search", "").strip()

    enrollments = CourseEnrollment.objects.select_related(
        "department",
        "student",
        "course",
        "regulation",
    ).filter(enroll=True)

    if department_id:
        enrollments = enrollments.filter(department_id=department_id)

    if section:
        enrollments = enrollments.filter(section=section)

    if batch:
        enrollments = enrollments.filter(batch=batch)

    if course_id:
        enrollments = enrollments.filter(course_id=course_id)

    if regulation_id:
        enrollments = enrollments.filter(regulation_id=regulation_id)

    if academic_year:
        enrollments = enrollments.filter(academic_year=academic_year)

    if search:
        enrollments = enrollments.filter(
            Q(department__Department_code__icontains=search) |
            Q(department__Department__icontains=search) |
            Q(student__reg_no__icontains=search) |
            Q(student__reg_no__icontains=search) |
            Q(section__icontains=search) |
            Q(batch__icontains=search) |
            Q(course__course_code__icontains=search) |
            Q(course__title__icontains=search) |
            Q(regulation__year__icontains=search) |
            Q(academic_year__icontains=search)
        )

    return enrollments.order_by("-id")


def all_course_enrollments(request):
    departments = Add_Department.objects.all().order_by("Department_code")

    sections = (
        CourseEnrollment.objects.filter(enroll=True)
        .exclude(section__isnull=True)
        .exclude(section__exact="")
        .values_list("section", flat=True)
        .distinct()
        .order_by("section")
    )

    batches = (
        CourseEnrollment.objects.filter(enroll=True)
        .exclude(batch__isnull=True)
        .exclude(batch__exact="")
        .values_list("batch", flat=True)
        .distinct()
        .order_by("batch")
    )

    academic_years = (
        CourseEnrollment.objects.filter(enroll=True)
        .exclude(academic_year__isnull=True)
        .exclude(academic_year__exact="")
        .values_list("academic_year", flat=True)
        .distinct()
        .order_by("academic_year")
    )

    courses = (
        CourseEnrollment.objects.filter(enroll=True, course__isnull=False)
        .values("course_id", "course__course_code", "course__title")
        .distinct()
        .order_by("course__course_code")
    )

    regulations = (
        CourseEnrollment.objects.filter(enroll=True, regulation__isnull=False)
        .values("regulation_id", "regulation__year")
        .distinct()
        .order_by("regulation__year")
    )

    bulk_batches = (
        StudentDetails.objects
        .exclude(batch__isnull=True)
        .exclude(batch__exact="")
        .values_list("batch", flat=True)
        .distinct()
        .order_by("batch")
    )

    course_years = (
        Course.objects.filter(is_active=True)
        .exclude(year__isnull=True)
        .exclude(year__exact="")
        .values_list("year", flat=True)
        .distinct()
        .order_by("year")
    )

    course_semesters = (
        Course.objects.filter(is_active=True)
        .exclude(semester__isnull=True)
        .exclude(semester__exact="")
        .values_list("semester", flat=True)
        .distinct()
        .order_by("semester")
    )

    context = {
        "departments": departments,
        "sections": sections,
        "batches": batches,
        "academic_years": academic_years,
        "courses": courses,
        "regulations": regulations,
        "bulk_batches": bulk_batches,
        "course_years": course_years,
        "course_semesters": course_semesters,
        "current_academic_year": get_academic_year(),
    }

    return render(
        request,
        "course_management/admin/all_course_enrollments.html",
        context,
    )


def all_course_enrollments_ajax(request):
    page_number = request.GET.get("page", 1)
    per_page = request.GET.get("per_page", 10)

    try:
        per_page = int(per_page)
    except Exception:
        per_page = 10

    all_enrollments = CourseEnrollment.objects.filter(enroll=True)
    filtered_enrollments = get_filtered_enrollments(request)

    paginator = Paginator(filtered_enrollments, per_page)
    page_obj = paginator.get_page(page_number)

    data = []
    for enrollment in page_obj.object_list:
        data.append({
            "department_code": getattr(enrollment.department, "Department_code", "") or getattr(enrollment.department, "code", "") or "",
            "reg_no": getattr(enrollment.student, "reg_no", "") or getattr(enrollment.student, "register_number", "") or "",
            "section": enrollment.section or "",
            "batch": enrollment.batch or "",
            "course_code": getattr(enrollment.course, "course_code", "") or "",
            "course_title": getattr(enrollment.course, "title", "") or "",
            "regulation": str(enrollment.regulation.year) if enrollment.regulation else "",
            "academic_year": enrollment.academic_year or "",
        })

    return JsonResponse({
        "data": data,
        "total_count": all_enrollments.count(),
        "filtered_count": filtered_enrollments.count(),
        "page": page_obj.number,
        "num_pages": paginator.num_pages,
        "has_previous": page_obj.has_previous(),
        "has_next": page_obj.has_next(),
        "previous_page_number": page_obj.previous_page_number() if page_obj.has_previous() else None,
        "next_page_number": page_obj.next_page_number() if page_obj.has_next() else None,
        "start_index": page_obj.start_index() if paginator.count else 0,
        "end_index": page_obj.end_index() if paginator.count else 0,
    })


def get_courses_by_department_ajax(request):
    department_id = request.GET.get("department_id", "").strip()

    courses_qs = CourseEnrollment.objects.filter(
        enroll=True,
        course__isnull=False
    )

    if department_id:
        courses_qs = courses_qs.filter(department_id=department_id)

    courses = (
        courses_qs.values("course_id", "course__course_code", "course__title")
        .distinct()
        .order_by("course__course_code")
    )

    data = []
    for course in courses:
        data.append({
            "id": course["course_id"],
            "course_code": course["course__course_code"] or "",
            "title": course["course__title"] or "",
        })

    return JsonResponse({"courses": data})


def export_course_enrollments_excel(request):
    enrollments = get_filtered_enrollments(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Course Enrollments"

    headers = [
        "department_code",
        "reg_no",
        "section",
        "batch",
        "course_code",
        "regulation",
        "academic_year",
        "year", "semester"
    ]
    ws.append(headers)

    for enrollment in enrollments:
        ws.append([
            getattr(enrollment.department, "Department_code", "") or getattr(enrollment.department, "code", "") or "",
            getattr(enrollment.student, "reg_no", "") or getattr(enrollment.student, "register_number", "") or "",
            enrollment.section or "",
            enrollment.batch or "",
            getattr(enrollment.course, "course_code", "") or "",
            str(enrollment.regulation.year) if enrollment.regulation else "",
            enrollment.academic_year or "",
            enrollment.year or "",
            enrollment.semester or "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="all_course_enrollments.xlsx"'
    wb.save(response)
    return response


from django.db import transaction
from django.http import JsonResponse

def sync_course_enrollments(request):
    enrollments = CourseEnrollment.objects.select_related(
        "course",
        "department",
    ).filter(
        enroll=True,
        course__isnull=False,
    )

    updated_count = 0
    skipped_count = 0

    with transaction.atomic():
        for enrollment in enrollments:
            course = enrollment.course

            # Check course belongs to same department as enrollment
            if (
                not course
                or not enrollment.department
                or course.department_id != enrollment.department_id
            ):
                skipped_count += 1
                continue

            update_fields = []

            if enrollment.year != course.year:
                enrollment.year = course.year
                update_fields.append("year")

            if enrollment.semester != course.semester:
                enrollment.semester = course.semester
                update_fields.append("semester")

            if not enrollment.academic_year:
                enrollment.academic_year = get_academic_year()
                update_fields.append("academic_year")

            if update_fields:
                enrollment.save(update_fields=update_fields)
                updated_count += 1
            else:
                skipped_count += 1

    return JsonResponse({
        "status": "success",
        "message": "Course enrollments synced successfully",
        "updated_count": updated_count,
        "skipped_count": skipped_count,
    })


def bulk_enroll_course_enrollments(request):
    if request.method != "POST":
        return JsonResponse({
            "success": False,
            "message": "POST request required.",
        }, status=405)

    department_id = (request.POST.get("department_id") or "").strip()
    batch = (request.POST.get("batch") or "").strip()
    course_year = (request.POST.get("course_year") or "").strip()
    semester = (request.POST.get("semester") or "").strip()

    if not department_id or not batch or not course_year or not semester:
        return JsonResponse({
            "success": False,
            "message": "Department, batch, year, and semester are required.",
        }, status=400)

    department = Add_Department.objects.filter(id=department_id).first()
    if not department:
        return JsonResponse({
            "success": False,
            "message": "Selected department was not found.",
        }, status=404)

    students = list(
        StudentDetails.objects.filter(
            department=department,
            batch=batch,
        )
        .filter(Q(is_active=True) | Q(is_active__isnull=True))
        .exclude(reg_no__isnull=True)
        .exclude(reg_no__exact="")
        .filter(Q(is_break_of_study=False) | Q(is_break_of_study__isnull=True))
        .order_by("section", "reg_no")
    )

    if not students:
        return JsonResponse({
            "success": False,
            "message": "No active students found for the selected department and batch.",
        }, status=404)

    regulation_years = sorted({
        str(student.regulation).strip()
        for student in students
        if str(student.regulation or "").strip()
    })
    regulations = {
        reg.year: reg
        for reg in Regulations.objects.filter(year__in=regulation_years)
    }

    assignment_rows = (
        AssignSubjectFaculty.objects
        .filter(
            department=department,
            batch=batch,
            is_active=True,
            faculty__isnull=False,
            course__is_active=True,
            course__year=course_year,
            course__semester=semester,
        )
        .select_related("course", "faculty", "regulation")
        .order_by("regulation_id", "section", "course__course_code", "id")
    )

    assignments_by_context = {}
    for assignment in assignment_rows:
        if not assignment.course_id or not assignment.regulation_id:
            continue

        context_key = (
            assignment.regulation_id,
            (assignment.section or "").strip(),
        )
        course_map = assignments_by_context.setdefault(context_key, {})
        course_map.setdefault(assignment.course_id, assignment)

    if not assignments_by_context:
        return JsonResponse({
            "success": False,
            "message": "No active subject-faculty assignments found for the selected department, batch, year, and semester.",
        }, status=404)

    academic_year = get_academic_year()
    today = timezone.localdate()

    created_count = 0
    updated_count = 0
    reactivated_count = 0
    skipped_students = 0
    missing_regulation_count = 0
    missing_assignment_count = 0
    matched_students = 0
    enrolled_course_rows = 0
    missing_assignment_contexts = set()

    with transaction.atomic():
        for student in students:
            regulation_year = str(student.regulation or "").strip()
            regulation = regulations.get(regulation_year)
            if not regulation:
                missing_regulation_count += 1
                skipped_students += 1
                continue

            section = (student.section or "").strip()
            course_map = (
                assignments_by_context.get((regulation.id, section))
                or assignments_by_context.get((regulation.id, ""))
            )

            if not course_map:
                missing_assignment_count += 1
                skipped_students += 1
                missing_assignment_contexts.add(
                    f"Regulation {regulation.year} / Section {section or '-'}"
                )
                continue

            matched_students += 1

            for assignment in course_map.values():
                enrollment = (
                    CourseEnrollment.objects.filter(
                        student=student,
                        course=assignment.course,
                        batch=batch,
                        section=section,
                        regulation=regulation,
                    )
                    .filter(
                        Q(academic_year__isnull=True) | Q(academic_year__exact="") |
                        Q(year__isnull=True) | Q(year__exact="") |
                        Q(semester__isnull=True) | Q(semester__exact="")
                    )
                    .first()
                )

                if enrollment is None:
                    enrollment, created = CourseEnrollment.objects.get_or_create(
                        student=student,
                        course=assignment.course,
                        batch=batch,
                        section=section,
                        regulation=regulation,
                        academic_year=academic_year,
                        year=course_year,
                        semester=semester,
                        defaults={
                            "department": department,
                            "faculty": assignment.faculty,
                            "enrollment_date": today,
                            "enroll": True,
                        },
                    )
                    if created:
                        created_count += 1
                        enrolled_course_rows += 1
                        continue

                previous_enroll_state = bool(enrollment.enroll)
                has_changes = False

                if enrollment.department_id != department.id:
                    enrollment.department = department
                    has_changes = True

                if enrollment.faculty_id != assignment.faculty_id:
                    enrollment.faculty = assignment.faculty
                    has_changes = True

                if enrollment.batch != batch:
                    enrollment.batch = batch
                    has_changes = True

                if enrollment.section != section:
                    enrollment.section = section
                    has_changes = True

                if enrollment.regulation_id != regulation.id:
                    enrollment.regulation = regulation
                    has_changes = True

                if enrollment.academic_year != academic_year:
                    enrollment.academic_year = academic_year
                    has_changes = True

                if str(enrollment.year or "") != course_year:
                    enrollment.year = course_year
                    has_changes = True

                if str(enrollment.semester or "") != semester:
                    enrollment.semester = semester
                    has_changes = True

                if not enrollment.enrollment_date or not previous_enroll_state:
                    enrollment.enrollment_date = today
                    has_changes = True

                if not enrollment.enroll:
                    enrollment.enroll = True
                    has_changes = True

                if has_changes:
                    enrollment.save()
                    updated_count += 1
                    if not previous_enroll_state and enrollment.enroll:
                        reactivated_count += 1

                enrolled_course_rows += 1

    summary_bits = [
        f"{matched_students} student(s) matched",
        f"{enrolled_course_rows} course enrollment row(s) processed",
        f"{created_count} created",
        f"{updated_count} updated",
    ]

    if reactivated_count:
        summary_bits.append(f"{reactivated_count} reactivated")
    if missing_regulation_count:
        summary_bits.append(f"{missing_regulation_count} missing regulation")
    if missing_assignment_count:
        summary_bits.append(f"{missing_assignment_count} without assignments")

    message = "Bulk enrollment completed: " + ", ".join(summary_bits) + "."

    if missing_assignment_contexts:
        sample_contexts = ", ".join(sorted(list(missing_assignment_contexts))[:3])
        message += f" Missing assignment contexts: {sample_contexts}."

    return JsonResponse({
        "success": True,
        "message": message,
        "matched_students": matched_students,
        "processed_enrollments": enrolled_course_rows,
        "created_count": created_count,
        "updated_count": updated_count,
        "reactivated_count": reactivated_count,
        "missing_regulation_count": missing_regulation_count,
        "missing_assignment_count": missing_assignment_count,
        "skipped_students": skipped_students,
    })




def attendance_percentage(request):
    slabs = AttendancePercentageMaster.objects.all().order_by("percentage_from")

    if request.method == "POST":
        action = request.POST.get("action")
        current_faculty = general_information.objects.filter(
            faculty_id=getattr(request.user, "Employee_id", None)
        ).first()

        if action == "save":
            row_ids = request.POST.getlist("row_id[]")
            percentage_from_list = request.POST.getlist("percentage_from[]")
            percentage_to_list = request.POST.getlist("percentage_to[]")
            attendance_mark_list = request.POST.getlist("attendance_mark[]")

            saved_any = False

            for i in range(len(percentage_from_list)):
                row_id = row_ids[i].strip() if i < len(row_ids) and row_ids[i] else ""
                percentage_from_raw = percentage_from_list[i].strip() if i < len(percentage_from_list) else ""
                percentage_to_raw = percentage_to_list[i].strip() if i < len(percentage_to_list) else ""
                attendance_mark_raw = attendance_mark_list[i].strip() if i < len(attendance_mark_list) else ""

                if not any([percentage_from_raw, percentage_to_raw, attendance_mark_raw]):
                    continue

                try:
                    percentage_from_val = Decimal(percentage_from_raw)
                    percentage_to_val = Decimal(percentage_to_raw)
                    attendance_mark_val = Decimal(attendance_mark_raw)
                except InvalidOperation:
                    messages.error(request, f"Invalid number in row {i + 1}.")
                    return redirect("attendance_percentage")

                if percentage_from_val > percentage_to_val:
                    messages.error(request, f"Row {i + 1}: 'Percentage From' cannot be greater than 'Percentage To'.")
                    return redirect("attendance_percentage")

                if row_id:
                    obj = AttendancePercentageMaster.objects.filter(id=row_id).first()
                    if not obj:
                        continue
                    obj.percentage_from = percentage_from_val
                    obj.percentage_to = percentage_to_val
                    obj.attendance_mark = attendance_mark_val
                    obj.updated_by = current_faculty
                    obj.save()
                else:
                    AttendancePercentageMaster.objects.create(
                        percentage_from=percentage_from_val,
                        percentage_to=percentage_to_val,
                        attendance_mark=attendance_mark_val,
                        created_by=current_faculty,
                        updated_by=current_faculty,
                    )

                saved_any = True

            if saved_any:
                messages.success(request, "Attendance percentage mark mapping saved successfully.")
            else:
                messages.warning(request, "No data entered.")

            return redirect("attendance_percentage")

        elif action == "delete":
            row_id = request.POST.get("row_id")
            AttendancePercentageMaster.objects.filter(id=row_id).delete()
            messages.success(request, "Row deleted successfully.")
            return redirect("attendance_percentage")

    context = {
        "slabs": slabs,
    }
    return render(request, "course_management/admin/attendance_percentage.html", context)
  

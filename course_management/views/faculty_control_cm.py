import re
from django.shortcuts import render, redirect
from django.shortcuts import redirect
from course_management.decorators import course_management
from user_accounts.decorators import faculty_login_required, check_permission
from student_management.models import *
from examination_management.models import *
from course_management.models import *
from django.core.paginator import Paginator
from django.utils import timezone


# @faculty_login_required
@course_management
def cm_home(request):
    request.session['current_page'] = 'cm_home'

    return redirect('home')

@course_management
@check_permission("hello")
def hello(request):
    return render(request, "hello.html")


from django.shortcuts import render, redirect
from course_management.models import   StudentLeaveOdApplication
from user_accounts.models import USER, Department
from student_management.models import Student_cgpa
from django.contrib import messages
from user_accounts.models import StudentDetails, Add_Department


DEPT_DICT = {
    "CSE": "CS",
    "IT": "IT",
    "EEE": "EE",
    "CSBS": "CB",
    "AIML": "AL",
    "MECH": "ME",
    "ECE": "EC",
    "CIVIL": "CE",
    "AD": "AD"
}

DEPT_NAME_MAP = {
    "CSE": "COMPUTER SCIENCE AND ENGINEERING",
    "IT": "INFORMATION TECHNOLOGY",
    "EEE": "ELECTRICAL AND ELECTRONICS ENGINEERING",
    "CSBS": "COMPUTER SCIENCE AND BUSINESS SYSTEMS",
    "AIML": "ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING",
    "MECH": "MECHANICAL ENGINEERING",
    "ECE": "ELECTRONICS AND COMMUNICATION ENGINEERING",
    "CIVIL": "CIVIL ENGINEERING",
    "AD": "ARTIFICIAL INTELLIGENCE AND DATA SCIENCE"
}



DEPT_NAME_TO_CODE = {
    "COMPUTER SCIENCE AND ENGINEERING": "CSE",
    "INFORMATION TECHNOLOGY": "IT",
    "ELECTRICAL AND ELECTRONICS ENGINEERING": "EEE",
    "COMPUTER SCIENCE AND BUSINESS SYSTEMS": "CSBS",
    "ARTIFICIAL INTELLIGENCE AND MACHINE LEARNING": "AIML",
    "MECHANICAL ENGINEERING": "MECH",
    "ELECTRONICS AND COMMUNICATION ENGINEERING": "ECE",
    "CIVIL ENGINEERING": "CIVIL",
    "ARTIFICIAL INTELLIGENCE AND DATA SCIENCE": "AD"
}

@check_permission("assign_mentor")
def assign_mentor(request):
    department = request.user.Department.Department

    emp_id = request.user.Employee_id
    faculty = general_information.objects.get(faculty_id=emp_id)
    department_obj = faculty.department
    student_department = faculty.department
    # 
    faculty_list = general_information.objects.all().order_by('name')
    batches = StudentDetails.objects.filter(department=student_department)\
                                     .values_list('batch', flat=True)\
                                     .distinct()

    # Selected batch
    selected_batch = request.GET.get('batch')

    # Filter students
    students = StudentDetails.objects.filter(department=student_department)
    if selected_batch:
        students = students.filter(batch=selected_batch)

    # Handle form submission
    if request.method == "POST":
        mentor_id = request.POST.get('mentor')
        student_regnos = request.POST.getlist('students')

        if not mentor_id:
            messages.error(request, "Please select a faculty to assign.")
        elif not student_regnos:
            messages.error(request, "Please select at least one student.")
        else:
            try:
                mentor_obj = general_information.objects.get(id=mentor_id)
            except general_information.DoesNotExist:
                messages.error(request, "Selected faculty does not exist.")
                return redirect('assign_mentor')

            assigned_count = 0
            for reg_no in student_regnos:
                student_obj = students.filter(reg_no=reg_no).first()
                if not student_obj:
                    messages.warning(request, f"No student found for Reg No {reg_no}")
                    continue

                # Assign mentor as ForeignKey
                student_obj.mentor = mentor_obj
                student_obj.save(update_fields=["mentor"])
                assigned_count += 1

            if assigned_count:
                messages.success(request, f"Mentor assigned to {assigned_count} student(s) successfully.")
            else:
                messages.error(request, "No students were assigned due to missing data.")

        return redirect('assign_mentor')

    # Existing mentor assignments
    assigned_dict = {
        student.reg_no: student.mentor.id
        for student in students if student.mentor
    }

    # Count stats
    total_students = students.count()
    assigned_count = len(assigned_dict)
    available_count = total_students - assigned_count

    # Get mentor details for display
    mentor_ids = set(assigned_dict.values())
    mentor_details = {}
    if mentor_ids:
        mentors = general_information.objects.filter(id__in=mentor_ids).values(
            'id', 'name', 'faculty_id', 'designation'
        )
        for mentor in mentors:
            mentor_details[mentor['id']] = {
                'username': mentor['name'],
                'employee_id': mentor['faculty_id'],
                'role': mentor['designation'] or 'Faculty'
            }

    context = {
        "faculty_list": faculty_list,
        "batches": batches,
        "students": students,
        "assigned_dict": assigned_dict,
        "mentor_details": mentor_details,
        "selected_batch": selected_batch,
        "department_obj": department_obj,
        "total_students": total_students,
        "assigned_count": assigned_count,
        "available_count": available_count,
    }

    return render(request, "course_management/hod/assign_mentor.html", context)


@check_permission("assign_class_advisor")
def assign_class_advisor(request):
    user_department = getattr(request.user, "Department", None)
    faculty = (
        general_information.objects
        .filter(faculty_id=request.user.Employee_id)
        .select_related("department", "department__degree")
        .first()
    )
    department = faculty.department if faculty else None

    if department is None and user_department:
        department_code = getattr(user_department, "Department_code", None)
        department_name = getattr(user_department, "Department", None)
        if department_code:
            department = Add_Department.objects.filter(
                Department_code=department_code
            ).first()
        if department is None and department_name:
            department = Add_Department.objects.filter(
                Department=department_name
            ).first()

    if department is None:
        messages.error(
            request,
            "No faculty profile or academic department mapping was found for your login."
        )
    
    # faculty_list = general_information.objects.filter(
    #     department=department,
    #     # approval='Approved'
    # ).order_by('name')
    
    faculty_list = general_information.objects.all().order_by('name')

    # Batch, regulation, section lists
    student_scope = StudentDetails.objects.filter(department=department) if department else StudentDetails.objects.none()
    batches = student_scope.values_list("batch", flat=True).distinct()
    regulations = Regulations.objects.values_list("year", flat=True).distinct()
    sections = student_scope.values_list("section", flat=True).distinct()

    # Selected filters
    selected_batch = request.GET.get("batch")
    selected_section = request.GET.get("section")

    # Filter students
    students = student_scope
    if selected_batch:
        students = students.filter(batch=selected_batch)
    if selected_section:
        students = students.filter(section=selected_section)

    # Handle POST: Assign/Remove CA
    if request.method == "POST":
        ca_id = request.POST.get("ca")  # ID from general_information
        student_regnos = request.POST.getlist("students")
        remove_ca = request.POST.get("remove_ca")

        # Remove CA
        if remove_ca:
            # remove CA for a single student reg_no in this department
            StudentDetails.objects.filter(
                reg_no=remove_ca, department=department
            ).update(ca=None)
            messages.success(request, f"CA assignment removed for student {remove_ca}.")
            return redirect("assign_class_advisor")

        # Assign CA
        if not ca_id:
            messages.error(request, "Please select a Class Advisor.")
        else:
            try:
                ca_obj = general_information.objects.get(id=ca_id)
            except general_information.DoesNotExist:
                messages.error(request, "Selected faculty does not exist.")
                return redirect("assign_class_advisor")

            if not student_regnos:
                student_regnos = list(students.values_list("reg_no", flat=True))

            assigned_count = 0
            for reg_no in student_regnos:
                student_obj = students.filter(reg_no=reg_no).first()
                if not student_obj:
                    continue
                student_obj.ca = ca_obj  # Assign ForeignKey
                student_obj.save(update_fields=["ca"])
                assigned_count += 1

            if assigned_count:
                messages.success(
                    request,
                    f"CA assigned to {assigned_count} student(s) in Batch {selected_batch or 'ALL'}, "
                    f"Section {selected_section or 'ALL'} successfully."
                )
            else:
                messages.error(request, "No students were assigned (check data).")

        return redirect("assign_class_advisor")

    # Existing assignments
    assigned_dict = {
        student.reg_no: student.ca_id
        for student in students
        if student.ca_id
    }

    total_students = students.count()
    assigned_count = len(assigned_dict)
    available_count = total_students - assigned_count

    # Fetch CA details
    ca_ids = set(assigned_dict.values())
    ca_details = {}
    if ca_ids:
        cas = general_information.objects.filter(id__in=ca_ids).values(
            'id', 'name', 'faculty_id', 'designation'
        )
        for ca in cas:
            ca_details[ca['id']] = {
                'username': ca['name'],
                'employee_id': ca['faculty_id'],
                'role': ca['designation'] or 'Faculty'
            }

    context = {
        "faculty_list": faculty_list,
        "batches": batches,
        "sections": sections,
        "students": students,
        "assigned_dict": assigned_dict,
        "ca_details": ca_details,
        "selected_batch": selected_batch,
        "selected_section": selected_section,
        "department_obj": department,
        "total_students": total_students,
        "assigned_count": assigned_count,
        "available_count": available_count,
    }

    return render(request, "course_management/hod/assign_ca.html", context)



from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect, render
from faculty_management.models import Vision, general_information
from django.db import transaction
from django.urls import reverse
from django.core.paginator import Paginator

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.contrib import messages
from django.db.models import Q, Prefetch, Case, When, Value, IntegerField, Max


def get_academic_year():
    """
    Dynamically returns academic year string.
    Example:
      If current month >= June → '2025-2026'
      Else (Jan–May) → '2024-2025'
    """
    today = date.today()
    current_year = today.year
    if today.month >= 6:  # June or later
        return f"{current_year}-{current_year + 1}"
    else:  # Before June → part of previous cycle
        return f"{current_year - 1}-{current_year}"
    

def _create_assign_subject_faculty(**kwargs):
    next_id = (AssignSubjectFaculty.objects.aggregate(max_id=Max("id"))["max_id"] or 0) + 1
    return AssignSubjectFaculty.objects.create(id=next_id, **kwargs)


def _safe_int(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _course_hour_info_from_course_hours(course_hours):
    cfg = getattr(course_hours, "hour_config", None) if course_hours else None
    lecture_hours = _safe_int(getattr(cfg, "lecture_hours", None)) if cfg else _safe_int(getattr(course_hours, "leture_npwk", None))
    laboratory_hours = _safe_int(getattr(cfg, "laboratory_hours", None)) if cfg else _safe_int(getattr(course_hours, "laboratory_npwk", None))

    if lecture_hours > 0 and laboratory_hours > 0:
        course_type = "Integrated"
    elif laboratory_hours > 0:
        course_type = "Laboratory"
    elif lecture_hours > 0:
        course_type = "Theory"
    else:
        course_type = "Other"

    return {
        "lecture_hours": lecture_hours,
        "laboratory_hours": laboratory_hours,
        "needs_skilled_faculty": laboratory_hours > 0,
        "course_type": course_type,
    }


def _course_hour_info(course):
    course_hours = (
        CourseHours.objects
        .select_related("hour_config")
        .filter(course=course)
        .first()
    )
    return _course_hour_info_from_course_hours(course_hours)


@check_permission("assign_subject_faculty")
def assign_subject_faculty(request):
    try:
        schedule_state = get_subject_allocation_window_status()
        user_dept = getattr(request.user, "Department", None)
        faculty = general_information.objects.get(faculty_id=request.user.Employee_id)
        department = faculty.department
        current_academic_year = get_academic_year()
        if not department:
            messages.error(request, "Your department record was not found.")
            return redirect("assign_subject_faculty")

        selected_regulation = request.GET.get("regulation", "").strip()
        selected_semester = request.GET.get("semester", "").strip()
        # ---------------- WILLINGNESS FILTERS ----------------
        w_year       = (request.GET.get("w_year") or current_academic_year).strip()
        w_regulation = (request.GET.get("w_regulation") or "").strip()
        w_semester   = (request.GET.get("w_semester") or "").strip()
        w_batch      = (request.GET.get("w_batch") or "").strip()
        w_section    = (request.GET.get("w_section") or "").strip()

        # ---------------- ASSIGNMENT FILTERS ----------------
        a_year       = (request.GET.get("a_year") or current_academic_year).strip()
        a_regulation = (request.GET.get("a_regulation") or "").strip()
        a_semester   = (request.GET.get("a_semester") or "").strip()
        a_batch      = (request.GET.get("a_batch") or "").strip()
        a_section    = (request.GET.get("a_section") or "").strip()
        faculties = general_information.objects.all()
        batches = StudentDetails.objects.values_list("batch", flat=True)\
                    .exclude(batch__isnull=True).exclude(batch__exact="")\
                    .distinct().order_by("batch")
        sections = SectionMaster.objects.values_list("section", flat=True)\
                    .exclude(section__isnull=True).exclude(section__exact="")\
                    .distinct().order_by("section")
        regulations = Regulations.objects.all().order_by("-year")

        base_courses = Course.objects.filter(department=department, is_active=True)
        course_qs = base_courses

        if selected_regulation:
            course_qs = course_qs.filter(regulation_id=int(selected_regulation))

        semesters = list(course_qs.order_by("semester").values_list("semester", flat=True).distinct()) \
                    if selected_regulation else []
        filter_semesters = list(
            base_courses.exclude(semester__isnull=True)
            .exclude(semester__exact="")
            .order_by("semester")
            .values_list("semester", flat=True)
            .distinct()
        )

        if selected_semester:
            course_qs = course_qs.filter(semester=int(selected_semester))

        courses_for_form = list(course_qs) if (selected_regulation and selected_semester) else []
        course_hours_by_course_id = {
            ch.course_id: ch
            for ch in CourseHours.objects.select_related("hour_config").filter(
                course_id__in=[course.id for course in courses_for_form]
            )
        }
        for course_obj in courses_for_form:
            hour_info = _course_hour_info_from_course_hours(course_hours_by_course_id.get(course_obj.id))
            course_obj.needs_skilled_faculty = hour_info["needs_skilled_faculty"]
            course_obj.course_type_label = hour_info["course_type"]
            course_obj.lecture_hours_value = hour_info["lecture_hours"]
            course_obj.laboratory_hours_value = hour_info["laboratory_hours"]

        if request.method == "POST" and request.POST.get("toggle_assignment"):
            assign_id = request.POST.get("toggle_assignment")
            assign = get_object_or_404(AssignSubjectFaculty, id=assign_id, department=department)
            assign.is_active = not assign.is_active
            assign.save()
            messages.success(request, "Status updated.")
            return redirect(request.get_full_path())

        if request.method == "POST" and not request.POST.get("toggle_assignment"):
            if not schedule_state["can_act"]:
                messages.error(request, schedule_state["status_message"])
                return redirect(request.get_full_path())
            course_id = request.POST.get("course")
            faculty_id = request.POST.get("faculty")
            skilled_faculty_id = request.POST.get("skilled_faculty")
            batch = request.POST.get("batch")
            section = request.POST.get("section")
            reason_choice = (request.POST.get("reason") or "").strip()
            custom_reason = (request.POST.get("custom_reason") or "").strip()
            is_active = bool(request.POST.get("is_active"))

            if not selected_regulation or not selected_semester:
                messages.error(request, "Please select regulation and semester before assigning.")
                return redirect(request.get_full_path())

            if not all([course_id, faculty_id, batch, section, reason_choice]):
                messages.error(request, "Please complete all required fields before assigning.")
                return redirect(request.get_full_path())

            if reason_choice == "OTHER" and not custom_reason:
                messages.error(request, "Please enter the custom reason.")
                return redirect(request.get_full_path())

            regulation = Regulations.objects.get(id=int(selected_regulation))
            course = Course.objects.get(id=int(course_id))
            faculty = general_information.objects.get(faculty_id=int(faculty_id))
            hour_info = _course_hour_info(course)
            skilled_faculty = None

            if hour_info["needs_skilled_faculty"]:
                if not skilled_faculty_id:
                    messages.error(
                        request,
                        "Please select a skilled faculty for integrated and laboratory courses."
                    )
                    return redirect(request.get_full_path())
                skilled_faculty = general_information.objects.filter(faculty_id=skilled_faculty_id).first()
                if not skilled_faculty:
                    messages.error(request, "Selected skilled faculty does not exist.")
                    return redirect(request.get_full_path())
                if skilled_faculty.id == faculty.id:
                    messages.error(request, "Main faculty and skilled faculty must be different.")
                    return redirect(request.get_full_path())

            reason = custom_reason if reason_choice == "OTHER" and custom_reason else reason_choice

            existing_assignment = AssignSubjectFaculty.objects.filter(
                regulation=regulation,
                course=course,
                batch=batch,
                section=section,
                department=department,
                academic_year=current_academic_year,
            ).order_by("-id").first()

            if existing_assignment:
                existing_assignment.faculty = faculty
                existing_assignment.skilled_faculty = skilled_faculty
                existing_assignment.is_active = is_active
                existing_assignment.reason = reason
                existing_assignment.academic_year = current_academic_year
                existing_assignment.save()
                messages.success(request, "Assignment updated successfully.")
            else:
                with transaction.atomic():
                    _create_assign_subject_faculty(
                        regulation=regulation,
                        course=course,
                        batch=batch,
                        section=section,
                        faculty=faculty,
                        skilled_faculty=skilled_faculty,
                        department=department,
                        is_active=is_active,
                        reason=reason,
                        academic_year=current_academic_year,
                    )
                messages.success(request, "Assignment created successfully.")

            return redirect(request.get_full_path())

        willingness_qs = FacultySubjectWillingness.objects.filter(
            department=department
        ).select_related("faculty", "course", "regulation")

        if w_year:
            willingness_qs = willingness_qs.filter(academic_year=w_year)

        if w_regulation:
            willingness_qs = willingness_qs.filter(regulation_id=w_regulation)

        if w_semester:
            willingness_qs = willingness_qs.filter(semester=w_semester)

        if w_batch:
            willingness_qs = willingness_qs.filter(batch=w_batch)

        if w_section:
            willingness_qs = willingness_qs.filter(section=w_section)
        
        willingness_qs = willingness_qs.order_by("-created_at", "-id")

        w_page = request.GET.get("w_page", 1)
        w_paginator = Paginator(willingness_qs, 15)
        willingness = w_paginator.get_page(w_page)

        assignments_qs = AssignSubjectFaculty.objects.filter(
            department=department
        ).select_related("faculty", "skilled_faculty", "course", "regulation")

        current_assignment_lookup_qs = (
            AssignSubjectFaculty.objects.filter(
                department=department,
                academic_year=current_academic_year,
            )
            .select_related("faculty", "skilled_faculty", "course", "regulation")
            .order_by("-id")
        )

        if a_year:
            assignments_qs = assignments_qs.filter(academic_year=a_year)

        if a_regulation:
            assignments_qs = assignments_qs.filter(regulation_id=a_regulation)

        if a_semester:
            assignments_qs = assignments_qs.filter(course__semester=a_semester)

        if a_batch:
            assignments_qs = assignments_qs.filter(batch=a_batch)

        if a_section:
            assignments_qs = assignments_qs.filter(section=a_section)

        assignments_qs = assignments_qs.order_by(
            Case(
                When(is_active=True, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            ),
            "-id",
        )

        a_page = request.GET.get("a_page", 1)
        a_paginator = Paginator(assignments_qs, 15)
        assignments = a_paginator.get_page(a_page)

        for w in willingness:
            assigned = AssignSubjectFaculty.objects.filter(
                course=w.course, batch=w.batch, regulation=w.regulation,
                department=department, section=w.section
            ).first()
            w.assigned_section = assigned.section if assigned else None
            hour_info = _course_hour_info(w.course) if w.course_id else {
                "needs_skilled_faculty": False,
                "course_type": "Other",
            }
            w.needs_skilled_faculty = hour_info["needs_skilled_faculty"]
            w.course_type_label = hour_info["course_type"]

        normalized_reason_options = []
        raw_choices = getattr(AssignSubjectFaculty, "REASON_CHOICES", [])
        for opt in raw_choices:
            value, label = (opt if isinstance(opt, (list, tuple)) else (opt, opt))
            if label.lower() in ("other", "others"):
                value, label = "OTHER", "Other"
            normalized_reason_options.append({"value": value, "label": label})

        def build_query_string(overrides=None, remove_keys=None):
            params = request.GET.copy()
            for key in remove_keys or []:
                params.pop(key, None)
            for key, value in (overrides or {}).items():
                if value in (None, ""):
                    params.pop(key, None)
                else:
                    params[key] = str(value)
            return params.urlencode()

        current_query_string = request.GET.urlencode()
        willingness_page_query = build_query_string(
            overrides={"a_page": assignments.number},
            remove_keys=["w_page"],
        )
        assignments_page_query = build_query_string(
            overrides={"w_page": willingness.number},
            remove_keys=["a_page"],
        )

        academic_years = sorted(
            {
                year for year in AssignSubjectFaculty.objects.exclude(academic_year__isnull=True)
                .exclude(academic_year__exact="")
                .values_list("academic_year", flat=True)
            }
            |
            {
                year for year in FacultySubjectWillingness.objects.exclude(academic_year__isnull=True)
                .exclude(academic_year__exact="")
                .values_list("academic_year", flat=True)
            }
            |
            {current_academic_year}
        )

        assignment_lookup = {}
        for assignment in current_assignment_lookup_qs:
            key = "||".join([
                str(assignment.course_id or ""),
                str(assignment.regulation_id or ""),
                str(assignment.batch or ""),
                str(assignment.section or ""),
                str(assignment.academic_year or ""),
            ])
            if key not in assignment_lookup:
                assignment_lookup[key] = {
                    "faculty_name": getattr(assignment.faculty, "name", "") or "-",
                    "faculty_id": str(getattr(assignment.faculty, "faculty_id", "") or "-"),
                    "skilled_faculty_name": getattr(assignment.skilled_faculty, "name", "") or "-",
                    "skilled_faculty_id": str(getattr(assignment.skilled_faculty, "faculty_id", "") or "-"),
                    "course_code": getattr(assignment.course, "course_code", "") or "-",
                    "course_title": getattr(assignment.course, "title", "") or "-",
                    "regulation": getattr(assignment.regulation, "year", "") or "-",
                    "semester": getattr(assignment.course, "semester", "") or "-",
                    "batch": assignment.batch or "-",
                    "section": assignment.section or "-",
                    "academic_year": assignment.academic_year or "-",
                    "reason": assignment.reason or "-",
                    "is_active": bool(assignment.is_active),
                }

        return render(request, "course_management/hod/assign_subject_faculty.html", {
            "courses": courses_for_form,
            "faculties": faculties,
            "regulations": regulations,
            "department": department,
            "batches": batches,
            "sections": sections,
            "semesters": semesters,
            "filter_semesters": filter_semesters,
            "selected_semester": selected_semester,
            "selected_regulation": selected_regulation,
            "assignments": assignments,
            "willingness": willingness,
            "reason_options": normalized_reason_options,
            "academic_years": academic_years,
            "current_academic_year": current_academic_year,
            "current_query_string": current_query_string,
            "willingness_page_query": willingness_page_query,
            "assignments_page_query": assignments_page_query,
            "assignment_lookup": assignment_lookup,

            "w_year": w_year,
            "w_regulation": w_regulation,
            "w_semester": w_semester,
            "w_batch": w_batch,
            "w_section": w_section,

            "a_year": a_year,
            "a_regulation": a_regulation,
            "a_semester": a_semester,
            "a_batch": a_batch,
            "a_section": a_section,
            "schedule_state": schedule_state,
        })

    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
        return redirect("faculty_dashboard")



@check_permission("assign_subject_faculty")
def review_subject_willingness(request):
    schedule_state = get_subject_allocation_window_status()
    if not schedule_state["can_act"]:
        messages.error(request, schedule_state["status_message"])
        return redirect("assign_subject_faculty")

    action = (request.POST.get("action") or "").strip().lower()
    row_id = request.POST.get("id", "").strip()

    return_query = (request.POST.get("return_query") or "").strip()

    def back_redirect():
        base_url = reverse("assign_subject_faculty")
        return redirect(f"{base_url}?{return_query}" if return_query else base_url)

    if action not in {"approve", "reject"} or not row_id.isdigit():
        messages.error(request, "Invalid action.")
        return back_redirect()

    user_dept = getattr(request.user, "Department", None)
    department = Add_Department.objects.filter(
        Department__iexact=getattr(user_dept, "Department", ""), is_active=True
    ).first()
    if not department:
        messages.error(request, "Department not found.")
        return redirect("dashboard")

    w = get_object_or_404(FacultySubjectWillingness, id=int(row_id), department=department)

    if w.status != "Pending":
        messages.info(request, "Only pending entries can be reviewed.")
        return back_redirect()

    if action == "reject":
        w.status = "Rejected"
        w.save(update_fields=["status", "updated_at"])
        messages.success(request, "Willingness rejected.")
        return back_redirect()

    # Approve
    reason_choice = request.POST.get("reason", "").strip()
    custom_reason = request.POST.get("custom_reason", "").strip()
    allotted_section = request.POST.get("allotted_section", "").strip()
    skilled_faculty_id = (request.POST.get("skilled_faculty") or "").strip()

    if not reason_choice:
        messages.error(request, "Please select a reason.")
        return back_redirect()

    if reason_choice == "OTHER" and not custom_reason:
        messages.error(request, "Custom reason is required.")
        return back_redirect()

    if not allotted_section:
        allotted_section = w.section
    else:
        # Validate only if user selected a section
        if not SectionMaster.objects.filter(section=allotted_section).exists():
            messages.error(request, "Invalid section selected.")
            return back_redirect()

    reason = custom_reason if reason_choice == "OTHER" else reason_choice
    current_academic_year = get_academic_year()
    hour_info = _course_hour_info(w.course) if w.course_id else {"needs_skilled_faculty": False}
    skilled_faculty = None

    if hour_info["needs_skilled_faculty"]:
        if not skilled_faculty_id:
            messages.error(
                request,
                "Please select a skilled faculty for integrated and laboratory courses."
            )
            return back_redirect()
        skilled_faculty = general_information.objects.filter(faculty_id=skilled_faculty_id).first()
        if not skilled_faculty:
            messages.error(request, "Selected skilled faculty does not exist.")
            return back_redirect()
        if w.faculty_id and skilled_faculty.id == w.faculty_id:
            messages.error(request, "Main faculty and skilled faculty must be different.")
            return back_redirect()

    with transaction.atomic():
        w.status = "Approved"
        w.reason = reason
        w.save(update_fields=["status", "reason", "updated_at"])

        # Reuse the existing class-slot assignment when present.
        existing_assignment = (
            AssignSubjectFaculty.objects.filter(
                course=w.course,
                batch=w.batch,
                section=allotted_section,
                regulation=w.regulation,
                department=department,
                academic_year=current_academic_year,
            )
            .order_by("-id")
            .first()
        )

        if existing_assignment:
            existing_assignment.faculty = w.faculty
            existing_assignment.skilled_faculty = skilled_faculty
            existing_assignment.is_active = True
            existing_assignment.reason = reason
            existing_assignment.academic_year = current_academic_year
            existing_assignment.save()
        else:
            _create_assign_subject_faculty(
                course=w.course,
                batch=w.batch,
                section=allotted_section,
                regulation=w.regulation,
                department=department,
                faculty=w.faculty,
                skilled_faculty=skilled_faculty,
                is_active=True,
                reason=reason,
                academic_year=current_academic_year,
            )

    messages.success(request, f"Approved. Assigned to Section {allotted_section}.")
    return back_redirect()


from django.http import HttpResponse
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import BaseDocTemplate, PageTemplate, Frame, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from django.conf import settings
from django.contrib.staticfiles import finders
import os
from datetime import datetime
from django.shortcuts import redirect
from django.contrib import messages

# Import your models
from course_management.models import AssignSubjectFaculty
from user_accounts.models import Add_Department  # Make sure this import is correct


# =========================================================
# ✅ PROFESSIONAL COLOR PALETTE (Same as willingness PDF)
# =========================================================
PRIMARY_BLUE = colors.HexColor("#0f2f57")
SECONDARY_BLUE = colors.HexColor("#1a4b8c")
ACCENT_RED = colors.HexColor("#b91c1c")
DARK_GRAY = colors.HexColor("#1f2937")
MEDIUM_GRAY = colors.HexColor("#4b5563")
LIGHT_GRAY = colors.HexColor("#9ca3af")
BG_GRAY = colors.HexColor("#f8fafc")
BORDER_GRAY = colors.HexColor("#e5e7eb")


def safe_str(value):
    """Safely convert any value to string."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


# =========================================================
# ✅ PROFESSIONAL HEADER/FOOTER DESIGN (Same as willingness PDF)
# =========================================================
def _professional_header_footer(canvas, doc, title="Subject Allocation Report"):
    """Professional header with logo and institute details"""
    canvas.saveState()
    
    page_w, page_h = A4
    left = doc.leftMargin
    right = page_w - doc.rightMargin
    top_y = page_h - 8 * mm
    
    # Resolve Logo
    logo_rel = "images/ritlogo.png"
    logo_path = finders.find(logo_rel)
    if not logo_path:
        static_root = getattr(settings, "STATIC_ROOT", "")
        if static_root:
            cand = os.path.join(static_root, logo_rel)
            if os.path.exists(cand):
                logo_path = cand
    
    # Draw Logo
    if logo_path and os.path.exists(logo_path):
        try:
            canvas.drawImage(
                ImageReader(logo_path),
                left,
                top_y - 20 * mm,
                width=32 * mm,
                height=20 * mm,
                preserveAspectRatio=True,
                mask="auto"
            )
        except Exception:
            pass
    
    # Header Text
    center_x = (left + right) / 2
    
    canvas.setFillColor(PRIMARY_BLUE)
    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawCentredString(center_x, top_y - 6 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")
    
    canvas.setFillColor(ACCENT_RED)
    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawCentredString(center_x, top_y - 13 * mm, "An Autonomous Institution")
    
    canvas.setFillColor(MEDIUM_GRAY)
    canvas.setFont("Helvetica", 8.5)
    canvas.drawCentredString(center_x, top_y - 18.5 * mm,
                             "Approved by AICTE, New Delhi ")
    canvas.setFont("Helvetica", 8)
    canvas.drawCentredString(center_x, top_y - 23.5 * mm,
                             "Accredited by NAAC & ISO 9001:2015 Certified Institution")
    canvas.drawCentredString(center_x, top_y - 28 * mm,
                             "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")
    
    
    # Footer with page number
    footer_y = 22 * mm
    canvas.setStrokeColor(BORDER_GRAY)
    canvas.setLineWidth(0.8)
    canvas.line(left, footer_y + 8 * mm, right, footer_y + 8 * mm)
    
    canvas.setFillColor(LIGHT_GRAY)
    canvas.setFont("Helvetica", 8)
    
    gen_time = datetime.now().strftime('%d %b %Y, %I:%M %p')
    canvas.drawString(left, footer_y, f"Generated: {gen_time}")
    
    if title:
        canvas.drawCentredString(center_x, footer_y, title)
    
    canvas.drawRightString(right, footer_y, f"Page {doc.page}")
    
    canvas.restoreState()


# =========================================================
# ✅ PROFESSIONAL TABLE HELPER (Same as willingness PDF)
# =========================================================
def create_table(data, col_widths, header_bg=None, zebra=True):
    """Create professional tables"""
    if header_bg is None:
        header_bg = PRIMARY_BLUE
    
    t = Table(data, repeatRows=1, colWidths=col_widths)
    
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9.5),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
        
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
        ("FONTSIZE", (0, 1), (-1, -1), 9.5),
        ("VALIGN", (0, 1), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])
    
    if zebra:
        style.add("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY])
    
    t.setStyle(style)
    return t


# =========================================================
# ✅ MAIN PDF GENERATION FUNCTION
# =========================================================
def assign_subject_faculty_pdf(request):
    """Professional Subject Allocation PDF matching willingness PDF style"""
    
    # Get user's department safely
    try:
        faculty_obj = general_information.objects.get(faculty_id=request.user.Employee_id)
        department = faculty_obj.department
    except Exception:
        department = None
    
    if not department:
        messages.error(request, "Your department record was not found.")
        return redirect("assign_subject_faculty")
    
    # Filter parameters (default to current academic year to match the page)
    current_academic_year = get_academic_year()
    a_year = (request.GET.get("a_year") or current_academic_year).strip()
    a_regulation = (request.GET.get("a_regulation") or "").strip()
    a_semester = (request.GET.get("a_semester") or "").strip()
    a_batch = (request.GET.get("a_batch") or "").strip()
    a_section = (request.GET.get("a_section") or "").strip()

    assignments = AssignSubjectFaculty.objects.filter(
        department=department,
        is_active=True
    ).select_related("faculty", "skilled_faculty", "course", "regulation").order_by("-id")

    def _valid(p):
        return p is not None and str(p).strip() != "" and str(p).strip().lower() != "none"

    if _valid(a_year):
        assignments = assignments.filter(academic_year=a_year)
    if _valid(a_regulation):
        try:
            assignments = assignments.filter(regulation_id=int(a_regulation))
        except Exception:
            pass
    if _valid(a_semester):
        assignments = assignments.filter(course__semester=a_semester)
    if _valid(a_batch):
        assignments = assignments.filter(batch=a_batch)
    if _valid(a_section):
        assignments = assignments.filter(section=a_section)
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    filename = f"Subject_Allocation_Report_{datetime.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    
    try:
        # =========================================================
        # ✅ DOCUMENT SETUP (Same as willingness PDF)
        # =========================================================
        doc = BaseDocTemplate(
            response,
            pagesize=A4,
            title=f"Subject Allocation Report - {safe_str(department.Department)}",
            leftMargin=16 * mm,
            rightMargin=16 * mm,
            topMargin=18 * mm,
            bottomMargin=22 * mm,
            showBoundary=0
        )
        
        frame = Frame(
            doc.leftMargin,
            doc.bottomMargin + 10 * mm,
            doc.width,
            doc.height - 38 * mm + 8 * mm,
            leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
            id="normal"
        )
        doc.addPageTemplates([
            PageTemplate(
                id="All", 
                frames=[frame], 
                onPage=lambda canvas, doc: _professional_header_footer(
                    canvas, doc, 
                    "Subject Allocation Report"
                )
            )
        ])
        
        # =========================================================
        # ✅ ENHANCED STYLES (Same as willingness PDF)
        # =========================================================
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            "title_style",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=PRIMARY_BLUE,
            alignment=TA_CENTER,
            spaceAfter=10,
            spaceBefore=5,
            fontName="Helvetica-Bold",
            leading=20
        )
        
        section_style = ParagraphStyle(
            "section_style",
            parent=styles["Heading2"],
            fontSize=13,
            textColor=PRIMARY_BLUE,
            alignment=TA_LEFT,
            spaceBefore=18,
            spaceAfter=10,
            fontName="Helvetica-Bold",
            borderWidth=0,
            leftIndent=0,
            leading=16
        )
        
        info_label_style = ParagraphStyle(
            "info_label_style",
            parent=styles["Normal"],
            fontSize=10,
            textColor=MEDIUM_GRAY,
            alignment=TA_LEFT,
            fontName="Helvetica-Bold",
            leading=12
        )
        
        info_value_style = ParagraphStyle(
            "info_value_style",
            parent=styles["Normal"],
            fontSize=10,
            textColor=DARK_GRAY,
            alignment=TA_LEFT,
            fontName="Helvetica",
            leading=12
        )
        
        table_header_style = ParagraphStyle(
            "table_header_style",
            parent=styles["Normal"],
            fontSize=9.5,
            textColor=colors.white,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            leading=11
        )
        
        table_cell_style = ParagraphStyle(
            "table_cell_style",
            parent=styles["Normal"],
            fontSize=9.5,
            textColor=DARK_GRAY,
            leading=11,
            alignment=TA_LEFT,
            wordWrap="CJK",
            fontName="Helvetica"
        )
        
        table_cell_center_style = ParagraphStyle(
            "table_cell_center_style",
            parent=table_cell_style,
            alignment=TA_CENTER,
            fontName="Helvetica"
        )
        
        empty_state_style = ParagraphStyle(
            "empty_state_style",
            parent=styles["Heading3"],
            fontSize=12,
            textColor=MEDIUM_GRAY,
            alignment=TA_CENTER,
            spaceBefore=20,
            spaceAfter=20
        )
        
        # =========================================================
        # ✅ BUILD DOCUMENT ELEMENTS
        # =========================================================
        elements = []
        
        # Add initial spacer
        elements.append(Spacer(1, 5 * mm))
        
        # =========================================================
        # ✅ GROUP BY SECTION with Professional Layout
        # =========================================================
        grouped = {}
        for a in assignments:
            academic_year = safe_str(a.academic_year) if a.academic_year else "-"
            reg = a.regulation.year if a.regulation else "Unknown Regulation"
            year = a.course.year if a.course else "-"
            sem = a.course.semester if a.course else "-"
            batch = safe_str(a.batch) if a.batch else "-"
            section = safe_str(a.section) if a.section else "-"
            grouped.setdefault(academic_year, {}).setdefault(reg, {}).setdefault(year, {}).setdefault(sem, {}).setdefault(batch, {}).setdefault(section, []).append(a)
        
        has_content = False
        
        for academic_year, regulations in grouped.items():
            for reg, years in regulations.items():
                for year, semesters in years.items():
                    for sem, batches in semesters.items():
                        for batch, sections in batches.items():
                            for section, records in sections.items():
                                has_content = True
                            
                            # Main Title for each section
                            elements.append(Paragraph(
                                f"SUBJECT ALLOCATION REPORT", 
                                title_style
                            ))
                            
                            # Department name
                            elements.append(Paragraph(
                                f"Department: {safe_str(department.Department)}",
                                ParagraphStyle(
                                    "dept_style",
                                    parent=styles["Normal"],
                                    fontSize=12,
                                    textColor=SECONDARY_BLUE,
                                    alignment=TA_CENTER,
                                    spaceAfter=8,
                                    fontName="Helvetica-Bold"
                                )
                            ))
                            elements.append(Spacer(1, 10 * mm))
                            
                            # Section Header
                            elements.append(Paragraph(
                                f"Section: {section}",
                                section_style
                            ))
                            
                            # Academic Details Table
                            academic_data = [
                                [
                                    Paragraph("Academic Year:", info_label_style),
                                    Paragraph(safe_str(academic_year), info_value_style),
                                    Paragraph("Regulation:", info_label_style),
                                    Paragraph(safe_str(reg), info_value_style)
                                ],
                                [
                                    Paragraph("Year:", info_label_style),
                                    Paragraph(safe_str(year), info_value_style),
                                    Paragraph("Semester:", info_label_style),
                                    Paragraph(safe_str(sem), info_value_style)
                                ],
                                [
                                    Paragraph("Batch:", info_label_style),
                                    Paragraph(safe_str(batch), info_value_style),
                                    Paragraph("Section:", info_label_style),
                                    Paragraph(safe_str(section), info_value_style)
                                ],
                                [
                                    Paragraph("Total Subjects:", info_label_style),
                                    Paragraph(str(len(records)), info_value_style),
                                    Paragraph("", info_label_style),
                                    Paragraph("", info_value_style)
                                ]
                            ]
                            
                            academic_table = Table(
                                academic_data,
                                colWidths=[30*mm, 35*mm, 30*mm, doc.width - (30*mm + 35*mm + 30*mm)]
                            )
                            academic_table.setStyle(TableStyle([
                                ("BACKGROUND", (0, 0), (-1, -1), BG_GRAY),
                                ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
                                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                                ("TOPPADDING", (0, 0), (-1, -1), 6),
                                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                            ]))
                            
                            elements.append(academic_table)
                            elements.append(Spacer(1, 12 * mm))
                            
                            # Main Allocation Table
                            table_data = [[
                                Paragraph("SL No", table_header_style),
                                Paragraph("Main Faculty", table_header_style),
                                Paragraph("Skilled Faculty", table_header_style),
                                Paragraph("Course Code", table_header_style),
                                Paragraph("Course Title", table_header_style),
                                Paragraph("Allocation Reason", table_header_style),
                            ]]
                            
                            for i, a in enumerate(records, start=1):
                                faculty_name = safe_str(a.faculty.name) if a.faculty else "-"
                                skilled_faculty_name = safe_str(a.skilled_faculty.name) if a.skilled_faculty else "-"
                                course_code = safe_str(a.course.course_code) if a.course else "-"
                                course_title = safe_str(a.course.title) if a.course else "-"
                                reason = safe_str(a.reason) or "Regular Allocation"
                                
                                table_data.append([
                                    Paragraph(str(i), table_cell_center_style),
                                    Paragraph(faculty_name, table_cell_style),
                                    Paragraph(skilled_faculty_name, table_cell_style),
                                    Paragraph(course_code, table_cell_style),
                                    Paragraph(course_title, table_cell_style),
                                    Paragraph(reason, table_cell_style),
                                ])
                            
                            # Create table with professional styling
                            table_widths = [
                                15*mm,    # SL No
                                35*mm,    # Main Faculty
                                35*mm,    # Skilled Faculty
                                25*mm,    # Course Code
                                45*mm,    # Course Title
                                doc.width - (15+35+35+25+45)*mm  # Allocation Reason
                            ]
                            
                            allocation_table = create_table(
                                data=table_data,
                                col_widths=table_widths,
                                header_bg=SECONDARY_BLUE,
                                zebra=True
                            )
                            
                            elements.append(allocation_table)
                            elements.append(Spacer(1, 15 * mm))
                            
                            # Force new page for next section if not the last one
                            elements.append(PageBreak())
        
        # =========================================================
        # ✅ HANDLE EMPTY DATA
        # =========================================================
        if not has_content:
            elements.append(Spacer(1, 20 * mm))
            elements.append(Paragraph(
                "No subject allocation data found for the selected filters.",
                empty_state_style
            ))
        
        # =========================================================
        # ✅ BUILD DOCUMENT
        # =========================================================
        doc.build(elements)
        
    except Exception as e:
        # print(f"PDF generation error: {str(e)}")
        import traceback
        traceback.print_exc()
        return HttpResponse(
            f"<h3>PDF Generation Failed</h3><p>Error: {str(e)}</p>"
            "<p>Please contact system administrator.</p>",
            status=500
        )
    
    return response





from django.http import HttpResponse
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import BaseDocTemplate, PageTemplate, Frame, Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from django.conf import settings
from django.contrib.staticfiles import finders
import os
from datetime import datetime
from django.shortcuts import redirect
from django.contrib import messages

# Import your models
from faculty_management.models import general_information
from course_management.models import AssignSubjectFaculty,FacultySubjectWillingness


# =========================================================
# ✅ PROFESSIONAL COLOR PALETTE
# =========================================================
PRIMARY_BLUE = colors.HexColor("#0f2f57")
SECONDARY_BLUE = colors.HexColor("#1a4b8c")
ACCENT_RED = colors.HexColor("#b91c1c")
DARK_GRAY = colors.HexColor("#1f2937")
MEDIUM_GRAY = colors.HexColor("#4b5563")
LIGHT_GRAY = colors.HexColor("#9ca3af")
BG_GRAY = colors.HexColor("#f8fafc")
BORDER_GRAY = colors.HexColor("#e5e7eb")


def safe_str(value):
    """Safely convert any value to string."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


# =========================================================
# ✅ PROFESSIONAL HEADER/FOOTER DESIGN
# =========================================================
def _professional_header_footer(canvas, doc, title="Subject Willingness Report"):
    """Professional header with logo and institute details"""
    canvas.saveState()
    
    page_w, page_h = A4
    left = doc.leftMargin
    right = page_w - doc.rightMargin
    top_y = page_h - 8 * mm
    
    # Resolve Logo
    logo_rel = "images/ritlogo.png"
    logo_path = finders.find(logo_rel)
    if not logo_path:
        static_root = getattr(settings, "STATIC_ROOT", "")
        if static_root:
            cand = os.path.join(static_root, logo_rel)
            if os.path.exists(cand):
                logo_path = cand
    
    # Draw Logo
    if logo_path and os.path.exists(logo_path):
        try:
            canvas.drawImage(
                ImageReader(logo_path),
                left,
                top_y - 20 * mm,
                width=32 * mm,
                height=20 * mm,
                preserveAspectRatio=True,
                mask="auto"
            )
        except Exception:
            pass
    
    # Header Text
    center_x = (left + right) / 2
    
    canvas.setFillColor(PRIMARY_BLUE)
    canvas.setFont("Helvetica-Bold", 14)
    canvas.drawCentredString(center_x, top_y - 6 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")
    
    canvas.setFillColor(ACCENT_RED)
    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawCentredString(center_x, top_y - 13 * mm, "(An Autonomous Institution)")
    
    canvas.setFillColor(MEDIUM_GRAY)
    canvas.setFont("Helvetica", 8.5)
    canvas.drawCentredString(center_x, top_y - 18.5 * mm,
                             "Rajapalayam - 626 117")
    canvas.setFont("Helvetica", 8)
    canvas.drawCentredString(center_x, top_y - 23.5 * mm,
                             "Subject Allocation Report")
    # canvas.drawCentredString(center_x, top_y - 28 * mm,
    #                          "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")
    
    # Decorative line under header
    canvas.setStrokeColor(PRIMARY_BLUE)
    canvas.setLineWidth(1.2)
    canvas.line(left + 10 * mm, page_h - 38 * mm, right - 10 * mm, page_h - 38 * mm)
    
    # Footer with page number
    footer_y = 22 * mm
    canvas.setStrokeColor(BORDER_GRAY)
    canvas.setLineWidth(0.8)
    canvas.line(left, footer_y + 8 * mm, right, footer_y + 8 * mm)
    
    canvas.setFillColor(LIGHT_GRAY)
    canvas.setFont("Helvetica", 8)
    
    gen_time = datetime.now().strftime('%d %b %Y, %I:%M %p')
    canvas.drawString(left, footer_y, f"Generated: {gen_time}")
    
    if title:
        canvas.drawCentredString(center_x, footer_y, title)
    
    canvas.drawRightString(right, footer_y, f"Page {doc.page}")
    
    canvas.restoreState()


# =========================================================
# ✅ PROFESSIONAL TABLE HELPER
# =========================================================
def create_table(data, col_widths, header_bg=None, zebra=True):
    """Create professional tables"""
    if header_bg is None:
        header_bg = PRIMARY_BLUE
    
    t = Table(data, repeatRows=1, colWidths=col_widths)
    
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9.5),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
        
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
        ("FONTSIZE", (0, 1), (-1, -1), 9.5),
        ("VALIGN", (0, 1), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])
    
    if zebra:
        style.add("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY])
    
    t.setStyle(style)
    return t


# =========================================================
# ✅ MAIN PDF GENERATION FUNCTION
# =========================================================





def my_willingness_pdf(request):
    try:
        employee_id = getattr(request.user, "Employee_id", None)
        faculty = general_information.objects.filter(
            faculty_id=employee_id
        ).select_related("department").first()
    except Exception:
        faculty = None

    if not faculty:
        messages.error(request, "Faculty profile not found.")
        return redirect("subject_willingness")

    try:
        qs = (
            FacultySubjectWillingness.objects
            .filter(faculty=faculty)
            .select_related("course", "regulation")
            .order_by("-created_at")
        )

        qs, selected_filters = apply_willingness_filters(qs, request)

    except Exception:
        messages.error(request, "Error retrieving data.")
        return redirect("subject_willingness")

    response = HttpResponse(content_type="application/pdf")
    filename = f"Faculty_Willingness_Report_{datetime.now().strftime('%Y%m%d')}.pdf"
    response["Content-Disposition"] = f'inline; filename="{filename}"'

    try:
        doc = BaseDocTemplate(
            response,
            pagesize=A4,
            title=f"Faculty Subject Willingness - {safe_str(faculty.name)}",
            leftMargin=16 * mm,
            rightMargin=16 * mm,
            topMargin=18 * mm,
            bottomMargin=22 * mm,
            showBoundary=0
        )

        frame = Frame(
            doc.leftMargin,
            doc.bottomMargin + 10 * mm,
            doc.width,
            doc.height - 38 * mm + 8 * mm,
            leftPadding=0,
            rightPadding=0,
            topPadding=0,
            bottomPadding=0,
            id="normal"
        )

        doc.addPageTemplates([
            PageTemplate(
                id="All",
                frames=[frame],
                onPage=lambda canvas, doc: _professional_header_footer(
                    canvas,
                    doc,
                    "Subject Willingness Report"
                )
            )
        ])

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "title_style",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=PRIMARY_BLUE,
            alignment=TA_CENTER,
            spaceAfter=10,
            spaceBefore=5,
            fontName="Helvetica-Bold",
            leading=20
        )

        section_style = ParagraphStyle(
            "section_style",
            parent=styles["Heading2"],
            fontSize=13,
            textColor=PRIMARY_BLUE,
            alignment=TA_LEFT,
            spaceBefore=18,
            spaceAfter=10,
            fontName="Helvetica-Bold",
            leading=16
        )

        info_label_style = ParagraphStyle(
            "info_label_style",
            parent=styles["Normal"],
            fontSize=10,
            textColor=MEDIUM_GRAY,
            alignment=TA_LEFT,
            fontName="Helvetica-Bold",
            leading=12
        )

        info_value_style = ParagraphStyle(
            "info_value_style",
            parent=styles["Normal"],
            fontSize=10,
            textColor=DARK_GRAY,
            alignment=TA_LEFT,
            fontName="Helvetica",
            leading=12
        )

        table_header_style = ParagraphStyle(
            "table_header_style",
            parent=styles["Normal"],
            fontSize=8.5,
            textColor=colors.white,
            alignment=TA_CENTER,
            fontName="Helvetica-Bold",
            leading=10
        )

        table_cell_style = ParagraphStyle(
            "table_cell_style",
            parent=styles["Normal"],
            fontSize=8.5,
            textColor=DARK_GRAY,
            leading=10,
            alignment=TA_LEFT,
            wordWrap="CJK",
            fontName="Helvetica"
        )

        table_cell_center_style = ParagraphStyle(
            "table_cell_center_style",
            parent=table_cell_style,
            alignment=TA_CENTER
        )

        summary_style = ParagraphStyle(
            "summary_style",
            parent=styles["Normal"],
            fontSize=11,
            textColor=SECONDARY_BLUE,
            alignment=TA_CENTER,
            spaceAfter=10,
            fontName="Helvetica-Bold"
        )

        empty_state_style = ParagraphStyle(
            "empty_state_style",
            parent=styles["Heading3"],
            fontSize=12,
            textColor=MEDIUM_GRAY,
            alignment=TA_CENTER,
            spaceBefore=20,
            spaceAfter=20
        )

        elements = []
        elements.append(Spacer(1, 5 * mm))
        elements.append(Paragraph("FACULTY SUBJECT WILLINGNESS REPORT", title_style))
        elements.append(Spacer(1, 8 * mm))

        elements.append(Paragraph("FACULTY PROFILE", section_style))

        fac_id = safe_str(faculty.faculty_id) if faculty.faculty_id else ""

        designation = ""
        if getattr(faculty, "present_designation", None):
            designation = safe_str(faculty.present_designation)
        elif getattr(faculty, "designation", None):
            designation = safe_str(faculty.designation)

        dept_name = ""
        if getattr(faculty, "department", None):
            try:
                dept_name = safe_str(faculty.department.Department)
            except Exception:
                dept_name = safe_str(getattr(faculty, "department", ""))

        faculty_name = safe_str(faculty.name) if faculty.name else ""

        faculty_info_data = [
            [
                Paragraph("Faculty Name", info_label_style),
                Paragraph(faculty_name, info_value_style),
                Paragraph("Faculty ID", info_label_style),
                Paragraph(fac_id, info_value_style),
            ],
            [
                Paragraph("Designation", info_label_style),
                Paragraph(designation, info_value_style),
                Paragraph("Department", info_label_style),
                Paragraph(dept_name, info_value_style),
            ],
            [
                Paragraph("Generated On", info_label_style),
                Paragraph(datetime.now().strftime("%d-%b-%Y"), info_value_style),
                Paragraph("Total Records", info_label_style),
                Paragraph(str(qs.count()), info_value_style),
            ],
        ]

        faculty_table = Table(
            faculty_info_data,
            colWidths=[
                25 * mm,
                60 * mm,
                25 * mm,
                doc.width - (25 * mm + 60 * mm + 25 * mm)
            ]
        )

        faculty_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), BG_GRAY),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))

        elements.append(faculty_table)
        elements.append(Spacer(1, 10 * mm))

        active_filters = []
        for key, value in selected_filters.items():
            if value:
                active_filters.append(f"{key.replace('_', ' ').title()}: {safe_str(value)}")

        if active_filters:
            elements.append(Paragraph("APPLIED FILTERS", section_style))
            elements.append(Paragraph(", ".join(active_filters), info_value_style))
            elements.append(Spacer(1, 8 * mm))

        if qs.exists():
            total_courses = qs.count()
            elements.append(Paragraph(
                f"Total Courses Submitted: {total_courses}",
                summary_style
            ))
            elements.append(Spacer(1, 8 * mm))

            if total_courses > 10:
                elements.append(PageBreak())

            table_data = [[
                Paragraph("Course<br/>Code / Title", table_header_style),
                Paragraph("Regulation", table_header_style),
                Paragraph("Year /<br/>Semester", table_header_style),
                Paragraph("Section /<br/>Batch", table_header_style),
                Paragraph("Times<br/>Handled", table_header_style),
                Paragraph("Pass<br/>%", table_header_style),
                Paragraph("Status", table_header_style),
                Paragraph("Reason", table_header_style),
            ]]

            for w in qs:
                course_code = safe_str(w.course.course_code) if w.course and w.course.course_code else "-"
                course_title = safe_str(w.course.title) if w.course and w.course.title else ""
                course_txt = f"<b>{course_code}</b><br/>{course_title}"

                reg_txt = "-"
                if getattr(w, "regulation", None):
                    reg_txt = safe_str(getattr(w.regulation, "year", w.regulation))

                year_txt = safe_str(w.year) if w.year else "-"
                sem_txt = safe_str(w.semester) if w.semester else "-"
                yearsem = f"{year_txt}<br/>{sem_txt}"

                section_txt = safe_str(w.section) if w.section else "-"
                batch_txt = safe_str(w.batch) if w.batch else "-"
                secbatch = f"{section_txt}<br/>{batch_txt}"

                handled_txt = safe_str(w.No_of_time_handled) if w.No_of_time_handled else "-"

                pass_val = safe_str(w.pass_percentage_obtained) if w.pass_percentage_obtained else "-"
                pass_percent = f"{pass_val}%" if pass_val != "-" else "-"

                status_txt = safe_str(w.status) if w.status else "-"
                reason_txt = safe_str(w.reason) if w.reason else "Not specified"

                table_data.append([
                    Paragraph(course_txt, table_cell_style),
                    Paragraph(reg_txt, table_cell_center_style),
                    Paragraph(yearsem, table_cell_center_style),
                    Paragraph(secbatch, table_cell_center_style),
                    Paragraph(handled_txt, table_cell_center_style),
                    Paragraph(pass_percent, table_cell_center_style),
                    Paragraph(status_txt, table_cell_center_style),
                    Paragraph(reason_txt, table_cell_style),
                ])

            table_widths = [
                36 * mm,
                20 * mm,
                18 * mm,
                22 * mm,
                16 * mm,
                15 * mm,
                18 * mm,
                doc.width - (36 + 20 + 18 + 22 + 16 + 15 + 18) * mm,
            ]

            willingness_table = create_table(
                data=table_data,
                col_widths=table_widths,
                header_bg=SECONDARY_BLUE,
                zebra=True
            )

            elements.append(willingness_table)

        else:
            elements.append(Spacer(1, 20 * mm))
            elements.append(Paragraph(
                "No subject willingness entries found for the selected filters.",
                empty_state_style
            ))

        doc.build(elements)

    except Exception as e:
        import traceback
        traceback.print_exc()
        return HttpResponse(
            f"<h3>PDF Generation Failed</h3><p>Error: {str(e)}</p>"
            "<p>Please contact system administrator.</p>",
            status=500
        )

    return response

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from user_accounts.models import USER  , StudentDetails
from course_management.models import StudentLeaveOdApplication
from faculty_management.models import general_information
from collections import defaultdict
from django.db.models import Q
from user_accounts.decorators import check_permission, is_super_user, no_cache
from student_management.models import *


@no_cache
@check_permission("approve_leave_od")
def approve_leave_od(request):

    # Resolve the approver by the EXACT logged-in USER row (id), not by
    # Employee_id. A single person can hold multiple role accounts sharing the
    # same Employee_id (e.g. Faculty + Hostel Warden). The approver chain stores
    # the specific role account's id in approver_id_id, and login/switch_user set
    # request.user to that exact row, so filtering by Employee_id + .first() could
    # pick the wrong account and hide pending approvals.
    approver_user = (
        USER.objects
        .using("rit_approval_system")
        .filter(
            id=request.user.id,
            is_active=True
        )
        .first()
    )

    if not approver_user:
        messages.error(request, "Approver not found.")
        return redirect("home")

    approver_user_id = approver_user.id

    request_data = request.POST if request.method == "POST" else request.GET

    q = (request_data.get("q") or "").strip()
    level_filter = (request_data.get("level") or "").strip()
    application_type_filter = (request_data.get("application_type") or "").strip()
    study_year_filter = (request_data.get("study_year") or "").strip()
    tab = (request_data.get("tab") or "pending").strip().lower()
    page_number = request_data.get("page", 1)

    if tab not in ["pending", "approved", "rejected"]:
        tab = "pending"

    per_page = 50

    my_rows = Student_LeaveApproversData.objects.filter(
        approver_id_id=approver_user_id
    )

    eligible_pending_ids = []

    pending_rows = (
        my_rows
        .filter(status=Student_LeaveApproversData.Status.PENDING)
        .values("id", "leave_application_id", "approver_level")
    )

    for row in pending_rows:

        lower_pending = (
            Student_LeaveApproversData.objects
            .filter(
                leave_application_id=row["leave_application_id"],
                approver_level__lt=row["approver_level"],
            )
            .exclude(status=Student_LeaveApproversData.Status.APPROVED)
            .exists()
        )

        if not lower_pending:
            eligible_pending_ids.append(row["id"])

    pending_base_qs = Student_LeaveApproversData.objects.filter(
        id__in=eligible_pending_ids
    )

    approved_base_qs = my_rows.filter(
        status=Student_LeaveApproversData.Status.APPROVED
    )

    rejected_base_qs = my_rows.filter(
        status=Student_LeaveApproversData.Status.REJECTED
    )

    pending_count = pending_base_qs.count()
    approved_count = approved_base_qs.count()
    rejected_count = rejected_base_qs.count()

    if tab == "approved":
        approvals_qs = approved_base_qs
    elif tab == "rejected":
        approvals_qs = rejected_base_qs
    else:
        approvals_qs = pending_base_qs

    approvals_qs = (
    approvals_qs
    .select_related(
        "leave_application",
        "leave_application__student",
        "leave_application__department",
        "creator_id",
    )
    .defer("approver_id")
    .order_by(
        "approver_level",
        "-leave_application__created_at",
        "-id"
    )
)

    if q:
        approvals_qs = approvals_qs.filter(
            Q(leave_application__student__name__icontains=q) |
            Q(leave_application__student__reg_no__icontains=q) |
            Q(leave_application__application_type__icontains=q) |
            Q(leave_application__reason__icontains=q) |
            Q(leave_application__from_date__icontains=q) |
            Q(leave_application__to_date__icontains=q) |
            Q(leave_application__status__icontains=q)
        )

    if level_filter:
        try:
            approvals_qs = approvals_qs.filter(
                approver_level=int(level_filter)
            )
        except ValueError:
            pass

    if application_type_filter:
        approvals_qs = approvals_qs.filter(
            leave_application__application_type=application_type_filter
        )

    if study_year_filter:
        approvals_qs = approvals_qs.filter(
            leave_application__study_year=study_year_filter
        )

    available_levels = (
        my_rows
        .values_list("approver_level", flat=True)
        .distinct()
        .order_by("approver_level")
    )

    available_study_years = (
        my_rows
        .values_list("leave_application__study_year", flat=True)
        .distinct()
        .order_by("leave_application__study_year")
    )

    if request.method == "POST":

        action = request.POST.get("action")
        selected_ids = request.POST.getlist("selected_ids")

        if action == "approve_all":

            all_pending = (
                Student_LeaveApproversData.objects
                .filter(
                    id__in=eligible_pending_ids,
                    approver_id_id=approver_user_id,
                    status=Student_LeaveApproversData.Status.PENDING,
                )
                .values("id", "leave_application_id", "approver_level")
            )

            approved_total = 0

            with transaction.atomic():

                for row in all_pending:

                    lower_pending = (
                        Student_LeaveApproversData.objects
                        .filter(
                            leave_application_id=row["leave_application_id"],
                            approver_level__lt=row["approver_level"],
                        )
                        .exclude(status=Student_LeaveApproversData.Status.APPROVED)
                        .exists()
                    )

                    if lower_pending:
                        continue

                    Student_LeaveApproversData.objects.filter(
                        id=row["id"]
                    ).update(
                        status=Student_LeaveApproversData.Status.APPROVED,
                        approved_date=timezone.now(),
                    )

                    # OR semantics for a level: the faculty level may hold two
                    # approvers (mentor + CA). A single approval satisfies the
                    # level, so auto-approve the co-approver(s) still pending at
                    # the same level for this application.
                    Student_LeaveApproversData.objects.filter(
                        leave_application_id=row["leave_application_id"],
                        approver_level=row["approver_level"],
                        status=Student_LeaveApproversData.Status.PENDING,
                    ).exclude(
                        id=row["id"]
                    ).update(
                        status=Student_LeaveApproversData.Status.APPROVED,
                        approved_date=timezone.now(),
                    )

                    all_approved = not Student_LeaveApproversData.objects.filter(
                        leave_application_id=row["leave_application_id"]
                    ).exclude(
                        status=Student_LeaveApproversData.Status.APPROVED
                    ).exists()

                    if all_approved:
                        StudentLeaveOdApplication.objects.filter(
                            id=row["leave_application_id"]
                        ).update(
                            status=StudentLeaveOdApplication.Status.APPROVED
                        )

                    approved_total += 1

            messages.success(
                request,
                f"{approved_total} application(s) approved successfully."
            )
            return redirect(f"{reverse('approve_leave_od')}?tab=pending")

        if not selected_ids:
            messages.warning(request, "No applications selected.")
            return redirect(f"{reverse('approve_leave_od')}?tab=pending")

        selected_rows = (
            Student_LeaveApproversData.objects
            .filter(
                id__in=selected_ids,
                approver_id_id=approver_user_id,
            )
            .values("id", "leave_application_id", "approver_level", "status")
        )

        if action == "approve":

            with transaction.atomic():

                for row in selected_rows:

                    if row["status"] != Student_LeaveApproversData.Status.PENDING:
                        continue

                    lower_pending = (
                        Student_LeaveApproversData.objects
                        .filter(
                            leave_application_id=row["leave_application_id"],
                            approver_level__lt=row["approver_level"],
                        )
                        .exclude(status=Student_LeaveApproversData.Status.APPROVED)
                        .exists()
                    )

                    if lower_pending:
                        continue

                    Student_LeaveApproversData.objects.filter(
                        id=row["id"]
                    ).update(
                        status=Student_LeaveApproversData.Status.APPROVED,
                        approved_date=timezone.now(),
                    )

                    # OR semantics for a level: the faculty level may hold two
                    # approvers (mentor + CA). A single approval satisfies the
                    # level, so auto-approve the co-approver(s) still pending at
                    # the same level for this application.
                    Student_LeaveApproversData.objects.filter(
                        leave_application_id=row["leave_application_id"],
                        approver_level=row["approver_level"],
                        status=Student_LeaveApproversData.Status.PENDING,
                    ).exclude(
                        id=row["id"]
                    ).update(
                        status=Student_LeaveApproversData.Status.APPROVED,
                        approved_date=timezone.now(),
                    )

                    all_approved = not Student_LeaveApproversData.objects.filter(
                        leave_application_id=row["leave_application_id"]
                    ).exclude(
                        status=Student_LeaveApproversData.Status.APPROVED
                    ).exists()

                    if all_approved:
                        StudentLeaveOdApplication.objects.filter(
                            id=row["leave_application_id"]
                        ).update(
                            status=StudentLeaveOdApplication.Status.APPROVED
                        )

            messages.success(request, "Selected applications approved successfully.")
            return redirect(f"{reverse('approve_leave_od')}?tab=pending")

        elif action == "reject":

            reason = (request.POST.get("rejection_reason") or "").strip()

            if not reason:
                messages.error(request, "Rejection reason is required.")
                return redirect(f"{reverse('approve_leave_od')}?tab=pending")

            with transaction.atomic():

                for row in selected_rows:

                    if row["status"] != Student_LeaveApproversData.Status.PENDING:
                        continue

                    lower_pending = (
                        Student_LeaveApproversData.objects
                        .filter(
                            leave_application_id=row["leave_application_id"],
                            approver_level__lt=row["approver_level"],
                        )
                        .exclude(status=Student_LeaveApproversData.Status.APPROVED)
                        .exists()
                    )

                    if lower_pending:
                        continue

                    StudentLeaveOdApplication.objects.filter(
                        id=row["leave_application_id"]
                    ).update(
                        status=StudentLeaveOdApplication.Status.REJECTED,
                        remarks=reason,
                    )

                    Student_LeaveApproversData.objects.filter(
                        id=row["id"]
                    ).update(
                        status=Student_LeaveApproversData.Status.REJECTED,
                        reason=reason,
                        approved_date=timezone.now(),
                    )

                    # OR semantics: a rejection by one approver at this level
                    # decides the level, so clear any co-approver row(s) still
                    # pending at the same level for this application.
                    Student_LeaveApproversData.objects.filter(
                        leave_application_id=row["leave_application_id"],
                        approver_level=row["approver_level"],
                        status=Student_LeaveApproversData.Status.PENDING,
                    ).exclude(
                        id=row["id"]
                    ).update(
                        status=Student_LeaveApproversData.Status.REJECTED,
                        reason=reason,
                        approved_date=timezone.now(),
                    )

            messages.success(request, "Selected applications rejected successfully.")
            return redirect(f"{reverse('approve_leave_od')}?tab=pending")

        messages.error(request, "Invalid action.")
        return redirect("approve_leave_od")

    paginator = Paginator(approvals_qs, per_page)
    page_obj = paginator.get_page(page_number)

    page_application_ids = [
        approval.leave_application_id
        for approval in page_obj.object_list
        if approval.leave_application_id
    ]
    page_approval_rows = list(
        Student_LeaveApproversData.objects
        .filter(leave_application_id__in=page_application_ids)
        .order_by("approver_level", "id")
    )
    page_approver_user_ids = {
        row.approver_id_id
        for row in page_approval_rows
        if row.approver_id_id
    }
    page_approver_user_map = {
        user.id: user
        for user in USER.objects.using("rit_approval_system").filter(id__in=page_approver_user_ids)
    }
    page_approval_status_map = {}
    for row in page_approval_rows:
        approver_user = page_approver_user_map.get(row.approver_id_id)
        page_approval_status_map.setdefault(row.leave_application_id, []).append({
            "level": row.approver_level,
            "status": row.status,
            "name": getattr(approver_user, "username", "") or "-",
            "employee_id": getattr(approver_user, "Employee_id", "") or "-",
        })

    grouped_approvals = {}

    for approval in page_obj.object_list:

        app = approval.leave_application
        if app:
            app.approval_status_rows = page_approval_status_map.get(app.id, [])
            app.duration_days = app.display_total_days

        grouped_approvals.setdefault(
            approval.approver_level,
            []
        ).append(approval)

    return render(
        request,
        "course_management/faculty/approve_leave_od.html",
        {
            "page_obj": page_obj,
            "grouped_approvals": grouped_approvals,
            "q": q,
            "level_filter": level_filter,
            "application_type_filter": application_type_filter,
            "study_year_filter": study_year_filter,
            "available_levels": available_levels,
            "available_study_years": available_study_years,
            "tab": tab,
            "pending_count": pending_count,
            "approved_count": approved_count,
            "rejected_count": rejected_count,
        },
    )




@is_super_user('student_management')
def student_leave_approval_management(request):
    condition_options = [
        {"mode": "TRANSPORT", "gender": "ANY", "label": "Transport Students"},
        {"mode": "HOSTEL", "gender": "MALE", "label": "Hostel Students - Boys"},
        {"mode": "HOSTEL", "gender": "FEMALE", "label": "Hostel Students - Girls"},
    ]
    student_role = _get_student_leave_creator_role()

    if request.method == 'GET':
        departments = Add_Department.objects.all()

        return render(request, "student_management/admin/student_leave_management.html", {
            'student_role': student_role,
            'departments': departments,
            'condition_options': condition_options,
        })

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            creator_role_id_raw = data.get('creatorRole')
            role_hierarchy = data.get('roleHierarchy', [])
            applicant_mode = _normalize_student_leave_mode(data.get("conditionMode"))
            applicant_gender = _normalize_student_leave_gender(data.get("conditionGender"))

            if applicant_mode == "HOSTEL" and applicant_gender == "ANY":
                return JsonResponse({
                    'error': 'Select Boys or Girls for hostel student hierarchy'
                }, status=400)

            try:
                creator_role_id = int(creator_role_id_raw)
            except:
                return JsonResponse({'error': 'Invalid creatorRole'}, status=400)

            # ✅ delete old
            if not student_role or creator_role_id != student_role.id:
                return JsonResponse({
                    'error': 'Student leave hierarchy must use Student as creator role'
                }, status=400)

            allowed_role_ids = {
                role["id"]
                for role in _get_student_leave_allowed_approver_roles(
                    applicant_mode,
                    applicant_gender,
                )
            }

            parsed_hierarchy = []
            for role_data in role_hierarchy:
                try:
                    approver_role_id = int(role_data.get('id'))
                except:
                    return JsonResponse({'error': 'Invalid role id'}, status=400)

                if approver_role_id not in allowed_role_ids:
                    return JsonResponse({
                        'error': 'Only Faculty, HOD, and the matching hostel warden role are allowed'
                    }, status=400)

                is_cross_department = bool(role_data.get('isCrossDepartment', False))

                dept_id = role_data.get('departmentId')
                if dept_id in ["", None, "null", "undefined"]:
                    dept_id = None
                else:
                    dept_id = int(dept_id)

                # validation
                if is_cross_department and not dept_id:
                    return JsonResponse({
                        'error': f'Department required for role {approver_role_id}'
                    }, status=400)

                department_obj = Add_Department.objects.filter(id=dept_id).first() if dept_id else None

                parsed_hierarchy.append({
                    "approver_role_id": approver_role_id,
                    "is_cross_department": is_cross_department,
                    "department_obj": department_obj,
                })

            Student_LeaveApprovers.objects.filter(
                creator_role_id=creator_role_id,
                applicant_mode=applicant_mode,
                applicant_gender=applicant_gender,
            ).delete()

            for index, role_data in enumerate(parsed_hierarchy):
                Student_LeaveApprovers.objects.create(
                    creator_role_id=creator_role_id,
                    applicant_mode=applicant_mode,
                    applicant_gender=applicant_gender,
                    approver_role_id=role_data["approver_role_id"],
                    approver_level=index + 1,
                    is_cross_department_approver="YES" if role_data["is_cross_department"] else "NO",
                    approver_department=role_data["department_obj"]
                )

            return JsonResponse({'message': 'Roles submitted successfully'}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request'}, status=405) 




from django.http import JsonResponse
from django.views.decorators.http import require_GET

def _compact_role_name(value):
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def _student_leave_role_key(role_name):
    compact = _compact_role_name(role_name)
    if compact == "STUDENT":
        return "STUDENT"
    if compact == "FACULTY":
        return "FACULTY"
    if compact in {"HOD", "HEADOFTHEDEPARTMENT"}:
        return "HOD"
    if "HOSTEL" in compact and "WARDEN" in compact:
        if "BOY" in compact or "HSB" in compact:
            return "HOSTEL_BOYS"
        if "GIRL" in compact or "HG" in compact:
            return "HOSTEL_GIRLS"
    return None


def _get_student_leave_creator_role():
    for role in Role.objects.using("rit_approval_system").all():
        if _student_leave_role_key(role.role) == "STUDENT":
            return role
    return None


def _get_student_leave_allowed_role_keys(applicant_mode, applicant_gender):
    if applicant_mode == "TRANSPORT":
        return ["FACULTY", "HOD"]
    if applicant_mode == "HOSTEL" and applicant_gender == "MALE":
        return ["FACULTY", "HOD", "HOSTEL_BOYS"]
    if applicant_mode == "HOSTEL" and applicant_gender == "FEMALE":
        return ["FACULTY", "HOD", "HOSTEL_GIRLS"]
    return ["FACULTY", "HOD"]


def _get_student_leave_allowed_approver_roles(applicant_mode, applicant_gender):
    allowed_keys = _get_student_leave_allowed_role_keys(applicant_mode, applicant_gender)
    role_by_key = {}

    for role in Role.objects.using("rit_approval_system").all():
        key = _student_leave_role_key(role.role)
        if key in allowed_keys and key not in role_by_key:
            role_by_key[key] = {"id": role.id, "role": role.role, "key": key}

    return [role_by_key[key] for key in allowed_keys if key in role_by_key]


def _normalize_student_leave_mode(value):
    value = str(value or "").strip().upper()
    if "HOSTEL" in value:
        return "HOSTEL"
    if "TRANSPORT" in value:
        return "TRANSPORT"
    return "DEFAULT"


def _normalize_student_leave_gender(value):
    value = str(value or "").strip().upper()
    if value in {"M", "MALE", "BOY", "BOYS"}:
        return "MALE"
    if value in {"F", "FEMALE", "GIRL", "GIRLS"}:
        return "FEMALE"
    return "ANY"


@is_super_user('student_management')
def student_api_leave_roles(request, creatorRoleId):
    try:
        # ✅ creatorRoleId from URL is string, cast to int safely
        try:
            creator_role_id = int(creatorRoleId)
        except (TypeError, ValueError):
            return JsonResponse({"error": "Invalid creatorRoleId"}, status=400)

        applicant_mode = _normalize_student_leave_mode(request.GET.get("conditionMode"))
        applicant_gender = _normalize_student_leave_gender(request.GET.get("conditionGender"))

        ROLE_DB = "rit_approval_system"

        # ✅ Ensure creator role exists in ROLE_DB
        creator_role = (
            Role.objects.using(ROLE_DB)
            .filter(id=creator_role_id)
            .values("id", "role")
            .first()
        )
        if not creator_role:
            return JsonResponse({"error": "Creator role not found"}, status=404)

        # ✅ All roles except creator (from ROLE_DB)
        if _student_leave_role_key(creator_role["role"]) != "STUDENT":
            return JsonResponse({"error": "Creator role must be Student"}, status=400)

        # Available Roles panel shows every role in the Role table (except the
        # Student creator role), not just the curated Faculty/HOD/Warden subset.
        allowed_roles = [
            {"id": r["id"], "role": r["role"]}
            for r in Role.objects.using(ROLE_DB).all().values("id", "role")
            if r["id"] != creator_role_id
            and _student_leave_role_key(r["role"]) != "STUDENT"
        ]
        allowed_role_ids = [role["id"] for role in allowed_roles]

        # ✅ Approver mappings (from your default DB unless Student_LeaveApprovers is routed)
        # If Student_LeaveApprovers is also in a different DB, add .using("that_db") here.
        approvers_qs = Student_LeaveApprovers.objects.filter(
            creator_role_id=creator_role_id,
            applicant_mode=applicant_mode,
            applicant_gender=applicant_gender,
            approver_role_id__in=allowed_role_ids,
        )

        matched_roles = []
        matched_ids = set()

        # ✅ Collect role_ids first (avoid N queries)
        approver_role_ids = list(
            approvers_qs.values_list("approver_role_id", flat=True)
        )

        # ✅ Map role_id -> role_name from ROLE_DB
        role_name_map = {
            r["id"]: r["role"]
            for r in Role.objects.using(ROLE_DB)
                .filter(id__in=approver_role_ids)
                .values("id", "role")
        }

        for item in approvers_qs:
            rid = item.approver_role_id
            matched_ids.add(rid)

            matched_roles.append({
                "id": rid,
                "role": role_name_map.get(rid, ""),  # ✅ no cross-db FK deref
                "is_cross_department": str(getattr(item, "is_cross_department_approver", "")).upper() == "YES",
                "approver_department_id": item.approver_department_id,
                "approver_department_name": item.approver_department.Department if item.approver_department else None,

            })

        unmatched_roles = [
            {"id": role["id"], "role": role["role"]}
            for role in allowed_roles
            if role["id"] not in matched_ids
        ]

        return JsonResponse({
            "matched_roles": matched_roles,
            "unmatched_roles": unmatched_roles
        })

    except Exception as e:
        # helpful during dev; you can log traceback too
        return JsonResponse({"error": str(e)}, status=500)
 







from user_accounts.models import USER  # your external user model
from collections import defaultdict

def _academic_year_of(d):
    """Academic year label for a date using the June boundary (e.g. 2024-2025)."""
    if not d:
        return None
    return f"{d.year}-{d.year + 1}" if d.month >= 6 else f"{d.year - 1}-{d.year}"


def _academic_year_range(ay):
    """Return (start_date, end_date) for an 'YYYY-YYYY' academic year label."""
    from datetime import date
    try:
        y1 = int(str(ay).split("-")[0])
    except (ValueError, AttributeError, IndexError):
        return None
    return date(y1, 6, 1), date(y1 + 1, 5, 31)


@check_permission("leave_od_applications")
def leave_od_applications(request):
    """Read-only dashboard of ALL student Leave/OD applications for the viewer's
    department, including the full approval chain (who approved, at which level,
    with what status and when).

    Access is governed only by the ``leave_od_applications`` permission key — it
    is NOT hardcoded to HOD. Any role granted that permission (HOD, Principal,
    Dean, etc.) sees the applications for their own department, resolved purely
    from the faculty→department link.
    """
    employee_id = request.user.Employee_id

    # --- Resolve viewer's department from the faculty record (no role check) ---
    gi_self = (
        general_information.objects
        .select_related("department")
        .filter(faculty_id=employee_id)
        .first()
    )

    if not gi_self:
        messages.error(request, "Faculty profile not found.")
        return redirect("home")

    if not gi_self.department:
        messages.error(request, "Your profile is missing a department. Contact admin.")
        return redirect("home")

    dept = gi_self.department

    # --- Filters ---
    q = (request.GET.get("q") or "").strip()
    type_filter = (request.GET.get("application_type") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()
    study_year_filter = (request.GET.get("study_year") or "").strip()
    academic_year_filter = (request.GET.get("academic_year") or "").strip()

    applications_qs = (
        StudentLeaveOdApplication.objects
        .select_related("student", "student__mentor", "student__ca", "department")
        .filter(department=dept)
        .order_by("-created_at")
    )

    if q:
        applications_qs = applications_qs.filter(
            Q(student__name__icontains=q) | Q(student__reg_no__icontains=q)
        )
    if type_filter:
        applications_qs = applications_qs.filter(application_type=type_filter)
    if status_filter:
        applications_qs = applications_qs.filter(status=status_filter)
    if study_year_filter:
        applications_qs = applications_qs.filter(study_year=study_year_filter)
    if academic_year_filter:
        ay_range = _academic_year_range(academic_year_filter)
        if ay_range:
            applications_qs = applications_qs.filter(from_date__date__range=ay_range)

    # --- Summary counts (whole department, ignoring the active filters) ---
    dept_base = StudentLeaveOdApplication.objects.filter(department=dept)
    total_count = dept_base.count()
    pending_count = dept_base.filter(status=StudentLeaveOdApplication.Status.PENDING).count()
    mentor_approved_count = dept_base.filter(status=StudentLeaveOdApplication.Status.MENTOR_APPROVED).count()
    approved_count = dept_base.filter(status=StudentLeaveOdApplication.Status.APPROVED).count()
    rejected_count = dept_base.filter(status=StudentLeaveOdApplication.Status.REJECTED).count()

    available_study_years = list(
        dept_base.exclude(study_year__isnull=True).exclude(study_year="")
        .values_list("study_year", flat=True).distinct()
    )

    # Academic years present in this department, derived from from_date (June boundary)
    from django.db.models import Min, Max
    bounds = dept_base.aggregate(min_d=Min("from_date"), max_d=Max("from_date"))
    available_academic_years = []
    if bounds["min_d"] and bounds["max_d"]:
        first_year = bounds["min_d"].year if bounds["min_d"].month >= 6 else bounds["min_d"].year - 1
        last_year = bounds["max_d"].year if bounds["max_d"].month >= 6 else bounds["max_d"].year - 1
        available_academic_years = [f"{y}-{y + 1}" for y in range(last_year, first_year - 1, -1)]

    # --- Pagination ---
    paginator = Paginator(applications_qs, 15)
    page_obj = paginator.get_page(request.GET.get("page"))
    page_apps = list(page_obj.object_list)

    # --- Real approval chain for the current page (who approved / level / when) ---
    page_application_ids = [a.id for a in page_apps]
    approval_rows = list(
        Student_LeaveApproversData.objects
        .filter(leave_application_id__in=page_application_ids)
        .order_by("approver_level", "id")
    )
    approver_user_ids = {r.approver_id_id for r in approval_rows if r.approver_id_id}
    approver_user_map = {
        u.id: u
        for u in USER.objects.using("rit_approval_system")
        .select_related("role")
        .filter(id__in=approver_user_ids)
    }

    approval_map = {}
    for r in approval_rows:
        u = approver_user_map.get(r.approver_id_id)
        role_name = getattr(getattr(u, "role", None), "role", "") if u else ""
        approval_map.setdefault(r.leave_application_id, []).append({
            "level": r.approver_level,
            "status": r.status,
            "name": getattr(u, "username", "") or "-",
            "employee_id": getattr(u, "Employee_id", "") or "-",
            "role": role_name or "-",
            "approved_date": r.approved_date,
        })

    # --- Enrich rows on the current page for display ---
    for app in page_apps:
        s = app.student  # StudentDetails or None
        app.student_name = getattr(s, "name", "-") if s else "-"
        app.student_reg_no = getattr(s, "reg_no", "-") if s else "-"
        app.section = getattr(s, "section", "-") if s else "-"
        app.batch = getattr(s, "batch", "-") if s else "-"

        mentor_gi = getattr(s, "mentor", None) if s else None
        ca_gi = getattr(s, "ca", None) if s else None
        app.mentor_name = getattr(mentor_gi, "name", "-") if mentor_gi else "-"
        app.ca_name = getattr(ca_gi, "name", "-") if ca_gi else "-"

        app.duration_days = app.display_total_days

        app.academic_year = _academic_year_of(app.from_date) or "-"
        app.approval_rows = approval_map.get(app.id, [])

    return render(
        request,
        "course_management/faculty/student_leave_od_applications.html",
        {
            "today": timezone.now(),
            "viewer_department": dept,
            "page_obj": page_obj,
            "applications": page_apps,
            "total_count": total_count,
            "pending_count": pending_count,
            "mentor_approved_count": mentor_approved_count,
            "approved_count": approved_count,
            "rejected_count": rejected_count,
            "q": q,
            "application_type_filter": type_filter,
            "status_filter": status_filter,
            "study_year_filter": study_year_filter,
            "academic_year_filter": academic_year_filter,
            "available_study_years": available_study_years,
            "available_academic_years": available_academic_years,
        },
    )






import json
import re

from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q

from django.db.models import Case, When, Value, IntegerField
from django.db.models.functions import Cast, Mod

# @check_permission("faculty_courses")
# def faculty_courses(request):
#     faculty_id = getattr(request.user, "Employee_id", None)

#     try:
#         faculty = general_information.objects.get(faculty_id=faculty_id)
#     except general_information.DoesNotExist:
#         faculty = None
#     academic_year = get_academic_year()
#     assigned_courses = (
#         AssignSubjectFaculty.objects
#         .filter(
#             faculty__faculty_id=faculty_id,
#             is_active=True,
#             academic_year=academic_year,
#         )
#         .select_related(
#             "course",
#             "course__department",
#             "course__regulation",
#             "faculty",
#             "department",
#         )
#         .annotate(
#             # Convert semester to int
#             semester_int=Cast("course__semester", IntegerField()),

#             # semester % 2 → Even = 0, Odd = 1
#             semester_mod=Mod(
#                 Cast("course__semester", IntegerField()),
#                 Value(2)
#             ),

#             # ✅ EVEN FIRST (0), ODD NEXT (1)
#             semester_order=Case(
#                 When(semester_mod=0, then=Value(0)),  # EVEN
#                 default=Value(1),                     # ODD
#                 output_field=IntegerField(),
#             )
#         )
#         .order_by(
#             "-course__regulation__year",  # Academic Year DESC
#             "semester_order",             # EVEN → ODD
#             "semester_int",               # 2,4,6 then 1,3,5
#             "course__year",
#         )
#     )

#     context = {
#         "faculty": faculty,
#         "assigned_courses": assigned_courses,
#     }

#     return render(
#         request,
#         "course_management/faculty/faculty_courses.html",
#         context
#     )
 

@check_permission("faculty_courses")
def faculty_courses(request):
    faculty_id = getattr(request.user, "Employee_id", None)

    try:
        faculty = general_information.objects.get(faculty_id=faculty_id)
    except general_information.DoesNotExist:
        faculty = None
    current_academic_year = get_academic_year()
    assigned_courses = (
        AssignSubjectFaculty.objects
        # No academic_year filter — show every year the faculty has ever been
        # assigned courses in, not just the current one.
        .filter(
            faculty__faculty_id=faculty_id,
            is_active=True,
        )
        .select_related(
            "course",
            "course__department",
            "course__regulation",
            "faculty",
            "department",
        )
        .annotate(
            # Convert semester to int
            semester_int=Cast("course__semester", IntegerField()),

            # semester % 2 → Even = 0, Odd = 1
            semester_mod=Mod(
                Cast("course__semester", IntegerField()),
                Value(2)
            ),

            # ✅ EVEN FIRST (0), ODD NEXT (1)
            semester_order=Case(
                When(semester_mod=0, then=Value(0)),  # EVEN
                default=Value(1),                     # ODD
                output_field=IntegerField(),
            ),

            # Current academic year always sorts first, regardless of any
            # future-dated rows; every other year then falls in below it.
            is_current_year=Case(
                When(academic_year=current_academic_year, then=Value(0)),
                default=Value(1),
                output_field=IntegerField(),
            ),
        )
        .order_by(
            "is_current_year",   # Current year first
            "-academic_year",    # Then previous years, most recent to oldest
            "semester_order",    # EVEN → ODD
            "semester_int",      # 2,4,6 then 1,3,5
            "course__year",
        )
    )

    context = {
        "faculty": faculty,
        "assigned_courses": assigned_courses,
        "current_academic_year": current_academic_year,
    }

    return render(
        request,
        "course_management/faculty/faculty_courses.html",
        context
    )
 


import json
import re
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, render
from django.views.decorators.http import require_http_methods
from django.db import transaction

def _po_number_to_int(program_number):
    if not program_number:
        return None
    m = re.search(r"(\d+)", str(program_number))
    return int(m.group(1)) if m else None


def _clean_level(val):
    """
    Accept only '', 1,2,3. Treat 0/blank/invalid as ''.
    """
    if val is None:
        return ""
    s = str(val).strip()
    if s in ("", "0"):
        return ""
    if s in ("1", "2", "3"):
        return s
    try:
        i = int(float(s))
        return str(i) if i in (1, 2, 3) else ""
    except Exception:
        return ""


def _po_type_to_bool(po_type_val) -> bool:
    """
    Convert UI value (0/1 or true/false) to boolean mapping field.
    Revised=1 => True, Non-Revised=0 => False
    """
    return str(po_type_val).strip().lower() in ("1", "true", "yes")


def _get_po_type_from_mapping(m: "Co_Po_Mapping") -> int:
    """
    mapping is BooleanField:
      True  => Revised (1)
      False => Non-Revised (0)
    """
    return 1 if bool(getattr(m, "mapping", False)) else 0


def _get_po_type_label_from_mapping(m: "Co_Po_Mapping") -> str:
    return "Revised" if _get_po_type_from_mapping(m) == 1 else "Non-Revised"


def _mapping_to_dict(m: "Co_Po_Mapping"):
    co_obj = m.co_number
    co_number = (
        getattr(co_obj, "co_number", None)
        or getattr(co_obj, "co", None)
        or getattr(co_obj, "co_code", None)
        or f"CO{getattr(co_obj, 'id', '')}"
    )

    def lvl(v):
        s = ("" if v is None else str(v)).strip()
        return s if s in ("1", "2", "3") else ""

    return {
        "id": m.id,
        "co_id": m.co_number_id,
        "co_number": co_number,
        "description": m.co_description or "",

        # ✅ Revised/Non-Revised from BOOLEAN mapping field
        "po_type": _get_po_type_from_mapping(m),                 # 1/0
        "po_type_label": _get_po_type_label_from_mapping(m),     # text

        # PO
        "po1": lvl(m.po_number_1),
        "po2": lvl(m.po_number_2),
        "po3": lvl(m.po_number_3),
        "po4": lvl(m.po_number_4),
        "po5": lvl(m.po_number_5),
        "po6": lvl(m.po_number_6),
        "po7": lvl(m.po_number_7),
        "po8": lvl(m.po_number_8),
        "po9": lvl(m.po_number_9),
        "po10": lvl(m.po_number_10),
        "po11": lvl(m.po_number_11),
        "po12": lvl(m.po_number_12),
        "po13": lvl(m.po_number_13),

        # PSO
        "pso1": lvl(m.pso_number_1),
        "pso2": lvl(m.pso_number_2),
        "pso3": lvl(m.pso_number_3),
        "pso4": lvl(m.pso_number_4),
        "pso5": lvl(m.pso_number_5),
    }


# ==========================================================
# ✅ PSO API (Loads PSO list by department + optional year)
# ==========================================================

@require_http_methods(["GET"])
def api_program_specific_outcomes(request):
    department_id = (request.GET.get("department_id") or "").strip()
    year_raw = (request.GET.get("year") or "").strip()
    batch = (request.GET.get("batch") or "").strip()
    print(f"API: department_id={department_id}, year={year_raw}, batch={batch}")

    if not department_id:
        return JsonResponse({"success": False, "error": "department_id is required"}, status=400)

    # base queryset
    base_qs = Program_specific_Outcomes.objects.filter(is_active=True, department_id=department_id, batch=batch)

    # optional year
    year = None
    if year_raw:
        try:
            year = int(year_raw)
        except Exception:
            year = None

    qs = base_qs
    used_year_filter = False

    # try strict year first
    if year is not None:
        qs_year = base_qs.filter(year=year)
        if qs_year.exists():
            qs = qs_year
            used_year_filter = True
        else:
            # fallback: if nothing for that year, return without year filter
            # (also optionally include year=None rows first)
            qs = base_qs

    qs = qs.order_by("id")

    items, nums = [], []
    for idx, pso in enumerate(qs[:5], start=1):
        nums.append(idx)
        items.append({
            "num": idx,
            "statement": pso.pso_statement,
            "db_year": pso.year,                # helps you debug
            "db_department_id": pso.department_id
        })

    return JsonResponse({
        "success": True,
        "department_id": department_id,
        "requested_year": year,
        "used_year_filter": used_year_filter,
        "count": qs.count(),
        "nums": nums,
        "items": items
    })


# ==========================================
# ✅ Program Outcomes API (Revised / Non)
# ==========================================
@require_http_methods(["GET"])
def api_program_outcomes(request):
    is_rev = request.GET.get("is_revised", "0")
    is_revised = str(is_rev).lower() in ("1", "true", "yes")

    qs = Program_outcomes.objects.filter(is_active=True, is_revised=is_revised)
    try:
        qs = qs.order_by("program_number")
    except Exception:
        pass

    items, nums = [], []
    for po in qs:
        program_number = getattr(po, "program_number", None) or getattr(po, "po_number", None) or ""
        num = _po_number_to_int(program_number)
        if not num or num < 1 or num > 13:
            continue

        items.append({
            "num": num,
            "program_number": program_number or f"PO{num}",
            "name": getattr(po, "program_name", "") or "",
            "description": getattr(po, "program_description", "") or "",
        })
        nums.append(num)

    nums = sorted(list(set(nums)))
    return JsonResponse({"success": True, "is_revised": is_revised, "nums": nums, "items": items})


# =========================================================
# ✅ Faculty Course Students Page
# =========================================================
def faculty_course_students(request, year, semester, course_id, batch, section, regulation_id):
    faculty_id = getattr(request.user, "Employee_id", None)
    faculty = get_object_or_404(general_information, faculty_id=faculty_id)

    course = get_object_or_404(Course, id=course_id)
    regulation = get_object_or_404(Regulations, id=regulation_id)

    course_enrollments = (
        CourseEnrollment.objects
        .filter(
            course_id=course_id,
            batch=batch,
            section=section,
            regulation_id=regulation_id,
            faculty=faculty,
            enroll=True,
            student__is_active=True,
            student__is_discontinued=False,
        )
        .select_related("student", "course", "faculty", "department__degree")
        .order_by("student__reg_no")
    )

    students = [e.student for e in course_enrollments]

    assigned = AssignSubjectFaculty.objects.filter(
        course=course,
        regulation=regulation,
        faculty=faculty
    ).first()

    co_list = []
    try:
        co_qs = CourseOutcome.objects.filter(
            regulation=regulation.year
        ).order_by("id")
    except Exception:
        co_qs = CourseOutcome.objects.all().order_by("id")

    for co in co_qs:
        co_list.append({
            "id": co.id,
            "co_number": f"{getattr(co, 'co_code', '')}".strip()
                         + (" - " if getattr(co, 'co_code', None) else "")
                         + f"{getattr(co, 'co_name', '')}".strip(),
            "description": getattr(co, "description", "") or "",
        })

    existing = []
    if assigned:
        existing_qs = (
            Co_Po_Mapping.objects
            .filter(course=course, assigned_faculty=assigned)
            .select_related("co_number")
            .order_by("co_number_id")
        )
        existing = [_mapping_to_dict(m) for m in existing_qs]

    department_id = getattr(course, "department_id", "") or ""

    context = {
        "faculty": faculty,
        "course": course,
        "regulation": regulation,

        "course_enrollments": course_enrollments,
        "students": students,

        "year": year,
        "semester": semester,
        "batch": batch,
        "section": section,
        "regulation_id": regulation_id,

        "department_id": department_id,
        "batch_year": batch,

        "co_list_json": json.dumps(co_list),
        "existing_mappings_json": json.dumps(existing),
    }

    return render(
        request,
        "course_management/faculty/faculty_course_students.html",
        context
    )


@require_http_methods(["POST"])
def co_po_mapping_manage(request):
    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)

    action = (data.get("action") or "").strip().lower()
    course_id = data.get("course_id")
    regulation_id = data.get("regulation_id")

    if not course_id or not regulation_id:
        return JsonResponse({"success": False, "error": "course_id and regulation_id are required"}, status=400)

    course = get_object_or_404(Course, id=course_id)
    regulation = get_object_or_404(Regulations, id=regulation_id)

    faculty_id = getattr(request.user, "Employee_id", None)
    faculty = get_object_or_404(general_information, faculty_id=faculty_id)

    assigned = AssignSubjectFaculty.objects.filter(
        course=course,
        regulation=regulation,
        faculty=faculty
    ).first()

    if not assigned:
        return JsonResponse({
            "success": False,
            "error": "No faculty assignment found for this course & regulation"
        }, status=400)

    def refresh_rows():
        qs = Co_Po_Mapping.objects.filter(
            course=course,
            assigned_faculty=assigned
        ).select_related("co_number").order_by("co_number_id")
        return [_mapping_to_dict(m) for m in qs]

    if action == "list":
        return JsonResponse({"success": True, "rows": refresh_rows()})

    if action == "save":
        mapping_id = data.get("mapping_id")
        co_id = data.get("co_id")

        if not co_id:
            return JsonResponse({"success": False, "error": "co_id is required"}, status=400)

        try:
            co = CourseOutcome.objects.get(id=co_id)
        except CourseOutcome.DoesNotExist:
            return JsonResponse({"success": False, "error": "Invalid Course Outcome ID"}, status=400)

        description = (data.get("description") or "").strip()

        # PO Type: 1 → Revised (True), 0 → Non-Revised (False)
        mapping_bool = _po_type_to_bool(data.get("po_type", "0"))

        # ────────────────────────────────────────────────
        #   NO RESTRICTIONS on number of POs or PSOs
        #   Any count is allowed (including 0)
        # ────────────────────────────────────────────────

        po_levels = {}
        for i in range(1, 14):
            lv = _clean_level(data.get(f"po{i}"))
            po_levels[i] = lv

        pso_levels = {}
        for i in range(1, 6):
            lv = _clean_level(data.get(f"pso{i}"))
            pso_levels[i] = lv

        defaults = {
            "mapping": mapping_bool,
            "co_description": description or (co.description or ""),

            # PO fields
            "po_number_1": po_levels[1],
            "po_number_2": po_levels[2],
            "po_number_3": po_levels[3],
            "po_number_4": po_levels[4],
            "po_number_5": po_levels[5],
            "po_number_6": po_levels[6],
            "po_number_7": po_levels[7],
            "po_number_8": po_levels[8],
            "po_number_9": po_levels[9],
            "po_number_10": po_levels[10],
            "po_number_11": po_levels[11],
            "po_number_12": po_levels[12],
            "po_number_13": po_levels[13],

            # PSO fields
            "pso_number_1": pso_levels[1],
            "pso_number_2": pso_levels[2],
            "pso_number_3": pso_levels[3],
            "pso_number_4": pso_levels[4],
            "pso_number_5": pso_levels[5],
        }

        with transaction.atomic():
            if mapping_id:
                # Update existing mapping
                mapping = get_object_or_404(
                    Co_Po_Mapping,
                    id=mapping_id,
                    course=course,
                    assigned_faculty=assigned
                )
                mapping.co_number = co
                for key, value in defaults.items():
                    setattr(mapping, key, value)
                mapping.save()
            else:
                # Create new or update if same CO already exists
                Co_Po_Mapping.objects.update_or_create(
                    course=course,
                    assigned_faculty=assigned,
                    co_number=co,
                    defaults=defaults
                )

        return JsonResponse({
            "success": True,
            "message": "Mapping saved successfully.",
            "rows": refresh_rows()
        })

    if action == "delete":
        mapping_id = data.get("mapping_id")
        if not mapping_id:
            return JsonResponse({"success": False, "error": "mapping_id is required"}, status=400)

        deleted, _ = Co_Po_Mapping.objects.filter(
            id=mapping_id,
            course=course,
            assigned_faculty=assigned
        ).delete()

        if deleted == 0:
            return JsonResponse({
                "success": False,
                "error": "Mapping not found or already deleted"
            }, status=404)

        return JsonResponse({
            "success": True,
            "message": "Mapping deleted successfully.",
            "rows": refresh_rows()
        })

    return JsonResponse({"success": False, "error": "Invalid action"}, status=400)


import io
from collections import defaultdict
from datetime import datetime

from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape

def _safe(v):
    return "" if v is None else str(v).strip()


def faculty_course_students_attendance_view(request, year, semester, course_id, batch, section, regulation_id):
    faculty_id = getattr(request.user, "Employee_id", None)
    faculty = get_object_or_404(general_information, faculty_id=faculty_id)
    course = get_object_or_404(Course, id=course_id)
    regulation = get_object_or_404(Regulations, id=regulation_id)

    course_enrollments = (
        CourseEnrollment.objects
        .filter(
            course_id=course_id,
            batch=batch,
            section=section,
            regulation_id=regulation_id,
            faculty=faculty,
            enroll=True
        )
        .select_related("student", "course", "faculty", "department__degree")
        .order_by("student__reg_no")
    )

    students = [e.student for e in course_enrollments]
    student_ids = [s.id for s in students]

    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    parsed_from = None
    parsed_to = None

    try:
        if date_from:
            parsed_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        if date_to:
            parsed_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Invalid date format. Use YYYY-MM-DD.", status=400)

    if parsed_from and parsed_to and parsed_from > parsed_to:
        return HttpResponse("'From Date' cannot be greater than 'To Date'.", status=400)

    att_qs = HourAttendance.objects.filter(
        course=course,
        batch=str(batch),
        section=str(section),
        year=str(year),
        semester=str(semester),
    )

    if student_ids:
        att_qs = att_qs.filter(student_id__in=student_ids)
    else:
        att_qs = att_qs.none()

    if parsed_from:
        att_qs = att_qs.filter(date__gte=parsed_from)
    if parsed_to:
        att_qs = att_qs.filter(date__lte=parsed_to)

    att_qs = att_qs.select_related("student").order_by("date", "student__reg_no", "period")

    grouped = defaultdict(lambda: defaultdict(dict))
    all_periods_by_date = defaultdict(set)

    for row in att_qs:
        if not row.student_id:
            continue
        grouped[row.student_id][row.date][row.period] = row.status
        if row.period is not None:
            all_periods_by_date[row.date].add(row.period)

    sorted_dates = sorted(all_periods_by_date.keys())
    students_sorted = sorted(students, key=lambda s: _safe(getattr(s, "reg_no", "")))

    date_columns = []
    for dt in sorted_dates:
        periods = sorted(all_periods_by_date[dt])
        date_columns.append({
            "date": dt,
            "date_str": dt.strftime("%d-%m-%Y"),
            "periods": periods,
            "period_count": len(periods),
        })

    student_rows = []
    for idx, student in enumerate(students_sorted, start=1):
        values = []
        for dt_info in date_columns:
            dt = dt_info["date"]
            for period in dt_info["periods"]:
                status_value = grouped[student.id].get(dt, {}).get(period, "-")
                values.append({
                    "value": status_value,
                    "is_absent": str(status_value).strip().lower() == "absent",
                })

        student_rows.append({
            "sno": idx,
            "reg_no": _safe(getattr(student, "reg_no", "")),
            "name": _safe(getattr(student, "name", "")),
            "values": values,
        })

    return render(
        request,
        "course_management/faculty/attendance_detail_view.html",
        {
            "faculty": faculty,
            "course": course,
            "regulation": regulation,
            "year": year,
            "semester": semester,
            "batch": batch,
            "section": section,
            "date_from": date_from,
            "date_to": date_to,
            "date_columns": date_columns,
            "student_rows": student_rows,
        }
    )



from reportlab.lib.pagesizes import landscape

def faculty_course_students_attendance_pdf(request, year, semester, course_id, batch, section, regulation_id):
    faculty_id = getattr(request.user, "Employee_id", None)
    faculty = get_object_or_404(general_information, faculty_id=faculty_id)
    course = get_object_or_404(Course, id=course_id)
    regulation = get_object_or_404(Regulations, id=regulation_id)

    faculty_name = _safe(getattr(faculty, "name", "")) or "Faculty"
    faculty_code = faculty_id or ""

    course_enrollments = (
        CourseEnrollment.objects
        .filter(
            course_id=course_id,
            batch=batch,
            section=section,
            regulation_id=regulation_id,
            faculty=faculty,
            enroll=True
        )
        .select_related("student", "course", "faculty", "department__degree")
        .order_by("student__reg_no")
    )

    students = [e.student for e in course_enrollments]
    student_ids = [s.id for s in students]
    total_students = len(students)
    enrollment = course_enrollments.first()

    degree_department = ""
    if enrollment and enrollment.department and enrollment.department.degree:
        degree_code = _safe(enrollment.department.degree.degree_code)
        department_name = _safe(enrollment.department.Department)
        degree_department = f"{degree_code} - {department_name}"

    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    parsed_from = None
    parsed_to = None

    try:
        if date_from:
            parsed_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        if date_to:
            parsed_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Invalid date format. Use YYYY-MM-DD.", status=400)

    if parsed_from and parsed_to and parsed_from > parsed_to:
        return HttpResponse("'From Date' cannot be greater than 'To Date'.", status=400)

    att_qs = HourAttendance.objects.filter(
        course=course,
        batch=str(batch),
        section=str(section),
        year=str(year),
        semester=str(semester),
    )

    if student_ids:
        att_qs = att_qs.filter(student_id__in=student_ids)
    else:
        att_qs = att_qs.none()

    if parsed_from:
        att_qs = att_qs.filter(date__gte=parsed_from)
    if parsed_to:
        att_qs = att_qs.filter(date__lte=parsed_to)

    total_hours = att_qs.values("date", "period").distinct().count()

    agg = (
        att_qs.values("student_id")
        .annotate(
            present=Count("id", filter=Q(status="Present")),
            absent=Count("id", filter=Q(status="Absent")),
            od=Count("id", filter=Q(status="On Duty")),
        )
    )
    counts_by_student = {r["student_id"]: r for r in agg}

    styles = getSampleStyleSheet()

    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1a4b8c")
    ACCENT_RED = colors.HexColor("#b91c1c")
    DARK_GRAY = colors.HexColor("#111827")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BG_GRAY = colors.HexColor("#f8fafc")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle(
        "title_style",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=PRIMARY_BLUE,
        alignment=TA_CENTER,
        spaceAfter=4,
        fontName="Helvetica-Bold",
    )

    sub_style = ParagraphStyle(
        "sub_style",
        parent=styles["Normal"],
        fontSize=10,
        textColor=MEDIUM_GRAY,
        alignment=TA_CENTER,
        spaceAfter=8,
    )

    section_style = ParagraphStyle(
        "section_style",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=PRIMARY_BLUE,
        alignment=TA_LEFT,
        spaceBefore=8,
        spaceAfter=5,
        fontName="Helvetica-Bold",
    )

    info_style = ParagraphStyle(
        "info_style",
        parent=styles["Normal"],
        fontSize=10,
        textColor=DARK_GRAY,
        alignment=TA_LEFT,
        leading=13,
        fontName="Helvetica",
    )

    table_header = ParagraphStyle(
        "table_header",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.white,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        leading=12,
        wordWrap="CJK",
    )

    cell_left = ParagraphStyle(
        "cell_left",
        parent=styles["Normal"],
        fontSize=10,
        textColor=DARK_GRAY,
        alignment=TA_LEFT,
        leading=12,
        wordWrap="CJK",
    )

    cell_center = ParagraphStyle(
        "cell_center",
        parent=cell_left,
        alignment=TA_CENTER,
    )

    def calc_col_widths(page_width):
        fixed = {
            "sno": 14 * mm,
            "reg": 34 * mm,
            "p": 16 * mm,
            "od": 16 * mm,
            "a": 16 * mm,
            "total": 20 * mm,
            "perc": 24 * mm,
            "status": 22 * mm,
        }
        fixed_sum = sum(fixed.values())
        name_w = max(90 * mm, page_width - fixed_sum)

        return [
            fixed["sno"],
            fixed["reg"],
            name_w,
            fixed["p"],
            fixed["od"],
            fixed["a"],
            fixed["total"],
            fixed["perc"],
            fixed["status"],
        ]

    def make_table(table_data, col_widths, header_bg=SECONDARY_BLUE, zebra=True):
        t = Table(table_data, colWidths=col_widths, repeatRows=1)

        ts = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.7, BORDER_GRAY),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("ALIGN", (2, 0), (2, -1), "LEFT"),
            ("ALIGN", (3, 0), (-1, -1), "CENTER"),
        ])

        if zebra and len(table_data) > 2:
            ts.add("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY])

        t.setStyle(ts)
        return t

    buffer = io.BytesIO()
    PAGE = landscape(A4)

    doc = BaseDocTemplate(
        buffer,
        pagesize=PAGE,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Course Attendance - {course.course_code}",
        showBoundary=0
    )

    HEADER_HEIGHT = 26 * mm

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = PAGE
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 4 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(
                ImageReader(logo_path),
                left, top_y - 15 * mm,
                width=24 * mm, height=14 * mm,
                preserveAspectRatio=True, mask="auto"
            )

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(center_x, top_y - 4 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawCentredString(center_x, top_y - 9 * mm, "An Autonomous Institution")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 7.5)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "Approved by AICTE, New Delhi")
        canvas.setFont("Helvetica", 7)
        canvas.drawCentredString(center_x, top_y - 16.5 * mm, "Accredited by NAAC & ISO 9001:2015 Certified")
        canvas.drawCentredString(center_x, top_y - 20 * mm, "NBA Accredited: CSE, EEE, ECE, MECH")

        footer_y = 9 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.6)
        canvas.line(left, footer_y + 5 * mm, right, footer_y + 5 * mm)

        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 7)
        gen_time = datetime.now().strftime("%d %b %Y, %H:%M")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawCentredString(center_x, footer_y, f"Faculty: {faculty_code}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")

        canvas.restoreState()

    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin + 3 * mm,
        doc.width,
        doc.height - HEADER_HEIGHT + 4 * mm,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        id="normal"
    )
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

    def p_left(txt):
        return Paragraph(_safe(txt), cell_left)

    def p_center(txt):
        return Paragraph(_safe(txt), cell_center)

    rows = []
    for idx, s in enumerate(students, start=1):
        r = counts_by_student.get(s.id, {})
        present = int(r.get("present") or 0)
        absent = int(r.get("absent") or 0)
        od = int(r.get("od") or 0)
        attended = present + od

        perc = (attended / total_hours * 100.0) if total_hours else 0.0

        if perc >= 75:
            status_text = "Good"
        elif perc >= 50:
            status_text = "Moderate"
        else:
            status_text = "Low"

        reg_no = _safe(getattr(s, "reg_no", ""))
        student_name = _safe(getattr(s, "name", ""))

        rows.append([
            p_center(str(idx)),
            p_center(reg_no),
            p_left(student_name),
            p_center(str(present)),
            p_center(str(od)),
            p_center(str(absent)),
            p_center(str(total_hours or 0)),
            p_center(f"{perc:.2f}%"),
            p_center(status_text),
        ])

    elements = []
    elements.append(Spacer(1, 2 * mm))
    elements.append(Paragraph("COURSE ATTENDANCE REPORT", title_style))
    elements.append(Paragraph(f"{faculty_name} ({faculty_code})", sub_style))

    info_data = [
        [Paragraph("<b>Course:</b>", info_style),
         Paragraph(f"{course.course_code} - {course.title}", info_style),
         Paragraph("<b>Regulation:</b>", info_style),
         Paragraph(str(regulation), info_style)],

        [Paragraph("<b>Department:</b>", info_style),
         Paragraph(degree_department, info_style),
         Paragraph("<b>Year / Sem:</b>", info_style),
         Paragraph(f"{year} / {semester}", info_style)],

        [Paragraph("<b>Batch/Sec:</b>", info_style),
         Paragraph(f"{batch} / {section}", info_style),
         Paragraph("<b>Students:</b>", info_style),
         Paragraph(str(total_students), info_style)],
    ]

    if parsed_from or parsed_to:
        date_range = f"{date_from or '...'} to {date_to or '...'}"
        info_data.append([
            Paragraph("<b>Date Range:</b>", info_style),
            Paragraph(date_range, info_style),
            Paragraph("<b>Total Hrs:</b>", info_style),
            Paragraph(str(total_hours), info_style),
        ])
    else:
        info_data.append([
            Paragraph("<b>Total Hrs:</b>", info_style),
            Paragraph(str(total_hours), info_style),
            Paragraph("<b>Students:</b>", info_style),
            Paragraph(str(total_students), info_style),
        ])

    info_table = Table(info_data, colWidths=[26 * mm, doc.width - 86 * mm, 26 * mm, 34 * mm])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 5 * mm))
    elements.append(Paragraph("STUDENT ATTENDANCE DETAILS", section_style))

    if not rows:
        elements.append(Paragraph("No attendance records found for the selected criteria.", info_style))
    else:
        header_row = [
            Paragraph("S.No", table_header),
            Paragraph("Reg No", table_header),
            Paragraph("Student Name", table_header),
            Paragraph("Hours", table_header),
            Paragraph("OD", table_header),
            Paragraph("A", table_header),
            Paragraph("Total", table_header),
            Paragraph("%", table_header),
            Paragraph("Status", table_header),
        ]

        table_data = [header_row] + rows
        col_widths = calc_col_widths(doc.width)
        attendance_table = make_table(table_data, col_widths, header_bg=SECONDARY_BLUE, zebra=True)
        elements.append(attendance_table)

    try:
        doc.build(elements)
    except Exception as e:
        return HttpResponse(f"PDF generation failed: {e}", status=500)

    buffer.seek(0)
    date_suffix = f"_{date_from}_{date_to}" if (date_from or date_to) else ""
    filename = f"Attendance_{course.course_code}_{batch}_{section}_Y{year}_S{semester}{date_suffix}.pdf"
    return FileResponse(buffer, as_attachment=False, filename=filename)
 

def faculty_course_students_attendance_datewise_pdf(request, year, semester, course_id, batch, section, regulation_id):
    import io
    import os
    from collections import defaultdict
    from datetime import datetime

    from django.http import FileResponse, HttpResponse
    from django.shortcuts import get_object_or_404
    from django.conf import settings
    from django.contrib.staticfiles import finders

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import (
        BaseDocTemplate, PageTemplate, Frame,
        Table, TableStyle, Paragraph, Spacer, PageBreak
    )

    def _safe(v):
        return "" if v is None else str(v).strip()

    class AttendanceDateDocTemplate(BaseDocTemplate):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.current_attendance_date = ""

        def afterFlowable(self, flowable):
            if isinstance(flowable, Paragraph):
                try:
                    txt = flowable.getPlainText().strip()
                except Exception:
                    txt = ""
                if txt.startswith("Attendance Date:"):
                    self.current_attendance_date = txt.replace("Attendance Date:", "").strip()

    faculty_id = getattr(request.user, "Employee_id", None)
    faculty = get_object_or_404(general_information, faculty_id=faculty_id)
    course = get_object_or_404(Course, id=course_id)
    regulation = get_object_or_404(Regulations, id=regulation_id)

    faculty_name = _safe(getattr(faculty, "name", "")) or "Faculty"
    faculty_code = faculty_id or ""

    course_enrollments = (
        CourseEnrollment.objects
        .filter(
            course_id=course_id,
            batch=batch,
            section=section,
            regulation_id=regulation_id,
            faculty=faculty,
            enroll=True
        )
        .select_related("student", "course", "faculty", "department__degree")
        .order_by("student__reg_no")
    )

    students = [e.student for e in course_enrollments]
    student_ids = [s.id for s in students]
    total_students = len(students)
    enrollment = course_enrollments.first()

    degree_department = ""
    if enrollment and enrollment.department and enrollment.department.degree:
        degree_code = _safe(enrollment.department.degree.degree_code)
        department_name = _safe(enrollment.department.Department)
        degree_department = f"{degree_code} - {department_name}"

    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    parsed_from = None
    parsed_to = None

    try:
        if date_from:
            parsed_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        if date_to:
            parsed_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Invalid date format. Use YYYY-MM-DD.", status=400)

    if parsed_from and parsed_to and parsed_from > parsed_to:
        return HttpResponse("'From Date' cannot be greater than 'To Date'.", status=400)

    att_qs = HourAttendance.objects.filter(
        course=course,
        batch=str(batch),
        section=str(section),
        year=str(year),
        semester=str(semester),
    )

    if student_ids:
        att_qs = att_qs.filter(student_id__in=student_ids)
    else:
        att_qs = att_qs.none()

    if parsed_from:
        att_qs = att_qs.filter(date__gte=parsed_from)
    if parsed_to:
        att_qs = att_qs.filter(date__lte=parsed_to)

    att_qs = att_qs.select_related("student").order_by("date", "student__reg_no", "period")

    if not att_qs.exists():
        return HttpResponse("No attendance records found for the selected criteria.", status=404)

    grouped = defaultdict(lambda: defaultdict(dict))
    all_periods = set()

    for row in att_qs:
        if not row.student_id:
            continue
        grouped[row.date][row.student_id][row.period] = row.status
        if row.period:
            all_periods.add(row.period)

    period_list = sorted(all_periods)
    sorted_dates = sorted(grouped.keys())
    students_sorted = sorted(students, key=lambda s: _safe(getattr(s, "reg_no", "")))

    styles = getSampleStyleSheet()

    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1a4b8c")
    ACCENT_RED = colors.HexColor("#b91c1c")
    DARK_GRAY = colors.HexColor("#111827")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BG_GRAY = colors.HexColor("#f8fafc")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle(
        "title_style",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=PRIMARY_BLUE,
        alignment=TA_CENTER,
        spaceAfter=4,
        fontName="Helvetica-Bold",
    )

    sub_style = ParagraphStyle(
        "sub_style",
        parent=styles["Normal"],
        fontSize=10,
        textColor=MEDIUM_GRAY,
        alignment=TA_CENTER,
        spaceAfter=8,
    )

    date_marker_style = ParagraphStyle(
        "date_marker_style",
        parent=styles["Normal"],
        fontSize=1,
        textColor=colors.white,
        alignment=TA_LEFT,
        leading=1,
        spaceBefore=0,
        spaceAfter=0,
    )

    info_style = ParagraphStyle(
        "info_style",
        parent=styles["Normal"],
        fontSize=10,
        textColor=DARK_GRAY,
        alignment=TA_LEFT,
        leading=13,
        fontName="Helvetica",
    )

    table_header = ParagraphStyle(
        "table_header",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.white,
        alignment=TA_CENTER,
        fontName="Helvetica-Bold",
        leading=11,
        wordWrap="CJK",
    )

    cell_left = ParagraphStyle(
        "cell_left",
        parent=styles["Normal"],
        fontSize=8.5,
        textColor=DARK_GRAY,
        alignment=TA_LEFT,
        leading=10,
        wordWrap="CJK",
    )

    cell_center = ParagraphStyle(
        "cell_center",
        parent=cell_left,
        alignment=TA_CENTER,
    )

    buffer = io.BytesIO()
    PAGE = landscape(A4)

    doc = AttendanceDateDocTemplate(
        buffer,
        pagesize=PAGE,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
        title=f"Datewise Attendance - {course.course_code}",
        showBoundary=0
    )

    HEADER_HEIGHT = 26 * mm

    # This fixes page 1 date
    if sorted_dates:
        doc.current_attendance_date = sorted_dates[0].strftime("%d-%m-%Y")

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = PAGE
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 4 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(
                ImageReader(logo_path),
                left, top_y - 15 * mm,
                width=24 * mm, height=14 * mm,
                preserveAspectRatio=True, mask="auto"
            )

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(center_x, top_y - 4 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 8.5)
        canvas.drawCentredString(center_x, top_y - 9 * mm, "An Autonomous Institution")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 7.5)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "Approved by AICTE, New Delhi")
        canvas.setFont("Helvetica", 7)
        canvas.drawCentredString(center_x, top_y - 16.5 * mm, "Accredited by NAAC & ISO 9001:2015 Certified")
        canvas.drawCentredString(center_x, top_y - 20 * mm, "NBA Accredited: CSE, EEE, ECE, MECH")

        # Date on left side of every page, including page 1
        current_date = getattr(doc_, "current_attendance_date", "")
        if current_date:
            canvas.setFillColor(ACCENT_RED)
            canvas.setFont("Helvetica-Bold", 10)
            canvas.drawString(left, top_y - 24.5 * mm, f"Date: {current_date}")

        footer_y = 9 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.6)
        canvas.line(left, footer_y + 5 * mm, right, footer_y + 5 * mm)

        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 7)
        gen_time = datetime.now().strftime("%d %b %Y, %H:%M")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")
        canvas.drawCentredString(center_x, footer_y, f"Faculty: {faculty_code}")
        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")

        canvas.restoreState()

    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin + 3 * mm,
        doc.width,
        doc.height - HEADER_HEIGHT + 4 * mm,
        leftPadding=0,
        rightPadding=0,
        topPadding=0,
        bottomPadding=0,
        id="normal"
    )
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

    def p_left(txt):
        return Paragraph(_safe(txt), cell_left)

    def p_center(txt):
        return Paragraph(_safe(txt), cell_center)

    elements = []
    elements.append(Spacer(1, 2 * mm))
    elements.append(Paragraph("DATE-WISE HOUR ATTENDANCE REPORT", title_style))
    elements.append(Paragraph(f"{faculty_name} ({faculty_code})", sub_style))

    info_data = [
        [Paragraph("<b>Course:</b>", info_style),
         Paragraph(f"{course.course_code} - {course.title}", info_style),
         Paragraph("<b>Regulation:</b>", info_style),
         Paragraph(str(regulation), info_style)],

        [Paragraph("<b>Department:</b>", info_style),
         Paragraph(degree_department, info_style),
         Paragraph("<b>Year / Sem:</b>", info_style),
         Paragraph(f"{year} / {semester}", info_style)],

        [Paragraph("<b>Batch/Sec:</b>", info_style),
         Paragraph(f"{batch} / {section}", info_style),
         Paragraph("<b>Students:</b>", info_style),
         Paragraph(str(total_students), info_style)],

        [Paragraph("<b>Date Range:</b>", info_style),
         Paragraph(f"{date_from or '...'} to {date_to or '...'}", info_style),
         Paragraph("<b>Total Dates:</b>", info_style),
         Paragraph(str(len(sorted_dates)), info_style)],
    ]

    info_table = Table(info_data, colWidths=[26 * mm, doc.width - 86 * mm, 26 * mm, 34 * mm])
    info_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 5 * mm))

    if not sorted_dates:
        elements.append(Paragraph("No attendance records found for the selected criteria.", info_style))
    else:
        for idx, dt in enumerate(sorted_dates, start=1):
            # invisible marker only for updating header date
            elements.append(Paragraph(f"Attendance Date: {dt.strftime('%d-%m-%Y')}", date_marker_style))
            elements.append(Spacer(1, 2 * mm))

            header_row = [
                Paragraph("S.No", table_header),
                Paragraph("Reg No", table_header),
                Paragraph("Student Name", table_header),
            ] + [Paragraph(f"Hour {p}", table_header) for p in period_list]

            rows = [header_row]

            for sno, student in enumerate(students_sorted, start=1):
                period_status_map = grouped[dt].get(student.id, {})

                row = [
                    p_center(str(sno)),
                    p_center(_safe(getattr(student, "reg_no", ""))),
                    p_left(_safe(getattr(student, "name", ""))),
                ]

                for period in period_list:
                    row.append(p_center(_safe(period_status_map.get(period, "-"))))

                rows.append(row)

            base_width = doc.width
            sno_w = 12 * mm
            reg_w = 28 * mm
            name_w = 50 * mm
            remaining = base_width - (sno_w + reg_w + name_w)
            period_w = remaining / max(len(period_list), 1)

            col_widths = [sno_w, reg_w, name_w] + [period_w] * len(period_list)

            date_table = Table(rows, colWidths=col_widths, repeatRows=1)
            date_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.6, BORDER_GRAY),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
                ("ALIGN", (0, 0), (1, -1), "CENTER"),
                ("ALIGN", (3, 1), (-1, -1), "CENTER"),
            ]))

            elements.append(date_table)

            if idx < len(sorted_dates):
                elements.append(PageBreak())

    try:
        doc.build(elements)
    except Exception as e:
        return HttpResponse(f"PDF generation failed: {e}", status=500)

    buffer.seek(0)
    date_suffix = f"_{date_from}_{date_to}" if (date_from or date_to) else ""
    filename = f"Attendance_Datewise_{course.course_code}_{batch}_{section}_Y{year}_S{semester}{date_suffix}.pdf"
    return FileResponse(buffer, as_attachment=False, filename=filename)
 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def faculty_course_students_attendance_datewise_excel(request, year, semester, course_id, batch, section, regulation_id): 
    faculty_id = getattr(request.user, "Employee_id", None)
    faculty = get_object_or_404(general_information, faculty_id=faculty_id)
    course = get_object_or_404(Course, id=course_id)
    regulation = get_object_or_404(Regulations, id=regulation_id)

    course_enrollments = (
        CourseEnrollment.objects
        .filter(
            course_id=course_id,
            batch=batch,
            section=section,
            regulation_id=regulation_id,
            faculty=faculty,
            enroll=True
        )
        .select_related("student")
        .order_by("student__reg_no")
    )

    students = [e.student for e in course_enrollments]
    student_ids = [s.id for s in students]

    date_from = (request.GET.get("from") or "").strip()
    date_to = (request.GET.get("to") or "").strip()

    parsed_from = None
    parsed_to = None

    try:
        if date_from:
            parsed_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        if date_to:
            parsed_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Invalid date format. Use YYYY-MM-DD.", status=400)

    if parsed_from and parsed_to and parsed_from > parsed_to:
        return HttpResponse("'From Date' cannot be greater than 'To Date'.", status=400)

    att_qs = HourAttendance.objects.filter(
        course=course,
        batch=str(batch),
        section=str(section),
        year=str(year),
        semester=str(semester),
    )

    if student_ids:
        att_qs = att_qs.filter(student_id__in=student_ids)
    else:
        att_qs = att_qs.none()

    if parsed_from:
        att_qs = att_qs.filter(date__gte=parsed_from)
    if parsed_to:
        att_qs = att_qs.filter(date__lte=parsed_to)

    att_qs = att_qs.select_related("student").order_by("date", "student__reg_no", "period")

    if not att_qs.exists():
        return HttpResponse("No attendance records found for the selected criteria.", status=404)

    grouped = defaultdict(lambda: defaultdict(dict))
    all_periods_by_date = defaultdict(set)

    for row in att_qs:
        if not row.student_id:
            continue
        grouped[row.student_id][row.date][row.period] = row.status
        if row.period is not None:
            all_periods_by_date[row.date].add(row.period)

    sorted_dates = sorted(all_periods_by_date.keys())
    students_sorted = sorted(students, key=lambda s: _safe(getattr(s, "reg_no", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # highlight for Absent
    absent_fill = PatternFill(fill_type="solid", fgColor="FDE2E1")

    ws.merge_cells("A1:D1")
    ws["A1"] = "DATE-WISE HOUR ATTENDANCE REPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Course"
    ws["B2"] = f"{course.course_code} - {course.title}"
    ws["A3"] = "Faculty"
    ws["B3"] = f"{_safe(getattr(faculty, 'name', ''))} ({_safe(faculty_id)})"
    ws["A4"] = "Batch / Section"
    ws["B4"] = f"{batch} / {section}"
    ws["C2"] = "Regulation"
    ws["D2"] = str(regulation)
    ws["C3"] = "Year / Semester"
    ws["D3"] = f"{year} / {semester}"
    ws["C4"] = "Date Range"
    ws["D4"] = f"{date_from or '...'} to {date_to or '...'}"

    for cell in ["A2", "A3", "A4", "C2", "C3", "C4"]:
        ws[cell].font = bold_font

    header_row_1 = 6
    header_row_2 = 7

    ws.merge_cells(start_row=header_row_1, start_column=1, end_row=header_row_2, end_column=1)
    ws.merge_cells(start_row=header_row_1, start_column=2, end_row=header_row_2, end_column=2)
    ws.merge_cells(start_row=header_row_1, start_column=3, end_row=header_row_2, end_column=3)

    ws.cell(header_row_1, 1).value = "S.No"
    ws.cell(header_row_1, 2).value = "Reg No"
    ws.cell(header_row_1, 3).value = "Student Name"

    current_col = 4
    date_column_map = []

    for dt in sorted_dates:
        periods = sorted(all_periods_by_date[dt])
        start_col = current_col

        for period in periods:
            ws.cell(header_row_2, current_col).value = f"Hour {period}"
            current_col += 1

        end_col = current_col - 1

        if start_col <= end_col:
            ws.merge_cells(
                start_row=header_row_1,
                start_column=start_col,
                end_row=header_row_1,
                end_column=end_col
            )
            ws.cell(header_row_1, start_col).value = dt.strftime("%d-%m-%Y")
            date_column_map.append((dt, periods))

    for row_num in [header_row_1, header_row_2]:
        for col in range(1, current_col):
            cell = ws.cell(row_num, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    data_row = 8
    for idx, student in enumerate(students_sorted, start=1):
        ws.cell(data_row, 1).value = idx
        ws.cell(data_row, 2).value = _safe(getattr(student, "reg_no", ""))
        ws.cell(data_row, 3).value = _safe(getattr(student, "name", ""))

        ws.cell(data_row, 1).alignment = center_align
        ws.cell(data_row, 2).alignment = center_align
        ws.cell(data_row, 3).alignment = left_align

        for fixed_col in [1, 2, 3]:
            ws.cell(data_row, fixed_col).border = thin_border

        col_ptr = 4
        for dt, periods in date_column_map:
            for period in periods:
                value = grouped[student.id].get(dt, {}).get(period, "-")
                cell = ws.cell(data_row, col_ptr)
                cell.value = _safe(value)
                cell.alignment = center_align
                cell.border = thin_border

                if str(value).strip().lower() == "absent":
                    cell.fill = absent_fill

                col_ptr += 1

        data_row += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28

    for col in range(4, current_col):
        ws.column_dimensions[get_column_letter(col)].width = 14

    ws.freeze_panes = "D8"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Attendance_Datewise_{course.course_code}_{batch}_{section}_Y{year}_S{semester}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
 

from django.http import HttpResponse, JsonResponse
from user_accounts.decorators import check_permission
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from datetime import date, datetime
from course_management.models import Course, CourseEnrollment, AssignSubjectFaculty, CoursePlan, Regulations, SectionMaster, FacultySubjectWillingness
from faculty_management.models import general_information
from user_accounts.models import Degree, Add_Department, StudentDetails
from course_management.decorators import course_management
from course_management.decorators import course_management



def _to_int_or_none(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None



def apply_willingness_filters(qs, request):
    filters = {
        "academic_year": request.GET.get("academic_year", "").strip(),
        "regulation": request.GET.get("regulation", "").strip(),
        "year": request.GET.get("year", "").strip(),
        "semester": request.GET.get("semester", "").strip(),
        "status": request.GET.get("status", "").strip(),
        "course": request.GET.get("course", "").strip(),
    }

    if filters["academic_year"]:
        qs = qs.filter(academic_year=filters["academic_year"])

    if filters["regulation"]:
        qs = qs.filter(regulation_id=filters["regulation"])

    if filters["year"]:
        qs = qs.filter(year=filters["year"])

    if filters["semester"]:
        qs = qs.filter(semester=filters["semester"])

    if filters["status"]:
        qs = qs.filter(status=filters["status"])

    if filters["course"]:
        qs = qs.filter(course_id=filters["course"])

    return qs, filters



def apply_willingness_filters(qs, request):
    selected_filters = {
        "academic_year": (request.GET.get("academic_year") or "").strip(),
        "regulation": (request.GET.get("regulation") or "").strip(),
        "year": (request.GET.get("year") or "").strip(),
        "semester": (request.GET.get("semester") or "").strip(),
        "status": (request.GET.get("status") or "").strip(),
        "course": (request.GET.get("course") or "").strip(),
    }

    if selected_filters["academic_year"]:
        qs = qs.filter(academic_year=selected_filters["academic_year"])

    if selected_filters["regulation"]:
        qs = qs.filter(regulation_id=selected_filters["regulation"])

    if selected_filters["year"]:
        qs = qs.filter(year=selected_filters["year"])

    if selected_filters["semester"]:
        qs = qs.filter(semester=selected_filters["semester"])

    if selected_filters["status"]:
        qs = qs.filter(status=selected_filters["status"])

    if selected_filters["course"]:
        qs = qs.filter(course_id=selected_filters["course"])

    return qs, selected_filters


def get_subject_allocation_window_status():
    today = timezone.localdate()
    schedules = (
        SubjectAllocationSchedule.objects
        .filter(is_active=True)
        .order_by("-updated_at")
    )
    schedule = schedules.first()

    if not schedule:
        return {
            "schedule": None,
            "today": today,
            "can_act": False,
            "status_message": "Subject allocation schedule is not configured.",
        }

    can_act = schedule.start_date <= today <= schedule.end_date
    if can_act:
        ay = f" ({schedule.academic_year})" if schedule.academic_year else ""
        status_message = f"Actions are enabled from {schedule.start_date} to {schedule.end_date}{ay}."
    elif today < schedule.start_date:
        status_message = f"Actions will start on {schedule.start_date}."
    else:
        status_message = f"Action window closed on {schedule.end_date}."

    return {
        "schedule": schedule,
        "today": today,
        "can_act": can_act,
        "status_message": status_message,
    }


@check_permission("subject_willingness")
def subject_willingness(request):
    faculty = general_information.objects.filter(
        faculty_id=request.user.Employee_id
    ).select_related("department").first()
    schedule_state = get_subject_allocation_window_status()

    if request.method == "POST":
        if not schedule_state["can_act"]:
            messages.error(request, schedule_state["status_message"])
            return redirect("subject_willingness")

        if not faculty:
            messages.error(request, "Your faculty profile was not found.")
            return redirect("subject_willingness")

        edit_id = request.POST.get("edit_id")

        academic_year = (request.POST.get("academic_year") or "").strip()
        degree_id = _to_int_or_none(request.POST.get("degree"))
        department_id = _to_int_or_none(request.POST.get("department"))
        regulation_id = _to_int_or_none(request.POST.get("regulation"))
        year = (request.POST.get("year") or "").strip()
        semester = (request.POST.get("semester") or "").strip()
        section = (request.POST.get("section") or "").strip()
        batch = (request.POST.get("batch") or "").strip()
        course_id = _to_int_or_none(request.POST.get("course"))
        reason = (request.POST.get("reason") or "").strip()
        handled = (request.POST.get("No_of_time_handled") or "").strip()
        handled_rit = (request.POST.get("No_of_time_handled_in_RIT") or "").strip()
        pass_pct = (request.POST.get("pass_percentage_obtained") or "").strip()

        required = [
            ("academic_year", academic_year),
            ("degree", degree_id),
            ("department", department_id),
            ("regulation", regulation_id),
            ("year", year),
            ("semester", semester),
            ("course", course_id),
        ]

        missing = [k for k, v in required if not v]
        if missing:
            messages.error(request, f"Missing fields: {', '.join(missing)}")
            return redirect("subject_willingness")

        degree_obj = Degree.objects.filter(id=degree_id, is_active=True).first()
        dept_obj = Add_Department.objects.filter(id=department_id, is_active=True).first()
        reg_obj = Regulations.objects.filter(id=regulation_id).first()
        course_obj = Course.objects.filter(id=course_id, is_active=True).first()

        if not all([degree_obj, dept_obj, reg_obj, course_obj]):
            messages.error(request, "Invalid Degree/Department/Regulation/Course selection.")
            return redirect("subject_willingness")

        handled = _to_int_or_none(handled)
        handled_rit = _to_int_or_none(handled_rit)

        with transaction.atomic():
            if edit_id:
                sw = FacultySubjectWillingness.objects.filter(
                    id=edit_id,
                    faculty=faculty
                ).select_for_update().first()

                if not sw:
                    messages.error(request, "Record not found or not permitted to edit.")
                    return redirect("subject_willingness")

                if sw.status == "Approved":
                    messages.warning(request, "Approved records cannot be edited.")
                    return redirect("subject_willingness")

                sw.academic_year = academic_year
                sw.degree = degree_obj
                sw.department = dept_obj
                sw.regulation = reg_obj
                sw.year = year
                sw.semester = semester
                sw.section = section or None
                sw.batch = batch or None
                sw.course = course_obj
                sw.reason = reason
                sw.No_of_time_handled = handled
                sw.No_of_time_handled_in_RIT = handled_rit
                sw.pass_percentage_obtained = pass_pct or None
                sw.status = "Pending"
                sw.save()

                messages.success(request, "Willingness updated.")
                return redirect("subject_willingness")

            existing = FacultySubjectWillingness.objects.filter(
                faculty=faculty,
                course_id=course_id,
                academic_year=academic_year,
                section=section or None,
            ).order_by("-created_at").first()

            if existing and existing.status == "Approved":
                messages.warning(
                    request,
                    "Already submitted and approved for this course and academic year."
                )
                return redirect("subject_willingness")

            if not existing:
                # Provide explicit id when DB tables lack AUTO_INCREMENT
                try:
                    max_id = FacultySubjectWillingness.objects.aggregate(max_id=Max('id'))['max_id'] or 0
                    next_id = int(max_id) + 1
                except Exception:
                    next_id = None

                if next_id:
                    FacultySubjectWillingness.objects.create(
                        id=next_id,
                        faculty=faculty,
                        academic_year=academic_year,
                        degree=degree_obj,
                        department=dept_obj,
                        regulation=reg_obj,
                        year=year,
                        semester=semester,
                        section=section or None,
                        batch=batch or None,
                        course=course_obj,
                        reason=reason,
                        No_of_time_handled=handled,
                        No_of_time_handled_in_RIT=handled_rit,
                        pass_percentage_obtained=pass_pct or None,
                        status="Pending",
                    )
                else:
                    FacultySubjectWillingness.objects.create(
                        faculty=faculty,
                        academic_year=academic_year,
                        degree=degree_obj,
                        department=dept_obj,
                        regulation=reg_obj,
                        year=year,
                        semester=semester,
                        section=section or None,
                        batch=batch or None,
                        course=course_obj,
                        reason=reason,
                        No_of_time_handled=handled,
                        No_of_time_handled_in_RIT=handled_rit,
                        pass_percentage_obtained=pass_pct or None,
                        status="Pending",
                    )
            else:
                existing.degree = degree_obj
                existing.department = dept_obj
                existing.regulation = reg_obj
                existing.year = year
                existing.semester = semester
                existing.section = section or None
                existing.batch = batch or None
                existing.course = course_obj
                existing.reason = reason
                existing.No_of_time_handled = handled
                existing.No_of_time_handled_in_RIT = handled_rit
                existing.pass_percentage_obtained = pass_pct or None
                existing.status = "Pending"
                existing.save()

            messages.success(request, "Willingness submitted.")
            return redirect("subject_willingness")

    cur_year = date.today().year
    ay_options = [f"{cur_year - 1}-{cur_year}", f"{cur_year}-{cur_year + 1}"]

    degrees = Degree.objects.filter(is_active=True).order_by("degree")
    regulations = Regulations.objects.all().order_by("-year")
    sections = list(
        SectionMaster.objects.values_list("section", flat=True)
        .distinct()
        .order_by("section")
    )
    batches = list(
        StudentDetails.objects.values_list("batch", flat=True)
        .distinct()
        .order_by("batch")
    )

    willingness_qs = FacultySubjectWillingness.objects.none()
    selected_filters = {
        "academic_year": "",
        "regulation": "",
        "year": "",
        "semester": "",
        "status": "",
        "course": "",
    }

    filter_options = {
        "academic_years": [],
        "regulations": [],
        "years": [],
        "semesters": [],
        "statuses": [],
        "courses": [],
    }

    if faculty:
        base_qs = (
            FacultySubjectWillingness.objects
            .filter(faculty=faculty)
            .select_related("degree", "department", "regulation", "course")
            .order_by("-created_at")
        )

        filter_options = {
            "academic_years": (
                base_qs.exclude(academic_year__isnull=True)
                .exclude(academic_year="")
                .values_list("academic_year", flat=True)
                .distinct()
                .order_by("-academic_year")
            ),
            "regulations": (
                Regulations.objects.filter(
                    id__in=base_qs.exclude(regulation_id__isnull=True)
                    .values_list("regulation_id", flat=True)
                    .distinct()
                )
                .order_by("-year")
            ),
            "years": (
                base_qs.exclude(year__isnull=True)
                .exclude(year="")
                .values_list("year", flat=True)
                .distinct()
                .order_by("year")
            ),
            "semesters": (
                base_qs.exclude(semester__isnull=True)
                .exclude(semester="")
                .values_list("semester", flat=True)
                .distinct()
                .order_by("semester")
            ),
            "statuses": (
                base_qs.exclude(status__isnull=True)
                .exclude(status="")
                .values_list("status", flat=True)
                .distinct()
                .order_by("status")
            ),
            "courses": (
                Course.objects.filter(
                    id__in=base_qs.exclude(course_id__isnull=True)
                    .values_list("course_id", flat=True)
                    .distinct()
                )
                .order_by("course_code")
            ),
        }

        willingness_qs, selected_filters = apply_willingness_filters(base_qs, request)

    context = {
        "academic_years": ay_options,
        "degrees": degrees,
        "regulations": regulations,
        "sections": sections,
        "batches": batches,
        "willingness_list": list(willingness_qs),
        "filter_options": filter_options,
        "selected_filters": selected_filters,
        "schedule_state": schedule_state,
    }

    return render(request, "course_management/faculty/subject_willingness.html", context)


def sw_get_departments(request):
    degree_id = request.GET.get("degree_id")
    deps = Add_Department.objects.filter(is_active=True)

    if degree_id:
        deps = deps.filter(degree_id=degree_id)

    data = [
        {"id": r["id"], "name": r["Department"]}
        for r in deps.values("id", "Department").order_by("Department")
    ]

    return JsonResponse({"departments": data})


def sw_get_years_semesters(request):
    department_id = request.GET.get("department_id")
    regulation_id = request.GET.get("regulation_id")

    if not department_id or not regulation_id:
        return JsonResponse({"years": [], "semesters": []})

    qs = Course.objects.filter(
        department_id=department_id,
        regulation_id=regulation_id,
        is_active=True
    )

    years = [y for y in qs.values_list("year", flat=True).distinct() if y]
    semesters = [s for s in qs.values_list("semester", flat=True).distinct() if s]

    years_sorted = sorted(years, key=lambda x: int(x) if str(x).isdigit() else str(x))
    sem_sorted = sorted(semesters, key=lambda x: int(x) if str(x).isdigit() else str(x))

    return JsonResponse({"years": years_sorted, "semesters": sem_sorted})


def sw_get_courses(request):
    department_id = request.GET.get("department_id")
    regulation_id = request.GET.get("regulation_id")
    year = request.GET.get("year")
    semester = request.GET.get("semester")

    if not (department_id and regulation_id and year and semester):
        return JsonResponse({"courses": []})

    qs = Course.objects.filter(
        department_id=department_id,
        regulation_id=regulation_id,
        year=year,
        semester=semester,
        is_active=True
    ).values("id", "course_code", "title").distinct().order_by("course_code")

    data = [
        {
            "id": c["id"],
            "label": f"{c['course_code']} - {c['title']}"
        }
        for c in qs if c["id"]
    ]

    return JsonResponse({"courses": data})



from django.urls import reverse
from urllib.parse import urlencode
@check_permission("course_plan")
def course_plan(request):
    faculty = get_object_or_404(general_information, faculty_id=request.user.Employee_id)

    cur_year = date.today().year
    ay_options = [f"{cur_year-1}-{cur_year}", f"{cur_year}-{cur_year+1}"]

    # ✅ values restored after reload (from GET params)
    selected = {
        "academic_year": request.GET.get("academic_year", ""),
        "degree": request.GET.get("degree", ""),
        "department": request.GET.get("department", ""),
        "regulation": request.GET.get("regulation", ""),
        "year": request.GET.get("year", ""),
        "semester": request.GET.get("semester", ""),
        "course": request.GET.get("course", ""),
        "page": request.GET.get("page", "1"),  # ✅ pagination
    }

    def _redirect_with_selected(post_data=None):
        post_data = post_data or {}

        data = {
            "academic_year": post_data.get("academic_year") or selected["academic_year"],
            "degree": post_data.get("degree") or selected["degree"],
            "department": post_data.get("department") or selected["department"],
            "regulation": post_data.get("regulation") or selected["regulation"],
            "year": post_data.get("year") or selected["year"],
            "semester": post_data.get("semester") or selected["semester"],
            "course": post_data.get("course") or selected["course"],

            # ✅ keep current page after edit/delete/create
            "page": post_data.get("page") or selected["page"] or "1",
        }

        data = {k: v for k, v in data.items() if v}
        base_url = reverse("course_plan")
        qs = urlencode(data)
        return redirect(f"{base_url}?{qs}" if qs else base_url)

    # -------------------- DELETE --------------------
    if request.method == "POST" and request.POST.get("action") == "delete":
        plan_id = request.POST.get("plan_id")
        plan = get_object_or_404(CoursePlan, id=plan_id, faculty=faculty)
        plan.delete()
        messages.success(request, "Course plan deleted.")
        return _redirect_with_selected(request.POST)

    # -------------------- EDIT --------------------
    if request.method == "POST" and request.POST.get("action") == "edit":
        plan_id = request.POST.get("plan_id")
        plan = get_object_or_404(CoursePlan, id=plan_id, faculty=faculty)

        plan.academic_year = request.POST.get("academic_year")
        plan.unit_module_no = request.POST.get("unit_module_no")
        plan.co_no = request.POST.get("co_no")
        plan.delivery_method = request.POST.get("delivery_method")
        plan.period_no = request.POST.get("period_no")
        plan.topic = request.POST.get("topic")
        plan.content_beyond_syllabus = request.POST.get("content_beyond_syllabus")
        plan.innovative_practice = request.POST.get("innovative_practice")
        plan.justify = request.POST.get("justify")
        plan.save()

        messages.success(request, "Course plan updated successfully.")
        return _redirect_with_selected(request.POST)

    # -------------------- CREATE --------------------
    if request.method == "POST" and request.POST.get("action") == "create":
        academic_year = request.POST.get("academic_year")
        degree_id = request.POST.get("degree")
        department_id = request.POST.get("department")
        regulation_id = request.POST.get("regulation")
        year = request.POST.get("year")
        semester = request.POST.get("semester")
        course_id = request.POST.get("course")

        unit_module_no = request.POST.get("unit_module_no")
        co_no = request.POST.get("co_no")
        delivery_method = request.POST.get("delivery_method")
        topic = request.POST.get("topic")
        content_beyond_syllabus = request.POST.get("content_beyond_syllabus")
        innovative_practice = request.POST.get("innovative_practice")
        period_no = request.POST.get("period_no")
        justify = request.POST.get("justify")

        missing = [k for k, v in {
            "academic_year": academic_year,
            "degree": degree_id,
            "department": department_id,
            "regulation": regulation_id,
            "year": year,
            "semester": semester,
            "course": course_id,
            "unit_module_no": unit_module_no,
            "co_no": co_no,
            "delivery_method": delivery_method,
            "topic": topic,
        }.items() if not v]

        if missing:
            messages.error(request, f"Missing fields: {', '.join(missing)}")
            return _redirect_with_selected(request.POST)

        course = get_object_or_404(Course, id=course_id)
        dept = get_object_or_404(Add_Department, id=department_id)

        CoursePlan.objects.create(
            faculty=faculty,
            faculty_department=dept,
            course=course,
            academic_year=academic_year,
            unit_module_no=unit_module_no,
            co_no=co_no,
            delivery_method=delivery_method,
            topic=topic,
            content_beyond_syllabus=content_beyond_syllabus,
            innovative_practice=innovative_practice,
            period_no=period_no,
            justify=justify,
        )

        messages.success(request, "Course plan submitted.")
        return _redirect_with_selected(request.POST)

    # -------------------- LOAD PAGE --------------------
    saved_plans_qs = CoursePlan.objects.filter(faculty=faculty).order_by("-id")

    # ✅ Pagination: 5 per page
    paginator = Paginator(saved_plans_qs, 5)
    page_obj = paginator.get_page(selected.get("page") or 1)

    # ✅ base_qs = filters only (no page) to build pagination links in template
    base_params = {
        "academic_year": selected.get("academic_year", ""),
        "degree": selected.get("degree", ""),
        "department": selected.get("department", ""),
        "regulation": selected.get("regulation", ""),
        "year": selected.get("year", ""),
        "semester": selected.get("semester", ""),
        "course": selected.get("course", ""),
    }
    base_params = {k: v for k, v in base_params.items() if v}
    base_qs = urlencode(base_params)

    # ✅ compact page range (nice)
    total_pages = paginator.num_pages
    current = page_obj.number
    start = max(current - 2, 1)
    end = min(current + 2, total_pages)
    page_range = list(range(start, end + 1))

    context = {
        "degrees": Degree.objects.filter(is_active=True),
        "regulations": Regulations.objects.all().order_by("-year"),
        "academic_years": ay_options,

        # ✅ page-wise plans
        "page_obj": page_obj,

        # ✅ pagination qs helpers
        "base_qs": base_qs,
        "page_range": page_range,

        # ✅ pass selected values to JS restore
        "selected": selected,
    }
    return render(request, "course_management/faculty/course_plan.html", context)


def cp_get_departments(request):
    degree_id = request.GET.get("degree_id")
    departments = Add_Department.objects.filter(degree_id=degree_id, is_active=True).values("id", "Department")
    data = [{"id": d["id"], "name": d["Department"]} for d in departments]
    return JsonResponse({"departments": data})


def cp_get_years_semesters(request):
    department_id = request.GET.get("department_id")
    regulation_id = request.GET.get("regulation_id")
    faculty_id = request.user.Employee_id

    qs = AssignSubjectFaculty.objects.filter(
        Q(faculty__faculty_id=faculty_id) | Q(skilled_faculty__faculty_id=faculty_id),
        is_active=True
    )

    if department_id:
        qs = qs.filter(department_id=department_id)
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)

    years = sorted(list(qs.values_list("course__year", flat=True).distinct()))
    semesters = sorted(list(qs.values_list("course__semester", flat=True).distinct()))

    return JsonResponse({"years": years, "semesters": semesters})


def cp_get_courses(request):
    faculty_id = request.user.Employee_id
    department_id = request.GET.get("department_id")
    regulation_id = request.GET.get("regulation_id")
    year = request.GET.get("year")
    semester = request.GET.get("semester")

    qs = AssignSubjectFaculty.objects.filter(
        Q(faculty__faculty_id=faculty_id) | Q(skilled_faculty__faculty_id=faculty_id),
        is_active=True
    )

    if department_id:
        qs = qs.filter(department_id=department_id)
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if year:
        qs = qs.filter(course__year=year)
    if semester:
        qs = qs.filter(course__semester=semester)

    courses = qs.values("course__id", "course__course_code", "course__title").distinct()
    data = [{"id": c["course__id"], "label": f"{c['course__course_code']} - {c['course__title']}"} for c in courses]
    return JsonResponse({"courses": data})
import io
import os
from datetime import datetime
from itertools import groupby

from django.conf import settings
from django.contrib import messages
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.staticfiles import finders

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame,
    Table, TableStyle, Paragraph, Spacer, PageBreak, KeepTogether
)

from faculty_management.models import *


# ----------------------------
# Helpers
# ----------------------------
def safe_str(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def is_filled(val) -> bool:
    s = safe_str(val).strip()
    return s not in {"", "-", "0", "0.0"}


def digits_int(s: str, default=9999):
    d = "".join(ch for ch in safe_str(s) if ch.isdigit())
    return int(d) if d else default


def normalize_code(prefix: str, raw: str, fallback_index: int):
    s = safe_str(raw).upper().replace(" ", "")
    digits = "".join(ch for ch in s if ch.isdigit())
    if digits:
        return f"{prefix}{int(digits)}"
    return f"{prefix}{fallback_index}"


def get_po_val_raw(mapping_obj, n: int):
    return safe_str(getattr(mapping_obj, f"po_number_{n}", ""))


def get_pso_val_raw(mapping_obj, n: int):
    return safe_str(getattr(mapping_obj, f"pso_number_{n}", ""))


def get_co_desc_from_mapping(m):
    return safe_str(getattr(m, "co_description", "")) or safe_str(getattr(m, "co_descrption", ""))


def get_co_code_from_mapping(m):
    return safe_str(getattr(getattr(m, "co_number", None), "co_code", "")) or safe_str(getattr(m, "co_number_id", ""))


def first_filled(*vals):
    for v in vals:
        if is_filled(v):
            return safe_str(v)
    return ""


def parse_level_int(val):
    """
    Mapping values usually like 1/2/3. This converts to int.
    Ignores '-', '', '0', '0.0', non-numeric.
    """
    s = safe_str(val)
    if not is_filled(s):
        return None
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits:
        return None
    try:
        return int(digits)
    except Exception:
        return None


def fmt_avg(x):
    """show like 1.5 (one decimal); if integer show without .0"""
    try:
        if x is None:
            return "-"
        if abs(x - int(x)) < 1e-9:
            return str(int(x))
        return f"{x:.1f}"
    except Exception:
        return "-"


# ----------------------------
# Main PDF View
# ----------------------------
def course_plan_pdf(request):
    faculty = general_information.objects.filter(
        faculty_id=request.user.Employee_id
    ).select_related("department").first()

    if not faculty:
        messages.error(request, "Faculty profile not found.")
        return HttpResponse("Faculty profile not found.", status=404)

    course_id = request.GET.get("course_id")
    if not course_id:
        messages.error(request, "Please select a course to download PDF.")
        return HttpResponse("Course not selected.", status=400)

    course_obj = get_object_or_404(Course, id=course_id)
    dept = getattr(faculty, "department", None)
    academic_year = safe_str(request.GET.get("academic_year", ""))

    show_revised_only = bool(getattr(course_obj, "mapping", False))

    # ✅ CourseHours (YOUR model)
    course_hours = (
        CourseHours.objects
        .select_related("hour_config")
        .filter(course=course_obj)
        .order_by("-id")
        .first()
    )

    # ------------------------------------------
    # Queries
    # -----------------------------------------
    visions = Vision.objects.none()
    missions = Mission.objects.none()
    peos = Program_Educational_Objective.objects.none()
    psos = Program_specific_Outcomes.objects.none()

    if dept:
        visions = Vision.objects.filter(department=dept, is_active=True).order_by("-year", "-created_at")
        missions = Mission.objects.filter(department=dept, is_active=True).order_by("-year", "-created_at")
        peos = Program_Educational_Objective.objects.filter(department=dept, is_active=True).order_by("-year", "id")
        psos = Program_specific_Outcomes.objects.filter(department=dept, is_active=True).order_by("-year", "id")

    pos_all = Program_outcomes.objects.filter(is_active=True).order_by("id")
    pos_revised = pos_all.filter(is_revised=True)[:13]
    pos_non_revised = pos_all.filter(is_revised=False)[:13]

    mappings = Co_Po_Mapping.objects.filter(course=course_obj).select_related("co_number").order_by("id")

    assessments = (
        Assessment_master.objects
        .filter(course=course_obj)
        .select_related("internal_assessment", "co_code", "level_code", "assessment")
        .order_by("internal_assessment_id", "module", "id")
    )

    # -----------------------------------------
    # Styles / Colors
    # -----------------------------------------
    styles = getSampleStyleSheet()

    PRIMARY_BLUE = colors.HexColor("#0f2f57")
    SECONDARY_BLUE = colors.HexColor("#1a4b8c")
    ACCENT_RED = colors.HexColor("#b91c1c")
    DARK_GRAY = colors.HexColor("#111827")
    MEDIUM_GRAY = colors.HexColor("#4b5563")
    LIGHT_GRAY = colors.HexColor("#9ca3af")
    BG_GRAY = colors.HexColor("#f8fafc")
    BORDER_GRAY = colors.HexColor("#e5e7eb")

    title_style = ParagraphStyle(
        "title_style", parent=styles["Heading1"],
        fontSize=16, textColor=PRIMARY_BLUE, alignment=TA_CENTER,
        spaceAfter=8, fontName="Helvetica-Bold", leading=20
    )
    course_code_style = ParagraphStyle(
        "course_code_style", parent=styles["Normal"],
        fontSize=11, textColor=SECONDARY_BLUE, alignment=TA_CENTER,
        spaceAfter=2, fontName="Helvetica-Bold"
    )
    course_title_style = ParagraphStyle(
        "course_title_style", parent=styles["Normal"],
        fontSize=12, textColor=DARK_GRAY, alignment=TA_CENTER,
        spaceAfter=6, fontName="Helvetica-Bold"
    )

    section_style = ParagraphStyle(
        "section_style", parent=styles["Heading2"],
        fontSize=12.6, textColor=PRIMARY_BLUE, alignment=TA_LEFT,
        spaceBefore=8, spaceAfter=6, fontName="Helvetica-Bold", leading=16
    )
    subsection_style = ParagraphStyle(
        "subsection_style", parent=styles["Heading3"],
        fontSize=11, textColor=SECONDARY_BLUE, alignment=TA_LEFT,
        spaceBefore=6, spaceAfter=5, fontName="Helvetica-Bold", leading=14
    )
    body_style = ParagraphStyle(
        "body_style", parent=styles["Normal"],
        fontSize=9.8, leading=13, textColor=DARK_GRAY,
        alignment=TA_JUSTIFY, spaceAfter=4
    )

    table_header_style = ParagraphStyle(
        "table_header_style", parent=styles["Normal"],
        fontSize=9.2, textColor=colors.white, alignment=TA_CENTER,
        fontName="Helvetica-Bold", leading=11
    )
    table_cell_style = ParagraphStyle(
        "table_cell_style", parent=styles["Normal"],
        fontSize=9.2, textColor=DARK_GRAY, alignment=TA_LEFT,
        leading=11, wordWrap="CJK"
    )
    table_cell_center_style = ParagraphStyle(
        "table_cell_center_style", parent=table_cell_style, alignment=TA_CENTER
    )

    tiny_header_style = ParagraphStyle(
        "tiny_header_style", parent=styles["Normal"],
        fontSize=7.3, textColor=colors.white, alignment=TA_CENTER,
        fontName="Helvetica-Bold", leading=9
    )
    tiny_cell_center = ParagraphStyle(
        "tiny_cell_center", parent=styles["Normal"],
        fontSize=7.1, textColor=DARK_GRAY, alignment=TA_CENTER, leading=9
    )
    tiny_cell_left = ParagraphStyle(
        "tiny_cell_left", parent=styles["Normal"],
        fontSize=7.1, textColor=DARK_GRAY, alignment=TA_LEFT,
        leading=9, wordWrap="CJK"
    )

    info_label_style = ParagraphStyle(
        "info_label_style", parent=styles["Normal"],
        fontSize=9.6, textColor=MEDIUM_GRAY, fontName="Helvetica-Bold", leading=12
    )
    info_value_style = ParagraphStyle(
        "info_value_style", parent=styles["Normal"],
        fontSize=9.6, textColor=DARK_GRAY, leading=12
    )

    ass_module_heading_style = ParagraphStyle(
        "ass_module_heading_style",
        parent=subsection_style,
        fontSize=10.5,
        textColor=SECONDARY_BLUE,
        spaceBefore=4,
        spaceAfter=4
    )

    summary_cell_style = ParagraphStyle(
        "summary_cell_style",
        parent=styles["Normal"],
        fontSize=8.8,
        leading=10.5,
        textColor=DARK_GRAY,
        alignment=TA_CENTER
    )

    note_style = ParagraphStyle(
        "note_style",
        parent=styles["Normal"],
        fontSize=9.2,
        leading=12,
        textColor=MEDIUM_GRAY,
        alignment=TA_LEFT,
        spaceAfter=4
    )

    # -----------------------------------------
    # Document Setup
    # -----------------------------------------
    buffer = io.BytesIO()
    doc = BaseDocTemplate(
        buffer,
        pagesize=A4,
        title=f"Course Plan - {safe_str(getattr(course_obj, 'course_code', ''))}",
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        showBoundary=0
    )

    # -----------------------------------------
    # Table Helpers
    # -----------------------------------------
    def create_table(data, col_widths, header_bg=PRIMARY_BLUE, zebra=True):
        t = Table(data, repeatRows=1, colWidths=col_widths)
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9.2),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),

            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("FONTSIZE", (0, 1), (-1, -1), 9.2),
            ("VALIGN", (0, 1), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ])
        if zebra and len(data) > 1:
            style.add("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY])
        t.setStyle(style)
        return t

    def create_tiny_matrix(data, col_widths, header_bg=PRIMARY_BLUE):
        t = Table(data, repeatRows=1, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7.3),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.35, BORDER_GRAY),
            ("FONTSIZE", (0, 1), (-1, -1), 7.1),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        return t

    def build_mapping_avg_box(
        title: str,
        co_rows: list,
        keep_indices: list,
        col_labels: list,
        denom: int,
        col_widths: list,
    ):
        """
        ✅ ONLY AVERAGE (like 1.5)
        avg = (sum of entered levels in each column) / (total number of CO rows entered in that mapping)
        """
        totals = [0] * len(keep_indices)
        for _, vals in co_rows:
            for j, idx in enumerate(keep_indices):
                v = vals[idx] if idx < len(vals) else ""
                n = parse_level_int(v)
                if n is not None:
                    totals[j] += n

        avgs = []
        for tval in totals:
            avgs.append((tval / denom) if denom > 0 else 0)

        head = [Paragraph(title, tiny_header_style)] + [Paragraph(lbl, tiny_header_style) for lbl in col_labels]
        row = [Paragraph(f"Avg ", table_cell_center_style)]
        for a in avgs:
            row.append(Paragraph(f"<b>{fmt_avg(a)}</b>", summary_cell_style))

        data = [head, row]

        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7.3),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.7, BORDER_GRAY),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("BOX", (0, 0), (-1, -1), 1.2, colors.HexColor("#111827")),
        ]))
        return tbl

    def build_simple_list_table(title: str, items, text_getter, max_rows=4):
        data = [[Paragraph("S.No", table_header_style), Paragraph("Statement", table_header_style)]]
        added = 0
        for obj in list(items)[: max_rows * 3]:
            if added >= max_rows:
                break
            txt = safe_str(text_getter(obj))
            if not is_filled(txt):
                continue
            added += 1
            data.append([
                Paragraph(str(added), table_cell_center_style),
                Paragraph(txt, table_cell_style),
            ])

        block = [Paragraph(title, subsection_style)]
        if added == 0:
            block.append(Paragraph(f"No {title} data available.", body_style))
        else:
            block.append(create_table(data, [14 * mm, doc.width - 14 * mm], header_bg=SECONDARY_BLUE))
        block.append(Spacer(1, 4 * mm))
        return block

    def build_code_list_table(title: str, prefix: str, items, code_getter, text_getter, max_rows=5):
        data = [[Paragraph(prefix, table_header_style), Paragraph("Statement", table_header_style)]]
        added = 0
        sorted_items = sorted(list(items), key=lambda x: digits_int(code_getter(x), default=9999))

        for obj in sorted_items:
            if added >= max_rows:
                break
            stmt = safe_str(text_getter(obj))
            if not is_filled(stmt):
                continue
            added += 1
            code = normalize_code(prefix, safe_str(code_getter(obj)), added)
            data.append([
                Paragraph(code, table_cell_center_style),
                Paragraph(stmt, table_cell_style)
            ])

        block = [Paragraph(title, subsection_style)]
        if added == 0:
            block.append(Paragraph(f"No {title} data available.", body_style))
        else:
            block.append(create_table(data, [26 * mm, doc.width - 26 * mm], header_bg=SECONDARY_BLUE))
        block.append(Spacer(1, 4 * mm))
        return block

    def kv_table(rows_2col, col_widths=None, header_title="Field"):
        data = [[Paragraph(header_title, table_header_style), Paragraph("Value", table_header_style)]]
        for k, v in rows_2col:
            data.append([Paragraph(safe_str(k), info_label_style), Paragraph(safe_str(v) or "-", info_value_style)])

        if not col_widths:
            col_widths = [40 * mm, doc.width - 40 * mm]

        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
        ]))
        return t

    def note_box(lines):
        data = [[Paragraph("NOTE", table_header_style)]]
        for ln in lines:
            data.append([Paragraph(safe_str(ln), note_style)])

        t = Table(data, colWidths=[doc.width], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6b7280")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#fff7ed")),
        ]))
        return t

    # -----------------------------------------
    # Header/Footer
    # -----------------------------------------
    HEADER_HEIGHT = 36 * mm

    def draw_header_footer(canvas, doc_):
        canvas.saveState()
        page_w, page_h = A4
        left = doc_.leftMargin
        right = page_w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = page_h - 8 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand

        if logo_path and os.path.exists(logo_path):
            canvas.drawImage(
                ImageReader(logo_path),
                left, top_y - 20 * mm,
                width=30 * mm, height=18 * mm,
                preserveAspectRatio=True, mask="auto"
            )

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawCentredString(center_x, top_y - 6 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawCentredString(center_x, top_y - 13 * mm, "An Autonomous Institution")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 8.2)
        canvas.drawCentredString(center_x, top_y - 18.5 * mm, "Approved by AICTE, New Delhi")
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(center_x, top_y - 23 * mm, "Accredited by NAAC & ISO 9001:2015 Certified Institution")
        canvas.drawCentredString(center_x, top_y - 27.5 * mm, "NBA Accredited UG Programs: CSE, EEE, ECE and MECH")

        footer_y = 18 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.8)
        canvas.line(left, footer_y + 7 * mm, right, footer_y + 7 * mm)

        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 8)
        gen_time = datetime.now().strftime("%d %b %Y, %I:%M %p")
        canvas.drawString(left, footer_y, f"Generated: {gen_time}")

        ccode = safe_str(getattr(course_obj, "course_code", ""))
        if ccode:
            canvas.drawCentredString(center_x, footer_y, f"Course: {ccode}")

        canvas.drawRightString(right, footer_y, f"Page {doc_.page}")
        canvas.restoreState()

    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin + 6 * mm,
        doc.width,
        doc.height - HEADER_HEIGHT + 8 * mm,
        leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
        id="normal"
    )
    doc.addPageTemplates([PageTemplate(id="All", frames=[frame], onPage=draw_header_footer)])

    elements = []

    # =========================================================
    # PAGE 1: GENERAL INFO (YOUR MODELS)
    # =========================================================
    elements.append(Spacer(1, 5 * mm))
    elements.append(Paragraph("COURSE PLAN REPORT", title_style))

    course_code = safe_str(course_obj.course_code) or "N/A"
    course_title = safe_str(course_obj.title) or "N/A"

    elements.append(Paragraph(course_code.upper(), course_code_style))
    elements.append(Paragraph(course_title, course_title_style))

    if academic_year:
        elements.append(Paragraph(
            f"Academic Year: {academic_year}",
            ParagraphStyle(
                "academic_year_style",
                parent=styles["Normal"],
                fontSize=10,
                textColor=MEDIUM_GRAY,
                alignment=TA_CENTER,
                spaceAfter=6
            )
        ))

    dept_name = safe_str(getattr(dept, "Department", "")) if dept else "-"
    fac_name = safe_str(getattr(faculty, "name", "")) or "-"
    fac_id = safe_str(getattr(faculty, "faculty_id", "") or getattr(faculty, "Employee_id", "")) or "-"
    fac_desig = safe_str(getattr(faculty, "designation", "") or getattr(faculty, "Designation", "") or getattr(faculty, "desgination", "")) or "-"

    regulation_str = safe_str(course_obj.regulation) if course_obj.regulation else "-"
    course_year = safe_str(course_obj.year) or "-"
    course_semester = safe_str(course_obj.semester) or "-"

    category_name = safe_str(course_obj.elective.Course_category_name) if course_obj.elective else "-"
    category_code = safe_str(course_obj.elective.category_code) if course_obj.elective else "-"
    category_desc = safe_str(course_obj.elective.category_description) if course_obj.elective else "-"

    L = safe_str(getattr(course_hours, "leture_npwk", "")) if course_hours else "-"
    T = safe_str(getattr(course_hours, "tutorial_npwk", "")) if course_hours else "-"
    P = safe_str(getattr(course_hours, "laboratory_npwk", "")) if course_hours else "-"
    TOTAL_HRS = safe_str(getattr(course_hours, "total_hours", "")) if course_hours else "-"
    CREDITS = safe_str(getattr(course_hours, "credits", "")) if course_hours else "-"

    hour_cfg = safe_str(getattr(getattr(course_hours, "hour_config", None), "config_name", "")) if course_hours else ""
    if not is_filled(hour_cfg):
        hour_cfg = safe_str(getattr(getattr(course_hours, "hour_config", None), "name", "")) if course_hours else ""

    elements.append(Paragraph("GENERAL DETAILS", section_style))

    faculty_rows = [
        ("Department", dept_name),
        ("Faculty Name", fac_name),
        ("Faculty ID", fac_id),
        ("Designation", fac_desig),
        ("Date", datetime.now().strftime("%d-%b-%Y")),
    ]
    elements.append(kv_table(faculty_rows, header_title="Faculty / Department"))
    elements.append(Spacer(1, 4 * mm))

    course_rows = [
        ("Course Code", course_code),
        ("Course Title", course_title),
        ("Regulation", regulation_str),
        ("Year", course_year),
        ("Semester", course_semester),
        ("Category Name", category_name),
        ("L - T - P (per week)", f"{L} - {T} - {P}"),
        ("Total Hours", TOTAL_HRS),
        ("Credits", CREDITS),
    ]
    if is_filled(hour_cfg):
        course_rows.insert(8, ("Hour Configuration", hour_cfg))

    elements.append(kv_table(course_rows, header_title="Course Information"))
    elements.append(Spacer(1, 4 * mm))

   

    elements.append(PageBreak())

    # =========================================================
    # PAGE 2: VISION + MISSION + PEO + PSO
    # =========================================================
    elements.append(Paragraph("DEPARTMENT VISION / MISSION / PEO / PSO", section_style))
    elements.append(Paragraph(
        f"Department: {safe_str(faculty.department)}",
        ParagraphStyle("dept_line", parent=styles["Normal"], fontSize=10, textColor=MEDIUM_GRAY, spaceAfter=6)
    ))

    page2_blocks = []
    page2_blocks += build_simple_list_table("VISION", list(visions)[:10], lambda v: getattr(v, "vision_statement", ""), max_rows=4)
    page2_blocks += build_simple_list_table("MISSION", list(missions)[:15], lambda m: getattr(m, "mission_statement", ""), max_rows=4)
    page2_blocks += build_code_list_table(
        "PROGRAM EDUCATIONAL OBJECTIVES (PEO)", "PEO", list(peos),
        code_getter=lambda p: getattr(p, "peo_code", ""),
        text_getter=lambda p: getattr(p, "peo_statement", ""),
        max_rows=5
    )
    page2_blocks += build_code_list_table(
        "PROGRAM SPECIFIC OUTCOMES (PSO)", "PSO", list(psos),
        code_getter=lambda p: getattr(p, "pso_code", ""),
        text_getter=lambda p: getattr(p, "pso_statement", ""),
        max_rows=5
    )

    elements.append(KeepTogether(page2_blocks))
    elements.append(PageBreak())

    # =========================================================
    # PAGE 3: PO TABLE
    # =========================================================
    if show_revised_only:
        elements.append(Paragraph("PROGRAM OUTCOMES (REVISED)", section_style))
        pos = pos_revised
        no_msg = "No Revised PO data available."
    else:
        elements.append(Paragraph("PROGRAM OUTCOMES (NON-REVISED)", section_style))
        pos = pos_non_revised
        no_msg = "No Non-Revised PO data available."

    po_data = [[Paragraph("PO Code", table_header_style), Paragraph("Description", table_header_style)]]
    po_index = 0
    for po in pos:
        desc = safe_str(getattr(po, "program_description", ""))
        if not is_filled(desc):
            continue
        po_index += 1
        po_data.append([Paragraph(f"PO{po_index}", table_cell_center_style), Paragraph(desc, table_cell_style)])
        if po_index >= 13:
            break

    if po_index == 0:
        elements.append(Paragraph(no_msg, body_style))
    else:
        elements.append(create_table(po_data, [24 * mm, doc.width - 24 * mm], header_bg=SECONDARY_BLUE))

    elements.append(PageBreak())

    # =========================================================
    # CO + CO-PO/PSO MAPPING (SAME PAGE) + ✅ AVG BOXES
    # =========================================================
    elements.append(Paragraph("COURSE OUTCOMES (CO)", section_style))

    co_data = [[Paragraph("CO Code", table_header_style), Paragraph("Description", table_header_style)]]
    co_added = 0
    seen = set()

    for m in mappings:
        co_code = get_co_code_from_mapping(m)
        co_desc = get_co_desc_from_mapping(m)
        if not is_filled(co_desc):
            continue
        key = (co_code, co_desc)
        if key in seen:
            continue
        seen.add(key)
        co_data.append([Paragraph(co_code or "", table_cell_center_style), Paragraph(co_desc, table_cell_style)])
        co_added += 1

    if co_added == 0:
        elements.append(Paragraph("No CO description found for this course.", body_style))
    else:
        elements.append(create_table(co_data, [26 * mm, doc.width - 26 * mm], header_bg=SECONDARY_BLUE))

    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph("CO - PO / PSO MAPPING", section_style))

    if not mappings.exists():
        elements.append(Paragraph("No CO-PO/PSO mapping found for this course.", body_style))
    else:
        # -------------------------
        # A) CO - PO Mapping
        # -------------------------
        elements.append(Paragraph("A) CO - PO Mapping", subsection_style))

        raw_rows_po = []
        for m in mappings:
            co_txt = get_co_code_from_mapping(m) or safe_str(m.co_number_id)
            po_vals = [get_po_val_raw(m, n) for n in range(1, 14)]
            if any(is_filled(v) for v in po_vals):
                raw_rows_po.append((co_txt, po_vals))

        if not raw_rows_po:
            elements.append(Paragraph("No CO-PO entries available.", body_style))
        else:
            po_keep = [i for i in range(13) if any(is_filled(r[1][i]) for r in raw_rows_po)]
            header = [Paragraph("CO", tiny_header_style)] + [Paragraph(f"PO{i+1}", tiny_header_style) for i in po_keep]
            matrix = [header]

            for co_txt, po_vals in raw_rows_po:
                row = [Paragraph(co_txt or "", tiny_cell_left)]
                for i in po_keep:
                    row.append(Paragraph(safe_str(po_vals[i]), tiny_cell_center))
                matrix.append(row)

            denom_po = len(raw_rows_po)  # ✅ number of CO entered (for this mapping)

            co_w = 22 * mm
            col_w = (doc.width - co_w) / max(1, len(po_keep))
            widths = [co_w] + [col_w] * len(po_keep)

            elements.append(create_tiny_matrix(matrix, widths, header_bg=SECONDARY_BLUE))
            elements.append(Spacer(1, 2 * mm))

            # ✅ ONLY AVG values (1.5 etc.)
            po_labels = [f"PO{i+1}" for i in po_keep]
            elements.append(build_mapping_avg_box(
                title="Average",
                co_rows=raw_rows_po,
                keep_indices=po_keep,
                col_labels=po_labels,
                denom=denom_po if denom_po > 0 else 1,
                col_widths=widths
            ))
            elements.append(Spacer(1, 6 * mm))

        # -------------------------
        # B) CO - PSO Mapping
        # -------------------------
        elements.append(Paragraph("B) CO - PSO Mapping", subsection_style))

        raw_rows_pso = []
        for m in mappings:
            co_txt = get_co_code_from_mapping(m) or safe_str(m.co_number_id)
            pso_vals = [get_pso_val_raw(m, n) for n in range(1, 6)]
            if any(is_filled(v) for v in pso_vals):
                raw_rows_pso.append((co_txt, pso_vals))

        if not raw_rows_pso:
            elements.append(Paragraph("No CO-PSO entries available.", body_style))
        else:
            pso_keep = [i for i in range(5) if any(is_filled(r[1][i]) for r in raw_rows_pso)]
            header = [Paragraph("CO", tiny_header_style)] + [Paragraph(f"PSO{i+1}", tiny_header_style) for i in pso_keep]
            matrix = [header]

            for co_txt, pso_vals in raw_rows_pso:
                row = [Paragraph(co_txt or "", tiny_cell_left)]
                for i in pso_keep:
                    row.append(Paragraph(safe_str(pso_vals[i]), tiny_cell_center))
                matrix.append(row)

            denom_pso = len(raw_rows_pso)  # ✅ number of CO entered (for this mapping)

            co_w = 30 * mm
            col_w = (doc.width - co_w) / max(1, len(pso_keep))
            widths = [co_w] + [col_w] * len(pso_keep)

            elements.append(create_tiny_matrix(matrix, widths, header_bg=SECONDARY_BLUE))
            elements.append(Spacer(1, 2 * mm))

            # ✅ ONLY AVG values (1.5 etc.)
            pso_labels = [f"PSO{i+1}" for i in pso_keep]
            elements.append(build_mapping_avg_box(
                title="Average",
                co_rows=raw_rows_pso,
                keep_indices=pso_keep,
                col_labels=pso_labels,
                denom=denom_pso if denom_pso > 0 else 1,
                col_widths=widths
            ))

    elements.append(PageBreak())

    # =========================================================
    # COURSE PLAN (GROUPED)
    # =========================================================
    elements.append(Paragraph("COURSE PLAN", section_style))

    cp_qs = CoursePlan.objects.filter(faculty=faculty, course=course_obj).order_by(
        "unit_module_no", "co_no", "id"
    )

    cp_rows = []
    for p in cp_qs:
        unit = safe_str(getattr(p, "unit_module_no", ""))
        co_no = safe_str(getattr(p, "co_no", ""))
        delivery = safe_str(getattr(p, "delivery_method", ""))
        topic = safe_str(getattr(p, "topic", ""))
        period = safe_str(getattr(p, "period_no", ""))

        if not any(is_filled(x) for x in [unit, co_no, delivery, topic, period]):
            continue

        cp_rows.append((unit.strip(), co_no.strip(), delivery, topic, period))

    if not cp_rows:
        elements.append(Paragraph("No Course Plan entries found for this course.", body_style))
    else:
        cp_data = [[
            Paragraph("S.No", table_header_style),
            Paragraph("Unit/Module", table_header_style),
            Paragraph("CO", table_header_style),
            Paragraph("Delivery Method", table_header_style),
            Paragraph("Topic", table_header_style),
            Paragraph("Period No", table_header_style),
        ]]

        sno = 0
        for (unit_key, co_key), group_items in groupby(cp_rows, key=lambda x: (x[0], x[1])):
            group_items = list(group_items)
            for idx, (_, __, delivery, topic, period) in enumerate(group_items):
                sno += 1
                unit_cell = unit_key if idx == 0 else ""
                co_cell = co_key if idx == 0 else ""

                cp_data.append([
                    Paragraph(str(sno), table_cell_center_style),
                    Paragraph(unit_cell, table_cell_center_style),
                    Paragraph(co_cell, table_cell_center_style),
                    Paragraph(delivery or "", table_cell_style),
                    Paragraph(topic or "", table_cell_style),
                    Paragraph(period or "", table_cell_center_style),
                ])

        elements.append(create_table(
            cp_data,
            col_widths=[12 * mm, 22 * mm, 14 * mm, 38 * mm, doc.width - (12 + 22 + 14 + 38 + 16) * mm, 16 * mm],
            header_bg=SECONDARY_BLUE
        ))

    elements.append(PageBreak())

    # =========================================================
    # ASSESSMENT PLAN (IA -> Module -> separate table)
    # =========================================================
    elements.append(Paragraph("ASSESSMENT PLAN", section_style))

    if not assessments.exists():
        elements.append(Paragraph("No Assessment entries found for this course.", body_style))
    else:
        ia_groups = {}
        for a in assessments:
            ia = getattr(a, "internal_assessment", None)
            iat_raw = safe_str(getattr(ia, "iat", ""))
            d = "".join(ch for ch in iat_raw if ch.isdigit())

            if d:
                ia_name = f"Internal Assessment {int(d)}"
            else:
                ia_name = (
                    safe_str(getattr(ia, "internal_name", "")) or
                    safe_str(getattr(ia, "internal_assessment_name", "")) or
                    safe_str(getattr(a, "internal_assessment_id", "")) or
                    "Internal Assessment"
                )

            ia_groups.setdefault(ia_name, []).append(a)

        def ia_sort_key(name: str):
            d = "".join(ch for ch in name if ch.isdigit())
            return int(d) if d else 9999

        any_added = 0

        for ia_name in sorted(ia_groups.keys(), key=ia_sort_key):
            rows = ia_groups[ia_name]

            cleaned = []
            for a in rows:
                module = safe_str(getattr(a, "module", "")).strip()

                ass_name = (
                    safe_str(getattr(a, "customAssessmentname", "")) or
                    safe_str(getattr(a, "Assessmentname", "")) or
                    safe_str(getattr(getattr(a, "assessment", None), "assessment_name", ""))
                )
                co_code = safe_str(getattr(getattr(a, "co_code", None), "co_code", ""))
                bloom = (
                    safe_str(getattr(getattr(a, "level_code", None), "level_code", "")) or
                    safe_str(getattr(getattr(a, "level_code", None), "level_name", ""))
                )
                maxm = safe_str(getattr(a, "Maxmarks", ""))
                wt = safe_str(getattr(a, "weightage", ""))

                if not any(is_filled(x) for x in [module, ass_name, co_code, bloom, maxm, wt]):
                    continue

                cleaned.append((module, ass_name, co_code, bloom, maxm, wt))

            if not cleaned:
                continue

            any_added += 1
            elements.append(Paragraph(f"{any_added}) {ia_name}", subsection_style))
            elements.append(Spacer(1, 2 * mm))

            def mod_key(item):
                m = item[0]
                d = "".join(ch for ch in m if ch.isdigit())
                return (int(d) if d else 9999, m)

            cleaned.sort(key=mod_key)

            for module_name, module_rows in groupby(cleaned, key=lambda x: x[0]):
                module_rows = list(module_rows)

                elements.append(Paragraph(f"Module: {module_name or '-'}", ass_module_heading_style))

                ass_data = [[
                    Paragraph("S.No", table_header_style),
                    Paragraph("Assessment", table_header_style),
                    Paragraph("CO", table_header_style),
                    Paragraph("Bloom", table_header_style),
                    Paragraph("Max", table_header_style),
                    Paragraph("Weightage", table_header_style),
                ]]

                s = 0
                for _, ass_name, co_code, bloom, maxm, wt in module_rows:
                    s += 1
                    ass_data.append([
                        Paragraph(str(s), table_cell_center_style),
                        Paragraph(ass_name or "", table_cell_style),
                        Paragraph(co_code or "", table_cell_center_style),
                        Paragraph(bloom or "", table_cell_center_style),
                        Paragraph(maxm or "", table_cell_center_style),
                        Paragraph(wt or "", table_cell_center_style),
                    ])

                elements.append(create_table(
                    ass_data,
                    col_widths=[
                        12 * mm,
                        doc.width - (12 + 18 + 18 + 14 + 16) * mm,
                        18 * mm,
                        18 * mm,
                        14 * mm,
                        16 * mm,
                    ],
                    header_bg=SECONDARY_BLUE
                ))
                elements.append(Spacer(1, 5 * mm))

            elements.append(Spacer(1, 3 * mm))

        if any_added == 0:
            elements.append(Paragraph("No Assessment entries found for this course.", body_style))

    # -----------------------------------------
    # Build PDF
    # -----------------------------------------
    try:
        doc.build(elements)
    except Exception as e:
        # print("PDF Generation Error:", e)
        return HttpResponse("PDF generation failed.", status=500)

    buffer.seek(0)
    filename = f"CoursePlan_{course_code.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename)

from course_management.models import SubjectRequest
from django.http import JsonResponse, FileResponse



@check_permission("create_subject_request")
def create_subject_request(request):
    faculty = get_object_or_404(general_information, faculty_id=request.user.Employee_id)
    schedule_state = get_subject_allocation_window_status()

    own_department = faculty.department
    departments = Add_Department.objects.filter(is_active=True)
    regulations = Regulations.objects.all().order_by("-year")
    sections = SectionMaster.objects.all().order_by("section")

    batches = StudentDetails.objects.exclude(batch__isnull=True)\
        .values_list("batch", flat=True).distinct().order_by("batch")

    requests_list = SubjectRequest.objects.filter(faculty=faculty).order_by("-requested_on")
    academic_years = (
        requests_list.exclude(academic_year__isnull=True)
        .exclude(academic_year="")
        .values_list("academic_year", flat=True)
        .distinct()
        .order_by("-academic_year")
    )
    report_semesters_by_year = {}
    for year in academic_years:
        report_semesters_by_year[year] = list(
            requests_list.filter(academic_year=year)
            .exclude(semester__isnull=True)
            .exclude(semester="")
            .values_list("semester", flat=True)
            .distinct()
            .order_by("semester")
        )

    # Assigned faculty map
    assigned_mapping = {}
    for req in requests_list:
        assigned = AssignSubjectFaculty.objects.filter(
            course=req.course,
            regulation=req.regulation,
            batch=req.batch,
            section=req.section,
            academic_year=req.academic_year,
            is_active=True
        ).select_related("faculty").first()
        assigned_mapping[req.id] = assigned.faculty.name if assigned and assigned.faculty else None
        req.assigned_faculty = assigned.faculty if assigned and assigned.faculty else None

    # EDIT MODE
    edit_obj = None
    if request.GET.get("edit_id"):
        edit_obj = SubjectRequest.objects.filter(id=request.GET.get("edit_id")).first()

    if request.method == "POST":
        if not schedule_state["can_act"]:
            messages.error(request, schedule_state["status_message"])
            return redirect("create_subject_request")

        if request.POST.get("form_type") == "edit":
            edit_obj = get_object_or_404(
                SubjectRequest,
                id=request.POST.get("id"),
                faculty=faculty,
                status=SubjectRequest.Status.PENDING
            )
            edit_obj.requested_to_department_id = request.POST.get("requested_to_department")
            edit_obj.batch = request.POST.get("batch")
            edit_obj.section = request.POST.get("section")
            edit_obj.academic_year = request.POST.get("academic_year")
            edit_obj.reason = request.POST.get("reason")
            edit_obj.save()
            messages.success(request, "Request updated successfully.")
            return redirect("create_subject_request")

        SubjectRequest.objects.create(
            faculty=faculty,
            faculty_department=own_department,
            requested_department=own_department,
            requested_to_department_id=request.POST.get("requested_to_department"),
            course_id=request.POST.get("course"),
            regulation_id=request.POST.get("regulation"),
            semester=request.POST.get("semester"),
            batch=request.POST.get("batch"),
            section=request.POST.get("section"),
            academic_year=request.POST.get("academic_year"),
            reason=request.POST.get("reason"),
        )
        messages.success(request, "Request submitted successfully.")
        return redirect("create_subject_request")

    # DELETE REQUEST
    if request.GET.get("delete_id"):
        if not schedule_state["can_act"]:
            messages.error(request, schedule_state["status_message"])
            return redirect("create_subject_request")

        obj = get_object_or_404(
            SubjectRequest,
            id=request.GET.get("delete_id"),
            faculty=faculty
        )
        if obj.status == SubjectRequest.Status.PENDING:
            obj.delete()
            messages.success(request, "Request deleted successfully.")
        else:
            messages.error(request, "Only pending requests can be deleted.")
        return redirect("create_subject_request")

    return render(request, "course_management/faculty/create_request.html", {
        "departments": departments,
        "regulations": regulations,
        "sections": sections,
        "batches": batches,
        "requests": requests_list,
        "own_department": own_department,
        "edit_obj": edit_obj,
        "assigned_mapping": assigned_mapping,
        "schedule_state": schedule_state,
        "academic_years": academic_years,
        "report_semesters_by_year": report_semesters_by_year,
    })


# Ajax
@check_permission("create_subject_request")
def sr_get_years(request):
    faculty = get_object_or_404(general_information, faculty_id=request.user.Employee_id)
    dept = Add_Department.objects.filter(Department=faculty.department).first()
    return JsonResponse({"years": list(range(1, dept.degree.effective_duration + 1))})


@check_permission("create_subject_request")
def sr_get_courses(request):
    faculty = get_object_or_404(general_information, faculty_id=request.user.Employee_id)
    dept = Add_Department.objects.filter(Department=faculty.department).first()

    qs = Course.objects.filter(is_active=True, department=dept)
    if request.GET.get("regulation_id"): qs = qs.filter(regulation_id=request.GET.get("regulation_id"))
    if request.GET.get("year"): qs = qs.filter(year=request.GET.get("year"))
    if request.GET.get("semester"): qs = qs.filter(semester=request.GET.get("semester"))

    return JsonResponse({"courses": [
        {"id": c.id, "label": f"{c.course_code} - {c.title}", "semester": c.semester}
        for c in qs
    ]})
    # faculty = get_object_or_404(general_information, faculty_id=request.user.Employee_id)
    # dept = Add_Department.objects.filter(Department=faculty.department).first()

    # qs = Course.objects.filter(is_active=True, department=dept)

    # if request.GET.get("regulation_id"): qs = qs.filter(regulation_id=request.GET.get("regulation_id"))
    # if request.GET.get("year"): qs = qs.filter(year=request.GET.get("year"))
    # if request.GET.get("semester"): qs = qs.filter(semester=request.GET.get("semester"))

    # courses = [{"id": c.id, "label": f"{c.course_code} - {c.title}", "semester": c.semester} for c in qs]
    # return JsonResponse({"courses": courses})

import io
from django.http import FileResponse
from django.shortcuts import get_object_or_404
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.utils import ImageReader
from django.conf import settings
from django.contrib.staticfiles import finders
import os
from datetime import datetime



# ---------- BULK PDF DOWNLOAD ----------
@check_permission("create_subject_request")
def download_subject_request_pdf(request):
    faculty = get_object_or_404(general_information, faculty_id=request.user.Employee_id)
    selected_academic_year = (request.GET.get("academic_year") or "").strip()
    selected_semester = (request.GET.get("semester") or "").strip()

    if not selected_academic_year or not selected_semester:
        messages.error(request, "Please select academic year and semester before downloading the report.")
        return redirect("create_subject_request")

    requests = SubjectRequest.objects.filter(
        faculty=faculty,
        academic_year=selected_academic_year,
        semester=selected_semester
    ).order_by("requested_to_department__Department", "requested_on")

    # Professional Header Function
    def _header_on_page(canvas, doc):
        canvas.saveState()
        page_w, page_h = A4
        left = 18 * mm
        right = 18 * mm

        # Professional Logo Styling
        try:
            logo_rel = "images/ritlogo.png"
            logo_path = finders.find(logo_rel)
            if logo_path and os.path.exists(logo_path):
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                target_h = 16 * mm
                target_w = target_h * (iw / float(ih))
                canvas.drawImage(img, left, page_h - (target_h + 8 * mm),
                                width=target_w, height=target_h, 
                                preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

        # Header Text with Professional Styling
        canvas.setFillColor(colors.HexColor("#2C3E50"))
        canvas.setFont('Helvetica-Bold', 16)
        canvas.drawCentredString(page_w / 2.0, page_h - 14 * mm, 'RAMCO INSTITUTE OF TECHNOLOGY')
        
        # canvas.setFont('Helvetica-Bold', 11)
        # canvas.setFillColor(colors.HexColor("#E74C3C"))
        # canvas.drawCentredString(page_w / 2.0, page_h - 21 * mm, 'An Autonomous Institution')
        
        canvas.setFont('Helvetica', 10)
        canvas.setFillColor(colors.HexColor("#2C3E50"))
        canvas.drawCentredString(page_w / 2.0, page_h - 27 * mm, 'RAJAPALAYAM - 626117 | Tamil Nadu')
        
        # Report Title
        canvas.setFont('Helvetica-Bold', 12)
        canvas.setFillColor(colors.HexColor("#3498DB"))
        canvas.drawCentredString(page_w / 2.0, page_h - 34 * mm, 'Subject Requests Report')

        # Footer with Decorative Line
        canvas.setStrokeColor(colors.HexColor("#3498DB"))
        canvas.setLineWidth(0.5)
        canvas.line(18 * mm, 15 * mm, page_w - 18 * mm, 15 * mm)
        
        canvas.setFont('Helvetica', 9)
        canvas.setFillColor(colors.HexColor("#7F8C8D"))
        canvas.drawRightString(page_w - right, 10 * mm, f"Page {canvas.getPageNumber()}")
        
        # Footer Text
        footer_text = f"Generated on: {datetime.now().strftime('%d-%m-%Y at %H:%M')}"
        canvas.drawString(left, 10 * mm, footer_text)

        canvas.restoreState()

    buffer = io.BytesIO()
    
    # Create document with proper margins
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        leftMargin=18*mm,
        rightMargin=18*mm,
        topMargin=50*mm,
        bottomMargin=25*mm,
        title="Subject Requests Report"
    )

    styles = getSampleStyleSheet()
    
    # Professional Style Definitions
    title_style = ParagraphStyle(
        'title_style',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor("#2C3E50"),
        alignment=1,
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )
    
    dept_header_style = ParagraphStyle(
        'dept_header_style',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=colors.HexColor("#3498DB"),
        spaceAfter=8,
        fontName='Helvetica-Bold'
    )
    
    header_style = ParagraphStyle(
        'header_style', 
        parent=styles['Normal'], 
        alignment=1,
        fontName='Helvetica-Bold', 
        fontSize=9, 
        textColor=colors.white,
        leading=10
    )
    
    cell_style = ParagraphStyle(
        'cell_style', 
        parent=styles['Normal'], 
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#2C3E50")
    )
    
    faculty_info_style = ParagraphStyle(
        'faculty_info_style',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor("#34495E"),
        alignment=1,
        spaceAfter=15
    )

    elements = []
    
    # Main Title
    elements.append(Paragraph("SUBJECT REQUESTS REPORT", title_style))
    
    # Decorative Line
    elements.append(Paragraph("<hr width='70%' color='#3498DB'/>", styles['Normal']))
    elements.append(Spacer(1, 8))

    # Faculty Information
    faculty_info = [
        f"<b>Faculty Name:</b> {getattr(faculty, 'name', '')}",
        f"<b>Faculty ID:</b> {getattr(faculty, 'faculty_id', '')}",
        f"<b>Designation:</b> {getattr(faculty, 'present_designation', '') or getattr(faculty, 'designation', '')}",
        f"<b>Department:</b> {getattr(faculty.department, 'Department', '') if faculty.department else ''}",
        f"<b>Academic Year:</b> {selected_academic_year}",
        f"<b>Semester:</b> {selected_semester}",
    ]
    
    faculty_info_text = " &nbsp;&nbsp; • &nbsp;&nbsp; ".join(faculty_info)
    elements.append(Paragraph(faculty_info_text, faculty_info_style))
    elements.append(Spacer(1, 15))

    # Summary Statistics
    total_requests = len(requests)
    elements.append(Paragraph(
        f"<b>📊 SUMMARY:</b> Total {total_requests} subject request(s) submitted",
        ParagraphStyle(
            'summary_style',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor("#27AE60"),
            alignment=1,
            spaceAfter=15
        )
    ))

    # Group requests by department and add page breaks
    current_department = None
    department_requests = []
    first_department = True

    for req in requests:
        if current_department != req.requested_to_department:
            if current_department is not None:
                # Add page break before new department (except for first one)
                if not first_department:
                    elements.append(PageBreak())
                else:
                    first_department = False
                
                # Add table for previous department
                _add_department_table(elements, current_department, department_requests, styles, header_style, cell_style)
                
            current_department = req.requested_to_department
            department_requests = []
        department_requests.append(req)
    
    # Add the last department
    if current_department is not None and department_requests:
        if not first_department:
            elements.append(PageBreak())
        _add_department_table(elements, current_department, department_requests, styles, header_style, cell_style)

    # Empty state message
    if not requests:
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(
            "No subject requests found.",
            ParagraphStyle(
                'empty_style',
                parent=styles['Heading3'],
                fontSize=12,
                textColor=colors.HexColor("#95A5A6"),
                alignment=1
            )
        ))

    # Build PDF
    try:
        doc.build(elements, onFirstPage=_header_on_page, onLaterPages=_header_on_page)
    except Exception as e:
        # Fallback simple PDF
        simple_buffer = io.BytesIO()
        simple_doc = SimpleDocTemplate(simple_buffer, pagesize=A4)
        simple_elements = [Paragraph("Error generating PDF", styles['Heading1'])]
        simple_doc.build(simple_elements)
        simple_buffer.seek(0)
        return FileResponse(simple_buffer, as_attachment=True, filename="Subject_Requests_All.pdf")

    buffer.seek(0)
    filename = f"Subject_Requests_{faculty.faculty_id}_{selected_academic_year}_Sem{selected_semester}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename)
 

def _add_department_table(elements, department, requests, styles, header_style, cell_style):
    """Helper function to add a department's requests as a table"""
    # Department Header
    elements.append(Paragraph(
        f"📋 REQUESTS TO: {department.Department.upper()}",
        ParagraphStyle(
            'dept_header',
            parent=styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor("#2C3E50"),
            backColor=colors.HexColor("#F8F9FA"),
            spaceAfter=12,
            fontName='Helvetica-Bold',
            alignment=0
        )
    ))

    # Department summary
    dept_total = len(requests)
    status_count = {}
    for req in requests:
        status = req.status or "Pending"
        status_count[status] = status_count.get(status, 0) + 1
    
    status_summary = []
    for status, count in status_count.items():
        status_color = {
            "approved": "#27AE60",
            "rejected": "#E74C3C",
            "pending": "#F39C12"
        }.get(status.lower(), "#7F8C8D")
        status_summary.append(f"<font color='{status_color}'><b>{status.title()}:</b> {count}</font>")
    
    elements.append(Paragraph(
        f"<b>Total Requests:</b> {dept_total} &nbsp;&nbsp; | &nbsp;&nbsp; " + " &nbsp;&nbsp; | &nbsp;&nbsp; ".join(status_summary),
        ParagraphStyle(
            'dept_summary',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor("#34495E"),
            spaceAfter=15,
            alignment=0
        )
    ))

    # Table data
    data = [[
        Paragraph("Course", header_style),
        Paragraph("Semester", header_style),
        Paragraph("Batch", header_style),
        Paragraph("Section", header_style),
        Paragraph("Academic Year", header_style),
        Paragraph("Status", header_style),
        Paragraph("Reason", header_style),
    ]]

    for req in requests:
        # Status with color coding
        status_text = req.status or "Pending"
        status_color = {
            "approved": "#27AE60",
            "rejected": "#E74C3C",
            "pending": "#F39C12"
        }.get(status_text.lower(), "#7F8C8D")
        
        status_cell = Paragraph(
            f"<font color='{status_color}'><b>{status_text.upper()}</b></font>",
            cell_style
        )

        data.append([
            Paragraph(f"<b>{req.course}</b>" if req.course else "-", cell_style),
            Paragraph(str(req.semester) if req.semester else "-", cell_style),
            Paragraph(str(req.batch) if req.batch else "-", cell_style),
            Paragraph(str(req.section) if req.section else "-", cell_style),
            Paragraph(str(req.academic_year) if req.academic_year else "-", cell_style),
            status_cell,
            Paragraph(str(req.reason) if req.reason else "Not specified", cell_style),
        ])

    # Table column widths
    col_widths = [35*mm, 20*mm, 20*mm, 20*mm, 25*mm, 25*mm, 55*mm]

    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        
        # Grid and borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 1), (4, -1), 'CENTER'),  # Center align numeric/text columns
        
        # Font styling for content
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),  # Bold course names
        
        # Padding
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        
        # Alternate row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 10))

@check_permission("approve_subject_requests")
@check_permission("approve_subject_requests")
def approve_subject_requests(request):
    faculty = get_object_or_404(general_information, faculty_id=request.user.Employee_id)
    schedule_state = get_subject_allocation_window_status()
    hod_department = faculty.department

    # Get filter parameters
    filter_academic_year = request.GET.get("academic_year", "")
    filter_regulation = request.GET.get("regulation", "")
    filter_semester = request.GET.get("semester", "")
    filter_batch = request.GET.get("batch", "")
    filter_section = request.GET.get("section", "")

    # Base queryset
    requests_list = SubjectRequest.objects.filter(
        requested_to_department=hod_department
    ).select_related("faculty", "course", "regulation", "requested_department")
    print("Request list => ", requests_list)

    # Apply filters
    if filter_academic_year:
        requests_list = requests_list.filter(academic_year=filter_academic_year)
    if filter_regulation:
        requests_list = requests_list.filter(regulation_id=filter_regulation)
    if filter_semester:
        requests_list = requests_list.filter(semester=filter_semester)
    if filter_batch:
        requests_list = requests_list.filter(batch=filter_batch)
    if filter_section:
        requests_list = requests_list.filter(section=filter_section)

    requests_list = requests_list.order_by("-requested_on")

    # Get distinct values for filters
    all_requests = SubjectRequest.objects.filter(requested_to_department=hod_department)
    
    academic_years = (
        all_requests.exclude(academic_year__isnull=True)
        .exclude(academic_year="")
        .values_list("academic_year", flat=True)
        .distinct()
        .order_by("-academic_year")
    )
    
    regulations = (
        Regulations.objects.filter(
            id__in=all_requests.exclude(regulation__isnull=True)
            .values_list("regulation_id", flat=True)
            .distinct()
        ).order_by("-year")
    )
    
    semesters = (
        all_requests.exclude(semester__isnull=True)
        .exclude(semester="")
        .values_list("semester", flat=True)
        .distinct()
        .order_by("semester")
    )
    
    batches = (
        all_requests.exclude(batch__isnull=True)
        .exclude(batch="")
        .values_list("batch", flat=True)
        .distinct()
        .order_by("-batch")
    )
    
    sections = (
        all_requests.exclude(section__isnull=True)
        .exclude(section="")
        .values_list("section", flat=True)
        .distinct()
        .order_by("section")
    )

    faculty_list = general_information.objects.filter(
        department=hod_department
    ).order_by("name")

    # ✅ Attach assigned faculty name directly
    for req in requests_list:
        assigned = AssignSubjectFaculty.objects.filter(
            course=req.course,
            batch=req.batch,
            section=req.section,
            academic_year=req.academic_year,
            is_active=True
        ).select_related("faculty").first()
        req.assigned_faculty = assigned.faculty.name if assigned and assigned.faculty else None
        req.assigned_faculty_id = assigned.faculty.faculty_id if assigned and assigned.faculty else None

    if request.method == "POST":
        if not schedule_state["can_act"]:
            messages.error(request, schedule_state["status_message"])
            return redirect("approve_subject_requests")

        req_obj = get_object_or_404(
            SubjectRequest,
            id=request.POST.get("req_id"),
            requested_to_department=hod_department
        )
        action = request.POST.get("action")

        if action in ["approve", "update_faculty"]:
            f_id = request.POST.get("selected_faculty")
            chosen = get_object_or_404(
                general_information,
                faculty_id=f_id,
                department=hod_department
            )

            assignment = AssignSubjectFaculty.objects.filter(
                course=req_obj.course,
                batch=req_obj.batch,
                section=req_obj.section,
                academic_year=req_obj.academic_year,
                is_active=True
            ).first()
            if assignment:
                assignment.department = req_obj.requested_department
                assignment.faculty = chosen
                assignment.regulation = req_obj.regulation
                assignment.reason = req_obj.reason
                assignment.save(update_fields=[
                    "department",
                    "faculty",
                    "regulation",
                    "reason",
                ])
            else:
                AssignSubjectFaculty.objects.create(
                    department=req_obj.requested_department,
                    faculty=chosen,
                    course=req_obj.course,
                    regulation=req_obj.regulation,
                    batch=req_obj.batch,
                    section=req_obj.section,
                    academic_year=req_obj.academic_year,
                    reason=req_obj.reason,
                    is_active=True
                )
            req_obj.status = SubjectRequest.Status.APPROVED
            messages.success(
                request,
                "Subject faculty assignment updated successfully."
            )

        elif action == "reject":
            req_obj.status = SubjectRequest.Status.REJECTED
            messages.success(request, "Subject request rejected successfully.")

        req_obj.save()
        return redirect("approve_subject_requests")

    return render(request, "course_management/hod/approve_requests.html", {
        "requests": requests_list,
        "faculty_list": faculty_list,
        "department": hod_department,
        "academic_years": academic_years,
        "regulations": regulations,
        "semesters": semesters,
        "batches": batches,
        "sections": sections,
        "schedule_state": schedule_state,
        # Pass current filter values back to template
        "filter_academic_year": filter_academic_year,
        "filter_regulation": filter_regulation,
        "filter_semester": filter_semester,
        "filter_batch": filter_batch,
        "filter_section": filter_section,
    })


@check_permission("approve_subject_requests")
def download_hod_subject_request_pdf(request):
    faculty = get_object_or_404(general_information, faculty_id=request.user.Employee_id)
    hod_department = faculty.department
    
    # Get filter parameters from query string (same as page filters)
    selected_academic_year = (request.GET.get("academic_year") or "").strip()
    selected_regulation = (request.GET.get("regulation") or "").strip()
    selected_semester = (request.GET.get("semester") or "").strip()
    selected_batch = (request.GET.get("batch") or "").strip()
    selected_section = (request.GET.get("section") or "").strip()

    # Build filter queryset
    requests = SubjectRequest.objects.filter(
        requested_to_department=hod_department
    ).select_related("faculty", "course", "regulation", "requested_department")

    # Apply filters (same logic as main view)
    if selected_academic_year:
        requests = requests.filter(academic_year=selected_academic_year)
    if selected_regulation:
        requests = requests.filter(regulation_id=selected_regulation)
    if selected_semester:
        requests = requests.filter(semester=selected_semester)
    if selected_batch:
        requests = requests.filter(batch=selected_batch)
    if selected_section:
        requests = requests.filter(section=selected_section)

    requests = requests.order_by("requested_on")

    if not requests.exists():
        messages.warning(request, "No requests found matching the current filters.")
        return redirect("approve_subject_requests")
    page_size = landscape(A4)

    # Professional Header Function
    def _header_on_page(canvas, doc):
        canvas.saveState()
        page_w, page_h = page_size
        left = doc.leftMargin
        right = doc.rightMargin

        # Professional Logo Styling
        try:
            logo_rel = "images/ritlogo.png"
            logo_path = finders.find(logo_rel)
            if logo_path and os.path.exists(logo_path):
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                target_h = 14 * mm
                target_w = target_h * (iw / float(ih))
                canvas.drawImage(img, left, page_h - (target_h + 8 * mm),
                                width=target_w, height=target_h, 
                                preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

        # Header Text with Professional Styling
        canvas.setFillColor(colors.HexColor("#17365D"))
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawCentredString(page_w / 2.0, page_h - 11 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#34495E"))
        canvas.drawCentredString(page_w / 2.0, page_h - 17 * mm, "RAJAPALAYAM - 626117 | Tamil Nadu")

        canvas.setFont("Helvetica-Bold", 8)
        canvas.setFillColor(colors.HexColor("#1D75BD"))
        canvas.drawCentredString(page_w / 2.0, page_h - 22 * mm, "HOD - Subject Requests Report")

        canvas.setStrokeColor(colors.HexColor("#D8E6F3"))
        canvas.setLineWidth(0.7)
        canvas.line(left, page_h - 27 * mm, page_w - right, page_h - 27 * mm)

        # Footer with Decorative Line
        canvas.setStrokeColor(colors.HexColor("#3498DB"))
        canvas.setLineWidth(0.5)
        canvas.line(left, 12 * mm, page_w - right, 12 * mm)

        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#7F8C8D"))
        canvas.drawRightString(page_w - right, 10 * mm, f"Page {canvas.getPageNumber()}")

        # Footer Text
        footer_text = f"Generated on: {datetime.now().strftime('%d-%m-%Y at %H:%M')}"
        canvas.drawString(left, 10 * mm, footer_text)

        canvas.restoreState()

    buffer = io.BytesIO()
    
    # Create document with proper margins
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=page_size,
        leftMargin=14*mm,
        rightMargin=14*mm,
        topMargin=32*mm,
        bottomMargin=18*mm,
        title="HOD Subject Requests Report"
    )

    styles = getSampleStyleSheet()
    
    # Professional Style Definitions
    title_style = ParagraphStyle(
        'title_style',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor("#17365D"),
        alignment=1,
        spaceAfter=8,
        fontName='Helvetica-Bold'
    )

    meta_label_style = ParagraphStyle(
        'meta_label_style',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
        textColor=colors.HexColor("#64748B"),
        fontName='Helvetica-Bold',
        alignment=1
    )

    meta_value_style = ParagraphStyle(
        'meta_value_style',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#0F172A"),
        fontName='Helvetica-Bold',
        alignment=1
    )

    dept_style = ParagraphStyle(
        'dept_style',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#17365D"),
        alignment=1,
        spaceAfter=5,
        fontName='Helvetica-Bold'
    )
    
    header_style = ParagraphStyle(
        'header_style', 
        parent=styles['Normal'], 
        alignment=1,
        fontName='Helvetica-Bold', 
        fontSize=7,
        textColor=colors.white,
        leading=8
    )
    
    cell_style = ParagraphStyle(
        'cell_style', 
        parent=styles['Normal'], 
        fontSize=7,
        leading=8.5,
        textColor=colors.HexColor("#2C3E50")
    )

    center_cell_style = ParagraphStyle(
        'center_cell_style',
        parent=styles['Normal'],
        fontSize=7,
        leading=8.5,
        textColor=colors.HexColor("#2C3E50"),
        alignment=1
    )

    summary_style = ParagraphStyle(
        'summary_style',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#17365D"),
        alignment=1,
        spaceAfter=8
    )

    elements = []
    
    # Main Title
    elements.append(Paragraph("HOD SUBJECT REQUESTS REPORT", title_style))
    
    # Decorative Line
    elements.append(Paragraph("<hr width='70%' color='#3498DB'/>", styles['Normal']))
    elements.append(Spacer(1, 8))

    # Department Information
    elements.append(Paragraph(
        f"<b>Department:</b> {hod_department.Department}",
        dept_style
    ))
    
    # Build filter info dynamically
    filter_parts = []
    if selected_academic_year:
        filter_parts.append(f"<b>Academic Year:</b> {selected_academic_year}")
    if selected_regulation:
        reg_obj = Regulations.objects.filter(id=selected_regulation).first()
        reg_text = f"R-{reg_obj.year}" if reg_obj else selected_regulation
        filter_parts.append(f"<b>Regulation:</b> {reg_text}")
    if selected_semester:
        filter_parts.append(f"<b>Semester:</b> {selected_semester}")
    if selected_batch:
        filter_parts.append(f"<b>Batch:</b> {selected_batch}")
    if selected_section:
        filter_parts.append(f"<b>Section:</b> {selected_section}")
    
    if filter_parts:
        elements.append(Paragraph(" &nbsp;&nbsp; • &nbsp;&nbsp; ".join(filter_parts), dept_style))
    else:
        elements.append(Paragraph("<b>Filters:</b> All Requests", dept_style))
    
    # HOD Information
    hod_info = [
        f"<b>HOD Name:</b> {getattr(faculty, 'name', '')}",
        f"<b>HOD ID:</b> {getattr(faculty, 'faculty_id', '')}",
        f"<b>Designation:</b> {getattr(faculty, 'present_designation', '') or getattr(faculty, 'designation', '')}"
    ]
    
    hod_info_text = " &nbsp;&nbsp; • &nbsp;&nbsp; ".join(hod_info)
    elements.append(Paragraph(hod_info_text, dept_style))
    elements.append(Spacer(1, 15))

    # Summary Statistics
    total_requests = len(requests)
    
    # Calculate status counts
    status_count = {
        'PENDING': 0,
        'APPROVED': 0,
        'REJECTED': 0
    }
    for req in requests:
        status = req.status or 'PENDING'
        status_count[status] = status_count.get(status, 0) + 1

    summary_text = (
        f"<b>📊 SUMMARY:</b> Total {total_requests} request(s) &nbsp;&nbsp; | &nbsp;&nbsp; "
        f"<font color='#F39C12'><b>Pending:</b> {status_count['PENDING']}</font> &nbsp;&nbsp; | &nbsp;&nbsp; "
        f"<font color='#27AE60'><b>Approved:</b> {status_count['APPROVED']}</font> &nbsp;&nbsp; | &nbsp;&nbsp; "
        f"<font color='#E74C3C'><b>Rejected:</b> {status_count['REJECTED']}</font>"
    )
    
    elements.append(Paragraph(summary_text, summary_style))
    elements.append(Spacer(1, 15))

    # Table Data with Professional Headers
    data = [[
        Paragraph("Requesting<br/>Faculty", header_style),
        Paragraph("Course", header_style),
        Paragraph("Batch", header_style),
        Paragraph("Section", header_style),
        Paragraph("Semester", header_style),
        Paragraph("Academic<br/>Year", header_style),
        Paragraph("Currently<br/>Assigned", header_style),
        Paragraph("Status", header_style),
        Paragraph("Reason", header_style),
    ]]

    for req in requests:
        # Get assigned faculty
        assigned = AssignSubjectFaculty.objects.filter(
            course=req.course,
            batch=req.batch,
            section=req.section,
            academic_year=req.academic_year,
            is_active=True
        ).first()

        assigned_name = assigned.faculty.name if assigned and assigned.faculty else "Not Assigned"

        # Status with color coding
        status_text = req.status or "PENDING"
        status_color = {
            "APPROVED": "#27AE60",
            "REJECTED": "#E74C3C",
            "PENDING": "#F39C12"
        }.get(status_text, "#7F8C8D")
        
        status_cell = Paragraph(
            f"<font color='{status_color}'><b>{status_text}</b></font>",
            cell_style
        )

        data.append([
            Paragraph(f"<b>{req.faculty.name}</b><br/>{req.faculty.faculty_id}" if req.faculty else "-", cell_style),
            Paragraph(f"<b>{getattr(req.course, 'course_code', '') or '-'}</b><br/>{getattr(req.course, 'title', '') or req.course}" if req.course else "-", cell_style),
            Paragraph(str(req.batch) if req.batch else "-", center_cell_style),
            Paragraph(str(req.section) if req.section else "-", center_cell_style),
            Paragraph(str(req.semester) if req.semester else "-", center_cell_style),
            Paragraph(str(req.academic_year) if req.academic_year else "-", center_cell_style),
            Paragraph(assigned_name, cell_style),
            status_cell,
            Paragraph(str(req.reason) if req.reason else "Not specified", cell_style),
        ])

    # Table column widths tuned for landscape A4.
    col_widths = [28*mm, 48*mm, 18*mm, 18*mm, 20*mm, 24*mm, 34*mm, 23*mm, doc.width - (28 + 48 + 18 + 18 + 20 + 24 + 34 + 23) * mm]

    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1D75BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        
        # Grid and borders
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#C8D7E5')),
        ('LINEBELOW', (0, 0), (-1, 0), 0.8, colors.HexColor('#0F5E9C')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (2, 1), (5, -1), 'CENTER'),
        
        # Font styling for content
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),  # Bold faculty names
        ('FONTNAME', (1, 1), (1, -1), 'Helvetica-Bold'),  # Bold course names
        
        # Padding
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        
        # Alternate row colors
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFC')]),
    ]))

    elements.append(table)
    
    # Empty state message
    if not requests:
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(
            "No subject requests found for your department.",
            ParagraphStyle(
                'empty_style',
                parent=styles['Heading3'],
                fontSize=12,
                textColor=colors.HexColor("#95A5A6"),
                alignment=1
            )
        ))

    # Build PDF with professional header
    try:
        doc.build(elements, onFirstPage=_header_on_page, onLaterPages=_header_on_page)
    except Exception as e:
        # Fallback simple PDF
        simple_buffer = io.BytesIO()
        simple_doc = SimpleDocTemplate(simple_buffer, pagesize=A4)
        simple_elements = [Paragraph("Error generating PDF", styles['Heading1'])]
        simple_doc.build(simple_elements)
        simple_buffer.seek(0)
        return FileResponse(simple_buffer, as_attachment=True, filename="HOD_Subject_Requests.pdf")

    buffer.seek(0)
    filename = f"HOD_Subject_Requests_{hod_department.Department}_{selected_academic_year}_Sem{selected_semester}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename)


    faculty = get_object_or_404(general_information, faculty_id=request.user.Employee_id)
    hod_department = Add_Department.objects.filter(Department=faculty.department).first()

    # Fetch all requests (Pending + Approved + Rejected)
    requests_list = SubjectRequest.objects.filter(
        requested_to_department=hod_department
    ).order_by("-requested_on")

    # Faculty choices (same department only)
    faculty_list = general_information.objects.filter(department=hod_department)

    # Map Assigned Faculty Names
    assigned_mapping = {}
    for sr in requests_list:
        assigned = AssignSubjectFaculty.objects.filter(
            course=sr.course,
            batch=sr.batch,
            section=sr.section,
            academic_year=sr.academic_year,
            is_active=True
        ).first()
        assigned_mapping[sr.id] = assigned.faculty.name if assigned and assigned.faculty else None

    # Approve / Reject Actions
    if request.method == "POST":
        req_id = request.POST.get("req_id")
        action = request.POST.get("action")
        obj = get_object_or_404(SubjectRequest, id=req_id)

        if action == "approve":
            selected_faculty_id = request.POST.get("selected_faculty")
            selected_faculty = get_object_or_404(general_information, faculty_id=selected_faculty_id)

            AssignSubjectFaculty.objects.create(
                department=obj.requested_department,
                faculty=selected_faculty,
                course=obj.course,
                regulation=obj.regulation,
                batch=obj.batch,
                section=obj.section,
                academic_year=obj.academic_year,
                reason=obj.reason,
                is_active=True
            )
            obj.status = SubjectRequest.Status.APPROVED

        elif action == "reject":
            obj.status = SubjectRequest.Status.REJECTED

        obj.save()
        messages.success(request, f"Request has been {obj.status.lower()}.")
        return redirect("approve_subject_requests")

    return render(request, "course_management/hod/approve_requests.html", {
        "requests": requests_list,
        "faculty_list": faculty_list,
        "department": hod_department,
        "assigned_mapping": assigned_mapping,
    })
    faculty = get_object_or_404(general_information, faculty_id=request.user.Employee_id)
    hod_department = Add_Department.objects.filter(Department=faculty.department).first()

    # Fetch ALL requests (pending, approved, rejected)
    requests_list = SubjectRequest.objects.filter(
        requested_to_department=hod_department
    ).order_by("-requested_on")

    faculty_list = general_information.objects.filter(department=hod_department)

    # Build dictionary of assigned faculty for approved requests
    assigned_mapping = {}
    for sr in requests_list:
        assigned = AssignSubjectFaculty.objects.filter(
            course=sr.course,
            batch=sr.batch,
            section=sr.section,
            academic_year=sr.academic_year,
            is_active=True
        ).first()
        assigned_mapping[sr.id] = assigned.faculty.name if assigned else None

    if request.method == "POST":
        req_id = request.POST.get("req_id")
        action = request.POST.get("action")
        obj = get_object_or_404(SubjectRequest, id=req_id)

        if action == "approve":
            selected_faculty_id = request.POST.get("selected_faculty")
            selected_faculty = get_object_or_404(general_information, faculty_id=selected_faculty_id)

            AssignSubjectFaculty.objects.create(
                department=obj.requested_department,
                faculty=selected_faculty,
                course=obj.course,
                regulation=obj.regulation,
                batch=obj.batch,
                section=obj.section,
                academic_year=obj.academic_year,
                reason=obj.reason,
                is_active=True
            )
            obj.status = SubjectRequest.Status.APPROVED

        elif action == "reject":
            obj.status = SubjectRequest.Status.REJECTED

        obj.save()
        messages.success(request, f"Request has been {obj.status.lower()}.")
        return redirect("approve_subject_requests")

    return render(request, "course_management/hod/approve_requests.html", {
        "requests": requests_list,
        "department": hod_department,
        "faculty_list": faculty_list,
        "assigned_mapping": assigned_mapping,
    })
    # HOD/Approver user → Get their department
    faculty = get_object_or_404(general_information, faculty_id=request.user.Employee_id)
    hod_department = Add_Department.objects.filter(Department=faculty.department).first()

    # Fetch ALL requests sent to this department (not only pending)
    requests_list = SubjectRequest.objects.filter(
        requested_to_department=hod_department
    ).order_by("-requested_on")

    # Faculty list for dropdown (only same department as HOD)
    faculty_list = general_information.objects.filter(department=hod_department)

    if request.method == "POST":
        req_id = request.POST.get("req_id")
        action = request.POST.get("action")
        obj = get_object_or_404(SubjectRequest, id=req_id)

        if action == "approve":
            selected_faculty_id = request.POST.get("selected_faculty")
            selected_faculty = get_object_or_404(general_information, faculty_id=selected_faculty_id)

            # Create AssignSubjectFaculty entry
            assign = AssignSubjectFaculty.objects.create(
                department=obj.requested_department,
                faculty=selected_faculty,
                course=obj.course,
                regulation=obj.regulation,
                batch=obj.batch,
                section=obj.section,
                academic_year=obj.academic_year,
                reason=obj.reason,
                is_active=True
            )

            obj.status = SubjectRequest.Status.APPROVED
            obj.assigned_faculty = selected_faculty.name  # Store assigned faculty name for display
        
        elif action == "reject":
            obj.status = SubjectRequest.Status.REJECTED
        
        obj.save()
        messages.success(request, f"Request has been {obj.status.lower()}.")
        return redirect("approve_subject_requests")

    return render(request, "course_management/hod/approve_requests.html", {
        "requests": requests_list,
        "department": hod_department,
        "faculty_list": faculty_list,
    })





from django.shortcuts import get_object_or_404, render


def Student_Internal_Exam_timetable(request):
    reg_no = request.user.Employee_id 
    student = get_object_or_404(StudentDetails, reg_no=reg_no)

    # Fetching required details
    semester = request.GET.get("semester") or student.semester
    batch = student.batch
    department = student.department
    regulation = student.regulation
    degree = department.degree if department else None 

  
    department_id = department.id if department else None
    degree_id = degree.id if degree else None 

    regulation_obj = None
    if regulation:
        regulation_obj = Regulations.objects.filter(year=regulation).first()

    timetable_qs = InternalTimeTable.objects.select_related(
        "degree", "department", "regulation", "course", "internal_assessment"
    ).filter(
        department=department,
        degree=degree,
        regulation=regulation_obj,
        batch=batch,
        semester=semester
    ).order_by("internal_assessment_id"
               )



    # Printing the details for debugging
    # print(f"Student: {student.name}, Semester: {semester}, Batch: {batch}, Department ID: {department_id}, Degree: {degree_id}, regulation: {regulation}")

    return render(request, "course_management/faculty/Student_Internal_Exam_timetable.html", {
        "student": student,
        "semester": semester,
        "batch": batch,
        "degree": degree, 
        "department_id": department_id,  
        "degree_id": degree_id ,
        "timetable": timetable_qs,
    })
from course_management.models import Program_outcomes



def Student_Semester_Exam_timetable(request):

    reg_no = request.user.Employee_id
    student = get_object_or_404(StudentDetails, reg_no=reg_no)

    department = student.department
    degree = department.degree if department else None

    # Total semesters
    total_semesters = (degree.duration or 0) * 2 if degree else 0
    semester_list = list(range(1, total_semesters + 1))

    selected_sem = request.GET.get("semester")

    if selected_sem:
        current_sem = int(selected_sem)
    else:
        current_sem = int(student.semester or 0)

    # --------------------------------------------------
    # ✅ 1. Fetch regular schedules for selected semester
    # --------------------------------------------------
    regular_schedules = SemesterExamScheduletimetable.objects.filter(
        department=department,
        semester=str(current_sem),
        is_failed=False
    ).select_related("course")

    # --------------------------------------------------
    # ✅ 2. Get publish session / examsession
    # --------------------------------------------------
    exam_sessions = regular_schedules.values_list(
        "examsession", flat=True
    ).distinct()

    # --------------------------------------------------
    # ✅ 3. Find student's failed courses
    # --------------------------------------------------
    failed_courses = Result.objects.filter(
        student=student,
        grade__iexact="U"
    ).values_list("course_id", flat=True)

    # --------------------------------------------------
    # ✅ 4. Fetch failed timetable schedules
    # (Same examsession + student's failed courses)
    # --------------------------------------------------
    failed_schedules = SemesterExamScheduletimetable.objects.filter(
        department=department,
        is_failed=True,
        examsession__in=exam_sessions,
        course_id__in=failed_courses
    ).select_related("course")

    # --------------------------------------------------
    # ✅ 5. Combine both
    # --------------------------------------------------
    schedules = (regular_schedules | failed_schedules).order_by(
        "exam_date", "session"
    )

    context = {
        "student": student,
        "semester_list": semester_list,
        "current_sem": current_sem,
        "schedules": schedules
    }

    return render(
        request,
        "course_management/faculty/semesterexamschedulestudent.html",
        context
    )


from collections import OrderedDict
from decimal import Decimal, InvalidOperation

from django.shortcuts import render
from django.db.models import Prefetch



def _to_float(value, default=0):
    try:
        if value in [None, ""]:
            return float(default)
        return float(Decimal(str(value)))
    except (InvalidOperation, ValueError, TypeError):
        return float(default)


def _clean_num(value):
    value = float(value or 0)
    if value.is_integer():
        return int(value)
    return round(value, 2)


def _percent(part, whole):
    part = _to_float(part)
    whole = _to_float(whole)
    if whole <= 0:
        return 0
    return round((part / whole) * 100, 2)


def _roman_semester(value):
    mapping = {
        "1": "I", "2": "II", "3": "III", "4": "IV",
        "5": "V", "6": "VI", "7": "VII", "8": "VIII",
        "I": "I", "II": "II", "III": "III", "IV": "IV",
        "V": "V", "VI": "VI", "VII": "VII", "VIII": "VIII",
    }
    return mapping.get(str(value).strip().upper(), str(value).strip())


def _get_regulation_short_name(reg):
    year = str(reg.year or "")
    short = year[-2:] if len(year) >= 2 else year
    return f"RIT (R-{short})"


def _get_course_category_label(course):
    if course.elective:
        if course.elective.category_code:
            return str(course.elective.category_code).strip().upper()
        if course.elective.Course_category_name:
            return str(course.elective.Course_category_name).strip().upper()
    return "OTH"


def _build_regulation_course_queryset(regulation, filters=None):
    coursehours_qs = CourseHours.objects.select_related("hour_config").all()
    qs = (
        Course.objects
        .filter(regulation=regulation, is_active=True)
        .select_related("department", "regulation", "elective")
        .prefetch_related(Prefetch("semesters", queryset=coursehours_qs))
        .order_by("department__Department", "year", "semester", "course_code")
    )
    filters = filters or {}

    if filters.get("department_id"):
        qs = qs.filter(department_id=filters["department_id"])
    if filters.get("year"):
        qs = qs.filter(year=str(filters["year"]))
    if filters.get("semester"):
        qs = qs.filter(semester=str(filters["semester"]))
    if filters.get("category_id"):
        qs = qs.filter(elective_id=filters["category_id"])
    if filters.get("is_active") in {"active", "inactive"}:
        qs = qs.filter(is_active=(filters["is_active"] == "active"))
    if filters.get("q"):
        q = filters["q"].strip()
        qs = qs.filter(Q(course_code__icontains=q) | Q(title__icontains=q))

    return qs


def _build_regulation_analytics(regulation, filters=None):
    courses = _build_regulation_course_queryset(regulation, filters=filters)

    total_courses = courses.count()
    total_departments = courses.values("department").distinct().count()

    total_credits_all = 0
    total_hours_all = 0

    all_categories_set = set()
    all_semesters_set = set()

    for course in courses:
        all_categories_set.add(_get_course_category_label(course))
        sem_roman = _roman_semester(course.semester)
        if sem_roman:
            all_semesters_set.add(sem_roman)

    roman_order = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII"]
    semester_labels = [r for r in roman_order if r in all_semesters_set]
    extra_semesters = sorted([s for s in all_semesters_set if s not in roman_order])
    semester_labels += extra_semesters

    dynamic_categories = sorted(all_categories_set)

    regulation_category_totals = OrderedDict((cat, 0) for cat in dynamic_categories)

    semester_category_matrix = OrderedDict()
    for cat in dynamic_categories:
        semester_category_matrix[cat] = OrderedDict((sem, 0) for sem in semester_labels)

    semester_stats = OrderedDict()
    for sem in semester_labels:
        semester_stats[f"Sem {sem}"] = {"credits": 0, "count": 0}

    semester_detailed_rows = OrderedDict((sem, []) for sem in semester_labels)

    department_stats = OrderedDict()
    category_stats = OrderedDict((cat, 0) for cat in dynamic_categories)

    theory_total = 0
    laboratory_total = 0
    tutorial_total = 0
    integrated_total = 0

    table_rows = []

    for course in courses:
        dept_id = course.department.id if course.department else 0
        dept_name = course.department.Department if course.department else "No Department"
        category_name = _get_course_category_label(course)
        sem_roman = _roman_semester(course.semester)
        semester_key = f"Sem {sem_roman}" if sem_roman else "Sem Unknown"

        if dept_id not in department_stats:
            department_stats[dept_id] = {
                "name": dept_name,
                "total_credits": 0,
                "theory_credits": 0,
                "laboratory_credits": 0,
                "tutorial_credits": 0,
                "integrated_credits": 0,
                "course_count": 0,
                "category_credits": OrderedDict((cat, 0) for cat in dynamic_categories),
            }

        department_stats[dept_id]["course_count"] += 1

        if semester_key not in semester_stats:
            semester_stats[semester_key] = {"credits": 0, "count": 0}
        semester_stats[semester_key]["count"] += 1

        course_hours_rows = list(course.semesters.all())

        if not course_hours_rows:
            table_rows.append({
                "department": dept_name,
                "course_code": course.course_code or "-",
                "title": course.title or "-",
                "year": course.year or "-",
                "semester": course.semester or "-",
                "category": category_name,
                "lecture": 0,
                "tutorial": 0,
                "lab": 0,
                "total_hours": 0,
                "credits": 0,
                "is_active": course.is_active,
            })
            continue

        for hour in course_hours_rows:
            lecture = _to_float(hour.leture_npwk)
            tutorial = _to_float(hour.tutorial_npwk)
            lab = _to_float(hour.laboratory_npwk)
            total_hours = _to_float(hour.total_hours)
            credits = _to_float(hour.credits)

            total_hours_all += total_hours
            total_credits_all += credits

            if semester_key in semester_stats:
                semester_stats[semester_key]["credits"] += credits

            if category_name not in category_stats:
                category_stats[category_name] = 0
            category_stats[category_name] += credits

            if category_name not in regulation_category_totals:
                regulation_category_totals[category_name] = 0
            regulation_category_totals[category_name] += credits

            if category_name not in semester_category_matrix:
                semester_category_matrix[category_name] = OrderedDict((sem, 0) for sem in semester_labels)

            if sem_roman:
                if sem_roman not in semester_category_matrix[category_name]:
                    semester_category_matrix[category_name][sem_roman] = 0
                semester_category_matrix[category_name][sem_roman] += credits

            if lecture > 0 and lab > 0:
                integrated_total += credits
                department_stats[dept_id]["integrated_credits"] += credits
            elif lab > 0 and lecture == 0:
                laboratory_total += credits
                department_stats[dept_id]["laboratory_credits"] += credits
            elif lecture > 0 and lab == 0:
                theory_total += credits
                department_stats[dept_id]["theory_credits"] += credits

            if tutorial > 0:
                tutorial_total += credits
                department_stats[dept_id]["tutorial_credits"] += credits

            department_stats[dept_id]["total_credits"] += credits

            if category_name not in department_stats[dept_id]["category_credits"]:
                department_stats[dept_id]["category_credits"][category_name] = 0
            department_stats[dept_id]["category_credits"][category_name] += credits

            if sem_roman in semester_detailed_rows:
                semester_detailed_rows[sem_roman].append({
                    "sno": len(semester_detailed_rows[sem_roman]) + 1,
                    "course_code": course.course_code or "-",
                    "course_title": course.title or "-",
                    "category": category_name,
                    "l": _clean_num(lecture),
                    "t": _clean_num(tutorial),
                    "p": _clean_num(lab),
                    "tcp": _clean_num(total_hours),
                    "credits": _clean_num(credits),
                    "row_type": (
                        "THEORY" if lecture > 0 and lab == 0 else
                        "LABORATORY" if lab > 0 and lecture == 0 else
                        "INTEGRATED" if lecture > 0 and lab > 0 else
                        "OTHERS"
                    )
                })

            table_rows.append({
                "department": dept_name,
                "course_code": course.course_code or "-",
                "title": course.title or "-",
                "year": course.year or "-",
                "semester": course.semester or "-",
                "category": category_name,
                "lecture": _clean_num(lecture),
                "tutorial": _clean_num(tutorial),
                "lab": _clean_num(lab),
                "total_hours": _clean_num(total_hours),
                "credits": _clean_num(credits),
                "is_active": course.is_active,
            })

    regulation_category_display = OrderedDict()
    for cat, value in regulation_category_totals.items():
        regulation_category_display[cat] = {
            "credits": _clean_num(value),
            "percent": _percent(value, total_credits_all)
        }

    semester_total_row = OrderedDict()
    for sem in semester_labels:
        semester_total_row[sem] = _clean_num(
            sum(semester_category_matrix[cat].get(sem, 0) for cat in semester_category_matrix)
        )

    all_semester_sections = OrderedDict()
    for sem in semester_labels:
        rows = semester_detailed_rows.get(sem, [])
        grouped = OrderedDict([
            ("THEORY", [r for r in rows if r["row_type"] == "THEORY"]),
            ("LABORATORY", [r for r in rows if r["row_type"] == "LABORATORY"]),
            ("INTEGRATED", [r for r in rows if r["row_type"] == "INTEGRATED"]),
            ("OTHERS", [r for r in rows if r["row_type"] == "OTHERS"]),
        ])
        semester_total_hours = _clean_num(sum(_to_float(r["tcp"]) for r in rows))
        semester_total_credits = _clean_num(sum(_to_float(r["credits"]) for r in rows))

        all_semester_sections[sem] = {
            "rows": rows,
            "grouped_rows": grouped,
            "total_hours": semester_total_hours,
            "total_credits": semester_total_credits,
        }

    for dept_id, stats in department_stats.items():
        stats["category_rows"] = []
        for cat, cat_value in stats["category_credits"].items():
            stats["category_rows"].append({
                "name": cat,
                "credits": _clean_num(cat_value),
                "percent": _percent(cat_value, stats["total_credits"])
            })

    dept_labels = [stats["name"] for stats in department_stats.values()]
    dept_credit_values = [_clean_num(stats["total_credits"]) for stats in department_stats.values()]
    dept_course_values = [stats["course_count"] for stats in department_stats.values()]

    category_labels = list(category_stats.keys())
    category_values = [_clean_num(v) for v in category_stats.values()]

    semester_chart_labels = list(semester_total_row.keys())
    semester_credit_values = [_clean_num(semester_total_row[sem]) for sem in semester_chart_labels]
    semester_course_values = [
        semester_stats.get(f"Sem {sem}", {}).get("count", 0) for sem in semester_chart_labels
    ]

    return {
        "courses": courses,
        "dynamic_categories": list(regulation_category_display.keys()),
        "dynamic_semesters": semester_labels,
        "all_semester_sections": all_semester_sections,

        "total_courses": total_courses,
        "total_departments": total_departments,
        "total_credits_all": _clean_num(total_credits_all),
        "total_hours_all": _clean_num(total_hours_all),

        "department_stats": department_stats,
        "semester_stats": semester_stats,
        "category_stats": category_stats,

        "theory_total": _clean_num(theory_total),
        "laboratory_total": _clean_num(laboratory_total),
        "tutorial_total": _clean_num(tutorial_total),
        "integrated_total": _clean_num(integrated_total),

        "table_rows": table_rows,

        "dept_labels": dept_labels,
        "dept_credit_values": dept_credit_values,
        "dept_course_values": dept_course_values,
        "category_labels": category_labels,
        "category_values": category_values,
        "semester_labels": semester_chart_labels,
        "semester_credit_values": semester_credit_values,
        "semester_course_values": semester_course_values,

        "regulation_category_display": regulation_category_display,
        "semester_category_matrix": semester_category_matrix,
        "semester_total_row": semester_total_row,
    }

@check_permission("course_analysis_dashboard")
def course_analysis_dashboard(request):
    regulations = Regulations.objects.all().order_by("year")
    regulations = Regulations.objects.all().order_by("year")
    selected_regulation_id = request.GET.get("regulation_id")
    selected_filters = {
        "department_id": (request.GET.get("department_id") or "").strip(),
        "year": (request.GET.get("year") or "").strip(),
        "semester": (request.GET.get("semester") or "").strip(),
        "category_id": (request.GET.get("category_id") or "").strip(),
        "is_active": (request.GET.get("is_active") or "").strip().lower(),
        "q": (request.GET.get("q") or "").strip(),
    }

    # comparison table should also be dynamic
    all_comparison_categories = set()
    regulation_comparison_columns = []

    for reg in regulations:
        reg_data = _build_regulation_analytics(reg)
        regulation_comparison_columns.append({
            "id": reg.id,
            "name": _get_regulation_short_name(reg),
            "regulation": reg,
            "total_credits": reg_data["total_credits_all"],
            "category_display": reg_data["regulation_category_display"],
        })
        all_comparison_categories.update(reg_data["regulation_category_display"].keys())

    all_comparison_categories = sorted(all_comparison_categories)

    comparison_table_rows = []
    for cat in all_comparison_categories:
        row = {"category": cat, "values": []}
        for reg_col in regulation_comparison_columns:
            item = reg_col["category_display"].get(cat)
            if item and _to_float(item["credits"]) > 0:
                row["values"].append(f'{item["credits"]} ({item["percent"]})')
            else:
                row["values"].append("-")
        comparison_table_rows.append(row)

    comparison_total_row = [col["total_credits"] for col in regulation_comparison_columns]

    selected_regulation = None
    selected_data = {
        "courses": Course.objects.none(),
        "dynamic_categories": [],
        "dynamic_semesters": [],
        "total_courses": 0,
        "total_departments": 0,
        "total_credits_all": 0,
        "total_hours_all": 0,
        "department_stats": OrderedDict(),
        "semester_stats": OrderedDict(),
        "category_stats": OrderedDict(),
        "theory_total": 0,
        "laboratory_total": 0,
        "tutorial_total": 0,
        "integrated_total": 0,
        "table_rows": [],
        "dept_labels": [],
        "dept_credit_values": [],
        "dept_course_values": [],
        "category_labels": [],
        "category_values": [],
        "semester_labels": [],
        "semester_credit_values": [],
        "semester_course_values": [],
        "regulation_category_display": OrderedDict(),
        "semester_category_matrix": OrderedDict(),
        "semester_total_row": OrderedDict(),
        "selected_semester": "",
        "selected_semester_rows": [],
        "semester_grouped_rows": OrderedDict(),
    }

    if selected_regulation_id:
        selected_regulation = Regulations.objects.filter(id=selected_regulation_id).first()
        if selected_regulation:
            selected_data = _build_regulation_analytics(selected_regulation, filters=selected_filters)

    filter_options = {
        "departments": [],
        "years": [],
        "semesters": [],
        "categories": [],
    }
    if selected_regulation:
        base_qs = Course.objects.filter(regulation=selected_regulation).select_related("department", "elective")
        filter_options["departments"] = (
            Add_Department.objects.filter(
                id__in=base_qs.exclude(department_id__isnull=True).values_list("department_id", flat=True).distinct()
            ).order_by("Department")
        )
        filter_options["years"] = (
            base_qs.exclude(year__isnull=True).exclude(year="").values_list("year", flat=True).distinct().order_by("year")
        )
        filter_options["semesters"] = (
            base_qs.exclude(semester__isnull=True).exclude(semester="").values_list("semester", flat=True).distinct().order_by("semester")
        )
        filter_options["categories"] = (
            Course_category.objects.filter(
                id__in=base_qs.exclude(elective_id__isnull=True).values_list("elective_id", flat=True).distinct()
            ).order_by("Course_category_name")
        )

    context = {
        "regulations": regulations,
        "selected_regulation": selected_regulation,

        "regulation_comparison_columns": regulation_comparison_columns,
        "comparison_table_rows": comparison_table_rows,
        "comparison_total_row": comparison_total_row,
        "selected_filters": selected_filters,
        "filter_options": filter_options,
        **selected_data,
    }
    return render(request, "course_management/course_analysis_dashboard.html", context)


# ==================================================================
# Course Enrollment Dashboard
# Regulation -> Degree -> Department -> Year -> Semester cascading filters,
# scoped by viewer identity:
#   - Global user (GlobalUsers.global_user, same mechanism as fee_view)  -> sees everything
#   - HOD                                                                -> locked to own department
#   - Ordinary faculty                                                   -> only their assigned subjects
# ==================================================================
from django.db.models import Count
from user_accounts.models import GlobalUsers


def _is_global_course_user(user):
    """
    Same 'global user' mechanism as fee_view (student_management) — set via
    the Create Global Users admin screen. This is granted per (employee,
    role), not per employee — the same person can hold multiple accounts
    under different roles (e.g. HOD on one, Vice Principal on another), and
    global access on one role must not leak into the others. Only the role
    the viewer is actually logged in as is checked.
    """
    emp_id = str(getattr(user, "Employee_id", "") or "").strip()
    role_id = getattr(user, "role_id", None)
    if not emp_id or role_id is None:
        return False
    return GlobalUsers.objects.filter(
        employee_id=emp_id, role_id=str(role_id), global_user=True
    ).exists()


def _ce_role_text(user):
    role = getattr(user, "role", None)
    if role is None:
        return ""
    if isinstance(role, str):
        return role.strip().lower()
    return str(getattr(role, "role", "") or "").strip().lower()


def _ce_is_hod(user):
    rt = _ce_role_text(user)
    return (rt == "hod") or rt.startswith("hod") or ("head of department" in rt)


def _ce_resolve_faculty(user):
    emp_id = getattr(user, "Employee_id", None)
    if not emp_id:
        return None
    return (
        general_information.objects
        .filter(faculty_id=emp_id)
        .select_related("department", "department__degree")
        .first()
    )


def _ce_resolve_mode(request):
    """Returns (mode, faculty, hod_dept). mode is one of principal/hod/faculty/none."""
    faculty = _ce_resolve_faculty(request.user)

    if _is_global_course_user(request.user):
        return "principal", faculty, None

    if faculty and _ce_is_hod(request.user) and faculty.department:
        return "hod", faculty, faculty.department

    if faculty:
        return "faculty", faculty, None

    return "none", faculty, None


def _ce_safe_str(v):
    return (v or "").strip() if isinstance(v, str) else ("" if v is None else str(v).strip())


def _ce_scoped_course_queryset(mode, faculty, hod_dept):
    """The base Course queryset for a given viewer, before any of the
    regulation/degree/department/year/semester filters are applied."""
    qs = Course.objects.filter(is_active=True).select_related(
        "department", "department__degree", "regulation", "elective"
    )
    if mode == "hod" and hod_dept:
        return qs.filter(department_id=hod_dept.id)
    if mode == "faculty" and faculty:
        assigned_ids = (
            AssignSubjectFaculty.objects
            .filter(faculty=faculty, is_active=True)
            .values_list("course_id", flat=True)
        )
        return qs.filter(id__in=assigned_ids)
    if mode == "none":
        return Course.objects.none()
    return qs  # principal: unrestricted


def _ce_apply_filters(qs, regulation_id, degree_id, department_id, year, semester, q, type_filter=None, academic_year=None):
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if degree_id:
        qs = qs.filter(department__degree_id=degree_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    if year:
        qs = qs.filter(year=year)
    if semester:
        qs = qs.filter(semester=semester)
    if q:
        qs = qs.filter(Q(course_code__icontains=q) | Q(title__icontains=q))
    if type_filter in ("honours", "regular"):
        honours_ids = HonoursCourse.objects.filter(
            academic_year=academic_year or get_academic_year()
        ).values_list("course_id", flat=True)
        if type_filter == "honours":
            qs = qs.filter(id__in=honours_ids)
        else:
            qs = qs.exclude(id__in=honours_ids)
    return qs


def _ce_recent_academic_years(n=3):
    """Current academic year (per get_academic_year()) plus the previous
    (n-1), newest first — e.g. ['2026-2027', '2025-2026', '2024-2025']."""
    start_year = int(get_academic_year().split("-")[0])
    return [f"{start_year - i}-{start_year - i + 1}" for i in range(n)]


def _ce_build_rows(courses, mode, faculty, regulation_id, academic_year=None):
    """
    One row per (course, assigned faculty, section) combination — batched (a
    fixed handful of queries no matter how many courses match) instead of
    querying per course. A course with nobody assigned yet still appears
    once, with faculty "Not Assigned" and its raw enrollment count as-is
    (principal/hod views only; the faculty view only ever shows that
    faculty's own subjects, so an unassigned course can never appear there).
    Enrollment counts and the Honours flag are both scoped to academic_year.
    """
    courses = list(courses)
    course_ids = [c.id for c in courses]
    if not course_ids:
        return []

    academic_year = academic_year or get_academic_year()

    asf_qs = (
        AssignSubjectFaculty.objects
        .filter(course_id__in=course_ids, is_active=True)
        .select_related("faculty")
    )
    if mode == "faculty" and faculty:
        asf_qs = asf_qs.filter(faculty=faculty)
    if regulation_id:
        asf_qs = asf_qs.filter(regulation_id=regulation_id)

    asf_by_course = {}
    for a in asf_qs:
        asf_by_course.setdefault(a.course_id, []).append(a)

    enroll_qs = CourseEnrollment.objects.filter(
        course_id__in=course_ids, enroll=True, academic_year=academic_year
    )
    if regulation_id:
        enroll_qs = enroll_qs.filter(regulation_id=regulation_id)
    count_rows = (
        enroll_qs.values("course_id", "faculty_id", "section").annotate(cnt=Count("id"))
    )
    count_by_course_faculty_section = {}
    total_by_course = {}
    for r in count_rows:
        sec = _ce_safe_str(r["section"]).upper()
        key = (r["course_id"], r["faculty_id"], sec)
        count_by_course_faculty_section[key] = count_by_course_faculty_section.get(key, 0) + r["cnt"]
        total_by_course[r["course_id"]] = total_by_course.get(r["course_id"], 0) + r["cnt"]

    honours_course_ids = set(
        HonoursCourse.objects
        .filter(course_id__in=course_ids, academic_year=academic_year)
        .values_list("course_id", flat=True)
    )

    rows = []
    for course in courses:
        is_honours = course.id in honours_course_ids
        assignments = asf_by_course.get(course.id, [])
        if assignments:
            seen = set()
            for a in assignments:
                fac = a.faculty
                fid = fac.id if fac else None
                section = _ce_safe_str(a.section).upper()
                dedup_key = (fid, section)
                if dedup_key in seen:
                    continue
                seen.add(dedup_key)
                rows.append({
                    "course": course,
                    "faculty": fac,
                    "section": section,
                    "enrolled_count": count_by_course_faculty_section.get((course.id, fid, section), 0),
                    "is_honours": is_honours,
                })
        elif mode != "faculty":
            rows.append({
                "course": course,
                "faculty": None,
                "section": "",
                "enrolled_count": total_by_course.get(course.id, 0),
                "is_honours": is_honours,
            })

    return rows


def _ce_filters_from_request(request, hod_dept=None):
    regulation_id = _ce_safe_str(request.GET.get("regulation"))
    degree_id = _ce_safe_str(request.GET.get("degree"))
    department_id = _ce_safe_str(request.GET.get("department"))
    year = _ce_safe_str(request.GET.get("year"))
    semester = _ce_safe_str(request.GET.get("semester"))
    q = _ce_safe_str(request.GET.get("q"))
    academic_year = _ce_safe_str(request.GET.get("academic_year")) or get_academic_year()
    type_filter = _ce_safe_str(request.GET.get("type")).lower()
    if type_filter not in ("honours", "regular"):
        type_filter = ""

    if hod_dept:
        department_id = str(hod_dept.id)
        degree_id = str(hod_dept.degree_id) if hod_dept.degree_id else degree_id

    return regulation_id, degree_id, department_id, year, semester, q, academic_year, type_filter


@course_management
@check_permission("course_enrollment_dashboard")
def course_enrollment_dashboard(request):
    mode, faculty, hod_dept = _ce_resolve_mode(request)
    regulation_id, degree_id, department_id, year, semester, q, academic_year, type_filter = _ce_filters_from_request(request, hod_dept)
    section = _ce_safe_str(request.GET.get("section")).upper()

    base_qs = _ce_scoped_course_queryset(mode, faculty, hod_dept)
    courses = _ce_apply_filters(
        base_qs, regulation_id, degree_id, department_id, year, semester, q, type_filter, academic_year
    ).order_by("course_code")

    rows = _ce_build_rows(courses, mode, faculty, regulation_id, academic_year)
    if section:
        rows = [r for r in rows if (r.get("section") or "").upper() == section]

    section_options = list(
        SectionMaster.objects.values_list("section", flat=True)
        .exclude(section__isnull=True).exclude(section__exact="")
        .distinct().order_by("section")
    )

    paginator = Paginator(rows, 50)
    page_number = request.GET.get("page") or 1
    try:
        page_obj = paginator.page(page_number)
    except Exception:
        page_obj = paginator.page(1)

    # ---- cascading dropdown seed options (regulation list is always full;
    # everything after it is filled in by JS via the api_ce_* endpoints) ----
    regulation_ids = base_qs.exclude(regulation_id__isnull=True).values_list("regulation_id", flat=True).distinct()
    regulations = Regulations.objects.filter(id__in=regulation_ids).order_by("year")

    degree_options = []
    department_options = Add_Department.objects.none()
    year_options = []
    semester_options = []

    if mode == "hod" and hod_dept:
        department_options = Add_Department.objects.filter(id=hod_dept.id)
        year_options = list(
            base_qs.exclude(year__isnull=True).exclude(year="").values_list("year", flat=True).distinct().order_by("year")
        )
        semester_options = list(
            base_qs.exclude(semester__isnull=True).exclude(semester="").values_list("semester", flat=True).distinct().order_by("semester")
        )
    else:
        # Each level is scoped by whatever filters are already set above it,
        # but none of them require the deeper ones to be picked first — e.g.
        # Year/Semester are usable even if Degree/Department were left as
        # "All". Only what's actually selected narrows the next level down.
        scoped = base_qs
        if regulation_id:
            scoped = scoped.filter(regulation_id=regulation_id)

        degree_ids = scoped.exclude(department__degree_id__isnull=True).values_list("department__degree_id", flat=True).distinct()
        degree_options = list(Degree.objects.filter(id__in=degree_ids).order_by("degree").values("id", "degree", "degree_code"))
        if degree_id:
            scoped = scoped.filter(department__degree_id=degree_id)

        dept_ids = scoped.exclude(department_id__isnull=True).values_list("department_id", flat=True).distinct()
        department_options = Add_Department.objects.filter(id__in=dept_ids).order_by("Department")
        if department_id:
            scoped = scoped.filter(department_id=department_id)

        year_options = list(scoped.exclude(year__isnull=True).exclude(year="").values_list("year", flat=True).distinct().order_by("year"))
        if year:
            scoped = scoped.filter(year=year)

        semester_options = list(scoped.exclude(semester__isnull=True).exclude(semester="").values_list("semester", flat=True).distinct().order_by("semester"))

    context = {
        "mode": mode,
        "faculty": faculty,
        "hod_dept": hod_dept,

        "regulations": regulations,
        "degree_options": degree_options,
        "department_options": department_options,
        "year_options": year_options,
        "semester_options": semester_options,

        "selected_regulation_id": regulation_id,
        "selected_degree_id": degree_id,
        "selected_department_id": department_id,
        "selected_year": year,
        "selected_semester": semester,
        "q": q,
        "selected_type": type_filter,

        "section_options": section_options,
        "selected_section": section,

        "academic_year_options": _ce_recent_academic_years(),
        "selected_academic_year": academic_year,

        "page_obj": page_obj,
        "total_rows": paginator.count,
    }
    return render(request, "course_management/course_enrollment_dashboard.html", context)


def _ce_dropdown_base_qs(request):
    mode, faculty, hod_dept = _ce_resolve_mode(request)
    return _ce_scoped_course_queryset(mode, faculty, hod_dept), mode, hod_dept


@check_permission("course_enrollment_dashboard")
def api_ce_degrees(request):
    qs, mode, hod_dept = _ce_dropdown_base_qs(request)
    regulation_id = _ce_safe_str(request.GET.get("regulation_id"))
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if mode == "hod" and hod_dept:
        degrees = Degree.objects.filter(id=hod_dept.degree_id) if hod_dept.degree_id else Degree.objects.none()
    else:
        degree_ids = qs.exclude(department__degree_id__isnull=True).values_list("department__degree_id", flat=True).distinct()
        degrees = Degree.objects.filter(id__in=degree_ids).order_by("degree")
    return JsonResponse({"results": list(degrees.values("id", "degree", "degree_code"))})


@check_permission("course_enrollment_dashboard")
def api_ce_departments(request):
    qs, mode, hod_dept = _ce_dropdown_base_qs(request)
    regulation_id = _ce_safe_str(request.GET.get("regulation_id"))
    degree_id = _ce_safe_str(request.GET.get("degree_id"))
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if degree_id:
        qs = qs.filter(department__degree_id=degree_id)
    if mode == "hod" and hod_dept:
        departments = Add_Department.objects.filter(id=hod_dept.id)
    else:
        dept_ids = qs.exclude(department_id__isnull=True).values_list("department_id", flat=True).distinct()
        departments = Add_Department.objects.filter(id__in=dept_ids).order_by("Department")
    return JsonResponse({"results": list(departments.values("id", "Department", "Department_code"))})


@check_permission("course_enrollment_dashboard")
def api_ce_years(request):
    qs, mode, hod_dept = _ce_dropdown_base_qs(request)
    regulation_id = _ce_safe_str(request.GET.get("regulation_id"))
    degree_id = _ce_safe_str(request.GET.get("degree_id"))
    department_id = _ce_safe_str(request.GET.get("department_id"))
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if degree_id:
        qs = qs.filter(department__degree_id=degree_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    years = qs.exclude(year__isnull=True).exclude(year="").values_list("year", flat=True).distinct().order_by("year")
    return JsonResponse({"results": list(years)})


@check_permission("course_enrollment_dashboard")
def api_ce_semesters(request):
    qs, mode, hod_dept = _ce_dropdown_base_qs(request)
    regulation_id = _ce_safe_str(request.GET.get("regulation_id"))
    degree_id = _ce_safe_str(request.GET.get("degree_id"))
    department_id = _ce_safe_str(request.GET.get("department_id"))
    year = _ce_safe_str(request.GET.get("year"))
    if regulation_id:
        qs = qs.filter(regulation_id=regulation_id)
    if degree_id:
        qs = qs.filter(department__degree_id=degree_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    if year:
        qs = qs.filter(year=year)
    semesters = qs.exclude(semester__isnull=True).exclude(semester="").values_list("semester", flat=True).distinct().order_by("semester")
    return JsonResponse({"results": list(semesters)})


@check_permission("course_enrollment_dashboard")
def api_ce_sections(request):
    sections = list(
        SectionMaster.objects.values_list("section", flat=True)
        .exclude(section__isnull=True).exclude(section__exact="")
        .distinct().order_by("section")
    )
    response = JsonResponse({"results": sections})
    response["Cache-Control"] = "no-store"
    return response


@check_permission("course_enrollment_dashboard")
def course_enrollment_dashboard_pdf(request):
    from reportlab.platypus import SimpleDocTemplate

    mode, faculty, hod_dept = _ce_resolve_mode(request)
    regulation_id, degree_id, department_id, year, semester, q, academic_year, type_filter = _ce_filters_from_request(request, hod_dept)
    section = _ce_safe_str(request.GET.get("section")).upper()

    base_qs = _ce_scoped_course_queryset(mode, faculty, hod_dept)
    courses = _ce_apply_filters(
        base_qs, regulation_id, degree_id, department_id, year, semester, q, type_filter, academic_year
    ).order_by("course_code")
    rows = _ce_build_rows(courses, mode, faculty, regulation_id, academic_year)
    if section:
        rows = [r for r in rows if (r.get("section") or "").upper() == section]

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="Course_Enrollment_Report.pdf"'

    doc = SimpleDocTemplate(
        response, pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=32 * mm, bottomMargin=18 * mm,
        title="Course Enrollment Report",
    )

    styles = getSampleStyleSheet()
    kpi_style = ParagraphStyle("ceKpi", parent=styles["Normal"], fontSize=10, textColor=DARK_GRAY, spaceAfter=6)
    head_cell = ParagraphStyle("ceHead", parent=styles["Normal"], fontSize=8.5, alignment=TA_CENTER,
                                textColor=colors.white, fontName="Helvetica-Bold")
    cell = ParagraphStyle("ceCell", parent=styles["Normal"], fontSize=8.5, alignment=TA_CENTER, textColor=DARK_GRAY)
    cell_left = ParagraphStyle("ceCellLeft", parent=cell, alignment=TA_LEFT)

    elements = []

    regulation_label = "All Regulations"
    if regulation_id:
        reg = Regulations.objects.filter(id=regulation_id).first()
        if reg:
            regulation_label = str(reg)

    degree_label = "All Degrees"
    if degree_id:
        deg = Degree.objects.filter(id=degree_id).first()
        if deg:
            degree_label = deg.degree

    department_label = "All Departments"
    if department_id:
        dept = Add_Department.objects.filter(id=department_id).first()
        if dept:
            department_label = dept.Department

    year_label = _pdf_text(year) or "All Years"
    semester_label = _pdf_text(semester) or "All Semesters"

    label_style = ParagraphStyle("ceFilterLabel", parent=styles["Normal"], fontSize=9,
                                  textColor=MEDIUM_GRAY, fontName="Helvetica-Bold")
    value_style = ParagraphStyle("ceFilterValue", parent=styles["Normal"], fontSize=9.5, textColor=DARK_GRAY)

    type_label = {"honours": "Honours", "regular": "Regular"}.get(type_filter, "All (Honours + Regular)")

    filters_data = [
        [Paragraph("Regulation", label_style), Paragraph(_pdf_text(regulation_label), value_style),
         Paragraph("Degree", label_style), Paragraph(_pdf_text(degree_label), value_style)],
        [Paragraph("Department", label_style), Paragraph(_pdf_text(department_label), value_style),
         Paragraph("Year", label_style), Paragraph(year_label, value_style)],
        [Paragraph("Semester", label_style), Paragraph(semester_label, value_style),
         Paragraph("Academic Year", label_style), Paragraph(_pdf_text(academic_year), value_style)],
        [Paragraph("Type", label_style), Paragraph(_pdf_text(type_label), value_style),
         Paragraph("Section", label_style), Paragraph(_pdf_text(section) or "All Sections", value_style)],
    ]
    filters_table = Table(filters_data, colWidths=[26 * mm, 65 * mm, 26 * mm, 65 * mm])
    filters_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
        ("BACKGROUND", (0, 0), (0, -1), BG_GRAY),
        ("BACKGROUND", (2, 0), (2, -1), BG_GRAY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(Spacer(1, 2 * mm))
    elements.append(filters_table)
    elements.append(Spacer(1, 3 * mm))

    elements.append(Paragraph("<b>Total Rows:</b> " + str(len(rows)), kpi_style))
    elements.append(Spacer(1, 4 * mm))

    data = [[
        Paragraph("<b>S.No</b>", head_cell),
        Paragraph("<b>Course Code</b>", head_cell),
        Paragraph("<b>Course Name</b>", head_cell),
        Paragraph("<b>Type</b>", head_cell),
        Paragraph("<b>Category</b>", head_cell),
        Paragraph("<b>Faculty</b>", head_cell),
        Paragraph("<b>Section</b>", head_cell),
        Paragraph("<b>Enrolled</b>", head_cell),
    ]]
    if rows:
        for idx, r in enumerate(rows, start=1):
            course = r["course"]
            category = course.elective.Course_category_name if course.elective else "-"
            faculty_name = r["faculty"].name if r["faculty"] else "Not Assigned"
            course_type = "H" if r.get("is_honours") else "R"
            data.append([
                Paragraph(str(idx), cell),
                Paragraph(_pdf_text(course.course_code) or "-", cell),
                Paragraph(_pdf_text(course.title) or "-", cell_left),
                Paragraph(course_type, cell),
                Paragraph(_pdf_text(category) or "-", cell_left),
                Paragraph(_pdf_text(faculty_name), cell_left),
                Paragraph(_pdf_text(r.get("section")) or "-", cell),
                Paragraph(str(r["enrolled_count"]), cell),
            ])
    else:
        data.append([Paragraph("No courses found for the selected filters.", cell_left), "", "", "", "", "", "", ""])

    table = Table(data, colWidths=[10 * mm, 22 * mm, 42 * mm, 13 * mm, 25 * mm, 32 * mm, 17 * mm, 16 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)

    header_cb = _make_ramco_pdf_header("COURSE ENROLLMENT REPORT")
    doc.build(elements, onFirstPage=header_cb, onLaterPages=header_cb)
    return response


# ==================================================================
# Attendance Lag Dashboard
# Regulation -> Degree -> Department -> Year -> Semester -> Academic Year
# (current + past 3), same cascading filters and global-user scoping as
# course_enrollment_dashboard. For each (course, batch, section) offering,
# shows how many enrolled students are lagging in attendance, and lets you
# drill into their reg no / name / percentage / mark.
#
# "Lagging" is read directly off the AttendancePercentageMaster bands (not a
# hardcoded 75%/50% cutoff like the older attendance PDF): a student's
# percentage is matched against the active bands exactly like
# faculty_course_students_attendance_pdf computes it (present + On Duty,
# out of distinct (date, period) slots held), and they count as lagging if
# their matched band is the lowest configured mark (or no band matches at
# all, i.e. below the lowest band).
# ==================================================================
import operator
from functools import reduce
from user_accounts.models import AttendancePercentageMaster


def _al_lag_bands():
    """
    Active AttendancePercentageMaster bands, plus the lag cutoff percentage.

    The master table's lowest band (e.g. 0-74 => mark 0) *is* the "lagging"
    zone itself — so the cutoff is the highest percentage_to among whichever
    band(s) share the minimum configured mark (0), not simply the lowest
    percentage_from (which is 0 and would flag everyone).

    Lagging is decided as perc <= cutoff, not by exact band matching:
    adjacent bands are often configured with small gaps between them (e.g.
    90-91%), and a student landing in such a gap has perfectly fine
    attendance — treating "no band matched" as automatically lagging would
    misclassify them.
    """
    bands = list(AttendancePercentageMaster.objects.filter(is_active=True).order_by("percentage_from"))
    if not bands:
        return bands, Decimal("75")
    min_mark = min(b.attendance_mark for b in bands)
    lag_cutoff = max(b.percentage_to for b in bands if b.attendance_mark == min_mark)
    return bands, lag_cutoff


def _al_match_mark(perc, bands):
    for b in bands:
        if b.percentage_from <= perc <= b.percentage_to:
            return b.attendance_mark
    return None


def _al_group_filter_q(groups):
    clauses = [
        Q(course_id=g["course_id"], batch=g["batch"], section=g["section"], year=g["year"], semester=g["semester"])
        for g in groups
    ]
    return reduce(operator.or_, clauses) if clauses else Q(pk__in=[])


def _al_dropdown_base_qs(request):
    return _ce_dropdown_base_qs(request)


@check_permission("attendance_lag_dashboard")
def api_al_degrees(request):
    return api_ce_degrees(request)


@check_permission("attendance_lag_dashboard")
def api_al_departments(request):
    return api_ce_departments(request)


@check_permission("attendance_lag_dashboard")
def api_al_years(request):
    return api_ce_years(request)


@check_permission("attendance_lag_dashboard")
def api_al_semesters(request):
    return api_ce_semesters(request)


@check_permission("attendance_lag_dashboard")
def api_al_sections(request):
    return api_ce_sections(request)


@course_management
@check_permission("attendance_lag_dashboard")
def attendance_lag_dashboard(request):
    mode, faculty, hod_dept = _ce_resolve_mode(request)
    regulation_id, degree_id, department_id, year, semester, q, academic_year, _type_unused = (
        _ce_filters_from_request(request, hod_dept)
    )
    section = _ce_safe_str(request.GET.get("section")).upper()

    base_qs = _ce_scoped_course_queryset(mode, faculty, hod_dept)
    courses = _ce_apply_filters(base_qs, regulation_id, degree_id, department_id, year, semester, q).order_by("course_code")
    course_ids = list(courses.values_list("id", flat=True))

    enroll_qs = CourseEnrollment.objects.filter(course_id__in=course_ids, enroll=True, academic_year=academic_year)
    if mode == "faculty" and faculty:
        enroll_qs = enroll_qs.filter(faculty=faculty)

    all_groups = list(
        enroll_qs.values("course_id", "batch", "section", "year", "semester")
        .annotate(total_students=Count("id"))
        .order_by("course_id", "batch", "section")
    )
    if section:
        all_groups = [g for g in all_groups if (g.get("section") or "").upper() == section]

    section_options = list(
        SectionMaster.objects.values_list("section", flat=True)
        .exclude(section__isnull=True).exclude(section__exact="")
        .distinct().order_by("section")
    )

    paginator = Paginator(all_groups, 50)
    page_number = request.GET.get("page") or 1
    try:
        page_obj = paginator.page(page_number)
    except Exception:
        page_obj = paginator.page(1)

    page_groups = list(page_obj.object_list)
    bands, lag_cutoff_percentage = _al_lag_bands()

    rows = []
    if page_groups:
        courses_by_id = {
            c.id: c for c in
            Course.objects.filter(id__in={g["course_id"] for g in page_groups}).select_related("department", "elective")
        }
        group_q = _al_group_filter_q(page_groups)

        roster_qs = CourseEnrollment.objects.filter(group_q, enroll=True, academic_year=academic_year).select_related("student")
        if mode == "faculty" and faculty:
            roster_qs = roster_qs.filter(faculty=faculty)

        roster_by_group = defaultdict(list)
        for e in roster_qs:
            key = (e.course_id, e.batch, e.section, e.year, e.semester)
            roster_by_group[key].append(e.student)

        att_rows = HourAttendance.objects.filter(group_q, academic_year=academic_year).values(
            "course_id", "batch", "section", "year", "semester", "student_id", "status", "date", "period"
        )
        dates_by_group = defaultdict(set)
        counts_by_group_student = defaultdict(lambda: {"present": 0, "absent": 0, "od": 0})
        for r in att_rows:
            key = (r["course_id"], r["batch"], r["section"], r["year"], r["semester"])
            dates_by_group[key].add((r["date"], r["period"]))
            bucket = counts_by_group_student[(key, r["student_id"])]
            if r["status"] == "Present":
                bucket["present"] += 1
            elif r["status"] == "Absent":
                bucket["absent"] += 1
            elif r["status"] == "On Duty":
                bucket["od"] += 1

        for g in page_groups:
            key = (g["course_id"], g["batch"], g["section"], g["year"], g["semester"])
            total_hours = len(dates_by_group.get(key, ()))
            roster = roster_by_group.get(key, [])

            lagging = []
            if total_hours and roster:
                for st in roster:
                    counts = counts_by_group_student.get((key, st.id))
                    if not counts:
                        continue
                    attended = counts["present"] + counts["od"]
                    perc = round((attended / total_hours) * 100.0, 2)
                    mark = _al_match_mark(Decimal(str(perc)), bands)
                    is_lag = Decimal(str(perc)) <= lag_cutoff_percentage
                    if is_lag:
                        lagging.append({
                            "reg_no": _ce_safe_str(getattr(st, "reg_no", "")) or "-",
                            "name": _ce_safe_str(getattr(st, "name", "")) or "-",
                            "present": counts["present"],
                            "absent": counts["absent"],
                            "od": counts["od"],
                            "total_hours": total_hours,
                            "percentage": perc,
                            "mark": str(mark) if mark is not None else "-",
                        })

            lagging.sort(key=lambda r: r["percentage"])

            rows.append({
                "course": courses_by_id.get(g["course_id"]),
                "batch": g["batch"],
                "section": g["section"],
                "year": g["year"],
                "semester": g["semester"],
                "total_students": g["total_students"],
                "total_hours": total_hours,
                "lag_count": len(lagging),
                "lagging_students_json": json.dumps(lagging),
            })

    # ---- cascading dropdown seed options (identical pattern to course_enrollment_dashboard) ----
    regulation_ids = base_qs.exclude(regulation_id__isnull=True).values_list("regulation_id", flat=True).distinct()
    regulations = Regulations.objects.filter(id__in=regulation_ids).order_by("year")

    degree_options = []
    department_options = Add_Department.objects.none()
    year_options = []
    semester_options = []

    if mode == "hod" and hod_dept:
        department_options = Add_Department.objects.filter(id=hod_dept.id)
        year_options = list(
            base_qs.exclude(year__isnull=True).exclude(year="").values_list("year", flat=True).distinct().order_by("year")
        )
        semester_options = list(
            base_qs.exclude(semester__isnull=True).exclude(semester="").values_list("semester", flat=True).distinct().order_by("semester")
        )
    else:
        scoped = base_qs
        if regulation_id:
            scoped = scoped.filter(regulation_id=regulation_id)

        degree_ids = scoped.exclude(department__degree_id__isnull=True).values_list("department__degree_id", flat=True).distinct()
        degree_options = list(Degree.objects.filter(id__in=degree_ids).order_by("degree").values("id", "degree", "degree_code"))
        if degree_id:
            scoped = scoped.filter(department__degree_id=degree_id)

        dept_ids = scoped.exclude(department_id__isnull=True).values_list("department_id", flat=True).distinct()
        department_options = Add_Department.objects.filter(id__in=dept_ids).order_by("Department")
        if department_id:
            scoped = scoped.filter(department_id=department_id)

        year_options = list(scoped.exclude(year__isnull=True).exclude(year="").values_list("year", flat=True).distinct().order_by("year"))
        if year:
            scoped = scoped.filter(year=year)

        semester_options = list(scoped.exclude(semester__isnull=True).exclude(semester="").values_list("semester", flat=True).distinct().order_by("semester"))

    context = {
        "mode": mode,
        "faculty": faculty,
        "hod_dept": hod_dept,

        "regulations": regulations,
        "degree_options": degree_options,
        "department_options": department_options,
        "year_options": year_options,
        "semester_options": semester_options,

        "selected_regulation_id": regulation_id,
        "selected_degree_id": degree_id,
        "selected_department_id": department_id,
        "selected_year": year,
        "selected_semester": semester,
        "q": q,

        "section_options": section_options,
        "selected_section": section,

        "academic_year_options": _ce_recent_academic_years(4),
        "selected_academic_year": academic_year,
        "min_required_percentage": lag_cutoff_percentage,

        "rows": rows,
        "page_obj": page_obj,
        "total_groups": paginator.count,
    }
    return render(request, "course_management/attendance_lag_dashboard.html", context)


@check_permission("attendance_lag_dashboard")
def attendance_lag_dashboard_pdf(request):
    from reportlab.platypus import SimpleDocTemplate

    mode, faculty, hod_dept = _ce_resolve_mode(request)
    regulation_id, degree_id, department_id, year, semester, q, academic_year, _type_unused = (
        _ce_filters_from_request(request, hod_dept)
    )
    section = _ce_safe_str(request.GET.get("section")).upper()

    base_qs = _ce_scoped_course_queryset(mode, faculty, hod_dept)
    courses = _ce_apply_filters(base_qs, regulation_id, degree_id, department_id, year, semester, q).order_by("course_code")
    course_ids = list(courses.values_list("id", flat=True))

    enroll_qs = CourseEnrollment.objects.filter(course_id__in=course_ids, enroll=True, academic_year=academic_year)
    if mode == "faculty" and faculty:
        enroll_qs = enroll_qs.filter(faculty=faculty)

    all_groups = list(
        enroll_qs.values("course_id", "batch", "section", "year", "semester")
        .annotate(total_students=Count("id"))
        .order_by("course_id", "batch", "section")
    )
    if section:
        all_groups = [g for g in all_groups if (g.get("section") or "").upper() == section]

    bands, lag_cutoff_percentage = _al_lag_bands()

    rows = []
    if all_groups:
        courses_by_id = {
            c.id: c for c in
            Course.objects.filter(id__in={g["course_id"] for g in all_groups}).select_related("department", "elective")
        }
        group_q = _al_group_filter_q(all_groups)

        roster_qs = CourseEnrollment.objects.filter(group_q, enroll=True, academic_year=academic_year).select_related("student")
        if mode == "faculty" and faculty:
            roster_qs = roster_qs.filter(faculty=faculty)

        roster_by_group = defaultdict(list)
        for e in roster_qs:
            key = (e.course_id, e.batch, e.section, e.year, e.semester)
            roster_by_group[key].append(e.student)

        att_rows = HourAttendance.objects.filter(group_q, academic_year=academic_year).values(
            "course_id", "batch", "section", "year", "semester", "student_id", "status", "date", "period"
        )
        dates_by_group = defaultdict(set)
        counts_by_group_student = defaultdict(lambda: {"present": 0, "absent": 0, "od": 0})
        for r in att_rows:
            key = (r["course_id"], r["batch"], r["section"], r["year"], r["semester"])
            dates_by_group[key].add((r["date"], r["period"]))
            bucket = counts_by_group_student[(key, r["student_id"])]
            if r["status"] == "Present":
                bucket["present"] += 1
            elif r["status"] == "Absent":
                bucket["absent"] += 1
            elif r["status"] == "On Duty":
                bucket["od"] += 1

        for g in all_groups:
            key = (g["course_id"], g["batch"], g["section"], g["year"], g["semester"])
            total_hours = len(dates_by_group.get(key, ()))
            roster = roster_by_group.get(key, [])

            lag_count = 0
            if total_hours and roster:
                for st in roster:
                    counts = counts_by_group_student.get((key, st.id))
                    if not counts:
                        continue
                    attended = counts["present"] + counts["od"]
                    perc = round((attended / total_hours) * 100.0, 2)
                    if Decimal(str(perc)) <= lag_cutoff_percentage:
                        lag_count += 1

            rows.append({
                "course": courses_by_id.get(g["course_id"]),
                "batch": g["batch"],
                "section": g["section"],
                "year": g["year"],
                "semester": g["semester"],
                "total_students": g["total_students"],
                "total_hours": total_hours,
                "lag_count": lag_count,
            })

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="Attendance_Lag_Report.pdf"'

    doc = SimpleDocTemplate(
        response, pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=32 * mm, bottomMargin=18 * mm,
        title="Attendance Lag Report",
    )

    styles = getSampleStyleSheet()
    kpi_style = ParagraphStyle("alKpi", parent=styles["Normal"], fontSize=10, textColor=DARK_GRAY, spaceAfter=6)
    head_cell = ParagraphStyle("alHead", parent=styles["Normal"], fontSize=8.5, alignment=TA_CENTER,
                                textColor=colors.white, fontName="Helvetica-Bold")
    cell = ParagraphStyle("alCell", parent=styles["Normal"], fontSize=8.5, alignment=TA_CENTER, textColor=DARK_GRAY)
    cell_left = ParagraphStyle("alCellLeft", parent=cell, alignment=TA_LEFT)

    elements = []

    regulation_label = "All Regulations"
    if regulation_id:
        reg = Regulations.objects.filter(id=regulation_id).first()
        if reg:
            regulation_label = str(reg)

    degree_label = "All Degrees"
    if degree_id:
        deg = Degree.objects.filter(id=degree_id).first()
        if deg:
            degree_label = deg.degree

    department_label = "All Departments"
    if department_id:
        dept = Add_Department.objects.filter(id=department_id).first()
        if dept:
            department_label = dept.Department

    year_label = _pdf_text(year) or "All Years"
    semester_label = _pdf_text(semester) or "All Semesters"

    label_style = ParagraphStyle("alFilterLabel", parent=styles["Normal"], fontSize=9,
                                  textColor=MEDIUM_GRAY, fontName="Helvetica-Bold")
    value_style = ParagraphStyle("alFilterValue", parent=styles["Normal"], fontSize=9.5, textColor=DARK_GRAY)

    filters_data = [
        [Paragraph("Regulation", label_style), Paragraph(_pdf_text(regulation_label), value_style),
         Paragraph("Degree", label_style), Paragraph(_pdf_text(degree_label), value_style)],
        [Paragraph("Department", label_style), Paragraph(_pdf_text(department_label), value_style),
         Paragraph("Year", label_style), Paragraph(year_label, value_style)],
        [Paragraph("Semester", label_style), Paragraph(semester_label, value_style),
         Paragraph("Academic Year", label_style), Paragraph(_pdf_text(academic_year), value_style)],
        [Paragraph("Lag Cutoff", label_style), Paragraph("<= " + _pdf_text(lag_cutoff_percentage) + "%", value_style),
         Paragraph("Section", label_style), Paragraph(_pdf_text(section) or "All Sections", value_style)],
    ]
    filters_table = Table(filters_data, colWidths=[26 * mm, 65 * mm, 26 * mm, 65 * mm])
    filters_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
        ("BACKGROUND", (0, 0), (0, -1), BG_GRAY),
        ("BACKGROUND", (2, 0), (2, -1), BG_GRAY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(Spacer(1, 2 * mm))
    elements.append(filters_table)
    elements.append(Spacer(1, 3 * mm))

    elements.append(Paragraph("<b>Total Offerings:</b> " + str(len(rows)), kpi_style))
    elements.append(Spacer(1, 4 * mm))

    data = [[
        Paragraph("<b>S.No</b>", head_cell),
        Paragraph("<b>Course Code</b>", head_cell),
        Paragraph("<b>Course Name</b>", head_cell),
        Paragraph("<b>Batch</b>", head_cell),
        Paragraph("<b>Section</b>", head_cell),
        Paragraph("<b>Year</b>", head_cell),
        Paragraph("<b>Sem</b>", head_cell),
        Paragraph("<b>Enrolled</b>", head_cell),
        Paragraph("<b>Hours Held</b>", head_cell),
        Paragraph("<b>Lagging</b>", head_cell),
    ]]
    if rows:
        for idx, r in enumerate(rows, start=1):
            course = r["course"]
            data.append([
                Paragraph(str(idx), cell),
                Paragraph(_pdf_text(course.course_code) if course else "-", cell),
                Paragraph(_pdf_text(course.title) if course else "-", cell_left),
                Paragraph(_pdf_text(r.get("batch")) or "-", cell),
                Paragraph(_pdf_text(r.get("section")) or "-", cell),
                Paragraph(_pdf_text(r.get("year")) or "-", cell),
                Paragraph(_pdf_text(r.get("semester")) or "-", cell),
                Paragraph(str(r["total_students"]), cell),
                Paragraph(str(r["total_hours"]), cell),
                Paragraph(str(r["lag_count"]), cell),
            ])
    else:
        data.append([Paragraph("No course offerings found for the selected filters.", cell_left)] + [""] * 9)

    table = Table(data, colWidths=[9 * mm, 20 * mm, 42 * mm, 16 * mm, 15 * mm, 12 * mm, 11 * mm, 16 * mm, 18 * mm, 15 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)

    header_cb = _make_ramco_pdf_header("ATTENDANCE LAG REPORT")
    doc.build(elements, onFirstPage=header_cb, onLaterPages=header_cb)
    return response


@check_permission("subject_allocation_schedule")
def subject_allocation_schedule(request):
    edit_id = request.GET.get("edit_id")
    edit_schedule = None
    if edit_id and str(edit_id).isdigit():
        edit_schedule = SubjectAllocationSchedule.objects.filter(id=int(edit_id)).first()

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip().lower()

        if action == "delete":
            schedule_id = request.POST.get("schedule_id")
            if not schedule_id or not str(schedule_id).isdigit():
                messages.error(request, "Invalid schedule for delete.")
                return redirect("subject_allocation_schedule")
            SubjectAllocationSchedule.objects.filter(id=int(schedule_id)).delete()
            messages.success(request, "Schedule deleted successfully.")
            return redirect("subject_allocation_schedule")

        schedule_id = request.POST.get("schedule_id")
        academic_year = (request.POST.get("academic_year") or "").strip()
        start_date_raw = (request.POST.get("start_date") or "").strip()
        end_date_raw = (request.POST.get("end_date") or "").strip()
        is_active = bool(request.POST.get("is_active"))

        if not academic_year:
            messages.error(request, "Academic year is required.")
            return redirect("subject_allocation_schedule")

        try:
            start_date = date.fromisoformat(start_date_raw)
            end_date = date.fromisoformat(end_date_raw)
        except ValueError:
            messages.error(request, "Please provide valid start and end dates.")
            return redirect("subject_allocation_schedule")

        if start_date > end_date:
            messages.error(request, "Start date cannot be after end date.")
            return redirect("subject_allocation_schedule")

        if schedule_id and str(schedule_id).isdigit():
            schedule = SubjectAllocationSchedule.objects.filter(id=int(schedule_id)).first()
            if not schedule:
                messages.error(request, "Schedule not found for edit.")
                return redirect("subject_allocation_schedule")
            schedule.academic_year = academic_year
            schedule.start_date = start_date
            schedule.end_date = end_date
            schedule.is_active = is_active
            schedule.save(update_fields=["academic_year", "start_date", "end_date", "is_active", "updated_at"])
            messages.success(request, "Schedule updated successfully.")
        else:
            SubjectAllocationSchedule.objects.create(
                academic_year=academic_year,
                start_date=start_date,
                end_date=end_date,
                is_active=is_active,
            )
            messages.success(request, "Schedule added successfully.")
        return redirect("subject_allocation_schedule")

    schedules = SubjectAllocationSchedule.objects.order_by("-updated_at", "-id")
    if not edit_schedule:
        edit_schedule = schedules.first()

    schedule_state = get_subject_allocation_window_status()
    return render(
        request,
        "course_management/subject_allocation_schedule.html",
        {
            "schedule": edit_schedule,
            "schedule_state": schedule_state,
            "schedules": schedules,
        },
    )



# ==================================================================
# Lab Timetable / Lab Assignment / Lab Utilization
# ==================================================================
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.core.mail import send_mail
from django.conf import settings as _dj_settings
from faculty_management.models import general_information
from course_management.models import (
    Lab, LabTimetable, LabTimetableSlot, LabUtilityLog, Course, Regulations,
)

LAB_DAYS = [("MON", "Monday"), ("TUE", "Tuesday"), ("WED", "Wednesday"),
            ("THU", "Thursday"), ("FRI", "Friday")]
LAB_PERIODS = [1, 2, 3, 4, 5, 6, 7, 8]
LAB_TOTAL_WEEKLY_PERIODS = len(LAB_DAYS) * len(LAB_PERIODS)  # 5 days x 8 periods = 40


def _lab_scheduled_count(lab):
    """Distinct (day, period) slots that carry a course, across all of this lab's
    timetables. Using .distinct() on (day, period) means the count can never exceed
    LAB_TOTAL_WEEKLY_PERIODS even if a lab happens to have more than one timetable
    (e.g. different sections) — it reflects real weekly occupancy of the physical lab."""
    return (
        LabTimetableSlot.objects
        .filter(timetable__lab=lab, course__isnull=False)
        .values("day", "period").distinct().count()
    )


def _lab_utilization_summary(lab):
    """(scheduled_periods, total_periods, percentage) for one lab's weekly timetable,
    e.g. 30 scheduled out of 40 total periods = 75.0%."""
    scheduled = _lab_scheduled_count(lab)
    pct = round((scheduled / LAB_TOTAL_WEEKLY_PERIODS) * 100, 1) if LAB_TOTAL_WEEKLY_PERIODS else 0
    return scheduled, LAB_TOTAL_WEEKLY_PERIODS, pct


def _course_faculty_map(course_ids, department=None, section=None, regulation=None):
    """One-query lookup: {course_id: general_information} of the faculty assigned
    to teach each course, from course_management_assignsubjectfaculty
    (AssignSubjectFaculty) — the real mapping key is course_id -> faculty_id.

    department is intentionally NOT a hard filter: this app has duplicate
    Add_Department rows for the same department name (verified: e.g. "Computer
    Science and Engineering" exists as two separate rows), so requiring the
    assignment's department to be the exact same row as the lab timetable's
    department can silently miss a real, correct assignment. Instead, when a
    course has more than one active assignment row (different sections/years/
    departments), department/section/regulation/academic-year are used only as
    tie-break PREFERENCES to pick the most relevant one."""
    course_ids = [cid for cid in set(course_ids) if cid]
    if not course_ids:
        return {}
    department_id = getattr(department, "id", department)
    regulation_id = getattr(regulation, "id", regulation)
    current_year = getattr(settings, "ACADEMIC_YEAR", None)
    qs = (
        AssignSubjectFaculty.objects
        .filter(course_id__in=course_ids, is_active=True, faculty__isnull=False)
        .select_related("faculty")
    )
    best = {}
    for a in qs:
        score = 0
        if section and a.section == section:
            score += 8
        if department_id and a.department_id == department_id:
            score += 4
        if regulation_id and a.regulation_id == regulation_id:
            score += 2
        if current_year and a.academic_year == current_year:
            score += 1
        current = best.get(a.course_id)
        if not current or score > current[0]:
            best[a.course_id] = (score, a.faculty)
    return {cid: fac for cid, (score, fac) in best.items()}


def _pdf_safe(text):
    """Replace typographic punctuation that can render as a missing/garbled glyph
    with base-14 Helvetica (WinAnsi) fonts, e.g. en/em dashes, smart quotes."""
    if not text:
        return ""
    text = str(text)
    for bad, good in {
        "–": "-", "—": "-",
        "‘": "'", "’": "'",
        "“": '"', "”": '"',
        "…": "...", " ": " ",
    }.items():
        text = text.replace(bad, good)
    return text


def _pdf_text(text):
    """Sanitize + XML-escape a value for safe embedding inside a ReportLab
    Paragraph string that also contains raw markup like <b>/<br/>."""
    from xml.sax.saxutils import escape as _xml_escape
    return _xml_escape(_pdf_safe(text))


def _cm_lab_faculty(request):
    email = (getattr(request.user, "email", "") or "").strip()
    if not email:
        return None
    return (general_information.objects
            .filter(college_email__iexact=email)
            .select_related("department").first())


from user_accounts.decorators import custom_forbidden


def _can_manage_lab_tt(request):
    """The lab_timetable_create permission (or lab_timetable_edit) allows edit + delete."""
    if getattr(request.user, "is_superuser", False):
        return True
    perms = request.session.get("permissions", {}) or {}
    return bool(perms.get("lab_timetable_create") or perms.get("lab_timetable_edit"))


def _lab_semesters():
    return [str(i) for i in range(1, 9)]


def _lab_courses_for(department, semester):
    """Courses in this dept + semester that have a Practical (P / laboratory) value > 0
    in the per-course LTP table (CourseHours.laboratory_npwk). Only these belong in a
    LAB timetable. Each returned course gets a `.p_value` attribute for display."""
    result = []
    courses = (
        Course.objects
        .filter(department=department, semester=semester, is_active=True)
        .prefetch_related("semesters")
        .order_by("course_code")
    )
    for c in courses:
        p = 0
        for h in c.semesters.all():  # CourseHours rows (related_name="semesters")
            try:
                v = int(float(str(h.laboratory_npwk).strip()))
            except (TypeError, ValueError):
                v = 0
            if v > p:
                p = v
        if p > 0:
            c.p_value = p
            result.append(c)
    return result


def _send_timetable_email(tt):
    tech = getattr(tt.lab, "technician", None)
    email = getattr(tech, "college_email", None)
    if not email:
        return
    try:
        send_mail(
            subject="Lab Timetable saved: " + str(tt.lab.lab_code),
            message=("The lab timetable for " + str(tt.lab.lab_name) + " (" + str(tt.lab.lab_code) +
                     "), Semester " + str(tt.semester) + ", Section " + str(tt.section or "-") +
                     " has been saved/updated."),
            from_email=getattr(_dj_settings, "DEFAULT_FROM_EMAIL", None),
            recipient_list=[email],
            fail_silently=True,
        )
    except Exception:
        pass


# ------------------------------------------------------------------ VIEW
@course_management
@check_permission("lab_timetable_view")
def lab_timetable_view(request):
    departments = Add_Department.objects.filter(is_active=True)
    faculty = _cm_lab_faculty(request)
    is_super = bool(getattr(request.user, "is_superuser", False))
    user_dept = getattr(faculty, "department", None) if faculty else None
    lock_dept = bool(user_dept) and not is_super

    sel_dept = request.GET.get("dept") or ""
    if lock_dept:
        sel_dept = str(user_dept.id)  # mapped to the logged-in person's own department, enforced server-side
    sel_lab = request.GET.get("lab") or ""
    sel_sem = request.GET.get("semester") or ""
    sel_section = request.GET.get("section") or ""

    labs = Lab.objects.select_related("department").filter(department_id=sel_dept, is_active=True) if sel_dept else Lab.objects.none()

    timetable = None
    grid_rows = []
    lab_obj = None
    scheduled_periods = total_periods = weekly_utilization_pct = 0
    if sel_lab and sel_sem:
        lab_obj = Lab.objects.select_related("department", "technician").filter(id=sel_lab).first()
        # Guard: a locked (non-superuser) user cannot pull up a lab from another department.
        if lock_dept and lab_obj and lab_obj.department_id != user_dept.id:
            lab_obj = None
        if lab_obj:
            scheduled_periods, total_periods, weekly_utilization_pct = _lab_utilization_summary(lab_obj)
            tt_qs = LabTimetable.objects.filter(lab_id=sel_lab, semester=sel_sem)
            if sel_section:
                tt_qs = tt_qs.filter(section=sel_section)
            timetable = tt_qs.select_related("lab", "department", "regulation").first()
            if timetable:
                slots = list(timetable.slots.select_related("course").all())
                slot_map = {(s.day, s.period): s.course for s in slots}
                faculty_map = _course_faculty_map(
                    (s.course_id for s in slots if s.course_id),
                    timetable.department, timetable.section, timetable.regulation,
                )
                for code, label in LAB_DAYS:
                    cells = []
                    for p in LAB_PERIODS:
                        course = slot_map.get((code, p))
                        cells.append({
                            "period": p,
                            "course": course,
                            "faculty": faculty_map.get(course.id) if course else None,
                        })
                    grid_rows.append({"day": code, "label": label, "cells": cells})

    return render(request, "course_management/lab/lab_timetable_view.html", {
        "departments": departments, "labs": labs,
        "sel_dept": sel_dept, "sel_lab": sel_lab, "sel_sem": sel_sem, "sel_section": sel_section,
        "semesters": _lab_semesters(), "periods": LAB_PERIODS,
        "timetable": timetable, "grid_rows": grid_rows, "lab_obj": lab_obj,
        "user_dept": user_dept, "lock_dept": lock_dept,
        "scheduled_periods": scheduled_periods, "total_periods": total_periods,
        "weekly_utilization_pct": weekly_utilization_pct,
    })


def _grid_for_form(courses, existing_map):
    rows = []
    for code, label in LAB_DAYS:
        rows.append({
            "day": code, "label": label,
            "cells": [{"period": p, "course_id": existing_map.get((code, p), "")} for p in LAB_PERIODS],
        })
    return rows


def _save_slots(request, timetable):
    timetable.slots.all().delete()
    bulk = []
    for code, _ in LAB_DAYS:
        for p in LAB_PERIODS:
            cid = request.POST.get("slot_" + code + "_" + str(p)) or None
            if cid:
                bulk.append(LabTimetableSlot(timetable=timetable, day=code, period=p, course_id=cid))
    if bulk:
        LabTimetableSlot.objects.bulk_create(bulk)


# ------------------------------------------------------------------ CREATE
@course_management
@check_permission("lab_timetable_create")
def lab_timetable_create(request):
    faculty = _cm_lab_faculty(request)
    labs = Lab.objects.select_related("department").filter(is_active=True)
    regulations = Regulations.objects.all()

    if request.method == "POST":
        lab_id = request.POST.get("lab") or None
        semester = (request.POST.get("semester") or "").strip()
        section = (request.POST.get("section") or "").strip()
        regulation_id = request.POST.get("regulation") or None
        year = (request.POST.get("year") or "").strip()

        lab = Lab.objects.filter(id=lab_id).select_related("department").first()
        if not lab or not semester or not section:
            messages.error(request, "Lab, semester and section are required.")
            return redirect("lab_timetable_create")

        if LabTimetable.objects.filter(lab=lab, semester=semester, section=section).exists():
            messages.warning(request, "A timetable already exists for this lab/semester/section. Opening it for edit.")
            tt = LabTimetable.objects.get(lab=lab, semester=semester, section=section)
            return redirect("lab_timetable_edit", pk=tt.id)

        try:
            with transaction.atomic():
                tt = LabTimetable.objects.create(
                    lab=lab, department=lab.department, regulation_id=regulation_id or None,
                    year=year or None, semester=semester, section=section or None,
                    created_by=faculty,
                )
                _save_slots(request, tt)
            _send_timetable_email(tt)
            messages.success(request, "Lab timetable created successfully.")
            return redirect("/course_management/course_examination/lab_timetable_view/?lab=" + str(lab.id) +
                            "&semester=" + semester + "&dept=" + str(lab.department_id) + "&section=" + section)
        except Exception as e:
            messages.error(request, "Error: " + str(e))

    sel_lab = request.GET.get("lab") or ""
    sel_sem = request.GET.get("semester") or ""
    sel_section = request.GET.get("section") or ""
    show_grid = False
    courses = []
    grid_rows = []
    lab_obj = None
    if sel_lab and sel_sem and sel_section:
        lab_obj = Lab.objects.select_related("department").filter(id=sel_lab).first()
        if lab_obj:
            if LabTimetable.objects.filter(lab=lab_obj, semester=sel_sem, section=sel_section).exists():
                tt = LabTimetable.objects.get(lab=lab_obj, semester=sel_sem, section=sel_section)
                return redirect("lab_timetable_edit", pk=tt.id)
            courses = _lab_courses_for(lab_obj.department, sel_sem)
            grid_rows = _grid_for_form(courses, {})
            show_grid = True

    return render(request, "course_management/lab/lab_timetable_form.html", {
        "labs": labs, "regulations": regulations, "semesters": _lab_semesters(),
        "sel_lab": sel_lab, "sel_sem": sel_sem, "sel_section": sel_section,
        "show_grid": show_grid, "courses": courses, "grid_rows": grid_rows,
        "lab_obj": lab_obj, "is_edit": False, "timetable": None,
    })


# ------------------------------------------------------------------ EDIT
@course_management
def lab_timetable_edit(request, pk):
    if not _can_manage_lab_tt(request):
        return custom_forbidden(request)
    tt = get_object_or_404(
        LabTimetable.objects.select_related("lab", "lab__department", "department", "regulation"), pk=pk
    )
    if request.method == "POST":
        tt.regulation_id = request.POST.get("regulation") or tt.regulation_id
        tt.year = (request.POST.get("year") or tt.year)
        try:
            with transaction.atomic():
                tt.save()
                _save_slots(request, tt)
            _send_timetable_email(tt)
            messages.success(request, "Lab timetable updated successfully.")
            return redirect("/course_management/course_examination/lab_timetable_view/?lab=" + str(tt.lab_id) +
                            "&semester=" + str(tt.semester) + "&dept=" + str(tt.department_id) +
                            "&section=" + str(tt.section or ""))
        except Exception as e:
            messages.error(request, "Error: " + str(e))

    courses = _lab_courses_for(tt.lab.department, tt.semester)
    existing_map = {(s.day, s.period): s.course_id for s in tt.slots.all()}
    grid_rows = _grid_for_form(courses, existing_map)

    return render(request, "course_management/lab/lab_timetable_form.html", {
        "labs": Lab.objects.filter(is_active=True), "regulations": Regulations.objects.all(),
        "semesters": _lab_semesters(), "sel_lab": str(tt.lab_id), "sel_sem": tt.semester,
        "sel_section": tt.section or "", "show_grid": True, "courses": courses,
        "grid_rows": grid_rows, "lab_obj": tt.lab, "is_edit": True, "timetable": tt,
    })


# ------------------------------------------------------------------ DELETE timetable
@course_management
def lab_timetable_delete(request, pk):
    if not _can_manage_lab_tt(request):
        return custom_forbidden(request)
    tt = get_object_or_404(LabTimetable, pk=pk)
    lab_id, dept_id, sem, section = tt.lab_id, tt.department_id, tt.semester, (tt.section or "")
    if request.method == "POST":
        tt.delete()  # slots cascade
        messages.success(request, "Lab timetable deleted successfully.")
        return redirect("/course_management/course_examination/lab_timetable_view/?lab=" + str(lab_id) +
                        "&semester=" + str(sem) + "&dept=" + str(dept_id) + "&section=" + section)
    return redirect("lab_timetable_view")


# ------------------------------------------------------------------ ASSIGN (create lab + technician)
@course_management
@check_permission("lab_timetable_assign")
def lab_timetable_assign(request):
    departments = Add_Department.objects.filter(is_active=True)
    faculty = _cm_lab_faculty(request)
    is_super = bool(getattr(request.user, "is_superuser", False))
    user_dept = getattr(faculty, "department", None) if faculty else None
    lock_dept = bool(user_dept) and not is_super

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add_lab":
            department_id = request.POST.get("department") or None
            if lock_dept:
                department_id = user_dept.id  # enforce the user's own department
            lab_name = (request.POST.get("lab_name") or "").strip()
            lab_code = (request.POST.get("lab_code") or "").strip()
            technician_id = request.POST.get("technician") or None
            if not (department_id and lab_name and lab_code):
                messages.error(request, "Department, lab name and lab code are required.")
            elif Lab.objects.filter(lab_code__iexact=lab_code).exists():
                messages.error(request, "A lab with this code already exists.")
            else:
                Lab.objects.create(department_id=department_id, lab_name=lab_name,
                                   lab_code=lab_code, technician_id=technician_id or None)
                messages.success(request, "Lab created successfully.")
        elif action == "assign_tech":
            lab_id = request.POST.get("lab_id") or None
            technician_id = request.POST.get("technician") or None
            lab = Lab.objects.filter(id=lab_id).first()
            if lab and (not lock_dept or lab.department_id == user_dept.id):
                lab.technician_id = technician_id or None
                lab.save(update_fields=["technician"])
                messages.success(request, "Technician assigned.")
        return redirect("lab_timetable_assign")

    labs = Lab.objects.select_related("department", "technician").order_by("lab_code")
    if lock_dept:
        labs = labs.filter(department=user_dept)
    return render(request, "course_management/lab/lab_assign.html", {
        "departments": departments, "labs": labs,
        "user_dept": user_dept, "lock_dept": lock_dept,
    })


# ------------------------------------------------------------------ UTILITY LOG
@course_management
@check_permission("lab_utility_log")
def lab_utility_log(request):
    faculty = _cm_lab_faculty(request)
    is_super = bool(getattr(request.user, "is_superuser", False))

    # Viewing this page's details is open to everyone with the permission, regardless
    # of department or whether they are the lab's assigned technician. Only MARKING a
    # day (POST, below) stays restricted to the assigned technician / superuser.
    all_labs = (
        Lab.objects.select_related("department", "technician")
        .filter(is_active=True).order_by("lab_code")
    )

    labs_overview = []
    for lab in all_labs:
        scheduled, total, pct = _lab_utilization_summary(lab)
        labs_overview.append({"lab": lab, "scheduled": scheduled, "total": total, "pct": pct})

    if request.method == "POST":
        lab_id = request.POST.get("lab") or None
        log_date = request.POST.get("log_date") or None
        is_utilized = request.POST.get("is_utilized") == "yes"
        remarks = (request.POST.get("remarks") or "").strip()
        lab = Lab.objects.filter(id=lab_id).first()

        if not lab or not log_date:
            messages.error(request, "Lab and date are required.")
        elif not is_super and (not faculty or lab.technician_id != faculty.id):
            messages.error(request, "Only the lab's assigned technician can mark utilization for it.")
        else:
            LabUtilityLog.objects.update_or_create(
                lab=lab, log_date=log_date,
                defaults={"is_utilized": is_utilized, "remarks": remarks or None, "marked_by": faculty},
            )
            messages.success(request, "Utilization marked.")
        return redirect(request.path + "?lab=" + str(lab_id or ""))

    sel_lab = request.GET.get("lab") or ""
    lab_obj = (
        Lab.objects.select_related("department", "technician").filter(id=sel_lab).first()
        if sel_lab else None
    )
    logs = LabUtilityLog.objects.filter(lab=lab_obj).select_related("marked_by").order_by("-log_date")[:60] if lab_obj else []
    total_days = LabUtilityLog.objects.filter(lab=lab_obj).count() if lab_obj else 0
    utilized_days = LabUtilityLog.objects.filter(lab=lab_obj, is_utilized=True).count() if lab_obj else 0
    daily_pct = round((utilized_days / total_days) * 100, 1) if total_days else 0

    # Weekly timetable occupancy: e.g. 30 scheduled out of 40 total periods = 75%.
    scheduled_periods = total_periods = weekly_utilization_pct = 0
    lab_hours_by_course = []
    if lab_obj:
        scheduled_periods, total_periods, weekly_utilization_pct = _lab_utilization_summary(lab_obj)
        slots = (LabTimetableSlot.objects
                 .filter(timetable__lab=lab_obj, course__isnull=False)
                 .select_related("course"))
        agg = {}
        for s in slots:
            row = agg.setdefault(s.course_id, {"code": s.course.course_code, "title": s.course.title, "hours": 0})
            row["hours"] += 1
        lab_hours_by_course = sorted(agg.values(), key=lambda x: x["code"] or "")

    can_mark = bool(is_super or (faculty and lab_obj and lab_obj.technician_id == faculty.id))

    return render(request, "course_management/lab/lab_utility_log.html", {
        "labs": all_labs, "labs_overview": labs_overview,
        "sel_lab": sel_lab, "lab_obj": lab_obj, "logs": logs,
        "total_days": total_days, "utilized_days": utilized_days, "utilization_pct": daily_pct,
        "scheduled_periods": scheduled_periods, "total_periods": total_periods,
        "weekly_utilization_pct": weekly_utilization_pct,
        "lab_hours_by_course": lab_hours_by_course, "can_mark": can_mark,
        "today": timezone.localdate(),
    })


# ------------------------------------------------------------------ AJAX
def lab_get_technicians(request):
    dept = request.GET.get("dept")
    qs = general_information.objects.select_related("designation").filter(department_id=dept).order_by("name") if dept else general_information.objects.none()
    data = [{"id": f.id, "name": f.name or "",
             "designation": getattr(f.designation, "designation_name", "") or ""} for f in qs]
    return JsonResponse({"technicians": data})


def lab_get_labs(request):
    dept = request.GET.get("dept")
    qs = Lab.objects.filter(department_id=dept, is_active=True).order_by("lab_code") if dept else Lab.objects.none()
    data = [{"id": l.id, "lab_code": l.lab_code, "lab_name": l.lab_name} for l in qs]
    return JsonResponse({"labs": data})


def lab_get_sections(request):
    """Sections actually in use for a department (from StudentDetails.section) —
    a department may have just one section, or two, or three; never a fixed global list."""
    dept = request.GET.get("dept")
    if not dept:
        return JsonResponse({"sections": []})
    from user_accounts.models import StudentDetails
    sections = (
        StudentDetails.objects.filter(department_id=dept)
        .exclude(section__isnull=True).exclude(section="")
        .values_list("section", flat=True).distinct().order_by("section")
    )
    return JsonResponse({"sections": list(sections)})


# ==================================================================
# PDF exports — Lab Timetable (landscape grid) & Lab Utilization Report
# House style mirrors the existing Ramco Institute of Technology letterhead
# used elsewhere in this file (see _professional_header_footer above), but
# uses doc_.pagesize (not a hardcoded portrait A4) so it positions correctly
# on a landscape page.
# ==================================================================
def _make_ramco_pdf_header(title, meta_text_fn=None):
    """Returns an onPage callback that draws the standard Ramco Institute of
    Technology letterhead + footer for a landscape PDF report."""
    def _on_page(canvas, doc_):
        canvas.saveState()
        w, h = doc_.pagesize
        left = doc_.leftMargin
        right = w - doc_.rightMargin
        center_x = (left + right) / 2
        top_y = h - 8 * mm

        logo_rel = "images/ritlogo.png"
        logo_path = finders.find(logo_rel)
        if not logo_path:
            static_root = getattr(settings, "STATIC_ROOT", "")
            if static_root:
                cand = os.path.join(static_root, logo_rel)
                if os.path.exists(cand):
                    logo_path = cand
        if logo_path and os.path.exists(logo_path):
            try:
                canvas.drawImage(
                    ImageReader(logo_path), left, top_y - 16 * mm,
                    width=26 * mm, height=16 * mm,
                    preserveAspectRatio=True, mask="auto"
                )
            except Exception:
                pass

        canvas.setFillColor(PRIMARY_BLUE)
        canvas.setFont("Helvetica-Bold", 15)
        canvas.drawCentredString(center_x, top_y - 5 * mm, "RAMCO INSTITUTE OF TECHNOLOGY")

        canvas.setFillColor(ACCENT_RED)
        canvas.setFont("Helvetica-Bold", 9)
        canvas.drawCentredString(center_x, top_y - 10.5 * mm, "An Autonomous Institution")

        canvas.setFillColor(MEDIUM_GRAY)
        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(
            center_x, top_y - 15 * mm,
            "Approved by AICTE, New Delhi  |  Accredited by NAAC & ISO 9001:2015 Certified Institution"
        )

        canvas.setFillColor(DARK_GRAY)
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(center_x, top_y - 21 * mm, title)

        canvas.setStrokeColor(BORDER_GRAY)
        canvas.setLineWidth(0.8)
        canvas.line(left, top_y - 24 * mm, right, top_y - 24 * mm)

        footer_y = 10 * mm
        canvas.setStrokeColor(BORDER_GRAY)
        canvas.line(left, footer_y + 5 * mm, right, footer_y + 5 * mm)
        canvas.setFillColor(LIGHT_GRAY)
        canvas.setFont("Helvetica", 7.5)
        gen_time = datetime.now().strftime("%d %b %Y, %I:%M %p")
        canvas.drawString(left, footer_y, "Generated: " + gen_time)
        if meta_text_fn:
            canvas.drawCentredString(center_x, footer_y, meta_text_fn())
        canvas.drawRightString(right, footer_y, "Page " + str(doc_.page))
        canvas.restoreState()
    return _on_page


def _lab_timetable_grid_elements(tt):
    """Shared builder: [info_table, Spacer, grid_table] flowables for a single
    LabTimetable — the lab/dept/semester/technician info bar + weekly utilization,
    then the 5-day x 8-period grid with course code, full title and assigned faculty
    per cell. Used by both lab_timetable_pdf and the Principal/HOD dashboard PDFs."""
    slots = list(tt.slots.select_related("course").all())
    slot_map = {(s.day, s.period): s.course for s in slots}
    faculty_map = _course_faculty_map(
        (s.course_id for s in slots if s.course_id), tt.department, tt.section, tt.regulation,
    )
    scheduled, total, pct = _lab_utilization_summary(tt.lab)

    styles = getSampleStyleSheet()
    info_style = ParagraphStyle("ttInfo", parent=styles["Normal"], fontSize=9, textColor=DARK_GRAY)
    head_cell_style = ParagraphStyle("ttHeadCell", parent=styles["Normal"], fontSize=8.5,
                                      alignment=TA_CENTER, textColor=colors.white, leading=10)
    cell_style = ParagraphStyle("ttCell", parent=styles["Normal"], fontSize=8,
                                 alignment=TA_CENTER, textColor=DARK_GRAY, leading=10)
    free_style = ParagraphStyle("ttFree", parent=cell_style, textColor=LIGHT_GRAY)

    dept_name = tt.lab.department.Department if tt.lab.department else "-"
    tech_name = tt.lab.technician.name if tt.lab.technician else "Not assigned"
    section_txt = (" / Section " + _pdf_text(tt.section)) if tt.section else ""

    info_data = [
        [
            Paragraph("<b>Lab:</b> " + _pdf_text(tt.lab.lab_code) + " - " + _pdf_text(tt.lab.lab_name), info_style),
            Paragraph("<b>Department:</b> " + _pdf_text(dept_name), info_style),
            Paragraph("<b>Semester:</b> " + str(tt.semester) + section_txt, info_style),
            Paragraph("<b>Technician:</b> " + _pdf_text(tech_name), info_style),
        ],
        [
            Paragraph(
                "<b>Weekly Utilization:</b> " + str(scheduled) + "/" + str(total) +
                " periods scheduled (" + str(pct) + "%)",
                info_style
            ),
            "", "", "",
        ],
    ]
    info_table = Table(info_data, colWidths=[68 * mm, 78 * mm, 55 * mm, 62 * mm])
    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BG_GRAY),
        ("BOX", (0, 0), (-1, -1), 0.6, BORDER_GRAY),
        ("SPAN", (0, 1), (-1, 1)),
        ("LINEABOVE", (0, 1), (-1, 1), 0.4, BORDER_GRAY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))

    header_row = [Paragraph("<b>Day / Period</b>", head_cell_style)]
    for p in LAB_PERIODS:
        header_row.append(Paragraph("<b>Period " + str(p) + "</b>", head_cell_style))
    data = [header_row]

    for code, label in LAB_DAYS:
        row = [Paragraph("<b>" + label + "</b>", cell_style)]
        for p in LAB_PERIODS:
            course = slot_map.get((code, p))
            if course:
                cell_faculty = faculty_map.get(course.id)
                faculty_line = _pdf_text(cell_faculty.name) if cell_faculty else "No faculty assigned"
                row.append(Paragraph(
                    "<b>" + _pdf_text(course.course_code) + "</b><br/>"
                    "<font size=6.5>" + _pdf_text(course.title) + "</font><br/>"
                    "<font size=8 color='#1a4b8c'><b><i>" + faculty_line + "</i></b></font>",
                    cell_style
                ))
            else:
                row.append(Paragraph("&mdash;", free_style))
        data.append(row)

    col_widths = [30 * mm] + [30.8 * mm for _ in LAB_PERIODS]
    grid_table = Table(data, colWidths=col_widths, repeatRows=1)
    grid_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
        ("BACKGROUND", (0, 1), (0, -1), BG_GRAY),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (1, 1), (-1, -1), [colors.white, BG_GRAY]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    return [info_table, Spacer(1, 6 * mm), grid_table]


@course_management
@check_permission("lab_timetable_view")
def lab_timetable_pdf(request, pk):
    """Landscape PDF of a single lab timetable's weekly 5-day x 8-period grid."""
    from reportlab.lib.pagesizes import landscape
    from reportlab.platypus import SimpleDocTemplate

    tt = get_object_or_404(
        LabTimetable.objects.select_related("lab", "lab__department", "lab__technician", "regulation"),
        pk=pk
    )

    page_size = landscape(A4)
    response = HttpResponse(content_type="application/pdf")
    safe_code = (tt.lab.lab_code or "lab").replace(" ", "_")
    filename = "Lab_Timetable_" + safe_code + "_Sem" + str(tt.semester) + ".pdf"
    response["Content-Disposition"] = 'attachment; filename="' + filename + '"'

    doc = SimpleDocTemplate(
        response, pagesize=page_size,
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=32 * mm, bottomMargin=18 * mm,
    )

    elements = _lab_timetable_grid_elements(tt)

    header_cb = _make_ramco_pdf_header(
        "LAB TIME TABLE",
        lambda: (tt.lab.lab_code or "") + " - " + (tt.lab.lab_name or "")
    )
    doc.build(elements, onFirstPage=header_cb, onLaterPages=header_cb)
    return response


@course_management
@check_permission("lab_utility_log")
def lab_utility_log_pdf(request):
    """Landscape PDF report: either the all-labs weekly utilization overview
    (no `lab` param), or a single lab's detailed utilization report (with `lab`)."""
    from reportlab.lib.pagesizes import landscape
    from reportlab.platypus import SimpleDocTemplate

    sel_lab = request.GET.get("lab") or ""
    lab_obj = (
        Lab.objects.select_related("department", "technician").filter(id=sel_lab).first()
        if sel_lab else None
    )

    page_size = landscape(A4)
    response = HttpResponse(content_type="application/pdf")
    if lab_obj:
        safe_code = (lab_obj.lab_code or "lab").replace(" ", "_")
        filename = "Lab_Utilization_" + safe_code + ".pdf"
    else:
        filename = "Lab_Utilization_Overview.pdf"
    response["Content-Disposition"] = 'attachment; filename="' + filename + '"'

    doc = SimpleDocTemplate(
        response, pagesize=page_size,
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=32 * mm, bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()
    section_title_style = ParagraphStyle("secTitle", parent=styles["Normal"], fontSize=11,
                                          textColor=PRIMARY_BLUE, spaceAfter=4, fontName="Helvetica-Bold")
    info_style = ParagraphStyle("infoP", parent=styles["Normal"], fontSize=9, textColor=DARK_GRAY)
    head_cell = ParagraphStyle("headCell", parent=styles["Normal"], fontSize=8.5,
                                alignment=TA_CENTER, textColor=colors.white, fontName="Helvetica-Bold")
    cell = ParagraphStyle("bodyCell", parent=styles["Normal"], fontSize=8.5,
                           alignment=TA_CENTER, textColor=DARK_GRAY)
    cell_left = ParagraphStyle("bodyCellLeft", parent=cell, alignment=TA_LEFT)

    elements = []

    if lab_obj:
        scheduled, total, pct = _lab_utilization_summary(lab_obj)
        dept_name = lab_obj.department.Department if lab_obj.department else "-"
        tech_name = lab_obj.technician.name if lab_obj.technician else "Not assigned"

        info_data = [[
            Paragraph("<b>Lab:</b> " + _pdf_text(lab_obj.lab_code) + " - " + _pdf_text(lab_obj.lab_name), info_style),
            Paragraph("<b>Department:</b> " + _pdf_text(dept_name), info_style),
            Paragraph("<b>Technician:</b> " + _pdf_text(tech_name), info_style),
            Paragraph("<b>Weekly Utilization:</b> " + str(scheduled) + "/" + str(total) +
                      " periods (" + str(pct) + "%)", info_style),
        ]]
        info_table = Table(info_data, colWidths=[65 * mm, 68 * mm, 60 * mm, 70 * mm])
        info_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), BG_GRAY),
            ("BOX", (0, 0), (-1, -1), 0.6, BORDER_GRAY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 6 * mm))

        # Course-wise weekly lab hours
        slots = (LabTimetableSlot.objects
                 .filter(timetable__lab=lab_obj, course__isnull=False)
                 .select_related("course"))
        agg = {}
        for s in slots:
            row = agg.setdefault(s.course_id, {"code": s.course.course_code, "title": s.course.title, "hours": 0})
            row["hours"] += 1
        course_rows = sorted(agg.values(), key=lambda x: x["code"] or "")

        elements.append(Paragraph("Course-wise Weekly Lab Hours", section_title_style))
        cdata = [[Paragraph("<b>Course Code</b>", head_cell), Paragraph("<b>Title</b>", head_cell),
                  Paragraph("<b>Hours / Week</b>", head_cell)]]
        if course_rows:
            for r in course_rows:
                cdata.append([
                    Paragraph(_pdf_text(r["code"]) or "-", cell),
                    Paragraph(_pdf_text(r["title"]) or "-", cell_left),
                    Paragraph(str(r["hours"]), cell),
                ])
        else:
            cdata.append([Paragraph("No lab hours scheduled in the timetable yet.", cell_left), "", ""])
        ct = Table(cdata, colWidths=[45 * mm, 150 * mm, 40 * mm], repeatRows=1)
        ct.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(ct)
        elements.append(Spacer(1, 6 * mm))

        # Daily utilization log
        logs = (LabUtilityLog.objects.filter(lab=lab_obj)
                .select_related("marked_by").order_by("-log_date")[:90])
        elements.append(Paragraph("Daily Utilization Log", section_title_style))
        ldata = [[Paragraph("<b>Date</b>", head_cell), Paragraph("<b>Status</b>", head_cell),
                  Paragraph("<b>Remarks</b>", head_cell), Paragraph("<b>Marked By</b>", head_cell)]]
        if logs:
            for lg in logs:
                status = "Utilized" if lg.is_utilized else "Not Utilized"
                ldata.append([
                    Paragraph(lg.log_date.strftime("%d %b %Y"), cell),
                    Paragraph(status, cell),
                    Paragraph(_pdf_text(lg.remarks) or "-", cell_left),
                    Paragraph(_pdf_text(lg.marked_by.name) if lg.marked_by else "-", cell),
                ])
        else:
            ldata.append([Paragraph("No log entries yet.", cell_left), "", "", ""])
        lt = Table(ldata, colWidths=[35 * mm, 35 * mm, 125 * mm, 40 * mm], repeatRows=1)
        lt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(lt)

        header_cb = _make_ramco_pdf_header(
            "LAB UTILIZATION REPORT",
            lambda: (lab_obj.lab_code or "") + " - " + (lab_obj.lab_name or "")
        )
    else:
        all_labs = Lab.objects.select_related("department").filter(is_active=True).order_by("lab_code")
        elements.append(Paragraph("All Labs - Weekly Utilization Overview", section_title_style))
        odata = [[
            Paragraph("<b>Lab Code</b>", head_cell), Paragraph("<b>Lab Name</b>", head_cell),
            Paragraph("<b>Department</b>", head_cell), Paragraph("<b>Scheduled</b>", head_cell),
            Paragraph("<b>Utilization</b>", head_cell),
        ]]
        if all_labs:
            for lab in all_labs:
                scheduled, total, pct = _lab_utilization_summary(lab)
                odata.append([
                    Paragraph(_pdf_text(lab.lab_code) or "-", cell),
                    Paragraph(_pdf_text(lab.lab_name) or "-", cell_left),
                    Paragraph(_pdf_text(lab.department.Department) if lab.department else "-", cell_left),
                    Paragraph(str(scheduled) + "/" + str(total), cell),
                    Paragraph(str(pct) + "%", cell),
                ])
        else:
            odata.append([Paragraph("No labs found.", cell_left), "", "", "", ""])
        ot = Table(odata, colWidths=[35 * mm, 75 * mm, 90 * mm, 35 * mm, 40 * mm], repeatRows=1)
        ot.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(ot)

        header_cb = _make_ramco_pdf_header("LAB UTILIZATION REPORT — ALL LABS")

    doc.build(elements, onFirstPage=header_cb, onLaterPages=header_cb)
    return response


# ==================================================================
# Lab Utilization Dashboards — Principal (all departments, filterable) and
# HOD (locked to the viewer's own department). No role name is hard-coded
# anywhere below: access is controlled purely by whichever permission
# ("principal_lab_dashboard" / "hod_lab_dashboard") a role has been granted
# via the normal permission modal, exactly like every other page in this app.
# The HOD page is "locked" simply because its view always passes the VIEWER'S
# OWN department as `locked_department` — the same pattern already used by
# lab_timetable_view / lab_timetable_assign.
# ==================================================================
def _lab_dashboard_data(request, locked_department=None):
    """Shared context for both dashboards. If locked_department is given, the
    department filter is force-set to it (no switching); otherwise the caller
    (Principal) can filter across every department via ?dept=."""
    departments = Add_Department.objects.filter(is_active=True).order_by("Department")

    sel_dept = str(locked_department.id) if locked_department else (request.GET.get("dept") or "")
    sel_regulation = request.GET.get("regulation") or ""
    sel_year = request.GET.get("academic_year") or ""
    sel_sem = request.GET.get("semester") or ""

    # Department-wise lab counts (tiles). When locked to one department, only that
    # department's tile is produced — an HOD dashboard should not expose other
    # departments' lab counts.
    dept_scope = [locked_department] if locked_department else list(departments)
    dept_lab_counts = [
        {"department": d, "count": Lab.objects.filter(department=d, is_active=True).count()}
        for d in dept_scope
    ]

    labs_qs = Lab.objects.select_related("department", "technician").filter(is_active=True)
    if sel_dept:
        labs_qs = labs_qs.filter(department_id=sel_dept)

    if sel_regulation:
        labs_qs = labs_qs.filter(timetables__regulation_id=sel_regulation)
    if sel_year:
        labs_qs = labs_qs.filter(timetables__year=sel_year)
    if sel_sem:
        labs_qs = labs_qs.filter(timetables__semester=sel_sem)
    if sel_regulation or sel_year or sel_sem:
        labs_qs = labs_qs.distinct()
    labs_qs = labs_qs.order_by("lab_code")

    rows = []
    total_scheduled = 0
    total_capacity = 0
    for lab in labs_qs:
        scheduled, total, pct = _lab_utilization_summary(lab)
        total_scheduled += scheduled
        total_capacity += total
        rows.append({"lab": lab, "scheduled": scheduled, "total": total, "pct": pct})

    overall_pct = round((total_scheduled / total_capacity) * 100, 1) if total_capacity else 0

    regulations = Regulations.objects.all().order_by("-year")
    academic_years = (
        LabTimetable.objects.exclude(year__isnull=True).exclude(year="")
        .values_list("year", flat=True).distinct().order_by("-year")
    )

    return {
        "departments": departments,
        "dept_lab_counts": dept_lab_counts,
        "sel_dept": sel_dept,
        "sel_regulation": sel_regulation,
        "sel_year": sel_year,
        "sel_sem": sel_sem,
        "regulations": regulations,
        "academic_years": academic_years,
        "semesters": _lab_semesters(),
        "rows": rows,
        "total_labs": labs_qs.count(),
        "overall_pct": overall_pct,
        "total_scheduled": total_scheduled,
        "total_capacity": total_capacity,
        "locked_department": locked_department,
    }


def _lab_dashboard_detail(lab, sel_regulation, sel_year, sel_sem):
    """Full drill-down for one lab: its most relevant timetable (matching the
    dashboard's active filters, falling back to its latest timetable), the
    weekly grid with course + assigned faculty per cell, weekly utilization,
    and the course-wise weekly lab-hours breakdown."""
    tt_qs = LabTimetable.objects.filter(lab=lab).select_related("department", "regulation")
    if sel_regulation:
        tt_qs = tt_qs.filter(regulation_id=sel_regulation)
    if sel_year:
        tt_qs = tt_qs.filter(year=sel_year)
    if sel_sem:
        tt_qs = tt_qs.filter(semester=sel_sem)
    timetable = tt_qs.order_by("-id").first()

    grid_rows = []
    if timetable:
        slots = list(timetable.slots.select_related("course").all())
        slot_map = {(s.day, s.period): s.course for s in slots}
        faculty_map = _course_faculty_map(
            (s.course_id for s in slots if s.course_id),
            timetable.department, timetable.section, timetable.regulation,
        )
        for code, label in LAB_DAYS:
            cells = []
            for p in LAB_PERIODS:
                course = slot_map.get((code, p))
                cells.append({
                    "period": p,
                    "course": course,
                    "faculty": faculty_map.get(course.id) if course else None,
                })
            grid_rows.append({"day": code, "label": label, "cells": cells})

    scheduled, total, pct = _lab_utilization_summary(lab)
    slots_all = (
        LabTimetableSlot.objects.filter(timetable__lab=lab, course__isnull=False)
        .select_related("course")
    )
    agg = {}
    for s in slots_all:
        row = agg.setdefault(s.course_id, {"code": s.course.course_code, "title": s.course.title, "hours": 0})
        row["hours"] += 1
    lab_hours_by_course = sorted(agg.values(), key=lambda x: x["code"] or "")

    return {
        "timetable": timetable,
        "grid_rows": grid_rows,
        "scheduled_periods": scheduled,
        "total_periods": total,
        "weekly_utilization_pct": pct,
        "lab_hours_by_course": lab_hours_by_course,
    }


def _dashboard_lab_obj(request, locked_department):
    sel_lab = request.GET.get("lab") or ""
    lab_obj = (
        Lab.objects.select_related("department", "technician").filter(id=sel_lab).first()
        if sel_lab else None
    )
    # A locked (HOD) viewer cannot drill into another department's lab.
    if locked_department and lab_obj and lab_obj.department_id != locked_department.id:
        lab_obj = None
    return sel_lab, lab_obj


def _hod_locked_department(request):
    """The viewer's own department, or None (unlocked) for a superuser / a role
    with no department profile — mirrors the lock_dept pattern already used by
    lab_timetable_view / lab_timetable_assign. No role name is ever checked."""
    faculty = _cm_lab_faculty(request)
    is_super = bool(getattr(request.user, "is_superuser", False))
    user_dept = getattr(faculty, "department", None) if faculty else None
    return user_dept if (user_dept and not is_super) else None


@course_management
@check_permission("principal_lab_dashboard")
def principal_lab_dashboard(request):
    ctx = _lab_dashboard_data(request, locked_department=None)
    sel_lab, lab_obj = _dashboard_lab_obj(request, locked_department=None)
    detail = _lab_dashboard_detail(lab_obj, ctx["sel_regulation"], ctx["sel_year"], ctx["sel_sem"]) if lab_obj else None
    ctx.update({"sel_lab": sel_lab, "lab_obj": lab_obj, "detail": detail})
    return render(request, "course_management/lab/principal_lab_dashboard.html", ctx)


@course_management
@check_permission("hod_lab_dashboard")
def hod_lab_dashboard(request):
    locked_department = _hod_locked_department(request)
    ctx = _lab_dashboard_data(request, locked_department=locked_department)
    sel_lab, lab_obj = _dashboard_lab_obj(request, locked_department)
    detail = _lab_dashboard_detail(lab_obj, ctx["sel_regulation"], ctx["sel_year"], ctx["sel_sem"]) if lab_obj else None
    ctx.update({"sel_lab": sel_lab, "lab_obj": lab_obj, "detail": detail, "user_dept": locked_department})
    return render(request, "course_management/lab/hod_lab_dashboard.html", ctx)


def _dashboard_pdf_response(request, locked_department, report_title, filename_prefix):
    """Shared landscape PDF for both dashboards: a single-lab full detail report
    when ?lab= is given, otherwise the department-wise + all-labs overview."""
    from reportlab.lib.pagesizes import landscape
    from reportlab.platypus import SimpleDocTemplate

    ctx = _lab_dashboard_data(request, locked_department=locked_department)
    sel_lab, lab_obj = _dashboard_lab_obj(request, locked_department)

    page_size = landscape(A4)
    response = HttpResponse(content_type="application/pdf")
    if lab_obj:
        filename = filename_prefix + "_" + (lab_obj.lab_code or "lab").replace(" ", "_") + ".pdf"
    else:
        filename = filename_prefix + "_Overview.pdf"
    response["Content-Disposition"] = 'attachment; filename="' + filename + '"'

    doc = SimpleDocTemplate(
        response, pagesize=page_size,
        leftMargin=10 * mm, rightMargin=10 * mm, topMargin=32 * mm, bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()
    section_title_style = ParagraphStyle("dashSecTitle", parent=styles["Normal"], fontSize=11,
                                          textColor=PRIMARY_BLUE, spaceAfter=4, fontName="Helvetica-Bold")
    kpi_style = ParagraphStyle("dashKpi", parent=styles["Normal"], fontSize=10, textColor=DARK_GRAY)
    head_cell = ParagraphStyle("dashHead", parent=styles["Normal"], fontSize=8.5,
                                alignment=TA_CENTER, textColor=colors.white, fontName="Helvetica-Bold")
    cell = ParagraphStyle("dashCell", parent=styles["Normal"], fontSize=8.5,
                           alignment=TA_CENTER, textColor=DARK_GRAY)
    cell_left = ParagraphStyle("dashCellLeft", parent=cell, alignment=TA_LEFT)

    elements = []

    if lab_obj:
        detail = _lab_dashboard_detail(lab_obj, ctx["sel_regulation"], ctx["sel_year"], ctx["sel_sem"])
        if detail["timetable"]:
            elements.extend(_lab_timetable_grid_elements(detail["timetable"]))
        else:
            elements.append(Paragraph(
                "<b>Lab:</b> " + _pdf_text(lab_obj.lab_code) + " - " + _pdf_text(lab_obj.lab_name) +
                "  |  <b>Weekly Utilization:</b> " + str(detail["scheduled_periods"]) + "/" +
                str(detail["total_periods"]) + " (" + str(detail["weekly_utilization_pct"]) + "%)",
                kpi_style
            ))
        elements.append(Spacer(1, 6 * mm))

        elements.append(Paragraph("Course-wise Weekly Lab Hours", section_title_style))
        cdata = [[Paragraph("<b>Course Code</b>", head_cell), Paragraph("<b>Title</b>", head_cell),
                  Paragraph("<b>Hours / Week</b>", head_cell)]]
        if detail["lab_hours_by_course"]:
            for r in detail["lab_hours_by_course"]:
                cdata.append([
                    Paragraph(_pdf_text(r["code"]) or "-", cell),
                    Paragraph(_pdf_text(r["title"]) or "-", cell_left),
                    Paragraph(str(r["hours"]), cell),
                ])
        else:
            cdata.append([Paragraph("No lab hours scheduled in the timetable yet.", cell_left), "", ""])
        ct = Table(cdata, colWidths=[45 * mm, 150 * mm, 40 * mm], repeatRows=1)
        ct.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(ct)

        header_cb = _make_ramco_pdf_header(
            report_title,
            lambda: (lab_obj.lab_code or "") + " - " + (lab_obj.lab_name or "")
        )
    else:
        kpi_line = (
            "<b>Total Labs:</b> " + str(ctx["total_labs"]) +
            "     <b>Overall Utilization:</b> " + str(ctx["total_scheduled"]) + "/" +
            str(ctx["total_capacity"]) + " (" + str(ctx["overall_pct"]) + "%)"
        )
        elements.append(Paragraph(kpi_line, kpi_style))
        elements.append(Spacer(1, 4 * mm))

        elements.append(Paragraph("Department-wise Lab Count", section_title_style))
        ddata = [[Paragraph("<b>Department</b>", head_cell), Paragraph("<b>Labs</b>", head_cell)]]
        for row in ctx["dept_lab_counts"]:
            ddata.append([
                Paragraph(_pdf_text(row["department"].Department), cell_left),
                Paragraph(str(row["count"]), cell),
            ])
        dt = Table(ddata, colWidths=[200 * mm, 35 * mm], repeatRows=1)
        dt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(dt)
        elements.append(Spacer(1, 6 * mm))

        elements.append(Paragraph("Labs - Weekly Utilization", section_title_style))
        odata = [[
            Paragraph("<b>Lab Code</b>", head_cell), Paragraph("<b>Lab Name</b>", head_cell),
            Paragraph("<b>Department</b>", head_cell), Paragraph("<b>Scheduled</b>", head_cell),
            Paragraph("<b>Utilization</b>", head_cell),
        ]]
        if ctx["rows"]:
            for r in ctx["rows"]:
                odata.append([
                    Paragraph(_pdf_text(r["lab"].lab_code) or "-", cell),
                    Paragraph(_pdf_text(r["lab"].lab_name) or "-", cell_left),
                    Paragraph(_pdf_text(r["lab"].department.Department) if r["lab"].department else "-", cell_left),
                    Paragraph(str(r["scheduled"]) + "/" + str(r["total"]), cell),
                    Paragraph(str(r["pct"]) + "%", cell),
                ])
        else:
            odata.append([Paragraph("No labs found.", cell_left), "", "", "", ""])
        ot = Table(odata, colWidths=[35 * mm, 70 * mm, 85 * mm, 35 * mm, 40 * mm], repeatRows=1)
        ot.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), SECONDARY_BLUE),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BG_GRAY]),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(ot)

        header_cb = _make_ramco_pdf_header(report_title)

    doc.build(elements, onFirstPage=header_cb, onLaterPages=header_cb)
    return response


@course_management
@check_permission("principal_lab_dashboard")
def principal_lab_dashboard_pdf(request):
    return _dashboard_pdf_response(request, None, "PRINCIPAL LAB UTILIZATION DASHBOARD", "Principal_Lab_Dashboard")


@course_management
@check_permission("hod_lab_dashboard")
def hod_lab_dashboard_pdf(request):
    locked_department = _hod_locked_department(request)
    return _dashboard_pdf_response(request, locked_department, "HOD LAB UTILIZATION DASHBOARD", "HOD_Lab_Dashboard")

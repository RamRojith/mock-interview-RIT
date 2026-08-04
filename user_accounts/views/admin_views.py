from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse

from django.contrib.auth import (
    authenticate, 
    login, 
    logout as auth_logout
    )
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils.timezone import now
from django.urls import get_resolver
from django.urls import URLPattern, URLResolver
from django.http import HttpResponse
from user_accounts.urls import admin_urls
from user_accounts.decorators import *
from user_accounts.models import *

def admin_view_names():
    
    resolver = get_resolver(admin_urls)
    view_names = []

    for pattern in resolver.url_patterns:
        if isinstance(pattern, URLPattern):
            view_names.append(pattern.name)              
    return view_names

def admin_login(request):
    # request.session.flush()
    if request.method == 'POST':
        Employee_id = request.POST.get('Employee_id')
        password = request.POST.get('password')
        
        # Authenticate user
        user = authenticate(request, Employee_id=Employee_id, password=password)

        if user is not None:
            # Check if the user is a superuser
            if user.is_superuser:
                
                login(request, user) 
                # permissions={}
                # role_permissions = RolePermissionFunction.objects.filter(role=user.role)
                # for permission in role_permissions:
                #     permissions[permission.function] = permission.permission  
                # request.session['permissions'] = permissions
                request.session['app_name'] ="Admin Portal" 
                request.session['pages'] = list(
                    sorted(
                        {word.replace('_', ' ').title(): word 
                        for word in set(
                            # [
                            # permission.function 
                            #             for permission in role_permissions 
                            #             if permission.permission] + 
                                        admin_view_names())
                        }.items(),
                        key=lambda item: len(item[0]), 
                        reverse=False
                    )
                )
              
                messages.success(request,"Authentication successful. You are now logged in.")
                return redirect('home')  
            else:
                return render(request, 'admin_authentication/login.html', {'error': 'Not an admin user'})
        else:
            return render(request, 'admin_authentication/login.html', {'error': 'Invalid credentials'})

    return render(request, 'admin_authentication/login.html')


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth import login as auth_login
from user_accounts.models import USER, Role, Department
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.db.models import Q, Count
from django.core.paginator import Paginator, EmptyPage
from django.contrib.auth.hashers import make_password
from django.views.decorators.http import require_POST


def staff_user_accounts(request):
    """
    Page render only.
    Users + stats are loaded via AJAX from staff_user_accounts_data().
    """
    departments = Department.objects.using("rit_approval_system").all().order_by("Department")

    # ✅ login-as (POST) stays here
    if request.method == "POST":
        user_id = request.POST.get("login_as")
        if user_id:
            user = get_object_or_404(USER.objects.using("rit_approval_system"), pk=user_id)
            user.backend = "user_accounts.auth.EmployeeIDBackend"
            auth_login(request, user)
            messages.success(request, f"You are now logged in as '{user.username}'.")
            return redirect("faculty_dashboard")

        return redirect("staff_user_accounts")

    return render(request, "admin/staff_user_accounts.html", {
        "departments": departments
    })
 

@login_required
@require_POST
def reset_staff_user_password(request, id):
    user = get_object_or_404(
        USER.objects.using("rit_approval_system").filter(is_student=False, is_parent=False),
        pk=id,
    )
    employee_id = user.Employee_id
    updated_count = USER.objects.using("rit_approval_system").filter(
        Employee_id=employee_id,
        is_student=False,
        is_parent=False,
    ).update(password=make_password("123"))

    return JsonResponse({
        "success": True,
        "message": (
            f"Password reset to 123 for {user.username} "
            f"({updated_count} account record(s) updated)."
        ),
    })


def staff_user_accounts_data(request):
    """
    AJAX endpoint:
    Params:
      q        : global search
      dept     : department name
      role     : role name
      page     : page number (1..)
      page_size: items per page

    Returns:
      results, pagination meta, stats, role_options
    """
    q = (request.GET.get("q") or "").strip()
    dept = (request.GET.get("dept") or "").strip()
    role = (request.GET.get("role") or "").strip()

    try:
        page = int(request.GET.get("page", 1))
    except Exception:
        page = 1

    try:
        page_size = int(request.GET.get("page_size", 25))
    except Exception:
        page_size = 25

    # ✅ base queryset (all staff/admin users)
    base_qs = (
        USER.objects.using("rit_approval_system")
        .filter(is_student=False, is_parent=False)
        .select_related("Department", "role")
    )

    # ✅ role dropdown options (from ALL staff users)
    role_options = list(
        base_qs.exclude(role__role__isnull=True)
               .exclude(role__role__exact="")
               .values_list("role__role", flat=True)
               .distinct()
               .order_by("role__role")
    )

    # ✅ apply filters on working qs
    qs = base_qs

    if dept:
        qs = qs.filter(Department__Department__iexact=dept)

    if role:
        qs = qs.filter(role__role__iexact=role)

    if q:
        qs = qs.filter(
            Q(username__icontains=q) |
            Q(Employee_id__icontains=q) |
            Q(email__icontains=q) |
            Q(Department__Department__icontains=q) |
            Q(Department__Department_code__icontains=q) |
            Q(role__role__icontains=q)
        )

    # ✅ ordering (active first)
    qs = qs.order_by("-is_active", "id")

    # ✅ stats (based on FILTERED qs)
    total_count = qs.count()
    active_count = qs.filter(is_active=True).count()
    inactive_count = qs.filter(is_active=False).count()

    role_counts_qs = (
        qs.values("role__role")
          .annotate(count=Count("id"))
          .order_by("-count", "role__role")
    )

    role_counts = []
    for rc in role_counts_qs:
        role_name = rc.get("role__role") or "No Role"
        role_counts.append({"role": role_name, "count": rc.get("count", 0)})

    # ✅ pagination
    paginator = Paginator(qs, page_size)
    try:
        page_obj = paginator.page(page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages if paginator.num_pages else 1)

    results = []
    start_index = (page_obj.number - 1) * page_size

    for idx, u in enumerate(page_obj.object_list, start=1):
        # profile image
        try:
            profile_url = u.profile_img.url if getattr(u, "profile_img", None) else ""
        except Exception:
            profile_url = ""

        # last login
        last_login_str = "-"
        if u.last_login:
            last_login_str = u.last_login.strftime("%Y-%m-%d %H:%M")

        dept_name = ""
        dept_code = ""
        if getattr(u, "Department", None):
            dept_name = getattr(u.Department, "Department", "") or ""
            dept_code = getattr(u.Department, "Department_code", "") or ""

        role_name = ""
        if getattr(u, "role", None):
            role_name = getattr(u.role, "role", "") or ""

        results.append({
            "sno": start_index + idx,
            "id": u.id,
            "username": u.username or "",
            "employee_id": u.Employee_id or "",
            "email": u.email or "",
            "department": dept_name,
            "department_code": dept_code,
            "role": role_name,
            "is_active": bool(u.is_active),
            "last_login": last_login_str,
            "profile": profile_url,
        })

    return JsonResponse({
        "results": results,
        "role_options": role_options,
        "stats": {
            "total": total_count,
            "active": active_count,
            "inactive": inactive_count,
            "role_counts": role_counts,  # list of {role,count}
        },
        "pagination": {
            "page": page_obj.number,
            "page_size": page_size,
            "total": total_count,
            "total_pages": paginator.num_pages,
            "has_next": page_obj.has_next(),
            "has_prev": page_obj.has_previous(),
        }
    })




def edit_staff_user_accounts(request, id):
    # Fetch the user from external DB
    user = get_object_or_404(USER.objects.using("rit_approval_system"), pk=id)

    if request.method == "POST":
        # Editable fields
        username = request.POST.get("username")
        email = request.POST.get("email")
        department_id = request.POST.get("department")
        role_id = request.POST.get("role")
        profile_img = request.FILES.get("profile_img")
        is_active = request.POST.get("is_active")  # checkbox or dropdown

        # Validation
        if not username or not email:
            messages.error(request, "Username and Email are required.")
            return redirect("edit_staff_user_accounts", id=id)

        # Update fields
        user.username = username
        user.email = email

        if department_id:
            user.Department_id = department_id
        if role_id:
            user.role_id = role_id
        if profile_img:
            user.profile_img = profile_img

        # Handle is_active checkbox
        user.is_active = True if is_active == "on" or is_active == "1" else False

        # Save to external DB
        user.save(using="rit_approval_system")

        messages.success(request, f"User '{username}' updated successfully.")
        return redirect("staff_user_accounts")

    # Send departments and roles for dropdown
    departments = Department.objects.using("rit_approval_system").all()
    roles = Role.objects.using("rit_approval_system").all()

    return render(request, "admin/edit_staff_user_accounts.html", {
        "user": user,
        "departments": departments,
        "roles": roles,
    })



def student_user_accounts(request):
    """
    Page render only.
    Student list & stats are loaded via AJAX from student_user_accounts_data().
    """
    departments = Department.objects.using("rit_approval_system").all().order_by("Department")

    if request.method == "POST":
        user_id = request.POST.get("login_as")
        if user_id:
            user = get_object_or_404(USER.objects.using("rit_approval_system"), pk=user_id)
            user.backend = "user_accounts.auth.EmployeeIDBackend"
            auth_login(request, user)
            messages.success(request, f"You are now logged in as '{user.username}'.")
            return redirect("student_dashboard")

        return redirect("student_user_accounts")

    return render(request, "admin/student_user_accounts.html", {
        "departments": departments
    })


@login_required
@require_POST
def reset_student_user_password(request, id):
    user = get_object_or_404(
        USER.objects.using("rit_approval_system").filter(is_student=True),
        pk=id,
    )
    student_id = user.Employee_id
    updated_count = USER.objects.using("rit_approval_system").filter(
        Employee_id=student_id,
        is_student=True,
    ).update(password=make_password("123"))

    return JsonResponse({
        "success": True,
        "message": (
            f"Password reset to 123 for {user.username} "
            f"({updated_count} account record(s) updated)."
        ),
    })


def student_user_accounts_data(request):
    """
    AJAX endpoint:
    Params:
      q        : global search (any field)
      dept     : department name filter
      role     : role filter
      page     : page number
      page_size: page size

    Returns:
      results, pagination meta, stats, role_options
    """
    q = (request.GET.get("q") or "").strip()
    dept = (request.GET.get("dept") or "").strip()
    role = (request.GET.get("role") or "").strip()

    try:
        page = int(request.GET.get("page", 1))
    except Exception:
        page = 1

    try:
        page_size = int(request.GET.get("page_size", 25))
    except Exception:
        page_size = 25

    base_qs = (
        USER.objects.using("rit_approval_system")
        .filter(is_student=True)
        .select_related("Department", "role")
    )

    # ✅ Role dropdown options (from ALL students)
    role_options = list(
        base_qs.exclude(role__role__isnull=True)
               .exclude(role__role__exact="")
               .values_list("role__role", flat=True)
               .distinct()
               .order_by("role__role")
    )

    qs = base_qs

    if dept:
        qs = qs.filter(Department__Department__iexact=dept)

    if role:
        qs = qs.filter(role__role__iexact=role)

    if q:
        qs = qs.filter(
            Q(username__icontains=q) |
            Q(Employee_id__icontains=q) |
            Q(email__icontains=q) |
            Q(Department__Department__icontains=q) |
            Q(Department__Department_code__icontains=q) |
            Q(role__role__icontains=q)
        )

    # ✅ Active first, then id
    qs = qs.order_by("-is_active", "id")

    # ✅ Stats (FILTERED)
    total_count = qs.count()
    active_count = qs.filter(is_active=True).count()
    inactive_count = qs.filter(is_active=False).count()

    role_counts_qs = (
        qs.values("role__role")
          .annotate(count=Count("id"))
          .order_by("-count", "role__role")
    )

    role_counts = []
    for rc in role_counts_qs:
        role_name = rc.get("role__role") or "No Role"
        role_counts.append({"role": role_name, "count": rc.get("count", 0)})

    # ✅ Pagination
    paginator = Paginator(qs, page_size)
    try:
        page_obj = paginator.page(page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages if paginator.num_pages else 1)

    results = []
    start_index = (page_obj.number - 1) * page_size

    for idx, u in enumerate(page_obj.object_list, start=1):
        try:
            profile_url = u.profile_img.url if getattr(u, "profile_img", None) else ""
        except Exception:
            profile_url = ""

        last_login_str = "-"
        if u.last_login:
            last_login_str = u.last_login.strftime("%Y-%m-%d %H:%M")

        dept_name = ""
        dept_code = ""
        if getattr(u, "Department", None):
            dept_name = getattr(u.Department, "Department", "") or ""
            dept_code = getattr(u.Department, "Department_code", "") or ""

        role_name = ""
        if getattr(u, "role", None):
            role_name = getattr(u.role, "role", "") or ""

        results.append({
            "sno": start_index + idx,
            "id": u.id,
            "username": u.username or "",
            "student_id": u.Employee_id or "",   # ✅ keep Student ID naming for JS
            "email": u.email or "",
            "department": dept_name,
            "department_code": dept_code,
            "role": role_name,
            "is_active": bool(u.is_active),
            "last_login": last_login_str,
            "profile": profile_url,
        })

    return JsonResponse({
        "results": results,
        "role_options": role_options,
        "stats": {
            "total": total_count,
            "active": active_count,
            "inactive": inactive_count,
            "role_counts": role_counts,
        },
        "pagination": {
            "page": page_obj.number,
            "page_size": page_size,
            "total": total_count,
            "total_pages": paginator.num_pages,
            "has_next": page_obj.has_next(),
            "has_prev": page_obj.has_previous(),
        }
    })


from django.contrib import messages
from django.core.files.base import ContentFile
from django.shortcuts import get_object_or_404, redirect, render

def edit_student_user_accounts(request, id):
    user = get_object_or_404(USER.objects.using("rit_approval_system"), pk=id)

    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        department_id = request.POST.get("department")
        role_id = request.POST.get("role")
        profile_img = request.FILES.get("profile_img")
        is_active = request.POST.get("is_active")

        if not username or not email:
            messages.error(request, "Username and Email are required.")
            return redirect("edit_student_user_accounts", id=id)

        user.username = username
        user.email = email

        if department_id:
            user.Department_id = department_id

        if role_id:
            user.role_id = role_id

        user.is_active = is_active in ["on", "1", "true", "True"]

        uploaded_file_name = None
        uploaded_file_bytes = None

        if profile_img:
            uploaded_file_name = profile_img.name
            uploaded_file_bytes = profile_img.read()

            # save image to USER model
            user.profile_img.save(
                uploaded_file_name,
                ContentFile(uploaded_file_bytes),
                save=False
            )

        user.save(using="rit_approval_system")

        student = None

        if email:
            student = StudentDetails.objects.filter(email=email).first()

        if not student and username:
            student = StudentDetails.objects.filter(reg_no=username).first()

        if student:
            student.email = email
            student.is_active = user.is_active

            if uploaded_file_name and uploaded_file_bytes:
                student.profile_img.save(
                    uploaded_file_name,
                    ContentFile(uploaded_file_bytes),
                    save=False
                )

            student.save()
            messages.success(request, f"User '{username}' and StudentDetails updated successfully.")
        else:
            messages.warning(
                request,
                f"User '{username}' updated successfully, but matching StudentDetails record was not found."
            )

        return redirect("student_user_accounts")

    departments = Department.objects.using("rit_approval_system").all()
    roles = Role.objects.using("rit_approval_system").all()

    return render(request, "admin/edit_student_user_accounts.html", {
        "user": user,
        "departments": departments,
        "roles": roles,
    })




# ============================
# ✅ PARENT ACCOUNTS (AJAX)
# (updated from your old DataTables page)
# ============================
def parent_user_accounts(request):
    """
    Page render only.
    Parent list & stats are loaded via AJAX from parent_user_accounts_data().
    """
    departments = Department.objects.using("rit_approval_system").all().order_by("Department")

    if request.method == "POST":
        user_id = request.POST.get("login_as")
        if user_id:
            user = get_object_or_404(USER.objects.using("rit_approval_system"), pk=user_id)
            user.backend = "user_accounts.auth.EmployeeIDBackend"
            auth_login(request, user)
            messages.success(request, f"You are now logged in as '{user.username}'.")
            return redirect("faculty_dashboard")
        return redirect("parent_user_accounts")

    return render(request, "admin/parent_user_accounts.html", {
        "departments": departments
    })


@login_required
@require_POST
def reset_parent_user_password(request, id):
    user = get_object_or_404(
        USER.objects.using("rit_approval_system").filter(is_parent=True, is_student=False),
        pk=id,
    )
    parent_id = user.Employee_id
    updated_count = USER.objects.using("rit_approval_system").filter(
        Employee_id=parent_id,
        is_parent=True,
        is_student=False,
    ).update(password=make_password("123"))

    return JsonResponse({
        "success": True,
        "message": (
            f"Password reset to 123 for {user.username} "
            f"({updated_count} account record(s) updated)."
        ),
    })


def parent_user_accounts_data(request):
    """
    AJAX endpoint:
    Params:
      q        : global search (any field)
      dept     : department name filter
      role     : role filter
      page     : page number
      page_size: page size

    Returns:
      results, pagination meta, stats, role_options
    """
    q = (request.GET.get("q") or "").strip()
    dept = (request.GET.get("dept") or "").strip()
    role = (request.GET.get("role") or "").strip()

    try:
        page = int(request.GET.get("page", 1))
    except Exception:
        page = 1

    try:
        page_size = int(request.GET.get("page_size", 25))
    except Exception:
        page_size = 25

    base_qs = (
        USER.objects.using("rit_approval_system")
        .filter(is_parent=True, is_student=False)   # ✅ parents only
        .select_related("Department", "role")
    )

    role_options = list(
        base_qs.exclude(role__role__isnull=True)
               .exclude(role__role__exact="")
               .values_list("role__role", flat=True)
               .distinct()
               .order_by("role__role")
    )

    qs = base_qs

    if dept:
        qs = qs.filter(Department__Department__iexact=dept)

    if role:
        qs = qs.filter(role__role__iexact=role)

    if q:
        qs = qs.filter(
            Q(username__icontains=q) |
            Q(Employee_id__icontains=q) |
            Q(email__icontains=q) |
            Q(Department__Department__icontains=q) |
            Q(Department__Department_code__icontains=q) |
            Q(role__role__icontains=q)
        )

    qs = qs.order_by("-is_active", "id")

    total_count = qs.count()
    active_count = qs.filter(is_active=True).count()
    inactive_count = qs.filter(is_active=False).count()

    role_counts_qs = (
        qs.values("role__role")
          .annotate(count=Count("id"))
          .order_by("-count", "role__role")
    )

    role_counts = []
    for rc in role_counts_qs:
        role_name = rc.get("role__role") or "No Role"
        role_counts.append({"role": role_name, "count": rc.get("count", 0)})

    paginator = Paginator(qs, page_size)
    try:
        page_obj = paginator.page(page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages if paginator.num_pages else 1)

    results = []
    start_index = (page_obj.number - 1) * page_size

    for idx, u in enumerate(page_obj.object_list, start=1):
        try:
            profile_url = u.profile_img.url if getattr(u, "profile_img", None) else ""
        except Exception:
            profile_url = ""

        last_login_str = "-"
        if u.last_login:
            last_login_str = u.last_login.strftime("%Y-%m-%d %H:%M")

        dept_name = ""
        dept_code = ""
        if getattr(u, "Department", None):
            dept_name = getattr(u.Department, "Department", "") or ""
            dept_code = getattr(u.Department, "Department_code", "") or ""

        role_name = ""
        if getattr(u, "role", None):
            role_name = getattr(u.role, "role", "") or ""

        results.append({
            "sno": start_index + idx,
            "id": u.id,
            "username": u.username or "",
            "parent_id": u.Employee_id or "",     # ✅ parent id naming for JS
            "email": u.email or "",
            "department": dept_name,
            "department_code": dept_code,
            "role": role_name,
            "is_active": bool(u.is_active),
            "last_login": last_login_str,
            "profile": profile_url,
        })

    return JsonResponse({
        "results": results,
        "role_options": role_options,
        "stats": {
            "total": total_count,
            "active": active_count,
            "inactive": inactive_count,
            "role_counts": role_counts,
        },
        "pagination": {
            "page": page_obj.number,
            "page_size": page_size,
            "total": total_count,
            "total_pages": paginator.num_pages,
            "has_next": page_obj.has_next(),
            "has_prev": page_obj.has_previous(),
        }
    })



def edit_parent_user_accounts(request, id):
    # Fetch the user from external DB
    user = get_object_or_404(USER.objects.using("rit_approval_system"), pk=id)

    if request.method == "POST":
        # Editable fields
        username = request.POST.get("username")
        email = request.POST.get("email")
        department_id = request.POST.get("department")
        role_id = request.POST.get("role")
        profile_img = request.FILES.get("profile_img")
        is_active = request.POST.get("is_active")  # checkbox or dropdown

        # Validation
        if not username or not email:
            messages.error(request, "Username and Email are required.")
            return redirect("edit_staff_user_accounts", id=id)

        # Update fields
        user.username = username
        user.email = email

        if department_id:
            user.Department_id = department_id
        if role_id:
            user.role_id = role_id
        if profile_img:
            user.profile_img = profile_img

        # Handle is_active checkbox
        user.is_active = True if is_active == "on" or is_active == "1" else False

        # Save to external DB
        user.save(using="rit_approval_system")

        messages.success(request, f"User '{username}' updated successfully.")
        return redirect("parent_user_accounts")

    # Send departments and roles for dropdown
    departments = Department.objects.using("rit_approval_system").all()
    roles = Role.objects.using("rit_approval_system").all()

    return render(request, "admin/edit_parent_user_accounts.html", {
        "user": user,
        "departments": departments,
        "roles": roles,
    })








from user_accounts.decorators import faculty_login_required

@no_cache
# @faculty_login_required
def switch_user(request):
    if request.method == 'POST':
        unique_id = request.POST.get('unique_id')
        try:
            super_user = request.user.is_superuser
            selected_user = USER.objects.using("rit_approval_system").get(unique_id=unique_id)

            if request.user.unique_id != selected_user.unique_id:
                app = request.session.get('app_name', 'Approval System')

                auth_logout(request)
                # request.session.flush()
                request.session["app_name"] = app

                selected_user.backend = 'user_accounts.auth.EmployeeIDBackend'
                login(request, selected_user)
                messages.success(request, "Authentication successful. You are now logged in.")

                # Redirect based on role
                if selected_user.is_student:
                    if super_user:
                        return redirect('home')
                    else:
                        return redirect('student_dashboard')  # create a student home page
                else:
                    if super_user:
                        request.session["employee_id"] = selected_user.Employee_id
                        return redirect('home')
                    else:
                        return redirect('faculty_dashboard')

            else:
                messages.error(request, "You're already logged into this account.")

        except USER.DoesNotExist:
            return HttpResponse("User not found", status=404)

    previous_url = request.META.get('HTTP_REFERER', '/')
    # previous_url = "faculty_dashboard" if request.user.is_staff or request.user.is_superuser else "student_dashboard"
    return redirect(previous_url)


def degree_departments(request):
    # Fetch all departments
    departments = Department.objects.using("rit_approval_system").all().order_by('Department')
    degree_departments = DegreeDepartment.objects.all().order_by("degree_department")

    # Handle Add/Edit form submission
    if request.method == "POST":
        action = request.POST.get("action")
        degree_code = request.POST.get("degree_code")
        selected_departments = request.POST.getlist("department_ids")  # Checkbox list
        dd_id = request.POST.get("dd_id")  # For edit

        if action == "bulk_add":
            if not degree_code or not selected_departments:
                messages.error(request, "Please select a degree and at least one department.")
                return redirect("degree_departments")

            count = 0
            for dept_id in selected_departments:
                obj, created = DegreeDepartment.objects.update_or_create(
                    degree_code=degree_code,
                    department_id=dept_id,
                    defaults={}
                )
                count += 1
            messages.success(request, f"✅ {count} Degree Department(s) added/updated successfully.")
            return redirect("degree_departments")

        elif action == "edit":
            if dd_id and degree_code and selected_departments:
                dept_id = selected_departments[0]  # Only one department for edit
                obj = get_object_or_404(DegreeDepartment, id=dd_id)
                obj.degree_code = degree_code
                obj.department_id = dept_id
                obj.save()
                messages.success(request, "✅ Degree Department updated successfully.")
            return redirect("degree_departments")

    # Handle Delete
    delete_id = request.GET.get("delete_id")
    if delete_id:
        obj = get_object_or_404(DegreeDepartment, id=delete_id)
        obj.delete()
        messages.warning(request, "🗑️ Degree Department deleted successfully.")
        return redirect("degree_departments")

    context = {
        "departments": departments,
        "degree_departments": degree_departments,
    }
    return render(request, "admin/degree_departments.html", context)



@is_super_user('admin_management')
def add_degree(request):
    """
    Page render + Add/Edit/Delete (POST).
    List is loaded via AJAX from add_degree_data().
    """
    if request.method == "POST":
        # ✅ ADD / EDIT
        if ("add_degree" in request.POST) or ("edit_degree" in request.POST):
            degree_id = (request.POST.get("degree_id") or "").strip()  # empty for add
            degree_code = (request.POST.get("degree_code") or "").strip()
            degree_name = (request.POST.get("degree") or "").strip()
            duration = (request.POST.get("duration") or "").strip()
            degree_graduate = (request.POST.get("degree_graduate") or "").strip()
            is_active = True if request.POST.get("is_active") == "on" else False

            try:
                if degree_id:
                    Degree.objects.filter(id=degree_id).update(
                        degree_code=degree_code,
                        degree=degree_name,
                        duration=duration,
                        degree_graduate=degree_graduate,
                        is_active=is_active,
                    )
                    messages.success(request, f"Degree '{degree_name}' updated successfully.")
                else:
                    obj, created = Degree.objects.update_or_create(
                        degree_code=degree_code,
                        defaults={
                            "degree": degree_name,
                            "duration": duration,
                            "degree_graduate": degree_graduate,
                            "is_active": is_active,
                        }
                    )
                    if created:
                        messages.success(request, f"Degree '{degree_name}' added successfully.")
                    else:
                        messages.info(request, f"Degree '{degree_name}' already exists. Updated the details.")

            except Exception as e:
                messages.error(request, f"Error saving degree: {str(e)}")

            return redirect("add_degree")

        # 🗑️ DELETE
        if "delete_degree" in request.POST:
            degree_id = request.POST.get("degree_id")
            try:
                degree = Degree.objects.get(id=degree_id)
                degree_name = degree.degree
                degree.delete()
                messages.success(request, f"Degree '{degree_name}' deleted successfully.")
            except Degree.DoesNotExist:
                messages.error(request, "Degree not found.")
            except Exception as e:
                messages.error(request, f"Error deleting degree: {str(e)}")
            return redirect("add_degree")

    # ✅ Page render only
    return render(request, "admin/add_degree.html")
 


@is_super_user('admin_management')
def add_degree_data(request):
    """
    AJAX endpoint for degrees list:
    Params:
      q         : search any field
      active    : all / 1 / 0
      graduate  : graduate type filter (exact)
      page      : page number
      page_size : page size

    Returns:
      results, pagination, filters (graduate_options)
    """
    q = (request.GET.get("q") or "").strip()
    active = (request.GET.get("active") or "").strip()     # "", "1", "0"
    graduate = (request.GET.get("graduate") or "").strip()

    try:
        page = int(request.GET.get("page", 1))
    except Exception:
        page = 1

    try:
        page_size = int(request.GET.get("page_size", 25))
    except Exception:
        page_size = 25

    base_qs = Degree.objects.all().order_by("-is_active", "degree_code", "id")

    graduate_options = list(
        base_qs.exclude(degree_graduate__isnull=True)
              .exclude(degree_graduate__exact="")
              .values_list("degree_graduate", flat=True)
              .distinct()
              .order_by("degree_graduate")
    )

    qs = base_qs

    if active == "1":
        qs = qs.filter(is_active=True)
    elif active == "0":
        qs = qs.filter(is_active=False)

    if graduate:
        qs = qs.filter(degree_graduate__iexact=graduate)

    if q:
        qs = qs.filter(
            Q(degree_code__icontains=q) |
            Q(degree__icontains=q) |
            Q(duration__icontains=q) |
            Q(degree_graduate__icontains=q) |
            Q(is_active__icontains=q)
        )

    total = qs.count()

    paginator = Paginator(qs, page_size)
    try:
        page_obj = paginator.page(page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages if paginator.num_pages else 1)

    results = []
    start_index = (page_obj.number - 1) * page_size

    for idx, d in enumerate(page_obj.object_list, start=1):
        results.append({
            "sno": start_index + idx,
            "id": d.id,
            "degree_code": d.degree_code or "",
            "degree": d.degree or "",
            "duration": d.duration,
            "degree_graduate": d.degree_graduate or "",
            "is_active": bool(d.is_active),
        })

    return JsonResponse({
        "results": results,
        "graduate_options": graduate_options,
        "pagination": {
            "page": page_obj.number,
            "page_size": page_size,
            "total": total,
            "total_pages": paginator.num_pages,
            "has_next": page_obj.has_next(),
            "has_prev": page_obj.has_previous(),
        }
    })
 
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator, EmptyPage
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
 
@is_super_user('admin_management')
def export_degree_excel(request):
    q = (request.GET.get("q") or "").strip()
    active = (request.GET.get("active") or "").strip()
    graduate = (request.GET.get("graduate") or "").strip()

    qs = Degree.objects.all().order_by("-is_active", "degree_code", "id")

    if active == "1":
        qs = qs.filter(is_active=True)
    elif active == "0":
        qs = qs.filter(is_active=False)

    if graduate:
        qs = qs.filter(degree_graduate__iexact=graduate)

    if q:
        qs = qs.filter(
            Q(degree_code__icontains=q) |
            Q(degree__icontains=q) |
            Q(duration__icontains=q) |
            Q(degree_graduate__icontains=q) |
            Q(is_active__icontains=q)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Degrees"

    headers = ["degree_code", "degree", "duration", "graduate"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for d in qs:
        ws.append([
            d.degree_code or "",
            d.degree or "",
            d.duration if d.duration is not None else "",
            d.degree_graduate or "",
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_length:
                max_length = len(val)
        ws.column_dimensions[col_letter].width = max_length + 3

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="degree_export.xlsx"'

    wb.save(response)
    return response

from django.db import transaction
from django.db import transaction, IntegrityError


def _maybe_replace_file(instance, field_name, new_file):
    """Replace a FileField if a new file is provided and remove the old file."""
    if new_file:
        old = getattr(instance, field_name)
        if old and hasattr(old, "delete"):
            old.delete(save=False)
        setattr(instance, field_name, new_file)

@is_super_user('admin_management')
def add_department(request):
    """
    Page render + Add/Edit/Delete (POST).
    List is loaded via AJAX from add_department_data().
    """
    degrees = Degree.objects.filter(is_active=True).order_by("degree_code", "degree")

    # ADD / EDIT
    if request.method == "POST" and ("add_department" in request.POST or "edit_department" in request.POST):
        department_id = (request.POST.get("department_id") or "").strip()
        degree_id = (request.POST.get("degree") or "").strip()

        department_name = (request.POST.get("Department") or "").strip()
        department_code = (request.POST.get("Department_code") or "").strip()
        department_label = (request.POST.get("department_label") or "").strip()
        is_active = True if request.POST.get("is_active") == "on" else False
        is_academic = True if request.POST.get("is_academic") == "on" else False

        degree = None
        if degree_id:
            degree = get_object_or_404(Degree, id=degree_id, is_active=True)

        try:
            with transaction.atomic():
                if department_id:
                    dept = get_object_or_404(Add_Department, id=department_id)
                    dept.degree = degree
                    dept.Department = department_name
                    dept.Department_code = department_code
                    dept.department_label = department_label
                    dept.is_active = is_active
                    dept.is_academic = is_academic
                    dept.save()

                    messages.success(request, f"Program '{dept.Department or dept.Department_code}' updated successfully.")
                else:
                    if not department_code:
                        messages.error(request, "Program Code is required.")
                        return redirect("add_department")

                    obj, created = Add_Department.objects.update_or_create(
                        degree=degree,
                        Department_code=department_code,
                        defaults={
                            "Department": department_name,
                            "department_label": department_label,
                            "is_active": is_active,
                            "is_academic": is_academic,
                        }
                    )
                    if created:
                        messages.success(request, f"Program '{obj.Department or obj.Department_code}' created successfully.")
                    else:
                        messages.success(request, f"Program '{obj.Department or obj.Department_code}' updated successfully.")

            if department_code:
                try:
                    Department.objects.using("rit_approval_system").update_or_create(
                        Department_code=department_code,
                        defaults={"Department": department_name or department_code},
                    )
                except Exception:
                    pass

        except Exception as e:
            messages.error(request, f"Error saving program: {e}")

        return redirect("add_department")

    # DELETE
    if request.method == "POST" and "delete_department" in request.POST:
        dep_id = request.POST.get("department_id")
        try:
            dept = get_object_or_404(Add_Department, id=dep_id)
            name = dept.Department or dept.Department_code or f"ID {dept.id}"
            dept.delete()
            messages.success(request, f"Program '{name}' deleted successfully.")
        except Exception as e:
            messages.error(request, f"Error deleting program: {e}")
        return redirect("add_department")

    return render(request, "admin/add_department.html", {
        "degrees": degrees
    })


@is_super_user('admin_management')
def add_department_data(request):
    q = (request.GET.get("q") or "").strip()
    degree_id = (request.GET.get("degree") or "").strip()
    active = (request.GET.get("active") or "").strip()

    try:
        page = int(request.GET.get("page", 1))
    except Exception:
        page = 1

    try:
        page_size = int(request.GET.get("page_size", 25))
    except Exception:
        page_size = 25

    qs = Add_Department.objects.select_related("degree").all().order_by(
    "-is_academic",   # Academic first
    "-is_active",     # Active first
    "Department",     # Then alphabetically
    "Department_code",
    "id"
)

    if degree_id:
        qs = qs.filter(degree__id=degree_id)

    if active == "1":
        qs = qs.filter(is_active=True)
    elif active == "0":
        qs = qs.filter(is_active=False)

    if q:
        qs = qs.filter(
            Q(Department__icontains=q) |
            Q(Department_code__icontains=q) |
            Q(department_label__icontains=q) |
            Q(degree__degree__icontains=q) |
            Q(degree__degree_code__icontains=q)
        )

    total = qs.count()

    paginator = Paginator(qs, page_size)
    try:
        page_obj = paginator.page(page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages if paginator.num_pages else 1)

    start_index = (page_obj.number - 1) * page_size
    results = []

    for idx, dep in enumerate(page_obj.object_list, start=1):
        degree_text = "-"
        degree_id_val = ""
        if dep.degree:
            degree_text = f"{dep.degree.degree_code} - {dep.degree.degree}"
            degree_id_val = dep.degree.id

        results.append({
            "sno": start_index + idx,
            "id": dep.id,
            "degree_id": degree_id_val,
            "degree_text": degree_text,
            "Department": dep.Department or "",
            "Department_code": dep.Department_code or "",
            "department_label": dep.department_label or "",
            "is_active": bool(dep.is_active),
            "is_academic": bool(dep.is_academic),
            "created_at": dep.created_at.strftime("%Y-%m-%d %H:%M") if dep.created_at else "-",
            "updated_at": dep.updated_at.strftime("%Y-%m-%d %H:%M") if dep.updated_at else "-",
        })

    return JsonResponse({
        "results": results,
        "pagination": {
            "page": page_obj.number,
            "page_size": page_size,
            "total": total,
            "total_pages": paginator.num_pages,
            "has_next": page_obj.has_next(),
            "has_prev": page_obj.has_previous(),
        }
    })
   
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

@is_super_user('admin_management')
def export_department_excel(request):
    q = (request.GET.get("q") or "").strip()
    degree_id = (request.GET.get("degree") or "").strip()
    active = (request.GET.get("active") or "").strip()

    # ONLY Academic departments
    qs = Add_Department.objects.select_related("degree").filter(is_academic=True).order_by(
        "Department",
        "Department_code",
        "id"
    )

    if degree_id:
        qs = qs.filter(degree__id=degree_id)

    if active == "1":
        qs = qs.filter(is_active=True)
    elif active == "0":
        qs = qs.filter(is_active=False)

    if q:
        qs = qs.filter(
            Q(Department__icontains=q) |
            Q(Department_code__icontains=q) |
            Q(department_label__icontains=q) |
            Q(degree__degree__icontains=q) |
            Q(degree__degree_code__icontains=q)
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Academic Departments"

    # ✅ Correct 4 headers
    headers = ["degree", "department_code", "department_name", "department_label"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    # ✅ Only 4 values (matching headers)
    for dep in qs:
        degree_text = ""
        if dep.degree:
            degree_text = f"{dep.degree.degree_code}"

        ws.append([
            degree_text,
            dep.Department_code or "",
            dep.Department or "",
            dep.department_label or "",
        ])

    # Auto width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = max_length + 3

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="academic_department_export.xlsx"'

    wb.save(response)
    return response


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.contrib.contenttypes.models import ContentType
from django.db import IntegrityError

EXTERNAL_DB = "rit_approval_system"

@require_http_methods(["GET", "POST"])
def curd_approvals(request):

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()

        if action in ("create", "update"):
            rule_id = request.POST.get("rule_id")

            content_type_id = request.POST.get("content_type_id")
            action_code = request.POST.get("action_code")
            requester_role_id = request.POST.get("requester_role_id")
            approver_role_id = request.POST.get("approver_role_id")
            is_active = True if request.POST.get("is_active") == "1" else False

            if not all([content_type_id, action_code, requester_role_id, approver_role_id]):
                messages.error(request, "Please fill all required fields.")
                return redirect("curd_approvals")

            # cast ids
            try:
                requester_role_id_int = int(requester_role_id)
                approver_role_id_int = int(approver_role_id)
            except ValueError:
                messages.error(request, "Invalid role id.")
                return redirect("curd_approvals")

            ct = get_object_or_404(ContentType, id=content_type_id)

            # ✅ validate role ids exist in external DB
            if not Role.objects.using(EXTERNAL_DB).filter(id=requester_role_id_int).exists():
                messages.error(request, "Requester role not found in external DB.")
                return redirect("curd_approvals")

            if not Role.objects.using(EXTERNAL_DB).filter(id=approver_role_id_int).exists():
                messages.error(request, "Approver role not found in external DB.")
                return redirect("curd_approvals")

            try:
                if action == "create":
                    CURD_Approval.objects.create(
                        content_type=ct,
                        action=action_code,
                        requester_role_id=requester_role_id_int,
                        approver_role_id=approver_role_id_int,
                        is_active=is_active,
                    )
                    messages.success(request, "Approval rule created successfully.")
                else:
                    obj = get_object_or_404(CURD_Approval, id=rule_id)
                    obj.content_type = ct
                    obj.action = action_code
                    obj.requester_role_id = requester_role_id_int
                    obj.approver_role_id = approver_role_id_int
                    obj.is_active = is_active
                    obj.save()
                    messages.success(request, "Approval rule updated successfully.")

            except IntegrityError:
                messages.error(
                    request,
                    "This rule already exists (content type + action + requester + approver must be unique)."
                )

            return redirect("curd_approvals")

        if action == "toggle":
            rule_id = request.POST.get("rule_id")
            obj = get_object_or_404(CURD_Approval, id=rule_id)
            obj.is_active = not obj.is_active
            obj.save(update_fields=["is_active"])
            messages.success(request, f"Rule {'activated' if obj.is_active else 'deactivated'} successfully.")
            return redirect("curd_approvals")

        if action == "delete":
            rule_id = request.POST.get("rule_id")
            obj = get_object_or_404(CURD_Approval, id=rule_id)
            obj.delete()
            messages.success(request, "Rule deleted successfully.")
            return redirect("curd_approvals")

        messages.error(request, "Invalid action.")
        return redirect("curd_approvals")

    # GET
    content_types = ContentType.objects.all().order_by("app_label", "model")

    roles_qs = Role.objects.using(EXTERNAL_DB).all().order_by("role_name")
    role_map = {r.id: str(r) for r in roles_qs}  # or r.role_name

    rules = CURD_Approval.objects.select_related("content_type").order_by("-id")

    return render(request, "admin/curd_approvals.html", {
        "content_types": content_types,
        "roles": roles_qs,          # for dropdown
        "role_map": role_map,       # for table display
        "rules": rules,
        "ACTION_CHOICES": CURD_Approval.ACTION_CHOICES,
    })





from django.shortcuts import render, redirect
from django.contrib import messages
from user_accounts.models import GlobalUsers
from user_accounts.models import USER

def create_global_users(request):
    
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        is_global = request.POST.get("is_global")

        user = (
            USER.objects.using("rit_approval_system")
            .select_related("role")
            .get(id=user_id)
        )
        employee_id = str(user.Employee_id or "").strip()
        role_id = str(user.role_id or "").strip()
        is_global_user = is_global == "1"

        updated_count = GlobalUsers.objects.filter(
            employee_id=employee_id,
            role_id=role_id,
        ).update(global_user=is_global_user)

        if updated_count == 0:
            GlobalUsers.objects.create(
                employee_id=employee_id,
                role_id=role_id,
                global_user=is_global_user,
            )

        messages.success(request, "Global user status updated successfully!")
        return redirect("create_global_users")

    global_user_keys = {
        (str(employee_id or "").strip(), str(role_id or "").strip())
        for employee_id, role_id in GlobalUsers.objects.filter(global_user=True)
        .values_list("employee_id", "role_id")
    }

    users = list(
        USER.objects.using("rit_approval_system")
        .select_related("Department", "role")
        .filter(is_student=False, is_parent=False)
    )

    for user in users:
        user.is_global_user = (
            str(user.Employee_id or "").strip(),
            str(user.role_id or "").strip(),
        ) in global_user_keys

    users.sort(
        key=lambda u: (
            not u.is_global_user,
            str(u.Department.Department if u.Department else ""),
            str(u.username or "")
        )
    )

    departments = Add_Department.objects.all().order_by("Department")

    return render(request, "admin/create_global_users.html", {
        "users": users,
        "departments": departments,
        "global_user_keys": global_user_keys,
    })


from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator, EmptyPage
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


@is_super_user('admin_management')
def role_view(request):
    return render(request, "admin/role_view.html")


@is_super_user('admin_management')
def role_view_data(request):
    q = (request.GET.get("q") or "").strip()

    try:
        page = int(request.GET.get("page", 1))
    except Exception:
        page = 1

    try:
        page_size = int(request.GET.get("page_size", 25))
    except Exception:
        page_size = 25

    qs = Role.objects.using("rit_approval_system").all().order_by("role", "id")

    if q:
        qs = qs.filter(Q(role__icontains=q))

    total = qs.count()

    paginator = Paginator(qs, page_size)
    try:
        page_obj = paginator.page(page)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages if paginator.num_pages else 1)

    start_index = (page_obj.number - 1) * page_size
    results = []

    for idx, r in enumerate(page_obj.object_list, start=1):
        results.append({
            "sno": start_index + idx,
            "id": r.id,
            "role": r.role or "",
        })

    return JsonResponse({
        "results": results,
        "pagination": {
            "page": page_obj.number,
            "page_size": page_size,
            "total": total,
            "total_pages": paginator.num_pages,
            "has_next": page_obj.has_next(),
            "has_prev": page_obj.has_previous(),
        }
    })


@is_super_user('admin_management')
def export_role_excel(request):
    q = (request.GET.get("q") or "").strip()

    qs = Role.objects.using("rit_approval_system").all().order_by("role", "id")

    if q:
        qs = qs.filter(Q(role__icontains=q))

    wb = Workbook()
    ws = wb.active
    ws.title = "Roles"

    headers = ["role"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in qs:
        ws.append([r.role or ""])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = max_length + 3

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="role_export.xlsx"'
    wb.save(response)
    return response




